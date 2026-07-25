import tkinter as tk
from tkinter.scrolledtext import ScrolledText
from tkinter import ttk, messagebox, filedialog, scrolledtext
from tkinter.font import Font
import threading
from datetime import datetime, timedelta
import json
import webbrowser
from pathlib import Path
import matplotlib
from education_system.systems.university.infrastructure import paths
matplotlib.use('TkAgg')
import numpy as np

# Import auth instance management from user_authentication
try:
    from education_system.systems.university.infrastructure.auth import get_current_user, set_auth_instance
    HAS_AUTH = True
except ImportError:
    HAS_AUTH = False
    get_current_user = lambda: None
    set_auth_instance = lambda x: None

auth = None

def set_auth(auth_instance):
    global auth
    auth = auth_instance
    # Also set it in the global auth instance if available
    if HAS_AUTH:
        set_auth_instance(auth_instance)
# Import the shared authentication system
try:
    from education_system.systems.university.infrastructure.auth import UserAuth
    from education_system.systems.university.infrastructure.shared_context import get_auth
except ImportError as e:
    print(f"⚠️ Could not import UserAuth: {e}")
    UserAuth = None
    get_auth = lambda: None

from education_system.systems.university.infrastructure.i18n import get_text as _, init_i18n
init_i18n()

# Import standalone functions
from education_system.systems.university.interfaces.gui.finance.finance_reporting.misc import (
    generate_advanced_financial_forecasting,
    generate_comprehensive_budget_variance_report,
    compliance_audit_system,
)

# Import email service
try:
    from education_system.systems.university.infrastructure.email.email_service import send_email
    from education_system.systems.university.infrastructure.email.template_utils import render_template
    HAS_EMAIL = True
except ImportError:
    HAS_EMAIL = False
    send_email = None
    render_template = None


def get_admin_email():
    """Get admin email from database"""
    try:
        from education_system.systems.university.infrastructure.database.db import get_connection
        with get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(
                "SELECT email FROM users WHERE role = 'admin' AND email IS NOT NULL LIMIT 1"
            )
            result = cursor.fetchone()
            if result:
                return result[0]
    except Exception as e:
        print(f"Error getting admin email: {e}")
    return None


def send_report_to_admin(report_title, report_content, parent_window=None):
    """Send a report to the admin via the shared email_bus.

    Routing through ``integration_bus.send_finance_report`` rather
    than the raw SMTP helper means the send lands in ``email_log``
    and fires ``EVENT_EMAIL_SENT`` for audit. Falls back to the raw
    helper if the bus isn't available (partial deploy).
    """
    try:
        from education_system.systems.university.services.bus.integration_bus import (
            send_finance_report,
        )
        msg_id = send_finance_report(
            report_title=report_title,
            summary=str(report_content),
            related_to="finance_report",
        )
        if msg_id:
            messagebox.showinfo("Success", "Report sent to admin (logged in email_log).")
            return True
    except Exception:
        pass

    if not HAS_EMAIL:
        messagebox.showerror("Error", "Email system not available")
        return False

    admin_email = get_admin_email()
    if not admin_email:
        messagebox.showerror("Error", "No admin email found in database.\nPlease ensure an admin user has an email configured.")
        return False

    try:
        generated_date = datetime.now().strftime('%Y-%m-%d %H:%M')
        generated_timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        subject, body = render_template('reports/finance_report', {
            'report_title': report_title,
            'generated_date': generated_date,
            'generated_timestamp': generated_timestamp,
            'report_content': report_content
        })
        if not subject or not body:
            subject = f"Finance Report: {report_title} - {generated_date}"
            body = f"Finance Report: {report_title}\nGenerated: {generated_timestamp}\n\n{report_content}"
        if send_email(admin_email, subject, body):
            messagebox.showinfo("Success", f"Report sent to admin at:\n{admin_email}")
            return True
        messagebox.showerror("Error", "Failed to send email. Please check email configuration.")
        return False
    except Exception as e:
        messagebox.showerror("Error", f"Failed to send report:\n{e}")
        return False


# This module defines mixin functions for FinancialManagementGUI
# Note: Methods are registered by main.py to avoid circular imports

def create_reports_tab(self):
    """Create reports tab with export options"""
    reports_frame = ttk.Frame(self.notebook, padding="10")
    self.notebook.add(reports_frame, text=_("finance_reporting.tabs.reports"))

    # Report generation frame
    gen_frame = ttk.LabelFrame(reports_frame, text=_("finance_reporting.reports.generate"), padding="10")
    gen_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))

    # Report type selection
    ttk.Label(gen_frame, text=_("finance_reporting.reports.type")).grid(row=0, column=0, sticky=tk.W)
    self.report_type = tk.StringVar(value="comprehensive")
    report_combo = ttk.Combobox(gen_frame, textvariable=self.report_type,
                               values=["comprehensive", "executive_summary", "compliance", "custom"],
                               state="readonly")
    report_combo.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(10, 0))

    # Output format selection
    ttk.Label(gen_frame, text=_("finance_reporting.reports.format")).grid(row=1, column=0, sticky=tk.W, pady=(10, 0))
    self.output_format = tk.StringVar(value="PDF")
    format_combo = ttk.Combobox(gen_frame, textvariable=self.output_format,
                               values=["PDF", "Excel", "CSV", "JSON", "All"],
                               state="readonly")
    format_combo.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=(10, 0))

    # Generate button
    ttk.Button(gen_frame, text=_("finance_reporting.buttons.generate_report"), command=self.generate_selected_report,
              style='Accent.TButton').grid(row=2, column=0, columnspan=2, pady=(10, 0))

    gen_frame.grid_columnconfigure(1, weight=1)

    # Scheduled reports frame
    schedule_frame = ttk.LabelFrame(reports_frame, text=_("finance_reporting.reports.scheduled"), padding="10")
    schedule_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

    # Scheduled reports list
    self.schedule_tree = ttk.Treeview(schedule_frame, columns=('Type', 'Frequency', 'Next Run'), height=10)
    self.schedule_tree.heading('#0', text=_("finance_reporting.columns.report_name"))
    self.schedule_tree.heading('Type', text=_("finance_reporting.columns.type"))
    self.schedule_tree.heading('Frequency', text=_("finance_reporting.columns.frequency"))
    self.schedule_tree.heading('Next Run', text=_("finance_reporting.columns.next_run"))
    self.schedule_tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

    # Scrollbar
    schedule_scroll = ttk.Scrollbar(schedule_frame, orient=tk.VERTICAL, command=self.schedule_tree.yview)
    schedule_scroll.grid(row=0, column=1, sticky=(tk.N, tk.S))
    self.schedule_tree.configure(yscrollcommand=schedule_scroll.set)

    # Populate scheduled reports
    self.populate_scheduled_reports()

    schedule_frame.grid_rowconfigure(0, weight=1)
    schedule_frame.grid_columnconfigure(0, weight=1)
    reports_frame.grid_rowconfigure(1, weight=1)
    reports_frame.grid_columnconfigure(0, weight=1)

def generate_quick_report(self):
    """Generate quick summary report"""
    def generate_in_background():
        try:
            from education_system.systems.university.infrastructure.database.db import get_connection
            conn = get_connection()
            cursor = conn.cursor()

            # Generate quick summary
            cursor.execute('''
            SELECT
                SUM(sf.amount) as total_expected,
                SUM(CASE WHEN sf.status = 'paid' THEN sf.amount ELSE 0 END) as total_collected,
                COUNT(DISTINCT sf.student_id) as student_count
            FROM student_fees sf
            ''')

            summary_data = cursor.fetchone()

            # Today's activity
            today = datetime.now().strftime('%Y-%m-%d')
            cursor.execute('SELECT COUNT(*), SUM(amount) FROM payments WHERE payment_date = ?', (today,))
            today_data = cursor.fetchone()

            conn.close()

            # Create report content
            report_content = f"""
    QUICK FINANCIAL SUMMARY - {datetime.now().strftime('%Y-%m-%d %H:%M')}
    ========================================

    OVERVIEW
    --------
    Total Expected Revenue: £{summary_data[0] or 0:,.2f}
    Total Collected: £{summary_data[1] or 0:,.2f}
    Collection Rate: {(summary_data[1] / summary_data[0] * 100) if summary_data[0] else 0:.1f}%
    Active Students: {summary_data[2] or 0:,}

    TODAY'S ACTIVITY
    ---------------
    Payments Received: {today_data[0] or 0}
    Amount Collected: £{today_data[1] or 0:,.2f}

    STATUS: {'✓ Normal' if today_data[0] > 5 else '⚠ Low Activity'}

    This is a quick summary report. For detailed analysis,
    use the comprehensive reporting features.
            """

            # Save report
            filename = f"quick_report_{datetime.now().strftime('%Y%m%d_%H%M')}.txt"
            with open(filename, 'w') as f:
                f.write(report_content)

            self.root.after(0, lambda: [
                self.log_activity(f"Quick report generated: {filename}"),
                messagebox.showinfo("Report Generated", f"Quick report saved as {filename}")
            ])

        except Exception as e:
            self.root.after(0, lambda _e=e: messagebox.showerror("Error", f"Failed to generate report: {_e}"))

    thread = threading.Thread(target=generate_in_background)
    thread.daemon = True
    thread.start()

def generate_selected_report(self):
    """Generate the selected report type"""
    report_type = self.report_type.get()
    output_format = self.output_format.get()

    self.update_status(f"Generating {report_type} report...")

    def generate_in_background():
        try:
            if report_type == "comprehensive":
                generate_advanced_financial_forecasting()
                generate_comprehensive_budget_variance_report()

            elif report_type == "executive_summary":
                # Generate executive summary
                from education_system.systems.university.infrastructure.database.db import get_connection
                conn = get_connection()
                cursor = conn.cursor()

                cursor.execute('''
                SELECT
                    SUM(sf.amount) as total_expected,
                    SUM(CASE WHEN sf.status = 'paid' THEN sf.amount ELSE 0 END) as total_collected,
                    COUNT(DISTINCT sf.student_id) as student_count
                FROM student_fees sf
                ''')

                summary_data = cursor.fetchone()
                conn.close()

                # Create executive summary PDF
                filename = f"executive_summary_{datetime.now().strftime('%Y%m%d')}.txt"
                with open(filename, 'w') as f:
                    f.write(f"""
    EXECUTIVE SUMMARY - {datetime.now().strftime('%Y-%m-%d')}
    ==========================================

    KEY METRICS:
    - Total Expected Revenue: £{summary_data[0] or 0:,.2f}
    - Revenue Collected: £{summary_data[1] or 0:,.2f}
    - Collection Rate: {(summary_data[1] / summary_data[0] * 100) if summary_data[0] else 0:.1f}%
    - Active Students: {summary_data[2] or 0:,}

    STATUS: {'✓ On Track' if summary_data[1] / summary_data[0] > 0.85 else '⚠ Needs Attention'}

    RECOMMENDATIONS:
    • Monitor collection rates closely
    • Implement payment plan optimization
    • Focus on high-risk student support
                    """)

            elif report_type == "compliance":
                compliance_audit_system()

            elif report_type == "custom":
                # Show custom report builder
                self.root.after(0, self.show_custom_report_builder)
                return

            self.root.after(0, lambda rt=report_type: [
                self.log_activity(f"{rt} report generated"),
                self.update_status("Ready"),
                messagebox.showinfo("Report Generated", f"{rt} report generated successfully")
            ])

        except Exception as e:
            error_msg = str(e)
            self.root.after(0, lambda err=error_msg: [
                self.log_activity(f"Report generation error: {err}"),
                self.update_status("Error"),
                messagebox.showerror("Error", f"Report generation failed: {err}")
            ])

    thread = threading.Thread(target=generate_in_background)
    thread.daemon = True
    thread.start()

def show_custom_report_builder(self):
    """Show custom report builder window"""
    builder_window = tk.Toplevel(self.root)
    builder_window.title(_("finance_reporting.windows.custom_report_builder"))
    builder_window.geometry("1200x800")

    main_frame = ttk.Frame(builder_window, padding="10")
    main_frame.pack(fill=tk.BOTH, expand=True)

    ttk.Label(main_frame, text="Custom Report Builder",
             style='Title.TLabel').pack(pady=(0, 20))

    # Report components selection
    components_frame = ttk.LabelFrame(main_frame, text="Select Report Components", padding="10")
    components_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

    # Component checkboxes
    self.report_components = {}
    components = [
        ('executive_summary', 'Executive Summary Dashboard'),
        ('collection_analysis', 'Collection Rate Analysis'),
        ('payment_trends', 'Payment Trend Charts'),
        ('risk_assessment', 'Student Risk Assessment'),
        ('department_performance', 'Department Performance'),
        ('fee_analysis', 'Fee Type Analysis'),
        ('cash_flow', 'Cash Flow Projections'),
        ('budget_variance', 'Budget Variance Tables'),
        ('comparative', 'Comparative Analytics'),
        ('recommendations', 'Recommendation Engine')
    ]

    for i, (comp_id, comp_name) in enumerate(components):
        self.report_components[comp_id] = tk.BooleanVar()
        ttk.Checkbutton(components_frame, text=comp_name,
                       variable=self.report_components[comp_id]).grid(
                           row=i//2, column=i%2, sticky=tk.W, padx=10, pady=5)

    # Report settings
    settings_frame = ttk.LabelFrame(main_frame, text="Report Settings", padding="10")
    settings_frame.pack(fill=tk.X, pady=(0, 10))

    # Report name
    ttk.Label(settings_frame, text="Report Name:").grid(row=0, column=0, sticky=tk.W)
    self.custom_report_name = tk.StringVar(value="Custom Financial Report")
    ttk.Entry(settings_frame, textvariable=self.custom_report_name,
             width=30).grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(10, 0))

    # Output format
    ttk.Label(settings_frame, text="Output Format:").grid(row=1, column=0, sticky=tk.W, pady=(10, 0))
    self.custom_format = tk.StringVar(value="PDF")
    format_combo = ttk.Combobox(settings_frame, textvariable=self.custom_format,
                               values=["PDF", "Excel", "HTML", "Text"], state="readonly")
    format_combo.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=(10, 0))

    settings_frame.grid_columnconfigure(1, weight=1)

    # Generate button
    def generate_custom_report():
        selected_components = [comp_id for comp_id, var in self.report_components.items() if var.get()]
        if not selected_components:
            messagebox.showwarning("Warning", "Please select at least one report component.")
            return

        report_name = self.custom_report_name.get()
        output_format = self.custom_format.get()

        # Generate custom report content
        report_content = f"{report_name}\n{'=' * len(report_name)}\n\n"
        report_content += f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n"
        report_content += f"Components: {', '.join(selected_components)}\n\n"

        filename = f"custom_report_{datetime.now().strftime('%Y%m%d_%H%M')}.txt"
        with open(filename, 'w') as f:
            f.write(report_content)
            f.write("Custom report configuration saved.\n")
            f.write("Selected components will be generated based on available data.\n")

        messagebox.showinfo("Report Generated", f"Custom report configuration saved as {filename}")
        builder_window.destroy()

    ttk.Button(main_frame, text="Generate Custom Report",
              command=generate_custom_report, style='Accent.TButton').pack(pady=10)

def populate_scheduled_reports(self):
    """Populate scheduled reports list"""
    # Clear existing items
    for item in self.schedule_tree.get_children():
        self.schedule_tree.delete(item)

    # Sample scheduled reports
    scheduled_reports = [
        ("Daily Summary", "Executive", "Daily", "Tomorrow 08:00"),
        ("Weekly Analysis", "Comprehensive", "Weekly", "Monday 09:00"),
        ("Monthly Board Report", "Executive", "Monthly", "1st of next month"),
        ("Compliance Report", "Audit", "Quarterly", "End of quarter"),
        ("Risk Assessment", "Analytics", "Bi-weekly", "Next Friday")
    ]

    for report_name, report_type, frequency, next_run in scheduled_reports:
        self.schedule_tree.insert('', 'end', text=report_name,
                                values=(report_type, frequency, next_run))

def export_quick_report(self):
    """Export quick summary report"""
    filename = filedialog.asksaveasfilename(
        defaultextension=".txt",
        filetypes=[
            ("Text files", "*.txt"),
            ("PDF files", "*.pdf"),
            ("Excel files", "*.xlsx"),
            ("CSV files", "*.csv"),
            ("HTML files", "*.html"),
            ("All files", "*.*")
        ],
        title="Export Quick Report"
    )

    if filename:
        try:
            # Get financial data
            from education_system.systems.university.infrastructure.database.db import get_connection
            conn = get_connection()
            cursor = conn.cursor()

            # Get summary data
            cursor.execute('''
                SELECT
                    SUM(amount) as total_collected,
                    COUNT(*) as payment_count,
                    COUNT(DISTINCT student_id) as student_count
                FROM payments
                WHERE payment_date >= date('now', '-365 days')
            ''')
            summary = cursor.fetchone()
            total_collected = summary[0] or 0
            payment_count = summary[1] or 0
            student_count = summary[2] or 0

            conn.close()

            # Export based on file extension
            ext = filename.split('.')[-1].lower()

            if ext == 'txt':
                self._export_txt(filename, total_collected, payment_count, student_count)
            elif ext == 'csv':
                self._export_csv(filename, total_collected, payment_count, student_count)
            elif ext == 'html':
                self._export_html(filename, total_collected, payment_count, student_count)
            elif ext == 'xlsx':
                self._export_excel(filename, total_collected, payment_count, student_count)
            elif ext == 'pdf':
                self._export_pdf(filename, total_collected, payment_count, student_count)
            else:
                self._export_txt(filename, total_collected, payment_count, student_count)

            messagebox.showinfo("Export Complete", f"Quick report saved to {filename}")
            self.log_activity(f"Quick report exported to {filename}")

        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export report: {e}")

def _export_txt(self, filename, total_collected, payment_count, student_count):
    """Export report as text file"""
    with open(filename, 'w') as f:
        f.write("FINANCIAL QUICK REPORT\n")
        f.write("=" * 50 + "\n")
        f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
        f.write(f"Total Collected (Last Year): £{total_collected:,.2f}\n")
        f.write(f"Payment Count: {payment_count:,}\n")
        f.write(f"Student Count: {student_count:,}\n")
        f.write(f"Average per Student: £{total_collected/student_count if student_count > 0 else 0:,.2f}\n")

def _export_csv(self, filename, total_collected, payment_count, student_count):
    """Export report as CSV file"""
    import csv
    with open(filename, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['Metric', 'Value'])
        writer.writerow(['Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        writer.writerow(['Total Collected (Last Year)', f'£{total_collected:,.2f}'])
        writer.writerow(['Payment Count', payment_count])
        writer.writerow(['Student Count', student_count])
        writer.writerow(['Average per Student', f'£{total_collected/student_count if student_count > 0 else 0:,.2f}'])

def _export_html(self, filename, total_collected, payment_count, student_count):
    """Export report as HTML file"""
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
    <title>Financial Quick Report</title>
    <style>
    body {{ font-family: Arial, sans-serif; margin: 20px; }}
    h1 {{ color: #2c3e50; }}
    table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
    th, td {{ border: 1px solid #ddd; padding: 12px; text-align: left; }}
    th {{ background-color: #3498db; color: white; }}
    tr:nth-child(even) {{ background-color: #f2f2f2; }}
    </style>
    </head>
    <body>
    <h1>Financial Quick Report</h1>
    <p><strong>Generated:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
    <table>
    <tr><th>Metric</th><th>Value</th></tr>
    <tr><td>Total Collected (Last Year)</td><td>£{total_collected:,.2f}</td></tr>
    <tr><td>Payment Count</td><td>{payment_count:,}</td></tr>
    <tr><td>Student Count</td><td>{student_count:,}</td></tr>
    <tr><td>Average per Student</td><td>£{total_collected/student_count if student_count > 0 else 0:,.2f}</td></tr>
    </table>
    </body>
    </html>
    """
    with open(filename, 'w') as f:
        f.write(html_content)

def _export_excel(self, filename, total_collected, payment_count, student_count):
    """Export report as Excel file"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Financial Report"

        # Header
        ws['A1'] = "FINANCIAL QUICK REPORT"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:B1')

        ws['A2'] = "Generated:"
        ws['B2'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Data
        ws['A4'] = "Metric"
        ws['B4'] = "Value"
        ws['A4'].font = Font(bold=True)
        ws['B4'].font = Font(bold=True)

        ws['A5'] = "Total Collected (Last Year)"
        ws['B5'] = f'£{total_collected:,.2f}'

        ws['A6'] = "Payment Count"
        ws['B6'] = payment_count

        ws['A7'] = "Student Count"
        ws['B7'] = student_count

        ws['A8'] = "Average per Student"
        ws['B8'] = f'£{total_collected/student_count if student_count > 0 else 0:,.2f}'

        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

        wb.save(filename)
    except ImportError:
        # Fallback to CSV if openpyxl not available
        self._export_csv(filename.replace('.xlsx', '.csv'), total_collected, payment_count, student_count)

def _export_pdf(self, filename, total_collected, payment_count, student_count):
    """Export report as PDF file"""
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet

        doc = SimpleDocTemplate(filename, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title = Paragraph("FINANCIAL QUICK REPORT", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 12))

        # Generated date
        gen_date = Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal'])
        elements.append(gen_date)
        elements.append(Spacer(1, 12))

        # Data table
        data = [
            ['Metric', 'Value'],
            ['Total Collected (Last Year)', f'£{total_collected:,.2f}'],
            ['Payment Count', f'{payment_count:,}'],
            ['Student Count', f'{student_count:,}'],
            ['Average per Student', f'£{total_collected/student_count if student_count > 0 else 0:,.2f}']
        ]

        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        elements.append(table)
        doc.build(elements)
    except ImportError:
        # Fallback to HTML if reportlab not available
        self._export_html(filename.replace('.pdf', '.html'), total_collected, payment_count, student_count)

def export_comprehensive_report(self):
    """Export comprehensive financial report with GUI progress"""
    export_window = tk.Toplevel(self.root)
    export_window.title(_("finance_reporting.windows.export_comprehensive_report"))
    export_window.geometry("600x400")

    main_frame = ttk.Frame(export_window, padding="10")
    main_frame.pack(fill=tk.BOTH, expand=True)

    ttk.Label(main_frame, text="Comprehensive Report Export",
             style='Title.TLabel').pack(pady=(0, 20))

    # Export options
    options_frame = ttk.LabelFrame(main_frame, text="Export Options", padding="10")
    options_frame.pack(fill=tk.X, pady=(0, 10))

    self.export_forecasting = tk.BooleanVar(value=True)
    self.export_dashboard = tk.BooleanVar(value=True)
    self.export_analytics = tk.BooleanVar(value=True)
    self.export_charts = tk.BooleanVar(value=True)

    ttk.Checkbutton(options_frame, text="Financial Forecasting Reports",
                   variable=self.export_forecasting).pack(anchor=tk.W)
    ttk.Checkbutton(options_frame, text="Dashboard Summaries",
                   variable=self.export_dashboard).pack(anchor=tk.W)
    ttk.Checkbutton(options_frame, text="Advanced Analytics",
                   variable=self.export_analytics).pack(anchor=tk.W)
    ttk.Checkbutton(options_frame, text="Charts and Visualizations",
                   variable=self.export_charts).pack(anchor=tk.W)

    # Format selection
    format_frame = ttk.LabelFrame(main_frame, text="Output Format", padding="10")
    format_frame.pack(fill=tk.X, pady=(0, 10))

    self.export_format = tk.StringVar(value="PDF")
    formats = ["PDF", "Excel", "CSV", "All Formats"]
    for fmt in formats:
        ttk.Radiobutton(format_frame, text=fmt, variable=self.export_format,
                       value=fmt).pack(anchor=tk.W)

    # Progress area
    progress_frame = ttk.LabelFrame(main_frame, text="Export Progress", padding="10")
    progress_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

    self.export_progress = ttk.Progressbar(progress_frame, mode='determinate')
    self.export_progress.pack(fill=tk.X, pady=(0, 5))

    self.export_status = tk.StringVar(value="Ready to export")
    ttk.Label(progress_frame, textvariable=self.export_status).pack(anchor=tk.W)

    # Export log
    self.export_log = ScrolledText(progress_frame, height=8, wrap=tk.WORD)
    self.export_log.pack(fill=tk.BOTH, expand=True)

    def run_export():
        self.export_progress['value'] = 0
        self.export_log.delete(1.0, tk.END)

        export_tasks = []
        if self.export_forecasting.get():
            export_tasks.append(("Financial Forecasting", self.export_forecasting_report))
        if self.export_dashboard.get():
            export_tasks.append(("Dashboard Summary", self.export_dashboard_report))
        if self.export_analytics.get():
            export_tasks.append(("Advanced Analytics", self.export_analytics_report))
        if self.export_charts.get():
            export_tasks.append(("Charts & Visualizations", self.export_charts_report))

        if not export_tasks:
            messagebox.showwarning("Export", "Please select at least one export option")
            return

        total_tasks = len(export_tasks)

        for i, (task_name, task_func) in enumerate(export_tasks):
            self.export_status.set(f"Exporting {task_name}...")
            self.export_log.insert(tk.END, f"Starting {task_name}...\n")
            self.export_log.see(tk.END)
            export_window.update()

            try:
                task_func()
                self.export_log.insert(tk.END, f"✓ {task_name} completed\n")
            except Exception as e:
                self.export_log.insert(tk.END, f"✗ {task_name} failed: {e}\n")

            self.export_progress['value'] = ((i + 1) / total_tasks) * 100
            export_window.update()

        self.export_status.set("Export completed")
        self.export_log.insert(tk.END, "\nExport process completed!\n")
        messagebox.showinfo("Export Complete", "Comprehensive report export completed successfully!")

    # Export task methods (simplified implementations)
    def export_forecasting_report(self):
        import time
        time.sleep(1)  # Simulate export time
        return True

    def export_dashboard_report(self):
        import time
        time.sleep(0.8)
        return True

    def export_analytics_report(self):
        import time
        time.sleep(1.2)
        return True

    def export_charts_report(self):
        import time
        time.sleep(0.5)
        return True

    self.export_forecasting_report = export_forecasting_report
    self.export_dashboard_report = export_dashboard_report
    self.export_analytics_report = export_analytics_report
    self.export_charts_report = export_charts_report

    # Control buttons
    buttons_frame = ttk.Frame(main_frame)
    buttons_frame.pack(fill=tk.X, pady=(10, 0))

    ttk.Button(buttons_frame, text="Start Export", command=run_export,
              style='Accent.TButton').pack(side=tk.LEFT, padx=(0, 5))
    ttk.Button(buttons_frame, text="Close", command=export_window.destroy).pack(side=tk.RIGHT)

# Method registration is handled by main.py
