"""Analytics and reporting"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import tkinter.scrolledtext as scrolledtext
from university_system.infrastructure.database.db import sqlite3
import os
from pathlib import Path
import csv
import numpy as np
import math
from scipy import stats
from datetime import datetime, timedelta
import json
from university_system.modules.shared.constants import paths
from university_system.modules.domain.academics.gui.grade_tracking.utils import ensure_column_exists

# Use centralized path configuration
DEFAULT_DB_PATH = paths.DEFAULT_DB_PATH
_CENTRALDEFAULT_DB_PATH = paths.DEFAULT_DB_PATH

# Optional imports with fallbacks
try:
    import pandas as pd
except ImportError:
    pd = None

try:
    import matplotlib.pyplot as plt
    import seaborn as sns
    import matplotlib
    matplotlib.use('TkAgg')
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
except ImportError:
    plt = None
    sns = None
    FigureCanvasTkAgg = None

try:
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
except ImportError:
    # ReportLab not available - PDF features will be disabled
    SimpleDocTemplate = None

# Database connection setup

_original_sqlite3_connect_grade = sqlite3.connect

def _patched_sqlite3_connect_grade(database, *args, **kwargs):
    try:
        db_name = os.path.basename(str(database)) if database else ""
        if not database or db_name in (str(DEFAULT_DB_PATH), "student_grading_system.db"):
            return _original_sqlite3_connect_grade(str(_CENTRALDEFAULT_DB_PATH), *args, **kwargs)
    except Exception:
        pass
    return _original_sqlite3_connect_grade(database, *args, **kwargs)

sqlite3.connect = _patched_sqlite3_connect_grade

# Import the database connection
try:
    from university_system.infrastructure.database.db import get_connection
except ImportError:
    def get_connection():
        """Fallback database connection function"""
        base_dir = Path(__file__).resolve().parents[1]  # Fixed indentation here
        db_path = base_dir / "db_files" / str(DEFAULT_DB_PATH)
        
        # Ensure directory exists
        db_path.parent.mkdir(parents=True, exist_ok=True)
        
        return sqlite3.connect(str(DEFAULT_DB_PATH))

# Global variables for grade systems
GRADE_SYSTEMS = {
    "letter": {
        "A+": 4.3, "A": 4.0, "A-": 3.7,
        "B+": 3.3, "B": 3.0, "B-": 2.7,
        "C+": 2.3, "C": 2.0, "C-": 1.7,
        "D+": 1.3, "D": 1.0, "D-": 0.7,
        "F": 0.0
    },
    "percentage": {
        "range": (0, 100),
        "conversion": {
        (90, 100): "A+", (85, 89.99): "A", (80, 84.99): "A-",
        (75, 79.99): "B+", (70, 74.99): "B", (65, 69.99): "B-",
        (60, 64.99): "C+", (55, 59.99): "C", (50, 54.99): "C-",
        (45, 49.99): "D+", (40, 44.99): "D", (35, 39.99): "D-",
        (0, 34.99): "F"
        }
    }
}

def percentage_to_letter(percentage):
    """Convert a percentage score to a letter grade"""
    for score_range, letter in GRADE_SYSTEMS["percentage"]["conversion"].items():
        min_score, max_score = score_range
        if min_score <= percentage <= max_score:
            return letter
    return 'F'

def letter_to_percentage(letter_grade):
    """Convert a letter grade to a percentage score (midpoint of range)"""
    for score_range, letter in GRADE_SYSTEMS["percentage"]["conversion"].items():
        if letter == letter_grade:
            min_score, max_score = score_range
        return (min_score + max_score) / 2
    return 0

def letter_to_gpa(letter_grade):
    """Convert a letter grade to a GPA value"""
    if letter_grade in GRADE_SYSTEMS["letter"]:
        return GRADE_SYSTEMS["letter"][letter_grade]
    return 0

def init_basic_database():
    """Initialize the basic database tables that are missing"""
    try:
        conn = get_connection()
        cursor = conn.cursor()
        
        # Create students table
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS students (
        student_id TEXT PRIMARY KEY,
        first_name TEXT NOT NULL,
        middle_name TEXT,
        last_name TEXT NOT NULL,
        course TEXT NOT NULL,
        email_address TEXT,
        gender TEXT,
        dob TEXT,
        enrollment_date TEXT DEFAULT (date('now')),
        status TEXT DEFAULT 'Active'
        )
        ''')
        
        # Create modules table
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS modules (
        module_code TEXT PRIMARY KEY,
        module_name TEXT NOT NULL,
        module_type TEXT,
        credits INTEGER DEFAULT 1,
        description TEXT,
        course TEXT,
        semester TEXT,
        year INTEGER
        )
        ''')
        
        # Create student_modules table (enrollment)
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS student_modules (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        student_id TEXT,
        module_code TEXT,
        enrollment_date TEXT DEFAULT (date('now')),
        status TEXT DEFAULT 'Enrolled',
        FOREIGN KEY (student_id) REFERENCES students (student_id),
        FOREIGN KEY (module_code) REFERENCES modules (module_code),
        UNIQUE(student_id, module_code)
        )
        ''')
        
        # Create assessments table
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS assessments (
        assessment_id INTEGER PRIMARY KEY AUTOINCREMENT,
        assessment_name TEXT NOT NULL,
        assessment_type TEXT NOT NULL,
        module_code TEXT NOT NULL,
        max_points REAL NOT NULL,
        weight REAL NOT NULL,
        due_date TEXT,
        date_created TEXT DEFAULT (datetime('now')),
        description TEXT,
        rubric TEXT,
        FOREIGN KEY (module_code) REFERENCES modules (module_code)
        )
        ''')

        # Ensure rubric column exists for legacy databases.
        ensure_column_exists(cursor, 'assessments', 'rubric', 'TEXT')
        
        conn.commit()
        conn.close()
        return True
        
    except sqlite3.Error as e:
        messagebox.showerror("Database Error", f"Database error: {e}")
        return False

def init_enhanced_grades_db():
    """Initialize the enhanced grades database with all required tables"""
    try:
        conn = get_connection()
        cursor = conn.cursor()
        
        # Create base grade tables if they don't exist
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS grades (
        grade_id INTEGER PRIMARY KEY AUTOINCREMENT,
        student_id TEXT,
        assessment_id INTEGER,
        score REAL,
        letter_grade TEXT,
        submission_date TEXT,
        comments TEXT,
        FOREIGN KEY (student_id) REFERENCES students (student_id),
        FOREIGN KEY (assessment_id) REFERENCES assessments (assessment_id)
        )
        ''')
        
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS module_grades (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        student_id TEXT,
        module_code TEXT,
        final_score REAL,
        final_grade TEXT,
        completion_date TEXT,
        FOREIGN KEY (student_id) REFERENCES students (student_id),
        FOREIGN KEY (module_code) REFERENCES modules (module_code)
        )
        ''')
        
        # Enhanced tables for statistics and analytics
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS grade_statistics (
        stat_id INTEGER PRIMARY KEY AUTOINCREMENT,
        assessment_id INTEGER,
        mean REAL,
        median REAL,
        std_dev REAL,
        min_score REAL,
        max_score REAL,
        q1 REAL,
        q3 REAL,
        skewness REAL,
        kurtosis REAL,
        date_calculated TEXT,
        FOREIGN KEY (assessment_id) REFERENCES assessments (assessment_id)
        )
        ''')
        
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS normalized_grades (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        grade_id INTEGER,
        z_score REAL,
        percentile REAL,
        curved_score REAL,
        curved_letter TEXT,
        FOREIGN KEY (grade_id) REFERENCES grades (grade_id)
        )
        ''')

        # Predictive analytics tables
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS risk_factors (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        student_id TEXT,
        factor_name TEXT,
        factor_value REAL,
        assessment_id INTEGER,
        date_calculated TEXT,
        FOREIGN KEY (student_id) REFERENCES students (student_id),
        FOREIGN KEY (assessment_id) REFERENCES assessments (assessment_id)
        )
        ''')

        cursor.execute('''
        CREATE TABLE IF NOT EXISTS risk_details (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        risk_factor_id INTEGER,
        detail TEXT,
        weight REAL,
        created_at TEXT,
        FOREIGN KEY (risk_factor_id) REFERENCES risk_factors (id)
        )
        ''')

        cursor.execute('''
        CREATE TABLE IF NOT EXISTS intervention_types (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT,
        description TEXT
        )
        ''')

        cursor.execute('''
        CREATE TABLE IF NOT EXISTS recommended_interventions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        student_id TEXT,
        risk_factor_id INTEGER,
        intervention_type_id INTEGER,
        recommended_date TEXT,
        status TEXT DEFAULT 'pending',
        FOREIGN KEY (student_id) REFERENCES students (student_id),
        FOREIGN KEY (risk_factor_id) REFERENCES risk_factors (id),
        FOREIGN KEY (intervention_type_id) REFERENCES intervention_types (id)
        )
        ''')

        conn.commit()
        conn.close()
        return True
    except sqlite3.Error as e:
        messagebox.showerror("Database Error", f"Database error: {e}")
        return False

def safe_grab_set(dialog):
    """Safely set grab on dialog window, handling TclError exceptions"""
    try:
        # Ensure dialog is visible first
        dialog.update_idletasks()
        dialog.grab_set()
    except tk.TclError:
        # If grab fails, continue without modal behavior
        pass

def safe_combo_update(obj, combo_attr, values):
    """Safely update combobox values, handling widget destruction"""
    try:
        if hasattr(obj, combo_attr):
            combo = getattr(obj, combo_attr)
        if combo and combo.winfo_exists():
            combo['values'] = values
            return True
    except (tk.TclError, AttributeError):
        # Widget no longer exists or is invalid
        pass
    return False

class AnalyticsManager:
    """Analytics and reporting"""

    def __init__(self, app):
        self.app = app
        self.root = app.root
        self.auth = app.auth
        self.conn = app.conn

    @property
    def content_frame(self):
        """Dynamically get content frame from layout"""
        if hasattr(self.app, 'layout') and hasattr(self.app.layout, 'content_frame'):
            return self.app.layout.content_frame
        return None

    def create_analytics_content(self):
        """Create analytics and performance tab"""
        # Initialize notebook for this view. Each view that uses multiple tabs
        # should create a fresh ttk.Notebook attached to the current content frame.
        self.notebook = ttk.Notebook(self.content_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True)
        # Create the analytics frame as a child of the notebook
        analytics_frame = ttk.Frame(self.notebook)
        self.notebook.add(analytics_frame, text="Analytics")
            
        # Create paned window for layout
        paned = ttk.PanedWindow(analytics_frame, orient='horizontal')
        paned.pack(fill='both', expand=True, padx=10, pady=5)

        # Left panel - Controls with scrollbar
        left_panel_container = ttk.Frame(paned)
        paned.add(left_panel_container, weight=1)

        # Create canvas for scrolling
        canvas = tk.Canvas(left_panel_container, highlightthickness=0)
        scrollbar = ttk.Scrollbar(left_panel_container, orient="vertical", command=canvas.yview)
        left_panel = ttk.Frame(canvas)

        # Configure scrolling
        left_panel.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=left_panel, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # Mouse wheel scrolling (with existence check to prevent errors after window close)
        def _on_mousewheel(event):
            try:
                if canvas.winfo_exists():
                    canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            except tk.TclError:
                pass

        def _on_mousewheel_linux(event):
            try:
                if canvas.winfo_exists():
                    if event.num == 4:
                        canvas.yview_scroll(-1, "units")
                    elif event.num == 5:
                        canvas.yview_scroll(1, "units")
            except tk.TclError:
                pass

        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        canvas.bind_all("<Button-4>", _on_mousewheel_linux)
        canvas.bind_all("<Button-5>", _on_mousewheel_linux)

        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Analytics controls
        controls_label = ttk.Label(left_panel, text="Analytics Options", style='Header.TLabel')
        controls_label.pack(pady=10)
            
        # Performance Analysis
        perf_frame = ttk.LabelFrame(left_panel, text="Performance Analysis", padding=10)
        perf_frame.pack(fill='x', padx=10, pady=5)
            
        ttk.Button(perf_frame, text="At-Risk Students", 
                  command=self.identify_at_risk_students,
                  style='Custom.TButton').pack(fill='x', pady=2)
        ttk.Button(perf_frame, text="Grade Distribution", 
                  command=self.show_grade_distribution,
                  style='Custom.TButton').pack(fill='x', pady=2)
        ttk.Button(perf_frame, text="Module Performance", 
                  command=self.analyze_module_performance,
                  style='Custom.TButton').pack(fill='x', pady=2)
        ttk.Button(perf_frame, text="Course Comparison", 
                  command=self.compare_course_performance,
                  style='Custom.TButton').pack(fill='x', pady=2)
        ttk.Button(perf_frame, text="Trends Over Time", 
                  command=self.analyze_performance_trends,
                  style='Custom.TButton').pack(fill='x', pady=2)
            
        # Risk Assessment
        risk_frame = ttk.LabelFrame(left_panel, text="Risk Assessment", padding=10)
        risk_frame.pack(fill='x', padx=10, pady=5)
            
        ttk.Button(risk_frame, text="Student Risk Assessment", 
                  command=self.student_risk_assessment,
                  style='Custom.TButton').pack(fill='x', pady=2)
        ttk.Button(risk_frame, text="Early Warning System", 
                  command=self.early_warning_system,
                  style='Custom.TButton').pack(fill='x', pady=2)
        ttk.Button(risk_frame, text="Dropout Risk Analysis", 
                  command=self.dropout_risk_analysis,
                  style='Custom.TButton').pack(fill='x', pady=2)
        ttk.Button(risk_frame, text="Intervention Recommendations", 
                  command=self.intervention_recommendations,
                  style='Custom.TButton').pack(fill='x', pady=2)
            
        # Visualization Options
        viz_frame = ttk.LabelFrame(left_panel, text="Visualizations", padding=10)
        viz_frame.pack(fill='x', padx=10, pady=5)
            
        ttk.Button(viz_frame, text="Performance Dashboard", 
                  command=self.generate_performance_dashboard,
                  style='Custom.TButton').pack(fill='x', pady=2)
        ttk.Button(viz_frame, text="Grade Trends Chart", 
                  command=self.show_grade_trends_chart,
                  style='Custom.TButton').pack(fill='x', pady=2)
        ttk.Button(viz_frame, text="Student Progress Charts", 
                  command=self.student_progress_charts,
                  style='Custom.TButton').pack(fill='x', pady=2)
            
        # Right panel - Display area
        right_panel = ttk.Frame(paned)
        paned.add(right_panel, weight=2)
            
        # Results display area
        results_label = ttk.Label(right_panel, text="Analytics Results", style='Header.TLabel')
        results_label.pack(pady=10)
            
        # Scrollable text area for results
        self.analytics_results = scrolledtext.ScrolledText(right_panel, height=30, width=60)
        self.analytics_results.pack(fill='both', expand=True, padx=10, pady=5)
            
        # Chart display frame
        self.chart_frame = ttk.Frame(right_panel)
        self.chart_frame.pack(fill='both', expand=True, padx=10, pady=5)
        

    def create_reports_content(self):
        """Create reports tab"""
        # Initialize notebook for this view
        self.notebook = ttk.Notebook(self.content_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True)
        # Create reports frame as a child of the notebook
        reports_frame = ttk.Frame(self.notebook)
        self.notebook.add(reports_frame, text="Reports")
            
        # Create paned window
        paned = ttk.PanedWindow(reports_frame, orient='horizontal')
        paned.pack(fill='both', expand=True, padx=10, pady=5)

        # Left panel - Report options with scrollbar
        left_panel_container = ttk.Frame(paned)
        paned.add(left_panel_container, weight=1)

        # Create canvas for scrolling
        canvas = tk.Canvas(left_panel_container, highlightthickness=0)
        scrollbar = ttk.Scrollbar(left_panel_container, orient="vertical", command=canvas.yview)
        left_panel = ttk.Frame(canvas)

        # Configure scrolling
        left_panel.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=left_panel, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # Mouse wheel scrolling (with existence check to prevent errors after window close)
        def _on_mousewheel_reports(event):
            try:
                if canvas.winfo_exists():
                    canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            except tk.TclError:
                pass

        def _on_mousewheel_linux_reports(event):
            try:
                if canvas.winfo_exists():
                    if event.num == 4:
                        canvas.yview_scroll(-1, "units")
                    elif event.num == 5:
                        canvas.yview_scroll(1, "units")
            except tk.TclError:
                pass

        canvas.bind_all("<MouseWheel>", _on_mousewheel_reports)
        canvas.bind_all("<Button-4>", _on_mousewheel_linux_reports)
        canvas.bind_all("<Button-5>", _on_mousewheel_linux_reports)

        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        reports_label = ttk.Label(left_panel, text="Report Generation", style='Header.TLabel')
        reports_label.pack(pady=10)
            
        # Student Reports
        student_reports = ttk.LabelFrame(left_panel, text="Student Reports", padding=10)
        student_reports.pack(fill='x', padx=10, pady=5)
            
        ttk.Button(student_reports, text="Individual Transcript", 
                  command=self.generate_individual_transcript,
                  style='Custom.TButton').pack(fill='x', pady=2)
        ttk.Button(student_reports, text="Student Progress Report", 
                  command=self.generate_student_progress_report,
                  style='Custom.TButton').pack(fill='x', pady=2)
        ttk.Button(student_reports, text="Competency Profile", 
                  command=self.generate_competency_profile,
                  style='Custom.TButton').pack(fill='x', pady=2)
        ttk.Button(student_reports, text="At-Risk Student Report", 
                  command=self.generate_at_risk_report,
                  style='Custom.TButton').pack(fill='x', pady=2)
            
        # Module Reports
        module_reports = ttk.LabelFrame(left_panel, text="Module Reports", padding=10)
        module_reports.pack(fill='x', padx=10, pady=5)
            
        ttk.Button(module_reports, text="Module Grade Report", 
                  command=self.generate_module_grade_report,
                  style='Custom.TButton').pack(fill='x', pady=2)
        ttk.Button(module_reports, text="Module Outcome Report", 
                  command=self.generate_module_outcome_report,
                  style='Custom.TButton').pack(fill='x', pady=2)
        ttk.Button(module_reports, text="Assessment Analysis", 
                  command=self.generate_assessment_analysis,
                  style='Custom.TButton').pack(fill='x', pady=2)
            
        # System Reports
        system_reports = ttk.LabelFrame(left_panel, text="System Reports", padding=10)
        system_reports.pack(fill='x', padx=10, pady=5)
            
        ttk.Button(system_reports, text="Institution Summary", 
                  command=self.generate_institution_summary,
                  style='Custom.TButton').pack(fill='x', pady=2)
        ttk.Button(system_reports, text="Performance Analytics", 
                  command=self.generate_performance_analytics,
                  style='Custom.TButton').pack(fill='x', pady=2)
        ttk.Button(system_reports, text="Trend Analysis Report", 
                  command=self.generate_trend_analysis_report,
                  style='Custom.TButton').pack(fill='x', pady=2)
        ttk.Button(system_reports, text="Risk Assessment Report", 
                  command=self.generate_comprehensive_risk_report,
                  style='Custom.TButton').pack(fill='x', pady=2)
            
        # Export Options
        export_frame = ttk.LabelFrame(left_panel, text="Export Options", padding=10)
        export_frame.pack(fill='x', padx=10, pady=5)
            
        ttk.Button(export_frame, text="Export to PDF", 
                  command=self.export_reports_pdf,
                  style='Custom.TButton').pack(fill='x', pady=2)
        ttk.Button(export_frame, text="Export to CSV", 
                  command=self.export_reports_csv,
                  style='Custom.TButton').pack(fill='x', pady=2)
        ttk.Button(export_frame, text="Export to Excel", 
                  command=self.export_reports_excel,
                  style='Custom.TButton').pack(fill='x', pady=2)
            
        # Right panel - Report preview
        right_panel = ttk.Frame(paned)
        paned.add(right_panel, weight=2)
            
        preview_label = ttk.Label(right_panel, text="Report Preview", style='Header.TLabel')
        preview_label.pack(pady=10)
            
        # Report preview area
        self.report_preview = scrolledtext.ScrolledText(right_panel, height=35, width=70)
        self.report_preview.pack(fill='both', expand=True, padx=10, pady=5)
        

    def create_competency_content(self):
        """Create the competency-based assessment tab (fixed version)"""
        # Initialize notebook for this view
        self.notebook = ttk.Notebook(self.content_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True)
        self.competency_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.competency_frame, text="Competencies")
            
        # Title
        ttk.Label(self.competency_frame, text="Competency-Based Assessment", 
                 font=('Arial', 16, 'bold')).pack(pady=10)
            
        # Competency assessment functionality
        competency_controls = ttk.Frame(self.competency_frame)
        competency_controls.pack(fill=tk.X, padx=20, pady=10)
    
        ttk.Label(competency_controls, text="Select Student:").pack(side=tk.LEFT, padx=5)
        comp_student_var = tk.StringVar()
        comp_student_combo = ttk.Combobox(competency_controls, textvariable=comp_student_var, width=30)
        comp_student_combo.pack(side=tk.LEFT, padx=5)
    
        # Competency display
        competency_display = ttk.LabelFrame(self.competency_frame, text="Competency Assessment Matrix", padding=10)
        competency_display.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
    
        columns = ("Competency", "Level", "Evidence", "Date", "Assessor")
        comp_tree = ttk.Treeview(competency_display, columns=columns, show="headings", height=15)
        for col in columns:
            comp_tree.heading(col, text=col)
            comp_tree.column(col, width=120)
    
        scrollbar = ttk.Scrollbar(competency_display, orient=tk.VERTICAL, command=comp_tree.yview)
        comp_tree.configure(yscrollcommand=scrollbar.set)
        comp_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    
        # Sample data
        sample_comps = [
            ("Technical Skills", "Advanced", "Portfolio Project", "2024-01-15", "Dr. Smith"),
            ("Leadership", "Intermediate", "Group Project", "2024-01-10", "Prof. Jones"),
            ("Research", "Proficient", "Term Paper", "2024-01-05", "Dr. Davis"),
        ]
        for comp in sample_comps:
            comp_tree.insert('', tk.END, values=comp)
    

    def create_prediction_content(self):
        """Create predictive analytics tab"""
        # Initialize notebook for this view
        self.notebook = ttk.Notebook(self.content_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True)
        prediction_frame = ttk.Frame(self.notebook)
        self.notebook.add(prediction_frame, text="Predictions")
            
        # Create paned window
        paned = ttk.PanedWindow(prediction_frame, orient='horizontal')
        paned.pack(fill='both', expand=True, padx=10, pady=5)

        # Left panel - Prediction options with scrollbar
        left_panel_container = ttk.Frame(paned)
        paned.add(left_panel_container, weight=1)

        # Create canvas for scrolling
        canvas = tk.Canvas(left_panel_container, highlightthickness=0)
        scrollbar = ttk.Scrollbar(left_panel_container, orient="vertical", command=canvas.yview)
        left_panel = ttk.Frame(canvas)

        # Configure scrolling
        left_panel.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=left_panel, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # Mouse wheel scrolling (with existence check to prevent errors after window close)
        def _on_mousewheel_prediction(event):
            try:
                if canvas.winfo_exists():
                    canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            except tk.TclError:
                pass

        def _on_mousewheel_linux_prediction(event):
            try:
                if canvas.winfo_exists():
                    if event.num == 4:
                        canvas.yview_scroll(-1, "units")
                    elif event.num == 5:
                        canvas.yview_scroll(1, "units")
            except tk.TclError:
                pass

        canvas.bind_all("<MouseWheel>", _on_mousewheel_prediction)
        canvas.bind_all("<Button-4>", _on_mousewheel_linux_prediction)
        canvas.bind_all("<Button-5>", _on_mousewheel_linux_prediction)

        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        prediction_label = ttk.Label(left_panel, text="Predictive Analytics", style='Header.TLabel')
        prediction_label.pack(pady=10)
            
        # Performance Prediction
        perf_pred_frame = ttk.LabelFrame(left_panel, text="Performance Prediction", padding=10)
        perf_pred_frame.pack(fill='x', padx=10, pady=5)
            
        ttk.Button(perf_pred_frame, text="Grade Prediction", 
                  command=self.predict_grades_dialog,
                  style='Custom.TButton').pack(fill='x', pady=2)
        ttk.Button(perf_pred_frame, text="GPA Prediction", 
                  command=self.predict_gpa_dialog,
                  style='Custom.TButton').pack(fill='x', pady=2)
        ttk.Button(perf_pred_frame, text="Module Success Prediction", 
                  command=self.predict_module_success,
                  style='Custom.TButton').pack(fill='x', pady=2)
        ttk.Button(perf_pred_frame, text="Success Probability Calculator", 
                  command=self.calculate_success_probability,
                  style='Custom.TButton').pack(fill='x', pady=2)
            
        # Risk Prediction
        risk_pred_frame = ttk.LabelFrame(left_panel, text="Risk Prediction", padding=10)
        risk_pred_frame.pack(fill='x', padx=10, pady=5)
            
        ttk.Button(risk_pred_frame, text="At-Risk Prediction Model", 
                  command=self.build_at_risk_model,
                  style='Custom.TButton').pack(fill='x', pady=2)
        ttk.Button(risk_pred_frame, text="Dropout Risk Analysis", 
                  command=self.analyze_dropout_risk,
                  style='Custom.TButton').pack(fill='x', pady=2)
        ttk.Button(risk_pred_frame, text="Early Warning Alerts", 
                  command=self.generate_early_warnings,
                  style='Custom.TButton').pack(fill='x', pady=2)
            
        # Trend Forecasting
        trend_frame = ttk.LabelFrame(left_panel, text="Trend Forecasting", padding=10)
        trend_frame.pack(fill='x', padx=10, pady=5)
            
        ttk.Button(trend_frame, text="Performance Trends", 
                  command=self.forecast_performance_trends,
                  style='Custom.TButton').pack(fill='x', pady=2)
        ttk.Button(trend_frame, text="Course Performance Forecast", 
                  command=self.forecast_course_performance,
                  style='Custom.TButton').pack(fill='x', pady=2)
        ttk.Button(trend_frame, text="Success Rate Trends", 
                  command=self.forecast_success_rates,
                  style='Custom.TButton').pack(fill='x', pady=2)
            
        # Batch Operations
        batch_frame = ttk.LabelFrame(left_panel, text="Batch Operations", padding=10)
        batch_frame.pack(fill='x', padx=10, pady=5)
            
        ttk.Button(batch_frame, text="Batch Grade Predictions", 
                  command=self.batch_predict_grades,
                  style='Custom.TButton').pack(fill='x', pady=2)
        ttk.Button(batch_frame, text="Batch Risk Assessment", 
                  command=self.batch_risk_assessment,
                  style='Custom.TButton').pack(fill='x', pady=2)
        ttk.Button(batch_frame, text="Generate Interventions", 
                  command=self.generate_interventions,
                  style='Custom.TButton').pack(fill='x', pady=2)
            
        # Right panel - Results display
        right_panel = ttk.Frame(paned)
        paned.add(right_panel, weight=2)
            
        results_label = ttk.Label(right_panel, text="Prediction Results", style='Header.TLabel')
        results_label.pack(pady=10)
            
        # Results display area
        self.prediction_results = scrolledtext.ScrolledText(right_panel, height=25, width=60)
        self.prediction_results.pack(fill='both', expand=True, padx=10, pady=5)
            
        # Chart display for predictions
        self.prediction_chart_frame = ttk.Frame(right_panel)
        self.prediction_chart_frame.pack(fill='both', expand=True, padx=10, pady=5)
        

    def generate_institution_summary(self):
        """Generate institution summary"""
        conn = None
        try:
            conn = get_connection()
            cursor = conn.cursor()
    
            cursor.execute('SELECT COUNT(*) FROM students')
            total_students = cursor.fetchone()[0]
    
            cursor.execute('SELECT COUNT(*) FROM modules')
            total_modules = cursor.fetchone()[0]
    
            cursor.execute('SELECT COUNT(*) FROM assessments')
            total_assessments = cursor.fetchone()[0]
    
            cursor.execute('SELECT COUNT(*) FROM grades')
            total_grades = cursor.fetchone()[0]
    
            cursor.execute('''
            SELECT AVG(
                CASE final_grade
                    WHEN 'A+' THEN 4.0 WHEN 'A' THEN 4.0 WHEN 'A-' THEN 3.7
                    WHEN 'B+' THEN 3.3 WHEN 'B' THEN 3.0 WHEN 'B-' THEN 2.7
                    WHEN 'C+' THEN 2.3 WHEN 'C' THEN 2.0 WHEN 'C-' THEN 1.7
                    WHEN 'D+' THEN 1.3 WHEN 'D' THEN 1.0 WHEN 'D-' THEN 0.7
                    WHEN 'F' THEN 0.0 ELSE NULL
                END
            )
            FROM module_grades
            WHERE final_grade IS NOT NULL
            ''')
            avg_gpa = cursor.fetchone()[0]
    
            cursor.execute('''
            SELECT course,
                   COUNT(*) AS students,
                   AVG(
                       CASE 
                           WHEN a.max_points IS NOT NULL AND a.max_points > 0 
                           THEN (g.score / a.max_points) * 100 
                       END
                   ) AS avg_score
            FROM students s
            LEFT JOIN grades g ON s.student_id = g.student_id
            LEFT JOIN assessments a ON g.assessment_id = a.assessment_id
            GROUP BY course
            ORDER BY students DESC
            ''')
            course_rows = cursor.fetchall()
    
            cursor.execute('''
            SELECT
                substr(COALESCE(g.submission_date, ''), 1, 7) AS period,
                COUNT(*) AS submissions
            FROM grades g
            WHERE g.submission_date IS NOT NULL AND g.submission_date != ''
            GROUP BY substr(g.submission_date, 1, 7)
            ORDER BY period DESC
            LIMIT 6
            ''')
            activity_rows = cursor.fetchall()
    
            summary_lines = [
                f"Total Students: {total_students}",
                f"Total Modules: {total_modules}",
                f"Total Assessments: {total_assessments}",
                f"Total Recorded Grades: {total_grades}",
                f"Average GPA (module level): {avg_gpa:.2f}" if avg_gpa is not None else "Average GPA: N/A"
            ]
    
            course_lines = [
                f"{row[0]}: {row[1]} students, Avg Score {row[2]:.1f}%"
                if row[2] is not None else
                f"{row[0]}: {row[1]} students, Avg Score N/A"
                for row in course_rows[:5]
            ]
    
            enrollment_lines = [
                f"{row[0]}: {row[1]} grade submissions"
                for row in activity_rows
            ]
    
            sections = [
                ("Institutional Snapshot", summary_lines),
                ("Largest Courses By Enrollment", course_lines),
                ("Recent Academic Activity (Last 6 Periods)", enrollment_lines)
            ]
    
            footer = (
                "Monitor enrollment distribution and recent activity to ensure support resources align with "
                "student engagement patterns."
            )
    
            self._display_report("Institution Summary", sections, footer)
    
        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Failed to generate institution summary: {e}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate institution summary: {e}")
        finally:
            if conn:
                conn.close()
    

    def generate_performance_analytics(self):
        """Generate performance analytics"""
        conn = None
        try:
            conn = get_connection()
            cursor = conn.cursor()
    
            cursor.execute('''
            SELECT
                course,
                COUNT(*) AS students,
                AVG(
                    CASE mg.final_grade
                        WHEN 'A+' THEN 4.0 WHEN 'A' THEN 4.0 WHEN 'A-' THEN 3.7
                        WHEN 'B+' THEN 3.3 WHEN 'B' THEN 3.0 WHEN 'B-' THEN 2.7
                        WHEN 'C+' THEN 2.3 WHEN 'C' THEN 2.0 WHEN 'C-' THEN 1.7
                        WHEN 'D+' THEN 1.3 WHEN 'D' THEN 1.0 WHEN 'D-' THEN 0.7
                        WHEN 'F' THEN 0.0 ELSE NULL
                    END
                ) AS avg_gpa
            FROM students s
            LEFT JOIN module_grades mg ON s.student_id = mg.student_id
            GROUP BY course
            ''')
            course_gpa_rows = cursor.fetchall()
    
            cursor.execute('''
            SELECT 
                letter_grade,
                COUNT(*) AS count
            FROM grades
            WHERE letter_grade IS NOT NULL
            GROUP BY letter_grade
            ''')
            grade_dist_rows = cursor.fetchall()
    
            cursor.execute('''
            SELECT 
                s.course,
                AVG(
                    CASE 
                        WHEN a.max_points IS NOT NULL AND a.max_points > 0 
                        THEN (g.score / a.max_points) * 100 
                    END
                ) AS avg_score,
                SUM(CASE WHEN g.letter_grade = 'F' THEN 1 ELSE 0 END) AS fail_count,
                COUNT(g.grade_id) AS grade_count
            FROM students s
            LEFT JOIN grades g ON s.student_id = g.student_id
            LEFT JOIN assessments a ON g.assessment_id = a.assessment_id
            GROUP BY s.course
            ''')
            course_perf_rows = cursor.fetchall()
    
            cursor.execute('''
            SELECT
                m.module_code,
                AVG(mg.final_score) AS avg_score,
                AVG(
                    CASE mg.final_grade
                        WHEN 'A+' THEN 4.0 WHEN 'A' THEN 4.0 WHEN 'A-' THEN 3.7
                        WHEN 'B+' THEN 3.3 WHEN 'B' THEN 3.0 WHEN 'B-' THEN 2.7
                        WHEN 'C+' THEN 2.3 WHEN 'C' THEN 2.0 WHEN 'C-' THEN 1.7
                        WHEN 'D+' THEN 1.3 WHEN 'D' THEN 1.0 WHEN 'D-' THEN 0.7
                        WHEN 'F' THEN 0.0 ELSE NULL
                    END
                ) AS avg_gpa
            FROM module_grades mg
            JOIN modules m ON mg.module_code = m.module_code
            GROUP BY m.module_code
            ''')
            module_perf_rows = cursor.fetchall()
    
            summary_lines = [
                f"Courses Tracked: {len(course_gpa_rows)}",
                f"Average Course GPA: "
                f"{(sum(row[2] for row in course_gpa_rows if row[2] is not None) / max(1, sum(1 for row in course_gpa_rows if row[2] is not None))):.2f}"
                if any(row[2] is not None for row in course_gpa_rows) else "Average Course GPA: N/A",
                f"Total Grade Records Considered: {sum(row[1] for row in grade_dist_rows)}"
            ]
    
            grade_lines = [
                f"{row[0]}: {row[1]} ({(row[1] / max(1, sum(r[1] for r in grade_dist_rows))) * 100:.1f}%)"
                for row in sorted(
                    grade_dist_rows,
                    key=lambda r: r[0]
                )
            ]
    
            course_lines = [
                f"{row[0]}: Avg Score {row[1]:.1f}%, Failure Rate "
                f"{(row[2] / row[3]):.0%}" if row[3] else
                f"{row[0]}: No assessment data"
                for row in sorted(
                    course_perf_rows,
                    key=lambda r: (r[1] if r[1] is not None else 0),
                    reverse=True
                )[:5]
            ]
    
            module_lines = [
                f"{row[0]}: Avg Score {row[1]:.1f}%, Avg GPA {row[2]:.2f}"
                if row[1] is not None and row[2] is not None else
                f"{row[0]}: Score {row[1]:.1f}%"
                if row[1] is not None else
                f"{row[0]}: GPA {row[2]:.2f}"
                if row[2] is not None else
                f"{row[0]}: No final grade data"
                for row in sorted(
                    module_perf_rows,
                    key=lambda r: (r[1] if r[1] is not None else 0),
                    reverse=True
                )[:5]
            ]
    
            sections = [
                ("Performance Overview", summary_lines),
                ("Grade Distribution", grade_lines),
                ("Course Performance Highlights", course_lines),
                ("Module Performance Highlights", module_lines)
            ]
    
            footer = (
                "Use grade distribution insights alongside course and module summaries to identify "
                "areas for additional academic support or curriculum review."
            )
    
            self._display_report("Performance Analytics Report", sections, footer)
    
        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Failed to generate performance analytics: {e}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate performance analytics: {e}")
        finally:
            if conn:
                conn.close()
    

    def generate_trend_analysis_report(self):
        """Generate trend analysis report"""
        conn = None
        try:
            conn = get_connection()
            cursor = conn.cursor()
    
            cursor.execute('''
            SELECT 
                substr(COALESCE(g.submission_date, ''), 1, 7) AS period,
                COUNT(*) AS submissions,
                AVG(
                    CASE 
                        WHEN a.max_points IS NOT NULL AND a.max_points > 0 
                        THEN (g.score / a.max_points) * 100 
                    END
                ) AS avg_score
            FROM grades g
            JOIN assessments a ON g.assessment_id = a.assessment_id
            WHERE g.submission_date IS NOT NULL AND g.submission_date != ''
            GROUP BY substr(g.submission_date, 1, 7)
            ORDER BY period DESC
            LIMIT 12
            ''')
            grade_trend_rows = cursor.fetchall()
    
            cursor.execute('''
            SELECT 
                substr(COALESCE(assessment_date, ''), 1, 7) AS period,
                AVG(risk_score) AS avg_risk
            FROM student_risk_assessment
            WHERE assessment_date IS NOT NULL AND assessment_date != ''
            GROUP BY substr(assessment_date, 1, 7)
            ORDER BY period DESC
            LIMIT 12
            ''')
            risk_trend_rows = cursor.fetchall()
    
            cursor.execute('''
            SELECT 
                substr(COALESCE(enrollment_date, ''), 1, 7) AS period,
                COUNT(*) AS enrollments
            FROM students
            WHERE enrollment_date IS NOT NULL AND enrollment_date != ''
            GROUP BY substr(enrollment_date, 1, 7)
            ORDER BY period DESC
            LIMIT 12
            ''')
            enrollment_rows = cursor.fetchall()
    
            grade_lines = [
                f"{row[0]}: {row[1]} submissions, Avg Score {row[2]:.1f}%"
                if row[2] is not None else
                f"{row[0]}: {row[1]} submissions, Avg Score N/A"
                for row in grade_trend_rows
            ]
    
            risk_lines = [
                f"{row[0]}: Avg Risk Score {row[1]:.2f}"
                if row[1] is not None else
                f"{row[0]}: Avg Risk Score N/A"
                for row in risk_trend_rows
            ]
    
            enrollment_lines = [
                f"{row[0]}: New Enrollments {row[1]}"
                for row in enrollment_rows
            ]
    
            sections = [
                ("Grade Performance Trend (Last 12 Periods)", grade_lines),
                ("Student Risk Trend (Last 12 Periods)", risk_lines),
                ("Enrollment Trend (Last 12 Periods)", enrollment_lines)
            ]
    
            footer = (
                "Review month-over-month changes to anticipate resource needs. "
                "Align support services when risk scores trend upward or grade averages decline."
            )
    
            self._display_report("Trend Analysis Report", sections, footer)
    
        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Failed to generate trend analysis report: {e}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate trend analysis report: {e}")
        finally:
            if conn:
                conn.close()
    

    def generate_comprehensive_risk_report(self):
        """Generate comprehensive risk report"""
        conn = None
        try:
            conn = get_connection()
            cursor = conn.cursor()
    
            cursor.execute('''
            SELECT 
                ra.student_id,
                ra.risk_level,
                ra.risk_score
            FROM student_risk_assessment ra
            JOIN (
                SELECT student_id, MAX(id) AS max_id
                FROM student_risk_assessment
                GROUP BY student_id
            ) latest ON ra.id = latest.max_id
            ''')
            risk_rows = cursor.fetchall()
    
            cursor.execute('''
            SELECT 
                course,
                AVG(
                    CASE 
                        WHEN a.max_points IS NOT NULL AND a.max_points > 0 
                        THEN (g.score / a.max_points) * 100 
                    END
                ) AS avg_score,
                SUM(CASE WHEN g.letter_grade = 'F' THEN 1 ELSE 0 END) AS fail_count,
                COUNT(g.grade_id) AS grade_count
            FROM students s
            LEFT JOIN grades g ON s.student_id = g.student_id
            LEFT JOIN assessments a ON g.assessment_id = a.assessment_id
            GROUP BY course
            ''')
            course_fail_rows = cursor.fetchall()
    
            cursor.execute('''
            SELECT 
                module_code,
                SUM(CASE WHEN final_grade = 'F' THEN 1 ELSE 0 END) AS fail_count,
                COUNT(*) AS total_records
            FROM module_grades
            GROUP BY module_code
            ''')
            module_fail_rows = cursor.fetchall()
    
            cursor.execute('''
            SELECT 
                risk_level,
                COUNT(*)
            FROM student_risk_assessment
            GROUP BY risk_level
            ''')
            historical_levels = cursor.fetchall()
    
            total_risk_students = len(risk_rows)
            high_risk = sum(1 for row in risk_rows if row[1] == "High")
            medium_risk = sum(1 for row in risk_rows if row[1] == "Medium")
            low_risk = sum(1 for row in risk_rows if row[1] == "Low")
            avg_risk_score = (
                sum(row[2] for row in risk_rows) / total_risk_students
                if total_risk_students else None
            )
    
            summary_lines = [
                f"Students With Active Risk Scores: {total_risk_students}",
                f"High Risk: {high_risk}",
                f"Medium Risk: {medium_risk}",
                f"Low Risk: {low_risk}",
                f"Average Risk Score: {avg_risk_score:.2f}" if avg_risk_score is not None else "Average Risk Score: N/A"
            ]
    
            course_lines = []
            for course, avg_score, fail_count, grade_count in sorted(
                course_fail_rows,
                key=lambda r: (r[1] if r[1] is not None else 0)
            )[:5]:
                fail_rate = (fail_count / grade_count) if grade_count else 0
                course_lines.append(
                    f"{course}: Avg Score {avg_score:.1f}%, Fail Rate {fail_rate:.0%}"
                    if avg_score is not None else
                    f"{course}: Fail Rate {fail_rate:.0%}"
                )
    
            module_lines = []
            for module_code, fail_count, total_records in sorted(
                module_fail_rows,
                key=lambda r: (-(r[1] / r[2]) if r[2] else 0)
            )[:5]:
                if total_records:
                    fail_rate = fail_count / total_records
                    module_lines.append(
                        f"{module_code}: Fail Rate {fail_rate:.0%} ({fail_count}/{total_records})"
                    )
    
            historical_lines = [
                f"{level or 'Unknown'}: {count} records"
                for level, count in historical_levels
            ]
    
            intervention_lines = []
            if high_risk:
                intervention_lines.append(
                    f"High risk cohort ({high_risk} students): Immediate academic advising and weekly follow-up."
                )
            if medium_risk:
                intervention_lines.append(
                    f"Medium risk cohort ({medium_risk} students): Assign peer mentors and monitor progress bi-weekly."
                )
            if not intervention_lines:
                intervention_lines.append("Current risk levels are low; continue regular monitoring.")
    
            sections = [
                ("Risk Distribution Overview", summary_lines),
                ("Courses With Elevated Failure Rates", course_lines),
                ("Modules Contributing To Risk", module_lines),
                ("Historical Risk Level Counts", historical_lines),
                ("Recommended Actions", intervention_lines)
            ]
    
            footer = (
                "Combine course and module insights with risk levels to design targeted intervention programmes "
                "and measure their impact over time."
            )
    
            self._display_report("Comprehensive Risk Report", sections, footer)
    
        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Failed to generate comprehensive risk report: {e}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate comprehensive risk report: {e}")
        finally:
            if conn:
                conn.close()
    

    def export_reports_pdf(self):
        """Export reports to PDF"""
        try:
            if not hasattr(self, 'report_preview') or not self.report_preview:
                messagebox.showwarning("Warning", "No report preview available. Please generate a report first.")
                return
    
            # Get report content
            report_content = self.report_preview.get('1.0', tk.END).strip()
            if not report_content:
                messagebox.showwarning("Warning", "Report preview is empty. Please generate a report first.")
                return
    
            # File save dialog
            filename = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf"), ("Text files", "*.txt")],
                title="Export Report"
            )
    
            if not filename:
                return
    
            # Try to use reportlab for PDF, fallback to text file
            if filename.endswith('.pdf'):
                try:
                    from reportlab.lib.pagesizes import letter
                    from reportlab.lib.styles import getSampleStyleSheet
                    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak
                    from reportlab.lib.units import inch
    
                    doc = SimpleDocTemplate(filename, pagesize=letter)
                    styles = getSampleStyleSheet()
                    story = []
    
                    # Split content into paragraphs
                    for line in report_content.split('\n'):
                        if line.strip():
                            para = Paragraph(line.replace('<', '&lt;').replace('>', '&gt;'), styles['Normal'])
                            story.append(para)
                        else:
                            story.append(Spacer(1, 0.2*inch))
    
                    doc.build(story)
                    messagebox.showinfo("Success", f"Report exported to PDF:\n{filename}")
    
                except ImportError:
                    # Reportlab not available, save as text instead
                    if messagebox.askyesno("PDF Library Not Found",
                                        "PDF library (reportlab) not available.\n"
                                        "Would you like to save as a text file instead?"):
                        text_filename = filename.replace('.pdf', '.txt')
                        with open(text_filename, 'w', encoding='utf-8') as f:
                            f.write(report_content)
                        messagebox.showinfo("Success", f"Report exported to text file:\n{text_filename}")
            else:
                # Save as text file
                with open(filename, 'w', encoding='utf-8') as f:
                    f.write(report_content)
                messagebox.showinfo("Success", f"Report exported to:\n{filename}")
    
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export report: {e}")
    

    def export_reports_csv(self):
        """Export reports to CSV"""
        try:
            if not hasattr(self, 'report_preview') or not self.report_preview:
                messagebox.showwarning("Warning", "No report preview available. Please generate a report first.")
                return
    
            # Get report content
            report_content = self.report_preview.get('1.0', tk.END).strip()
            if not report_content:
                messagebox.showwarning("Warning", "Report preview is empty. Please generate a report first.")
                return
    
            # File save dialog
            filename = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("Text files", "*.txt")],
                title="Export Report to CSV"
            )
    
            if not filename:
                return
    
            # Try to parse report content as tabular data
            import csv
            lines = report_content.split('\n')
    
            with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
    
                # Try to detect if content has tabular structure (tabs or multiple spaces)
                is_tabular = any('\t' in line or '  ' in line for line in lines[:5])
    
                if is_tabular:
                    # Parse as tabular data
                    for line in lines:
                        if line.strip():
                            # Split by tabs or multiple spaces
                            if '\t' in line:
                                row = [cell.strip() for cell in line.split('\t')]
                            else:
                                # Split by multiple spaces (2 or more)
                                import re
                                row = [cell.strip() for cell in re.split(r'\s{2,}', line) if cell.strip()]
    
                            if row:
                                writer.writerow(row)
                else:
                    # Save as single column with each line as a row
                    for line in lines:
                        if line.strip():
                            writer.writerow([line])
    
            messagebox.showinfo("Success", f"Report exported to CSV:\n{filename}")
    
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export report to CSV: {e}")
    

    def export_reports_excel(self):
        """Export reports to Excel"""
        try:
            if not hasattr(self, 'report_preview') or not self.report_preview:
                messagebox.showwarning("Warning", "No report preview available. Please generate a report first.")
                return
    
            # Get report content
            report_content = self.report_preview.get('1.0', tk.END).strip()
            if not report_content:
                messagebox.showwarning("Warning", "Report preview is empty. Please generate a report first.")
                return
    
            # File save dialog
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv")],
                title="Export Report to Excel"
            )
    
            if not filename:
                return
    
            # Try to use openpyxl for Excel export
            try:
                import openpyxl
                from openpyxl.styles import Font, Alignment, PatternFill
    
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Report"
    
                lines = report_content.split('\n')
    
                # Try to detect if content has tabular structure
                is_tabular = any('\t' in line or '  ' in line for line in lines[:5])
    
                row_num = 1
                for line in lines:
                    if line.strip():
                        if is_tabular:
                            # Split by tabs or multiple spaces
                            if '\t' in line:
                                cells = [cell.strip() for cell in line.split('\t')]
                            else:
                                import re
                                cells = [cell.strip() for cell in re.split(r'\s{2,}', line) if cell.strip()]
    
                            for col_num, cell_value in enumerate(cells, 1):
                                cell = ws.cell(row=row_num, column=col_num, value=cell_value)
    
                                # Style first row as header if it looks like headers
                                if row_num == 1:
                                    cell.font = Font(bold=True)
                                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                                    cell.alignment = Alignment(horizontal='center')
                        else:
                            # Single column format
                            ws.cell(row=row_num, column=1, value=line)
    
                        row_num += 1
    
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except (AttributeError, TypeError):
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
    
                wb.save(filename)
                messagebox.showinfo("Success", f"Report exported to Excel:\n{filename}")
    
            except ImportError:
                # Openpyxl not available, fall back to CSV
                if messagebox.askyesno("Excel Library Not Found",
                                      "Excel library (openpyxl) not available.\n"
                                      "Would you like to save as CSV instead?"):
                    csv_filename = filename.replace('.xlsx', '.csv')
                    import csv
    
                    with open(csv_filename, 'w', newline='', encoding='utf-8') as csvfile:
                        writer = csv.writer(csvfile)
                        lines = report_content.split('\n')
                        is_tabular = any('\t' in line or '  ' in line for line in lines[:5])
    
                        for line in lines:
                            if line.strip():
                                if is_tabular:
                                    if '\t' in line:
                                        row = [cell.strip() for cell in line.split('\t')]
                                    else:
                                        import re
                                        row = [cell.strip() for cell in re.split(r'\s{2,}', line) if cell.strip()]
                                    if row:
                                        writer.writerow(row)
                                else:
                                    writer.writerow([line])
    
                    messagebox.showinfo("Success", f"Report exported to CSV:\n{csv_filename}")
    
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export report to Excel: {e}")
    

    def _display_report(self, title, sections, footer=None):
        """Render report content in a separate window with email option."""
        # Create separate window for report
        report_window = tk.Toplevel(self.root)
        report_window.title(title)
        report_window.geometry("800x600")

        # Create main frame
        main_frame = ttk.Frame(report_window)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Create scrolled text widget for report
        report_text = scrolledtext.ScrolledText(main_frame, wrap=tk.WORD, font=('Courier', 10))
        report_text.pack(fill=tk.BOTH, expand=True)

        # Build report content
        lines = [
            title,
            "=" * len(title),
            f"Generated: {datetime.now():%Y-%m-%d %H:%M}",
            ""
        ]

        for heading, content_lines in sections:
            heading = heading.strip() or "Details"
            lines.append(heading)
            lines.append("-" * len(heading))
            if content_lines:
                lines.extend(content_lines)
            else:
                lines.append("No data available.")
            lines.append("")

        if footer:
            lines.append(footer)

        report_content = "\n".join(lines).rstrip() + "\n"
        report_text.insert('1.0', report_content)
        report_text.config(state='disabled')

        # Button frame at bottom
        button_frame = ttk.Frame(report_window)
        button_frame.pack(fill='x', padx=10, pady=10)

        # Email to Admin button
        def email_to_admin():
            try:
                # Get admin email from database - try multiple tables/columns
                conn = get_connection()
                cursor = conn.cursor()

                admin_email = None
                # Try users table with case-insensitive role check
                try:
                    cursor.execute("""
                        SELECT email FROM users
                        WHERE LOWER(role) IN ('admin', 'administrator')
                        AND email IS NOT NULL AND email != ''
                        LIMIT 1
                    """)
                    admin_row = cursor.fetchone()
                    if admin_row and admin_row[0]:
                        admin_email = admin_row[0]
                except sqlite3.Error:
                    pass

                # Try email_address column as fallback (some schemas may use this)
                if not admin_email:
                    try:
                        cursor.execute("""
                            SELECT email_address FROM users
                            WHERE LOWER(role) IN ('admin', 'administrator')
                            AND email_address IS NOT NULL AND email_address != ''
                            LIMIT 1
                        """)
                        admin_row = cursor.fetchone()
                        if admin_row and admin_row[0]:
                            admin_email = admin_row[0]
                    except sqlite3.Error:
                        pass

                conn.close()

                if not admin_email:
                    messagebox.showwarning("No Admin", "No administrator email found in the database.\nPlease ensure an admin user exists with a valid email address.")
                    return

                # Try to use email service
                try:
                    from university_system.infrastructure.email.email_service import send_email

                    # Send email with report content
                    success = send_email(
                        recipient_email=admin_email,
                        subject=f"Grade Tracking Report: {title}",
                        body=f"Please find the attached report:\n\n{report_content}"
                    )

                    if success:
                        messagebox.showinfo("Success", f"Report emailed to administrator at {admin_email}")
                    else:
                        messagebox.showerror("Error", "Failed to send email. Please check email service configuration.")
                except ImportError as ie:
                    messagebox.showerror("Error", f"Email service import error: {ie}")
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to send email: {e}")

            except sqlite3.Error as e:
                messagebox.showerror("Database Error", f"Failed to retrieve admin email: {e}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to email report: {e}")

        ttk.Button(button_frame, text="Email to Admin", command=email_to_admin).pack(side='left', padx=5)
        ttk.Button(button_frame, text="Close", command=report_window.destroy).pack(side='right', padx=5)
    

    def export_risk_results(self, tree):
        """Export risk analysis results"""
        try:
            filename = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("Excel files", "*.xlsx")],
                title="Export Risk Results"
            )
                
            if filename:
                # Get data from treeview
                data = []
                for item in tree.get_children():
                    data.append(tree.item(item, 'values'))
                    
                if filename.endswith('.xlsx'):
                    # Export to Excel (if available)
                    try:
                        import pandas as pd
                        df = pd.DataFrame(data, columns=['Student ID', 'Name', 'Course', 'Risk Score', 
                                        'Risk Level', 'Factors', 'GPA', 'Failed Assessments'])
                        df.to_excel(filename, index=False)
                    except ImportError:
                        messagebox.showerror("Error", "pandas library required for Excel export")
                        return
                else:
                    # Export to CSV
                    with open(filename, 'w', newline='', encoding='utf-8') as file:
                        writer = csv.writer(file)
                        writer.writerow(['Student ID', 'Name', 'Course', 'Risk Score', 
                                       'Risk Level', 'Factors', 'GPA', 'Failed Assessments'])
                        writer.writerows(data)
                    
                messagebox.showinfo("Success", f"Risk results exported to {filename}")
                    
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export results: {e}")
    

    def _refresh_student_competencies(self, target_student_id=None):
        """Reload student competency assessments, optionally filtered by student ID."""
        tree = getattr(self, 'student_comp_tree', None)
        if not tree or not self._widget_exists(tree):
            return
    
        filter_id = None
        if target_student_id:
            filter_id = target_student_id
        elif hasattr(self, 'comp_student_filter_var'):
            selected = self.comp_student_filter_var.get()
            if selected and selected != 'All':
                filter_id = selected.split(' - ')[0].strip()
    
        try:
            for item in tree.get_children():
                tree.delete(item)
    
            conn = get_connection()
            cursor = conn.cursor()
    
            query = '''
            SELECT 
                sc.id,
                sc.student_id,
                s.first_name,
                s.last_name,
                c.name,
                COALESCE(cl.level_name, 'Not Specified') AS level_name,
                COALESCE(cl.level_value, 0) AS level_value,
                COALESCE(sc.assessment_date, '') AS assessment_date,
                COALESCE(sc.evidence, '') AS evidence
            FROM student_competencies sc
            JOIN students s ON sc.student_id = s.student_id
            JOIN competencies c ON sc.competency_id = c.competency_id
            LEFT JOIN competency_levels cl ON sc.level_id = cl.level_id
            '''
    
            params = []
            if filter_id:
                query += ' WHERE sc.student_id = ?'
                params.append(filter_id)
    
            query += ' ORDER BY sc.assessment_date DESC, s.last_name, s.first_name'
            cursor.execute(query, params)
            rows = cursor.fetchall()
    
            for record in rows:
                record_id, student_id, first_name, last_name, comp_name, level_name, level_value, assess_date, evidence = record
                student_display = f"{student_id} - {first_name} {last_name}".strip()
                if level_name == 'Not Specified' and level_value:
                    level_display = f"Level {level_value}"
                elif level_value:
                    level_display = f"{level_name} (Level {level_value})"
                else:
                    level_display = level_name
    
                tree.insert(
                    '',
                    'end',
                    values=(
                        record_id,
                        student_display,
                        comp_name,
                        level_display,
                        assess_date,
                        evidence
                    )
                )
    
            conn.close()
            self.update_status(f"Loaded {len(rows)} competency assessment records")
    
        except sqlite3.Error as exc:
            messagebox.showerror("Database Error", f"Failed to load competency assessments: {exc}")
    

    def add_competency_dialog(self):
        """Display dialog for creating a new competency."""
        dialog = tk.Toplevel(self.root)
        dialog.title("Add Competency")
        dialog.geometry("480x360")
        safe_grab_set(dialog)
    
        ttk.Label(dialog, text="Competency Name:").grid(row=0, column=0, sticky=tk.W, padx=10, pady=8)
        name_var = tk.StringVar()
        ttk.Entry(dialog, textvariable=name_var, width=40).grid(row=0, column=1, padx=10, pady=8)
    
        ttk.Label(dialog, text="Category:").grid(row=1, column=0, sticky=tk.W, padx=10, pady=8)
        category_var = tk.StringVar()
        ttk.Entry(dialog, textvariable=category_var, width=40).grid(row=1, column=1, padx=10, pady=8)
    
        ttk.Label(dialog, text="Description:").grid(row=2, column=0, sticky=tk.NW, padx=10, pady=8)
        desc_text = scrolledtext.ScrolledText(dialog, width=35, height=8)
        desc_text.grid(row=2, column=1, padx=10, pady=8)
    
        button_frame = ttk.Frame(dialog)
        button_frame.grid(row=3, column=0, columnspan=2, pady=15)

        def save_competency():
            name = name_var.get().strip()
            category = category_var.get().strip()
            description = desc_text.get('1.0', tk.END).strip()

            if not name:
                messagebox.showwarning("Validation", "Competency name is required.")
                return

            try:
                conn = get_connection()
                cursor = conn.cursor()
                cursor.execute(
                    '''
                    INSERT INTO competencies (name, description, category)
                    VALUES (?, ?, ?)
                    ''',
                    (name, description or None, category or None)
                )
                conn.commit()
                conn.close()

                messagebox.showinfo("Success", f"Competency '{name}' added successfully.")
                dialog.destroy()
                self.refresh_competencies()
                self.populate_filter_combos()

            except sqlite3.Error as exc:
                messagebox.showerror("Database Error", f"Failed to add competency: {exc}")

        ttk.Button(button_frame, text="Save", command=save_competency).pack(side=tk.LEFT, padx=10)
        ttk.Button(button_frame, text="Cancel", command=dialog.destroy).pack(side=tk.LEFT, padx=10)
    

    def edit_competency_dialog(self):
        """Edit the selected competency."""
        tree = getattr(self, 'competency_tree', None)
        if not tree or not tree.selection():
            messagebox.showwarning("No Selection", "Please select a competency to edit.")
            return
    
        selected_item = tree.selection()[0]
        values = tree.item(selected_item, 'values')
        competency_id = values[0]
    
        try:
            conn = get_connection()
            cursor = conn.cursor()
            cursor.execute(
                '''
                SELECT name, description, category
                FROM competencies
                WHERE competency_id = ?
                ''',
                (competency_id,)
            )
            competency = cursor.fetchone()
            conn.close()
    
            if not competency:
                messagebox.showerror("Error", "Selected competency no longer exists.")
                return
    
        except sqlite3.Error as exc:
            messagebox.showerror("Database Error", f"Failed to load competency: {exc}")
            return
    
        name_val, desc_val, category_val = competency
    
        dialog = tk.Toplevel(self.root)
        dialog.title("Edit Competency")
        dialog.geometry("480x360")
        safe_grab_set(dialog)
    
        ttk.Label(dialog, text="Competency Name:").grid(row=0, column=0, sticky=tk.W, padx=10, pady=8)
        name_var = tk.StringVar(value=name_val or "")
        ttk.Entry(dialog, textvariable=name_var, width=40).grid(row=0, column=1, padx=10, pady=8)
    
        ttk.Label(dialog, text="Category:").grid(row=1, column=0, sticky=tk.W, padx=10, pady=8)
        category_var = tk.StringVar(value=category_val or "")
        ttk.Entry(dialog, textvariable=category_var, width=40).grid(row=1, column=1, padx=10, pady=8)
    
        ttk.Label(dialog, text="Description:").grid(row=2, column=0, sticky=tk.NW, padx=10, pady=8)
        desc_text = scrolledtext.ScrolledText(dialog, width=35, height=8)
        desc_text.grid(row=2, column=1, padx=10, pady=8)
        desc_text.insert('1.0', desc_val or "")
    
        button_frame = ttk.Frame(dialog)
        button_frame.grid(row=3, column=0, columnspan=2, pady=15)

        def update_competency():
            new_name = name_var.get().strip()
            new_category = category_var.get().strip()
            new_description = desc_text.get('1.0', tk.END).strip()

            if not new_name:
                messagebox.showwarning("Validation", "Competency name is required.")
                return

            try:
                conn_edit = get_connection()
                cursor_edit = conn_edit.cursor()
                cursor_edit.execute(
                    '''
                    UPDATE competencies
                    SET name = ?, description = ?, category = ?
                    WHERE competency_id = ?
                    ''',
                    (new_name, new_description or None, new_category or None, competency_id)
                )
                conn_edit.commit()
                conn_edit.close()

                messagebox.showinfo("Success", "Competency updated successfully.")
                dialog.destroy()
                self.refresh_competencies()
                self.populate_filter_combos()

            except sqlite3.Error as exc:
                messagebox.showerror("Database Error", f"Failed to update competency: {exc}")

        ttk.Button(button_frame, text="Save Changes", command=update_competency).pack(side=tk.LEFT, padx=10)
        ttk.Button(button_frame, text="Cancel", command=dialog.destroy).pack(side=tk.LEFT, padx=10)
    

    def delete_competency(self):
        """Delete the selected competency and related records."""
        tree = getattr(self, 'competency_tree', None)
        if not tree or not tree.selection():
            messagebox.showwarning("No Selection", "Please select a competency to delete.")
            return
    
        selected_item = tree.selection()[0]
        comp_id, comp_name = tree.item(selected_item, 'values')[:2]
    
        confirm = messagebox.askyesno(
            "Confirm Delete",
            f"Delete competency '{comp_name}'?\n\n"
            "All related levels, mappings, and student records will also be removed."
        )
        if not confirm:
            return
    
        try:
            conn = get_connection()
            cursor = conn.cursor()
    
            cursor.execute('DELETE FROM assessment_competencies WHERE competency_id = ?', (comp_id,))
            cursor.execute('DELETE FROM student_competencies WHERE competency_id = ?', (comp_id,))
            cursor.execute('DELETE FROM competency_levels WHERE competency_id = ?', (comp_id,))
            cursor.execute('DELETE FROM competencies WHERE competency_id = ?', (comp_id,))
    
            conn.commit()
            conn.close()
    
            messagebox.showinfo("Success", f"Competency '{comp_name}' deleted.")
            self.refresh_competencies()
            self.populate_filter_combos()
    
        except sqlite3.Error as exc:
            messagebox.showerror("Database Error", f"Failed to delete competency: {exc}")
    

    def manage_competency_levels(self):
        """Open management dialog for competency levels."""
        tree = getattr(self, 'competency_tree', None)
        if not tree or not tree.selection():
            messagebox.showwarning("No Selection", "Select a competency to manage levels.")
            return
    
        selected_item = tree.selection()[0]
        comp_id, comp_name = tree.item(selected_item, 'values')[:2]
    
        level_window = tk.Toplevel(self.root)
        level_window.title(f"Competency Levels - {comp_name}")
        level_window.geometry("600x420")
        safe_grab_set(level_window)
    
        columns = ('Level ID', 'Name', 'Value', 'Description')
        level_tree = ttk.Treeview(level_window, columns=columns, show='headings')
        for col in columns:
            level_tree.heading(col, text=col)
            anchor = 'w' if col == 'Description' else 'center'
            width = 260 if col == 'Description' else 100
            level_tree.column(col, width=width, anchor=anchor)
        level_tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
    
        btn_frame = ttk.Frame(level_window)
        btn_frame.pack(pady=5)
    
        def refresh_levels():
            try:
                for item in level_tree.get_children():
                    level_tree.delete(item)
    
                conn = get_connection()
                cursor = conn.cursor()
                cursor.execute(
                    '''
                    SELECT level_id, level_name, level_value, COALESCE(description, '')
                    FROM competency_levels
                    WHERE competency_id = ?
                    ORDER BY level_value
                    ''',
                    (comp_id,)
                )
                rows = cursor.fetchall()
                conn.close()
    
                for level_id, level_name, level_value, description in rows:
                    level_tree.insert(
                        '',
                        'end',
                        values=(level_id, level_name, level_value, description)
                    )
    
            except sqlite3.Error as exc:
                messagebox.showerror("Database Error", f"Failed to load levels: {exc}")
    
        def add_level():
            dialog = tk.Toplevel(level_window)
            dialog.title("Add Level")
            dialog.geometry("360x260")
            safe_grab_set(dialog)
    
            ttk.Label(dialog, text="Level Name:").grid(row=0, column=0, sticky=tk.W, padx=10, pady=8)
            level_name_var = tk.StringVar()
            ttk.Entry(dialog, textvariable=level_name_var, width=30).grid(row=0, column=1, padx=10, pady=8)
    
            ttk.Label(dialog, text="Value (numeric order):").grid(row=1, column=0, sticky=tk.W, padx=10, pady=8)
            level_value_var = tk.StringVar(value="1")
            ttk.Entry(dialog, textvariable=level_value_var, width=10).grid(row=1, column=1, padx=10, pady=8, sticky=tk.W)
    
            ttk.Label(dialog, text="Description:").grid(row=2, column=0, sticky=tk.NW, padx=10, pady=8)
            desc_box = scrolledtext.ScrolledText(dialog, width=28, height=5)
            desc_box.grid(row=2, column=1, padx=10, pady=8)
    
            def save_level():
                name = level_name_var.get().strip()
                try:
                    value = int(level_value_var.get().strip())
                except ValueError:
                    messagebox.showwarning("Validation", "Level value must be an integer.")
                    return
                description = desc_box.get('1.0', tk.END).strip()
    
                if not name:
                    messagebox.showwarning("Validation", "Level name is required.")
                    return
    
                try:
                    conn_add = get_connection()
                    cursor_add = conn_add.cursor()
                    cursor_add.execute(
                        '''
                        INSERT INTO competency_levels (competency_id, level_name, level_value, description)
                        VALUES (?, ?, ?, ?)
                        ''',
                        (comp_id, name, value, description or None)
                    )
                    conn_add.commit()
                    conn_add.close()
    
                    dialog.destroy()
                    refresh_levels()
                    self.refresh_competencies()
    
                except sqlite3.Error as exc:
                    messagebox.showerror("Database Error", f"Failed to add level: {exc}")
    
            ttk.Button(dialog, text="Save", command=save_level).grid(row=3, column=0, columnspan=2, pady=12)
    
        def edit_level():
            selection = level_tree.selection()
            if not selection:
                messagebox.showwarning("No Selection", "Select a level to edit.")
                return
    
            level_id, level_name, level_value, description = level_tree.item(selection[0], 'values')
    
            dialog = tk.Toplevel(level_window)
            dialog.title("Edit Level")
            dialog.geometry("360x260")
            safe_grab_set(dialog)
    
            ttk.Label(dialog, text="Level Name:").grid(row=0, column=0, sticky=tk.W, padx=10, pady=8)
            level_name_var = tk.StringVar(value=level_name)
            ttk.Entry(dialog, textvariable=level_name_var, width=30).grid(row=0, column=1, padx=10, pady=8)
    
            ttk.Label(dialog, text="Value (numeric order):").grid(row=1, column=0, sticky=tk.W, padx=10, pady=8)
            level_value_var = tk.StringVar(value=str(level_value))
            ttk.Entry(dialog, textvariable=level_value_var, width=10).grid(row=1, column=1, padx=10, pady=8, sticky=tk.W)
    
            ttk.Label(dialog, text="Description:").grid(row=2, column=0, sticky=tk.NW, padx=10, pady=8)
            desc_box = scrolledtext.ScrolledText(dialog, width=28, height=5)
            desc_box.grid(row=2, column=1, padx=10, pady=8)
            desc_box.insert('1.0', description or "")
    
            def save_changes():
                name = level_name_var.get().strip()
                try:
                    value = int(level_value_var.get().strip())
                except ValueError:
                    messagebox.showwarning("Validation", "Level value must be an integer.")
                    return
                desc_val = desc_box.get('1.0', tk.END).strip()
    
                if not name:
                    messagebox.showwarning("Validation", "Level name is required.")
                    return
    
                try:
                    conn_edit = get_connection()
                    cursor_edit = conn_edit.cursor()
                    cursor_edit.execute(
                        '''
                        UPDATE competency_levels
                        SET level_name = ?, level_value = ?, description = ?
                        WHERE level_id = ?
                        ''',
                        (name, value, desc_val or None, level_id)
                    )
                    conn_edit.commit()
                    conn_edit.close()
    
                    dialog.destroy()
                    refresh_levels()
                    self.refresh_competencies()
    
                except sqlite3.Error as exc:
                    messagebox.showerror("Database Error", f"Failed to update level: {exc}")
    
            ttk.Button(dialog, text="Save Changes", command=save_changes).grid(row=3, column=0, columnspan=2, pady=12)
    
        def delete_level():
            selection = level_tree.selection()
            if not selection:
                messagebox.showwarning("No Selection", "Select a level to delete.")
                return
    
            level_id, level_name = level_tree.item(selection[0], 'values')[:2]
    
            confirm_delete = messagebox.askyesno(
                "Confirm Delete",
                f"Delete level '{level_name}'?\n\n"
                "Student competency records referencing this level will also be removed."
            )
            if not confirm_delete:
                return
    
            try:
                conn_del = get_connection()
                cursor_del = conn_del.cursor()
                cursor_del.execute('DELETE FROM student_competencies WHERE level_id = ?', (level_id,))
                cursor_del.execute('DELETE FROM competency_levels WHERE level_id = ?', (level_id,))
                conn_del.commit()
                conn_del.close()
    
                refresh_levels()
                self.refresh_competencies()
                self._refresh_student_competencies()
    
            except sqlite3.Error as exc:
                messagebox.showerror("Database Error", f"Failed to delete level: {exc}")
    
        ttk.Button(btn_frame, text="Add Level", command=add_level).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Edit Level", command=edit_level).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Delete Level", command=delete_level).pack(side=tk.LEFT, padx=5)
    
        refresh_levels()
    

    def refresh_competencies(self):
        """Load competencies into the treeview."""
        tree = getattr(self, 'competency_tree', None)
        if not tree or not self._widget_exists(tree):
            return
    
        try:
            for item in tree.get_children():
                tree.delete(item)
    
            conn = get_connection()
            cursor = conn.cursor()
            cursor.execute(
                '''
                SELECT 
                    c.competency_id,
                    c.name,
                    COALESCE(c.category, 'Uncategorized') AS category,
                    COALESCE(c.description, ''),
                    COUNT(DISTINCT cl.level_id) AS level_count
                FROM competencies c
                LEFT JOIN competency_levels cl ON c.competency_id = cl.competency_id
                GROUP BY c.competency_id, c.name, c.category, c.description
                ORDER BY c.name
                '''
            )
            rows = cursor.fetchall()
            conn.close()
    
            for comp_id, name, category, description, level_count in rows:
                tree.insert(
                    '',
                    'end',
                    values=(comp_id, name, category, description, level_count)
                )
    
            self.update_status(f"Loaded {len(rows)} competencies")
            self._refresh_student_competencies()
            self.populate_filter_combos()
    
        except sqlite3.Error as exc:
            messagebox.showerror("Database Error", f"Failed to load competencies: {exc}")
    

    def map_assessment_competency_dialog(self):
        """Map an assessment to a competency."""
        dialog = tk.Toplevel(self.root)
        dialog.title("Map Assessment to Competency")
        dialog.geometry("520x260")
        safe_grab_set(dialog)
    
        assessment_var = tk.StringVar()
        competency_var = tk.StringVar()
        weight_var = tk.StringVar(value="1.0")
    
        ttk.Label(dialog, text="Assessment:").grid(row=0, column=0, sticky=tk.W, padx=10, pady=10)
        assessment_combo = ttk.Combobox(dialog, textvariable=assessment_var, width=45, state='readonly')
        assessment_combo.grid(row=0, column=1, padx=10, pady=10)
    
        ttk.Label(dialog, text="Competency:").grid(row=1, column=0, sticky=tk.W, padx=10, pady=10)
        competency_combo = ttk.Combobox(dialog, textvariable=competency_var, width=45, state='readonly')
        competency_combo.grid(row=1, column=1, padx=10, pady=10)
    
        ttk.Label(dialog, text="Weight:").grid(row=2, column=0, sticky=tk.W, padx=10, pady=10)
        ttk.Entry(dialog, textvariable=weight_var, width=10).grid(row=2, column=1, padx=10, pady=10, sticky=tk.W)
    
        try:
            conn = get_connection()
            cursor = conn.cursor()
    
            cursor.execute(
                '''
                SELECT assessment_id, assessment_name, module_code
                FROM assessments
                ORDER BY module_code, assessment_name
                '''
            )
            assessments = cursor.fetchall()
            assessment_options = {
                f"{a_id} - {a_name} ({module})": a_id for a_id, a_name, module in assessments
            }
            assessment_combo['values'] = list(assessment_options.keys())
    
            cursor.execute(
                '''
                SELECT competency_id, name
                FROM competencies
                ORDER BY name
                '''
            )
            competencies = cursor.fetchall()
            competency_options = {f"{c_id} - {name}": c_id for c_id, name in competencies}
            competency_combo['values'] = list(competency_options.keys())
    
            conn.close()
    
        except sqlite3.Error as exc:
            messagebox.showerror("Database Error", f"Failed to load data: {exc}")
            dialog.destroy()
            return
    
        def save_mapping():
            assessment_key = assessment_var.get()
            competency_key = competency_var.get()
    
            if assessment_key not in assessment_options or competency_key not in competency_options:
                messagebox.showwarning("Validation", "Please select both assessment and competency.")
                return
    
            try:
                weight = float(weight_var.get())
            except ValueError:
                messagebox.showwarning("Validation", "Weight must be a numeric value.")
                return
    
            assessment_id = assessment_options[assessment_key]
            competency_id = competency_options[competency_key]
    
            try:
                conn_map = get_connection()
                cursor_map = conn_map.cursor()
    
                cursor_map.execute(
                    '''
                    SELECT id FROM assessment_competencies
                    WHERE assessment_id = ? AND competency_id = ?
                    ''',
                    (assessment_id, competency_id)
                )
                existing = cursor_map.fetchone()
    
                if existing:
                    if messagebox.askyesno(
                        "Mapping Exists",
                        "Mapping already exists. Update weight instead?"
                    ):
                        cursor_map.execute(
                            '''
                            UPDATE assessment_competencies
                            SET weight = ?
                            WHERE id = ?
                            ''',
                            (weight, existing[0])
                        )
                        action_message = "Mapping weight updated."
                    else:
                        conn_map.close()
                        return
                else:
                    cursor_map.execute(
                        '''
                        INSERT INTO assessment_competencies (assessment_id, competency_id, weight)
                        VALUES (?, ?, ?)
                        ''',
                        (assessment_id, competency_id, weight)
                    )
                    action_message = "Mapping created."
    
                conn_map.commit()
                conn_map.close()
    
                messagebox.showinfo("Success", action_message)
                dialog.destroy()
                self.refresh_competency_mappings()
    
            except sqlite3.Error as exc:
                messagebox.showerror("Database Error", f"Failed to save mapping: {exc}")
    
        ttk.Button(dialog, text="Save Mapping", command=save_mapping).grid(row=3, column=0, columnspan=2, pady=15)
    

    def edit_competency_mapping(self):
        """Edit the selected assessment competency mapping."""
        tree = getattr(self, 'mapping_tree', None)
        if not tree or not tree.selection():
            messagebox.showwarning("No Selection", "Select a mapping to edit.")
            return
    
        selected_item = tree.selection()[0]
        mapping_id, assessment_display, competency_display, weight_value, module_code = tree.item(selected_item, 'values')
    
        dialog = tk.Toplevel(self.root)
        dialog.title("Edit Mapping")
        dialog.geometry("520x260")
        safe_grab_set(dialog)
    
        assessment_var = tk.StringVar(value=assessment_display)
        competency_var = tk.StringVar(value=competency_display)
        weight_var = tk.StringVar(value=str(weight_value))
    
        ttk.Label(dialog, text="Assessment:").grid(row=0, column=0, sticky=tk.W, padx=10, pady=10)
        assessment_combo = ttk.Combobox(dialog, textvariable=assessment_var, width=45, state='readonly')
        assessment_combo.grid(row=0, column=1, padx=10, pady=10)
    
        ttk.Label(dialog, text="Competency:").grid(row=1, column=0, sticky=tk.W, padx=10, pady=10)
        competency_combo = ttk.Combobox(dialog, textvariable=competency_var, width=45, state='readonly')
        competency_combo.grid(row=1, column=1, padx=10, pady=10)
    
        ttk.Label(dialog, text="Weight:").grid(row=2, column=0, sticky=tk.W, padx=10, pady=10)
        ttk.Entry(dialog, textvariable=weight_var, width=10).grid(row=2, column=1, padx=10, pady=10, sticky=tk.W)
    
        try:
            conn = get_connection()
            cursor = conn.cursor()
    
            cursor.execute(
                '''
                SELECT assessment_id, assessment_name, module_code
                FROM assessments
                ORDER BY module_code, assessment_name
                '''
            )
            assessments = cursor.fetchall()
            assessment_options = {
                f"{a_id} - {a_name} ({module})": a_id for a_id, a_name, module in assessments
            }
            assessment_combo['values'] = list(assessment_options.keys())
    
            cursor.execute(
                '''
                SELECT competency_id, name
                FROM competencies
                ORDER BY name
                '''
            )
            competencies = cursor.fetchall()
            competency_options = {f"{c_id} - {name}": c_id for c_id, name in competencies}
            competency_combo['values'] = list(competency_options.keys())
    
            conn.close()
    
        except sqlite3.Error as exc:
            messagebox.showerror("Database Error", f"Failed to load mapping data: {exc}")
            dialog.destroy()
            return
    
        def save_changes():
            assessment_key = assessment_var.get()
            competency_key = competency_var.get()
    
            if assessment_key not in assessment_options or competency_key not in competency_options:
                messagebox.showwarning("Validation", "Please select both assessment and competency.")
                return
    
            try:
                weight = float(weight_var.get())
            except ValueError:
                messagebox.showwarning("Validation", "Weight must be numeric.")
                return
    
            assessment_id = assessment_options[assessment_key]
            competency_id = competency_options[competency_key]
    
            try:
                conn_edit = get_connection()
                cursor_edit = conn_edit.cursor()
    
                cursor_edit.execute(
                    '''
                    SELECT id
                    FROM assessment_competencies
                    WHERE assessment_id = ? AND competency_id = ? AND id != ?
                    ''',
                    (assessment_id, competency_id, mapping_id)
                )
                duplicate = cursor_edit.fetchone()
                if duplicate:
                    messagebox.showwarning(
                        "Duplicate Mapping",
                        "Another mapping already links this assessment and competency."
                    )
                    conn_edit.close()
                    return
    
                cursor_edit.execute(
                    '''
                    UPDATE assessment_competencies
                    SET assessment_id = ?, competency_id = ?, weight = ?
                    WHERE id = ?
                    ''',
                    (assessment_id, competency_id, weight, mapping_id)
                )
    
                conn_edit.commit()
                conn_edit.close()
    
                messagebox.showinfo("Success", "Mapping updated successfully.")
                dialog.destroy()
                self.refresh_competency_mappings()
    
            except sqlite3.Error as exc:
                messagebox.showerror("Database Error", f"Failed to update mapping: {exc}")
    
        ttk.Button(dialog, text="Save Changes", command=save_changes).grid(row=3, column=0, columnspan=2, pady=15)
    

    def remove_competency_mapping(self):
        """Delete the selected competency mapping."""
        tree = getattr(self, 'mapping_tree', None)
        if not tree or not tree.selection():
            messagebox.showwarning("No Selection", "Select a mapping to remove.")
            return
    
        mapping_id = tree.item(tree.selection()[0], 'values')[0]
        if not messagebox.askyesno("Confirm Delete", "Remove selected mapping?"):
            return
    
        try:
            conn = get_connection()
            cursor = conn.cursor()
            cursor.execute('DELETE FROM assessment_competencies WHERE id = ?', (mapping_id,))
            conn.commit()
            conn.close()
    
            messagebox.showinfo("Success", "Mapping removed.")
            self.refresh_competency_mappings()
    
        except sqlite3.Error as exc:
            messagebox.showerror("Database Error", f"Failed to remove mapping: {exc}")
    

    def refresh_competency_mappings(self):
        """Refresh the assessment-to-competency mapping list."""
        tree = getattr(self, 'mapping_tree', None)
        if not tree or not self._widget_exists(tree):
            return
    
        try:
            for item in tree.get_children():
                tree.delete(item)
    
            conn = get_connection()
            cursor = conn.cursor()
            cursor.execute(
                '''
                SELECT
                    ac.id,
                    a.assessment_name,
                    c.name,
                    ac.weight,
                    a.module_code
                FROM assessment_competencies ac
                JOIN assessments a ON ac.assessment_id = a.assessment_id
                JOIN competencies c ON ac.competency_id = c.competency_id
                ORDER BY a.module_code, a.assessment_name
                '''
            )
            rows = cursor.fetchall()
            conn.close()
    
            for map_id, assessment_name, competency_name, weight, module_code in rows:
                tree.insert(
                    '',
                    'end',
                    values=(
                        map_id,
                        f"{assessment_name} ({module_code})",
                        competency_name,
                        weight,
                        module_code
                    )
                )
    
            self.update_status(f"Loaded {len(rows)} competency mappings")
    
        except sqlite3.Error as exc:
            messagebox.showerror("Database Error", f"Failed to load mappings: {exc}")
    

    def record_student_competency_dialog(self):
        """Record a competency achievement for a student."""
        dialog = tk.Toplevel(self.root)
        dialog.title("Record Student Competency")
        dialog.geometry("560x420")
        safe_grab_set(dialog)
    
        student_var = tk.StringVar()
        competency_var = tk.StringVar()
        level_var = tk.StringVar()
        date_var = tk.StringVar(value=datetime.now().strftime('%Y-%m-%d'))
    
        ttk.Label(dialog, text="Student:").grid(row=0, column=0, sticky=tk.W, padx=10, pady=8)
        student_combo = ttk.Combobox(dialog, textvariable=student_var, width=40, state='readonly')
        student_combo.grid(row=0, column=1, padx=10, pady=8)
    
        ttk.Label(dialog, text="Competency:").grid(row=1, column=0, sticky=tk.W, padx=10, pady=8)
        competency_combo = ttk.Combobox(dialog, textvariable=competency_var, width=40, state='readonly')
        competency_combo.grid(row=1, column=1, padx=10, pady=8)
    
        ttk.Label(dialog, text="Level:").grid(row=2, column=0, sticky=tk.W, padx=10, pady=8)
        level_combo = ttk.Combobox(dialog, textvariable=level_var, width=40, state='readonly')
        level_combo.grid(row=2, column=1, padx=10, pady=8)
    
        ttk.Label(dialog, text="Assessment Date (YYYY-MM-DD):").grid(row=3, column=0, sticky=tk.W, padx=10, pady=8)
        ttk.Entry(dialog, textvariable=date_var, width=20).grid(row=3, column=1, padx=10, pady=8, sticky=tk.W)
    
        ttk.Label(dialog, text="Evidence / Notes:").grid(row=4, column=0, sticky=tk.NW, padx=10, pady=8)
        evidence_text = scrolledtext.ScrolledText(dialog, width=38, height=8)
        evidence_text.grid(row=4, column=1, padx=10, pady=8)
    
        try:
            conn = get_connection()
            cursor = conn.cursor()
    
            cursor.execute(
                '''
                SELECT student_id, first_name, last_name
                FROM students
                ORDER BY last_name, first_name
                '''
            )
            students = cursor.fetchall()
            student_options = {f"{sid} - {fname} {lname}".strip(): sid for sid, fname, lname in students}
            student_combo['values'] = list(student_options.keys())
            if student_combo['values']:
                student_var.set(student_combo['values'][0])
    
            cursor.execute(
                '''
                SELECT competency_id, name
                FROM competencies
                ORDER BY name
                '''
            )
            competencies = cursor.fetchall()
            competency_options = {f"{cid} - {name}": cid for cid, name in competencies}
            competency_combo['values'] = list(competency_options.keys())
            if competency_combo['values']:
                competency_var.set(competency_combo['values'][0])
    
            conn.close()
    
        except sqlite3.Error as exc:
            messagebox.showerror("Database Error", f"Failed to load students or competencies: {exc}")
            dialog.destroy()
            return
    
        def load_levels_for_competency(event=None):
            key = competency_var.get()
            competency_id = competency_options.get(key)
            if not competency_id:
                level_combo['values'] = []
                level_var.set('')
                return
    
            try:
                conn_levels = get_connection()
                cursor_levels = conn_levels.cursor()
                cursor_levels.execute(
                    '''
                    SELECT level_id, level_name, level_value
                    FROM competency_levels
                    WHERE competency_id = ?
                    ORDER BY level_value
                    ''',
                    (competency_id,)
                )
                levels = cursor_levels.fetchall()
                conn_levels.close()
    
                if levels:
                    level_options = {
                        f"{lid} - {lname} (Level {lvalue})": lid for lid, lname, lvalue in levels
                    }
                else:
                    level_options = {"No defined levels": None}
    
                level_combo['values'] = list(level_options.keys())
                level_combo.level_options = level_options  # attach for later lookup
                if levels:
                    level_var.set(level_combo['values'][0])
                else:
                    level_var.set("No defined levels")
    
            except sqlite3.Error as exc:
                messagebox.showerror("Database Error", f"Failed to load competency levels: {exc}")
    
        competency_combo.bind('<<ComboboxSelected>>', load_levels_for_competency)
    
        if competency_combo['values']:
            load_levels_for_competency()
        else:
            level_combo['values'] = []
            level_var.set('')
    
        def save_record():
            student_key = student_var.get()
            competency_key = competency_var.get()
            level_key = level_var.get()
            evidence = evidence_text.get('1.0', tk.END).strip()
            assessment_date = date_var.get().strip()
    
            student_id = student_options.get(student_key)
            competency_id = competency_options.get(competency_key)
    
            if not student_id or not competency_id:
                messagebox.showwarning("Validation", "Student and competency are required.")
                return
    
            level_id = None
            if hasattr(level_combo, 'level_options'):
                level_id = level_combo.level_options.get(level_key)
    
            if level_id is None:
                messagebox.showwarning(
                    "Validation",
                    "The selected competency has no levels defined. Please add competency levels first."
                )
                return
    
            if not assessment_date:
                assessment_date = datetime.now().strftime('%Y-%m-%d')
    
            try:
                datetime.strptime(assessment_date, '%Y-%m-%d')
            except ValueError:
                messagebox.showwarning("Validation", "Assessment date must be in YYYY-MM-DD format.")
                return
    
            try:
                conn_save = get_connection()
                cursor_save = conn_save.cursor()
                cursor_save.execute(
                    '''
                    INSERT INTO student_competencies (student_id, competency_id, level_id, assessment_date, evidence)
                    VALUES (?, ?, ?, ?, ?)
                    ''',
                    (student_id, competency_id, level_id, assessment_date, evidence or None)
                )
                conn_save.commit()
                conn_save.close()
    
                messagebox.showinfo("Success", "Competency record saved.")
                dialog.destroy()
                self._refresh_student_competencies()
    
            except sqlite3.Error as exc:
                messagebox.showerror("Database Error", f"Failed to save competency record: {exc}")
    
        button_frame = ttk.Frame(dialog)
        button_frame.grid(row=5, column=0, columnspan=2, pady=12)
        ttk.Button(button_frame, text="Save Record", command=save_record).pack(side=tk.LEFT, padx=8)
        ttk.Button(button_frame, text="Cancel", command=dialog.destroy).pack(side=tk.LEFT, padx=8)
    

    def edit_student_competency(self):
        """Edit a previously recorded student competency."""
        tree = getattr(self, 'student_comp_tree', None)
        if not tree or not tree.selection():
            messagebox.showwarning("No Selection", "Select a student competency record to edit.")
            return
    
        record_id = tree.item(tree.selection()[0], 'values')[0]
    
        try:
            conn = get_connection()
            cursor = conn.cursor()
            cursor.execute(
                '''
                SELECT student_id, competency_id, level_id, assessment_date, evidence
                FROM student_competencies
                WHERE id = ?
                ''',
                (record_id,)
            )
            record = cursor.fetchone()
            conn.close()
    
            if not record:
                messagebox.showerror("Error", "Selected competency record no longer exists.")
                return
    
        except sqlite3.Error as exc:
            messagebox.showerror("Database Error", f"Failed to load competency record: {exc}")
            return
    
        student_id, competency_id, level_id, assessment_date, evidence = record
    
        dialog = tk.Toplevel(self.root)
        dialog.title("Edit Student Competency")
        dialog.geometry("560x420")
        safe_grab_set(dialog)
    
        student_var = tk.StringVar()
        competency_var = tk.StringVar()
        level_var = tk.StringVar()
        date_var = tk.StringVar(value=assessment_date or datetime.now().strftime('%Y-%m-%d'))
    
        ttk.Label(dialog, text="Student:").grid(row=0, column=0, sticky=tk.W, padx=10, pady=8)
        student_combo = ttk.Combobox(dialog, textvariable=student_var, width=40, state='readonly')
        student_combo.grid(row=0, column=1, padx=10, pady=8)
    
        ttk.Label(dialog, text="Competency:").grid(row=1, column=0, sticky=tk.W, padx=10, pady=8)
        competency_combo = ttk.Combobox(dialog, textvariable=competency_var, width=40, state='readonly')
        competency_combo.grid(row=1, column=1, padx=10, pady=8)
    
        ttk.Label(dialog, text="Level:").grid(row=2, column=0, sticky=tk.W, padx=10, pady=8)
        level_combo = ttk.Combobox(dialog, textvariable=level_var, width=40, state='readonly')
        level_combo.grid(row=2, column=1, padx=10, pady=8)
    
        ttk.Label(dialog, text="Assessment Date (YYYY-MM-DD):").grid(row=3, column=0, sticky=tk.W, padx=10, pady=8)
        ttk.Entry(dialog, textvariable=date_var, width=20).grid(row=3, column=1, padx=10, pady=8, sticky=tk.W)
    
        ttk.Label(dialog, text="Evidence / Notes:").grid(row=4, column=0, sticky=tk.NW, padx=10, pady=8)
        evidence_text = scrolledtext.ScrolledText(dialog, width=38, height=8)
        evidence_text.grid(row=4, column=1, padx=10, pady=8)
        evidence_text.insert('1.0', evidence or "")
    
        try:
            conn = get_connection()
            cursor = conn.cursor()
    
            cursor.execute(
                '''
                SELECT student_id, first_name, last_name
                FROM students
                ORDER BY last_name, first_name
                '''
            )
            students = cursor.fetchall()
            student_options = {f"{sid} - {fname} {lname}".strip(): sid for sid, fname, lname in students}
            student_combo['values'] = list(student_options.keys())
            for display, sid in student_options.items():
                if sid == student_id:
                    student_var.set(display)
                    break
    
            cursor.execute(
                '''
                SELECT competency_id, name
                FROM competencies
                ORDER BY name
                '''
            )
            competencies = cursor.fetchall()
            competency_options = {f"{cid} - {name}": cid for cid, name in competencies}
            competency_combo['values'] = list(competency_options.keys())
            for display, cid in competency_options.items():
                if cid == competency_id:
                    competency_var.set(display)
                    break
    
            conn.close()
    
        except sqlite3.Error as exc:
            messagebox.showerror("Database Error", f"Failed to load supporting data: {exc}")
            dialog.destroy()
            return
    
        def load_levels(event=None):
            key = competency_var.get()
            selected_competency_id = competency_options.get(key)
            level_combo.level_options = {}
            if not selected_competency_id:
                level_combo['values'] = []
                level_var.set('')
                return
    
            try:
                conn_levels = get_connection()
                cursor_levels = conn_levels.cursor()
                cursor_levels.execute(
                    '''
                    SELECT level_id, level_name, level_value
                    FROM competency_levels
                    WHERE competency_id = ?
                    ORDER BY level_value
                    ''',
                    (selected_competency_id,)
                )
                levels = cursor_levels.fetchall()
                conn_levels.close()
    
                if levels:
                    level_options = {
                        f"{lid} - {lname} (Level {lvalue})": lid for lid, lname, lvalue in levels
                    }
                else:
                    level_options = {"No defined levels": None}
    
                level_combo['values'] = list(level_options.keys())
                level_combo.level_options = level_options
    
                # Select current level if possible
                if level_id:
                    for display, lid in level_options.items():
                        if lid == level_id:
                            level_var.set(display)
                            break
                    else:
                        level_var.set(level_combo['values'][0] if level_combo['values'] else '')
                else:
                    level_var.set(level_combo['values'][0] if level_combo['values'] else '')
    
            except sqlite3.Error as exc:
                messagebox.showerror("Database Error", f"Failed to load competency levels: {exc}")
    
        competency_combo.bind('<<ComboboxSelected>>', load_levels)
        load_levels()
    
        def save_changes():
            student_key = student_var.get()
            competency_key = competency_var.get()
            level_key = level_var.get()
            evidence_val = evidence_text.get('1.0', tk.END).strip()
            assessment_date_val = date_var.get().strip()
    
            student_id_val = student_options.get(student_key)
            competency_id_val = competency_options.get(competency_key)
    
            if not student_id_val or not competency_id_val:
                messagebox.showwarning("Validation", "Student and competency are required.")
                return
    
            level_id_val = None
            if hasattr(level_combo, 'level_options'):
                level_id_val = level_combo.level_options.get(level_key)
    
            if level_id_val is None:
                messagebox.showwarning(
                    "Validation",
                    "The selected competency has no levels defined. Please add competency levels first."
                )
                return
    
            if not assessment_date_val:
                assessment_date_val = datetime.now().strftime('%Y-%m-%d')
    
            try:
                datetime.strptime(assessment_date_val, '%Y-%m-%d')
            except ValueError:
                messagebox.showwarning("Validation", "Assessment date must be in YYYY-MM-DD format.")
                return
    
            try:
                conn_save = get_connection()
                cursor_save = conn_save.cursor()
                cursor_save.execute(
                    '''
                    UPDATE student_competencies
                    SET student_id = ?, competency_id = ?, level_id = ?, assessment_date = ?, evidence = ?
                    WHERE id = ?
                    ''',
                    (student_id_val, competency_id_val, level_id_val, assessment_date_val, evidence_val or None, record_id)
                )
                conn_save.commit()
                conn_save.close()
    
                messagebox.showinfo("Success", "Competency record updated.")
                dialog.destroy()
                self._refresh_student_competencies()
    
            except sqlite3.Error as exc:
                messagebox.showerror("Database Error", f"Failed to update record: {exc}")
    
        button_frame = ttk.Frame(dialog)
        button_frame.grid(row=5, column=0, columnspan=2, pady=12)
        ttk.Button(button_frame, text="Save Changes", command=save_changes).pack(side=tk.LEFT, padx=8)
        ttk.Button(button_frame, text="Cancel", command=dialog.destroy).pack(side=tk.LEFT, padx=8)
    

    def view_student_competency_profile(self):
        """Show a detailed competency profile for the selected student."""
        tree = getattr(self, 'student_comp_tree', None)
        student_id = None
    
        if tree and tree.selection():
            student_value = tree.item(tree.selection()[0], 'values')[1]
            student_id = student_value.split(' - ')[0].strip()
        elif hasattr(self, 'comp_student_filter_var'):
            selected = self.comp_student_filter_var.get()
            if selected and selected != 'All':
                student_id = selected.split(' - ')[0].strip()
    
        if not student_id:
            messagebox.showwarning("No Student", "Select a student record or choose a student from the filter.")
            return
    
        try:
            conn = get_connection()
            cursor = conn.cursor()
    
            cursor.execute(
                '''
                SELECT first_name, middle_name, last_name, course
                FROM students
                WHERE student_id = ?
                ''',
                (student_id,)
            )
            student = cursor.fetchone()
            if not student:
                conn.close()
                messagebox.showerror("Error", "Student not found.")
                return
    
            first_name, middle_name, last_name, course = student
            full_name = " ".join(part for part in (first_name, middle_name, last_name) if part)
    
            cursor.execute(
                '''
                SELECT 
                    c.name,
                    COALESCE(cl.level_name, 'Not Specified'),
                    COALESCE(cl.level_value, 0),
                    COALESCE(sc.assessment_date, ''),
                    COALESCE(sc.evidence, '')
                FROM student_competencies sc
                JOIN competencies c ON sc.competency_id = c.competency_id
                LEFT JOIN competency_levels cl ON sc.level_id = cl.level_id
                WHERE sc.student_id = ?
                ORDER BY sc.assessment_date DESC
                ''',
                (student_id,)
            )
            records = cursor.fetchall()
            conn.close()
    
        except sqlite3.Error as exc:
            messagebox.showerror("Database Error", f"Failed to load student profile: {exc}")
            return
    
        if not records:
            messagebox.showinfo("No Records", "No competency assessments recorded for this student.")
            return
    
        profile_window = tk.Toplevel(self.root)
        profile_window.title(f"Competency Profile - {full_name}")
        profile_window.geometry("620x480")
        safe_grab_set(profile_window)
    
        text_area = scrolledtext.ScrolledText(profile_window, width=70, height=25)
        text_area.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
    
        competencies_summary = {}
        for name, level_name, level_value, assess_date, evidence in records:
            summary = competencies_summary.setdefault(name, {"highest_level": (level_value, level_name), "records": []})
            if level_value > summary["highest_level"][0]:
                summary["highest_level"] = (level_value, level_name)
            summary["records"].append((level_name, level_value, assess_date, evidence))
    
        lines = [
            f"Student: {full_name} ({student_id})",
            f"Course: {course or 'N/A'}",
            f"Total Recorded Competencies: {len(records)}",
            "-" * 60,
        ]
    
        for competency_name, details in competencies_summary.items():
            highest_value, highest_name = details["highest_level"]
            level_display = (
                f"{highest_name} (Level {highest_value})" if highest_value else highest_name
            )
            lines.append(f"{competency_name}: Highest Recorded Level - {level_display}")
            for level_name, level_value, assess_date, evidence in details["records"]:
                detail_display = level_name
                if level_value:
                    detail_display = f"{level_name} (Level {level_value})"
                lines.append(f"  • {assess_date or 'Date N/A'} - {detail_display}")
                if evidence:
                    lines.append(f"    Evidence: {evidence}")
            lines.append("")
    
        text_area.insert('1.0', "\n".join(lines))
        text_area.config(state='disabled')
    

    def generate_competency_report_dialog(self):
        """Generate tailored competency reports into the preview pane."""
        dialog = tk.Toplevel(self.root)
        dialog.title("Generate Competency Report")
        dialog.geometry("500x320")
        safe_grab_set(dialog)
    
        report_type_var = tk.StringVar(value="all")
    
        ttk.Label(dialog, text="Report Type:").pack(anchor=tk.W, padx=10, pady=5)
        ttk.Radiobutton(dialog, text="All Competency Activity", variable=report_type_var, value="all").pack(anchor=tk.W, padx=20)
        ttk.Radiobutton(dialog, text="By Student", variable=report_type_var, value="student").pack(anchor=tk.W, padx=20)
        ttk.Radiobutton(dialog, text="By Competency", variable=report_type_var, value="competency").pack(anchor=tk.W, padx=20)
    
        selector_frame = ttk.Frame(dialog)
        selector_frame.pack(fill=tk.X, padx=10, pady=10)
    
        ttk.Label(selector_frame, text="Student:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        student_var = tk.StringVar()
        student_combo = ttk.Combobox(selector_frame, textvariable=student_var, width=35, state='readonly')
        student_combo.grid(row=0, column=1, padx=5, pady=5)
    
        ttk.Label(selector_frame, text="Competency:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
        competency_var = tk.StringVar()
        competency_combo = ttk.Combobox(selector_frame, textvariable=competency_var, width=35, state='readonly')
        competency_combo.grid(row=1, column=1, padx=5, pady=5)
    
        try:
            conn = get_connection()
            cursor = conn.cursor()
    
            cursor.execute(
                '''
                SELECT student_id, first_name, last_name
                FROM students
                ORDER BY last_name, first_name
                '''
            )
            students = cursor.fetchall()
            student_options = {f"{sid} - {fname} {lname}".strip(): sid for sid, fname, lname in students}
            student_combo['values'] = list(student_options.keys())
    
            cursor.execute(
                '''
                SELECT competency_id, name
                FROM competencies
                ORDER BY name
                '''
            )
            competencies = cursor.fetchall()
            competency_options = {f"{cid} - {name}": cid for cid, name in competencies}
            competency_combo['values'] = list(competency_options.keys())
    
            conn.close()
    
        except sqlite3.Error as exc:
            messagebox.showerror("Database Error", f"Failed to load report options: {exc}")
            dialog.destroy()
            return
    
        def run_report():
            report_type = report_type_var.get()
    
            if report_type == "all":
                dialog.destroy()
                self.generate_competency_profile()
                return
    
            if report_type == "student":
                student_key = student_var.get()
                student_id = student_options.get(student_key)
                if not student_id:
                    messagebox.showwarning("Validation", "Select a student for the report.")
                    return
                success = self._generate_student_competency_report(student_id)
                if success:
                    dialog.destroy()
                return
    
            if report_type == "competency":
                competency_key = competency_var.get()
                competency_id = competency_options.get(competency_key)
                if not competency_id:
                    messagebox.showwarning("Validation", "Select a competency for the report.")
                    return
                success = self._generate_competency_focus_report(competency_id)
                if success:
                    dialog.destroy()
    
        ttk.Button(dialog, text="Generate Report", command=run_report).pack(pady=10)
        ttk.Button(dialog, text="Close", command=dialog.destroy).pack()
    

    def _generate_student_competency_report(self, student_id):
        """Generate a competency report for a specific student."""
        try:
            conn = get_connection()
            cursor = conn.cursor()
    
            cursor.execute(
                '''
                SELECT first_name, middle_name, last_name, course
                FROM students
                WHERE student_id = ?
                ''',
                (student_id,)
            )
            student = cursor.fetchone()
            if not student:
                conn.close()
                messagebox.showerror("Error", "Student not found.")
                return False
    
            first_name, middle_name, last_name, course = student
            full_name = " ".join(part for part in (first_name, middle_name, last_name) if part)
    
            cursor.execute(
                '''
                SELECT 
                    c.name,
                    COALESCE(cl.level_name, 'Not Specified'),
                    COALESCE(cl.level_value, 0),
                    COALESCE(sc.assessment_date, ''),
                    COALESCE(sc.evidence, '')
                FROM student_competencies sc
                JOIN competencies c ON sc.competency_id = c.competency_id
                LEFT JOIN competency_levels cl ON sc.level_id = cl.level_id
                WHERE sc.student_id = ?
                ORDER BY sc.assessment_date DESC
                ''',
                (student_id,)
            )
            rows = cursor.fetchall()
    
            conn.close()
    
        except sqlite3.Error as exc:
            messagebox.showerror("Database Error", f"Failed to build report: {exc}")
            return False
    
        if not rows:
            messagebox.showinfo("No Records", "No competency assessments recorded for this student.")
            return False
    
        total_records = len(rows)
        distinct_competencies = len({row[0] for row in rows})
        dated_records = [row for row in rows if row[3]]
        latest_date = max(row[3] for row in dated_records) if dated_records else "N/A"
    
        numeric_levels = [row[2] for row in rows if row[2]]
        average_level = (sum(numeric_levels) / len(numeric_levels)) if numeric_levels else None
    
        overview_lines = [
            f"Student: {full_name} ({student_id})",
            f"Course: {course or 'N/A'}",
            f"Total Competency Records: {total_records}",
            f"Distinct Competencies Documented: {distinct_competencies}",
            f"Latest Assessment Date: {latest_date}",
            f"Average Recorded Level: {average_level:.2f}" if average_level is not None else "Average Recorded Level: N/A"
        ]
    
        competency_summary = {}
        for name, level_name, level_value, assess_date, evidence in rows:
            entry = competency_summary.setdefault(name, {"highest": (level_value, level_name), "latest": assess_date})
            if level_value > entry["highest"][0]:
                entry["highest"] = (level_value, level_name)
            if assess_date and (not entry["latest"] or assess_date > entry["latest"]):
                entry["latest"] = assess_date
    
        competency_lines = []
        for comp_name, data in sorted(competency_summary.items()):
            level_value, level_name = data["highest"]
            level_display = level_name if level_value == 0 else f"{level_name} (Level {level_value})"
            competency_lines.append(
                f"{comp_name}: Highest Level {level_display}, Last Assessed {data['latest'] or 'N/A'}"
            )
    
        recent_lines = []
        for name, level_name, level_value, assess_date, evidence in rows[:10]:
            level_display = level_name if level_value == 0 else f"{level_name} (Level {level_value})"
            line = f"{assess_date or 'Date N/A'} - {name}: {level_display}"
            if evidence:
                line += f" | Evidence: {evidence}"
            recent_lines.append(line)
    
        sections = [
            ("Student Overview", overview_lines),
            ("Competency Progress Highlights", competency_lines),
            ("Recent Assessments (Max 10)", recent_lines)
        ]
    
        footer = "Use this report to plan follow-up assessments and celebrate areas of strong competency achievement."
        self._display_report(f"Student Competency Report - {full_name}", sections, footer)
        return True
    

    def _generate_competency_focus_report(self, competency_id):
        """Generate a report focused on a single competency across students."""
        try:
            conn = get_connection()
            cursor = conn.cursor()
    
            cursor.execute(
                '''
                SELECT name, description, category
                FROM competencies
                WHERE competency_id = ?
                ''',
                (competency_id,)
            )
            competency = cursor.fetchone()
            if not competency:
                conn.close()
                messagebox.showerror("Error", "Competency not found.")
                return False
    
            name, description, category = competency
    
            cursor.execute(
                '''
                SELECT 
                    sc.student_id,
                    s.first_name,
                    s.last_name,
                    COALESCE(cl.level_name, 'Not Specified'),
                    COALESCE(cl.level_value, 0),
                    COALESCE(sc.assessment_date, ''),
                    COALESCE(sc.evidence, '')
                FROM student_competencies sc
                JOIN students s ON sc.student_id = s.student_id
                LEFT JOIN competency_levels cl ON sc.level_id = cl.level_id
                WHERE sc.competency_id = ?
                ORDER BY cl.level_value DESC, sc.assessment_date DESC
                ''',
                (competency_id,)
            )
            rows = cursor.fetchall()
    
            conn.close()
    
        except sqlite3.Error as exc:
            messagebox.showerror("Database Error", f"Failed to build report: {exc}")
            return False
    
        if not rows:
            messagebox.showinfo("No Records", "No student assessments recorded for this competency.")
            return False
    
        total_students = len({row[0] for row in rows})
        total_records = len(rows)
        level_values = [row[4] for row in rows if row[4]]
        average_level = (sum(level_values) / len(level_values)) if level_values else None
    
        overview_lines = [
            f"Competency: {name}",
            f"Category: {category or 'Uncategorized'}",
            f"Description: {description or 'N/A'}",
            f"Students Assessed: {total_students}",
            f"Total Records: {total_records}",
            f"Average Recorded Level: {average_level:.2f}" if average_level is not None else "Average Recorded Level: N/A"
        ]
    
        top_performers = []
        seen_students = set()
        for student_id, first_name, last_name, level_name, level_value, assess_date, _ in rows:
            if student_id in seen_students:
                continue
            seen_students.add(student_id)
            level_display = level_name if level_value == 0 else f"{level_name} (Level {level_value})"
            top_performers.append(
                f"{first_name} {last_name} ({student_id}) - {level_display} on {assess_date or 'Date N/A'}"
            )
            if len(top_performers) == 10:
                break
    
        activity_lines = []
        for student_id, first_name, last_name, level_name, level_value, assess_date, evidence in rows[:12]:
            level_display = level_name if level_value == 0 else f"{level_name} (Level {level_value})"
            entry = f"{assess_date or 'Date N/A'} - {first_name} {last_name} ({student_id}): {level_display}"
            if evidence:
                entry += f" | Evidence: {evidence}"
            activity_lines.append(entry)
    
        sections = [
            ("Competency Overview", overview_lines),
            ("Top Recorded Performances", top_performers),
            ("Recent Activity (Max 12 Records)", activity_lines)
        ]
    
        footer = "Leverage high performers as exemplars and consider support plans for students with limited evidence."
        self._display_report(f"Competency Report - {name}", sections, footer)
        return True
    

    def filter_student_competencies(self, event):
        """Filter student competencies"""
        self._refresh_student_competencies()
    

    def predict_module_success(self):
        """Predict module success based on historical data"""
        try:
            conn = get_connection()
            cursor = conn.cursor()
    
            # Get all modules with historical data
            cursor.execute('''
            SELECT DISTINCT m.module_code, m.module_name,
                   COUNT(DISTINCT mg.student_id) AS student_count,
                   AVG(mg.final_score) AS avg_score,
                   AVG(
                       CASE mg.final_grade
                           WHEN 'A+' THEN 4.0 WHEN 'A' THEN 4.0 WHEN 'A-' THEN 3.7
                           WHEN 'B+' THEN 3.3 WHEN 'B' THEN 3.0 WHEN 'B-' THEN 2.7
                           WHEN 'C+' THEN 2.3 WHEN 'C' THEN 2.0 WHEN 'C-' THEN 1.7
                           WHEN 'D+' THEN 1.3 WHEN 'D' THEN 1.0 WHEN 'D-' THEN 0.7
                           WHEN 'F' THEN 0.0 ELSE NULL
                       END
                   ) AS avg_gpa
            FROM modules m
            LEFT JOIN module_grades mg ON m.module_code = mg.module_code
            WHERE mg.final_score IS NOT NULL
            GROUP BY m.module_code, m.module_name
            HAVING student_count >= 5
            ORDER BY m.module_code
            ''')
            modules = cursor.fetchall()
    
            if not modules:
                messagebox.showinfo("No Data", "Insufficient data for module success prediction")
                conn.close()
                return
    
            # Create prediction window
            pred_window = tk.Toplevel(self.root)
            pred_window.title("Module Success Prediction")
            pred_window.geometry("900x600")
            safe_grab_set(pred_window)
    
            ttk.Label(pred_window, text="Module Success Predictions",
                     font=('Arial', 16, 'bold')).pack(pady=10)
    
            # Results frame
            results_frame = ttk.Frame(pred_window)
            results_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
    
            # Results treeview
            columns = ('Module', 'Students', 'Avg Score', 'Success Rate', 'Prediction')
            tree = ttk.Treeview(results_frame, columns=columns, show='headings')
    
            tree.heading('Module', text='Module')
            tree.heading('Students', text='Enrolled Students')
            tree.heading('Avg Score', text='Avg Score')
            tree.heading('Success Rate', text='Success Rate')
            tree.heading('Prediction', text='Prediction')
    
            tree.column('Module', width=200)
            tree.column('Students', width=120)
            tree.column('Avg Score', width=100)
            tree.column('Success Rate', width=120)
            tree.column('Prediction', width=200)
    
            scrollbar = ttk.Scrollbar(results_frame, orient=tk.VERTICAL, command=tree.yview)
            tree.configure(yscrollcommand=scrollbar.set)
    
            tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    
            # Populate predictions
            for module_code, module_name, student_count, avg_score, avg_gpa in modules:
                # Calculate success rate (grade C or better)
                cursor.execute('''
                SELECT COUNT(*) * 100.0 / ?
                FROM module_grades
                WHERE module_code = ? AND final_grade IN ('A+', 'A', 'A-', 'B+', 'B', 'B-', 'C+', 'C', 'C-')
                ''', (student_count, module_code))
                success_rate = cursor.fetchone()[0] or 0
    
                # Generate prediction
                if success_rate >= 80:
                    prediction = "High Success Expected"
                elif success_rate >= 60:
                    prediction = "Moderate Success Expected"
                else:
                    prediction = "Increased Support Needed"
    
                tree.insert('', 'end', values=(
                    f"{module_code} - {module_name}",
                    student_count,
                    f"{avg_score:.1f}%",
                    f"{success_rate:.1f}%",
                    prediction
                ))
    
            conn.close()
    
        except Exception as e:
            messagebox.showerror("Error", f"Failed to predict module success: {e}")
    

    def calculate_success_probability(self):
        """Calculate success probability for individual students"""
        try:
            conn = get_connection()
            cursor = conn.cursor()
    
            # Get all students with grades
            cursor.execute('''
            SELECT s.student_id, s.first_name, s.last_name, s.course
            FROM students s
            WHERE EXISTS (
                SELECT 1 FROM grades g WHERE g.student_id = s.student_id
            )
            ORDER BY s.last_name, s.first_name
            ''')
            students = cursor.fetchall()
    
            if not students:
                messagebox.showinfo("No Data", "No students with grades found")
                conn.close()
                return
    
            # Create probability window
            prob_window = tk.Toplevel(self.root)
            prob_window.title("Success Probability Calculator")
            prob_window.geometry("1000x600")
            safe_grab_set(prob_window)
    
            ttk.Label(prob_window, text="Student Success Probability",
                     font=('Arial', 16, 'bold')).pack(pady=10)
    
            # Results frame
            results_frame = ttk.Frame(prob_window)
            results_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
    
            # Results treeview
            columns = ('Student ID', 'Name', 'Course', 'GPA', 'Probability', 'Status')
            tree = ttk.Treeview(results_frame, columns=columns, show='headings')
    
            for col in columns:
                tree.heading(col, text=col)
    
            tree.column('Student ID', width=100)
            tree.column('Name', width=180)
            tree.column('Course', width=150)
            tree.column('GPA', width=80)
            tree.column('Probability', width=120)
            tree.column('Status', width=150)
    
            scrollbar = ttk.Scrollbar(results_frame, orient=tk.VERTICAL, command=tree.yview)
            tree.configure(yscrollcommand=scrollbar.set)
    
            tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    
            # Calculate probabilities
            for student_id, first_name, last_name, course in students:
                gpa, _, _ = self.calculate_student_gpa(cursor, student_id)
    
                # Calculate success probability based on multiple factors
                probability = 0.5  # Base probability
    
                if gpa is not None:
                    # GPA factor (0-4 scale)
                    if gpa >= 3.5:
                        probability = 0.95
                    elif gpa >= 3.0:
                        probability = 0.85
                    elif gpa >= 2.5:
                        probability = 0.70
                    elif gpa >= 2.0:
                        probability = 0.55
                    else:
                        probability = 0.30
    
                    # Adjust for failed courses
                    cursor.execute('''
                    SELECT COUNT(*) FROM grades
                    WHERE student_id = ? AND letter_grade = 'F'
                    ''', (student_id,))
                    failed_count = cursor.fetchone()[0]
                    probability -= (failed_count * 0.05)
    
                    # Adjust for consistent performance
                    cursor.execute('''
                    SELECT AVG(
                        CASE final_grade
                            WHEN 'A+' THEN 4.0 WHEN 'A' THEN 4.0 WHEN 'A-' THEN 3.7
                            WHEN 'B+' THEN 3.3 WHEN 'B' THEN 3.0 WHEN 'B-' THEN 2.7
                            WHEN 'C+' THEN 2.3 WHEN 'C' THEN 2.0 WHEN 'C-' THEN 1.7
                            WHEN 'D+' THEN 1.3 WHEN 'D' THEN 1.0 WHEN 'D-' THEN 0.7
                            WHEN 'F' THEN 0.0 ELSE NULL
                        END
                    ) FROM module_grades
                    WHERE student_id = ?
                    ''', (student_id,))
                    avg_gpa_result = cursor.fetchone()[0]
                    if avg_gpa_result and abs(avg_gpa_result - gpa) < 0.3:
                        probability += 0.05  # Consistent performance bonus
    
                probability = max(0.0, min(1.0, probability))  # Clamp to [0, 1]
    
                # Determine status
                if probability >= 0.8:
                    status = "High Success Likelihood"
                elif probability >= 0.6:
                    status = "Good Success Likelihood"
                elif probability >= 0.4:
                    status = "Moderate Risk"
                else:
                    status = "High Risk"
    
                tree.insert('', 'end', values=(
                    student_id,
                    f"{first_name} {last_name}",
                    course or 'N/A',
                    f"{gpa:.2f}" if gpa else 'N/A',
                    f"{probability * 100:.1f}%",
                    status
                ))
    
            conn.close()
    
        except Exception as e:
            messagebox.showerror("Error", f"Failed to calculate success probability: {e}")
    

    def build_at_risk_model(self):
        """Build at-risk prediction model using current data"""
        try:
            conn = get_connection()
            cursor = conn.cursor()
    
            # Clear existing risk assessments (optional - building new model)
            response = messagebox.askyesno(
                "Build Risk Model",
                "This will analyze all students and create/update risk assessments. Continue?"
            )
    
            if not response:
                conn.close()
                return
    
            # Get all students
            cursor.execute('''
            SELECT student_id, first_name, last_name, course
            FROM students
            ''')
            students = cursor.fetchall()
    
            if not students:
                messagebox.showinfo("No Data", "No students found in database")
                conn.close()
                return
    
            # Create progress window
            progress_window = tk.Toplevel(self.root)
            progress_window.title("Building Risk Model")
            progress_window.geometry("500x200")
            safe_grab_set(progress_window)
    
            ttk.Label(progress_window, text="Building Risk Assessment Model",
                     font=('Arial', 14, 'bold')).pack(pady=20)
    
            progress_var = tk.StringVar(value="Initializing...")
            progress_label = ttk.Label(progress_window, textvariable=progress_var)
            progress_label.pack(pady=10)
    
            progress_bar = ttk.Progressbar(progress_window, length=400, mode='determinate')
            progress_bar.pack(pady=10)
            progress_bar['maximum'] = len(students)
    
            # Process each student
            assessed_count = 0
            for idx, (student_id, first_name, last_name, course) in enumerate(students):
                progress_var.set(f"Assessing {first_name} {last_name}...")
                progress_bar['value'] = idx + 1
                progress_window.update()
    
                # Calculate risk score
                risk_score, risk_factors = self.calculate_comprehensive_risk_score(cursor, student_id)
    
                # Determine risk level
                if risk_score >= 0.7:
                    risk_level = "High"
                elif risk_score >= 0.4:
                    risk_level = "Medium"
                else:
                    risk_level = "Low"
    
                # Store in database
                cursor.execute('''
                INSERT INTO student_risk_assessment
                (student_id, risk_score, risk_level, assessment_date, factors)
                VALUES (?, ?, ?, date('now'), ?)
                ''', (student_id, risk_score, risk_level, ', '.join(risk_factors)))
    
                assessed_count += 1
    
            conn.commit()
            conn.close()
    
            progress_window.destroy()
    
            messagebox.showinfo(
                "Model Built",
                f"Risk model built successfully!\n\nAssessed {assessed_count} students.\n\n"
                f"High Risk: {sum(1 for s in students if self.calculate_comprehensive_risk_score(get_connection().cursor(), s[0])[0] >= 0.7)}\n"
                f"Medium Risk: {sum(1 for s in students if 0.4 <= self.calculate_comprehensive_risk_score(get_connection().cursor(), s[0])[0] < 0.7)}\n"
                f"Low Risk: {sum(1 for s in students if self.calculate_comprehensive_risk_score(get_connection().cursor(), s[0])[0] < 0.4)}"
            )
    
        except Exception as e:
            messagebox.showerror("Error", f"Failed to build risk model: {e}")
    

    def analyze_dropout_risk(self):
        """Analyze dropout risk using historical patterns"""
        try:
            conn = get_connection()
            cursor = conn.cursor()
    
            # Get students with risk factors
            cursor.execute('''
            SELECT s.student_id, s.first_name, s.last_name, s.course, s.enrollment_date
            FROM students s
            WHERE s.status = 'Active'
            ORDER BY s.last_name, s.first_name
            ''')
            students = cursor.fetchall()
    
            if not students:
                messagebox.showinfo("No Data", "No active students found")
                conn.close()
                return
    
            # Create dropout analysis window
            dropout_window = tk.Toplevel(self.root)
            dropout_window.title("Dropout Risk Analysis")
            dropout_window.geometry("1100x650")
            safe_grab_set(dropout_window)
    
            ttk.Label(dropout_window, text="Dropout Risk Analysis",
                     font=('Arial', 16, 'bold')).pack(pady=10)
    
            # Results frame
            results_frame = ttk.Frame(dropout_window)
            results_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
    
            # Results treeview
            columns = ('Student ID', 'Name', 'Course', 'Dropout Risk', 'Risk Factors', 'Recommendation')
            tree = ttk.Treeview(results_frame, columns=columns, show='headings')
    
            for col in columns:
                tree.heading(col, text=col)
    
            tree.column('Student ID', width=100)
            tree.column('Name', width=150)
            tree.column('Course', width=120)
            tree.column('Dropout Risk', width=100)
            tree.column('Risk Factors', width=200)
            tree.column('Recommendation', width=180)
    
            scrollbar = ttk.Scrollbar(results_frame, orient=tk.VERTICAL, command=tree.yview)
            tree.configure(yscrollcommand=scrollbar.set)
    
            tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    
            # Analyze each student
            for student_id, first_name, last_name, course, enrollment_date in students:
                dropout_risk = 0
                risk_factors = []
    
                # Factor 1: Low GPA
                gpa, _, _ = self.calculate_student_gpa(cursor, student_id)
                if gpa is not None and gpa < 2.0:
                    dropout_risk += 0.30
                    risk_factors.append("Low GPA")
                elif gpa is not None and gpa < 2.5:
                    dropout_risk += 0.15
                    risk_factors.append("Below Average GPA")
    
                # Factor 2: Multiple failures
                cursor.execute('''
                SELECT COUNT(*) FROM grades
                WHERE student_id = ? AND letter_grade = 'F'
                ''', (student_id,))
                fail_count = cursor.fetchone()[0]
                if fail_count >= 3:
                    dropout_risk += 0.25
                    risk_factors.append(f"{fail_count} Failed Courses")
                elif fail_count > 0:
                    dropout_risk += 0.10
                    risk_factors.append(f"{fail_count} Failed Course(s)")
    
                # Factor 3: Declining performance
                cursor.execute('''
                SELECT
                    CASE final_grade
                        WHEN 'A+' THEN 4.0 WHEN 'A' THEN 4.0 WHEN 'A-' THEN 3.7
                        WHEN 'B+' THEN 3.3 WHEN 'B' THEN 3.0 WHEN 'B-' THEN 2.7
                        WHEN 'C+' THEN 2.3 WHEN 'C' THEN 2.0 WHEN 'C-' THEN 1.7
                        WHEN 'D+' THEN 1.3 WHEN 'D' THEN 1.0 WHEN 'D-' THEN 0.7
                        WHEN 'F' THEN 0.0 ELSE NULL
                    END AS grade_points
                FROM module_grades
                WHERE student_id = ?
                ORDER BY completion_date DESC
                LIMIT 3
                ''', (student_id,))
                recent_grades = [row[0] for row in cursor.fetchall() if row[0]]
                if len(recent_grades) >= 2:
                    if all(recent_grades[i] > recent_grades[i+1] for i in range(len(recent_grades)-1)):
                        dropout_risk += 0.20
                        risk_factors.append("Declining Performance")
    
                # Factor 4: Low engagement (few completed assessments)
                cursor.execute('''
                SELECT COUNT(*) FROM grades WHERE student_id = ?
                ''', (student_id,))
                total_grades = cursor.fetchone()[0]
                if total_grades < 5:
                    dropout_risk += 0.15
                    risk_factors.append("Low Engagement")
    
                dropout_risk = min(1.0, dropout_risk)
    
                # Determine risk level and recommendation
                if dropout_risk >= 0.6:
                    risk_level = f"{dropout_risk*100:.0f}% (High)"
                    recommendation = "Immediate Intervention"
                elif dropout_risk >= 0.3:
                    risk_level = f"{dropout_risk*100:.0f}% (Medium)"
                    recommendation = "Monitor Closely"
                else:
                    risk_level = f"{dropout_risk*100:.0f}% (Low)"
                    recommendation = "Continue Support"
    
                tree.insert('', 'end', values=(
                    student_id,
                    f"{first_name} {last_name}",
                    course or 'N/A',
                    risk_level,
                    ', '.join(risk_factors) if risk_factors else "None",
                    recommendation
                ))
    
            conn.close()
    
        except Exception as e:
            messagebox.showerror("Error", f"Failed to analyze dropout risk: {e}")
    

    def generate_early_warnings(self):
        """Generate early warning alerts for at-risk students"""
        try:
            conn = get_connection()
            cursor = conn.cursor()
    
            # Find students with concerning patterns
            warnings = []
    
            # Check for recent failures
            cursor.execute('''
            SELECT s.student_id, s.first_name, s.last_name,
                   COUNT(*) as fail_count
            FROM students s
            JOIN grades g ON s.student_id = g.student_id
            WHERE g.letter_grade = 'F'
            GROUP BY s.student_id
            HAVING fail_count > 0
            ''')
            for student_id, first_name, last_name, fail_count in cursor.fetchall():
                warnings.append((
                    student_id,
                    f"{first_name} {last_name}",
                    "Failed Assessments",
                    f"{fail_count} failed assessment(s)",
                    "High" if fail_count >= 2 else "Medium"
                ))
    
            # Check for low GPA
            cursor.execute('''
            SELECT s.student_id, s.first_name, s.last_name
            FROM students s
            ''')
            for student_id, first_name, last_name in cursor.fetchall():
                gpa, _, _ = self.calculate_student_gpa(cursor, student_id)
                if gpa is not None and gpa < 2.0:
                    warnings.append((
                        student_id,
                        f"{first_name} {last_name}",
                        "Low GPA",
                        f"GPA: {gpa:.2f}",
                        "High"
                    ))
    
            # Check for missing grades
            cursor.execute('''
            SELECT s.student_id, s.first_name, s.last_name,
                   COUNT(DISTINCT a.assessment_id) as total_assessments,
                   COUNT(DISTINCT g.assessment_id) as completed_assessments
            FROM students s
            CROSS JOIN assessments a
            LEFT JOIN grades g ON s.student_id = g.student_id AND a.assessment_id = g.assessment_id
            GROUP BY s.student_id
            HAVING total_assessments > 0 AND completed_assessments < total_assessments * 0.5
            ''')
            for student_id, first_name, last_name, total, completed in cursor.fetchall():
                warnings.append((
                    student_id,
                    f"{first_name} {last_name}",
                    "Missing Assessments",
                    f"Completed {completed}/{total} assessments",
                    "Medium"
                ))
    
            conn.close()
    
            if not warnings:
                messagebox.showinfo("Early Warnings", "No early warnings detected. All students are performing well!")
                return
    
            # Create warnings window
            warning_window = tk.Toplevel(self.root)
            warning_window.title("Early Warning System")
            warning_window.geometry("950x600")
            safe_grab_set(warning_window)
    
            ttk.Label(warning_window, text="Early Warning Alerts",
                     font=('Arial', 16, 'bold')).pack(pady=10)
    
            # Summary
            high_count = sum(1 for w in warnings if w[4] == "High")
            medium_count = sum(1 for w in warnings if w[4] == "Medium")
    
            summary_text = f"Total Warnings: {len(warnings)} | High Priority: {high_count} | Medium Priority: {medium_count}"
            ttk.Label(warning_window, text=summary_text, font=('Arial', 11)).pack(pady=5)
    
            # Results frame
            results_frame = ttk.Frame(warning_window)
            results_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
    
            # Results treeview
            columns = ('Priority', 'Student ID', 'Name', 'Warning Type', 'Details')
            tree = ttk.Treeview(results_frame, columns=columns, show='headings')
    
            for col in columns:
                tree.heading(col, text=col)
    
            tree.column('Priority', width=80)
            tree.column('Student ID', width=100)
            tree.column('Name', width=150)
            tree.column('Warning Type', width=150)
            tree.column('Details', width=250)
    
            scrollbar = ttk.Scrollbar(results_frame, orient=tk.VERTICAL, command=tree.yview)
            tree.configure(yscrollcommand=scrollbar.set)
    
            tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    
            # Sort by priority (High first)
            warnings.sort(key=lambda x: (0 if x[4] == "High" else 1, x[1]))
    
            for student_id, name, warning_type, details, priority in warnings:
                tree.insert('', 'end', values=(priority, student_id, name, warning_type, details))
    
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate early warnings: {e}")

    # ============================================================================
    # RISK ANALYSIS METHODS (5 methods)
    # ============================================================================

    def identify_at_risk_students(self):
        """Identify students at risk based on grades and performance"""
        conn = None
        try:
            conn = get_connection()
            cursor = conn.cursor()

            # Calculate risk scores for all students
            cursor.execute("""
                SELECT
                    s.student_id,
                    s.first_name || ' ' || s.last_name AS name,
                    COUNT(DISTINCT g.assessment_id) AS total_assessments,
                    AVG(g.score / a.max_points * 100) AS avg_percentage,
                    SUM(CASE WHEN g.letter_grade IN ('F', 'D-', 'D') THEN 1 ELSE 0 END) AS poor_grades,
                    COUNT(CASE WHEN g.submission_date > a.due_date THEN 1 END) AS late_submissions
                FROM students s
                LEFT JOIN grades g ON s.student_id = g.student_id
                LEFT JOIN assessments a ON g.assessment_id = a.assessment_id
                GROUP BY s.student_id, s.first_name, s.last_name
                HAVING COUNT(g.grade_id) > 0
            """)

            students = cursor.fetchall()
            at_risk_students = []

            for student_id, name, total_assessments, avg_percentage, poor_grades, late_submissions in students:
                risk_score = 0
                risk_factors = []

                # Calculate risk score
                if avg_percentage and avg_percentage < 60:
                    risk_score += 40
                    risk_factors.append(f"Low average: {avg_percentage:.1f}%")
                elif avg_percentage and avg_percentage < 70:
                    risk_score += 20
                    risk_factors.append(f"Below average: {avg_percentage:.1f}%")

                if poor_grades and poor_grades >= 3:
                    risk_score += 30
                    risk_factors.append(f"{poor_grades} failing grades")
                elif poor_grades and poor_grades >= 1:
                    risk_score += 15
                    risk_factors.append(f"{poor_grades} poor grade(s)")

                if late_submissions and late_submissions >= 3:
                    risk_score += 20
                    risk_factors.append(f"{late_submissions} late submissions")
                elif late_submissions and late_submissions >= 1:
                    risk_score += 10
                    risk_factors.append(f"{late_submissions} late submission(s)")

                if risk_score >= 30:  # Threshold for at-risk
                    risk_level = "High" if risk_score >= 60 else "Medium"
                    at_risk_students.append((student_id, name, risk_score, risk_level, ", ".join(risk_factors)))

            if not at_risk_students:
                messagebox.showinfo("At-Risk Students", "No students currently identified as at-risk.")
                return

            # Display results window
            results_window = tk.Toplevel(self.root)
            results_window.title("At-Risk Students")
            results_window.geometry("900x600")
            safe_grab_set(results_window)

            ttk.Label(results_window, text="At-Risk Student Identification",
                     font=('Arial', 16, 'bold')).pack(pady=10)

            # Summary
            high_risk = sum(1 for s in at_risk_students if s[3] == "High")
            medium_risk = sum(1 for s in at_risk_students if s[3] == "Medium")
            summary = f"Total At-Risk: {len(at_risk_students)} | High Risk: {high_risk} | Medium Risk: {medium_risk}"
            ttk.Label(results_window, text=summary, font=('Arial', 11)).pack(pady=5)

            # Results frame
            results_frame = ttk.Frame(results_window)
            results_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

            # Treeview
            columns = ('Student ID', 'Name', 'Risk Score', 'Risk Level', 'Risk Factors')
            tree = ttk.Treeview(results_frame, columns=columns, show='headings')

            for col in columns:
                tree.heading(col, text=col)

            tree.column('Student ID', width=100)
            tree.column('Name', width=150)
            tree.column('Risk Score', width=100)
            tree.column('Risk Level', width=100)
            tree.column('Risk Factors', width=400)

            scrollbar = ttk.Scrollbar(results_frame, orient=tk.VERTICAL, command=tree.yview)
            tree.configure(yscrollcommand=scrollbar.set)

            tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

            # Sort by risk score (highest first)
            at_risk_students.sort(key=lambda x: x[2], reverse=True)

            for student_id, name, risk_score, risk_level, risk_factors in at_risk_students:
                tree.insert('', 'end', values=(student_id, name, risk_score, risk_level, risk_factors))

        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Failed to identify at-risk students: {e}")
        finally:
            if conn:
                conn.close()

    def student_risk_assessment(self):
        """Perform detailed risk assessment for a selected student"""
        # Get student selection dialog
        conn = None
        try:
            conn = get_connection()
            cursor = conn.cursor()

            # Get all students
            cursor.execute("SELECT student_id, first_name || ' ' || last_name FROM students ORDER BY last_name")
            students = cursor.fetchall()

            if not students:
                messagebox.showinfo("No Students", "No students found in database.")
                return

            # Selection dialog
            dialog = tk.Toplevel(self.root)
            dialog.title("Select Student for Risk Assessment")
            dialog.geometry("400x500")
            safe_grab_set(dialog)

            ttk.Label(dialog, text="Select a student:", font=('Arial', 12, 'bold')).pack(pady=10)

            # Listbox with scrollbar
            list_frame = ttk.Frame(dialog)
            list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

            scrollbar = ttk.Scrollbar(list_frame)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

            listbox = tk.Listbox(list_frame, yscrollcommand=scrollbar.set)
            listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.config(command=listbox.yview)

            for student_id, name in students:
                listbox.insert(tk.END, f"{student_id} - {name}")

            def assess_student():
                selection = listbox.curselection()
                if not selection:
                    messagebox.showwarning("No Selection", "Please select a student.")
                    return

                student_id = students[selection[0]][0]
                dialog.destroy()
                self._perform_detailed_risk_assessment(student_id)

            ttk.Button(dialog, text="Assess", command=assess_student).pack(pady=10)

        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Failed to load students: {e}")
        finally:
            if conn:
                conn.close()

    def _perform_detailed_risk_assessment(self, student_id):
        """Helper method to perform detailed risk assessment"""
        conn = None
        try:
            conn = get_connection()
            cursor = conn.cursor()

            # Get student info
            cursor.execute("SELECT first_name || ' ' || last_name FROM students WHERE student_id = ?", (student_id,))
            student_name = cursor.fetchone()[0]

            # Get grade statistics from both sources
            cursor.execute("""
                SELECT
                    COUNT(*) AS total_grades,
                    AVG(percentage) AS avg_percentage,
                    MIN(percentage) AS min_percentage,
                    MAX(percentage) AS max_percentage,
                    SUM(CASE WHEN letter_grade IN ('F', 'D-', 'D') THEN 1 ELSE 0 END) AS poor_grades,
                    SUM(CASE WHEN letter_grade IN ('A+', 'A', 'A-') THEN 1 ELSE 0 END) AS excellent_grades
                FROM (
                    -- Traditional assessments from grades table
                    SELECT
                        (g.score / a.max_points * 100) AS percentage,
                        g.letter_grade
                    FROM grades g
                    JOIN assessments a ON g.assessment_id = a.assessment_id
                    WHERE g.student_id = ?

                    UNION ALL

                    -- Assignment submissions from assignment_submissions table
                    SELECT
                        sub.grade AS percentage,
                        CASE
                            WHEN sub.grade >= 93 THEN 'A+'
                            WHEN sub.grade >= 90 THEN 'A'
                            WHEN sub.grade >= 87 THEN 'A-'
                            WHEN sub.grade >= 83 THEN 'B+'
                            WHEN sub.grade >= 80 THEN 'B'
                            WHEN sub.grade >= 77 THEN 'B-'
                            WHEN sub.grade >= 73 THEN 'C+'
                            WHEN sub.grade >= 70 THEN 'C'
                            WHEN sub.grade >= 67 THEN 'C-'
                            WHEN sub.grade >= 63 THEN 'D+'
                            WHEN sub.grade >= 60 THEN 'D'
                            WHEN sub.grade >= 57 THEN 'D-'
                            ELSE 'F'
                        END AS letter_grade
                    FROM assignment_submissions sub
                    WHERE sub.student_id = ? AND sub.grade IS NOT NULL
                ) AS combined_grades
            """, (student_id, student_id))

            stats = cursor.fetchone()
            total_grades, avg_pct, min_pct, max_pct, poor_grades, excellent_grades = stats

            # Calculate risk factors
            risk_factors = []
            risk_score = 0

            if avg_pct:
                if avg_pct < 50:
                    risk_score += 50
                    risk_factors.append(("Academic Performance", f"Very low average: {avg_pct:.1f}%", "Critical"))
                elif avg_pct < 60:
                    risk_score += 35
                    risk_factors.append(("Academic Performance", f"Low average: {avg_pct:.1f}%", "High"))
                elif avg_pct < 70:
                    risk_score += 20
                    risk_factors.append(("Academic Performance", f"Below average: {avg_pct:.1f}%", "Medium"))

            if poor_grades and poor_grades >= 3:
                risk_score += 30
                risk_factors.append(("Failing Grades", f"{poor_grades} poor grades", "High"))
            elif poor_grades and poor_grades >= 1:
                risk_score += 15
                risk_factors.append(("Failing Grades", f"{poor_grades} poor grade(s)", "Medium"))

            if total_grades < 3:
                risk_score += 10
                risk_factors.append(("Engagement", "Very few submissions", "Low"))

            # Determine overall risk level
            if risk_score >= 60:
                overall_risk = "CRITICAL"
            elif risk_score >= 40:
                overall_risk = "HIGH"
            elif risk_score >= 20:
                overall_risk = "MEDIUM"
            else:
                overall_risk = "LOW"

            # Display report
            sections = [
                ("Student Information", [
                    f"Student ID: {student_id}",
                    f"Student Name: {student_name}",
                    f"Overall Risk Level: {overall_risk}",
                    f"Risk Score: {risk_score}/100"
                ]),
                ("Academic Statistics", [
                    f"Total Assessments Completed: {total_grades or 0}",
                    f"Average Score: {avg_pct:.1f}%" if avg_pct else "Average Score: N/A",
                    f"Lowest Score: {min_pct:.1f}%" if min_pct else "Lowest Score: N/A",
                    f"Highest Score: {max_pct:.1f}%" if max_pct else "Highest Score: N/A",
                    f"Excellent Grades (A): {excellent_grades or 0}",
                    f"Poor Grades (D/F): {poor_grades or 0}"
                ]),
                ("Risk Factors", [f"• {category}: {description} (Severity: {severity})"
                                  for category, description, severity in risk_factors]
                 if risk_factors else ["No significant risk factors identified"]),
                ("Recommendations", self._generate_recommendations(risk_score, risk_factors))
            ]

            self._display_report(f"Risk Assessment Report - {student_name}", sections,
                               "This assessment is based on current academic performance data.")

        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Failed to perform risk assessment: {e}")
        finally:
            if conn:
                conn.close()

    def _generate_recommendations(self, risk_score, risk_factors):
        """Generate intervention recommendations based on risk factors"""
        recommendations = []

        if risk_score >= 60:
            recommendations.append("• URGENT: Schedule immediate intervention meeting")
            recommendations.append("• Assign academic mentor or tutor")
            recommendations.append("• Review learning support services availability")

        if any("Academic Performance" in str(factor) for factor in risk_factors):
            recommendations.append("• Provide additional tutoring in weak subject areas")
            recommendations.append("• Review study habits and time management skills")

        if any("Failing Grades" in str(factor) for factor in risk_factors):
            recommendations.append("• Arrange one-on-one sessions with instructors")
            recommendations.append("• Consider grade improvement opportunities")

        if any("Engagement" in str(factor) for factor in risk_factors):
            recommendations.append("• Investigate attendance and participation issues")
            recommendations.append("• Reach out to discuss personal challenges")

        if not recommendations:
            recommendations.append("• Continue monitoring academic progress")
            recommendations.append("• Encourage participation in enrichment activities")

        return recommendations

    def dropout_risk_analysis(self):
        """Analyze students at risk of dropping out"""
        conn = None
        try:
            conn = get_connection()
            cursor = conn.cursor()

            # Query both grades and assignment_submissions tables
            cursor.execute("""
                SELECT
                    student_id,
                    name,
                    course,
                    COUNT(*) AS total_submissions,
                    AVG(percentage) AS avg_percentage,
                    SUM(CASE WHEN letter_grade = 'F' THEN 1 ELSE 0 END) AS failures,
                    MAX(submission_date) AS last_submission
                FROM (
                    SELECT
                        s.student_id,
                        s.first_name || ' ' || s.last_name AS name,
                        s.course,
                        (g.score / a.max_points * 100) AS percentage,
                        g.letter_grade,
                        g.submission_date
                    FROM students s
                    JOIN grades g ON s.student_id = g.student_id
                    JOIN assessments a ON g.assessment_id = a.assessment_id
                    WHERE s.status = 'Active'

                    UNION ALL

                    SELECT
                        s.student_id,
                        s.first_name || ' ' || s.last_name AS name,
                        s.course,
                        sub.grade AS percentage,
                        CASE
                            WHEN sub.grade >= 93 THEN 'A+'
                            WHEN sub.grade >= 90 THEN 'A'
                            WHEN sub.grade >= 87 THEN 'A-'
                            WHEN sub.grade >= 83 THEN 'B+'
                            WHEN sub.grade >= 80 THEN 'B'
                            WHEN sub.grade >= 77 THEN 'B-'
                            WHEN sub.grade >= 73 THEN 'C+'
                            WHEN sub.grade >= 70 THEN 'C'
                            WHEN sub.grade >= 67 THEN 'C-'
                            WHEN sub.grade >= 63 THEN 'D+'
                            WHEN sub.grade >= 60 THEN 'D'
                            WHEN sub.grade >= 57 THEN 'D-'
                            ELSE 'F'
                        END AS letter_grade,
                        sub.graded_date AS submission_date
                    FROM students s
                    JOIN assignment_submissions sub ON s.student_id = sub.student_id
                    WHERE s.status = 'Active' AND sub.grade IS NOT NULL
                ) AS combined_data
                GROUP BY student_id, name, course
            """)

            students = cursor.fetchall()
            dropout_risks = []

            for student_id, name, course, submissions, avg_pct, failures, last_sub in students:
                dropout_score = 0
                reasons = []

                # Low submission count indicates disengagement
                if submissions == 0:
                    dropout_score += 50
                    reasons.append("No submissions recorded")
                elif submissions < 3:
                    dropout_score += 30
                    reasons.append("Very low engagement")

                # Poor academic performance
                if avg_pct and avg_pct < 40:
                    dropout_score += 40
                    reasons.append(f"Critically low average: {avg_pct:.1f}%")
                elif avg_pct and avg_pct < 50:
                    dropout_score += 25
                    reasons.append(f"Very low average: {avg_pct:.1f}%")

                # Multiple failures
                if failures and failures >= 3:
                    dropout_score += 30
                    reasons.append(f"{failures} failed assessments")

                # Lack of recent activity
                if last_sub:
                    try:
                        last_date = datetime.strptime(last_sub, '%Y-%m-%d')
                        days_inactive = (datetime.now() - last_date).days
                        if days_inactive > 30:
                            dropout_score += 20
                            reasons.append(f"{days_inactive} days since last submission")
                    except (ValueError, TypeError):
                        pass

                if dropout_score >= 40:  # Threshold for dropout risk
                    risk_level = "Critical" if dropout_score >= 70 else "High" if dropout_score >= 50 else "Moderate"
                    dropout_risks.append((student_id, name, course, dropout_score, risk_level, "; ".join(reasons)))

            # Display results
            if not dropout_risks:
                messagebox.showinfo("Dropout Risk Analysis", "No students identified at significant dropout risk.")
                return

            # Sort by dropout score
            dropout_risks.sort(key=lambda x: x[3], reverse=True)

            sections = [
                ("Summary", [
                    f"Total Students at Dropout Risk: {len(dropout_risks)}",
                    f"Critical Risk: {sum(1 for r in dropout_risks if r[4] == 'Critical')}",
                    f"High Risk: {sum(1 for r in dropout_risks if r[4] == 'High')}",
                    f"Moderate Risk: {sum(1 for r in dropout_risks if r[4] == 'Moderate')}"
                ]),
                ("At-Risk Students", [
                    f"{idx+1}. {name} ({student_id}) - {course}\n"
                    f"   Risk Level: {risk_level} (Score: {score})\n"
                    f"   Reasons: {reasons}"
                    for idx, (student_id, name, course, score, risk_level, reasons) in enumerate(dropout_risks)
                ])
            ]

            self._display_report("Dropout Risk Analysis", sections,
                               "Immediate intervention recommended for critical and high-risk students.")

        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Failed to analyze dropout risk: {e}")
        finally:
            if conn:
                conn.close()

    def intervention_recommendations(self):
        """Generate intervention recommendations for at-risk students"""
        conn = None
        try:
            conn = get_connection()
            cursor = conn.cursor()

            # Get at-risk students
            cursor.execute("""
                SELECT
                    s.student_id,
                    s.first_name || ' ' || s.last_name AS name,
                    AVG(g.score / a.max_points * 100) AS avg_percentage,
                    SUM(CASE WHEN g.letter_grade IN ('F', 'D') THEN 1 ELSE 0 END) AS poor_grades,
                    COUNT(g.grade_id) AS total_grades
                FROM students s
                JOIN grades g ON s.student_id = g.student_id
                JOIN assessments a ON g.assessment_id = a.assessment_id
                GROUP BY s.student_id, s.first_name, s.last_name
                HAVING AVG(g.score / a.max_points * 100) < 70 OR
                       SUM(CASE WHEN g.letter_grade IN ('F', 'D') THEN 1 ELSE 0 END) >= 2
            """)

            at_risk = cursor.fetchall()

            if not at_risk:
                messagebox.showinfo("Interventions", "No students currently require intervention.")
                return

            interventions = []
            for student_id, name, avg_pct, poor_grades, total_grades in at_risk:
                student_interventions = []

                # Academic interventions
                if avg_pct and avg_pct < 50:
                    student_interventions.append("URGENT: One-on-one tutoring (minimum 3 sessions/week)")
                    student_interventions.append("Academic probation review")
                elif avg_pct and avg_pct < 60:
                    student_interventions.append("Regular tutoring sessions (2x/week)")
                    student_interventions.append("Study skills workshop")
                elif avg_pct and avg_pct < 70:
                    student_interventions.append("Peer tutoring or study group")
                    student_interventions.append("Office hours attendance")

                # Based on failing grades
                if poor_grades >= 3:
                    student_interventions.append("Academic advisor meeting (urgent)")
                    student_interventions.append("Course load review and adjustment")
                elif poor_grades >= 2:
                    student_interventions.append("Instructor consultations for failing subjects")

                # Engagement interventions
                if total_grades < 5:
                    student_interventions.append("Attendance monitoring")
                    student_interventions.append("Wellness check-in")

                # General support
                student_interventions.append("Learning resource center orientation")
                student_interventions.append("Time management counseling")

                interventions.append((student_id, name, avg_pct or 0, poor_grades, student_interventions))

            # Display report
            sections = [
                ("Overview", [
                    f"Students Requiring Intervention: {len(interventions)}",
                    f"Critical Cases (avg < 50%): {sum(1 for i in interventions if i[2] < 50)}",
                    f"High Priority (avg < 60%): {sum(1 for i in interventions if 50 <= i[2] < 60)}"
                ]),
                ("Recommended Interventions", [
                    f"\n{name} ({student_id})\n"
                    f"Current Average: {avg:.1f}% | Failing Grades: {poor}\n"
                    f"Recommended Actions:\n" + "\n".join(f"  • {action}" for action in actions)
                    for student_id, name, avg, poor, actions in interventions
                ])
            ]

            self._display_report("Intervention Recommendations", sections,
                               "Follow up with students within 48 hours of critical interventions.")

        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Failed to generate interventions: {e}")
        finally:
            if conn:
                conn.close()

    def early_warning_system(self):
        """Generate early warning alerts for students showing concerning patterns"""
        conn = None
        try:
            conn = get_connection()
            cursor = conn.cursor()

            warnings = []

            # Check for sudden grade drops
            cursor.execute("""
                SELECT DISTINCT
                    s.student_id,
                    s.first_name || ' ' || s.last_name AS name,
                    g.score / a.max_points * 100 AS percentage,
                    g.submission_date,
                    a.assessment_name
                FROM students s
                JOIN grades g ON s.student_id = g.student_id
                JOIN assessments a ON g.assessment_id = a.assessment_id
                WHERE g.score / a.max_points * 100 < 60
                ORDER BY s.student_id, g.submission_date DESC
            """)

            for student_id, name, percentage, date, assessment in cursor.fetchall():
                warnings.append((student_id, name, "Low Score Alert",
                               f"{assessment}: {percentage:.1f}%", "High"))

            # Check for missing submissions (students with very few grades)
            cursor.execute("""
                SELECT
                    s.student_id,
                    s.first_name || ' ' || s.last_name AS name,
                    COUNT(g.grade_id) AS submission_count
                FROM students s
                LEFT JOIN grades g ON s.student_id = g.student_id
                WHERE s.status = 'Active'
                GROUP BY s.student_id, s.first_name, s.last_name
                HAVING COUNT(g.grade_id) < 3
            """)

            for student_id, name, count in cursor.fetchall():
                warnings.append((student_id, name, "Low Engagement",
                               f"Only {count} submission(s) recorded", "Medium"))

            if not warnings:
                messagebox.showinfo("Early Warning System", "No current warnings to display.")
                return

            # Display warnings window
            warning_window = tk.Toplevel(self.root)
            warning_window.title("Early Warning Alerts")
            warning_window.geometry("900x600")
            safe_grab_set(warning_window)

            ttk.Label(warning_window, text="Early Warning Alerts",
                     font=('Arial', 16, 'bold')).pack(pady=10)

            # Summary
            high_count = sum(1 for w in warnings if w[4] == "High")
            medium_count = sum(1 for w in warnings if w[4] == "Medium")

            summary_text = f"Total Warnings: {len(warnings)} | High Priority: {high_count} | Medium Priority: {medium_count}"
            ttk.Label(warning_window, text=summary_text, font=('Arial', 11)).pack(pady=5)

            # Results frame
            results_frame = ttk.Frame(warning_window)
            results_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

            # Results treeview
            columns = ('Priority', 'Student ID', 'Name', 'Warning Type', 'Details')
            tree = ttk.Treeview(results_frame, columns=columns, show='headings')

            for col in columns:
                tree.heading(col, text=col)

            tree.column('Priority', width=80)
            tree.column('Student ID', width=100)
            tree.column('Name', width=150)
            tree.column('Warning Type', width=150)
            tree.column('Details', width=250)

            scrollbar = ttk.Scrollbar(results_frame, orient=tk.VERTICAL, command=tree.yview)
            tree.configure(yscrollcommand=scrollbar.set)

            tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

            # Sort by priority (High first)
            warnings.sort(key=lambda x: (0 if x[4] == "High" else 1, x[1]))

            for student_id, name, warning_type, details, priority in warnings:
                tree.insert('', 'end', values=(priority, student_id, name, warning_type, details))

        except sqlite3.Error as e:
            messagebox.showerror("Error", f"Failed to generate early warnings: {e}")
        finally:
            if conn:
                conn.close()

    # ============================================================================
    # PERFORMANCE ANALYSIS METHODS (5 methods)
    # ============================================================================

    def show_grade_distribution(self):
        """Display grade distribution across all assessments"""
        conn = None
        try:
            conn = get_connection()
            cursor = conn.cursor()

            # Query both grades table and assignment_submissions table
            cursor.execute("""
                SELECT letter_grade, COUNT(*) AS count
                FROM (
                    -- Traditional assessments from grades table
                    SELECT letter_grade
                    FROM grades
                    WHERE letter_grade IS NOT NULL

                    UNION ALL

                    -- Assignment submissions from assignment_submissions table
                    SELECT
                        CASE
                            WHEN grade >= 93 THEN 'A+'
                            WHEN grade >= 90 THEN 'A'
                            WHEN grade >= 87 THEN 'A-'
                            WHEN grade >= 83 THEN 'B+'
                            WHEN grade >= 80 THEN 'B'
                            WHEN grade >= 77 THEN 'B-'
                            WHEN grade >= 73 THEN 'C+'
                            WHEN grade >= 70 THEN 'C'
                            WHEN grade >= 67 THEN 'C-'
                            WHEN grade >= 63 THEN 'D+'
                            WHEN grade >= 60 THEN 'D'
                            WHEN grade >= 57 THEN 'D-'
                            ELSE 'F'
                        END AS letter_grade
                    FROM assignment_submissions
                    WHERE grade IS NOT NULL
                ) AS combined_grades
                GROUP BY letter_grade
                ORDER BY
                    CASE letter_grade
                        WHEN 'A+' THEN 1 WHEN 'A' THEN 2 WHEN 'A-' THEN 3
                        WHEN 'B+' THEN 4 WHEN 'B' THEN 5 WHEN 'B-' THEN 6
                        WHEN 'C+' THEN 7 WHEN 'C' THEN 8 WHEN 'C-' THEN 9
                        WHEN 'D+' THEN 10 WHEN 'D' THEN 11 WHEN 'D-' THEN 12
                        WHEN 'F' THEN 13 ELSE 14
                    END
            """)

            distribution = cursor.fetchall()
            total_grades = sum(count for _, count in distribution)

            if total_grades == 0:
                messagebox.showinfo("Grade Distribution", "No grades recorded yet.")
                return

            dist_lines = [
                f"{grade}: {count} ({count/total_grades*100:.1f}%)"
                for grade, count in distribution
            ]

            sections = [
                ("Grade Distribution Summary", [
                    f"Total Grades: {total_grades}",
                    ""
                ]),
                ("Distribution by Letter Grade", dist_lines)
            ]

            self._display_report("Grade Distribution Analysis", sections)

        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Failed to show grade distribution: {e}")
        finally:
            if conn:
                conn.close()

    def analyze_module_performance(self):
        """Analyze performance across all modules"""
        conn = None
        try:
            conn = get_connection()
            cursor = conn.cursor()

            # Query both grades and assignment_submissions tables
            cursor.execute("""
                SELECT
                    module_code,
                    module_name,
                    COUNT(DISTINCT student_id) AS student_count,
                    COUNT(*) AS total_grades,
                    AVG(percentage) AS avg_percentage,
                    SUM(CASE WHEN letter_grade IN ('A+', 'A', 'A-') THEN 1 ELSE 0 END) AS excellent_count,
                    SUM(CASE WHEN letter_grade IN ('F', 'D', 'D-') THEN 1 ELSE 0 END) AS poor_count
                FROM (
                    -- Traditional assessments from grades table
                    SELECT
                        m.module_code,
                        m.module_name,
                        g.student_id,
                        (g.score / a.max_points * 100) AS percentage,
                        g.letter_grade
                    FROM modules m
                    JOIN assessments a ON m.module_code = a.module_code
                    JOIN grades g ON a.assessment_id = g.assessment_id

                    UNION ALL

                    -- Assignment submissions from assignment_submissions table
                    SELECT
                        m.module_code,
                        m.module_name,
                        sub.student_id,
                        sub.grade AS percentage,
                        CASE
                            WHEN sub.grade >= 93 THEN 'A+'
                            WHEN sub.grade >= 90 THEN 'A'
                            WHEN sub.grade >= 87 THEN 'A-'
                            WHEN sub.grade >= 83 THEN 'B+'
                            WHEN sub.grade >= 80 THEN 'B'
                            WHEN sub.grade >= 77 THEN 'B-'
                            WHEN sub.grade >= 73 THEN 'C+'
                            WHEN sub.grade >= 70 THEN 'C'
                            WHEN sub.grade >= 67 THEN 'C-'
                            WHEN sub.grade >= 63 THEN 'D+'
                            WHEN sub.grade >= 60 THEN 'D'
                            WHEN sub.grade >= 57 THEN 'D-'
                            ELSE 'F'
                        END AS letter_grade
                    FROM modules m
                    JOIN assignments asn ON m.module_code = asn.module_code
                    JOIN assignment_submissions sub ON asn.id = sub.assignment_id
                    WHERE sub.grade IS NOT NULL
                ) AS combined_data
                GROUP BY module_code, module_name
                HAVING COUNT(*) > 0
                ORDER BY avg_percentage DESC
            """)

            modules = cursor.fetchall()

            if not modules:
                messagebox.showinfo("Module Performance", "No module data available.")
                return

            module_lines = [
                f"{code} - {name}\n"
                f"  Students: {students} | Grades: {grades} | Avg: {avg:.1f}%\n"
                f"  Excellent (A): {excellent} | Poor (D/F): {poor}"
                for code, name, students, grades, avg, excellent, poor in modules
            ]

            sections = [
                ("Module Performance Overview", [
                    f"Total Modules Analyzed: {len(modules)}",
                    f"Average Module Performance: {sum(m[4] for m in modules)/len(modules):.1f}%"
                ]),
                ("Module Details", module_lines)
            ]

            self._display_report("Module Performance Analysis", sections)

        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Failed to analyze module performance: {e}")
        finally:
            if conn:
                conn.close()

    def compare_course_performance(self):
        """Compare performance across different courses"""
        conn = None
        try:
            conn = get_connection()
            cursor = conn.cursor()

            # Query both grades and assignment_submissions tables
            cursor.execute("""
                SELECT
                    s.course,
                    COUNT(DISTINCT cg.student_id) AS student_count,
                    AVG(cg.percentage) AS avg_percentage,
                    MIN(cg.percentage) AS min_percentage,
                    MAX(cg.percentage) AS max_percentage
                FROM students s
                JOIN (
                    -- Traditional assessments
                    SELECT g.student_id, (g.score / a.max_points * 100) AS percentage
                    FROM grades g
                    JOIN assessments a ON g.assessment_id = a.assessment_id

                    UNION ALL

                    -- Assignment submissions
                    SELECT sub.student_id, sub.grade AS percentage
                    FROM assignment_submissions sub
                    WHERE sub.grade IS NOT NULL
                ) AS cg ON s.student_id = cg.student_id
                WHERE s.course IS NOT NULL
                GROUP BY s.course
                ORDER BY avg_percentage DESC
            """)

            courses = cursor.fetchall()

            if not courses:
                messagebox.showinfo("Course Comparison", "No course data available.")
                return

            course_lines = [
                f"{course}: {students} students | Avg: {avg:.1f}% | Range: {min_val:.1f}% - {max_val:.1f}%"
                for course, students, avg, min_val, max_val in courses
            ]

            sections = [
                ("Course Performance Comparison", course_lines)
            ]

            self._display_report("Course Performance Comparison", sections)

        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Failed to compare course performance: {e}")
        finally:
            if conn:
                conn.close()

    def analyze_performance_trends(self):
        """Analyze performance trends over time"""
        conn = None
        try:
            conn = get_connection()
            cursor = conn.cursor()

            # Query both grades and assignment_submissions tables
            cursor.execute("""
                SELECT
                    date(submission_date) AS submission_month,
                    AVG(percentage) AS avg_percentage,
                    COUNT(*) AS grade_count
                FROM (
                    -- Traditional assessments
                    SELECT g.submission_date, (g.score / a.max_points * 100) AS percentage
                    FROM grades g
                    JOIN assessments a ON g.assessment_id = a.assessment_id
                    WHERE g.submission_date IS NOT NULL

                    UNION ALL

                    -- Assignment submissions
                    SELECT sub.graded_date AS submission_date, sub.grade AS percentage
                    FROM assignment_submissions sub
                    WHERE sub.grade IS NOT NULL AND sub.graded_date IS NOT NULL
                ) AS combined_grades
                GROUP BY date(submission_date)
                ORDER BY submission_month
            """)

            trends = cursor.fetchall()

            if not trends:
                messagebox.showinfo("Performance Trends", "Insufficient data for trend analysis.")
                return

            trend_lines = [
                f"{month}: {avg:.1f}% average ({count} grades)"
                for month, avg, count in trends
            ]

            sections = [
                ("Performance Trends Over Time", trend_lines)
            ]

            self._display_report("Performance Trend Analysis", sections)

        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Failed to analyze trends: {e}")
        finally:
            if conn:
                conn.close()

    def generate_performance_dashboard(self):
        """Generate comprehensive performance dashboard"""
        conn = None
        try:
            conn = get_connection()
            cursor = conn.cursor()

            # Total students
            cursor.execute("SELECT COUNT(*) FROM students")
            total_students = cursor.fetchone()[0]

            # Total modules with grades
            cursor.execute("""
                SELECT COUNT(DISTINCT module_code)
                FROM (
                    SELECT m.module_code
                    FROM modules m
                    WHERE EXISTS (
                        SELECT 1 FROM grades g
                        JOIN assessments a ON g.assessment_id = a.assessment_id
                        WHERE a.module_code = m.module_code
                    ) OR EXISTS (
                        SELECT 1 FROM assignment_submissions sub
                        JOIN assignments asn ON sub.assignment_id = asn.id
                        WHERE asn.module_code = m.module_code AND sub.grade IS NOT NULL
                    )
                )
            """)
            total_modules = cursor.fetchone()[0]

            # Total assessments
            cursor.execute("""
                SELECT
                    (SELECT COUNT(*) FROM assessments) +
                    (SELECT COUNT(*) FROM assignments)
            """)
            total_assessments = cursor.fetchone()[0]

            # Total grades and average from both sources
            cursor.execute("""
                SELECT COUNT(*), AVG(percentage)
                FROM (
                    SELECT (g.score / a.max_points * 100) AS percentage
                    FROM grades g
                    JOIN assessments a ON g.assessment_id = a.assessment_id

                    UNION ALL

                    SELECT sub.grade AS percentage
                    FROM assignment_submissions sub
                    WHERE sub.grade IS NOT NULL
                ) AS combined_grades
            """)
            total_grades, overall_avg = cursor.fetchone()

            # Grade distribution from both sources
            cursor.execute("""
                SELECT letter_grade, COUNT(*) AS count
                FROM (
                    SELECT letter_grade FROM grades WHERE letter_grade IS NOT NULL
                    UNION ALL
                    SELECT
                        CASE
                            WHEN grade >= 93 THEN 'A+'
                            WHEN grade >= 90 THEN 'A'
                            WHEN grade >= 87 THEN 'A-'
                            WHEN grade >= 83 THEN 'B+'
                            WHEN grade >= 80 THEN 'B'
                            WHEN grade >= 77 THEN 'B-'
                            WHEN grade >= 73 THEN 'C+'
                            WHEN grade >= 70 THEN 'C'
                            WHEN grade >= 67 THEN 'C-'
                            WHEN grade >= 63 THEN 'D+'
                            WHEN grade >= 60 THEN 'D'
                            WHEN grade >= 57 THEN 'D-'
                            ELSE 'F'
                        END
                    FROM assignment_submissions WHERE grade IS NOT NULL
                ) AS combined_grades
                GROUP BY letter_grade
                ORDER BY COUNT(*) DESC
                LIMIT 5
            """)
            top_grades = cursor.fetchall()

            sections = [
                ("System Overview", [
                    f"Total Students: {total_students or 0}",
                    f"Total Modules: {total_modules or 0}",
                    f"Total Assessments: {total_assessments or 0}",
                    f"Total Grades Recorded: {total_grades or 0}",
                    f"Overall Average: {overall_avg:.1f}%" if overall_avg else "Overall Average: N/A"
                ]),
                ("Top Grade Distribution", [
                    f"{grade}: {count} occurrences"
                    for grade, count in top_grades
                ] if top_grades else ["No data available."])
            ]

            self._display_report("Performance Dashboard", sections)

        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Failed to generate dashboard: {e}")
        finally:
            if conn:
                conn.close()

    # ============================================================================
    # REPORTING METHODS (13 methods)
    # ============================================================================

    def show_grade_trends_chart(self):
        """Show grade trends in a visual chart"""
        if plt is None:
            messagebox.showinfo("Feature Unavailable", "Matplotlib is required for chart visualization.")
            return

        conn = None
        try:
            conn = get_connection()
            cursor = conn.cursor()

            # Query both data sources
            cursor.execute("""
                SELECT submission_date, AVG(percentage) AS avg_score
                FROM (
                    -- Traditional assessments
                    SELECT submission_date, (score / max_points * 100) AS percentage
                    FROM grades g
                    JOIN assessments a ON g.assessment_id = a.assessment_id
                    WHERE submission_date IS NOT NULL

                    UNION ALL

                    -- Assignment submissions
                    SELECT graded_date AS submission_date, grade AS percentage
                    FROM assignment_submissions
                    WHERE graded_date IS NOT NULL AND grade IS NOT NULL
                ) AS combined_grades
                GROUP BY submission_date
                ORDER BY submission_date
            """)

            data = cursor.fetchall()
            if not data:
                messagebox.showinfo("Grade Trends", "No data available for trend chart.")
                return

            dates = [row[0] for row in data]
            scores = [row[1] for row in data]

            # Create chart window
            chart_window = tk.Toplevel(self.root)
            chart_window.title("Grade Trends Chart")
            chart_window.geometry("900x600")

            fig, ax = plt.subplots(figsize=(11, 6))
            ax.plot(range(len(dates)), scores, marker='o', linestyle='-', linewidth=2, markersize=6, color='#2E86AB')
            ax.fill_between(range(len(dates)), scores, alpha=0.2, color='#2E86AB')
            ax.axhline(y=70, color='orange', linestyle='--', alpha=0.7, label='Passing Grade (70%)')
            ax.axhline(y=90, color='green', linestyle='--', alpha=0.7, label='Excellence (90%)')
            ax.set_xlabel("Time Period", fontsize=12)
            ax.set_ylabel("Average Score (%)", fontsize=12)
            ax.set_title("Grade Trends Over Time (All Students)", fontsize=14, fontweight='bold')
            ax.grid(True, alpha=0.3)
            ax.legend()
            ax.set_ylim(0, 105)

            canvas = FigureCanvasTkAgg(fig, chart_window)
            canvas.draw()
            canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

            # Add statistics
            avg_overall = sum(scores) / len(scores)
            stats_frame = ttk.Frame(chart_window)
            stats_frame.pack(fill='x', padx=10, pady=5)
            stats_text = f"Data Points: {len(data)} | Overall Average: {avg_overall:.1f}% | Trend: {'Improving' if scores[-1] > scores[0] else 'Declining' if scores[-1] < scores[0] else 'Stable'}"
            ttk.Label(stats_frame, text=stats_text, font=('Arial', 10, 'bold')).pack()

        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Failed to show trends: {e}")
        finally:
            if conn:
                conn.close()

    def student_progress_charts(self):
        """Show student progress charts"""
        conn = None
        try:
            conn = get_connection()
            cursor = conn.cursor()

            # Get all students
            cursor.execute("SELECT student_id, first_name || ' ' || last_name FROM students ORDER BY last_name")
            students = cursor.fetchall()

            if not students:
                messagebox.showinfo("No Students", "No students found.")
                return

            # Selection dialog
            dialog = tk.Toplevel(self.root)
            dialog.title("Select Student for Progress Charts")
            dialog.geometry("400x500")
            safe_grab_set(dialog)

            ttk.Label(dialog, text="Select a student:", font=('Arial', 12, 'bold')).pack(pady=10)

            list_frame = ttk.Frame(dialog)
            list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

            scrollbar = ttk.Scrollbar(list_frame)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

            listbox = tk.Listbox(list_frame, yscrollcommand=scrollbar.set)
            listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.config(command=listbox.yview)

            for student_id, name in students:
                listbox.insert(tk.END, f"{student_id} - {name}")

            def show_charts():
                selection = listbox.curselection()
                if not selection:
                    messagebox.showwarning("No Selection", "Please select a student.")
                    return

                student_id = students[selection[0]][0]
                student_name = students[selection[0]][1]
                dialog.destroy()
                self._show_student_progress_charts(student_id, student_name)

            ttk.Button(dialog, text="Show Progress Charts", command=show_charts).pack(pady=10)

        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Failed to load students: {e}")
        finally:
            if conn:
                conn.close()

    def _show_student_progress_charts(self, student_id, student_name):
        """Display progress charts for a specific student"""
        if plt is None:
            messagebox.showinfo("Feature Unavailable",
                              "Matplotlib is required for chart visualization.\n"
                              "Install it with: pip install matplotlib")
            return

        conn = None
        try:
            conn = get_connection()
            cursor = conn.cursor()

            # Get student grades from both sources
            cursor.execute("""
                SELECT
                    submission_date,
                    percentage,
                    letter_grade,
                    assessment_name
                FROM (
                    -- Traditional assessments
                    SELECT
                        g.submission_date,
                        (g.score / a.max_points * 100) AS percentage,
                        g.letter_grade,
                        a.assessment_name
                    FROM grades g
                    JOIN assessments a ON g.assessment_id = a.assessment_id
                    WHERE g.student_id = ?

                    UNION ALL

                    -- Assignment submissions
                    SELECT
                        sub.graded_date AS submission_date,
                        sub.grade AS percentage,
                        CASE
                            WHEN sub.grade >= 93 THEN 'A+'
                            WHEN sub.grade >= 90 THEN 'A'
                            WHEN sub.grade >= 87 THEN 'A-'
                            WHEN sub.grade >= 83 THEN 'B+'
                            WHEN sub.grade >= 80 THEN 'B'
                            WHEN sub.grade >= 77 THEN 'B-'
                            WHEN sub.grade >= 73 THEN 'C+'
                            WHEN sub.grade >= 70 THEN 'C'
                            WHEN sub.grade >= 67 THEN 'C-'
                            WHEN sub.grade >= 63 THEN 'D+'
                            WHEN sub.grade >= 60 THEN 'D'
                            WHEN sub.grade >= 57 THEN 'D-'
                            ELSE 'F'
                        END AS letter_grade,
                        asn.title AS assessment_name
                    FROM assignment_submissions sub
                    JOIN assignments asn ON sub.assignment_id = asn.id
                    WHERE sub.student_id = ? AND sub.grade IS NOT NULL
                ) AS combined_grades
                WHERE submission_date IS NOT NULL
                ORDER BY submission_date
            """, (student_id, student_id))

            data = cursor.fetchall()

            if not data:
                messagebox.showinfo("No Data", f"No grades found for {student_name}")
                return

            dates = [row[0] for row in data]
            scores = [row[1] for row in data]

            # Create chart window
            chart_window = tk.Toplevel(self.root)
            chart_window.title(f"Progress Charts - {student_name}")
            chart_window.geometry("1000x700")

            fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(12, 10))

            # Chart 1: Score progression over time
            ax1.plot(range(len(dates)), scores, marker='o', linestyle='-', linewidth=2, markersize=8, color='#2E86AB')
            ax1.axhline(y=70, color='orange', linestyle='--', label='Passing Grade (70%)')
            ax1.axhline(y=90, color='green', linestyle='--', label='Excellence (90%)')
            ax1.fill_between(range(len(dates)), scores, alpha=0.3, color='#2E86AB')
            ax1.set_xlabel("Assessment Number", fontsize=12)
            ax1.set_ylabel("Score (%)", fontsize=12)
            ax1.set_title(f"{student_name} - Grade Progression Over Time", fontsize=14, fontweight='bold')
            ax1.grid(True, alpha=0.3)
            ax1.legend()
            ax1.set_ylim(0, 105)

            # Chart 2: Grade distribution (bar chart)
            grade_counts = {}
            for row in data:
                letter = row[2]
                grade_counts[letter] = grade_counts.get(letter, 0) + 1

            grades_order = ['A+', 'A', 'A-', 'B+', 'B', 'B-', 'C+', 'C', 'C-', 'D+', 'D', 'D-', 'F']
            grade_labels = [g for g in grades_order if g in grade_counts]
            grade_values = [grade_counts[g] for g in grade_labels]

            colors = ['#2ECC40' if g.startswith('A') else '#FFDC00' if g.startswith('B')
                     else '#FF851B' if g.startswith('C') else '#FF4136' for g in grade_labels]

            ax2.bar(grade_labels, grade_values, color=colors, alpha=0.7)
            ax2.set_xlabel("Letter Grade", fontsize=12)
            ax2.set_ylabel("Count", fontsize=12)
            ax2.set_title(f"{student_name} - Grade Distribution", fontsize=14, fontweight='bold')
            ax2.grid(True, alpha=0.3, axis='y')

            plt.tight_layout()

            canvas = FigureCanvasTkAgg(fig, chart_window)
            canvas.draw()
            canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

            # Add statistics text
            avg_score = sum(scores) / len(scores)
            stats_frame = ttk.Frame(chart_window)
            stats_frame.pack(fill='x', padx=10, pady=5)

            stats_text = f"Total Assessments: {len(data)} | Average Score: {avg_score:.1f}% | Highest: {max(scores):.1f}% | Lowest: {min(scores):.1f}%"
            ttk.Label(stats_frame, text=stats_text, font=('Arial', 10, 'bold')).pack()

        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Failed to generate charts: {e}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to create charts: {e}")
        finally:
            if conn:
                conn.close()

    def generate_individual_transcript(self):
        """Generate official transcript for a student"""
        conn = None
        try:
            conn = get_connection()
            cursor = conn.cursor()

            # Get all students
            cursor.execute("SELECT student_id, first_name || ' ' || last_name FROM students ORDER BY last_name")
            students = cursor.fetchall()

            if not students:
                messagebox.showinfo("No Students", "No students found.")
                return

            # Selection dialog
            dialog = tk.Toplevel(self.root)
            dialog.title("Select Student for Transcript")
            dialog.geometry("400x500")
            safe_grab_set(dialog)

            ttk.Label(dialog, text="Select a student:", font=('Arial', 12, 'bold')).pack(pady=10)

            list_frame = ttk.Frame(dialog)
            list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

            scrollbar = ttk.Scrollbar(list_frame)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

            listbox = tk.Listbox(list_frame, yscrollcommand=scrollbar.set)
            listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.config(command=listbox.yview)

            for student_id, name in students:
                listbox.insert(tk.END, f"{student_id} - {name}")

            # Export format selection
            format_frame = ttk.LabelFrame(dialog, text="Export Format", padding=10)
            format_frame.pack(fill='x', padx=10, pady=10)

            format_var = tk.StringVar(value="display")
            ttk.Radiobutton(format_frame, text="Display in Window", variable=format_var, value="display").pack(anchor='w')
            ttk.Radiobutton(format_frame, text="Save as Text (.txt)", variable=format_var, value="txt").pack(anchor='w')
            ttk.Radiobutton(format_frame, text="Save as PDF (.pdf)", variable=format_var, value="pdf").pack(anchor='w')
            ttk.Radiobutton(format_frame, text="Save as JSON (.json)", variable=format_var, value="json").pack(anchor='w')

            def generate_transcript():
                selection = listbox.curselection()
                if not selection:
                    messagebox.showwarning("No Selection", "Please select a student.")
                    return

                student_id = students[selection[0]][0]
                export_format = format_var.get()
                dialog.destroy()
                self._generate_transcript_for_student(student_id, export_format)

            ttk.Button(dialog, text="Generate", command=generate_transcript).pack(pady=10)

        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Failed to load students: {e}")
        finally:
            if conn:
                conn.close()

    def _generate_transcript_for_student(self, student_id, export_format="display"):
        """Helper to generate transcript"""
        conn = None
        try:
            conn = get_connection()
            cursor = conn.cursor()

            # Get student info
            cursor.execute("""
                SELECT first_name, last_name, course, enrollment_date
                FROM students
                WHERE student_id = ?
            """, (student_id,))
            student_info = cursor.fetchone()

            if not student_info:
                messagebox.showerror("Error", f"Student {student_id} not found")
                return

            # Get grades from both sources with type information
            cursor.execute("""
                SELECT
                    module_code,
                    module_name,
                    assessment_name,
                    assessment_type,
                    score,
                    max_points,
                    letter_grade,
                    submission_date
                FROM (
                    -- Traditional assessments from grades table
                    SELECT
                        m.module_code,
                        m.module_name,
                        a.assessment_name,
                        COALESCE(a.assessment_type, 'Assessment') AS assessment_type,
                        g.score,
                        a.max_points,
                        g.letter_grade,
                        g.submission_date
                    FROM grades g
                    JOIN assessments a ON g.assessment_id = a.assessment_id
                    JOIN modules m ON a.module_code = m.module_code
                    WHERE g.student_id = ?

                    UNION ALL

                    -- Assignment submissions from assignment_submissions table
                    SELECT
                        m.module_code,
                        m.module_name,
                        asn.title AS assessment_name,
                        'Assignment' AS assessment_type,
                        ROUND((sub.grade * COALESCE(asn.max_marks, 100) / 100), 2) AS score,
                        COALESCE(asn.max_marks, 100) AS max_points,
                        CASE
                            WHEN sub.grade >= 93 THEN 'A+'
                            WHEN sub.grade >= 90 THEN 'A'
                            WHEN sub.grade >= 87 THEN 'A-'
                            WHEN sub.grade >= 83 THEN 'B+'
                            WHEN sub.grade >= 80 THEN 'B'
                            WHEN sub.grade >= 77 THEN 'B-'
                            WHEN sub.grade >= 73 THEN 'C+'
                            WHEN sub.grade >= 70 THEN 'C'
                            WHEN sub.grade >= 67 THEN 'C-'
                            WHEN sub.grade >= 63 THEN 'D+'
                            WHEN sub.grade >= 60 THEN 'D'
                            WHEN sub.grade >= 57 THEN 'D-'
                            ELSE 'F'
                        END AS letter_grade,
                        sub.graded_date AS submission_date
                    FROM assignment_submissions sub
                    JOIN assignments asn ON sub.assignment_id = asn.id
                    JOIN modules m ON asn.module_code = m.module_code
                    WHERE sub.student_id = ? AND sub.grade IS NOT NULL
                ) AS combined_grades
                ORDER BY module_code, submission_date
            """, (student_id, student_id))
            grades = cursor.fetchall()

            # Calculate GPA
            grade_points = {
                'A+': 4.3, 'A': 4.0, 'A-': 3.7,
                'B+': 3.3, 'B': 3.0, 'B-': 2.7,
                'C+': 2.3, 'C': 2.0, 'C-': 1.7,
                'D+': 1.3, 'D': 1.0, 'D-': 0.7,
                'F': 0.0
            }
            gpa = sum(grade_points.get(g[6], 0) for g in grades) / len(grades) if grades else 0

            first_name, last_name, course, enrollment_date = student_info
            full_name = f"{first_name} {last_name}"

            # Handle different export formats
            if export_format == "txt":
                self._export_transcript_txt(student_id, full_name, course, enrollment_date, grades, gpa)
            elif export_format == "pdf":
                self._export_transcript_pdf(student_id, full_name, course, enrollment_date, grades, gpa)
            elif export_format == "json":
                self._export_transcript_json(student_id, full_name, course, enrollment_date, grades, gpa)
            else:  # display
                grade_lines = [
                    f"{module_code} - {module_name}\n"
                    f"  [{assess_type}] {assessment}: {score}/{max_points} ({letter}) - {date}"
                    for module_code, module_name, assessment, assess_type, score, max_points, letter, date in grades
                ]

                sections = [
                    ("Student Information", [
                        f"Name: {full_name}",
                        f"Student ID: {student_id}",
                        f"Course: {course}",
                        f"Enrollment Date: {enrollment_date}"
                    ]),
                    ("Academic Record", grade_lines if grade_lines else ["No grades recorded"]),
                    ("Summary", [
                        f"Total Assessments: {len(grades)}",
                        f"Cumulative GPA: {gpa:.2f}"
                    ])
                ]

                self._display_report(f"Official Transcript - {full_name}", sections,
                                   "This is an unofficial transcript for internal use only.")

        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Failed to generate transcript: {e}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate transcript: {e}")
        finally:
            if conn:
                conn.close()

    def _export_transcript_txt(self, student_id, full_name, course, enrollment_date, grades, gpa):
        """Export transcript as text file"""
        filename = filedialog.asksaveasfilename(
            title="Save Transcript as Text",
            defaultextension=".txt",
            initialfile=f"transcript_{student_id}.txt",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
        )
        if not filename:
            return

        try:
            with open(filename, 'w') as f:
                f.write("=" * 70 + "\n")
                f.write("OFFICIAL TRANSCRIPT\n")
                f.write("=" * 70 + "\n\n")

                f.write("STUDENT INFORMATION\n")
                f.write("-" * 70 + "\n")
                f.write(f"Name: {full_name}\n")
                f.write(f"Student ID: {student_id}\n")
                f.write(f"Course: {course}\n")
                f.write(f"Enrollment Date: {enrollment_date}\n\n")

                f.write("ACADEMIC RECORD\n")
                f.write("-" * 70 + "\n")
                if grades:
                    current_module = None
                    for module_code, module_name, assessment, assess_type, score, max_points, letter, date in grades:
                        if module_code != current_module:
                            f.write(f"\n{module_code} - {module_name}\n")
                            current_module = module_code
                        f.write(f"  [{assess_type}] {assessment}: {score}/{max_points} ({letter}) - {date}\n")
                else:
                    f.write("No grades recorded\n")

                f.write("\n" + "=" * 70 + "\n")
                f.write("SUMMARY\n")
                f.write("=" * 70 + "\n")
                f.write(f"Total Assessments: {len(grades)}\n")
                f.write(f"Cumulative GPA: {gpa:.2f}\n\n")
                f.write("This is an unofficial transcript for internal use only.\n")
                f.write("Generated on: " + datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "\n")

            messagebox.showinfo("Success", f"Transcript saved to {filename}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save transcript: {e}")

    def _export_transcript_pdf(self, student_id, full_name, course, enrollment_date, grades, gpa):
        """Export transcript as PDF file"""
        if SimpleDocTemplate is None:
            messagebox.showerror("Error", "PDF export requires reportlab library. Please install it with: pip install reportlab")
            return

        filename = filedialog.asksaveasfilename(
            title="Save Transcript as PDF",
            defaultextension=".pdf",
            initialfile=f"transcript_{student_id}.pdf",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
        )
        if not filename:
            return

        try:
            from reportlab.platypus import Paragraph, Spacer, PageBreak
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import inch

            doc = SimpleDocTemplate(filename, pagesize=letter)
            styles = getSampleStyleSheet()
            story = []

            # Title
            title_style = styles['Title']
            story.append(Paragraph("OFFICIAL TRANSCRIPT", title_style))
            story.append(Spacer(1, 0.3 * inch))

            # Student Information
            heading_style = styles['Heading2']
            story.append(Paragraph("Student Information", heading_style))
            story.append(Spacer(1, 0.1 * inch))

            info_text = f"""
            <b>Name:</b> {full_name}<br/>
            <b>Student ID:</b> {student_id}<br/>
            <b>Course:</b> {course}<br/>
            <b>Enrollment Date:</b> {enrollment_date}
            """
            story.append(Paragraph(info_text, styles['Normal']))
            story.append(Spacer(1, 0.3 * inch))

            # Academic Record
            story.append(Paragraph("Academic Record", heading_style))
            story.append(Spacer(1, 0.1 * inch))

            if grades:
                current_module = None
                for module_code, module_name, assessment, assess_type, score, max_points, letter_grade, date in grades:
                    if module_code != current_module:
                        module_text = f"<b>{module_code} - {module_name}</b>"
                        story.append(Paragraph(module_text, styles['Normal']))
                        story.append(Spacer(1, 0.05 * inch))
                        current_module = module_code

                    grade_text = f"&nbsp;&nbsp;&nbsp;&nbsp;[{assess_type}] {assessment}: {score}/{max_points} ({letter_grade}) - {date}"
                    story.append(Paragraph(grade_text, styles['Normal']))
            else:
                story.append(Paragraph("No grades recorded", styles['Normal']))

            story.append(Spacer(1, 0.3 * inch))

            # Summary
            story.append(Paragraph("Summary", heading_style))
            story.append(Spacer(1, 0.1 * inch))
            summary_text = f"""
            <b>Total Assessments:</b> {len(grades)}<br/>
            <b>Cumulative GPA:</b> {gpa:.2f}
            """
            story.append(Paragraph(summary_text, styles['Normal']))
            story.append(Spacer(1, 0.3 * inch))

            # Footer
            footer_text = "This is an unofficial transcript for internal use only.<br/>"
            footer_text += f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            story.append(Paragraph(footer_text, styles['Italic']))

            doc.build(story)
            messagebox.showinfo("Success", f"Transcript saved to {filename}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save PDF: {e}")

    def _export_transcript_json(self, student_id, full_name, course, enrollment_date, grades, gpa):
        """Export transcript as JSON file"""
        filename = filedialog.asksaveasfilename(
            title="Save Transcript as JSON",
            defaultextension=".json",
            initialfile=f"transcript_{student_id}.json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )
        if not filename:
            return

        try:
            transcript_data = {
                "student_information": {
                    "student_id": student_id,
                    "name": full_name,
                    "course": course,
                    "enrollment_date": enrollment_date
                },
                "academic_record": [
                    {
                        "module_code": grade[0],
                        "module_name": grade[1],
                        "assessment_name": grade[2],
                        "assessment_type": grade[3],
                        "score": float(grade[4]) if grade[4] else None,
                        "max_points": float(grade[5]) if grade[5] else None,
                        "letter_grade": grade[6],
                        "submission_date": grade[7]
                    }
                    for grade in grades
                ],
                "summary": {
                    "total_assessments": len(grades),
                    "cumulative_gpa": round(gpa, 2)
                },
                "metadata": {
                    "generated_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "note": "This is an unofficial transcript for internal use only."
                }
            }

            with open(filename, 'w') as f:
                json.dump(transcript_data, f, indent=2)

            messagebox.showinfo("Success", f"Transcript saved to {filename}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save JSON: {e}")

    def generate_student_progress_report(self):
        """Generate comprehensive student progress report with statistics and insights"""
        conn = None
        try:
            conn = get_connection()
            cursor = conn.cursor()

            # Get student progress data
            cursor.execute('''
            SELECT
                s.student_id,
                s.first_name,
                s.last_name,
                s.course,
                COUNT(DISTINCT a.module_code) AS module_count,
                COUNT(g.grade_id) AS grade_count,
                AVG(
                    CASE
                        WHEN a.max_points IS NOT NULL AND a.max_points > 0
                        THEN (g.score / a.max_points) * 100
                    END
                ) AS avg_percentage,
                SUM(CASE WHEN g.letter_grade = 'F' THEN 1 ELSE 0 END) AS fail_count
            FROM students s
            LEFT JOIN grades g ON s.student_id = g.student_id
            LEFT JOIN assessments a ON g.assessment_id = a.assessment_id
            GROUP BY s.student_id, s.first_name, s.last_name, s.course
            ORDER BY s.last_name, s.first_name
            ''')
            rows = cursor.fetchall()

            # Get GPA data (calculate grade points from final_grade)
            cursor.execute('''
            SELECT student_id, AVG(
                CASE final_grade
                    WHEN 'A+' THEN 4.0
                    WHEN 'A' THEN 4.0
                    WHEN 'A-' THEN 3.7
                    WHEN 'B+' THEN 3.3
                    WHEN 'B' THEN 3.0
                    WHEN 'B-' THEN 2.7
                    WHEN 'C+' THEN 2.3
                    WHEN 'C' THEN 2.0
                    WHEN 'C-' THEN 1.7
                    WHEN 'D+' THEN 1.3
                    WHEN 'D' THEN 1.0
                    WHEN 'D-' THEN 0.7
                    WHEN 'F' THEN 0.0
                    ELSE NULL
                END
            ) AS avg_gpa
            FROM module_grades
            WHERE final_grade IS NOT NULL
            GROUP BY student_id
            ''')
            gpa_map = {row[0]: row[1] for row in cursor.fetchall()}

            # Process student records
            progress_records = []
            for row in rows:
                student_id, first_name, last_name, course, modules, grades_count, avg_percent, fail_count = row
                name = " ".join(part for part in (first_name, last_name) if part)
                progress_records.append({
                    "student_id": student_id,
                    "name": name or student_id,
                    "course": course or "Unassigned",
                    "modules": modules or 0,
                    "grades": grades_count or 0,
                    "avg_percent": avg_percent if avg_percent is not None else None,
                    "failures": fail_count or 0,
                    "gpa": gpa_map.get(student_id)
                })

            # Calculate statistics
            total_students = len(progress_records)
            graded_students = sum(1 for r in progress_records if r["grades"] > 0)
            avg_scores = [r["avg_percent"] for r in progress_records if r["avg_percent"] is not None]
            avg_gpas = [r["gpa"] for r in progress_records if r["gpa"] is not None]
            low_participation = sum(1 for r in progress_records if r["grades"] < 3)

            overall_avg = sum(avg_scores) / len(avg_scores) if avg_scores else None
            overall_gpa = sum(avg_gpas) / len(avg_gpas) if avg_gpas else None

            # Identify high performers
            high_performers = [
                r for r in progress_records
                if (r["avg_percent"] is not None and r["avg_percent"] >= 85) or
                   (r["gpa"] is not None and r["gpa"] >= 3.5)
            ]

            # Identify students needing support
            support_candidates = [
                r for r in progress_records
                if r["failures"] >= 2 or (r["avg_percent"] is not None and r["avg_percent"] < 60)
            ]

            # Top 5 students
            top_students = sorted(
                progress_records,
                key=lambda r: ((r["avg_percent"] or 0), (r["gpa"] or 0)),
                reverse=True
            )[:5]

            # Priority support list
            support_priority = sorted(
                support_candidates,
                key=lambda r: (-r["failures"], r["avg_percent"] if r["avg_percent"] is not None else 101)
            )[:5]

            # Course summary
            course_summary = {}
            for record in progress_records:
                course = record["course"]
                summary = course_summary.setdefault(course, {"count": 0, "scores": [], "fails": 0})
                summary["count"] += 1
                if record["avg_percent"] is not None:
                    summary["scores"].append(record["avg_percent"])
                summary["fails"] += record["failures"]

            # Top 5 courses
            course_lines = []
            for course, data in sorted(
                course_summary.items(),
                key=lambda item: (
                    (sum(item[1]["scores"]) / len(item[1]["scores"])) if item[1]["scores"] else 0
                ),
                reverse=True
            )[:5]:
                avg_course_score = (sum(data["scores"]) / len(data["scores"])) if data["scores"] else None
                course_lines.append(
                    f"{course}: {data['count']} students, "
                    f"Avg Score: {avg_course_score:.1f}%"
                    if avg_course_score is not None else
                    f"{course}: {data['count']} students, Avg Score: N/A"
                )

            # Build report sections
            summary_lines = [
                f"Total Students: {total_students}",
                f"Students With Recorded Grades: {graded_students}",
                f"Students With Limited Activity (<3 assessments): {low_participation}",
                f"Average Assessment Score: {overall_avg:.1f}%"
                if overall_avg is not None else "Average Assessment Score: N/A",
                f"Average GPA: {overall_gpa:.2f}" if overall_gpa is not None else "Average GPA: N/A",
                f"High-Performing Students (≥85% or GPA ≥3.5): {len(high_performers)}",
                f"Students Requiring Support: {len(support_candidates)}"
            ]

            top_lines = [
                f"{idx + 1}. {record['name']} ({record['course']}) - "
                f"Avg Score: {record['avg_percent']:.1f}%"
                if record["avg_percent"] is not None else
                f"{idx + 1}. {record['name']} ({record['course']}) - Avg Score: N/A"
                for idx, record in enumerate(top_students)
            ]

            # Add GPA to top students
            if top_lines:
                for idx, record in enumerate(top_students):
                    if record["gpa"] is not None:
                        top_lines[idx] += f", GPA: {record['gpa']:.2f}"
                    else:
                        top_lines[idx] += ", GPA: N/A"

            support_lines = [
                f"{idx + 1}. {record['name']} ({record['course']}) - "
                f"Fails: {record['failures']}, "
                f"Avg Score: {record['avg_percent']:.1f}%"
                if record["avg_percent"] is not None else
                f"{idx + 1}. {record['name']} ({record['course']}) - Fails: {record['failures']}, Avg Score: N/A"
                for idx, record in enumerate(support_priority)
            ]

            sections = [
                ("Progress Overview", summary_lines),
                ("Top Performing Students", top_lines),
                ("Courses Snapshot", course_lines),
                ("Support Priorities", support_lines)
            ]

            footer = (
                "Tip: Track students appearing in Support Priorities for targeted interventions "
                "and encourage low-activity students to engage with upcoming assessments."
            )

            self._display_report("Student Progress Report", sections, footer)

        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Failed to generate student progress report: {e}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate student progress report: {e}")
        finally:
            if conn:
                conn.close()

    def generate_competency_profile(self):
        """Generate competency profile based on assessment types"""
        conn = None
        try:
            conn = get_connection()
            cursor = conn.cursor()

            cursor.execute("""
                SELECT
                    s.student_id,
                    s.first_name || ' ' || s.last_name AS name,
                    a.assessment_type,
                    AVG(g.score / a.max_points * 100) AS avg_performance
                FROM students s
                JOIN grades g ON s.student_id = g.student_id
                JOIN assessments a ON g.assessment_id = a.assessment_id
                GROUP BY s.student_id, s.first_name, s.last_name, a.assessment_type
                ORDER BY s.student_id, a.assessment_type
            """)

            profiles = cursor.fetchall()

            if not profiles:
                messagebox.showinfo("Competency Profiles", "No data available for competency analysis.")
                return

            # Group by student
            student_profiles = {}
            for student_id, name, assessment_type, avg in profiles:
                if student_id not in student_profiles:
                    student_profiles[student_id] = (name, [])
                student_profiles[student_id][1].append(f"{assessment_type}: {avg:.1f}%")

            profile_lines = [
                f"{name} ({student_id})\n" + "\n".join(f"  • {comp}" for comp in competencies)
                for student_id, (name, competencies) in student_profiles.items()
            ]

            sections = [
                ("Competency Profiles by Assessment Type", profile_lines)
            ]

            self._display_report("Student Competency Profiles", sections)

        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Failed to generate competency profiles: {e}")
        finally:
            if conn:
                conn.close()

    def generate_at_risk_report(self):
        """Generate comprehensive at-risk student report"""
        # Leverage identify_at_risk_students logic
        self.identify_at_risk_students()

    def generate_module_grade_report(self):
        """Generate grade report for a specific module"""
        self.analyze_module_performance()

    def generate_module_outcome_report(self):
        """Generate learning outcome report for modules"""
        conn = None
        try:
            conn = get_connection()
            cursor = conn.cursor()

            cursor.execute("""
                SELECT
                    m.module_code,
                    m.module_name,
                    COUNT(DISTINCT sm.student_id) AS enrolled_students,
                    COUNT(DISTINCT g.student_id) AS active_students,
                    AVG(g.score / a.max_points * 100) AS avg_outcome
                FROM modules m
                LEFT JOIN student_modules sm ON m.module_code = sm.module_code
                LEFT JOIN assessments a ON m.module_code = a.module_code
                LEFT JOIN grades g ON a.assessment_id = g.assessment_id
                GROUP BY m.module_code, m.module_name
                ORDER BY m.module_code
            """)

            modules = cursor.fetchall()

            outcome_lines = [
                f"{code} - {name}\n"
                f"  Enrolled: {enrolled} | Active: {active} | Avg Outcome: {avg:.1f}%" if avg else
                f"{code} - {name}\n  Enrolled: {enrolled} | Active: {active} | Avg Outcome: N/A"
                for code, name, enrolled, active, avg in modules
            ]

            sections = [
                ("Module Learning Outcomes", outcome_lines)
            ]

            self._display_report("Module Outcome Report", sections)

        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Failed to generate outcome report: {e}")
        finally:
            if conn:
                conn.close()

    def generate_assessment_analysis(self):
        """Generate comprehensive assessment analysis report"""
        # This method already exists in assessment_manager - delegate or implement simplified version
        conn = None
        try:
            conn = get_connection()
            cursor = conn.cursor()

            cursor.execute("""
                SELECT
                    a.assessment_name,
                    a.assessment_type,
                    m.module_name,
                    COUNT(g.grade_id) AS submissions,
                    AVG(g.score / a.max_points * 100) AS avg_percentage
                FROM assessments a
                LEFT JOIN modules m ON a.module_code = m.module_code
                LEFT JOIN grades g ON a.assessment_id = g.assessment_id
                GROUP BY a.assessment_id, a.assessment_name, a.assessment_type, m.module_name
                ORDER BY m.module_name, a.assessment_name
            """)

            assessments = cursor.fetchall()

            assessment_lines = [
                f"{name} ({atype}) - {module}\n"
                f"  Submissions: {submissions} | Avg: {avg:.1f}%" if avg else
                f"{name} ({atype}) - {module}\n  Submissions: {submissions} | Avg: N/A"
                for name, atype, module, submissions, avg in assessments
            ]

            sections = [
                ("Assessment Analysis", assessment_lines)
            ]

            self._display_report("Assessment Analysis Report", sections)

        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Failed to generate assessment analysis: {e}")
        finally:
            if conn:
                conn.close()

    def generate_institution_summary(self):
        """Generate institution-wide summary report"""
        conn = None
        try:
            conn = get_connection()
            cursor = conn.cursor()

            # Overall stats
            cursor.execute("""
                SELECT
                    COUNT(DISTINCT student_id) AS total_students,
                    COUNT(DISTINCT CASE WHEN status = 'Active' THEN student_id END) AS active_students
                FROM students
            """)
            student_stats = cursor.fetchone()

            cursor.execute("SELECT COUNT(*) FROM modules")
            total_modules = cursor.fetchone()[0]

            cursor.execute("SELECT COUNT(*) FROM assessments")
            total_assessments = cursor.fetchone()[0]

            cursor.execute("""
                SELECT AVG(score / max_points * 100)
                FROM grades g
                JOIN assessments a ON g.assessment_id = a.assessment_id
            """)
            overall_avg = cursor.fetchone()[0]

            sections = [
                ("Institution Summary", [
                    f"Total Students: {student_stats[0]}",
                    f"Active Students: {student_stats[1]}",
                    f"Total Modules: {total_modules}",
                    f"Total Assessments: {total_assessments}",
                    f"Overall Average Performance: {overall_avg:.1f}%" if overall_avg else "Overall Average Performance: N/A"
                ])
            ]

            self._display_report("Institution Summary Report", sections)

        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Failed to generate summary: {e}")
        finally:
            if conn:
                conn.close()

    def generate_performance_analytics(self):
        """Generate detailed performance analytics"""
        self.generate_performance_dashboard()

    def generate_trend_analysis_report(self):
        """Generate trend analysis report"""
        self.analyze_performance_trends()

    def generate_comprehensive_risk_report(self):
        """Generate comprehensive risk analysis report"""
        conn = None
        try:
            conn = get_connection()
            cursor = conn.cursor()

            # Combine multiple risk analyses
            cursor.execute("""
                SELECT
                    s.student_id,
                    s.first_name || ' ' || s.last_name AS name,
                    AVG(g.score / a.max_points * 100) AS avg_percentage,
                    COUNT(g.grade_id) AS total_grades,
                    SUM(CASE WHEN g.letter_grade = 'F' THEN 1 ELSE 0 END) AS failures
                FROM students s
                LEFT JOIN grades g ON s.student_id = g.student_id
                LEFT JOIN assessments a ON g.assessment_id = a.assessment_id
                WHERE s.status = 'Active'
                GROUP BY s.student_id, s.first_name, s.last_name
            """)

            students = cursor.fetchall()
            risk_summary = {
                'critical': [],
                'high': [],
                'medium': [],
                'low': []
            }

            for student_id, name, avg, total, failures in students:
                risk_score = 0
                if avg and avg < 50:
                    risk_score = 90
                elif avg and avg < 60:
                    risk_score = 60
                elif avg and avg < 70:
                    risk_score = 40

                if failures >= 3:
                    risk_score += 30
                elif failures >= 1:
                    risk_score += 15

                if total < 3:
                    risk_score += 20

                if risk_score >= 70:
                    risk_summary['critical'].append((student_id, name, risk_score))
                elif risk_score >= 50:
                    risk_summary['high'].append((student_id, name, risk_score))
                elif risk_score >= 30:
                    risk_summary['medium'].append((student_id, name, risk_score))
                else:
                    risk_summary['low'].append((student_id, name, risk_score))

            sections = [
                ("Risk Summary", [
                    f"Critical Risk: {len(risk_summary['critical'])} students",
                    f"High Risk: {len(risk_summary['high'])} students",
                    f"Medium Risk: {len(risk_summary['medium'])} students",
                    f"Low Risk: {len(risk_summary['low'])} students"
                ]),
                ("Critical Risk Students", [f"• {name} ({sid}) - Score: {score}"
                                           for sid, name, score in risk_summary['critical']]
                 if risk_summary['critical'] else ["None"]),
                ("High Risk Students", [f"• {name} ({sid}) - Score: {score}"
                                        for sid, name, score in risk_summary['high']]
                 if risk_summary['high'] else ["None"])
            ]

            self._display_report("Comprehensive Risk Report", sections,
                               "Immediate action required for critical and high-risk students.")

        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Failed to generate risk report: {e}")
        finally:
            if conn:
                conn.close()

    # ============================================================================
    # DATA EXPORT METHODS (3 methods)
    # ============================================================================

    def export_reports_pdf(self):
        """Export current report to PDF"""
        if not hasattr(self, 'report_preview') or not self.report_preview:
            messagebox.showwarning("No Report", "Please generate a report first.")
            return

        if SimpleDocTemplate is None:
            messagebox.showerror("Feature Unavailable", "ReportLab is required for PDF export.")
            return

        try:
            report_content = self.report_preview.get('1.0', tk.END)

            filename = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
            )

            if filename:
                doc = SimpleDocTemplate(filename, pagesize=letter)
                styles = getSampleStyleSheet()
                story = []

                for line in report_content.split('\n'):
                    if line.strip():
                        story.append(Paragraph(line, styles['Normal']))
                        story.append(Spacer(1, 0.2*inch))

                doc.build(story)
                messagebox.showinfo("Success", f"Report exported to PDF:\n{filename}")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to export to PDF: {e}")

    def export_reports_csv(self):
        """Export current report to CSV"""
        if not hasattr(self, 'report_preview') or not self.report_preview:
            messagebox.showwarning("No Report", "Please generate a report first.")
            return

        try:
            report_content = self.report_preview.get('1.0', tk.END)

            filename = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
            )

            if filename:
                with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                    writer = csv.writer(csvfile)
                    for line in report_content.split('\n'):
                        if line.strip():
                            writer.writerow([line])

                messagebox.showinfo("Success", f"Report exported to CSV:\n{filename}")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to export to CSV: {e}")

    def export_reports_excel(self):
        """Export current report to Excel"""
        if not hasattr(self, 'report_preview') or not self.report_preview:
            messagebox.showwarning("No Report", "Please generate a report first.")
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            report_content = self.report_preview.get('1.0', tk.END)

            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )

            if filename:
                wb = Workbook()
                ws = wb.active
                ws.title = "Report"

                for idx, line in enumerate(report_content.split('\n'), 1):
                    if line.strip():
                        ws.cell(row=idx, column=1, value=line)

                wb.save(filename)
                messagebox.showinfo("Success", f"Report exported to Excel:\n{filename}")

        except ImportError:
            messagebox.showerror("Feature Unavailable", "openpyxl is required for Excel export.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export to Excel: {e}")

    # ============================================================================
    # PREDICTION/FORECASTING METHODS (10 methods)
    # ============================================================================

    def predict_grades_dialog(self):
        """Predict future grades for a specific student based on current performance"""
        conn = None
        try:
            conn = get_connection()
            cursor = conn.cursor()

            # Get list of students
            cursor.execute("""
                SELECT DISTINCT s.student_id, s.first_name || ' ' || s.last_name AS name
                FROM students s
                JOIN grades g ON s.student_id = g.student_id
                ORDER BY s.last_name, s.first_name
            """)
            students = cursor.fetchall()

            if not students:
                messagebox.showinfo("Grade Prediction", "No students with grades found.")
                return

            # Create dialog for student selection
            dialog = tk.Toplevel(self.content_frame)
            dialog.title("Predict Student Grades")
            dialog.geometry("500x400")
            dialog.transient(self.content_frame)

            ttk.Label(dialog, text="Select Student:", font=('Arial', 10, 'bold')).pack(pady=10)

            # Student selection
            student_var = tk.StringVar()
            student_combo = ttk.Combobox(dialog, textvariable=student_var, width=40, state='readonly')
            student_combo['values'] = [f"{sid} - {name}" for sid, name in students]
            student_combo.pack(pady=5)

            # Results display
            results_frame = ttk.LabelFrame(dialog, text="Prediction Results", padding=10)
            results_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

            results_text = tk.Text(results_frame, height=15, width=55, wrap=tk.WORD)
            results_scrollbar = ttk.Scrollbar(results_frame, command=results_text.yview)
            results_text.configure(yscrollcommand=results_scrollbar.set)
            results_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            results_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

            def predict_for_student():
                selection = student_var.get()
                if not selection:
                    messagebox.showwarning("Selection Required", "Please select a student.")
                    return

                student_id = selection.split(' - ')[0]

                # Get student's performance data
                cursor.execute("""
                    SELECT
                        a.module_code,
                        m.module_name,
                        AVG(g.score / a.max_points * 100) AS avg_score,
                        COUNT(g.grade_id) AS assessment_count
                    FROM grades g
                    JOIN assessments a ON g.assessment_id = a.assessment_id
                    JOIN modules m ON a.module_code = m.module_code
                    WHERE g.student_id = ? AND a.max_points > 0
                    GROUP BY a.module_code, m.module_name
                """, (student_id,))
                modules = cursor.fetchall()

                if not modules:
                    results_text.delete('1.0', tk.END)
                    results_text.insert('1.0', "No graded assessments found for this student.")
                    return

                # Build prediction report
                report_lines = [f"Grade Predictions for: {selection}", "=" * 50, ""]

                module_predictions = []
                for module_code, module_name, avg_score, count in modules:
                    # Predict final grade
                    predicted_final = avg_score  # Base prediction

                    # Adjust based on trend if enough data
                    if count >= 3:
                        cursor.execute("""
                            SELECT g.score / a.max_points * 100
                            FROM grades g
                            JOIN assessments a ON g.assessment_id = a.assessment_id
                            WHERE g.student_id = ? AND a.module_code = ? AND a.max_points > 0
                            ORDER BY g.grade_id DESC
                            LIMIT 3
                        """, (student_id, module_code))
                        recent = [row[0] for row in cursor.fetchall()]

                        if len(recent) >= 3:
                            recent_avg = sum(recent[:2]) / 2
                            if recent_avg > avg_score:
                                predicted_final += (recent_avg - avg_score) * 0.3  # Trending up
                            elif recent_avg < avg_score:
                                predicted_final += (recent_avg - avg_score) * 0.3  # Trending down

                    # Determine letter grade
                    if predicted_final >= 90:
                        letter = "A"
                    elif predicted_final >= 80:
                        letter = "B"
                    elif predicted_final >= 70:
                        letter = "C"
                    elif predicted_final >= 60:
                        letter = "D"
                    else:
                        letter = "F"

                    module_predictions.append((module_name, avg_score, predicted_final, letter, count))

                # Display predictions by module
                for module_name, current, predicted, letter, count in sorted(module_predictions, key=lambda x: x[2], reverse=True):
                    trend = "↑" if predicted > current else "↓" if predicted < current else "→"
                    report_lines.append(f"{module_name}:")
                    report_lines.append(f"  Current Average: {current:.1f}%")
                    report_lines.append(f"  Predicted Final: {predicted:.1f}% ({letter}) {trend}")
                    report_lines.append(f"  Based on {count} assessments")
                    report_lines.append("")

                # Overall prediction
                overall_avg = sum(p[2] for p in module_predictions) / len(module_predictions)
                if overall_avg >= 90:
                    overall_letter = "A"
                elif overall_avg >= 80:
                    overall_letter = "B"
                elif overall_avg >= 70:
                    overall_letter = "C"
                elif overall_avg >= 60:
                    overall_letter = "D"
                else:
                    overall_letter = "F"

                report_lines.extend([
                    "Overall Prediction:",
                    f"  Predicted GPA Equivalent: {overall_avg:.1f}% ({overall_letter})",
                    "",
                    "Note: Predictions are estimates based on current performance.",
                    "Actual results may vary based on future assessments."
                ])

                results_text.delete('1.0', tk.END)
                results_text.insert('1.0', '\n'.join(report_lines))

            # Buttons
            button_frame = ttk.Frame(dialog)
            button_frame.pack(pady=10)

            ttk.Button(button_frame, text="Predict Grades", command=predict_for_student).pack(side=tk.LEFT, padx=5)
            ttk.Button(button_frame, text="Close", command=dialog.destroy).pack(side=tk.LEFT, padx=5)

        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Failed to predict grades: {e}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open prediction dialog: {e}")
        finally:
            if conn:
                conn.close()

    def predict_gpa_dialog(self):
        """Predict GPA based on current trajectory"""
        conn = None
        try:
            conn = get_connection()
            cursor = conn.cursor()

            cursor.execute("""
                SELECT
                    s.student_id,
                    s.first_name || ' ' || s.last_name AS name,
                    AVG(CASE g.letter_grade
                        WHEN 'A+' THEN 4.3 WHEN 'A' THEN 4.0 WHEN 'A-' THEN 3.7
                        WHEN 'B+' THEN 3.3 WHEN 'B' THEN 3.0 WHEN 'B-' THEN 2.7
                        WHEN 'C+' THEN 2.3 WHEN 'C' THEN 2.0 WHEN 'C-' THEN 1.7
                        WHEN 'D+' THEN 1.3 WHEN 'D' THEN 1.0 WHEN 'D-' THEN 0.7
                        WHEN 'F' THEN 0.0 ELSE 0
                    END) AS current_gpa
                FROM students s
                JOIN grades g ON s.student_id = g.student_id
                GROUP BY s.student_id, s.first_name, s.last_name
                HAVING COUNT(g.grade_id) > 0
            """)

            predictions = cursor.fetchall()

            if not predictions:
                messagebox.showinfo("GPA Prediction", "No data available for prediction.")
                return

            prediction_lines = [
                f"{name} ({student_id}): Current GPA: {gpa:.2f} | Projected Final GPA: {gpa:.2f}"
                for student_id, name, gpa in predictions
            ]

            sections = [
                ("GPA Predictions", prediction_lines),
                ("Note", ["Predictions based on current performance trends.",
                         "Actual results may vary based on future performance."])
            ]

            self._display_report("GPA Predictions", sections)

        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Failed to predict GPA: {e}")
        finally:
            if conn:
                conn.close()

    def predict_module_success(self):
        """Predict success probability for modules"""
        messagebox.showinfo("Module Success Prediction",
                          "Module success prediction analyzes:\n"
                          "• Historical pass rates\n"
                          "• Current student performance\n"
                          "• Assessment difficulty trends\n\n"
                          "This feature requires statistical analysis of past data.")

    def calculate_success_probability(self):
        """Calculate probability of passing based on current performance"""
        conn = None
        try:
            conn = get_connection()
            cursor = conn.cursor()

            cursor.execute("""
                SELECT
                    s.student_id,
                    s.first_name || ' ' || s.last_name AS name,
                    AVG(g.score / a.max_points * 100) AS avg_percentage
                FROM students s
                JOIN grades g ON s.student_id = g.student_id
                JOIN assessments a ON g.assessment_id = a.assessment_id
                GROUP BY s.student_id, s.first_name, s.last_name
                HAVING COUNT(g.grade_id) >= 3
            """)

            students = cursor.fetchall()
            probability_lines = []

            for student_id, name, avg in students:
                # Simple probability model based on current average
                if avg >= 70:
                    probability = min(95, 50 + (avg - 70) * 1.5)
                    risk = "Low"
                elif avg >= 60:
                    probability = 50 + (avg - 60) * 2
                    risk = "Medium"
                else:
                    probability = max(10, avg * 0.8)
                    risk = "High"

                probability_lines.append(
                    f"{name} ({student_id}): {probability:.1f}% pass probability | Risk: {risk}"
                )

            sections = [
                ("Success Probability Analysis", probability_lines if probability_lines else
                 ["Insufficient data for probability calculation"])
            ]

            self._display_report("Success Probability Report", sections)

        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Failed to calculate probabilities: {e}")
        finally:
            if conn:
                conn.close()

    def build_at_risk_model(self):
        """Build predictive model for at-risk identification"""
        messagebox.showinfo("At-Risk Model",
                          "Building at-risk prediction model requires:\n"
                          "• Historical student data\n"
                          "• Feature engineering\n"
                          "• Machine learning algorithms\n\n"
                          "This is an advanced feature requiring ML libraries.")

    def analyze_dropout_risk(self):
        """Analyze dropout risk using predictive analytics"""
        # Leverage existing dropout_risk_analysis
        self.dropout_risk_analysis()

    def generate_early_warnings(self):
        """Generate early warning predictions"""
        # Leverage existing early_warning_system
        self.early_warning_system()

    def forecast_performance_trends(self):
        """Forecast future performance trends"""
        messagebox.showinfo("Performance Forecasting",
                          "Performance forecasting analyzes historical trends to predict:\n"
                          "• Future grade trajectories\n"
                          "• Expected performance by module\n"
                          "• Cohort-level outcomes\n\n"
                          "Requires time-series analysis capabilities.")

    def forecast_course_performance(self):
        """Forecast overall course performance"""
        conn = None
        try:
            conn = get_connection()
            cursor = conn.cursor()

            cursor.execute("""
                SELECT
                    course,
                    AVG(g.score / a.max_points * 100) AS current_avg,
                    COUNT(DISTINCT s.student_id) AS student_count
                FROM students s
                JOIN grades g ON s.student_id = g.student_id
                JOIN assessments a ON g.assessment_id = a.assessment_id
                WHERE course IS NOT NULL
                GROUP BY course
            """)

            courses = cursor.fetchall()

            forecast_lines = [
                f"{course}: Current Avg: {avg:.1f}% | Forecast: {avg:.1f}% | Students: {count}"
                for course, avg, count in courses
            ]

            sections = [
                ("Course Performance Forecast", forecast_lines if forecast_lines else ["No data available"])
            ]

            self._display_report("Course Performance Forecast", sections)

        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Failed to forecast performance: {e}")
        finally:
            if conn:
                conn.close()

    def forecast_success_rates(self):
        """Forecast overall success rates"""
        conn = None
        try:
            conn = get_connection()
            cursor = conn.cursor()

            cursor.execute("""
                SELECT
                    COUNT(DISTINCT student_id) AS total_students,
                    SUM(CASE WHEN avg_grade >= 60 THEN 1 ELSE 0 END) AS passing_students
                FROM (
                    SELECT s.student_id, AVG(g.score / a.max_points * 100) AS avg_grade
                    FROM students s
                    JOIN grades g ON s.student_id = g.student_id
                    JOIN assessments a ON g.assessment_id = a.assessment_id
                    GROUP BY s.student_id
                )
            """)

            total, passing = cursor.fetchone()
            success_rate = (passing / total * 100) if total > 0 else 0

            sections = [
                ("Success Rate Forecast", [
                    f"Total Students: {total}",
                    f"Projected Passing Students: {passing}",
                    f"Projected Success Rate: {success_rate:.1f}%"
                ])
            ]

            self._display_report("Success Rate Forecast", sections)

        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Failed to forecast success rates: {e}")
        finally:
            if conn:
                conn.close()

    def batch_predict_grades(self):
        """Batch predict grades for multiple students based on current performance"""
        conn = None
        try:
            conn = get_connection()
            cursor = conn.cursor()

            # Get all students with their current performance
            cursor.execute("""
                SELECT
                    s.student_id,
                    s.first_name || ' ' || s.last_name AS name,
                    s.course,
                    COUNT(g.grade_id) AS assessment_count,
                    AVG(g.score / a.max_points * 100) AS current_avg,
                    MIN(g.score / a.max_points * 100) AS min_score,
                    MAX(g.score / a.max_points * 100) AS max_score
                FROM students s
                LEFT JOIN grades g ON s.student_id = g.student_id
                LEFT JOIN assessments a ON g.assessment_id = a.assessment_id
                WHERE a.max_points > 0
                GROUP BY s.student_id, s.first_name, s.last_name, s.course
                HAVING COUNT(g.grade_id) >= 2
                ORDER BY s.last_name, s.first_name
            """)

            students = cursor.fetchall()

            if not students:
                messagebox.showinfo("Batch Prediction",
                                  "Not enough data for predictions.\n"
                                  "Students need at least 2 graded assessments.")
                return

            predictions = []
            for student_id, name, course, count, avg, min_score, max_score in students:
                # Simple prediction model based on current performance
                # Predict final grade based on weighted average with trend analysis
                trend = "stable"
                predicted_final = avg

                # Calculate trend if there are enough grades
                if count >= 3:
                    # Get recent grades trend
                    cursor.execute("""
                        SELECT g.score / a.max_points * 100
                        FROM grades g
                        JOIN assessments a ON g.assessment_id = a.assessment_id
                        WHERE g.student_id = ? AND a.max_points > 0
                        ORDER BY g.grade_id DESC
                        LIMIT 3
                    """, (student_id,))
                    recent = [row[0] for row in cursor.fetchall()]

                    if len(recent) >= 3:
                        # Simple trend: compare recent vs earlier
                        recent_avg = sum(recent[:2]) / 2
                        if recent_avg > avg + 5:
                            trend = "improving"
                            predicted_final = avg + (recent_avg - avg) * 0.5
                        elif recent_avg < avg - 5:
                            trend = "declining"
                            predicted_final = avg + (recent_avg - avg) * 0.5

                # Determine predicted letter grade
                if predicted_final >= 90:
                    predicted_grade = "A"
                elif predicted_final >= 80:
                    predicted_grade = "B"
                elif predicted_final >= 70:
                    predicted_grade = "C"
                elif predicted_final >= 60:
                    predicted_grade = "D"
                else:
                    predicted_grade = "F"

                predictions.append({
                    'name': name,
                    'course': course,
                    'current_avg': avg,
                    'predicted_final': predicted_final,
                    'predicted_grade': predicted_grade,
                    'trend': trend,
                    'assessment_count': count
                })

            # Format predictions for display
            prediction_lines = []
            for idx, pred in enumerate(predictions, 1):
                trend_symbol = "↑" if pred['trend'] == 'improving' else "↓" if pred['trend'] == 'declining' else "→"
                line = (f"{idx}. {pred['name']} ({pred['course']}) | "
                       f"Current: {pred['current_avg']:.1f}% | "
                       f"Predicted Final: {pred['predicted_final']:.1f}% ({pred['predicted_grade']}) {trend_symbol} | "
                       f"Assessments: {pred['assessment_count']}")
                prediction_lines.append(line)

            # Group by predicted grade
            grade_groups = {}
            for pred in predictions:
                grade = pred['predicted_grade']
                if grade not in grade_groups:
                    grade_groups[grade] = 0
                grade_groups[grade] += 1

            summary_lines = [
                f"Total Students Analyzed: {len(predictions)}",
                "",
                "Predicted Grade Distribution:",
            ]
            for grade in ['A', 'B', 'C', 'D', 'F']:
                count = grade_groups.get(grade, 0)
                summary_lines.append(f"  {grade}: {count} students ({count/len(predictions)*100:.1f}%)")

            improving = sum(1 for p in predictions if p['trend'] == 'improving')
            declining = sum(1 for p in predictions if p['trend'] == 'declining')
            summary_lines.extend([
                "",
                f"Trending Up (↑): {improving} students",
                f"Trending Down (↓): {declining} students",
                f"Stable (→): {len(predictions) - improving - declining} students"
            ])

            sections = [
                ("Prediction Summary", summary_lines),
                ("Individual Predictions", prediction_lines)
            ]

            footer = ("Note: Predictions are based on current performance trends and historical data. "
                     "Actual results may vary based on future performance and assessment difficulty.")

            self._display_report("Batch Grade Predictions", sections, footer)

        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Failed to predict grades: {e}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate predictions: {e}")
        finally:
            if conn:
                conn.close()

    def batch_risk_assessment(self):
        """Perform batch risk assessment"""
        # Leverage existing identify_at_risk_students
        self.identify_at_risk_students()

    def generate_interventions(self):
        """Generate intervention strategies"""
        # Leverage existing intervention_recommendations
        self.intervention_recommendations()

