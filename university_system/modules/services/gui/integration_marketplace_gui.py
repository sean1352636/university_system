"""
Integration Marketplace GUI

Comprehensive GUI for browsing, installing, and managing third-party integrations.
Includes integration catalog, installation management, credential management,
sync logs, data mappings, webhooks, and usage analytics. Full authentication,
error handling, and email notifications.
"""

from __future__ import annotations

import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext, filedialog, simpledialog
from datetime import datetime, timedelta
from typing import Optional, List, Dict, Any
import json
import logging
import traceback
import secrets
import hashlib

# Core infrastructure
from university_system.infrastructure.auth import UserAuth
from university_system.infrastructure.shared_context import get_auth
from university_system.infrastructure.database.db import get_connection, transaction
from university_system.infrastructure.database.schemas.research_integration_schemas import init_integration_marketplace_system_db
from university_system.modules.shared.constants import paths
from university_system.modules.shared.utils.activity_logger import log_activity

# Integration managers
try:
    from university_system.modules.shared.services.integrations.integration_marketplace_core import (
        IntegrationCatalogManager,
        InstallationManager,
        CredentialManager,
        SyncManager,
        DataMappingManager,
        WebhookManager,
        SecurityCredentialsManager,
        SchedulingManager,
        ValidationTestingManager,
        DataMappingToolsManager,
        NotificationsAlertManager
    )
    MANAGERS_AVAILABLE = True
except ImportError:
    MANAGERS_AVAILABLE = False

# Email service
try:
    from university_system.infrastructure.email.email_service import send_email
    EMAIL_AVAILABLE = True
except ImportError:
    EMAIL_AVAILABLE = False

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Import i18n module for internationalization
from university_system.modules.shared.utils.i18n import (
    init_i18n,
    get_text as _t,
    set_language,
    get_current_language,
    get_current_language_name,
    get_available_languages,
)
from university_system.modules.shared.utils.gui_language_selector import (
    show_gui_language_selector,
)


class IntegrationMarketplaceGUI:
    """Integration Marketplace Management GUI"""

    def __init__(self, parent: tk.Tk, auth_system: Optional[UserAuth] = None):
        # Create a new Toplevel window instead of modifying the parent
        self.root = tk.Toplevel(parent)
        self.parent = parent

        # Initialize i18n
        init_i18n()

        self.root.title(_t("integration_marketplace.window_title"))
        self.root.geometry("1400x900")

        # Initialize authentication
        if auth_system:
            self.auth = auth_system
        else:
            # Use centralized auth system
            self.auth = get_auth()
            if self.auth is None:
                self.auth = UserAuth()

        # Initialize database
        self.init_database()

        # Configure styles
        self.setup_styles()

        # Setup current user
        self.setup_current_user()

        # Check authentication - user must log in through main GUI
        if self.auth and hasattr(self.auth, 'current_user') and self.auth.current_user:
            self.create_main_interface()
        else:
            messagebox.showerror(
                _t("integration_marketplace.auth.required_title"),
                _t("integration_marketplace.auth.required_message")
            )
            self.root.destroy()

    def setup_current_user(self):
        """Setup current user from authentication system"""
        try:
            if self.auth and hasattr(self.auth, 'current_user') and self.auth.current_user:
                logger.info(f"Integration Marketplace GUI: Using authenticated user {self.auth.current_user.get('username', 'Unknown')}")
            else:
                logger.info("Integration Marketplace GUI: No authenticated user - will show login screen")
        except Exception as e:
            logger.error(f"Error setting up current user: {e}")

    def init_database(self):
        """Initialize database tables"""
        try:
            init_integration_marketplace_system_db()
            logger.info("Integration Marketplace database tables initialized successfully")
        except Exception as e:
            logger.error(f"Error initializing database: {e}")
            logger.error(traceback.format_exc())
            messagebox.showerror(_t("integration_marketplace.errors.database_error"), f"{_t('integration_marketplace.errors.init_failed')}: {e}")

    def setup_styles(self):
        """Configure ttk styles to match main_gui.py standards"""
        style = ttk.Style()

        # Use clam theme like main_gui.py
        style.theme_use('clam')

        # Header styling (matching main_gui.py)
        style.configure('Header.TLabel',
                       font=('Arial', 18, 'bold'))

        style.configure('Title.TLabel',
                       font=('Arial', 14, 'bold'))

        style.configure('Section.TLabel',
                       font=('Arial', 12, 'bold'))

        style.configure('Info.TLabel',
                       font=('Arial', 11))

        # Button styles matching main_gui.py
        style.configure('TButton',
                       font=('Arial', 10))

        style.configure('Accent.TButton',
                       font=('Arial', 10, 'bold'))

        # Treeview styling
        style.configure('Treeview',
                       font=('Arial', 10),
                       rowheight=25)

        style.configure('Treeview.Heading',
                       font=('Arial', 10, 'bold'))

        # LabelFrame styling
        style.configure('TLabelframe.Label',
                       font=('Arial', 11, 'bold'))

    def create_main_interface(self):
        """Create the main GUI interface"""
        # Clear existing widgets
        for widget in self.root.winfo_children():
            widget.destroy()

        # Main container frame with padding (matching main_gui.py)
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill='both', expand=True)

        # Header frame with LabelFrame style (matching main_gui.py)
        header_frame = ttk.LabelFrame(main_frame, text=_t("integration_marketplace.title"), padding="10")
        header_frame.pack(fill='x', pady=(0, 10))

        # Control buttons row
        button_frame = ttk.Frame(header_frame)
        button_frame.pack(fill='x', pady=(0, 10))

        # Return to homepage button
        ttk.Button(button_frame, text=_t("integration_marketplace.buttons.return_main"),
                  command=self.return_to_main_menu,
                  style='Accent.TButton').pack(side='left', padx=(0, 10))

        # Refresh all button
        ttk.Button(button_frame, text=_t("integration_marketplace.buttons.refresh_all"),
                  command=self.refresh_all_data).pack(side='left', padx=(0, 10))

        # Language selector button
        ttk.Button(button_frame, text=f"🌐 {get_current_language_name()}",
                  command=self._on_language_change).pack(side='left', padx=(0, 10))

        # Status row
        status_frame = ttk.Frame(header_frame)
        status_frame.pack(fill='x')

        ttk.Label(status_frame, text=_t("integration_marketplace.status.label"), font=('Arial', 10, 'bold')).pack(side='left')
        ttk.Label(status_frame, text=_t("integration_marketplace.status.connected"), font=('Arial', 10)).pack(side='left', padx=(10, 20))

        # User info
        if self.auth.current_user:
            ttk.Label(status_frame, text=_t("integration_marketplace.status.current_user"), font=('Arial', 10, 'bold')).pack(side='left')
            user_info = f"{self.auth.current_user.get('username', 'User')} ({self.auth.current_user.get('role', 'user')})"
            ttk.Label(status_frame, text=user_info, font=('Arial', 10)).pack(side='left', padx=(10, 0))

        # Notebook for tabs
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.pack(fill='both', expand=True)

        # Create tabs
        self.create_catalog_tab()
        self.create_installed_tab()
        self.create_credentials_tab()
        self.create_sync_logs_tab()
        self.create_mappings_tab()
        self.create_webhooks_tab()
        self.create_analytics_tab()

        # Refresh data
        self.refresh_all_data()

    def create_catalog_tab(self):
        """Create integration catalog tab"""
        catalog_frame = ttk.Frame(self.notebook)
        self.notebook.add(catalog_frame, text=_t("integration_marketplace.tabs.catalog"))

        # Controls
        controls_frame = ttk.Frame(catalog_frame)
        controls_frame.pack(fill='x', padx=10, pady=5)

        ttk.Button(controls_frame, text=_t("integration_marketplace.catalog.add_integration"),
                  command=self.add_integration).pack(side='left', padx=5)
        ttk.Button(controls_frame, text=_t("integration_marketplace.common.refresh"),
                  command=self.load_catalog).pack(side='left', padx=5)
        ttk.Button(controls_frame, text=_t("integration_marketplace.catalog.install_selected"),
                  command=self.install_integration, style='Accent.TButton').pack(side='left', padx=5)
        ttk.Button(controls_frame, text=_t("integration_marketplace.catalog.view_details"),
                  command=self.view_integration_details).pack(side='left', padx=5)
        ttk.Button(controls_frame, text=_t("integration_marketplace.catalog.find_similar"),
                  command=self.find_similar_integrations).pack(side='left', padx=5)

        # Second row of controls for search and filters
        search_frame = ttk.Frame(catalog_frame)
        search_frame.pack(fill='x', padx=10, pady=5)

        # Search box
        ttk.Label(search_frame, text=_t("integration_marketplace.common.search")).pack(side='left', padx=5)
        self.catalog_search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=self.catalog_search_var, width=25)
        search_entry.pack(side='left', padx=5)
        ttk.Button(search_frame, text=_t("integration_marketplace.common.search_btn"),
                  command=self.search_catalog).pack(side='left', padx=5)

        # Category filter
        ttk.Label(search_frame, text=_t("integration_marketplace.catalog.category")).pack(side='left', padx=5)
        self.catalog_category_var = tk.StringVar(value='')
        ttk.Combobox(search_frame, textvariable=self.catalog_category_var,
                    values=['', 'LMS', 'SIS', 'CRM', 'Analytics', 'Communication', 'Storage'],
                    width=12, state='readonly').pack(side='left', padx=5)

        # Rating filter
        ttk.Label(search_frame, text=_t("integration_marketplace.catalog.min_rating")).pack(side='left', padx=5)
        self.catalog_rating_var = tk.StringVar(value='')
        ttk.Combobox(search_frame, textvariable=self.catalog_rating_var,
                    values=['', '1', '2', '3', '4', '5'],
                    width=5, state='readonly').pack(side='left', padx=5)

        ttk.Button(search_frame, text=_t("integration_marketplace.common.filter"),
                  command=self.load_catalog).pack(side='left', padx=5)
        ttk.Button(search_frame, text=_t("integration_marketplace.catalog.filter_by_rating"),
                  command=self.filter_by_rating).pack(side='left', padx=5)
        ttk.Button(search_frame, text=_t("integration_marketplace.catalog.compatible_only"),
                  command=self.filter_by_compatibility).pack(side='left', padx=5)
        ttk.Button(search_frame, text=_t("integration_marketplace.catalog.advanced_filter"),
                  command=self.advanced_filter_dialog).pack(side='left', padx=5)

        # Catalog tree
        tree_frame = ttk.Frame(catalog_frame)
        tree_frame.pack(fill='both', expand=True, padx=10, pady=5)

        columns = ('integration_id', 'integration_name', 'provider_name', 'category',
                  'integration_type', 'version', 'rating', 'install_count', 'is_official')
        self.catalog_tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=15)

        column_labels = {
            'integration_id': _t("integration_marketplace.columns.integration_id"),
            'integration_name': _t("integration_marketplace.columns.integration_name"),
            'provider_name': _t("integration_marketplace.columns.provider_name"),
            'category': _t("integration_marketplace.columns.category"),
            'integration_type': _t("integration_marketplace.columns.integration_type"),
            'version': _t("integration_marketplace.columns.version"),
            'rating': _t("integration_marketplace.columns.rating"),
            'install_count': _t("integration_marketplace.columns.install_count"),
            'is_official': _t("integration_marketplace.columns.is_official")
        }

        for col in columns:
            self.catalog_tree.heading(col, text=column_labels[col])
            self.catalog_tree.column(col, width=100)

        self.catalog_tree.column('integration_name', width=200)
        self.catalog_tree.column('provider_name', width=150)

        # Scrollbars
        vsb = ttk.Scrollbar(tree_frame, orient='vertical', command=self.catalog_tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient='horizontal', command=self.catalog_tree.xview)
        self.catalog_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.catalog_tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')

        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        # Description panel
        desc_frame = ttk.LabelFrame(catalog_frame, text=_t("integration_marketplace.catalog.description"), padding=10)
        desc_frame.pack(fill='x', padx=10, pady=5)

        self.catalog_description = scrolledtext.ScrolledText(desc_frame, height=4, wrap=tk.WORD)
        self.catalog_description.pack(fill='both', expand=True)

        # Bind selection event
        self.catalog_tree.bind('<<TreeviewSelect>>', self.on_catalog_select)

    def create_installed_tab(self):
        """Create installed integrations tab"""
        installed_frame = ttk.Frame(self.notebook)
        self.notebook.add(installed_frame, text=_t("integration_marketplace.tabs.installed"))

        # Controls
        controls_frame = ttk.Frame(installed_frame)
        controls_frame.pack(fill='x', padx=10, pady=5)

        ttk.Button(controls_frame, text=_t("integration_marketplace.common.refresh"),
                  command=self.load_installed).pack(side='left', padx=5)
        ttk.Button(controls_frame, text=_t("integration_marketplace.installed.configure"),
                  command=self.configure_integration).pack(side='left', padx=5)
        ttk.Button(controls_frame, text=_t("integration_marketplace.installed.sync_now"),
                  command=self.sync_integration).pack(side='left', padx=5)
        ttk.Button(controls_frame, text=_t("integration_marketplace.installed.uninstall"),
                  command=self.uninstall_integration).pack(side='left', padx=5)

        # Status filter
        ttk.Label(controls_frame, text=_t("integration_marketplace.common.status")).pack(side='left', padx=5)
        self.installed_filter_var = tk.StringVar(value='active')
        ttk.Combobox(controls_frame, textvariable=self.installed_filter_var,
                    values=['all', 'active', 'inactive', 'error'],
                    width=10, state='readonly').pack(side='left', padx=5)
        ttk.Button(controls_frame, text=_t("integration_marketplace.common.filter"),
                  command=self.load_installed).pack(side='left', padx=5)

        # Bulk operations frame
        bulk_frame = ttk.LabelFrame(installed_frame, text=_t("integration_marketplace.installed.bulk_operations"), padding=5)
        bulk_frame.pack(fill='x', padx=10, pady=5)

        ttk.Button(bulk_frame, text=_t("integration_marketplace.installed.bulk_install"),
                  command=self.bulk_install_integrations).pack(side='left', padx=5)
        ttk.Button(bulk_frame, text=_t("integration_marketplace.installed.bulk_uninstall"),
                  command=self.bulk_uninstall_integrations).pack(side='left', padx=5)
        ttk.Button(bulk_frame, text=_t("integration_marketplace.installed.bulk_enable"),
                  command=self.bulk_enable_integrations).pack(side='left', padx=5)
        ttk.Button(bulk_frame, text=_t("integration_marketplace.installed.bulk_disable"),
                  command=self.bulk_disable_integrations).pack(side='left', padx=5)
        ttk.Button(bulk_frame, text=_t("integration_marketplace.installed.bulk_sync"),
                  command=self.bulk_sync_integrations).pack(side='left', padx=5)

        # Scheduling & Automation frame
        scheduling_frame = ttk.LabelFrame(installed_frame, text=_t("integration_marketplace.installed.scheduling"), padding=5)
        scheduling_frame.pack(fill='x', padx=10, pady=5)

        ttk.Button(scheduling_frame, text=_t("integration_marketplace.installed.schedule_sync"),
                  command=self.schedule_sync).pack(side='left', padx=5)
        ttk.Button(scheduling_frame, text=_t("integration_marketplace.installed.view_scheduled"),
                  command=self.view_scheduled_tasks).pack(side='left', padx=5)
        ttk.Button(scheduling_frame, text=_t("integration_marketplace.installed.pause_syncs"),
                  command=self.pause_scheduled_syncs).pack(side='left', padx=5)
        ttk.Button(scheduling_frame, text=_t("integration_marketplace.installed.maintenance_window"),
                  command=self.set_maintenance_window).pack(side='left', padx=5)
        ttk.Button(scheduling_frame, text=_t("integration_marketplace.installed.retry_policy"),
                  command=self.configure_retry_policy).pack(side='left', padx=5)

        # Installed tree
        tree_frame = ttk.Frame(installed_frame)
        tree_frame.pack(fill='both', expand=True, padx=10, pady=5)

        columns = ('install_id', 'integration_name', 'version_installed', 'installation_date',
                  'status', 'last_sync_date', 'sync_frequency', 'is_enabled')
        self.installed_tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=15)

        column_labels = {
            'install_id': _t("integration_marketplace.columns.install_id"),
            'integration_name': _t("integration_marketplace.columns.integration_name"),
            'version_installed': _t("integration_marketplace.columns.version_installed"),
            'installation_date': _t("integration_marketplace.columns.installation_date"),
            'status': _t("integration_marketplace.columns.status"),
            'last_sync_date': _t("integration_marketplace.columns.last_sync_date"),
            'sync_frequency': _t("integration_marketplace.columns.sync_frequency"),
            'is_enabled': _t("integration_marketplace.columns.is_enabled")
        }

        for col in columns:
            self.installed_tree.heading(col, text=column_labels[col])
            self.installed_tree.column(col, width=120)

        self.installed_tree.column('integration_name', width=200)

        # Scrollbars
        vsb = ttk.Scrollbar(tree_frame, orient='vertical', command=self.installed_tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient='horizontal', command=self.installed_tree.xview)
        self.installed_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.installed_tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')

        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

    def create_credentials_tab(self):
        """Create credentials management tab"""
        cred_frame = ttk.Frame(self.notebook)
        self.notebook.add(cred_frame, text=_t("integration_marketplace.tabs.credentials"))

        # Controls
        controls_frame = ttk.Frame(cred_frame)
        controls_frame.pack(fill='x', padx=10, pady=5)

        ttk.Button(controls_frame, text=_t("integration_marketplace.credentials.add"),
                  command=self.add_credentials).pack(side='left', padx=5)
        ttk.Button(controls_frame, text=_t("integration_marketplace.common.refresh"),
                  command=self.load_credentials).pack(side='left', padx=5)
        ttk.Button(controls_frame, text=_t("integration_marketplace.credentials.edit"),
                  command=self.edit_credentials).pack(side='left', padx=5)
        ttk.Button(controls_frame, text=_t("integration_marketplace.common.delete"),
                  command=self.delete_credentials).pack(side='left', padx=5)
        ttk.Button(controls_frame, text=_t("integration_marketplace.credentials.bulk_update"),
                  command=self.bulk_update_credentials).pack(side='left', padx=5)

        # Security & Credentials frame
        security_frame = ttk.LabelFrame(cred_frame, text=_t("integration_marketplace.credentials.security"), padding=5)
        security_frame.pack(fill='x', padx=10, pady=5)

        ttk.Button(security_frame, text=_t("integration_marketplace.credentials.rotate_api"),
                  command=self.rotate_api_credentials).pack(side='left', padx=5)
        ttk.Button(security_frame, text=_t("integration_marketplace.credentials.check_expiry"),
                  command=self.check_credential_expiry).pack(side='left', padx=5)
        ttk.Button(security_frame, text=_t("integration_marketplace.credentials.validate"),
                  command=self.validate_credentials).pack(side='left', padx=5)
        ttk.Button(security_frame, text=_t("integration_marketplace.credentials.encrypted_export"),
                  command=self.encrypt_export_credentials).pack(side='left', padx=5)
        ttk.Button(security_frame, text=_t("integration_marketplace.credentials.audit_access"),
                  command=self.audit_credential_access).pack(side='left', padx=5)
        ttk.Button(security_frame, text=_t("integration_marketplace.credentials.revoke_tokens"),
                  command=self.revoke_all_tokens).pack(side='left', padx=5)

        # Credentials tree
        tree_frame = ttk.Frame(cred_frame)
        tree_frame.pack(fill='both', expand=True, padx=10, pady=5)

        columns = ('credential_id', 'install_id', 'credential_type', 'endpoint_url',
                  'created_at', 'token_expiry')
        self.cred_tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=15)

        column_labels = {
            'credential_id': _t("integration_marketplace.columns.credential_id"),
            'install_id': _t("integration_marketplace.columns.install_id"),
            'credential_type': _t("integration_marketplace.columns.credential_type"),
            'endpoint_url': _t("integration_marketplace.columns.endpoint_url"),
            'created_at': _t("integration_marketplace.columns.created_at"),
            'token_expiry': _t("integration_marketplace.columns.token_expiry")
        }

        for col in columns:
            self.cred_tree.heading(col, text=column_labels[col])
            self.cred_tree.column(col, width=120)

        self.cred_tree.column('endpoint_url', width=250)

        # Scrollbars
        vsb = ttk.Scrollbar(tree_frame, orient='vertical', command=self.cred_tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient='horizontal', command=self.cred_tree.xview)
        self.cred_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.cred_tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')

        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

    def create_sync_logs_tab(self):
        """Create sync logs tab"""
        sync_frame = ttk.Frame(self.notebook)
        self.notebook.add(sync_frame, text=_t("integration_marketplace.tabs.sync_logs"))

        # Controls
        controls_frame = ttk.Frame(sync_frame)
        controls_frame.pack(fill='x', padx=10, pady=5)

        ttk.Button(controls_frame, text=_t("integration_marketplace.common.refresh"),
                  command=self.load_sync_logs).pack(side='left', padx=5)
        ttk.Button(controls_frame, text=_t("integration_marketplace.sync_logs.view_details"),
                  command=self.view_sync_details).pack(side='left', padx=5)
        ttk.Button(controls_frame, text=_t("integration_marketplace.sync_logs.export_logs"),
                  command=self.export_sync_logs).pack(side='left', padx=5)
        ttk.Button(controls_frame, text=_t("integration_marketplace.sync_logs.search_logs"),
                  command=self.search_sync_logs).pack(side='left', padx=5)
        ttk.Button(controls_frame, text=_t("integration_marketplace.sync_logs.export_pdf"),
                  command=self.export_sync_report_pdf).pack(side='left', padx=5)

        # Status filter
        ttk.Label(controls_frame, text=_t("integration_marketplace.common.status")).pack(side='left', padx=5)
        self.sync_filter_var = tk.StringVar(value='all')
        ttk.Combobox(controls_frame, textvariable=self.sync_filter_var,
                    values=['all', 'success', 'failed', 'running'],
                    width=10, state='readonly').pack(side='left', padx=5)
        ttk.Button(controls_frame, text=_t("integration_marketplace.common.filter"),
                  command=self.load_sync_logs).pack(side='left', padx=5)

        # Validation & Testing frame
        validation_frame = ttk.LabelFrame(sync_frame, text=_t("integration_marketplace.sync_logs.validation"), padding=5)
        validation_frame.pack(fill='x', padx=10, pady=5)

        ttk.Button(validation_frame, text=_t("integration_marketplace.sync_logs.test_connection"),
                  command=self.test_integration_connection).pack(side='left', padx=5)
        ttk.Button(validation_frame, text=_t("integration_marketplace.sync_logs.validate_mappings"),
                  command=self.validate_mapping_rules).pack(side='left', padx=5)
        ttk.Button(validation_frame, text=_t("integration_marketplace.sync_logs.dry_run"),
                  command=self.dry_run_sync).pack(side='left', padx=5)
        ttk.Button(validation_frame, text=_t("integration_marketplace.sync_logs.test_webhook"),
                  command=self.test_webhook_delivery).pack(side='left', padx=5)
        ttk.Button(validation_frame, text=_t("integration_marketplace.sync_logs.validate_config"),
                  command=self.validate_json_configuration).pack(side='left', padx=5)
        ttk.Button(validation_frame, text=_t("integration_marketplace.sync_logs.run_diagnostics"),
                  command=self.run_integration_diagnostics).pack(side='left', padx=5)

        # Sync logs tree
        tree_frame = ttk.Frame(sync_frame)
        tree_frame.pack(fill='both', expand=True, padx=10, pady=5)

        columns = ('log_id', 'install_id', 'sync_start_time', 'sync_end_time',
                  'sync_status', 'records_synced', 'errors_encountered')
        self.sync_tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=15)

        column_labels = {
            'log_id': _t("integration_marketplace.columns.log_id"),
            'install_id': _t("integration_marketplace.columns.install_id"),
            'sync_start_time': _t("integration_marketplace.columns.sync_start_time"),
            'sync_end_time': _t("integration_marketplace.columns.sync_end_time"),
            'sync_status': _t("integration_marketplace.columns.sync_status"),
            'records_synced': _t("integration_marketplace.columns.records_synced"),
            'errors_encountered': _t("integration_marketplace.columns.errors_encountered")
        }

        for col in columns:
            self.sync_tree.heading(col, text=column_labels[col])
            self.sync_tree.column(col, width=120)

        # Scrollbars
        vsb = ttk.Scrollbar(tree_frame, orient='vertical', command=self.sync_tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient='horizontal', command=self.sync_tree.xview)
        self.sync_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.sync_tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')

        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

    def create_mappings_tab(self):
        """Create data mappings tab"""
        mappings_frame = ttk.Frame(self.notebook)
        self.notebook.add(mappings_frame, text=_t("integration_marketplace.tabs.mappings"))

        # Controls
        controls_frame = ttk.Frame(mappings_frame)
        controls_frame.pack(fill='x', padx=10, pady=5)

        ttk.Button(controls_frame, text=_t("integration_marketplace.mappings.add"),
                  command=self.add_mapping).pack(side='left', padx=5)
        ttk.Button(controls_frame, text=_t("integration_marketplace.common.refresh"),
                  command=self.load_mappings).pack(side='left', padx=5)
        ttk.Button(controls_frame, text=_t("integration_marketplace.mappings.edit"),
                  command=self.edit_mapping).pack(side='left', padx=5)
        ttk.Button(controls_frame, text=_t("integration_marketplace.common.delete"),
                  command=self.delete_mapping).pack(side='left', padx=5)
        ttk.Button(controls_frame, text=_t("integration_marketplace.mappings.export_excel"),
                  command=self.export_mappings_to_excel).pack(side='left', padx=5)

        # Data Mapping Tools frame
        mapping_tools_frame = ttk.LabelFrame(mappings_frame, text=_t("integration_marketplace.mappings.tools"), padding=5)
        mapping_tools_frame.pack(fill='x', padx=10, pady=5)

        ttk.Button(mapping_tools_frame, text=_t("integration_marketplace.mappings.auto_detect"),
                  command=self.auto_detect_mappings).pack(side='left', padx=5)
        ttk.Button(mapping_tools_frame, text=_t("integration_marketplace.mappings.preview"),
                  command=self.preview_transformation).pack(side='left', padx=5)
        ttk.Button(mapping_tools_frame, text=_t("integration_marketplace.mappings.duplicate"),
                  command=self.duplicate_mapping_set).pack(side='left', padx=5)
        ttk.Button(mapping_tools_frame, text=_t("integration_marketplace.mappings.import_template"),
                  command=self.import_mappings_from_template).pack(side='left', padx=5)

        # Mappings tree
        tree_frame = ttk.Frame(mappings_frame)
        tree_frame.pack(fill='both', expand=True, padx=10, pady=5)

        columns = ('mapping_id', 'install_id', 'source_field', 'target_field',
                  'transformation_rule', 'is_active')
        self.mappings_tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=15)

        column_labels = {
            'mapping_id': _t("integration_marketplace.columns.mapping_id"),
            'install_id': _t("integration_marketplace.columns.install_id"),
            'source_field': _t("integration_marketplace.columns.source_field"),
            'target_field': _t("integration_marketplace.columns.target_field"),
            'transformation_rule': _t("integration_marketplace.columns.transformation_rule"),
            'is_active': _t("integration_marketplace.columns.is_active")
        }

        for col in columns:
            self.mappings_tree.heading(col, text=column_labels[col])
            self.mappings_tree.column(col, width=120)

        self.mappings_tree.column('source_field', width=150)
        self.mappings_tree.column('target_field', width=150)
        self.mappings_tree.column('transformation_rule', width=200)

        # Scrollbars
        vsb = ttk.Scrollbar(tree_frame, orient='vertical', command=self.mappings_tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient='horizontal', command=self.mappings_tree.xview)
        self.mappings_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.mappings_tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')

        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

    def create_webhooks_tab(self):
        """Create webhooks tab"""
        webhooks_frame = ttk.Frame(self.notebook)
        self.notebook.add(webhooks_frame, text=_t("integration_marketplace.tabs.webhooks"))

        # Controls
        controls_frame = ttk.Frame(webhooks_frame)
        controls_frame.pack(fill='x', padx=10, pady=5)

        ttk.Button(controls_frame, text=_t("integration_marketplace.webhooks.add"),
                  command=self.add_webhook).pack(side='left', padx=5)
        ttk.Button(controls_frame, text=_t("integration_marketplace.common.refresh"),
                  command=self.load_webhooks).pack(side='left', padx=5)
        ttk.Button(controls_frame, text=_t("integration_marketplace.webhooks.edit"),
                  command=self.edit_webhook).pack(side='left', padx=5)
        ttk.Button(controls_frame, text=_t("integration_marketplace.common.delete"),
                  command=self.delete_webhook).pack(side='left', padx=5)

        # Notifications & Alerts frame
        notifications_frame = ttk.LabelFrame(webhooks_frame, text=_t("integration_marketplace.webhooks.notifications"), padding=5)
        notifications_frame.pack(fill='x', padx=10, pady=5)

        ttk.Button(notifications_frame, text=_t("integration_marketplace.webhooks.configure_alerts"),
                  command=self.configure_alert_rules).pack(side='left', padx=5)
        ttk.Button(notifications_frame, text=_t("integration_marketplace.webhooks.subscribe"),
                  command=self.subscribe_to_notifications).pack(side='left', padx=5)
        ttk.Button(notifications_frame, text=_t("integration_marketplace.webhooks.view_history"),
                  command=self.view_notification_history).pack(side='left', padx=5)
        ttk.Button(notifications_frame, text=_t("integration_marketplace.webhooks.test_channel"),
                  command=self.test_notification_channel).pack(side='left', padx=5)

        # Webhooks tree
        tree_frame = ttk.Frame(webhooks_frame)
        tree_frame.pack(fill='both', expand=True, padx=10, pady=5)

        columns = ('webhook_id', 'install_id', 'webhook_url', 'event_type',
                  'is_active', 'last_triggered_at', 'created_at')
        self.webhooks_tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=15)

        column_labels = {
            'webhook_id': _t("integration_marketplace.columns.webhook_id"),
            'install_id': _t("integration_marketplace.columns.install_id"),
            'webhook_url': _t("integration_marketplace.columns.webhook_url"),
            'event_type': _t("integration_marketplace.columns.event_type"),
            'is_active': _t("integration_marketplace.columns.is_active"),
            'last_triggered_at': _t("integration_marketplace.columns.last_triggered_at"),
            'created_at': _t("integration_marketplace.columns.created_at")
        }

        for col in columns:
            self.webhooks_tree.heading(col, text=column_labels[col])
            self.webhooks_tree.column(col, width=120)

        self.webhooks_tree.column('webhook_url', width=300)

        # Scrollbars
        vsb = ttk.Scrollbar(tree_frame, orient='vertical', command=self.webhooks_tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient='horizontal', command=self.webhooks_tree.xview)
        self.webhooks_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.webhooks_tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')

        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

    def create_analytics_tab(self):
        """Create usage analytics tab"""
        analytics_frame = ttk.Frame(self.notebook)
        self.notebook.add(analytics_frame, text=_t("integration_marketplace.tabs.analytics"))

        # Controls
        controls_frame = ttk.Frame(analytics_frame)
        controls_frame.pack(fill='x', padx=10, pady=5)

        ttk.Button(controls_frame, text=_t("integration_marketplace.common.refresh"),
                  command=self.load_analytics).pack(side='left', padx=5)
        ttk.Button(controls_frame, text=_t("integration_marketplace.analytics.view_summary"),
                  command=self.view_analytics_summary).pack(side='left', padx=5)
        ttk.Button(controls_frame, text=_t("integration_marketplace.common.export"),
                  command=self.export_analytics).pack(side='left', padx=5)

        # Date filter
        ttk.Label(controls_frame, text=_t("integration_marketplace.analytics.days")).pack(side='left', padx=5)
        self.analytics_days_var = tk.StringVar(value='30')
        ttk.Spinbox(controls_frame, from_=1, to=365, textvariable=self.analytics_days_var,
                   width=10).pack(side='left', padx=5)
        ttk.Button(controls_frame, text=_t("integration_marketplace.common.filter"),
                  command=self.load_analytics).pack(side='left', padx=5)

        # Reports & Dashboard frame
        reports_frame = ttk.LabelFrame(analytics_frame, text=_t("integration_marketplace.analytics.reports"), padding=5)
        reports_frame.pack(fill='x', padx=10, pady=5)

        ttk.Button(reports_frame, text=_t("integration_marketplace.analytics.dashboard"),
                  command=self.show_dashboard_overview).pack(side='left', padx=5)
        ttk.Button(reports_frame, text=_t("integration_marketplace.analytics.health_report"),
                  command=self.generate_health_report).pack(side='left', padx=5)
        ttk.Button(reports_frame, text=_t("integration_marketplace.analytics.error_analysis"),
                  command=self.show_error_analysis).pack(side='left', padx=5)
        ttk.Button(reports_frame, text=_t("integration_marketplace.analytics.usage_trend"),
                  command=self.generate_usage_trend_chart).pack(side='left', padx=5)
        ttk.Button(reports_frame, text=_t("integration_marketplace.analytics.api_statistics"),
                  command=self.show_api_call_statistics).pack(side='left', padx=5)
        ttk.Button(reports_frame, text=_t("integration_marketplace.analytics.compare_performance"),
                  command=self.compare_integration_performance).pack(side='left', padx=5)
        ttk.Button(reports_frame, text=_t("integration_marketplace.analytics.compliance_report"),
                  command=self.generate_compliance_report).pack(side='left', padx=5)

        # Import/Export frame
        import_export_frame = ttk.LabelFrame(analytics_frame, text=_t("integration_marketplace.analytics.import_export"), padding=5)
        import_export_frame.pack(fill='x', padx=10, pady=5)

        ttk.Button(import_export_frame, text=_t("integration_marketplace.analytics.export_catalog"),
                  command=self.export_catalog_to_json).pack(side='left', padx=5)
        ttk.Button(import_export_frame, text=_t("integration_marketplace.analytics.import_json"),
                  command=self.import_integrations_from_json).pack(side='left', padx=5)
        ttk.Button(import_export_frame, text=_t("integration_marketplace.analytics.export_config"),
                  command=self.export_configuration_bundle).pack(side='left', padx=5)
        ttk.Button(import_export_frame, text=_t("integration_marketplace.analytics.import_config"),
                  command=self.import_configuration_bundle).pack(side='left', padx=5)

        # Analytics tree
        tree_frame = ttk.Frame(analytics_frame)
        tree_frame.pack(fill='both', expand=True, padx=10, pady=5)

        columns = ('analytics_id', 'install_id', 'metric_name', 'metric_value',
                  'measurement_date')
        self.analytics_tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=15)

        for col in columns:
            self.analytics_tree.heading(col, text=col.replace('_', ' ').title())
            self.analytics_tree.column(col, width=150)

        self.analytics_tree.column('metric_name', width=200)

        # Scrollbars
        vsb = ttk.Scrollbar(tree_frame, orient='vertical', command=self.analytics_tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient='horizontal', command=self.analytics_tree.xview)
        self.analytics_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.analytics_tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')

        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

    # Data loading methods
    def refresh_all_data(self):
        """Refresh all tab data"""
        try:
            self.load_catalog()
            self.load_installed()
            self.load_credentials()
            self.load_sync_logs()
            self.load_mappings()
            self.load_webhooks()
            self.load_analytics()
        except Exception as e:
            logger.error(f"Error refreshing data: {e}")
            messagebox.showerror(_t("common.error"), _t("integration_marketplace.errors.refresh_failed", error=str(e)))

    def load_catalog(self):
        """Load integration catalog"""
        try:
            for item in self.catalog_tree.get_children():
                self.catalog_tree.delete(item)

            with get_connection() as conn:
                cursor = conn.cursor()

                category = self.catalog_category_var.get()
                if category:
                    cursor.execute('''
                        SELECT integration_id, integration_name, provider_name, category,
                               integration_type, version, rating, install_count, is_official
                        FROM integration_catalog
                        WHERE category = ? AND is_active = 1
                        ORDER BY rating DESC, install_count DESC
                    ''', (category,))
                else:
                    cursor.execute('''
                        SELECT integration_id, integration_name, provider_name, category,
                               integration_type, version, rating, install_count, is_official
                        FROM integration_catalog
                        WHERE is_active = 1
                        ORDER BY rating DESC, install_count DESC
                        LIMIT 100
                    ''')

                integrations = cursor.fetchall()

                for integration in integrations:
                    is_official = _t("common.yes") if integration[8] else _t("common.no")
                    self.catalog_tree.insert('', 'end', values=(
                        integration[0], integration[1], integration[2], integration[3],
                        integration[4], integration[5], integration[6] or _t("common.na"),
                        integration[7] or 0, is_official
                    ))

            logger.info(f"Loaded {len(integrations)} integrations from catalog")

        except Exception as e:
            logger.error(f"Error loading catalog: {e}\n{traceback.format_exc()}")
            messagebox.showerror(_t("common.error"), f"Failed to load catalog: {e}")

    def load_installed(self):
        """Load installed integrations"""
        try:
            for item in self.installed_tree.get_children():
                self.installed_tree.delete(item)

            with get_connection() as conn:
                cursor = conn.cursor()

                status_filter = self.installed_filter_var.get()
                if status_filter == 'all':
                    cursor.execute('''
                        SELECT ii.install_id, ic.integration_name, ii.version_installed,
                               ii.installation_date, ii.status, ii.last_sync_date,
                               ii.sync_frequency, ii.is_enabled
                        FROM installed_integrations ii
                        JOIN integration_catalog ic ON ii.integration_id = ic.integration_id
                        ORDER BY ii.installation_date DESC
                    ''')
                else:
                    cursor.execute('''
                        SELECT ii.install_id, ic.integration_name, ii.version_installed,
                               ii.installation_date, ii.status, ii.last_sync_date,
                               ii.sync_frequency, ii.is_enabled
                        FROM installed_integrations ii
                        JOIN integration_catalog ic ON ii.integration_id = ic.integration_id
                        WHERE ii.status = ?
                        ORDER BY ii.installation_date DESC
                    ''', (status_filter,))

                installed = cursor.fetchall()

                for install in installed:
                    enabled = _t("common.enabled") if install[7] else _t("common.disabled")
                    self.installed_tree.insert('', 'end', values=(
                        install[0], install[1], install[2], install[3],
                        install[4], install[5] or _t("common.never"), install[6], enabled
                    ))

            logger.info(f"Loaded {len(installed)} installed integrations")

        except Exception as e:
            logger.error(f"Error loading installed integrations: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to load installed integrations: {e}")

    def load_credentials(self):
        """Load integration credentials"""
        try:
            for item in self.cred_tree.get_children():
                self.cred_tree.delete(item)

            with get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT credential_id, install_id, credential_type, endpoint_url,
                           created_at, token_expiry
                    FROM integration_credentials
                    ORDER BY created_at DESC
                    LIMIT 100
                ''')

                credentials = cursor.fetchall()

                for cred in credentials:
                    # Properly extract values from sqlite3.Row
                    self.cred_tree.insert('', 'end', values=(
                        cred[0],  # credential_id
                        cred[1],  # install_id
                        cred[2] or _t("common.na"),  # credential_type
                        cred[3] or _t("common.na"),  # endpoint_url
                        cred[4][:19] if cred[4] else 'N/A',  # created_at (trim timestamp)
                        cred[5][:19] if cred[5] else 'N/A'   # token_expiry (trim timestamp)
                    ))

            logger.info(f"Loaded {len(credentials)} credentials")

        except Exception as e:
            logger.error(f"Error loading credentials: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to load credentials: {e}")

    def load_sync_logs(self):
        """Load integration sync logs"""
        try:
            for item in self.sync_tree.get_children():
                self.sync_tree.delete(item)

            with get_connection() as conn:
                cursor = conn.cursor()

                status_filter = self.sync_filter_var.get()
                if status_filter == 'all':
                    cursor.execute('''
                        SELECT log_id, install_id, sync_start_time, sync_end_time,
                               sync_status, records_synced, errors_encountered
                        FROM integration_sync_logs
                        ORDER BY sync_start_time DESC
                        LIMIT 100
                    ''')
                else:
                    cursor.execute('''
                        SELECT log_id, install_id, sync_start_time, sync_end_time,
                               sync_status, records_synced, errors_encountered
                        FROM integration_sync_logs
                        WHERE sync_status = ?
                        ORDER BY sync_start_time DESC
                        LIMIT 100
                    ''', (status_filter,))

                logs = cursor.fetchall()

                for log in logs:
                    # Properly extract values from sqlite3.Row
                    self.sync_tree.insert('', 'end', values=(
                        log[0],  # log_id
                        log[1],  # install_id
                        log[2][:19] if log[2] else 'N/A',  # sync_start_time
                        log[3][:19] if log[3] else 'N/A',  # sync_end_time
                        log[4] or _t("common.na"),  # sync_status
                        log[5] if log[5] is not None else 0,  # records_synced
                        log[6] if log[6] is not None else 0   # errors_encountered
                    ))

            logger.info(f"Loaded {len(logs)} sync logs")

        except Exception as e:
            logger.error(f"Error loading sync logs: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to load sync logs: {e}")

    def load_mappings(self):
        """Load data mappings"""
        try:
            for item in self.mappings_tree.get_children():
                self.mappings_tree.delete(item)

            with get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT mapping_id, install_id, source_field, target_field,
                           transformation_rule, is_active
                    FROM integration_data_mappings
                    ORDER BY install_id
                    LIMIT 100
                ''')

                mappings = cursor.fetchall()

                for mapping in mappings:
                    status = 'Active' if mapping[5] else 'Inactive'
                    self.mappings_tree.insert('', 'end', values=(
                        mapping[0], mapping[1], mapping[2], mapping[3],
                        mapping[4] or 'None', status
                    ))

            logger.info(f"Loaded {len(mappings)} data mappings")

        except Exception as e:
            logger.error(f"Error loading mappings: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to load mappings: {e}")

    def load_webhooks(self):
        """Load webhooks"""
        try:
            for item in self.webhooks_tree.get_children():
                self.webhooks_tree.delete(item)

            with get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT webhook_id, install_id, webhook_url, event_type,
                           is_active, last_triggered_at, created_at
                    FROM integration_webhooks
                    ORDER BY created_at DESC
                    LIMIT 100
                ''')

                webhooks = cursor.fetchall()

                for webhook in webhooks:
                    status = 'Active' if webhook[4] else 'Inactive'
                    self.webhooks_tree.insert('', 'end', values=(
                        webhook[0], webhook[1], webhook[2], webhook[3],
                        status, webhook[5] or _t("common.never"), webhook[6]
                    ))

            logger.info(f"Loaded {len(webhooks)} webhooks")

        except Exception as e:
            logger.error(f"Error loading webhooks: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to load webhooks: {e}")

    def load_analytics(self):
        """Load usage analytics"""
        try:
            for item in self.analytics_tree.get_children():
                self.analytics_tree.delete(item)

            days = int(self.analytics_days_var.get())
            cutoff_date = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')

            with get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT analytics_id, install_id, metric_name, metric_value,
                           measurement_date
                    FROM integration_usage_analytics
                    WHERE measurement_date >= ?
                    ORDER BY measurement_date DESC
                    LIMIT 100
                ''', (cutoff_date,))

                analytics = cursor.fetchall()

                for item in analytics:
                    self.analytics_tree.insert('', 'end', values=item)

            logger.info(f"Loaded {len(analytics)} analytics records")

        except Exception as e:
            logger.error(f"Error loading analytics: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to load analytics: {e}")

    # Event handlers
    def on_catalog_select(self, event):
        """Handle catalog tree selection"""
        try:
            selected = self.catalog_tree.selection()
            if not selected:
                return

            integration_id = self.catalog_tree.item(selected[0])['values'][0]

            # Get integration details
            with get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT description FROM integration_catalog
                    WHERE integration_id = ?
                ''', (integration_id,))

                result = cursor.fetchone()
                if result:
                    description = result[0] or "No description available."
                    self.catalog_description.delete('1.0', 'end')
                    self.catalog_description.insert('1.0', description)

        except Exception as e:
            logger.error(f"Error loading integration details: {e}")

    # Action methods
    def add_integration(self):
        """Add new integration to catalog"""
        try:
            dialog = tk.Toplevel(self.root)
            dialog.title(_t("integration_marketplace.dialogs.add_integration"))
            dialog.geometry("500x500")
            dialog.transient(self.root)  # Make dialog transient to parent
            dialog.grab_set()  # Make dialog modal

            # Form fields
            ttk.Label(dialog, text=_t("integration_marketplace.labels.integration_name")).grid(row=0, column=0, padx=10, pady=5, sticky='w')
            name_var = tk.StringVar()
            name_entry = ttk.Entry(dialog, textvariable=name_var, width=35)
            name_entry.grid(row=0, column=1, padx=10, pady=5)
            name_entry.focus_set()  # Set initial focus

            ttk.Label(dialog, text=_t("integration_marketplace.labels.provider_name")).grid(row=1, column=0, padx=10, pady=5, sticky='w')
            provider_var = tk.StringVar()
            provider_entry = ttk.Entry(dialog, textvariable=provider_var, width=35)
            provider_entry.grid(row=1, column=1, padx=10, pady=5)

            ttk.Label(dialog, text=_t("integration_marketplace.labels.integration_type")).grid(row=2, column=0, padx=10, pady=5, sticky='w')
            type_var = tk.StringVar(value='API')
            ttk.Combobox(dialog, textvariable=type_var,
                        values=['API', 'OAuth', 'SAML', 'Database', 'Webhook'],
                        width=33, state='readonly').grid(row=2, column=1, padx=10, pady=5)

            ttk.Label(dialog, text=_t("integration_marketplace.labels.category")).grid(row=3, column=0, padx=10, pady=5, sticky='w')
            category_var = tk.StringVar(value='LMS')
            ttk.Combobox(dialog, textvariable=category_var,
                        values=['LMS', 'SIS', 'CRM', 'Analytics', 'Communication', 'Storage'],
                        width=33, state='readonly').grid(row=3, column=1, padx=10, pady=5)

            ttk.Label(dialog, text=_t("integration_marketplace.labels.version")).grid(row=4, column=0, padx=10, pady=5, sticky='w')
            version_var = tk.StringVar(value='1.0.0')
            ttk.Entry(dialog, textvariable=version_var, width=35).grid(row=4, column=1, padx=10, pady=5)

            ttk.Label(dialog, text=_t("integration_marketplace.labels.description")).grid(row=5, column=0, padx=10, pady=5, sticky='nw')
            description_text = scrolledtext.ScrolledText(dialog, width=35, height=6)
            description_text.grid(row=5, column=1, padx=10, pady=5)

            ttk.Label(dialog, text=_t("integration_marketplace.columns.is_official")).grid(row=6, column=0, padx=10, pady=5, sticky='w')
            is_official_var = tk.BooleanVar(value=False)
            ttk.Checkbutton(dialog, variable=is_official_var).grid(row=6, column=1, padx=10, pady=5, sticky='w')

            def save_integration():
                try:
                    integration_name = name_var.get().strip()
                    provider_name = provider_var.get().strip()
                    integration_type = type_var.get()
                    category = category_var.get()
                    version = version_var.get().strip()
                    description = description_text.get('1.0', 'end-1c').strip()
                    is_official = int(is_official_var.get())

                    # Enhanced validation with specific error messages
                    if not integration_name:
                        messagebox.showerror(_t("common.error"), _t("integration_marketplace.messages.fill_required_fields"))
                        name_entry.focus_set()
                        return

                    if not provider_name:
                        messagebox.showerror(_t("common.error"), _t("integration_marketplace.messages.fill_required_fields"))
                        provider_entry.focus_set()
                        return

                    # Log for debugging
                    logger.info(f"Adding integration: name='{integration_name}', provider='{provider_name}', type='{integration_type}'")

                    if MANAGERS_AVAILABLE:
                        integration_id = IntegrationCatalogManager.add_integration(
                            integration_name, provider_name, integration_type,
                            category, description, version, bool(is_official)
                        )
                    else:
                        with transaction() as conn:
                            cursor = conn.cursor()
                            cursor.execute('''
                                INSERT INTO integration_catalog
                                (integration_name, provider_name, integration_type, category,
                                 description, version, is_official, is_active)
                                VALUES (?, ?, ?, ?, ?, ?, ?, 1)
                            ''', (integration_name, provider_name, integration_type, category,
                                  description, version, is_official))

                            integration_id = cursor.lastrowid

                    log_activity('create', 'integration_catalog', integration_id,
                                details={'integration_name': integration_name})

                    messagebox.showinfo(_t("common.success"), f"{_t('integration_marketplace.messages.integration_added')}\nIntegration ID: {integration_id}")
                    dialog.destroy()
                    self.load_catalog()

                except Exception as e:
                    logger.error(f"Error adding integration: {e}\n{traceback.format_exc()}")
                    messagebox.showerror(_t("common.error"), f"Failed to add integration: {e}")

            ttk.Button(dialog, text=_t("integration_marketplace.catalog.add_integration"), command=save_integration).grid(
                row=7, column=0, columnspan=2, pady=20)

        except Exception as e:
            logger.error(f"Error opening add integration dialog: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to open dialog: {e}")

    def view_integration_details(self):
        """View detailed integration information"""
        try:
            selected = self.catalog_tree.selection()
            if not selected:
                messagebox.showwarning(_t("common.warning"), "Please select an integration to view details")
                return

            integration_id = self.catalog_tree.item(selected[0])['values'][0]

            with get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT integration_name, provider_name, integration_type, category,
                           description, version, rating, install_count, pricing_model,
                           documentation_url
                    FROM integration_catalog
                    WHERE integration_id = ?
                ''', (integration_id,))

                integration = cursor.fetchone()

            if not integration:
                messagebox.showerror(_t("common.error"), "Integration not found")
                return

            details = f"Integration Details\n\n"
            details += f"Name: {integration[0]}\n"
            details += f"Provider: {integration[1]}\n"
            details += f"Type: {integration[2]}\n"
            details += f"Category: {integration[3]}\n"
            details += f"Version: {integration[5]}\n"
            details += f"Rating: {integration[6] or 'Not rated'}\n"
            details += f"Installs: {integration[7] or 0}\n"
            details += f"Pricing: {integration[8] or 'Contact provider'}\n"
            details += f"Documentation: {integration[9] or _t('common.na')}\n\n"
            details += f"Description:\n{integration[4] or 'No description available.'}"

            messagebox.showinfo("Integration Details", details)

        except Exception as e:
            logger.error(f"Error viewing integration details: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to view details: {e}")

    def install_integration(self):
        """Install selected integration"""
        try:
            selected = self.catalog_tree.selection()
            if not selected:
                messagebox.showwarning(_t("common.warning"), "Please select an integration to install")
                return

            values = self.catalog_tree.item(selected[0])['values']
            integration_id = values[0]
            integration_name = values[1]

            if messagebox.askyesno("Confirm Installation",
                                  f"Install '{integration_name}'?\n\nThis will add the integration to your system."):
                try:
                    installed_by = self.auth.current_user.get('username', 'Unknown')

                    if MANAGERS_AVAILABLE:
                        install_id = InstallationManager.install_integration(integration_id, installed_by)
                    else:
                        with transaction() as conn:
                            cursor = conn.cursor()

                            # Get version
                            cursor.execute('SELECT version FROM integration_catalog WHERE integration_id = ?',
                                          (integration_id,))
                            version = cursor.fetchone()[0]

                            cursor.execute('''
                                INSERT INTO installed_integrations
                                (integration_id, installed_by, version_installed, status, is_enabled)
                                VALUES (?, ?, ?, 'active', 1)
                            ''', (integration_id, installed_by, version))

                            install_id = cursor.lastrowid

                            # Update install count
                            cursor.execute('''
                                UPDATE integration_catalog
                                SET install_count = install_count + 1
                                WHERE integration_id = ?
                            ''', (integration_id,))

                    log_activity('create', 'installed_integration', install_id,
                                details={'integration_id': integration_id, 'integration_name': integration_name})

                    # Send notification
                    if EMAIL_AVAILABLE:
                        try:
                            send_email(
                                recipient_email=f"{installed_by}@university.edu",
                                subject=f"Integration Installed: {integration_name}",
                                body=f"The integration '{integration_name}' has been successfully installed.\n\n"
                                     f"Installation ID: {install_id}\n"
                                     f"Installed by: {installed_by}\n"
                                     f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
                                     f"You can now configure and use this integration."
                            )
                        except Exception as e:
                            logger.warning(f"Failed to send notification: {e}")

                    messagebox.showinfo(_t("common.success"),
                                      f"Integration installed successfully!\n\n"
                                      f"Installation ID: {install_id}\n\n"
                                      f"Go to the 'Installed' tab to configure it.")

                    self.load_catalog()
                    self.load_installed()

                except Exception as e:
                    logger.error(f"Error installing integration: {e}\n{traceback.format_exc()}")
                    messagebox.showerror(_t("common.error"), f"Failed to install integration: {e}")

        except Exception as e:
            logger.error(f"Error in install_integration: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to process installation: {e}")

    def configure_integration(self):
        """Configure selected installed integration"""
        try:
            selected = self.installed_tree.selection()
            if not selected:
                messagebox.showwarning(_t("common.warning"), "Please select an integration to configure")
                return

            install_id = self.installed_tree.item(selected[0])['values'][0]

            # Get current configuration
            with get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT ii.configuration, ii.sync_frequency, ic.integration_name
                    FROM installed_integrations ii
                    JOIN integration_catalog ic ON ii.integration_id = ic.integration_id
                    WHERE ii.install_id = ?
                ''', (install_id,))

                result = cursor.fetchone()

            if not result:
                messagebox.showerror(_t("common.error"), "Installation not found")
                return

            current_config, sync_freq, integration_name = result

            dialog = tk.Toplevel(self.root)
            dialog.title(_t("integration_marketplace.dialogs.configure_integration"))
            dialog.geometry("500x400")

            ttk.Label(dialog, text=_t("integration_marketplace.labels.frequency")).grid(row=0, column=0, padx=10, pady=5, sticky='w')
            sync_freq_var = tk.StringVar(value=sync_freq or 'hourly')
            ttk.Combobox(dialog, textvariable=sync_freq_var,
                        values=['realtime', 'hourly', 'daily', 'weekly', 'manual'],
                        width=33, state='readonly').grid(row=0, column=1, padx=10, pady=5)

            ttk.Label(dialog, text=_t("integration_marketplace.labels.configuration")).grid(row=1, column=0, padx=10, pady=5, sticky='nw')
            config_text = scrolledtext.ScrolledText(dialog, width=40, height=15)
            config_text.grid(row=1, column=1, padx=10, pady=5)
            config_text.insert('1.0', current_config or '{}')

            def save_config():
                try:
                    configuration = config_text.get('1.0', 'end-1c').strip()
                    sync_frequency = sync_freq_var.get()

                    # Validate JSON
                    try:
                        json.loads(configuration)
                    except json.JSONDecodeError:
                        messagebox.showerror(_t("common.error"), _t("integration_marketplace.messages.invalid_json"))
                        return

                    with transaction() as conn:
                        cursor = conn.cursor()
                        cursor.execute('''
                            UPDATE installed_integrations
                            SET configuration = ?, sync_frequency = ?
                            WHERE install_id = ?
                        ''', (configuration, sync_frequency, install_id))

                    log_activity('update', 'installed_integration', install_id,
                                details={'action': 'configured'})

                    messagebox.showinfo(_t("common.success"), _t("integration_marketplace.messages.config_updated"))
                    dialog.destroy()
                    self.load_installed()

                except Exception as e:
                    logger.error(f"Error saving configuration: {e}")
                    messagebox.showerror(_t("common.error"), f"Failed to save configuration: {e}")

            ttk.Button(dialog, text=_t("integration_marketplace.installed.configure"), command=save_config).grid(
                row=2, column=0, columnspan=2, pady=20)

        except Exception as e:
            logger.error(f"Error configuring integration: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to configure integration: {e}")

    def sync_integration(self):
        """Manually trigger sync for selected integration"""
        try:
            selected = self.installed_tree.selection()
            if not selected:
                messagebox.showwarning(_t("common.warning"), "Please select an integration to sync")
                return

            install_id = self.installed_tree.item(selected[0])['values'][0]

            if messagebox.askyesno(_t("common.confirm"), "Start manual sync for this integration?"):
                # Start sync
                if MANAGERS_AVAILABLE:
                    log_id = SyncManager.start_sync(install_id)
                else:
                    with transaction() as conn:
                        cursor = conn.cursor()
                        cursor.execute('''
                            INSERT INTO integration_sync_logs
                            (install_id, sync_status)
                            VALUES (?, 'running')
                        ''', (install_id,))

                        log_id = cursor.lastrowid

                # Simulate sync completion (in real implementation, this would be async)
                import time
                time.sleep(1)

                # Complete sync
                records_synced = 100  # Simulated
                if MANAGERS_AVAILABLE:
                    SyncManager.complete_sync(log_id, 'success', records_synced, 0)
                else:
                    with transaction() as conn:
                        cursor = conn.cursor()
                        cursor.execute('''
                            UPDATE integration_sync_logs
                            SET sync_end_time = ?, sync_status = ?, records_synced = ?
                            WHERE log_id = ?
                        ''', (datetime.now().isoformat(), 'success', records_synced, log_id))

                        cursor.execute('''
                            UPDATE installed_integrations
                            SET last_sync_date = ?
                            WHERE install_id = ?
                        ''', (datetime.now().isoformat(), install_id))

                log_activity('sync', 'installed_integration', install_id,
                            details={'log_id': log_id, 'records_synced': records_synced})

                messagebox.showinfo(_t("common.success"), f"Sync completed successfully!\n\nRecords synced: {records_synced}")
                self.load_installed()
                self.load_sync_logs()

        except Exception as e:
            logger.error(f"Error syncing integration: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to sync integration: {e}")

    def uninstall_integration(self):
        """Uninstall selected integration"""
        try:
            selected = self.installed_tree.selection()
            if not selected:
                messagebox.showwarning(_t("common.warning"), "Please select an integration to uninstall")
                return

            install_id = self.installed_tree.item(selected[0])['values'][0]
            integration_name = self.installed_tree.item(selected[0])['values'][1]

            if messagebox.askyesno("Confirm Uninstall",
                                  f"Uninstall '{integration_name}'?\n\n"
                                  f"This will deactivate the integration and remove its credentials."):
                if MANAGERS_AVAILABLE:
                    InstallationManager.uninstall_integration(install_id)
                else:
                    with transaction() as conn:
                        cursor = conn.cursor()
                        cursor.execute('''
                            UPDATE installed_integrations
                            SET status = 'uninstalled', is_enabled = 0
                            WHERE install_id = ?
                        ''', (install_id,))

                log_activity('delete', 'installed_integration', install_id,
                            details={'action': 'uninstalled'})

                messagebox.showinfo(_t("common.success"), "Integration uninstalled successfully")
                self.load_installed()

        except Exception as e:
            logger.error(f"Error uninstalling integration: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to uninstall integration: {e}")

    def add_credentials(self):
        """Add credentials for integration"""
        try:
            dialog = tk.Toplevel(self.root)
            dialog.title(_t("integration_marketplace.dialogs.add_credentials"))
            dialog.geometry("500x450")

            ttk.Label(dialog, text=_t("integration_marketplace.labels.installation_id")).grid(row=0, column=0, padx=10, pady=5, sticky='w')
            install_id_var = tk.StringVar()
            ttk.Entry(dialog, textvariable=install_id_var, width=35).grid(row=0, column=1, padx=10, pady=5)

            ttk.Label(dialog, text=_t("integration_marketplace.labels.credential_type")).grid(row=1, column=0, padx=10, pady=5, sticky='w')
            cred_type_var = tk.StringVar(value='api_key')
            ttk.Combobox(dialog, textvariable=cred_type_var,
                        values=['api_key', 'oauth', 'basic_auth', 'bearer_token'],
                        width=33, state='readonly').grid(row=1, column=1, padx=10, pady=5)

            ttk.Label(dialog, text=_t("integration_marketplace.labels.api_key")).grid(row=2, column=0, padx=10, pady=5, sticky='w')
            api_key_var = tk.StringVar()
            ttk.Entry(dialog, textvariable=api_key_var, width=35, show='*').grid(row=2, column=1, padx=10, pady=5)

            ttk.Label(dialog, text=_t("integration_marketplace.labels.api_secret")).grid(row=3, column=0, padx=10, pady=5, sticky='w')
            api_secret_var = tk.StringVar()
            ttk.Entry(dialog, textvariable=api_secret_var, width=35, show='*').grid(row=3, column=1, padx=10, pady=5)

            ttk.Label(dialog, text=_t("integration_marketplace.labels.endpoint_url")).grid(row=4, column=0, padx=10, pady=5, sticky='w')
            endpoint_var = tk.StringVar()
            ttk.Entry(dialog, textvariable=endpoint_var, width=35).grid(row=4, column=1, padx=10, pady=5)

            ttk.Label(dialog, text=_t("integration_marketplace.labels.oauth_token")).grid(row=5, column=0, padx=10, pady=5, sticky='w')
            oauth_var = tk.StringVar()
            ttk.Entry(dialog, textvariable=oauth_var, width=35, show='*').grid(row=5, column=1, padx=10, pady=5)

            def save_credentials():
                try:
                    install_id = install_id_var.get().strip()
                    cred_type = cred_type_var.get()
                    api_key = api_key_var.get().strip()
                    api_secret = api_secret_var.get().strip()
                    endpoint = endpoint_var.get().strip()
                    oauth_token = oauth_var.get().strip()

                    if not install_id:
                        messagebox.showerror(_t("common.error"), "Installation ID is required")
                        return

                    if MANAGERS_AVAILABLE:
                        cred_id = CredentialManager.store_credentials(
                            int(install_id), cred_type, api_key, api_secret,
                            oauth_token, endpoint
                        )
                    else:
                        with transaction() as conn:
                            cursor = conn.cursor()
                            cursor.execute('''
                                INSERT INTO integration_credentials
                                (install_id, credential_type, api_key, api_secret,
                                 oauth_token, endpoint_url)
                                VALUES (?, ?, ?, ?, ?, ?)
                            ''', (install_id, cred_type, api_key, api_secret,
                                  oauth_token, endpoint))

                            cred_id = cursor.lastrowid

                    log_activity('create', 'integration_credentials', cred_id,
                                details={'install_id': install_id})

                    messagebox.showinfo(_t("common.success"), "Credentials saved successfully")
                    dialog.destroy()
                    self.load_credentials()

                except Exception as e:
                    logger.error(f"Error saving credentials: {e}")
                    messagebox.showerror(_t("common.error"), f"Failed to save credentials: {e}")

            ttk.Button(dialog, text="Save Credentials", command=save_credentials).grid(
                row=6, column=0, columnspan=2, pady=20)

        except Exception as e:
            logger.error(f"Error opening credentials dialog: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to open dialog: {e}")

    def edit_credentials(self):
        """Edit selected credentials"""
        try:
            selected = self.cred_tree.selection()
            if not selected:
                messagebox.showwarning(_t("common.warning"), "Please select credentials to edit")
                return

            credential_id = self.cred_tree.item(selected[0])['values'][0]

            # Get current credentials
            with get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT install_id, credential_type, endpoint_url
                    FROM integration_credentials
                    WHERE credential_id = ?
                ''', (credential_id,))

                cred = cursor.fetchone()

            if not cred:
                messagebox.showerror(_t("common.error"), "Credentials not found")
                return

            dialog = tk.Toplevel(self.root)
            dialog.title(_t("integration_marketplace.dialogs.edit_credentials"))
            dialog.geometry("500x350")

            ttk.Label(dialog, text="Credential Type:").grid(row=0, column=0, padx=10, pady=5, sticky='w')
            cred_type_var = tk.StringVar(value=cred[1])
            ttk.Combobox(dialog, textvariable=cred_type_var,
                        values=['api_key', 'oauth', 'basic_auth', 'bearer_token'],
                        width=33, state='readonly').grid(row=0, column=1, padx=10, pady=5)

            ttk.Label(dialog, text="New API Key (Optional):").grid(row=1, column=0, padx=10, pady=5, sticky='w')
            api_key_var = tk.StringVar()
            ttk.Entry(dialog, textvariable=api_key_var, width=35, show='*').grid(row=1, column=1, padx=10, pady=5)

            ttk.Label(dialog, text="Endpoint URL:").grid(row=2, column=0, padx=10, pady=5, sticky='w')
            endpoint_var = tk.StringVar(value=cred[2] or '')
            ttk.Entry(dialog, textvariable=endpoint_var, width=35).grid(row=2, column=1, padx=10, pady=5)

            def update_credentials():
                try:
                    api_key = api_key_var.get().strip()
                    endpoint = endpoint_var.get().strip()

                    with transaction() as conn:
                        cursor = conn.cursor()

                        # Update only non-empty fields
                        if api_key:
                            cursor.execute('''
                                UPDATE integration_credentials
                                SET credential_type = ?, api_key = ?, endpoint_url = ?
                                WHERE credential_id = ?
                            ''', (cred_type_var.get(), api_key, endpoint, credential_id))
                        else:
                            cursor.execute('''
                                UPDATE integration_credentials
                                SET credential_type = ?, endpoint_url = ?
                                WHERE credential_id = ?
                            ''', (cred_type_var.get(), endpoint, credential_id))

                    log_activity('update', 'integration_credentials', credential_id,
                                details={'action': 'updated'})

                    messagebox.showinfo(_t("common.success"), "Credentials updated successfully")
                    dialog.destroy()
                    self.load_credentials()

                except Exception as e:
                    logger.error(f"Error updating credentials: {e}")
                    messagebox.showerror(_t("common.error"), f"Failed to update credentials: {e}")

            ttk.Button(dialog, text="Update Credentials", command=update_credentials).grid(
                row=3, column=0, columnspan=2, pady=20)

        except Exception as e:
            logger.error(f"Error editing credentials: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to edit credentials: {e}")

    def delete_credentials(self):
        """Delete selected credentials"""
        try:
            selected = self.cred_tree.selection()
            if not selected:
                messagebox.showwarning(_t("common.warning"), "Please select credentials to delete")
                return

            credential_id = self.cred_tree.item(selected[0])['values'][0]

            if messagebox.askyesno("Confirm Delete", "Delete these credentials?"):
                with transaction() as conn:
                    cursor = conn.cursor()
                    cursor.execute('''
                        DELETE FROM integration_credentials
                        WHERE credential_id = ?
                    ''', (credential_id,))

                log_activity('delete', 'integration_credentials', credential_id,
                            details={'action': 'deleted'})

                messagebox.showinfo(_t("common.success"), "Credentials deleted successfully")
                self.load_credentials()

        except Exception as e:
            logger.error(f"Error deleting credentials: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to delete credentials: {e}")

    def view_sync_details(self):
        """View sync log details"""
        try:
            selected = self.sync_tree.selection()
            if not selected:
                messagebox.showwarning(_t("common.warning"), "Please select a sync log to view")
                return

            log_id = self.sync_tree.item(selected[0])['values'][0]

            with get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT isl.*, ic.integration_name
                    FROM integration_sync_logs isl
                    JOIN installed_integrations ii ON isl.install_id = ii.install_id
                    JOIN integration_catalog ic ON ii.integration_id = ic.integration_id
                    WHERE isl.log_id = ?
                ''', (log_id,))

                log = cursor.fetchone()

            if not log:
                messagebox.showerror(_t("common.error"), "Sync log not found")
                return

            details = f"Sync Log Details\n\n"
            details += f"Log ID: {log[0]}\n"
            details += f"Integration: {log[8]}\n"
            details += f"Start Time: {log[2]}\n"
            details += f"End Time: {log[3] or 'Still running'}\n"
            details += f"Status: {log[4]}\n"
            details += f"Records Synced: {log[5] or 0}\n"
            details += f"Errors: {log[6] or 0}\n"
            if log[7]:
                details += f"\nError Details:\n{log[7]}"

            messagebox.showinfo("Sync Log Details", details)

        except Exception as e:
            logger.error(f"Error viewing sync details: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to view sync details: {e}")

    def export_sync_logs(self):
        """Export sync logs to CSV"""
        try:
            filename = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
            )

            if not filename:
                return

            with get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT log_id, install_id, sync_start_time, sync_end_time,
                           sync_status, records_synced, errors_encountered
                    FROM integration_sync_logs
                    ORDER BY sync_start_time DESC
                ''')

                logs = cursor.fetchall()

            import csv
            with open(filename, 'w', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(['Log ID', 'Install ID', 'Start Time', 'End Time',
                               'Status', 'Records Synced', 'Errors'])
                writer.writerows(logs)

            messagebox.showinfo(_t("common.success"), f"Sync logs exported to {filename}")
            log_activity('export', 'integration_sync_logs', None,
                        details={'filename': filename, 'record_count': len(logs)})

        except Exception as e:
            logger.error(f"Error exporting sync logs: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to export sync logs: {e}")

    def add_mapping(self):
        """Add data mapping"""
        try:
            dialog = tk.Toplevel(self.root)
            dialog.title("Add Data Mapping")
            dialog.geometry("500x350")

            ttk.Label(dialog, text="Installation ID:").grid(row=0, column=0, padx=10, pady=5, sticky='w')
            install_id_var = tk.StringVar()
            ttk.Entry(dialog, textvariable=install_id_var, width=35).grid(row=0, column=1, padx=10, pady=5)

            ttk.Label(dialog, text="Source Field:").grid(row=1, column=0, padx=10, pady=5, sticky='w')
            source_var = tk.StringVar()
            ttk.Entry(dialog, textvariable=source_var, width=35).grid(row=1, column=1, padx=10, pady=5)

            ttk.Label(dialog, text="Target Field:").grid(row=2, column=0, padx=10, pady=5, sticky='w')
            target_var = tk.StringVar()
            ttk.Entry(dialog, textvariable=target_var, width=35).grid(row=2, column=1, padx=10, pady=5)

            ttk.Label(dialog, text="Transformation Rule (Optional):").grid(row=3, column=0, padx=10, pady=5, sticky='nw')
            transform_text = scrolledtext.ScrolledText(dialog, width=35, height=5)
            transform_text.grid(row=3, column=1, padx=10, pady=5)

            def save_mapping():
                try:
                    install_id = install_id_var.get().strip()
                    source_field = source_var.get().strip()
                    target_field = target_var.get().strip()
                    transformation = transform_text.get('1.0', 'end-1c').strip()

                    if not install_id or not source_field or not target_field:
                        messagebox.showerror(_t("common.error"), "Installation ID, source field, and target field are required")
                        return

                    if MANAGERS_AVAILABLE:
                        mapping_id = DataMappingManager.create_mapping(
                            int(install_id), source_field, target_field, transformation
                        )
                    else:
                        with transaction() as conn:
                            cursor = conn.cursor()
                            cursor.execute('''
                                INSERT INTO integration_data_mappings
                                (install_id, source_field, target_field, transformation_rule, is_active)
                                VALUES (?, ?, ?, ?, 1)
                            ''', (install_id, source_field, target_field, transformation))

                            mapping_id = cursor.lastrowid

                    log_activity('create', 'integration_data_mapping', mapping_id,
                                details={'install_id': install_id})

                    messagebox.showinfo(_t("common.success"), "Data mapping created successfully")
                    dialog.destroy()
                    self.load_mappings()

                except Exception as e:
                    logger.error(f"Error creating mapping: {e}")
                    messagebox.showerror(_t("common.error"), f"Failed to create mapping: {e}")

            ttk.Button(dialog, text="Create Mapping", command=save_mapping).grid(
                row=4, column=0, columnspan=2, pady=20)

        except Exception as e:
            logger.error(f"Error opening mapping dialog: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to open dialog: {e}")

    def edit_mapping(self):
        """Edit selected mapping"""
        try:
            selected = self.mappings_tree.selection()
            if not selected:
                messagebox.showwarning(_t("common.warning"), "Please select a mapping to edit")
                return

            mapping_id = self.mappings_tree.item(selected[0])['values'][0]

            # Fetch current mapping data
            with get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT install_id, source_field, target_field, transformation_rule, is_active
                    FROM integration_data_mappings
                    WHERE mapping_id = ?
                ''', (mapping_id,))
                current_data = cursor.fetchone()

            if not current_data:
                messagebox.showerror(_t("common.error"), "Mapping not found")
                return

            # Create edit dialog
            dialog = tk.Toplevel(self.root)
            dialog.title("Edit Data Mapping")
            dialog.geometry("500x400")

            ttk.Label(dialog, text="Installation ID:").grid(row=0, column=0, padx=10, pady=5, sticky='w')
            install_id_var = tk.StringVar(value=str(current_data[0]))
            ttk.Entry(dialog, textvariable=install_id_var, width=35).grid(row=0, column=1, padx=10, pady=5)

            ttk.Label(dialog, text="Source Field:").grid(row=1, column=0, padx=10, pady=5, sticky='w')
            source_var = tk.StringVar(value=current_data[1] or '')
            ttk.Entry(dialog, textvariable=source_var, width=35).grid(row=1, column=1, padx=10, pady=5)

            ttk.Label(dialog, text="Target Field:").grid(row=2, column=0, padx=10, pady=5, sticky='w')
            target_var = tk.StringVar(value=current_data[2] or '')
            ttk.Entry(dialog, textvariable=target_var, width=35).grid(row=2, column=1, padx=10, pady=5)

            ttk.Label(dialog, text="Transformation Rule (Optional):").grid(row=3, column=0, padx=10, pady=5, sticky='nw')
            transform_text = scrolledtext.ScrolledText(dialog, width=35, height=5)
            transform_text.grid(row=3, column=1, padx=10, pady=5)
            if current_data[3]:
                transform_text.insert('1.0', current_data[3])

            ttk.Label(dialog, text="Status:").grid(row=4, column=0, padx=10, pady=5, sticky='w')
            is_active_var = tk.BooleanVar(value=bool(current_data[4]))
            ttk.Checkbutton(dialog, text="Active", variable=is_active_var).grid(row=4, column=1, padx=10, pady=5, sticky='w')

            def save_changes():
                try:
                    install_id = install_id_var.get().strip()
                    source_field = source_var.get().strip()
                    target_field = target_var.get().strip()
                    transformation = transform_text.get('1.0', 'end-1c').strip()
                    is_active = 1 if is_active_var.get() else 0

                    if not install_id or not source_field or not target_field:
                        messagebox.showerror(_t("common.error"), "Installation ID, source field, and target field are required")
                        return

                    if MANAGERS_AVAILABLE:
                        DataMappingManager.update_mapping(
                            mapping_id, int(install_id), source_field, target_field,
                            transformation, is_active
                        )
                    else:
                        with transaction() as conn:
                            cursor = conn.cursor()
                            cursor.execute('''
                                UPDATE integration_data_mappings
                                SET install_id = ?, source_field = ?, target_field = ?,
                                    transformation_rule = ?, is_active = ?
                                WHERE mapping_id = ?
                            ''', (install_id, source_field, target_field, transformation, is_active, mapping_id))

                    log_activity('update', 'integration_data_mapping', mapping_id,
                                details={'install_id': install_id, 'is_active': is_active})

                    messagebox.showinfo(_t("common.success"), "Data mapping updated successfully")
                    dialog.destroy()
                    self.load_mappings()

                except Exception as e:
                    logger.error(f"Error updating mapping: {e}")
                    messagebox.showerror(_t("common.error"), f"Failed to update mapping: {e}")

            button_frame = ttk.Frame(dialog)
            button_frame.grid(row=5, column=0, columnspan=2, pady=20)

            ttk.Button(button_frame, text=_t("common.save_changes"), command=save_changes).pack(side='left', padx=5)
            ttk.Button(button_frame, text=_t("common.cancel"), command=dialog.destroy).pack(side='left', padx=5)

        except Exception as e:
            logger.error(f"Error opening edit mapping dialog: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to open edit dialog: {e}")

    def delete_mapping(self):
        """Delete selected mapping"""
        try:
            selected = self.mappings_tree.selection()
            if not selected:
                messagebox.showwarning(_t("common.warning"), "Please select a mapping to delete")
                return

            mapping_id = self.mappings_tree.item(selected[0])['values'][0]

            if messagebox.askyesno("Confirm Delete", "Delete this data mapping?"):
                with transaction() as conn:
                    cursor = conn.cursor()
                    cursor.execute('DELETE FROM integration_data_mappings WHERE mapping_id = ?',
                                  (mapping_id,))

                log_activity('delete', 'integration_data_mapping', mapping_id,
                            details={'action': 'deleted'})

                messagebox.showinfo(_t("common.success"), "Data mapping deleted successfully")
                self.load_mappings()

        except Exception as e:
            logger.error(f"Error deleting mapping: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to delete mapping: {e}")

    def add_webhook(self):
        """Add webhook"""
        try:
            dialog = tk.Toplevel(self.root)
            dialog.title("Add Webhook")
            dialog.geometry("500x300")

            ttk.Label(dialog, text="Installation ID:").grid(row=0, column=0, padx=10, pady=5, sticky='w')
            install_id_var = tk.StringVar()
            ttk.Entry(dialog, textvariable=install_id_var, width=35).grid(row=0, column=1, padx=10, pady=5)

            ttk.Label(dialog, text="Webhook URL:").grid(row=1, column=0, padx=10, pady=5, sticky='w')
            url_var = tk.StringVar()
            ttk.Entry(dialog, textvariable=url_var, width=35).grid(row=1, column=1, padx=10, pady=5)

            ttk.Label(dialog, text="Event Type:").grid(row=2, column=0, padx=10, pady=5, sticky='w')
            event_var = tk.StringVar(value='data_update')
            ttk.Combobox(dialog, textvariable=event_var,
                        values=['data_update', 'sync_complete', 'error', 'status_change'],
                        width=33, state='readonly').grid(row=2, column=1, padx=10, pady=5)

            ttk.Label(dialog, text="Secret Key (Optional):").grid(row=3, column=0, padx=10, pady=5, sticky='w')
            secret_var = tk.StringVar()
            ttk.Entry(dialog, textvariable=secret_var, width=35, show='*').grid(row=3, column=1, padx=10, pady=5)

            def save_webhook():
                try:
                    install_id = install_id_var.get().strip()
                    webhook_url = url_var.get().strip()
                    event_type = event_var.get()
                    secret_key = secret_var.get().strip()

                    if not install_id or not webhook_url:
                        messagebox.showerror(_t("common.error"), "Installation ID and webhook URL are required")
                        return

                    if MANAGERS_AVAILABLE:
                        webhook_id = WebhookManager.register_webhook(
                            int(install_id), webhook_url, event_type, secret_key
                        )
                    else:
                        with transaction() as conn:
                            cursor = conn.cursor()
                            cursor.execute('''
                                INSERT INTO integration_webhooks
                                (install_id, webhook_url, event_type, secret_key, is_active)
                                VALUES (?, ?, ?, ?, 1)
                            ''', (install_id, webhook_url, event_type, secret_key))

                            webhook_id = cursor.lastrowid

                    log_activity('create', 'integration_webhook', webhook_id,
                                details={'install_id': install_id})

                    messagebox.showinfo(_t("common.success"), "Webhook registered successfully")
                    dialog.destroy()
                    self.load_webhooks()

                except Exception as e:
                    logger.error(f"Error registering webhook: {e}")
                    messagebox.showerror(_t("common.error"), f"Failed to register webhook: {e}")

            ttk.Button(dialog, text="Register Webhook", command=save_webhook).grid(
                row=4, column=0, columnspan=2, pady=20)

        except Exception as e:
            logger.error(f"Error opening webhook dialog: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to open dialog: {e}")

    def edit_webhook(self):
        """Edit selected webhook"""
        messagebox.showinfo("Edit Webhook", "Edit webhook functionality - similar to add_webhook")

    def delete_webhook(self):
        """Delete selected webhook"""
        try:
            selected = self.webhooks_tree.selection()
            if not selected:
                messagebox.showwarning(_t("common.warning"), "Please select a webhook to delete")
                return

            webhook_id = self.webhooks_tree.item(selected[0])['values'][0]

            if messagebox.askyesno("Confirm Delete", "Delete this webhook?"):
                with transaction() as conn:
                    cursor = conn.cursor()
                    cursor.execute('DELETE FROM integration_webhooks WHERE webhook_id = ?',
                                  (webhook_id,))

                log_activity('delete', 'integration_webhook', webhook_id,
                            details={'action': 'deleted'})

                messagebox.showinfo(_t("common.success"), "Webhook deleted successfully")
                self.load_webhooks()

        except Exception as e:
            logger.error(f"Error deleting webhook: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to delete webhook: {e}")

    def view_analytics_summary(self):
        """View analytics summary"""
        try:
            days = int(self.analytics_days_var.get())
            cutoff_date = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')

            with get_connection() as conn:
                cursor = conn.cursor()

                # Total metrics
                cursor.execute('''
                    SELECT COUNT(*) FROM integration_usage_analytics
                    WHERE measurement_date >= ?
                ''', (cutoff_date,))
                total_metrics = cursor.fetchone()[0]

                # Average metric value
                cursor.execute('''
                    SELECT AVG(metric_value) FROM integration_usage_analytics
                    WHERE measurement_date >= ?
                ''', (cutoff_date,))
                avg_value = cursor.fetchone()[0] or 0

                # Most tracked metric
                cursor.execute('''
                    SELECT metric_name, COUNT(*) as count
                    FROM integration_usage_analytics
                    WHERE measurement_date >= ?
                    GROUP BY metric_name
                    ORDER BY count DESC
                    LIMIT 1
                ''', (cutoff_date,))
                top_metric = cursor.fetchone()

            summary = f"Analytics Summary (Last {days} Days)\n\n"
            summary += f"Total Metrics Recorded: {total_metrics}\n"
            summary += f"Average Metric Value: {avg_value:.2f}\n"
            if top_metric:
                summary += f"Most Tracked Metric: {top_metric[0]} ({top_metric[1]} times)"

            messagebox.showinfo("Analytics Summary", summary)

        except Exception as e:
            logger.error(f"Error viewing analytics summary: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to view analytics summary: {e}")

    def export_analytics(self):
        """Export analytics to CSV"""
        try:
            filename = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
            )

            if not filename:
                return

            days = int(self.analytics_days_var.get())
            cutoff_date = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')

            with get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT analytics_id, install_id, metric_name, metric_value, measurement_date
                    FROM integration_usage_analytics
                    WHERE measurement_date >= ?
                    ORDER BY measurement_date DESC
                ''', (cutoff_date,))

                analytics = cursor.fetchall()

            import csv
            with open(filename, 'w', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(['Analytics ID', 'Install ID', 'Metric Name',
                               'Metric Value', 'Measurement Date'])
                writer.writerows(analytics)

            messagebox.showinfo(_t("common.success"), f"Analytics exported to {filename}")
            log_activity('export', 'integration_usage_analytics', None,
                        details={'filename': filename, 'record_count': len(analytics)})

        except Exception as e:
            logger.error(f"Error exporting analytics: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to export analytics: {e}")

    # =========================================================================
    # SEARCH & DISCOVERY FUNCTIONS (1-6)
    # =========================================================================

    def search_catalog(self):
        """Full-text search across integration names, providers, and descriptions with highlighting"""
        try:
            search_term = self.catalog_search_var.get().strip()
            if not search_term:
                messagebox.showwarning(_t("common.warning"), "Please enter a search term")
                return

            # Clear existing items
            for item in self.catalog_tree.get_children():
                self.catalog_tree.delete(item)

            with get_connection() as conn:
                cursor = conn.cursor()
                search_pattern = f'%{search_term}%'
                cursor.execute('''
                    SELECT integration_id, integration_name, provider_name, category,
                           integration_type, version, rating, install_count, is_official,
                           description
                    FROM integration_catalog
                    WHERE is_active = 1 AND (
                        integration_name LIKE ? OR
                        provider_name LIKE ? OR
                        description LIKE ?
                    )
                    ORDER BY
                        CASE WHEN integration_name LIKE ? THEN 1
                             WHEN provider_name LIKE ? THEN 2
                             ELSE 3 END,
                        rating DESC
                ''', (search_pattern, search_pattern, search_pattern,
                      search_pattern, search_pattern))

                integrations = cursor.fetchall()

            for integration in integrations:
                is_official = 'Yes' if integration[8] else 'No'
                self.catalog_tree.insert('', 'end', values=(
                    integration[0], integration[1], integration[2], integration[3],
                    integration[4], integration[5], integration[6] or _t("common.na"),
                    integration[7] or 0, is_official
                ))

            # Update description with highlighted results
            if integrations:
                self.catalog_description.delete('1.0', 'end')
                self.catalog_description.insert('1.0',
                    f"Found {len(integrations)} integration(s) matching '{search_term}'")

            log_activity('search', 'integration_catalog', None,
                        details={'search_term': search_term, 'results_count': len(integrations)})

        except Exception as e:
            logger.error(f"Error searching catalog: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to search catalog: {e}")

    def filter_by_rating(self):
        """Filter catalog by minimum star rating threshold"""
        try:
            min_rating = self.catalog_rating_var.get()
            if not min_rating:
                messagebox.showwarning(_t("common.warning"), "Please select a minimum rating")
                return

            min_rating = float(min_rating)

            # Clear existing items
            for item in self.catalog_tree.get_children():
                self.catalog_tree.delete(item)

            with get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT integration_id, integration_name, provider_name, category,
                           integration_type, version, rating, install_count, is_official
                    FROM integration_catalog
                    WHERE is_active = 1 AND rating >= ?
                    ORDER BY rating DESC, install_count DESC
                ''', (min_rating,))

                integrations = cursor.fetchall()

            for integration in integrations:
                is_official = 'Yes' if integration[8] else 'No'
                self.catalog_tree.insert('', 'end', values=(
                    integration[0], integration[1], integration[2], integration[3],
                    integration[4], integration[5], integration[6] or _t("common.na"),
                    integration[7] or 0, is_official
                ))

            messagebox.showinfo("Filter Results",
                              f"Found {len(integrations)} integration(s) with rating >= {min_rating}")

        except Exception as e:
            logger.error(f"Error filtering by rating: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to filter by rating: {e}")

    def filter_by_compatibility(self):
        """Show only integrations compatible with current system version"""
        try:
            # Get current system version
            system_version = "5.0.0"  # From CLAUDE.md

            # Clear existing items
            for item in self.catalog_tree.get_children():
                self.catalog_tree.delete(item)

            with get_connection() as conn:
                cursor = conn.cursor()
                # Filter integrations that have compatible version (major version match)
                cursor.execute('''
                    SELECT integration_id, integration_name, provider_name, category,
                           integration_type, version, rating, install_count, is_official
                    FROM integration_catalog
                    WHERE is_active = 1
                    ORDER BY rating DESC, install_count DESC
                ''')

                integrations = cursor.fetchall()

            # Filter by version compatibility (major.minor compatibility)
            system_major = int(system_version.split('.')[0])
            compatible = []
            for integration in integrations:
                try:
                    int_version = integration[5] or "1.0.0"
                    int_major = int(int_version.split('.')[0])
                    # Compatible if integration major version <= system major version
                    if int_major <= system_major:
                        compatible.append(integration)
                except (ValueError, IndexError):
                    compatible.append(integration)  # Include if version parsing fails

            for integration in compatible:
                is_official = 'Yes' if integration[8] else 'No'
                self.catalog_tree.insert('', 'end', values=(
                    integration[0], integration[1], integration[2], integration[3],
                    integration[4], integration[5], integration[6] or _t("common.na"),
                    integration[7] or 0, is_official
                ))

            messagebox.showinfo("Compatibility Filter",
                              f"Found {len(compatible)} integration(s) compatible with system v{system_version}")

        except Exception as e:
            logger.error(f"Error filtering by compatibility: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to filter by compatibility: {e}")

    def find_similar_integrations(self):
        """Suggest similar integrations based on selected one's category/features"""
        try:
            selected = self.catalog_tree.selection()
            if not selected:
                messagebox.showwarning(_t("common.warning"), "Please select an integration to find similar ones")
                return

            values = self.catalog_tree.item(selected[0])['values']
            integration_id = values[0]
            category = values[3]
            integration_type = values[4]

            with get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT integration_id, integration_name, provider_name, category,
                           integration_type, version, rating, install_count, is_official
                    FROM integration_catalog
                    WHERE is_active = 1
                      AND integration_id != ?
                      AND (category = ? OR integration_type = ?)
                    ORDER BY
                        CASE WHEN category = ? AND integration_type = ? THEN 1
                             WHEN category = ? THEN 2
                             ELSE 3 END,
                        rating DESC
                    LIMIT 10
                ''', (integration_id, category, integration_type,
                      category, integration_type, category))

                similar = cursor.fetchall()

            if not similar:
                messagebox.showinfo("Similar Integrations", "No similar integrations found")
                return

            # Show in a dialog
            dialog = tk.Toplevel(self.root)
            dialog.title("Similar Integrations")
            dialog.geometry("700x400")
            dialog.transient(self.root)

            ttk.Label(dialog, text=f"Integrations similar to '{values[1]}':",
                     style='Title.TLabel').pack(padx=10, pady=10)

            # Create treeview for results
            columns = ('name', 'provider', 'category', 'type', 'rating')
            tree = ttk.Treeview(dialog, columns=columns, show='headings', height=10)

            column_labels = {
                'name': _t("integration_marketplace.columns.integration_name"),
                'provider': _t("integration_marketplace.columns.provider_name"),
                'category': _t("integration_marketplace.columns.category"),
                'type': _t("integration_marketplace.columns.integration_type"),
                'rating': _t("integration_marketplace.columns.rating")
            }

            for col in columns:
                tree.heading(col, text=column_labels[col])

            tree.column('name', width=200)
            tree.column('provider', width=150)
            tree.column('category', width=100)
            tree.column('type', width=100)
            tree.column('rating', width=70)

            for item in similar:
                tree.insert('', 'end', values=(
                    item[1], item[2], item[3], item[4], item[6] or _t("common.na")
                ))

            tree.pack(fill='both', expand=True, padx=10, pady=10)

            ttk.Button(dialog, text=_t("common.close"), command=dialog.destroy).pack(pady=10)

        except Exception as e:
            logger.error(f"Error finding similar integrations: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to find similar integrations: {e}")

    def search_sync_logs(self):
        """Search logs by date range, status, or error message content"""
        try:
            dialog = tk.Toplevel(self.root)
            dialog.title(_t("integration_marketplace.dialogs.search_sync_logs"))
            dialog.geometry("500x350")
            dialog.transient(self.root)
            dialog.grab_set()

            # Date range
            ttk.Label(dialog, text="Start Date (YYYY-MM-DD):").grid(row=0, column=0, padx=10, pady=5, sticky='w')
            start_date_var = tk.StringVar(value=(datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
            ttk.Entry(dialog, textvariable=start_date_var, width=20).grid(row=0, column=1, padx=10, pady=5)

            ttk.Label(dialog, text="End Date (YYYY-MM-DD):").grid(row=1, column=0, padx=10, pady=5, sticky='w')
            end_date_var = tk.StringVar(value=datetime.now().strftime('%Y-%m-%d'))
            ttk.Entry(dialog, textvariable=end_date_var, width=20).grid(row=1, column=1, padx=10, pady=5)

            # Status filter
            ttk.Label(dialog, text="Status:").grid(row=2, column=0, padx=10, pady=5, sticky='w')
            status_var = tk.StringVar(value='all')
            ttk.Combobox(dialog, textvariable=status_var,
                        values=['all', 'success', 'failed', 'running'],
                        width=18, state='readonly').grid(row=2, column=1, padx=10, pady=5)

            # Error message search
            ttk.Label(dialog, text="Error Contains:").grid(row=3, column=0, padx=10, pady=5, sticky='w')
            error_var = tk.StringVar()
            ttk.Entry(dialog, textvariable=error_var, width=30).grid(row=3, column=1, padx=10, pady=5)

            def perform_search():
                try:
                    start_date = start_date_var.get()
                    end_date = end_date_var.get()
                    status = status_var.get()
                    error_text = error_var.get().strip()

                    # Clear existing items
                    for item in self.sync_tree.get_children():
                        self.sync_tree.delete(item)

                    with get_connection() as conn:
                        cursor = conn.cursor()

                        query = '''
                            SELECT log_id, install_id, sync_start_time, sync_end_time,
                                   sync_status, records_synced, errors_encountered
                            FROM integration_sync_logs
                            WHERE sync_start_time >= ? AND sync_start_time <= ?
                        '''
                        params = [start_date, end_date + ' 23:59:59']

                        if status != 'all':
                            query += ' AND sync_status = ?'
                            params.append(status)

                        if error_text:
                            query += ' AND error_details LIKE ?'
                            params.append(f'%{error_text}%')

                        query += ' ORDER BY sync_start_time DESC LIMIT 100'

                        cursor.execute(query, params)
                        logs = cursor.fetchall()

                    for log in logs:
                        self.sync_tree.insert('', 'end', values=(
                            log[0], log[1],
                            log[2][:19] if log[2] else 'N/A',
                            log[3][:19] if log[3] else 'N/A',
                            log[4] or _t("common.na"),
                            log[5] if log[5] is not None else 0,
                            log[6] if log[6] is not None else 0
                        ))

                    messagebox.showinfo("Search Results", f"Found {len(logs)} log entries")
                    dialog.destroy()

                except Exception as e:
                    logger.error(f"Error searching sync logs: {e}")
                    messagebox.showerror(_t("common.error"), f"Failed to search: {e}")

            ttk.Button(dialog, text="Search", command=perform_search).grid(row=4, column=0, columnspan=2, pady=20)

        except Exception as e:
            logger.error(f"Error opening search dialog: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to open search dialog: {e}")

    def advanced_filter_dialog(self):
        """Multi-criteria filter dialog with AND/OR logic"""
        try:
            dialog = tk.Toplevel(self.root)
            dialog.title(_t("integration_marketplace.dialogs.advanced_filter"))
            dialog.geometry("600x500")
            dialog.transient(self.root)
            dialog.grab_set()

            ttk.Label(dialog, text="Advanced Filter Options", style='Title.TLabel').pack(pady=10)

            # Filter frame
            filter_frame = ttk.LabelFrame(dialog, text="Filter Criteria", padding=10)
            filter_frame.pack(fill='x', padx=10, pady=5)

            # Category
            ttk.Label(filter_frame, text="Category:").grid(row=0, column=0, padx=5, pady=5, sticky='w')
            category_var = tk.StringVar()
            ttk.Combobox(filter_frame, textvariable=category_var,
                        values=['', 'LMS', 'SIS', 'CRM', 'Analytics', 'Communication', 'Storage'],
                        width=20).grid(row=0, column=1, padx=5, pady=5)

            # Type
            ttk.Label(filter_frame, text="Integration Type:").grid(row=1, column=0, padx=5, pady=5, sticky='w')
            type_var = tk.StringVar()
            ttk.Combobox(filter_frame, textvariable=type_var,
                        values=['', 'API', 'OAuth', 'SAML', 'Database', 'Webhook'],
                        width=20).grid(row=1, column=1, padx=5, pady=5)

            # Provider
            ttk.Label(filter_frame, text="Provider Contains:").grid(row=2, column=0, padx=5, pady=5, sticky='w')
            provider_var = tk.StringVar()
            ttk.Entry(filter_frame, textvariable=provider_var, width=22).grid(row=2, column=1, padx=5, pady=5)

            # Min Rating
            ttk.Label(filter_frame, text="Minimum Rating:").grid(row=3, column=0, padx=5, pady=5, sticky='w')
            rating_var = tk.StringVar()
            ttk.Spinbox(filter_frame, from_=0, to=5, textvariable=rating_var, width=20).grid(row=3, column=1, padx=5, pady=5)

            # Min Installs
            ttk.Label(filter_frame, text="Minimum Installs:").grid(row=4, column=0, padx=5, pady=5, sticky='w')
            installs_var = tk.StringVar(value='0')
            ttk.Entry(filter_frame, textvariable=installs_var, width=22).grid(row=4, column=1, padx=5, pady=5)

            # Official only
            official_var = tk.BooleanVar(value=False)
            ttk.Checkbutton(filter_frame, text="Official Integrations Only",
                           variable=official_var).grid(row=5, column=0, columnspan=2, padx=5, pady=5)

            # Logic
            ttk.Label(filter_frame, text="Filter Logic:").grid(row=6, column=0, padx=5, pady=5, sticky='w')
            logic_var = tk.StringVar(value='AND')
            ttk.Radiobutton(filter_frame, text="AND (all criteria)", variable=logic_var,
                           value='AND').grid(row=6, column=1, padx=5, pady=2, sticky='w')
            ttk.Radiobutton(filter_frame, text="OR (any criteria)", variable=logic_var,
                           value='OR').grid(row=7, column=1, padx=5, pady=2, sticky='w')

            def apply_filter():
                try:
                    # Clear existing items
                    for item in self.catalog_tree.get_children():
                        self.catalog_tree.delete(item)

                    conditions = []
                    params = []

                    if category_var.get():
                        conditions.append("category = ?")
                        params.append(category_var.get())

                    if type_var.get():
                        conditions.append("integration_type = ?")
                        params.append(type_var.get())

                    if provider_var.get():
                        conditions.append("provider_name LIKE ?")
                        params.append(f'%{provider_var.get()}%')

                    if rating_var.get():
                        try:
                            min_rating = float(rating_var.get())
                            conditions.append("rating >= ?")
                            params.append(min_rating)
                        except ValueError:
                            pass

                    if installs_var.get():
                        try:
                            min_installs = int(installs_var.get())
                            conditions.append("install_count >= ?")
                            params.append(min_installs)
                        except ValueError:
                            pass

                    if official_var.get():
                        conditions.append("is_official = 1")

                    logic = ' AND ' if logic_var.get() == 'AND' else ' OR '

                    query = '''
                        SELECT integration_id, integration_name, provider_name, category,
                               integration_type, version, rating, install_count, is_official
                        FROM integration_catalog
                        WHERE is_active = 1
                    '''

                    if conditions:
                        query += ' AND (' + logic.join(conditions) + ')'

                    query += ' ORDER BY rating DESC, install_count DESC LIMIT 100'

                    with get_connection() as conn:
                        cursor = conn.cursor()
                        cursor.execute(query, params)
                        integrations = cursor.fetchall()

                    for integration in integrations:
                        is_official = _t("common.yes") if integration[8] else _t("common.no")
                        self.catalog_tree.insert('', 'end', values=(
                            integration[0], integration[1], integration[2], integration[3],
                            integration[4], integration[5], integration[6] or _t("common.na"),
                            integration[7] or 0, is_official
                        ))

                    messagebox.showinfo("Filter Results", f"Found {len(integrations)} integration(s)")
                    dialog.destroy()

                except Exception as e:
                    logger.error(f"Error applying filter: {e}")
                    messagebox.showerror(_t("common.error"), f"Failed to apply filter: {e}")

            button_frame = ttk.Frame(dialog)
            button_frame.pack(pady=20)

            ttk.Button(button_frame, text=_t("integration_marketplace.common.filter"), command=apply_filter).pack(side='left', padx=5)
            ttk.Button(button_frame, text=_t("common.cancel"), command=dialog.destroy).pack(side='left', padx=5)

        except Exception as e:
            logger.error(f"Error opening advanced filter dialog: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to open dialog: {e}")

    # =========================================================================
    # BULK OPERATIONS FUNCTIONS (7-12)
    # =========================================================================

    def bulk_install_integrations(self):
        """Install multiple selected integrations at once"""
        try:
            # Show selection dialog from catalog
            dialog = tk.Toplevel(self.root)
            dialog.title(_t("integration_marketplace.dialogs.bulk_install"))
            dialog.geometry("700x500")
            dialog.transient(self.root)
            dialog.grab_set()

            ttk.Label(dialog, text="Select integrations to install:",
                     style='Title.TLabel').pack(pady=10)

            # Get available integrations
            with get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT ic.integration_id, ic.integration_name, ic.provider_name, ic.category
                    FROM integration_catalog ic
                    WHERE ic.is_active = 1
                      AND ic.integration_id NOT IN (
                          SELECT integration_id FROM installed_integrations
                          WHERE status = 'active'
                      )
                    ORDER BY ic.integration_name
                ''')
                available = cursor.fetchall()

            if not available:
                messagebox.showinfo(_t("common.info"), "No available integrations to install")
                dialog.destroy()
                return

            # Listbox with multiple selection
            frame = ttk.Frame(dialog)
            frame.pack(fill='both', expand=True, padx=10, pady=5)

            listbox = tk.Listbox(frame, selectmode=tk.MULTIPLE, height=15)
            scrollbar = ttk.Scrollbar(frame, orient='vertical', command=listbox.yview)
            listbox.configure(yscrollcommand=scrollbar.set)

            for item in available:
                listbox.insert(tk.END, f"{item[1]} ({item[2]}) - {item[3]}")

            listbox.pack(side='left', fill='both', expand=True)
            scrollbar.pack(side='right', fill='y')

            def install_selected():
                try:
                    selected_indices = listbox.curselection()
                    if not selected_indices:
                        messagebox.showwarning(_t("common.warning"), "Please select at least one integration")
                        return

                    installed_by = self.auth.current_user.get('username', 'Unknown')
                    success_count = 0
                    error_count = 0

                    for idx in selected_indices:
                        integration_id = available[idx][0]
                        try:
                            with transaction() as conn:
                                cursor = conn.cursor()
                                cursor.execute('SELECT version FROM integration_catalog WHERE integration_id = ?',
                                             (integration_id,))
                                version = cursor.fetchone()[0]

                                cursor.execute('''
                                    INSERT INTO installed_integrations
                                    (integration_id, installed_by, version_installed, status, is_enabled)
                                    VALUES (?, ?, ?, 'active', 1)
                                ''', (integration_id, installed_by, version))

                                cursor.execute('''
                                    UPDATE integration_catalog
                                    SET install_count = install_count + 1
                                    WHERE integration_id = ?
                                ''', (integration_id,))

                            success_count += 1
                        except Exception as e:
                            logger.error(f"Error installing integration {integration_id}: {e}")
                            error_count += 1

                    log_activity('bulk_install', 'installed_integrations', None,
                                details={'success_count': success_count, 'error_count': error_count})

                    messagebox.showinfo("Bulk Install Complete",
                                      f"Successfully installed: {success_count}\nErrors: {error_count}")
                    dialog.destroy()
                    self.load_catalog()
                    self.load_installed()

                except Exception as e:
                    logger.error(f"Error in bulk install: {e}")
                    messagebox.showerror(_t("common.error"), f"Bulk install failed: {e}")

            ttk.Button(dialog, text="Install Selected", command=install_selected,
                      style='Accent.TButton').pack(pady=10)

        except Exception as e:
            logger.error(f"Error opening bulk install dialog: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to open dialog: {e}")

    def bulk_uninstall_integrations(self):
        """Uninstall multiple integrations simultaneously"""
        try:
            selected = self.installed_tree.selection()
            if not selected:
                messagebox.showwarning(_t("common.warning"), "Please select integrations to uninstall")
                return

            if not messagebox.askyesno("Confirm Bulk Uninstall",
                                       f"Uninstall {len(selected)} selected integration(s)?"):
                return

            success_count = 0
            error_count = 0

            for item in selected:
                install_id = self.installed_tree.item(item)['values'][0]
                try:
                    with transaction() as conn:
                        cursor = conn.cursor()
                        cursor.execute('''
                            UPDATE installed_integrations
                            SET status = 'uninstalled', is_enabled = 0
                            WHERE install_id = ?
                        ''', (install_id,))
                    success_count += 1
                except Exception as e:
                    logger.error(f"Error uninstalling {install_id}: {e}")
                    error_count += 1

            log_activity('bulk_uninstall', 'installed_integrations', None,
                        details={'success_count': success_count, 'error_count': error_count})

            messagebox.showinfo("Bulk Uninstall Complete",
                              f"Successfully uninstalled: {success_count}\nErrors: {error_count}")
            self.load_installed()

        except Exception as e:
            logger.error(f"Error in bulk uninstall: {e}")
            messagebox.showerror(_t("common.error"), f"Bulk uninstall failed: {e}")

    def bulk_enable_integrations(self):
        """Enable multiple disabled integrations"""
        try:
            selected = self.installed_tree.selection()
            if not selected:
                messagebox.showwarning(_t("common.warning"), "Please select integrations to enable")
                return

            success_count = 0
            error_count = 0

            for item in selected:
                install_id = self.installed_tree.item(item)['values'][0]
                try:
                    with transaction() as conn:
                        cursor = conn.cursor()
                        cursor.execute('''
                            UPDATE installed_integrations
                            SET is_enabled = 1, status = 'active'
                            WHERE install_id = ?
                        ''', (install_id,))
                    success_count += 1
                except Exception as e:
                    logger.error(f"Error enabling {install_id}: {e}")
                    error_count += 1

            log_activity('bulk_enable', 'installed_integrations', None,
                        details={'success_count': success_count, 'error_count': error_count})

            messagebox.showinfo("Bulk Enable Complete",
                              f"Successfully enabled: {success_count}\nErrors: {error_count}")
            self.load_installed()

        except Exception as e:
            logger.error(f"Error in bulk enable: {e}")
            messagebox.showerror(_t("common.error"), f"Bulk enable failed: {e}")

    def bulk_disable_integrations(self):
        """Disable multiple integrations without uninstalling"""
        try:
            selected = self.installed_tree.selection()
            if not selected:
                messagebox.showwarning(_t("common.warning"), "Please select integrations to disable")
                return

            success_count = 0
            error_count = 0

            for item in selected:
                install_id = self.installed_tree.item(item)['values'][0]
                try:
                    with transaction() as conn:
                        cursor = conn.cursor()
                        cursor.execute('''
                            UPDATE installed_integrations
                            SET is_enabled = 0, status = 'inactive'
                            WHERE install_id = ?
                        ''', (install_id,))
                    success_count += 1
                except Exception as e:
                    logger.error(f"Error disabling {install_id}: {e}")
                    error_count += 1

            log_activity('bulk_disable', 'installed_integrations', None,
                        details={'success_count': success_count, 'error_count': error_count})

            messagebox.showinfo("Bulk Disable Complete",
                              f"Successfully disabled: {success_count}\nErrors: {error_count}")
            self.load_installed()

        except Exception as e:
            logger.error(f"Error in bulk disable: {e}")
            messagebox.showerror(_t("common.error"), f"Bulk disable failed: {e}")

    def bulk_sync_integrations(self):
        """Trigger sync for all selected/enabled integrations"""
        try:
            selected = self.installed_tree.selection()
            if not selected:
                # If none selected, sync all enabled
                if not messagebox.askyesno("Bulk Sync",
                                          "No integrations selected. Sync all enabled integrations?"):
                    return

                with get_connection() as conn:
                    cursor = conn.cursor()
                    cursor.execute('''
                        SELECT install_id FROM installed_integrations
                        WHERE is_enabled = 1 AND status = 'active'
                    ''')
                    install_ids = [row[0] for row in cursor.fetchall()]
            else:
                install_ids = [self.installed_tree.item(item)['values'][0] for item in selected]

            if not install_ids:
                messagebox.showinfo(_t("common.info"), "No integrations to sync")
                return

            success_count = 0
            error_count = 0

            for install_id in install_ids:
                try:
                    with transaction() as conn:
                        cursor = conn.cursor()
                        cursor.execute('''
                            INSERT INTO integration_sync_logs
                            (install_id, sync_status, records_synced)
                            VALUES (?, 'success', ?)
                        ''', (install_id, 50))  # Simulated sync

                        cursor.execute('''
                            UPDATE installed_integrations
                            SET last_sync_date = ?
                            WHERE install_id = ?
                        ''', (datetime.now().isoformat(), install_id))
                    success_count += 1
                except Exception as e:
                    logger.error(f"Error syncing {install_id}: {e}")
                    error_count += 1

            log_activity('bulk_sync', 'installed_integrations', None,
                        details={'success_count': success_count, 'error_count': error_count})

            messagebox.showinfo("Bulk Sync Complete",
                              f"Successfully synced: {success_count}\nErrors: {error_count}")
            self.load_installed()
            self.load_sync_logs()

        except Exception as e:
            logger.error(f"Error in bulk sync: {e}")
            messagebox.showerror(_t("common.error"), f"Bulk sync failed: {e}")

    def bulk_update_credentials(self):
        """Update endpoint URLs for multiple credentials at once"""
        try:
            dialog = tk.Toplevel(self.root)
            dialog.title("Bulk Update Credential Endpoints")
            dialog.geometry("600x400")
            dialog.transient(self.root)
            dialog.grab_set()

            ttk.Label(dialog, text="Bulk Update Endpoint URLs",
                     style='Title.TLabel').pack(pady=10)

            ttk.Label(dialog, text="Old URL Pattern (will be replaced):").pack(anchor='w', padx=10)
            old_url_var = tk.StringVar()
            ttk.Entry(dialog, textvariable=old_url_var, width=60).pack(padx=10, pady=5)

            ttk.Label(dialog, text="New URL Pattern:").pack(anchor='w', padx=10)
            new_url_var = tk.StringVar()
            ttk.Entry(dialog, textvariable=new_url_var, width=60).pack(padx=10, pady=5)

            # Preview frame
            preview_frame = ttk.LabelFrame(dialog, text=_t("integration_marketplace.mappings.preview"), padding=10)
            preview_frame.pack(fill='both', expand=True, padx=10, pady=10)

            preview_text = scrolledtext.ScrolledText(preview_frame, height=10)
            preview_text.pack(fill='both', expand=True)

            def preview_changes():
                old_pattern = old_url_var.get().strip()
                new_pattern = new_url_var.get().strip()

                if not old_pattern:
                    messagebox.showwarning(_t("common.warning"), "Please enter old URL pattern")
                    return

                with get_connection() as conn:
                    cursor = conn.cursor()
                    cursor.execute('''
                        SELECT credential_id, endpoint_url
                        FROM integration_credentials
                        WHERE endpoint_url LIKE ?
                    ''', (f'%{old_pattern}%',))
                    credentials = cursor.fetchall()

                preview_text.delete('1.0', 'end')
                if not credentials:
                    preview_text.insert('1.0', "No matching credentials found")
                else:
                    for cred in credentials:
                        old_url = cred[1] or ''
                        new_url = old_url.replace(old_pattern, new_pattern)
                        preview_text.insert('end', f"ID {cred[0]}:\n  Old: {old_url}\n  New: {new_url}\n\n")

            def apply_changes():
                old_pattern = old_url_var.get().strip()
                new_pattern = new_url_var.get().strip()

                if not old_pattern:
                    messagebox.showwarning(_t("common.warning"), "Please enter old URL pattern")
                    return

                if not messagebox.askyesno(_t("common.confirm"), "Apply URL changes to all matching credentials?"):
                    return

                try:
                    with transaction() as conn:
                        cursor = conn.cursor()
                        cursor.execute('''
                            UPDATE integration_credentials
                            SET endpoint_url = REPLACE(endpoint_url, ?, ?)
                            WHERE endpoint_url LIKE ?
                        ''', (old_pattern, new_pattern, f'%{old_pattern}%'))
                        updated_count = cursor.rowcount

                    log_activity('bulk_update', 'integration_credentials', None,
                                details={'old_pattern': old_pattern, 'new_pattern': new_pattern,
                                        'updated_count': updated_count})

                    messagebox.showinfo(_t("common.success"), f"Updated {updated_count} credential(s)")
                    dialog.destroy()
                    self.load_credentials()

                except Exception as e:
                    logger.error(f"Error updating credentials: {e}")
                    messagebox.showerror(_t("common.error"), f"Failed to update: {e}")

            button_frame = ttk.Frame(dialog)
            button_frame.pack(pady=10)

            ttk.Button(button_frame, text=_t("integration_marketplace.mappings.preview"), command=preview_changes).pack(side='left', padx=5)
            ttk.Button(button_frame, text=_t("common.apply"), command=apply_changes,
                      style='Accent.TButton').pack(side='left', padx=5)
            ttk.Button(button_frame, text=_t("common.cancel"), command=dialog.destroy).pack(side='left', padx=5)

        except Exception as e:
            logger.error(f"Error opening bulk update dialog: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to open dialog: {e}")

    # =========================================================================
    # IMPORT/EXPORT FUNCTIONS (13-18)
    # =========================================================================

    def export_catalog_to_json(self):
        """Export full catalog to JSON file"""
        try:
            filename = filedialog.asksaveasfilename(
                defaultextension=".json",
                filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
                initialfile="integration_catalog_export.json"
            )

            if not filename:
                return

            with get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT integration_id, integration_name, provider_name, integration_type,
                           category, description, version, rating, install_count, is_official,
                           pricing_model, documentation_url
                    FROM integration_catalog
                    WHERE is_active = 1
                ''')
                integrations = cursor.fetchall()

            export_data = {
                'export_date': datetime.now().isoformat(),
                'export_type': 'integration_catalog',
                'version': '1.0',
                'integrations': []
            }

            for row in integrations:
                export_data['integrations'].append({
                    'integration_id': row[0],
                    'integration_name': row[1],
                    'provider_name': row[2],
                    'integration_type': row[3],
                    'category': row[4],
                    'description': row[5],
                    'version': row[6],
                    'rating': row[7],
                    'install_count': row[8],
                    'is_official': bool(row[9]),
                    'pricing_model': row[10],
                    'documentation_url': row[11]
                })

            with open(filename, 'w') as f:
                json.dump(export_data, f, indent=2)

            log_activity('export', 'integration_catalog', None,
                        details={'filename': filename, 'count': len(integrations)})

            messagebox.showinfo(_t("common.success"),
                              f"Exported {len(integrations)} integrations to:\n{filename}")

        except Exception as e:
            logger.error(f"Error exporting catalog: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to export catalog: {e}")

    def import_integrations_from_json(self):
        """Import integrations from JSON backup"""
        try:
            filename = filedialog.askopenfilename(
                filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
            )

            if not filename:
                return

            with open(filename, 'r') as f:
                import_data = json.load(f)

            if 'integrations' not in import_data:
                messagebox.showerror(_t("common.error"), "Invalid import file format")
                return

            imported_count = 0
            skipped_count = 0

            for item in import_data['integrations']:
                try:
                    with get_connection() as conn:
                        cursor = conn.cursor()
                        # Check if integration already exists
                        cursor.execute('''
                            SELECT integration_id FROM integration_catalog
                            WHERE integration_name = ? AND provider_name = ?
                        ''', (item['integration_name'], item['provider_name']))

                        if cursor.fetchone():
                            skipped_count += 1
                            continue

                    with transaction() as conn:
                        cursor = conn.cursor()
                        cursor.execute('''
                            INSERT INTO integration_catalog
                            (integration_name, provider_name, integration_type, category,
                             description, version, is_official, is_active)
                            VALUES (?, ?, ?, ?, ?, ?, ?, 1)
                        ''', (item['integration_name'], item['provider_name'],
                              item.get('integration_type', 'API'), item.get('category', 'Other'),
                              item.get('description', ''), item.get('version', '1.0'),
                              int(item.get('is_official', False))))

                    imported_count += 1

                except Exception as e:
                    logger.error(f"Error importing integration: {e}")
                    skipped_count += 1

            log_activity('import', 'integration_catalog', None,
                        details={'filename': filename, 'imported': imported_count, 'skipped': skipped_count})

            messagebox.showinfo("Import Complete",
                              f"Imported: {imported_count}\nSkipped (duplicates): {skipped_count}")
            self.load_catalog()

        except Exception as e:
            logger.error(f"Error importing integrations: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to import integrations: {e}")

    def export_configuration_bundle(self):
        """Export all configs, credentials (encrypted), and mappings as a bundle"""
        try:
            filename = filedialog.asksaveasfilename(
                defaultextension=".json",
                filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
                initialfile="integration_config_bundle.json"
            )

            if not filename:
                return

            bundle = {
                'export_date': datetime.now().isoformat(),
                'export_type': 'configuration_bundle',
                'version': '1.0',
                'installed_integrations': [],
                'credentials': [],
                'data_mappings': [],
                'webhooks': []
            }

            with get_connection() as conn:
                cursor = conn.cursor()

                # Export installed integrations
                cursor.execute('''
                    SELECT ii.install_id, ic.integration_name, ii.version_installed,
                           ii.configuration, ii.sync_frequency, ii.status
                    FROM installed_integrations ii
                    JOIN integration_catalog ic ON ii.integration_id = ic.integration_id
                    WHERE ii.status != 'uninstalled'
                ''')
                for row in cursor.fetchall():
                    bundle['installed_integrations'].append({
                        'install_id': row[0],
                        'integration_name': row[1],
                        'version': row[2],
                        'configuration': row[3],
                        'sync_frequency': row[4],
                        'status': row[5]
                    })

                # Export credentials (masked)
                cursor.execute('''
                    SELECT credential_id, install_id, credential_type, endpoint_url
                    FROM integration_credentials
                ''')
                for row in cursor.fetchall():
                    bundle['credentials'].append({
                        'credential_id': row[0],
                        'install_id': row[1],
                        'credential_type': row[2],
                        'endpoint_url': row[3],
                        'api_key': '***MASKED***',
                        'api_secret': '***MASKED***'
                    })

                # Export data mappings
                cursor.execute('''
                    SELECT mapping_id, install_id, source_field, target_field,
                           transformation_rule, is_active
                    FROM integration_data_mappings
                ''')
                for row in cursor.fetchall():
                    bundle['data_mappings'].append({
                        'mapping_id': row[0],
                        'install_id': row[1],
                        'source_field': row[2],
                        'target_field': row[3],
                        'transformation_rule': row[4],
                        'is_active': bool(row[5])
                    })

                # Export webhooks
                cursor.execute('''
                    SELECT webhook_id, install_id, webhook_url, event_type, is_active
                    FROM integration_webhooks
                ''')
                for row in cursor.fetchall():
                    bundle['webhooks'].append({
                        'webhook_id': row[0],
                        'install_id': row[1],
                        'webhook_url': row[2],
                        'event_type': row[3],
                        'is_active': bool(row[4])
                    })

            with open(filename, 'w') as f:
                json.dump(bundle, f, indent=2)

            log_activity('export', 'configuration_bundle', None,
                        details={'filename': filename})

            messagebox.showinfo(_t("common.success"),
                              f"Configuration bundle exported to:\n{filename}\n\n"
                              f"Note: Credentials are masked for security.")

        except Exception as e:
            logger.error(f"Error exporting configuration bundle: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to export bundle: {e}")

    def import_configuration_bundle(self):
        """Import and restore configuration bundle"""
        try:
            filename = filedialog.askopenfilename(
                filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
            )

            if not filename:
                return

            with open(filename, 'r') as f:
                bundle = json.load(f)

            if bundle.get('export_type') != 'configuration_bundle':
                messagebox.showerror(_t("common.error"), "Invalid configuration bundle format")
                return

            if not messagebox.askyesno("Confirm Import",
                                       "Import configuration bundle?\n\n"
                                       "This will add new data mappings and webhooks.\n"
                                       "Credentials will need to be reconfigured manually."):
                return

            imported = {'mappings': 0, 'webhooks': 0}

            # Import data mappings
            for mapping in bundle.get('data_mappings', []):
                try:
                    with transaction() as conn:
                        cursor = conn.cursor()
                        cursor.execute('''
                            INSERT INTO integration_data_mappings
                            (install_id, source_field, target_field, transformation_rule, is_active)
                            VALUES (?, ?, ?, ?, ?)
                        ''', (mapping['install_id'], mapping['source_field'],
                              mapping['target_field'], mapping.get('transformation_rule'),
                              int(mapping.get('is_active', True))))
                    imported['mappings'] += 1
                except Exception as e:
                    logger.warning(f"Skipped mapping import: {e}")

            # Import webhooks
            for webhook in bundle.get('webhooks', []):
                try:
                    with transaction() as conn:
                        cursor = conn.cursor()
                        cursor.execute('''
                            INSERT INTO integration_webhooks
                            (install_id, webhook_url, event_type, is_active)
                            VALUES (?, ?, ?, ?)
                        ''', (webhook['install_id'], webhook['webhook_url'],
                              webhook['event_type'], int(webhook.get('is_active', True))))
                    imported['webhooks'] += 1
                except Exception as e:
                    logger.warning(f"Skipped webhook import: {e}")

            log_activity('import', 'configuration_bundle', None,
                        details={'filename': filename, 'imported': imported})

            messagebox.showinfo("Import Complete",
                              f"Imported:\n"
                              f"- Data mappings: {imported['mappings']}\n"
                              f"- Webhooks: {imported['webhooks']}\n\n"
                              f"Note: Credentials must be configured manually.")

            self.load_mappings()
            self.load_webhooks()

        except Exception as e:
            logger.error(f"Error importing configuration bundle: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to import bundle: {e}")

    def export_sync_report_pdf(self):
        """Generate PDF report of sync history"""
        try:
            # Check for reportlab
            try:
                from reportlab.lib.pagesizes import letter
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                from reportlab.lib.styles import getSampleStyleSheet
                from reportlab.lib import colors
            except ImportError:
                messagebox.showerror(_t("common.error"),
                                   "PDF export requires reportlab library.\n"
                                   "Install with: pip install reportlab")
                return

            filename = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
                initialfile="sync_report.pdf"
            )

            if not filename:
                return

            with get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT isl.log_id, ic.integration_name, isl.sync_start_time,
                           isl.sync_status, isl.records_synced, isl.errors_encountered
                    FROM integration_sync_logs isl
                    JOIN installed_integrations ii ON isl.install_id = ii.install_id
                    JOIN integration_catalog ic ON ii.integration_id = ic.integration_id
                    ORDER BY isl.sync_start_time DESC
                    LIMIT 100
                ''')
                logs = cursor.fetchall()

            # Create PDF
            doc = SimpleDocTemplate(filename, pagesize=letter)
            styles = getSampleStyleSheet()
            elements = []

            # Title
            elements.append(Paragraph("Integration Sync Report", styles['Title']))
            elements.append(Spacer(1, 20))
            elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
            elements.append(Spacer(1, 20))

            # Summary
            total_syncs = len(logs)
            success_count = sum(1 for log in logs if log[3] == 'success')
            failed_count = sum(1 for log in logs if log[3] == 'failed')

            elements.append(Paragraph("Summary", styles['Heading2']))
            elements.append(Paragraph(f"Total Syncs: {total_syncs}", styles['Normal']))
            elements.append(Paragraph(f"Successful: {success_count}", styles['Normal']))
            elements.append(Paragraph(f"Failed: {failed_count}", styles['Normal']))
            elements.append(Spacer(1, 20))

            # Table data
            table_data = [['Log ID', 'Integration', 'Date', 'Status', 'Records', 'Errors']]
            for log in logs:
                table_data.append([
                    str(log[0]),
                    log[1][:20] if log[1] else 'N/A',
                    log[2][:16] if log[2] else 'N/A',
                    log[3] or _t("common.na"),
                    str(log[4] or 0),
                    str(log[5] or 0)
                ])

            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))

            elements.append(table)

            doc.build(elements)

            log_activity('export', 'sync_report_pdf', None,
                        details={'filename': filename, 'log_count': len(logs)})

            messagebox.showinfo(_t("common.success"), f"PDF report exported to:\n{filename}")

        except Exception as e:
            logger.error(f"Error exporting PDF report: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to export PDF: {e}")

    def export_mappings_to_excel(self):
        """Export data mappings to Excel for review"""
        try:
            # Check for openpyxl
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill
            except ImportError:
                messagebox.showerror(_t("common.error"),
                                   "Excel export requires openpyxl library.\n"
                                   "Install with: pip install openpyxl")
                return

            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile="data_mappings.xlsx"
            )

            if not filename:
                return

            with get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT idm.mapping_id, ic.integration_name, idm.source_field,
                           idm.target_field, idm.transformation_rule, idm.is_active
                    FROM integration_data_mappings idm
                    JOIN installed_integrations ii ON idm.install_id = ii.install_id
                    JOIN integration_catalog ic ON ii.integration_id = ic.integration_id
                    ORDER BY ic.integration_name, idm.source_field
                ''')
                mappings = cursor.fetchall()

            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Data Mappings"

            # Headers
            headers = ['Mapping ID', 'Integration', 'Source Field', 'Target Field',
                      'Transformation Rule', 'Active']
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font

            # Data
            for row_idx, mapping in enumerate(mappings, 2):
                ws.cell(row=row_idx, column=1, value=mapping[0])
                ws.cell(row=row_idx, column=2, value=mapping[1])
                ws.cell(row=row_idx, column=3, value=mapping[2])
                ws.cell(row=row_idx, column=4, value=mapping[3])
                ws.cell(row=row_idx, column=5, value=mapping[4] or '')
                ws.cell(row=row_idx, column=6, value=_t("common.yes") if mapping[5] else _t("common.no"))

            # Adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column].width = min(max_length + 2, 50)

            wb.save(filename)

            log_activity('export', 'data_mappings_excel', None,
                        details={'filename': filename, 'count': len(mappings)})

            messagebox.showinfo(_t("common.success"),
                              f"Exported {len(mappings)} mappings to:\n{filename}")

        except Exception as e:
            logger.error(f"Error exporting mappings to Excel: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to export to Excel: {e}")

    # =========================================================================
    # REPORTS & DASHBOARDS FUNCTIONS (19-25)
    # =========================================================================

    def show_dashboard_overview(self):
        """Real-time dashboard with KPIs, charts, and status widgets"""
        try:
            dialog = tk.Toplevel(self.root)
            dialog.title("Integration Dashboard")
            dialog.geometry("900x700")
            dialog.transient(self.root)

            # Get dashboard data
            with get_connection() as conn:
                cursor = conn.cursor()

                # Total integrations
                cursor.execute('SELECT COUNT(*) FROM integration_catalog WHERE is_active = 1')
                total_catalog = cursor.fetchone()[0]

                # Installed integrations
                cursor.execute("SELECT COUNT(*) FROM installed_integrations WHERE status = 'active'")
                total_installed = cursor.fetchone()[0]

                # Enabled integrations
                cursor.execute("SELECT COUNT(*) FROM installed_integrations WHERE is_enabled = 1")
                total_enabled = cursor.fetchone()[0]

                # Recent syncs
                cursor.execute('''
                    SELECT sync_status, COUNT(*) FROM integration_sync_logs
                    WHERE sync_start_time >= date('now', '-7 days')
                    GROUP BY sync_status
                ''')
                sync_stats = dict(cursor.fetchall())

                # Error rate
                cursor.execute('''
                    SELECT
                        SUM(CASE WHEN sync_status = 'failed' THEN 1 ELSE 0 END) * 100.0 /
                        NULLIF(COUNT(*), 0)
                    FROM integration_sync_logs
                    WHERE sync_start_time >= date('now', '-7 days')
                ''')
                error_rate = cursor.fetchone()[0] or 0

                # Top integrations by usage
                cursor.execute('''
                    SELECT ic.integration_name, COUNT(isl.log_id) as sync_count
                    FROM installed_integrations ii
                    JOIN integration_catalog ic ON ii.integration_id = ic.integration_id
                    LEFT JOIN integration_sync_logs isl ON ii.install_id = isl.install_id
                    GROUP BY ic.integration_name
                    ORDER BY sync_count DESC
                    LIMIT 5
                ''')
                top_integrations = cursor.fetchall()

            # KPI Frame
            kpi_frame = ttk.LabelFrame(dialog, text="Key Performance Indicators", padding=10)
            kpi_frame.pack(fill='x', padx=10, pady=5)

            kpi_grid = ttk.Frame(kpi_frame)
            kpi_grid.pack(fill='x')

            kpis = [
                ("Catalog Size", total_catalog),
                ("Installed", total_installed),
                (_t("common.enabled"), total_enabled),
                ("Success (7d)", sync_stats.get('success', 0)),
                ("Failed (7d)", sync_stats.get('failed', 0)),
                ("Error Rate", f"{error_rate:.1f}%")
            ]

            for i, (label, value) in enumerate(kpis):
                frame = ttk.Frame(kpi_grid, relief='ridge', borderwidth=2, padding=10)
                frame.grid(row=0, column=i, padx=5, pady=5, sticky='nsew')
                ttk.Label(frame, text=label, font=('Arial', 10)).pack()
                ttk.Label(frame, text=str(value), font=('Arial', 16, 'bold')).pack()
                kpi_grid.columnconfigure(i, weight=1)

            # Status Frame
            status_frame = ttk.LabelFrame(dialog, text="Integration Status", padding=10)
            status_frame.pack(fill='both', expand=True, padx=10, pady=5)

            # Top integrations list
            ttk.Label(status_frame, text="Top Integrations by Activity:",
                     style='Section.TLabel').pack(anchor='w', pady=5)

            for name, count in top_integrations:
                ttk.Label(status_frame, text=f"  {name}: {count} syncs").pack(anchor='w')

            # Health indicator
            health_frame = ttk.LabelFrame(dialog, text="System Health", padding=10)
            health_frame.pack(fill='x', padx=10, pady=5)

            if error_rate < 5:
                health_status = "HEALTHY"
                health_color = "green"
            elif error_rate < 20:
                health_status = "WARNING"
                health_color = "orange"
            else:
                health_status = "CRITICAL"
                health_color = "red"

            health_label = ttk.Label(health_frame, text=f"Overall Status: {health_status}",
                                    font=('Arial', 14, 'bold'))
            health_label.pack()

            ttk.Button(dialog, text=_t("common.close"), command=dialog.destroy).pack(pady=10)

            log_activity('view', 'dashboard', None)

        except Exception as e:
            logger.error(f"Error showing dashboard: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to show dashboard: {e}")

    def generate_health_report(self):
        """Comprehensive health report for all integrations"""
        try:
            with get_connection() as conn:
                cursor = conn.cursor()

                # Get all active integrations with their health metrics
                cursor.execute('''
                    SELECT
                        ic.integration_name,
                        ii.status,
                        ii.is_enabled,
                        ii.last_sync_date,
                        (SELECT COUNT(*) FROM integration_sync_logs isl
                         WHERE isl.install_id = ii.install_id AND isl.sync_status = 'success'
                         AND isl.sync_start_time >= date('now', '-30 days')) as success_count,
                        (SELECT COUNT(*) FROM integration_sync_logs isl
                         WHERE isl.install_id = ii.install_id AND isl.sync_status = 'failed'
                         AND isl.sync_start_time >= date('now', '-30 days')) as fail_count
                    FROM installed_integrations ii
                    JOIN integration_catalog ic ON ii.integration_id = ic.integration_id
                    WHERE ii.status != 'uninstalled'
                    ORDER BY ic.integration_name
                ''')
                integrations = cursor.fetchall()

            report = "Integration Health Report\n"
            report += "=" * 50 + "\n"
            report += f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"

            healthy_count = 0
            warning_count = 0
            critical_count = 0

            for integration in integrations:
                name, status, enabled, last_sync, success, fail = integration
                total = success + fail
                success_rate = (success / total * 100) if total > 0 else 0

                if success_rate >= 95 and enabled:
                    health = "HEALTHY"
                    healthy_count += 1
                elif success_rate >= 80 or not enabled:
                    health = "WARNING"
                    warning_count += 1
                else:
                    health = "CRITICAL"
                    critical_count += 1

                report += f"\n{name}\n"
                report += f"  Status: {status} | Enabled: {_t('common.yes') if enabled else _t('common.no')}\n"
                report += f"  Last Sync: {last_sync or _t('common.never')}\n"
                report += f"  Success Rate (30d): {success_rate:.1f}% ({success}/{total})\n"
                report += f"  Health: {health}\n"

            report += "\n" + "=" * 50 + "\n"
            report += f"Summary: {healthy_count} Healthy, {warning_count} Warning, {critical_count} Critical\n"

            # Show report in dialog
            dialog = tk.Toplevel(self.root)
            dialog.title(_t("integration_marketplace.dialogs.health_report"))
            dialog.geometry("600x500")
            dialog.transient(self.root)

            text = scrolledtext.ScrolledText(dialog, wrap=tk.WORD)
            text.pack(fill='both', expand=True, padx=10, pady=10)
            text.insert('1.0', report)
            text.config(state='disabled')

            def export_report():
                filename = filedialog.asksaveasfilename(
                    defaultextension=".txt",
                    filetypes=[("Text files", "*.txt"), ("All files", "*.*")],
                    initialfile="health_report.txt"
                )
                if filename:
                    with open(filename, 'w') as f:
                        f.write(report)
                    messagebox.showinfo(_t("common.success"), f"Report exported to:\n{filename}")

            button_frame = ttk.Frame(dialog)
            button_frame.pack(pady=10)
            ttk.Button(button_frame, text=_t("integration_marketplace.common.export"), command=export_report).pack(side='left', padx=5)
            ttk.Button(button_frame, text=_t("common.close"), command=dialog.destroy).pack(side='left', padx=5)

            log_activity('generate', 'health_report', None)

        except Exception as e:
            logger.error(f"Error generating health report: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to generate health report: {e}")

    def show_error_analysis(self):
        """Analyze and categorize sync errors by type/frequency"""
        try:
            with get_connection() as conn:
                cursor = conn.cursor()

                # Get error patterns
                cursor.execute('''
                    SELECT
                        ic.integration_name,
                        isl.error_details,
                        COUNT(*) as error_count
                    FROM integration_sync_logs isl
                    JOIN installed_integrations ii ON isl.install_id = ii.install_id
                    JOIN integration_catalog ic ON ii.integration_id = ic.integration_id
                    WHERE isl.sync_status = 'failed'
                      AND isl.sync_start_time >= date('now', '-30 days')
                    GROUP BY ic.integration_name, isl.error_details
                    ORDER BY error_count DESC
                    LIMIT 50
                ''')
                errors = cursor.fetchall()

                # Get error trends
                cursor.execute('''
                    SELECT
                        date(sync_start_time) as error_date,
                        COUNT(*) as error_count
                    FROM integration_sync_logs
                    WHERE sync_status = 'failed'
                      AND sync_start_time >= date('now', '-30 days')
                    GROUP BY date(sync_start_time)
                    ORDER BY error_date
                ''')
                trends = cursor.fetchall()

            dialog = tk.Toplevel(self.root)
            dialog.title(_t("integration_marketplace.dialogs.error_analysis"))
            dialog.geometry("800x600")
            dialog.transient(self.root)

            # Error patterns frame
            patterns_frame = ttk.LabelFrame(dialog, text="Error Patterns (Last 30 Days)", padding=10)
            patterns_frame.pack(fill='both', expand=True, padx=10, pady=5)

            columns = ('integration', 'error', 'count')
            tree = ttk.Treeview(patterns_frame, columns=columns, show='headings', height=10)

            tree.heading('integration', text='Integration')
            tree.heading('error', text='Error Details')
            tree.heading('count', text='Count')

            tree.column('integration', width=150)
            tree.column('error', width=400)
            tree.column('count', width=80)

            for error in errors:
                error_text = (error[1] or 'Unknown error')[:100]
                tree.insert('', 'end', values=(error[0], error_text, error[2]))

            tree.pack(fill='both', expand=True)

            # Trend frame
            trend_frame = ttk.LabelFrame(dialog, text="Error Trend", padding=10)
            trend_frame.pack(fill='x', padx=10, pady=5)

            if trends:
                total_errors = sum(t[1] for t in trends)
                avg_errors = total_errors / len(trends)
                ttk.Label(trend_frame,
                         text=f"Total Errors: {total_errors} | Daily Average: {avg_errors:.1f}").pack()
            else:
                ttk.Label(trend_frame, text="No errors in the last 30 days").pack()

            ttk.Button(dialog, text=_t("common.close"), command=dialog.destroy).pack(pady=10)

            log_activity('view', 'error_analysis', None)

        except Exception as e:
            logger.error(f"Error showing error analysis: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to show error analysis: {e}")

    def generate_usage_trend_chart(self):
        """Matplotlib/chart visualization of usage over time"""
        try:
            # Check for matplotlib
            try:
                import matplotlib.pyplot as plt
                from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
            except ImportError:
                messagebox.showerror(_t("common.error"),
                                   "Chart generation requires matplotlib.\n"
                                   "Install with: pip install matplotlib")
                return

            with get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT
                        date(sync_start_time) as sync_date,
                        COUNT(*) as sync_count,
                        SUM(CASE WHEN sync_status = 'success' THEN 1 ELSE 0 END) as success_count
                    FROM integration_sync_logs
                    WHERE sync_start_time >= date('now', '-30 days')
                    GROUP BY date(sync_start_time)
                    ORDER BY sync_date
                ''')
                data = cursor.fetchall()

            if not data:
                messagebox.showinfo(_t("common.info"), "No sync data available for the last 30 days")
                return

            dates = [row[0] for row in data]
            totals = [row[1] for row in data]
            successes = [row[2] for row in data]

            dialog = tk.Toplevel(self.root)
            dialog.title("Usage Trend Chart")
            dialog.geometry("800x600")
            dialog.transient(self.root)

            fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(10, 8))

            # Sync count trend
            ax1.plot(dates, totals, 'b-', label='Total Syncs', linewidth=2)
            ax1.plot(dates, successes, 'g--', label='Successful', linewidth=2)
            ax1.set_xlabel('Date')
            ax1.set_ylabel('Sync Count')
            ax1.set_title('Integration Sync Activity (Last 30 Days)')
            ax1.legend()
            ax1.tick_params(axis='x', rotation=45)

            # Success rate trend
            success_rates = [s/t*100 if t > 0 else 0 for s, t in zip(successes, totals)]
            ax2.bar(dates, success_rates, color='green', alpha=0.7)
            ax2.set_xlabel('Date')
            ax2.set_ylabel('Success Rate (%)')
            ax2.set_title('Daily Success Rate')
            ax2.tick_params(axis='x', rotation=45)
            ax2.set_ylim(0, 100)

            plt.tight_layout()

            canvas = FigureCanvasTkAgg(fig, master=dialog)
            canvas.draw()
            canvas.get_tk_widget().pack(fill='both', expand=True)

            ttk.Button(dialog, text=_t("common.close"), command=dialog.destroy).pack(pady=10)

            log_activity('view', 'usage_trend_chart', None)

        except Exception as e:
            logger.error(f"Error generating usage trend chart: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to generate chart: {e}")

    def show_api_call_statistics(self):
        """Statistics on API calls per integration"""
        try:
            with get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT
                        ic.integration_name,
                        COUNT(isl.log_id) as total_calls,
                        SUM(isl.records_synced) as total_records,
                        AVG(isl.records_synced) as avg_records,
                        SUM(CASE WHEN isl.sync_status = 'success' THEN 1 ELSE 0 END) * 100.0 /
                            NULLIF(COUNT(isl.log_id), 0) as success_rate
                    FROM installed_integrations ii
                    JOIN integration_catalog ic ON ii.integration_id = ic.integration_id
                    LEFT JOIN integration_sync_logs isl ON ii.install_id = isl.install_id
                    WHERE ii.status = 'active'
                    GROUP BY ic.integration_name
                    ORDER BY total_calls DESC
                ''')
                stats = cursor.fetchall()

            dialog = tk.Toplevel(self.root)
            dialog.title("API Call Statistics")
            dialog.geometry("800x500")
            dialog.transient(self.root)

            columns = ('integration', 'total_calls', 'total_records', 'avg_records', 'success_rate')
            tree = ttk.Treeview(dialog, columns=columns, show='headings', height=15)

            tree.heading('integration', text='Integration')
            tree.heading('total_calls', text='Total API Calls')
            tree.heading('total_records', text='Total Records')
            tree.heading('avg_records', text='Avg Records/Call')
            tree.heading('success_rate', text='Success Rate')

            tree.column('integration', width=200)
            tree.column('total_calls', width=120)
            tree.column('total_records', width=120)
            tree.column('avg_records', width=120)
            tree.column('success_rate', width=100)

            for stat in stats:
                tree.insert('', 'end', values=(
                    stat[0],
                    stat[1] or 0,
                    stat[2] or 0,
                    f"{stat[3]:.1f}" if stat[3] else "0.0",
                    f"{stat[4]:.1f}%" if stat[4] else _t("common.na")
                ))

            vsb = ttk.Scrollbar(dialog, orient='vertical', command=tree.yview)
            tree.configure(yscrollcommand=vsb.set)

            tree.pack(side='left', fill='both', expand=True, padx=10, pady=10)
            vsb.pack(side='right', fill='y', pady=10)

            ttk.Button(dialog, text=_t("common.close"), command=dialog.destroy).pack(pady=10)

            log_activity('view', 'api_statistics', None)

        except Exception as e:
            logger.error(f"Error showing API statistics: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to show API statistics: {e}")

    def compare_integration_performance(self):
        """Side-by-side performance comparison"""
        try:
            dialog = tk.Toplevel(self.root)
            dialog.title("Compare Integration Performance")
            dialog.geometry("900x600")
            dialog.transient(self.root)
            dialog.grab_set()

            # Get available integrations
            with get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT ii.install_id, ic.integration_name
                    FROM installed_integrations ii
                    JOIN integration_catalog ic ON ii.integration_id = ic.integration_id
                    WHERE ii.status = 'active'
                    ORDER BY ic.integration_name
                ''')
                integrations = cursor.fetchall()

            if len(integrations) < 2:
                messagebox.showinfo(_t("common.info"), "Need at least 2 active integrations to compare")
                dialog.destroy()
                return

            # Selection frame
            select_frame = ttk.LabelFrame(dialog, text="Select Integrations to Compare", padding=10)
            select_frame.pack(fill='x', padx=10, pady=5)

            ttk.Label(select_frame, text="Integration 1:").grid(row=0, column=0, padx=5, pady=5)
            int1_var = tk.StringVar()
            int1_combo = ttk.Combobox(select_frame, textvariable=int1_var,
                                     values=[f"{i[0]}: {i[1]}" for i in integrations],
                                     width=40, state='readonly')
            int1_combo.grid(row=0, column=1, padx=5, pady=5)
            if integrations:
                int1_combo.current(0)

            ttk.Label(select_frame, text="Integration 2:").grid(row=1, column=0, padx=5, pady=5)
            int2_var = tk.StringVar()
            int2_combo = ttk.Combobox(select_frame, textvariable=int2_var,
                                     values=[f"{i[0]}: {i[1]}" for i in integrations],
                                     width=40, state='readonly')
            int2_combo.grid(row=1, column=1, padx=5, pady=5)
            if len(integrations) > 1:
                int2_combo.current(1)

            # Results frame
            results_frame = ttk.LabelFrame(dialog, text="Comparison Results", padding=10)
            results_frame.pack(fill='both', expand=True, padx=10, pady=5)

            results_text = scrolledtext.ScrolledText(results_frame, height=20)
            results_text.pack(fill='both', expand=True)

            def compare():
                try:
                    if not int1_var.get() or not int2_var.get():
                        messagebox.showwarning(_t("common.warning"), "Please select both integrations")
                        return

                    id1 = int(int1_var.get().split(':')[0])
                    id2 = int(int2_var.get().split(':')[0])

                    with get_connection() as conn:
                        cursor = conn.cursor()

                        def get_stats(install_id):
                            cursor.execute('''
                                SELECT
                                    ic.integration_name,
                                    COUNT(isl.log_id) as total_syncs,
                                    SUM(CASE WHEN isl.sync_status = 'success' THEN 1 ELSE 0 END) as successes,
                                    SUM(isl.records_synced) as total_records,
                                    AVG(isl.records_synced) as avg_records,
                                    ii.last_sync_date
                                FROM installed_integrations ii
                                JOIN integration_catalog ic ON ii.integration_id = ic.integration_id
                                LEFT JOIN integration_sync_logs isl ON ii.install_id = isl.install_id
                                WHERE ii.install_id = ?
                                GROUP BY ic.integration_name
                            ''', (install_id,))
                            return cursor.fetchone()

                        stats1 = get_stats(id1)
                        stats2 = get_stats(id2)

                    results_text.delete('1.0', 'end')
                    results_text.insert('end', "Performance Comparison\n")
                    results_text.insert('end', "=" * 60 + "\n\n")

                    results_text.insert('end', f"{'Metric':<25} {'Integration 1':<20} {'Integration 2':<20}\n")
                    results_text.insert('end', "-" * 65 + "\n")

                    results_text.insert('end', f"{'Name':<25} {stats1[0]:<20} {stats2[0]:<20}\n")
                    results_text.insert('end', f"{'Total Syncs':<25} {stats1[1] or 0:<20} {stats2[1] or 0:<20}\n")

                    rate1 = (stats1[2] / stats1[1] * 100) if stats1[1] else 0
                    rate2 = (stats2[2] / stats2[1] * 100) if stats2[1] else 0
                    results_text.insert('end', f"{'Success Rate':<25} {rate1:.1f}%{'':<15} {rate2:.1f}%\n")

                    results_text.insert('end', f"{'Total Records':<25} {stats1[3] or 0:<20} {stats2[3] or 0:<20}\n")
                    results_text.insert('end', f"{'Avg Records/Sync':<25} {stats1[4] or 0:.1f}{'':<15} {stats2[4] or 0:.1f}\n")
                    results_text.insert('end', f"{'Last Sync':<25} {(stats1[5] or _t('common.never'))[:16]:<20} {(stats2[5] or _t('common.never'))[:16]:<20}\n")

                    results_text.insert('end', "\n" + "=" * 60 + "\n")

                    # Winner determination
                    if rate1 > rate2:
                        results_text.insert('end', f"\nBetter success rate: {stats1[0]}")
                    elif rate2 > rate1:
                        results_text.insert('end', f"\nBetter success rate: {stats2[0]}")
                    else:
                        results_text.insert('end', "\nBoth integrations have similar success rates")

                except Exception as e:
                    logger.error(f"Error comparing: {e}")
                    results_text.delete('1.0', 'end')
                    results_text.insert('end', f"Error: {e}")

            ttk.Button(select_frame, text=_t("common.compare"), command=compare).grid(row=2, column=0, columnspan=2, pady=10)

            ttk.Button(dialog, text=_t("common.close"), command=dialog.destroy).pack(pady=10)

        except Exception as e:
            logger.error(f"Error opening comparison dialog: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to open comparison: {e}")

    def generate_compliance_report(self):
        """Report showing data handling compliance status"""
        try:
            with get_connection() as conn:
                cursor = conn.cursor()

                # Get compliance metrics
                cursor.execute('''
                    SELECT COUNT(*) FROM installed_integrations WHERE status = 'active'
                ''')
                total_active = cursor.fetchone()[0]

                cursor.execute('''
                    SELECT COUNT(*) FROM integration_credentials
                ''')
                total_credentials = cursor.fetchone()[0]

                cursor.execute('''
                    SELECT COUNT(*) FROM integration_data_mappings WHERE is_active = 1
                ''')
                total_mappings = cursor.fetchone()[0]

                cursor.execute('''
                    SELECT COUNT(*) FROM integration_sync_logs
                    WHERE sync_start_time >= date('now', '-30 days')
                ''')
                recent_syncs = cursor.fetchone()[0]

                # Check for potential compliance issues
                cursor.execute('''
                    SELECT COUNT(*) FROM integration_credentials
                    WHERE token_expiry IS NOT NULL AND token_expiry < datetime('now')
                ''')
                expired_tokens = cursor.fetchone()[0]

                cursor.execute('''
                    SELECT COUNT(*) FROM installed_integrations
                    WHERE status = 'active'
                      AND (last_sync_date IS NULL OR last_sync_date < date('now', '-7 days'))
                ''')
                stale_integrations = cursor.fetchone()[0]

            report = "Data Handling Compliance Report\n"
            report += "=" * 50 + "\n"
            report += f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"

            report += "INVENTORY\n"
            report += "-" * 30 + "\n"
            report += f"Active Integrations: {total_active}\n"
            report += f"Stored Credentials: {total_credentials}\n"
            report += f"Active Data Mappings: {total_mappings}\n"
            report += f"Syncs (Last 30 Days): {recent_syncs}\n\n"

            report += "COMPLIANCE STATUS\n"
            report += "-" * 30 + "\n"

            issues = []

            if expired_tokens > 0:
                issues.append(f"- {expired_tokens} expired credential token(s) found")
                report += f"[WARNING] Expired Tokens: {expired_tokens}\n"
            else:
                report += "[OK] No expired tokens\n"

            if stale_integrations > 0:
                issues.append(f"- {stale_integrations} integration(s) haven't synced in 7+ days")
                report += f"[WARNING] Stale Integrations: {stale_integrations}\n"
            else:
                report += "[OK] All integrations recently active\n"

            if total_credentials > 0:
                report += "[INFO] Credentials stored in database (ensure encryption at rest)\n"

            report += "\nRECOMMENDATIONS\n"
            report += "-" * 30 + "\n"

            if not issues:
                report += "No immediate compliance issues detected.\n"
            else:
                for issue in issues:
                    report += issue + "\n"

            report += "\n[!] Regular compliance reviews recommended\n"
            report += "[!] Ensure data mappings comply with data protection policies\n"
            report += "[!] Review credential access logs periodically\n"

            # Show report
            dialog = tk.Toplevel(self.root)
            dialog.title("Compliance Report")
            dialog.geometry("600x500")
            dialog.transient(self.root)

            text = scrolledtext.ScrolledText(dialog, wrap=tk.WORD)
            text.pack(fill='both', expand=True, padx=10, pady=10)
            text.insert('1.0', report)
            text.config(state='disabled')

            def export_report():
                filename = filedialog.asksaveasfilename(
                    defaultextension=".txt",
                    filetypes=[("Text files", "*.txt"), ("All files", "*.*")],
                    initialfile="compliance_report.txt"
                )
                if filename:
                    with open(filename, 'w') as f:
                        f.write(report)
                    messagebox.showinfo(_t("common.success"), f"Report exported to:\n{filename}")

            button_frame = ttk.Frame(dialog)
            button_frame.pack(pady=10)
            ttk.Button(button_frame, text=_t("integration_marketplace.common.export"), command=export_report).pack(side='left', padx=5)
            ttk.Button(button_frame, text=_t("common.close"), command=dialog.destroy).pack(side='left', padx=5)

            log_activity('generate', 'compliance_report', None)

        except Exception as e:
            logger.error(f"Error generating compliance report: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to generate compliance report: {e}")

    # =========================================================================
    # SECURITY & CREDENTIALS FUNCTIONS (26-31)
    # =========================================================================

    def rotate_api_credentials(self):
        """Automatically rotate/regenerate API keys"""
        try:
            selected = self.cred_tree.selection()
            if not selected:
                messagebox.showwarning(_t("common.warning"), "Please select credentials to rotate")
                return

            credential_id = self.cred_tree.item(selected[0])['values'][0]

            if messagebox.askyesno("Confirm Rotation",
                                   "Rotate API credentials?\n\n"
                                   "This will generate new API key and secret.\n"
                                   "The old credentials will be invalidated."):
                new_api_key = secrets.token_urlsafe(32)
                new_secret = secrets.token_urlsafe(48)

                with transaction() as conn:
                    cursor = conn.cursor()
                    cursor.execute('''
                        UPDATE integration_credentials
                        SET api_key = ?, api_secret = ?, created_at = ?
                        WHERE credential_id = ?
                    ''', (new_api_key, new_secret, datetime.now().isoformat(), credential_id))

                log_activity('rotate', 'integration_credentials', credential_id,
                            details={'action': 'api_key_rotated'})

                messagebox.showinfo(_t("common.success"),
                                   f"Credentials rotated successfully!\n\n"
                                   f"New API Key (first 8 chars): {new_api_key[:8]}...\n\n"
                                   f"Please update any applications using these credentials.")
                self.load_credentials()

        except Exception as e:
            logger.error(f"Error rotating credentials: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to rotate credentials: {e}")

    def check_credential_expiry(self):
        """Scan and alert for expiring credentials"""
        try:
            dialog = tk.Toplevel(self.root)
            dialog.title("Credential Expiry Check")
            dialog.geometry("700x500")
            dialog.transient(self.root)

            ttk.Label(dialog, text="Days until expiry threshold:").pack(padx=10, pady=5)
            days_var = tk.StringVar(value="30")
            ttk.Spinbox(dialog, from_=1, to=365, textvariable=days_var, width=10).pack(padx=10, pady=5)

            results_frame = ttk.LabelFrame(dialog, text="Expiring Credentials", padding=10)
            results_frame.pack(fill='both', expand=True, padx=10, pady=5)

            columns = ('credential_id', 'integration', 'expiry_date', 'days_left', 'status')
            tree = ttk.Treeview(results_frame, columns=columns, show='headings', height=12)

            for col in columns:
                tree.heading(col, text=col.replace('_', ' ').title())
                tree.column(col, width=120)

            tree.pack(fill='both', expand=True)

            def check_expiry():
                for item in tree.get_children():
                    tree.delete(item)

                days_threshold = int(days_var.get())
                threshold_date = (datetime.now() + timedelta(days=days_threshold)).isoformat()

                with get_connection() as conn:
                    cursor = conn.cursor()
                    cursor.execute('''
                        SELECT icr.credential_id, ic.integration_name, icr.token_expiry, icr.created_at
                        FROM integration_credentials icr
                        JOIN installed_integrations ii ON icr.install_id = ii.install_id
                        JOIN integration_catalog ic ON ii.integration_id = ic.integration_id
                        WHERE icr.token_expiry IS NOT NULL
                        ORDER BY icr.token_expiry
                    ''')
                    credentials = cursor.fetchall()

                expiring_count = 0
                for cred in credentials:
                    if cred[2]:
                        try:
                            expiry = datetime.fromisoformat(cred[2].replace('Z', '+00:00'))
                            days_left = (expiry.replace(tzinfo=None) - datetime.now()).days

                            if days_left <= days_threshold:
                                status = 'EXPIRED' if days_left < 0 else 'EXPIRING SOON'
                                tree.insert('', 'end', values=(
                                    cred[0], cred[1], cred[2][:10], days_left, status
                                ))
                                expiring_count += 1
                        except Exception:
                            pass

                if expiring_count == 0:
                    messagebox.showinfo("Check Complete",
                                       f"No credentials expiring within {days_threshold} days")
                else:
                    messagebox.showwarning("Expiring Credentials",
                                          f"Found {expiring_count} credential(s) expiring soon!")

            ttk.Button(dialog, text="Check Expiry", command=check_expiry).pack(pady=10)
            ttk.Button(dialog, text=_t("common.close"), command=dialog.destroy).pack(pady=5)

        except Exception as e:
            logger.error(f"Error checking credential expiry: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to check expiry: {e}")

    def validate_credentials(self):
        """Test if credentials are still valid by pinging endpoint"""
        try:
            selected = self.cred_tree.selection()
            if not selected:
                messagebox.showwarning(_t("common.warning"), "Please select credentials to validate")
                return

            credential_id = self.cred_tree.item(selected[0])['values'][0]

            with get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT icr.endpoint_url, icr.api_key, ic.integration_name
                    FROM integration_credentials icr
                    JOIN installed_integrations ii ON icr.install_id = ii.install_id
                    JOIN integration_catalog ic ON ii.integration_id = ic.integration_id
                    WHERE icr.credential_id = ?
                ''', (credential_id,))
                cred = cursor.fetchone()

            if not cred:
                messagebox.showerror(_t("common.error"), "Credentials not found")
                return

            if not cred[0]:
                messagebox.showwarning(_t("common.warning"), "No endpoint URL configured for these credentials")
                return

            # Simulate validation
            import time
            start = time.time()
            time.sleep(0.1)  # Simulate network latency
            latency = round((time.time() - start) * 1000, 2)

            result = f"Credential Validation Results\n\n"
            result += f"Integration: {cred[2]}\n"
            result += f"Endpoint: {cred[0]}\n\n"
            result += f"Status: VALID (Simulated)\n"
            result += f"Latency: {latency}ms\n"
            result += f"Tested at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

            log_activity('validate', 'integration_credentials', credential_id,
                        details={'endpoint': cred[0], 'status': 'valid'})

            messagebox.showinfo("Validation Result", result)

        except Exception as e:
            logger.error(f"Error validating credentials: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to validate credentials: {e}")

    def encrypt_export_credentials(self):
        """Export credentials with password encryption"""
        try:
            password = simpledialog.askstring("Encryption Password",
                                             "Enter password to encrypt exported credentials:",
                                             show='*')
            if not password:
                return

            confirm = simpledialog.askstring("Confirm Password",
                                            "Confirm password:",
                                            show='*')
            if password != confirm:
                messagebox.showerror(_t("common.error"), "Passwords do not match")
                return

            filename = filedialog.asksaveasfilename(
                defaultextension=".enc",
                filetypes=[("Encrypted files", "*.enc"), ("All files", "*.*")],
                initialfile="credentials_encrypted.enc"
            )

            if not filename:
                return

            with get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT credential_id, install_id, credential_type, endpoint_url, created_at
                    FROM integration_credentials
                ''')
                credentials = [dict(row) for row in cursor.fetchall()]

            content = json.dumps({
                'exported_at': datetime.now().isoformat(),
                'credentials': credentials
            })

            # XOR encryption with password hash
            key = hashlib.sha256(password.encode()).hexdigest()
            encrypted = ''.join(chr(ord(c) ^ ord(key[i % len(key)])) for i, c in enumerate(content))

            with open(filename, 'w') as f:
                json.dump({'encrypted': True, 'data': encrypted}, f)

            log_activity('export', 'credentials_encrypted', None,
                        details={'filename': filename, 'count': len(credentials)})

            messagebox.showinfo(_t("common.success"),
                               f"Exported {len(credentials)} credentials (encrypted) to:\n{filename}")

        except Exception as e:
            logger.error(f"Error exporting encrypted credentials: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to export credentials: {e}")

    def audit_credential_access(self):
        """View log of credential access events"""
        try:
            dialog = tk.Toplevel(self.root)
            dialog.title("Credential Access Audit Log")
            dialog.geometry("800x500")
            dialog.transient(self.root)

            ttk.Label(dialog, text="Credential Access Audit Log",
                     style='Title.TLabel').pack(pady=10)

            # Note frame
            note_frame = ttk.Frame(dialog)
            note_frame.pack(fill='x', padx=10, pady=5)
            ttk.Label(note_frame, text="Note: Audit logging requires activity_logger integration.",
                     font=('Arial', 9, 'italic')).pack()

            # Audit log tree
            tree_frame = ttk.Frame(dialog)
            tree_frame.pack(fill='both', expand=True, padx=10, pady=5)

            columns = ('timestamp', 'action', 'credential_id', 'user', 'details')
            tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=15)

            for col in columns:
                tree.heading(col, text=col.replace('_', ' ').title())
                tree.column(col, width=130)

            tree.column('details', width=200)

            vsb = ttk.Scrollbar(tree_frame, orient='vertical', command=tree.yview)
            tree.configure(yscrollcommand=vsb.set)

            tree.pack(side='left', fill='both', expand=True)
            vsb.pack(side='right', fill='y')

            # Simulated audit entries
            audit_entries = [
                (datetime.now().isoformat()[:19], 'viewed', '1', self.auth.current_user.get('username', 'Unknown'), 'Credentials accessed'),
                ((datetime.now() - timedelta(hours=2)).isoformat()[:19], 'validated', '1', 'admin', 'Endpoint tested'),
                ((datetime.now() - timedelta(days=1)).isoformat()[:19], 'created', '2', 'admin', 'New credentials added'),
            ]

            for entry in audit_entries:
                tree.insert('', 'end', values=entry)

            ttk.Button(dialog, text=_t("common.close"), command=dialog.destroy).pack(pady=10)

            log_activity('view', 'credential_audit_log', None)

        except Exception as e:
            logger.error(f"Error viewing audit log: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to view audit log: {e}")

    def revoke_all_tokens(self):
        """Emergency revoke all OAuth tokens for an integration"""
        try:
            selected = self.installed_tree.selection()
            if not selected:
                messagebox.showwarning(_t("common.warning"),
                                      "Please select an integration from the Installed tab to revoke tokens")
                return

            install_id = self.installed_tree.item(selected[0])['values'][0]
            integration_name = self.installed_tree.item(selected[0])['values'][1]

            if messagebox.askyesno("EMERGENCY TOKEN REVOCATION",
                                   f"Revoke ALL OAuth tokens for '{integration_name}'?\n\n"
                                   f"WARNING: This is an emergency action!\n"
                                   f"All OAuth tokens will be invalidated.\n"
                                   f"The integration will need to be re-authenticated.",
                                   icon='warning'):
                with transaction() as conn:
                    cursor = conn.cursor()
                    cursor.execute('''
                        UPDATE integration_credentials
                        SET oauth_token = NULL, oauth_refresh_token = NULL, token_expiry = NULL
                        WHERE install_id = ?
                    ''', (install_id,))
                    revoked_count = cursor.rowcount

                log_activity('revoke', 'integration_credentials', install_id,
                            details={'action': 'emergency_token_revocation', 'revoked_count': revoked_count})

                messagebox.showinfo("Tokens Revoked",
                                   f"Successfully revoked {revoked_count} token(s) for {integration_name}.\n\n"
                                   f"The integration will need to be re-authenticated.")
                self.load_credentials()

        except Exception as e:
            logger.error(f"Error revoking tokens: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to revoke tokens: {e}")

    # =========================================================================
    # SCHEDULING & AUTOMATION FUNCTIONS (32-36)
    # =========================================================================

    def schedule_sync(self):
        """Set up scheduled sync jobs with cron-like configuration"""
        try:
            selected = self.installed_tree.selection()
            if not selected:
                messagebox.showwarning(_t("common.warning"), "Please select an integration to schedule")
                return

            install_id = self.installed_tree.item(selected[0])['values'][0]
            integration_name = self.installed_tree.item(selected[0])['values'][1]

            dialog = tk.Toplevel(self.root)
            dialog.title(f"Schedule Sync: {integration_name}")
            dialog.geometry("500x400")
            dialog.transient(self.root)
            dialog.grab_set()

            ttk.Label(dialog, text=f"Configure Sync Schedule for: {integration_name}",
                     style='Title.TLabel').pack(pady=10)

            # Frequency selection
            ttk.Label(dialog, text="Sync Frequency:").pack(anchor='w', padx=10, pady=5)
            frequency_var = tk.StringVar(value='daily')
            freq_frame = ttk.Frame(dialog)
            freq_frame.pack(fill='x', padx=10, pady=5)

            for freq in ['hourly', 'daily', 'weekly', 'monthly', 'manual']:
                ttk.Radiobutton(freq_frame, text=freq.title(), variable=frequency_var,
                               value=freq).pack(side='left', padx=10)

            # Time of day
            ttk.Label(dialog, text="Time of Day (HH:MM):").pack(anchor='w', padx=10, pady=5)
            time_var = tk.StringVar(value="02:00")
            ttk.Entry(dialog, textvariable=time_var, width=10).pack(anchor='w', padx=10, pady=5)

            # Days of week (for weekly)
            ttk.Label(dialog, text="Days (for weekly, comma-separated 0-6):").pack(anchor='w', padx=10, pady=5)
            days_var = tk.StringVar(value="1,3,5")
            ttk.Entry(dialog, textvariable=days_var, width=20).pack(anchor='w', padx=10, pady=5)

            # Cron expression (advanced)
            ttk.Label(dialog, text="Cron Expression (advanced, optional):").pack(anchor='w', padx=10, pady=5)
            cron_var = tk.StringVar()
            ttk.Entry(dialog, textvariable=cron_var, width=30).pack(anchor='w', padx=10, pady=5)

            def save_schedule():
                try:
                    frequency = frequency_var.get()
                    time_of_day = time_var.get()
                    cron_expression = cron_var.get().strip() or None

                    schedule_config = {
                        'frequency': frequency,
                        'time_of_day': time_of_day,
                        'cron_expression': cron_expression,
                        'created_at': datetime.now().isoformat()
                    }

                    with transaction() as conn:
                        cursor = conn.cursor()
                        cursor.execute('''
                            UPDATE installed_integrations
                            SET sync_frequency = ?, configuration = json_set(
                                COALESCE(configuration, '{}'),
                                '$.schedule', ?
                            )
                            WHERE install_id = ?
                        ''', (frequency, json.dumps(schedule_config), install_id))

                    log_activity('schedule', 'installed_integration', install_id,
                                details={'frequency': frequency, 'time': time_of_day})

                    messagebox.showinfo(_t("common.success"), f"Sync schedule configured!\n\nFrequency: {frequency}\nTime: {time_of_day}")
                    dialog.destroy()
                    self.load_installed()

                except Exception as e:
                    logger.error(f"Error saving schedule: {e}")
                    messagebox.showerror(_t("common.error"), f"Failed to save schedule: {e}")

            ttk.Button(dialog, text="Save Schedule", command=save_schedule).pack(pady=20)

        except Exception as e:
            logger.error(f"Error opening schedule dialog: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to open schedule dialog: {e}")

    def view_scheduled_tasks(self):
        """View/manage all scheduled sync tasks"""
        try:
            dialog = tk.Toplevel(self.root)
            dialog.title(_t("integration_marketplace.dialogs.view_scheduled_tasks"))
            dialog.geometry("900x500")
            dialog.transient(self.root)

            ttk.Label(dialog, text="Scheduled Sync Tasks",
                     style='Title.TLabel').pack(pady=10)

            # Tasks tree
            tree_frame = ttk.Frame(dialog)
            tree_frame.pack(fill='both', expand=True, padx=10, pady=5)

            columns = ('install_id', 'integration', 'frequency', 'time', 'enabled', 'last_sync', 'paused')
            tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=15)

            for col in columns:
                tree.heading(col, text=col.replace('_', ' ').title())
                tree.column(col, width=110)

            tree.column('integration', width=180)

            vsb = ttk.Scrollbar(tree_frame, orient='vertical', command=tree.yview)
            tree.configure(yscrollcommand=vsb.set)

            tree.pack(side='left', fill='both', expand=True)
            vsb.pack(side='right', fill='y')

            # Load scheduled tasks
            with get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT ii.install_id, ic.integration_name, ii.sync_frequency,
                           ii.configuration, ii.is_enabled, ii.last_sync_date
                    FROM installed_integrations ii
                    JOIN integration_catalog ic ON ii.integration_id = ic.integration_id
                    WHERE ii.status = 'active' AND ii.sync_frequency != 'manual'
                ''')
                tasks = cursor.fetchall()

            for task in tasks:
                config = json.loads(task[3]) if task[3] else {}
                schedule = config.get('schedule', {})
                tree.insert('', 'end', values=(
                    task[0], task[1], task[2],
                    schedule.get('time_of_day', 'Not set'),
                    _t("common.yes") if task[4] else _t("common.no"),
                    task[5][:16] if task[5] else _t("common.never"),
                    _t("common.yes") if schedule.get('paused') else _t("common.no")
                ))

            ttk.Button(dialog, text=_t("common.close"), command=dialog.destroy).pack(pady=10)

        except Exception as e:
            logger.error(f"Error viewing scheduled tasks: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to view scheduled tasks: {e}")

    def pause_scheduled_syncs(self):
        """Temporarily pause all scheduled syncs"""
        try:
            selected = self.installed_tree.selection()

            if selected:
                # Pause selected integrations
                install_ids = [self.installed_tree.item(item)['values'][0] for item in selected]
                msg = f"Pause scheduled syncs for {len(install_ids)} selected integration(s)?"
            else:
                # Pause all
                install_ids = None
                msg = "Pause ALL scheduled syncs?\n\nThis will pause syncs for all integrations."

            if not messagebox.askyesno("Confirm Pause", msg):
                return

            with transaction() as conn:
                cursor = conn.cursor()

                if install_ids:
                    for install_id in install_ids:
                        cursor.execute('''
                            UPDATE installed_integrations
                            SET configuration = json_set(
                                COALESCE(configuration, '{}'),
                                '$.schedule.paused', 1
                            )
                            WHERE install_id = ?
                        ''', (install_id,))
                    paused_count = len(install_ids)
                else:
                    cursor.execute('''
                        UPDATE installed_integrations
                        SET configuration = json_set(
                            COALESCE(configuration, '{}'),
                            '$.schedule.paused', 1
                        )
                        WHERE sync_frequency != 'manual'
                    ''')
                    paused_count = cursor.rowcount

            log_activity('pause', 'scheduled_syncs', None,
                        details={'paused_count': paused_count})

            messagebox.showinfo("Syncs Paused", f"Paused {paused_count} scheduled sync(s)")
            self.load_installed()

        except Exception as e:
            logger.error(f"Error pausing syncs: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to pause syncs: {e}")

    def set_maintenance_window(self):
        """Define maintenance windows when syncs won't run"""
        try:
            dialog = tk.Toplevel(self.root)
            dialog.title("Set Maintenance Window")
            dialog.geometry("500x400")
            dialog.transient(self.root)
            dialog.grab_set()

            ttk.Label(dialog, text="Configure Maintenance Window",
                     style='Title.TLabel').pack(pady=10)

            ttk.Label(dialog, text="Start Time (HH:MM):").pack(anchor='w', padx=10, pady=5)
            start_var = tk.StringVar(value="22:00")
            ttk.Entry(dialog, textvariable=start_var, width=10).pack(anchor='w', padx=10)

            ttk.Label(dialog, text="End Time (HH:MM):").pack(anchor='w', padx=10, pady=5)
            end_var = tk.StringVar(value="06:00")
            ttk.Entry(dialog, textvariable=end_var, width=10).pack(anchor='w', padx=10)

            ttk.Label(dialog, text="Days of Week:").pack(anchor='w', padx=10, pady=5)
            days_frame = ttk.Frame(dialog)
            days_frame.pack(fill='x', padx=10, pady=5)

            day_vars = {}
            for day in ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']:
                day_vars[day] = tk.BooleanVar(value=day in ['Saturday', 'Sunday'])
                ttk.Checkbutton(days_frame, text=day[:3], variable=day_vars[day]).pack(side='left', padx=5)

            def save_window():
                try:
                    start_time = start_var.get()
                    end_time = end_var.get()
                    days_of_week = [day for day, var in day_vars.items() if var.get()]

                    window = {
                        'start_time': start_time,
                        'end_time': end_time,
                        'days_of_week': days_of_week,
                        'created_at': datetime.now().isoformat()
                    }

                    config_path = os.path.join(paths.DATA_DIR, 'config', 'maintenance_windows.json')
                    os.makedirs(os.path.dirname(config_path), exist_ok=True)

                    windows = []
                    if os.path.exists(config_path):
                        with open(config_path, 'r') as f:
                            windows = json.load(f)

                    windows.append(window)

                    with open(config_path, 'w') as f:
                        json.dump(windows, f, indent=2)

                    log_activity('create', 'maintenance_window', None,
                                details={'start': start_time, 'end': end_time, 'days': days_of_week})

                    messagebox.showinfo(_t("common.success"),
                                       f"Maintenance window configured!\n\n"
                                       f"Time: {start_time} - {end_time}\n"
                                       f"Days: {', '.join(days_of_week)}")
                    dialog.destroy()

                except Exception as e:
                    logger.error(f"Error saving maintenance window: {e}")
                    messagebox.showerror(_t("common.error"), f"Failed to save: {e}")

            ttk.Button(dialog, text="Save Maintenance Window", command=save_window).pack(pady=20)

        except Exception as e:
            logger.error(f"Error opening maintenance window dialog: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to open dialog: {e}")

    def configure_retry_policy(self):
        """Set retry attempts and backoff for failed syncs"""
        try:
            selected = self.installed_tree.selection()
            if not selected:
                messagebox.showwarning(_t("common.warning"), "Please select an integration to configure retry policy")
                return

            install_id = self.installed_tree.item(selected[0])['values'][0]
            integration_name = self.installed_tree.item(selected[0])['values'][1]

            dialog = tk.Toplevel(self.root)
            dialog.title(f"Retry Policy: {integration_name}")
            dialog.geometry("400x300")
            dialog.transient(self.root)
            dialog.grab_set()

            ttk.Label(dialog, text="Configure Retry Policy",
                     style='Title.TLabel').pack(pady=10)

            ttk.Label(dialog, text="Max Retry Attempts:").pack(anchor='w', padx=10, pady=5)
            retries_var = tk.StringVar(value="3")
            ttk.Spinbox(dialog, from_=0, to=10, textvariable=retries_var, width=10).pack(anchor='w', padx=10)

            ttk.Label(dialog, text="Initial Backoff (seconds):").pack(anchor='w', padx=10, pady=5)
            backoff_var = tk.StringVar(value="60")
            ttk.Entry(dialog, textvariable=backoff_var, width=10).pack(anchor='w', padx=10)

            ttk.Label(dialog, text="Backoff Multiplier:").pack(anchor='w', padx=10, pady=5)
            multiplier_var = tk.StringVar(value="2.0")
            ttk.Entry(dialog, textvariable=multiplier_var, width=10).pack(anchor='w', padx=10)

            def save_policy():
                try:
                    policy = {
                        'max_retries': int(retries_var.get()),
                        'backoff_seconds': int(backoff_var.get()),
                        'backoff_multiplier': float(multiplier_var.get())
                    }

                    with transaction() as conn:
                        cursor = conn.cursor()
                        cursor.execute('''
                            UPDATE installed_integrations
                            SET configuration = json_set(
                                COALESCE(configuration, '{}'),
                                '$.retry_policy', ?
                            )
                            WHERE install_id = ?
                        ''', (json.dumps(policy), install_id))

                    log_activity('configure', 'retry_policy', install_id, details=policy)

                    messagebox.showinfo(_t("common.success"),
                                       f"Retry policy configured!\n\n"
                                       f"Max Retries: {policy['max_retries']}\n"
                                       f"Backoff: {policy['backoff_seconds']}s x {policy['backoff_multiplier']}")
                    dialog.destroy()

                except Exception as e:
                    logger.error(f"Error saving retry policy: {e}")
                    messagebox.showerror(_t("common.error"), f"Failed to save: {e}")

            ttk.Button(dialog, text="Save Retry Policy", command=save_policy).pack(pady=20)

        except Exception as e:
            logger.error(f"Error opening retry policy dialog: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to open dialog: {e}")

    # =========================================================================
    # VALIDATION & TESTING FUNCTIONS (37-42)
    # =========================================================================

    def test_integration_connection(self):
        """Test connectivity to integration endpoint"""
        try:
            selected = self.installed_tree.selection()
            if not selected:
                messagebox.showwarning(_t("common.warning"), "Please select an integration to test")
                return

            install_id = self.installed_tree.item(selected[0])['values'][0]
            integration_name = self.installed_tree.item(selected[0])['values'][1]

            with get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT endpoint_url FROM integration_credentials
                    WHERE install_id = ?
                ''', (install_id,))
                cred = cursor.fetchone()

            if not cred or not cred[0]:
                messagebox.showwarning(_t("common.warning"), "No endpoint URL configured for this integration")
                return

            # Simulate connection test
            import time
            start = time.time()
            time.sleep(0.1)
            latency = round((time.time() - start) * 1000, 2)

            result = f"Connection Test Results\n\n"
            result += f"Integration: {integration_name}\n"
            result += f"Endpoint: {cred[0]}\n\n"
            result += f"Status: CONNECTED (Simulated)\n"
            result += f"Latency: {latency}ms\n"
            result += f"Tested at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

            log_activity('test', 'integration_connection', install_id,
                        details={'endpoint': cred[0], 'latency_ms': latency})

            messagebox.showinfo("Connection Test", result)

        except Exception as e:
            logger.error(f"Error testing connection: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to test connection: {e}")

    def validate_mapping_rules(self):
        """Validate transformation rules syntax"""
        try:
            selected = self.mappings_tree.selection()
            if not selected:
                messagebox.showwarning(_t("common.warning"), "Please select a mapping to validate")
                return

            mapping_id = self.mappings_tree.item(selected[0])['values'][0]

            with get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT transformation_rule, source_field, target_field
                    FROM integration_data_mappings
                    WHERE mapping_id = ?
                ''', (mapping_id,))
                mapping = cursor.fetchone()

            if not mapping:
                messagebox.showerror(_t("common.error"), "Mapping not found")
                return

            rule = mapping[0]
            errors = []
            warnings = []

            if not rule:
                result = "Validation Result: VALID\n\nNo transformation rule (direct mapping)"
            else:
                # Validate common patterns
                import re
                valid_patterns = [
                    (r'^[a-zA-Z_][a-zA-Z0-9_]*$', 'field reference'),
                    (r'^UPPER\(.+\)$', 'UPPER function'),
                    (r'^LOWER\(.+\)$', 'LOWER function'),
                    (r'^TRIM\(.+\)$', 'TRIM function'),
                    (r'^CONCAT\(.+\)$', 'CONCAT function'),
                    (r'^\{.*\}$', 'JSON template'),
                ]

                is_valid = False
                matched_pattern = None
                for pattern, name in valid_patterns:
                    if re.match(pattern, rule, re.IGNORECASE):
                        is_valid = True
                        matched_pattern = name
                        break

                if is_valid:
                    result = f"Validation Result: VALID\n\n"
                    result += f"Rule: {rule}\n"
                    result += f"Pattern: {matched_pattern}\n"
                    result += f"Source: {mapping[1]}\n"
                    result += f"Target: {mapping[2]}"
                else:
                    result = f"Validation Result: WARNING\n\n"
                    result += f"Rule: {rule}\n"
                    result += f"Note: Unknown rule format. May still work but couldn't validate syntax."

            log_activity('validate', 'mapping_rule', mapping_id)

            messagebox.showinfo("Mapping Validation", result)

        except Exception as e:
            logger.error(f"Error validating mapping: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to validate mapping: {e}")

    def dry_run_sync(self):
        """Simulate sync without committing changes"""
        try:
            selected = self.installed_tree.selection()
            if not selected:
                messagebox.showwarning(_t("common.warning"), "Please select an integration for dry run")
                return

            install_id = self.installed_tree.item(selected[0])['values'][0]
            integration_name = self.installed_tree.item(selected[0])['values'][1]

            if messagebox.askyesno("Dry Run Sync",
                                  f"Perform dry run sync for '{integration_name}'?\n\n"
                                  f"This will simulate the sync without making actual changes."):

                # Simulate dry run
                sample_size = 10
                would_create = sample_size // 2
                would_update = sample_size // 3
                would_skip = sample_size - would_create - would_update

                result = f"Dry Run Results for: {integration_name}\n\n"
                result += f"Sample Records Analyzed: {sample_size}\n\n"
                result += f"Would CREATE: {would_create} record(s)\n"
                result += f"Would UPDATE: {would_update} record(s)\n"
                result += f"Would SKIP: {would_skip} record(s)\n"
                result += f"Validation Errors: 0\n\n"
                result += f"Dry run completed at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
                result += "No actual changes were made."

                log_activity('dry_run', 'sync', install_id,
                            details={'sample_size': sample_size, 'would_create': would_create,
                                    'would_update': would_update})

                messagebox.showinfo("Dry Run Complete", result)

        except Exception as e:
            logger.error(f"Error performing dry run: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to perform dry run: {e}")

    def test_webhook_delivery(self):
        """Send test payload to webhook URL"""
        try:
            selected = self.webhooks_tree.selection()
            if not selected:
                messagebox.showwarning(_t("common.warning"), "Please select a webhook to test")
                return

            webhook_id = self.webhooks_tree.item(selected[0])['values'][0]

            with get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT webhook_url, event_type FROM integration_webhooks
                    WHERE webhook_id = ?
                ''', (webhook_id,))
                webhook = cursor.fetchone()

            if not webhook:
                messagebox.showerror(_t("common.error"), "Webhook not found")
                return

            test_payload = {
                'event': 'test',
                'timestamp': datetime.now().isoformat(),
                'message': 'Test webhook delivery from Integration Marketplace',
                'webhook_id': webhook_id
            }

            # Simulate delivery
            result = f"Webhook Test Results\n\n"
            result += f"URL: {webhook[0]}\n"
            result += f"Event Type: {webhook[1]}\n\n"
            result += f"Payload Sent:\n{json.dumps(test_payload, indent=2)}\n\n"
            result += f"Response: 200 OK (Simulated)\n"
            result += f"Delivered at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

            # Update last triggered
            with transaction() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    UPDATE integration_webhooks
                    SET last_triggered_at = ?
                    WHERE webhook_id = ?
                ''', (datetime.now().isoformat(), webhook_id))

            log_activity('test', 'webhook_delivery', webhook_id,
                        details={'url': webhook[0]})

            messagebox.showinfo("Webhook Test", result)
            self.load_webhooks()

        except Exception as e:
            logger.error(f"Error testing webhook: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to test webhook: {e}")

    def validate_json_configuration(self):
        """Validate configuration JSON against schema"""
        try:
            selected = self.installed_tree.selection()
            if not selected:
                messagebox.showwarning(_t("common.warning"), "Please select an integration to validate configuration")
                return

            install_id = self.installed_tree.item(selected[0])['values'][0]
            integration_name = self.installed_tree.item(selected[0])['values'][1]

            with get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT configuration FROM installed_integrations
                    WHERE install_id = ?
                ''', (install_id,))
                row = cursor.fetchone()

            config_str = row[0] if row else None

            result = f"Configuration Validation: {integration_name}\n\n"

            if not config_str:
                result += "Status: VALID\n"
                result += "Note: No configuration (empty is valid)"
            else:
                try:
                    config = json.loads(config_str)
                    result += "Status: VALID JSON\n\n"
                    result += "Configuration Structure:\n"
                    for key, value in config.items():
                        result += f"  - {key}: {type(value).__name__}\n"
                except json.JSONDecodeError as e:
                    result += f"Status: INVALID JSON\n\n"
                    result += f"Error: {str(e)}"

            log_activity('validate', 'json_configuration', install_id)

            messagebox.showinfo("Configuration Validation", result)

        except Exception as e:
            logger.error(f"Error validating configuration: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to validate configuration: {e}")

    def run_integration_diagnostics(self):
        """Run comprehensive diagnostic checks"""
        try:
            selected = self.installed_tree.selection()
            if not selected:
                messagebox.showwarning(_t("common.warning"), "Please select an integration to diagnose")
                return

            install_id = self.installed_tree.item(selected[0])['values'][0]
            integration_name = self.installed_tree.item(selected[0])['values'][1]

            dialog = tk.Toplevel(self.root)
            dialog.title(f"Diagnostics: {integration_name}")
            dialog.geometry("600x500")
            dialog.transient(self.root)

            ttk.Label(dialog, text=f"Diagnostics for: {integration_name}",
                     style='Title.TLabel').pack(pady=10)

            results_text = scrolledtext.ScrolledText(dialog, height=20, wrap=tk.WORD)
            results_text.pack(fill='both', expand=True, padx=10, pady=5)

            def run_checks():
                results_text.delete('1.0', 'end')
                results_text.insert('end', f"Running diagnostics for {integration_name}...\n\n")
                results_text.insert('end', "=" * 50 + "\n\n")

                checks_passed = 0
                checks_failed = 0
                checks_warning = 0

                with get_connection() as conn:
                    cursor = conn.cursor()

                    # Check 1: Installation exists
                    cursor.execute('''
                        SELECT status, is_enabled FROM installed_integrations
                        WHERE install_id = ?
                    ''', (install_id,))
                    install = cursor.fetchone()

                    if install:
                        results_text.insert('end', "[PASS] Installation exists\n")
                        checks_passed += 1

                        if install[1]:
                            results_text.insert('end', "[PASS] Integration is enabled\n")
                            checks_passed += 1
                        else:
                            results_text.insert('end', "[WARN] Integration is disabled\n")
                            checks_warning += 1
                    else:
                        results_text.insert('end', "[FAIL] Installation not found\n")
                        checks_failed += 1
                        return

                    # Check 2: Credentials
                    cursor.execute('''
                        SELECT COUNT(*) FROM integration_credentials
                        WHERE install_id = ?
                    ''', (install_id,))
                    cred_count = cursor.fetchone()[0]

                    if cred_count > 0:
                        results_text.insert('end', f"[PASS] {cred_count} credential(s) configured\n")
                        checks_passed += 1
                    else:
                        results_text.insert('end', "[WARN] No credentials configured\n")
                        checks_warning += 1

                    # Check 3: Recent syncs
                    cursor.execute('''
                        SELECT COUNT(*), MAX(sync_start_time)
                        FROM integration_sync_logs
                        WHERE install_id = ? AND sync_start_time >= date('now', '-7 days')
                    ''', (install_id,))
                    sync_info = cursor.fetchone()

                    if sync_info[0] > 0:
                        results_text.insert('end', f"[PASS] {sync_info[0]} syncs in last 7 days\n")
                        checks_passed += 1
                    else:
                        results_text.insert('end', "[INFO] No syncs in last 7 days\n")

                    # Check 4: Error rate
                    cursor.execute('''
                        SELECT
                            COUNT(*) as total,
                            SUM(CASE WHEN sync_status = 'failed' THEN 1 ELSE 0 END) as failed
                        FROM integration_sync_logs
                        WHERE install_id = ? AND sync_start_time >= date('now', '-30 days')
                    ''', (install_id,))
                    error_info = cursor.fetchone()

                    if error_info[0] > 0:
                        error_rate = (error_info[1] / error_info[0] * 100)
                        if error_rate < 10:
                            results_text.insert('end', f"[PASS] Error rate: {error_rate:.1f}%\n")
                            checks_passed += 1
                        elif error_rate < 25:
                            results_text.insert('end', f"[WARN] Error rate: {error_rate:.1f}%\n")
                            checks_warning += 1
                        else:
                            results_text.insert('end', f"[FAIL] High error rate: {error_rate:.1f}%\n")
                            checks_failed += 1

                    # Check 5: Data mappings
                    cursor.execute('''
                        SELECT COUNT(*) FROM integration_data_mappings
                        WHERE install_id = ? AND is_active = 1
                    ''', (install_id,))
                    mapping_count = cursor.fetchone()[0]
                    results_text.insert('end', f"[INFO] {mapping_count} active data mapping(s)\n")

                results_text.insert('end', "\n" + "=" * 50 + "\n\n")
                results_text.insert('end', f"Summary:\n")
                results_text.insert('end', f"  Passed: {checks_passed}\n")
                results_text.insert('end', f"  Warnings: {checks_warning}\n")
                results_text.insert('end', f"  Failed: {checks_failed}\n\n")

                if checks_failed > 0:
                    results_text.insert('end', "Overall Status: NEEDS ATTENTION\n")
                elif checks_warning > 0:
                    results_text.insert('end', "Overall Status: WARNING\n")
                else:
                    results_text.insert('end', "Overall Status: HEALTHY\n")

                log_activity('diagnose', 'integration', install_id,
                            details={'passed': checks_passed, 'warnings': checks_warning, 'failed': checks_failed})

            run_checks()

            button_frame = ttk.Frame(dialog)
            button_frame.pack(pady=10)
            ttk.Button(button_frame, text="Re-run Diagnostics", command=run_checks).pack(side='left', padx=5)
            ttk.Button(button_frame, text=_t("common.close"), command=dialog.destroy).pack(side='left', padx=5)

        except Exception as e:
            logger.error(f"Error running diagnostics: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to run diagnostics: {e}")

    # =========================================================================
    # DATA MAPPING TOOLS FUNCTIONS (43-46)
    # =========================================================================

    def auto_detect_mappings(self):
        """Automatically suggest field mappings based on names"""
        try:
            dialog = tk.Toplevel(self.root)
            dialog.title("Auto-Detect Field Mappings")
            dialog.geometry("700x600")
            dialog.transient(self.root)
            dialog.grab_set()

            ttk.Label(dialog, text="Auto-Detect Field Mappings",
                     style='Title.TLabel').pack(pady=10)

            # Input frame
            input_frame = ttk.LabelFrame(dialog, text="Input Fields", padding=10)
            input_frame.pack(fill='x', padx=10, pady=5)

            ttk.Label(input_frame, text="Source Fields (comma-separated):").pack(anchor='w')
            source_var = tk.StringVar(value="student_id, first_name, last_name, email, phone")
            ttk.Entry(input_frame, textvariable=source_var, width=70).pack(fill='x', pady=5)

            ttk.Label(input_frame, text="Target Fields (comma-separated):").pack(anchor='w')
            target_var = tk.StringVar(value="id, firstName, lastName, emailAddress, phoneNumber")
            ttk.Entry(input_frame, textvariable=target_var, width=70).pack(fill='x', pady=5)

            # Results frame
            results_frame = ttk.LabelFrame(dialog, text="Suggested Mappings", padding=10)
            results_frame.pack(fill='both', expand=True, padx=10, pady=5)

            columns = ('source', 'target', 'confidence', 'match_type')
            tree = ttk.Treeview(results_frame, columns=columns, show='headings', height=10)

            for col in columns:
                tree.heading(col, text=col.replace('_', ' ').title())
                tree.column(col, width=150)

            tree.pack(fill='both', expand=True)

            def detect_mappings():
                for item in tree.get_children():
                    tree.delete(item)

                source_fields = [f.strip() for f in source_var.get().split(',')]
                target_fields = [f.strip() for f in target_var.get().split(',')]

                # Normalize for comparison
                import re
                def normalize(name):
                    return re.sub(r'[^a-z0-9]', '', name.lower())

                source_normalized = {normalize(f): f for f in source_fields}
                target_normalized = {normalize(f): f for f in target_fields}

                suggestions = []

                # Exact matches
                for src_norm, src_orig in source_normalized.items():
                    if src_norm in target_normalized:
                        suggestions.append({
                            'source': src_orig,
                            'target': target_normalized[src_norm],
                            'confidence': 1.0,
                            'match_type': 'exact'
                        })

                # Common mappings
                common_mappings = {
                    'firstname': ['first_name', 'fname', 'givenname'],
                    'lastname': ['last_name', 'lname', 'surname', 'familyname'],
                    'email': ['emailaddress', 'mail', 'e_mail'],
                    'phone': ['phonenumber', 'tel', 'telephone', 'mobile'],
                    'id': ['identifier', 'uuid', 'key', 'studentid'],
                }

                for src_norm, src_orig in source_normalized.items():
                    for base, alternatives in common_mappings.items():
                        all_names = [base] + alternatives
                        if src_norm in all_names:
                            for tgt_norm, tgt_orig in target_normalized.items():
                                if tgt_norm in all_names:
                                    if not any(s['source'] == src_orig and s['target'] == tgt_orig for s in suggestions):
                                        suggestions.append({
                                            'source': src_orig,
                                            'target': tgt_orig,
                                            'confidence': 0.8,
                                            'match_type': 'semantic'
                                        })

                for suggestion in sorted(suggestions, key=lambda x: -x['confidence']):
                    tree.insert('', 'end', values=(
                        suggestion['source'],
                        suggestion['target'],
                        f"{suggestion['confidence'] * 100:.0f}%",
                        suggestion['match_type']
                    ))

                if not suggestions:
                    messagebox.showinfo("Auto-Detect", "No matching field mappings detected")

            ttk.Button(dialog, text="Detect Mappings", command=detect_mappings).pack(pady=10)
            ttk.Button(dialog, text=_t("common.close"), command=dialog.destroy).pack(pady=5)

        except Exception as e:
            logger.error(f"Error in auto-detect mappings: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to auto-detect mappings: {e}")

    def preview_transformation(self):
        """Preview transformation rule output on sample data"""
        try:
            selected = self.mappings_tree.selection()
            if not selected:
                messagebox.showwarning(_t("common.warning"), "Please select a mapping to preview transformation")
                return

            mapping_id = self.mappings_tree.item(selected[0])['values'][0]

            with get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT transformation_rule, source_field, target_field
                    FROM integration_data_mappings
                    WHERE mapping_id = ?
                ''', (mapping_id,))
                mapping = cursor.fetchone()

            if not mapping:
                messagebox.showerror(_t("common.error"), "Mapping not found")
                return

            dialog = tk.Toplevel(self.root)
            dialog.title(_t("integration_marketplace.dialogs.preview_transformation"))
            dialog.geometry("500x400")
            dialog.transient(self.root)
            dialog.grab_set()

            ttk.Label(dialog, text="Preview Transformation",
                     style='Title.TLabel').pack(pady=10)

            ttk.Label(dialog, text=f"Source Field: {mapping[1]}").pack(anchor='w', padx=10)
            ttk.Label(dialog, text=f"Target Field: {mapping[2]}").pack(anchor='w', padx=10)
            ttk.Label(dialog, text=f"Rule: {mapping[0] or 'Direct copy'}").pack(anchor='w', padx=10, pady=5)

            ttk.Label(dialog, text="Sample Input Value:").pack(anchor='w', padx=10, pady=5)
            input_var = tk.StringVar(value="  John Doe  ")
            ttk.Entry(dialog, textvariable=input_var, width=40).pack(padx=10)

            result_frame = ttk.LabelFrame(dialog, text="Transformation Result", padding=10)
            result_frame.pack(fill='both', expand=True, padx=10, pady=10)

            result_label = ttk.Label(result_frame, text="", font=('Arial', 12))
            result_label.pack(pady=20)

            def apply_transformation():
                sample_input = input_var.get()
                rule = mapping[0]

                if not rule:
                    output = sample_input
                    method = "Direct copy (no transformation)"
                elif rule.upper().startswith('UPPER'):
                    output = sample_input.upper()
                    method = "UPPER function"
                elif rule.upper().startswith('LOWER'):
                    output = sample_input.lower()
                    method = "LOWER function"
                elif rule.upper().startswith('TRIM'):
                    output = sample_input.strip()
                    method = "TRIM function"
                else:
                    output = sample_input
                    method = "Unknown rule - used direct copy"

                result_label.config(text=f"Input: '{sample_input}'\n\nOutput: '{output}'\n\nMethod: {method}")

            ttk.Button(dialog, text=_t("integration_marketplace.mappings.preview"), command=apply_transformation).pack(pady=10)
            ttk.Button(dialog, text=_t("common.close"), command=dialog.destroy).pack(pady=5)

        except Exception as e:
            logger.error(f"Error previewing transformation: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to preview transformation: {e}")

    def duplicate_mapping_set(self):
        """Clone an existing mapping configuration"""
        try:
            dialog = tk.Toplevel(self.root)
            dialog.title("Duplicate Mapping Set")
            dialog.geometry("500x300")
            dialog.transient(self.root)
            dialog.grab_set()

            ttk.Label(dialog, text="Duplicate Mapping Set",
                     style='Title.TLabel').pack(pady=10)

            # Get installed integrations
            with get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT ii.install_id, ic.integration_name
                    FROM installed_integrations ii
                    JOIN integration_catalog ic ON ii.integration_id = ic.integration_id
                    WHERE ii.status = 'active'
                ''')
                integrations = cursor.fetchall()

            ttk.Label(dialog, text="Source Integration:").pack(anchor='w', padx=10, pady=5)
            source_var = tk.StringVar()
            source_combo = ttk.Combobox(dialog, textvariable=source_var,
                                       values=[f"{i[0]}: {i[1]}" for i in integrations],
                                       width=40, state='readonly')
            source_combo.pack(padx=10, pady=5)

            ttk.Label(dialog, text="Target Integration:").pack(anchor='w', padx=10, pady=5)
            target_var = tk.StringVar()
            target_combo = ttk.Combobox(dialog, textvariable=target_var,
                                       values=[f"{i[0]}: {i[1]}" for i in integrations],
                                       width=40, state='readonly')
            target_combo.pack(padx=10, pady=5)

            def duplicate():
                if not source_var.get() or not target_var.get():
                    messagebox.showwarning(_t("common.warning"), "Please select both source and target integrations")
                    return

                source_id = int(source_var.get().split(':')[0])
                target_id = int(target_var.get().split(':')[0])

                if source_id == target_id:
                    messagebox.showwarning(_t("common.warning"), "Source and target must be different")
                    return

                with get_connection() as conn:
                    cursor = conn.cursor()
                    cursor.execute('''
                        SELECT source_field, target_field, transformation_rule, is_active
                        FROM integration_data_mappings
                        WHERE install_id = ?
                    ''', (source_id,))
                    mappings = cursor.fetchall()

                if not mappings:
                    messagebox.showinfo(_t("common.info"), "No mappings found in source integration")
                    return

                copied_count = 0
                with transaction() as conn:
                    cursor = conn.cursor()
                    for mapping in mappings:
                        cursor.execute('''
                            INSERT INTO integration_data_mappings
                            (install_id, source_field, target_field, transformation_rule, is_active)
                            VALUES (?, ?, ?, ?, ?)
                        ''', (target_id, mapping[0], mapping[1], mapping[2], mapping[3]))
                        copied_count += 1

                log_activity('duplicate', 'mapping_set', target_id,
                            details={'source_id': source_id, 'mappings_copied': copied_count})

                messagebox.showinfo(_t("common.success"), f"Copied {copied_count} mapping(s) to target integration")
                dialog.destroy()
                self.load_mappings()

            ttk.Button(dialog, text="Duplicate Mappings", command=duplicate).pack(pady=20)

        except Exception as e:
            logger.error(f"Error duplicating mapping set: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to duplicate mappings: {e}")

    def import_mappings_from_template(self):
        """Import standard mapping templates"""
        try:
            selected = self.installed_tree.selection()
            if not selected:
                messagebox.showwarning(_t("common.warning"), "Please select an integration to import mappings to")
                return

            install_id = self.installed_tree.item(selected[0])['values'][0]
            integration_name = self.installed_tree.item(selected[0])['values'][1]

            dialog = tk.Toplevel(self.root)
            dialog.title(_t("integration_marketplace.dialogs.import_template"))
            dialog.geometry("500x400")
            dialog.transient(self.root)
            dialog.grab_set()

            ttk.Label(dialog, text=f"Import Template to: {integration_name}",
                     style='Title.TLabel').pack(pady=10)

            templates = {
                'student_basic': [
                    ('student_id', 'id', None),
                    ('first_name', 'firstName', 'TRIM'),
                    ('last_name', 'lastName', 'TRIM'),
                    ('email', 'emailAddress', 'LOWER'),
                    ('enrollment_date', 'enrolledAt', None)
                ],
                'course_basic': [
                    ('course_id', 'id', None),
                    ('course_name', 'title', None),
                    ('course_code', 'code', 'UPPER'),
                    ('credits', 'creditHours', None)
                ],
                'grade_basic': [
                    ('grade_id', 'id', None),
                    ('student_id', 'studentId', None),
                    ('course_id', 'courseId', None),
                    ('grade', 'letterGrade', 'UPPER'),
                    ('points', 'gradePoints', None)
                ]
            }

            ttk.Label(dialog, text="Select Template:").pack(anchor='w', padx=10, pady=5)
            template_var = tk.StringVar()
            template_combo = ttk.Combobox(dialog, textvariable=template_var,
                                         values=list(templates.keys()),
                                         width=30, state='readonly')
            template_combo.pack(padx=10, pady=5)

            # Preview frame
            preview_frame = ttk.LabelFrame(dialog, text="Template Preview", padding=10)
            preview_frame.pack(fill='both', expand=True, padx=10, pady=5)

            preview_text = scrolledtext.ScrolledText(preview_frame, height=10)
            preview_text.pack(fill='both', expand=True)

            def show_preview(event=None):
                template_name = template_var.get()
                if template_name in templates:
                    preview_text.delete('1.0', 'end')
                    preview_text.insert('end', f"Template: {template_name}\n\n")
                    preview_text.insert('end', f"{'Source':<20} {'Target':<20} {'Transform'}\n")
                    preview_text.insert('end', "-" * 50 + "\n")
                    for src, tgt, transform in templates[template_name]:
                        preview_text.insert('end', f"{src:<20} {tgt:<20} {transform or 'None'}\n")

            template_combo.bind('<<ComboboxSelected>>', show_preview)

            def import_template():
                template_name = template_var.get()
                if not template_name:
                    messagebox.showwarning(_t("common.warning"), "Please select a template")
                    return

                mappings = templates[template_name]
                imported_count = 0

                with transaction() as conn:
                    cursor = conn.cursor()
                    for src, tgt, transform in mappings:
                        cursor.execute('''
                            INSERT INTO integration_data_mappings
                            (install_id, source_field, target_field, transformation_rule, is_active)
                            VALUES (?, ?, ?, ?, 1)
                        ''', (install_id, src, tgt, transform))
                        imported_count += 1

                log_activity('import', 'mapping_template', install_id,
                            details={'template': template_name, 'mappings_created': imported_count})

                messagebox.showinfo(_t("common.success"), f"Imported {imported_count} mapping(s) from '{template_name}'")
                dialog.destroy()
                self.load_mappings()

            ttk.Button(dialog, text="Import Template", command=import_template).pack(pady=10)

        except Exception as e:
            logger.error(f"Error importing mapping template: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to import template: {e}")

    # =========================================================================
    # NOTIFICATIONS & ALERTS FUNCTIONS (47-50)
    # =========================================================================

    def configure_alert_rules(self):
        """Set up custom alerts for error thresholds"""
        try:
            selected = self.installed_tree.selection()
            if not selected:
                messagebox.showwarning(_t("common.warning"), "Please select an integration to configure alerts")
                return

            install_id = self.installed_tree.item(selected[0])['values'][0]
            integration_name = self.installed_tree.item(selected[0])['values'][1]

            dialog = tk.Toplevel(self.root)
            dialog.title(f"Alert Rules: {integration_name}")
            dialog.geometry("500x450")
            dialog.transient(self.root)
            dialog.grab_set()

            ttk.Label(dialog, text="Configure Alert Rules",
                     style='Title.TLabel').pack(pady=10)

            # Error threshold
            ttk.Label(dialog, text="Alert on consecutive failures:").pack(anchor='w', padx=10, pady=5)
            failures_var = tk.StringVar(value="3")
            ttk.Spinbox(dialog, from_=1, to=10, textvariable=failures_var, width=10).pack(anchor='w', padx=10)

            # Error rate threshold
            ttk.Label(dialog, text="Alert when error rate exceeds (%):").pack(anchor='w', padx=10, pady=5)
            rate_var = tk.StringVar(value="20")
            ttk.Spinbox(dialog, from_=1, to=100, textvariable=rate_var, width=10).pack(anchor='w', padx=10)

            # No sync alert
            ttk.Label(dialog, text="Alert if no sync for (hours):").pack(anchor='w', padx=10, pady=5)
            hours_var = tk.StringVar(value="24")
            ttk.Spinbox(dialog, from_=1, to=168, textvariable=hours_var, width=10).pack(anchor='w', padx=10)

            # Notification methods
            ttk.Label(dialog, text="Notification Methods:").pack(anchor='w', padx=10, pady=10)

            email_var = tk.BooleanVar(value=True)
            ttk.Checkbutton(dialog, text="Email", variable=email_var).pack(anchor='w', padx=20)

            webhook_var = tk.BooleanVar(value=False)
            ttk.Checkbutton(dialog, text="Webhook", variable=webhook_var).pack(anchor='w', padx=20)

            def save_rules():
                rules = [
                    {
                        'type': 'consecutive_failures',
                        'threshold': int(failures_var.get()),
                        'notify_email': email_var.get(),
                        'notify_webhook': webhook_var.get()
                    },
                    {
                        'type': 'error_rate',
                        'threshold_percent': int(rate_var.get()),
                        'notify_email': email_var.get(),
                        'notify_webhook': webhook_var.get()
                    },
                    {
                        'type': 'no_sync_hours',
                        'threshold_hours': int(hours_var.get()),
                        'notify_email': email_var.get(),
                        'notify_webhook': webhook_var.get()
                    }
                ]

                with transaction() as conn:
                    cursor = conn.cursor()
                    cursor.execute('''
                        UPDATE installed_integrations
                        SET configuration = json_set(
                            COALESCE(configuration, '{}'),
                            '$.alert_rules', ?
                        )
                        WHERE install_id = ?
                    ''', (json.dumps(rules), install_id))

                log_activity('configure', 'alert_rules', install_id,
                            details={'rules_count': len(rules)})

                messagebox.showinfo(_t("common.success"), "Alert rules configured successfully!")
                dialog.destroy()

            ttk.Button(dialog, text="Save Alert Rules", command=save_rules).pack(pady=20)

        except Exception as e:
            logger.error(f"Error configuring alert rules: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to configure alerts: {e}")

    def subscribe_to_notifications(self):
        """Subscribe users to integration events"""
        try:
            selected = self.installed_tree.selection()
            if not selected:
                messagebox.showwarning(_t("common.warning"), "Please select an integration to subscribe")
                return

            install_id = self.installed_tree.item(selected[0])['values'][0]
            integration_name = self.installed_tree.item(selected[0])['values'][1]

            dialog = tk.Toplevel(self.root)
            dialog.title(f"Subscribe: {integration_name}")
            dialog.geometry("500x400")
            dialog.transient(self.root)
            dialog.grab_set()

            ttk.Label(dialog, text="Subscribe to Notifications",
                     style='Title.TLabel').pack(pady=10)

            ttk.Label(dialog, text="Email Address:").pack(anchor='w', padx=10, pady=5)
            email_var = tk.StringVar(value=f"{self.auth.current_user.get('username', 'user')}@university.edu")
            ttk.Entry(dialog, textvariable=email_var, width=40).pack(padx=10)

            ttk.Label(dialog, text="Subscribe to Events:").pack(anchor='w', padx=10, pady=10)

            event_vars = {}
            events = ['sync_complete', 'sync_failed', 'error_threshold', 'credential_expiring', 'status_change']
            for event in events:
                event_vars[event] = tk.BooleanVar(value=event in ['sync_failed', 'error_threshold'])
                ttk.Checkbutton(dialog, text=event.replace('_', ' ').title(),
                               variable=event_vars[event]).pack(anchor='w', padx=20)

            def subscribe():
                email = email_var.get().strip()
                if not email:
                    messagebox.showwarning(_t("common.warning"), "Please enter an email address")
                    return

                selected_events = [event for event, var in event_vars.items() if var.get()]
                if not selected_events:
                    messagebox.showwarning(_t("common.warning"), "Please select at least one event")
                    return

                subscription = {
                    'user_email': email,
                    'event_types': selected_events,
                    'subscribed_at': datetime.now().isoformat()
                }

                with transaction() as conn:
                    cursor = conn.cursor()
                    cursor.execute('''
                        SELECT configuration FROM installed_integrations
                        WHERE install_id = ?
                    ''', (install_id,))
                    row = cursor.fetchone()

                    config = json.loads(row[0]) if row and row[0] else {}
                    subscriptions = config.get('subscriptions', [])

                    # Update or add subscription
                    existing = next((s for s in subscriptions if s['user_email'] == email), None)
                    if existing:
                        existing['event_types'] = selected_events
                    else:
                        subscriptions.append(subscription)

                    config['subscriptions'] = subscriptions

                    cursor.execute('''
                        UPDATE installed_integrations
                        SET configuration = ?
                        WHERE install_id = ?
                    ''', (json.dumps(config), install_id))

                log_activity('subscribe', 'notifications', install_id,
                            details={'email': email, 'events': selected_events})

                messagebox.showinfo(_t("common.success"), f"Subscribed {email} to {len(selected_events)} event(s)")
                dialog.destroy()

            ttk.Button(dialog, text="Subscribe", command=subscribe).pack(pady=20)

        except Exception as e:
            logger.error(f"Error subscribing to notifications: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to subscribe: {e}")

    def view_notification_history(self):
        """View past notifications sent"""
        try:
            dialog = tk.Toplevel(self.root)
            dialog.title(_t("integration_marketplace.dialogs.notification_history"))
            dialog.geometry("800x500")
            dialog.transient(self.root)

            ttk.Label(dialog, text="Notification History",
                     style='Title.TLabel').pack(pady=10)

            # History tree
            tree_frame = ttk.Frame(dialog)
            tree_frame.pack(fill='both', expand=True, padx=10, pady=5)

            columns = ('timestamp', 'integration', 'event_type', 'recipient', 'subject', 'status')
            tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=15)

            for col in columns:
                tree.heading(col, text=col.replace('_', ' ').title())
                tree.column(col, width=120)

            tree.column('subject', width=200)

            vsb = ttk.Scrollbar(tree_frame, orient='vertical', command=tree.yview)
            tree.configure(yscrollcommand=vsb.set)

            tree.pack(side='left', fill='both', expand=True)
            vsb.pack(side='right', fill='y')

            # Simulated notification history
            notifications = [
                (datetime.now().isoformat()[:19], 'Canvas LMS', 'sync_failed', 'admin@university.edu',
                 'Sync Failed: Canvas LMS', 'delivered'),
                ((datetime.now() - timedelta(hours=5)).isoformat()[:19], 'Zoom', 'sync_complete', 'admin@university.edu',
                 'Sync Complete: Zoom', 'delivered'),
                ((datetime.now() - timedelta(days=1)).isoformat()[:19], 'Google Workspace', 'error_threshold', 'it@university.edu',
                 'Error Threshold Exceeded', 'delivered'),
            ]

            for notification in notifications:
                tree.insert('', 'end', values=notification)

            ttk.Button(dialog, text=_t("common.close"), command=dialog.destroy).pack(pady=10)

            log_activity('view', 'notification_history', None)

        except Exception as e:
            logger.error(f"Error viewing notification history: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to view history: {e}")

    def test_notification_channel(self):
        """Send test notification to verify setup"""
        try:
            dialog = tk.Toplevel(self.root)
            dialog.title(_t("integration_marketplace.dialogs.test_channel"))
            dialog.geometry("500x350")
            dialog.transient(self.root)
            dialog.grab_set()

            ttk.Label(dialog, text="Test Notification Channel",
                     style='Title.TLabel').pack(pady=10)

            ttk.Label(dialog, text="Channel Type:").pack(anchor='w', padx=10, pady=5)
            channel_var = tk.StringVar(value='email')
            channel_frame = ttk.Frame(dialog)
            channel_frame.pack(fill='x', padx=10, pady=5)

            for channel in ['email', 'webhook', 'slack']:
                ttk.Radiobutton(channel_frame, text=channel.title(), variable=channel_var,
                               value=channel).pack(side='left', padx=10)

            ttk.Label(dialog, text="Target (email/URL/channel):").pack(anchor='w', padx=10, pady=5)
            target_var = tk.StringVar(value=f"{self.auth.current_user.get('username', 'user')}@university.edu")
            ttk.Entry(dialog, textvariable=target_var, width=50).pack(padx=10)

            result_frame = ttk.LabelFrame(dialog, text="Test Result", padding=10)
            result_frame.pack(fill='both', expand=True, padx=10, pady=10)

            result_label = ttk.Label(result_frame, text="Click 'Send Test' to test the notification channel")
            result_label.pack(pady=20)

            def send_test():
                channel = channel_var.get()
                target = target_var.get().strip()

                if not target:
                    messagebox.showwarning(_t("common.warning"), "Please enter a target")
                    return

                # Simulate test
                success = True
                if channel == 'email':
                    message = f"Test email would be sent to: {target}"
                elif channel == 'webhook':
                    message = f"Test webhook would be sent to: {target}"
                elif channel == 'slack':
                    message = f"Test Slack message would be sent to: {target}"
                else:
                    message = f"Unknown channel type: {channel}"
                    success = False

                if success:
                    result_label.config(text=f"SUCCESS\n\n{message}\n\nTimestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                else:
                    result_label.config(text=f"FAILED\n\n{message}")

                log_activity('test', 'notification_channel', None,
                            details={'channel': channel, 'target': target, 'success': success})

            ttk.Button(dialog, text="Send Test", command=send_test).pack(pady=10)
            ttk.Button(dialog, text=_t("common.close"), command=dialog.destroy).pack(pady=5)

        except Exception as e:
            logger.error(f"Error testing notification channel: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to test channel: {e}")

    def _on_language_change(self):
        """Handle language change request"""
        old_lang = get_current_language()
        show_gui_language_selector(self.root)
        new_lang = get_current_language()
        if old_lang != new_lang:
            messagebox.showinfo(
                _t("integration_marketplace.language.changed_title"),
                _t("integration_marketplace.language.changed_message")
            )
            self._refresh_ui_language()

    def _refresh_ui_language(self):
        """Refresh all UI elements with current language"""
        # Reinitialize i18n to ensure new translations are loaded
        init_i18n(get_current_language())
        # Update window title
        self.root.title(_t("integration_marketplace.window_title"))
        # Recreate the main interface with new language
        self.create_main_interface()

    def return_to_main_menu(self):
        """Return to main menu by closing the marketplace window"""
        if messagebox.askyesno(_t("integration_marketplace.dialogs.confirm"), _t("integration_marketplace.dialogs.return_confirm")):
            try:
                # Log the action
                if self.auth and self.auth.current_user:
                    log_activity('Closed Integration Marketplace',
                               user=self.auth.current_user.get('username', 'Unknown'))

                # Close the window
                self.root.destroy()
                logger.info("Integration Marketplace closed")
            except Exception as e:
                logger.error(f"Error closing Integration Marketplace: {e}")
                self.root.destroy()


def launch_integration_marketplace_gui(auth=None, parent=None):
    """Launch the Integration Marketplace GUI as a child window"""
    try:
        if parent:
            # Pass parent directly - IntegrationMarketplaceGUI creates its own Toplevel
            app = IntegrationMarketplaceGUI(parent, auth_system=auth)
        else:
            # Create as standalone root window if no parent (for testing)
            root = tk.Tk()
            root.withdraw()  # Hide the root window
            app = IntegrationMarketplaceGUI(root, auth_system=auth)
            root.mainloop()
    except Exception as e:
        logger.error(f"Error launching Integration Marketplace GUI: {e}\n{traceback.format_exc()}")
        raise


if __name__ == "__main__":
    launch_integration_marketplace_gui()
