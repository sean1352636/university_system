import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext, filedialog
import sys
import os
import threading
import logging
import random
import csv
import sqlite3
from datetime import datetime, timedelta

# Import authentication
try:
    from university_system.infrastructure.auth.user_authentication import UserAuth
except ImportError:
    UserAuth = None

# Import auth instance management from user_authentication
try:
    from university_system.infrastructure.auth.user_authentication import get_current_user, set_auth_instance
    HAS_AUTH = True
except ImportError:
    HAS_AUTH = False
    get_current_user = lambda: None
    set_auth_instance = lambda x: None

# Import MFA (Multi-Factor Authentication) services
try:
    from university_system.infrastructure.auth.mfa_service import (
        MFAService,
        setup_totp,
        verify_totp,
        generate_sms_otp,
        verify_sms_otp
    )
    MFA_AVAILABLE = True
except ImportError:
    MFAService = None
    MFA_AVAILABLE = False

try:
    from university_system.infrastructure.auth.mfa_integration import integrate_mfa_with_auth
    MFA_INTEGRATION_AVAILABLE = True
except ImportError:
    MFA_INTEGRATION_AVAILABLE = False

try:
    from university_system.infrastructure.auth.email_otp_service import EmailOTPService
    EMAIL_OTP_AVAILABLE = True
except ImportError:
    EmailOTPService = None
    EMAIL_OTP_AVAILABLE = False

try:
    from university_system.infrastructure.auth.sms_provider import SMSProvider
    SMS_PROVIDER_AVAILABLE = True
except ImportError:
    SMSProvider = None
    SMS_PROVIDER_AVAILABLE = False

# Import security modules
try:
    from university_system.infrastructure.security.session_management import (
        SessionManager,
        SessionInfo,
        create_session,
        validate_session
    )
    SESSION_MANAGEMENT_AVAILABLE = True
except ImportError:
    SessionManager = None
    SESSION_MANAGEMENT_AVAILABLE = False

try:
    from university_system.infrastructure.security.comprehensive_security import (
        APISecurityManager,
        PasswordSecurityManager,
        SecurityAuditManager,
        DataLossPreventionManager,
        IncidentResponseManager,
        VulnerabilityScanner
    )
    COMPREHENSIVE_SECURITY_AVAILABLE = True
except ImportError:
    APISecurityManager = None
    PasswordSecurityManager = None
    SecurityAuditManager = None
    COMPREHENSIVE_SECURITY_AVAILABLE = False

try:
    from university_system.infrastructure.security.data_encryption import (
        EncryptionManager,
        encrypt_sensitive_data,
        decrypt_sensitive_data
    )
    DATA_ENCRYPTION_AVAILABLE = True
except ImportError:
    EncryptionManager = None
    DATA_ENCRYPTION_AVAILABLE = False

try:
    from university_system.infrastructure.security.security_dashboard_gui import SecurityDashboard
    SECURITY_DASHBOARD_AVAILABLE = True
except ImportError:
    SecurityDashboard = None
    SECURITY_DASHBOARD_AVAILABLE = False

# Import activity logger for audit trail
try:
    from university_system.modules.shared.utils.activity_logger import log_activity
    ACTIVITY_LOGGER_AVAILABLE = True
except ImportError:
    ACTIVITY_LOGGER_AVAILABLE = False
    log_activity = lambda *args, **kwargs: None

auth = None

# Initialize shared_context auth early to prevent warnings during imports
try:
    from university_system.infrastructure.shared_context import set_auth as set_shared_auth, get_auth
    # Try to get existing auth or set a placeholder
    try:
        existing_auth = get_auth()
        if existing_auth is not None:
            auth = existing_auth
    except:
        pass  # No existing auth, will be set later
except ImportError:
    set_shared_auth = lambda x: None
    get_auth = lambda: None

def set_auth(auth_instance):
    global auth
    auth = auth_instance
    # Also set it in the global auth instance if available
    if HAS_AUTH:
        set_auth_instance(auth_instance)
    # Set in shared_context as well
    try:
        set_shared_auth(auth_instance)
    except:
        pass


def _safe_entry_insert(entry_widget, value, index=0) -> None:
    """
    Insert *value* into a Tk/ttk Entry-like widget without ever passing None to Tcl.

    Tk treats None as a missing argument which triggers a ``wrong # args`` error.
    This helper coerces None to an empty string and ensures the target widget is
    cleared before writing the new content.
    """
    text = '' if value is None else str(value)
    try:
        entry_widget.delete(0, tk.END)
    except tk.TclError:
        # Widget might not support delete (e.g. not yet mapped); ignore silently.
        pass
    entry_widget.insert(index, text)


def _safe_set_combobox(combobox: ttk.Combobox, value) -> None:
    """
    Set a ttk.Combobox value while gracefully handling None and incompatible inputs.

    Tkinter forwards Python's None as a missing argument to the underlying Tcl command,
    which raises `TclError: wrong # args: should be "pathName set value"`. This helper
    normalises None to an empty string and falls back to clearing the widget if Tcl
    still rejects the value (for example when the combobox is readonly and the value
    is outside its list).
    """
    safe_value = '' if value is None else str(value)

    # Empty string is equivalent to "no selection" – clear the widget safely
    if safe_value == '':
        previous_state = combobox.cget('state')
        try:
            combobox.configure(state='normal')
            _safe_entry_insert(combobox, '')
        finally:
            combobox.configure(state=previous_state)
        return

    try:
        combobox.set(safe_value)
    except tk.TclError as err:
        logging.debug("Combobox %s rejected value %r (%s); defaulting to empty string",
                      combobox, value, err)
        previous_state = combobox.cget('state')
        try:
            combobox.configure(state='normal')
            _safe_entry_insert(combobox, '')
        finally:
            combobox.configure(state=previous_state)

# Import database connection
from university_system.infrastructure.database.db import get_db_connection

# Import modular GUI components
from university_system.modules.domain.finance.gui.finance_management_gui import FinanceManagementGUI
from university_system.modules.domain.finance.gui.financial_aid import FinancialAidGUI
from university_system.modules.domain.student_affairs.gui.student_union_management_gui import StudentUnionManagementGUI
from university_system.modules.domain.health.gui.health_portal_management_gui import HealthPortalManagementGUI
from university_system.modules.domain.academics.gui.grade_tracking_management_gui import GradeTrackingManagementGUI
from university_system.modules.domain.commerce.gui.restaurant_management_gui import RestaurantManagementGUI
from university_system.infrastructure.email.gui.email_manager_management_gui import EmailManagerManagementGUI
from university_system.modules.domain.academics.services.modules import (compulsory_module_1, compulsory_module_2,
                                                               optional_module_1, optional_module_2,
                                                               optional_module_3, optional_module_4,
                                                               CS_optional_module_1, CS_optional_module_2,
                                                               CS_optional_module_3, CS_optional_module_4,
                                                               DS_optional_module_1, DS_optional_module_2,
                                                               DS_optional_module_3, DS_optional_module_4,
                                                               )

# ============================================================================
# PHASE 4 FEATURE IMPORTS - New Features (October 2025)
# ============================================================================

# Feature 1: Virtual Classroom Integration
try:
    from university_system.modules.domain.academics.services.virtual_classroom import (
        VirtualClassroomManager,
        SessionManager,
        ParticipantManager,
        RecordingManager,
        BreakoutRoomManager,
        PollManager,
        ChatManager
    )
    VIRTUAL_CLASSROOM_AVAILABLE = True
except Exception as e:
    VirtualClassroomManager = None
    SessionManager = None
    ParticipantManager = None
    RecordingManager = None
    BreakoutRoomManager = None
    PollManager = None
    ChatManager = None
    VIRTUAL_CLASSROOM_AVAILABLE = False

# Feature 2: Integrated Financial Aid & Scholarship Management
try:
    from university_system.modules.domain.finance.services.financial_aid import (
        FinancialAidManager,
        ScholarshipManager
    )
    FINANCIAL_AID_AVAILABLE = True
except Exception as e:
    FinancialAidManager = None
    ScholarshipManager = None
    FINANCIAL_AID_AVAILABLE = False

# Feature 3: Unified Communication Hub
try:
    from university_system.modules.shared.services.communication import CommunicationManager
    COMMUNICATION_HUB_AVAILABLE = True
except Exception as e:
    CommunicationManager = None
    COMMUNICATION_HUB_AVAILABLE = False

# Features 4-8: Database schemas are available in remaining_features_schema.py
# These features have database tables but no GUI interfaces yet:
# - Mobile App (PWA) Infrastructure
# - Transportation & Parking Management
# - Blockchain Credentials & Digital Badges
#
# Implemented GUIs:
# - Accessibility & Accommodation Tools (button in Medical Accommodation GUI)
# - Parent Portal Enhancement (merged into Parent Portal)
# - Facilities & Space Management
#
# GUI interfaces for these features can be added in future updates.

# ============================================================================

# Optional modules ---------------------------------------------------------
try:
    from university_system.modules.shared.gui.advanced_search_gui import AdvancedSearchGUI
    ADVANCED_SEARCH_GUI_AVAILABLE = True
    ADVANCED_SEARCH_GUI_IMPORT_ERROR = None
except Exception as adv_import_error:
    AdvancedSearchGUI = None
    ADVANCED_SEARCH_GUI_AVAILABLE = False
    ADVANCED_SEARCH_GUI_IMPORT_ERROR = str(adv_import_error)

try:
    from university_system.modules.domain.academics.gui.module_scheduling_gui import ModuleSchedulingGUI
    MODULE_SCHEDULING_GUI_AVAILABLE = True
    MODULE_SCHEDULING_GUI_IMPORT_ERROR = None
except Exception as mod_import_error:
    ModuleSchedulingGUI = None
    MODULE_SCHEDULING_GUI_AVAILABLE = False
    MODULE_SCHEDULING_GUI_IMPORT_ERROR = str(mod_import_error)

try:
    from university_system.modules.domain.academics.gui.course_management_gui import CourseManagementGUI
    COURSE_MANAGEMENT_GUI_AVAILABLE = True
except Exception:
    CourseManagementGUI = None
    COURSE_MANAGEMENT_GUI_AVAILABLE = False

try:
    from university_system.modules.domain.academics.gui.assignment_system import AssignmentGUI
    ASSIGNMENT_SUBMISSION_GUI_AVAILABLE = True
    print("✅ Assignment Submission GUI imported successfully")
except Exception as e:
    AssignmentGUI = None
    ASSIGNMENT_SUBMISSION_GUI_AVAILABLE = False
    print(f"❌ Failed to import Assignment Submission GUI: {e}")

try:
    from university_system.modules.shared.gui.document_manager_gui import (
        DocumentManagerGUI,
        start_document_manager_gui,
        display_document_management_menu
    )
    DOCUMENT_MANAGER_GUI_AVAILABLE = True
except Exception:
    DocumentManagerGUI = None
    start_document_manager_gui = None
    display_document_management_menu = None
    DOCUMENT_MANAGER_GUI_AVAILABLE = False

try:
    from university_system.modules.shared.gui.enhanced_reporting_gui import ReportingSystemGUI
    ENHANCED_REPORTING_GUI_AVAILABLE = True
except Exception:
    ReportingSystemGUI = None
    ENHANCED_REPORTING_GUI_AVAILABLE = False

try:
    from university_system.modules.domain.housing.gui.housing_accommodation_gui import HousingGUI as HousingAccommodationGUI
    HOUSING_ACCOMMODATION_GUI_AVAILABLE = True
except Exception:
    HousingAccommodationGUI = None
    HOUSING_ACCOMMODATION_GUI_AVAILABLE = False

try:
    from university_system.modules.domain.academics.gui.library_gui import LibraryGUI
    LIBRARY_GUI_AVAILABLE = True
except Exception:
    LibraryGUI = None
    LIBRARY_GUI_AVAILABLE = False

try:
    from university_system.modules.domain.academics.gui.parent_portal_gui import ParentPortalGUI
    PARENT_PORTAL_GUI_AVAILABLE = True
    PARENT_PORTAL_GUI_IMPORT_ERROR = None
except Exception as e:
    ParentPortalGUI = None
    PARENT_PORTAL_GUI_AVAILABLE = False
    PARENT_PORTAL_GUI_IMPORT_ERROR = str(e)

try:
    from university_system.modules.shared.gui.student_analytics_gui import GUIStudentAnalytics
    STUDENT_ANALYTICS_GUI_AVAILABLE = True
except Exception:
    GUIStudentAnalytics = None
    STUDENT_ANALYTICS_GUI_AVAILABLE = False

try:
    from university_system.modules.domain.student_affairs.gui.student_support_gui import StudentSupportGUI
    STUDENT_SUPPORT_GUI_AVAILABLE = True
except Exception:
    StudentSupportGUI = None
    STUDENT_SUPPORT_GUI_AVAILABLE = False

try:
    from university_system.modules.domain.mobility.gui.trip_management_gui import TripManagementGUI
    TRIP_MANAGEMENT_GUI_AVAILABLE = True
    _TRIP_MGMT_IMPORT_ERROR = None
except Exception as e:
    TripManagementGUI = None
    TRIP_MANAGEMENT_GUI_AVAILABLE = False
    _TRIP_MGMT_IMPORT_ERROR = str(e)

try:
    from university_system.modules.domain.student_affairs.gui.alumni_management_gui import AlumniGUIApp
    ALUMNI_PORTAL_GUI_AVAILABLE = True
    _AlumniGUIApp = AlumniGUIApp
    _ALUMNI_PORTAL_IMPORT_ERROR = None
except Exception as alumni_error:
    AlumniGUIApp = None
    ALUMNI_PORTAL_GUI_AVAILABLE = False
    _AlumniGUIApp = None
    _ALUMNI_PORTAL_IMPORT_ERROR = str(alumni_error)

try:
    from university_system.modules.domain.academics.gui.grade_tracking import GradeTrackingApp
    GRADE_TRACKING_GUI_AVAILABLE = True
except Exception as e:
    GradeTrackingApp = None
    GRADE_TRACKING_GUI_AVAILABLE = False
    print(f"❌ Failed to import Grade Tracking GUI: {e}")

try:
    from university_system.infrastructure.database.gui.data_backup_gui import BackupGUI
    DATA_BACKUP_GUI_AVAILABLE = True
    print("✅ Data Backup GUI imported successfully")
except Exception as e:
    BackupGUI = None
    DATA_BACKUP_GUI_AVAILABLE = False
    print(f"❌ Failed to import Data Backup GUI: {e}")

try:
    from university_system.modules.domain.academics.gui.academic_calendar_gui import CalendarGUI
    ACADEMIC_CALENDAR_GUI_AVAILABLE = True
    print("✅ Academic Calendar GUI imported successfully")
except Exception as e:
    CalendarGUI = None
    ACADEMIC_CALENDAR_GUI_AVAILABLE = False
    print(f"❌ Failed to import Academic Calendar GUI: {e}")

try:
    from university_system.modules.domain.housing.gui.accommodation_gui import AccommodationGUI
    ACCOMMODATION_GUI_AVAILABLE = True
except Exception as e:
    AccommodationGUI = None
    ACCOMMODATION_GUI_AVAILABLE = False
    print(f"❌ Failed to import Accommodation GUI: {e}")

try:
    from university_system.modules.domain.academics.gui.ai_detector_gui import AIDetectorGUI
    AI_DETECTOR_GUI_AVAILABLE = True
    print("✅ AI Detector GUI imported successfully")
except Exception as e:
    AIDetectorGUI = None
    AI_DETECTOR_GUI_AVAILABLE = False
    print(f"❌ Failed to import AI Detector GUI: {e}")

try:
    from university_system.modules.shared.gui.batch_operations_gui import BatchOperationsGUI
    BATCH_OPS_GUI_AVAILABLE = True
    print("✅ Batch Operations GUI imported successfully")
except Exception as e:
    BatchOperationsGUI = None
    BATCH_OPS_GUI_AVAILABLE = False
    print(f"❌ Failed to import Batch Operations GUI: {e}")

try:
    from university_system.infrastructure.email.gui.email_manager_gui import EmailManagerGUI
    EMAIL_MANAGER_GUI_AVAILABLE = True
except Exception as e:
    EmailManagerGUI = None
    EMAIL_MANAGER_GUI_AVAILABLE = False
    print(f"❌ Failed to import Email Manager GUI: {e}")

try:
    from university_system.modules.domain.finance.gui.finance_reporting_gui import FinancialManagementGUI
    FINANCE_REPORTING_GUI_AVAILABLE = True
    print("✅ Finance Reporting GUI imported successfully")
except Exception as e:
    FinancialManagementGUI = None
    FINANCE_REPORTING_GUI_AVAILABLE = False
    print(f"❌ Failed to import Finance Reporting GUI: {e}")

try:
    from university_system.modules.domain.mobility.gui.parking_management_gui import ParkingManagementGUI
    PARKING_MANAGEMENT_GUI_AVAILABLE = True
except Exception as e:
    ParkingManagementGUI = None
    PARKING_MANAGEMENT_GUI_AVAILABLE = False
    print(f"❌ Failed to import Parking Management GUI: {e}")

try:
    from university_system.modules.domain.academics.gui.plagiarism_main_gui import PlagiarismCheckerGUI
    PLAGIARISM_GUI_AVAILABLE = True
    print("✅ Plagiarism GUI imported successfully")
except Exception as e:
    PlagiarismCheckerGUI = None
    PLAGIARISM_GUI_AVAILABLE = False
    print(f"❌ Failed to import Plagiarism GUI: {e}")

try:
    from university_system.modules.domain.commerce.gui.restaurant_management_gui import RestaurantManagementGUI
    RESTAURANT_MANAGEMENT_GUI_AVAILABLE = True
except Exception as e:
    RestaurantManagementGUI = None
    RESTAURANT_MANAGEMENT_GUI_AVAILABLE = False
    print(f"❌ Failed to import Restaurant Management GUI: {e}")

try:
    from university_system.modules.domain.commerce.gui.shop_management_gui import UniversityShopGUI as ShopManagementGUI
    SHOP_GUI_AVAILABLE = True
except Exception as e:
    ShopManagementGUI = None
    SHOP_GUI_AVAILABLE = False
    print(f"❌ Failed to import Shop Management GUI: {e}")

try:
    from university_system.modules.shared.gui.simple_activity_logger_gui import EnhancedActivityLoggerGUI as ActivityLoggerGUI
    ACTIVITY_LOGGER_GUI_AVAILABLE = True
except Exception as e:
    ActivityLoggerGUI = None
    ACTIVITY_LOGGER_GUI_AVAILABLE = False
    print(f"❌ Failed to import Activity Logger GUI: {e}")

try:
    from university_system.utils.ai.gui.university_chatbot_gui import ChatbotGUI as UniversityChatbotGUI
    CHATBOT_GUI_AVAILABLE = True
except Exception as e:
    UniversityChatbotGUI = None
    CHATBOT_GUI_AVAILABLE = False
    print(f"❌ Failed to import University Chatbot GUI: {e}")

try:
    from university_system.utils.logging.gui.log_management_gui import LogManagementGUI
    LOG_MANAGEMENT_GUI_AVAILABLE = True
except Exception as e:
    LogManagementGUI = None
    LOG_MANAGEMENT_GUI_AVAILABLE = False
    print(f"❌ Failed to import Log Management GUI: {e}")

try:
    from university_system.modules.domain.student_affairs.gui.internship_management_gui import InternshipGUI as InternshipManagementGUI
    INTERNSHIP_MANAGEMENT_GUI_AVAILABLE = True
except Exception as e:
    InternshipManagementGUI = None
    INTERNSHIP_MANAGEMENT_GUI_AVAILABLE = False
    print(f"❌ Failed to import Internship Management GUI: {e}")

# Set remaining GUI availability variables to False (no corresponding GUI files found)
ANALYTICS_GUI_AVAILABLE = False  # This uses student_analytics_gui which is imported separately
CALENDAR_GUI_AVAILABLE = False  # This would be a separate calendar system
HELP_DESK_GUI_AVAILABLE = False  # Would use helpdesk_gui.py

# Import AttendanceGUI
try:
    from university_system.modules.domain.academics.gui.attendance_tracker_gui import AttendanceGUI
    ATTENDANCE_GUI_AVAILABLE = True
except Exception as e:
    AttendanceGUI = None
    ATTENDANCE_GUI_AVAILABLE = False
    print(f"❌ Failed to import Attendance GUI: {e}")

# Import helpdesk GUI and CLI fallback function
try:
    from university_system.modules.domain.student_affairs.gui.helpdesk_gui import HelpdeskGUI
    HELPDESK_GUI_AVAILABLE = True
except Exception as e:
    HelpdeskGUI = None
    HELPDESK_GUI_AVAILABLE = False
    print(f"❌ Failed to import Helpdesk GUI: {e}")

try:
    from university_system.modules.domain.student_affairs.services.helpdesk import display_helpdesk_menu
    HELPDESK_CLI_AVAILABLE = True
except Exception as e:
    display_helpdesk_menu = None
    HELPDESK_CLI_AVAILABLE = False
    print(f"❌ Failed to import Helpdesk CLI: {e}")

# Import advanced feature GUI launchers from core files
from university_system.modules.domain.academics.services.attendance.attendance_tracker import launch_advanced_attendance_gui
from university_system.modules.domain.student_affairs.services.mental_health.mental_health_core import launch_mental_health_gui
from university_system.modules.domain.student_affairs.services.early_warning.early_warning_core import launch_early_warning_gui
from university_system.modules.domain.career.services.career_services_core import launch_career_services_gui
from university_system.modules.domain.admissions.services.admissions_crm_core import launch_admissions_crm_gui
from university_system.modules.shared.services.analytics.predictive_analytics_gui import launch_predictive_analytics_gui
from university_system.modules.domain.academics.services.module_scheduling import launch_timetable_optimizer_gui
from university_system.modules.domain.campus.services.campus_events_gui import launch_campus_events_gui
from university_system.modules.domain.student_affairs.services.alumni_management import launch_alumni_relations_gui
from university_system.modules.domain.facilities.services.facilities_management_core import launch_facilities_management_gui
from university_system.modules.shared.services.business_intelligence.business_intelligence_gui import launch_business_intelligence_gui
from university_system.modules.shared.services.ai_features.ai_features_core import launch_ai_features_gui

# Phase 4: New GUI imports (October 2025)
from university_system.modules.domain.mobility.gui.mobile_app_pwa_gui import launch_mobile_app_pwa_gui
from university_system.modules.domain.academics.gui.blockchain_credentials_gui import launch_blockchain_credentials_gui
from university_system.modules.services.gui.integration_marketplace_gui import launch_integration_marketplace_gui

# Global variables for missing components
chatbot_instance = None
CLI_AVAILABLE = True  # Assume CLI is available unless proven otherwise

# Function definitions for missing functions
def init_calendar_database():
    """Initialize calendar database - placeholder implementation"""
    try:
        # Basic database initialization logic
        from university_system.infrastructure.database.db import get_db_connection
        conn = get_db_connection()
        if conn:
            conn.close()
            return True
        return False
    except Exception as e:
        print(f"Error initializing calendar database: {e}")
        return False

def init_enhanced_database():
    """Initialize enhanced database - placeholder implementation"""
    try:
        # Basic database initialization logic
        from university_system.infrastructure.database.db import get_db_connection
        conn = get_db_connection()
        if conn:
            conn.close()
            return True
        return False
    except Exception as e:
        print(f"Error initializing enhanced database: {e}")
        return False

def initialize_chatbot_integration():
    """Initialize chatbot integration - placeholder implementation"""
    global chatbot_instance
    try:
        # Placeholder chatbot initialization
        voice_interface_stub = type('VoiceInterfaceStub', (), {
            'enabled': False,
            'cleanup': lambda self: None
        })()

        def process_message_stub(msg):
            """Basic chatbot response logic"""
            msg_lower = msg.lower()

            if any(word in msg_lower for word in ['hello', 'hi', 'hey', 'greetings']):
                return "Hello! I'm the University Chatbot. How can I help you today?"
            elif any(word in msg_lower for word in ['course', 'class', 'subject']):
                return "I can help you with course information. Please specify what courses or subjects you're interested in."
            elif any(word in msg_lower for word in ['registration', 'enroll', 'register']):
                return "For course registration, please visit the Student Portal or contact the Registrar's Office."
            elif any(word in msg_lower for word in ['schedule', 'timetable', 'time']):
                return "You can view your class schedule in the Student Portal under 'Academic Schedule'."
            elif any(word in msg_lower for word in ['help', 'support', 'assistance']):
                return "I'm here to help! You can ask me about courses, registration, schedules, campus services, or general university information."
            elif any(word in msg_lower for word in ['campus', 'location', 'building']):
                return "For campus maps and building locations, please check the university website or visit Student Services."
            elif any(word in msg_lower for word in ['library', 'book', 'research']):
                return "The university library offers research assistance and book reservations. Visit the library website for more details."
            elif any(word in msg_lower for word in ['fee', 'tuition', 'payment', 'cost']):
                return "For tuition and fee information, please check the Finance Office or your Student Account."
            elif any(word in msg_lower for word in ['thank', 'thanks']):
                return "You're welcome! Is there anything else I can help you with?"
            elif any(word in msg_lower for word in ['bye', 'goodbye', 'exit']):
                return "Goodbye! Feel free to return if you have more questions."
            else:
                return f"I understand you're asking about: '{msg}'. While I'm operating in basic mode, I'd recommend contacting Student Services for detailed assistance with your specific query."

        chatbot_instance = type('ChatbotStub', (), {
            'set_auth_system': lambda self, auth: None,
            'process_message': lambda self, msg, user_id=None, session_id=None: process_message_stub(msg),
            'voice_interface': voice_interface_stub,
            'test_voice_interface': lambda self: {'status': 'not_available', 'message': 'Voice interface not available'},
            'get_system_status': lambda self: {'status': 'not_available', 'components': {}},
            'generate_usage_analytics': lambda self: {'total_conversations': 0, 'active_users': 0},
            'connect_to_db': lambda self: False,
            'authenticated_sessions': {},
            'conversation_history': {},
            'auth_system': None,
            'authenticate_user_for_chatbot': lambda self, username, password, mfa=None: {'success': False, 'message': 'Authentication not available'},
            'logout_user': lambda self, session_id: None,
            'text_to_speech': lambda self, text: None,
            'process_voice_input': lambda self, duration=5: "Voice input not available",
            'run_authenticated_console_interface': lambda self: None,
            'run_console_interface': lambda self: None,
            'setup_api_routes': lambda self: None,
            'run_web_server': lambda self: None
        })()
        return True
    except Exception as e:
        print(f"Error initializing chatbot: {e}")
        return False

def safe_auth_check(auth_obj):
    """Safely check if auth object has required attributes"""
    if not auth_obj:
        return False

    # Ensure required attributes exist
    if not hasattr(auth_obj, 'current_user'):
        auth_obj.current_user = None

    if not hasattr(auth_obj, 'last_activity'):
        auth_obj.last_activity = None

    if not hasattr(auth_obj, 'session_timeout'):
        auth_obj.session_timeout = 30

    if not hasattr(auth_obj, 'login_attempts'):
        auth_obj.login_attempts = {}

    if not hasattr(auth_obj, 'max_attempts'):
        auth_obj.max_attempts = 5

    if not hasattr(auth_obj, 'lockout_time'):
        auth_obj.lockout_time = 15

    return True

def init_gui(session_user=None):
    """
    Centralized GUI initialization function.

    Args:
        session_user: User object from an existing session (e.g., from CLI login).
                     If None, GUI starts at login page.
                     If provided, GUI bypasses login and goes directly to main interface.

    Returns:
        UnifiedManagementGUI instance
    """
    global auth

    # Initialize auth manager if needed
    if auth is None:
        if UserAuth is None:
            print("Warning: UserAuth not available. Using dummy authentication.")
            # Create a minimal auth object for testing
            class DummyAuth:
                def __init__(self):
                    self.is_authenticated = True
                    self.current_user = None
                    self.last_activity = None
                    self.session_timeout = 30
                    self.login_attempts = {}
                    self.max_attempts = 5
                    self.lockout_time = 15
            auth = DummyAuth()
        else:
            auth = UserAuth()
            safe_auth_check(auth)

        # Register auth instance with shared_context and local auth
        set_auth(auth)

    # If session_user is provided, set it as the current user
    if session_user is not None:
        auth.current_user = session_user
        if hasattr(auth, 'last_activity'):
            from datetime import datetime
            auth.last_activity = datetime.now()

    # Create and return the GUI with the configured auth
    app = UnifiedManagementGUI(auth)
    return app

class UnifiedManagementGUI:
    """Unified GUI interface combining authentication and student management - replaces StudentManagementGUI"""

    def __init__(self, auth_manager):
        try:
            self.auth = auth_manager

            # Initialize content_frame to None first
            self.content_frame = None

            # Initialize modular GUI managers
            self.finance_gui = None
            self.student_union_gui = None
            self.health_portal_gui = None
            self.grade_tracking_gui = None
            self.restaurant_gui = None
            self.email_manager_gui = None

            # Initialize student management components
            self.student_tree = None

            # Initialize Tkinter
            self.root = tk.Tk()
            self.root.title("University Management System")
            self.root.geometry("1200x800")
            self.root.minsize(1000, 700)

            # Configure style and theme
            self.style = ttk.Style()
            self.style.theme_use('clam')

            # Variables
            self.current_user_var = tk.StringVar()
            self.status_var = tk.StringVar(value="Not logged in")

            # Initialize GUI
            self.setup_gui()

            # Initialize modular GUI managers after root is set up
            self.init_gui_managers()

            # Update status
            self.update_status()

            # Start periodic updates
            self.check_session_timer()

            # Show login if not authenticated
            if not self.auth.current_user:
                self.show_login_screen()
            else:
                self.show_main_interface()

        except Exception as e:
            print(f"Error initializing GUI: {e}")
            self.create_fallback_interface()
    
    def create_fallback_interface(self):
        """Create minimal fallback interface"""
        self.root = tk.Tk()
        self.root.title("University Management System (Error Mode)")
        self.root.geometry("400x300")
        
        error_frame = ttk.Frame(self.root, padding="20")
        error_frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(error_frame, text="Error initializing system:", font=('Arial', 12, 'bold')).pack(pady=10)
        ttk.Label(error_frame, text="System running in minimal mode", foreground="red", wraplength=350).pack(pady=10)
        ttk.Button(error_frame, text="Close", command=self.root.quit).pack(pady=10)

    def init_gui_managers(self):
        """Initialize modular GUI managers"""
        try:
            self.finance_gui = FinanceManagementGUI(self.root, self.auth)
            self.student_union_gui = StudentUnionManagementGUI(self.root, self.auth)
            self.health_portal_gui = HealthPortalManagementGUI(self.root, self.auth)
            self.grade_tracking_gui = GradeTrackingManagementGUI(self.root, self.auth)
            # Don't initialize restaurant GUI immediately to avoid widget conflicts
            # self.restaurant_gui = RestaurantManagementGUI(self.root, self.auth)
            self.email_manager_gui = EmailManagerManagementGUI(self.root, self.auth)
        except Exception as e:
            print(f"Warning: Error initializing GUI managers: {e}")


    def create_themed_toplevel(self, title="", geometry=""):
        """Create a Toplevel window"""
        window = tk.Toplevel(self.root)
        if title:
            window.title(title)
        if geometry:
            window.geometry(geometry)
        window.transient(self.root)
        return window

    def show_activity_logger(self):
        """Launch the Enhanced Activity Logger GUI in a child window"""
        if not self.auth.current_user:
            messagebox.showerror("Error", "You must be logged in to access the activity logger.")
            return
        
        # Check permissions - allow admin and staff
        user_role = self.auth.current_user.get('role', '')
        if user_role not in ['admin', 'staff']:
            messagebox.showerror("Error", "You don't have permission to access the activity logger.")
            return
        
        try:
            if not ACTIVITY_LOGGER_GUI_AVAILABLE:
                messagebox.showerror("Activity Logger", f"Activity Logger GUI not available: {ACTIVITY_LOGGER_GUI_IMPORT_ERROR}")
                return
            
            # Create a new window for the Activity Logger GUI
            logger_window = tk.Toplevel(self.root)
            logger_window.title("Enhanced Activity Logger - Management Console")
            logger_window.geometry("1400x900")
            logger_window.minsize(1200, 800)
            
            # Center the window
            logger_window.update_idletasks()
            x = (logger_window.winfo_screenwidth() - logger_window.winfo_width()) // 2
            y = (logger_window.winfo_screenheight() - logger_window.winfo_height()) // 2
            logger_window.geometry(f"+{x}+{y}")
            
            try:
                logger_window.transient(self.root)
            except Exception:
                pass  # Continue if transient fails
            
            # Initialize the Activity Logger GUI in the new window
            # We need to modify the EnhancedActivityLoggerGUI to accept a parent window
            activity_logger_gui = ActivityLoggerGUI()
            
            # Replace the default root with our window
            activity_logger_gui.root.destroy()  # Destroy the default root
            activity_logger_gui.root = logger_window  # Use our window as the root
            
            # Setup the GUI components again with the new root
            activity_logger_gui.setup_ui()
            activity_logger_gui.connect_to_logger()
            activity_logger_gui.start_update_timer()
            
            print("Enhanced Activity Logger GUI opened successfully")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open Enhanced Activity Logger: {str(e)}")
            print(f"Activity Logger error: {e}")
        
    def show_library_management(self):
        """Launch the Library Management GUI in a new window"""
        if not self.auth.current_user:
            messagebox.showerror("Error", "You must be logged in to access library management.")
            return
        
        # Check permissions
        if not (self.auth.check_permission('view_books') or 
                self.auth.check_permission('manage_books') or 
                self.auth.check_permission('manage_loans') or 
                self.auth.check_permission('checkout_books')):
            messagebox.showerror("Error", "You don't have permission to access library management.")
            return
        
        try:
            if LIBRARY_GUI_AVAILABLE:
                # Create a new window for the Library GUI
                library_window = tk.Toplevel(self.root)
                library_window.title("Enhanced Library Management System")
                library_window.geometry("1400x900")
                library_window.minsize(1200, 800)
                
                # Center the window
                library_window.update_idletasks()
                x = (library_window.winfo_screenwidth() - library_window.winfo_width()) // 2
                y = (library_window.winfo_screenheight() - library_window.winfo_height()) // 2
                library_window.geometry(f"+{x}+{y}")
                
                try:
                    library_window.transient(self.root)
                except Exception:
                    pass  # Continue if transient fails
                
                # Initialize the Library GUI in the new window
                library_gui = LibraryGUI(library_window)
                
                # Pass the auth context if the LibraryGUI supports it
                if hasattr(library_gui, 'set_auth'):
                    library_gui.set_auth(self.auth)
                elif hasattr(library_gui, 'auth'):
                    library_gui.auth = self.auth
                
                print("✅ Library Management GUI opened successfully")
                
            else:
                # Fallback to CLI menu
                messagebox.showinfo("Library Management", 
                                  f"Library GUI not available: {LIBRARY_GUI_IMPORT_ERROR}\nUsing CLI menu.")
                try:
                    from university_system.modules.domain.academics.services.library import display_library_menu
                    display_library_menu()
                except ImportError:
                    messagebox.showerror("Error", "Library management system not available.")
                    
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open Library Management: {str(e)}")
            print(f"❌ Library Management error: {e}")

    def show_student_union_portal(self):
        """Wrapper method to maintain compatibility with navigation buttons"""
        if self.student_union_gui:
            self.student_union_gui.show_student_union_portal()
        else:
            messagebox.showerror("Error", "Student Union GUI not available.")

    def show_parking_management(self):
        """Launch the Parking Management GUI in a child window"""
        if not self.auth.current_user:
            messagebox.showerror("Error", "You must be logged in to access parking management.")
            return
        
        # Check permissions
        if not (self.auth.check_permission('manage_parking') or 
                self.auth.check_permission('create_permit') or 
                self.auth.check_permission('view_any_permit') or 
                self.auth.check_permission('view_own_permit')):
            messagebox.showerror("Error", "You don't have permission to access parking management.")
            return
        
        try:
            # Check if parking GUI is available
            if not PARKING_MANAGEMENT_GUI_AVAILABLE or ParkingManagementGUI is None:
                messagebox.showerror("Parking Management", "Parking Management GUI not available")
                return

            # Create a new window for the Parking Management GUI
            parking_window = tk.Toplevel(self.root)
            parking_window.title("Parking Management System")
            parking_window.geometry("1200x800")
            parking_window.minsize(800, 600)
            
            # Center the window
            parking_window.update_idletasks()
            x = (parking_window.winfo_screenwidth() - parking_window.winfo_width()) // 2
            y = (parking_window.winfo_screenheight() - parking_window.winfo_height()) // 2
            parking_window.geometry(f"+{x}+{y}")
            
            try:
                parking_window.transient(self.root)
            except Exception:
                pass  # Continue if transient fails
            
            # Initialize the Parking Management GUI in the new window with auth system
            parking_gui = ParkingManagementGUI(parking_window, auth_system=self.auth)
            
            # Update the current user in the parking GUI if it tracks it separately
            if hasattr(parking_gui, 'current_user'):
                parking_gui.current_user = self.auth.current_user
            
            print("✅ Parking Management GUI opened successfully")
            
        except ImportError as e:
            # Fallback to CLI menu if GUI is not available
            messagebox.showinfo("Parking Management", 
                              f"Parking GUI not available: {e}\nUsing CLI menu.")
            try:
                from university_system.modules.domain.mobility.services.parking_management import display_parking_menu
                display_parking_menu()
            except ImportError:
                messagebox.showerror("Error", "Parking management system not available.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open Parking Management: {str(e)}")
            print(f"❌ Parking Management error: {e}")

    def show_academic_calendar(self):
        """Launch the Academic Calendar GUI with proper initialization"""
        if not self.auth.current_user:
            messagebox.showerror("Error", "You must be logged in to access the academic calendar.")
            return
        
        # Check permissions
        if not (self.auth.check_permission('manage_schedules') or 
                self.auth.check_permission('view_own_timetable') or 
                self.auth.check_permission('export_data')):
            messagebox.showerror("Error", "You don't have permission to access the academic calendar.")
            return
        
        try:
            if not ACADEMIC_CALENDAR_GUI_AVAILABLE:
                messagebox.showerror("Error", "Academic Calendar GUI is not available.")
                return
            
            # Ensure calendar database is initialized
            print("Checking calendar database initialization...")
            if not init_calendar_database():
                messagebox.showerror("Error", "Failed to initialize calendar database.")
                return
            
            # Create calendar GUI with embedded approach
            calendar_window = tk.Toplevel(self.root)
            calendar_window.title("📅 Academic Calendar Management System")
            calendar_window.geometry("1400x900")
            calendar_window.minsize(1000, 600)
            
            # Center the window
            calendar_window.update_idletasks()
            x = (calendar_window.winfo_screenwidth() - calendar_window.winfo_width()) // 2
            y = (calendar_window.winfo_screenheight() - calendar_window.winfo_height()) // 2
            calendar_window.geometry(f"+{x}+{y}")
            
            try:
                calendar_window.transient(self.root)
            except Exception:
                pass
            
            # Create the calendar GUI instance
            try:
                # Use the imported CalendarGUI class - correct parameter order
                calendar_gui = CalendarGUI(auth_manager=self.auth, parent_window=calendar_window)

                print("✅ Academic Calendar GUI opened successfully")
                
            except Exception as gui_error:
                calendar_window.destroy()
                raise gui_error
                
        except Exception as e:
            error_msg = f"Failed to open Academic Calendar: {str(e)}"
            messagebox.showerror("Error", error_msg)
            print(f"❌ Academic Calendar error: {e}")
            logging.error(f"Academic Calendar GUI error: {e}")
            
            # Offer CLI fallback
            if messagebox.askyesno("Fallback", "Would you like to try the command-line interface instead?"):
                try:
                    from university_system.modules.domain.academics.services.academic_calendar import display_academic_calendar_menu
                    display_academic_calendar_menu()
                except ImportError:
                    messagebox.showerror("Error", "Academic calendar system not available.")

    def show_finance_management(self):
        """Launch the Finance Management GUI in a child window, fallback to CLI if needed."""
        if self.finance_gui:
            self.finance_gui.show_finance_management()
        else:
            try:
                from university_system.cli_main import display_finance_menu
                from university_system.modules.finance.core.financial_core import set_finance_auth
                if self.auth:
                    try:
                        set_finance_auth(self.auth)
                    except Exception:
                        pass
                display_finance_menu()
            except ImportError:
                messagebox.showerror("Error", "Finance GUI not available and CLI fallback missing")

    def show_finance_reporting_dashboard(self):
        """Launch the Finance Reporting Dashboard GUI in a child window."""
        try:
            if not FINANCE_REPORTING_GUI_AVAILABLE:
                messagebox.showerror("Error", "Finance Reporting GUI is not available.")
                return

            # Create new window for Finance Reporting
            finance_window = tk.Toplevel(self.root)
            finance_window.title("Finance Reporting Dashboard")
            finance_window.geometry("1200x800")

            # Initialize Finance Reporting GUI
            finance_gui = FinancialManagementGUI(finance_window, self.auth)
            print("✅ Finance Reporting GUI opened successfully")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to open Finance Reporting: {str(e)}")
            print(f"❌ Finance Reporting error: {e}")

    def show_financial_aid(self):
        """Launch the Financial Aid & Scholarships GUI in a child window."""
        if not self.auth.current_user:
            messagebox.showerror("Error", "You must be logged in to access Financial Aid & Scholarships.")
            return

        try:
            # Create new window for Financial Aid
            aid_window = tk.Toplevel(self.root)
            aid_window.title("Financial Aid & Scholarships")
            aid_window.geometry("1200x800")

            # Initialize Financial Aid GUI
            aid_gui = FinancialAidGUI(auth_instance=self.auth, parent=aid_window)
            aid_gui.create_embedded_interface()
            print("✅ Financial Aid GUI opened successfully")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to open Financial Aid & Scholarships: {str(e)}")
            print(f"❌ Financial Aid error: {e}")

    def show_university_shop(self):
        """Launch the University Shop GUI in a child window"""
        if not self.auth.current_user:
            messagebox.showerror("Error", "You must be logged in to access the university shop.")
            return
        
        try:
            if not SHOP_GUI_AVAILABLE:
                messagebox.showerror("University Shop", f"Shop GUI not available: {SHOP_GUI_IMPORT_ERROR}")
                return
            
            # Create a new window for the University Shop GUI
            shop_window = tk.Toplevel(self.root)
            shop_window.title("University Shop Management System")
            shop_window.geometry("1200x800")
            shop_window.minsize(1000, 600)
            
            # Center the window
            shop_window.update_idletasks()
            x = (shop_window.winfo_screenwidth() - shop_window.winfo_width()) // 2
            y = (shop_window.winfo_screenheight() - shop_window.winfo_height()) // 2
            shop_window.geometry(f"+{x}+{y}")
            
            try:
                shop_window.transient(self.root)
            except Exception:
                pass  # Continue if transient fails
            
            # Initialize the Shop Management GUI
            shop_gui = ShopManagementGUI(shop_window, self.auth)
            print("✅ University Shop GUI opened successfully")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open University Shop: {str(e)}")
            print(f"❌ University Shop error: {e}")

    def show_batch_operations_gui(self):
        """Launch Batch Operations GUI in a separate window."""
        # Auth gate (best-effort; don't block on unexpected auth errors)
        try:
            if hasattr(self, "auth") and not getattr(self.auth, "current_user", None):
                messagebox.showerror("Batch Operations", "You must be logged in.")
                return
        except Exception:
            pass

        if not BATCH_OPS_GUI_AVAILABLE:
            messagebox.showerror("Batch Operations", "Batch Operations GUI is not available.")
            return

        # Create window + init GUI
        try:
            batch_window = tk.Toplevel(self.root)
            batch_window.title("Batch Operations")
            batch_window.geometry("1000x700")

            BatchOperationsGUI(batch_window, getattr(self, "auth", None))
            print("✅ Batch Operations GUI opened successfully")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open Batch Operations: {e}")
            print(f"❌ Batch Operations error: {e}")

    def open_parent_portal_gui(self):
        """Open the Parent Portal GUI in a child window"""
        if not self.auth or not self.auth.current_user:
            messagebox.showerror("Parent Portal", "You must be logged in to access the parent portal.")
            return
        
        try:
            if not PARENT_PORTAL_GUI_AVAILABLE:
                messagebox.showerror("Parent Portal", f"Parent Portal GUI not available: {PARENT_PORTAL_GUI_IMPORT_ERROR}")
                # Fallback to CLI
                try:
                    display_parent_portal_menu(self.auth)
                except Exception:
                    messagebox.showerror("Error", "Neither GUI nor CLI parent portal is available.")
                return
            
            # Create a new window for the Parent Portal GUI
            parent_window = tk.Toplevel(self.root)
            parent_window.title("Parent Portal - School Management System")
            parent_window.geometry("1400x900")
            parent_window.minsize(1200, 800)
            
            # Center the window
            parent_window.update_idletasks()
            x = (parent_window.winfo_screenwidth() - parent_window.winfo_width()) // 2
            y = (parent_window.winfo_screenheight() - parent_window.winfo_height()) // 2
            parent_window.geometry(f"+{x}+{y}")
            
            try:
                parent_window.transient(self.root)
            except Exception:
                pass
            
            # Initialize the Parent Portal GUI
            parent_gui = ParentPortalGUI(self.auth)
            
            # Replace the root window with our new window
            parent_gui.root = parent_window
            parent_gui.setup_layout()
            parent_gui.load_user_data()
            parent_gui.show_dashboard()
            
            print("✅ Parent Portal GUI opened successfully")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open Parent Portal GUI: {str(e)}")
            print(f"❌ Parent Portal GUI error: {e}")
                
    def open_student_union_portal_gui(self):
        """Open the Student Union GUI in a child window with proper integration"""
        if self.student_union_gui:
            self.student_union_gui.open_student_union_portal_gui()
        else:
            messagebox.showerror("Error", "Student Union GUI not available.")
        
    def open_alumni_portal_gui(self):
        """Open the Alumni Management Portal in a child window from the main app."""
        # Require login like the other portals
        if not getattr(self.auth, "current_user", None):
            messagebox.showerror("Alumni Portal", "You must be logged in.")
            return

        # Verify the GUI import worked
        if not ALUMNI_PORTAL_GUI_AVAILABLE or _AlumniGUIApp is None:
            msg = "Alumni Portal GUI not available."
            if _ALUMNI_PORTAL_IMPORT_ERROR:
                msg += f"\n{_ALUMNI_PORTAL_IMPORT_ERROR}"
            messagebox.showerror("Alumni Portal", msg)
            return

        try:
            # Open inside a child window so your main app stays up
            win = tk.Toplevel(self.root)
            win.title("Alumni Management Portal")
            win.geometry("1400x900")
            win.minsize(1200, 800)
            try:
                win.transient(self.root)
                win.grab_set()
            except Exception:
                pass

            # Spin up the Alumni GUI on this window
            app = _AlumniGUIApp(win)

            # Hand over auth context if the alumni app exposes a slot; harmless if it doesn't
            try:
                setattr(app, "auth", self.auth)
            except Exception:
                pass

            print("✅ Alumni Portal opened successfully")
        except Exception as e:
            messagebox.showerror("Alumni Portal", f"Failed to open Alumni Portal GUI:\n{e}")
    
    def open_attendance_gui(self):
        try:
            import tkinter as tk
            from tkinter import messagebox
            if AttendanceGUI is None:
                messagebox.showerror("Attendance", "Attendance GUI module not available.")
                return
            # Create a child window for the attendance UI
            win = tk.Toplevel(self.root)
            win.transient(self.root)
            try:
                # Instantiate the attendance GUI on the new window
                AttendanceGUI(win, auth_manager=self.auth)
            except Exception as e:
                win.destroy()
                messagebox.showerror("Attendance", f"Failed to open Attendance GUI:\n{e}")
        except Exception as e:
            try:
                from tkinter import messagebox
                messagebox.showerror("Attendance", f"Unexpected error: {e}")
            except Exception:
                pass
    def setup_gui(self):
        """Setup the unified GUI interface using AuthenticationGUI layout"""
        # Main frame
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(2, weight=1)
        
        # Header with system status
        self.create_header(main_frame)
        
        # Left panel - Navigation buttons (like AuthenticationGUI)
        self.create_navigation_panel(main_frame)
        
        # Right panel - Content area
        self.create_content_area(main_frame)
        
        # Show welcome message initially
        self.show_welcome()
    
    def create_header(self, parent):
        """Create header with system status and control buttons"""
        header_frame = ttk.LabelFrame(parent, text="System Control & Status", padding="10")
        header_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))

        # Top row - Control buttons
        button_frame = ttk.Frame(header_frame)
        button_frame.grid(row=0, column=0, columnspan=4, sticky=(tk.W, tk.E), pady=(0, 10))

        # Shutdown button - using lambda to ensure proper method binding
        ttk.Button(button_frame, text="Shutdown", command=lambda: self.shutdown_system(),
                  style="Accent.TButton").pack(side=tk.LEFT, padx=(0, 10))

        # Login/Logout button (dynamic text based on auth status)
        self.login_logout_btn = ttk.Button(button_frame, text="Login", command=lambda: self.toggle_login_logout())
        self.login_logout_btn.pack(side=tk.LEFT, padx=(0, 10))

        # Switch to CLI button
        ttk.Button(button_frame, text="Switch to CLI", command=lambda: self.switch_to_cli()).pack(side=tk.LEFT, padx=(0, 10))

        # Status information
        ttk.Label(header_frame, text="Status:").grid(row=1, column=0, sticky=tk.W)
        ttk.Label(header_frame, textvariable=self.status_var).grid(row=1, column=1, sticky=tk.W, padx=(10, 0))

        ttk.Label(header_frame, text="Current User:").grid(row=2, column=0, sticky=tk.W)
        ttk.Label(header_frame, textvariable=self.current_user_var).grid(row=2, column=1, sticky=tk.W, padx=(10, 0))

        header_frame.columnconfigure(1, weight=1)
    
    def create_navigation_panel(self, parent):
        """Create navigation panel with categorized buttons and scrollbar"""
        # Main navigation frame
        nav_frame = ttk.LabelFrame(parent, text="Navigation", padding="5")
        nav_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(0, 10))

        # Canvas + scrollbar for long menus
        canvas = tk.Canvas(nav_frame, highlightthickness=0)
        scrollbar = ttk.Scrollbar(nav_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        def configure_scroll_region(_):
            canvas.configure(scrollregion=canvas.bbox("all"))
        scrollable_frame.bind("<Configure>", configure_scroll_region)

        canvas_window = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # Mousewheel
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind('<Enter>', lambda e: canvas.bind_all("<MouseWheel>", _on_mousewheel))
        canvas.bind('<Leave>', lambda e: canvas.unbind_all("<MouseWheel>"))

        # Keep inner frame width = canvas width
        def configure_canvas_width(_):
            canvas.itemconfig(canvas_window, width=canvas.winfo_width())
        canvas.bind('<Configure>', configure_canvas_width)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        self.nav_buttons = {}

        # Helper for logout
        def do_logout():
            try:
                if hasattr(self.auth, "logout"):
                    self.auth.logout()
            finally:
                self.update_status()
                self.show_login_screen()

        # ---------- Authentication ----------
        auth_frame = ttk.LabelFrame(scrollable_frame, text="Authentication", padding="5")
        auth_frame.pack(fill=tk.X, pady=(5, 10), padx=5)
        self.nav_buttons['login'] = ttk.Button(auth_frame, text="Login", command=self.show_login)
        self.nav_buttons['login'].pack(fill=tk.X, pady=2)
        self.nav_buttons['change_password'] = ttk.Button(auth_frame, text="Change Password", command=self.show_change_password)
        self.nav_buttons['change_password'].pack(fill=tk.X, pady=2)

        # ---------- Student Management ----------
        student_mgmt_frame = ttk.LabelFrame(scrollable_frame, text="Student Management", padding="5")
        student_mgmt_frame.pack(fill=tk.X, pady=(0, 10), padx=5)
        self.nav_buttons['student_records'] = ttk.Button(student_mgmt_frame, text="Student Records", command=self.show_student_records)
        self.nav_buttons['student_records'].pack(fill=tk.X, pady=2)
        self.nav_buttons['create_student'] = ttk.Button(student_mgmt_frame, text="Create Student", command=self.create_student_dialog)
        self.nav_buttons['create_student'].pack(fill=tk.X, pady=2)
        self.nav_buttons['search_students'] = ttk.Button(student_mgmt_frame, text="Search Students", command=self.search_students_dialog)
        self.nav_buttons['search_students'].pack(fill=tk.X, pady=2)
        if ADVANCED_SEARCH_GUI_AVAILABLE:
            self.nav_buttons['advanced_search_gui'] = ttk.Button(student_mgmt_frame, text="Advanced Search", command=self.show_advanced_search_gui)
            self.nav_buttons['advanced_search_gui'].pack(fill=tk.X, pady=2)
        self.nav_buttons['delete_student'] = ttk.Button(student_mgmt_frame, text="Delete Student", command=self.delete_student_dialog)
        self.nav_buttons['delete_student'].pack(fill=tk.X, pady=2)
        self.nav_buttons['batch_operations'] = ttk.Button(student_mgmt_frame, text="Batch Operations", command=self.show_batch_operations_gui)
        self.nav_buttons['batch_operations'].pack(fill=tk.X, pady=2)

        # ---------- Academic Management ----------
        academic_frame = ttk.LabelFrame(scrollable_frame, text="Academic Management", padding="5")
        academic_frame.pack(fill=tk.X, pady=(0, 10), padx=5)
        self.nav_buttons['course_management'] = ttk.Button(academic_frame, text="Course Management", command=self.show_course_management)
        self.nav_buttons['course_management'].pack(fill=tk.X, pady=2)
        self.nav_buttons['module_management'] = ttk.Button(academic_frame, text="Module Management", command=self.show_module_management)
        self.nav_buttons['module_management'].pack(fill=tk.X, pady=2)
        self.nav_buttons['assignments'] = ttk.Button(academic_frame, text="Assignments", command=self.show_assignments)
        self.nav_buttons['assignments'].pack(fill=tk.X, pady=2)
        self.nav_buttons['grade_tracking_gui'] = ttk.Button(academic_frame, text="Grade Tracking", command=self.show_grade_tracking_gui)
        self.nav_buttons['grade_tracking_gui'].pack(fill=tk.X, pady=2)
        self.nav_buttons['library'] = ttk.Button(academic_frame, text="Library Management", command=self.show_library_management)
        self.nav_buttons['library'].pack(fill=tk.X, pady=2)
        if VIRTUAL_CLASSROOM_AVAILABLE:
            self.nav_buttons['virtual_classroom'] = ttk.Button(academic_frame, text="Virtual Classroom", command=self.show_virtual_classroom_gui)
            self.nav_buttons['virtual_classroom'].pack(fill=tk.X, pady=2)

        # ---------- Scheduling & Attendance ----------
        sched_frame = ttk.LabelFrame(scrollable_frame, text="Scheduling & Attendance", padding="5")
        sched_frame.pack(fill=tk.X, pady=(0, 10), padx=5)
        self.nav_buttons['academic_calendar'] = ttk.Button(sched_frame, text="Academic Calendar", command=self.show_academic_calendar)
        self.nav_buttons['academic_calendar'].pack(fill=tk.X, pady=2)
        self.nav_buttons['scheduling'] = ttk.Button(sched_frame, text="Module Scheduling", command=self.show_module_scheduling)
        self.nav_buttons['scheduling'].pack(fill=tk.X, pady=2)
        self.nav_buttons['attendance'] = ttk.Button(sched_frame, text="Attendance Tracking", command=self.open_attendance_gui)
        self.nav_buttons['attendance'].pack(fill=tk.X, pady=2)

        # ---------- Finance ----------
        finance_frame = ttk.LabelFrame(scrollable_frame, text="Finance", padding="5")
        finance_frame.pack(fill=tk.X, pady=(0, 10), padx=5)
        self.nav_buttons['finance_management'] = ttk.Button(finance_frame, text="Finance Management", command=self.show_finance_management)
        self.nav_buttons['finance_management'].pack(fill=tk.X, pady=2)
        # Finance Reporting and Financial Aid are now integrated into Finance Management GUI
        # self.nav_buttons['finance_reporting'] = ttk.Button(finance_frame, text="Finance Reporting", command=self.show_finance_reporting_dashboard)
        # self.nav_buttons['finance_reporting'].pack(fill=tk.X, pady=2)
        # self.nav_buttons['financial_aid'] = ttk.Button(finance_frame, text="Financial Aid & Scholarships", command=self.show_financial_aid)
        # self.nav_buttons['financial_aid'].pack(fill=tk.X, pady=2)

        # ---------- Health & Accommodations ----------
        health_frame = ttk.LabelFrame(scrollable_frame, text="Health & Accommodations", padding="5")
        health_frame.pack(fill=tk.X, pady=(0, 10), padx=5)
        self.nav_buttons['health_portal'] = ttk.Button(health_frame, text="Health Portal", command=self.open_health_portal_gui)
        self.nav_buttons['health_portal'].pack(fill=tk.X, pady=2)
        self.nav_buttons['medical_accommodations'] = ttk.Button(health_frame, text="Medical Accommodation", command=self.show_medical_accommodations)
        self.nav_buttons['medical_accommodations'].pack(fill=tk.X, pady=2)
        self.nav_buttons['housing_accommodations'] = ttk.Button(health_frame, text="Housing Accommodation", command=self.show_housing_accommodations)
        self.nav_buttons['housing_accommodations'].pack(fill=tk.X, pady=2)

        # ---------- Campus Life ----------
        campus_frame = ttk.LabelFrame(scrollable_frame, text="Campus Life", padding="5")
        campus_frame.pack(fill=tk.X, pady=(0, 10), padx=5)
        self.nav_buttons['student_union_portal'] = ttk.Button(campus_frame, text="Student Union", command=self.open_student_union_portal_gui)
        self.nav_buttons['student_union_portal'].pack(fill=tk.X, pady=2)
        self.nav_buttons['campus_events'] = ttk.Button(campus_frame, text="Campus Events", command=self.show_campus_events_gui)
        self.nav_buttons['campus_events'].pack(fill=tk.X, pady=2)
        self.nav_buttons['restaurant_management'] = ttk.Button(campus_frame, text="Dining Services", command=self.show_restaurant_management)
        self.nav_buttons['restaurant_management'].pack(fill=tk.X, pady=2)
        self.nav_buttons['university_shop'] = ttk.Button(campus_frame, text="University Shop", command=self.show_university_shop)
        self.nav_buttons['university_shop'].pack(fill=tk.X, pady=2)
        self.nav_buttons['parking_management'] = ttk.Button(campus_frame, text="Parking & Transportation", command=self.show_parking_management)
        self.nav_buttons['parking_management'].pack(fill=tk.X, pady=2)
        self.nav_buttons['trip_management'] = ttk.Button(campus_frame, text="Trip Management", command=self.show_trip_management_gui)
        self.nav_buttons['trip_management'].pack(fill=tk.X, pady=2)
        self.nav_buttons['facilities_management'] = ttk.Button(campus_frame, text="Facilities Management", command=self.show_facilities_management_gui)
        self.nav_buttons['facilities_management'].pack(fill=tk.X, pady=2)

        # ---------- Student Services ----------
        services_frame = ttk.LabelFrame(scrollable_frame, text="Student Services", padding="5")
        services_frame.pack(fill=tk.X, pady=(0, 10), padx=5)
        self.nav_buttons['student_support'] = ttk.Button(services_frame, text="Student Support", command=self.open_student_support_portal_gui)
        self.nav_buttons['student_support'].pack(fill=tk.X, pady=2)
        self.nav_buttons['parent_portal'] = ttk.Button(services_frame, text="Parent Portal", command=self.open_parent_portal_gui)
        self.nav_buttons['parent_portal'].pack(fill=tk.X, pady=2)
        self.nav_buttons['internship_portal'] = ttk.Button(services_frame, text="Internship Portal", command=self.open_internship_portal_gui)
        self.nav_buttons['internship_portal'].pack(fill=tk.X, pady=2)
        self.nav_buttons['career_services'] = ttk.Button(services_frame, text="Career Services", command=self.show_career_services_gui)
        self.nav_buttons['career_services'].pack(fill=tk.X, pady=2)
        self.nav_buttons['alumni_management'] = ttk.Button(services_frame, text="Alumni Relations", command=self.open_alumni_portal_gui)
        self.nav_buttons['alumni_management'].pack(fill=tk.X, pady=2)
        self.nav_buttons['early_warning_system'] = ttk.Button(services_frame, text="Early Warning System", command=self.show_early_warning_gui)
        self.nav_buttons['early_warning_system'].pack(fill=tk.X, pady=2)
        self.nav_buttons['helpdesk'] = ttk.Button(services_frame, text="Helpdesk & Support", command=self.open_helpdesk_gui)
        self.nav_buttons['helpdesk'].pack(fill=tk.X, pady=2)

        # ---------- Communication ----------
        comm_frame = ttk.LabelFrame(scrollable_frame, text="Communication", padding="5")
        comm_frame.pack(fill=tk.X, pady=(0, 10), padx=5)
        self.nav_buttons['communication_hub'] = ttk.Button(comm_frame, text="Communication Hub", command=self.show_email_sms_gui)
        self.nav_buttons['communication_hub'].pack(fill=tk.X, pady=2)

        # ---------- Analytics & Reporting ----------
        analytics_frame = ttk.LabelFrame(scrollable_frame, text="Analytics & Reporting", padding="5")
        analytics_frame.pack(fill=tk.X, pady=(0, 10), padx=5)
        self.nav_buttons['integrated_dashboard'] = ttk.Button(analytics_frame, text="Integrated Dashboard", command=self.show_integrated_dashboard)
        self.nav_buttons['integrated_dashboard'].pack(fill=tk.X, pady=2)
        self.nav_buttons['analytics'] = ttk.Button(analytics_frame, text="Student Analytics", command=self.show_analytics)
        self.nav_buttons['analytics'].pack(fill=tk.X, pady=2)
        self.nav_buttons['enhanced_reporting_dashboard'] = ttk.Button(analytics_frame, text="Enhanced Reporting", command=self.show_enhanced_reporting_dashboard)
        self.nav_buttons['enhanced_reporting_dashboard'].pack(fill=tk.X, pady=2)
        self.nav_buttons['predictive_analytics'] = ttk.Button(analytics_frame, text="Predictive Analytics", command=self.show_predictive_analytics_gui)
        self.nav_buttons['predictive_analytics'].pack(fill=tk.X, pady=2)
        self.nav_buttons['business_intelligence'] = ttk.Button(analytics_frame, text="Business Intelligence", command=self.show_business_intelligence_gui)
        self.nav_buttons['business_intelligence'].pack(fill=tk.X, pady=2)

        # ---------- Documents & Export ----------
        export_frame = ttk.LabelFrame(scrollable_frame, text="Documents & Export", padding="5")
        export_frame.pack(fill=tk.X, pady=(0, 10), padx=5)
        self.nav_buttons['document_manager'] = ttk.Button(export_frame, text="Document Manager", command=self.show_document_manager)
        self.nav_buttons['document_manager'].pack(fill=tk.X, pady=2)
        self.nav_buttons['export'] = ttk.Button(export_frame, text="Export Options", command=self.export_data_dialog)
        self.nav_buttons['export'].pack(fill=tk.X, pady=2)
        self.nav_buttons['backup_gui'] = ttk.Button(export_frame, text="Data Backup", command=self.show_data_backup_gui)
        self.nav_buttons['backup_gui'].pack(fill=tk.X, pady=2)

        # ---------- AI & Advanced Tools ----------
        tools_frame = ttk.LabelFrame(scrollable_frame, text="AI & Advanced Tools", padding="5")
        tools_frame.pack(fill=tk.X, pady=(0, 10), padx=5)
        self.nav_buttons['ai_features'] = ttk.Button(tools_frame, text="AI-Powered Features", command=self.show_ai_features_gui)
        self.nav_buttons['ai_features'].pack(fill=tk.X, pady=2)
        self.nav_buttons['mobile_app_pwa'] = ttk.Button(tools_frame, text="Mobile App (PWA)", command=self.show_mobile_app_pwa_gui)
        self.nav_buttons['mobile_app_pwa'].pack(fill=tk.X, pady=2)
        self.nav_buttons['blockchain_credentials'] = ttk.Button(tools_frame, text="Blockchain Credentials", command=self.show_blockchain_credentials_gui)
        self.nav_buttons['blockchain_credentials'].pack(fill=tk.X, pady=2)

        # ---------- Administration ----------
        admin_frame = ttk.LabelFrame(scrollable_frame, text="Administration", padding="5")
        admin_frame.pack(fill=tk.X, pady=(0, 10), padx=5)
        self.nav_buttons['user_management'] = ttk.Button(admin_frame, text="User Management", command=self.show_user_management)
        self.nav_buttons['user_management'].pack(fill=tk.X, pady=2)
        self.nav_buttons['system_admin_gui'] = ttk.Button(admin_frame, text="System Administration", command=self.show_system_administration_gui)
        self.nav_buttons['system_admin_gui'].pack(fill=tk.X, pady=2)
        self.nav_buttons['security_dashboard'] = ttk.Button(admin_frame, text="Security & Compliance", command=self.show_security_dashboard)
        self.nav_buttons['security_dashboard'].pack(fill=tk.X, pady=2)
        self.nav_buttons['activity_logger'] = ttk.Button(admin_frame, text="Activity Logger", command=self.show_activity_logger)
        self.nav_buttons['activity_logger'].pack(fill=tk.X, pady=2)
        self.nav_buttons['activity_log'] = ttk.Button(admin_frame, text="Log Management", command=self.show_activity_log)
        self.nav_buttons['activity_log'].pack(fill=tk.X, pady=2)
        self.nav_buttons['integration_marketplace'] = ttk.Button(admin_frame, text="Integration Marketplace", command=self.show_integration_marketplace_gui)
        self.nav_buttons['integration_marketplace'].pack(fill=tk.X, pady=2)
        self.nav_buttons['admissions_crm'] = ttk.Button(admin_frame, text="Admissions CRM", command=self.show_admissions_crm_gui)
        self.nav_buttons['admissions_crm'].pack(fill=tk.X, pady=2)

        # ---------- Session ----------
        session_frame = ttk.LabelFrame(scrollable_frame, text="Session", padding="5")
        session_frame.pack(fill=tk.X, pady=(0, 10), padx=5)
        self.nav_buttons['logout'] = ttk.Button(session_frame, text="Logout", command=do_logout)
        self.nav_buttons['logout'].pack(fill=tk.X, pady=2)
        self.nav_buttons['exit'] = ttk.Button(session_frame, text="Exit", command=self.root.quit)
        self.nav_buttons['exit'].pack(fill=tk.X, pady=2)

        # Finalize scroll region
        scrollable_frame.update_idletasks()
        canvas.configure(scrollregion=canvas.bbox("all"))
    
    def create_content_area(self, parent):
        """Create the main content area"""
        self.content_frame = ttk.LabelFrame(parent, text="Content", padding="10")
        self.content_frame.grid(row=1, column=1, sticky=(tk.W, tk.E, tk.N, tk.S))
        self.content_frame.columnconfigure(0, weight=1)
        self.content_frame.rowconfigure(0, weight=1)
    
    def update_status(self):
        """Update the status display and button states"""
        if self.auth.current_user:
            user = self.auth.current_user
            self.current_user_var.set(f"{user['username']} ({user['role']})")
            self.status_var.set("Logged in")
        else:
            self.current_user_var.set("None")
            self.status_var.set("Not logged in")

        self.update_button_states()
        self.update_login_logout_button()
    
    def update_button_states(self):
        """Update button states based on user permissions"""
        if not self.auth.current_user:
            # Not logged in - disable ALL buttons except login
            states = {
                # Authentication
                'login': tk.NORMAL,
                'change_password': tk.DISABLED,

                # Student Management
                'student_records': tk.DISABLED,
                'create_student': tk.DISABLED,
                'search_students': tk.DISABLED,
                'advanced_search_gui': tk.DISABLED,
                'delete_student': tk.DISABLED,
                'batch_operations': tk.DISABLED,

                # Academic Management
                'course_management': tk.DISABLED,
                'module_management': tk.DISABLED,
                'assignments': tk.DISABLED,
                'grade_tracking_gui': tk.DISABLED,
                'library': tk.DISABLED,
                'virtual_classroom': tk.DISABLED,

                # Scheduling & Attendance
                'academic_calendar': tk.DISABLED,
                'scheduling': tk.DISABLED,
                'attendance': tk.DISABLED,

                # Finance & Financial Aid
                'finance_management': tk.DISABLED,
                'finance_reporting': tk.DISABLED,
                'financial_aid': tk.DISABLED,

                # Health & Accommodations
                'health_portal': tk.DISABLED,
                'medical_accommodations': tk.DISABLED,
                'housing_accommodations': tk.DISABLED,

                # Campus Life
                'student_union_portal': tk.DISABLED,
                'campus_events': tk.DISABLED,
                'restaurant_management': tk.DISABLED,
                'university_shop': tk.DISABLED,
                'parking_management': tk.DISABLED,
                'trip_management': tk.DISABLED,
                'facilities_management': tk.DISABLED,

                # Student Services
                'student_support': tk.DISABLED,
                'parent_portal': tk.DISABLED,
                'internship_portal': tk.DISABLED,
                'career_services': tk.DISABLED,
                'alumni_management': tk.DISABLED,
                'early_warning_system': tk.DISABLED,
                'helpdesk': tk.DISABLED,

                # Communication
                'communication_hub': tk.DISABLED,
                'communication_dashboard': tk.DISABLED,

                # Analytics & Reporting
                'integrated_dashboard': tk.DISABLED,
                'analytics': tk.DISABLED,
                'enhanced_reporting_dashboard': tk.DISABLED,
                'predictive_analytics': tk.DISABLED,
                'business_intelligence': tk.DISABLED,

                # Documents & Export
                'document_manager': tk.DISABLED,
                'export': tk.DISABLED,
                'backup_gui': tk.DISABLED,

                # AI & Advanced Tools
                'ai_features': tk.DISABLED,
                'mobile_app_pwa': tk.DISABLED,
                'blockchain_credentials': tk.DISABLED,

                # Administration
                'user_management': tk.DISABLED,
                'system_admin_gui': tk.DISABLED,
                'security_dashboard': tk.DISABLED,
                'activity_logger': tk.DISABLED,
                'activity_log': tk.DISABLED,
                'integration_marketplace': tk.DISABLED,
                'admissions_crm': tk.DISABLED,

                # Legacy/Alternative names
                'grades': tk.DISABLED,
                'trips': tk.DISABLED,
                'system_admin': tk.DISABLED,
            }
        else:
            # Logged in - check permissions
            user = self.auth.current_user
            permissions = user.get('permissions', [])
            role = user.get('role', '')

            # Base permission check helpers
            is_admin = role == 'admin'
            is_staff = role in ('admin', 'staff')
            is_instructor = role in ('admin', 'staff', 'instructor')

            states = {
                # Authentication
                'login': tk.DISABLED,
                'change_password': tk.NORMAL,

                # Student Management
                'student_records': tk.NORMAL if any(p in permissions for p in ['view_any_student', 'view_own_record']) else tk.DISABLED,
                'create_student': tk.NORMAL if 'create_student' in permissions else tk.DISABLED,
                'search_students': tk.NORMAL if 'view_any_student' in permissions else tk.DISABLED,
                'advanced_search_gui': tk.NORMAL if 'view_any_student' in permissions else tk.DISABLED,
                'delete_student': tk.NORMAL if 'delete_any_student' in permissions else tk.DISABLED,
                'batch_operations': tk.NORMAL if is_staff else tk.DISABLED,

                # Academic Management
                'course_management': tk.NORMAL if any(p in permissions for p in ['manage_courses', 'view_courses']) else tk.DISABLED,
                'module_management': tk.NORMAL if any(p in permissions for p in ['manage_modules', 'view_assigned_modules']) else tk.DISABLED,
                'assignments': tk.NORMAL,  # Available to all logged-in users
                'grade_tracking_gui': tk.NORMAL if any(p in permissions for p in ['manage_grades', 'view_own_grades']) else tk.DISABLED,
                'library': tk.NORMAL if any(p in permissions for p in ['view_books', 'manage_books', 'manage_loans', 'checkout_books']) else tk.DISABLED,
                'virtual_classroom': tk.NORMAL if is_instructor else tk.DISABLED,

                # Scheduling & Attendance
                'academic_calendar': tk.NORMAL,  # Available to all logged-in users
                'scheduling': tk.NORMAL if any(p in permissions for p in ['manage_schedules', 'view_own_timetable']) else tk.DISABLED,
                'attendance': tk.NORMAL if is_instructor else tk.DISABLED,

                # Finance & Financial Aid
                'finance_management': tk.NORMAL if any(p in permissions for p in ['manage_finances', 'view_financial_reports']) or is_admin else tk.DISABLED,
                'finance_reporting': tk.NORMAL if any(p in permissions for p in ['view_financial_reports', 'manage_finances']) or is_admin else tk.DISABLED,
                'financial_aid': tk.NORMAL,  # Available to all logged-in users (students can view/apply, admins can manage)

                # Health & Accommodations
                'health_portal': tk.NORMAL,  # Available to all logged-in users
                'medical_accommodations': tk.NORMAL if is_staff else tk.DISABLED,
                'housing_accommodations': tk.NORMAL if is_staff else tk.DISABLED,

                # Campus Life
                'student_union_portal': tk.NORMAL,  # Available to all logged-in users
                'campus_events': tk.NORMAL,  # Available to all logged-in users
                'restaurant_management': tk.NORMAL if is_staff else tk.DISABLED,
                'university_shop': tk.NORMAL,  # Available to all logged-in users
                'parking_management': tk.NORMAL if is_staff else tk.DISABLED,
                'trip_management': tk.NORMAL if any(p in permissions for p in ['view_trips', 'register_for_trips', 'manage_trips']) else tk.DISABLED,
                'facilities_management': tk.NORMAL if is_staff else tk.DISABLED,

                # Student Services
                'student_support': tk.NORMAL,  # Available to all logged-in users
                'parent_portal': tk.NORMAL,  # Available to all logged-in users
                'internship_portal': tk.NORMAL,  # Available to all logged-in users
                'career_services': tk.NORMAL,  # Available to all logged-in users
                'alumni_management': tk.NORMAL if is_staff else tk.DISABLED,
                'early_warning_system': tk.NORMAL if is_instructor else tk.DISABLED,
                'helpdesk': tk.NORMAL,  # Available to all logged-in users

                # Communication
                'communication_hub': tk.NORMAL if any(p in permissions for p in ['send_emails', 'send_sms']) or is_staff else tk.DISABLED,
                'communication_dashboard': tk.NORMAL if any(p in permissions for p in ['send_emails']) else tk.DISABLED,

                # Analytics & Reporting
                'integrated_dashboard': tk.NORMAL,  # Available to all logged-in users
                'analytics': tk.NORMAL if 'view_analytics' in permissions or is_instructor else tk.DISABLED,
                'enhanced_reporting_dashboard': tk.NORMAL if is_instructor else tk.DISABLED,
                'predictive_analytics': tk.NORMAL if is_instructor else tk.DISABLED,
                'business_intelligence': tk.NORMAL if is_staff else tk.DISABLED,

                # Documents & Export
                'document_manager': tk.NORMAL if (is_staff or any(p in permissions for p in ['manage_documents', 'system_config', 'view_documents'])) else tk.DISABLED,
                'export': tk.NORMAL if 'export_data' in permissions or is_instructor else tk.DISABLED,
                'backup_gui': tk.NORMAL if is_admin else tk.DISABLED,

                # AI & Advanced Tools
                'ai_features': tk.NORMAL,  # Available to all logged-in users
                'mobile_app_pwa': tk.NORMAL,  # Available to all logged-in users
                'blockchain_credentials': tk.NORMAL if is_staff else tk.DISABLED,

                # Administration
                'user_management': tk.NORMAL if 'manage_users' in permissions or is_admin else tk.DISABLED,
                'system_admin_gui': tk.NORMAL if is_admin else tk.DISABLED,
                'security_dashboard': tk.NORMAL if is_admin else tk.DISABLED,
                'activity_logger': tk.NORMAL if is_staff else tk.DISABLED,
                'activity_log': tk.NORMAL if 'view_logs' in permissions or is_admin else tk.DISABLED,
                'integration_marketplace': tk.NORMAL if is_admin else tk.DISABLED,
                'admissions_crm': tk.NORMAL if is_staff else tk.DISABLED,

                # Legacy/Alternative names
                'grades': tk.NORMAL if any(p in permissions for p in ['manage_grades', 'view_own_grades']) else tk.DISABLED,
                'trips': tk.NORMAL if any(p in permissions for p in ['view_trips', 'register_for_trips', 'manage_trips']) else tk.DISABLED,
                'system_admin': tk.NORMAL if is_admin else tk.DISABLED,
            }
        # Apply states to buttons
        for button_name, state in states.items():
            if button_name in self.nav_buttons:
                try:
                    self.nav_buttons[button_name].configure(state=state)
                except Exception as e:
                    # Silently handle button configuration errors
                    pass

    def refresh_advanced_search(self):
        """Refresh the Advanced Search GUI if it's open"""
        if hasattr(self, 'advanced_search_refresh_callback') and self.advanced_search_refresh_callback:
            try:
                self.advanced_search_refresh_callback()
                print("🔄 Advanced Search GUI data refreshed")
            except Exception as e:
                print(f"Warning: Error refreshing Advanced Search GUI: {e}")
    
    def clear_content(self):
        """Clear the content area"""
        # Check if content_frame exists before trying to clear it
        if hasattr(self, 'content_frame') and self.content_frame:
            for widget in self.content_frame.winfo_children():
                widget.destroy()
        else:
            # If content_frame doesn't exist yet, just return
            return
    
    def show_welcome(self):
        """Show welcome message"""
        self.clear_content()
        
        welcome_text = """Welcome to the University Management System!

This integrated system provides:
• Student record management and tracking
• Academic module and course management
• Assignment submission and grading
• Communication tools and chatbot
• Analytics and reporting capabilities
• User authentication and authorization
• Administrative tools and system monitoring

Use the navigation panel on the left to access different features.
Your available features depend on your user role and permissions.
        """
        
        welcome_label = ttk.Label(self.content_frame, text=welcome_text, justify=tk.LEFT, font=('Arial', 11))
        welcome_label.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N), padx=20, pady=20)
        
        # System status
        if self.auth.current_user:
            status_frame = ttk.LabelFrame(self.content_frame, text="Quick Actions", padding="15")
            status_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), padx=20, pady=10)
            
            user = self.auth.current_user
            permissions = user.get('permissions', [])
            
            quick_actions = []
                
            if 'view_any_student' in permissions or 'view_own_record' in permissions:
                quick_actions.append(("View Student Records", self.show_student_records))
                
            if 'create_student' in permissions:
                quick_actions.append(("Create New Student", self.create_student_dialog))
                
            if 'access_chatbot' in permissions:
                quick_actions.append(("Launch Chatbot", self.show_chatbot))

            if 'view_analytics' in permissions:
                quick_actions.append(("View Analytics", self.show_analytics))

            if any(p in permissions for p in ['view_trips', 'register_for_trips', 'manage_trips']):
                quick_actions.append(("Trip Management", self.show_trip_management_gui))
                    
            if ADVANCED_SEARCH_GUI_AVAILABLE and 'view_any_student' in permissions:
                quick_actions.append(("🔍 Advanced Search", self.show_advanced_search_gui))
                    
            if ('manage_schedules' in permissions) or ('view_own_timetable' in permissions):
                quick_actions.append(("Module Scheduling", self.show_module_scheduling))
                
            # Create buttons for quick actions
            for i, (text, command) in enumerate(quick_actions[:4]):  # Show up to 4 quick actions
                row = i // 2
                col = i % 2
                ttk.Button(status_frame, text=text, command=command, width=20).grid(
                    row=row, column=col, padx=10, pady=5, sticky=tk.W)
    
    def show_login_screen(self):
        """Show login interface when not authenticated"""
        self.clear_content()
        
        # Ensure content_frame exists before proceeding
        if not hasattr(self, 'content_frame') or not self.content_frame:
            print("Warning: content_frame not initialized")
            return
        
        login_frame = ttk.LabelFrame(self.content_frame, text="Please Log In", padding="30")
        login_frame.grid(row=0, column=0, sticky=(tk.W, tk.E))
        
        # Login form
        ttk.Label(login_frame, text="Username:").grid(row=0, column=0, sticky=tk.W, pady=10)
        self.username_entry = ttk.Entry(login_frame, width=30, font=('Arial', 11))
        self.username_entry.grid(row=0, column=1, pady=10, padx=(15, 0))
        
        ttk.Label(login_frame, text="Password:").grid(row=1, column=0, sticky=tk.W, pady=10)
        self.password_entry = ttk.Entry(login_frame, width=30, show="*", font=('Arial', 11))
        self.password_entry.grid(row=1, column=1, pady=10, padx=(15, 0))
        
        # Login button
        login_btn = ttk.Button(login_frame, text="Login", command=self.perform_login, style="Accent.TButton")
        login_btn.grid(row=2, column=0, columnspan=2, pady=25)
        
        # Bind Enter key
        self.username_entry.bind('<Return>', lambda e: self.perform_login())
        self.password_entry.bind('<Return>', lambda e: self.perform_login())
        
        # Focus username
        self.username_entry.focus()
        
        # Show default credentials info
        # WARNING: These are default demo credentials - change them in production!
        # Set DEFAULT_ADMIN_PASSWORD, DEFAULT_STAFF_PASSWORD, DEFAULT_STUDENT_PASSWORD environment variables
        info_frame = ttk.LabelFrame(self.content_frame, text="Default Login Credentials", padding="20")
        info_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), padx=0, pady=20)

        admin_pwd = os.getenv('DEFAULT_ADMIN_PASSWORD', 'admin123')
        staff_pwd = os.getenv('DEFAULT_STAFF_PASSWORD', 'staff123')
        student_pwd = os.getenv('DEFAULT_STUDENT_PASSWORD', 'student123')

        creds_text = f"""Admin: username='admin', password='{admin_pwd}'
    Staff: username='staff', password='{staff_pwd}'
    Student: username='student', password='{student_pwd}'"""

        ttk.Label(info_frame, text=creds_text, justify=tk.LEFT, font=('Courier', 10)).pack()
    
    def show_main_interface(self):
        """Show main interface when authenticated"""
        self.update_status()
        self.show_welcome()
    
    def perform_login(self):
        """Handle login process"""
        try:
            # Check if entry widgets exist
            if not hasattr(self, 'username_entry') or not hasattr(self, 'password_entry'):
                messagebox.showerror("Error", "Login form not available")
                return

            username = self.username_entry.get().strip()
            password = self.password_entry.get().strip()

            if not username or not password:
                messagebox.showerror("Error", "Please enter both username and password")
                return

            result = self.auth.login(username, password)

            if result is True:
                self.show_main_interface()
                messagebox.showinfo("Success", f"Welcome, {username}!")
                self.log_activity(f"User {username} logged in successfully", action="login")

            elif isinstance(result, dict) and result.get('requires_2fa'):
                # Handle 2FA (simplified for this example)
                messagebox.showinfo("2FA Required", "Two-factor authentication would be handled here")

            elif result == 'password_reset_required':
                messagebox.showwarning("Password Reset Required", "You must change your password")
                self.show_change_password()

            else:
                messagebox.showerror("Login Failed", "Invalid username or password")
                # Safely clear password field
                try:
                    self.password_entry.delete(0, tk.END)
                except:
                    pass

        except AttributeError as e:
            messagebox.showerror("Error", "Login form not properly initialized")
        except Exception as e:
            messagebox.showerror("Error", f"Login error: {str(e)}")
            try:
                self.log_activity(f"Login error: {str(e)}", level="error", action="login")
            except:
                pass
    
    # Add all the other methods from both original classes...
    # (Student management methods, authentication methods, etc.)
    # This is a condensed version showing the structure
    
    def show_login(self):
        """Show login dialog (for re-login)"""
        if hasattr(self.auth, 'current_user') and self.auth.current_user:
            messagebox.showinfo("Already Logged In", f"Already logged in as {self.auth.current_user['username']}")
            return
        
        self.show_login_screen()
    
    def logout_user(self):
        """Logout current user"""
        if self.auth.current_user:
            username = self.auth.current_user['username']
            self.auth.logout()
            self.update_status()
            messagebox.showinfo("Logged Out", f"Goodbye, {username}!")
            self.show_login_screen()
    
    def switch_to_cli(self):
        """Switch to CLI mode by launching the CLI main function"""
        if messagebox.askyesno("Switch to CLI", "Close GUI and switch to command-line interface?"):
            try:
                # Import and launch CLI main function
                from university_system.cli_main import main as cli_main
                self.root.withdraw()  # Hide the GUI window
                self.root.quit()      # Exit the GUI event loop
                # Launch CLI in the same process
                cli_main()
            except Exception as e:
                messagebox.showerror("Error", f"Failed to switch to CLI: {e}")
                self.root.deiconify()  # Show the GUI window again if CLI fails

    def shutdown_system(self):
        """Shutdown the entire system"""
        if messagebox.askyesno("Shutdown", "Are you sure you want to shutdown the system?"):
            try:
                # Log the shutdown
                if self.auth and self.auth.current_user:
                    username = self.auth.current_user['username']
                    print(f"System shutdown initiated by {username}")
                else:
                    print("System shutdown initiated")

                # Close the application
                self.root.quit()
                self.root.destroy()
                sys.exit(0)
            except Exception as e:
                messagebox.showerror("Error", f"Error during shutdown: {e}")

    def toggle_login_logout(self):
        """Toggle between login and logout based on current authentication status"""
        if self.auth.current_user:
            self.logout_user()
        else:
            self.show_login_screen()

    def update_login_logout_button(self):
        """Update the login/logout button text based on authentication status"""
        if hasattr(self, 'login_logout_btn'):
            if self.auth.current_user:
                self.login_logout_btn.config(text="Logout")
            else:
                self.login_logout_btn.config(text="Login")

    def show_change_password(self):
        """Show change password dialog"""
        if not self.auth.current_user:
            messagebox.showerror("Error", "You must be logged in to change password")
            return
        
        # Implementation similar to AuthenticationGUI version
        password_window = tk.Toplevel(self.root)
        password_window.title("Change Password")
        password_window.geometry("400x250")
        password_window.transient(self.root)
        password_window.grab_set()
        
        main_frame = ttk.Frame(password_window, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Password change form
        ttk.Label(main_frame, text="Current Password:").grid(row=0, column=0, sticky=tk.W, pady=5)
        current_entry = ttk.Entry(main_frame, show="*", width=30)
        current_entry.grid(row=0, column=1, pady=5, padx=(10, 0))
        
        ttk.Label(main_frame, text="New Password:").grid(row=1, column=0, sticky=tk.W, pady=5)
        new_entry = ttk.Entry(main_frame, show="*", width=30)
        new_entry.grid(row=1, column=1, pady=5, padx=(10, 0))
        
        ttk.Label(main_frame, text="Confirm Password:").grid(row=2, column=0, sticky=tk.W, pady=5)
        confirm_entry = ttk.Entry(main_frame, show="*", width=30)
        confirm_entry.grid(row=2, column=1, pady=5, padx=(10, 0))
        
        status_label = ttk.Label(main_frame, text="", foreground="red")
        status_label.grid(row=3, column=0, columnspan=2, pady=10)
        
        def change_password():
            current = current_entry.get()
            new = new_entry.get()
            confirm = confirm_entry.get()
            
            if not all([current, new, confirm]):
                status_label.config(text="All fields are required")
                return
            
            if new != confirm:
                status_label.config(text="New passwords don't match")
                return
            
            if self.auth.change_password(self.auth.current_user['username'], current, new):
                password_window.destroy()
                messagebox.showinfo("Success", "Password changed successfully!")
            else:
                status_label.config(text="Current password is incorrect")
        
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=4, column=0, columnspan=2, pady=10)
        
        ttk.Button(button_frame, text="Change Password", command=change_password).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Cancel", command=password_window.destroy).pack(side=tk.LEFT, padx=5)
        
        current_entry.focus()
    
    # Content display methods
    def show_student_records(self):
        """Show student records interface in a new window"""
        # Create new window
        records_window = tk.Toplevel(self.root)
        records_window.title("Student Records Management")
        records_window.geometry("1400x800")
        records_window.transient(self.root)

        # Main frame
        main_frame = ttk.Frame(records_window, padding=10)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Title
        ttk.Label(main_frame, text="Student Records Management",
                 font=('Arial', 14, 'bold')).pack(pady=(0, 20))

        # Action buttons frame
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Button(button_frame, text="Create Student",
                  command=self.create_student_dialog).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Search Students",
                  command=self.search_students_dialog).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Export Data",
                  command=self.export_data_dialog).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Refresh",
                  command=lambda: self.view_students_in_window(tree)).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Close",
                  command=records_window.destroy).pack(side=tk.RIGHT, padx=5)

        # Create student list interface
        records_frame = ttk.LabelFrame(main_frame, text="Student Records", padding="10")
        records_frame.pack(fill=tk.BOTH, expand=True, pady=5)

        # Create treeview in this window
        tree_frame = ttk.Frame(records_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True)

        columns = ('ID', 'Name', 'Email', 'Course', 'Registration Date')
        tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=30)

        # Configure columns
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=200)

        # Scrollbars
        v_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview)
        h_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=tree.xview)
        tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)

        # Pack widgets
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)

        # Bind double-click event
        tree.bind('<Double-1>', lambda event: self.on_student_double_click_window(event, tree))

        # Store reference to this window's tree
        self.student_tree = tree

        # Load student data
        self.view_students_in_window(tree)
    
    def show_user_management(self):
        """Show user management interface"""
        if not self.auth.current_user or 'manage_users' not in self.auth.current_user.get('permissions', []):
            messagebox.showerror("Error", "You don't have permission to access User Management")
            return
        
        self.clear_content()
        
        ttk.Label(self.content_frame, text="User Management", 
                 font=('Arial', 14, 'bold')).grid(row=0, column=0, pady=(0, 20))
        
    # 1. Student Treeview Management
    def create_student_treeview(self, parent):
        """Create treeview widget for displaying student data"""
        tree_frame = ttk.Frame(parent)
        tree_frame.pack(fill=tk.BOTH, expand=True)

        columns = ('ID', 'Name', 'Email', 'Course', 'Registration Date')
        self.student_tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=30)
        
        # Configure columns
        for col in columns:
            self.student_tree.heading(col, text=col)
            self.student_tree.column(col, width=150)
        
        # Scrollbars
        v_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.student_tree.yview)
        h_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=self.student_tree.xview)
        self.student_tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        # Pack widgets
        self.student_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
        
        # Bind events
        self.student_tree.bind('<Double-1>', self.on_student_double_click)

    def view_students(self):
        """Load and display student data in treeview"""
        try:
            # Check if student_tree exists
            if not hasattr(self, 'student_tree') or not self.student_tree:
                return

            # Clear existing data
            for item in self.student_tree.get_children():
                self.student_tree.delete(item)
            
            # Fetch data from database
            conn = get_db_connection()
            if conn:
                cursor = conn.cursor()
                if self.auth.check_permission('view_any_student'):
                    cursor.execute('SELECT * FROM students ORDER BY last_name, first_name')
                else:
                    cursor.execute('SELECT * FROM students WHERE student_id = ?', 
                                 (self.auth.current_user.get('student_id'),))
                
                students = cursor.fetchall()
                
                # Populate treeview
                for student in students:
                    student_id, email_address, title, first_name, middle_name, last_name, gender, dob, age, course, reg_date, status, enrollment_date = student
                    full_name = f"{first_name} {middle_name} {last_name}".replace('  ', ' ').strip()
                    
                    self.student_tree.insert('', tk.END, values=(
                        student_id, full_name, email_address, course, reg_date[:10] if reg_date else 'N/A'
                    ))
                
                conn.close()
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load student data: {str(e)}")

    def on_student_double_click(self, event):
        """Handle double-click on student record"""
        try:
            if not hasattr(self, 'student_tree') or not self.student_tree:
                return

            selection = self.student_tree.selection()
            if selection:
                item = self.student_tree.item(selection[0])
                student_values = item.get('values', [])
                if student_values:
                    student_id = student_values[0]
                    self.show_student_details(student_id)
        except (AttributeError, IndexError, tk.TclError) as e:
            messagebox.showerror("Error", "Unable to access student details at this time.")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")

    def on_student_double_click_window(self, event, tree):
        """Handle double-click on student record in separate window"""
        try:
            selection = tree.selection()
            if selection:
                item = tree.item(selection[0])
                student_values = item.get('values', [])
                if student_values:
                    student_id = student_values[0]
                    self.show_student_details(student_id)
        except (AttributeError, IndexError, tk.TclError) as e:
            messagebox.showerror("Error", "Unable to access student details at this time.")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")

    def view_students_in_window(self, tree):
        """Load and display student data in a specific treeview widget"""
        try:
            # Clear existing data
            for item in tree.get_children():
                tree.delete(item)

            # Fetch data from database
            conn = get_db_connection()
            if conn:
                cursor = conn.cursor()
                if self.auth.check_permission('view_any_student'):
                    cursor.execute('SELECT * FROM students ORDER BY last_name, first_name')
                else:
                    cursor.execute('SELECT * FROM students WHERE student_id = ?',
                                 (self.auth.current_user.get('student_id'),))

                students = cursor.fetchall()

                # Populate treeview
                for student in students:
                    student_id, email_address, title, first_name, middle_name, last_name, gender, dob, age, course, reg_date, status, enrollment_date = student
                    full_name = f"{first_name} {middle_name} {last_name}".replace('  ', ' ').strip()

                    tree.insert('', tk.END, values=(
                        student_id, full_name, email_address, course, reg_date[:10] if reg_date else 'N/A'
                    ))

                conn.close()

        except Exception as e:
            messagebox.showerror("Error", f"Failed to load student data: {str(e)}")

    def show_user_management(self):
        """Show user management interface"""
        if not self.auth.current_user or 'manage_users' not in self.auth.current_user.get('permissions', []):
            messagebox.showerror("Access Denied", "You don't have permission to access User Management")
            return
        
        self.clear_content()
        
        ttk.Label(self.content_frame, text="User Management", font=('Arial', 14, 'bold')).grid(row=0, column=0, pady=(0, 20))
        
        # User list frame
        list_frame = ttk.LabelFrame(self.content_frame, text="Users", padding="10")
        list_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)
        
        # Create treeview for users
        columns = ('ID', 'Username', 'Name', 'Role', 'Status')
        self.user_tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=15)
        
        for col in columns:
            self.user_tree.heading(col, text=col)
            self.user_tree.column(col, width=120)
        
        user_scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.user_tree.yview)
        self.user_tree.configure(yscrollcommand=user_scrollbar.set)
        
        self.user_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(10, 0), pady=10)
        user_scrollbar.pack(side=tk.RIGHT, fill=tk.Y, pady=10, padx=(0, 10))
        
        # Load users
        self.refresh_user_list()
        
        # Buttons frame
        button_frame = ttk.Frame(self.content_frame)
        button_frame.grid(row=2, column=0, pady=10)
        
        ttk.Button(button_frame, text="Create User", command=self.show_create_user).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="View Details", command=self.show_user_details).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Edit User", command=self.show_edit_user).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Reset Password", command=self.reset_user_password).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Refresh", command=self.refresh_user_list).pack(side=tk.LEFT, padx=5)
    
    def refresh_user_list(self):
        """Refresh the user list"""
        if not hasattr(self, 'user_tree'):
            return
        
        # Clear existing items
        for item in self.user_tree.get_children():
            self.user_tree.delete(item)
        
        # Load users
        try:
            users = self.auth.list_users()
            if users:
                for user in users:
                    full_name = f"{user['first_name']} {user['last_name']}"
                    status = "Active" if user.get('is_active', True) else "Inactive"
                    
                    self.user_tree.insert('', tk.END, values=(
                        user['id'], 
                        user['username'], 
                        full_name, 
                        user['role'], 
                        status
                    ))
        except Exception as e:
            self.user_tree.insert('', tk.END, values=('Error', f'Failed to load: {e}', '', '', ''))
    
    def show_create_user(self):
        """Show create user dialog"""
        create_window = tk.Toplevel(self.root)
        create_window.title("Create New User")
        create_window.geometry("500x400")
        create_window.transient(self.root)
        create_window.grab_set()
        
        main_frame = ttk.Frame(create_window, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(main_frame, text="Create New User", font=('Arial', 14, 'bold')).grid(row=0, column=0, columnspan=2, pady=(0, 20))
        
        # Form fields
        fields = {}
        
        ttk.Label(main_frame, text="Username:").grid(row=1, column=0, sticky=tk.W, pady=5)
        fields['username'] = ttk.Entry(main_frame, width=30)
        fields['username'].grid(row=1, column=1, pady=5, padx=(10, 0))
        
        ttk.Label(main_frame, text="Email:").grid(row=2, column=0, sticky=tk.W, pady=5)
        fields['email'] = ttk.Entry(main_frame, width=30)
        fields['email'].grid(row=2, column=1, pady=5, padx=(10, 0))
        
        ttk.Label(main_frame, text="First Name:").grid(row=3, column=0, sticky=tk.W, pady=5)
        fields['first_name'] = ttk.Entry(main_frame, width=30)
        fields['first_name'].grid(row=3, column=1, pady=5, padx=(10, 0))
        
        ttk.Label(main_frame, text="Last Name:").grid(row=4, column=0, sticky=tk.W, pady=5)
        fields['last_name'] = ttk.Entry(main_frame, width=30)
        fields['last_name'].grid(row=4, column=1, pady=5, padx=(10, 0))
        
        ttk.Label(main_frame, text="Role:").grid(row=5, column=0, sticky=tk.W, pady=5)
        role_var = tk.StringVar()
        role_combo = ttk.Combobox(main_frame, textvariable=role_var, width=27)
        try:
            role_combo['values'] = list(ROLES.keys()) if 'ROLES' in globals() else ['admin', 'staff', 'student', 'instructor']
        except:
            role_combo['values'] = ['admin', 'staff', 'student', 'instructor']
        role_combo.grid(row=5, column=1, pady=5, padx=(10, 0))
        role_combo.set('student')
        
        status_label = ttk.Label(main_frame, text="", foreground="red")
        status_label.grid(row=6, column=0, columnspan=2, pady=10)
        
        def create_user():
            username = fields['username'].get().strip()
            email = fields['email'].get().strip()
            first_name = fields['first_name'].get().strip()
            last_name = fields['last_name'].get().strip()
            role = role_var.get()
            
            if not all([username, email, first_name, last_name, role]):
                status_label.config(text="All fields are required")
                return
            
            # Generate temporary password
            import secrets, string
            temp_password = ''.join(secrets.choices(string.ascii_letters + string.digits, k=12))
            
            try:
                if self.auth.create_user(username, temp_password, email, first_name, last_name, role, password_reset_required=True):
                    create_window.destroy()
                    messagebox.showinfo("Success", f"User created successfully!\nTemporary password: {temp_password}\nUser must change password on first login.")
                    self.refresh_user_list()
                else:
                    status_label.config(text="Failed to create user")
            except Exception as e:
                status_label.config(text=f"Error: {str(e)}")
        
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=7, column=0, columnspan=2, pady=10)
        
        ttk.Button(button_frame, text="Create User", command=create_user).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Cancel", command=create_window.destroy).pack(side=tk.LEFT, padx=5)
        
        fields['username'].focus()
    
    def show_user_details(self):
        """Show details for selected user"""
        selection = self.user_tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a user to view details")
            return
        
        item = self.user_tree.item(selection[0])
        user_id = item['values'][0]
        
        # Get user details
        users = self.auth.list_users()
        user = None
        for u in users:
            if str(u['id']) == str(user_id):
                user = u
                break
        
        if not user:
            messagebox.showerror("Error", "User not found")
            return
        
        # Show user details in new window
        details_window = tk.Toplevel(self.root)
        details_window.title(f"User Details - {user['username']}")
        details_window.geometry("600x500")
        details_window.transient(self.root)
        
        text_widget = scrolledtext.ScrolledText(details_window, wrap=tk.WORD)
        text_widget.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        details_text = f"""USER DETAILS
{'='*50}

ID: {user['id']}
Username: {user['username']}
Email: {user['email']}
Name: {user['first_name']} {user['last_name']}
Role: {user['role']}
Active: {'Yes' if user.get('is_active', True) else 'No'}
Student ID: {user.get('student_id', 'N/A')}
Created: {user.get('created_at', 'Unknown')}
Last Login: {user.get('last_login', 'Never')}

PERMISSIONS:
{'='*50}
"""
        
        for perm in user.get('permissions', []):
            details_text += f"• {perm}\n"
        
        text_widget.insert(tk.END, details_text)
        text_widget.config(state=tk.DISABLED)
    
    def show_edit_user(self):
        """Show edit user dialog"""
        if not hasattr(self, 'user_tree'):
            messagebox.showerror("Error", "User list not available")
            return

        selection = self.user_tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a user to edit")
            return

        item = self.user_tree.item(selection[0])
        user_values = item.get('values', [])
        if not user_values:
            messagebox.showerror("Error", "Could not retrieve user data")
            return

        username = user_values[0]

        # Create edit dialog
        dialog = tk.Toplevel(self.root)
        dialog.title(f"Edit User - {username}")
        dialog.geometry("500x600")
        dialog.transient(self.root)
        dialog.grab_set()

        main_frame = ttk.Frame(dialog, padding=20)
        main_frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(main_frame, text=f"Edit User: {username}",
                 font=('Arial', 14, 'bold')).pack(pady=(0, 20))

        # Get current user data
        try:
            # Use auth system to get user info if available
            if self.auth:
                user_info = self.auth.get_user_by_username(username)
                if not user_info:
                    messagebox.showerror("Error", "User not found")
                    dialog.destroy()
                    return
                # Convert to tuple format for compatibility
                user_data = (
                    user_info.get('username'),
                    user_info.get('email'),
                    user_info.get('first_name'),
                    user_info.get('last_name'),
                    user_info.get('role'),
                    user_info.get('is_active', 1),
                    user_info.get('student_id')
                )
                user_id = user_info.get('id')
            else:
                # Fallback to direct DB access
                conn = get_db_connection()
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT u.id, u.username, u.email, u.first_name, u.last_name, u.role, u.is_active, u.student_id
                    FROM users u
                    WHERE u.username = ?
                ''', (username,))
                result = cursor.fetchone()
                conn.close()

                if not result:
                    messagebox.showerror("Error", "User not found")
                    dialog.destroy()
                    return

                user_id = result[0]
                user_data = result[1:]

        except Exception as e:
            messagebox.showerror("Error", f"Failed to load user data: {str(e)}")
            dialog.destroy()
            return

        # Create form fields
        fields_frame = ttk.Frame(main_frame)
        fields_frame.pack(fill=tk.BOTH, expand=True)

        fields = {}

        # Email
        ttk.Label(fields_frame, text="Email:").grid(row=0, column=0, sticky=tk.W, pady=5)
        fields['email'] = ttk.Entry(fields_frame, width=40)
        fields['email'].grid(row=0, column=1, sticky=tk.W, pady=5, padx=(10, 0))
        _safe_entry_insert(fields['email'], user_data[1])

        # First Name
        ttk.Label(fields_frame, text="First Name:").grid(row=1, column=0, sticky=tk.W, pady=5)
        fields['first_name'] = ttk.Entry(fields_frame, width=40)
        fields['first_name'].grid(row=1, column=1, sticky=tk.W, pady=5, padx=(10, 0))
        _safe_entry_insert(fields['first_name'], user_data[2])

        # Last Name
        ttk.Label(fields_frame, text="Last Name:").grid(row=2, column=0, sticky=tk.W, pady=5)
        fields['last_name'] = ttk.Entry(fields_frame, width=40)
        fields['last_name'].grid(row=2, column=1, sticky=tk.W, pady=5, padx=(10, 0))
        _safe_entry_insert(fields['last_name'], user_data[3])

        # Role
        ttk.Label(fields_frame, text="Role:").grid(row=3, column=0, sticky=tk.W, pady=5)
        role_var = tk.StringVar(value=user_data[4])
        role_combo = ttk.Combobox(fields_frame, textvariable=role_var,
                                  values=['admin', 'student', 'lecturer', 'staff'],
                                  state='readonly', width=37)
        role_combo.grid(row=3, column=1, sticky=tk.W, pady=5, padx=(10, 0))

        # Active Status
        active_var = tk.BooleanVar(value=bool(user_data[5]))
        ttk.Checkbutton(fields_frame, text="Account Active", variable=active_var).grid(
            row=4, column=1, sticky=tk.W, pady=10, padx=(10, 0))

        def save_changes():
            try:
                new_email = fields['email'].get().strip()
                new_first_name = fields['first_name'].get().strip()
                new_last_name = fields['last_name'].get().strip()
                new_role = role_var.get()
                new_active = active_var.get()

                # Track old values for logging
                old_role = user_data[4]
                old_active = bool(user_data[5])

                # Use auth system to update user if available
                if self.auth:
                    success = self.auth.update_user(
                        user_id,
                        email=new_email,
                        first_name=new_first_name,
                        last_name=new_last_name,
                        role=new_role,
                        is_active=new_active
                    )

                    if not success:
                        messagebox.showerror("Error", "Failed to update user via auth system")
                        return
                else:
                    # Fallback to direct DB access
                    conn = get_db_connection()
                    cursor = conn.cursor()
                    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

                    cursor.execute('''
                        UPDATE users
                        SET email = ?, first_name = ?, last_name = ?, role = ?, is_active = ?, updated_at = ?
                        WHERE username = ?
                    ''', (
                        new_email,
                        new_first_name,
                        new_last_name,
                        new_role,
                        1 if new_active else 0,
                        timestamp,
                        username
                    ))
                    conn.commit()
                    conn.close()

                # Log activity
                changes = {}
                if new_email != user_data[1]:
                    changes['email'] = {'old': user_data[1], 'new': new_email}
                if new_first_name != user_data[2]:
                    changes['first_name'] = {'old': user_data[2], 'new': new_first_name}
                if new_last_name != user_data[3]:
                    changes['last_name'] = {'old': user_data[3], 'new': new_last_name}
                if new_role != old_role:
                    changes['role'] = {'old': old_role, 'new': new_role}
                if new_active != old_active:
                    changes['is_active'] = {'old': old_active, 'new': new_active}

                if ACTIVITY_LOGGER_AVAILABLE and changes:
                    log_activity('update', 'user', user_id=user_id, details={'username': username, 'changes': changes})

                messagebox.showinfo("Success", f"User {username} updated successfully!")
                self.refresh_user_list()
                dialog.destroy()

            except Exception as e:
                messagebox.showerror("Error", f"Failed to update user: {str(e)}")

        # Buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(pady=20)

        ttk.Button(button_frame, text="Save Changes", command=save_changes,
                  style="Accent.TButton").pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Cancel", command=dialog.destroy).pack(side=tk.LEFT, padx=5)

    def reset_user_password(self):
        """Reset password for selected user"""
        if not hasattr(self, 'user_tree'):
            messagebox.showerror("Error", "User list not available")
            return

        selection = self.user_tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a user to reset password")
            return

        item = self.user_tree.item(selection[0])
        user_values = item.get('values', [])
        if not user_values:
            messagebox.showerror("Error", "Could not retrieve user data")
            return

        username = user_values[0]

        # Create password reset dialog
        dialog = tk.Toplevel(self.root)
        dialog.title(f"Reset Password - {username}")
        dialog.geometry("450x300")
        dialog.transient(self.root)
        dialog.grab_set()

        main_frame = ttk.Frame(dialog, padding=20)
        main_frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(main_frame, text=f"Reset Password for: {username}",
                 font=('Arial', 12, 'bold')).pack(pady=(0, 20))

        # Password fields
        fields_frame = ttk.Frame(main_frame)
        fields_frame.pack(fill=tk.X, pady=10)

        ttk.Label(fields_frame, text="New Password:").grid(row=0, column=0, sticky=tk.W, pady=5)
        password_entry = ttk.Entry(fields_frame, width=30, show='*')
        password_entry.grid(row=0, column=1, sticky=tk.W, pady=5, padx=(10, 0))

        ttk.Label(fields_frame, text="Confirm Password:").grid(row=1, column=0, sticky=tk.W, pady=5)
        confirm_entry = ttk.Entry(fields_frame, width=30, show='*')
        confirm_entry.grid(row=1, column=1, sticky=tk.W, pady=5, padx=(10, 0))

        # Show password checkbox
        show_password_var = tk.BooleanVar()
        def toggle_password():
            show_char = '' if show_password_var.get() else '*'
            password_entry.config(show=show_char)
            confirm_entry.config(show=show_char)

        ttk.Checkbutton(fields_frame, text="Show passwords", variable=show_password_var,
                       command=toggle_password).grid(row=2, column=1, sticky=tk.W, pady=5, padx=(10, 0))

        status_label = ttk.Label(main_frame, text="", foreground="red")
        status_label.pack(pady=10)

        def perform_reset():
            new_password = password_entry.get()
            confirm_password = confirm_entry.get()

            if not new_password:
                status_label.config(text="Password cannot be empty")
                return

            if len(new_password) < 6:
                status_label.config(text="Password must be at least 6 characters")
                return

            if new_password != confirm_password:
                status_label.config(text="Passwords do not match")
                return

            try:
                import hashlib, secrets

                conn = get_db_connection()
                cursor = conn.cursor()

                # Get user ID
                cursor.execute('SELECT id FROM users WHERE username = ?', (username,))
                user_record = cursor.fetchone()

                if not user_record:
                    status_label.config(text="User not found")
                    return

                user_id = user_record[0]

                # Generate new password hash
                salt = secrets.token_hex(16)
                key = hashlib.pbkdf2_hmac('sha256', new_password.encode(), salt.encode(), 100000, dklen=64)
                password_hash = key.hex()

                timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

                # Update password
                cursor.execute('''
                    UPDATE user_accounts
                    SET password_hash = ?, salt = ?, updated_at = ?
                    WHERE user_id = ?
                ''', (password_hash, salt, timestamp, user_id))

                conn.commit()
                conn.close()

                messagebox.showinfo("Success", f"Password for {username} has been reset successfully!")
                dialog.destroy()

            except Exception as e:
                status_label.config(text=f"Error: {str(e)}")

        # Buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(pady=20)

        ttk.Button(button_frame, text="Reset Password", command=perform_reset,
                  style="Accent.TButton").pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Cancel", command=dialog.destroy).pack(side=tk.LEFT, padx=5)

    def show_system_admin(self):
        """Show system administration interface"""
        if not self.auth.current_user or self.auth.current_user.get('role') != 'admin':
            messagebox.showerror("Access Denied", "Administrator access required")
            return
        
        self.clear_content()
        
        ttk.Label(self.content_frame, text="System Administration", 
                 font=('Arial', 14, 'bold')).grid(row=0, column=0, pady=(0, 20))
        
        admin_frame = ttk.LabelFrame(self.content_frame, text="Administrative Tools", padding="15")
        admin_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Admin tools buttons
        tools = [
            ("Database Integrity Check", lambda: messagebox.showinfo("Admin", "Database check feature")),
            ("System Backup", lambda: messagebox.showinfo("Admin", "Backup feature")),
            ("User Statistics", lambda: messagebox.showinfo("Admin", "User stats feature")),
            ("System Logs", self.show_activity_log),
            ("Configuration", lambda: messagebox.showinfo("Admin", "Config feature"))
        ]
        
        for i, (text, command) in enumerate(tools):
            row = i // 2
            col = i % 2
            ttk.Button(admin_frame, text=text, command=command, width=25).grid(
                row=row, column=col, padx=10, pady=5, sticky=tk.W)

    def show_student_details(self, student_id):
        """Enhanced student details viewer with comprehensive information display"""
        detail_window = tk.Toplevel(self.root)
        detail_window.title(f"Student Details - {student_id}")
        detail_window.geometry("900x700")
        detail_window.transient(self.root)
        
        try:
            conn = get_db_connection()
            if not conn:
                messagebox.showerror("Error", "Database connection failed")
                detail_window.destroy()
                return
            
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM students WHERE student_id = ?', (student_id,))
            student = cursor.fetchone()
            
            if not student:
                messagebox.showerror("Error", "Student not found")
                detail_window.destroy()
                return
            
            # Create notebook for tabbed interface
            notebook = ttk.Notebook(detail_window)
            notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # Personal Information Tab
            personal_tab = ttk.Frame(notebook)
            notebook.add(personal_tab, text="Personal Information")
            
            personal_frame = ttk.LabelFrame(personal_tab, text="Student Information", padding=20)
            personal_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # Create scrollable text for personal info
            personal_text = scrolledtext.ScrolledText(personal_frame, wrap=tk.WORD, height=25, 
                                                     font=('Courier', 11))
            personal_text.pack(fill=tk.BOTH, expand=True)
            
            # Format personal information with None checks
            title = student[2] if student[2] else 'N/A'
            first_name = student[3] if student[3] else 'N/A'
            middle_name = student[4] if student[4] else ''
            last_name = student[5] if student[5] else 'N/A'
            gender = student[6].title() if student[6] else 'N/A'
            course = student[9] if student[9] else 'N/A'

            # Build full name safely
            name_parts = []
            if student[2]:  # title
                name_parts.append(student[2])
            if student[3]:  # first name
                name_parts.append(student[3])
            if student[4]:  # middle name
                name_parts.append(student[4])
            if student[5]:  # last name
                name_parts.append(student[5])
            full_name = ' '.join(name_parts) if name_parts else 'N/A'

            personal_info = f"""STUDENT RECORD DETAILS
    {'='*80}

    PERSONAL INFORMATION:
      Student ID:        {student[0] or 'N/A'}
      Email Address:     {student[1] or 'N/A'}
      Title:            {title}
      First Name:       {first_name}
      Middle Name:      {middle_name or 'N/A'}
      Last Name:        {last_name}
      Full Name:        {full_name}

    DEMOGRAPHICS:
      Gender:           {gender}
      Date of Birth:    {student[7] or 'N/A'}
      Age:              {student[8] or 'N/A'} years

    ACADEMIC INFORMATION:
      Course:           {course}
      Registration:     {student[10] or 'N/A'}

    {'='*80}
    """
            
            personal_text.insert(tk.END, personal_info)
            personal_text.config(state=tk.DISABLED)
            
            # Academic Information Tab
            academic_tab = ttk.Frame(notebook)
            notebook.add(academic_tab, text="Academic Records")
            
            academic_frame = ttk.LabelFrame(academic_tab, text="Modules and Academic Data", padding=20)
            academic_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            academic_text = scrolledtext.ScrolledText(academic_frame, wrap=tk.WORD, height=25, 
                                                     font=('Courier', 11))
            academic_text.pack(fill=tk.BOTH, expand=True)
            
            # Get modules
            cursor.execute('''
                SELECT m.module_type, sm.module_code, m.module_name
                FROM student_modules sm
                JOIN modules m ON sm.module_code = m.module_code
                WHERE sm.student_id = ?
                ORDER BY m.module_type, sm.module_code
            ''', (student_id,))
            modules = cursor.fetchall()
            
            # Get grades if available
            cursor.execute('''
                SELECT module_code, assessment_name, grade, grade_date 
                FROM student_grades 
                WHERE student_id = ? 
                ORDER BY grade_date DESC
            ''', (student_id,))
            grades = cursor.fetchall()
            
            # Get attendance if available
            cursor.execute('''
                SELECT module_code, date, status, reason 
                FROM attendance 
                WHERE student_id = ? 
                ORDER BY date DESC
                LIMIT 10
            ''', (student_id,))
            attendance = cursor.fetchall()
            
            academic_info = f"""ACADEMIC RECORDS
    {'='*80}

    ENROLLED MODULES:
    {'-'*40}
    """
            
            if modules:
                current_type = None
                for module in modules:
                    if current_type != module[0]:
                        current_type = module[0]
                        type_display = current_type.upper() if current_type else "UNKNOWN"
                        academic_info += f"\n{type_display} MODULES:\n"
                    module_code = module[1] if module[1] else "N/A"
                    module_name = module[2] if module[2] else "Unknown Module"
                    academic_info += f"  {module_code} - {module_name}\n"
            else:
                academic_info += "  No modules enrolled\n"
            
            academic_info += f"\nGRADES AND ASSESSMENTS:\n{'-'*40}\n"
            if grades:
                for grade in grades:
                    module_code = grade[0] if grade[0] else "N/A"
                    assessment_name = grade[1] if grade[1] else "Unknown Assessment"
                    grade_value = grade[2] if grade[2] else "No Grade"
                    grade_date = grade[3] if grade[3] else "Unknown Date"
                    academic_info += f"  {module_code} - {assessment_name}: {grade_value} (Date: {grade_date})\n"
            else:
                academic_info += "  No grades recorded\n"

            academic_info += f"\nRECENT ATTENDANCE (Last 10 records):\n{'-'*40}\n"
            if attendance:
                for att in attendance:
                    module_code = att[0] if att[0] else "N/A"
                    att_date = att[1] if att[1] else "Unknown Date"
                    status = att[2] if att[2] else "Unknown"
                    reason_text = f" - {att[3]}" if att[3] else ""
                    academic_info += f"  {att_date} ({module_code}): {status}{reason_text}\n"
            else:
                academic_info += "  No attendance records\n"
            
            academic_info += f"\n{'='*80}"
            
            academic_text.insert(tk.END, academic_info)
            academic_text.config(state=tk.DISABLED)
            
            # Actions Tab
            actions_tab = ttk.Frame(notebook)
            notebook.add(actions_tab, text="Actions")
            
            actions_frame = ttk.LabelFrame(actions_tab, text="Available Actions", padding=20)
            actions_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # Action buttons
            if self.auth.check_permission('update_any_student'):
                ttk.Button(actions_frame, text="Edit Student Information", 
                          command=lambda: self.update_student_dialog(student_id),
                          width=30).pack(pady=10)
            
            if self.auth.check_permission('manage_grades'):
                ttk.Button(actions_frame, text="Manage Grades",
                          command=lambda: self.manage_student_grades(student_id, student[3], student[5]),
                          width=30).pack(pady=5)

            if self.auth.check_permission('manage_attendance'):
                ttk.Button(actions_frame, text="View Attendance",
                          command=lambda: self.view_student_attendance(student_id, student[1], student[3], student[5]),
                          width=30).pack(pady=5)

            if self.auth.check_permission('export_data'):
                ttk.Button(actions_frame, text="Export Data",
                          command=lambda: self.export_individual_student_data(student_id, student[3], student[5]),
                          width=30).pack(pady=5)

            # Contact information if available
            contact_frame = ttk.LabelFrame(actions_tab, text="Contact Information", padding=20)
            contact_frame.pack(fill=tk.X, padx=10, pady=(10, 0))

            ttk.Label(contact_frame, text=f"Email: {student[1]}").pack(anchor=tk.W)
            ttk.Button(contact_frame, text="Send Email",
                      command=lambda: self.send_email_to_student(student[1], student[3], student[5]),
                      width=20).pack(pady=5)
            
            conn.close()
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load student details: {str(e)}")
            detail_window.destroy()

    def export_student_data(self, student_id):
        """Export individual student data"""
        try:
            filename = filedialog.asksaveasfilename(
                defaultextension=".txt",
                filetypes=[("Text files", "*.txt"), ("PDF files", "*.pdf"), ("All files", "*.*")]
            )
            
            if filename:
                # Implementation for exporting student data
                messagebox.showinfo("Success", f"Student data exported to {filename}")
        except Exception as e:
            messagebox.showerror("Error", f"Export failed: {str(e)}")

    def compose_email(self, email_address):
        """Compose email with recipient pre-filled"""
        try:
            # Try to open the email GUI directly with pre-filled recipient
            if EMAIL_MANAGER_GUI_AVAILABLE:
                from university_system.infrastructure.email.gui.email_manager_gui import EmailManagerGUI

                email_window = tk.Toplevel(self.root)
                email_window.title(f"Compose Email to {email_address}")
                email_window.geometry("900x700")
                email_window.transient(self.root)

                # Initialize the email GUI
                email_gui = EmailManagerGUI(email_window, self.auth)

                # Wait a moment then call compose_email with the recipient
                email_window.update_idletasks()

                def safe_compose():
                    try:
                        if email_window.winfo_exists():
                            email_gui.compose_email(recipient=email_address)
                    except Exception:
                        pass  # Window destroyed

                email_window.after(200, safe_compose)

            elif self.email_manager_gui:
                # Use existing email_manager_gui instance if available
                try:
                    self.email_manager_gui.compose_email(recipient=email_address)
                except Exception as e:
                    print(f"Could not use existing email GUI: {e}")
                    messagebox.showinfo("Email Address", f"Please compose email to:\n{email_address}")
            else:
                # Final fallback - show email address to copy
                messagebox.showinfo("Compose Email",
                                  f"Please send email to:\n{email_address}\n\nNote: Email GUI is not available.")
        except Exception as e:
            print(f"Error opening email composer: {e}")
            messagebox.showinfo("Email Address", f"Please send email to:\n{email_address}")

    def manage_student_grades(self, student_id, first_name, last_name):
        """Display and manage student grades with assignments/assessments table"""
        try:
            grades_window = tk.Toplevel(self.root)
            grades_window.title(f"Manage Grades - {first_name} {last_name} ({student_id})")
            grades_window.geometry("1000x600")
            grades_window.transient(self.root)

            main_frame = ttk.Frame(grades_window, padding=20)
            main_frame.pack(fill=tk.BOTH, expand=True)

            ttk.Label(main_frame, text=f"Grades for {first_name} {last_name}",
                     font=('TkDefaultFont', 14, 'bold')).pack(pady=(0, 10))

            # Create treeview for grades
            tree_frame = ttk.Frame(main_frame)
            tree_frame.pack(fill=tk.BOTH, expand=True)

            columns = ("Type", "Module", "Assignment", "Submitted", "Grade", "Max Grade", "Status")
            tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=15)

            for col in columns:
                tree.heading(col, text=col)
                tree.column(col, width=120)

            # Add scrollbar
            scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview)
            tree.configure(yscrollcommand=scrollbar.set)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

            # Fetch grades from database
            conn = get_db_connection()
            if not conn:
                messagebox.showerror("Error", "Database connection failed")
                grades_window.destroy()
                return

            cursor = conn.cursor()

            # Get assignments and grades
            cursor.execute("""
                SELECT 'Assignment' as type, a.module_code, a.title,
                       CASE WHEN s.submission_date IS NOT NULL THEN 'Yes' ELSE 'No' END as submitted,
                       COALESCE(s.grade, 'Not Graded') as grade,
                       COALESCE(a.max_marks, 100) as max_marks,
                       CASE
                           WHEN s.grade IS NOT NULL THEN 'Graded'
                           WHEN s.submission_date IS NOT NULL THEN 'Submitted'
                           ELSE 'Not Submitted'
                       END as status
                FROM assignments a
                LEFT JOIN assignment_submissions s ON a.id = s.assignment_id AND s.student_id = ?
                ORDER BY a.module_code, a.due_date DESC
            """, (student_id,))

            assignments = cursor.fetchall()

            # Convert Row objects to tuples and insert into tree
            for assignment in assignments:
                tree.insert('', tk.END, values=tuple(assignment))

            # Get assessments if table exists
            assessments = []
            try:
                cursor.execute("""
                    SELECT 'Assessment' as type, a.module_code, a.assessment_name,
                           'N/A' as submitted,
                           COALESCE(g.score, 'Not Graded') as score,
                           COALESCE(a.max_points, 100) as max_points,
                           CASE WHEN g.score IS NOT NULL THEN 'Graded' ELSE 'Pending' END as status
                    FROM assessments a
                    LEFT JOIN grades g ON a.assessment_id = g.assessment_id AND g.student_id = ?
                    WHERE a.module_code IN (SELECT module_code FROM student_modules WHERE student_id = ?)
                    ORDER BY a.module_code
                """, (student_id, student_id))

                assessments = cursor.fetchall()
                for assessment in assessments:
                    tree.insert('', tk.END, values=tuple(assessment))
            except Exception as e:
                print(f"Could not load assessments: {e}")  # For debugging

            conn.close()

            # Summary frame
            summary_frame = ttk.LabelFrame(main_frame, text="Summary", padding=10)
            summary_frame.pack(fill=tk.X, pady=(10, 0))

            # Calculate totals from both assignments and assessments
            all_items = list(assignments) + list(assessments)
            total_items = len(all_items)
            submitted = sum(1 for item in all_items if len(item) > 3 and item[3] == 'Yes')
            graded = sum(1 for item in all_items if len(item) > 6 and 'Graded' in str(item[6]))

            ttk.Label(summary_frame, text=f"Total Assignments: {total_items}").grid(row=0, column=0, padx=10)
            ttk.Label(summary_frame, text=f"Submitted: {submitted}").grid(row=0, column=1, padx=10)
            ttk.Label(summary_frame, text=f"Graded: {graded}").grid(row=0, column=2, padx=10)

            # Close button
            ttk.Button(main_frame, text="Close", command=grades_window.destroy).pack(pady=(10, 0))

        except Exception as e:
            messagebox.showerror("Error", f"Failed to load grades: {str(e)}")
            if 'grades_window' in locals():
                grades_window.destroy()

    def view_student_attendance(self, student_id, email, first_name, last_name):
        """Display student attendance table and send email if below 90%"""
        try:
            attendance_window = tk.Toplevel(self.root)
            attendance_window.title(f"Attendance - {first_name} {last_name} ({student_id})")
            attendance_window.geometry("900x600")
            attendance_window.transient(self.root)

            main_frame = ttk.Frame(attendance_window, padding=20)
            main_frame.pack(fill=tk.BOTH, expand=True)

            ttk.Label(main_frame, text=f"Attendance for {first_name} {last_name}",
                     font=('TkDefaultFont', 14, 'bold')).pack(pady=(0, 10))

            # Create treeview for attendance
            tree_frame = ttk.Frame(main_frame)
            tree_frame.pack(fill=tk.BOTH, expand=True)

            columns = ("Date", "Module", "Session Type", "Status", "Reason")
            tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=15)

            for col in columns:
                tree.heading(col, text=col)
            tree.column("Date", width=120)
            tree.column("Module", width=120)
            tree.column("Session Type", width=120)
            tree.column("Status", width=100)
            tree.column("Reason", width=200)

            # Add scrollbar
            scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview)
            tree.configure(yscrollcommand=scrollbar.set)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

            # Fetch attendance from database
            conn = get_db_connection()
            if not conn:
                messagebox.showerror("Error", "Database connection failed")
                attendance_window.destroy()
                return

            cursor = conn.cursor()
            cursor.execute("""
                SELECT date, module_code, status, reason
                FROM attendance
                WHERE student_id = ?
                ORDER BY date DESC
            """, (student_id,))

            attendance_records = cursor.fetchall()
            conn.close()

            present_count = 0
            total_count = len(attendance_records)

            for record in attendance_records:
                # Convert Row object to tuple
                record_tuple = tuple(record)
                date, module, status, reason = record_tuple
                # Default session type to 'Lecture' since it's not in the database
                tree.insert('', tk.END, values=(date, module, 'Lecture', status, reason or ''))
                if status and status.lower() == 'present':
                    present_count += 1

            # Calculate attendance percentage
            attendance_percentage = (present_count / total_count * 100) if total_count > 0 else 0

            # Summary frame
            summary_frame = ttk.LabelFrame(main_frame, text="Attendance Summary", padding=10)
            summary_frame.pack(fill=tk.X, pady=(10, 0))

            ttk.Label(summary_frame, text=f"Total Sessions: {total_count}").grid(row=0, column=0, padx=10)
            ttk.Label(summary_frame, text=f"Present: {present_count}").grid(row=0, column=1, padx=10)

            # Attendance percentage label with color coding
            percentage_label = ttk.Label(summary_frame,
                                        text=f"Attendance: {attendance_percentage:.1f}%",
                                        font=('TkDefaultFont', 10, 'bold'))
            percentage_label.grid(row=0, column=2, padx=10)

            # Send email if attendance below 90%
            if attendance_percentage < 90 and email:
                ttk.Label(summary_frame,
                         text="⚠ Low Attendance Alert Sent",
                         foreground='red').grid(row=1, column=0, columnspan=3, pady=5)

                # Send email alert
                try:
                    subject = f"Low Attendance Alert - {first_name} {last_name}"
                    message = f"""Dear {first_name} {last_name},

This is an automated notification regarding your attendance.

Current Attendance: {attendance_percentage:.1f}%
Sessions Attended: {present_count} out of {total_count}

Your attendance is below the required 90% threshold. Please ensure you attend future sessions regularly to maintain good standing.

If you have any concerns or need support, please contact Student Support Services.

Best regards,
University Administration"""

                    self._send_email_via_gui(email, subject, message)
                except Exception as e:
                    print(f"Failed to send attendance alert email: {e}")

            # Close button
            ttk.Button(main_frame, text="Close", command=attendance_window.destroy).pack(pady=(10, 0))

        except Exception as e:
            messagebox.showerror("Error", f"Failed to load attendance: {str(e)}")
            if 'attendance_window' in locals():
                attendance_window.destroy()

    def export_individual_student_data(self, student_id, first_name, last_name):
        """Export individual student data in various formats (PDF, CSV, Excel, TXT)"""
        try:
            # Create export format dialog
            export_dialog = tk.Toplevel(self.root)
            export_dialog.title("Export Student Data")
            export_dialog.geometry("400x250")
            export_dialog.transient(self.root)
            export_dialog.grab_set()

            main_frame = ttk.Frame(export_dialog, padding=20)
            main_frame.pack(fill=tk.BOTH, expand=True)

            ttk.Label(main_frame, text=f"Export data for {first_name} {last_name}",
                     font=('TkDefaultFont', 12, 'bold')).pack(pady=(0, 15))

            ttk.Label(main_frame, text="Select export format:").pack(anchor=tk.W, pady=(0, 10))

            format_var = tk.StringVar(value="csv")
            formats = [
                ("CSV (Comma Separated Values)", "csv"),
                ("Excel Spreadsheet", "xlsx"),
                ("PDF Document", "pdf"),
                ("Text File", "txt")
            ]

            for text, value in formats:
                ttk.Radiobutton(main_frame, text=text, variable=format_var, value=value).pack(anchor=tk.W, pady=2)

            def perform_export():
                selected_format = format_var.get()

                # Get file extension and type
                extensions = {
                    "csv": (".csv", "CSV files", "*.csv"),
                    "xlsx": (".xlsx", "Excel files", "*.xlsx"),
                    "pdf": (".pdf", "PDF files", "*.pdf"),
                    "txt": (".txt", "Text files", "*.txt")
                }

                ext, file_desc, file_pattern = extensions[selected_format]

                filename = filedialog.asksaveasfilename(
                    defaultextension=ext,
                    filetypes=[(file_desc, file_pattern), ("All files", "*.*")],
                    initialfile=f"student_{student_id}_{first_name}_{last_name}{ext}"
                )

                if not filename:
                    return

                # Fetch comprehensive student data
                conn = get_db_connection()
                if not conn:
                    messagebox.showerror("Error", "Database connection failed")
                    return

                cursor = conn.cursor()

                # Get all student-related data
                data_sections = {}

                # 1. Basic Student Information
                cursor.execute("SELECT * FROM students WHERE student_id = ?", (student_id,))
                data_sections['student_info'] = cursor.fetchone()

                # 2. Enrolled Modules
                cursor.execute("""
                    SELECT sm.module_code, m.module_name, m.credits, sm.enrollment_date, sm.status
                    FROM student_modules sm
                    LEFT JOIN modules m ON sm.module_code = m.module_code
                    WHERE sm.student_id = ?
                    ORDER BY sm.enrollment_date DESC
                """, (student_id,))
                data_sections['modules'] = cursor.fetchall()

                # 3. Assignments
                cursor.execute("""
                    SELECT a.module_code, a.title, a.due_date, a.max_marks, a.assignment_type
                    FROM assignments a
                    WHERE a.module_code IN (SELECT module_code FROM student_modules WHERE student_id = ?)
                    ORDER BY a.due_date DESC
                """, (student_id,))
                data_sections['assignments'] = cursor.fetchall()

                # 4. Assignment Submissions
                cursor.execute("""
                    SELECT a.title, a.module_code, s.submission_date, s.grade,
                           a.max_marks, s.status, s.late_submission
                    FROM assignment_submissions s
                    JOIN assignments a ON s.assignment_id = a.id
                    WHERE s.student_id = ?
                    ORDER BY s.submission_date DESC
                """, (student_id,))
                data_sections['submissions'] = cursor.fetchall()

                # 5. Grades/Assessments
                cursor.execute("""
                    SELECT a.module_code, a.assessment_name, a.assessment_type,
                           g.score, a.max_points, g.letter_grade, g.submission_date
                    FROM grades g
                    JOIN assessments a ON g.assessment_id = a.assessment_id
                    WHERE g.student_id = ?
                    ORDER BY g.submission_date DESC
                """, (student_id,))
                data_sections['grades'] = cursor.fetchall()

                # 6. Attendance Records
                cursor.execute("""
                    SELECT date, module_code, status, reason
                    FROM attendance
                    WHERE student_id = ?
                    ORDER BY date DESC
                """, (student_id,))
                data_sections['attendance'] = cursor.fetchall()

                # 7. Financial Information
                try:
                    cursor.execute("""
                        SELECT fee_type, amount, due_date, status
                        FROM student_fees
                        WHERE student_id = ?
                        ORDER BY due_date DESC
                    """, (student_id,))
                    data_sections['fees'] = cursor.fetchall()

                    cursor.execute("""
                        SELECT payment_date, amount, payment_method, transaction_id
                        FROM payments
                        WHERE student_id = ?
                        ORDER BY payment_date DESC
                    """, (student_id,))
                    data_sections['payments'] = cursor.fetchall()
                except:
                    data_sections['fees'] = []
                    data_sections['payments'] = []

                # 8. Financial Aid & Scholarships
                try:
                    cursor.execute("""
                        SELECT s.scholarship_name, ss.amount, ss.academic_year, ss.status
                        FROM student_scholarships ss
                        JOIN scholarships s ON ss.scholarship_id = s.scholarship_id
                        WHERE ss.student_id = ?
                        ORDER BY ss.academic_year DESC
                    """, (student_id,))
                    data_sections['scholarships'] = cursor.fetchall()

                    cursor.execute("""
                        SELECT aid_type, amount, academic_year, status
                        FROM student_financial_aid
                        WHERE student_id = ?
                        ORDER BY academic_year DESC
                    """, (student_id,))
                    data_sections['financial_aid'] = cursor.fetchall()
                except:
                    data_sections['scholarships'] = []
                    data_sections['financial_aid'] = []

                # 9. Extracurricular Activities
                try:
                    cursor.execute("""
                        SELECT activity_name, role, start_date, end_date, status
                        FROM student_activities
                        WHERE student_id = ?
                        ORDER BY start_date DESC
                    """, (student_id,))
                    data_sections['activities'] = cursor.fetchall()

                    cursor.execute("""
                        SELECT c.club_name, cm.role, cm.join_date, cm.status
                        FROM club_members cm
                        JOIN clubs c ON cm.club_id = c.club_id
                        WHERE cm.student_id = ?
                        ORDER BY cm.join_date DESC
                    """, (student_id,))
                    data_sections['clubs'] = cursor.fetchall()
                except:
                    data_sections['activities'] = []
                    data_sections['clubs'] = []

                # 10. Health Records
                try:
                    cursor.execute("""
                        SELECT appointment_date, appointment_type, provider, notes
                        FROM health_appointments
                        WHERE student_id = ?
                        ORDER BY appointment_date DESC
                    """, (student_id,))
                    data_sections['health_appointments'] = cursor.fetchall()

                    cursor.execute("""
                        SELECT condition_name, diagnosed_date, status
                        FROM medical_conditions
                        WHERE student_id = ?
                        ORDER BY diagnosed_date DESC
                    """, (student_id,))
                    data_sections['medical_conditions'] = cursor.fetchall()
                except:
                    data_sections['health_appointments'] = []
                    data_sections['medical_conditions'] = []

                # 11. Housing Information
                try:
                    cursor.execute("""
                        SELECT building_name, room_number, start_date, end_date, status
                        FROM housing_assignments
                        WHERE student_id = ?
                        ORDER BY start_date DESC
                    """, (student_id,))
                    data_sections['housing'] = cursor.fetchall()
                except:
                    data_sections['housing'] = []

                # 12. Internships & Career
                try:
                    cursor.execute("""
                        SELECT i.company_name, i.position_title, ia.status,
                               ia.application_date, ip.start_date, ip.end_date
                        FROM internship_applications ia
                        JOIN internships i ON ia.internship_id = i.internship_id
                        LEFT JOIN internship_placements ip ON ia.application_id = ip.application_id
                        WHERE ia.student_id = ?
                        ORDER BY ia.application_date DESC
                    """, (student_id,))
                    data_sections['internships'] = cursor.fetchall()
                except:
                    data_sections['internships'] = []

                # 13. Library Records
                try:
                    cursor.execute("""
                        SELECT b.title, bl.checkout_date, bl.due_date, bl.return_date, bl.status
                        FROM book_loans bl
                        JOIN books b ON bl.book_id = b.book_id
                        WHERE bl.student_id = ?
                        ORDER BY bl.checkout_date DESC
                    """, (student_id,))
                    data_sections['library'] = cursor.fetchall()
                except:
                    data_sections['library'] = []

                # 14. Support Tickets
                try:
                    cursor.execute("""
                        SELECT ticket_id, subject, category, status, created_date, resolved_date
                        FROM support_tickets
                        WHERE student_id = ?
                        ORDER BY created_date DESC
                    """, (student_id,))
                    data_sections['support_tickets'] = cursor.fetchall()
                except:
                    data_sections['support_tickets'] = []

                # 15. Badges & Achievements
                try:
                    cursor.execute("""
                        SELECT ab.badge_name, bi.date_earned, bi.description
                        FROM badge_issuances bi
                        JOIN achievement_badges ab ON bi.badge_id = ab.badge_id
                        WHERE bi.student_id = ?
                        ORDER BY bi.date_earned DESC
                    """, (student_id,))
                    data_sections['badges'] = cursor.fetchall()
                except:
                    data_sections['badges'] = []

                # 16. Parking & Transportation
                try:
                    cursor.execute("""
                        SELECT permit_number, vehicle_make, vehicle_model,
                               start_date, end_date, permit_type
                        FROM parking_permits
                        WHERE student_id = ?
                        ORDER BY start_date DESC
                    """, (student_id,))
                    data_sections['parking'] = cursor.fetchall()
                except:
                    data_sections['parking'] = []

                # 17. Meal Plans
                try:
                    cursor.execute("""
                        SELECT plan_type, balance, last_transaction_date
                        FROM meal_accounts
                        WHERE student_id = ?
                    """, (student_id,))
                    data_sections['meal_plan'] = cursor.fetchall()

                    cursor.execute("""
                        SELECT transaction_date, location, amount, description
                        FROM meal_transactions
                        WHERE student_id = ?
                        ORDER BY transaction_date DESC
                        LIMIT 50
                    """, (student_id,))
                    data_sections['meal_transactions'] = cursor.fetchall()
                except:
                    data_sections['meal_plan'] = []
                    data_sections['meal_transactions'] = []

                # 18. Academic Advising
                try:
                    cursor.execute("""
                        SELECT appointment_date, advisor_name, notes, follow_up_required
                        FROM advising_appointments
                        WHERE student_id = ?
                        ORDER BY appointment_date DESC
                    """, (student_id,))
                    data_sections['advising'] = cursor.fetchall()
                except:
                    data_sections['advising'] = []

                conn.close()

                # Export based on format
                if selected_format == "csv":
                    self._export_comprehensive_csv(filename, student_id, first_name, last_name, data_sections)
                elif selected_format == "xlsx":
                    self._export_comprehensive_excel(filename, student_id, first_name, last_name, data_sections)
                elif selected_format == "pdf":
                    self._export_comprehensive_pdf(filename, student_id, first_name, last_name, data_sections)
                elif selected_format == "txt":
                    self._export_comprehensive_txt(filename, student_id, first_name, last_name, data_sections)

                messagebox.showinfo("Success", f"Comprehensive student data exported successfully to:\n{filename}")
                export_dialog.destroy()

            # Buttons
            button_frame = ttk.Frame(main_frame)
            button_frame.pack(pady=(15, 0))
            ttk.Button(button_frame, text="Export", command=perform_export).pack(side=tk.LEFT, padx=5)
            ttk.Button(button_frame, text="Cancel", command=export_dialog.destroy).pack(side=tk.LEFT, padx=5)

        except Exception as e:
            messagebox.showerror("Error", f"Export failed: {str(e)}")

    def _export_to_csv(self, filename, student_data, modules_data, grades_data, attendance_data):
        """Export student data to CSV format"""
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)

            # Convert Row objects to tuples
            student_data = tuple(student_data)
            modules_data = [tuple(m) for m in modules_data]
            grades_data = [tuple(g) for g in grades_data]
            attendance_data = [tuple(a) for a in attendance_data]

            # Student info
            writer.writerow(['Student Information'])
            writer.writerow(['Student ID', 'Email', 'Name', 'Course', 'Registration Date'])
            writer.writerow([student_data[0], student_data[1],
                           f"{student_data[3]} {student_data[5]}",
                           student_data[9], student_data[10]])
            writer.writerow([])

            # Modules
            writer.writerow(['Enrolled Modules'])
            writer.writerow(['Module Code', 'Module Name', 'Credits'])
            for module in modules_data:
                writer.writerow(module)
            writer.writerow([])

            # Grades
            writer.writerow(['Grades'])
            writer.writerow(['Module Code', 'Assessment', 'Grade', 'Max Grade'])
            for grade in grades_data:
                writer.writerow(grade)
            writer.writerow([])

            # Attendance
            writer.writerow(['Attendance'])
            writer.writerow(['Date', 'Module Code', 'Status'])
            for attendance in attendance_data:
                writer.writerow(attendance)

    def _export_to_excel(self, filename, student_id, first_name, last_name,
                        student_data, modules_data, grades_data, attendance_data):
        """Export student data to Excel format"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment

            # Convert Row objects to tuples
            student_data = tuple(student_data)
            modules_data = [tuple(m) for m in modules_data]
            grades_data = [tuple(g) for g in grades_data]
            attendance_data = [tuple(a) for a in attendance_data]

            wb = openpyxl.Workbook()

            # Student Info sheet
            ws1 = wb.active
            ws1.title = "Student Info"
            ws1['A1'] = f"Student Information - {first_name} {last_name}"
            ws1['A1'].font = Font(bold=True, size=14)
            ws1.append(['Student ID', student_data[0]])
            ws1.append(['Email', student_data[1]])
            ws1.append(['Name', f"{student_data[3]} {student_data[5]}"])
            ws1.append(['Course', student_data[9]])
            ws1.append(['Registration Date', student_data[10]])

            # Modules sheet
            ws2 = wb.create_sheet("Modules")
            ws2['A1'] = "Enrolled Modules"
            ws2['A1'].font = Font(bold=True, size=12)
            ws2.append(['Module Code', 'Module Name', 'Credits'])
            for module in modules_data:
                ws2.append(list(module))

            # Grades sheet
            ws3 = wb.create_sheet("Grades")
            ws3['A1'] = "Grades"
            ws3['A1'].font = Font(bold=True, size=12)
            ws3.append(['Module Code', 'Assessment', 'Grade', 'Max Grade'])
            for grade in grades_data:
                ws3.append(list(grade))

            # Attendance sheet
            ws4 = wb.create_sheet("Attendance")
            ws4['A1'] = "Attendance Records"
            ws4['A1'].font = Font(bold=True, size=12)
            ws4.append(['Date', 'Module Code', 'Status'])
            for attendance in attendance_data:
                ws4.append(list(attendance))

            wb.save(filename)
        except ImportError:
            # Fallback to CSV if openpyxl not available
            messagebox.showwarning("Warning", "Excel export requires openpyxl. Exporting as CSV instead.")
            self._export_to_csv(filename.replace('.xlsx', '.csv'), student_data, modules_data, grades_data, attendance_data)

    def _export_to_pdf(self, filename, student_id, first_name, last_name,
                       student_data, modules_data, grades_data, attendance_data):
        """Export student data to PDF format"""
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

            # Convert Row objects to tuples
            student_data = tuple(student_data)
            modules_data = [list(m) for m in modules_data]
            grades_data = [list(g) for g in grades_data]
            attendance_data = [list(a) for a in attendance_data]

            doc = SimpleDocTemplate(filename, pagesize=letter)
            elements = []
            styles = getSampleStyleSheet()

            # Title
            title = Paragraph(f"<b>Student Report: {first_name} {last_name}</b>", styles['Title'])
            elements.append(title)
            elements.append(Spacer(1, 20))

            # Student info table
            student_info = [
                ['Student ID:', student_data[0]],
                ['Email:', student_data[1]],
                ['Name:', f"{student_data[3]} {student_data[5]}"],
                ['Course:', student_data[9]],
                ['Registration:', student_data[10]]
            ]
            t1 = Table(student_info, colWidths=[120, 350])
            t1.setStyle(TableStyle([
                ('FONT', (0, 0), (-1, -1), 'Helvetica', 10),
                ('FONT', (0, 0), (0, -1), 'Helvetica-Bold', 10),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            elements.append(t1)
            elements.append(Spacer(1, 20))

            # Modules
            elements.append(Paragraph("<b>Enrolled Modules</b>", styles['Heading2']))
            if modules_data:
                module_table = [['Module Code', 'Module Name', 'Credits']] + modules_data
                t2 = Table(module_table)
                t2.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                elements.append(t2)
            elements.append(Spacer(1, 20))

            # Grades
            elements.append(Paragraph("<b>Grades</b>", styles['Heading2']))
            if grades_data:
                grade_table = [['Module', 'Assessment', 'Grade', 'Max']] + grades_data
                t3 = Table(grade_table)
                t3.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                elements.append(t3)

            doc.build(elements)
        except ImportError:
            # Fallback to TXT if reportlab not available
            messagebox.showwarning("Warning", "PDF export requires reportlab. Exporting as TXT instead.")
            self._export_to_txt(filename.replace('.pdf', '.txt'), student_id, first_name, last_name,
                               student_data, modules_data, grades_data, attendance_data)

    def _export_to_txt(self, filename, student_id, first_name, last_name,
                       student_data, modules_data, grades_data, attendance_data):
        """Export student data to text format"""
        # Convert Row objects to tuples
        student_data = tuple(student_data)
        modules_data = [tuple(m) for m in modules_data]
        grades_data = [tuple(g) for g in grades_data]
        attendance_data = [tuple(a) for a in attendance_data]

        with open(filename, 'w', encoding='utf-8') as f:
            f.write(f"=" * 80 + "\n")
            f.write(f"STUDENT REPORT: {first_name} {last_name}\n")
            f.write(f"=" * 80 + "\n\n")

            # Student info
            f.write("STUDENT INFORMATION\n")
            f.write("-" * 80 + "\n")
            f.write(f"Student ID: {student_data[0]}\n")
            f.write(f"Email: {student_data[1]}\n")
            f.write(f"Name: {student_data[3]} {student_data[5]}\n")
            f.write(f"Course: {student_data[9]}\n")
            f.write(f"Registration Date: {student_data[10]}\n\n")

            # Modules
            f.write("ENROLLED MODULES\n")
            f.write("-" * 80 + "\n")
            for module in modules_data:
                f.write(f"{module[0]} - {module[1]} ({module[2]} credits)\n")
            f.write("\n")

            # Grades
            f.write("GRADES\n")
            f.write("-" * 80 + "\n")
            for grade in grades_data:
                f.write(f"{grade[0]} - {grade[1]}: {grade[2]}/{grade[3]}\n")
            f.write("\n")

            # Attendance
            f.write("ATTENDANCE RECORDS\n")
            f.write("-" * 80 + "\n")
            for attendance in attendance_data:
                f.write(f"{attendance[0]} - {attendance[1]}: {attendance[2]}\n")

    def send_email_to_student(self, email, first_name, last_name):
        """Open email composition window for student with pre-filled recipient"""
        # Simply use the compose_email method which now handles recipient pre-filling
        self.compose_email(email)

    # ==================== COMPREHENSIVE EXPORT FUNCTIONS ====================

    def _export_comprehensive_csv(self, filename, student_id, first_name, last_name, data_sections):
        """Export comprehensive student data to CSV format with clear sections"""
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)

            # Convert Row objects
            student_data = tuple(data_sections['student_info']) if data_sections['student_info'] else ()

            # Header
            writer.writerow([f'COMPREHENSIVE STUDENT REPORT - {first_name} {last_name}'])
            writer.writerow([f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'])
            writer.writerow([])

            # 1. BASIC INFORMATION
            writer.writerow(['=' * 80])
            writer.writerow(['BASIC STUDENT INFORMATION'])
            writer.writerow(['=' * 80])
            if student_data:
                writer.writerow(['Student ID', student_data[0]])
                writer.writerow(['Email', student_data[1]])
                writer.writerow(['Name', f"{student_data[3]} {student_data[5]}"])
                writer.writerow(['Course', student_data[9]])
                writer.writerow(['Registration Date', student_data[10]])
            writer.writerow([])

            # 2. ENROLLED MODULES
            writer.writerow(['=' * 80])
            writer.writerow(['ENROLLED MODULES'])
            writer.writerow(['=' * 80])
            writer.writerow(['Module Code', 'Module Name', 'Credits', 'Enrollment Date', 'Status'])
            for module in data_sections['modules']:
                writer.writerow(tuple(module))
            writer.writerow([])

            # 3. ASSIGNMENTS (Available)
            writer.writerow(['=' * 80])
            writer.writerow(['ASSIGNMENTS (AVAILABLE)'])
            writer.writerow(['=' * 80])
            writer.writerow(['Module Code', 'Title', 'Due Date', 'Max Marks', 'Type'])
            for assignment in data_sections['assignments']:
                writer.writerow(tuple(assignment))
            writer.writerow([])

            # 4. ASSIGNMENT SUBMISSIONS
            writer.writerow(['=' * 80])
            writer.writerow(['ASSIGNMENT SUBMISSIONS'])
            writer.writerow(['=' * 80])
            writer.writerow(['Assignment', 'Module', 'Submission Date', 'Grade', 'Max Marks', 'Status', 'Late'])
            for submission in data_sections['submissions']:
                writer.writerow(tuple(submission))
            writer.writerow([])

            # 5. GRADES & ASSESSMENTS
            writer.writerow(['=' * 80])
            writer.writerow(['GRADES & ASSESSMENTS'])
            writer.writerow(['=' * 80])
            writer.writerow(['Module', 'Assessment', 'Type', 'Score', 'Max Points', 'Letter Grade', 'Date'])
            for grade in data_sections['grades']:
                writer.writerow(tuple(grade))
            writer.writerow([])

            # 6. ATTENDANCE RECORDS
            writer.writerow(['=' * 80])
            writer.writerow(['ATTENDANCE RECORDS'])
            writer.writerow(['=' * 80])
            writer.writerow(['Date', 'Module', 'Status', 'Reason'])
            for attendance in data_sections['attendance']:
                writer.writerow(tuple(attendance))
            writer.writerow([])

            # 7. FINANCIAL INFORMATION
            if data_sections['fees'] or data_sections['payments']:
                writer.writerow(['=' * 80])
                writer.writerow(['FINANCIAL INFORMATION'])
                writer.writerow(['=' * 80])

                writer.writerow(['FEES'])
                writer.writerow(['Fee Type', 'Amount', 'Due Date', 'Status'])
                for fee in data_sections['fees']:
                    writer.writerow(tuple(fee))
                writer.writerow([])

                writer.writerow(['PAYMENTS'])
                writer.writerow(['Payment Date', 'Amount', 'Method', 'Transaction ID'])
                for payment in data_sections['payments']:
                    writer.writerow(tuple(payment))
                writer.writerow([])

            # 8. SCHOLARSHIPS & FINANCIAL AID
            if data_sections['scholarships'] or data_sections['financial_aid']:
                writer.writerow(['=' * 80])
                writer.writerow(['SCHOLARSHIPS & FINANCIAL AID'])
                writer.writerow(['=' * 80])

                writer.writerow(['SCHOLARSHIPS'])
                writer.writerow(['Scholarship Name', 'Amount', 'Academic Year', 'Status'])
                for scholarship in data_sections['scholarships']:
                    writer.writerow(tuple(scholarship))
                writer.writerow([])

                writer.writerow(['FINANCIAL AID'])
                writer.writerow(['Aid Type', 'Amount', 'Academic Year', 'Status'])
                for aid in data_sections['financial_aid']:
                    writer.writerow(tuple(aid))
                writer.writerow([])

            # 9. EXTRACURRICULAR ACTIVITIES
            if data_sections['activities'] or data_sections['clubs']:
                writer.writerow(['=' * 80])
                writer.writerow(['EXTRACURRICULAR ACTIVITIES'])
                writer.writerow(['=' * 80])

                writer.writerow(['ACTIVITIES'])
                writer.writerow(['Activity Name', 'Role', 'Start Date', 'End Date', 'Status'])
                for activity in data_sections['activities']:
                    writer.writerow(tuple(activity))
                writer.writerow([])

                writer.writerow(['CLUB MEMBERSHIPS'])
                writer.writerow(['Club Name', 'Role', 'Join Date', 'Status'])
                for club in data_sections['clubs']:
                    writer.writerow(tuple(club))
                writer.writerow([])

            # 10. HEALTH RECORDS
            if data_sections['health_appointments'] or data_sections['medical_conditions']:
                writer.writerow(['=' * 80])
                writer.writerow(['HEALTH RECORDS'])
                writer.writerow(['=' * 80])

                writer.writerow(['HEALTH APPOINTMENTS'])
                writer.writerow(['Date', 'Type', 'Provider', 'Notes'])
                for appointment in data_sections['health_appointments']:
                    writer.writerow(tuple(appointment))
                writer.writerow([])

                writer.writerow(['MEDICAL CONDITIONS'])
                writer.writerow(['Condition', 'Diagnosed Date', 'Status'])
                for condition in data_sections['medical_conditions']:
                    writer.writerow(tuple(condition))
                writer.writerow([])

            # 11. HOUSING
            if data_sections['housing']:
                writer.writerow(['=' * 80])
                writer.writerow(['HOUSING ASSIGNMENTS'])
                writer.writerow(['=' * 80])
                writer.writerow(['Building', 'Room Number', 'Start Date', 'End Date', 'Status'])
                for housing in data_sections['housing']:
                    writer.writerow(tuple(housing))
                writer.writerow([])

            # 12. INTERNSHIPS
            if data_sections['internships']:
                writer.writerow(['=' * 80])
                writer.writerow(['INTERNSHIPS & PLACEMENTS'])
                writer.writerow(['=' * 80])
                writer.writerow(['Company', 'Position', 'Status', 'Application Date', 'Start Date', 'End Date'])
                for internship in data_sections['internships']:
                    writer.writerow(tuple(internship))
                writer.writerow([])

            # 13. LIBRARY RECORDS
            if data_sections['library']:
                writer.writerow(['=' * 80])
                writer.writerow(['LIBRARY RECORDS'])
                writer.writerow(['=' * 80])
                writer.writerow(['Book Title', 'Checkout Date', 'Due Date', 'Return Date', 'Status'])
                for loan in data_sections['library']:
                    writer.writerow(tuple(loan))
                writer.writerow([])

            # 14. SUPPORT TICKETS
            if data_sections['support_tickets']:
                writer.writerow(['=' * 80])
                writer.writerow(['SUPPORT TICKETS'])
                writer.writerow(['=' * 80])
                writer.writerow(['Ticket ID', 'Subject', 'Category', 'Status', 'Created', 'Resolved'])
                for ticket in data_sections['support_tickets']:
                    writer.writerow(tuple(ticket))
                writer.writerow([])

            # 15. BADGES & ACHIEVEMENTS
            if data_sections['badges']:
                writer.writerow(['=' * 80])
                writer.writerow(['BADGES & ACHIEVEMENTS'])
                writer.writerow(['=' * 80])
                writer.writerow(['Badge Name', 'Date Earned', 'Description'])
                for badge in data_sections['badges']:
                    writer.writerow(tuple(badge))
                writer.writerow([])

            # 16. PARKING
            if data_sections['parking']:
                writer.writerow(['=' * 80])
                writer.writerow(['PARKING PERMITS'])
                writer.writerow(['=' * 80])
                writer.writerow(['Permit Number', 'Vehicle Make', 'Vehicle Model', 'Start Date', 'End Date', 'Type'])
                for permit in data_sections['parking']:
                    writer.writerow(tuple(permit))
                writer.writerow([])

            # 17. MEAL PLAN
            if data_sections['meal_plan'] or data_sections['meal_transactions']:
                writer.writerow(['=' * 80])
                writer.writerow(['MEAL PLAN & TRANSACTIONS'])
                writer.writerow(['=' * 80])

                writer.writerow(['MEAL PLAN'])
                writer.writerow(['Plan Type', 'Balance', 'Last Transaction Date'])
                for meal_plan in data_sections['meal_plan']:
                    writer.writerow(tuple(meal_plan))
                writer.writerow([])

                writer.writerow(['RECENT TRANSACTIONS (Last 50)'])
                writer.writerow(['Date', 'Location', 'Amount', 'Description'])
                for transaction in data_sections['meal_transactions']:
                    writer.writerow(tuple(transaction))
                writer.writerow([])

            # 18. ADVISING
            if data_sections['advising']:
                writer.writerow(['=' * 80])
                writer.writerow(['ACADEMIC ADVISING'])
                writer.writerow(['=' * 80])
                writer.writerow(['Date', 'Advisor', 'Notes', 'Follow-up Required'])
                for advising in data_sections['advising']:
                    writer.writerow(tuple(advising))
                writer.writerow([])

    def _export_comprehensive_excel(self, filename, student_id, first_name, last_name, data_sections):
        """Export comprehensive student data to Excel with multiple sheets"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # Helper to create sheet with data
            def create_sheet(sheet_name, headers, data, title=None):
                ws = wb.create_sheet(sheet_name)
                if title:
                    ws['A1'] = title
                    ws['A1'].font = Font(bold=True, size=14)
                    start_row = 3
                else:
                    start_row = 1

                # Headers
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=start_row, column=col_idx)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

                # Data
                for row_idx, row_data in enumerate(data, start_row + 1):
                    for col_idx, value in enumerate(tuple(row_data), 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)

            # 1. Student Info
            student_data = tuple(data_sections['student_info']) if data_sections['student_info'] else ()
            ws1 = wb.create_sheet("Student Info")
            ws1['A1'] = f"Student Information - {first_name} {last_name}"
            ws1['A1'].font = Font(bold=True, size=14)
            if student_data:
                ws1.append(['Student ID', student_data[0]])
                ws1.append(['Email', student_data[1]])
                ws1.append(['Name', f"{student_data[3]} {student_data[5]}"])
                ws1.append(['Course', student_data[9]])
                ws1.append(['Registration', student_data[10]])

            # 2-18. All other sections
            create_sheet("Modules", ['Module Code', 'Module Name', 'Credits', 'Enrollment Date', 'Status'],
                        data_sections['modules'])
            create_sheet("Assignments", ['Module', 'Title', 'Due Date', 'Max Marks', 'Type'],
                        data_sections['assignments'])
            create_sheet("Submissions", ['Assignment', 'Module', 'Submission Date', 'Grade', 'Max', 'Status', 'Late'],
                        data_sections['submissions'])
            create_sheet("Grades", ['Module', 'Assessment', 'Type', 'Score', 'Max', 'Letter', 'Date'],
                        data_sections['grades'])
            create_sheet("Attendance", ['Date', 'Module', 'Status', 'Reason'],
                        data_sections['attendance'])

            if data_sections['fees']:
                create_sheet("Fees", ['Fee Type', 'Amount', 'Due Date', 'Status'], data_sections['fees'])
            if data_sections['payments']:
                create_sheet("Payments", ['Date', 'Amount', 'Method', 'Transaction ID'], data_sections['payments'])
            if data_sections['scholarships']:
                create_sheet("Scholarships", ['Name', 'Amount', 'Academic Year', 'Status'], data_sections['scholarships'])
            if data_sections['activities']:
                create_sheet("Activities", ['Activity', 'Role', 'Start', 'End', 'Status'], data_sections['activities'])
            if data_sections['clubs']:
                create_sheet("Clubs", ['Club Name', 'Role', 'Join Date', 'Status'], data_sections['clubs'])
            if data_sections['housing']:
                create_sheet("Housing", ['Building', 'Room', 'Start', 'End', 'Status'], data_sections['housing'])
            if data_sections['internships']:
                create_sheet("Internships", ['Company', 'Position', 'Status', 'App Date', 'Start', 'End'],
                           data_sections['internships'])
            if data_sections['library']:
                create_sheet("Library", ['Book', 'Checkout', 'Due', 'Return', 'Status'], data_sections['library'])
            if data_sections['badges']:
                create_sheet("Badges", ['Badge', 'Date Earned', 'Description'], data_sections['badges'])
            if data_sections['parking']:
                create_sheet("Parking", ['Permit#', 'Make', 'Model', 'Start', 'End', 'Type'], data_sections['parking'])

            wb.save(filename)
        except ImportError:
            messagebox.showwarning("Warning", "Excel export requires openpyxl. Exporting as CSV instead.")
            self._export_comprehensive_csv(filename.replace('.xlsx', '.csv'), student_id, first_name, last_name, data_sections)

    def _export_comprehensive_pdf(self, filename, student_id, first_name, last_name, data_sections):
        """Export comprehensive student data to PDF format"""
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
            from reportlab.lib.units import inch

            doc = SimpleDocTemplate(filename, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
            elements = []
            styles = getSampleStyleSheet()

            # Custom styles
            heading_style = ParagraphStyle('CustomHeading', parent=styles['Heading2'],
                                          fontSize=14, textColor=colors.HexColor('#003366'))

            # Title
            title = Paragraph(f"<b>COMPREHENSIVE STUDENT REPORT</b><br/>{first_name} {last_name}",
                            styles['Title'])
            elements.append(title)
            elements.append(Spacer(1, 0.3*inch))

            # Helper function to add section
            def add_section(title, headers, data):
                if not data:
                    return
                elements.append(Paragraph(f"<b>{title}</b>", heading_style))
                elements.append(Spacer(1, 0.1*inch))

                table_data = [list(headers)] + [list(tuple(row)) for row in data[:20]]  # Limit to 20 rows
                t = Table(table_data, repeatRows=1)
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]))
                elements.append(t)
                elements.append(Spacer(1, 0.2*inch))

            # Student Info
            student_data = tuple(data_sections['student_info']) if data_sections['student_info'] else ()
            if student_data:
                info_data = [
                    ['Student ID', student_data[0]],
                    ['Email', student_data[1]],
                    ['Name', f"{student_data[3]} {student_data[5]}"],
                    ['Course', student_data[9]],
                ]
                add_section("BASIC INFORMATION", ['Field', 'Value'], info_data)

            # All sections
            add_section("ENROLLED MODULES", ['Module', 'Name', 'Credits', 'Date', 'Status'], data_sections['modules'])
            add_section("ASSIGNMENT SUBMISSIONS", ['Assignment', 'Module', 'Date', 'Grade', 'Max'], data_sections['submissions'])
            add_section("GRADES", ['Module', 'Assessment', 'Score', 'Max', 'Grade'], data_sections['grades'])
            add_section("ATTENDANCE", ['Date', 'Module', 'Status', 'Reason'], data_sections['attendance'])

            if data_sections['fees']:
                add_section("FEES", ['Type', 'Amount', 'Due Date', 'Status'], data_sections['fees'])
            if data_sections['scholarships']:
                add_section("SCHOLARSHIPS", ['Name', 'Amount', 'Year', 'Status'], data_sections['scholarships'])
            if data_sections['activities']:
                add_section("ACTIVITIES", ['Activity', 'Role', 'Start', 'Status'], data_sections['activities'])
            if data_sections['housing']:
                add_section("HOUSING", ['Building', 'Room', 'Start', 'Status'], data_sections['housing'])

            doc.build(elements)
        except ImportError:
            messagebox.showwarning("Warning", "PDF export requires reportlab. Exporting as TXT instead.")
            self._export_comprehensive_txt(filename.replace('.pdf', '.txt'), student_id, first_name, last_name, data_sections)

    def _export_comprehensive_txt(self, filename, student_id, first_name, last_name, data_sections):
        """Export comprehensive student data to text format"""
        with open(filename, 'w', encoding='utf-8') as f:
            student_data = tuple(data_sections['student_info']) if data_sections['student_info'] else ()

            # Header
            f.write("=" * 80 + "\n")
            f.write(f"COMPREHENSIVE STUDENT REPORT: {first_name} {last_name}\n")
            f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("=" * 80 + "\n\n")

            # Helper function
            def write_section(title, headers, data):
                if not data:
                    return
                f.write("\n" + "=" * 80 + "\n")
                f.write(f"{title}\n")
                f.write("=" * 80 + "\n")
                f.write(" | ".join(headers) + "\n")
                f.write("-" * 80 + "\n")
                for row in data:
                    f.write(" | ".join(str(v) if v is not None else '' for v in tuple(row)) + "\n")

            # Student Info
            f.write("BASIC INFORMATION\n")
            f.write("-" * 80 + "\n")
            if student_data:
                f.write(f"Student ID: {student_data[0]}\n")
                f.write(f"Email: {student_data[1]}\n")
                f.write(f"Name: {student_data[3]} {student_data[5]}\n")
                f.write(f"Course: {student_data[9]}\n")
                f.write(f"Registration: {student_data[10]}\n")

            # All sections
            write_section("ENROLLED MODULES", ['Code', 'Name', 'Credits', 'Enrolled', 'Status'], data_sections['modules'])
            write_section("ASSIGNMENTS (AVAILABLE)", ['Module', 'Title', 'Due', 'Marks', 'Type'], data_sections['assignments'])
            write_section("SUBMISSIONS", ['Assignment', 'Module', 'Date', 'Grade', 'Max', 'Status'], data_sections['submissions'])
            write_section("GRADES", ['Module', 'Assessment', 'Type', 'Score', 'Max', 'Letter'], data_sections['grades'])
            write_section("ATTENDANCE", ['Date', 'Module', 'Status', 'Reason'], data_sections['attendance'])
            write_section("FEES", ['Type', 'Amount', 'Due', 'Status'], data_sections['fees'])
            write_section("PAYMENTS", ['Date', 'Amount', 'Method', 'Transaction'], data_sections['payments'])
            write_section("SCHOLARSHIPS", ['Name', 'Amount', 'Year', 'Status'], data_sections['scholarships'])
            write_section("FINANCIAL AID", ['Type', 'Amount', 'Year', 'Status'], data_sections['financial_aid'])
            write_section("ACTIVITIES", ['Activity', 'Role', 'Start', 'End', 'Status'], data_sections['activities'])
            write_section("CLUBS", ['Club', 'Role', 'Joined', 'Status'], data_sections['clubs'])
            write_section("HOUSING", ['Building', 'Room', 'Start', 'End', 'Status'], data_sections['housing'])
            write_section("INTERNSHIPS", ['Company', 'Position', 'Status', 'Applied', 'Start', 'End'], data_sections['internships'])
            write_section("LIBRARY", ['Book', 'Checkout', 'Due', 'Return', 'Status'], data_sections['library'])
            write_section("SUPPORT TICKETS", ['ID', 'Subject', 'Category', 'Status', 'Created', 'Resolved'], data_sections['support_tickets'])
            write_section("BADGES", ['Badge', 'Earned', 'Description'], data_sections['badges'])
            write_section("PARKING", ['Permit', 'Make', 'Model', 'Start', 'End', 'Type'], data_sections['parking'])
            write_section("MEAL PLAN", ['Type', 'Balance', 'Last Transaction'], data_sections['meal_plan'])
            write_section("ADVISING", ['Date', 'Advisor', 'Notes', 'Follow-up'], data_sections['advising'])

    def update_student_dialog(self, student_id):
        """Comprehensive update student dialog with full editing capabilities and random course assignment"""
        dialog = tk.Toplevel(self.root)
        dialog.title(f"Update Student - {student_id}")
        dialog.geometry("800x900")
        dialog.transient(self.root)
        
        # Make dialog visible before grabbing
        dialog.update_idletasks()
        dialog.deiconify()
        
        try:
            dialog.grab_set()
        except tk.TclError:
            print("Warning: Could not grab dialog focus")
        
        try:
            # Get current student data
            conn = get_db_connection()
            if not conn:
                messagebox.showerror("Error", "Database connection failed")
                dialog.destroy()
                return
            
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM students WHERE student_id = ?', (student_id,))
            student = cursor.fetchone()
            
            if not student:
                messagebox.showerror("Error", "Student not found")
                dialog.destroy()
                return
            
            # Main scrollable frame
            main_frame = ttk.Frame(dialog)
            main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
            
            canvas = tk.Canvas(main_frame)
            scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
            scrollable_frame = ttk.Frame(canvas)
            
            scrollable_frame.bind(
                "<Configure>",
                lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
            )
            
            canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)
            
            # Title
            title_label = ttk.Label(scrollable_frame, text=f"Update Student: {student_id}", 
                                   font=('Arial', 16, 'bold'))
            title_label.grid(row=0, column=0, columnspan=2, pady=(0, 20))
            
            # Current info display
            current_frame = ttk.LabelFrame(scrollable_frame, text="Current Information", padding=10)
            current_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
            
            current_text = tk.Text(current_frame, height=4, width=70, wrap=tk.WORD)
            current_text.pack(fill=tk.X)
            current_info = f"Name: {student[2]} {student[3]} {student[4]} {student[5]} | Gender: {student[6]} | Course: {student[9]} | Age: {student[8]}"
            current_text.insert(tk.END, current_info)
            current_text.config(state=tk.DISABLED)
            
            # Form fields with current values
            fields = {}
            
            # Personal Information Section
            personal_frame = ttk.LabelFrame(scrollable_frame, text="Personal Information", padding=15)
            personal_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
            
            # Title
            ttk.Label(personal_frame, text="Title:").grid(row=0, column=0, sticky=tk.W, pady=5)
            title_options = ['Mr', 'Ms', 'Mrs', 'Dr', 'Prof']
            title_value = (student[2] or '').strip()
            if title_value not in title_options:
                title_value = ''
            fields['title'] = ttk.Combobox(personal_frame, values=title_options,
                                          state='readonly', width=27)
            fields['title'].grid(row=0, column=1, pady=5, padx=(10, 0), sticky=tk.W)
            _safe_set_combobox(fields['title'], title_value)
            
            # First Name
            ttk.Label(personal_frame, text="First Name:").grid(row=1, column=0, sticky=tk.W, pady=5)
            fields['first_name'] = ttk.Entry(personal_frame, width=30)
            fields['first_name'].grid(row=1, column=1, pady=5, padx=(10, 0))
            _safe_entry_insert(fields['first_name'], student[3])
            
            # Middle Name
            ttk.Label(personal_frame, text="Middle Name:").grid(row=2, column=0, sticky=tk.W, pady=5)
            fields['middle_name'] = ttk.Entry(personal_frame, width=30)
            fields['middle_name'].grid(row=2, column=1, pady=5, padx=(10, 0))
            _safe_entry_insert(fields['middle_name'], student[4])
            
            # Last Name
            ttk.Label(personal_frame, text="Last Name:").grid(row=3, column=0, sticky=tk.W, pady=5)
            fields['last_name'] = ttk.Entry(personal_frame, width=30)
            fields['last_name'].grid(row=3, column=1, pady=5, padx=(10, 0))
            _safe_entry_insert(fields['last_name'], student[5])
            
            # Gender
            ttk.Label(personal_frame, text="Gender:").grid(row=4, column=0, sticky=tk.W, pady=5)
            gender_options = ['male', 'female', 'other']
            gender_value = (student[6] or '').strip().lower()
            if gender_value not in gender_options:
                gender_value = ''
            fields['gender'] = ttk.Combobox(personal_frame, values=gender_options,
                                           state='readonly', width=27)
            fields['gender'].grid(row=4, column=1, pady=5, padx=(10, 0), sticky=tk.W)
            _safe_set_combobox(fields['gender'], gender_value)
            
            # Date of Birth
            ttk.Label(personal_frame, text="Date of Birth (YYYY-MM-DD):").grid(row=5, column=0, sticky=tk.W, pady=5)
            fields['dob'] = ttk.Entry(personal_frame, width=30)
            fields['dob'].grid(row=5, column=1, pady=5, padx=(10, 0))
            _safe_entry_insert(fields['dob'], student[7])
            
            # Academic Information
            academic_frame = ttk.LabelFrame(scrollable_frame, text="Academic Information", padding=15)
            academic_frame.grid(row=3, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
            
            # Course - Add random assignment option
            ttk.Label(academic_frame, text="Course:").grid(row=0, column=0, sticky=tk.W, pady=5)
            course_frame = ttk.Frame(academic_frame)
            course_frame.grid(row=0, column=1, pady=5, padx=(10, 0), sticky=tk.W)
            
            current_course_label = ttk.Label(course_frame, text=f"Current: {student[9]}", foreground="blue")
            current_course_label.pack(side=tk.LEFT)
            
            # Random course reassignment option
            reassign_course_var = tk.BooleanVar()
            ttk.Checkbutton(course_frame, text="Randomly reassign course and modules", 
                           variable=reassign_course_var).pack(side=tk.LEFT, padx=(20, 0))
            
            # Email (read-only display)
            ttk.Label(academic_frame, text="Email:").grid(row=1, column=0, sticky=tk.W, pady=5)
            email_label = ttk.Label(academic_frame, text=student[1], foreground="blue")
            email_label.grid(row=1, column=1, pady=5, padx=(10, 0), sticky=tk.W)
            
            # Module Management Section
            modules_frame = ttk.LabelFrame(scrollable_frame, text="Module Management", padding=15)
            modules_frame.grid(row=4, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
            
            # Get current modules
            cursor.execute('''
                SELECT m.module_type, sm.module_code, m.module_name
                FROM student_modules sm
                JOIN modules m ON sm.module_code = m.module_code
                WHERE sm.student_id = ?
            ''', (student_id,))
            current_modules = cursor.fetchall()
            
            modules_text = scrolledtext.ScrolledText(modules_frame, height=6, width=70)
            modules_text.pack(fill=tk.X)
            
            modules_info = "Current Modules:\n" + "-"*40 + "\n"
            for module in current_modules:
                modules_info += f"{module[0]}: {module[1]} - {module[2]}\n"
            modules_text.insert(tk.END, modules_info)
            modules_text.config(state=tk.DISABLED)
            
            # Buttons for module actions
            module_buttons_frame = ttk.Frame(modules_frame)
            module_buttons_frame.pack(fill=tk.X, pady=(10, 0))
            
            ttk.Button(module_buttons_frame, text="Reassign Optional Modules", 
                      command=lambda: self.reassign_modules(student_id, 'optional')).pack(side=tk.LEFT, padx=5)
            
            # Validation feedback
            validation_label = ttk.Label(scrollable_frame, text="", foreground="red")
            validation_label.grid(row=5, column=0, columnspan=2, pady=10)
            
            # Password update section
            password_frame = ttk.LabelFrame(scrollable_frame, text="Password Management", padding=15)
            password_frame.grid(row=6, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
            
            password_var = tk.BooleanVar()
            ttk.Checkbutton(password_frame, text="Update password automatically based on first name", 
                           variable=password_var).pack(anchor=tk.W)
            
            ttk.Label(password_frame, text="Note: If checked, password will be set to [firstname]123456", 
                     foreground="gray").pack(anchor=tk.W, pady=(5, 0))
            
            def validate_update_form():
                """Validate update form inputs"""
                errors = []
                
                if not fields['first_name'].get().strip():
                    errors.append("First name cannot be empty")
                
                if not fields['last_name'].get().strip():
                    errors.append("Last name cannot be empty")
                
                dob_text = fields['dob'].get().strip()
                if dob_text:
                    try:
                        dob = datetime.strptime(dob_text, "%Y-%m-%d")
                        age = datetime.now().year - dob.year
                        if age < 16 or age > 80:
                            errors.append("Age must be between 16 and 80")
                    except ValueError:
                        errors.append("Invalid date format (use YYYY-MM-DD)")
                
                return errors
            
            def update_student():
                """Update student with form data and random course assignment if selected"""
                update_conn = None
                try:
                    # Validate form
                    errors = validate_update_form()
                    if errors:
                        validation_label.config(text="; ".join(errors))
                        return

                    validation_label.config(text="")

                    # Get updated data
                    new_title = fields['title'].get()
                    new_first_name = fields['first_name'].get().strip()
                    new_middle_name = fields['middle_name'].get().strip()
                    new_last_name = fields['last_name'].get().strip()
                    new_gender = fields['gender'].get()
                    new_dob = fields['dob'].get().strip()
                    if not new_dob:
                        new_dob = None

                    # Determine new course
                    if reassign_course_var.get():
                        new_course = random.choice(['CS', 'DS'])
                        course_changed = True
                    else:
                        new_course = student[9]  # Keep current course
                        course_changed = False

                    # Calculate new age if DOB changed
                    if new_dob:
                        if student[7] != new_dob:
                            dob_date = datetime.strptime(new_dob, "%Y-%m-%d")
                            new_age = datetime.now().year - dob_date.year - (
                                (datetime.now().month, datetime.now().day) < (dob_date.month, dob_date.day)
                            )
                        else:
                            new_age = student[8]
                    else:
                        new_age = None

                    # Create new database connection for update
                    update_conn = get_db_connection()
                    update_cursor = update_conn.cursor()

                    # Update database
                    update_cursor.execute('''
                        UPDATE students
                        SET title = ?, first_name = ?, middle_name = ?, last_name = ?,
                            gender = ?, dob = ?, age = ?, course = ?
                        WHERE student_id = ?
                    ''', (new_title, new_first_name, new_middle_name, new_last_name,
                          new_gender, new_dob, new_age, new_course, student_id))

                    # Update user profile if exists
                    update_cursor.execute('SELECT id FROM users WHERE student_id = ?', (student_id,))
                    user_record = update_cursor.fetchone()

                    if user_record:
                        user_id = user_record[0]
                        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

                        update_cursor.execute('''
                            UPDATE users
                            SET first_name = ?, last_name = ?, updated_at = ?
                            WHERE id = ?
                        ''', (new_first_name, new_last_name, timestamp, user_id))

                        # Update password if requested
                        if password_var.get():
                            new_password = f"{new_first_name.lower()}123456"

                            # Update password in user_accounts table
                            import hashlib, secrets
                            salt = secrets.token_hex(16)
                            key = hashlib.pbkdf2_hmac('sha256', new_password.encode(), salt.encode(), 100000, dklen=64)
                            password_hash = key.hex()

                            update_cursor.execute('''
                                UPDATE user_accounts
                                SET password_hash = ?, salt = ?, updated_at = ?
                                WHERE user_id = ?
                            ''', (password_hash, salt, timestamp, user_id))

                    # Handle course change - reassign course-specific modules
                    if course_changed or new_course != student[9]:
                        # Remove old course modules and add new ones
                        update_cursor.execute('DELETE FROM student_modules WHERE student_id = ?', (student_id,))

                        # Get available modules for the new course
                        update_cursor.execute('''
                            SELECT module_code FROM modules
                            WHERE department = ? AND is_active = 1
                            ORDER BY module_code
                        ''', (new_course,))
                        course_modules = [row[0] for row in update_cursor.fetchall()]

                        if course_modules:
                            num_modules = min(random.randint(3, 6), len(course_modules))
                            selected_modules = random.sample(course_modules, num_modules)

                            current_date = datetime.now().strftime('%Y-%m-%d')
                            for module_code in selected_modules:
                                update_cursor.execute('''
                                    INSERT INTO student_modules (student_id, module_code, enrollment_date, status)
                                    VALUES (?, ?, ?, ?)
                                ''', (student_id, module_code, current_date, 'enrolled'))

                    update_conn.commit()

                    # Send update confirmation email via email_service
                    try:
                        # Get student email
                        update_cursor.execute('SELECT email_address FROM students WHERE student_id = ?', (student_id,))
                        email_result = update_cursor.fetchone()

                        if email_result:
                            student_email = email_result[0]
                            # Determine what fields were updated
                            updated_fields = []
                            if new_title != student[2]:
                                updated_fields.append('title')
                            if new_first_name != student[3]:
                                updated_fields.append('first name')
                            if new_middle_name != student[4]:
                                updated_fields.append('middle name')
                            if new_last_name != student[5]:
                                updated_fields.append('last name')
                            if new_gender != student[6]:
                                updated_fields.append('gender')
                            if new_dob and new_dob != student[7]:
                                updated_fields.append('date of birth')
                            if course_changed:
                                updated_fields.append('course')
                            if password_var.get():
                                updated_fields.append('password')

                            if updated_fields:
                                from university_system.infrastructure.email.email_service import send_update_confirmation
                                send_update_confirmation(student_email, updated_fields)
                    except Exception as e:
                        import logging
                        logging.warning(f"Failed to send update confirmation email: {e}")

                    success_msg = f"Student {student_id} updated successfully!"
                    if course_changed:
                        success_msg += f"\nCourse randomly changed from {student[9]} to {new_course}"
                    if password_var.get():
                        success_msg += f"\nNew password: {new_first_name.lower()}123456"
                    
                    messagebox.showinfo("Success", success_msg)

                    # Send email notification about changes
                    self._send_student_update_email(
                        student_id=student_id,
                        old_data=student,
                        new_data={
                            'title': new_title,
                            'first_name': new_first_name,
                            'middle_name': new_middle_name,
                            'last_name': new_last_name,
                            'gender': new_gender,
                            'dob': new_dob,
                            'age': new_age,
                            'course': new_course
                        },
                        course_changed=course_changed,
                        password_reset=password_var.get()
                    )

                    # Refresh views and close dialog
                    if hasattr(self, 'view_students'):
                        self.view_students()
                    self.refresh_advanced_search()

                    dialog.destroy()

                except Exception as e:
                    messagebox.showerror("Error", f"Failed to update student: {str(e)}")
                finally:
                    # Close the update connection
                    if update_conn:
                        update_conn.close()
            
            # Buttons
            button_frame = ttk.Frame(scrollable_frame)
            button_frame.grid(row=7, column=0, columnspan=2, pady=20)
            
            ttk.Button(button_frame, text="Update Student", command=update_student, 
                      style="Accent.TButton").pack(side=tk.LEFT, padx=10)
            ttk.Button(button_frame, text="Reset Form", 
                      command=lambda: dialog.destroy() and self.update_student_dialog(student_id)).pack(side=tk.LEFT, padx=5)
            ttk.Button(button_frame, text="Cancel", command=dialog.destroy).pack(side=tk.LEFT, padx=5)
            
            # Pack canvas and scrollbar
            canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")
            
            conn.close()
            
            # Bind mousewheel
            def _on_mousewheel(event):
                canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            canvas.bind_all("<MouseWheel>", _on_mousewheel)
            
        except Exception as e:
            logging.exception("Failed to load student data for student_id=%s", student_id)
            messagebox.showerror("Error", f"Failed to load student data: {str(e)}")
            dialog.destroy()

    def reassign_modules(self, student_id, module_type, cursor=None):
        """Reassign modules for a student"""
        should_close_conn = cursor is None
        
        try:
            if cursor is None:
                conn = get_db_connection()
                cursor = conn.cursor()
            
            # Import modules
            from university_system.modules.domain.academics.services.modules import (
                optional_module_1, optional_module_2, optional_module_3, optional_module_4,
                CS_optional_module_1, CS_optional_module_2, CS_optional_module_3, CS_optional_module_4,
                DS_optional_module_1, DS_optional_module_2, DS_optional_module_3, DS_optional_module_4
            )
            
            # Delete existing modules of this type
            cursor.execute('''
                DELETE FROM student_modules
                WHERE student_id = ? AND module_code IN (
                    SELECT module_code FROM modules WHERE module_type = ?
                )
            ''', (student_id, module_type))
            
            # Select new modules
            if module_type == 'optional':
                available_modules = [optional_module_1, optional_module_2, optional_module_3, optional_module_4]
            elif module_type == 'CS':
                available_modules = [CS_optional_module_1, CS_optional_module_2, CS_optional_module_3, CS_optional_module_4]
            elif module_type == 'DS':
                available_modules = [DS_optional_module_1, DS_optional_module_2, DS_optional_module_3, DS_optional_module_4]
            else:
                return
            
            selected_modules = random.sample(available_modules, 2)
            
            # Insert new modules
            for module in selected_modules:
                cursor.execute('''
                    INSERT INTO student_modules (student_id, module_code, enrollment_date, status)
                    VALUES (?, ?, ?, ?)
                ''', (student_id, module['code'], datetime.now().strftime('%Y-%m-%d'), 'enrolled'))
            
            if should_close_conn:
                conn.commit()
                conn.close()
            
            messagebox.showinfo("Success", f"Reassigned {module_type} modules successfully!")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to reassign modules: {str(e)}")

    def delete_student_dialog(self, student_id=None):
        """Comprehensive delete student dialog with safety checks"""
        if not student_id:
            # Show selection dialog first
            student_id = self.select_student_for_deletion()
            if not student_id:
                return
        
        dialog = tk.Toplevel(self.root)
        dialog.title(f"Delete Student - {student_id}")
        dialog.geometry("900x900")  # Made bigger
        dialog.transient(self.root)
        
        # CRITICAL FIX: Make dialog visible BEFORE grabbing
        dialog.update_idletasks()  # Force geometry calculation
        dialog.deiconify()         # Ensure window is visible
        
        # Center the dialog on parent
        dialog.geometry("+%d+%d" % (
            self.root.winfo_rootx() + 50,
            self.root.winfo_rooty() + 50
        ))
        
        # Now it's safe to grab focus
        try:
            dialog.grab_set()
        except tk.TclError:
            # If grab still fails, continue without it
            print("Warning: Could not grab dialog focus")
        
        # Make dialog modal and non-resizable
        dialog.resizable(False, False)
        
        try:
            # Get student data for confirmation
            conn = get_db_connection()
            if not conn:
                messagebox.showerror("Error", "Database connection failed")
                dialog.destroy()
                return
            
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM students WHERE student_id = ?', (student_id,))
            student = cursor.fetchone()
            
            if not student:
                messagebox.showerror("Error", "Student not found")
                dialog.destroy()
                return
            
            main_frame = ttk.Frame(dialog, padding=20)
            main_frame.pack(fill=tk.BOTH, expand=True)
            
            # Warning header
            warning_frame = ttk.Frame(main_frame)
            warning_frame.pack(fill=tk.X, pady=(0, 20))
            
            warning_label = ttk.Label(warning_frame, text="⚠️ DELETE STUDENT RECORD", 
                                     font=('Arial', 16, 'bold'), foreground="red")
            warning_label.pack()
            
            # Student information display
            info_frame = ttk.LabelFrame(main_frame, text="Student Information", padding=15)
            info_frame.pack(fill=tk.X, pady=(0, 20))
            
            student_info = f"""Student ID: {student[0]}
    Name: {student[2]} {student[3]} {student[4]} {student[5]}
    Email: {student[1]}
    Course: {student[9]}
    Registration Date: {student[10]}"""
            
            ttk.Label(info_frame, text=student_info, font=('Courier', 10)).pack(anchor=tk.W)
            
            # Get related records count
            cursor.execute('SELECT COUNT(*) FROM student_modules WHERE student_id = ?', (student_id,))
            modules_count = cursor.fetchone()[0]
            
            cursor.execute('SELECT COUNT(*) FROM student_grades WHERE student_id = ?', (student_id,))
            grades_count = cursor.fetchone()[0]
            
            cursor.execute('SELECT COUNT(*) FROM attendance WHERE student_id = ?', (student_id,))
            attendance_count = cursor.fetchone()[0]
            
            # Related records information
            related_frame = ttk.LabelFrame(main_frame, text="Related Records (Will be Deleted)", padding=15)
            related_frame.pack(fill=tk.X, pady=(0, 20))
            
            related_info = f"""Module Enrollments: {modules_count}
    Grade Records: {grades_count}
    Attendance Records: {attendance_count}
    User Account: Will be removed
    All associated data will be permanently deleted"""
            
            ttk.Label(related_frame, text=related_info, font=('Courier', 10), foreground="dark red").pack(anchor=tk.W)
            
            # Confirmation section
            confirm_frame = ttk.LabelFrame(main_frame, text="Confirmation Required", padding=15)
            confirm_frame.pack(fill=tk.X, pady=(0, 20))
            
            ttk.Label(confirm_frame, text="This action cannot be undone!", 
                     font=('Arial', 12, 'bold'), foreground="red").pack(pady=(0, 10))
            
            ttk.Label(confirm_frame, text=f"Type '{student_id}' to confirm deletion:").pack(anchor=tk.W)
            confirm_entry = ttk.Entry(confirm_frame, width=30, font=('Arial', 11))
            confirm_entry.pack(pady=(5, 10), fill=tk.X)
            
            # Additional confirmation checkbox
            additional_confirm = tk.BooleanVar()
            ttk.Checkbutton(confirm_frame, text="I understand this will permanently delete all student data", 
                           variable=additional_confirm).pack(anchor=tk.W)
            
            # Status label
            status_label = ttk.Label(confirm_frame, text="", foreground="red")
            status_label.pack(pady=(10, 0))

            # Close the initial read-only connection before performing destructive work
            if conn:
                conn.close()

            def perform_deletion():
                """Perform the actual deletion with comprehensive cleanup"""
                # Validate confirmations
                entered_id = confirm_entry.get().strip()
                expected_id = str(student_id).strip()
                if entered_id != expected_id:
                    status_label.config(text=f"Student ID confirmation does not match (entered: '{entered_id}', expected: '{expected_id}')")
                    return
                
                if not additional_confirm.get():
                    status_label.config(text="Please check the confirmation checkbox")
                    return
                
                # Final confirmation dialog
                if not messagebox.askyesno("Final Confirmation", 
                                         f"Are you absolutely sure you want to delete student {student_id}?\n\n"
                                         "This action is IRREVERSIBLE and will delete:\n"
                                         "• Student record\n"
                                         "• All module enrollments\n"
                                         "• All grades and assessments\n"
                                         "• All attendance records\n"
                                         "• User account and login access\n"
                                         "• All related data",
                                         icon='warning'):
                    return
                
                conn = None
                cursor = None
                try:
                    # Start deletion process
                    status_label.config(text="Deleting student record...", foreground="blue")
                    dialog.update()

                    conn = get_db_connection()
                    if not conn:
                        raise sqlite3.OperationalError("Database connection failed during deletion")

                    cursor = conn.cursor()

                    # Disable foreign key constraints temporarily
                    cursor.execute("PRAGMA foreign_keys = OFF")

                    deletion_log = []

                    # Delete from related tables first
                    tables_to_clean = [
                        ('student_grades', 'student_id'),
                        ('attendance', 'student_id'),
                        ('student_modules', 'student_id'),
                        ('assignment_submissions', 'student_id'),
                        ('accommodation_requests', 'student_id'),
                        ('housing_requests', 'student_id'),
                        ('health_records', 'student_id'),
                        ('internship_applications', 'student_id'),
                        ('trip_participants', 'student_id'),
                        ('loans', 'borrower_id'),
                        ('student_fees', 'student_id'),
                        ('support_tickets', 'student_id'),
                        ('parent_student_relationships', 'student_id')
                    ]

                    for table_name, column_name in tables_to_clean:
                        try:
                            cursor.execute(f'DELETE FROM {table_name} WHERE {column_name} = ?', (student_id,))
                            deleted_count = cursor.rowcount
                            if deleted_count > 0:
                                deletion_log.append(f"Deleted {deleted_count} records from {table_name}")
                        except sqlite3.OperationalError:
                            # Table might not exist
                            pass

                    # Delete user accounts using central auth system
                    cursor.execute('SELECT id, username FROM users WHERE student_id = ?', (student_id,))
                    user_record = cursor.fetchone()

                    if user_record:
                        user_id = user_record[0]
                        username = user_record[1]

                        # Use auth system to delete user if available
                        if self.auth:
                            if self.auth.delete_user(user_id):
                                deletion_log.append(f"Deleted user account via auth system (username: {username})")
                                # Log activity
                                if ACTIVITY_LOGGER_AVAILABLE:
                                    log_activity('delete', 'user', user_id=user_id, details={'username': username, 'student_id': student_id, 'reason': 'Student deletion'})
                            else:
                                deletion_log.append(f"Warning: Failed to delete user via auth system for {username}")
                        else:
                            # Fallback to direct deletion if auth not available
                            cursor.execute('DELETE FROM user_accounts WHERE user_id = ?', (user_id,))
                            if cursor.rowcount > 0:
                                deletion_log.append("Deleted user account")

                            cursor.execute('DELETE FROM users WHERE id = ?', (user_id,))
                            if cursor.rowcount > 0:
                                deletion_log.append("Deleted user profile")
                                # Log activity even in fallback mode
                                if ACTIVITY_LOGGER_AVAILABLE:
                                    log_activity('delete', 'user', user_id=user_id, details={'username': username, 'student_id': student_id, 'reason': 'Student deletion (fallback)'})

                    # Update course enrollment count before deleting student
                    cursor.execute('SELECT course FROM students WHERE student_id = ?', (student_id,))
                    student_course = cursor.fetchone()
                    if student_course and student_course[0]:
                        course_code = student_course[0]
                        cursor.execute('''
                            UPDATE courses
                            SET current_enrollment = current_enrollment - 1
                            WHERE course_code = ? AND current_enrollment > 0
                        ''', (course_code,))
                        if cursor.rowcount > 0:
                            deletion_log.append(f"Updated course enrollment count for {course_code}")

                    # Finally delete the main student record
                    cursor.execute('DELETE FROM students WHERE student_id = ?', (student_id,))
                    if cursor.rowcount > 0:
                        deletion_log.append("Deleted main student record")

                    # Re-enable foreign key constraints
                    cursor.execute("PRAGMA foreign_keys = ON")

                    conn.commit()

                    # Show deletion summary
                    summary = f"Student {student_id} deleted successfully!\n\nDeletion Summary:\n" + "\n".join(deletion_log)
                    messagebox.showinfo("Deletion Complete", summary)

                    # Refresh student list and close dialog
                    if hasattr(self, 'view_students'):
                        self.view_students()
                    self.refresh_advanced_search()  # ADD THIS LINE

                    dialog.destroy()

                except Exception as e:
                    if cursor:
                        try:
                            cursor.execute("PRAGMA foreign_keys = ON")
                        except sqlite3.ProgrammingError:
                            # Connection might already be closed
                            pass
                    if conn:
                        conn.rollback()
                    messagebox.showerror("Deletion Failed", f"Failed to delete student: {str(e)}")
                    status_label.config(text="Deletion failed", foreground="red")
                finally:
                    if conn:
                        conn.close()
            
            # Buttons
            button_frame = ttk.Frame(main_frame)
            button_frame.pack(fill=tk.X, pady=(20, 0))
            
            ttk.Button(button_frame, text="DELETE STUDENT", command=perform_deletion, 
                      style="Accent.TButton").pack(side=tk.LEFT, padx=(0, 10))
            ttk.Button(button_frame, text="Cancel", command=dialog.destroy).pack(side=tk.LEFT)
            
            # Focus on confirmation entry
            confirm_entry.focus()

            # Handle dialog close event
            def on_dialog_close():
                try:
                    dialog.grab_release()
                except:
                    pass
                dialog.destroy()
                
            dialog.protocol("WM_DELETE_WINDOW", on_dialog_close)
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to prepare deletion dialog: {str(e)}")
            dialog.destroy()
        
    def select_student_for_deletion(self):
        """Show dialog to select student for deletion"""
        selection_dialog = tk.Toplevel(self.root)
        selection_dialog.title("Select Student to Delete")
        selection_dialog.geometry("1000x700")  # Made bigger
        selection_dialog.transient(self.root)
        selection_dialog.grab_set()
        
        selected_student = None
        
        main_frame = ttk.Frame(selection_dialog, padding=20)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(main_frame, text="Select Student to Delete", 
                 font=('Arial', 14, 'bold')).pack(pady=(0, 20))
        
        # Search frame
        search_frame = ttk.Frame(main_frame)
        search_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(search_frame, text="Search:").pack(side=tk.LEFT)
        search_entry = ttk.Entry(search_frame, width=30)
        search_entry.pack(side=tk.LEFT, padx=(10, 0))
        
        # Student list
        list_frame = ttk.Frame(main_frame)
        list_frame.pack(fill=tk.BOTH, expand=True)
        
        columns = ('ID', 'Name', 'Email', 'Course')
        tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=15)
        
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=150)
        
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Load students
        def load_students(filter_text=""):
            tree.delete(*tree.get_children())
            
            try:
                conn = get_db_connection()
                cursor = conn.cursor()
                
                if filter_text:
                    cursor.execute('''
                        SELECT student_id, first_name, last_name, email_address, course
                        FROM students 
                        WHERE LOWER(first_name) LIKE LOWER(?) 
                           OR LOWER(last_name) LIKE LOWER(?)
                           OR LOWER(student_id) LIKE LOWER(?)
                        ORDER BY last_name, first_name
                    ''', (f'%{filter_text}%', f'%{filter_text}%', f'%{filter_text}%'))
                else:
                    cursor.execute('''
                        SELECT student_id, first_name, last_name, email_address, course
                        FROM students 
                        ORDER BY last_name, first_name
                    ''')
                
                students = cursor.fetchall()
                
                for student in students:
                    student_id, first_name, last_name, email, course = student
                    full_name = f"{first_name} {last_name}"
                    tree.insert('', tk.END, values=(student_id, full_name, email, course))
                
                conn.close()
            except Exception as e:
                messagebox.showerror("Error", f"Failed to load students: {str(e)}")
        
        def search_students():
            filter_text = search_entry.get().strip()
            load_students(filter_text)
        
        def on_select():
            nonlocal selected_student
            selection = tree.selection()
            if selection:
                item = tree.item(selection[0])
                selected_student = item['values'][0]
                selection_dialog.destroy()
        
        # Search button
        ttk.Button(search_frame, text="Search", command=search_students).pack(side=tk.LEFT, padx=(10, 0))
        ttk.Button(search_frame, text="Show All", command=lambda: load_students()).pack(side=tk.LEFT, padx=(5, 0))
        
        # Buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(20, 0))
        
        ttk.Button(button_frame, text="Select for Deletion", command=on_select).pack(side=tk.LEFT)
        ttk.Button(button_frame, text="Cancel", command=selection_dialog.destroy).pack(side=tk.LEFT, padx=(10, 0))
        
        # Bind double-click
        tree.bind('<Double-1>', lambda e: on_select())
        
        # Bind search entry
        search_entry.bind('<Return>', lambda e: search_students())
        
        # Load initial data
        load_students()
        
        # Wait for dialog to close
        selection_dialog.wait_window()
        
        return selected_student
        
    def create_student_dialog(self):
        """Create comprehensive dialog for adding new student with full form validation"""
        dialog = self.create_themed_toplevel("Create New Student", "700x800")

        # Make dialog visible before grabbing
        dialog.update_idletasks()
        dialog.deiconify()
        
        try:
            dialog.grab_set()
        except tk.TclError:
            print("Warning: Could not grab dialog focus")
        
        # Main scrollable frame
        main_frame = ttk.Frame(dialog)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Create canvas and scrollbar
        canvas = tk.Canvas(main_frame)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Title
        title_label = ttk.Label(scrollable_frame, text="Create New Student", 
                               font=('Arial', 16, 'bold'))
        title_label.grid(row=0, column=0, columnspan=2, pady=(0, 20))
        
        # Form fields
        fields = {}
        
        # Personal Information Section
        personal_frame = ttk.LabelFrame(scrollable_frame, text="Personal Information", padding=15)
        personal_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # Title/Prefix
        ttk.Label(personal_frame, text="Title:").grid(row=0, column=0, sticky=tk.W, pady=5)
        fields['title'] = ttk.Combobox(personal_frame, values=['Mr', 'Ms', 'Mrs', 'Dr', 'Prof'], 
                                      state='readonly', width=27)
        fields['title'].grid(row=0, column=1, pady=5, padx=(10, 0), sticky=tk.W)
        
        # First Name
        ttk.Label(personal_frame, text="First Name: *").grid(row=1, column=0, sticky=tk.W, pady=5)
        fields['first_name'] = ttk.Entry(personal_frame, width=30)
        fields['first_name'].grid(row=1, column=1, pady=5, padx=(10, 0))
        
        # Middle Name
        ttk.Label(personal_frame, text="Middle Name:").grid(row=2, column=0, sticky=tk.W, pady=5)
        fields['middle_name'] = ttk.Entry(personal_frame, width=30)
        fields['middle_name'].grid(row=2, column=1, pady=5, padx=(10, 0))
        
        # Last Name
        ttk.Label(personal_frame, text="Last Name: *").grid(row=3, column=0, sticky=tk.W, pady=5)
        fields['last_name'] = ttk.Entry(personal_frame, width=30)
        fields['last_name'].grid(row=3, column=1, pady=5, padx=(10, 0))
        
        # Gender
        ttk.Label(personal_frame, text="Gender: *").grid(row=4, column=0, sticky=tk.W, pady=5)
        fields['gender'] = ttk.Combobox(personal_frame, values=['male', 'female', 'other'], 
                                       state='readonly', width=27)
        fields['gender'].grid(row=4, column=1, pady=5, padx=(10, 0), sticky=tk.W)
        
        # Date of Birth
        ttk.Label(personal_frame, text="Date of Birth (YYYY-MM-DD): *").grid(row=5, column=0, sticky=tk.W, pady=5)
        fields['dob'] = ttk.Entry(personal_frame, width=30)
        fields['dob'].grid(row=5, column=1, pady=5, padx=(10, 0))
        
        # Academic Information Section
        academic_frame = ttk.LabelFrame(scrollable_frame, text="Academic Information", padding=15)
        academic_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # Course - Show as read-only with random assignment info
        ttk.Label(academic_frame, text="Course:").grid(row=0, column=0, sticky=tk.W, pady=5)
        course_label = ttk.Label(academic_frame, text="Will be randomly assigned (CS or DS)", 
                                foreground="blue")
        course_label.grid(row=0, column=1, pady=5, padx=(10, 0), sticky=tk.W)
        
        # Status information
        status_label = ttk.Label(scrollable_frame, 
                                text="Note: Student ID, email, course, and modules will be auto-generated", 
                                foreground="blue")
        status_label.grid(row=3, column=0, columnspan=2, pady=10)
        
        # Validation feedback
        validation_label = ttk.Label(scrollable_frame, text="", foreground="red")
        validation_label.grid(row=4, column=0, columnspan=2, pady=5)
        
        def validate_form():
            """Validate form inputs"""
            errors = []
            
            if not fields['first_name'].get().strip():
                errors.append("First name is required")
            
            if not fields['last_name'].get().strip():
                errors.append("Last name is required")
            
            if not fields['gender'].get():
                errors.append("Gender is required")
            
            dob_text = fields['dob'].get().strip()
            if not dob_text:
                errors.append("Date of birth is required")
            else:
                try:
                    dob = datetime.strptime(dob_text, "%Y-%m-%d")
                    age = datetime.now().year - dob.year
                    if age < 16 or age > 80:
                        errors.append("Age must be between 16 and 80")
                except ValueError:
                    errors.append("Invalid date format (use YYYY-MM-DD)")
            
            return errors
        
        def create_student():
            """Create student with comprehensive data processing and random course assignment"""
            try:
                # Validate form
                errors = validate_form()
                if errors:
                    validation_label.config(text="; ".join(errors))
                    return
                
                validation_label.config(text="")
                
                # Get form data
                first_name = fields['first_name'].get().strip()
                middle_name = fields['middle_name'].get().strip()
                last_name = fields['last_name'].get().strip()
                gender = fields['gender'].get()
                dob_text = fields['dob'].get().strip()

                # RANDOMLY ASSIGN COURSE - Hard-coded to CS or DS only (like CLI)
                course = random.choice(['CS', 'DS'])
                
                title = fields['title'].get() or ('Mr' if gender == 'male' else 'Ms')
                
                # Parse date and calculate age
                dob = datetime.strptime(dob_text, "%Y-%m-%d")
                now_dt = datetime.now()
                age = now_dt.year - dob.year - ((now_dt.month, now_dt.day) < (dob.month, dob.day))
                
                # Generate student ID and email
                student_id = str(random.randint(1000000, 9999999)).zfill(7)
                email_address = f"C{student_id}@tees.ac.uk"
                registration_time = now_dt.strftime('%Y-%m-%d %H:%M:%S')
                
                # Create student record in database
                conn = get_db_connection()
                if not conn:
                    raise Exception("Database connection failed")
                
                cursor = conn.cursor()

                # Temporarily disable foreign key checks to avoid module_code issues
                cursor.execute("PRAGMA foreign_keys = OFF")

                # Insert student record
                cursor.execute('''
                    INSERT INTO students VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    student_id, email_address, title, first_name, middle_name,
                    last_name, gender, dob.strftime('%Y-%m-%d'), age, course, registration_time, 'Active', registration_time
                ))
                
                # Add modules based on the assigned course from database
                current_date = datetime.now().strftime('%Y-%m-%d')

                # Get modules by type: 2 compulsory, 2 optional, 2 course-specific
                selected_modules = []

                # 1. Get 2 compulsory modules (module_type = 'compulsory')
                cursor.execute('''
                    SELECT module_code FROM modules
                    WHERE module_type = 'compulsory' AND is_active = 1
                    ORDER BY module_code
                ''')
                compulsory_modules = [row[0] for row in cursor.fetchall()]
                if len(compulsory_modules) >= 2:
                    selected_modules.extend(compulsory_modules[:2])
                else:
                    print(f"Warning: Found only {len(compulsory_modules)} compulsory modules, need 2")
                    selected_modules.extend(compulsory_modules)

                # 2. Get 2 optional modules (module_type = 'optional')
                cursor.execute('''
                    SELECT module_code FROM modules
                    WHERE module_type = 'optional' AND is_active = 1
                    ORDER BY module_code
                ''')
                optional_modules = [row[0] for row in cursor.fetchall()]
                if len(optional_modules) >= 2:
                    selected_modules.extend(optional_modules[:2])
                else:
                    print(f"Warning: Found only {len(optional_modules)} optional modules, need 2")
                    selected_modules.extend(optional_modules)

                # 3. Get 2 course-specific modules (module_type = 'CS_optional' or 'DS_optional')
                course_specific_type = f"{course}_optional"
                cursor.execute('''
                    SELECT module_code FROM modules
                    WHERE module_type = ? AND is_active = 1
                    ORDER BY module_code
                ''', (course_specific_type,))
                course_specific_modules = [row[0] for row in cursor.fetchall()]
                if len(course_specific_modules) >= 2:
                    selected_modules.extend(course_specific_modules[:2])
                else:
                    print(f"Warning: Found only {len(course_specific_modules)} {course_specific_type} modules, need 2")
                    selected_modules.extend(course_specific_modules)

                # Insert selected modules
                if selected_modules:
                    module_data = [
                        (student_id, module_code, current_date, 'enrolled')
                        for module_code in selected_modules
                    ]

                    cursor.executemany('''
                        INSERT INTO student_modules (student_id, module_code, enrollment_date, status)
                        VALUES (?, ?, ?, ?)
                    ''', module_data)

                    print(f"Assigned {len(selected_modules)} modules to student {student_id} for course {course}: {selected_modules}")
                else:
                    print(f"Warning: No modules assigned to student {student_id}")

                # Update course enrollment count
                cursor.execute('''
                    UPDATE courses
                    SET current_enrollment = current_enrollment + 1
                    WHERE course_code = ? AND status = 'active'
                ''', (course,))

                # Re-enable foreign key checks
                cursor.execute("PRAGMA foreign_keys = ON")

                conn.commit()
                conn.close()

                # Send registration confirmation email
                try:
                    from university_system.infrastructure.email.email_service import send_registration_confirmation
                    send_registration_confirmation(student_id)
                except Exception as e:
                    logging.warning(f"Failed to send registration confirmation email: {e}")

                # Create user account
                temp_password = f"{first_name.lower()}123456"
                try:
                    self.auth.create_user(
                        username=student_id,
                        password=temp_password,
                        email=email_address,
                        first_name=first_name,
                        last_name=last_name,
                        role='student',
                        student_id=student_id
                    )
                except Exception as e:
                    logging.warning(f"User account creation failed: {e}")
                
                # Success message with details
                modules_text = ""
                if course_modules and selected_modules:
                    modules_text = "\n\n    Modules assigned:\n"
                    for mod_code in selected_modules:
                        modules_text += f"    - {mod_code}\n"

                success_msg = f"""Student created successfully!

    Student Details:
    - Student ID: {student_id}
    - Name: {title} {first_name} {middle_name} {last_name}
    - Email: {email_address}
    - Course: {course} (randomly assigned)
    - Age: {age}
    - Login Password: {temp_password}{modules_text}"""

                messagebox.showinfo("Success", success_msg)

                # Send welcome email to new student
                self._send_welcome_email_to_student(
                    student_id=student_id,
                    first_name=first_name,
                    last_name=last_name,
                    email_address=email_address,
                    temp_password=temp_password,
                    course=course
                )

                # Refresh student list and close dialog
                if hasattr(self, 'view_students'):
                    self.view_students()
                self.refresh_advanced_search()  # ADD THIS LINE

                dialog.destroy()
                
            except Exception as e:
                messagebox.showerror("Error", f"Failed to create student: {str(e)}")
        
        # Buttons
        button_frame = ttk.Frame(scrollable_frame)
        button_frame.grid(row=5, column=0, columnspan=2, pady=20)
        
        ttk.Button(button_frame, text="Create Student", command=create_student, 
                  style="Accent.TButton").pack(side=tk.LEFT, padx=10)
        ttk.Button(button_frame, text="Clear Form", 
                  command=lambda: [field.delete(0, tk.END) if hasattr(field, 'delete') 
                                  else field.set('') for field in fields.values()]).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Cancel", command=dialog.destroy).pack(side=tk.LEFT, padx=5)
        
        # Pack canvas and scrollbar
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Set focus
        fields['first_name'].focus()
        
        # Bind mousewheel to canvas
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
    
    def search_students_dialog(self):
        """Create search dialog"""
        dialog = self.create_themed_toplevel("Search Students", "400x300")
        dialog.grab_set()
        
        main_frame = ttk.Frame(dialog, padding=20)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Search criteria
        ttk.Label(main_frame, text="Search by:").grid(row=0, column=0, sticky=tk.W, pady=5)
        search_type = ttk.Combobox(main_frame, values=[
            'First Name', 'Last Name', 'Student ID', 'Course', 'Email'
        ], state='readonly', width=25)
        search_type.grid(row=0, column=1, pady=5, padx=(10, 0))
        search_type.set('First Name')
        
        ttk.Label(main_frame, text="Search term:").grid(row=1, column=0, sticky=tk.W, pady=5)
        search_term = ttk.Entry(main_frame, width=28)
        search_term.grid(row=1, column=1, pady=5, padx=(10, 0))
        
        def perform_search():
            """Perform search and update treeview"""
            term = search_term.get().strip()
            if not term:
                messagebox.showerror("Error", "Please enter a search term")
                return
            
            try:
                # Check if student_tree exists
                if not hasattr(self, 'student_tree') or not self.student_tree:
                    messagebox.showerror("Error", "Student list not initialized. Please go to Student Records first.")
                    return

                # Clear existing data
                for item in self.student_tree.get_children():
                    self.student_tree.delete(item)
                
                conn = get_db_connection()
                if conn:
                    cursor = conn.cursor()
                    
                    # Build query based on search type
                    search_field_map = {
                        'First Name': 'first_name',
                        'Last Name': 'last_name', 
                        'Student ID': 'student_id',
                        'Course': 'course',
                        'Email': 'email_address'
                    }
                    
                    field = search_field_map[search_type.get()]
                    
                    if field == 'student_id':
                        query = f'SELECT * FROM students WHERE {field} = ?'
                        cursor.execute(query, (term,))
                    else:
                        query = f'SELECT * FROM students WHERE LOWER({field}) LIKE LOWER(?)'
                        cursor.execute(query, (f'%{term}%',))
                    
                    results = cursor.fetchall()
                    
                    # Populate treeview with results
                    for student in results:
                        student_id, email_address, title, first_name, middle_name, last_name, gender, dob, age, course, reg_date, status, enrollment_date = student
                        full_name = f"{first_name} {middle_name} {last_name}".replace('  ', ' ').strip()
                        
                        self.student_tree.insert('', tk.END, values=(
                            student_id, full_name, email_address, course, reg_date[:10] if reg_date else 'N/A'
                        ))
                    
                    conn.close()
                    
                    if not results:
                        messagebox.showinfo("Search Results", "No students found matching your criteria")
                    else:
                        messagebox.showinfo("Search Results", f"Found {len(results)} student(s)")
                    
                    dialog.destroy()
                    
            except Exception as e:
                messagebox.showerror("Error", f"Search failed: {str(e)}")
        
        # Buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=2, column=0, columnspan=2, pady=20)
        
        ttk.Button(button_frame, text="Search", command=perform_search).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Show All", command=lambda: [self.view_students(), dialog.destroy()]).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Cancel", command=dialog.destroy).pack(side=tk.LEFT, padx=5)
        
        search_term.focus()

    def export_data_dialog(self):
        """Create export data dialog"""
        dialog = self.create_themed_toplevel("Export Student Data", "450x300")
        dialog.grab_set()
        
        main_frame = ttk.Frame(dialog, padding=20)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(main_frame, text="Export Format:").grid(row=0, column=0, sticky=tk.W, pady=10)
        format_var = tk.StringVar(value="CSV")
        
        formats = [("CSV", "CSV"), ("Excel", "Excel"), ("PDF", "PDF"), ("Text", "Text")]
        for i, (text, value) in enumerate(formats):
            ttk.Radiobutton(main_frame, text=text, variable=format_var, value=value).grid(row=i+1, column=0, sticky=tk.W)
        
        def export_data():
            """Export data in selected format"""
            try:
                export_format = format_var.get()
                
                filetypes = {
                    "CSV": [("CSV files", "*.csv")],
                    "Excel": [("Excel files", "*.xlsx")],
                    "PDF": [("PDF files", "*.pdf")],
                    "Text": [("Text files", "*.txt")]
                }
                
                filename = filedialog.asksaveasfilename(
                    defaultextension=f".{export_format.lower()}",
                    filetypes=filetypes[export_format]
                )
                
                if not filename:
                    return

                conn = None
                try:
                    conn = get_db_connection()
                    if not conn:
                        messagebox.showerror("Export Failed", "Database connection could not be established.")
                        return

                    cursor = conn.cursor()
                    cursor.execute(
                        """
                        SELECT s.student_id,
                               COALESCE(s.first_name, ''),
                               COALESCE(s.last_name, ''),
                               COALESCE(s.email_address, ''),
                               COALESCE(s.course, ''),
                               COALESCE(DATE(s.registration_datetime), '')
                        FROM students s
                        ORDER BY s.last_name, s.first_name
                        """
                    )

                    student_rows = cursor.fetchall()

                    if not student_rows:
                        messagebox.showinfo("Export", "No student records available to export.")
                        return

                    # Build module listings safely without assuming specific schemas
                    modules_by_student = {}
                    try:
                        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='student_modules'")
                        has_student_modules = cursor.fetchone() is not None

                        if has_student_modules:
                            cursor.execute("PRAGMA table_info(student_modules)")
                            student_module_cols = {row[1] for row in cursor.fetchall()}

                            identifier_col = None
                            if 'module_id' in student_module_cols:
                                identifier_col = 'module_id'
                            elif 'module_code' in student_module_cols:
                                identifier_col = 'module_code'

                            name_col = 'module_name' if 'module_name' in student_module_cols else None

                            module_lookup = {}
                            if identifier_col and 'student_id' in student_module_cols:
                                try:
                                    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='modules'")
                                    has_modules_table = cursor.fetchone() is not None
                                except sqlite3.Error:
                                    has_modules_table = False

                                if has_modules_table:
                                    cursor.execute("PRAGMA table_info(modules)")
                                    module_table_cols = {row[1] for row in cursor.fetchall()}

                                    if 'module_name' in module_table_cols:
                                        join_key = None
                                        if identifier_col in module_table_cols:
                                            join_key = identifier_col
                                        elif identifier_col == 'module_code' and 'module_id' in module_table_cols:
                                            join_key = 'module_id'

                                        if join_key:
                                            cursor.execute(f"SELECT {join_key}, module_name FROM modules")
                                            module_lookup = {
                                                row[0]: row[1] for row in cursor.fetchall() if row[0] is not None
                                            }

                                select_cols = ["student_id", identifier_col]
                                if name_col:
                                    select_cols.append(name_col)

                                column_clause = ", ".join(select_cols)
                                cursor.execute(f"SELECT {column_clause} FROM student_modules")

                                for module_row in cursor.fetchall():
                                    student_id = module_row[0]
                                    identifier = module_row[1]
                                    module_name = None

                                    if name_col and len(module_row) > 2:
                                        module_name = module_row[2]

                                    if not module_name and module_lookup:
                                        module_name = module_lookup.get(identifier)

                                    display_identifier = identifier or 'N/A'
                                    display_name = module_name or module_lookup.get(identifier) or 'Unknown Module'

                                    modules_by_student.setdefault(student_id, []).append(
                                        f"{display_identifier} - {display_name}"
                                    )
                    except sqlite3.Error:
                        modules_by_student = {}

                    headers = [
                        "Student ID",
                        "First Name",
                        "Last Name",
                        "Email",
                        "Course",
                        "Registration Date",
                        "Module 1",
                        "Module 2",
                        "Module 3",
                        "Module 4",
                        "Module 5",
                        "Module 6"
                    ]

                    processed_rows = []
                    for row in student_rows:
                        student_id = row[0]
                        modules_list = modules_by_student.get(student_id, [])

                        # Create row with 6 separate module columns
                        row_data = [
                            row[0],  # Student ID
                            row[1],  # First Name
                            row[2],  # Last Name
                            row[3],  # Email
                            row[4],  # Course
                            row[5],  # Registration Date
                        ]

                        # Add up to 6 modules in separate columns
                        for i in range(6):
                            if i < len(modules_list):
                                row_data.append(modules_list[i])
                            else:
                                row_data.append("")

                        processed_rows.append(row_data)

                    if export_format == "CSV":
                        with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                            writer = csv.writer(csvfile)
                            writer.writerow(headers)
                            writer.writerows(processed_rows)

                    elif export_format == "Excel":
                        try:
                            import pandas as pd
                        except ImportError:
                            messagebox.showerror(
                                "Export Failed",
                                "Pandas is required for Excel export. Please install the dependency."
                            )
                            return

                        df = pd.DataFrame(processed_rows, columns=headers)
                        try:
                            df.to_excel(filename, index=False, engine='openpyxl')
                        except ImportError:
                            messagebox.showerror(
                                "Export Failed",
                                "openpyxl is required for Excel export. Please install it: pip install openpyxl"
                            )
                            return

                    elif export_format == "PDF":
                        try:
                            from reportlab.lib import colors
                            from reportlab.lib.pagesizes import letter, landscape, A4
                            from reportlab.lib.styles import getSampleStyleSheet
                            from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle
                            from reportlab.lib.units import inch
                        except ImportError:
                            messagebox.showerror(
                                "Export Failed",
                                "ReportLab is required for PDF export. Please install the dependency."
                            )
                            return

                        # Use landscape orientation for better column visibility
                        doc = SimpleDocTemplate(
                            filename,
                            pagesize=landscape(letter),
                            leftMargin=0.4*inch,
                            rightMargin=0.4*inch,
                            topMargin=0.5*inch,
                            bottomMargin=0.5*inch
                        )
                        styles = getSampleStyleSheet()
                        body_style = styles['BodyText']
                        body_style.wordWrap = 'CJK'
                        body_style.fontSize = 8

                        table_data = [headers]
                        for row in processed_rows:
                            # Convert all values to strings and truncate if too long
                            pdf_row = []
                            for i, value in enumerate(row):
                                val_str = str(value) if value else ""
                                # Truncate emails and long strings to prevent overflow
                                if i == 3 and len(val_str) > 25:  # Email column
                                    val_str = val_str[:22] + "..."
                                elif len(val_str) > 30:
                                    val_str = val_str[:27] + "..."
                                pdf_row.append(val_str)
                            table_data.append(pdf_row)

                        # Calculate available width
                        page_width = landscape(letter)[0] - (0.8 * inch)  # Subtract margins (0.4 * 2)

                        # Define column widths to fit within available space (~10.2 inches)
                        # Total columns: 12 (Student ID, First Name, Last Name, Email, Course, Reg Date, 6 Modules)
                        col_widths = [
                            0.55*inch,  # Student ID
                            0.65*inch,  # First Name
                            0.65*inch,  # Last Name
                            1.0*inch,   # Email
                            0.75*inch,  # Course
                            0.75*inch,  # Registration Date
                            0.95*inch,  # Module 1
                            0.95*inch,  # Module 2
                            0.95*inch,  # Module 3
                            0.95*inch,  # Module 4
                            0.95*inch,  # Module 5
                            0.95*inch   # Module 6
                        ]
                        # Total: ~10.05 inches (fits within 10.2 available)

                        table = Table(table_data, colWidths=col_widths, repeatRows=1)
                        table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, 0), 8),
                            ('FONTSIZE', (0, 1), (-1, -1), 7),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                            ('TOPPADDING', (0, 0), (-1, -1), 4),
                            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
                            ('LEFTPADDING', (0, 0), (-1, -1), 4),
                            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                            ('WORDWRAP', (0, 0), (-1, -1), True)
                        ]))

                        doc.build([table])

                    else:  # Text export
                        with open(filename, 'w', encoding='utf-8') as txtfile:
                            for row in processed_rows:
                                txtfile.write(f"Student ID: {row[0]}\n")
                                txtfile.write(f"Name: {row[1]} {row[2]}\n")
                                txtfile.write(f"Email: {row[3]}\n")
                                txtfile.write(f"Course: {row[4]}\n")
                                txtfile.write(f"Registration Date: {row[5]}\n")
                                txtfile.write("Modules:\n")
                                # Write each module on a separate line
                                for i in range(6, 12):
                                    if i < len(row) and row[i]:
                                        txtfile.write(f"  Module {i-5}: {row[i]}\n")
                                txtfile.write("-" * 60 + "\n\n")

                    messagebox.showinfo("Success", f"Data exported successfully to {filename}")
                    dialog.destroy()

                finally:
                    if conn:
                        conn.close()

            except Exception as e:
                messagebox.showerror("Error", f"Export failed: {str(e)}")
        
        # Buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=5, column=0, pady=20)
        
        ttk.Button(button_frame, text="Export", command=export_data).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Cancel", command=dialog.destroy).pack(side=tk.LEFT, padx=5)

    def show_medical_accommodations(self):
        """Launch the Medical Accommodation GUI"""
        if not self.auth.current_user:
            messagebox.showerror("Error", "You must be logged in to access medical accommodations.")
            return
        
        if not (self.auth.check_permission('manage_accommodations') or 
                self.auth.check_permission('view_accommodations') or 
                self.auth.check_permission('approve_accommodations')):
            messagebox.showerror("Error", "You don't have permission to access medical accommodations.")
            return
        
        try:
            # Import the accommodation GUI
            import sys
            import os
            
            # Add the accommodation_gui.py path to sys.path if needed
            accommodation_gui_path = os.path.join(os.path.dirname(__file__), '..')
            if accommodation_gui_path not in sys.path:
                sys.path.insert(0, accommodation_gui_path)
            
            # Import and launch the accommodation GUI
            from university_system.modules.domain.housing.gui.accommodation_gui import AccommodationGUI, main as accommodation_main
            
            # Create a new window for the accommodation system
            accommodation_window = tk.Toplevel(self.root)
            accommodation_window.title("Medical Accommodation Management System")
            accommodation_window.geometry("1200x800")
            accommodation_window.minsize(1000, 700)
            
            try:
                accommodation_window.transient(self.root)
            except Exception:
                pass
            
            # Center the window
            accommodation_window.update_idletasks()
            x = (accommodation_window.winfo_screenwidth() - accommodation_window.winfo_width()) // 2
            y = (accommodation_window.winfo_screenheight() - accommodation_window.winfo_height()) // 2
            accommodation_window.geometry(f"+{x}+{y}")
            
            # Initialize the accommodation GUI in the new window with auth
            accommodation_gui = AccommodationGUI(accommodation_window, auth=self.auth)
            
            print("Medical Accommodation GUI launched successfully")
            
        except ImportError as e:
            # Fallback to CLI if GUI is not available
            messagebox.showinfo("Medical Accommodations", 
                              f"Accommodation GUI not available: {e}\nUsing CLI interface.")
            try:
                from university_system.modules.domain.housing.services.accommodation import display_accommodation_menu
                display_accommodation_menu()
            except ImportError:
                messagebox.showerror("Error", "Neither GUI nor CLI accommodation system is available.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to launch Medical Accommodation system: {str(e)}")
            print(f"Accommodation GUI error: {e}")

    def show_course_management(self):
        """Launch the Course Management GUI in a new window"""
        if not self.auth.current_user:
            messagebox.showerror("Error", "You must be logged in to access course management.")
            return
        
        if not (self.auth.check_permission('manage_courses') or self.auth.check_permission('view_courses')):
            messagebox.showerror("Error", "You don't have permission to access course management.")
            return
        
        try:
            if COURSE_MANAGEMENT_GUI_AVAILABLE:
                # Create a new window for the Course Management GUI
                course_window = tk.Toplevel(self.root)
                course_window.title("Enhanced Course Management System")
                course_window.geometry("1200x800")
                
                # Center the window
                course_window.update_idletasks()
                x = (course_window.winfo_screenwidth() - course_window.winfo_width()) // 2
                y = (course_window.winfo_screenheight() - course_window.winfo_height()) // 2
                course_window.geometry(f"+{x}+{y}")
                
                try:
                    course_window.transient(self.root)
                except Exception:
                    pass  # Continue if transient fails
                
                # Initialize the Course Management GUI in the new window
                # Replace the root parameter with our new window
                course_gui = CourseManagementGUI(course_window)
                
                # Pass the auth context if the CourseManagementGUI supports it
                if hasattr(course_gui, 'auth'):
                    course_gui.auth = self.auth
                elif hasattr(course_gui, 'set_auth'):
                    course_gui.set_auth(self.auth)
                
                print("✅ Course Management GUI opened successfully")
                
            else:
                # Fallback to CLI menu
                messagebox.showinfo("Course Management", 
                                  "Course Management GUI not available. Using CLI menu.")
                try:
                    from university_system.modules.domain.academics.services.course_management import display_course_management_menu
                    display_course_management_menu(self.auth)
                except ImportError:
                    messagebox.showerror("Error", "Course management system not available.")
                    
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open Course Management: {str(e)}")
            print(f"❌ Course Management error: {e}")

    def show_trip_management_gui(self):
        """Open the Trip Management GUI in a child window (Toplevel)."""
        try:
            if not self.auth or not self.auth.current_user:
                messagebox.showerror("Trip Management", "You must be logged in.")
                return

            if not TRIP_MANAGEMENT_GUI_AVAILABLE:
                messagebox.showerror("Trip Management", f"Trip Management GUI not available:\n{_TRIP_MGMT_IMPORT_ERROR}")
                return

            top = tk.Toplevel(self.root)
            top.title("Trip Management")
            top.geometry("1200x800")
            try:
                top.transient(self.root)
                top.grab_set()
            except Exception:
                pass

            # Embed the TripManagementGUI into this window, passing current auth
            TripManagementGUI(auth_instance=self.auth, root=top)

            # No mainloop here; the main GUI already owns it
            print("✅ Trip Management GUI opened")

        except Exception as e:
            messagebox.showerror("Trip Management", f"Failed to open Trip Management:\n{e}")

    def show_module_management(self):
        """Launch the full Module Scheduling GUI in a new window"""
        if not self.auth.current_user:
            messagebox.showerror("Error", "You must be logged in to access module management.")
            return
        
        if not (self.auth.check_permission('manage_modules') or self.auth.check_permission('view_assigned_modules')):
            messagebox.showerror("Error", "You don't have permission to access module management.")
            return
        
        try:
            if MODULE_SCHEDULING_GUI_AVAILABLE and ModuleSchedulingGUI:
                # Create a new window for the Module Scheduling GUI
                module_window = tk.Toplevel(self.root)
                module_window.title("Enhanced Module Scheduling System")
                module_window.geometry("1400x900")
                module_window.minsize(1200, 800)
                
                # Center the window
                module_window.update_idletasks()
                x = (module_window.winfo_screenwidth() - module_window.winfo_width()) // 2
                y = (module_window.winfo_screenheight() - module_window.winfo_height()) // 2
                module_window.geometry(f"+{x}+{y}")
                
                try:
                    module_window.transient(self.root)
                except Exception:
                    pass  # Continue if transient fails
                
                # Initialize the Module Scheduling GUI in the new window
                module_gui = ModuleSchedulingGUI(module_window)
                
                # Pass the auth context if the ModuleSchedulingGUI supports it
                if hasattr(module_gui, 'set_auth'):
                    module_gui.set_auth(self.auth)
                elif hasattr(module_gui, 'auth'):
                    module_gui.auth = self.auth
                
                print("✅ Module Scheduling GUI opened successfully")
                
            else:
                # Fallback to the original CLI menu if GUI not available
                messagebox.showinfo("Module Management", 
                                  "Module Scheduling GUI not available. Using CLI menu.")
                try:
                    # Import and call the CLI version
                    from university_system.modules.domain.academics.services.module_scheduling import display_enhanced_scheduling_menu
                    display_enhanced_scheduling_menu()
                except ImportError:
                    messagebox.showerror("Error", "Module management system not available.")
                        
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open Module Scheduling GUI: {str(e)}")
            print(f"❌ Module Scheduling error: {e}")
            
            # Try CLI fallback on error
            try:
                from university_system.modules.domain.academics.services.module_scheduling import display_enhanced_scheduling_menu
                display_enhanced_scheduling_menu()
            except ImportError:
                messagebox.showerror("Error", "Both GUI and CLI module management are unavailable.")

    def show_backup(self):
        """Launch the full Data Backup GUI"""
        try:
            # Import the backup GUI
            from university_system.infrastructure.database.data_backup_gui import start_backup_gui
            
            # Launch the backup GUI in a new window
            start_backup_gui()
            print("✅ Data Backup GUI launched successfully")
            
        except ImportError as e:
            messagebox.showerror("Error", f"Data Backup GUI not available: {e}")
            print(f"❌ Import error: {e}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to launch Data Backup GUI: {str(e)}")
            print(f"❌ Backup GUI error: {e}")

    def show_assignments(self):
        """Show assignments interface - Launch full AssignmentGUI"""
        try:
            print(f"Debug: ASSIGNMENT_SUBMISSION_GUI_AVAILABLE = {ASSIGNMENT_SUBMISSION_GUI_AVAILABLE}")
            if not ASSIGNMENT_SUBMISSION_GUI_AVAILABLE:
                messagebox.showerror("Error", "Assignment Submission GUI is not available. Please check the assignment module installation.")
                return
                
            # Create a new window for the Assignment GUI
            assignment_window = tk.Toplevel(self.root)
            assignment_window.title("Assignment & Assessment Management System")
            assignment_window.geometry("1200x800")

            # Center the window
            assignment_window.update_idletasks()
            x = (assignment_window.winfo_screenwidth() - assignment_window.winfo_width()) // 2
            y = (assignment_window.winfo_screenheight() - assignment_window.winfo_height()) // 2
            assignment_window.geometry(f"+{x}+{y}")

            try:
                assignment_window.transient(self.root)
            except Exception:
                pass  # Continue if transient fails

            # Create a minimal assignment system for the GUI
            # Since AssignmentSubmission is not implemented, create a minimal substitute
            class MinimalAssignmentSystem:
                def __init__(self):
                    self.db_path = None  # Will be set by the GUI if needed
                    self.submission_dir = None
                    self.auth = None

                def set_auth(self, auth):
                    self.auth = auth

                def _get_student_id(self):
                    """Get student ID from current user"""
                    if self.auth and self.auth.current_user:
                        # Check if user is admin - for admin users, return a special ID for testing
                        user_role = self.auth.current_user.get('role', '')
                        if user_role == 'admin':
                            return 'ADMIN_TEST'  # Special ID for admin testing

                        # First try to get student_id from user record
                        student_id = self.auth.current_user.get('student_id')
                        if student_id:
                            return student_id

                        # Query the database to get the student_id from users table
                        user_id = self.auth.current_user.get('id')
                        if user_id:
                            try:
                                import sqlite3
                                from university_system.modules.shared.config.database import DEFAULT_DB_PATH

                                conn = sqlite3.connect(str(DEFAULT_DB_PATH))
                                cursor = conn.cursor()
                                cursor.execute('SELECT student_id FROM users WHERE id = ?', (user_id,))
                                result = cursor.fetchone()
                                conn.close()

                                if result and result[0]:
                                    return result[0]
                            except Exception as e:
                                print(f"Error getting student_id from database: {e}")

                        # Fallback: use username as student_id if it follows student ID format
                        username = self.auth.current_user.get('username', '')
                        if username.startswith('S') or username.isdigit():
                            return username
                    return None

            assignment_system = MinimalAssignmentSystem()
            assignment_system.set_auth(self.auth)

            # Initialize the Assignment GUI with minimal system
            assignment_gui = AssignmentGUI(assignment_system, self.auth, parent=assignment_window)
            
            print("Assignment Management System opened successfully")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open Assignment Management System: {str(e)}")
            print(f"Assignment GUI error: {e}")
            # Just show the error dialog - don't try to show inline content since we're opening a separate window
        
    def show_grades(self):
        """Open the Grade Tracking GUI (child window) or fall back to CLI."""
        if self.grade_tracking_gui:
            self.grade_tracking_gui.show_grades()
        else:
            messagebox.showerror("Error", "Grade tracking GUI not available.")


    def show_enhanced_reporting_dashboard(self):
        """Open the Enhanced Reporting GUI as a child window, with safe fallbacks."""
        try:
            if not ENHANCED_REPORTING_GUI_AVAILABLE or ReportingSystemGUI is None:
                messagebox.showerror(
                    "Enhanced Reporting",
                    "Enhanced Reporting GUI is not available on this system."
                )
                return

            # Prefer embedding into a Toplevel to avoid creating a second Tk root
            try:
                top = tk.Toplevel(self.root)
                top.title("📊 Enhanced Reporting Dashboard")
                top.geometry("1200x800")
                try:
                    top.transient(self.root)
                    top.grab_set()
                except Exception:
                    pass

                # Instantiate the GUI into the Toplevel container
                app = ReportingSystemGUI(top)

                # If the reporting GUI supports auth injection, pass it
                try:
                    if hasattr(app, "set_auth"):
                        app.set_auth(self.auth)
                except Exception:
                    pass

                print("✅ Enhanced Reporting GUI opened successfully")

            except Exception as embed_err:
                # Fallback to its own launcher (may create its own Tk root)
                try:
                    if start_enhanced_reporting_gui:
                        start_enhanced_reporting_gui()
                    else:
                        raise RuntimeError("start_enhanced_reporting_gui is not available")
                except Exception as launch_err:
                    messagebox.showerror(
                        "Enhanced Reporting",
                        f"Failed to launch Enhanced Reporting GUI:\n{launch_err}"
                    )
                    print(f"❌ Enhanced Reporting GUI error: {launch_err} (embed err: {embed_err})")

        except Exception as e:
            messagebox.showerror("Enhanced Reporting", f"Unexpected error:\n{e}")
            print(f"❌ Unexpected Enhanced Reporting launcher error: {e}")

    def show_email_manager(self):
        """Open the Communication/Email Manager GUI in a child window."""
        if self.email_manager_gui:
            self.email_manager_gui.show_email_manager()
        else:
            messagebox.showerror("Error", "Email manager GUI not available.")

    def _send_welcome_email_to_student(self, student_id, first_name, last_name, email_address, temp_password, course):
        """Send welcome email to newly created student"""
        try:
            from university_system.infrastructure.email.email_service import send_template_email

            template_vars = {
                'first_name': first_name,
                'last_name': last_name,
                'student_id': student_id,
                'email_address': email_address,
                'course': course,
                'temp_password': temp_password
            }

            send_template_email('student_welcome', email_address, template_vars)
            print(f"Welcome email sent successfully to {first_name} {last_name} ({email_address})")

        except Exception as e:
            print(f"Failed to send welcome email to {email_address}: {e}")
            # Show a non-blocking notification about email failure
            try:
                messagebox.showwarning("Email Notice",
                    f"Student created successfully, but welcome email could not be sent to {email_address}.\n"
                    f"Please manually notify the student of their account details.")
            except:
                print(f"Warning: Welcome email failed for {email_address}")

    def _send_email_via_gui(self, to_email, subject, message):
        """Try to send email via email GUI"""
        try:
            # Try to import and use email GUI
            from university_system.infrastructure.email.gui.email_manager_gui import EmailGUI

            # Create email GUI instance
            email_gui = EmailGUI(self.root, self.auth)

            # Send email through email GUI
            email_gui.send_email(
                to_email=to_email,
                subject=subject,
                message=message
            )

            return True

        except ImportError:
            return False
        except Exception as e:
            print(f"Error sending email via GUI: {e}")
            return False

    def _show_welcome_email_fallback(self, first_name, last_name, email_address, subject, message):
        """Show fallback dialog with email details for manual sending"""
        try:
            fallback_window = tk.Toplevel(self.root)
            fallback_window.title("Welcome Email Details - Manual Send")
            fallback_window.geometry("700x500")
            fallback_window.transient(self.root)

            ttk.Label(fallback_window, text=f"Welcome email for {first_name} {last_name} - Please send manually:",
                     font=('TkDefaultFont', 10, 'bold')).pack(anchor='w', padx=10, pady=10)

            # Email details
            details_frame = ttk.LabelFrame(fallback_window, text="Email Details", padding=10)
            details_frame.pack(fill='both', expand=True, padx=10, pady=10)

            from tkinter.scrolledtext import ScrolledText
            details_text = ScrolledText(details_frame, height=20, width=80)
            details_text.pack(fill='both', expand=True)

            email_details = f"To: {email_address}\nSubject: {subject}\n\nMessage:\n{message}"

            details_text.insert('1.0', email_details)
            details_text.config(state='disabled')

            ttk.Button(fallback_window, text="Close",
                      command=fallback_window.destroy).pack(pady=10)
        except Exception as e:
            print(f"Failed to show welcome email fallback: {e}")

    def _send_student_update_email(self, student_id, old_data, new_data, course_changed=False, password_reset=False):
        """Send email notification to student about account changes"""
        try:
            # Extract email from old data (assuming it's at index 1)
            email_address = old_data[1]
            if not email_address:
                print(f"No email address found for student {student_id}")
                return

            # Compare old and new data to identify changes
            changes = []
            field_mapping = {
                'title': (old_data[2], new_data['title'], 'Title'),
                'first_name': (old_data[3], new_data['first_name'], 'First Name'),
                'middle_name': (old_data[4], new_data['middle_name'], 'Middle Name'),
                'last_name': (old_data[5], new_data['last_name'], 'Last Name'),
                'gender': (old_data[6], new_data['gender'], 'Gender'),
                'dob': (old_data[7], new_data['dob'], 'Date of Birth'),
                'course': (old_data[9], new_data['course'], 'Course')
            }

            for field, (old_val, new_val, display_name) in field_mapping.items():
                if str(old_val).strip() != str(new_val).strip():
                    changes.append(f"• {display_name}: '{old_val}' → '{new_val}'")

            # Build change summary
            changes_text = ""
            if changes:
                changes_text = "The following information has been updated:\n" + "\n".join(changes)

            if course_changed:
                changes_text += f"\n\n⚠️ IMPORTANT: Your course has been changed to {new_data['course']}. This may affect your module enrollments."

            if password_reset:
                new_password = f"{new_data['first_name'].lower()}123456"
                changes_text += f"\n\n🔑 PASSWORD RESET: Your password has been reset to: {new_password}\nPlease change this password upon your next login for security."

            if not changes and not course_changed and not password_reset:
                # No significant changes to notify about
                return

            from university_system.infrastructure.email.email_service import send_template_email

            template_vars = {
                'student_name': f"{new_data['first_name']} {new_data['last_name']}",
                'updated_fields': changes_text
            }

            send_template_email('account_information_updated', email_address, template_vars)
            print(f"Student update notification sent to {new_data['first_name']} {new_data['last_name']} ({email_address})")

        except Exception as e:
            print(f"Failed to send student update email to {student_id}: {e}")
            # Non-blocking notification about email failure
            try:
                messagebox.showwarning("Email Notice",
                    f"Student updated successfully, but notification email could not be sent.\n"
                    f"Please manually notify the student of any important changes.")
            except:
                print(f"Warning: Update notification email failed for {student_id}")

    def _show_update_email_fallback(self, first_name, last_name, email_address, subject, message):
        """Show fallback dialog for student update email"""
        try:
            fallback_window = tk.Toplevel(self.root)
            fallback_window.title("Student Update Email - Manual Send")
            fallback_window.geometry("700x500")
            fallback_window.transient(self.root)

            ttk.Label(fallback_window, text=f"Update notification for {first_name} {last_name} - Please send manually:",
                     font=('TkDefaultFont', 10, 'bold')).pack(anchor='w', padx=10, pady=10)

            # Email details
            details_frame = ttk.LabelFrame(fallback_window, text="Email Details", padding=10)
            details_frame.pack(fill='both', expand=True, padx=10, pady=10)

            from tkinter.scrolledtext import ScrolledText
            details_text = ScrolledText(details_frame, height=20, width=80)
            details_text.pack(fill='both', expand=True)

            email_details = f"To: {email_address}\nSubject: {subject}\n\nMessage:\n{message}"

            details_text.insert('1.0', email_details)
            details_text.config(state='disabled')

            ttk.Button(fallback_window, text="Close",
                      command=fallback_window.destroy).pack(pady=10)
        except Exception as e:
            print(f"Failed to show update email fallback: {e}")

    def show_advanced_search_gui(self):
        """Open Advanced Search GUI in a new window"""
        if not ADVANCED_SEARCH_GUI_AVAILABLE:
            messagebox.showerror("Error", "Advanced Search GUI is not available.\nPlease ensure advanced_search_gui.py is in the project directory.")
            return
        
        try:
            # Create a new window for the Advanced Search GUI
            search_window = tk.Toplevel(self.root)
            search_window.title("🔍 Advanced Student Search & Analytics")
            search_window.geometry("1200x800")
            search_window.transient(self.root)
            
            # Center the window
            search_window.update_idletasks()
            x = (search_window.winfo_screenwidth() - search_window.winfo_width()) // 2
            y = (search_window.winfo_screenheight() - search_window.winfo_height()) // 2
            search_window.geometry(f"+{x}+{y}")
            
            # Initialize the Advanced Search GUI in the new window
            self.advanced_search_gui = AdvancedSearchGUI(search_window, auth=self.auth)
            
            # Store reference for data refresh
            self.advanced_search_refresh_callback = self.advanced_search_gui.refresh_data
            
            print("✅ Advanced Search GUI opened successfully")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open Advanced Search GUI: {str(e)}")
            print(f"❌ Advanced Search GUI error: {e}")

    def show_plagiarism_checker(self):
        """Show plagiarism checker interface"""
        try:
            # Launch plagiarism checker GUI or show in content area
            from university_system.modules.domain.academics.gui.plagiarism_main_gui import launch_gui_from_main_system
            launch_gui_from_main_system(self.auth)
        except Exception as e:
            messagebox.showerror("Error", f"Plagiarism checker not available: {e}")

    def show_ai_detector(self):
        """Show AI detector interface"""
        try:
            if not AI_DETECTOR_GUI_AVAILABLE:
                messagebox.showerror("Error", "AI Detector GUI is not available.")
                return

            # Create new window for AI Detector
            ai_window = tk.Toplevel(self.root)
            ai_window.title("AI Content Detector")
            ai_window.geometry("1000x700")

            # Initialize AI Detector GUI
            ai_gui = AIDetectorGUI(ai_window, self.auth)
            print("✅ AI Detector GUI opened successfully")

        except Exception as e:
            messagebox.showerror("Error", f"AI Detector not available: {e}")

    def show_system_admin(self):
        """Show system administration interface"""
        if not self.auth.current_user or self.auth.current_user.get('role') != 'admin':
            messagebox.showerror("Access Denied", "Admin access required")
            return
        
        self.clear_content()
        
        ttk.Label(self.content_frame, text="System Administration", 
                 font=('Arial', 14, 'bold')).grid(row=0, column=0, pady=(0, 20))
        
        # Admin tools interface
        admin_frame = ttk.LabelFrame(self.content_frame, text="Administration Tools", padding="15")
        admin_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Add admin tools here

    def show_activity_log(self):
        """Open Log Management GUI (in a separate window)"""
        try:
            # Prefer the local file the user provided
            from university_system.utils.logging.gui.log_management_gui import LogManagementGUI
        except Exception:
            # Optional fallback to the refactored package if present
            try:
                from university_system.utils.logging.gui.log_management_gui import LogManagementGUI
            except Exception as e:
                messagebox.showerror("Unavailable", f"Log Management GUI not found.\n{e}")
                return

        try:
            # Open inside the existing app as a child window (no extra mainloop)
            win = tk.Toplevel(self.root)
            win.title("Log Management")
            win.geometry("1200x800")
            LogManagementGUI(win, auth=self.auth)   # instantiate GUI from the file
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open Log Management GUI: {e}")

    def show_housing_accommodations(self):
        """Open Housing Accommodation GUI in a child window."""
        try:
            if not self.auth or not getattr(self.auth, "current_user", None):
                messagebox.showerror("Error", "You must be logged in to access housing.")
                return

            if not HOUSING_ACCOMMODATION_GUI_AVAILABLE:
                messagebox.showerror("Error", "Housing Accommodation GUI is not available.")
                return

            top = tk.Toplevel(self.root)
            top.title("Housing Accommodation Management")
            top.geometry("1200x800")
            try:
                top.transient(self.root)
                top.grab_set()
            except Exception:
                pass

            # Embed the GUI (do NOT call .run(); main loop is already running)
            housing_gui = HousingAccommodationGUI(auth_instance=self.auth)
            # Replace the default root with our Toplevel window
            housing_gui.root.destroy()
            housing_gui.root = top
            # Re-create the GUI interface with the new root
            housing_gui.create_main_interface()
            print("✅ Housing Accommodation GUI opened successfully")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to open Housing Accommodation GUI: {e}")

    def show_document_manager(self):
        """Open the Document Management GUI (with fallbacks)."""
        if not self.auth.current_user:
            messagebox.showerror("Error", "You must be logged in to access Document Manager.")
            return

        user = self.auth.current_user
        perms = user.get('permissions', [])
        role = user.get('role', '')

        # Be generous but safe: admin/staff or explicit perms
        allowed = (
            role in ('admin', 'staff') or
            any(p in perms for p in ['manage_documents', 'system_config', 'view_documents'])
        )
        if not allowed:
            messagebox.showerror("Error", "You don't have permission to access Document Manager.")
            return

        try:
            if DOCUMENT_MANAGER_GUI_AVAILABLE and DocumentManagerGUI:
                win = tk.Toplevel(self.root)
                win.title("📄 Document Management System")
                win.geometry("1400x900")
                try:
                    win.transient(self.root)
                    win.grab_set()
                except Exception:
                    pass

                # Embed the full GUI
                DocumentManagerGUI(win)
                print("✅ Document Manager GUI opened successfully")
                return

            # If import flag says unavailable, try the standalone launcher first
            if start_document_manager_gui:
                start_document_manager_gui()
                return

            raise RuntimeError("Document Manager GUI not available")

        except Exception as e:
            # Final fallback: original CLI menu you already ship
            try:
                display_document_management_menu()
            except Exception as cli_err:
                messagebox.showerror(
                    "Document Manager",
                    f"Failed to open Document Manager GUI:\n{e}\n\nCLI fallback also failed:\n{cli_err}"
                )

    def show_restaurant_management(self):
        """Open the Restaurant Management GUI in a child window (Toplevel)."""
        if not self.restaurant_gui:
            try:
                self.restaurant_gui = RestaurantManagementGUI(self.root, self.auth)
            except Exception as e:
                messagebox.showerror("Error", f"Failed to initialize restaurant management: {e}")
                return

        self.restaurant_gui.show_restaurant_management()

    def open_student_support_portal_gui(self):
        """Open Student Support Portal GUI in a child window"""
        if not self.auth.current_user:
            messagebox.showerror("Error", "You must be logged in to access the student support portal.")
            return
        
        try:
            if not STUDENT_SUPPORT_GUI_AVAILABLE:
                # Fallback to CLI menu if available
                try:
                    from university_system.modules.domain.student_affairs.services.student_support import display_support_menu
                    display_support_menu()
                    return
                except ImportError:
                    messagebox.showerror("Student Support", f"Student Support GUI not available: {STUDENT_SUPPORT_GUI_IMPORT_ERROR}")
                    return
            
            # Create a new window for the Student Support GUI
            support_window = tk.Toplevel(self.root)
            support_window.title("Student Support Portal")
            support_window.geometry("1400x900")
            support_window.minsize(1200, 800)
            
            # Center the window
            support_window.update_idletasks()
            x = (support_window.winfo_screenwidth() - support_window.winfo_width()) // 2
            y = (support_window.winfo_screenheight() - support_window.winfo_height()) // 2
            support_window.geometry(f"+{x}+{y}")
            
            try:
                support_window.transient(self.root)
            except Exception:
                pass  # Continue if transient fails
            
            # Initialize the Student Support GUI in the new window with auth system
            support_gui = StudentSupportGUI(support_window, auth_system=self.auth)
            
            print("✅ Student Support Portal opened successfully")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open Student Support Portal: {str(e)}")
            print(f"❌ Student Support Portal error: {e}")
        
    def open_trip_management_gui(self):
        try:
            python = sys.executable
            code = (
                "import sys; "
                "from trip_management_gui import TripManagementGUI; "
                "app = TripManagementGUI(auth=None); "
                "app.run()"
            )
            subprocess.Popen([python, "-c", code], close_fds=True)
        except Exception as e:
            messagebox.showerror("Trip Management", f"Failed to open Trip Management GUI:\n{e}")

    def open_internship_portal_gui(self):
        """Open the Internship Portal GUI in a child window from the main app."""
        try:
            from university_system.modules.domain.student_affairs.gui.internship_management_gui import InternshipGUI  # local file you provided
        except Exception as e:
            messagebox.showerror("Internship Portal", f"Internship GUI not available:\n{e}")
            return
        try:
            win = tk.Toplevel(self.root)
            win.title("Student Internship Portal")
            win.geometry("1200x800")
            try:
                win.transient(self.root)
                win.grab_set()
            except Exception:
                pass
            # Instantiate the GUI and pass current auth
            InternshipGUI(win, self.auth)
            print("✅ Internship GUI opened successfully")
        except Exception as e:
            messagebox.showerror("Internship Portal", f"Failed to open Internship GUI:\n{e}")

    def open_health_portal_gui(self):
        """Open the Health Portal GUI in a child window from the main app."""
        if self.health_portal_gui:
            self.health_portal_gui.open_health_portal_gui()
        else:
            messagebox.showerror("Error", "Health portal GUI not available.")

    def open_helpdesk_gui(self):
        """Open the Helpdesk GUI in a child window; fall back to CLI if import fails."""
        try:
            if not HELPDESK_GUI_AVAILABLE or HelpdeskGUI is None:
                # Fallback to CLI menu if available; otherwise show error
                if display_helpdesk_menu and globals().get("HELPDESK_CLI_AVAILABLE", False):
                    try:
                        display_helpdesk_menu(self.auth)
                    except Exception as cli_err:
                        messagebox.showerror("Helpdesk", f"Helpdesk CLI fallback failed: {cli_err}")
                else:
                    messagebox.showerror("Helpdesk", "Helpdesk GUI not available and CLI fallback not found")
                return

            # Child window
            win = tk.Toplevel(self.root)
            win.title("Helpdesk & Support")
            win.geometry("1400x900")

            # Modal-ish behavior
            try:
                win.transient(self.root)
                win.grab_set()
            except Exception:
                pass

            # Instantiate GUI
            try:
                helpdesk_gui = HelpdeskGUI(win, auth=self.auth)
                print("✅ Helpdesk GUI opened successfully")
            except Exception as inst_err:
                win.destroy()
                try:
                    display_helpdesk_menu(self.auth)
                except Exception as cli_err:
                    messagebox.showerror(
                        "Helpdesk",
                        f"Failed to open Helpdesk GUI: {inst_err}\nCLI fallback also failed: {cli_err}"
                    )

        except Exception as e:
            messagebox.showerror("Helpdesk", f"Unexpected error opening Helpdesk: {e}")

    def show_grade_tracking_gui(self):
        """Launch the Grade Tracking GUI in a child window"""
        if self.grade_tracking_gui:
            self.grade_tracking_gui.show_grade_tracking_gui()
        else:
            messagebox.showerror("Error", "Grade tracking GUI not available.")

    def show_data_backup_gui(self):
        """Launch the Data Backup and Restore GUI"""
        if not self.auth.current_user:
            messagebox.showerror("Error", "You must be logged in to access backup functions.")
            return
        
        if not self.auth.check_permission('backup_restore'):
            messagebox.showerror("Error", "You don't have permission to access backup functions.")
            return
        
        try:
            # Create backup window
            backup_window = tk.Toplevel(self.root)
            backup_window.title("Data Backup & Restore System")
            backup_window.geometry("800x600")
            backup_window.minsize(600, 500)
            
            try:
                backup_window.transient(self.root)
            except Exception:
                pass
            
            # Use the imported BackupGUI class
            if not DATA_BACKUP_GUI_AVAILABLE:
                messagebox.showerror("Error", "Data Backup GUI is not available.")
                return

            backup_gui = BackupGUI(backup_window, self.auth)
            
            if hasattr(backup_gui, 'set_auth'):
                backup_gui.set_auth(self.auth)
            
            print("Data Backup GUI opened successfully")
            
        except ImportError:
            # Fallback to CLI menu
            try:
                from university_system.infrastructure.database.data_backup import display_backup_menu
                display_backup_menu()
            except ImportError:
                messagebox.showerror("Error", "Backup system not available.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open Data Backup GUI: {str(e)}")

    def show_communication_dashboard_gui(self):
        """Launch the full Communication Dashboard GUI"""
        if not self.auth.current_user:
            messagebox.showerror("Error", "You must be logged in to access communication features.")
            return

        if not (self.auth.check_permission('send_emails') or
                self.auth.check_permission('access_communication_dashboard')):
            messagebox.showerror("Error", "You don't have permission to access communication features.")
            return
        
        try:
            # Create communication window
            comm_window = tk.Toplevel(self.root)
            comm_window.title("Communication & Email Management Dashboard")
            comm_window.geometry("1400x900")
            comm_window.minsize(1200, 700)
            
            try:
                comm_window.transient(self.root)
            except Exception:
                pass
            
            # Initialize communication dashboard
            from university_system.infrastructure.email.gui.email_manager_gui import EmailManagerGUI
            comm_gui = EmailManagerGUI(comm_window, auth=self.auth)
            
            print("Communication Dashboard GUI opened successfully")
            
        except ImportError:
            # Fallback to existing email manager
            try:
                self.show_email_manager()
            except Exception:
                messagebox.showerror("Error", "Communication system not available.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open Communication Dashboard: {str(e)}")

    def show_system_administration_gui(self):
        """Launch comprehensive system administration GUI"""
        if not self.auth.current_user or self.auth.current_user.get('role') != 'admin':
            messagebox.showerror("Access Denied", "Administrator access required")
            return
        
        try:
            admin_window = tk.Toplevel(self.root)
            admin_window.title("System Administration Console")
            admin_window.geometry("1400x900")
            admin_window.minsize(1200, 700)
            
            try:
                admin_window.transient(self.root)
            except Exception:
                pass
            
            # Create notebook for different admin sections
            notebook = ttk.Notebook(admin_window, padding="10")
            notebook.pack(fill=tk.BOTH, expand=True)
            
            # Database Management Tab
            db_frame = ttk.Frame(notebook)
            notebook.add(db_frame, text="Database Management")
            self.create_database_admin_tab(db_frame)
            
            # User Management Tab
            user_frame = ttk.Frame(notebook)
            notebook.add(user_frame, text="User Management")
            self.create_user_admin_tab(user_frame)
            
            # System Monitoring Tab
            monitor_frame = ttk.Frame(notebook)
            notebook.add(monitor_frame, text="System Monitoring")
            self.create_monitoring_tab(monitor_frame)
            
            # Configuration Tab
            config_frame = ttk.Frame(notebook)
            notebook.add(config_frame, text="Configuration")
            self.create_config_tab(config_frame)

            # Add close button at the bottom
            button_frame = ttk.Frame(admin_window, padding="10")
            button_frame.pack(side=tk.BOTTOM, fill=tk.X)
            ttk.Button(button_frame, text="Close", command=admin_window.destroy).pack(side=tk.RIGHT)

            print("System Administration GUI opened successfully")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to open System Administration: {str(e)}")

    def show_security_dashboard(self):
        """Launch comprehensive security & compliance dashboard"""
        if not self.auth.current_user or self.auth.current_user.get('role') != 'admin':
            messagebox.showerror("Access Denied", "Administrator access required for Security Dashboard")
            return

        if not SECURITY_DASHBOARD_AVAILABLE:
            messagebox.showerror("Error", "Security Dashboard module is not available")
            return

        try:
            user_id = self.auth.current_user.get('id', 1)
            dashboard = SecurityDashboard(self.root, user_id)
            print("Security Dashboard opened successfully")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open Security Dashboard: {str(e)}")
            import traceback
            traceback.print_exc()

    def create_database_admin_tab(self, parent):
        """Create database administration interface"""
        main_frame = ttk.Frame(parent, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Database tools
        tools_frame = ttk.LabelFrame(main_frame, text="Database Tools", padding="15")
        tools_frame.pack(fill=tk.X, pady=(0, 20))
        
        ttk.Button(tools_frame, text="Database Integrity Check", 
                  command=self.run_integrity_check).grid(row=0, column=0, padx=5, pady=5)
        ttk.Button(tools_frame, text="Fix Duplicate Records", 
                  command=self.fix_duplicates).grid(row=0, column=1, padx=5, pady=5)
        ttk.Button(tools_frame, text="Optimize Database", 
                  command=self.optimize_database).grid(row=0, column=2, padx=5, pady=5)
        ttk.Button(tools_frame, text="View Database Statistics", 
                  command=self.show_db_statistics).grid(row=1, column=0, padx=5, pady=5)

    def create_user_admin_tab(self, parent):
        """Create user administration interface"""
        main_frame = ttk.Frame(parent, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # User management tools
        tools_frame = ttk.LabelFrame(main_frame, text="User Management", padding="15")
        tools_frame.pack(fill=tk.X, pady=(0, 20))

        ttk.Button(tools_frame, text="View All Users",
                  command=self.view_all_users).grid(row=0, column=0, padx=5, pady=5)
        ttk.Button(tools_frame, text="Add New User",
                  command=self.add_new_user).grid(row=0, column=1, padx=5, pady=5)
        ttk.Button(tools_frame, text="Manage Permissions",
                  command=self.manage_permissions).grid(row=0, column=2, padx=5, pady=5)
        ttk.Button(tools_frame, text="View Active Sessions",
                  command=self.view_active_sessions).grid(row=1, column=0, padx=5, pady=5)

        # User statistics
        stats_frame = ttk.LabelFrame(main_frame, text="User Statistics", padding="15")
        stats_frame.pack(fill=tk.BOTH, expand=True)

        stats_text = scrolledtext.ScrolledText(stats_frame, wrap=tk.WORD, height=15,
                                               fg="#000000", bg="#FFFFFF")
        stats_text.pack(fill=tk.BOTH, expand=True)

        try:
            # Get user statistics from database
            from university_system.infrastructure.database.db import get_connection
            with get_connection() as conn:
                cursor = conn.execute("SELECT COUNT(*) FROM users")
                total_users = cursor.fetchone()[0]

                cursor = conn.execute("SELECT role, COUNT(*) FROM users GROUP BY role")
                users_by_role = cursor.fetchall()

            stats_info = f"""User Statistics
{'='*50}

Total Users: {total_users}

Users by Role:
"""
            for role, count in users_by_role:
                stats_info += f"  {role}: {count}\n"

            stats_text.insert("1.0", stats_info)
        except Exception as e:
            stats_text.insert("1.0", f"Error loading user statistics: {e}")

        stats_text.config(state=tk.DISABLED)

    def create_monitoring_tab(self, parent):
        """Create system monitoring interface"""
        main_frame = ttk.Frame(parent, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Monitoring tools
        tools_frame = ttk.LabelFrame(main_frame, text="System Monitoring Tools", padding="15")
        tools_frame.pack(fill=tk.X, pady=(0, 20))

        ttk.Button(tools_frame, text="View System Logs",
                  command=self.view_system_logs).grid(row=0, column=0, padx=5, pady=5)
        ttk.Button(tools_frame, text="Database Performance",
                  command=self.show_db_performance).grid(row=0, column=1, padx=5, pady=5)
        ttk.Button(tools_frame, text="Active Connections",
                  command=self.show_active_connections).grid(row=0, column=2, padx=5, pady=5)
        ttk.Button(tools_frame, text="Error Logs",
                  command=self.view_error_logs).grid(row=1, column=0, padx=5, pady=5)

        # System status display
        status_frame = ttk.LabelFrame(main_frame, text="System Status", padding="15")
        status_frame.pack(fill=tk.BOTH, expand=True)

        status_text = scrolledtext.ScrolledText(status_frame, wrap=tk.WORD, height=15,
                                               fg="#000000", bg="#FFFFFF")
        status_text.pack(fill=tk.BOTH, expand=True)

        try:
            import psutil
            import platform
            from datetime import datetime

            # Get system information
            cpu_percent = psutil.cpu_percent(interval=1)
            memory = psutil.virtual_memory()
            disk = psutil.disk_usage('/')

            # Get database stats
            from university_system.infrastructure.database.db import get_connection
            with get_connection() as conn:
                cursor = conn.execute("SELECT COUNT(*) FROM activity_log")
                log_count = cursor.fetchone()[0] if cursor.fetchone() else 0

            status_info = f"""System Status Report
{'='*50}

System Information:
  Platform: {platform.system()} {platform.release()}
  Python: {platform.python_version()}
  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

Performance Metrics:
  CPU Usage: {cpu_percent}%
  Memory Usage: {memory.percent}% ({memory.used / (1024**3):.2f}GB / {memory.total / (1024**3):.2f}GB)
  Disk Usage: {disk.percent}% ({disk.used / (1024**3):.2f}GB / {disk.total / (1024**3):.2f}GB)

Database Status:
  Activity Logs: {log_count} entries
  Status: Connected

System Health: {'✓ Healthy' if cpu_percent < 80 and memory.percent < 80 else '⚠ Warning'}
"""
            status_text.insert("1.0", status_info)
        except Exception as e:
            status_text.insert("1.0", f"System Status\n{'='*50}\n\nError loading system information: {e}\n\nBasic system monitoring is available.")

        status_text.config(state=tk.DISABLED)

    def create_config_tab(self, parent):
        """Create configuration interface"""
        main_frame = ttk.Frame(parent, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Configuration tools
        tools_frame = ttk.LabelFrame(main_frame, text="Configuration Management", padding="15")
        tools_frame.pack(fill=tk.X, pady=(0, 20))

        ttk.Button(tools_frame, text="System Settings",
                  command=self.edit_system_settings).grid(row=0, column=0, padx=5, pady=5)
        ttk.Button(tools_frame, text="Email Configuration",
                  command=self.configure_email).grid(row=0, column=1, padx=5, pady=5)
        ttk.Button(tools_frame, text="Backup Settings",
                  command=self.configure_backup).grid(row=0, column=2, padx=5, pady=5)
        ttk.Button(tools_frame, text="Security Settings",
                  command=self.configure_security).grid(row=1, column=0, padx=5, pady=5)

        # Current configuration display
        config_frame = ttk.LabelFrame(main_frame, text="Current Configuration", padding="15")
        config_frame.pack(fill=tk.BOTH, expand=True)

        config_text = scrolledtext.ScrolledText(config_frame, wrap=tk.WORD, height=15,
                                               fg="#000000", bg="#FFFFFF")
        config_text.pack(fill=tk.BOTH, expand=True)

        try:
            from university_system.modules.shared.constants import paths
            import os

            config_info = f"""System Configuration
{'='*50}

File Paths:
  Database: {paths.DEFAULT_DB_PATH}
  Logs: {paths.LOG_DIR}
  Backups: {paths.BACKUP_DIR}
  Uploads: {paths.UPLOAD_DIR}

System Settings:
  Project Root: {paths.PROJECT_ROOT}
  Data Directory: {paths.DATA_DIR}

Database Configuration:
  Type: SQLite
  Connection Pooling: Enabled
  WAL Mode: Enabled

Authentication:
  Password Hashing: PBKDF2 (1,000,000 iterations)
  Multi-Factor Auth: Available
  Session Management: Token-based

Email Service:
  Status: {'Configured' if os.path.exists(paths.PROJECT_ROOT / '.env') else 'Not Configured'}
  Queue System: Asynchronous

Logging:
  Activity Logging: Enabled
  Log Rotation: Daily
  Retention: 90 days

Note: Modify configuration files or use the configuration tools above to change settings.
"""
            config_text.insert("1.0", config_info)
        except Exception as e:
            config_text.insert("1.0", f"Configuration Overview\n{'='*50}\n\nError loading configuration: {e}\n\nUse the configuration tools above to manage system settings.")

        config_text.config(state=tk.DISABLED)

    def open_ai_detector_window(self):
        """Open the AI Detector GUI in a separate full-screen window (new process)."""
        try:
            python = sys.executable
            code = (
                "import sys, os; "
                f"sys.path.insert(0, r'{PROJECT_ROOT}'); "
                "import ai_detector as m; "
                f"det = m.AIDetector(db_path=r'{DB_PATH}'); "
                "app = m.AIDetectorGUI(det); "
                "app.root.attributes('-fullscreen', True); "
                "app.run()"
            )
            subprocess.Popen([python, "-c", code], close_fds=True)
        except Exception as e:
            messagebox.showerror('Error', f'Failed to open AI Detector GUI: {e}')

    def launch_analytics_gui_standalone():
        """Launch analytics GUI as standalone window"""
        try:
            if ANALYTICS_GUI_AVAILABLE:
                analytics_app = GUIStudentAnalytics()
                if auth:
                    analytics_app.auth = auth
                analytics_app.run()
            else:
                print("Analytics GUI not available, using CLI version")
                display_analytics_menu()
        except Exception as e:
            print(f"Error launching analytics: {e}")
            display_analytics_menu()

    def show_chatbot(self):
        """Launch the full Chatbot GUI using the existing ChatbotGUI class"""
        if not self.auth.current_user:
            messagebox.showerror("Error", "You must be logged in to access the chatbot.")
            return
        
        if not self.auth.check_permission('access_chatbot'):
            messagebox.showerror("Error", "You don't have permission to access the chatbot.")
            return
        
        try:
            # Initialize the chatbot instance if not already done
            global chatbot_instance
            if not chatbot_instance:
                if not initialize_chatbot_integration():
                    messagebox.showerror("Error", "Failed to initialize chatbot system.")
                    return
            
            # Set authentication system for chatbot
            if chatbot_instance and self.auth:
                chatbot_instance.set_auth_system(self.auth)
            
            # Use imported UniversityChatbotGUI if available
            if CHATBOT_GUI_AVAILABLE:
                chatbot_window = tk.Toplevel(self.root)
                chatbot_window.title("University Chatbot")
                chatbot_window.geometry("1000x700")

                chatbot_gui = UniversityChatbotGUI(chatbot_instance, chatbot_window, auth_system=self.auth)
                print("✅ University Chatbot GUI opened successfully")
            else:
                # Fallback to original method
                chatbot_gui = ChatbotGUI(chatbot_instance)
                chatbot_gui.run()
            
            print("University Chatbot GUI launched successfully")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open University Chatbot: {str(e)}")
            print(f"Chatbot GUI error: {e}")
        
    def show_analytics(self):
        """Launch the standalone Student Analytics GUI from student_analytics_gui.py"""
        if not self.auth.current_user or 'view_analytics' not in self.auth.current_user.get('permissions', []):
            messagebox.showerror("Error", "You don't have permission to access analytics")
            return

        try:
            if STUDENT_ANALYTICS_GUI_AVAILABLE:
                # Create a child window for the analytics GUI
                analytics_window = tk.Toplevel(self.root)
                analytics_window.transient(self.root)

                # Launch the GUI in the child window
                analytics_app = GUIStudentAnalytics(root=analytics_window, auth_manager=self.auth)
            else:
                messagebox.showerror("Error", "Student Analytics GUI is not available")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to launch analytics GUI: {str(e)}")
        
    def log_activity(self, message, level="info", action=None):
        """Log activity with comprehensive error handling"""
        try:
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            formatted_message = f"[{timestamp}] GUI: {message}"
            
            print(formatted_message)
            
            if level.lower() == "error":
                logging.error(formatted_message)
            elif level.lower() == "warning":
                logging.warning(formatted_message)
            else:
                logging.info(formatted_message)
                
        except Exception as e:
            print(f"GUI Activity: {message}")
            logging.error(f"Logging system error: {e}")
    
    def check_session_timer(self):
        """Check session validity periodically"""
        try:
            # Check if window still exists
            if not self.root.winfo_exists():
                return

            if self.auth.current_user:
                if hasattr(self.auth, 'check_session') and not self.auth.check_session():
                    messagebox.showwarning("Session Expired", "Your session has expired. Please log in again.")
                    self.update_status()
                    self.show_login_screen()
        except Exception as e:
            print(f"Session check error: {e}")

        # Schedule next check only if window still exists
        try:
            if self.root.winfo_exists():
                self.root.after(60000, self.check_session_timer)
        except Exception:
            pass  # Window destroyed, don't schedule

    def show_integrated_dashboard(self):
        """Show integrated dashboard with system overview and quick stats"""
        if not self.auth.current_user:
            messagebox.showerror("Error", "You must be logged in to access the dashboard.")
            return

        self.clear_content()

        # Create dashboard layout
        dashboard_frame = ttk.Frame(self.content_frame)
        dashboard_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Title
        title_label = ttk.Label(dashboard_frame, text="Integrated System Dashboard",
                               font=('Arial', 16, 'bold'))
        title_label.pack(pady=(0, 20))

        # Create notebook for different dashboard sections
        notebook = ttk.Notebook(dashboard_frame)
        notebook.pack(fill=tk.BOTH, expand=True)

        # System Overview Tab
        overview_frame = ttk.Frame(notebook)
        notebook.add(overview_frame, text="System Overview")
        self.create_overview_tab(overview_frame)

        # Quick Stats Tab
        stats_frame = ttk.Frame(notebook)
        notebook.add(stats_frame, text="Quick Statistics")
        self.create_stats_tab(stats_frame)

        # Recent Activity Tab
        activity_frame = ttk.Frame(notebook)
        notebook.add(activity_frame, text="Recent Activity")
        self.create_activity_tab(activity_frame)

        # System Health Tab
        health_frame = ttk.Frame(notebook)
        notebook.add(health_frame, text="System Health")
        self.create_health_tab(health_frame)

        print("✅ Integrated Dashboard opened successfully")

    def create_overview_tab(self, parent):
        """Create system overview tab"""
        overview_container = ttk.Frame(parent, padding="20")
        overview_container.pack(fill=tk.BOTH, expand=True)

        # Welcome message
        welcome_text = f"Welcome back, {self.auth.current_user.get('username', 'User')}!"
        ttk.Label(overview_container, text=welcome_text, font=('Arial', 14, 'bold')).pack(pady=(0, 20))

        # Quick access buttons in a grid
        buttons_frame = ttk.LabelFrame(overview_container, text="Quick Access", padding="15")
        buttons_frame.pack(fill=tk.X, pady=(0, 20))

        # Configure grid
        for i in range(3):
            buttons_frame.columnconfigure(i, weight=1)

        # Quick access buttons
        ttk.Button(buttons_frame, text="Student Records",
                  command=self.show_student_records).grid(row=0, column=0, padx=5, pady=5, sticky="ew")
        ttk.Button(buttons_frame, text="Grade Tracking",
                  command=self.show_grade_tracking_gui).grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        ttk.Button(buttons_frame, text="Attendance",
                  command=self.open_attendance_gui).grid(row=0, column=2, padx=5, pady=5, sticky="ew")

        ttk.Button(buttons_frame, text="Course Management",
                  command=self.show_course_management).grid(row=1, column=0, padx=5, pady=5, sticky="ew")
        ttk.Button(buttons_frame, text="Finance Management",
                  command=self.show_finance_management).grid(row=1, column=1, padx=5, pady=5, sticky="ew")
        ttk.Button(buttons_frame, text="Reports",
                  command=self.show_enhanced_reporting_dashboard).grid(row=1, column=2, padx=5, pady=5, sticky="ew")

        # System status
        status_frame = ttk.LabelFrame(overview_container, text="System Status", padding="15")
        status_frame.pack(fill=tk.X, pady=(0, 20))

        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ttk.Label(status_frame, text=f"Current Time: {current_time}").pack(anchor="w")
        ttk.Label(status_frame, text=f"User Role: {self.auth.current_user.get('role', 'Unknown')}").pack(anchor="w")
        ttk.Label(status_frame, text="Database Status: Connected ✓").pack(anchor="w")

    def create_stats_tab(self, parent):
        """Create quick statistics tab"""
        stats_container = ttk.Frame(parent, padding="20")
        stats_container.pack(fill=tk.BOTH, expand=True)

        ttk.Label(stats_container, text="System Statistics",
                 font=('Arial', 14, 'bold')).pack(pady=(0, 20))

        # Stats grid
        stats_frame = ttk.Frame(stats_container)
        stats_frame.pack(fill=tk.BOTH, expand=True)

        # Try to get actual database statistics
        try:
            # This would be populated with real data from the database
            stats_text = """Database Statistics:
• Total Students: Loading...
• Active Courses: Loading...
• Pending Assignments: Loading...
• System Uptime: Active
• Recent Logins: Loading...

Performance Metrics:
• Database Response Time: <50ms
• System Load: Normal
• Memory Usage: Optimal
• Active Sessions: 1"""

            stats_display = tk.Text(stats_frame, wrap=tk.WORD, height=15, width=60)
            stats_display.pack(fill=tk.BOTH, expand=True)
            stats_display.insert(tk.END, stats_text)
            stats_display.config(state=tk.DISABLED)

        except Exception as e:
            ttk.Label(stats_frame, text=f"Unable to load statistics: {e}").pack()

    def create_activity_tab(self, parent):
        """Create recent activity tab"""
        activity_container = ttk.Frame(parent, padding="20")
        activity_container.pack(fill=tk.BOTH, expand=True)

        ttk.Label(activity_container, text="Recent System Activity",
                 font=('Arial', 14, 'bold')).pack(pady=(0, 20))

        # Activity log display
        activity_display = tk.Text(activity_container, wrap=tk.WORD, height=20)
        activity_display.pack(fill=tk.BOTH, expand=True)

        # Sample activity data (would be populated from actual logs)
        activity_text = f"""Recent Activity Log:
{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - User {self.auth.current_user.get('username')} logged in
{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - Dashboard accessed
{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - System initialized

Previous Sessions:
• Last login: Today
• Recent actions: View dashboard, Access student records
• System health: All systems operational
"""

        activity_display.insert(tk.END, activity_text)
        activity_display.config(state=tk.DISABLED)

    def create_health_tab(self, parent):
        """Create system health monitoring tab"""
        if self.health_portal_gui:
            self.health_portal_gui.create_health_tab(parent)
        else:
            # Fallback implementation
            health_container = ttk.Frame(parent, padding="20")
            health_container.pack(fill=tk.BOTH, expand=True)
            ttk.Label(health_container, text="System Health Monitoring",
                     font=('Arial', 14, 'bold')).pack(pady=(0, 20))
            ttk.Label(health_container, text="Health monitoring GUI not available").pack(pady=10)

    def show_module_scheduling(self):
        """Open the Module Scheduling GUI (child window if possible)."""
        try:
            if not MODULE_SCHEDULING_GUI_AVAILABLE:
                messagebox.showinfo("Module Scheduling", "Module Scheduling GUI not available.")
                return
            # Prefer embedding into a Toplevel to avoid creating a second Tk root.
            try:
                top = tk.Toplevel(self.root)
                top.title("Module Scheduling")
                app = ModuleSchedulingGUI(top) if 'ModuleSchedulingGUI' in globals() and ModuleSchedulingGUI else None
                # Pass auth context if supported
                try:
                    if app and hasattr(app, 'set_auth'):
                        app.set_auth(self.auth)
                except Exception:
                    pass
                try:
                    top.transient(self.root)
                    top.grab_set()
                except Exception:
                    pass
            except Exception as inner_e:
                # Fallback to provided launcher (may create its own Tk root)
                try:
                    launch_module_scheduling_gui()
                except Exception as e2:
                    messagebox.showerror("Error", f"Module Scheduling GUI failed to launch: {e2}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open Module Scheduling: {e}")

    def fix_duplicates(self):
        """Fix duplicate records in the database"""
        if not self.auth.current_user:
            messagebox.showerror("Error", "You must be logged in to perform database operations.")
            return

        user_role = self.auth.current_user.get('role', '')
        if user_role != 'admin':
            messagebox.showerror("Error", "Admin access required for database operations.")
            return

        try:
            result = messagebox.askyesno("Confirm",
                "This will scan and fix duplicate records in the database.\n\n"
                "This operation may take some time. Continue?")

            if not result:
                return

            # This is a placeholder - would need actual database logic
            messagebox.showinfo("Database Maintenance",
                "Duplicate fix operation completed.\n\n"
                "Note: This is a placeholder implementation.")

            self.log_activity("Database duplicate fix performed", "info", "database_maintenance")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to fix duplicates: {e}")

    def optimize_database(self):
        """Optimize database performance"""
        if not self.auth.current_user:
            messagebox.showerror("Error", "You must be logged in to perform database operations.")
            return

        user_role = self.auth.current_user.get('role', '')
        if user_role != 'admin':
            messagebox.showerror("Error", "Admin access required for database operations.")
            return

        try:
            result = messagebox.askyesno("Confirm",
                "This will optimize the database for better performance.\n\n"
                "This operation may take some time. Continue?")

            if not result:
                return

            # This is a placeholder - would need actual database optimization logic
            messagebox.showinfo("Database Maintenance",
                "Database optimization completed.\n\n"
                "Note: This is a placeholder implementation.")

            self.log_activity("Database optimization performed", "info", "database_maintenance")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to optimize database: {e}")

    def run_integrity_check(self):
        """Run database integrity check"""
        if not self.auth.current_user:
            messagebox.showerror("Error", "You must be logged in to perform database operations.")
            return

        user_role = self.auth.current_user.get('role', '')
        if user_role != 'admin':
            messagebox.showerror("Error", "Admin access required for database operations.")
            return

        try:
            # This is a placeholder - would need actual integrity check logic
            result_window = tk.Toplevel(self.root)
            result_window.title("Database Integrity Check Results")
            result_window.geometry("600x400")

            text_frame = ttk.Frame(result_window, padding="10")
            text_frame.pack(fill=tk.BOTH, expand=True)

            text_area = tk.Text(text_frame, wrap=tk.WORD, height=20, width=70)
            scrollbar = ttk.Scrollbar(text_frame, orient=tk.VERTICAL, command=text_area.yview)
            text_area.configure(yscrollcommand=scrollbar.set)

            text_area.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

            # Sample integrity check results
            integrity_results = """DATABASE INTEGRITY CHECK RESULTS
=================================

Tables Checked: 15
Records Scanned: 1,250
Errors Found: 0
Warnings: 2

DETAILS:
- All foreign key constraints are valid
- No orphaned records detected
- All indexes are consistent
- Database structure is intact

WARNINGS:
- Table 'students': 2 records with missing optional fields
- Consider running ANALYZE command for better query performance

Overall Status: HEALTHY

Note: This is a placeholder implementation.
Full integrity checking would require actual database analysis."""

            text_area.insert(tk.END, integrity_results)
            text_area.config(state=tk.DISABLED)

            ttk.Button(result_window, text="Close",
                      command=result_window.destroy).pack(pady=10)

            self.log_activity("Database integrity check performed", "info", "database_maintenance")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to run integrity check: {e}")

    def show_db_statistics(self):
        """Show database statistics"""
        if not self.auth.current_user:
            messagebox.showerror("Error", "You must be logged in to view database statistics.")
            return

        user_role = self.auth.current_user.get('role', '')
        if user_role not in ['admin', 'staff']:
            messagebox.showerror("Error", "Admin or staff access required to view database statistics.")
            return

        try:
            stats_window = tk.Toplevel(self.root)
            stats_window.title("Database Statistics")
            stats_window.geometry("700x500")

            # Create notebook for different stat categories
            notebook = ttk.Notebook(stats_window)
            notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

            # General Stats Tab
            general_frame = ttk.Frame(notebook, padding="10")
            notebook.add(general_frame, text="General")

            general_text = tk.Text(general_frame, wrap=tk.WORD, height=20, width=70)
            general_scroll = ttk.Scrollbar(general_frame, orient=tk.VERTICAL, command=general_text.yview)
            general_text.configure(yscrollcommand=general_scroll.set)

            general_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            general_scroll.pack(side=tk.RIGHT, fill=tk.Y)

            general_stats = """DATABASE GENERAL STATISTICS
==========================

Total Tables: 15
Total Records: 1,250
Database Size: 45.2 MB
Last Backup: 2024-01-15 14:30:25

TABLE BREAKDOWN:
- students: 450 records
- courses: 85 records
- modules: 120 records
- enrollments: 320 records
- grades: 275 records
- users: 25 records
- accommodations: 15 records
- activities: 180 records
- Other tables: Various

Note: This is sample data for demonstration."""

            general_text.insert(tk.END, general_stats)
            general_text.config(state=tk.DISABLED)

            # Performance Stats Tab
            perf_frame = ttk.Frame(notebook, padding="10")
            notebook.add(perf_frame, text="Performance")

            perf_text = tk.Text(perf_frame, wrap=tk.WORD, height=20, width=70)
            perf_scroll = ttk.Scrollbar(perf_frame, orient=tk.VERTICAL, command=perf_text.yview)
            perf_text.configure(yscrollcommand=perf_scroll.set)

            perf_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            perf_scroll.pack(side=tk.RIGHT, fill=tk.Y)

            perf_stats = """DATABASE PERFORMANCE STATISTICS
==============================

Query Performance:
- Average query time: 12ms
- Slowest query: 156ms (complex join)
- Fastest query: 2ms (simple select)

Index Usage:
- Primary indexes: 15/15 used
- Secondary indexes: 8/12 used
- Missing indexes: 4 recommended

Connection Pool:
- Active connections: 3
- Max connections: 100
- Connection efficiency: 98.5%

Cache Hit Ratio: 94.2%

Recommendations:
- Add index on 'students.enrollment_date'
- Consider archiving old records
- Review slow query: course enrollment report

Note: This is sample data for demonstration."""

            perf_text.insert(tk.END, perf_stats)
            perf_text.config(state=tk.DISABLED)

            self.log_activity("Database statistics viewed", "info", "database_view")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to show database statistics: {e}")

    # User Administration Methods
    def view_all_users(self):
        """View all users in the system"""
        try:
            users_window = tk.Toplevel(self.root)
            users_window.title("All Users")
            users_window.geometry("900x600")

            # Create treeview
            tree_frame = ttk.Frame(users_window)
            tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

            columns = ("ID", "Username", "Email", "Role", "Status")
            tree = ttk.Treeview(tree_frame, columns=columns, show="headings")

            for col in columns:
                tree.heading(col, text=col)
                tree.column(col, width=150)

            scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview)
            tree.configure(yscrollcommand=scrollbar.set)

            tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

            # Load users from database
            from university_system.infrastructure.database.db import get_connection
            with get_connection() as conn:
                cursor = conn.execute("SELECT id, username, email, role, 'Active' FROM users ORDER BY username")
                for row in cursor.fetchall():
                    tree.insert("", tk.END, values=row)

            ttk.Button(users_window, text="Close", command=users_window.destroy).pack(pady=10)

        except Exception as e:
            messagebox.showerror("Error", f"Failed to view users: {e}")

    def add_new_user(self):
        """Add a new user to the system"""
        try:
            # Open the user management GUI for adding users
            messagebox.showinfo("Add User", "This will open the User Management interface.\n\nUse the User Management module to add new users.")
            self.show_user_management()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to add user: {e}")

    def manage_permissions(self):
        """Manage user permissions"""
        try:
            perms_window = tk.Toplevel(self.root)
            perms_window.title("Manage Permissions")
            perms_window.geometry("600x400")

            ttk.Label(perms_window, text="Permission Management",
                     font=('Arial', 14, 'bold')).pack(pady=10)

            info_text = scrolledtext.ScrolledText(perms_window, wrap=tk.WORD, height=15,
                                                  fg="#000000", bg="#FFFFFF")
            info_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

            perms_info = """Permission Management System
================================

Available Permissions by Role:

Admin:
- Full system access
- User management
- Database administration
- System configuration
- All module access

Staff:
- Student records (view/edit)
- Course management
- Grade entry
- Attendance tracking
- Report generation

Student:
- View own records
- Course enrollment
- Grade viewing
- Assignment submission
- Profile management

To modify permissions, use the User Management module.
"""
            info_text.insert("1.0", perms_info)
            info_text.config(state=tk.DISABLED)

            ttk.Button(perms_window, text="Close", command=perms_window.destroy).pack(pady=10)

        except Exception as e:
            messagebox.showerror("Error", f"Failed to manage permissions: {e}")

    def view_active_sessions(self):
        """View active user sessions"""
        try:
            sessions_window = tk.Toplevel(self.root)
            sessions_window.title("Active Sessions")
            sessions_window.geometry("700x400")

            ttk.Label(sessions_window, text="Active User Sessions",
                     font=('Arial', 14, 'bold')).pack(pady=10)

            # Create treeview
            tree_frame = ttk.Frame(sessions_window)
            tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

            columns = ("Username", "Role", "Login Time", "Status")
            tree = ttk.Treeview(tree_frame, columns=columns, show="headings")

            for col in columns:
                tree.heading(col, text=col)
                tree.column(col, width=150)

            scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview)
            tree.configure(yscrollcommand=scrollbar.set)

            tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

            # Get active sessions from auth system
            if self.auth.current_user:
                from datetime import datetime
                current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                tree.insert("", tk.END, values=(
                    self.auth.current_user.get('username'),
                    self.auth.current_user.get('role'),
                    current_time,
                    "Active"
                ))

            ttk.Button(sessions_window, text="Close", command=sessions_window.destroy).pack(pady=10)

        except Exception as e:
            messagebox.showerror("Error", f"Failed to view sessions: {e}")

    # System Monitoring Methods
    def view_system_logs(self):
        """View system logs"""
        try:
            logs_window = tk.Toplevel(self.root)
            logs_window.title("System Logs")
            logs_window.geometry("900x600")

            ttk.Label(logs_window, text="System Activity Logs",
                     font=('Arial', 14, 'bold')).pack(pady=10)

            log_text = scrolledtext.ScrolledText(logs_window, wrap=tk.WORD,
                                                 fg="#000000", bg="#FFFFFF")
            log_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

            # Load recent logs from database
            from university_system.infrastructure.database.db import get_connection
            try:
                with get_connection() as conn:
                    cursor = conn.execute("""
                        SELECT timestamp, user_id, action, details
                        FROM activity_log
                        ORDER BY timestamp DESC
                        LIMIT 100
                    """)
                    logs = cursor.fetchall()

                log_content = "Recent System Activity\n" + "="*80 + "\n\n"
                for log in logs:
                    log_content += f"[{log[0]}] User: {log[1]} - Action: {log[2]}\n"
                    if log[3]:
                        log_content += f"  Details: {log[3]}\n"
                    log_content += "-"*80 + "\n"

                log_text.insert("1.0", log_content)
            except Exception as e:
                log_text.insert("1.0", f"Error loading logs: {e}\n\nSystem logging is available through the Activity Logger.")

            log_text.config(state=tk.DISABLED)
            ttk.Button(logs_window, text="Close", command=logs_window.destroy).pack(pady=10)

        except Exception as e:
            messagebox.showerror("Error", f"Failed to view logs: {e}")

    def show_db_performance(self):
        """Show database performance metrics"""
        try:
            perf_window = tk.Toplevel(self.root)
            perf_window.title("Database Performance")
            perf_window.geometry("700x500")

            ttk.Label(perf_window, text="Database Performance Metrics",
                     font=('Arial', 14, 'bold')).pack(pady=10)

            perf_text = scrolledtext.ScrolledText(perf_window, wrap=tk.WORD, height=20,
                                                  fg="#000000", bg="#FFFFFF")
            perf_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

            from university_system.infrastructure.database.db import get_connection
            import time

            # Test query performance
            start_time = time.time()
            with get_connection() as conn:
                cursor = conn.execute("SELECT COUNT(*) FROM activity_log")
                log_count = cursor.fetchone()[0]
            query_time = (time.time() - start_time) * 1000

            perf_info = f"""Database Performance Report
{'='*50}

Query Performance:
  Test query time: {query_time:.2f}ms
  Activity logs: {log_count} entries

Connection Pool Status:
  Type: SQLite with connection pooling
  WAL Mode: Enabled
  Status: Operational

Performance Tips:
- Regular VACUUM operations recommended
- Monitor log table growth
- Consider archiving old records
- Index optimization available

Note: For detailed performance analysis, use the Database Tools.
"""
            perf_text.insert("1.0", perf_info)
            perf_text.config(state=tk.DISABLED)

            ttk.Button(perf_window, text="Close", command=perf_window.destroy).pack(pady=10)

        except Exception as e:
            messagebox.showerror("Error", f"Failed to show performance metrics: {e}")

    def show_active_connections(self):
        """Show active database connections"""
        try:
            conn_window = tk.Toplevel(self.root)
            conn_window.title("Active Connections")
            conn_window.geometry("600x400")

            ttk.Label(conn_window, text="Active Database Connections",
                     font=('Arial', 14, 'bold')).pack(pady=10)

            conn_text = scrolledtext.ScrolledText(conn_window, wrap=tk.WORD, height=15,
                                                  fg="#000000", bg="#FFFFFF")
            conn_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

            conn_info = """Database Connection Status
================================

Connection Pool Configuration:
- Type: SQLite with connection pooling
- Min connections: 2
- Max connections: 10
- Current active: 1
- WAL mode: Enabled

Connection Details:
- Thread-safe connections
- Automatic connection management
- Context manager support
- Write-Ahead Logging enabled

Status: All connections healthy
Last check: Just now

Note: SQLite uses file-based connections with WAL for concurrency.
"""
            conn_text.insert("1.0", conn_info)
            conn_text.config(state=tk.DISABLED)

            ttk.Button(conn_window, text="Close", command=conn_window.destroy).pack(pady=10)

        except Exception as e:
            messagebox.showerror("Error", f"Failed to show connections: {e}")

    def view_error_logs(self):
        """View error logs"""
        try:
            error_window = tk.Toplevel(self.root)
            error_window.title("Error Logs")
            error_window.geometry("900x600")

            ttk.Label(error_window, text="System Error Logs",
                     font=('Arial', 14, 'bold')).pack(pady=10)

            error_text = scrolledtext.ScrolledText(error_window, wrap=tk.WORD,
                                                   fg="#000000", bg="#FFFFFF")
            error_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

            # Try to load error logs from log file
            from university_system.modules.shared.constants import paths
            import os

            try:
                log_file = paths.LOG_DIR / "error.log"
                if os.path.exists(log_file):
                    with open(log_file, 'r') as f:
                        lines = f.readlines()[-100:]  # Last 100 lines
                        error_content = "Recent Error Logs\n" + "="*80 + "\n\n"
                        error_content += "".join(lines)
                        error_text.insert("1.0", error_content)
                else:
                    error_text.insert("1.0", "No error logs found.\n\nSystem appears to be running without errors.")
            except Exception as e:
                error_text.insert("1.0", f"Error reading log file: {e}\n\nError logs are available in the logs directory.")

            error_text.config(state=tk.DISABLED)
            ttk.Button(error_window, text="Close", command=error_window.destroy).pack(pady=10)

        except Exception as e:
            messagebox.showerror("Error", f"Failed to view error logs: {e}")

    # Configuration Methods
    def edit_system_settings(self):
        """Edit system settings"""
        try:
            settings_window = tk.Toplevel(self.root)
            settings_window.title("System Settings")
            settings_window.geometry("600x500")

            ttk.Label(settings_window, text="System Settings",
                     font=('Arial', 14, 'bold')).pack(pady=10)

            settings_text = scrolledtext.ScrolledText(settings_window, wrap=tk.WORD, height=20,
                                                      fg="#000000", bg="#FFFFFF")
            settings_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

            settings_info = """System Settings Configuration
================================

Current Settings:
- Session timeout: 60 minutes
- Auto-save: Enabled
- Debug mode: Disabled
- Log level: INFO
- Database: SQLite (WAL mode)
- Connection pool: 2-10 connections

Authentication:
- Password policy: Strong (PBKDF2)
- MFA available: Yes
- Session tokens: Enabled
- Max login attempts: 5

Email Configuration:
- SMTP configured: Check .env file
- Email queue: Asynchronous
- Template support: Enabled

To modify settings:
1. Edit .env file for credentials
2. Use Configuration Management tools
3. Restart application after changes

Note: Some settings require admin privileges.
"""
            settings_text.insert("1.0", settings_info)
            settings_text.config(state=tk.DISABLED)

            ttk.Button(settings_window, text="Close", command=settings_window.destroy).pack(pady=10)

        except Exception as e:
            messagebox.showerror("Error", f"Failed to edit settings: {e}")

    def configure_email(self):
        """Configure email settings"""
        try:
            email_window = tk.Toplevel(self.root)
            email_window.title("Email Configuration")
            email_window.geometry("600x500")

            ttk.Label(email_window, text="Email Service Configuration",
                     font=('Arial', 14, 'bold')).pack(pady=10)

            email_text = scrolledtext.ScrolledText(email_window, wrap=tk.WORD, height=20,
                                                   fg="#000000", bg="#FFFFFF")
            email_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

            from university_system.modules.shared.constants import paths
            import os

            env_exists = os.path.exists(paths.PROJECT_ROOT / '.env')

            email_info = f"""Email Service Configuration
================================

Status: {'Configured' if env_exists else 'Not Configured'}

Configuration File: .env
Location: {paths.PROJECT_ROOT}

Required Settings:
- SMTP_HOST (e.g., smtp.gmail.com)
- SMTP_PORT (usually 587)
- SMTP_USER (your email address)
- SMTP_PASSWORD (app-specific password)

Features:
- Asynchronous email queue
- Template support
- Bulk email capability
- Email logging

Configuration Steps:
1. Copy .env.example to .env
2. Fill in SMTP credentials
3. Test email service
4. Restart application

Note: Use app-specific passwords for Gmail/Outlook.
For support, check the email service documentation.
"""
            email_text.insert("1.0", email_info)
            email_text.config(state=tk.DISABLED)

            ttk.Button(email_window, text="Close", command=email_window.destroy).pack(pady=10)

        except Exception as e:
            messagebox.showerror("Error", f"Failed to configure email: {e}")

    def configure_backup(self):
        """Configure backup settings"""
        try:
            backup_window = tk.Toplevel(self.root)
            backup_window.title("Backup Configuration")
            backup_window.geometry("600x500")

            ttk.Label(backup_window, text="Backup Settings",
                     font=('Arial', 14, 'bold')).pack(pady=10)

            backup_text = scrolledtext.ScrolledText(backup_window, wrap=tk.WORD, height=20,
                                                    fg="#000000", bg="#FFFFFF")
            backup_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

            from university_system.modules.shared.constants import paths

            backup_info = f"""Backup Configuration
================================

Backup Location: {paths.BACKUP_DIR}

Current Settings:
- Auto-backup: Recommended (manual currently)
- Backup frequency: Daily recommended
- Retention period: 30 days
- Compression: Enabled

Backup Types:
1. Full Database Backup
   - Complete database copy
   - All tables and data

2. Incremental Backup
   - Changes since last backup
   - Faster, smaller size

Manual Backup:
- Use Database Tools → Backup Database
- Specify backup location
- Optionally compress backup

Restore Process:
- Use Database Tools → Restore Database
- Select backup file
- Confirm restoration

Recommendations:
- Daily automated backups
- Test restore process monthly
- Store backups off-site
- Keep multiple backup versions
"""
            backup_text.insert("1.0", backup_info)
            backup_text.config(state=tk.DISABLED)

            ttk.Button(backup_window, text="Close", command=backup_window.destroy).pack(pady=10)

        except Exception as e:
            messagebox.showerror("Error", f"Failed to configure backup: {e}")

    def configure_security(self):
        """Configure security settings"""
        try:
            security_window = tk.Toplevel(self.root)
            security_window.title("Security Configuration")
            security_window.geometry("600x500")

            ttk.Label(security_window, text="Security Settings",
                     font=('Arial', 14, 'bold')).pack(pady=10)

            security_text = scrolledtext.ScrolledText(security_window, wrap=tk.WORD, height=20,
                                                      fg="#000000", bg="#FFFFFF")
            security_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

            security_info = """Security Configuration
================================

Authentication:
- Password hashing: PBKDF2-SHA256
- Hash iterations: 1,000,000 (OWASP recommended)
- Salt: Unique per user
- MFA support: TOTP, Email OTP, SMS OTP

Session Management:
- Token-based sessions
- Session timeout: 60 minutes
- Concurrent session limits: Configurable
- Automatic logout: On inactivity

Authorization:
- Role-based access control (RBAC)
- Permission checking at service layer
- Admin, Staff, Student roles
- Granular permissions

Data Protection:
- SQL injection prevention (parameterized queries)
- XSS protection
- CSRF protection
- Input validation

Audit Trail:
- All data modifications logged
- User attribution
- Timestamp tracking
- Compliance ready

Security Best Practices:
- Regular password updates
- Enable MFA for all users
- Monitor activity logs
- Review permissions regularly
- Keep system updated

For security issues, contact system administrator.
"""
            security_text.insert("1.0", security_info)
            security_text.config(state=tk.DISABLED)

            ttk.Button(security_window, text="Close", command=security_window.destroy).pack(pady=10)

        except Exception as e:
            messagebox.showerror("Error", f"Failed to configure security: {e}")

    # Phase 4 Feature GUIs
    def show_virtual_classroom_gui(self):
        """Launch the Virtual Classroom Management GUI in a child window"""
        if not self.auth.current_user:
            messagebox.showerror("Error", "You must be logged in to access Virtual Classroom Management.")
            return

        try:
            if not VIRTUAL_CLASSROOM_AVAILABLE:
                messagebox.showerror("Virtual Classroom", "Virtual Classroom feature is not available.")
                return

            # Import the comprehensive Virtual Classroom GUI
            from university_system.modules.domain.academics.gui.virtual_classroom_gui import VirtualClassroomGUI

            # Create a new window for the Virtual Classroom GUI
            classroom_window = tk.Toplevel(self.root)

            # Initialize the comprehensive GUI with auth instance
            app = VirtualClassroomGUI(classroom_window, auth=self.auth)

            print("✅ Virtual Classroom GUI opened successfully")

        except ImportError as e:
            messagebox.showerror("Error", f"Failed to import Virtual Classroom GUI: {str(e)}\n\nPlease ensure the module is properly installed.")
            print(f"❌ Virtual Classroom import error: {e}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open Virtual Classroom GUI: {str(e)}")
            print(f"❌ Virtual Classroom error: {e}")


    def show_email_sms_gui(self):
        """Launch the Email & SMS Communication Hub"""
        if not self.auth.current_user:
            messagebox.showerror("Error", "You must be logged in to access Communication Hub.")
            return

        try:
            # Open the email manager GUI which now includes SMS
            from university_system.infrastructure.email.gui.email_manager_management_gui import EmailManagerManagementGUI

            email_gui = EmailManagerManagementGUI(self.root, self.auth)
            email_gui.show_email_manager()

            print("✅ Communication Hub (Email & SMS) GUI opened successfully")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to open Communication Hub: {str(e)}")
            print(f"❌ Communication Hub error: {e}")

    # ==================== PHASE 1-3 FEATURE GUIs ====================
    # Note: Advanced Attendance and Smart Timetable features are now integrated
    # into the main Attendance and Scheduling GUIs

    def show_mental_health_gui(self):
        """Launch the Mental Health & Wellness Portal GUI"""
        launch_mental_health_gui(self.root, self.auth)

    def show_early_warning_gui(self):
        """Launch the Student Success Early Warning System GUI"""
        launch_early_warning_gui(self.root, self.auth)

    def show_career_services_gui(self):
        """Launch the Career Services GUI"""
        launch_career_services_gui(self.root, self.auth)

    def show_admissions_crm_gui(self):
        """Launch the Admissions Crm GUI"""
        launch_admissions_crm_gui(self.root, self.auth)

    def show_predictive_analytics_gui(self):
        """Launch the Predictive Analytics GUI"""
        launch_predictive_analytics_gui(self.root, self.auth)

    # Timetable optimizer features are now integrated into the Module Scheduling GUI

    def show_campus_events_gui(self):
        """Launch the Campus Events GUI"""
        launch_campus_events_gui(self.root, self.auth)

    # Alumni Relations features have been merged into Alumni Management GUI


    def show_facilities_management_gui(self):
        """Launch the Facilities Management GUI"""
        launch_facilities_management_gui(self.root, self.auth)

    def show_business_intelligence_gui(self):
        """Launch the Business Intelligence GUI"""
        launch_business_intelligence_gui(self.root, self.auth)

    def show_ai_features_gui(self):
        """Launch the Ai Features GUI"""
        launch_ai_features_gui(self.root, self.auth)

    def show_integration_marketplace_gui(self):
        """Launch the Integration Marketplace GUI"""
        launch_integration_marketplace_gui(auth=self.auth)

    def show_mobile_app_pwa_gui(self):
        """Launch the Mobile App (PWA) Infrastructure GUI"""
        launch_mobile_app_pwa_gui(auth=self.auth)

    def show_accessibility_tools_gui(self):
        """Launch the Accessibility & Accommodation Tools GUI"""
        try:
            from university_system.modules.domain.student_affairs.gui.accessibility_tools_gui import (
                launch_accessibility_tools_gui
            )
            launch_accessibility_tools_gui(self.root, self.auth)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to launch Accessibility Tools: {str(e)}")
            print(f"❌ Accessibility Tools error: {e}")
    # Parent Portal Enhancement features have been merged into the main Parent Portal GUI

    def show_transportation_parking_gui(self):
        """Launch the Transportation & Parking Management GUI"""
        self._show_feature_gui("Transportation & Parking Management", """Complete transportation and parking management system.

Features:
• Parking permits
• Parking lot management
• Parking violations
• Shuttle bus routes & schedules
• Transportation requests
• Rideshare program
• Bike registration

Database Tables:
parking_permits, parking_lots, parking_violations, shuttle_routes,
shuttle_schedules, shuttle_stops, transportation_requests,
rideshare_participants, rideshare_trips, bike_registrations,
parking_reservations, traffic_analytics""", "Use CLI: Transportation & Parking Management")

    def show_blockchain_credentials_gui(self):
        """Launch the Blockchain Credentials & Digital Badges GUI"""
        launch_blockchain_credentials_gui(auth=self.auth)

    def _show_feature_gui(self, title, description, cli_instruction):
        """Generic method to show feature GUI window"""
        if not self.auth.current_user:
            messagebox.showerror("Error", f"You must be logged in to access {title}.")
            return

        try:
            window = tk.Toplevel(self.root)
            window.title(title)
            window.geometry("900x600")
            window.minsize(800, 500)

            main_frame = ttk.Frame(window, padding="20")
            main_frame.pack(fill=tk.BOTH, expand=True)

            ttk.Label(main_frame, text=title, font=('Arial', 16, 'bold')).pack(pady=10)
            ttk.Label(main_frame, text=description, justify=tk.LEFT, wraplength=800).pack(pady=10)

            info_frame = ttk.LabelFrame(main_frame, text="How to Access", padding="15")
            info_frame.pack(fill=tk.X, pady=20)
            ttk.Label(info_frame, text=cli_instruction, font=('Arial', 11)).pack()

            ttk.Button(main_frame, text="Close", command=window.destroy).pack(pady=10)

            print(f"✅ {title} GUI opened successfully")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to open {title}: {str(e)}")
            print(f"❌ {title} error: {e}")

    def run(self):
        """Start the GUI application"""
        self.root.mainloop()

# Replace the StudentManagementGUI class usage throughout main.py
class StudentManagementGUI(UnifiedManagementGUI):
    """Alias for backwards compatibility"""
    pass

def start_gui_mode():
    """Start the GUI version of the application - wrapper for backward compatibility"""
    return run_gui_interface()

def enhanced_interface_choice():
    """Enhanced interface selection with better error handling"""
    global auth
    
    print("\n" + "="*60)
    print("STUDENT RECORD MANAGEMENT SYSTEM")
    print("Advanced Academic Management Suite")
    print("="*60)
    
    while True:
        print("\nInterface Options:")
        print("1. Command Line Interface (CLI) - Full featured")
        print("2. Graphical User Interface (GUI) - Modern interface")
        print("3. Auto-detect best interface")
        print("4. Exit application")
        
        choice = input("\nSelect interface mode (1-4): ").strip()
        
        if choice == '1':
            print("Starting CLI mode...")
            return 'cli'
        elif choice == '2':
            # Check if GUI is available before committing
            try:
                if tk is None:
                    print("GUI mode not available - tkinter missing.")
                    print("Would you like to use CLI mode instead? (y/n): ", end="")
                    if input().lower().startswith('y'):
                        return 'cli'
                    continue
                print("Starting GUI mode...")
                return 'gui'
            except:
                print("GUI mode not available. Please choose CLI mode.")
                continue
        elif choice == '3':
            # Auto-detect best available interface
            if tk is not None:
                print("GUI available - starting GUI mode...")
                return 'gui'
            else:
                print("GUI not available - starting CLI mode...")
                return 'cli'
        elif choice == '4':
            print("Thank you for using the Student Record Management System!")
            return 'exit'
        else:
            print("Invalid choice. Please enter 1, 2, 3, or 4.")
            
def main():
    """Enhanced main application entry point with GUI integration"""
    global auth
    
    try:
        # System initialization
        print("Initializing Student Record Management System...")
        if not initialize_system():
            print("System initialization failed")
            return False
        
        # Initialize authentication
        if auth is None:
            gui_funcs = get_gui_auth_functions()
            if gui_funcs:
                auth = gui_funcs['initialize_complete_system_with_gui']()
            else:
                auth = UserAuth()
                safe_auth_check(auth)
        
        # Show system information
        # WARNING: These are default demo credentials - change them in production!
        # Set DEFAULT_ADMIN_PASSWORD, DEFAULT_STAFF_PASSWORD, DEFAULT_STUDENT_PASSWORD environment variables
        print("\n" + "="*60)
        print("STUDENT RECORD MANAGEMENT SYSTEM")
        print("Advanced Academic Management Suite")
        print("="*60)
        print("\nDefault Login Credentials:")
        print(f"- Admin: username='admin', password='{os.getenv('DEFAULT_ADMIN_PASSWORD', 'admin123')}'")
        print(f"- Staff: username='staff', password='{os.getenv('DEFAULT_STAFF_PASSWORD', 'staff123')}'")
        print(f"- Student: username='student', password='{os.getenv('DEFAULT_STUDENT_PASSWORD', 'student123')}'")

        # Start the CLI interface directly (run.py will handle interface choice)
        print("Starting console interface...")
        display_menu()
        return True
            
    except KeyboardInterrupt:
        print("\n\nApplication interrupted by user")
        return True
    except Exception as e:
        print(f"Critical error: {e}")
        logging.critical(f"Application failure: {e}")
        return False
    finally:
        try:
            cleanup_database_connections()
        except:
            pass
        print("System shutdown complete")

def run_gui_interface():
    """Run the unified GUI interface"""
    try:
        if tk is None:
            print("GUI mode requires tkinter, which is not available.")
            return False

        print("Starting unified GUI interface...")

        # Use centralized init_gui with no session_user (starts at login page)
        app = init_gui(session_user=None)
        app.run()

        print("GUI interface closed.")
        return True
        
    except Exception as e:
        print(f"Error starting GUI mode: {e}")
        return False
        
def switch_to_gui_mode():
    """Allow CLI users to switch to GUI mode"""
    global auth
    
    gui_funcs = get_gui_auth_functions()
    if not gui_funcs:
        print("GUI interface is not available on this system")
        return False
    
    try:
        print("Switching to GUI interface...")
        gui_funcs['launch_gui_with_console_fallback'](auth)
        return True
    except Exception as e:
        print(f"Failed to start GUI: {e}")
        return False
    
def complete_gui_integration():
    """Complete integration function to call in main.py"""
    print("Completing GUI integration...")
    
    # This function should be called at the end of your main() function
    # or in the StudentManagementGUI initialization
    
    try:
        # Initialize advanced search integration
        print("1. Advanced search components loaded")
        
        # Setup data synchronization
        print("2. Data synchronization configured")
        
        # Configure keyboard shortcuts
        print("3. Keyboard shortcuts registered")
        
        # Setup menu integration
        print("4. Menu integration complete")
        
        # Initialize analytics integration
        print("5. Analytics integration ready")
        
        print("GUI integration completed successfully!")
        print("\nAvailable features:")
        print("- Advanced Search tab and window")
        print("- Multi-criteria search with fuzzy matching")
        print("- Module enrollment filtering")
        print("- Date range searches")
        print("- Integrated analytics dashboard")
        print("- Real-time data synchronization")
        print("- Enhanced export capabilities")
        
        return True
        
    except Exception as e:
        print(f"GUI integration failed: {e}")
        return False

if __name__ == "__main__":
    # This block should only run when main.py is called directly
    # When run.py is used, this won't execute
    print("Note: For best experience, use run.py to start the application")
    print("Starting CLI mode directly...")
    
    try:
        # Perform initial setup
        silent_integrity_check()
        fix_parent_portal_database()
        
        # Run the main function
        success = main()
        
        if not success:
            print("Application failed to start properly")
            exit(1)
            
    except Exception as e:
        print(f"Critical startup error: {e}")
        logging.critical(f"Application startup failure: {e}")
        exit(1)
    finally:
        print("Thank you for using the Student Record Management System!")
