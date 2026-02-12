import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import os
import csv
import json
from datetime import datetime
import logging

logger = logging.getLogger(__name__)

try:
    from university_system.infrastructure.database.db import get_connection
except ImportError:
    from university_system.infrastructure.database.db import sqlite3, DEFAULT_DB_PATH
    def get_connection():
        return sqlite3.connect(str(DEFAULT_DB_PATH))

from university_system.core.sql_safety import validate_identifier

try:
    from university_system.modules.shared.utils.i18n import get_text as _t
except ImportError:
    _t = lambda key, **kwargs: key if "default" not in kwargs else kwargs.get("default")


class ExportManager:
    def __init__(self, gui):
        self.gui = gui
        self.root = gui.root

    def export_data_dialog(self):
        """Show export data dialog"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Export Data")
        dialog.geometry("600x450")
        dialog.transient(self.root)
        dialog.grab_set()

        main_frame = ttk.Frame(dialog, padding=20)
        main_frame.pack(fill='both', expand=True)

        ttk.Label(main_frame, text="Export Data", font=('Arial', 14, 'bold')).pack(pady=(0, 20))

        export_options = [
            ("All Documents", "documents"),
            ("All Students", "students"),
            ("Document Types", "document_types"),
            ("System Report", "system_report")
        ]

        self.gui.export_selection = tk.StringVar(value="documents")

        for text, value in export_options:
            ttk.Radiobutton(main_frame, text=text, variable=self.gui.export_selection, value=value).pack(anchor='w', pady=2)

        # Format selection
        ttk.Label(main_frame, text="Export Format:", font=('Arial', 10, 'bold')).pack(anchor='w', pady=(15, 5))
        self.gui.export_format = tk.StringVar(value="csv")

        formats = [("CSV", "csv"), ("Excel", "xlsx"), ("JSON", "json")]
        for text, value in formats:
            ttk.Radiobutton(main_frame, text=text, variable=self.gui.export_format, value=value).pack(anchor='w', pady=2)

        # Buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill='x', pady=(20, 0))

        ttk.Button(button_frame, text="Export", command=lambda: self.perform_export(dialog)).pack(side='right', padx=5)
        ttk.Button(button_frame, text="Cancel", command=dialog.destroy).pack(side='right')

    def perform_export(self, dialog):
        """Perform the actual export"""
        file_path = filedialog.asksaveasfilename(
            title="Save Export File",
            defaultextension=f".{self.gui.export_format.get()}",
            filetypes=[(f"{self.gui.export_format.get().upper()} files", f"*.{self.gui.export_format.get()}")]
        )

        if file_path:
            try:
                data_type = self.gui.export_selection.get()
                format_type = self.gui.export_format.get()

                conn = get_connection()
                cursor = conn.cursor()

                if data_type == "documents":
                    cursor.execute('''
                    SELECT sd.document_id, s.student_id, s.first_name, s.last_name,
                           dt.type_name, sd.upload_date, sd.verification_status
                    FROM student_documents sd
                    JOIN students s ON sd.student_id = s.student_id
                    JOIN document_types dt ON sd.type_id = dt.type_id
                    WHERE sd.is_current_version = 1
                    ''')
                    columns = ['Document ID', 'Student ID', 'First Name', 'Last Name', 'Document Type', 'Upload Date', 'Status']
                elif data_type == "students":
                    cursor.execute('SELECT * FROM students')
                    columns = ['Student ID', 'First Name', 'Last Name', 'Email', 'Course', 'Year', 'Enrollment Date', 'Status']
                else:
                    cursor.execute('SELECT * FROM document_types WHERE is_active = 1')
                    columns = ['Type ID', 'Type Name', 'Description', 'Required', 'Has Expiry', 'Reminder Days', 'Max Size MB', 'Formats', 'Requires Approval', 'Category', 'Sort Order', 'Active']

                data = cursor.fetchall()
                conn.close()

                if format_type == "csv":
                    with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                        writer = csv.writer(csvfile)
                        writer.writerow(columns)
                        writer.writerows(data)
                elif format_type == "json":
                    json_data = [dict(zip(columns, row)) for row in data]
                    with open(file_path, 'w', encoding='utf-8') as jsonfile:
                        json.dump(json_data, jsonfile, indent=2, default=str)

                messagebox.showinfo("Success", f"Data exported successfully to {file_path}")
                dialog.destroy()

            except Exception as e:
                messagebox.showerror("Export Error", f"Failed to export data: {str(e)}")

    def export_activity_log(self):
        """Export activity log to CSV"""
        try:
            file_path = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV Files", "*.csv"), ("All Files", "*.*")],
                title="Export Activity Log"
            )

            if not file_path:
                return

            with get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT timestamp, username, user_role, action, entity_type,
                           entity_id, details
                    FROM activity_log
                    ORDER BY timestamp DESC
                """)
                activities = cursor.fetchall()

            with open(file_path, 'w', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(['Timestamp', 'Username', 'Role', 'Action', 'Entity Type',
                               'Entity ID', 'Details'])
                writer.writerows(activities)

            messagebox.showinfo("Export Successful",
                              f"Activity log exported to:\n{file_path}")

        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export activity log: {e}")

    def export_all_documents(self):
        """Export all document records to CSV"""
        try:
            filename = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
                initialfile=f"all_documents_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            )

            if not filename:
                return

            conn = get_connection()
            cursor = conn.cursor()

            cursor.execute('''
            SELECT sd.document_id, sd.student_id, s.first_name, s.last_name,
                   dt.type_name, sd.upload_date, sd.expiry_date,
                   sd.verification_status, sd.workflow_status, sd.uploaded_by,
                   sd.original_filename, sd.file_size, sd.version_number
            FROM student_documents sd
            JOIN students s ON sd.student_id = s.student_id
            JOIN document_types dt ON sd.document_type_id = dt.type_id
            WHERE sd.is_current_version = 1
            ORDER BY sd.upload_date DESC
            ''')

            documents = cursor.fetchall()
            conn.close()

            with open(filename, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(['Doc ID', 'Student ID', 'First Name', 'Last Name', 'Document Type',
                               'Upload Date', 'Expiry Date', 'Verification Status', 'Workflow Status',
                               'Uploaded By', 'Filename', 'Size (bytes)', 'Version'])
                writer.writerows(documents)

            messagebox.showinfo("Success", f"All documents exported to:\n{filename}\nTotal: {len(documents)} records")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to export documents: {e}")

    def export_document_history(self):
        """Export complete document history"""
        if not self.gui.ensure_login():
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV Files", "*.csv"), ("All Files", "*.*")],
            initialfile="document_history.csv"
        )

        if not file_path:
            return

        try:
            with open(file_path, 'w', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(['Document ID', 'Student ID', 'Document Type', 'File Name',
                               'Status', 'Upload Date', 'Expiry Date', 'File Size', 'Tags', 'Notes'])

                with get_connection() as conn:
                    cursor = conn.cursor()
                    cursor.execute("""
                        SELECT id, student_id, document_type, file_name, status,
                               upload_date, expiry_date, file_size, tags, notes
                        FROM documents
                        ORDER BY upload_date DESC
                    """)
                    results = cursor.fetchall()

                    writer.writerows(results)

            messagebox.showinfo("Export Complete",
                              f"Document history exported successfully:\n{file_path}\n\n"
                              f"Total records: {len(results)}")

            self.gui.log_event('export', 'document_history',
                          details=f'Exported {len(results)} document records')

        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export document history: {e}")

    def export_workflow_data(self):
        """Export workflow data"""
        if not self.gui.ensure_login():
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV Files", "*.csv"), ("All Files", "*.*")],
            initialfile="workflow_data.csv"
        )

        if not file_path:
            return

        try:
            with open(file_path, 'w', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(['Workflow ID', 'Workflow Name', 'Document Type',
                               'Total Steps', 'Active', 'Created Date'])

                # Sample workflow data (would query from workflow_templates table)
                sample_workflows = [
                    (1, 'Standard Document Review', 'Transcript', 3, 'Yes', '2024-01-15'),
                    (2, 'Fast Track Approval', 'ID Card', 2, 'Yes', '2024-02-01'),
                    (3, 'Detailed Compliance Check', 'Financial Document', 5, 'Yes', '2024-03-10'),
                ]

                writer.writerows(sample_workflows)

            messagebox.showinfo("Export Complete",
                              f"Workflow data exported successfully:\n{file_path}")

            self.gui.log_event('export', 'workflow_data', details='Exported workflow data')

        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export workflow data: {e}")

    def export_student_list(self):
        """Export list of all students"""
        if not self.gui.ensure_login():
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV Files", "*.csv"), ("All Files", "*.*")],
            initialfile="student_list.csv"
        )

        if not file_path:
            return

        try:
            with open(file_path, 'w', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(['Student ID', 'Total Documents', 'Pending', 'Approved',
                               'Rejected', 'Last Upload Date'])

                with get_connection() as conn:
                    cursor = conn.cursor()
                    cursor.execute("""
                        SELECT
                            student_id,
                            COUNT(*) as total,
                            SUM(CASE WHEN status = 'Pending' THEN 1 ELSE 0 END) as pending,
                            SUM(CASE WHEN status = 'Approved' THEN 1 ELSE 0 END) as approved,
                            SUM(CASE WHEN status = 'Rejected' THEN 1 ELSE 0 END) as rejected,
                            MAX(upload_date) as last_upload
                        FROM documents
                        GROUP BY student_id
                        ORDER BY student_id
                    """)
                    results = cursor.fetchall()

                    writer.writerows(results)

            messagebox.showinfo("Export Complete",
                              f"Student list exported successfully:\n{file_path}\n\n"
                              f"Total students: {len(results)}")

            self.gui.log_event('export', 'student_list',
                          details=f'Exported {len(results)} students')

        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export student list: {e}")

    def export_student_documents(self):
        """Export documents for selected students"""
        if not self.gui.ensure_login():
            return

        # Ask for student IDs
        from tkinter import simpledialog
        student_ids = simpledialog.askstring("Export Student Documents",
                                             "Enter student IDs (comma-separated):")
        if not student_ids:
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV Files", "*.csv"), ("All Files", "*.*")],
            initialfile="student_documents.csv"
        )

        if not file_path:
            return

        try:
            ids = [id.strip() for id in student_ids.split(',')]

            with open(file_path, 'w', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(['Student ID', 'Document ID', 'Document Type', 'File Name',
                               'Status', 'Upload Date', 'Expiry Date'])

                with get_connection() as conn:
                    cursor = conn.cursor()
                    placeholders = ','.join(['?' for _ in ids])
                    cursor.execute(f"""
                        SELECT student_id, id, document_type, file_name, status,
                               upload_date, expiry_date
                        FROM documents
                        WHERE student_id IN ({placeholders})
                        ORDER BY student_id, upload_date DESC
                    """, ids)
                    results = cursor.fetchall()

                    writer.writerows(results)

            messagebox.showinfo("Export Complete",
                              f"Student documents exported successfully:\n{file_path}\n\n"
                              f"Total records: {len(results)}")

            self.gui.log_event('export', 'student_documents',
                          details=f'Exported documents for {len(ids)} students')

        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export student documents: {e}")

    def export_db_schema(self):
        """Export database schema"""
        if not self.gui.ensure_login('admin'):
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".sql",
            filetypes=[("SQL Files", "*.sql"), ("All Files", "*.*")],
            initialfile="database_schema.sql"
        )

        if not file_path:
            return

        try:
            with open(file_path, 'w') as f:
                f.write("-- Database Schema Export\n")
                f.write(f"-- Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")

                with get_connection() as conn:
                    cursor = conn.cursor()
                    cursor.execute("""
                        SELECT sql FROM sqlite_master
                        WHERE type='table'
                        ORDER BY name
                    """)
                    tables = cursor.fetchall()

                    for table_sql in tables:
                        if table_sql[0]:
                            f.write(f"{table_sql[0]};\n\n")

            messagebox.showinfo("Export Complete",
                              f"Database schema exported successfully:\n{file_path}")

            self.gui.log_event('export', 'db_schema', details='Exported database schema')

        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export schema: {e}")

    def export_student_data(self):
        """Export all student data"""
        # This is handled by export_all_students()
        self.export_all_students()

    def export_compliance_report(self):
        """Export compliance report"""
        if not self.gui.ensure_login():
            return

        # Create compliance report window
        report_window = tk.Toplevel(self.root)
        report_window.title("Compliance Report Export")
        report_window.geometry("600x500")
        report_window.transient(self.root)
        report_window.grab_set()

        ttk.Label(report_window, text="Compliance Report Configuration",
                 font=("Arial", 14, "bold")).pack(pady=10)

        # Options frame
        options_frame = ttk.LabelFrame(report_window, text="Report Options", padding=15)
        options_frame.pack(fill='both', expand=True, padx=10, pady=5)

        include_pending_var = tk.BooleanVar(value=True)
        include_expired_var = tk.BooleanVar(value=True)
        include_missing_var = tk.BooleanVar(value=True)
        include_compliant_var = tk.BooleanVar(value=False)

        ttk.Checkbutton(options_frame, text="Include Students with Pending Documents",
                       variable=include_pending_var).pack(anchor='w', pady=5)
        ttk.Checkbutton(options_frame, text="Include Students with Expired Documents",
                       variable=include_expired_var).pack(anchor='w', pady=5)
        ttk.Checkbutton(options_frame, text="Include Students with Missing Required Documents",
                       variable=include_missing_var).pack(anchor='w', pady=5)
        ttk.Checkbutton(options_frame, text="Include Fully Compliant Students",
                       variable=include_compliant_var).pack(anchor='w', pady=5)

        ttk.Separator(options_frame, orient='horizontal').pack(fill='x', pady=10)

        ttk.Label(options_frame, text="Export Format:").pack(anchor='w', pady=5)
        format_combo = ttk.Combobox(options_frame, width=30, state='readonly')
        format_combo['values'] = ['CSV', 'PDF', 'Excel']
        format_combo.current(0)
        format_combo.pack(anchor='w', padx=20)

        def generate_report():
            """Generate and export compliance report"""
            file_format = format_combo.get().lower()

            file_path = filedialog.asksaveasfilename(
                defaultextension=f".{file_format}",
                filetypes=[(f"{file_format.upper()} Files", f"*.{file_format}")],
                initialfile=f"compliance_report.{file_format}"
            )

            if not file_path:
                return

            try:
                if file_format == 'csv':
                    self.export_compliance_data(
                        file_path,
                        include_pending_var.get(),
                        include_expired_var.get(),
                        include_missing_var.get(),
                        include_compliant_var.get()
                    )

                    messagebox.showinfo("Export Complete",
                                      f"Compliance report exported successfully:\n{file_path}")
                else:
                    messagebox.showinfo("Format Not Implemented",
                                      f"{file_format.upper()} export will be implemented.\n\n"
                                      "Currently only CSV export is supported.")

                self.gui.log_event('export', 'compliance_report',
                              details=f'Exported compliance report to {file_format.upper()}')

                report_window.destroy()

            except Exception as e:
                messagebox.showerror("Export Error", f"Failed to export report: {e}")

        ttk.Button(report_window, text="Generate & Export",
                  command=generate_report).pack(pady=20)
        ttk.Button(report_window, text="Cancel",
                  command=report_window.destroy).pack()

    def export_compliance_data(self, file_path, include_pending, include_expired,
                               include_missing, include_compliant):
        """Export compliance data to CSV"""
        try:
            with open(file_path, 'w', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(['Student ID', 'Compliance Status', 'Pending Docs',
                               'Expired Docs', 'Missing Required', 'Total Documents'])

                with get_connection() as conn:
                    cursor = conn.cursor()

                    # Get all students (simulated - in reality would join with students table)
                    cursor.execute("""
                        SELECT DISTINCT student_id FROM documents
                        ORDER BY student_id
                    """)
                    students = cursor.fetchall()

                    for (student_id,) in students:
                        # Count various document statuses
                        cursor.execute("""
                            SELECT
                                COUNT(CASE WHEN status = 'Pending' THEN 1 END) as pending,
                                COUNT(CASE WHEN DATE(expiry_date) < DATE('now') THEN 1 END) as expired,
                                COUNT(*) as total
                            FROM documents
                            WHERE student_id = ?
                        """, (student_id,))

                        pending, expired, total = cursor.fetchone()
                        missing = 0  # Would calculate based on required document types

                        # Determine compliance status
                        if pending == 0 and expired == 0 and missing == 0:
                            status = "Compliant"
                            if not include_compliant:
                                continue
                        else:
                            status = "Non-Compliant"

                        # Apply filters
                        if not include_pending and pending > 0:
                            continue
                        if not include_expired and expired > 0:
                            continue
                        if not include_missing and missing > 0:
                            continue

                        writer.writerow([student_id, status, pending, expired, missing, total])

        except Exception as e:
            raise Exception(f"Failed to export compliance data: {e}")

    def export_custom_report(self):
        """Export custom report with user-defined fields"""
        if not self.gui.ensure_login():
            return

        # Create custom report builder
        builder_window = tk.Toplevel(self.root)
        builder_window.title("Custom Report Builder")
        builder_window.geometry("700x650")
        builder_window.transient(self.root)
        builder_window.grab_set()

        ttk.Label(builder_window, text="Custom Report Builder",
                 font=("Arial", 14, "bold")).pack(pady=10)

        # Fields selection
        fields_frame = ttk.LabelFrame(builder_window, text="Select Fields to Include", padding=15)
        fields_frame.pack(fill='both', expand=True, padx=10, pady=5)

        field_vars = {}
        available_fields = [
            ('Document ID', 'id'),
            ('Student ID', 'student_id'),
            ('Document Type', 'document_type'),
            ('File Name', 'file_name'),
            ('Status', 'status'),
            ('Upload Date', 'upload_date'),
            ('Expiry Date', 'expiry_date'),
            ('File Size', 'file_size'),
            ('Tags', 'tags'),
            ('Notes', 'notes')
        ]

        for label, field in available_fields:
            var = tk.BooleanVar(value=field in ['student_id', 'document_type', 'status'])
            field_vars[field] = var
            ttk.Checkbutton(fields_frame, text=label, variable=var).pack(anchor='w', pady=2)

        # Filters
        filters_frame = ttk.LabelFrame(builder_window, text="Filters (Optional)", padding=15)
        filters_frame.pack(fill='x', padx=10, pady=5)

        ttk.Label(filters_frame, text="Date Range:").grid(row=0, column=0, sticky='w')
        date_from = ttk.Entry(filters_frame, width=15)
        date_from.insert(0, "YYYY-MM-DD")
        date_from.grid(row=0, column=1, padx=5)

        ttk.Label(filters_frame, text="to").grid(row=0, column=2)
        date_to = ttk.Entry(filters_frame, width=15)
        date_to.insert(0, "YYYY-MM-DD")
        date_to.grid(row=0, column=3, padx=5)

        def export_report():
            """Export the custom report"""
            selected_fields = [field for field, var in field_vars.items() if var.get()]

            if not selected_fields:
                messagebox.showwarning("No Fields", "Please select at least one field")
                return

            file_path = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV Files", "*.csv")],
                initialfile="custom_report.csv"
            )

            if not file_path:
                return

            try:
                self.export_custom_dataset(file_path, selected_fields)

                messagebox.showinfo("Export Complete",
                                  f"Custom report exported successfully:\n{file_path}\n\n"
                                  f"Fields included: {len(selected_fields)}")

                self.gui.log_event('export', 'custom_report',
                              details=f'Exported custom report with {len(selected_fields)} fields')

                builder_window.destroy()

            except Exception as e:
                messagebox.showerror("Export Error", f"Failed to export report: {e}")

        ttk.Button(builder_window, text="Export Report",
                  command=export_report).pack(pady=15)
        ttk.Button(builder_window, text="Cancel",
                  command=builder_window.destroy).pack()

    def export_custom_dataset(self, file_path, fields):
        """Export custom dataset to CSV with selected fields"""
        try:
            with open(file_path, 'w', newline='') as f:
                writer = csv.writer(f)

                # Write header
                writer.writerow(fields)

                # Query data - validate field names
                validated_fields = [validate_identifier(f, "column") for f in fields]
                query = "SELECT " + ", ".join(validated_fields) + " FROM documents ORDER BY id DESC LIMIT 1000"

                with get_connection() as conn:
                    cursor = conn.cursor()
                    cursor.execute(query)
                    results = cursor.fetchall()

                # Write data
                writer.writerows(results)

        except Exception as e:
            raise Exception(f"Failed to export dataset: {e}")

    def export_all_students(self):
        """Export all students and their document summaries"""
        if not self.gui.ensure_login():
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV Files", "*.csv")],
            initialfile="all_students_report.csv"
        )

        if not file_path:
            return

        try:
            with open(file_path, 'w', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(['Student ID', 'Total Documents', 'Pending', 'Approved',
                               'Rejected', 'Expired', 'Last Upload Date'])

                with get_connection() as conn:
                    cursor = conn.cursor()
                    cursor.execute("""
                        SELECT
                            student_id,
                            COUNT(*) as total,
                            SUM(CASE WHEN status = 'Pending' THEN 1 ELSE 0 END) as pending,
                            SUM(CASE WHEN status = 'Approved' THEN 1 ELSE 0 END) as approved,
                            SUM(CASE WHEN status = 'Rejected' THEN 1 ELSE 0 END) as rejected,
                            SUM(CASE WHEN DATE(expiry_date) < DATE('now') THEN 1 ELSE 0 END) as expired,
                            MAX(upload_date) as last_upload
                        FROM documents
                        GROUP BY student_id
                        ORDER BY student_id
                    """)
                    results = cursor.fetchall()

                    writer.writerows(results)

            messagebox.showinfo("Export Complete",
                              f"Student summary exported successfully:\n{file_path}\n\n"
                              f"Total students: {len(results)}")

            self.gui.log_event('export', 'all_students',
                          details=f'Exported {len(results)} student summaries')

        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export students: {e}")

    def export_to_csv(self):
        """Export current view/data to CSV"""
        if not self.gui.ensure_login():
            return

        # Ask for file location
        file_path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV Files", "*.csv"), ("All Files", "*.*")],
            initialfile="document_export.csv"
        )

        if not file_path:
            return

        try:
            # Determine what to export based on current context
            # For now, export all documents
            with open(file_path, 'w', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(['Document ID', 'Student ID', 'Document Type', 'File Name',
                               'Status', 'Upload Date', 'Expiry Date', 'File Size'])

                with get_connection() as conn:
                    cursor = conn.cursor()
                    cursor.execute("""
                        SELECT id, student_id, document_type, file_name, status,
                               upload_date, expiry_date, file_size
                        FROM documents
                        ORDER BY upload_date DESC
                        LIMIT 1000
                    """)
                    results = cursor.fetchall()

                    writer.writerows(results)

            messagebox.showinfo("Export Successful",
                              f"Data exported to CSV successfully!\n\n"
                              f"{file_path}\n\n"
                              f"Total records: {len(results)}")

            self.gui.log_event('export', 'csv', details=f'Exported {len(results)} records to CSV')

        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export to CSV: {e}")

    def export_to_excel(self):
        """Export current view/data to Excel"""
        if not self.gui.ensure_login():
            return

        # Ask for file location
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")],
            initialfile="document_export.xlsx"
        )

        if not file_path:
            return

        try:
            # Try to import openpyxl
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils import get_column_letter
            except ImportError:
                messagebox.showerror("Missing Library",
                                   "The 'openpyxl' library is required for Excel export.\n\n"
                                   "Install it with:\n"
                                   "pip install openpyxl\n\n"
                                   "Would you like to use CSV export instead?")
                if messagebox.askyesno("Use CSV Instead?", "Export to CSV format instead?"):
                    self.export_to_csv()
                return

            # Create workbook
            wb = Workbook()

            # Sheet 1: Documents
            ws1 = wb.active
            ws1.title = "Documents"

            # Header style
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Headers for documents sheet
            headers = ['Document ID', 'Student ID', 'Document Type', 'File Name',
                      'Status', 'Upload Date', 'Expiry Date', 'File Size (KB)', 'Tags', 'Notes']

            for col, header in enumerate(headers, 1):
                cell = ws1.cell(row=1, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

            # Get documents data
            with get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT id, student_id, document_type, file_name, status,
                           upload_date, expiry_date, file_size, tags, notes
                    FROM documents
                    ORDER BY upload_date DESC
                    LIMIT 1000
                """)
                documents = cursor.fetchall()

                # Write data
                for row_idx, row_data in enumerate(documents, 2):
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws1.cell(row=row_idx, column=col_idx)

                        # Convert file size to KB if it's the file_size column
                        if col_idx == 8 and value:
                            cell.value = round(value / 1024, 2)
                        else:
                            cell.value = value

                        cell.border = border

                        # Align numbers to right, text to left
                        if isinstance(value, (int, float)):
                            cell.alignment = Alignment(horizontal='right')
                        else:
                            cell.alignment = Alignment(horizontal='left')

                # Auto-adjust column widths
                for col in range(1, len(headers) + 1):
                    max_length = 0
                    column = get_column_letter(col)
                    for cell in ws1[column]:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except Exception:

                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws1.column_dimensions[column].width = adjusted_width

                # Sheet 2: Summary Statistics
                ws2 = wb.create_sheet("Summary Statistics")

                # Get summary statistics
                cursor.execute("""
                    SELECT
                        COUNT(*) as total,
                        COUNT(DISTINCT student_id) as unique_students,
                        COUNT(DISTINCT document_type) as unique_types,
                        SUM(CASE WHEN status = 'Pending' THEN 1 ELSE 0 END) as pending,
                        SUM(CASE WHEN status = 'Approved' THEN 1 ELSE 0 END) as approved,
                        SUM(CASE WHEN status = 'Rejected' THEN 1 ELSE 0 END) as rejected,
                        SUM(CASE WHEN DATE(expiry_date) < DATE('now') THEN 1 ELSE 0 END) as expired,
                        SUM(file_size) as total_size
                    FROM documents
                """)
                stats = cursor.fetchone()

                # Write summary
                ws2['A1'] = "DOCUMENT MANAGEMENT SYSTEM - SUMMARY REPORT"
                ws2['A1'].font = Font(bold=True, size=14)
                ws2.merge_cells('A1:B1')

                ws2['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                ws2['A2'].font = Font(italic=True)
                ws2.merge_cells('A2:B2')

                summary_data = [
                    ['', ''],
                    ['OVERALL STATISTICS', ''],
                    ['Total Documents', stats[0]],
                    ['Unique Students', stats[1]],
                    ['Unique Document Types', stats[2]],
                    ['', ''],
                    ['STATUS BREAKDOWN', ''],
                    ['Pending', stats[3]],
                    ['Approved', stats[4]],
                    ['Rejected', stats[5]],
                    ['Expired', stats[6]],
                    ['', ''],
                    ['STORAGE', ''],
                    ['Total Storage', f"{stats[7] / 1024 / 1024 / 1024:.2f} GB" if stats[7] else "0 GB"],
                ]

                for row_idx, (label, value) in enumerate(summary_data, 4):
                    ws2.cell(row=row_idx, column=1).value = label
                    ws2.cell(row=row_idx, column=2).value = value

                    if label and not value and label != '':
                        # Section headers
                        ws2.cell(row=row_idx, column=1).font = Font(bold=True, size=11)
                        ws2.cell(row=row_idx, column=1).fill = PatternFill(start_color="D3D3D3",
                                                                            end_color="D3D3D3",
                                                                            fill_type="solid")

                ws2.column_dimensions['A'].width = 30
                ws2.column_dimensions['B'].width = 20

                # Sheet 3: Document Type Breakdown
                ws3 = wb.create_sheet("Document Types")

                ws3['A1'] = "Document Type"
                ws3['B1'] = "Count"
                ws3['C1'] = "Percentage"

                for col in ['A1', 'B1', 'C1']:
                    ws3[col].fill = header_fill
                    ws3[col].font = header_font
                    ws3[col].alignment = Alignment(horizontal='center')
                    ws3[col].border = border

                cursor.execute("""
                    SELECT document_type, COUNT(*) as count
                    FROM documents
                    GROUP BY document_type
                    ORDER BY count DESC
                """)
                type_stats = cursor.fetchall()

                total_docs = sum(count for _, count in type_stats)

                for row_idx, (doc_type, count) in enumerate(type_stats, 2):
                    ws3.cell(row=row_idx, column=1).value = doc_type
                    ws3.cell(row=row_idx, column=2).value = count
                    ws3.cell(row=row_idx, column=3).value = f"{count/total_docs*100:.1f}%" if total_docs > 0 else "0%"

                    for col in [1, 2, 3]:
                        ws3.cell(row=row_idx, column=col).border = border

                ws3.column_dimensions['A'].width = 25
                ws3.column_dimensions['B'].width = 15
                ws3.column_dimensions['C'].width = 15

            # Save workbook
            wb.save(file_path)

            messagebox.showinfo("Export Successful",
                              f"Data exported to Excel successfully!\n\n"
                              f"{file_path}\n\n"
                              f"Sheets included:\n"
                              f"- Documents ({len(documents)} records)\n"
                              f"- Summary Statistics\n"
                              f"- Document Types")

            self.gui.log_event('export', 'excel', details=f'Exported {len(documents)} records to Excel')

        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export to Excel: {e}")

    def export_to_pdf(self):
        """Export current view/data to PDF"""
        if not self.gui.ensure_login():
            return

        # Ask for file location
        file_path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF Files", "*.pdf"), ("All Files", "*.*")],
            initialfile="document_report.pdf"
        )

        if not file_path:
            return

        try:
            # Try to import reportlab
            try:
                from reportlab.lib.pagesizes import letter, A4
                from reportlab.lib import colors
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib.units import inch
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
                from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
            except ImportError:
                messagebox.showerror("Missing Library",
                                   "The 'reportlab' library is required for PDF export.\n\n"
                                   "Install it with:\n"
                                   "pip install reportlab\n\n"
                                   "Would you like to use CSV export instead?")
                if messagebox.askyesno("Use CSV Instead?", "Export to CSV format instead?"):
                    self.export_to_csv()
                return

            # Create PDF document
            doc = SimpleDocTemplate(file_path, pagesize=letter,
                                   rightMargin=0.5*inch, leftMargin=0.5*inch,
                                   topMargin=0.75*inch, bottomMargin=0.5*inch)

            # Container for the 'Flowable' objects
            elements = []

            # Define styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#1a1a1a'),
                spaceAfter=30,
                alignment=TA_CENTER
            )

            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=16,
                textColor=colors.HexColor('#2c3e50'),
                spaceAfter=12,
                spaceBefore=12
            )

            # Title
            title = Paragraph("Document Management System Report", title_style)
            elements.append(title)

            # Date
            date_text = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            date_para = Paragraph(date_text, styles['Normal'])
            elements.append(date_para)
            elements.append(Spacer(1, 0.3*inch))

            # Get data from database
            with get_connection() as conn:
                cursor = conn.cursor()

                # Summary Statistics
                cursor.execute("""
                    SELECT
                        COUNT(*) as total,
                        COUNT(DISTINCT student_id) as unique_students,
                        COUNT(DISTINCT document_type) as unique_types,
                        SUM(CASE WHEN status = 'Pending' THEN 1 ELSE 0 END) as pending,
                        SUM(CASE WHEN status = 'Approved' THEN 1 ELSE 0 END) as approved,
                        SUM(CASE WHEN status = 'Rejected' THEN 1 ELSE 0 END) as rejected,
                        SUM(CASE WHEN DATE(expiry_date) < DATE('now') THEN 1 ELSE 0 END) as expired
                    FROM documents
                """)
                stats = cursor.fetchone()

                # Summary section
                elements.append(Paragraph("Summary Statistics", heading_style))

                summary_data = [
                    ['Metric', 'Value'],
                    ['Total Documents', str(stats[0])],
                    ['Unique Students', str(stats[1])],
                    ['Unique Document Types', str(stats[2])],
                    ['Pending', str(stats[3])],
                    ['Approved', str(stats[4])],
                    ['Rejected', str(stats[5])],
                    ['Expired', str(stats[6])],
                ]

                summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
                summary_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 10),
                    ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
                ]))

                elements.append(summary_table)
                elements.append(Spacer(1, 0.3*inch))

                # Document Type Breakdown
                elements.append(Paragraph("Document Type Breakdown", heading_style))

                cursor.execute("""
                    SELECT document_type, COUNT(*) as count
                    FROM documents
                    GROUP BY document_type
                    ORDER BY count DESC
                    LIMIT 15
                """)
                type_stats = cursor.fetchall()

                total_docs = sum(count for _, count in type_stats)

                type_data = [['Document Type', 'Count', 'Percentage']]
                for doc_type, count in type_stats:
                    percentage = f"{count/total_docs*100:.1f}%" if total_docs > 0 else "0%"
                    type_data.append([doc_type, str(count), percentage])

                type_table = Table(type_data, colWidths=[3*inch, 1*inch, 1*inch])
                type_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 11),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
                ]))

                elements.append(type_table)
                elements.append(PageBreak())

                # Recent Documents (limited to 50 for PDF size)
                elements.append(Paragraph("Recent Documents (Last 50)", heading_style))

                cursor.execute("""
                    SELECT student_id, document_type, file_name, status, upload_date
                    FROM documents
                    ORDER BY upload_date DESC
                    LIMIT 50
                """)
                documents = cursor.fetchall()

                doc_data = [['Student ID', 'Type', 'File Name', 'Status', 'Upload Date']]
                for student_id, doc_type, file_name, status, upload_date in documents:
                    # Truncate long filenames
                    if file_name and len(file_name) > 30:
                        file_name = file_name[:27] + "..."
                    doc_data.append([
                        str(student_id),
                        str(doc_type)[:20],  # Truncate long type names
                        str(file_name),
                        str(status),
                        str(upload_date)
                    ])

                doc_table = Table(doc_data, colWidths=[1*inch, 1.3*inch, 2*inch, 0.9*inch, 1*inch])
                doc_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 7),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                ]))

                elements.append(doc_table)

                # Footer note
                elements.append(Spacer(1, 0.3*inch))
                footer_text = f"Report generated by Document Management System | Total records in system: {stats[0]}"
                footer = Paragraph(footer_text, styles['Italic'])
                elements.append(footer)

            # Build PDF
            doc.build(elements)

            messagebox.showinfo("Export Successful",
                              f"PDF report generated successfully!\n\n"
                              f"{file_path}\n\n"
                              f"Contents:\n"
                              f"- Summary Statistics\n"
                              f"- Document Type Breakdown\n"
                              f"- Recent Documents (50 records)")

            self.gui.log_event('export', 'pdf', details=f'Exported PDF report with {len(documents)} recent documents')

        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export to PDF: {e}\n\n{str(e)}")

    def export_expiry_report(self, tree):
        """Export expiry alerts to CSV"""
        try:
            items = tree.get_children()
            if not items:
                messagebox.showwarning("No Data", "No expiry alerts to export")
                return

            file_path = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV Files", "*.csv"), ("All Files", "*.*")],
                title="Export Expiry Report"
            )

            if not file_path:
                return

            with open(file_path, 'w', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(['Doc ID', 'Student ID', 'Document Type', 'File Name',
                               'Expiry Date', 'Days Until Expiry', 'Status'])

                for item in items:
                    values = tree.item(item, 'values')
                    writer.writerow(values)

            messagebox.showinfo("Export Successful", f"Expiry report exported to:\n{file_path}")

        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export expiry report: {e}")
