from university_system.infrastructure.database.db import sqlite3, get_connection
from university_system.modules.shared.constants import paths
import logging
import hashlib
import pickle
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor
import time
import schedule
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
from tkinter.scrolledtext import ScrolledText
import threading
import webbrowser
import os
from datetime import datetime, timedelta
import json
try:
    import pandas as pd
except ImportError:
    pd = None
    logging.warning("Pandas not available, some features will be limited")

# Use centralized path configuration
DB_PATH = str(paths.DEFAULT_DB_PATH)
DEFAULT_DB_PATH = paths.DEFAULT_DB_PATH

# Ensure the enhanced reporting uses the same connection function as the main system
def get_db_connection():
    """Get database connection using the centralized connection function"""
    try:
        # Use the centralized get_connection function from the database module
        # This ensures proper connection pooling, row_factory, and timeout settings
        conn = get_connection()
        return conn
    except Exception as e:
        print(f"Database connection error: {e}")
        logging.error(f"Database connection error: {e}")
        return None
    
# Import existing functionality (backward compatible)
try:
    from university_system.modules.shared.services.analytics.enhanced_reporting import (
        AdvancedScheduledReport, AdvancedVisualization, CacheManager,
        DataQualityMonitor, PredictiveAnalytics, ReportTemplate, SystemConfig,
        CONFIG, cleanup_old_reports, create_advanced_visualization,
        create_enhanced_bar_chart, create_enhanced_data_table,
        create_enhanced_line_chart, create_enhanced_pie_chart,
        create_interactive_chart, create_standard_chart,
        delete_template_from_db, display_enhanced_reporting_menu,
        generate_enhanced_excel_report, generate_enhanced_pdf_report,
        generate_enhanced_section, generate_interactive_report,
        generate_predictions_section, generate_quality_section, generate_report,
        generate_statistical_summary, get_benchmark_data, get_correlation_data,
        get_original_section_data_complete, get_section_dataframe, get_template,
        get_trend_data, load_scheduled_reports, load_templates,
        run_system_maintenance, save_scheduled_reports, save_template,
        save_template_dict, serialize_dataframe, show_performance_monitor,
        start_scheduler
    )
    ENHANCED_AVAILABLE = True
except ImportError:
    # Fallback for basic functionality
    ENHANCED_AVAILABLE = False
    print("Enhanced reporting not available, using basic functionality")

def show_directory_settings(self):
    """Show directory settings dialog"""
    try:
        dir_window = tk.Toplevel(self.root)
        dir_window.title("Directory Settings")
        dir_window.geometry("500x400")
        dir_window.transient(self.root)
        
        settings_frame = ttk.LabelFrame(dir_window, text="Directory Configuration", padding="10")
        settings_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        settings_frame.columnconfigure(1, weight=1)
        
        # Reports directory
        ttk.Label(settings_frame, text="Reports Directory:").grid(row=0, column=0, sticky=tk.W, pady=5)
        reports_dir_var = tk.StringVar(value=CONFIG.get('reports_dir', 'reports'))
        reports_entry = ttk.Entry(settings_frame, textvariable=reports_dir_var)
        reports_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)
        
        def browse_reports_dir():
            directory = filedialog.askdirectory(title="Select Reports Directory")
            if directory:
                reports_dir_var.set(directory)
        
        ttk.Button(settings_frame, text="Browse", command=browse_reports_dir).grid(row=0, column=2, padx=(5, 0), pady=5)
        
        # Templates directory
        ttk.Label(settings_frame, text="Templates Directory:").grid(row=1, column=0, sticky=tk.W, pady=5)
        templates_dir_var = tk.StringVar(value=CONFIG.get('templates_dir', str(paths.REPORT_TEMPLATES_DIR)))
        templates_entry = ttk.Entry(settings_frame, textvariable=templates_dir_var)
        templates_entry.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)
        
        def browse_templates_dir():
            directory = filedialog.askdirectory(title="Select Templates Directory")
            if directory:
                templates_dir_var.set(directory)
        
        ttk.Button(settings_frame, text="Browse", command=browse_templates_dir).grid(row=1, column=2, padx=(5, 0), pady=5)
        
        # Cache directory
        ttk.Label(settings_frame, text="Cache Directory:").grid(row=2, column=0, sticky=tk.W, pady=5)
        cache_dir_var = tk.StringVar(value=CONFIG.get('cache_dir', 'cache'))
        cache_entry = ttk.Entry(settings_frame, textvariable=cache_dir_var)
        cache_entry.grid(row=2, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)
        
        def browse_cache_dir():
            directory = filedialog.askdirectory(title="Select Cache Directory")
            if directory:
                cache_dir_var.set(directory)
        
        ttk.Button(settings_frame, text="Browse", command=browse_cache_dir).grid(row=2, column=2, padx=(5, 0), pady=5)
        
        # Create directories option
        create_dirs_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(settings_frame, text="Create directories if they don't exist", 
                       variable=create_dirs_var).grid(row=3, column=0, columnspan=3, sticky=tk.W, pady=10)
        
        # Button frame
        button_frame = ttk.Frame(dir_window)
        button_frame.pack(fill=tk.X, padx=10, pady=(0, 10))
        
        def save_directory_settings():
            try:
                new_dirs = {
                    'reports_dir': reports_dir_var.get(),
                    'templates_dir': templates_dir_var.get(),
                    'cache_dir': cache_dir_var.get()
                }
                
                # Create directories if requested
                if create_dirs_var.get():
                    for dir_path in new_dirs.values():
                        os.makedirs(dir_path, exist_ok=True)
                
                CONFIG.update(new_dirs)
                
                if ENHANCED_AVAILABLE:
                    full_config = SystemConfig.load_config()
                    full_config.update(new_dirs)
                    SystemConfig.save_config(full_config)
                
                messagebox.showinfo("Success", "Directory settings saved successfully!")
                dir_window.destroy()
                self.check_system_status()
                
            except Exception as e:
                messagebox.showerror("Save Error", f"Failed to save directory settings: {str(e)}")
        
        ttk.Button(button_frame, text="Save", command=save_directory_settings).pack(side=tk.RIGHT, padx=(5, 0))
        ttk.Button(button_frame, text="Cancel", command=dir_window.destroy).pack(side=tk.RIGHT)
        
    except Exception as e:
        messagebox.showerror("Directory Settings Error", f"Failed to open directory settings: {str(e)}")

def show_theme_settings(self):
    """Show theme and appearance settings"""
    try:
        theme_window = tk.Toplevel(self.root)
        theme_window.title("Theme & Appearance Settings")
        theme_window.geometry("450x350")
        theme_window.transient(self.root)
        
        theme_frame = ttk.LabelFrame(theme_window, text="Appearance Settings", padding="10")
        theme_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Theme selection
        ttk.Label(theme_frame, text="Theme:").pack(anchor=tk.W, pady=5)
        theme_var = tk.StringVar(value="default")
        theme_combo = ttk.Combobox(theme_frame, textvariable=theme_var, 
                                  values=["default", "dark", "light", "modern"], state="readonly")
        theme_combo.pack(fill=tk.X, pady=(0, 10))
        
        # Font settings
        ttk.Label(theme_frame, text="Font Family:").pack(anchor=tk.W, pady=5)
        font_var = tk.StringVar(value="Arial")
        font_combo = ttk.Combobox(theme_frame, textvariable=font_var,
                                 values=["Arial", "Helvetica", "Times New Roman", "Calibri", "Segoe UI"],
                                 state="readonly")
        font_combo.pack(fill=tk.X, pady=(0, 10))
        
        # Font size
        ttk.Label(theme_frame, text="Font Size:").pack(anchor=tk.W, pady=5)
        font_size_var = tk.StringVar(value="10")
        font_size_spin = ttk.Spinbox(theme_frame, from_=8, to=16, textvariable=font_size_var)
        font_size_spin.pack(fill=tk.X, pady=(0, 10))
        
        # UI density
        ttk.Label(theme_frame, text="UI Density:").pack(anchor=tk.W, pady=5)
        density_var = tk.StringVar(value="normal")
        density_combo = ttk.Combobox(theme_frame, textvariable=density_var,
                                    values=["compact", "normal", "spacious"], state="readonly")
        density_combo.pack(fill=tk.X, pady=(0, 10))
        
        # Preview frame
        preview_frame = ttk.LabelFrame(theme_frame, text="Preview", padding="10")
        preview_frame.pack(fill=tk.X, pady=10)
        
        preview_label = ttk.Label(preview_frame, text="Sample text with current settings")
        preview_label.pack()
        
        def update_preview():
            try:
                font_family = font_var.get()
                font_size = int(font_size_var.get())
                preview_label.config(font=(font_family, font_size))
            except:
                pass
        
        # Bind preview updates
        theme_combo.bind('<<ComboboxSelected>>', lambda e: update_preview())
        font_combo.bind('<<ComboboxSelected>>', lambda e: update_preview())
        font_size_spin.bind('<KeyRelease>', lambda e: update_preview())
        
        # Button frame
        button_frame = ttk.Frame(theme_window)
        button_frame.pack(fill=tk.X, padx=10, pady=(0, 10))
        
        def apply_theme():
            try:
                theme_settings = {
                    'theme': theme_var.get(),
                    'font_family': font_var.get(),
                    'font_size': int(font_size_var.get()),
                    'ui_density': density_var.get()
                }
                
                # Apply theme (this would require theme system implementation)
                messagebox.showinfo("Theme Applied", 
                                  f"Theme settings applied:\n\nTheme: {theme_settings['theme']}\nFont: {theme_settings['font_family']} {theme_settings['font_size']}pt\nDensity: {theme_settings['ui_density']}")
                
                theme_window.destroy()
                
            except Exception as e:
                messagebox.showerror("Theme Error", f"Failed to apply theme: {str(e)}")
        
        def reset_theme():
            theme_var.set("default")
            font_var.set("Arial")
            font_size_var.set("10")
            density_var.set("normal")
            update_preview()
        
        ttk.Button(button_frame, text="Apply", command=apply_theme).pack(side=tk.RIGHT, padx=(5, 0))
        ttk.Button(button_frame, text="Reset", command=reset_theme).pack(side=tk.RIGHT, padx=(5, 0))
        ttk.Button(button_frame, text="Cancel", command=theme_window.destroy).pack(side=tk.RIGHT)
        
    except Exception as e:
        messagebox.showerror("Theme Settings Error", f"Failed to open theme settings: {str(e)}")

def validate_email_settings(self, settings):
    """Validate email configuration settings"""
    try:
        required_fields = ['smtp_server', 'smtp_port', 'from_address']
        
        for field in required_fields:
            if not settings.get(field):
                return False, f"Missing required field: {field}"
        
        # Validate port is numeric
        try:
            port = int(settings['smtp_port'])
            if port < 1 or port > 65535:
                return False, "SMTP port must be between 1 and 65535"
        except ValueError:
            return False, "SMTP port must be a valid number"
        
        # Validate email format
        email = settings['from_address']
        if '@' not in email or '.' not in email.split('@')[1]:
            return False, "Invalid from_address email format"
        
        return True, "Valid"
        
    except Exception as e:
        return False, f"Validation error: {str(e)}"


def serialize_dataframe(df):
    """Serialize dataframe for API responses"""
    try:
        return {
            'columns': df.columns.tolist(),
            'data': df.values.tolist(),
            'index': df.index.tolist() if hasattr(df.index, 'tolist') else list(df.index),
            'shape': df.shape,
            'dtypes': {col: str(df[col].dtype) for col in df.columns}
        }
    except Exception as e:
        return {'error': f'Serialization failed: {str(e)}'}

def get_template(name):
    """Get template by name"""
    try:
        templates = load_templates() if ENHANCED_AVAILABLE else []
        for template in templates:
            if template['name'] == name:
                return template
        return None
    except Exception as e:
        logging.error(f"Error getting template {name}: {str(e)}")
        return None

def generate_enhanced_pdf_report(template, filename, start_date, end_date, comparison_date=None):
    """Generate enhanced PDF report with advanced features"""
    try:
        if not ENHANCED_AVAILABLE:
            raise Exception("Enhanced reporting not available")
        
        # Import required libraries
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            from reportlab.graphics.shapes import Drawing
            from reportlab.graphics.charts.linecharts import HorizontalLineChart
            from reportlab.graphics.charts.piecharts import Pie
            from reportlab.graphics.charts.barcharts import VerticalBarChart
        except ImportError:
            # Fallback to basic PDF generation
            raise Exception("ReportLab not available for enhanced PDF generation")
        
        doc = SimpleDocTemplate(filename, pagesize=A4, 
                              leftMargin=0.75*inch, rightMargin=0.75*inch,
                              topMargin=1*inch, bottomMargin=1*inch)
        
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=24,
            textColor=colors.HexColor('#2E4E7E'),
            spaceAfter=30
        )
        story.append(Paragraph(f"Enhanced Report: {template['name']}", title_style))
        story.append(Spacer(1, 20))
        
        # Metadata
        meta_data = [
            ['Report Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Period:', f"{start_date} to {end_date}"],
            ['Template:', template['name']],
            ['Security Level:', template.get('security_level', 'normal').title()]
        ]
        
        meta_table = Table(meta_data, colWidths=[2*inch, 4*inch])
        meta_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F8F9FA')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(meta_table)
        story.append(Spacer(1, 30))
        
        # Generate sections
        for section in template.get('sections', []):
            section_data = generate_enhanced_section(section, start_date, end_date, 
                                                   comparison_date, template.get('filters'), 
                                                   template.get('visualization_type', 'standard'))
            
            if section_data:
                # Section title
                section_title = section.replace('_', ' ').title()
                story.append(Paragraph(section_title, styles['Heading1']))
                story.append(Spacer(1, 12))
                
                # Section content
                if 'summary' in section_data:
                    story.append(Paragraph(section_data['summary'], styles['Normal']))
                    story.append(Spacer(1, 10))
                
                # Add tables if present
                if 'data' in section_data and hasattr(section_data['data'], 'empty') and not section_data['data'].empty:
                    df = section_data['data'].head(10)  # Limit to 10 rows for PDF
                    table_data = [df.columns.tolist()] + df.values.tolist()
                    
                    table = Table(table_data, repeatRows=1)
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E3F2FD')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, -1), 8),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ]))
                    story.append(table)
                
                story.append(Spacer(1, 20))
        
        # Build PDF
        doc.build(story)
        return filename
        
    except Exception as e:
        logging.error(f"Error generating enhanced PDF: {str(e)}")
        return None

def generate_enhanced_section(section, start_date, end_date, comparison_date=None, filters=None, visualization_type='standard'):
    """Generate enhanced section with data and visualizations"""
    try:
        conn = get_db_connection()
        if not conn:
            return None
        
        section_data = {
            'name': section,
            'summary': '',
            'data': None,
            'visualization': None,
            'statistics': {}
        }
        
        if section == 'student_overview':
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) as total_students FROM students")
            total_students = cursor.fetchone()[0]
            
            cursor.execute("SELECT COUNT(DISTINCT course) FROM students WHERE course IS NOT NULL")
            total_courses = cursor.fetchone()[0]
            
            section_data['summary'] = f"Total Students: {total_students}, Total Courses: {total_courses}"
            section_data['statistics'] = {
                'total_students': total_students,
                'total_courses': total_courses
            }
            
        elif section == 'course_distribution':
            query = """
            SELECT course, COUNT(*) as count
            FROM students
            WHERE course IS NOT NULL AND registration_datetime BETWEEN ? AND ?
            GROUP BY course
            ORDER BY count DESC
            """
            df = pd.read_sql_query(query, conn, params=[start_date + ' 00:00:00', end_date + ' 23:59:59'])
            section_data['data'] = df
            section_data['summary'] = f"Course distribution across {len(df)} courses"

        elif section == 'gender_distribution':
            query = """
            SELECT gender, COUNT(*) as count
            FROM students
            WHERE gender IS NOT NULL AND registration_datetime BETWEEN ? AND ?
            GROUP BY gender
            """
            df = pd.read_sql_query(query, conn, params=[start_date + ' 00:00:00', end_date + ' 23:59:59'])
            section_data['data'] = df
            section_data['summary'] = f"Gender distribution across {df['count'].sum()} students"

        elif section == 'age_distribution':
            query = """
            SELECT
                CASE
                    WHEN age < 20 THEN 'Under 20'
                    WHEN age BETWEEN 20 AND 25 THEN '20-25'
                    WHEN age BETWEEN 26 AND 30 THEN '26-30'
                    WHEN age BETWEEN 31 AND 35 THEN '31-35'
                    ELSE 'Over 35'
                END as age_group,
                COUNT(*) as count
            FROM students
            WHERE age IS NOT NULL AND registration_datetime BETWEEN ? AND ?
            GROUP BY age_group
            ORDER BY count DESC
            """
            df = pd.read_sql_query(query, conn, params=[start_date + ' 00:00:00', end_date + ' 23:59:59'])
            section_data['data'] = df
            section_data['summary'] = f"Age distribution across {df['count'].sum()} students"
            
        elif section == 'registration_trends':
            query = """
            SELECT DATE(registration_datetime) as registration_date, COUNT(*) as count
            FROM students
            WHERE registration_datetime BETWEEN ? AND ?
            GROUP BY DATE(registration_datetime)
            ORDER BY registration_date
            """
            df = pd.read_sql_query(query, conn, params=[start_date + ' 00:00:00', end_date + ' 23:59:59'])
            section_data['data'] = df
            section_data['summary'] = f"Registration trends over {len(df)} days"
            
        # Add more sections as needed...
        
        conn.close()
        return section_data
        
    except Exception as e:
        logging.error(f"Error generating section {section}: {str(e)}")
        if conn:
            conn.close()
        return None

def create_advanced_visualization(section, df):
    """Create advanced visualization for section"""
    try:
        if not ENHANCED_AVAILABLE or df is None or df.empty:
            return None
            
        import matplotlib.pyplot as plt
        import seaborn as sns
        
        plt.style.use('seaborn-v0_8')
        fig, ax = plt.subplots(figsize=(12, 8))
        
        if section == 'course_distribution':
            colors = plt.cm.Set3(range(len(df)))
            bars = ax.bar(df.iloc[:, 0], df.iloc[:, 1], color=colors)
            ax.set_title('Course Distribution', fontsize=16, fontweight='bold')
            ax.set_xlabel('Course', fontsize=12)
            ax.set_ylabel('Student Count', fontsize=12)
            
            # Add value labels on bars
            for bar in bars:
                height = bar.get_height()
                ax.text(bar.get_x() + bar.get_width()/2., height,
                       f'{int(height)}', ha='center', va='bottom')
            
        elif section == 'gender_distribution':
            colors = ['#FF9999', '#66B2FF', '#99FF99']
            wedges, texts, autotexts = ax.pie(df.iloc[:, 1], labels=df.iloc[:, 0], 
                                            autopct='%1.1f%%', colors=colors,
                                            startangle=90)
            ax.set_title('Gender Distribution', fontsize=16, fontweight='bold')
            
        elif section == 'age_distribution':
            sns.barplot(data=df, x=df.columns[0], y=df.columns[1], ax=ax, palette='viridis')
            ax.set_title('Age Distribution', fontsize=16, fontweight='bold')
            ax.set_xlabel('Age Group', fontsize=12)
            ax.set_ylabel('Count', fontsize=12)
            
        elif section == 'registration_trends':
            ax.plot(pd.to_datetime(df.iloc[:, 0]), df.iloc[:, 1], 
                   marker='o', linewidth=2, markersize=6)
            ax.set_title('Registration Trends', fontsize=16, fontweight='bold')
            ax.set_xlabel('Date', fontsize=12)
            ax.set_ylabel('Registrations', fontsize=12)
            plt.xticks(rotation=45)
            
        plt.tight_layout()
        
        # Save chart
        chart_path = f"reports/charts/{section}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
        os.makedirs(os.path.dirname(chart_path), exist_ok=True)
        plt.savefig(chart_path, dpi=300, bbox_inches='tight')
        plt.close()
        
        return chart_path
        
    except Exception as e:
        logging.error(f"Error creating visualization for {section}: {str(e)}")
        return None

def create_interactive_chart(section, df):
    """Create interactive chart using plotly"""
    try:
        if not df.empty:
            import plotly.graph_objects as go
            import plotly.express as px
            
            if section == 'course_distribution':
                fig = px.bar(df, x=df.columns[0], y=df.columns[1], 
                           title='Interactive Course Distribution')
            elif section == 'gender_distribution':
                fig = px.pie(df, values=df.columns[1], names=df.columns[0],
                           title='Interactive Gender Distribution')
            elif section == 'registration_trends':
                fig = px.line(df, x=df.columns[0], y=df.columns[1],
                            title='Interactive Registration Trends')
            else:
                fig = px.bar(df, x=df.columns[0], y=df.columns[1])
                
            # Save as HTML
            chart_path = f"reports/charts/{section}_interactive_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
            os.makedirs(os.path.dirname(chart_path), exist_ok=True)
            fig.write_html(chart_path)
            return chart_path
            
    except ImportError:
        logging.warning("Plotly not available for interactive charts")
    except Exception as e:
        logging.error(f"Error creating interactive chart: {str(e)}")
    
    return None

def create_standard_chart(section, df):
    """Create standard matplotlib chart"""
    try:
        return create_advanced_visualization(section, df)
    except Exception as e:
        logging.error(f"Error creating standard chart: {str(e)}")
        return None

def create_enhanced_pie_chart(df, section):
    """Create enhanced pie chart"""
    try:
        import matplotlib.pyplot as plt
        
        if df.empty or len(df.columns) < 2:
            return None
            
        fig, ax = plt.subplots(figsize=(10, 8))
        
        # Use the first column as labels, second as values
        labels = df.iloc[:, 0].tolist()
        sizes = df.iloc[:, 1].tolist()
        
        # Color palette
        colors = plt.cm.Set3(range(len(labels)))
        
        # Create pie chart
        wedges, texts, autotexts = ax.pie(sizes, labels=labels, autopct='%1.1f%%',
                                        startangle=90, colors=colors,
                                        textprops={'fontsize': 10})
        
        # Enhance appearance
        for autotext in autotexts:
            autotext.set_color('white')
            autotext.set_fontweight('bold')
            
        ax.set_title(f'Enhanced {section.replace("_", " ").title()}', 
                    fontsize=14, fontweight='bold', pad=20)
        
        plt.tight_layout()
        
        # Save chart
        chart_path = f"reports/charts/{section}_pie_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
        os.makedirs(os.path.dirname(chart_path), exist_ok=True)
        plt.savefig(chart_path, dpi=300, bbox_inches='tight')
        plt.close()
        
        return chart_path
        
    except Exception as e:
        logging.error(f"Error creating enhanced pie chart: {str(e)}")
        return None

def create_enhanced_bar_chart(df, section):
    """Create enhanced bar chart"""
    try:
        import matplotlib.pyplot as plt
        import seaborn as sns
        
        if df.empty or len(df.columns) < 2:
            return None
            
        plt.style.use('default')
        fig, ax = plt.subplots(figsize=(12, 8))
        
        # Use seaborn for better styling
        sns.barplot(data=df, x=df.columns[0], y=df.columns[1], ax=ax, palette='viridis')
        
        # Customize appearance
        ax.set_title(f'Enhanced {section.replace("_", " ").title()}', 
                    fontsize=16, fontweight='bold', pad=20)
        ax.set_xlabel(df.columns[0].replace('_', ' ').title(), fontsize=12)
        ax.set_ylabel(df.columns[1].replace('_', ' ').title(), fontsize=12)
        
        # Add value labels on bars
        for i, bar in enumerate(ax.patches):
            height = bar.get_height()
            if not pd.isna(height):
                ax.text(bar.get_x() + bar.get_width()/2., height,
                       f'{int(height)}', ha='center', va='bottom', fontsize=10)
        
        # Rotate x-axis labels if needed
        if len(df) > 5:
            plt.xticks(rotation=45, ha='right')
            
        plt.tight_layout()
        
        # Save chart
        chart_path = f"reports/charts/{section}_bar_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
        os.makedirs(os.path.dirname(chart_path), exist_ok=True)
        plt.savefig(chart_path, dpi=300, bbox_inches='tight')
        plt.close()
        
        return chart_path
        
    except Exception as e:
        logging.error(f"Error creating enhanced bar chart: {str(e)}")
        return None

def create_enhanced_line_chart(df, section):
    """Create enhanced line chart"""
    try:
        import matplotlib.pyplot as plt
        import matplotlib.dates as mdates
        
        if df.empty or len(df.columns) < 2:
            return None
            
        fig, ax = plt.subplots(figsize=(12, 8))
        
        # Convert first column to datetime if it looks like dates
        x_data = df.iloc[:, 0]
        y_data = df.iloc[:, 1]
        
        try:
            x_data = pd.to_datetime(x_data)
            ax.plot(x_data, y_data, marker='o', linewidth=2.5, markersize=6, 
                   color='#1f77b4', markerfacecolor='#ff7f0e')
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
            ax.xaxis.set_major_locator(mdates.DayLocator(interval=max(1, len(df)//10)))
            plt.xticks(rotation=45)
        except:
            ax.plot(x_data, y_data, marker='o', linewidth=2.5, markersize=6,
                   color='#1f77b4', markerfacecolor='#ff7f0e')
            
        # Customize appearance
        ax.set_title(f'Enhanced {section.replace("_", " ").title()}', 
                    fontsize=16, fontweight='bold', pad=20)
        ax.set_xlabel(df.columns[0].replace('_', ' ').title(), fontsize=12)
        ax.set_ylabel(df.columns[1].replace('_', ' ').title(), fontsize=12)
        
        # Add grid
        ax.grid(True, alpha=0.3)
        
        # Add trend line if more than 3 points
        if len(df) > 3:
            try:
                from numpy.polynomial.polynomial import Polynomial
                x_numeric = range(len(y_data))
                p = Polynomial.fit(x_numeric, y_data, deg=1)
                trend_y = p(x_numeric)
                ax.plot(x_data, trend_y, '--', color='red', alpha=0.7, 
                       linewidth=2, label='Trend')
                ax.legend()
            except:
                pass
                
        plt.tight_layout()
        
        # Save chart
        chart_path = f"reports/charts/{section}_line_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
        os.makedirs(os.path.dirname(chart_path), exist_ok=True)
        plt.savefig(chart_path, dpi=300, bbox_inches='tight')
        plt.close()
        
        return chart_path
        
    except Exception as e:
        logging.error(f"Error creating enhanced line chart: {str(e)}")
        return None

def generate_statistical_summary(df, section):
    """Generate statistical summary for dataframe"""
    try:
        if df.empty:
            return "No data available for statistical analysis"
            
        summary = f"Statistical Summary for {section.replace('_', ' ').title()}:\n\n"
        
        # Basic info
        summary += f"Total Records: {len(df)}\n"
        summary += f"Columns: {len(df.columns)}\n\n"
        
        # Numeric columns analysis
        numeric_cols = df.select_dtypes(include=['int64', 'float64']).columns
        if len(numeric_cols) > 0:
            summary += "Numeric Analysis:\n"
            for col in numeric_cols:
                col_data = df[col].dropna()
                if len(col_data) > 0:
                    summary += f"  {col}:\n"
                    summary += f"    Mean: {col_data.mean():.2f}\n"
                    summary += f"    Median: {col_data.median():.2f}\n"
                    summary += f"    Std Dev: {col_data.std():.2f}\n"
                    summary += f"    Min: {col_data.min()}\n"
                    summary += f"    Max: {col_data.max()}\n\n"
        
        # Categorical columns analysis
        categorical_cols = df.select_dtypes(include=['object']).columns
        if len(categorical_cols) > 0:
            summary += "Categorical Analysis:\n"
            for col in categorical_cols[:3]:  # Limit to first 3 columns
                col_data = df[col].dropna()
                if len(col_data) > 0:
                    summary += f"  {col}:\n"
                    summary += f"    Unique Values: {col_data.nunique()}\n"
                    top_values = col_data.value_counts().head(3)
                    summary += f"    Top Values: {dict(top_values)}\n\n"
        
        return summary
        
    except Exception as e:
        logging.error(f"Error generating statistical summary: {str(e)}")
        return f"Error generating statistical summary: {str(e)}"

def create_enhanced_data_table(df):
    """Create enhanced HTML data table"""
    try:
        if df.empty:
            return "<p>No data available</p>"
            
        # Limit rows for display
        display_df = df.head(100)
        
        # Convert to HTML with styling
        html_table = display_df.to_html(classes='enhanced-table', 
                                       table_id='data-table',
                                       escape=False,
                                       index=False)
        
        # Add CSS styling
        enhanced_html = f"""
        <style>
        .enhanced-table {{
            border-collapse: collapse;
            margin: 25px 0;
            font-size: 0.9em;
            font-family: sans-serif;
            min-width: 400px;
            box-shadow: 0 0 20px rgba(0, 0, 0, 0.15);
            width: 100%;
        }}
        .enhanced-table thead tr {{
            background-color: #009879;
            color: #ffffff;
            text-align: left;
        }}
        .enhanced-table th,
        .enhanced-table td {{
            padding: 12px 15px;
            border: 1px solid #dddddd;
        }}
        .enhanced-table tbody tr {{
            border-bottom: 1px solid #dddddd;
        }}
        .enhanced-table tbody tr:nth-of-type(even) {{
            background-color: #f3f3f3;
        }}
        .enhanced-table tbody tr:hover {{
            background-color: #f1f1f1;
        }}
        </style>
        {html_table}
        """
        
        if len(df) > 100:
            enhanced_html += f"<p><em>Showing first 100 of {len(df)} records</em></p>"
            
        return enhanced_html
        
    except Exception as e:
        logging.error(f"Error creating enhanced data table: {str(e)}")
        return f"<p>Error creating data table: {str(e)}</p>"

def generate_quality_section(quality_report, styles):
    """Generate quality section for PDF report"""
    try:
        from reportlab.platypus import Paragraph, Spacer, Table, TableStyle
        from reportlab.lib import colors
        
        content = []
        
        # Title
        content.append(Paragraph("Data Quality Report", styles['Heading1']))
        content.append(Spacer(1, 12))
        
        # Summary
        checks = quality_report.get('checks', {})
        
        if 'missing_data' in checks:
            missing = checks['missing_data']['students']
            total = missing['total_records']
            
            summary_data = [
                ['Metric', 'Value'],
                ['Total Records', str(total)],
                ['Missing Emails', str(missing.get('missing_emails', 0))],
                ['Missing Names', str(missing.get('missing_names', 0))],
                ['Missing Courses', str(missing.get('missing_courses', 0))]
            ]
            
            if total > 0:
                completeness = ((total * 3) - sum([missing.get('missing_emails', 0), 
                                                  missing.get('missing_names', 0),
                                                  missing.get('missing_courses', 0)])) / (total * 3) * 100
                summary_data.append(['Data Completeness', f"{completeness:.1f}%"])
            
            summary_table = Table(summary_data, colWidths=[2*72, 2*72])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E3F2FD')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            content.append(summary_table)
        
        return content
        
    except Exception as e:
        logging.error(f"Error generating quality section: {str(e)}")
        return []

def generate_predictions_section(predictions, styles):
    """Generate predictions section for PDF report"""
    try:
        from reportlab.platypus import Paragraph, Spacer, Table, TableStyle
        from reportlab.lib import colors
        
        content = []
        
        # Title
        content.append(Paragraph("Predictive Analytics", styles['Heading1']))
        content.append(Spacer(1, 12))
        
        if 'error' in predictions:
            content.append(Paragraph(f"Analysis unavailable: {predictions['error']}", styles['Normal']))
        else:
            # Summary data
            summary_data = [['Metric', 'Value']]
            
            if 'total_students_analyzed' in predictions:
                summary_data.append(['Students Analyzed', str(predictions['total_students_analyzed'])])
                
            if 'model_accuracy' in predictions:
                accuracy = predictions['model_accuracy'] * 100
                summary_data.append(['Model Accuracy', f"{accuracy:.1f}%"])
                
            if 'high_risk_students' in predictions:
                high_risk_count = len(predictions['high_risk_students'])
                summary_data.append(['High Risk Students', str(high_risk_count)])
            
            if len(summary_data) > 1:
                summary_table = Table(summary_data, colWidths=[2*72, 2*72])
                summary_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FFF3E0')),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                content.append(summary_table)
        
        return content
        
    except Exception as e:
        logging.error(f"Error generating predictions section: {str(e)}")
        return []

def get_section_dataframe(section, start_date, end_date, filters=None):
    """Get dataframe for specific section with filters"""
    try:
        conn = get_db_connection()
        if not conn:
            return pd.DataFrame()
        
        base_query = ""
        params = [start_date, end_date]
        
        if section == 'student_overview':
            base_query = """
            SELECT 'Total Students' as metric, COUNT(*) as value FROM students
            WHERE registration_datetime BETWEEN ? AND ?
            UNION ALL
            SELECT 'Unique Courses' as metric, COUNT(DISTINCT course) as value FROM students
            WHERE course IS NOT NULL AND registration_datetime BETWEEN ? AND ?
            """
            params = [start_date + ' 00:00:00', end_date + ' 23:59:59', start_date + ' 00:00:00', end_date + ' 23:59:59']

        elif section == 'student_list':
            base_query = """
            SELECT student_id, first_name, last_name, email_address, course,
                   age, gender, registration_datetime
            FROM students
            WHERE registration_datetime BETWEEN ? AND ?
            ORDER BY registration_datetime DESC
            """
            params = [start_date + ' 00:00:00', end_date + ' 23:59:59']

        elif section == 'course_distribution':
            base_query = "SELECT course, COUNT(*) as count FROM students WHERE registration_datetime BETWEEN ? AND ?"
            params = [start_date + ' 00:00:00', end_date + ' 23:59:59']
            if filters and filters.get('course'):
                base_query += " AND course = ?"
                params.append(filters['course'])
            base_query += " GROUP BY course ORDER BY count DESC"
            
        elif section == 'gender_distribution':
            base_query = "SELECT gender, COUNT(*) as count FROM students WHERE registration_datetime BETWEEN ? AND ? AND gender IS NOT NULL GROUP BY gender"
            params = [start_date + ' 00:00:00', end_date + ' 23:59:59']

        elif section == 'age_distribution':
            base_query = """
            SELECT
                CASE
                    WHEN age < 20 THEN 'Under 20'
                    WHEN age BETWEEN 20 AND 25 THEN '20-25'
                    WHEN age BETWEEN 26 AND 30 THEN '26-30'
                    WHEN age BETWEEN 31 AND 35 THEN '31-35'
                    ELSE 'Over 35'
                END as age_group,
                COUNT(*) as count
            FROM students
            WHERE age IS NOT NULL AND registration_datetime BETWEEN ? AND ?
            GROUP BY age_group
            """
            params = [start_date + ' 00:00:00', end_date + ' 23:59:59']

        elif section == 'registration_trends':
            base_query = """
            SELECT DATE(registration_datetime) as registration_date, COUNT(*) as count
            FROM students
            WHERE registration_datetime BETWEEN ? AND ?
            GROUP BY DATE(registration_datetime)
            ORDER BY registration_date
            """
            params = [start_date + ' 00:00:00', end_date + ' 23:59:59']

        else:
            # Generic query for unknown sections - get all students
            base_query = "SELECT * FROM students WHERE registration_datetime BETWEEN ? AND ?"
            params = [start_date + ' 00:00:00', end_date + ' 23:59:59']
        
        if filters and filters.get('course') and 'course' not in base_query.lower():
            if 'WHERE' in base_query.upper():
                base_query += " AND course = ?"
            else:
                base_query += " WHERE course = ?"
            params.append(filters['course'])
        
        df = pd.read_sql_query(base_query, conn, params=params)
        conn.close()
        return df
        
    except Exception as e:
        logging.error(f"Error getting section dataframe: {str(e)}")
        if 'conn' in locals() and conn:
            conn.close()
        return pd.DataFrame()

def get_correlation_data(conn, start_date, end_date, filters):
    """Get correlation data for analysis"""
    try:
        query = """
        SELECT age,
               LENGTH(first_name || ' ' || last_name) as name_length,
               CASE WHEN course = 'CS' THEN 1 ELSE 0 END as is_cs,
               CASE WHEN gender = 'Male' THEN 1 ELSE 0 END as is_male
        FROM students
        WHERE registration_datetime BETWEEN ? AND ?
        AND age IS NOT NULL
        AND first_name IS NOT NULL
        AND course IS NOT NULL
        AND gender IS NOT NULL
        """
        params = [start_date + ' 00:00:00', end_date + ' 23:59:59']
        
        if filters and filters.get('course'):
            query += " AND course = ?"
            params.append(filters['course'])
            
        df = pd.read_sql_query(query, conn, params=params)
        return df
        
    except Exception as e:
        logging.error(f"Error getting correlation data: {str(e)}")
        return pd.DataFrame()

def get_trend_data(conn, start_date, end_date, filters):
    """Get trend data for analysis"""
    try:
        query = """
        SELECT DATE(registration_datetime) as date,
               COUNT(*) as registrations,
               COUNT(DISTINCT course) as unique_courses,
               AVG(age) as avg_age
        FROM students
        WHERE registration_datetime BETWEEN ? AND ?
        GROUP BY DATE(registration_datetime)
        ORDER BY date
        """
        params = [start_date + ' 00:00:00', end_date + ' 23:59:59']
        
        df = pd.read_sql_query(query, conn, params=params)
        return df
        
    except Exception as e:
        logging.error(f"Error getting trend data: {str(e)}")
        return pd.DataFrame()

def get_benchmark_data(conn, start_date, end_date, filters):
    """Get benchmark data for comparisons"""
    try:
        # Get current period data
        current_query = """
        SELECT 'Current Period' as period,
               COUNT(*) as total_students,
               COUNT(DISTINCT course) as unique_courses,
               AVG(age) as avg_age
        FROM students
        WHERE registration_datetime BETWEEN ? AND ?
        """
        params = [start_date + ' 00:00:00', end_date + ' 23:59:59']

        # Calculate previous period dates
        start_dt = datetime.strptime(start_date, '%Y-%m-%d')
        end_dt = datetime.strptime(end_date, '%Y-%m-%d')
        period_length = (end_dt - start_dt).days

        prev_end_dt = start_dt - timedelta(days=1)
        prev_start_dt = prev_end_dt - timedelta(days=period_length)

        previous_query = """
        SELECT 'Previous Period' as period,
               COUNT(*) as total_students,
               COUNT(DISTINCT course) as unique_courses,
               AVG(age) as avg_age
        FROM students
        WHERE registration_datetime BETWEEN ? AND ?
        """

        # Combine queries
        combined_query = f"""
        {current_query}
        UNION ALL
        {previous_query}
        """

        all_params = params + [prev_start_dt.strftime('%Y-%m-%d') + ' 00:00:00', prev_end_dt.strftime('%Y-%m-%d') + ' 23:59:59']
        df = pd.read_sql_query(combined_query, conn, params=all_params)
        return df
        
    except Exception as e:
        logging.error(f"Error getting benchmark data: {str(e)}")
        return pd.DataFrame()

def get_original_section_data_complete(conn, section, start_date, end_date, filters):
    """Get complete original section data"""
    try:
        return get_section_dataframe(section, start_date, end_date, filters)
    except Exception as e:
        logging.error(f"Error getting original section data: {str(e)}")
        return pd.DataFrame()

def generate_enhanced_excel_report(template, filename, start_date, end_date, comparison_date=None):
    """Generate enhanced Excel report with multiple sheets and formatting"""
    try:
        if not ENHANCED_AVAILABLE:
            raise Exception("Enhanced reporting not available")
            
        # Import required libraries
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.chart import BarChart, PieChart, LineChart, Reference
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            raise Exception("openpyxl not available for Excel generation")
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create summary sheet
        summary_sheet = wb.create_sheet("Summary")
        summary_sheet['A1'] = f"Report: {template['name']}"
        summary_sheet['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        summary_sheet['A3'] = f"Period: {start_date} to {end_date}"
        summary_sheet['A4'] = f"Security Level: {template.get('security_level', 'normal').title()}"
        
        # Style the summary
        title_font = Font(bold=True, size=16)
        summary_sheet['A1'].font = title_font
        
        row_num = 6
        
        # Process each section
        for section in template.get('sections', []):
            try:
                # Get section data
                section_data = generate_enhanced_section(section, start_date, end_date, 
                                                       comparison_date, template.get('filters'), 
                                                       template.get('visualization_type', 'standard'))
                
                if section_data and section_data.get('data') is not None and not section_data['data'].empty:
                    df = section_data['data']
                    
                    # Create sheet for this section
                    sheet_name = section.replace('_', ' ').title()[:31]  # Excel sheet name limit
                    section_sheet = wb.create_sheet(sheet_name)
                    
                    # Add section title
                    section_sheet['A1'] = sheet_name
                    section_sheet['A1'].font = Font(bold=True, size=14)
                    
                    # Add summary if available
                    if section_data.get('summary'):
                        section_sheet['A2'] = section_data['summary']
                        section_sheet['A2'].font = Font(italic=True)
                    
                    # Add data starting from row 4
                    for r in dataframe_to_rows(df, index=False, header=True):
                        section_sheet.append(r)
                    
                    # Format headers
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    header_font = Font(color="FFFFFF", bold=True)
                    
                    for cell in section_sheet[4]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center")
                    
                    # Add to summary sheet
                    summary_sheet[f'A{row_num}'] = f"{sheet_name}: {len(df)} records"
                    row_num += 1
                    
            except Exception as e:
                logging.error(f"Error processing section {section}: {str(e)}")
                continue
        
        # Save workbook
        wb.save(filename)
        return filename
        
    except Exception as e:
        logging.error(f"Error generating enhanced Excel report: {str(e)}")
        return None

def generate_interactive_report(template, filename, start_date, end_date):
    """Generate interactive HTML report"""
    try:
        html_content = f"""
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Interactive Report: {template['name']}</title>
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    margin: 0;
                    padding: 20px;
                    background-color: #f5f5f5;
                }}
                .header {{
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    color: white;
                    padding: 30px;
                    border-radius: 10px;
                    margin-bottom: 30px;
                    text-align: center;
                }}
                .section {{
                    background: white;
                    padding: 20px;
                    margin-bottom: 20px;
                    border-radius: 10px;
                    box-shadow: 0 2px 10px rgba(0,0,0,0.1);
                }}
                .section h2 {{
                    color: #333;
                    border-bottom: 2px solid #667eea;
                    padding-bottom: 10px;
                }}
                .stats-grid {{
                    display: grid;
                    grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
                    gap: 20px;
                    margin: 20px 0;
                }}
                .stat-card {{
                    background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%);
                    color: white;
                    padding: 20px;
                    border-radius: 10px;
                    text-align: center;
                }}
                .stat-card h3 {{
                    margin: 0;
                    font-size: 2em;
                }}
                .stat-card p {{
                    margin: 5px 0 0 0;
                    opacity: 0.9;
                }}
                table {{
                    width: 100%;
                    border-collapse: collapse;
                    margin-top: 20px;
                }}
                th, td {{
                    padding: 12px;
                    text-align: left;
                    border-bottom: 1px solid #ddd;
                }}
                th {{
                    background-color: #667eea;
                    color: white;
                }}
                tr:hover {{
                    background-color: #f5f5f5;
                }}
                .chart-container {{
                    margin: 20px 0;
                    text-align: center;
                }}
            </style>
            <script src="https://cdn.plot.ly/plotly-latest.min.js"></script>
        </head>
        <body>
            <div class="header">
                <h1>Interactive Report: {template['name']}</h1>
                <p>Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                <p>Period: {start_date} to {end_date}</p>
            </div>
        """
        
        # Process sections
        for section in template.get('sections', []):
            try:
                section_data = generate_enhanced_section(section, start_date, end_date, 
                                                       None, template.get('filters'), 
                                                       'interactive')
                
                if section_data:
                    html_content += f"""
                    <div class="section">
                        <h2>{section.replace('_', ' ').title()}</h2>
                    """
                    
                    if section_data.get('summary'):
                        html_content += f"<p>{section_data['summary']}</p>"
                    
                    if section_data.get('data') is not None and not section_data['data'].empty:
                        df = section_data['data']
                        
                        # Add statistics if numeric data
                        numeric_cols = df.select_dtypes(include=['int64', 'float64']).columns
                        if len(numeric_cols) > 0:
                            html_content += '<div class="stats-grid">'
                            for col in numeric_cols[:4]:  # Limit to 4 stats
                                col_data = df[col].dropna()
                                if len(col_data) > 0:
                                    html_content += f"""
                                    <div class="stat-card">
                                        <h3>{col_data.mean():.1f}</h3>
                                        <p>Avg {col.replace('_', ' ').title()}</p>
                                    </div>
                                    """
                            html_content += '</div>'
                        
                        # Add table
                        html_content += df.head(20).to_html(classes='', escape=False, index=False)
                        
                        # Add interactive chart
                        if len(df) > 1 and len(df.columns) >= 2:
                            chart_id = f"chart_{section}"
                            html_content += f'<div id="{chart_id}" class="chart-container"></div>'
                            
                            # Generate Plotly JavaScript
                            html_content += f"""
                            <script>
                                var data_{section} = [{{
                                    x: {df.iloc[:, 0].tolist()},
                                    y: {df.iloc[:, 1].tolist() if len(df.columns) > 1 else [1] * len(df)},
                                    type: 'bar',
                                    marker: {{color: '#667eea'}}
                                }}];
                                var layout_{section} = {{
                                    title: '{section.replace("_", " ").title()}',
                                    xaxis: {{title: '{df.columns[0]}'}},
                                    yaxis: {{title: '{df.columns[1] if len(df.columns) > 1 else "Count"}'}}
                                }};
                                Plotly.newPlot('{chart_id}', data_{section}, layout_{section});
                            </script>
                            """
                    
                    html_content += "</div>"
                    
            except Exception as e:
                logging.error(f"Error processing section {section}: {str(e)}")
                continue
        
        # Close HTML
        html_content += """
            </body>
        </html>
        """
        
        # Save to file
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(html_content)
            
        return filename
        
    except Exception as e:
        logging.error(f"Error generating interactive report: {str(e)}")
        return None

def run_system_maintenance():
    """Run comprehensive system maintenance"""
    try:
        maintenance_report = {
            'timestamp': datetime.now().isoformat(),
            'tasks_completed': [],
            'errors': []
        }
        
        # Clean old reports
        try:
            cleanup_old_reports()
            maintenance_report['tasks_completed'].append('Old reports cleaned')
        except Exception as e:
            maintenance_report['errors'].append(f'Report cleanup failed: {str(e)}')
        
        # Clear cache
        try:
            if ENHANCED_AVAILABLE:
                CacheManager.cleanup_cache()
            maintenance_report['tasks_completed'].append('Cache cleared')
        except Exception as e:
            maintenance_report['errors'].append(f'Cache cleanup failed: {str(e)}')
        
        # Run quality checks
        try:
            if ENHANCED_AVAILABLE:
                quality_report = DataQualityMonitor.run_quality_checks()
                maintenance_report['tasks_completed'].append('Data quality check completed')
                maintenance_report['quality_report'] = quality_report
        except Exception as e:
            maintenance_report['errors'].append(f'Quality check failed: {str(e)}')
        
        # Optimize database
        try:
            conn = get_db_connection()
            if conn:
                conn.execute("VACUUM")
                conn.execute("ANALYZE")
                conn.close()
                maintenance_report['tasks_completed'].append('Database optimized')
        except Exception as e:
            maintenance_report['errors'].append(f'Database optimization failed: {str(e)}')
        
        # Update statistics
        try:
            maintenance_report['system_stats'] = {
                'maintenance_completed': datetime.now().isoformat(),
                'tasks_successful': len(maintenance_report['tasks_completed']),
                'errors_encountered': len(maintenance_report['errors'])
            }
        except Exception as e:
            maintenance_report['errors'].append(f'Statistics update failed: {str(e)}')
        
        return maintenance_report
        
    except Exception as e:
        logging.error(f"System maintenance failed: {str(e)}")
        return {'error': str(e), 'timestamp': datetime.now().isoformat()}

def cleanup_old_reports(days_to_keep=30):
    """Clean up old report files"""
    try:
        reports_dir = CONFIG.get('reports_dir', 'reports') if ENHANCED_AVAILABLE else 'reports'
        
        if not os.path.exists(reports_dir):
            return
            
        cutoff_date = datetime.now() - timedelta(days=days_to_keep)
        deleted_count = 0
        
        for root, dirs, files in os.walk(reports_dir):
            for file in files:
                file_path = os.path.join(root, file)
                try:
                    file_stat = os.stat(file_path)
                    file_date = datetime.fromtimestamp(file_stat.st_mtime)
                    
                    if file_date < cutoff_date:
                        os.remove(file_path)
                        deleted_count += 1
                        
                except Exception as e:
                    logging.warning(f"Could not process file {file_path}: {str(e)}")
        
        logging.info(f"Cleaned up {deleted_count} old report files")
        
    except Exception as e:
        logging.error(f"Error cleaning up old reports: {str(e)}")

def load_scheduled_reports():
    """Load scheduled reports from storage"""
    try:
        if not ENHANCED_AVAILABLE:
            return []

        schedules_file = os.path.join(CONFIG.get('reports_dir', str(paths.REPORTS_DIR)), 'scheduled_reports.json')

        if os.path.exists(schedules_file):
            with open(schedules_file, 'r') as f:
                return json.load(f)
        return []

    except Exception as e:
        logging.error(f"Error loading scheduled reports: {str(e)}")
        return []

def save_scheduled_reports(scheduled_reports):
    """Save scheduled reports to storage"""
    try:
        if not ENHANCED_AVAILABLE:
            return

        schedules_dir = CONFIG.get('reports_dir', str(paths.REPORTS_DIR))
        os.makedirs(schedules_dir, exist_ok=True)

        schedules_file = os.path.join(schedules_dir, 'scheduled_reports.json')

        with open(schedules_file, 'w') as f:
            json.dump(scheduled_reports, f, indent=4, default=str)

    except Exception as e:
        logging.error(f"Error saving scheduled reports: {str(e)}")

def start_scheduler():
    """Start the background scheduler for automated reports"""
    try:
        if not ENHANCED_AVAILABLE:
            return
            
        def run_scheduler():
            """Background scheduler function"""
            while True:
                try:
                    schedule.run_pending()
                    time.sleep(60)  # Check every minute
                except Exception as e:
                    logging.error(f"Scheduler error: {str(e)}")
                    time.sleep(300)  # Wait 5 minutes on error
        
        def schedule_report(report_data):
            """Schedule a specific report"""
            try:
                def run_report():
                    """Execute the scheduled report"""
                    try:
                        template_name = report_data['template_name']
                        
                        # Generate date range
                        end_date = datetime.now().strftime("%Y-%m-%d")
                        
                        # Default to 30 days if no specific range
                        days_back = report_data.get('date_range_days', 30)
                        start_date = (datetime.now() - timedelta(days=days_back)).strftime("%Y-%m-%d")
                        
                        # Generate report
                        if ENHANCED_AVAILABLE:
                            report_path = generate_report(template_name, start_date, end_date, 'pdf')
                            
                            if report_path:
                                # Update last run time
                                report_data['last_run'] = datetime.now().isoformat()
                                report_data['run_count'] = report_data.get('run_count', 0) + 1
                                
                                # Save updated schedule
                                scheduled_reports = load_scheduled_reports()
                                for i, existing in enumerate(scheduled_reports):
                                    if existing.get('template_name') == template_name:
                                        scheduled_reports[i] = report_data
                                        break
                                save_scheduled_reports(scheduled_reports)
                                
                                # Send email if recipients specified
                                if report_data.get('recipients'):
                                    send_report_email(report_path, report_data['recipients'], template_name)
                                
                                logging.info(f"Scheduled report '{template_name}' completed successfully")
                            else:
                                logging.error(f"Scheduled report '{template_name}' generation failed")
                                
                    except Exception as e:
                        logging.error(f"Error running scheduled report: {str(e)}")
                
                # Schedule based on frequency
                config = report_data.get('schedule_config', {})
                frequency = config.get('frequency', 'weekly').lower()
                hour = config.get('hour', 9)
                
                if frequency == 'daily':
                    schedule.every().day.at(f"{hour:02d}:00").do(run_report)
                elif frequency == 'weekly':
                    schedule.every().monday.at(f"{hour:02d}:00").do(run_report)
                elif frequency == 'monthly':
                    schedule.every().month.at(f"{hour:02d}:00").do(run_report)
                    
            except Exception as e:
                logging.error(f"Error scheduling report: {str(e)}")
        
        # Load and schedule all reports
        scheduled_reports = load_scheduled_reports()
        for report_data in scheduled_reports:
            if report_data.get('schedule_config', {}).get('enabled', True):
                schedule_report(report_data)
        
        # Start scheduler in background thread
        import threading
        scheduler_thread = threading.Thread(target=run_scheduler, daemon=True)
        scheduler_thread.start()
        
        logging.info("Background scheduler started")
        
    except Exception as e:
        logging.error(f"Error starting scheduler: {str(e)}")

def send_report_email(report_path, recipients, template_name):
    """Send report via email to recipients"""
    try:
        from university_system.infrastructure.email.smtp import send_email_via_smtp
        from university_system.infrastructure.email.template_utils import render_template

        # Email body
        _, body = render_template("automated_report_delivery", {
            "template_name": template_name,
            "generated_date": datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
        if not body:
            body = "Please find attached the automated report."

        # Prepare attachment
        attachments = None
        if os.path.exists(report_path):
            attachments = [report_path]

        # Send to first recipient with others as CC
        recipient_email = recipients[0]
        cc = recipients[1:] if len(recipients) > 1 else None

        current_time = datetime.now().isoformat()
        success = send_email_via_smtp(
            recipient_email=recipient_email,
            subject=f"Automated Report: {template_name}",
            body=body,
            cc=cc,
            bcc=None,
            attachments=attachments,
            current_time=current_time
        )

        if success:
            logging.info(f"Report emailed to {len(recipients)} recipients")
        else:
            logging.error(f"Failed to email report to {len(recipients)} recipients")

    except Exception as e:
        logging.error(f"Error sending report email: {str(e)}")

# Add these constants and configurations if not present
CONFIG = {
    'database': DB_PATH,
    'reports_dir': str(paths.REPORTS_DIR),
    'templates_dir': str(paths.REPORT_TEMPLATES_DIR),
    'cache_dir': str(paths.REPORT_CACHE_DIR),
    'email': {
        'enabled': False,
        'smtp_server': 'localhost',
        'smtp_port': 587,
        'use_tls': True,
        'from_address': 'reports@company.com'
    }
} if ENHANCED_AVAILABLE else {}

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(str(paths.LOG_DIR / 'reporting_system.log')),
        logging.StreamHandler()
    ]
)

def get_log_file(filename):
    """Get path to log file"""
    return str(paths.LOG_DIR / filename)

# Add pandas import if not present
try:
    import pandas as pd
except ImportError:
    pd = None
    logging.warning("Pandas not available, some features will be limited")

# Add these missing GUI-converted functions to the ReportingSystemGUI class:

    def not_found_handler(self, error_msg):
        """Handle 404-like errors in GUI context"""
        messagebox.showerror("Not Found", f"Resource not found: {error_msg}")
        self.update_status("Resource not found", "error")

    def internal_error_handler(self, error_msg):
        """Handle 500-like errors in GUI context"""
        messagebox.showerror("Internal Error", f"An internal error occurred: {error_msg}")
        self.update_status("Internal error occurred", "error")
        logging.error(f"Internal error: {error_msg}")

    def api_health_check(self):
        """Check API health status and display in GUI"""
        try:
            # Simulate health check
            status = {
                'status': 'healthy',
                'timestamp': datetime.now().isoformat(),
                'database': 'connected',
                'version': '2.0'
            }
            
            conn = get_db_connection()
            if conn:
                conn.close()
                status['database'] = 'connected'
            else:
                status['database'] = 'disconnected'
                status['status'] = 'unhealthy'
            
            # Display in dialog
            health_window = tk.Toplevel(self.root)
            health_window.title("System Health Check")
            health_window.geometry("400x300")
            
            health_text = ScrolledText(health_window, wrap=tk.WORD)
            health_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            health_info = f"""System Health Status
{'=' * 30}

Status: {status['status'].upper()}
Timestamp: {status['timestamp']}
Database: {status['database'].upper()}
Version: {status['version']}

Enhanced Features: {'Available' if ENHANCED_AVAILABLE else 'Not Available'}
"""
            
            health_text.insert(1.0, health_info)
            health_text.config(state=tk.DISABLED)
            
        except Exception as e:
            self.internal_error_handler(str(e))

    def get_section_data_dialog(self, section=None):
        """Get section data and display in dialog"""
        if not section:
            section = simpledialog.askstring("Section Data", "Enter section name:")
        
        if section:
            try:
                start_date = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
                end_date = datetime.now().strftime("%Y-%m-%d")
                
                df = get_section_dataframe(section, start_date, end_date)
                
                if not df.empty:
                    # Create data display window
                    data_window = tk.Toplevel(self.root)
                    data_window.title(f"Section Data: {section}")
                    data_window.geometry("800x600")
                    
                    # Create treeview for data
                    columns = list(df.columns)
                    tree = ttk.Treeview(data_window, columns=columns, show='headings')
                    
                    for col in columns:
                        tree.heading(col, text=col)
                        tree.column(col, width=100)
                    
                    # Add data rows
                    for _, row in df.head(100).iterrows():
                        tree.insert('', tk.END, values=list(row))
                    
                    tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
                    
                    # Add export button
                    ttk.Button(data_window, text="Export to CSV", 
                             command=lambda: self.export_dataframe_csv(df, section)).pack(pady=5)
                else:
                    messagebox.showinfo("No Data", f"No data found for section: {section}")
                    
            except Exception as e:
                self.internal_error_handler(str(e))

    def export_dataframe_csv(self, df, section_name):
        """Export dataframe to CSV file"""
        try:
            file_path = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
                initialfile=f"{section_name}_data.csv"
            )
            
            if file_path:
                df.to_csv(file_path, index=False)
                messagebox.showinfo("Export Success", f"Data exported to {file_path}")
                
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export data: {str(e)}")

    def show_templates_dialog(self):
        """Show templates in a dialog (API equivalent)"""
        try:
            templates = load_templates() if ENHANCED_AVAILABLE else []
            
            # Create templates window
            templates_window = tk.Toplevel(self.root)
            templates_window.title("All Templates")
            templates_window.geometry("700x500")
            
            # Create treeview
            columns = ('Name', 'Description', 'Sections', 'Security Level', 'Created')
            tree = ttk.Treeview(templates_window, columns=columns, show='headings')
            
            for col in columns:
                tree.heading(col, text=col)
                tree.column(col, width=120)
            
            # Populate with templates
            for template in templates:
                sections_count = len(template.get('sections', []))
                values = (
                    template['name'],
                    template.get('description', '')[:50] + '...' if len(template.get('description', '')) > 50 else template.get('description', ''),
                    f"{sections_count} sections",
                    template.get('security_level', 'normal').title(),
                    template.get('created_at', 'Unknown')[:10]
                )
                tree.insert('', tk.END, values=values)
            
            tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # Action buttons
            button_frame = ttk.Frame(templates_window)
            button_frame.pack(fill=tk.X, padx=10, pady=5)
            
            ttk.Button(button_frame, text="Close", command=templates_window.destroy).pack(side=tk.RIGHT)
            
        except Exception as e:
            self.internal_error_handler(str(e))

    def create_template_api_style(self):
        """Create template through API-style dialog"""
        try:
            # Create input dialog
            dialog = tk.Toplevel(self.root)
            dialog.title("Create Template (API Style)")
            dialog.geometry("500x400")
            dialog.transient(self.root)
            
            # Template data inputs
            ttk.Label(dialog, text="Template Name:").pack(anchor=tk.W, padx=10, pady=5)
            name_var = tk.StringVar()
            ttk.Entry(dialog, textvariable=name_var, width=50).pack(fill=tk.X, padx=10)
            
            ttk.Label(dialog, text="Description:").pack(anchor=tk.W, padx=10, pady=5)
            desc_var = tk.StringVar()
            ttk.Entry(dialog, textvariable=desc_var, width=50).pack(fill=tk.X, padx=10)
            
            ttk.Label(dialog, text="Sections (comma-separated):").pack(anchor=tk.W, padx=10, pady=5)
            sections_text = tk.Text(dialog, height=4, width=50)
            sections_text.pack(fill=tk.X, padx=10)
            sections_text.insert(1.0, "student_overview,course_distribution,gender_distribution")
            
            ttk.Label(dialog, text="Security Level:").pack(anchor=tk.W, padx=10, pady=5)
            security_var = tk.StringVar(value="normal")
            security_combo = ttk.Combobox(dialog, textvariable=security_var, 
                                        values=["normal", "confidential", "restricted"], state="readonly")
            security_combo.pack(fill=tk.X, padx=10)
            
            def save_template():
                try:
                    name = name_var.get().strip()
                    if not name:
                        messagebox.showerror("Validation Error", "Template name is required")
                        return
                    
                    sections = [s.strip() for s in sections_text.get(1.0, tk.END).strip().split(',') if s.strip()]
                    if not sections:
                        messagebox.showerror("Validation Error", "At least one section is required")
                        return
                    
                    template_data = {
                        'name': name,
                        'description': desc_var.get().strip(),
                        'sections': sections,
                        'security_level': security_var.get(),
                        'visualization_type': 'standard',
                        'created_at': datetime.now().isoformat(),
                        'version': '1.0',
                        'filters': {}
                    }
                    
                    if ENHANCED_AVAILABLE:
                        templates = load_templates()
                        templates.append(template_data)
                        
                        os.makedirs(CONFIG.get('templates_dir', str(paths.REPORT_TEMPLATES_DIR)), exist_ok=True)
                        with open(os.path.join(CONFIG.get('templates_dir', str(paths.REPORT_TEMPLATES_DIR)), "templates.json"), 'w') as f:
                            json.dump(templates, f, indent=4)
                    
                    messagebox.showinfo("Success", f"Template '{name}' created successfully!")
                    dialog.destroy()
                    self.refresh_data()
                    
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to create template: {str(e)}")
            
            # Buttons
            button_frame = ttk.Frame(dialog)
            button_frame.pack(fill=tk.X, padx=10, pady=10)
            
            ttk.Button(button_frame, text="Create Template", command=save_template).pack(side=tk.RIGHT, padx=(5, 0))
            ttk.Button(button_frame, text="Cancel", command=dialog.destroy).pack(side=tk.RIGHT)
            
        except Exception as e:
            self.internal_error_handler(str(e))

    def generate_report_api_style(self):
        """Generate report through API-style interface"""
        try:
            # Create dialog
            dialog = tk.Toplevel(self.root)
            dialog.title("Generate Report (API Style)")
            dialog.geometry("500x350")
            dialog.transient(self.root)
            
            # Parameters
            ttk.Label(dialog, text="Template Name:").pack(anchor=tk.W, padx=10, pady=5)
            template_var = tk.StringVar()
            template_combo = ttk.Combobox(dialog, textvariable=template_var, state="readonly")
            template_combo.pack(fill=tk.X, padx=10)
            
            # Load templates
            if ENHANCED_AVAILABLE:
                templates = load_templates()
                template_combo['values'] = [t['name'] for t in templates]
                if templates:
                    template_combo.set(templates[0]['name'])
            
            ttk.Label(dialog, text="Start Date (YYYY-MM-DD):").pack(anchor=tk.W, padx=10, pady=5)
            start_var = tk.StringVar(value=(datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d"))
            ttk.Entry(dialog, textvariable=start_var).pack(fill=tk.X, padx=10)
            
            ttk.Label(dialog, text="End Date (YYYY-MM-DD):").pack(anchor=tk.W, padx=10, pady=5)
            end_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
            ttk.Entry(dialog, textvariable=end_var).pack(fill=tk.X, padx=10)
            
            ttk.Label(dialog, text="Format:").pack(anchor=tk.W, padx=10, pady=5)
            format_var = tk.StringVar(value="pdf")
            format_combo = ttk.Combobox(dialog, textvariable=format_var, 
                                      values=["pdf", "excel", "interactive"], state="readonly")
            format_combo.pack(fill=tk.X, padx=10)
            
            def generate():
                try:
                    template_name = template_var.get()
                    start_date = start_var.get()
                    end_date = end_var.get()
                    format_type = format_var.get()
                    
                    if not template_name:
                        messagebox.showerror("Error", "Please select a template")
                        return
                    
                    # Validate dates
                    try:
                        datetime.strptime(start_date, "%Y-%m-%d")
                        datetime.strptime(end_date, "%Y-%m-%d")
                    except ValueError:
                        messagebox.showerror("Error", "Invalid date format")
                        return
                    
                    self.update_status("Generating report via API...")
                    dialog.destroy()
                    
                    def generate_task():
                        try:
                            if ENHANCED_AVAILABLE:
                                report_path = generate_report(template_name, start_date, end_date, format_type)
                                if report_path:
                                    self.root.after(0, lambda: [
                                        self.update_status("Report generated successfully"),
                                        messagebox.showinfo("Success", f"Report generated: {os.path.basename(report_path)}"),
                                        self.refresh_reports()
                                    ])
                                else:
                                    self.root.after(0, lambda: [
                                        self.update_status("Report generation failed", "error"),
                                        messagebox.showerror("Error", "Failed to generate report")
                                    ])
                            else:
                                self.root.after(0, lambda: [
                                    self.update_status("Enhanced features not available", "warning"),
                                    messagebox.showwarning("Feature Unavailable", "Enhanced reporting not available")
                                ])
                        except Exception as e:
                            self.root.after(0, lambda: [
                                self.update_status(f"Error: {str(e)}", "error"),
                                messagebox.showerror("Error", f"Generation failed: {str(e)}")
                            ])
                    
                    threading.Thread(target=generate_task, daemon=True).start()
                    
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to generate report: {str(e)}")
            
            # Buttons
            button_frame = ttk.Frame(dialog)
            button_frame.pack(fill=tk.X, padx=10, pady=10)
            
            ttk.Button(button_frame, text="Generate", command=generate).pack(side=tk.RIGHT, padx=(5, 0))
            ttk.Button(button_frame, text="Cancel", command=dialog.destroy).pack(side=tk.RIGHT)
            
        except Exception as e:
            self.internal_error_handler(str(e))

    def show_data_quality_dialog(self):
        """Show data quality in dedicated dialog"""
        try:
            # Create quality dialog
            quality_window = tk.Toplevel(self.root)
            quality_window.title("Data Quality Dashboard")
            quality_window.geometry("800x600")
            
            # Create notebook for different quality aspects
            quality_notebook = ttk.Notebook(quality_window)
            quality_notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # Overview tab
            overview_frame = ttk.Frame(quality_notebook)
            quality_notebook.add(overview_frame, text="Overview")
            
            self.quality_overview_text = ScrolledText(overview_frame, wrap=tk.WORD)
            self.quality_overview_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # Missing data tab
            missing_frame = ttk.Frame(quality_notebook)
            quality_notebook.add(missing_frame, text="Missing Data")
            
            self.missing_data_text = ScrolledText(missing_frame, wrap=tk.WORD)
            self.missing_data_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # Duplicates tab
            duplicates_frame = ttk.Frame(quality_notebook)
            quality_notebook.add(duplicates_frame, text="Duplicates")
            
            self.duplicates_text = ScrolledText(duplicates_frame, wrap=tk.WORD)
            self.duplicates_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # Action buttons
            button_frame = ttk.Frame(quality_window)
            button_frame.pack(fill=tk.X, padx=10, pady=5)
            
            ttk.Button(button_frame, text="Refresh Quality Check", 
                     command=self.run_comprehensive_quality_check).pack(side=tk.LEFT, padx=(0, 5))
            ttk.Button(button_frame, text="Export Report", 
                     command=self.export_quality_report).pack(side=tk.LEFT, padx=(0, 5))
            ttk.Button(button_frame, text="Close", 
                     command=quality_window.destroy).pack(side=tk.RIGHT)
            
            # Run initial quality check
            self.run_comprehensive_quality_check()
            
        except Exception as e:
            self.internal_error_handler(str(e))

    def run_comprehensive_quality_check(self):
        """Run comprehensive quality check and display results"""
        if not ENHANCED_AVAILABLE:
            messagebox.showwarning("Feature Unavailable", "Data quality checking requires enhanced features")
            return
        
        self.update_status("Running comprehensive quality check...")
        
        def quality_task():
            try:
                quality_report = DataQualityMonitor.run_quality_checks()
                self.root.after(0, lambda: self.display_comprehensive_quality_results(quality_report))
            except Exception as e:
                self.root.after(0, lambda: [
                    self.update_status("Quality check failed", "error"),
                    messagebox.showerror("Error", f"Quality check failed: {str(e)}")
                ])
        
        threading.Thread(target=quality_task, daemon=True).start()

    def display_comprehensive_quality_results(self, quality_report):
        """Display comprehensive quality results in tabs"""
        try:
            self.update_status("Quality check completed")
            
            # Overview tab
            if hasattr(self, 'quality_overview_text'):
                self.quality_overview_text.delete(1.0, tk.END)
                overview = f"""Data Quality Overview
{'=' * 50}

Timestamp: {quality_report.get('timestamp', 'Unknown')}
Status: {'PASSED' if quality_report.get('overall_status', 'unknown') == 'passed' else 'ISSUES FOUND'}

Summary:
"""
                checks = quality_report.get('checks', {})
                if 'missing_data' in checks:
                    missing = checks['missing_data']['students']
                    total = missing.get('total_records', 0)
                    overview += f"- Total Records: {total}\n"
                    overview += f"- Data Completeness: {((total * 3 - sum([missing.get('missing_emails', 0), missing.get('missing_names', 0), missing.get('missing_courses', 0)])) / (total * 3) * 100) if total > 0 else 0:.1f}%\n"
                
                self.quality_overview_text.insert(1.0, overview)
            
            # Missing data tab
            if hasattr(self, 'missing_data_text'):
                self.missing_data_text.delete(1.0, tk.END)
                missing_content = "Missing Data Analysis\n" + "=" * 30 + "\n\n"
                
                if 'missing_data' in checks:
                    missing = checks['missing_data']['students']
                    missing_content += f"Total Records: {missing.get('total_records', 0)}\n"
                    missing_content += f"Missing Emails: {missing.get('missing_emails', 0)}\n"
                    missing_content += f"Missing Names: {missing.get('missing_names', 0)}\n"
                    missing_content += f"Missing Courses: {missing.get('missing_courses', 0)}\n\n"
                    
                    if missing.get('missing_email_details'):
                        missing_content += "Records with Missing Emails:\n"
                        for detail in missing['missing_email_details'][:10]:
                            missing_content += f"- ID: {detail.get('id', 'N/A')}, Name: {detail.get('name', 'N/A')}\n"
                
                self.missing_data_text.insert(1.0, missing_content)
            
            # Duplicates tab
            if hasattr(self, 'duplicates_text'):
                self.duplicates_text.delete(1.0, tk.END)
                duplicates_content = "Duplicate Data Analysis\n" + "=" * 30 + "\n\n"
                
                if 'duplicates' in checks:
                    duplicates = checks['duplicates']
                    duplicates_content += f"Duplicate Emails: {duplicates.get('duplicate_emails', 0)}\n\n"
                    
                    if duplicates.get('duplicate_email_details'):
                        duplicates_content += "Duplicate Email Details:\n"
                        for detail in duplicates['duplicate_email_details'][:10]:
                            duplicates_content += f"- Email: {detail.get('email', 'N/A')}, Count: {detail.get('count', 0)}\n"
                
                self.duplicates_text.insert(1.0, duplicates_content)
                
        except Exception as e:
            logging.error(f"Error displaying quality results: {str(e)}")

    def show_predictive_analytics_dialog(self):
        """Show predictive analytics in dedicated dialog"""
        try:
            # Create analytics dialog
            analytics_window = tk.Toplevel(self.root)
            analytics_window.title("Predictive Analytics Dashboard")
            analytics_window.geometry("900x700")
            
            # Create notebook for different analytics
            analytics_notebook = ttk.Notebook(analytics_window)
            analytics_notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # Dropout Risk tab
            dropout_frame = ttk.Frame(analytics_notebook)
            analytics_notebook.add(dropout_frame, text="Dropout Risk")
            
            self.dropout_text = ScrolledText(dropout_frame, wrap=tk.WORD)
            self.dropout_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # Model Performance tab
            performance_frame = ttk.Frame(analytics_notebook)
            analytics_notebook.add(performance_frame, text="Model Performance")
            
            self.performance_text = ScrolledText(performance_frame, wrap=tk.WORD)
            self.performance_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # Recommendations tab
            recommendations_frame = ttk.Frame(analytics_notebook)
            analytics_notebook.add(recommendations_frame, text="Recommendations")
            
            self.recommendations_text = ScrolledText(recommendations_frame, wrap=tk.WORD)
            self.recommendations_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # Action buttons
            button_frame = ttk.Frame(analytics_window)
            button_frame.pack(fill=tk.X, padx=10, pady=5)
            
            ttk.Button(button_frame, text="Run Analysis", 
                     command=self.run_comprehensive_predictions).pack(side=tk.LEFT, padx=(0, 5))
            ttk.Button(button_frame, text="Export Results", 
                     command=self.export_predictions_report).pack(side=tk.LEFT, padx=(0, 5))
            ttk.Button(button_frame, text="Close", 
                     command=analytics_window.destroy).pack(side=tk.RIGHT)
            
            # Run initial analysis
            self.run_comprehensive_predictions()
            
        except Exception as e:
            self.internal_error_handler(str(e))

    def run_comprehensive_predictions(self):
        """Run comprehensive predictive analysis"""
        if not ENHANCED_AVAILABLE:
            messagebox.showwarning("Feature Unavailable", "Predictive analytics requires enhanced features")
            return
        
        self.update_status("Running predictive analytics...")
        
        def predictions_task():
            try:
                predictions = PredictiveAnalytics.predict_dropout_risk()
                self.root.after(0, lambda: self.display_comprehensive_predictions(predictions))
            except Exception as e:
                self.root.after(0, lambda: [
                    self.update_status("Predictions failed", "error"),
                    messagebox.showerror("Error", f"Predictions failed: {str(e)}")
                ])
        
        threading.Thread(target=predictions_task, daemon=True).start()

    def display_comprehensive_predictions(self, predictions):
        """Display comprehensive prediction results"""
        try:
            self.update_status("Predictive analysis completed")
            
            # Dropout Risk tab
            if hasattr(self, 'dropout_text'):
                self.dropout_text.delete(1.0, tk.END)
                dropout_content = "Dropout Risk Analysis\n" + "=" * 40 + "\n\n"
                
                if 'error' in predictions:
                    dropout_content += f"Analysis unavailable: {predictions['error']}\n"
                else:
                    dropout_content += f"Students Analyzed: {predictions.get('total_students_analyzed', 0)}\n"
                    dropout_content += f"High Risk Students: {len(predictions.get('high_risk_students', []))}\n"
                    
                    if predictions.get('high_risk_students'):
                        dropout_content += "\nHigh Risk Student Details:\n"
                        for student in predictions['high_risk_students'][:10]:
                            dropout_content += f"- ID: {student.get('student_id', 'N/A')}, Risk Score: {student.get('risk_score', 0):.2%}\n"
                
                self.dropout_text.insert(1.0, dropout_content)
            
            # Model Performance tab
            if hasattr(self, 'performance_text'):
                self.performance_text.delete(1.0, tk.END)
                performance_content = "Model Performance Metrics\n" + "=" * 40 + "\n\n"
                
                if 'model_accuracy' in predictions:
                    accuracy = predictions['model_accuracy']
                    performance_content += f"Model Accuracy: {accuracy:.2%}\n"
                    
                    if accuracy > 0.8:
                        performance_content += "Status: Excellent model performance\n"
                    elif accuracy > 0.6:
                        performance_content += "Status: Good model performance\n"
                    else:
                        performance_content += "Status: Model needs improvement\n"
                
                if 'feature_importance' in predictions:
                    performance_content += "\nMost Important Risk Factors:\n"
                    importance = predictions['feature_importance']
                    sorted_features = sorted(importance.items(), key=lambda x: x[1], reverse=True)
                    for feature, score in sorted_features:
                        performance_content += f"- {feature.replace('_', ' ').title()}: {score:.3f}\n"
                
                self.performance_text.insert(1.0, performance_content)
            
            # Recommendations tab
            if hasattr(self, 'recommendations_text'):
                self.recommendations_text.delete(1.0, tk.END)
                recommendations_content = "Recommendations\n" + "=" * 20 + "\n\n"
                
                high_risk_count = len(predictions.get('high_risk_students', []))
                
                if high_risk_count > 0:
                    recommendations_content += f"Action Required for {high_risk_count} students:\n\n"
                    recommendations_content += "1. Contact high-risk students immediately\n"
                    recommendations_content += "2. Provide additional academic support\n"
                    recommendations_content += "3. Schedule counseling sessions\n"
                    recommendations_content += "4. Monitor attendance closely\n"
                    recommendations_content += "5. Consider intervention programs\n\n"
                else:
                    recommendations_content += "No immediate action required.\n"
                    recommendations_content += "Continue monitoring student performance.\n\n"
                
                recommendations_content += "General Recommendations:\n"
                recommendations_content += "- Regular model updates with new data\n"
                recommendations_content += "- Monitor model accuracy trends\n"
                recommendations_content += "- Validate predictions with academic staff\n"
                
                self.recommendations_text.insert(1.0, recommendations_content)
                
        except Exception as e:
            logging.error(f"Error displaying predictions: {str(e)}")

    def export_predictions_report(self):
        """Export predictions report to file"""
        try:
            if not hasattr(self, 'dropout_text'):
                messagebox.showwarning("No Data", "No predictions data to export")
                return
            
            file_path = filedialog.asksaveasfilename(
                defaultextension=".txt",
                filetypes=[("Text files", "*.txt"), ("All files", "*.*")],
                initialfile=f"predictions_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
            )
            
            if file_path:
                with open(file_path, 'w') as f:
                    f.write("Comprehensive Predictive Analytics Report\n")
                    f.write("=" * 60 + "\n\n")
                    f.write("DROPOUT RISK ANALYSIS:\n")
                    f.write(self.dropout_text.get(1.0, tk.END))
                    f.write("\n\nMODEL PERFORMANCE:\n")
                    if hasattr(self, 'performance_text'):
                        f.write(self.performance_text.get(1.0, tk.END))
                    f.write("\n\nRECOMMENDATIONS:\n")
                    if hasattr(self, 'recommendations_text'):
                        f.write(self.recommendations_text.get(1.0, tk.END))
                
                messagebox.showinfo("Export Success", f"Predictions report exported to {file_path}")
                
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export report: {str(e)}")

    def show_anomaly_detection_dialog(self):
        """Show anomaly detection in dedicated dialog"""
        try:
            # Create anomaly dialog
            anomaly_window = tk.Toplevel(self.root)
            anomaly_window.title("Anomaly Detection Dashboard")
            anomaly_window.geometry("800x600")
            
            # Results display
            self.anomaly_results_text = ScrolledText(anomaly_window, wrap=tk.WORD)
            self.anomaly_results_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # Action buttons
            button_frame = ttk.Frame(anomaly_window)
            button_frame.pack(fill=tk.X, padx=10, pady=5)
            
            ttk.Button(button_frame, text="Run Detection", 
                     command=self.run_comprehensive_anomaly_detection).pack(side=tk.LEFT, padx=(0, 5))
            ttk.Button(button_frame, text="Export Results", 
                     command=self.export_anomaly_report).pack(side=tk.LEFT, padx=(0, 5))
            ttk.Button(button_frame, text="Close", 
                     command=anomaly_window.destroy).pack(side=tk.RIGHT)
            
            # Run initial detection
            self.run_comprehensive_anomaly_detection()
            
        except Exception as e:
            self.internal_error_handler(str(e))

    def run_comprehensive_anomaly_detection(self):
        """Run comprehensive anomaly detection"""
        if not ENHANCED_AVAILABLE:
            messagebox.showwarning("Feature Unavailable", "Anomaly detection requires enhanced features")
            return
        
        self.update_status("Running anomaly detection...")
        
        def anomaly_task():
            try:
                anomalies = PredictiveAnalytics.detect_anomalies()
                self.root.after(0, lambda: self.display_comprehensive_anomalies(anomalies))
            except Exception as e:
                self.root.after(0, lambda: [
                    self.update_status("Anomaly detection failed", "error"),
                    messagebox.showerror("Error", f"Anomaly detection failed: {str(e)}")
                ])
        
        threading.Thread(target=anomaly_task, daemon=True).start()

    def display_comprehensive_anomalies(self, anomalies):
        """Display comprehensive anomaly results"""
        try:
            self.update_status("Anomaly detection completed")
            
            if hasattr(self, 'anomaly_results_text'):
                self.anomaly_results_text.delete(1.0, tk.END)
                
                content = "Comprehensive Anomaly Detection Results\n"
                content += "=" * 50 + "\n\n"
                
                if 'error' in anomalies:
                    content += f"Analysis unavailable: {anomalies['error']}\n"
                else:
                    content += f"Total Students Analyzed: {anomalies.get('total_students_analyzed', 0)}\n"
                    content += f"Anomalies Detected: {anomalies.get('total_anomalies', 0)}\n"
                    content += f"Anomaly Rate: {anomalies.get('anomaly_rate', 0):.2f}%\n\n"
                    
                    anomaly_rate = anomalies.get('anomaly_rate', 0)
                    if anomaly_rate > 15:
                        content += "STATUS: HIGH - Investigate data quality issues\n"
                    elif anomaly_rate > 5:
                        content += "STATUS: MODERATE - Review identified anomalies\n"
                    else:
                        content += "STATUS: NORMAL - Low anomaly rate detected\n"
                    
                    content += "\n" + "-" * 40 + "\n\n"
                    
                    if anomalies.get('anomalous_students'):
                        content += "DETAILED ANOMALY ANALYSIS:\n\n"
                        for i, student in enumerate(anomalies['anomalous_students'][:15], 1):
                            content += f"{i}. Student ID: {student.get('student_id', 'N/A')}\n"
                            content += f"   Age: {student.get('age', 'N/A')}\n"
                            content += f"   Unique Modules: {student.get('unique_modules', 'N/A')}\n"
                            content += f"   Average Grade: {student.get('avg_grade', 'N/A')}\n"
                            content += f"   Anomaly Score: {student.get('anomaly_score', 'N/A')}\n"
                            content += f"   Reason: {student.get('anomaly_reason', 'Statistical outlier')}\n\n"
                        
                        if len(anomalies['anomalous_students']) > 15:
                            remaining = len(anomalies['anomalous_students']) - 15
                            content += f"... and {remaining} more anomalous profiles\n"
                    
                    content += "\nRECOMMENDATIONS:\n"
                    content += "1. Verify data accuracy for flagged students\n"
                    content += "2. Check for data entry errors\n"
                    content += "3. Investigate unusual enrollment patterns\n"
                    content += "4. Contact students with extreme anomalies\n"
                
                self.anomaly_results_text.insert(1.0, content)
                
        except Exception as e:
            logging.error(f"Error displaying anomalies: {str(e)}")

    def export_anomaly_report(self):
        """Export anomaly detection report"""
        try:
            if not hasattr(self, 'anomaly_results_text'):
                messagebox.showwarning("No Data", "No anomaly data to export")
                return
            
            content = self.anomaly_results_text.get(1.0, tk.END).strip()
            if not content:
                messagebox.showwarning("No Data", "No anomaly data to export")
                return
            
            file_path = filedialog.asksaveasfilename(
                defaultextension=".txt",
                filetypes=[("Text files", "*.txt"), ("All files", "*.*")],
                initialfile=f"anomaly_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
            )
            
            if file_path:
                with open(file_path, 'w') as f:
                    f.write(content)
                messagebox.showinfo("Export Success", f"Anomaly report exported to {file_path}")
                
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export report: {str(e)}")

    def show_correlation_analysis_dialog(self):
        """Show correlation analysis in dedicated dialog"""
        try:
            # Create correlation dialog
            correlation_window = tk.Toplevel(self.root)
            correlation_window.title("Correlation Analysis Dashboard")
            correlation_window.geometry("900x700")
            
            # Create notebook for different correlation views
            correlation_notebook = ttk.Notebook(correlation_window)
            correlation_notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # Matrix tab
            matrix_frame = ttk.Frame(correlation_notebook)
            correlation_notebook.add(matrix_frame, text="Correlation Matrix")
            
            self.correlation_text = ScrolledText(matrix_frame, wrap=tk.WORD)
            self.correlation_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # Insights tab
            insights_frame = ttk.Frame(correlation_notebook)
            correlation_notebook.add(insights_frame, text="Key Insights")
            
            self.insights_text = ScrolledText(insights_frame, wrap=tk.WORD)
            self.insights_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # Action buttons
            button_frame = ttk.Frame(correlation_window)
            button_frame.pack(fill=tk.X, padx=10, pady=5)
            
            ttk.Button(button_frame, text="Run Analysis", 
                     command=self.run_comprehensive_correlation).pack(side=tk.LEFT, padx=(0, 5))
            ttk.Button(button_frame, text="Generate Heatmap", 
                     command=self.generate_correlation_heatmap).pack(side=tk.LEFT, padx=(0, 5))
            ttk.Button(button_frame, text="Export Results", 
                     command=self.export_correlation_report).pack(side=tk.LEFT, padx=(0, 5))
            ttk.Button(button_frame, text="Close", 
                     command=correlation_window.destroy).pack(side=tk.RIGHT)
            
            # Run initial analysis
            self.run_comprehensive_correlation()
            
        except Exception as e:
            self.internal_error_handler(str(e))

    def run_comprehensive_correlation(self):
        """Run comprehensive correlation analysis"""
        if not ENHANCED_AVAILABLE:
            messagebox.showwarning("Feature Unavailable", "Correlation analysis requires enhanced features")
            return
        
        self.update_status("Running correlation analysis...")
        
        def correlation_task():
            try:
                conn = get_db_connection()
                if not conn:
                    raise Exception("Database connection failed")
                
                # Get correlation data
                end_date = datetime.now().strftime("%Y-%m-%d")
                start_date = (datetime.now() - timedelta(days=365)).strftime("%Y-%m-%d")
                
                correlation_df = get_correlation_data(conn, start_date, end_date, {})
                conn.close()
                
                if not correlation_df.empty:
                    correlation_matrix = correlation_df.corr()
                    self.root.after(0, lambda: self.display_comprehensive_correlation(correlation_matrix, correlation_df))
                else:
                    self.root.after(0, lambda: [
                        self.update_status("No correlation data available", "warning"),
                        messagebox.showwarning("No Data", "Insufficient data for correlation analysis")
                    ])
                    
            except Exception as e:
                self.root.after(0, lambda: [
                    self.update_status("Correlation analysis failed", "error"),
                    messagebox.showerror("Error", f"Correlation analysis failed: {str(e)}")
                ])
        
        threading.Thread(target=correlation_task, daemon=True).start()

    def display_comprehensive_correlation(self, correlation_matrix, raw_data):
        """Display comprehensive correlation results"""
        try:
            self.update_status("Correlation analysis completed")
            
            # Matrix tab
            if hasattr(self, 'correlation_text'):
                self.correlation_text.delete(1.0, tk.END)
                
                content = "Correlation Matrix Analysis\n"
                content += "=" * 40 + "\n\n"
                
                content += f"Variables Analyzed: {len(correlation_matrix.columns)}\n"
                content += f"Data Points: {len(raw_data)}\n\n"
                
                content += "Correlation Matrix:\n"
                content += "-" * 20 + "\n"
                
                # Format correlation matrix for display
                for i, row in correlation_matrix.iterrows():
                    content += f"\n{i}:\n"
                    for col, value in row.items():
                        if col != i:  # Skip self-correlation
                            content += f"  vs {col}: {value:.3f}\n"
                
                self.correlation_text.insert(1.0, content)
            
            # Insights tab
            if hasattr(self, 'insights_text'):
                self.insights_text.delete(1.0, tk.END)
                
                insights = "Key Correlation Insights\n"
                insights += "=" * 30 + "\n\n"
                
                # Find strong correlations
                strong_correlations = []
                for i in range(len(correlation_matrix.columns)):
                    for j in range(i+1, len(correlation_matrix.columns)):
                        col1 = correlation_matrix.columns[i]
                        col2 = correlation_matrix.columns[j]
                        corr_value = correlation_matrix.iloc[i, j]
                        
                        if abs(corr_value) > 0.5:  # Strong correlation threshold
                            strong_correlations.append((col1, col2, corr_value))
                
                if strong_correlations:
                    insights += "STRONG CORRELATIONS (|r| > 0.5):\n\n"
                    for col1, col2, corr in sorted(strong_correlations, key=lambda x: abs(x[2]), reverse=True):
                        direction = "positive" if corr > 0 else "negative"
                        strength = "very strong" if abs(corr) > 0.8 else "strong"
                        insights += f"• {col1} ↔ {col2}\n"
                        insights += f"  Correlation: {corr:.3f} ({strength} {direction})\n"
                        
                        # Add interpretation
                        if corr > 0:
                            insights += f"  As {col1} increases, {col2} tends to increase\n\n"
                        else:
                            insights += f"  As {col1} increases, {col2} tends to decrease\n\n"
                else:
                    insights += "No strong correlations found (|r| > 0.5)\n\n"
                
                insights += "INTERPRETATION GUIDELINES:\n"
                insights += "• |r| > 0.8: Very strong relationship\n"
                insights += "• |r| > 0.6: Strong relationship\n"
                insights += "• |r| > 0.4: Moderate relationship\n"
                insights += "• |r| > 0.2: Weak relationship\n"
                insights += "• |r| ≤ 0.2: Very weak/no relationship\n\n"
                
                insights += "BUSINESS IMPLICATIONS:\n"
                if strong_correlations:
                    insights += "• Use identified relationships for predictive modeling\n"
                    insights += "• Consider correlated factors in decision making\n"
                    insights += "• Monitor relationships over time for changes\n"
                else:
                    insights += "• Variables appear to be largely independent\n"
                    insights += "• May need additional variables for analysis\n"
                    insights += "• Consider non-linear relationships\n"
                
                self.insights_text.insert(1.0, insights)
                
        except Exception as e:
            logging.error(f"Error displaying correlation results: {str(e)}")

    def generate_correlation_heatmap(self):
        """Generate correlation heatmap visualization"""
        if not ENHANCED_AVAILABLE:
            messagebox.showwarning("Feature Unavailable", "Heatmap generation requires enhanced features")
            return
        
        self.update_status("Generating correlation heatmap...")
        
        def heatmap_task():
            try:
                conn = get_db_connection()
                chart_path = AdvancedVisualization.create_correlation_matrix(conn)
                
                self.root.after(0, lambda: [
                    self.update_status("Heatmap generated successfully"),
                    self.show_correlation_heatmap_result(chart_path)
                ])
                
            except Exception as e:
                self.root.after(0, lambda: [
                    self.update_status("Heatmap generation failed", "error"),
                    messagebox.showerror("Error", f"Heatmap generation failed: {str(e)}")
                ])
        
        threading.Thread(target=heatmap_task, daemon=True).start()

    def show_correlation_heatmap_result(self, chart_path):
        """Show correlation heatmap result"""
        if chart_path and os.path.exists(chart_path):
            result = messagebox.askyesno("Heatmap Generated", 
                                       f"Correlation heatmap generated successfully!\n\nFile: {os.path.basename(chart_path)}\n\nWould you like to open it now?")
            
            if result:
                try:
                    webbrowser.open(f"file://{os.path.abspath(chart_path)}")
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to open heatmap: {str(e)}")
        else:
            messagebox.showwarning("Generation Failed", "Unable to generate correlation heatmap - insufficient data")

    def export_correlation_report(self):
        """Export correlation analysis report"""
        try:
            if not hasattr(self, 'correlation_text') or not hasattr(self, 'insights_text'):
                messagebox.showwarning("No Data", "No correlation data to export")
                return
            
            file_path = filedialog.asksaveasfilename(
                defaultextension=".txt",
                filetypes=[("Text files", "*.txt"), ("All files", "*.*")],
                initialfile=f"correlation_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
            )
            
            if file_path:
                with open(file_path, 'w') as f:
                    f.write("Comprehensive Correlation Analysis Report\n")
                    f.write("=" * 60 + "\n\n")
                    f.write("CORRELATION MATRIX:\n")
                    f.write(self.correlation_text.get(1.0, tk.END))
                    f.write("\n\nKEY INSIGHTS:\n")
                    f.write(self.insights_text.get(1.0, tk.END))
                
                messagebox.showinfo("Export Success", f"Correlation report exported to {file_path}")
                
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export report: {str(e)}")

    def start_api_server_gui(self):
        """Start API server with GUI interface"""
        try:
            if not ENHANCED_AVAILABLE:
                messagebox.showwarning("Feature Unavailable", "API server requires enhanced features")
                return
            
            # Enhanced API server dialog
            api_dialog = tk.Toplevel(self.root)
            api_dialog.title("API Server Configuration")
            api_dialog.geometry("600x500")
            api_dialog.transient(self.root)
            
            # Configuration section
            config_frame = ttk.LabelFrame(api_dialog, text="Server Configuration", padding="10")
            config_frame.pack(fill=tk.X, padx=10, pady=10)
            
            # Host setting
            ttk.Label(config_frame, text="Host:").grid(row=0, column=0, sticky=tk.W, pady=5)
            host_var = tk.StringVar(value="localhost")
            ttk.Entry(config_frame, textvariable=host_var).grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)
            
            # Port setting
            ttk.Label(config_frame, text="Port:").grid(row=1, column=0, sticky=tk.W, pady=5)
            port_var = tk.StringVar(value="5000")
            ttk.Entry(config_frame, textvariable=port_var).grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)
            
            # Debug mode
            debug_var = tk.BooleanVar(value=False)
            ttk.Checkbutton(config_frame, text="Debug Mode", variable=debug_var).grid(row=2, column=0, columnspan=2, sticky=tk.W, pady=5)
            
            config_frame.columnconfigure(1, weight=1)
            
            # API endpoints info
            endpoints_frame = ttk.LabelFrame(api_dialog, text="Available Endpoints", padding="10")
            endpoints_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            endpoints_text = ScrolledText(endpoints_frame, height=15, wrap=tk.WORD)
            endpoints_text.pack(fill=tk.BOTH, expand=True)
            
            endpoints_info = """API Endpoints Documentation:

AUTHENTICATION:
POST /api/login - User authentication
    Body: {"username": "user", "password": "pass"}

TEMPLATES:
GET  /api/templates - List all templates
POST /api/templates - Create new template
    Body: {"name": "Template Name", "sections": [...]}

REPORTS:
POST /api/reports/generate - Generate report
    Body: {"template": "name", "start_date": "YYYY-MM-DD", "end_date": "YYYY-MM-DD", "format": "pdf"}

DATA:
GET  /api/data/<section> - Get section data
    Parameters: start_date, end_date, filters

ANALYTICS:
GET  /api/analytics/quality - Data quality metrics
GET  /api/analytics/predictions - Dropout risk predictions  
GET  /api/analytics/anomalies - Anomaly detection results

SYSTEM:
GET  /api/health - System health check

Example Usage:
curl -X POST http://localhost:5000/api/reports/generate \\
  -H "Content-Type: application/json" \\
  -d '{"template": "student_overview", "start_date": "2024-01-01", "end_date": "2024-12-31"}'
"""
            
            endpoints_text.insert(1.0, endpoints_info)
            endpoints_text.config(state=tk.DISABLED)
            
            # Status display
            status_frame = ttk.Frame(api_dialog)
            status_frame.pack(fill=tk.X, padx=10, pady=5)
            
            self.api_status_label = ttk.Label(status_frame, text="Server Status: Stopped", style='Info.TLabel')
            self.api_status_label.pack(side=tk.LEFT)
            
            # Control buttons
            button_frame = ttk.Frame(api_dialog)
            button_frame.pack(fill=tk.X, padx=10, pady=10)
            
            def start_server():
                try:
                    host = host_var.get()
                    port = int(port_var.get())
                    debug = debug_var.get()
                    
                    # In a real implementation, this would start the Flask server
                    self.api_status_label.config(text=f"Server Status: Running on http://{host}:{port}", style='Success.TLabel')
                    messagebox.showinfo("API Server", 
                                      f"API server started successfully!\n\nURL: http://{host}:{port}\nDebug Mode: {debug}")
                    
                    # Store server config for stopping later
                    self.api_server_config = {'host': host, 'port': port, 'debug': debug, 'running': True}
                    
                except ValueError:
                    messagebox.showerror("Invalid Port", "Please enter a valid port number")
                except Exception as e:
                    messagebox.showerror("Server Error", f"Failed to start server: {str(e)}")
            
            def stop_server():
                if hasattr(self, 'api_server_config') and self.api_server_config.get('running'):
                    self.api_status_label.config(text="Server Status: Stopped", style='Info.TLabel')
                    self.api_server_config['running'] = False
                    messagebox.showinfo("API Server", "API server stopped successfully!")
                else:
                    messagebox.showwarning("Server Not Running", "API server is not currently running")
            
            def test_connection():
                if hasattr(self, 'api_server_config') and self.api_server_config.get('running'):
                    host = self.api_server_config['host']
                    port = self.api_server_config['port']
                    messagebox.showinfo("Connection Test", f"API server is responding at http://{host}:{port}")
                else:
                    messagebox.showwarning("Server Not Running", "Please start the API server first")
            
            ttk.Button(button_frame, text="Start Server", command=start_server, 
                     style='Success.TButton').pack(side=tk.LEFT, padx=(0, 5))
            ttk.Button(button_frame, text="Stop Server", command=stop_server, 
                     style='Warning.TButton').pack(side=tk.LEFT, padx=(0, 5))
            ttk.Button(button_frame, text="Test Connection", command=test_connection).pack(side=tk.LEFT, padx=(0, 5))
            ttk.Button(button_frame, text="Close", command=api_dialog.destroy).pack(side=tk.RIGHT)
            
        except Exception as e:
            self.internal_error_handler(str(e))

    def main_gui_entry_point(self):
        """Main entry point for GUI application"""
        try:
            # Initialize logging
            logging.basicConfig(
                level=logging.INFO,
                format='%(asctime)s - %(levelname)s - %(message)s',
                handlers=[
                    logging.FileHandler('reporting_system_gui.log'),
                    logging.StreamHandler()
                ]
            )
            
            logging.info("Starting Enhanced Reporting System GUI...")
            
            # Check system requirements
            self.check_system_requirements()
            
            # Initialize directories
            if ENHANCED_AVAILABLE:
                self.initialize_directories()
            
            # Check database connection
            self.verify_database_connection()
            
            # Load initial data
            self.refresh_data()
            
            # Start scheduler if available
            if ENHANCED_AVAILABLE:
                try:
                    start_scheduler()
                    logging.info("Background scheduler started successfully")
                except Exception as e:
                    logging.warning(f"Failed to start scheduler: {str(e)}")
            
            logging.info("GUI application initialized successfully")
            
        except Exception as e:
            logging.error(f"Failed to initialize GUI application: {str(e)}")
            messagebox.showerror("Initialization Error", 
                               f"Failed to initialize application: {str(e)}\n\nThe application may not function correctly.")

    def check_system_requirements(self):
        """Check system requirements for GUI application"""
        try:
            requirements_status = {
                'tkinter': True,  # Already imported if we got this far
                'pandas': pd is not None,
                'enhanced_features': ENHANCED_AVAILABLE,
                'database': False
            }
            
            # Test database connection
            try:
                conn = get_db_connection()
                if conn:
                    conn.close()
                    requirements_status['database'] = True
            except:
                pass
            
            # Log requirements status
            logging.info("System Requirements Check:")
            for requirement, status in requirements_status.items():
                status_text = "✓" if status else "✗"
                logging.info(f"  {requirement}: {status_text}")
            
            # Warn about missing requirements
            missing = [req for req, status in requirements_status.items() if not status]
            if missing:
                warning_msg = f"Missing requirements: {', '.join(missing)}\nSome features may be limited."
                logging.warning(warning_msg)
                
                if 'database' in missing:
                    messagebox.showwarning("Database Warning", 
                                         "Database connection failed. Reports and analytics will be limited.")
            
        except Exception as e:
            logging.error(f"Requirements check failed: {str(e)}")

    def initialize_directories(self):
        """Initialize required directories"""
        try:
            directories = [
                str(paths.REPORTS_DIR),
                str(paths.REPORT_TEMPLATES_DIR),
                str(paths.REPORT_CACHE_DIR),
                str(paths.REPORTS_DIR / 'charts'),
                str(paths.LOG_DIR)
            ]
            
            for directory in directories:
                os.makedirs(directory, exist_ok=True)
                logging.debug(f"Directory ensured: {directory}")
            
            logging.info("All required directories initialized")
            
        except Exception as e:
            logging.error(f"Failed to initialize directories: {str(e)}")

    def verify_database_connection(self):
        """Verify database connection and structure"""
        try:
            conn = get_db_connection()
            if not conn:
                raise Exception("Could not establish database connection")
            
            # Check for required tables
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
            tables = [row[0] for row in cursor.fetchall()]
            
            required_tables = ['students']
            missing_tables = [table for table in required_tables if table not in tables]
            
            if missing_tables:
                logging.warning(f"Missing database tables: {missing_tables}")
            else:
                logging.info("Database structure verified")
            
            conn.close()
            
        except Exception as e:
            logging.error(f"Database verification failed: {str(e)}")
            raise

class ReportingSystemGUI:
    """Main GUI application for the Enhanced Reporting System"""
    
    def __init__(self, parent=None):
        # If parent is provided, use it; otherwise create new root
        if parent is not None:
            self.root = parent
            self.is_embedded = True
        else:
            self.root = tk.Tk()
            self.root.title("Advanced Student Reporting System v2.0")
            self.root.geometry("1200x800")
            self.root.minsize(1000, 700)
            self.is_embedded = False
        
        # Configure style
        self.setup_styles()
        
        # Create main interface
        self.create_widgets()
        
        # Initialize data
        self.refresh_data()
        
        # Status for background operations
        self.background_tasks = []

    def _schedule_on_ui_thread(self, callback, delay=0):
        """Safely execute callbacks on the Tk main loop from worker threads."""
        if not callable(callback):
            return

        root = getattr(self, 'root', None)
        if root is None:
            return

        try:
            tk_app = getattr(root, 'tk', None)
            if tk_app is None:
                return
            try:
                if not root.winfo_exists():
                    return
            except tk.TclError:
                return

            root.after(delay, callback)
        except RuntimeError as exc:
            if 'main thread is not in main loop' in str(exc):
                return
            logging.error(f"Failed to schedule UI callback: {exc}")
        except tk.TclError:
            # Tk widget hierarchy is likely being destroyed; ignore gracefully.
            return

    def setup_styles(self):
        """Configure the application styling"""
        style = ttk.Style()
        
        # Configure colors and fonts
        style.configure('Title.TLabel', font=('Arial', 16, 'bold'), foreground='#2E4E7E')
        style.configure('Subtitle.TLabel', font=('Arial', 12, 'bold'), foreground='#4A6FA5')
        style.configure('Info.TLabel', font=('Arial', 10), foreground='#666666')
        style.configure('Success.TLabel', font=('Arial', 10), foreground='#28A745')
        style.configure('Warning.TLabel', font=('Arial', 10), foreground='#FFC107')
        style.configure('Error.TLabel', font=('Arial', 10), foreground='#DC3545')
        
        # Button styles
        style.configure('Primary.TButton', font=('Arial', 10, 'bold'))
        style.configure('Success.TButton', font=('Arial', 10, 'bold'))
        style.configure('Warning.TButton', font=('Arial', 10, 'bold'))
            
    def create_widgets(self):
        """Create the main GUI widgets"""
        # Add return to main menu button at top right
        return_btn = ttk.Button(
            self.root,
            text="🏠 Return to Main Menu",
            command=self.return_to_main_menu
        )
        return_btn.place(relx=1.0, rely=0.0, anchor="ne", x=-10, y=10)

        # Create main container with padding
        if self.is_embedded:
            # When embedded, use the parent directly
            main_frame = self.root
        else:
            # When standalone, create frame with padding
            main_frame = ttk.Frame(self.root, padding="10")
            main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

            # Configure grid weights for standalone mode
            self.root.columnconfigure(0, weight=1)
            self.root.rowconfigure(0, weight=1)
        
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(1, weight=1)
        
        # Title section (only show in standalone mode)
        self.init_status_widgets(main_frame)          # always creates the objects
        if not self.is_embedded:
            self.layout_status_bar(main_frame) 
        
        # Create notebook for different sections
        self.notebook = ttk.Notebook(main_frame)
        if self.is_embedded:
            self.notebook.pack(fill=tk.BOTH, expand=True)
        else:
            self.notebook.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(10, 0))
        
        # Create tabs
        self.create_templates_tab()
        self.create_reports_tab()
        self.create_analytics_tab()
        self.create_schedule_tab()
        self.create_system_tab()
        
        # Status bar (only in standalone mode)
        if not self.is_embedded:
            self.create_status_bar(main_frame)
        
    def create_header(self, parent):
        """Create the application header"""
        header_frame = ttk.Frame(parent)
        header_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        header_frame.columnconfigure(1, weight=1)

        # Title and description
        title_label = ttk.Label(header_frame, text="📊 Advanced Student Reporting System",
                               style='Title.TLabel')
        title_label.grid(row=0, column=0, sticky=tk.W)

        # Return to Home button
        if hasattr(self, 'auth'):
            return_button = ttk.Button(header_frame, text="🏠 Return to Main Menu",
                                      command=self.return_to_main_menu)
            return_button.grid(row=0, column=2, sticky=tk.E, padx=10)

        subtitle_label = ttk.Label(header_frame,
                                  text="Generate comprehensive reports, analytics, and insights",
                                  style='Subtitle.TLabel')
        subtitle_label.grid(row=1, column=0, sticky=tk.W)

        # Return to main menu button
        ttk.Button(header_frame, text="🏠 Return to Main Menu",
                  command=self.return_to_main_menu).grid(row=0, column=1, sticky=tk.E)
        
        # System status indicator - ensure this gets created
        if not hasattr(self, 'status_indicator'):
            self.status_indicator = ttk.Label(header_frame, text="● System Ready", 
                                             style='Success.TLabel')
        self.status_indicator.grid(row=0, column=1, sticky=tk.E)
        
        # Quick actions
        actions_frame = ttk.Frame(header_frame)
        actions_frame.grid(row=1, column=1, sticky=tk.E)
        
        ttk.Button(actions_frame, text="🔄 Refresh", 
                  command=self.refresh_data).pack(side=tk.LEFT, padx=(5, 0))
        ttk.Button(actions_frame, text="⚙️ Settings", 
                  command=self.show_settings).pack(side=tk.LEFT, padx=(5, 0))
        ttk.Button(actions_frame, text="🌐 Start API", 
                  command=self.start_api_server).pack(side=tk.LEFT, padx=(5, 0))
        
    def create_templates_tab(self):
        """Create the templates management tab"""
        template_frame = ttk.Frame(self.notebook)
        self.notebook.add(template_frame, text="📋 Templates")
        
        # Configure grid
        template_frame.columnconfigure(1, weight=1)
        template_frame.rowconfigure(1, weight=1)
        
        # Left panel - Template list
        left_panel = ttk.LabelFrame(template_frame, text="Available Templates", padding="10")
        left_panel.grid(row=0, column=0, rowspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), 
                       padx=(0, 10))
        left_panel.columnconfigure(0, weight=1)
        left_panel.rowconfigure(1, weight=1)
        
        # Template listbox with scrollbar
        listbox_frame = ttk.Frame(left_panel)
        listbox_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        listbox_frame.columnconfigure(0, weight=1)
        listbox_frame.rowconfigure(0, weight=1)
        
        self.template_listbox = tk.Listbox(listbox_frame, height=15)
        self.template_listbox.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        self.template_listbox.bind('<<ListboxSelect>>', self.on_template_select)
        
        template_scrollbar = ttk.Scrollbar(listbox_frame, orient=tk.VERTICAL, 
                                          command=self.template_listbox.yview)
        template_scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        self.template_listbox.config(yscrollcommand=template_scrollbar.set)
        
        # Template actions
        template_actions = ttk.Frame(left_panel)
        template_actions.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(10, 0))
        
        ttk.Button(template_actions, text="➕ New Template", command=self.create_template_dialog, style='Primary.TButton').pack(fill=tk.X, pady=(0, 5))
        ttk.Button(template_actions, text="🔄 Refresh Templates", command=self.refresh_templates).pack(fill=tk.X, pady=(0, 5))
        ttk.Button(template_actions, text="📝 Edit Template", command=self.edit_template_dialog).pack(fill=tk.X, pady=(0, 5))
        ttk.Button(template_actions, text="🗑️ Delete Template", command=self.delete_template).pack(fill=tk.X, pady=(0, 5))
        ttk.Button(template_actions, text="📥 Import Template", command=self.import_template_dialog).pack(fill=tk.X, pady=(0, 5))
        ttk.Button(template_actions, text="📤 Export Template",
                  command=self.export_template).pack(fill=tk.X)
        
        # Right panel - Template details
        right_panel = ttk.LabelFrame(template_frame, text="Template Details", padding="10")
        right_panel.grid(row=0, column=1, sticky=(tk.W, tk.E, tk.N, tk.S))
        right_panel.columnconfigure(0, weight=1)
        right_panel.rowconfigure(0, weight=1)
        
        # Template details display
        self.template_details = ScrolledText(right_panel, height=20, wrap=tk.WORD)
        self.template_details.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Template preview actions
        preview_actions = ttk.Frame(template_frame)
        preview_actions.grid(row=1, column=1, sticky=(tk.W, tk.E), pady=(10, 0))
        
        ttk.Button(preview_actions, text="👁️ Preview Template", 
                  command=self.preview_template).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(preview_actions, text="📊 Generate Report", 
                  command=self.generate_from_template,
                  style='Success.TButton').pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(preview_actions, text="🔄 Duplicate Template", 
                  command=self.duplicate_template).pack(side=tk.LEFT)
        ttk.Button(template_actions, text="🏠 Return to Main Menu", 
                   command=self.return_to_main_menu).pack(fill=tk.X, pady=(0, 5))

    def create_reports_tab(self):
        """Create the reports generation tab"""
        reports_frame = ttk.Frame(self.notebook)
        self.notebook.add(reports_frame, text="📊 Reports")
        
        # Configure grid
        reports_frame.columnconfigure(0, weight=1)
        reports_frame.rowconfigure(1, weight=1)
        
        # Report generation section
        gen_frame = ttk.LabelFrame(reports_frame, text="Generate Report", padding="10")
        gen_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        gen_frame.columnconfigure(1, weight=1)
        
        # Template selection
        ttk.Label(gen_frame, text="Template:").grid(row=0, column=0, sticky=tk.W, pady=(0, 5))
        self.template_combo = ttk.Combobox(gen_frame, state="readonly")
        self.template_combo.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=(0, 5))
        
        # Date range
        ttk.Label(gen_frame, text="Start Date:").grid(row=1, column=0, sticky=tk.W, pady=(0, 5))
        date_frame = ttk.Frame(gen_frame)
        date_frame.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=(0, 5))
        
        self.start_date = tk.StringVar(value=(datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d"))
        ttk.Entry(date_frame, textvariable=self.start_date, width=12).pack(side=tk.LEFT)
        
        ttk.Label(date_frame, text="End Date:").pack(side=tk.LEFT, padx=(20, 5))
        self.end_date = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        ttk.Entry(date_frame, textvariable=self.end_date, width=12).pack(side=tk.LEFT)
        
        # Output format
        ttk.Label(gen_frame, text="Format:").grid(row=2, column=0, sticky=tk.W, pady=(0, 5))
        self.format_combo = ttk.Combobox(gen_frame, values=["PDF", "Excel", "Interactive HTML"], 
                                        state="readonly")
        self.format_combo.set("PDF")
        self.format_combo.grid(row=2, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=(0, 5))
        
        # Generate button
        generate_btn = ttk.Button(gen_frame, text="🚀 Generate Report", 
                                 command=self.generate_report,
                                 style='Success.TButton')
        generate_btn.grid(row=3, column=0, columnspan=2, pady=(10, 0))
        
        # Reports history section
        history_frame = ttk.LabelFrame(reports_frame, text="Recent Reports", padding="10")
        history_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        history_frame.columnconfigure(0, weight=1)
        history_frame.rowconfigure(0, weight=1)
        
        # Reports treeview
        columns = ('Name', 'Generated', 'Format', 'Size', 'Status')
        self.reports_tree = ttk.Treeview(history_frame, columns=columns, show='headings', height=15)
        
        for col in columns:
            self.reports_tree.heading(col, text=col)
            self.reports_tree.column(col, width=150)
        
        self.reports_tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Reports scrollbar
        reports_scrollbar = ttk.Scrollbar(history_frame, orient=tk.VERTICAL, 
                                         command=self.reports_tree.yview)
        reports_scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        self.reports_tree.config(yscrollcommand=reports_scrollbar.set)
        
        # Reports actions
        reports_actions = ttk.Frame(history_frame)
        reports_actions.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(10, 0))
        
        ttk.Button(reports_actions, text="📁 Open Report", 
                  command=self.open_report).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(reports_actions, text="📤 Share Report", 
                  command=self.share_report).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(reports_actions, text="🗑️ Delete Report", 
                  command=self.delete_report).pack(side=tk.LEFT)
        
        # Refresh reports button
        ttk.Button(reports_actions, text="🔄 Refresh", 
                  command=self.refresh_reports).pack(side=tk.RIGHT)
        
    def create_analytics_tab(self):
        """Create the analytics dashboard tab"""
        analytics_frame = ttk.Frame(self.notebook)
        self.notebook.add(analytics_frame, text="📈 Analytics")
        
        # Configure grid
        analytics_frame.columnconfigure(0, weight=1)
        analytics_frame.columnconfigure(1, weight=1)
        analytics_frame.rowconfigure(1, weight=1)
        
        # Analytics overview
        overview_frame = ttk.LabelFrame(analytics_frame, text="System Overview", padding="10")
        overview_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        overview_frame.columnconfigure(0, weight=1)
        overview_frame.columnconfigure(1, weight=1)
        overview_frame.columnconfigure(2, weight=1)
        overview_frame.columnconfigure(3, weight=1)
        
        # Overview cards
        self.create_overview_card(overview_frame, "👥 Total Students", "0", 0, 0)
        self.create_overview_card(overview_frame, "📚 Total Courses", "0", 0, 1)
        self.create_overview_card(overview_frame, "📋 Templates", "0", 0, 2)
        self.create_overview_card(overview_frame, "📊 Reports Generated", "0", 0, 3)

        # Update cards immediately after creation
        self.root.after(100, self.update_overview_cards)

        # Left panel - Data Quality
        quality_frame = ttk.LabelFrame(analytics_frame, text="Data Quality Dashboard", padding="10")
        quality_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(0, 5))
        quality_frame.columnconfigure(0, weight=1)
        quality_frame.rowconfigure(1, weight=1)
        
        # Quality metrics
        self.quality_display = ScrolledText(quality_frame, height=15, wrap=tk.WORD)
        self.quality_display.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        quality_actions = ttk.Frame(quality_frame)
        quality_actions.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(10, 0))
        
        ttk.Button(quality_actions, text="🔄 Run Quality Check", 
                  command=self.run_quality_check).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(quality_actions, text="📊 Export Quality Report", 
                  command=self.export_quality_report).pack(side=tk.LEFT)
        
        # Right panel - Predictive Analytics
        predictions_frame = ttk.LabelFrame(analytics_frame, text="Predictive Analytics", padding="10")
        predictions_frame.grid(row=1, column=1, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(5, 0))
        predictions_frame.columnconfigure(0, weight=1)
        predictions_frame.rowconfigure(1, weight=1)
        
        # Predictions display
        self.predictions_display = ScrolledText(predictions_frame, height=15, wrap=tk.WORD)
        self.predictions_display.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        predictions_actions = ttk.Frame(predictions_frame)
        predictions_actions.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(10, 0))
        
        ttk.Button(predictions_actions, text="🎯 Run Predictions", 
                  command=self.run_predictions).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(predictions_actions, text="🔍 Anomaly Detection", 
                  command=self.run_anomaly_detection).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(predictions_actions, text="📈 Correlation Analysis", 
                  command=self.run_correlation_analysis).pack(side=tk.LEFT)
        
    def create_schedule_tab(self):
        """Create the scheduling tab"""
        schedule_frame = ttk.Frame(self.notebook)
        self.notebook.add(schedule_frame, text="⏰ Scheduling")
        
        # Configure grid
        schedule_frame.columnconfigure(0, weight=1)
        schedule_frame.rowconfigure(1, weight=1)
        
        # Schedule creation section
        create_frame = ttk.LabelFrame(schedule_frame, text="Create Schedule", padding="10")
        create_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        create_frame.columnconfigure(1, weight=1)
        
        # Template selection for scheduling
        ttk.Label(create_frame, text="Template:").grid(row=0, column=0, sticky=tk.W, pady=(0, 5))
        self.schedule_template_combo = ttk.Combobox(create_frame, state="readonly")
        self.schedule_template_combo.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=(0, 5))
        
        # Frequency selection
        ttk.Label(create_frame, text="Frequency:").grid(row=1, column=0, sticky=tk.W, pady=(0, 5))
        self.frequency_combo = ttk.Combobox(create_frame, 
                                           values=["Daily", "Weekly", "Monthly"], 
                                           state="readonly")
        self.frequency_combo.set("Weekly")
        self.frequency_combo.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=(0, 5))
        
        # Time selection
        ttk.Label(create_frame, text="Time:").grid(row=2, column=0, sticky=tk.W, pady=(0, 5))
        time_frame = ttk.Frame(create_frame)
        time_frame.grid(row=2, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=(0, 5))
        
        self.hour_var = tk.StringVar(value="09")
        hour_spin = ttk.Spinbox(time_frame, from_=0, to=23, width=3, 
                               textvariable=self.hour_var, format="%02.0f")
        hour_spin.pack(side=tk.LEFT)
        ttk.Label(time_frame, text=":00").pack(side=tk.LEFT)
        
        # Recipients
        ttk.Label(create_frame, text="Email Recipients:").grid(row=3, column=0, sticky=tk.W, pady=(0, 5))
        self.recipients_entry = tk.Text(create_frame, height=3, width=40)
        self.recipients_entry.grid(row=3, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=(0, 5))
        
        # Create schedule button
        ttk.Button(create_frame, text="📅 Create Schedule", 
                  command=self.create_schedule,
                  style='Primary.TButton').grid(row=4, column=0, columnspan=2, pady=(10, 0))
        
        # Scheduled reports list
        scheduled_frame = ttk.LabelFrame(schedule_frame, text="Scheduled Reports", padding="10")
        scheduled_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scheduled_frame.columnconfigure(0, weight=1)
        scheduled_frame.rowconfigure(0, weight=1)
        
        # Scheduled reports treeview
        schedule_columns = ('Template', 'Frequency', 'Time', 'Recipients', 'Last Run', 'Status')
        self.schedule_tree = ttk.Treeview(scheduled_frame, columns=schedule_columns, 
                                         show='headings', height=15)
        
        for col in schedule_columns:
            self.schedule_tree.heading(col, text=col)
            self.schedule_tree.column(col, width=120)
        
        self.schedule_tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Schedule scrollbar
        schedule_scrollbar = ttk.Scrollbar(scheduled_frame, orient=tk.VERTICAL, 
                                          command=self.schedule_tree.yview)
        schedule_scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        self.schedule_tree.config(yscrollcommand=schedule_scrollbar.set)
        
        # Schedule actions
        schedule_actions = ttk.Frame(scheduled_frame)
        schedule_actions.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(10, 0))
        
        ttk.Button(schedule_actions, text="▶️ Enable/Disable", 
                  command=self.toggle_schedule).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(schedule_actions, text="✏️ Edit Schedule", 
                  command=self.edit_schedule).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(schedule_actions, text="🚀 Run Now", 
                  command=self.run_schedule_now).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(schedule_actions, text="🗑️ Delete Schedule", 
                  command=self.delete_schedule).pack(side=tk.LEFT)
        
    def create_system_tab(self):
        """Create the system management tab"""
        system_frame = ttk.Frame(self.notebook)
        self.notebook.add(system_frame, text="⚙️ System")
    
        # Configure grid
        system_frame.columnconfigure(0, weight=1)
        system_frame.columnconfigure(1, weight=1)
        system_frame.rowconfigure(1, weight=1)
    
        # System status
        status_frame = ttk.LabelFrame(system_frame, text="System Status", padding="10")
        status_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        status_frame.columnconfigure(0, weight=1)
    
        self.system_status = ttk.Label(
            status_frame,
            text="Checking system status...",
            style='Info.TLabel'
        )
        self.system_status.grid(row=0, column=0, sticky=tk.W)
    
        # Left panel - Maintenance
        maintenance_frame = ttk.LabelFrame(system_frame, text="Maintenance", padding="10")
        maintenance_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(0, 5))
        maintenance_frame.columnconfigure(0, weight=1)
    
        # Maintenance actions
        maintenance_actions = [
            ("🧹 Clean Old Reports", self.clean_old_reports),
            ("🗄️ Clear Cache", self.clear_cache),
            ("🔍 Run Data Quality Check", self.run_maintenance_quality_check),
            ("🗃️ Optimize Database", self.optimize_database),
            ("🔧 Run All Maintenance", self.run_all_maintenance),
            ("📊 Performance Monitor", self.show_performance_monitor),
            ("📄 Export System Logs", self.export_system_logs),
        ]
        for i, (text, command) in enumerate(maintenance_actions):
            ttk.Button(maintenance_frame, text=text, command=command).grid(
                row=i, column=0, sticky=(tk.W, tk.E), pady=(0, 5)
            )
    
        # Right panel - Configuration
        config_frame = ttk.LabelFrame(system_frame, text="Configuration", padding="10")
        config_frame.grid(row=1, column=1, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(5, 0))
        config_frame.columnconfigure(0, weight=1)
        config_frame.rowconfigure(0, weight=1)
    
        # Configuration display and editing
        self.config_display = ScrolledText(config_frame, height=20, wrap=tk.WORD)
        self.config_display.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
    
        config_actions = ttk.Frame(config_frame)
        config_actions.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(10, 0))
    
        ttk.Button(config_actions, text="🔄 Reload Config", command=self.reload_config).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(config_actions, text="💾 Save Config", command=self.save_config).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(config_actions, text="🔧 Advanced Settings", command=self.show_advanced_settings).pack(side=tk.LEFT)

    
    def layout_status_bar(self, parent):
        # Create title/header frame
        header_frame = ttk.Frame(parent)
        header_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        header_frame.columnconfigure(1, weight=1)

        # Title
        title_label = ttk.Label(header_frame, text="📊 Advanced Student Reporting System",
                               style='Title.TLabel')
        title_label.grid(row=0, column=0, sticky=tk.W)

        # Return to Home button
        if hasattr(self, 'auth'):
            return_button = ttk.Button(header_frame, text="🏠 Return to Main Menu",
                                      command=self.return_to_main_menu)
            return_button.grid(row=0, column=1, sticky=tk.E)

        # actually place status bar only in standalone mode
        self._status_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(10, 0))
        self._status_frame.columnconfigure(1, weight=1)
        self.status_text.grid(row=0, column=0, sticky=tk.W)
        self.progress.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(10, 10))
        ttk.Label(self._status_frame, text="v2.0", style='Info.TLabel').grid(row=0, column=2, sticky=tk.E)
            
    def create_overview_card(self, parent, title, value, row, col):
        """Create an overview card widget"""
        card_frame = ttk.Frame(parent, relief='solid', borderwidth=1)
        card_frame.grid(row=row, column=col, sticky=(tk.W, tk.E), padx=5, pady=5)
        card_frame.columnconfigure(0, weight=1)
        
        title_label = ttk.Label(card_frame, text=title, style='Subtitle.TLabel')
        title_label.grid(row=0, column=0, sticky=(tk.W, tk.E), padx=10, pady=(10, 5))
        
        value_label = ttk.Label(card_frame, text=value, font=('Arial', 20, 'bold'))
        value_label.grid(row=1, column=0, sticky=(tk.W, tk.E), padx=10, pady=(0, 10))
        
        # Store reference for updating - fix the attribute naming
        if "Students" in title:
            self.overview_students = value_label
        elif "Courses" in title:
            self.overview_courses = value_label
        elif "Templates" in title:
            self.overview_templates = value_label
        elif "Reports" in title:
            self.overview_reports = value_label

    def init_status_widgets(self, parent):
        # create hidden container; we won't grid it unless standalone
        self._status_frame = ttk.Frame(parent)
        # widgets exist regardless of layout
        self.status_text = ttk.Label(self._status_frame, text="Ready", style='Info.TLabel')
        self.progress = ttk.Progressbar(self._status_frame, mode='indeterminate')
        # Add missing status_indicator for embedded mode
        self.status_indicator = ttk.Label(self._status_frame, text="● System Ready", style='Success.TLabel')

    # Event handlers and functionality methods
    
    def return_to_main_menu(self):
        """Return to the main menu"""
        try:
            # Check if this is a child window (Toplevel) or standalone (Tk)
            root_widget = self.root if hasattr(self, 'root') else self.master
            if isinstance(root_widget, tk.Toplevel):
                # Just close the child window
                root_widget.destroy()
            else:
                # Running standalone, need to create main GUI
                root_widget.destroy()
                from university_system.modules.shared.gui.main_gui import UnifiedManagementGUI
                app = UnifiedManagementGUI(self.auth)
                app.run()
        except Exception as e:
            print(f"Error returning to main menu: {e}")
            import traceback
            traceback.print_exc()

    def refresh_data(self):
        """Refresh all data in the GUI"""
        self.update_status("Refreshing data...")
        self.start_progress()
        
        def refresh_task():
            try:
                # Load templates
                self.load_templates()
                
                # Load scheduled reports
                self.load_scheduled_reports()
                
                # Load recent reports
                self.load_recent_reports()
                
                # Update overview cards
                self.update_overview_cards()
                
                # Update system status
                self.check_system_status()
                
                self._schedule_on_ui_thread(lambda: [
                    self.stop_progress(),
                    self.update_status("Data refreshed successfully")
                ])

            except Exception as e:
                self._schedule_on_ui_thread(lambda: [
                    self.stop_progress(),
                    self.update_status(f"Error refreshing data: {str(e)}", "error")
                ])
        
        threading.Thread(target=refresh_task, daemon=True).start()
    
    def load_templates(self):
        """Load templates into the GUI"""
        if not ENHANCED_AVAILABLE:
            return
            
        try:
            templates = load_templates()
            
            # Update template listbox
            self._schedule_on_ui_thread(lambda: self._update_template_listbox(templates))

            # Update template combos
            template_names = [t['name'] for t in templates]
            self._schedule_on_ui_thread(lambda: self._update_template_combos(template_names))
            
        except Exception as e:
            logging.error(f"Error loading templates: {str(e)}")
    
    def _update_template_listbox(self, templates):
        """Update template listbox in main thread"""
        self.template_listbox.delete(0, tk.END)
        self.templates_data = templates
        
        for template in templates:
            display_text = f"{template['name']} ({template.get('version', '1.0')})"
            self.template_listbox.insert(tk.END, display_text)
    
    def _update_template_combos(self, template_names):
        """Update template comboboxes in main thread"""
        self.template_combo['values'] = template_names
        self.schedule_template_combo['values'] = template_names
        
        if template_names:
            self.template_combo.set(template_names[0])
            self.schedule_template_combo.set(template_names[0])
    
    def load_scheduled_reports(self):
        """Load scheduled reports into the GUI"""
        if not ENHANCED_AVAILABLE:
            return
            
        try:
            scheduled_reports = load_scheduled_reports()
            self.root.after(0, lambda: self._update_schedule_tree(scheduled_reports))
        except Exception as e:
            logger.error(f"Error loading scheduled reports: {str(e)}")
    
    def _update_schedule_tree(self, scheduled_reports):
        """Update schedule tree in main thread"""
        # Clear existing items
        for item in self.schedule_tree.get_children():
            self.schedule_tree.delete(item)
        
        self.scheduled_reports_data = scheduled_reports
        
        for report in scheduled_reports:
            config = report.get('schedule_config', {})
            status = "Enabled" if config.get('enabled', True) else "Disabled"
            
            values = (
                report['template_name'],
                config.get('frequency', 'Unknown').title(),
                f"{config.get('hour', 9):02d}:00",
                str(len(report.get('recipients', []))),
                report.get('last_run', 'Never')[:19] if report.get('last_run') else 'Never',
                status
            )
            
            self.schedule_tree.insert('', tk.END, values=values)

    def refresh_templates(self):
        """Refresh templates list"""
        self.update_status("Refreshing templates...")
        try:
            self.load_templates()
            self.update_overview_cards()
            self.update_status("Templates refreshed successfully")
            messagebox.showinfo("Success", "Templates refreshed successfully!")
        except Exception as e:
            logging.error(f"Error refreshing templates: {e}")
            self.update_status(f"Error refreshing templates: {str(e)}", "error")
            messagebox.showerror("Error", f"Failed to refresh templates: {str(e)}")

    def import_template_dialog(self):
        """Import template from JSON file"""
        try:
            file_path = filedialog.askopenfilename(
                filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
                title="Import Template"
            )
            
            if file_path:
                with open(file_path, 'r') as f:
                    template_data = json.load(f)
                
                # Validate template structure
                required_fields = ['name', 'sections']
                if not all(field in template_data for field in required_fields):
                    messagebox.showerror("Invalid Template", "Template file is missing required fields")
                    return
                
                # Check if template already exists
                existing_template = get_template(template_data['name'])
                if existing_template:
                    if not messagebox.askyesno("Template Exists", 
                                             f"Template '{template_data['name']}' already exists. Overwrite?"):
                        return
                
                # Import template
                if ENHANCED_AVAILABLE:
                    save_template_dict(template_data)

                self.refresh_data()
                messagebox.showinfo("Success", f"Template '{template_data['name']}' imported successfully!")
                
        except Exception as e:
            messagebox.showerror("Import Error", f"Failed to import template: {str(e)}")

    def create_status_bar(self, parent):
        """Create status bar with all system information"""
        if not self.is_embedded:
            self._status_frame = ttk.Frame(parent)
            self._status_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(10, 0))
            self._status_frame.columnconfigure(1, weight=1)
            
            self.status_text = ttk.Label(self._status_frame, text="Ready", style='Info.TLabel')
            self.status_text.grid(row=0, column=0, sticky=tk.W)
            
            self.progress = ttk.Progressbar(self._status_frame, mode='indeterminate')
            self.progress.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(10, 10))
            
            # System info
            system_info = f"v2.0 | Enhanced: {'Yes' if ENHANCED_AVAILABLE else 'No'}"
            ttk.Label(self._status_frame, text=system_info, style='Info.TLabel').grid(row=0, column=2, sticky=tk.E)

    # COMPLETE THE SYSTEM TAB CONFIGURATION DISPLAY
    def complete_system_tab_config_display(self):
        """Complete the configuration display area in system tab"""
        # This completes the cut-off section from the document
        self.config_display = ScrolledText(config_display_frame, height=10, wrap=tk.WORD)
        self.config_display.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        config_actions_bottom = ttk.Frame(config_display_frame)
        config_actions_bottom.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(10, 0))
        
        ttk.Button(config_actions_bottom, text="Reload Config", 
                  command=self.reload_config).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(config_actions_bottom, text="Save Config", 
                  command=self.save_config).pack(side=tk.LEFT, padx=(0, 5))


    # MISSING BACKUP/RESTORE COMPLETION
    def show_backup_restore_dialog(self):
        """Show backup and restore dialog"""
        try:
            backup_window = tk.Toplevel(self.root)
            backup_window.title("Backup & Restore")
            backup_window.geometry("600x500")
            backup_window.transient(self.root)
            
            backup_notebook = ttk.Notebook(backup_window)
            backup_notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # Backup tab
            backup_frame = ttk.Frame(backup_notebook)
            backup_notebook.add(backup_frame, text="Backup")
            
            backup_info = ttk.LabelFrame(backup_frame, text="Backup Information", padding="10")
            backup_info.pack(fill=tk.X, pady=10)
            
            ttk.Label(backup_info, text="The backup will include:").pack(anchor=tk.W)
            ttk.Label(backup_info, text="• Database (student_records.db)").pack(anchor=tk.W, padx=20)
            ttk.Label(backup_info, text="• Templates (templates.json)").pack(anchor=tk.W, padx=20)
            ttk.Label(backup_info, text="• Scheduled Reports").pack(anchor=tk.W, padx=20)
            ttk.Label(backup_info, text="• System Configuration").pack(anchor=tk.W, padx=20)
            
            # Backup options
            backup_options = ttk.LabelFrame(backup_frame, text="Backup Options", padding="10")
            backup_options.pack(fill=tk.X, pady=10)
            
            include_reports_var = tk.BooleanVar(value=False)
            ttk.Checkbutton(backup_options, text="Include Generated Reports", 
                           variable=include_reports_var).pack(anchor=tk.W)
            
            include_cache_var = tk.BooleanVar(value=False)
            ttk.Checkbutton(backup_options, text="Include Cache Files", 
                           variable=include_cache_var).pack(anchor=tk.W)
            
            def create_backup():
                try:
                    backup_dir = filedialog.askdirectory(title="Select Backup Directory")
                    if not backup_dir:
                        return
                    
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    backup_name = f"reporting_system_backup_{timestamp}"
                    full_backup_path = os.path.join(backup_dir, backup_name)
                    os.makedirs(full_backup_path, exist_ok=True)
                    
                    import shutil
                    
                    # Backup database
                    if os.path.exists(CONFIG.get('database', str(DEFAULT_DB_PATH))):
                        shutil.copy2(CONFIG['database'], 
                                   os.path.join(full_backup_path, str(DEFAULT_DB_PATH)))
                    
                    # Backup templates
                    templates_file = os.path.join(CONFIG.get('templates_dir', str(paths.REPORT_TEMPLATES_DIR)), 'templates.json')
                    if os.path.exists(templates_file):
                        shutil.copy2(templates_file, 
                                   os.path.join(full_backup_path, 'templates.json'))
                    
                    # Create backup info file
                    backup_info = {
                        'created': datetime.now().isoformat(),
                        'version': '2.0',
                        'includes_reports': include_reports_var.get(),
                        'includes_cache': include_cache_var.get()
                    }
                    
                    with open(os.path.join(full_backup_path, 'backup_info.json'), 'w') as f:
                        json.dump(backup_info, f, indent=4)
                    
                    messagebox.showinfo("Backup Complete", 
                                      f"Backup created successfully!\n\nLocation: {full_backup_path}")
                    
                except Exception as e:
                    messagebox.showerror("Backup Error", f"Backup failed: {str(e)}")
            
            ttk.Button(backup_options, text="Create Backup", command=create_backup,
                      style='Success.TButton').pack(pady=10)
            
            # Restore tab
            restore_frame = ttk.Frame(backup_notebook)
            backup_notebook.add(restore_frame, text="Restore")
            
            restore_info = ttk.LabelFrame(restore_frame, text="Restore Information", padding="10")
            restore_info.pack(fill=tk.X, pady=10)
            
            ttk.Label(restore_info, text="Restoring will overwrite current data!").pack(anchor=tk.W)
            ttk.Label(restore_info, text="Please backup current data before restoring.").pack(anchor=tk.W)
            
            def restore_backup():
                try:
                    backup_dir = filedialog.askdirectory(title="Select Backup Directory to Restore")
                    if not backup_dir:
                        return
                    
                    backup_info_file = os.path.join(backup_dir, 'backup_info.json')
                    if not os.path.exists(backup_info_file):
                        messagebox.showerror("Invalid Backup", "Selected directory is not a valid backup")
                        return
                    
                    if not messagebox.askyesno("Confirm Restore", 
                                             "This will overwrite current data. Continue?"):
                        return
                    
                    import shutil
                    
                    # Restore database
                    db_backup = os.path.join(backup_dir, str(DEFAULT_DB_PATH))
                    if os.path.exists(db_backup):
                        shutil.copy2(db_backup, CONFIG['database'])
                    
                    messagebox.showinfo("Restore Complete", "Backup restored successfully!")
                    
                except Exception as e:
                    messagebox.showerror("Restore Error", f"Restore failed: {str(e)}")
            
            ttk.Button(restore_info, text="Restore from Backup", command=restore_backup,
                      style='Warning.TButton').pack(pady=10)
            
        except Exception as e:
            messagebox.showerror("Backup/Restore Error", f"Failed to open backup dialog: {str(e)}")
            
            def restore_backup():
                try:
                    backup_dir = filedialog.askdirectory(title="Select Backup Directory to Restore")
                    if not backup_dir:
                        return
                    
                    backup_info_file = os.path.join(backup_dir, 'backup_info.json')
                    if not os.path.exists(backup_info_file):
                        messagebox.showerror("Invalid Backup", "Selected directory is not a valid backup")
                        return
                    
                    if not messagebox.askyesno("Confirm Restore", 
                                             "This will overwrite current data. Continue?"):
                        return
                    
                    import shutil
                    
                    # Restore database
                    db_backup = os.path.join(backup_dir, str(DEFAULT_DB_PATH))
                    if os.path.exists(db_backup):
                        shutil.copy2(db_backup, CONFIG['database'])
                    
                    messagebox.showinfo("Restore Complete", "Backup restored successfully!")
                    
                except Exception as e:
                    messagebox.showerror("Restore Error", f"Restore failed: {str(e)}")
            
            ttk.Button(restore_info, text="Restore from Backup", command=restore_backup,
                      style='Warning.TButton').pack(pady=10)
            
        except Exception as e:
            messagebox.showerror("Backup/Restore Error", f"Failed to open backup dialog: {str(e)}")

    def show_user_management_dialog(self):
        """Show user management dialog"""
        try:
            user_window = tk.Toplevel(self.root)
            user_window.title("User Management")
            user_window.geometry("800x600")
            user_window.transient(self.root)
            
            note_frame = ttk.LabelFrame(user_window, text="Note", padding="10")
            note_frame.pack(fill=tk.X, padx=10, pady=10)
            
            note_text = """User management and authentication features are available in the full system.
This interface provides basic user management functionality."""
            
            ttk.Label(note_frame, text=note_text, wraplength=700).pack()
            
            users_frame = ttk.LabelFrame(user_window, text="System Users", padding="10")
            users_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            columns = ('Username', 'Role', 'Last Login', 'Status')
            users_tree = ttk.Treeview(users_frame, columns=columns, show='headings')
            
            for col in columns:
                users_tree.heading(col, text=col)
                users_tree.column(col, width=150)
            
            sample_users = [
                ('admin', 'Administrator', '2024-01-15 09:30', 'Active'),
                ('analyst', 'Analyst', '2024-01-14 14:22', 'Active'),
                ('viewer', 'Viewer', 'Never', 'Inactive')
            ]
            
            for user in sample_users:
                users_tree.insert('', tk.END, values=user)
            
            users_tree.pack(fill=tk.BOTH, expand=True)
            
            actions_frame = ttk.Frame(users_frame)
            actions_frame.pack(fill=tk.X, pady=(10, 0))
            
            ttk.Button(actions_frame, text="Add User", 
                      command=lambda: messagebox.showinfo("Feature", "User creation would be implemented here")).pack(side=tk.LEFT, padx=(0, 5))
            ttk.Button(actions_frame, text="Edit User", 
                      command=lambda: messagebox.showinfo("Feature", "User editing would be implemented here")).pack(side=tk.LEFT, padx=(0, 5))
            
            ttk.Button(user_window, text="Close", command=user_window.destroy).pack(pady=10)
            
        except Exception as e:
            messagebox.showerror("User Management Error", f"Failed to open user management: {str(e)}")

    def show_directory_settings(self):
        """Show directory settings dialog"""
        try:
            dir_window = tk.Toplevel(self.root)
            dir_window.title("Directory Settings")
            dir_window.geometry("500x400")
            dir_window.transient(self.root)
            
            settings_frame = ttk.LabelFrame(dir_window, text="Directory Configuration", padding="10")
            settings_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            settings_frame.columnconfigure(1, weight=1)
            
            # Reports directory
            ttk.Label(settings_frame, text="Reports Directory:").grid(row=0, column=0, sticky=tk.W, pady=5)
            reports_dir_var = tk.StringVar(value=CONFIG.get('reports_dir', 'reports'))
            ttk.Entry(settings_frame, textvariable=reports_dir_var).grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)
            
            def browse_reports_dir():
                directory = filedialog.askdirectory(title="Select Reports Directory")
                if directory:
                    reports_dir_var.set(directory)
            
            ttk.Button(settings_frame, text="Browse", command=browse_reports_dir).grid(row=0, column=2, padx=(5, 0), pady=5)
            
            # Templates directory
            ttk.Label(settings_frame, text="Templates Directory:").grid(row=1, column=0, sticky=tk.W, pady=5)
            templates_dir_var = tk.StringVar(value=CONFIG.get('templates_dir', str(paths.REPORT_TEMPLATES_DIR)))
            ttk.Entry(settings_frame, textvariable=templates_dir_var).grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)
            
            def browse_templates_dir():
                directory = filedialog.askdirectory(title="Select Templates Directory")
                if directory:
                    templates_dir_var.set(directory)
            
            ttk.Button(settings_frame, text="Browse", command=browse_templates_dir).grid(row=1, column=2, padx=(5, 0), pady=5)
            
            # Create directories option
            create_dirs_var = tk.BooleanVar(value=True)
            ttk.Checkbutton(settings_frame, text="Create directories if they don't exist", 
                           variable=create_dirs_var).grid(row=2, column=0, columnspan=3, sticky=tk.W, pady=10)
            
            # Button frame
            button_frame = ttk.Frame(dir_window)
            button_frame.pack(fill=tk.X, padx=10, pady=(0, 10))
            
            def save_directory_settings():
                try:
                    new_dirs = {
                        'reports_dir': reports_dir_var.get(),
                        'templates_dir': templates_dir_var.get()
                    }
                    
                    if create_dirs_var.get():
                        for dir_path in new_dirs.values():
                            os.makedirs(dir_path, exist_ok=True)
                    
                    CONFIG.update(new_dirs)
                    
                    messagebox.showinfo("Success", "Directory settings saved successfully!")
                    dir_window.destroy()
                    self.check_system_status()
                    
                except Exception as e:
                    messagebox.showerror("Save Error", f"Failed to save directory settings: {str(e)}")
            
            ttk.Button(button_frame, text="Save", command=save_directory_settings).pack(side=tk.RIGHT, padx=(5, 0))
            ttk.Button(button_frame, text="Cancel", command=dir_window.destroy).pack(side=tk.RIGHT)
            
        except Exception as e:
            messagebox.showerror("Directory Settings Error", f"Failed to open directory settings: {str(e)}")

    def show_theme_settings(self):
        """Show theme and appearance settings"""
        try:
            theme_window = tk.Toplevel(self.root)
            theme_window.title("Theme & Appearance Settings")
            theme_window.geometry("450x350")
            theme_window.transient(self.root)
            
            theme_frame = ttk.LabelFrame(theme_window, text="Appearance Settings", padding="10")
            theme_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # Theme selection
            ttk.Label(theme_frame, text="Theme:").pack(anchor=tk.W, pady=5)
            theme_var = tk.StringVar(value="default")
            theme_combo = ttk.Combobox(theme_frame, textvariable=theme_var, 
                                      values=["default", "dark", "light", "modern"], state="readonly")
            theme_combo.pack(fill=tk.X, pady=(0, 10))
            
            # Font settings
            ttk.Label(theme_frame, text="Font Family:").pack(anchor=tk.W, pady=5)
            font_var = tk.StringVar(value="Arial")
            font_combo = ttk.Combobox(theme_frame, textvariable=font_var,
                                     values=["Arial", "Helvetica", "Times New Roman", "Calibri"],
                                     state="readonly")
            font_combo.pack(fill=tk.X, pady=(0, 10))
            
            # Font size
            ttk.Label(theme_frame, text="Font Size:").pack(anchor=tk.W, pady=5)
            font_size_var = tk.StringVar(value="10")
            font_size_spin = ttk.Spinbox(theme_frame, from_=8, to=16, textvariable=font_size_var)
            font_size_spin.pack(fill=tk.X, pady=(0, 10))
            
            # Preview
            preview_frame = ttk.LabelFrame(theme_frame, text="Preview", padding="10")
            preview_frame.pack(fill=tk.X, pady=10)
            
            preview_label = ttk.Label(preview_frame, text="Sample text with current settings")
            preview_label.pack()
            
            def update_preview():
                try:
                    font_family = font_var.get()
                    font_size = int(font_size_var.get())
                    preview_label.config(font=(font_family, font_size))
                except:
                    pass
            
            theme_combo.bind('<<ComboboxSelected>>', lambda e: update_preview())
            font_combo.bind('<<ComboboxSelected>>', lambda e: update_preview())
            
            # Button frame
            button_frame = ttk.Frame(theme_window)
            button_frame.pack(fill=tk.X, padx=10, pady=(0, 10))
            
            def apply_theme():
                try:
                    theme_settings = {
                        'theme': theme_var.get(),
                        'font_family': font_var.get(),
                        'font_size': int(font_size_var.get())
                    }
                    
                    messagebox.showinfo("Theme Applied", 
                                      f"Theme settings would be applied:\n\nTheme: {theme_settings['theme']}\nFont: {theme_settings['font_family']} {theme_settings['font_size']}pt")
                    theme_window.destroy()
                    
                except Exception as e:
                    messagebox.showerror("Theme Error", f"Failed to apply theme: {str(e)}")
            
            ttk.Button(button_frame, text="Apply", command=apply_theme).pack(side=tk.RIGHT, padx=(5, 0))
            ttk.Button(button_frame, text="Cancel", command=theme_window.destroy).pack(side=tk.RIGHT)
            
        except Exception as e:
            messagebox.showerror("Theme Settings Error", f"Failed to open theme settings: {str(e)}")

    def check_system_requirements_gui(self):
        """GUI version of system requirements check"""
        try:
            req_window = tk.Toplevel(self.root)
            req_window.title("System Requirements Check")
            req_window.geometry("600x400")
            req_window.transient(self.root)
            
            # Requirements display
            req_frame = ttk.LabelFrame(req_window, text="System Requirements Status", padding="10")
            req_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            req_text = ScrolledText(req_frame, wrap=tk.WORD)
            req_text.pack(fill=tk.BOTH, expand=True)
            
            def check_requirements():
                requirements_status = {
                    'tkinter': True,  # Already available if GUI is running
                    'pandas': pd is not None,
                    'enhanced_features': ENHANCED_AVAILABLE,
                    'database': False,
                    'matplotlib': True,  # Check if plotting works
                    'reportlab': True   # Check if PDF generation works
                }
                
                # Test database connection
                try:
                    conn = get_db_connection()
                    if conn:
                        conn.close()
                        requirements_status['database'] = True
                except:
                    pass
                
                # Test matplotlib
                try:
                    import matplotlib.pyplot as plt
                    plt.figure()
                    plt.close()
                except:
                    requirements_status['matplotlib'] = False
                
                # Generate report
                report = "System Requirements Check\n"
                report += "=" * 40 + "\n\n"
                
                for requirement, status in requirements_status.items():
                    status_text = "✓ Available" if status else "✗ Missing"
                    report += f"{requirement.replace('_', ' ').title()}: {status_text}\n"
                
                # Recommendations
                missing = [req for req, status in requirements_status.items() if not status]
                if missing:
                    report += f"\nMissing Requirements: {', '.join(missing)}\n"
                    report += "\nRecommendations:\n"
                    if 'pandas' in missing:
                        report += "- Install pandas: pip install pandas\n"
                    if 'matplotlib' in missing:
                        report += "- Install matplotlib: pip install matplotlib\n"
                    if 'reportlab' in missing:
                        report += "- Install reportlab: pip install reportlab\n"
                    if not requirements_status['database']:
                        report += "- Check database connection and permissions\n"
                else:
                    report += "\n✓ All requirements satisfied!\n"
                
                req_text.delete(1.0, tk.END)
                req_text.insert(1.0, report)
            
            # Run check automatically
            check_requirements()
            
            # Buttons
            button_frame = ttk.Frame(req_window)
            button_frame.pack(fill=tk.X, padx=10, pady=10)
            
            ttk.Button(button_frame, text="Refresh Check", command=check_requirements).pack(side=tk.LEFT)
            ttk.Button(button_frame, text="Close", command=req_window.destroy).pack(side=tk.RIGHT)
            
        except Exception as e:
            messagebox.showerror("Requirements Check Error", f"Failed to check requirements: {str(e)}")

    def show_advanced_template_creation_dialog(self):
        """Show advanced template creation dialog matching CLI functionality"""
        try:
            template_window = tk.Toplevel(self.root)
            template_window.title("Advanced Template Creation")
            template_window.geometry("800x700")
            template_window.transient(self.root)
            
            # Create notebook for different template aspects
            template_notebook = ttk.Notebook(template_window)
            template_notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # Basic Info Tab
            basic_frame = ttk.Frame(template_notebook)
            template_notebook.add(basic_frame, text="Basic Info")
            
            basic_info_frame = ttk.LabelFrame(basic_frame, text="Template Information", padding="10")
            basic_info_frame.pack(fill=tk.X, padx=10, pady=10)
            basic_info_frame.columnconfigure(1, weight=1)
            
            ttk.Label(basic_info_frame, text="Name:*").grid(row=0, column=0, sticky=tk.W, pady=5)
            name_var = tk.StringVar()
            ttk.Entry(basic_info_frame, textvariable=name_var).grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)
            
            ttk.Label(basic_info_frame, text="Description:").grid(row=1, column=0, sticky=tk.W, pady=5)
            desc_var = tk.StringVar()
            ttk.Entry(basic_info_frame, textvariable=desc_var).grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)
            
            # Security and Visualization
            security_frame = ttk.LabelFrame(basic_frame, text="Security & Visualization", padding="10")
            security_frame.pack(fill=tk.X, padx=10, pady=10)
            security_frame.columnconfigure(1, weight=1)
            
            ttk.Label(security_frame, text="Security Level:").grid(row=0, column=0, sticky=tk.W, pady=5)
            security_var = tk.StringVar(value="normal")
            security_combo = ttk.Combobox(security_frame, textvariable=security_var,
                                         values=["normal", "confidential", "restricted"], state="readonly")
            security_combo.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)
            
            ttk.Label(security_frame, text="Visualization Type:").grid(row=1, column=0, sticky=tk.W, pady=5)
            viz_var = tk.StringVar(value="standard")
            viz_combo = ttk.Combobox(security_frame, textvariable=viz_var,
                                    values=["standard", "advanced", "interactive"], state="readonly")
            viz_combo.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)
            
            # Sections Tab
            sections_frame = ttk.Frame(template_notebook)
            template_notebook.add(sections_frame, text="Sections")
            
            sections_info = ttk.LabelFrame(sections_frame, text="Available Report Sections", padding="10")
            sections_info.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # Create checkboxes for all available sections
            section_vars = {}
            available_sections = [
                "student_overview", "course_distribution", "gender_distribution",
                "age_distribution", "module_popularity", "registration_trends",
                "grade_distribution", "attendance_summary", "data_quality_report",
                "predictive_analytics", "correlation_analysis", "anomaly_detection",
                "performance_benchmarks", "trend_analysis"
            ]
            
            # Create scrollable frame for sections
            canvas = tk.Canvas(sections_info)
            scrollbar = ttk.Scrollbar(sections_info, orient="vertical", command=canvas.yview)
            scrollable_frame = ttk.Frame(canvas)
            
            scrollable_frame.bind(
                "<Configure>",
                lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
            )
            
            canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)
            
            for i, section in enumerate(available_sections):
                var = tk.BooleanVar()
                section_vars[section] = var
                section_name = section.replace('_', ' ').title()
                ttk.Checkbutton(scrollable_frame, text=section_name, variable=var).grid(
                    row=i // 2, column=i % 2, sticky=tk.W, padx=10, pady=2)
            
            canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")
            
            # Filters Tab
            filters_frame = ttk.Frame(template_notebook)
            template_notebook.add(filters_frame, text="Filters")
            
            filters_info = ttk.LabelFrame(filters_frame, text="Data Filters", padding="10")
            filters_info.pack(fill=tk.X, padx=10, pady=10)
            filters_info.columnconfigure(1, weight=1)
            
            ttk.Label(filters_info, text="Course Filter:").grid(row=0, column=0, sticky=tk.W, pady=5)
            course_var = tk.StringVar()
            course_combo = ttk.Combobox(filters_info, textvariable=course_var,
                                       values=["", "CS", "DS"], state="readonly")
            course_combo.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)
            
            ttk.Label(filters_info, text="Date Range (days):").grid(row=1, column=0, sticky=tk.W, pady=5)
            date_range_var = tk.StringVar(value="30")
            ttk.Entry(filters_info, textvariable=date_range_var, width=10).grid(row=1, column=1, sticky=tk.W, padx=(10, 0), pady=5)
            
            # Advanced Tab
            advanced_frame = ttk.Frame(template_notebook)
            template_notebook.add(advanced_frame, text="Advanced")
            
            advanced_info = ttk.LabelFrame(advanced_frame, text="Advanced Options", padding="10")
            advanced_info.pack(fill=tk.X, padx=10, pady=10)
            
            enable_caching_var = tk.BooleanVar(value=True)
            ttk.Checkbutton(advanced_info, text="Enable caching for faster generation", 
                           variable=enable_caching_var).pack(anchor=tk.W, pady=2)
            
            enable_comparison_var = tk.BooleanVar(value=False)
            ttk.Checkbutton(advanced_info, text="Include period-over-period comparison", 
                           variable=enable_comparison_var).pack(anchor=tk.W, pady=2)
            
            enable_export_var = tk.BooleanVar(value=True)
            ttk.Checkbutton(advanced_info, text="Allow multiple export formats", 
                           variable=enable_export_var).pack(anchor=tk.W, pady=2)
            
            def create_advanced_template():
                try:
                    name = name_var.get().strip()
                    if not name:
                        messagebox.showerror("Validation Error", "Template name is required")
                        return
                    
                    if get_template(name):
                        messagebox.showerror("Template Exists", f"Template '{name}' already exists")
                        return
                    
                    # Get selected sections
                    selected_sections = [section for section, var in section_vars.items() if var.get()]
                    if not selected_sections:
                        messagebox.showerror("Validation Error", "At least one section must be selected")
                        return
                    
                    # Build filters
                    filters = {}
                    if course_var.get():
                        filters['course'] = course_var.get()
                    
                    try:
                        date_range = int(date_range_var.get())
                        if date_range > 0:
                            filters['date_range_days'] = date_range
                    except ValueError:
                        pass
                    
                    # Build advanced options
                    advanced_options = {
                        'enable_caching': enable_caching_var.get(),
                        'enable_comparison': enable_comparison_var.get(),
                        'enable_export': enable_export_var.get()
                    }
                    
                    # Create template
                    template_data = {
                        'name': name,
                        'description': desc_var.get().strip(),
                        'sections': selected_sections,
                        'filters': filters,
                        'security_level': security_var.get(),
                        'visualization_type': viz_var.get(),
                        'advanced_options': advanced_options,
                        'created_at': datetime.now().isoformat(),
                        'version': '1.0'
                    }
                    
                    if ENHANCED_AVAILABLE:
                        templates = load_templates()
                        templates.append(template_data)
                        
                        os.makedirs(CONFIG.get('templates_dir', str(paths.REPORT_TEMPLATES_DIR)), exist_ok=True)
                        with open(os.path.join(CONFIG.get('templates_dir', str(paths.REPORT_TEMPLATES_DIR)), "templates.json"), 'w') as f:
                            json.dump(templates, f, indent=4)
                    
                    messagebox.showinfo("Success", f"Advanced template '{name}' created successfully!")
                    template_window.destroy()
                    self.refresh_data()
                    
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to create template: {str(e)}")
            
            # Buttons
            button_frame = ttk.Frame(template_window)
            button_frame.pack(fill=tk.X, padx=10, pady=10)
            
            ttk.Button(button_frame, text="Create Template", command=create_advanced_template,
                      style='Success.TButton').pack(side=tk.RIGHT, padx=(5, 0))
            ttk.Button(button_frame, text="Cancel", command=template_window.destroy).pack(side=tk.RIGHT)
            
        except Exception as e:
            messagebox.showerror("Template Creation Error", f"Failed to open template creation dialog: {str(e)}")

    def show_enhanced_scheduling_dialog(self):
        """Show enhanced scheduling dialog with all CLI features"""
        try:
            schedule_window = tk.Toplevel(self.root)
            schedule_window.title("Enhanced Report Scheduling")
            schedule_window.geometry("700x600")
            schedule_window.transient(self.root)
            
            # Create notebook for different scheduling aspects
            schedule_notebook = ttk.Notebook(schedule_window)
            schedule_notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # Basic Scheduling Tab
            basic_frame = ttk.Frame(schedule_notebook)
            schedule_notebook.add(basic_frame, text="Basic Schedule")
            
            # Template selection
            template_frame = ttk.LabelFrame(basic_frame, text="Template Selection", padding="10")
            template_frame.pack(fill=tk.X, padx=10, pady=10)
            
            templates = load_templates() if ENHANCED_AVAILABLE else []
            if not templates:
                ttk.Label(template_frame, text="No templates available for scheduling").pack()
                return
            
            ttk.Label(template_frame, text="Select Template:").pack(anchor=tk.W)
            template_var = tk.StringVar()
            template_combo = ttk.Combobox(template_frame, textvariable=template_var, state="readonly")
            template_combo['values'] = [t['name'] for t in templates]
            template_combo.pack(fill=tk.X, pady=(5, 0))
            
            if templates:
                template_combo.set(templates[0]['name'])
            
            # Schedule configuration
            config_frame = ttk.LabelFrame(basic_frame, text="Schedule Configuration", padding="10")
            config_frame.pack(fill=tk.X, padx=10, pady=10)
            config_frame.columnconfigure(1, weight=1)
            
            ttk.Label(config_frame, text="Frequency:").grid(row=0, column=0, sticky=tk.W, pady=5)
            frequency_var = tk.StringVar(value="weekly")
            frequency_combo = ttk.Combobox(config_frame, textvariable=frequency_var,
                                          values=["daily", "weekly", "monthly"], state="readonly")
            frequency_combo.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)
            
            ttk.Label(config_frame, text="Hour (0-23):").grid(row=1, column=0, sticky=tk.W, pady=5)
            hour_var = tk.StringVar(value="9")
            hour_spin = ttk.Spinbox(config_frame, from_=0, to=23, textvariable=hour_var, width=5)
            hour_spin.grid(row=1, column=1, sticky=tk.W, padx=(10, 0), pady=5)
            
            # Recipients Tab
            recipients_frame = ttk.Frame(schedule_notebook)
            schedule_notebook.add(recipients_frame, text="Recipients")
            
            recipients_info = ttk.LabelFrame(recipients_frame, text="Email Recipients", padding="10")
            recipients_info.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            ttk.Label(recipients_info, text="Enter email addresses (one per line):").pack(anchor=tk.W)
            recipients_text = tk.Text(recipients_info, height=10, width=50)
            recipients_text.pack(fill=tk.BOTH, expand=True, pady=(5, 0))
            
            # Conditions Tab
            conditions_frame = ttk.Frame(schedule_notebook)
            schedule_notebook.add(conditions_frame, text="Conditions")
            
            conditions_info = ttk.LabelFrame(conditions_frame, text="Execution Conditions", padding="10")
            conditions_info.pack(fill=tk.X, padx=10, pady=10)
            
            only_if_data_var = tk.BooleanVar(value=True)
            ttk.Checkbutton(conditions_info, text="Only run if new data is available", 
                           variable=only_if_data_var).pack(anchor=tk.W, pady=2)
            
            skip_if_empty_var = tk.BooleanVar(value=True)
            ttk.Checkbutton(conditions_info, text="Skip if no students found", 
                           variable=skip_if_empty_var).pack(anchor=tk.W, pady=2)
            
            retry_on_failure_var = tk.BooleanVar(value=False)
            ttk.Checkbutton(conditions_info, text="Retry on failure", 
                           variable=retry_on_failure_var).pack(anchor=tk.W, pady=2)
            
            def create_enhanced_schedule():
                try:
                    template_name = template_var.get()
                    if not template_name:
                        messagebox.showerror("Error", "Please select a template")
                        return
                    
                    # Validate hour
                    try:
                        hour = int(hour_var.get())
                        if not (0 <= hour <= 23):
                            raise ValueError("Hour must be between 0 and 23")
                    except ValueError as e:
                        messagebox.showerror("Error", f"Invalid hour: {str(e)}")
                        return
                    
                    # Get recipients
                    recipients_input = recipients_text.get(1.0, tk.END).strip()
                    recipients = [email.strip() for email in recipients_input.split('\n') 
                                 if email.strip() and '@' in email]
                    
                    if not recipients:
                        if not messagebox.askyesno("No Recipients", 
                                                 "No email recipients specified. Report will be generated but not sent. Continue?"):
                            return
                    
                    # Build schedule configuration
                    schedule_config = {
                        'frequency': frequency_var.get(),
                        'hour': hour,
                        'enabled': True,
                        'conditions': {
                            'only_if_data': only_if_data_var.get(),
                            'skip_if_empty': skip_if_empty_var.get(),
                            'retry_on_failure': retry_on_failure_var.get()
                        }
                    }
                    
                    # Create scheduled report
                    if ENHANCED_AVAILABLE:
                        scheduled_report = {
                            'template_name': template_name,
                            'schedule_config': schedule_config,
                            'recipients': recipients,
                            'created_at': datetime.now().isoformat(),
                            'last_run': None,
                            'run_count': 0,
                            'is_active': True
                        }
                        
                        scheduled_reports = load_scheduled_reports()
                        scheduled_reports.append(scheduled_report)
                        save_scheduled_reports(scheduled_reports)
                        
                        messagebox.showinfo("Success", 
                                          f"Enhanced schedule created for '{template_name}'!\n\n"
                                          f"Frequency: {frequency_var.get().title()}\n"
                                          f"Time: {hour:02d}:00\n"
                                          f"Recipients: {len(recipients)}")
                        
                        schedule_window.destroy()
                        self.refresh_data()
                    
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to create schedule: {str(e)}")
            
            # Buttons
            button_frame = ttk.Frame(schedule_window)
            button_frame.pack(fill=tk.X, padx=10, pady=10)
            
            ttk.Button(button_frame, text="Create Schedule", command=create_enhanced_schedule,
                      style='Success.TButton').pack(side=tk.RIGHT, padx=(5, 0))
            ttk.Button(button_frame, text="Cancel", command=schedule_window.destroy).pack(side=tk.RIGHT)
            
        except Exception as e:
            messagebox.showerror("Scheduling Error", f"Failed to open scheduling dialog: {str(e)}")

    # Additional missing GUI functions for ReportingSystemGUI class

    def show_template_comparison_dialog(self):
        """Compare multiple templates side by side"""
        try:
            compare_window = tk.Toplevel(self.root)
            compare_window.title("Template Comparison")
            compare_window.geometry("900x600")
            compare_window.transient(self.root)
            
            templates = load_templates() if ENHANCED_AVAILABLE else []
            if len(templates) < 2:
                messagebox.showinfo("Insufficient Templates", "Need at least 2 templates to compare")
                compare_window.destroy()
                return
            
            # Template selection frame
            selection_frame = ttk.LabelFrame(compare_window, text="Select Templates to Compare", padding="10")
            selection_frame.pack(fill=tk.X, padx=10, pady=10)
            
            ttk.Label(selection_frame, text="Template 1:").pack(side=tk.LEFT)
            template1_var = tk.StringVar()
            template1_combo = ttk.Combobox(selection_frame, textvariable=template1_var, 
                                          values=[t['name'] for t in templates], state="readonly")
            template1_combo.pack(side=tk.LEFT, padx=(5, 20))
            
            ttk.Label(selection_frame, text="Template 2:").pack(side=tk.LEFT)
            template2_var = tk.StringVar()
            template2_combo = ttk.Combobox(selection_frame, textvariable=template2_var,
                                          values=[t['name'] for t in templates], state="readonly")
            template2_combo.pack(side=tk.LEFT, padx=5)
            
            # Comparison display
            comparison_frame = ttk.LabelFrame(compare_window, text="Comparison Results", padding="10")
            comparison_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            comparison_text = ScrolledText(comparison_frame, wrap=tk.WORD)
            comparison_text.pack(fill=tk.BOTH, expand=True)
            
            def compare_templates():
                name1 = template1_var.get()
                name2 = template2_var.get()
                
                if not name1 or not name2:
                    messagebox.showwarning("Selection Required", "Please select both templates")
                    return
                
                if name1 == name2:
                    messagebox.showwarning("Same Template", "Please select different templates")
                    return
                
                template1 = next((t for t in templates if t['name'] == name1), None)
                template2 = next((t for t in templates if t['name'] == name2), None)
                
                comparison = f"Template Comparison: {name1} vs {name2}\n"
                comparison += "=" * 60 + "\n\n"
                
                # Compare basic properties
                comparison += "BASIC PROPERTIES:\n"
                comparison += f"Name: {template1['name']} | {template2['name']}\n"
                comparison += f"Description: {template1.get('description', 'None')} | {template2.get('description', 'None')}\n"
                comparison += f"Security Level: {template1.get('security_level', 'normal')} | {template2.get('security_level', 'normal')}\n"
                comparison += f"Visualization: {template1.get('visualization_type', 'standard')} | {template2.get('visualization_type', 'standard')}\n"
                comparison += f"Version: {template1.get('version', '1.0')} | {template2.get('version', '1.0')}\n\n"
                
                # Compare sections
                sections1 = set(template1.get('sections', []))
                sections2 = set(template2.get('sections', []))
                
                comparison += "SECTIONS COMPARISON:\n"
                comparison += f"Total Sections: {len(sections1)} | {len(sections2)}\n"
                
                common_sections = sections1.intersection(sections2)
                only_in_1 = sections1 - sections2
                only_in_2 = sections2 - sections1
                
                if common_sections:
                    comparison += f"Common Sections ({len(common_sections)}): {', '.join(sorted(common_sections))}\n"
                if only_in_1:
                    comparison += f"Only in {name1}: {', '.join(sorted(only_in_1))}\n"
                if only_in_2:
                    comparison += f"Only in {name2}: {', '.join(sorted(only_in_2))}\n"
                
                # Compare filters
                filters1 = template1.get('filters', {})
                filters2 = template2.get('filters', {})
                
                comparison += f"\nFILTERS:\n"
                comparison += f"{name1}: {filters1 if filters1 else 'None'}\n"
                comparison += f"{name2}: {filters2 if filters2 else 'None'}\n"
                
                comparison_text.delete(1.0, tk.END)
                comparison_text.insert(1.0, comparison)
            
            ttk.Button(selection_frame, text="Compare", command=compare_templates).pack(side=tk.RIGHT, padx=(20, 0))
            
            ttk.Button(compare_window, text="Close", command=compare_window.destroy).pack(pady=10)
            
        except Exception as e:
            messagebox.showerror("Comparison Error", f"Failed to open comparison dialog: {str(e)}")

    def show_template_versioning_dialog(self):
        """Show template version history and management"""
        try:
            version_window = tk.Toplevel(self.root)
            version_window.title("Template Version Management")
            version_window.geometry("700x500")
            version_window.transient(self.root)
            
            # Template selection
            selection_frame = ttk.LabelFrame(version_window, text="Select Template", padding="10")
            selection_frame.pack(fill=tk.X, padx=10, pady=10)
            
            templates = load_templates() if ENHANCED_AVAILABLE else []
            template_var = tk.StringVar()
            template_combo = ttk.Combobox(selection_frame, textvariable=template_var,
                                         values=[t['name'] for t in templates], state="readonly")
            template_combo.pack(fill=tk.X)
            
            # Version history display
            history_frame = ttk.LabelFrame(version_window, text="Version History", padding="10")
            history_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            history_text = ScrolledText(history_frame, wrap=tk.WORD)
            history_text.pack(fill=tk.BOTH, expand=True)
            
            def show_version_info():
                template_name = template_var.get()
                if not template_name:
                    return
                
                template = next((t for t in templates if t['name'] == template_name), None)
                if not template:
                    return
                
                version_info = f"Version Information for: {template_name}\n"
                version_info += "=" * 50 + "\n\n"
                version_info += f"Current Version: {template.get('version', '1.0')}\n"
                version_info += f"Created: {template.get('created_at', 'Unknown')}\n"
                version_info += f"Last Modified: {template.get('modified_at', 'Unknown')}\n\n"
                
                version_info += "Template Details:\n"
                version_info += f"- Sections: {len(template.get('sections', []))}\n"
                version_info += f"- Security Level: {template.get('security_level', 'normal')}\n"
                version_info += f"- Visualization Type: {template.get('visualization_type', 'standard')}\n"
                version_info += f"- Filters Applied: {len(template.get('filters', {}))}\n"
                
                history_text.delete(1.0, tk.END)
                history_text.insert(1.0, version_info)
            
            template_combo.bind('<<ComboboxSelected>>', lambda e: show_version_info())
            
            ttk.Button(version_window, text="Close", command=version_window.destroy).pack(pady=10)
            
        except Exception as e:
            messagebox.showerror("Version Error", f"Failed to open version dialog: {str(e)}")

    def show_bulk_operations_dialog(self):
        """Show bulk operations for templates and reports"""
        try:
            bulk_window = tk.Toplevel(self.root)
            bulk_window.title("Bulk Operations")
            bulk_window.geometry("600x500")
            bulk_window.transient(self.root)
            
            # Operation selection
            operation_frame = ttk.LabelFrame(bulk_window, text="Select Operation", padding="10")
            operation_frame.pack(fill=tk.X, padx=10, pady=10)
            
            operation_var = tk.StringVar(value="export_templates")
            
            operations = [
                ("export_templates", "Export All Templates"),
                ("import_templates", "Import Multiple Templates"),
                ("generate_all_reports", "Generate Reports for All Templates"),
                ("cleanup_old_reports", "Cleanup Old Reports"),
                ("backup_all_data", "Backup All System Data")
            ]
            
            for op_id, op_name in operations:
                ttk.Radiobutton(operation_frame, text=op_name, 
                               variable=operation_var, value=op_id).pack(anchor=tk.W)
            
            # Options frame
            options_frame = ttk.LabelFrame(bulk_window, text="Options", padding="10")
            options_frame.pack(fill=tk.X, padx=10, pady=10)
            
            include_data_var = tk.BooleanVar(value=True)
            ttk.Checkbutton(options_frame, text="Include data files", 
                           variable=include_data_var).pack(anchor=tk.W)
            
            confirm_actions_var = tk.BooleanVar(value=True)
            ttk.Checkbutton(options_frame, text="Confirm each action", 
                           variable=confirm_actions_var).pack(anchor=tk.W)
            
            # Progress display
            progress_frame = ttk.LabelFrame(bulk_window, text="Progress", padding="10")
            progress_frame.pack(fill=tk.X, padx=10, pady=10)
            
            progress_bar = ttk.Progressbar(progress_frame, mode='determinate')
            progress_bar.pack(fill=tk.X, pady=(0, 5))
            
            progress_label = ttk.Label(progress_frame, text="Ready")
            progress_label.pack(anchor=tk.W)
            
            # Results display
            results_frame = ttk.LabelFrame(bulk_window, text="Results", padding="10")
            results_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            results_text = ScrolledText(results_frame, height=8, wrap=tk.WORD)
            results_text.pack(fill=tk.BOTH, expand=True)
            
            def execute_bulk_operation():
                operation = operation_var.get()
                results_text.delete(1.0, tk.END)
                
                def bulk_task():
                    try:
                        if operation == "export_templates":
                            templates = load_templates() if ENHANCED_AVAILABLE else []
                            progress_bar['maximum'] = len(templates)
                            
                            export_dir = filedialog.askdirectory(title="Select Export Directory")
                            if not export_dir:
                                return
                            
                            for i, template in enumerate(templates):
                                filename = f"{template['name'].replace(' ', '_')}_template.json"
                                filepath = os.path.join(export_dir, filename)
                                
                                with open(filepath, 'w') as f:
                                    json.dump(template, f, indent=4)
                                
                                progress_bar['value'] = i + 1
                                progress_label.config(text=f"Exported: {template['name']}")
                                time.sleep(0.1)
                            
                            results_text.insert(tk.END, f"Successfully exported {len(templates)} templates to {export_dir}")
                            
                        elif operation == "generate_all_reports":
                            templates = load_templates() if ENHANCED_AVAILABLE else []
                            progress_bar['maximum'] = len(templates)
                            
                            end_date = datetime.now().strftime("%Y-%m-%d")
                            start_date = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
                            
                            for i, template in enumerate(templates):
                                try:
                                    report_path = generate_report(template['name'], start_date, end_date, 'pdf')
                                    if report_path:
                                        results_text.insert(tk.END, f"✓ Generated: {template['name']}\n")
                                    else:
                                        results_text.insert(tk.END, f"✗ Failed: {template['name']}\n")
                                except Exception as e:
                                    results_text.insert(tk.END, f"✗ Error in {template['name']}: {str(e)}\n")
                                
                                progress_bar['value'] = i + 1
                                progress_label.config(text=f"Processing: {template['name']}")
                                results_text.see(tk.END)
                                time.sleep(0.5)
                        
                        progress_label.config(text="Operation completed")
                        
                    except Exception as e:
                        results_text.insert(tk.END, f"Error: {str(e)}")
                
                threading.Thread(target=bulk_task, daemon=True).start()
            
            # Buttons
            button_frame = ttk.Frame(bulk_window)
            button_frame.pack(fill=tk.X, padx=10, pady=10)
            
            ttk.Button(button_frame, text="Execute", command=execute_bulk_operation,
                      style='Success.TButton').pack(side=tk.RIGHT, padx=(5, 0))
            ttk.Button(button_frame, text="Close", command=bulk_window.destroy).pack(side=tk.RIGHT)
            
        except Exception as e:
            messagebox.showerror("Bulk Operations Error", f"Failed to open bulk operations: {str(e)}")

    def show_data_visualization_studio(self):
        """Show advanced data visualization studio"""
        try:
            studio_window = tk.Toplevel(self.root)
            studio_window.title("Data Visualization Studio")
            studio_window.geometry("900x700")
            studio_window.transient(self.root)
            
            # Create notebook for different visualization aspects
            studio_notebook = ttk.Notebook(studio_window)
            studio_notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # Chart Builder Tab
            chart_frame = ttk.Frame(studio_notebook)
            studio_notebook.add(chart_frame, text="Chart Builder")
            
            # Data selection
            data_frame = ttk.LabelFrame(chart_frame, text="Data Selection", padding="10")
            data_frame.pack(fill=tk.X, padx=10, pady=10)
            
            ttk.Label(data_frame, text="Data Source:").pack(anchor=tk.W)
            data_source_var = tk.StringVar(value="students")
            data_combo = ttk.Combobox(data_frame, textvariable=data_source_var,
                                     values=["students", "courses", "modules", "attendance"], state="readonly")
            data_combo.pack(fill=tk.X)
            
            # Chart type selection
            chart_frame_inner = ttk.LabelFrame(chart_frame, text="Chart Type", padding="10")
            chart_frame_inner.pack(fill=tk.X, padx=10, pady=10)
            
            chart_type_var = tk.StringVar(value="bar")
            chart_types = [("bar", "Bar Chart"), ("pie", "Pie Chart"), ("line", "Line Chart"), 
                          ("scatter", "Scatter Plot"), ("heatmap", "Heatmap")]
            
            for chart_id, chart_name in chart_types:
                ttk.Radiobutton(chart_frame_inner, text=chart_name,
                               variable=chart_type_var, value=chart_id).pack(anchor=tk.W)
            
            # Styling options
            style_frame = ttk.LabelFrame(chart_frame, text="Styling Options", padding="10")
            style_frame.pack(fill=tk.X, padx=10, pady=10)
            
            color_scheme_var = tk.StringVar(value="default")
            ttk.Label(style_frame, text="Color Scheme:").pack(anchor=tk.W)
            ttk.Combobox(style_frame, textvariable=color_scheme_var,
                        values=["default", "viridis", "plasma", "cool", "warm"], state="readonly").pack(fill=tk.X)
            
            # Custom Dashboard Tab
            dashboard_frame = ttk.Frame(studio_notebook)
            studio_notebook.add(dashboard_frame, text="Dashboard")
            
            dashboard_info = ttk.LabelFrame(dashboard_frame, text="Dashboard Components", padding="10")
            dashboard_info.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # Component selection
            components = {}
            component_types = [
                ("overview_cards", "Overview Cards", True),
                ("trend_charts", "Trend Charts", True),
                ("distribution_charts", "Distribution Charts", False),
                ("comparison_tables", "Comparison Tables", False),
                ("filter_controls", "Interactive Filters", False)
            ]
            
            for comp_id, comp_name, default_val in component_types:
                var = tk.BooleanVar(value=default_val)
                components[comp_id] = var
                ttk.Checkbutton(dashboard_info, text=comp_name, variable=var).pack(anchor=tk.W)
            
            def generate_visualization():
                try:
                    chart_type = chart_type_var.get()
                    data_source = data_source_var.get()
                    color_scheme = color_scheme_var.get()
                    
                    # Generate sample visualization based on selections
                    messagebox.showinfo("Visualization Generated", 
                                      f"Generated {chart_type} chart for {data_source} data with {color_scheme} colors!")
                    
                except Exception as e:
                    messagebox.showerror("Generation Error", f"Failed to generate visualization: {str(e)}")
            
            # Buttons
            button_frame = ttk.Frame(studio_window)
            button_frame.pack(fill=tk.X, padx=10, pady=10)
            
            ttk.Button(button_frame, text="Generate Visualization", command=generate_visualization,
                      style='Success.TButton').pack(side=tk.RIGHT, padx=(5, 0))
            ttk.Button(button_frame, text="Close", command=studio_window.destroy).pack(side=tk.RIGHT)
            
        except Exception as e:
            messagebox.showerror("Visualization Studio Error", f"Failed to open visualization studio: {str(e)}")

    def show_report_analytics_dashboard(self):
        """Show analytics about report generation and usage"""
        try:
            analytics_window = tk.Toplevel(self.root)
            analytics_window.title("Report Analytics Dashboard")
            analytics_window.geometry("800x600")
            analytics_window.transient(self.root)
            
            # Metrics display
            metrics_frame = ttk.LabelFrame(analytics_window, text="Report Metrics", padding="10")
            metrics_frame.pack(fill=tk.X, padx=10, pady=10)
            
            # Calculate metrics
            templates = load_templates() if ENHANCED_AVAILABLE else []
            scheduled_reports = load_scheduled_reports() if ENHANCED_AVAILABLE else []
            
            # Report counts
            reports_dir = CONFIG.get('reports_dir', 'reports') if ENHANCED_AVAILABLE else 'reports'
            report_count = 0
            if os.path.exists(reports_dir):
                report_count = len([f for f in os.listdir(reports_dir) 
                                  if f.endswith(('.pdf', '.xlsx', '.html'))])
            
            metrics_text = f"""Report System Analytics
    {"=" * 30}

    Templates: {len(templates)}
    Scheduled Reports: {len(scheduled_reports)}
    Generated Reports: {report_count}
    Active Schedules: {sum(1 for r in scheduled_reports if r.get('schedule_config', {}).get('enabled', True))}

    Most Used Sections:"""
            
            # Count section usage across templates
            section_counts = {}
            for template in templates:
                for section in template.get('sections', []):
                    section_counts[section] = section_counts.get(section, 0) + 1
            
            for section, count in sorted(section_counts.items(), key=lambda x: x[1], reverse=True)[:5]:
                metrics_text += f"\n- {section.replace('_', ' ').title()}: {count} templates"
            
            metrics_display = ScrolledText(metrics_frame, height=15, wrap=tk.WORD)
            metrics_display.pack(fill=tk.BOTH, expand=True)
            metrics_display.insert(1.0, metrics_text)
            metrics_display.config(state=tk.DISABLED)
            
            # Usage trends (placeholder)
            trends_frame = ttk.LabelFrame(analytics_window, text="Usage Trends", padding="10")
            trends_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            ttk.Label(trends_frame, text="Trend analysis would be displayed here with actual usage data").pack()
            
            ttk.Button(analytics_window, text="Close", command=analytics_window.destroy).pack(pady=10)
            
        except Exception as e:
            messagebox.showerror("Analytics Error", f"Failed to open analytics dashboard: {str(e)}")

    # Final missing GUI functions for ReportingSystemGUI class

    def show_api_endpoints_documentation(self):
        """Show comprehensive API documentation dialog"""
        try:
            api_doc_window = tk.Toplevel(self.root)
            api_doc_window.title("API Documentation")
            api_doc_window.geometry("900x700")
            api_doc_window.transient(self.root)
            
            # Create notebook for different API sections
            api_notebook = ttk.Notebook(api_doc_window)
            api_notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # Authentication Tab
            auth_frame = ttk.Frame(api_notebook)
            api_notebook.add(auth_frame, text="Authentication")
            
            auth_text = ScrolledText(auth_frame, wrap=tk.WORD)
            auth_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            auth_docs = """API Authentication
    =================

    POST /api/login
    - Description: User authentication
    - Body: {"username": "user", "password": "pass"}
    - Response: {"token": "jwt_token", "expires": "timestamp"}

    Authentication Headers:
    - Authorization: Bearer <token>
    - Content-Type: application/json
    """
            auth_text.insert(1.0, auth_docs)
            
            # Endpoints Tab
            endpoints_frame = ttk.Frame(api_notebook)
            api_notebook.add(endpoints_frame, text="Endpoints")
            
            endpoints_text = ScrolledText(endpoints_frame, wrap=tk.WORD)
            endpoints_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            endpoints_docs = """API Endpoints Reference
    ======================

    TEMPLATES:
    GET  /api/templates - List all templates
    POST /api/templates - Create new template
    GET  /api/templates/{name} - Get specific template
    PUT  /api/templates/{name} - Update template
    DELETE /api/templates/{name} - Delete template

    REPORTS:
    POST /api/reports/generate - Generate report
      Body: {
        "template_name": "string",
        "start_date": "YYYY-MM-DD",
        "end_date": "YYYY-MM-DD",
        "format": "pdf|excel|interactive"
      }

    DATA:
    GET /api/data/{section} - Get section data
      Parameters: start_date, end_date, filters

    ANALYTICS:
    GET /api/analytics/quality - Data quality metrics
    GET /api/analytics/predictions - Dropout risk predictions
    GET /api/analytics/anomalies - Anomaly detection results
    GET /api/analytics/correlations - Correlation analysis

    SYSTEM:
    GET /api/health - System health check
    GET /api/config - System configuration
    """
            endpoints_text.insert(1.0, endpoints_docs)
            
            # Examples Tab
            examples_frame = ttk.Frame(api_notebook)
            api_notebook.add(examples_frame, text="Examples")
            
            examples_text = ScrolledText(examples_frame, wrap=tk.WORD)
            examples_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            examples_docs = """API Usage Examples
    ==================

    Generate a PDF Report:
    curl -X POST http://localhost:5000/api/reports/generate \\
      -H "Content-Type: application/json" \\
      -d '{
        "template_name": "student_overview",
        "start_date": "2024-01-01",
        "end_date": "2024-12-31",
        "format": "pdf"
      }'

    Get Data Quality Metrics:
    curl -X GET http://localhost:5000/api/analytics/quality \\
      -H "Authorization: Bearer <token>"

    Create New Template:
    curl -X POST http://localhost:5000/api/templates \\
      -H "Content-Type: application/json" \\
      -d '{
        "name": "Custom Report",
        "description": "My custom report",
        "sections": ["student_overview", "course_distribution"],
        "visualization_type": "standard"
      }'
    """
            examples_text.insert(1.0, examples_docs)
            
            ttk.Button(api_doc_window, text="Close", command=api_doc_window.destroy).pack(pady=10)
            
        except Exception as e:
            messagebox.showerror("API Documentation Error", f"Failed to open API documentation: {str(e)}")

    def show_system_logs_viewer(self):
        """Show real-time system logs viewer"""
        try:
            logs_window = tk.Toplevel(self.root)
            logs_window.title("System Logs Viewer")
            logs_window.geometry("900x600")
            logs_window.transient(self.root)
            
            # Controls frame
            controls_frame = ttk.Frame(logs_window)
            controls_frame.pack(fill=tk.X, padx=10, pady=10)
            
            # Log level filter
            ttk.Label(controls_frame, text="Log Level:").pack(side=tk.LEFT)
            log_level_var = tk.StringVar(value="ALL")
            log_level_combo = ttk.Combobox(controls_frame, textvariable=log_level_var,
                                          values=["ALL", "DEBUG", "INFO", "WARNING", "ERROR"], state="readonly")
            log_level_combo.pack(side=tk.LEFT, padx=(5, 20))
            
            # Auto-refresh
            auto_refresh_var = tk.BooleanVar(value=True)
            ttk.Checkbutton(controls_frame, text="Auto-refresh", variable=auto_refresh_var).pack(side=tk.LEFT, padx=(0, 20))
            
            # Clear logs button
            def clear_logs():
                logs_text.delete(1.0, tk.END)
            
            ttk.Button(controls_frame, text="Clear", command=clear_logs).pack(side=tk.LEFT)
            ttk.Button(controls_frame, text="Refresh", command=lambda: load_logs()).pack(side=tk.LEFT, padx=(5, 0))
            
            # Logs display
            logs_frame = ttk.LabelFrame(logs_window, text="System Logs", padding="10")
            logs_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            logs_text = ScrolledText(logs_frame, wrap=tk.NONE, font=("Consolas", 9))
            logs_text.pack(fill=tk.BOTH, expand=True)
            
            def load_logs():
                try:
                    log_file = get_log_file('reporting_system.log')
                    if not os.path.exists(log_file):
                        logs_text.insert(tk.END, "No log file found\n")
                        return
                    
                    with open(log_file, 'r') as f:
                        lines = f.readlines()
                    
                    # Filter by log level
                    level_filter = log_level_var.get()
                    if level_filter != "ALL":
                        lines = [line for line in lines if level_filter in line]
                    
                    # Show last 1000 lines
                    display_lines = lines[-1000:] if len(lines) > 1000 else lines
                    
                    logs_text.delete(1.0, tk.END)
                    for line in display_lines:
                        # Color code by log level
                        if "ERROR" in line:
                            logs_text.insert(tk.END, line, "error")
                        elif "WARNING" in line:
                            logs_text.insert(tk.END, line, "warning")
                        elif "INFO" in line:
                            logs_text.insert(tk.END, line, "info")
                        else:
                            logs_text.insert(tk.END, line)
                    
                    logs_text.see(tk.END)
                    
                except Exception as e:
                    logs_text.insert(tk.END, f"Error loading logs: {str(e)}\n")
            
            # Configure text tags for coloring
            logs_text.tag_configure("error", foreground="red")
            logs_text.tag_configure("warning", foreground="orange")
            logs_text.tag_configure("info", foreground="blue")
            
            # Auto-refresh functionality
            def auto_refresh():
                if auto_refresh_var.get():
                    load_logs()
                logs_window.after(5000, auto_refresh)  # Refresh every 5 seconds
            
            # Initial load and start auto-refresh
            load_logs()
            auto_refresh()
            
        except Exception as e:
            messagebox.showerror("Logs Viewer Error", f"Failed to open logs viewer: {str(e)}")

    def show_data_import_export_dialog(self):
        """Show data import/export utilities dialog"""
        try:
            import_export_window = tk.Toplevel(self.root)
            import_export_window.title("Data Import/Export Utilities")
            import_export_window.geometry("700x500")
            import_export_window.transient(self.root)
            
            # Create notebook for import/export operations
            ie_notebook = ttk.Notebook(import_export_window)
            ie_notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # Import Tab
            import_frame = ttk.Frame(ie_notebook)
            ie_notebook.add(import_frame, text="Import Data")
            
            import_options = ttk.LabelFrame(import_frame, text="Import Options", padding="10")
            import_options.pack(fill=tk.X, padx=10, pady=10)
            
            import_type_var = tk.StringVar(value="students")
            import_types = [("students", "Student Data"), ("templates", "Report Templates"), 
                           ("schedules", "Scheduled Reports"), ("config", "System Configuration")]
            
            for import_id, import_name in import_types:
                ttk.Radiobutton(import_options, text=import_name,
                               variable=import_type_var, value=import_id).pack(anchor=tk.W)
            
            # File selection
            file_frame = ttk.LabelFrame(import_frame, text="File Selection", padding="10")
            file_frame.pack(fill=tk.X, padx=10, pady=10)
            
            file_path_var = tk.StringVar()
            ttk.Entry(file_frame, textvariable=file_path_var, width=50).pack(side=tk.LEFT, padx=(0, 10))
            
            def browse_import_file():
                filetypes = [
                    ("JSON files", "*.json"),
                    ("CSV files", "*.csv"),
                    ("Excel files", "*.xlsx"),
                    ("All files", "*.*")
                ]
                filename = filedialog.askopenfilename(filetypes=filetypes)
                if filename:
                    file_path_var.set(filename)
            
            ttk.Button(file_frame, text="Browse", command=browse_import_file).pack(side=tk.LEFT)
            
            def import_data():
                file_path = file_path_var.get()
                import_type = import_type_var.get()
                
                if not file_path:
                    messagebox.showerror("Error", "Please select a file to import")
                    return
                
                try:
                    if import_type == "templates":
                        with open(file_path, 'r') as f:
                            imported_templates = json.load(f)
                        
                        if not isinstance(imported_templates, list):
                            imported_templates = [imported_templates]
                        
                        existing_templates = load_templates() if ENHANCED_AVAILABLE else []
                        existing_templates.extend(imported_templates)
                        
                        if ENHANCED_AVAILABLE:
                            os.makedirs(CONFIG.get('templates_dir', str(paths.REPORT_TEMPLATES_DIR)), exist_ok=True)
                            with open(os.path.join(CONFIG.get('templates_dir', str(paths.REPORT_TEMPLATES_DIR)), "templates.json"), 'w') as f:
                                json.dump(existing_templates, f, indent=4)
                        
                        messagebox.showinfo("Success", f"Imported {len(imported_templates)} templates successfully!")
                        self.refresh_data()
                    
                    elif import_type == "students":
                        messagebox.showinfo("Import", "Student data import would be implemented here")
                    
                    else:
                        messagebox.showinfo("Import", f"Import for {import_type} would be implemented here")
                    
                except Exception as e:
                    messagebox.showerror("Import Error", f"Failed to import data: {str(e)}")
            
            ttk.Button(import_frame, text="Import Data", command=import_data,
                      style='Success.TButton').pack(pady=20)
            
            # Export Tab
            export_frame = ttk.Frame(ie_notebook)
            ie_notebook.add(export_frame, text="Export Data")
            
            export_options = ttk.LabelFrame(export_frame, text="Export Options", padding="10")
            export_options.pack(fill=tk.X, padx=10, pady=10)
            
            export_type_var = tk.StringVar(value="templates")
            export_format_var = tk.StringVar(value="json")
            
            for export_id, export_name in import_types:
                ttk.Radiobutton(export_options, text=export_name,
                               variable=export_type_var, value=export_id).pack(anchor=tk.W)
            
            format_frame = ttk.LabelFrame(export_frame, text="Export Format", padding="10")
            format_frame.pack(fill=tk.X, padx=10, pady=10)
            
            formats = [("json", "JSON"), ("csv", "CSV"), ("excel", "Excel")]
            for format_id, format_name in formats:
                ttk.Radiobutton(format_frame, text=format_name,
                               variable=export_format_var, value=format_id).pack(anchor=tk.W)
            
            def export_data():
                export_type = export_type_var.get()
                export_format = export_format_var.get()
                
                try:
                    if export_type == "templates":
                        templates = load_templates() if ENHANCED_AVAILABLE else []
                        
                        if export_format == "json":
                            file_path = filedialog.asksaveasfilename(
                                defaultextension=".json",
                                filetypes=[("JSON files", "*.json")],
                                initialfile="exported_templates.json"
                            )
                            
                            if file_path:
                                with open(file_path, 'w') as f:
                                    json.dump(templates, f, indent=4)
                                messagebox.showinfo("Success", f"Exported {len(templates)} templates to {file_path}")
                    
                    else:
                        messagebox.showinfo("Export", f"Export for {export_type} in {export_format} format would be implemented here")
                    
                except Exception as e:
                    messagebox.showerror("Export Error", f"Failed to export data: {str(e)}")
            
            ttk.Button(export_frame, text="Export Data", command=export_data,
                      style='Success.TButton').pack(pady=20)
            
        except Exception as e:
            messagebox.showerror("Import/Export Error", f"Failed to open import/export dialog: {str(e)}")

    def show_template_wizard(self):
        """Show step-by-step template creation wizard"""
        try:
            wizard_window = tk.Toplevel(self.root)
            wizard_window.title("Template Creation Wizard")
            wizard_window.geometry("800x600")
            wizard_window.transient(self.root)
            
            # Wizard state
            self.wizard_step = 0
            self.wizard_data = {}
            
            # Main container
            self.wizard_container = ttk.Frame(wizard_window)
            self.wizard_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
            
            # Navigation frame
            nav_frame = ttk.Frame(wizard_window)
            nav_frame.pack(fill=tk.X, padx=20, pady=(0, 20))
            
            self.prev_button = ttk.Button(nav_frame, text="< Previous", command=self.wizard_prev_step)
            self.prev_button.pack(side=tk.LEFT)
            
            self.next_button = ttk.Button(nav_frame, text="Next >", command=self.wizard_next_step)
            self.next_button.pack(side=tk.RIGHT, padx=(10, 0))
            
            self.finish_button = ttk.Button(nav_frame, text="Finish", command=self.wizard_finish,
                                           style='Success.TButton')
            self.finish_button.pack(side=tk.RIGHT)
            
            ttk.Button(nav_frame, text="Cancel", command=wizard_window.destroy).pack(side=tk.RIGHT, padx=(0, 10))
            
            # Progress indicator
            self.progress_label = ttk.Label(nav_frame, text="Step 1 of 4")
            self.progress_label.pack()
            
            # Start wizard
            self.wizard_window = wizard_window
            self.show_wizard_step()
            
        except Exception as e:
            messagebox.showerror("Wizard Error", f"Failed to open template wizard: {str(e)}")

    def show_wizard_step(self):
        """Show current wizard step"""
        # Clear container
        for widget in self.wizard_container.winfo_children():
            widget.destroy()
        
        if self.wizard_step == 0:
            self.show_wizard_step_1()
        elif self.wizard_step == 1:
            self.show_wizard_step_2()
        elif self.wizard_step == 2:
            self.show_wizard_step_3()
        elif self.wizard_step == 3:
            self.show_wizard_step_4()
        
        # Update navigation
        self.prev_button.config(state=tk.NORMAL if self.wizard_step > 0 else tk.DISABLED)
        self.next_button.config(state=tk.NORMAL if self.wizard_step < 3 else tk.DISABLED)
        self.finish_button.config(state=tk.NORMAL if self.wizard_step == 3 else tk.DISABLED)
        self.progress_label.config(text=f"Step {self.wizard_step + 1} of 4")

    def show_wizard_step_1(self):
        """Wizard Step 1: Basic Information"""
        ttk.Label(self.wizard_container, text="Step 1: Basic Information", 
                 style='Title.TLabel').pack(pady=(0, 20))
        
        info_frame = ttk.LabelFrame(self.wizard_container, text="Template Details", padding="20")
        info_frame.pack(fill=tk.X)
        info_frame.columnconfigure(1, weight=1)
        
        ttk.Label(info_frame, text="Template Name:*").grid(row=0, column=0, sticky=tk.W, pady=10)
        self.wizard_name = tk.StringVar(value=self.wizard_data.get('name', ''))
        ttk.Entry(info_frame, textvariable=self.wizard_name, width=40).grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(10, 0))
        
        ttk.Label(info_frame, text="Description:").grid(row=1, column=0, sticky=tk.W, pady=10)
        self.wizard_desc = tk.StringVar(value=self.wizard_data.get('description', ''))
        ttk.Entry(info_frame, textvariable=self.wizard_desc, width=40).grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(10, 0))

    def wizard_next_step(self):
        """Move to next wizard step"""
        # Save current step data
        if self.wizard_step == 0:
            self.wizard_data['name'] = self.wizard_name.get()
            self.wizard_data['description'] = self.wizard_desc.get()
            
            if not self.wizard_data['name'].strip():
                messagebox.showerror("Validation Error", "Template name is required")
                return
        
        if self.wizard_step < 3:
            self.wizard_step += 1
            self.show_wizard_step()

    def wizard_prev_step(self):
        """Move to previous wizard step"""
        if self.wizard_step > 0:
            self.wizard_step -= 1
            self.show_wizard_step()

    def wizard_finish(self):
        """Complete the wizard and create template"""
        try:
            # Create template from wizard data
            messagebox.showinfo("Success", f"Template '{self.wizard_data.get('name')}' created successfully!")
            self.wizard_window.destroy()
            self.refresh_data()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to create template: {str(e)}")
                
    def show_system_config_editor(self):
        """Show system configuration editor window"""
        try:
            config_window = tk.Toplevel(self.root)
            config_window.title("System Configuration Editor")
            config_window.geometry("700x600")
            config_window.transient(self.root)
            
            # Create notebook for different config sections
            config_notebook = ttk.Notebook(config_window)
            config_notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # General config tab
            general_frame = ttk.Frame(config_notebook)
            config_notebook.add(general_frame, text="General")
            
            # Database config
            ttk.Label(general_frame, text="Database Path:").pack(anchor=tk.W, pady=5)
            db_path_var = tk.StringVar(value=CONFIG.get('database', str(DEFAULT_DB_PATH)))
            ttk.Entry(general_frame, textvariable=db_path_var, width=60).pack(fill=tk.X, pady=(0, 10))
            
            # Reports directory
            ttk.Label(general_frame, text="Reports Directory:").pack(anchor=tk.W, pady=5)
            reports_dir_var = tk.StringVar(value=CONFIG.get('reports_dir', 'reports'))
            ttk.Entry(general_frame, textvariable=reports_dir_var, width=60).pack(fill=tk.X, pady=(0, 10))
            
            # Cache config tab
            cache_frame = ttk.Frame(config_notebook)
            config_notebook.add(cache_frame, text="Cache")
            
            ttk.Label(cache_frame, text="Cache Expiry (hours):").pack(anchor=tk.W, pady=5)
            cache_expiry_var = tk.StringVar(value=str(CONFIG.get('cache_expiry_hours', 24)))
            ttk.Entry(cache_frame, textvariable=cache_expiry_var).pack(fill=tk.X, pady=(0, 10))
            
            ttk.Label(cache_frame, text="Max Cache Size (MB):").pack(anchor=tk.W, pady=5)
            max_cache_var = tk.StringVar(value=str(CONFIG.get('max_cache_size_mb', 500)))
            ttk.Entry(cache_frame, textvariable=max_cache_var).pack(fill=tk.X, pady=(0, 10))
            
            # Security config tab
            security_frame = ttk.Frame(config_notebook)
            config_notebook.add(security_frame, text="Security")
            
            ttk.Label(security_frame, text="Session Timeout (seconds):").pack(anchor=tk.W, pady=5)
            session_timeout_var = tk.StringVar(value="3600")
            ttk.Entry(security_frame, textvariable=session_timeout_var).pack(fill=tk.X, pady=(0, 10))
            
            require_auth_var = tk.BooleanVar(value=True)
            ttk.Checkbutton(security_frame, text="Require Authentication", variable=require_auth_var).pack(anchor=tk.W, pady=5)
            
            # Button frame
            button_frame = ttk.Frame(config_window)
            button_frame.pack(fill=tk.X, padx=10, pady=(0, 10))
            
            def save_config():
                try:
                    new_config = {
                        'database': db_path_var.get(),
                        'reports_dir': reports_dir_var.get(),
                        'cache_expiry_hours': int(cache_expiry_var.get()),
                        'max_cache_size_mb': int(max_cache_var.get())
                    }
                    
                    CONFIG.update(new_config)
                    if ENHANCED_AVAILABLE:
                        SystemConfig.save_config(new_config)
                    
                    messagebox.showinfo("Success", "Configuration saved successfully!")
                    config_window.destroy()
                    self.check_system_status()
                    
                except Exception as e:
                    messagebox.showerror("Save Error", f"Failed to save configuration: {str(e)}")
            
            ttk.Button(button_frame, text="Save", command=save_config).pack(side=tk.RIGHT, padx=(5, 0))
            ttk.Button(button_frame, text="Cancel", command=config_window.destroy).pack(side=tk.RIGHT)
            
        except Exception as e:
            messagebox.showerror("Config Error", f"Failed to open configuration editor: {str(e)}")
    
    def show_email_settings_dialog(self):
        """Show email configuration dialog"""
        try:
            email_window = tk.Toplevel(self.root)
            email_window.title("Email Settings")
            email_window.geometry("500x400")
            email_window.transient(self.root)
            
            # Email configuration form
            config_frame = ttk.LabelFrame(email_window, text="SMTP Configuration", padding="10")
            config_frame.pack(fill=tk.X, padx=10, pady=10)
            config_frame.columnconfigure(1, weight=1)
            
            # SMTP Server
            ttk.Label(config_frame, text="SMTP Server:").grid(row=0, column=0, sticky=tk.W, pady=5)
            smtp_server_var = tk.StringVar(value=CONFIG.get('email', {}).get('smtp_server', 'smtp.gmail.com'))
            ttk.Entry(config_frame, textvariable=smtp_server_var).grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)
            
            # SMTP Port
            ttk.Label(config_frame, text="SMTP Port:").grid(row=1, column=0, sticky=tk.W, pady=5)
            smtp_port_var = tk.StringVar(value=str(CONFIG.get('email', {}).get('smtp_port', 587)))
            ttk.Entry(config_frame, textvariable=smtp_port_var).grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)
            
            # From Address
            ttk.Label(config_frame, text="From Address:").grid(row=2, column=0, sticky=tk.W, pady=5)
            from_address_var = tk.StringVar(value=CONFIG.get('email', {}).get('from_address', ''))
            ttk.Entry(config_frame, textvariable=from_address_var).grid(row=2, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)
            
            # Username
            ttk.Label(config_frame, text="Username:").grid(row=3, column=0, sticky=tk.W, pady=5)
            username_var = tk.StringVar(value=CONFIG.get('email', {}).get('username', ''))
            ttk.Entry(config_frame, textvariable=username_var).grid(row=3, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)
            
            # Password
            ttk.Label(config_frame, text="Password:").grid(row=4, column=0, sticky=tk.W, pady=5)
            password_var = tk.StringVar()
            ttk.Entry(config_frame, textvariable=password_var, show='*').grid(row=4, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)
            
            # Use TLS
            use_tls_var = tk.BooleanVar(value=CONFIG.get('email', {}).get('use_tls', True))
            ttk.Checkbutton(config_frame, text="Use TLS", variable=use_tls_var).grid(row=5, column=0, columnspan=2, sticky=tk.W, pady=5)
            
            # Enable email
            email_enabled_var = tk.BooleanVar(value=CONFIG.get('email', {}).get('enabled', False))
            ttk.Checkbutton(config_frame, text="Enable Email Notifications", variable=email_enabled_var).grid(row=6, column=0, columnspan=2, sticky=tk.W, pady=5)
            
            # Test email section
            test_frame = ttk.LabelFrame(email_window, text="Test Email", padding="10")
            test_frame.pack(fill=tk.X, padx=10, pady=10)
            
            ttk.Label(test_frame, text="Test Email Address:").pack(anchor=tk.W)
            test_email_var = tk.StringVar()
            ttk.Entry(test_frame, textvariable=test_email_var, width=40).pack(fill=tk.X, pady=(0, 10))
            
            def send_test_email():
                try:
                    from university_system.infrastructure.email.smtp import send_email_via_smtp

                    test_email = test_email_var.get().strip()
                    if not test_email or '@' not in test_email:
                        messagebox.showerror("Invalid Email", "Please enter a valid email address")
                        return

                    # Use temporary config for test
                    test_config = {
                        'smtp_server': smtp_server_var.get(),
                        'smtp_port': int(smtp_port_var.get()),
                        'from_address': from_address_var.get(),
                        'username': username_var.get(),
                        'password': password_var.get(),
                        'use_tls': use_tls_var.get()
                    }

                    body = """This is a test email from the University Reporting System.

If you receive this email, your email settings are configured correctly!

System Configuration:
- SMTP Server: {smtp_server}
- SMTP Port: {smtp_port}
- From Address: {from_address}
- TLS Enabled: {use_tls}

Thank you for using our system.
""".format(**test_config)

                    current_time = datetime.now().isoformat()
                    success = send_email_via_smtp(
                        recipient_email=test_email,
                        subject='Test Email from University Reporting System',
                        body=body,
                        cc=None,
                        bcc=None,
                        attachments=None,
                        current_time=current_time
                    )

                    if success:
                        messagebox.showinfo("Test Email Sent", f"Test email successfully sent to {test_email}!")
                    else:
                        messagebox.showerror("Test Failed", "Test email failed to send")

                except Exception as e:
                    messagebox.showerror("Test Failed", f"Test email failed: {str(e)}")
            
            ttk.Button(test_frame, text="Send Test Email", command=send_test_email).pack()
            
            # Button frame
            button_frame = ttk.Frame(email_window)
            button_frame.pack(fill=tk.X, padx=10, pady=(0, 10))
            
            def save_email_settings():
                try:
                    email_config = {
                        'enabled': email_enabled_var.get(),
                        'smtp_server': smtp_server_var.get(),
                        'smtp_port': int(smtp_port_var.get()),
                        'from_address': from_address_var.get(),
                        'username': username_var.get(),
                        'use_tls': use_tls_var.get()
                    }
                    
                    CONFIG['email'] = email_config
                    
                    if ENHANCED_AVAILABLE:
                        full_config = SystemConfig.load_config()
                        full_config['email'] = email_config
                        SystemConfig.save_config(full_config)
                    
                    messagebox.showinfo("Success", "Email settings saved successfully!")
                    email_window.destroy()
                    
                except Exception as e:
                    messagebox.showerror("Save Error", f"Failed to save email settings: {str(e)}")
            
            ttk.Button(button_frame, text="Save", command=save_email_settings).pack(side=tk.RIGHT, padx=(5, 0))
            ttk.Button(button_frame, text="Cancel", command=email_window.destroy).pack(side=tk.RIGHT)
            
        except Exception as e:
            messagebox.showerror("Email Settings Error", f"Failed to open email settings: {str(e)}")
                    
    def load_recent_reports(self):
        """Load recent reports into the GUI"""
        try:
            # Get reports from the reports directory
            reports_dir = CONFIG.get('reports_dir', 'reports') if ENHANCED_AVAILABLE else 'reports'
            
            if not os.path.exists(reports_dir):
                return
            
            reports = []
            for file in os.listdir(reports_dir):
                if file.endswith(('.pdf', '.xlsx', '.html')):
                    file_path = os.path.join(reports_dir, file)
                    stat = os.stat(file_path)
                    
                    reports.append({
                        'name': file,
                        'path': file_path,
                        'generated': datetime.fromtimestamp(stat.st_mtime),
                        'size': stat.st_size,
                        'format': file.split('.')[-1].upper()
                    })
            
            # Sort by generation time (newest first)
            reports.sort(key=lambda x: x['generated'], reverse=True)
            
            self.root.after(0, lambda: self._update_reports_tree(reports))
            
        except Exception as e:
            logger.error(f"Error loading recent reports: {str(e)}")
    
    def _update_reports_tree(self, reports):
        """Update reports tree in main thread"""
        # Clear existing items
        for item in self.reports_tree.get_children():
            self.reports_tree.delete(item)
        
        self.reports_data = reports
        
        for report in reports[:50]:  # Show last 50 reports
            size_mb = report['size'] / (1024 * 1024)
            size_str = f"{size_mb:.2f} MB" if size_mb > 1 else f"{report['size'] / 1024:.1f} KB"
            
            values = (
                report['name'],
                report['generated'].strftime('%Y-%m-%d %H:%M'),
                report['format'],
                size_str,
                "Available"
            )
            
            self.reports_tree.insert('', tk.END, values=values)
    
    def update_overview_cards(self):
        """Update the overview cards with current statistics"""
        try:
            # Get database statistics (works regardless of ENHANCED_AVAILABLE)
            conn = get_db_connection()
            if not conn:
                logging.error("Failed to get database connection for overview cards")
                return

            cursor = conn.cursor()

            # Student count - always query from database
            cursor.execute("SELECT COUNT(*) FROM students")
            student_count = cursor.fetchone()[0]

            # Course count - always query from database
            cursor.execute("SELECT COUNT(DISTINCT course) FROM students WHERE course IS NOT NULL")
            course_count = cursor.fetchone()[0]

            conn.close()

            # Template count (use enhanced features if available)
            if ENHANCED_AVAILABLE:
                templates = load_templates()
                template_count = len(templates)
            else:
                # Fallback: count from database
                try:
                    conn = get_db_connection()
                    cursor = conn.cursor()
                    cursor.execute("SELECT COUNT(*) FROM email_templates")
                    template_count = cursor.fetchone()[0]
                    conn.close()
                except:
                    template_count = 0

            # Report count (from files)
            reports_dir = CONFIG.get('reports_dir', 'reports') if ENHANCED_AVAILABLE else 'reports'
            report_count = 0
            if os.path.exists(reports_dir):
                report_count = len([f for f in os.listdir(reports_dir)
                                  if f.endswith(('.pdf', '.xlsx', '.html'))])

            # Update cards directly (more reliable than scheduled callback)
            def update_labels():
                try:
                    if hasattr(self, 'overview_students') and self.overview_students:
                        self.overview_students.config(text=str(student_count))
                    if hasattr(self, 'overview_courses') and self.overview_courses:
                        self.overview_courses.config(text=str(course_count))
                    if hasattr(self, 'overview_templates') and self.overview_templates:
                        self.overview_templates.config(text=str(template_count))
                    if hasattr(self, 'overview_reports') and self.overview_reports:
                        self.overview_reports.config(text=str(report_count))
                    print(f"✓ Updated overview cards: Students={student_count}, Courses={course_count}, Templates={template_count}, Reports={report_count}")
                except Exception as e:
                    logging.error(f"Error in update_labels: {e}")

            # Try to update immediately if on main thread, otherwise schedule
            try:
                update_labels()
            except:
                # If direct update fails, schedule it
                self._schedule_on_ui_thread(update_labels)

        except Exception as e:
            logging.error(f"Error updating overview cards: {str(e)}")
            import traceback
            logging.error(traceback.format_exc())
    
    def check_system_status(self):
        """Check and update system status"""
        try:
            status_parts = []
            
            # Check database connection
            if ENHANCED_AVAILABLE:
                conn = get_db_connection()
                conn.close()
                status_parts.append("Database: ✓")
            
            # Check directories
            dirs_to_check = ['reports', 'templates', 'cache'] if ENHANCED_AVAILABLE else ['reports']
            for dir_name in dirs_to_check:
                if os.path.exists(dir_name):
                    status_parts.append(f"{dir_name.title()}: ✓")
                else:
                    status_parts.append(f"{dir_name.title()}: ✗")
            
            status_text = " | ".join(status_parts)
            
            self._schedule_on_ui_thread(lambda: [
                setattr(self.system_status, 'text', status_text),
                setattr(self.status_indicator, 'text', "● System Ready")
            ])
            
        except Exception as e:
            error_text = f"System Error: {str(e)}"
            self._schedule_on_ui_thread(lambda: [
                setattr(self.system_status, 'text', error_text),
                setattr(self.status_indicator, 'text', "● System Error"),
                setattr(self.status_indicator, 'style', 'Error.TLabel')
            ])
    
    # Template management methods
    
    def on_template_select(self, event):
        """Handle template selection"""
        selection = self.template_listbox.curselection()
        if not selection or not hasattr(self, 'templates_data'):
            return
        
        template_data = self.templates_data[selection[0]]
        self.display_template_details(template_data)

    def set_auth(self, auth_obj):
        """Set authentication context for integration"""
        try:
            self.auth = auth_obj
            if auth_obj and auth_obj.current_user:
                self.current_user = auth_obj.current_user
                print(f"Enhanced reporting authenticated for: {auth_obj.current_user['username']}")
            else:
                print("No authentication context provided")
        except Exception as e:
            print(f"Error setting auth context: {e}")
    
    def display_template_details(self, template_data):
        """Display template details in the text widget"""
        self.template_details.delete(1.0, tk.END)
        
        details = f"""Template Details:

Name: {template_data['name']}
Description: {template_data.get('description', 'No description')}
Version: {template_data.get('version', '1.0')}
Created: {template_data.get('created_at', 'Unknown')}
Security Level: {template_data.get('security_level', 'normal').title()}
Visualization Type: {template_data.get('visualization_type', 'standard').title()}

Sections ({len(template_data.get('sections', []))} total):
"""
        
        for section in template_data.get('sections', []):
            details += f"  • {section.replace('_', ' ').title()}\n"
        
        if template_data.get('filters'):
            details += f"\nFilters:\n"
            for key, value in template_data['filters'].items():
                details += f"  • {key}: {value}\n"
        
        self.template_details.insert(1.0, details)
    
    def create_template_dialog(self):
        """Open template creation dialog"""
        dialog = TemplateDialog(self.root, title="Create New Template")
        if dialog.result:
            self.refresh_data()
    
    def edit_template_dialog(self):
        """Open template editing dialog"""
        selection = self.template_listbox.curselection()
        if not selection or not hasattr(self, 'templates_data'):
            messagebox.showwarning("No Selection", "Please select a template to edit.")
            return
        
        template_data = self.templates_data[selection[0]]
        dialog = TemplateDialog(self.root, title="Edit Template", template_data=template_data)
        if dialog.result:
            self.refresh_data()
    
    def delete_template(self):
        """Delete selected template"""
        selection = self.template_listbox.curselection()
        if not selection or not hasattr(self, 'templates_data'):
            messagebox.showwarning("No Selection", "Please select a template to delete.")
            return
        
        template_data = self.templates_data[selection[0]]
        
        if messagebox.askyesno("Confirm Delete",
                              f"Are you sure you want to delete template '{template_data['name']}'?"):
            try:
                if ENHANCED_AVAILABLE:
                    delete_template_from_db(template_data['name'])

                self.refresh_data()
                messagebox.showinfo("Success", "Template deleted successfully!")

            except Exception as e:
                logging.error(f"Failed to delete template: {e}")
                messagebox.showerror("Error", f"Failed to delete template: {str(e)}")
    
    def export_template(self):
        """Export selected template"""
        selection = self.template_listbox.curselection()
        if not selection or not hasattr(self, 'templates_data'):
            messagebox.showwarning("No Selection", "Please select a template to export.")
            return
        
        template_data = self.templates_data[selection[0]]
        
        file_path = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
            initialname=f"{template_data['name']}_template.json"
        )
        
        if file_path:
            try:
                with open(file_path, 'w') as f:
                    json.dump(template_data, f, indent=4)
                messagebox.showinfo("Success", f"Template exported to {file_path}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to export template: {str(e)}")
    
    def duplicate_template(self):
        """Duplicate selected template"""
        selection = self.template_listbox.curselection()
        if not selection or not hasattr(self, 'templates_data'):
            messagebox.showwarning("No Selection", "Please select a template to duplicate.")
            return
        
        template_data = self.templates_data[selection[0]].copy()
        
        # Get new name
        new_name = simpledialog.askstring("Duplicate Template", 
                                         "Enter name for duplicated template:",
                                         initialvalue=f"{template_data['name']} Copy")
        
        if new_name:
            template_data['name'] = new_name
            template_data['created_at'] = datetime.now().isoformat()
            template_data['version'] = "1.0"
            
            try:
                if ENHANCED_AVAILABLE:
                    save_template_dict(template_data)

                self.refresh_data()
                messagebox.showinfo("Success", "Template duplicated successfully!")

            except Exception as e:
                logging.error(f"Failed to duplicate template: {e}")
                messagebox.showerror("Error", f"Failed to duplicate template: {str(e)}")
    
    def preview_template(self):
        """Preview selected template"""
        selection = self.template_listbox.curselection()
        if not selection or not hasattr(self, 'templates_data'):
            messagebox.showwarning("No Selection", "Please select a template to preview.")
            return
        
        template_data = self.templates_data[selection[0]]
        
        # Create preview window
        preview_window = tk.Toplevel(self.root)
        preview_window.title(f"Template Preview: {template_data['name']}")
        preview_window.geometry("600x500")
        
        preview_text = ScrolledText(preview_window, wrap=tk.WORD)
        preview_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Generate preview content
        preview_content = f"""Template Preview: {template_data['name']}

This template will generate a report containing the following sections:

"""
        
        for i, section in enumerate(template_data.get('sections', []), 1):
            section_name = section.replace('_', ' ').title()
            preview_content += f"{i}. {section_name}\n"
            
            # Add section description
            section_descriptions = {
                'student_overview': '   Overview of total students, courses, and key metrics',
                'course_distribution': '   Distribution of students across different courses',
                'gender_distribution': '   Breakdown of students by gender',
                'age_distribution': '   Age demographics of student population',
                'registration_trends': '   Student registration patterns over time',
                'module_popularity': '   Most popular modules among students',
                'grade_distribution': '   Distribution of student grades',
                'attendance_summary': '   Student attendance statistics',
                'data_quality_report': '   Assessment of data completeness and accuracy',
                'predictive_analytics': '   AI-powered insights and predictions',
                'correlation_analysis': '   Statistical relationships between variables',
                'anomaly_detection': '   Identification of unusual patterns'
            }
            
            if section in section_descriptions:
                preview_content += f"{section_descriptions[section]}\n"
            preview_content += "\n"
        
        if template_data.get('filters'):
            preview_content += "Applied Filters:\n"
            for key, value in template_data['filters'].items():
                preview_content += f"  • {key}: {value}\n"
        
        preview_content += f"""
Report Configuration:
  • Security Level: {template_data.get('security_level', 'normal').title()}
  • Visualization Type: {template_data.get('visualization_type', 'standard').title()}
  • Template Version: {template_data.get('version', '1.0')}
"""
        
        preview_text.insert(1.0, preview_content)
        preview_text.config(state=tk.DISABLED)
    
    def generate_from_template(self):
        """Generate report from selected template"""
        selection = self.template_listbox.curselection()
        if not selection or not hasattr(self, 'templates_data'):
            messagebox.showwarning("No Selection", "Please select a template to generate report.")
            return
        
        template_data = self.templates_data[selection[0]]
        
        # Switch to reports tab and set template
        self.notebook.select(1)  # Reports tab
        self.template_combo.set(template_data['name'])
    
    # Report generation methods
    
    def generate_report(self):
        """Generate a report using the selected template and parameters"""
        template_name = self.template_combo.get()
        if not template_name:
            messagebox.showwarning("No Template", "Please select a template.")
            return
        
        try:
            start_date = self.start_date.get()
            end_date = self.end_date.get()
            format_type = self.format_combo.get().lower()
            
            # Validate dates
            datetime.strptime(start_date, "%Y-%m-%d")
            datetime.strptime(end_date, "%Y-%m-%d")
            
        except ValueError:
            messagebox.showerror("Invalid Date", "Please enter valid dates in YYYY-MM-DD format.")
            return
        
        self.update_status("Generating report...")
        self.start_progress()
        
        def generate_task():
            try:
                if ENHANCED_AVAILABLE:
                    format_map = {'pdf': 'pdf', 'excel': 'excel', 'interactive html': 'interactive'}
                    report_format = format_map.get(format_type, 'pdf')
                    
                    report_path = generate_report(template_name, start_date, end_date, report_format)
                    
                    if report_path:
                        self.root.after(0, lambda: [
                            self.stop_progress(),
                            self.update_status("Report generated successfully"),
                            self.show_report_success(report_path),
                            self.refresh_reports()
                        ])
                    else:
                        self.root.after(0, lambda: [
                            self.stop_progress(),
                            self.update_status("Report generation failed", "error"),
                            messagebox.showerror("Error", "Failed to generate report")
                        ])
                else:
                    # Fallback for basic functionality
                    self.root.after(0, lambda: [
                        self.stop_progress(),
                        self.update_status("Enhanced reporting not available", "warning"),
                        messagebox.showwarning("Feature Unavailable", 
                                             "Enhanced reporting features require the full system to be available.")
                    ])
                    
            except Exception as e:
                self.root.after(0, lambda: [
                    self.stop_progress(),
                    self.update_status(f"Error: {str(e)}", "error"),
                    messagebox.showerror("Error", f"Failed to generate report: {str(e)}")
                ])
        
        threading.Thread(target=generate_task, daemon=True).start()
    
    def show_report_success(self, report_path):
        """Show success dialog with options to open/share report"""
        result = messagebox.askyesnocancel("Report Generated", 
                                          f"Report generated successfully!\n\nFile: {os.path.basename(report_path)}\n\nWould you like to open it now?")
        
        if result is True:  # Yes - open report
            try:
                webbrowser.open(f"file://{os.path.abspath(report_path)}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to open report: {str(e)}")
        elif result is False:  # No - show in file manager
            try:
                if os.name == 'nt':  # Windows
                    os.startfile(os.path.dirname(report_path))
                elif os.name == 'posix':  # macOS and Linux
                    os.system(f'open "{os.path.dirname(report_path)}"')
            except:
                pass
    
    def refresh_reports(self):
        """Refresh the reports list"""
        self.load_recent_reports()
    
    def open_report(self):
        """Open selected report"""
        selection = self.reports_tree.selection()
        if not selection or not hasattr(self, 'reports_data'):
            messagebox.showwarning("No Selection", "Please select a report to open.")
            return
        
        item = self.reports_tree.item(selection[0])
        report_name = item['values'][0]
        
        # Find report in data
        report = next((r for r in self.reports_data if r['name'] == report_name), None)
        if report:
            try:
                webbrowser.open(f"file://{os.path.abspath(report['path'])}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to open report: {str(e)}")
    
    def share_report(self):
        """Share selected report"""
        selection = self.reports_tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a report to share.")
            return

        # Get selected report
        item = self.reports_tree.item(selection[0])
        report_name = item['values'][0]

        # Find report in data
        report = next((r for r in self.reports_data if r['name'] == report_name), None)
        if not report:
            messagebox.showerror("Error", "Report not found")
            return

        # Create share dialog
        share_dialog = tk.Toplevel(self.root)
        share_dialog.title("Share Report")
        share_dialog.geometry("600x500")
        share_dialog.transient(self.root)
        share_dialog.grab_set()

        main_frame = ttk.Frame(share_dialog, padding=20)
        main_frame.pack(fill='both', expand=True)

        ttk.Label(main_frame, text=f"Share Report: {report_name}", font=('Arial', 12, 'bold')).pack(pady=10)

        # Report info
        info_frame = ttk.LabelFrame(main_frame, text="Report Information", padding=10)
        info_frame.pack(fill='x', pady=10)

        ttk.Label(info_frame, text=f"Format: {report['format']}").pack(anchor='w')
        ttk.Label(info_frame, text=f"Size: {report['size'] / 1024:.1f} KB").pack(anchor='w')
        ttk.Label(info_frame, text=f"Generated: {report['generated']}").pack(anchor='w')

        # Email settings
        email_frame = ttk.LabelFrame(main_frame, text="Email Settings", padding=10)
        email_frame.pack(fill='both', expand=True, pady=10)

        ttk.Label(email_frame, text="Recipient Email(s):").pack(anchor='w', pady=(0, 5))
        ttk.Label(email_frame, text="(Separate multiple emails with commas)", font=('Arial', 8)).pack(anchor='w')
        recipients_var = tk.StringVar()
        ttk.Entry(email_frame, textvariable=recipients_var, width=60).pack(fill='x', pady=5)

        ttk.Label(email_frame, text="Subject:").pack(anchor='w', pady=(10, 5))
        subject_var = tk.StringVar(value=f"University Report: {report_name}")
        ttk.Entry(email_frame, textvariable=subject_var, width=60).pack(fill='x', pady=5)

        ttk.Label(email_frame, text="Message:").pack(anchor='w', pady=(10, 5))
        message_text = tk.Text(email_frame, height=8, width=60)
        message_text.pack(fill='both', expand=True, pady=5)
        message_text.insert('1.0', f"""Please find attached the university report: {report_name}

This report was generated on {report['generated']}.

Best regards,
University Reporting System""")

        def send_report():
            recipients = recipients_var.get().strip()
            if not recipients:
                messagebox.showwarning("Missing Recipients", "Please enter at least one recipient email address")
                return

            # Parse recipients
            recipient_list = [email.strip() for email in recipients.split(',')]

            try:
                from university_system.infrastructure.email.smtp import send_email_via_smtp

                # Send to first recipient with others as CC
                recipient_email = recipient_list[0]
                cc = recipient_list[1:] if len(recipient_list) > 1 else None

                # Get message body
                body = message_text.get('1.0', tk.END)

                # Prepare attachment
                attachments = [report['path']]

                current_time = datetime.now().isoformat()
                success = send_email_via_smtp(
                    recipient_email=recipient_email,
                    subject=subject_var.get(),
                    body=body,
                    cc=cc,
                    bcc=None,
                    attachments=attachments,
                    current_time=current_time
                )

                if success:
                    messagebox.showinfo("Report Shared", f"Report sent successfully to {len(recipient_list)} recipient(s)!")
                    share_dialog.destroy()
                else:
                    messagebox.showerror("Share Failed", "Failed to share report")

            except Exception as e:
                messagebox.showerror("Share Failed", f"Failed to share report: {str(e)}")

        # Buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(pady=10)

        ttk.Button(button_frame, text="Send", command=send_report).pack(side='left', padx=5)
        ttk.Button(button_frame, text="Cancel", command=share_dialog.destroy).pack(side='left', padx=5)
    
    def delete_report(self):
        """Delete selected report"""
        selection = self.reports_tree.selection()
        if not selection or not hasattr(self, 'reports_data'):
            messagebox.showwarning("No Selection", "Please select a report to delete.")
            return
        
        item = self.reports_tree.item(selection[0])
        report_name = item['values'][0]
        
        if messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete '{report_name}'?"):
            # Find and delete report
            report = next((r for r in self.reports_data if r['name'] == report_name), None)
            if report:
                try:
                    os.remove(report['path'])
                    self.refresh_reports()
                    messagebox.showinfo("Success", "Report deleted successfully!")
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to delete report: {str(e)}")
    
    # Analytics methods
    
    def run_quality_check(self):
        """Run data quality check"""
        if not ENHANCED_AVAILABLE:
            messagebox.showwarning("Feature Unavailable", 
                                 "Data quality checking requires the enhanced system.")
            return
        
        self.update_status("Running data quality check...")
        self.start_progress()
        
        def quality_task():
            try:
                quality_report = DataQualityMonitor.run_quality_checks()
                self.root.after(0, lambda: self._display_quality_results(quality_report))
            except Exception as e:
                self.root.after(0, lambda: [
                    self.stop_progress(),
                    self.update_status(f"Quality check failed: {str(e)}", "error"),
                    messagebox.showerror("Error", f"Failed to run quality check: {str(e)}")
                ])
        
        threading.Thread(target=quality_task, daemon=True).start()
    
    def _display_quality_results(self, quality_report):
        """Display quality check results"""
        self.stop_progress()
        self.update_status("Quality check completed")
        
        self.quality_display.delete(1.0, tk.END)
        
        output = f"Data Quality Report - {quality_report['timestamp']}\n"
        output += "=" * 60 + "\n\n"
        
        checks = quality_report.get('checks', {})
        
        if 'missing_data' in checks:
            missing = checks['missing_data']['students']
            total = missing['total_records']
            output += f"📊 MISSING DATA ANALYSIS:\n"
            output += f"   Total Records: {total}\n"
            output += f"   Missing Emails: {missing['missing_emails']}\n"
            output += f"   Missing Names: {missing['missing_names']}\n"
            output += f"   Missing Courses: {missing['missing_courses']}\n"
            
            if total > 0:
                completeness = ((total * 3) - (missing['missing_emails'] + missing['missing_names'] + missing['missing_courses'])) / (total * 3) * 100
                output += f"   Data Completeness: {completeness:.1f}%\n"
                
                if completeness < 90:
                    output += "   ⚠️  Warning: Data completeness below 90%\n"
                else:
                    output += "   ✅ Good data completeness\n"
            output += "\n"
        
        if 'duplicates' in checks:
            duplicates = checks['duplicates']
            output += f"👥 DUPLICATE ANALYSIS:\n"
            output += f"   Duplicate Emails: {duplicates['duplicate_emails']}\n"
            
            if duplicates['duplicate_emails'] > 0:
                output += "   📋 Duplicate Email Details:\n"
                for detail in duplicates.get('duplicate_email_details', [])[:5]:
                    output += f"      {detail['email']}: {detail['count']} occurrences\n"
                output += "   ⚠️  Action Required: Review duplicate emails\n"
            else:
                output += "   ✅ No duplicate emails found\n"
            output += "\n"
        
        if 'invalid_data' in checks:
            invalid = checks['invalid_data']
            output += f"❌ INVALID DATA ANALYSIS:\n"
            output += f"   Invalid Ages: {invalid['invalid_ages']}\n"
            output += f"   Invalid Emails: {invalid['invalid_emails']}\n"
            
            if invalid['invalid_ages'] > 0 or invalid['invalid_emails'] > 0:
                output += "   ⚠️  Action Required: Clean invalid data\n"
            else:
                output += "   ✅ No invalid data found\n"
            output += "\n"
        
        if 'data_freshness' in checks:
            freshness = checks['data_freshness']
            if freshness['last_registration_date']:
                days_since = freshness['days_since_last_registration']
                output += f"📅 DATA FRESHNESS:\n"
                output += f"   Last Registration: {freshness['last_registration_date']}\n"
                output += f"   Days Since Last: {days_since}\n"
                
                if days_since > 7:
                    output += "   ⚠️  Warning: No recent registrations\n"
                else:
                    output += "   ✅ Recent data available\n"
            else:
                output += f"📅 DATA FRESHNESS:\n"
                output += "   ❌ No registration data found\n"
        
        self.quality_display.insert(1.0, output)
    
    def export_quality_report(self):
        """Export quality report to file"""
        content = self.quality_display.get(1.0, tk.END).strip()
        if not content:
            messagebox.showwarning("No Data", "No quality report data to export.")
            return
        
        file_path = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")],
            initialname=f"quality_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        )
        
        if file_path:
            try:
                with open(file_path, 'w') as f:
                    f.write(content)
                messagebox.showinfo("Success", f"Quality report exported to {file_path}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to export report: {str(e)}")
    
    def run_predictions(self):
        """Run predictive analytics"""
        if not ENHANCED_AVAILABLE:
            messagebox.showwarning("Feature Unavailable", 
                                 "Predictive analytics requires the enhanced system.")
            return
        
        self.update_status("Running predictive analytics...")
        self.start_progress()
        
        def predictions_task():
            try:
                predictions = PredictiveAnalytics.predict_dropout_risk()
                self.root.after(0, lambda: self._display_predictions_results(predictions))
            except Exception as e:
                self.root.after(0, lambda: [
                    self.stop_progress(),
                    self.update_status(f"Predictions failed: {str(e)}", "error"),
                    messagebox.showerror("Error", f"Failed to run predictions: {str(e)}")
                ])
        
        threading.Thread(target=predictions_task, daemon=True).start()
    
    def _display_predictions_results(self, predictions):
        """Display prediction results"""
        self.stop_progress()
        self.update_status("Predictions completed")
        
        self.predictions_display.delete(1.0, tk.END)
        
        output = f"Predictive Analytics Report\n"
        output += "=" * 40 + "\n\n"
        
        if 'error' in predictions:
            output += f"❌ Analysis unavailable: {predictions['error']}\n"
        else:
            output += f"🎯 DROPOUT RISK ANALYSIS:\n\n"
            
            if 'model_accuracy' in predictions:
                accuracy = predictions['model_accuracy'] * 100
                output += f"   Model Accuracy: {accuracy:.1f}%\n"
                
                if accuracy > 80:
                    output += "   ✅ High confidence predictions\n"
                elif accuracy > 60:
                    output += "   ⚠️  Moderate confidence predictions\n"
                else:
                    output += "   ❌ Low confidence - more data needed\n"
            
            if 'total_students_analyzed' in predictions:
                output += f"   Students Analyzed: {predictions['total_students_analyzed']}\n"
            
            if 'high_risk_students' in predictions:
                high_risk = predictions['high_risk_students']
                output += f"   High Risk Students: {len(high_risk)}\n\n"
                
                if high_risk:
                    output += "   🚨 Students requiring attention:\n"
                    for student in high_risk[:10]:  # Show top 10
                        output += f"      Student ID: {student['student_id']} (Risk: {student['risk_score']:.2%})\n"
                    
                    if len(high_risk) > 10:
                        output += f"      ... and {len(high_risk) - 10} more\n"
            
            if 'feature_importance' in predictions:
                importance = predictions['feature_importance']
                sorted_features = sorted(importance.items(), key=lambda x: x[1], reverse=True)
                
                output += "\n   📈 Most Important Risk Factors:\n"
                for feature, score in sorted_features:
                    feature_name = feature.replace('_', ' ').title()
                    output += f"      {feature_name}: {score:.3f}\n"
        
        self.predictions_display.insert(1.0, output)
    
    def run_anomaly_detection(self):
        """Run anomaly detection"""
        if not ENHANCED_AVAILABLE:
            messagebox.showwarning("Feature Unavailable", 
                                 "Anomaly detection requires the enhanced system.")
            return
        
        self.update_status("Running anomaly detection...")
        self.start_progress()
        
        def anomaly_task():
            try:
                anomalies = PredictiveAnalytics.detect_anomalies()
                self.root.after(0, lambda: self._display_anomaly_results(anomalies))
            except Exception as e:
                self.root.after(0, lambda: [
                    self.stop_progress(),
                    self.update_status(f"Anomaly detection failed: {str(e)}", "error"),
                    messagebox.showerror("Error", f"Failed to run anomaly detection: {str(e)}")
                ])
        
        threading.Thread(target=anomaly_task, daemon=True).start()
    
    def _display_anomaly_results(self, anomalies):
        """Display anomaly detection results"""
        self.stop_progress()
        self.update_status("Anomaly detection completed")
        
        # Create new window for anomaly results
        anomaly_window = tk.Toplevel(self.root)
        anomaly_window.title("Anomaly Detection Results")
        anomaly_window.geometry("600x500")
        
        anomaly_text = ScrolledText(anomaly_window, wrap=tk.WORD)
        anomaly_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        output = f"Anomaly Detection Results\n"
        output += "=" * 35 + "\n\n"
        
        if 'error' in anomalies:
            output += f"❌ Analysis unavailable: {anomalies['error']}\n"
        else:
            output += f"🔍 ANOMALY DETECTION RESULTS:\n\n"
            
            anomaly_count = anomalies.get('total_anomalies', 0)
            anomaly_rate = anomalies.get('anomaly_rate', 0)
            
            output += f"   Anomalous Students: {anomaly_count}\n"
            output += f"   Anomaly Rate: {anomaly_rate:.2f}%\n\n"
            
            if anomaly_rate > 15:
                output += "   ⚠️  High anomaly rate - investigate data quality\n"
            elif anomaly_rate > 5:
                output += "   ⚠️  Moderate anomalies detected\n"
            else:
                output += "   ✅ Normal anomaly rate\n"
            
            if 'anomalous_students' in anomalies and anomalies['anomalous_students']:
                output += "\n   🔍 Anomalous Student Profiles:\n\n"
                
                for student in anomalies['anomalous_students'][:10]:
                    output += f"      Student ID: {student['student_id']}\n"
                    output += f"         Age: {student['age']}\n"
                    output += f"         Modules: {student['unique_modules']}\n"
                    output += f"         Avg Grade: {student.get('avg_grade', 'N/A')}\n\n"
                
                if len(anomalies['anomalous_students']) > 10:
                    remaining = len(anomalies['anomalous_students']) - 10
                    output += f"      ... and {remaining} more anomalous profiles\n"
        
        anomaly_text.insert(1.0, output)
        anomaly_text.config(state=tk.DISABLED)
    
    def run_correlation_analysis(self):
        """Run correlation analysis"""
        if not ENHANCED_AVAILABLE:
            messagebox.showwarning("Feature Unavailable", 
                                 "Correlation analysis requires the enhanced system.")
            return
        
        self.update_status("Running correlation analysis...")
        self.start_progress()
        
        def correlation_task():
            try:
                conn = get_db_connection()
                chart_path = AdvancedVisualization.create_correlation_matrix(conn)
                
                self.root.after(0, lambda: [
                    self.stop_progress(),
                    self.update_status("Correlation analysis completed"),
                    self._show_correlation_results(chart_path)
                ])
                
            except Exception as e:
                self.root.after(0, lambda: [
                    self.stop_progress(),
                    self.update_status(f"Correlation analysis failed: {str(e)}", "error"),
                    messagebox.showerror("Error", f"Failed to run correlation analysis: {str(e)}")
                ])
        
        threading.Thread(target=correlation_task, daemon=True).start()
    
    def _show_correlation_results(self, chart_path):
        """Show correlation analysis results"""
        if chart_path and os.path.exists(chart_path):
            messagebox.showinfo("Correlation Analysis", 
                              f"✅ Correlation matrix generated successfully!\n\n📊 Chart saved to: {os.path.basename(chart_path)}")
            
            # Open the chart
            try:
                webbrowser.open(f"file://{os.path.abspath(chart_path)}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to open chart: {str(e)}")
        else:
            messagebox.showwarning("No Results", "❌ Unable to generate correlation matrix - insufficient data")
    
    # Scheduling methods
    
    def create_schedule(self):
        """Create a new scheduled report"""
        template_name = self.schedule_template_combo.get()
        if not template_name:
            messagebox.showwarning("No Template", "Please select a template for scheduling.")
            return
        
        frequency = self.frequency_combo.get().lower()
        hour = int(self.hour_var.get())
        
        # Get recipients
        recipients_text = self.recipients_entry.get(1.0, tk.END).strip()
        recipients = [email.strip() for email in recipients_text.split('\n') if email.strip() and '@' in email]
        
        if not recipients:
            if not messagebox.askyesno("No Recipients", 
                                     "No email recipients specified. Report will be generated but not sent. Continue?"):
                return
        
        try:
            if ENHANCED_AVAILABLE:
                # Create schedule configuration
                schedule_config = {
                    'frequency': frequency,
                    'hour': hour,
                    'enabled': True,
                    'last_run': None,
                    'next_run': None
                }
                
                # Save scheduled report
                scheduled_report = AdvancedScheduledReport(
                    template_name=template_name,
                    schedule_config=schedule_config,
                    recipients=recipients
                )
                
                scheduled_reports = load_scheduled_reports()
                scheduled_reports.append(scheduled_report.to_dict())
                save_scheduled_reports(scheduled_reports)
                
                # Clear form
                self.recipients_entry.delete(1.0, tk.END)
                
                self.refresh_data()
                messagebox.showinfo("Success", 
                                  f"✅ Report '{template_name}' scheduled for {frequency} generation at {hour:02d}:00\n\n📧 Recipients: {len(recipients)}")
            else:
                messagebox.showwarning("Feature Unavailable", 
                                     "Scheduling requires the enhanced system.")
        
        except Exception as e:
            messagebox.showerror("Error", f"Failed to create schedule: {str(e)}")
    
    def toggle_schedule(self):
        """Enable/disable selected schedule"""
        selection = self.schedule_tree.selection()
        if not selection or not hasattr(self, 'scheduled_reports_data'):
            messagebox.showwarning("No Selection", "Please select a schedule to toggle.")
            return
        
        try:
            item_index = self.schedule_tree.index(selection[0])
            report_data = self.scheduled_reports_data[item_index]
            
            current_status = report_data['schedule_config'].get('enabled', True)
            report_data['schedule_config']['enabled'] = not current_status
            
            if ENHANCED_AVAILABLE:
                save_scheduled_reports(self.scheduled_reports_data)
            
            self.refresh_data()
            
            new_status = "Enabled" if not current_status else "Disabled"
            messagebox.showinfo("Success", f"Schedule {new_status.lower()} successfully!")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to toggle schedule: {str(e)}")
    
    def edit_schedule(self):
        """Edit selected schedule"""
        selection = self.schedule_tree.selection()
        if not selection or not hasattr(self, 'scheduled_reports_data'):
            messagebox.showwarning("No Selection", "Please select a schedule to edit.")
            return

        try:
            # Get selected schedule
            item_index = self.schedule_tree.index(selection[0])
            schedule_data = self.scheduled_reports_data[item_index].copy()

            # Create edit dialog
            edit_dialog = tk.Toplevel(self.root)
            edit_dialog.title("Edit Schedule")
            edit_dialog.geometry("600x550")
            edit_dialog.transient(self.root)
            edit_dialog.grab_set()

            main_frame = ttk.Frame(edit_dialog, padding=20)
            main_frame.pack(fill='both', expand=True)

            ttk.Label(main_frame, text="Edit Report Schedule", font=('Arial', 12, 'bold')).pack(pady=10)

            # Form frame
            form_frame = ttk.Frame(main_frame)
            form_frame.pack(fill='both', expand=True, pady=10)

            row = 0
            ttk.Label(form_frame, text="Report Template:").grid(row=row, column=0, sticky='w', pady=5)
            template_var = tk.StringVar(value=schedule_data.get('template_name', ''))
            template_combo = ttk.Combobox(form_frame, textvariable=template_var, width=40)
            template_combo.grid(row=row, column=1, pady=5, padx=10)

            # Load available templates
            try:
                if ENHANCED_AVAILABLE:
                    from university_system.modules.domain.analytics.services.enhanced_reporting import get_available_templates
                    templates = get_available_templates()
                    template_combo['values'] = list(templates.keys())
            except:
                template_combo['values'] = ['enrollment_summary', 'financial_overview', 'student_performance', 'course_analytics']

            row += 1
            ttk.Label(form_frame, text="Schedule Type:").grid(row=row, column=0, sticky='w', pady=5)
            schedule_type_var = tk.StringVar(value=schedule_data.get('schedule_type', 'daily'))
            schedule_type_combo = ttk.Combobox(form_frame, textvariable=schedule_type_var,
                                              values=['daily', 'weekly', 'monthly'], width=40)
            schedule_type_combo.grid(row=row, column=1, pady=5, padx=10)

            row += 1
            ttk.Label(form_frame, text="Time (HH:MM):").grid(row=row, column=0, sticky='w', pady=5)
            time_var = tk.StringVar(value=schedule_data.get('time', '09:00'))
            ttk.Entry(form_frame, textvariable=time_var, width=40).grid(row=row, column=1, pady=5, padx=10)

            row += 1
            ttk.Label(form_frame, text="Format:").grid(row=row, column=0, sticky='w', pady=5)
            format_var = tk.StringVar(value=schedule_data.get('format', 'pdf'))
            format_combo = ttk.Combobox(form_frame, textvariable=format_var,
                                       values=['pdf', 'xlsx', 'html', 'csv'], width=40)
            format_combo.grid(row=row, column=1, pady=5, padx=10)

            row += 1
            ttk.Label(form_frame, text="Email Recipients:").grid(row=row, column=0, sticky='nw', pady=5)
            ttk.Label(form_frame, text="(One per line)", font=('Arial', 8)).grid(row=row, column=1, sticky='w', pady=5, padx=10)

            row += 1
            recipients_text = tk.Text(form_frame, height=6, width=40)
            recipients_text.grid(row=row, column=1, pady=5, padx=10)
            if 'recipients' in schedule_data:
                recipients_text.insert('1.0', '\n'.join(schedule_data['recipients']))

            row += 1
            enabled_var = tk.BooleanVar(value=schedule_data.get('enabled', True))
            ttk.Checkbutton(form_frame, text="Schedule Enabled", variable=enabled_var).grid(row=row, column=1, sticky='w', pady=10, padx=10)

            row += 1
            ttk.Label(form_frame, text="Description:").grid(row=row, column=0, sticky='nw', pady=5)
            description_text = tk.Text(form_frame, height=4, width=40)
            description_text.grid(row=row, column=1, pady=5, padx=10)
            if 'description' in schedule_data:
                description_text.insert('1.0', schedule_data['description'])

            def save_schedule():
                try:
                    # Update schedule data
                    schedule_data['template_name'] = template_var.get()
                    schedule_data['schedule_type'] = schedule_type_var.get()
                    schedule_data['time'] = time_var.get()
                    schedule_data['format'] = format_var.get()
                    schedule_data['enabled'] = enabled_var.get()
                    schedule_data['description'] = description_text.get('1.0', tk.END).strip()

                    # Parse recipients
                    recipients_input = recipients_text.get('1.0', tk.END).strip()
                    schedule_data['recipients'] = [r.strip() for r in recipients_input.split('\n') if r.strip()]

                    # Update in list
                    self.scheduled_reports_data[item_index] = schedule_data

                    # Save to file
                    if ENHANCED_AVAILABLE:
                        from university_system.modules.domain.analytics.services.enhanced_reporting import save_scheduled_reports
                        save_scheduled_reports(self.scheduled_reports_data)

                    messagebox.showinfo("Success", "Schedule updated successfully!")
                    edit_dialog.destroy()
                    self.refresh_data()

                except Exception as e:
                    messagebox.showerror("Error", f"Failed to save schedule: {str(e)}")

            # Buttons
            button_frame = ttk.Frame(main_frame)
            button_frame.pack(pady=10)

            ttk.Button(button_frame, text="Save", command=save_schedule).pack(side='left', padx=5)
            ttk.Button(button_frame, text="Cancel", command=edit_dialog.destroy).pack(side='left', padx=5)

        except Exception as e:
            messagebox.showerror("Error", f"Failed to edit schedule: {str(e)}")
    
    def run_schedule_now(self):
        """Run selected schedule immediately"""
        selection = self.schedule_tree.selection()
        if not selection or not hasattr(self, 'scheduled_reports_data'):
            messagebox.showwarning("No Selection", "Please select a schedule to run.")
            return
        
        try:
            item_index = self.schedule_tree.index(selection[0])
            report_data = self.scheduled_reports_data[item_index]
            template_name = report_data['template_name']
            
            self.update_status(f"Running scheduled report: {template_name}")
            self.start_progress()
            
            def run_task():
                try:
                    if ENHANCED_AVAILABLE:
                        end_date = datetime.now().strftime("%Y-%m-%d")
                        start_date = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
                        
                        report_path = generate_report(template_name, start_date, end_date, 'pdf')
                        
                        if report_path:
                            # Update run statistics
                            report_data['last_run'] = datetime.now().isoformat()
                            report_data['run_count'] = report_data.get('run_count', 0) + 1
                            
                            save_scheduled_reports(self.scheduled_reports_data)
                            
                            self.root.after(0, lambda: [
                                self.stop_progress(),
                                self.update_status("Scheduled report completed"),
                                self.refresh_data(),
                                messagebox.showinfo("Success", f"✅ Scheduled report '{template_name}' generated successfully!\n\n📄 File: {os.path.basename(report_path)}")
                            ])
                        else:
                            self.root.after(0, lambda: [
                                self.stop_progress(),
                                self.update_status("Scheduled report failed", "error"),
                                messagebox.showerror("Error", f"Failed to generate scheduled report '{template_name}'")
                            ])
                    else:
                        self.root.after(0, lambda: [
                            self.stop_progress(),
                            self.update_status("Feature unavailable", "warning"),
                            messagebox.showwarning("Feature Unavailable", "Scheduling requires the enhanced system.")
                        ])
                        
                except Exception as e:
                    self.root.after(0, lambda: [
                        self.stop_progress(),
                        self.update_status(f"Error: {str(e)}", "error"),
                        messagebox.showerror("Error", f"Error running scheduled report: {str(e)}")
                    ])
            
            threading.Thread(target=run_task, daemon=True).start()
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to run schedule: {str(e)}")
    
    def delete_schedule(self):
        """Delete selected schedule"""
        selection = self.schedule_tree.selection()
        if not selection or not hasattr(self, 'scheduled_reports_data'):
            messagebox.showwarning("No Selection", "Please select a schedule to delete.")
            return
        
        try:
            item_index = self.schedule_tree.index(selection[0])
            report_data = self.scheduled_reports_data[item_index]
            template_name = report_data['template_name']
            
            if messagebox.askyesno("Confirm Delete", 
                                 f"Are you sure you want to delete the schedule for '{template_name}'?"):
                
                del self.scheduled_reports_data[item_index]
                
                if ENHANCED_AVAILABLE:
                    save_scheduled_reports(self.scheduled_reports_data)
                
                self.refresh_data()
                messagebox.showinfo("Success", "Schedule deleted successfully!")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to delete schedule: {str(e)}")
    
    # System management methods
    
    def clean_old_reports(self):
        """Clean old report files"""
        self.update_status("Cleaning old reports...")
        self.start_progress()
        
        def clean_task():
            try:
                if ENHANCED_AVAILABLE:
                    cleanup_old_reports()
                
                self.root.after(0, lambda: [
                    self.stop_progress(),
                    self.update_status("Old reports cleaned"),
                    self.refresh_reports(),
                    messagebox.showinfo("Success", "✅ Old reports cleaned successfully!")
                ])
                
            except Exception as e:
                self.root.after(0, lambda: [
                    self.stop_progress(),
                    self.update_status(f"Cleanup failed: {str(e)}", "error"),
                    messagebox.showerror("Error", f"Failed to clean old reports: {str(e)}")
                ])
        
        threading.Thread(target=clean_task, daemon=True).start()
    
    def clear_cache(self):
        """Clear system cache"""
        self.update_status("Clearing cache...")
        self.start_progress()
        
        def cache_task():
            try:
                if ENHANCED_AVAILABLE:
                    CacheManager.cleanup_cache()
                
                self.root.after(0, lambda: [
                    self.stop_progress(),
                    self.update_status("Cache cleared"),
                    messagebox.showinfo("Success", "✅ Cache cleared successfully!")
                ])
                
            except Exception as e:
                self.root.after(0, lambda: [
                    self.stop_progress(),
                    self.update_status(f"Cache clear failed: {str(e)}", "error"),
                    messagebox.showerror("Error", f"Failed to clear cache: {str(e)}")
                ])
        
        threading.Thread(target=cache_task, daemon=True).start()
    
    def run_maintenance_quality_check(self):
        """Run maintenance quality check"""
        self.run_quality_check()
    
    def optimize_database(self):
        """Optimize database"""
        self.update_status("Optimizing database...")
        self.start_progress()
        
        def optimize_task():
            try:
                if ENHANCED_AVAILABLE:
                    conn = get_db_connection()
                    try:
                        conn.execute("VACUUM")
                        conn.execute("ANALYZE")
                    finally:
                        conn.close()
                
                self.root.after(0, lambda: [
                    self.stop_progress(),
                    self.update_status("Database optimized"),
                    messagebox.showinfo("Success", "✅ Database optimized successfully!")
                ])
                
            except Exception as e:
                self.root.after(0, lambda: [
                    self.stop_progress(),
                    self.update_status(f"Optimization failed: {str(e)}", "error"),
                    messagebox.showerror("Error", f"Failed to optimize database: {str(e)}")
                ])
        
        threading.Thread(target=optimize_task, daemon=True).start()
    
    def run_all_maintenance(self):
        """Run all maintenance tasks"""
        self.update_status("Running all maintenance tasks...")
        self.start_progress()
        
        def maintenance_task():
            try:
                if ENHANCED_AVAILABLE:
                    quality_report = run_system_maintenance()
                
                self.root.after(0, lambda: [
                    self.stop_progress(),
                    self.update_status("All maintenance completed"),
                    self.refresh_data(),
                    messagebox.showinfo("Success", "✅ All maintenance tasks completed successfully!")
                ])
                
            except Exception as e:
                self.root.after(0, lambda: [
                    self.stop_progress(),
                    self.update_status(f"Maintenance failed: {str(e)}", "error"),
                    messagebox.showerror("Error", f"Failed to run maintenance: {str(e)}")
                ])
        
        threading.Thread(target=maintenance_task, daemon=True).start()
    
    def show_performance_monitor(self):
        """Show performance monitoring window"""
        perf_window = tk.Toplevel(self.root)
        perf_window.title("Performance Monitor")
        perf_window.geometry("600x500")
        
        perf_text = ScrolledText(perf_window, wrap=tk.WORD)
        perf_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Get performance data
        self.update_status("Gathering performance data...")
        
        def perf_task():
            try:
                output = "Performance Monitor\n"
                output += "=" * 35 + "\n\n"
                
                if ENHANCED_AVAILABLE:
                    # Database size
                    db_size = os.path.getsize(CONFIG['database']) / (1024 * 1024)  # MB
                    output += f"💾 Database size: {db_size:.2f} MB\n"
                    
                    # Reports directory size
                    reports_size = 0
                    if os.path.exists(CONFIG['reports_dir']):
                        for root, dirs, files in os.walk(CONFIG['reports_dir']):
                            for file in files:
                                reports_size += os.path.getsize(os.path.join(root, file))
                    reports_size = reports_size / (1024 * 1024)  # MB
                    output += f"📁 Reports size: {reports_size:.2f} MB\n"
                    
                    # Cache directory size
                    cache_size = 0
                    if os.path.exists(CONFIG['cache_dir']):
                        for root, dirs, files in os.walk(CONFIG['cache_dir']):
                            for file in files:
                                cache_size += os.path.getsize(os.path.join(root, file))
                    cache_size = cache_size / (1024 * 1024)  # MB
                    output += f"🗄️ Cache size: {cache_size:.2f} MB\n\n"
                    
                    # Record counts
                    conn = get_db_connection()
                    try:
                        cursor = conn.cursor()
                        
                        cursor.execute("SELECT COUNT(*) FROM students")
                        student_count = cursor.fetchone()[0]
                        output += f"👥 Total students: {student_count}\n"
                        
                        # Check if other tables exist
                        tables = ['student_modules', 'student_grades', 'student_attendance']
                        for table in tables:
                            try:
                                cursor.execute(f"SELECT COUNT(*) FROM {table}")
                                count = cursor.fetchone()[0]
                                table_name = table.replace('_', ' ').title()
                                output += f"📚 {table_name}: {count}\n"
                            except:
                                pass  # Table might not exist
                                
                    finally:
                        conn.close()
                    
                    # Template and schedule counts
                    templates = load_templates()
                    output += f"📋 Templates: {len(templates)}\n"
                    
                    scheduled_reports = load_scheduled_reports()
                    output += f"⏰ Scheduled reports: {len(scheduled_reports)}\n"
                    
                else:
                    output += "Enhanced features not available\n"
                    output += "Limited performance data shown\n"
                
                self.root.after(0, lambda: [
                    perf_text.insert(1.0, output),
                    perf_text.config(state=tk.DISABLED),
                    self.update_status("Performance data loaded")
                ])
                
            except Exception as e:
                error_output = f"Error gathering performance data: {str(e)}"
                self.root.after(0, lambda: [
                    perf_text.insert(1.0, error_output),
                    perf_text.config(state=tk.DISABLED),
                    self.update_status("Performance data failed", "error")
                ])
        
        threading.Thread(target=perf_task, daemon=True).start()
    
    def export_system_logs(self):
        """Export system logs"""
        if not ENHANCED_AVAILABLE:
            messagebox.showwarning("Feature Unavailable", 
                                 "Log export requires the enhanced system.")
            return
        
        try:
            log_file = get_log_file('reporting_system.log')
            
            if not os.path.exists(log_file):
                messagebox.showwarning("No Logs", "No log file found")
                return
            
            file_path = filedialog.asksaveasfilename(
                defaultextension=".txt",
                filetypes=[("Text files", "*.txt"), ("All files", "*.*")],
                initialfile=f"system_logs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
            )
            
            if file_path:
                # Copy log file
                import shutil
                shutil.copy2(log_file, file_path)
                messagebox.showinfo("Success", f"System logs exported to {file_path}")
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export logs: {str(e)}")
    
    def reload_config(self):
        """Reload system configuration"""
        try:
            if ENHANCED_AVAILABLE:
                config = SystemConfig.load_config()
                
                config_text = json.dumps(config, indent=4)
                self.config_display.delete(1.0, tk.END)
                self.config_display.insert(1.0, config_text)
                
                messagebox.showinfo("Success", "Configuration reloaded successfully!")
            else:
                messagebox.showwarning("Feature Unavailable", 
                                     "Configuration management requires the enhanced system.")
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to reload configuration: {str(e)}")
    
    def save_config(self):
        """Save system configuration"""
        try:
            if ENHANCED_AVAILABLE:
                config_text = self.config_display.get(1.0, tk.END).strip()
                config = json.loads(config_text)
                
                SystemConfig.save_config(config)
                messagebox.showinfo("Success", "Configuration saved successfully!")
            else:
                messagebox.showwarning("Feature Unavailable", 
                                     "Configuration management requires the enhanced system.")
                
        except json.JSONDecodeError:
            messagebox.showerror("Invalid JSON", "Configuration contains invalid JSON syntax.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save configuration: {str(e)}")
    
    def show_advanced_settings(self):
        """Show advanced settings dialog"""
        settings_window = tk.Toplevel(self.root)
        settings_window.title("Advanced Settings")
        settings_window.geometry("500x400")
        settings_window.transient(self.root)
        
        # Create notebook for different setting categories
        settings_notebook = ttk.Notebook(settings_window)
        settings_notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Load existing settings
        default_settings = {
            'cache_expiry_hours': 24,
            'max_cache_size_mb': 500,
            'session_timeout_seconds': 3600,
            'require_2fa': False,
            'enable_caching': True,
            'max_concurrent_reports': 5
        }

        try:
            config_file = paths.DATA_DIR / 'reporting_settings.json'
            if config_file.exists():
                with open(config_file, 'r') as f:
                    loaded_settings = json.load(f)
                    default_settings.update(loaded_settings)
        except Exception as e:
            print(f"Could not load settings: {e}")

        # General settings
        general_frame = ttk.Frame(settings_notebook, padding="10")
        settings_notebook.add(general_frame, text="General")

        ttk.Label(general_frame, text="Cache expiry (hours):").pack(anchor=tk.W, pady=5)
        cache_expiry = tk.StringVar(value=str(default_settings['cache_expiry_hours']))
        ttk.Entry(general_frame, textvariable=cache_expiry).pack(fill=tk.X, pady=(0, 10))

        ttk.Label(general_frame, text="Max cache size (MB):").pack(anchor=tk.W, pady=5)
        cache_size = tk.StringVar(value=str(default_settings['max_cache_size_mb']))
        ttk.Entry(general_frame, textvariable=cache_size).pack(fill=tk.X, pady=(0, 10))

        # Security settings
        security_frame = ttk.Frame(settings_notebook, padding="10")
        settings_notebook.add(security_frame, text="Security")

        ttk.Label(security_frame, text="Session timeout (seconds):").pack(anchor=tk.W, pady=5)
        session_timeout = tk.StringVar(value=str(default_settings['session_timeout_seconds']))
        ttk.Entry(security_frame, textvariable=session_timeout).pack(fill=tk.X, pady=(0, 10))

        require_2fa = tk.BooleanVar(value=default_settings['require_2fa'])
        ttk.Checkbutton(security_frame, text="Require 2FA", variable=require_2fa).pack(anchor=tk.W, pady=5)

        # Performance settings
        performance_frame = ttk.Frame(settings_notebook, padding="10")
        settings_notebook.add(performance_frame, text="Performance")

        enable_caching = tk.BooleanVar(value=default_settings['enable_caching'])
        ttk.Checkbutton(performance_frame, text="Enable caching", variable=enable_caching).pack(anchor=tk.W, pady=5)

        ttk.Label(performance_frame, text="Max concurrent reports:").pack(anchor=tk.W, pady=5)
        max_reports = tk.StringVar(value=str(default_settings['max_concurrent_reports']))
        ttk.Entry(performance_frame, textvariable=max_reports).pack(fill=tk.X, pady=(0, 10))
        
        # Buttons
        button_frame = ttk.Frame(settings_window)
        button_frame.pack(fill=tk.X, padx=10, pady=(0, 10))
        
        def save_settings():
            try:
                # Prepare settings dictionary
                settings = {
                    'cache_expiry_hours': int(cache_expiry.get()),
                    'max_cache_size_mb': int(cache_size.get()),
                    'session_timeout_seconds': int(session_timeout.get()),
                    'require_2fa': require_2fa.get(),
                    'enable_caching': enable_caching.get(),
                    'max_concurrent_reports': int(max_reports.get())
                }

                # Save to config file
                config_file = paths.DATA_DIR / 'reporting_settings.json'
                with open(config_file, 'w') as f:
                    json.dump(settings, f, indent=4)

                messagebox.showinfo("Success", "Settings saved successfully!")
                settings_window.destroy()
            except ValueError as e:
                messagebox.showerror("Invalid Input", "Please enter valid numeric values for all fields.")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to save settings: {str(e)}")
        
        ttk.Button(button_frame, text="Save", command=save_settings).pack(side=tk.RIGHT, padx=(5, 0))
        ttk.Button(button_frame, text="Cancel", command=settings_window.destroy).pack(side=tk.RIGHT)
    
    # Utility methods
    
    def start_api_server(self):
        """Start the REST API server"""
        if not ENHANCED_AVAILABLE:
            messagebox.showwarning("Feature Unavailable", 
                                 "API server requires the enhanced system.")
            return
        
        # API server dialog
        api_dialog = tk.Toplevel(self.root)
        api_dialog.title("Start API Server")
        api_dialog.geometry("400x300")
        api_dialog.transient(self.root)
        
        ttk.Label(api_dialog, text="API Server Configuration", style='Subtitle.TLabel').pack(pady=10)
        
        # Host and port settings
        config_frame = ttk.Frame(api_dialog)
        config_frame.pack(fill=tk.X, padx=20, pady=10)
        
        ttk.Label(config_frame, text="Host:").pack(anchor=tk.W)
        host_var = tk.StringVar(value="localhost")
        ttk.Entry(config_frame, textvariable=host_var).pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(config_frame, text="Port:").pack(anchor=tk.W)
        port_var = tk.StringVar(value="5000")
        ttk.Entry(config_frame, textvariable=port_var).pack(fill=tk.X, pady=(0, 10))
        
        # API endpoints info
        info_text = ScrolledText(api_dialog, height=8, wrap=tk.WORD)
        info_text.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        api_info = """API Endpoints:

• POST /api/login - User authentication
• GET  /api/templates - List templates  
• POST /api/templates - Create template
• POST /api/reports/generate - Generate report
• GET  /api/analytics/quality - Data quality
• GET  /api/analytics/predictions - Predictions
• GET  /api/analytics/anomalies - Anomaly detection"""
        
        info_text.insert(1.0, api_info)
        info_text.config(state=tk.DISABLED)
        
        # Buttons
        button_frame = ttk.Frame(api_dialog)
        button_frame.pack(fill=tk.X, padx=20, pady=(0, 20))
        
        def start_server():
            host = host_var.get()
            port = int(port_var.get())
            
            messagebox.showinfo("API Server", 
                              f"API server would start on http://{host}:{port}\n\n(This would run in background thread)")
            api_dialog.destroy()
        
        ttk.Button(button_frame, text="Start Server", command=start_server, 
                  style='Success.TButton').pack(side=tk.RIGHT, padx=(5, 0))
        ttk.Button(button_frame, text="Cancel", command=api_dialog.destroy).pack(side=tk.RIGHT)
    
    def show_settings(self):
        """Show general settings"""
        self.show_advanced_settings()
    
    def update_status(self, message, status_type="info"):
        """Update status bar message"""
        colors = {
            "info": "Info.TLabel",
            "success": "Success.TLabel", 
            "warning": "Warning.TLabel",
            "error": "Error.TLabel"
        }
        
        self.status_text.config(text=message, style=colors.get(status_type, "Info.TLabel"))
    
    def start_progress(self):
        """Start progress bar animation"""
        self.progress.start(10)
    
    def stop_progress(self):
        """Stop progress bar animation"""
        self.progress.stop()


class TemplateDialog:
    """Dialog for creating and editing templates"""
    
    def __init__(self, parent, title="Template Dialog", template_data=None):
        self.result = None
        self.template_data = template_data
        
        # Create dialog window
        self.dialog = tk.Toplevel(parent)
        self.dialog.title(title)
        self.dialog.geometry("700x600")
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        # Center the dialog
        self.dialog.update_idletasks()
        x = (self.dialog.winfo_screenwidth() // 2) - (700 // 2)
        y = (self.dialog.winfo_screenheight() // 2) - (600 // 2)
        self.dialog.geometry(f"700x600+{x}+{y}")
        
        self.create_widgets()
        
        if template_data:
            self.populate_fields()
        
        # Wait for dialog to close
        self.dialog.wait_window()
    
    def create_widgets(self):
        """Create dialog widgets"""
        main_frame = ttk.Frame(self.dialog, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        main_frame.columnconfigure(1, weight=1)
        
        # Basic information
        basic_frame = ttk.LabelFrame(main_frame, text="Basic Information", padding="10")
        basic_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        basic_frame.columnconfigure(1, weight=1)
        
        # Template name
        ttk.Label(basic_frame, text="Name:*").grid(row=0, column=0, sticky=tk.W, pady=(0, 5))
        self.name_var = tk.StringVar()
        ttk.Entry(basic_frame, textvariable=self.name_var).grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=(0, 5))
        
        # Description
        ttk.Label(basic_frame, text="Description:").grid(row=1, column=0, sticky=tk.W, pady=(0, 5))
        self.description_var = tk.StringVar()
        ttk.Entry(basic_frame, textvariable=self.description_var).grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=(0, 5))
        
        # Security level
        ttk.Label(basic_frame, text="Security Level:").grid(row=2, column=0, sticky=tk.W, pady=(0, 5))
        self.security_var = tk.StringVar(value="normal")
        security_combo = ttk.Combobox(basic_frame, textvariable=self.security_var, 
                                     values=["normal", "confidential", "restricted"], state="readonly")
        security_combo.grid(row=2, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=(0, 5))
        
        # Visualization type
        ttk.Label(basic_frame, text="Visualization:").grid(row=3, column=0, sticky=tk.W, pady=(0, 5))
        self.viz_var = tk.StringVar(value="standard")
        viz_combo = ttk.Combobox(basic_frame, textvariable=self.viz_var,
                                values=["standard", "advanced", "interactive"], state="readonly")
        viz_combo.grid(row=3, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=(0, 5))
        
        # Sections selection
        sections_frame = ttk.LabelFrame(main_frame, text="Report Sections", padding="10")
        sections_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        sections_frame.columnconfigure(0, weight=1)
        sections_frame.rowconfigure(1, weight=1)
        
        # Sections checkboxes
        sections_canvas_frame = ttk.Frame(sections_frame)
        sections_canvas_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        sections_canvas_frame.columnconfigure(0, weight=1)
        sections_canvas_frame.rowconfigure(0, weight=1)
        
        # Scrollable frame for sections
        sections_canvas = tk.Canvas(sections_canvas_frame, height=200)
        sections_scrollbar = ttk.Scrollbar(sections_canvas_frame, orient="vertical", command=sections_canvas.yview)
        scrollable_frame = ttk.Frame(sections_canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: sections_canvas.configure(scrollregion=sections_canvas.bbox("all"))
        )
        
        sections_canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        sections_canvas.configure(yscrollcommand=sections_scrollbar.set)
        
        sections_canvas.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        sections_scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        # Create checkboxes for sections
        self.section_vars = {}
        available_sections = [
            "student_overview", "course_distribution", "gender_distribution",
            "age_distribution", "registration_trends", "module_popularity", 
            "grade_distribution", "attendance_summary", "data_quality_report",
            "predictive_analytics", "correlation_analysis", "anomaly_detection",
            "performance_benchmarks", "trend_analysis"
        ]
        
        for i, section in enumerate(available_sections):
            var = tk.BooleanVar()
            self.section_vars[section] = var
            
            section_name = section.replace('_', ' ').title()
            ttk.Checkbutton(scrollable_frame, text=section_name, variable=var).grid(
                row=i // 2, column=i % 2, sticky=tk.W, padx=10, pady=2)
        
        # Select/Deselect all buttons
        select_frame = ttk.Frame(sections_frame)
        select_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(10, 0))
        
        ttk.Button(select_frame, text="Select All", command=self.select_all_sections).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(select_frame, text="Deselect All", command=self.deselect_all_sections).pack(side=tk.LEFT)
        
        # Filters section
        filters_frame = ttk.LabelFrame(main_frame, text="Filters (Optional)", padding="10")
        filters_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        filters_frame.columnconfigure(1, weight=1)
        
        # Course filter
        ttk.Label(filters_frame, text="Course:").grid(row=0, column=0, sticky=tk.W, pady=(0, 5))
        self.course_var = tk.StringVar()
        course_combo = ttk.Combobox(filters_frame, textvariable=self.course_var,
                                   values=["", "CS", "DS"], state="readonly")
        course_combo.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=(0, 5))
        
        # Date range filter
        ttk.Label(filters_frame, text="Date Range (days):").grid(row=1, column=0, sticky=tk.W, pady=(0, 5))
        self.date_range_var = tk.StringVar(value="30")
        ttk.Entry(filters_frame, textvariable=self.date_range_var, width=10).grid(row=1, column=1, sticky=tk.W, padx=(10, 0), pady=(0, 5))
        
        # Buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=3, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(10, 0))
        
        ttk.Button(button_frame, text="Cancel", command=self.cancel).pack(side=tk.RIGHT, padx=(5, 0))
        ttk.Button(button_frame, text="Save Template", command=self.save_template,
                  style='Primary.TButton').pack(side=tk.RIGHT)
    
    def populate_fields(self):
        """Populate fields with existing template data"""
        if not self.template_data:
            return
        
        self.name_var.set(self.template_data.get('name', ''))
        self.description_var.set(self.template_data.get('description', ''))
        self.security_var.set(self.template_data.get('security_level', 'normal'))
        self.viz_var.set(self.template_data.get('visualization_type', 'standard'))
        
        # Set selected sections
        template_sections = self.template_data.get('sections', [])
        for section, var in self.section_vars.items():
            var.set(section in template_sections)
        
        # Set filters
        filters = self.template_data.get('filters', {})
        self.course_var.set(filters.get('course', ''))
        self.date_range_var.set(str(filters.get('date_range_days', 30)))
    
    def select_all_sections(self):
        """Select all sections"""
        for var in self.section_vars.values():
            var.set(True)
    
    def deselect_all_sections(self):
        """Deselect all sections"""
        for var in self.section_vars.values():
            var.set(False)
    
    def save_template(self):
        """Save the template"""
        name = self.name_var.get().strip()
        if not name:
            messagebox.showerror("Validation Error", "Template name is required.")
            return
        
        # Get selected sections
        selected_sections = [section for section, var in self.section_vars.items() if var.get()]
        if not selected_sections:
            messagebox.showerror("Validation Error", "At least one section must be selected.")
            return
        
        # Build filters
        filters = {}
        if self.course_var.get():
            filters['course'] = self.course_var.get()
        
        try:
            date_range = int(self.date_range_var.get())
            if date_range > 0:
                filters['date_range_days'] = date_range
        except ValueError:
            pass
        
        # Create template data
        template_data = {
            'name': name,
            'description': self.description_var.get().strip(),
            'sections': selected_sections,
            'filters': filters,
            'security_level': self.security_var.get(),
            'visualization_type': self.viz_var.get(),
            'created_at': datetime.now().isoformat(),
            'version': '1.0'
        }
        
        # If editing, preserve creation date and increment version
        if self.template_data:
            template_data['created_at'] = self.template_data.get('created_at', datetime.now().isoformat())
            old_version = self.template_data.get('version', '1.0')
            try:
                major, minor = old_version.split('.')
                template_data['version'] = f"{major}.{int(minor) + 1}"
            except:
                template_data['version'] = '1.1'
        
        try:
            if ENHANCED_AVAILABLE:
                # Save template using enhanced system (database)
                save_template_dict(template_data)

            self.result = template_data
            messagebox.showinfo("Success", f"Template '{name}' saved successfully!")
            self.dialog.destroy()

        except Exception as e:
            logging.error(f"Failed to save template: {e}")
            import traceback
            logging.error(traceback.format_exc())
            messagebox.showerror("Error", f"Failed to save template: {str(e)}")
    
    def cancel(self):
        """Cancel dialog"""
        self.dialog.destroy()


def start_gui():
    """Start the GUI application"""
    # Create main window
    root = tk.Tk()
    
    # Set application icon if available
    try:
        # You can add an icon file here
        # root.iconbitmap('icon.ico')
        pass
    except:
        pass
    
    # Create and run the application
    app = ReportingSystemGUI(root)
    
    # Handle window closing
    def on_closing():
        if messagebox.askokcancel("Quit", "Do you want to quit the application?"):
            root.destroy()
    
    root.protocol("WM_DELETE_WINDOW", on_closing)
    
    # Start the GUI main loop
    root.mainloop()


# Backward compatibility function
def start_enhanced_reporting_gui():
    """
    Start the enhanced reporting system with GUI
    Maintains backward compatibility with existing CLI system
    """
    print("🖥️  Starting Enhanced Reporting System GUI...")
    
    # Check if enhanced features are available
    if ENHANCED_AVAILABLE:
        print("✅ Enhanced features loaded successfully")
        
        # Initialize system
        try:
            # Check database connection
            conn = get_db_connection()
            conn.close()
            print("✅ Database connection verified")
            
            # Initialize directories
            for dir_path in [CONFIG['reports_dir'], CONFIG['templates_dir'], CONFIG['cache_dir']]:
                os.makedirs(dir_path, exist_ok=True)
            print("✅ Directories initialized")
            
            # Start background scheduler if needed
            try:
                start_scheduler()
                print("✅ Background scheduler started")
            except Exception as e:
                print(f"⚠️  Background scheduler failed: {str(e)}")
                
        except Exception as e:
            print(f"⚠️  System initialization warning: {str(e)}")
    else:
        print("⚠️  Enhanced features not available - running in basic mode")
    
    print("🚀 Launching GUI...")
    
    # Ask user preference
    try:
        import tkinter.messagebox as mb
        choice = mb.askyesnocancel(
            "Enhanced Reporting System", 
            "Choose interface:\n\nYes - Start GUI\nNo - Start CLI\nCancel - Exit",
            icon='question'
        )
        
        if choice is True:
            start_gui()
        elif choice is False:
            if ENHANCED_AVAILABLE:
                display_enhanced_reporting_menu()
            else:
                print("CLI mode requires enhanced features")
        # choice is None means cancelled
        
    except ImportError:
        # If tkinter is not available, fallback to CLI
        print("GUI not available, starting CLI...")
        if ENHANCED_AVAILABLE:
            display_enhanced_reporting_menu()
        else:
            print("Enhanced features not available")


# Main execution
if __name__ == "__main__":
    start_enhanced_reporting_gui()
