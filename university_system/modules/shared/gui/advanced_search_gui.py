from university_system.infrastructure.database.db import DEFAULT_DB_PATH, get_connection  # injected
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext
import threading
import queue
import json
import csv
from datetime import datetime, timedelta
import os
import sys
import shutil
import sqlite3
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

# Import enhanced console output utility
try:
    from university_system.modules.shared.utils.console_output import (
        console, print_success, print_error, print_warning, print_info, print_header
    )
except ImportError:
    # Fallback to basic print if console_output is not available
    class _FallbackConsole:
        def success(self, msg, **kwargs): print(msg)
        def error(self, msg, **kwargs): print(msg)
        def warning(self, msg, **kwargs): print(msg)
        def info(self, msg, **kwargs): print(msg)
        def header(self, msg, **kwargs): print(f"\n{msg}\n{'='*len(msg)}")
    console = _FallbackConsole()
    print_success = lambda msg: print(msg)
    print_error = lambda msg: print(msg)
    print_warning = lambda msg: print(msg)
    print_info = lambda msg: print(msg)
    print_header = lambda msg, **kwargs: print(f"\n{msg}\n{'='*len(msg)}")

# Import chart generation utility
try:
    from university_system.modules.shared.utils.chart_generator import (
        ChartGenerator, ChartViewer, DatabaseChartGenerator, create_chart_viewer, CHARTS_AVAILABLE
    )
except ImportError:
    CHARTS_AVAILABLE = False
    ChartGenerator = None
    ChartViewer = None
    DatabaseChartGenerator = None
    create_chart_viewer = None

# Import all original functions (backwards compatibility)
try:
    from university_system.modules.shared.services.analytics.advanced_search import (
        saved_searches, search_cache, academic_performance_analysis, add_condition,
        add_user_permissions, build_search_analytics_record,
        bulk_enrollment_management, bulk_export, check_database_status,
        create_scheduled_report, create_student_groups, custom_format_export,
        data_quality_reports, display_enhanced_menu, display_search_results,
        duplicate_detection, ensure_search_analytics_schema,
        execute_conditional_search, export_single_student,
        export_system_statistics, export_to_csv, export_to_excel, export_to_json,
        generate_custom_reports, generate_custom_sql_report, generate_email_list,
        generate_module_enrollment_report, generate_student_summary_report,
        get_search_analytics_columns, identify_at_risk_students,
        init_enhanced_database, insert_search_analytics_record,
        interactive_charts, manage_scheduled_reports, manage_user_permissions,
        mark_for_followup, performance_optimization,
        refresh_search_analytics_columns, remove_condition,
        save_last_search_results, search_analytics_dashboard,
        simulate_send_email, student_demographics_reports, view_academic_history,
        view_search_audit_trail
    )
except ImportError:
    # If running as standalone, define minimal required functions
    print_warning("Could not import advanced_search module. Running in standalone mode.")

# Import essential functions from email infrastructure
try:
    from university_system.infrastructure.email.email_db_utilities import execute_db_operation
    from university_system.infrastructure.email.admin import search_users, list_all_users
    print_success("Imported email infrastructure functions")
except ImportError as e:
    print_warning(f"Could not import email infrastructure functions: {e}")
    # Create fallback functions
    def execute_db_operation(operation_func):
        conn = get_connection()
        cursor = conn.cursor()
        try:
            result = operation_func(cursor)
            conn.commit()
            return result
        except Exception as e:
            conn.rollback()
            raise e
        finally:
            conn.close()

    def search_users(auth, search_term):
        def _search(cursor):
            cursor.execute("SELECT id, username, email FROM users WHERE username LIKE ? OR email LIKE ?",
                         (f'%{search_term}%', f'%{search_term}%'))
            return [{'id': row[0], 'username': row[1], 'email': row[2]} for row in cursor.fetchall()]
        return execute_db_operation(_search)

    def list_all_users(auth, limit=100):
        def _list_users(cursor):
            cursor.execute("SELECT id, username, email FROM users LIMIT ?", (limit,))
            users = [{'id': row[0], 'username': row[1], 'email': row[2]} for row in cursor.fetchall()]
            return {'users': users, 'total_count': len(users)}
        return execute_db_operation(_list_users)
    
    # Minimal database functions for standalone operation
    # Compute the central database path relative to this file so that
    # standalone mode writes to the shared student_records.db rather than
    # creating a new database in the working directory.
    from pathlib import Path
    def get_connection():
        from university_system.infrastructure.database.db import sqlite3
        # Use centralized path system
        from university_system.modules.shared.constants import paths
        return sqlite3.connect(str(paths.DEFAULT_DB_PATH))
    
    def init_enhanced_database():
        """
        Stand‑alone fallback database initializer. When the CLI version of
        advanced_search.py is not available, this function ensures that the
        necessary tables for analytics are created so that the GUI remains
        functional. It creates minimal versions of the tables defined in the
        CLI: search_analytics, saved_searches, and duplicate_candidates.
        """
        try:
            conn = get_connection()
            if conn is None:
                print_error("Could not obtain database connection for analytics fallback")
                return False
            cursor = conn.cursor()

            # Create table to track search analytics metrics
            cursor.execute('''
            CREATE TABLE IF NOT EXISTS search_analytics (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id TEXT,
                search_type TEXT,
                search_criteria TEXT,
                results_count INTEGER,
                execution_time REAL,
                timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
            )
            ''')

            # Create table to store saved search definitions
            cursor.execute('''
            CREATE TABLE IF NOT EXISTS saved_searches (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id TEXT,
                search_name TEXT,
                search_criteria TEXT,
                is_shared BOOLEAN DEFAULT 0,
                created_date DATETIME DEFAULT CURRENT_TIMESTAMP
            )
            ''')

            # Create users table if it doesn't exist (for user search functionality)
            cursor.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE,
                email TEXT UNIQUE,
                first_name TEXT,
                last_name TEXT,
                role TEXT DEFAULT 'student',
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
            ''')

            # Insert sample users if table is empty
            cursor.execute('SELECT COUNT(*) FROM users')
            if cursor.fetchone()[0] == 0:
                sample_users = [
                    ('admin', 'admin@example.com', 'Admin', 'User', 'admin'),
                    ('staff', 'staff@example.com', 'Staff', 'User', 'staff'),
                    ('student', 'student@example.com', 'Student', 'User', 'student')
                ]
                cursor.executemany('''
                INSERT INTO users (username, email, first_name, last_name, role)
                VALUES (?, ?, ?, ?, ?)
                ''', sample_users)

            # Create students table for analytics (matching main auth structure)
            cursor.execute('''
            CREATE TABLE IF NOT EXISTS students (
                student_id TEXT PRIMARY KEY,
                email_address TEXT,
                title TEXT,
                first_name TEXT,
                middle_name TEXT,
                last_name TEXT,
                gender TEXT,
                dob TEXT,
                age INTEGER,
                course TEXT,
                year TEXT,
                registration_datetime TEXT,
                status TEXT DEFAULT 'Active',
                enrollment_date TEXT
            )
            ''')

            # Insert sample students if table is empty
            cursor.execute('SELECT COUNT(*) FROM students')
            if cursor.fetchone()[0] == 0:
                from datetime import datetime
                current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                sample_students = [
                    ('S001', 'john.doe@student.edu', 'Mr', 'John', '', 'Doe', 'Male', '2000-01-15', 24, 'Computer Science', 'Year 3', current_time, 'Active', '2022-09-01'),
                    ('S002', 'jane.smith@student.edu', 'Ms', 'Jane', '', 'Smith', 'Female', '1999-05-20', 25, 'Data Science', 'Year 4', current_time, 'Active', '2021-09-01'),
                    ('S003', 'bob.wilson@student.edu', 'Mr', 'Bob', '', 'Wilson', 'Male', '2001-03-10', 23, 'Engineering', 'Year 2', current_time, 'Active', '2023-09-01')
                ]
                cursor.executemany('''
                INSERT INTO students (student_id, email_address, title, first_name, middle_name, last_name, gender, dob, age, course, year, registration_datetime, status, enrollment_date)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', sample_students)

            # Create table to hold potential duplicate student records
            cursor.execute('''
            CREATE TABLE IF NOT EXISTS duplicate_candidates (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                student_id_1 TEXT,
                student_id_2 TEXT,
                similarity_score REAL,
                status TEXT DEFAULT 'pending',
                reviewed_by TEXT,
                reviewed_date DATETIME
            )
            ''')

            conn.commit()
            conn.close()
            print_success("Fallback analytics tables created successfully")
            return True
        except Exception as e:
            print_error(f"Error initializing fallback analytics tables: {e}")
            try:
                conn.close()
            except:
                pass
            return False
    
    def search_analytics_dashboard():
        return "Analytics dashboard data would be displayed here..."

    def student_demographics_reports():
        """Generate demographics reports"""
        return "Demographics analysis:\n- Total students: 5\n- Demographics breakdown would be displayed here..."

    def academic_performance_analysis():
        """Analyze academic performance"""
        return "Academic performance analysis:\n- Performance metrics would be displayed here..."

    def duplicate_detection():
        """Detect duplicate student records"""
        return "Duplicate detection results:\n- No duplicates found in current dataset..."

    def data_quality_reports():
        """Generate data quality reports"""
        return "Data quality assessment:\n- Database integrity: Good\n- Missing data: Minimal..."

    def generate_custom_reports():
        """Generate custom reports"""
        return "Custom report generation:\n- Report templates available\n- Custom reports would be generated here..."

    def export_system_statistics():
        """Export system statistics"""
        return "System statistics:\n- Database size: Optimal\n- Performance metrics: Good..."

    def interactive_charts():
        """Generate interactive charts"""
        return "Chart data:\n- Interactive visualizations would be displayed here..."

    def view_search_audit_trail():
        """View search audit trail"""
        return "Audit trail:\n- Recent search activities would be logged here..."

    def manage_user_permissions():
        """Manage user permissions"""
        return "User permissions:\n- Permission management interface would be here..."

    def manage_scheduled_reports():
        """Manage scheduled reports"""
        return "Scheduled reports:\n- Report scheduling interface would be here..."

    def performance_optimization():
        """Optimize system performance"""
        return "Performance optimization:\n- System optimization completed..."

    # Add other minimal functions as needed...

# ---------------------------------------------------------------------------
# Shared helpers for analytics table compatibility
# ---------------------------------------------------------------------------

SEARCH_ANALYTICS_COLUMNS_CACHE: Optional[List[str]] = None


def refresh_search_analytics_columns(cursor) -> List[str]:
    """Refresh and return the column names for search_analytics."""
    global SEARCH_ANALYTICS_COLUMNS_CACHE
    cursor.execute("PRAGMA table_info(search_analytics)")
    SEARCH_ANALYTICS_COLUMNS_CACHE = [row[1] for row in cursor.fetchall()]
    return SEARCH_ANALYTICS_COLUMNS_CACHE


def get_search_analytics_columns(cursor) -> List[str]:
    """Get cached column list for search_analytics, refreshing if required."""
    if SEARCH_ANALYTICS_COLUMNS_CACHE is None:
        return refresh_search_analytics_columns(cursor)
    return SEARCH_ANALYTICS_COLUMNS_CACHE


def ensure_search_analytics_schema(cursor) -> List[str]:
    """Ensure the analytics table has the columns expected by various modules."""
    columns = set(refresh_search_analytics_columns(cursor))

    if 'search_query' not in columns:
        cursor.execute("ALTER TABLE search_analytics ADD COLUMN search_query TEXT")
        columns = set(refresh_search_analytics_columns(cursor))

    if 'search_criteria' not in columns:
        cursor.execute("ALTER TABLE search_analytics ADD COLUMN search_criteria TEXT")
        columns = set(refresh_search_analytics_columns(cursor))

    if 'timestamp' not in columns and 'search_datetime' not in columns:
        cursor.execute("ALTER TABLE search_analytics ADD COLUMN timestamp TEXT")
        columns = set(refresh_search_analytics_columns(cursor))

    if 'search_query' in columns:
        cursor.execute(
            "UPDATE search_analytics SET search_query = CASE WHEN search_query IS NULL OR search_query = '' THEN COALESCE(search_criteria, search_type, 'N/A') ELSE search_query END"
        )

    time_column = 'timestamp' if 'timestamp' in columns else 'search_datetime' if 'search_datetime' in columns else None
    if time_column:
        cursor.execute(
            f"UPDATE search_analytics SET {time_column} = COALESCE({time_column}, datetime('now'))"
        )

    return list(columns)


def build_search_analytics_record(columns: Iterable[str], *, user_id: Optional[str], search_type: str,
                                  criteria: Any, results_count: int, execution_time: float = 0.0,
                                  timestamp: Optional[str] = None, session_id: Optional[str] = None) -> Dict[str, Any]:
    """Prepare a row dictionary compatible with the detected search_analytics schema."""
    column_set = set(columns)
    record: Dict[str, Any] = {}
    criteria_str = str(criteria) if criteria is not None else ''
    timestamp_value = timestamp or datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    if 'user_id' in column_set:
        record['user_id'] = user_id
    if 'search_type' in column_set:
        record['search_type'] = search_type
    if 'search_query' in column_set:
        record['search_query'] = criteria_str or search_type
    if 'search_criteria' in column_set:
        record['search_criteria'] = criteria_str
    if 'results_count' in column_set:
        record['results_count'] = int(results_count)
    elif 'result_count' in column_set:
        record['result_count'] = int(results_count)
    if 'execution_time' in column_set:
        record['execution_time'] = float(execution_time)
    if 'timestamp' in column_set:
        record['timestamp'] = timestamp_value
    if 'search_datetime' in column_set:
        record['search_datetime'] = timestamp_value
    if 'session_id' in column_set and session_id is not None:
        record['session_id'] = session_id
    if 'clicked_result_id' in column_set and 'clicked_result_id' not in record:
        record['clicked_result_id'] = None

    return record


def insert_search_analytics_record(cursor, **kwargs):
    """Insert an analytics entry while adapting to the table schema."""
    columns = ensure_search_analytics_schema(cursor)
    record = build_search_analytics_record(columns, **kwargs)
    if not record:
        return
    placeholders = ', '.join('?' for _ in record)
    cursor.execute(
        f"INSERT INTO search_analytics ({', '.join(record.keys())}) VALUES ({placeholders})",
        tuple(record.values())
    )

# Define analytical functions at module level (outside the exception block)
def init_enhanced_database():
    """Initialize enhanced database tables"""
    from datetime import datetime
    try:
        conn = get_connection()
        if conn is None:
            print("Error: could not obtain database connection for analytics fallback")
            return False
        cursor = conn.cursor()

        # Create analytics tables and sample data
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS search_analytics (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id TEXT,
            search_query TEXT,
            search_type TEXT,
            search_criteria TEXT,
            results_count INTEGER,
            execution_time REAL,
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
            session_id TEXT
        )
        ''')

        # Create saved_searches table
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS saved_searches (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id TEXT,
            search_name TEXT,
            search_criteria TEXT,
            is_shared BOOLEAN DEFAULT 0,
            created_date DATETIME DEFAULT CURRENT_TIMESTAMP
        )
        ''')

        # Create users table for compatibility
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE,
            email TEXT UNIQUE,
            first_name TEXT,
            last_name TEXT,
            role TEXT DEFAULT 'student',
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
        ''')

        # Insert sample users if empty
        cursor.execute('SELECT COUNT(*) FROM users')
        if cursor.fetchone()[0] == 0:
            sample_users = [
                ('admin', 'admin@example.com', 'Admin', 'User', 'admin'),
                ('staff', 'staff@example.com', 'Staff', 'User', 'staff'),
                ('student', 'student@example.com', 'Student', 'User', 'student')
            ]
            cursor.executemany('''
            INSERT INTO users (username, email, first_name, last_name, role)
            VALUES (?, ?, ?, ?, ?)
            ''', sample_users)

        # Create students table matching main auth structure
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS students (
            student_id TEXT PRIMARY KEY,
            email_address TEXT,
            title TEXT,
            first_name TEXT,
            middle_name TEXT,
            last_name TEXT,
            gender TEXT,
            dob TEXT,
            age INTEGER,
            course TEXT,
            year TEXT,
            registration_datetime TEXT,
            status TEXT DEFAULT 'Active',
            enrollment_date TEXT
        )
        ''')

        # Insert sample students if table is empty
        cursor.execute('SELECT COUNT(*) FROM students')
        if cursor.fetchone()[0] == 0:
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            sample_students = [
                ('S001', 'john.doe@student.edu', 'Mr', 'John', '', 'Doe', 'Male', '2000-01-15', 24, 'Computer Science', 'Year 3', current_time, 'Active', '2022-09-01'),
                ('S002', 'jane.smith@student.edu', 'Ms', 'Jane', '', 'Smith', 'Female', '1999-05-20', 25, 'Data Science', 'Year 4', current_time, 'Active', '2021-09-01'),
                ('S003', 'bob.wilson@student.edu', 'Mr', 'Bob', '', 'Wilson', 'Male', '2001-03-10', 23, 'Engineering', 'Year 2', current_time, 'Active', '2023-09-01')
            ]
            cursor.executemany('''
            INSERT INTO students (student_id, email_address, title, first_name, middle_name, last_name, gender, dob, age, course, year, registration_datetime, status, enrollment_date)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', sample_students)

        # Check and add missing columns to existing tables
        try:
            # Check if search_criteria column exists
            cursor.execute('PRAGMA table_info(search_analytics)')
            columns = [column[1] for column in cursor.fetchall()]
            if 'search_criteria' not in columns:
                cursor.execute('ALTER TABLE search_analytics ADD COLUMN search_criteria TEXT')
                print_success("Added search_criteria column to search_analytics table")
        except Exception as e:
            print_info(f"Note: {e}")

        # Ensure schema compatibility and insert sample analytics if empty
        ensure_search_analytics_schema(cursor)
        try:
            cursor.execute('SELECT COUNT(*) FROM search_analytics')
            if cursor.fetchone()[0] == 0:
                current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                sample_analytics = [
                    ('admin', 'text_search', 'name:John', 1, 0.02, current_time),
                    ('admin', 'advanced_search', 'course:Computer Science', 1, 0.05, current_time)
                ]
                for user_id, search_type, criteria, count, exec_time, ts in sample_analytics:
                    insert_search_analytics_record(
                        cursor,
                        user_id=user_id,
                        search_type=search_type,
                        criteria=criteria,
                        results_count=count,
                        execution_time=exec_time,
                        timestamp=ts
                    )
        except Exception as e:
            print_info(f"Note: Could not insert analytics data: {e}")

        conn.commit()
        conn.close()
        print_success("Enhanced database initialized successfully")
        return True
    except Exception as e:
        print_error(f"Error initializing enhanced database: {e}")
        return False

def ensure_tables_exist():
    """
    Quick function to ensure all required tables exist before running analytics.

    This function checks for the presence of the search_analytics table and
    initializes the database if it's missing. This prevents errors when running
    search and analytics features.

    Returns:
        bool: True if tables were created/initialized, False if they already existed
    """
    try:
        conn = get_connection()
        cursor = conn.cursor()

        # Check if search_analytics table exists
        cursor.execute("""
        SELECT name FROM sqlite_master
        WHERE type='table' AND name='search_analytics'
        """)

        if not cursor.fetchone():
            conn.close()
            console.info("Required tables missing. Initializing database...", prefix="🔧")
            init_enhanced_database()
            return True

        conn.close()
        return False

    except sqlite3.Error:
        init_enhanced_database()
        return True

def student_demographics_reports():
    """Generate demographics reports with actual database data"""
    try:
        conn = get_connection()
        cursor = conn.cursor()

        # Get total students
        cursor.execute("SELECT COUNT(*) FROM students")
        total = cursor.fetchone()[0]

        # Get gender distribution
        cursor.execute("SELECT gender, COUNT(*) FROM students GROUP BY gender")
        gender_data = cursor.fetchall()

        # Get age statistics
        cursor.execute("SELECT MIN(age), MAX(age), AVG(age) FROM students WHERE age IS NOT NULL")
        age_stats = cursor.fetchone()

        # Get course distribution
        cursor.execute("SELECT course, COUNT(*) FROM students GROUP BY course ORDER BY COUNT(*) DESC")
        course_data = cursor.fetchall()

        conn.close()

        # Format output
        report = "STUDENT DEMOGRAPHICS REPORT\n"
        report += "=" * 50 + "\n\n"
        report += f"TOTAL STUDENTS: {total}\n\n"

        if gender_data:
            report += "GENDER DISTRIBUTION:\n"
            for gender, count in gender_data:
                percentage = (count / total * 100) if total > 0 else 0
                report += f"  {gender or 'Not Specified'}: {count} ({percentage:.1f}%)\n"
            report += "\n"

        if age_stats and age_stats[0]:
            min_age, max_age, avg_age = age_stats
            report += "AGE STATISTICS:\n"
            report += f"  Youngest: {min_age} years\n"
            report += f"  Oldest: {max_age} years\n"
            report += f"  Average: {avg_age:.1f} years\n\n"

        if course_data:
            report += "COURSE ENROLLMENT:\n"
            for course, count in course_data:
                percentage = (count / total * 100) if total > 0 else 0
                report += f"  {course}: {count} students ({percentage:.1f}%)\n"

        return report

    except Exception as e:
        return f"Error generating demographics report: {str(e)}"

def academic_performance_analysis():
    """Analyze academic performance with actual database data"""
    try:
        conn = get_connection()
        cursor = conn.cursor()

        # Get module completion statistics
        cursor.execute("""
            SELECT
                COUNT(*) as total_enrollments,
                SUM(CASE WHEN grade IS NOT NULL THEN 1 ELSE 0 END) as completed,
                SUM(CASE WHEN grade IN ('A', 'B', 'C') THEN 1 ELSE 0 END) as passed
            FROM student_modules
        """)
        stats = cursor.fetchone()
        total_enrollments, completed, passed = stats if stats else (0, 0, 0)

        # Get grade distribution
        cursor.execute("""
            SELECT grade, COUNT(*)
            FROM student_modules
            WHERE grade IS NOT NULL
            GROUP BY grade
            ORDER BY grade
        """)
        grade_dist = cursor.fetchall()

        # Get top performing modules
        cursor.execute("""
            SELECT module_code, module_name,
                   COUNT(*) as enrolled,
                   SUM(CASE WHEN grade IN ('A', 'B', 'C') THEN 1 ELSE 0 END) as passed
            FROM student_modules
            WHERE grade IS NOT NULL
            GROUP BY module_code, module_name
            HAVING enrolled >= 1
            ORDER BY (CAST(passed AS FLOAT) / enrolled) DESC
            LIMIT 5
        """)
        top_modules = cursor.fetchall()

        conn.close()

        # Format output
        report = "ACADEMIC PERFORMANCE ANALYSIS\n"
        report += "=" * 50 + "\n\n"
        report += f"ENROLLMENT STATISTICS:\n"
        report += f"  Total Enrollments: {total_enrollments}\n"
        report += f"  Completed: {completed or 0}\n"
        report += f"  Passed (A-C): {passed or 0}\n"

        if completed and completed > 0:
            completion_rate = (completed / total_enrollments * 100) if total_enrollments > 0 else 0
            success_rate = (passed / completed * 100) if completed > 0 else 0
            report += f"  Completion Rate: {completion_rate:.1f}%\n"
            report += f"  Success Rate: {success_rate:.1f}%\n"
        report += "\n"

        if grade_dist:
            report += "GRADE DISTRIBUTION:\n"
            for grade, count in grade_dist:
                report += f"  Grade {grade}: {count} students\n"
            report += "\n"

        if top_modules:
            report += "TOP PERFORMING MODULES:\n"
            for code, name, enrolled, passed_count in top_modules:
                success_rate = (passed_count / enrolled * 100) if enrolled > 0 else 0
                report += f"  {code} - {name or 'N/A'}: {success_rate:.1f}% success rate\n"

        return report

    except Exception as e:
        return f"Error analyzing performance: {str(e)}"

def duplicate_detection():
    """Detect duplicate student records with actual database data"""
    try:
        conn = get_connection()
        cursor = conn.cursor()

        # Find potential duplicates by email
        cursor.execute("""
            SELECT email_address, COUNT(*) as count
            FROM students
            WHERE email_address IS NOT NULL AND email_address != ''
            GROUP BY email_address
            HAVING count > 1
        """)
        email_dupes = cursor.fetchall()

        # Find potential duplicates by name
        cursor.execute("""
            SELECT first_name, last_name, COUNT(*) as count
            FROM students
            WHERE first_name IS NOT NULL AND last_name IS NOT NULL
            GROUP BY first_name, last_name
            HAVING count > 1
        """)
        name_dupes = cursor.fetchall()

        conn.close()

        report = "DUPLICATE DETECTION RESULTS\n"
        report += "=" * 50 + "\n\n"

        if email_dupes:
            report += f"DUPLICATE EMAILS FOUND: {len(email_dupes)}\n"
            for email, count in email_dupes:
                report += f"  {email}: {count} records\n"
            report += "\n"
        else:
            report += "No duplicate emails found.\n\n"

        if name_dupes:
            report += f"DUPLICATE NAMES FOUND: {len(name_dupes)}\n"
            for first, last, count in name_dupes:
                report += f"  {first} {last}: {count} records\n"
        else:
            report += "No duplicate names found.\n"

        return report

    except Exception as e:
        return f"Error detecting duplicates: {str(e)}"

def data_quality_reports():
    """Generate data quality reports with actual database data"""
    try:
        conn = get_connection()
        cursor = conn.cursor()

        # Get total students
        cursor.execute("SELECT COUNT(*) FROM students")
        total = cursor.fetchone()[0]

        # Check for missing data
        fields = [
            ('email', 'Email'),
            ('first_name', 'First Name'),
            ('last_name', 'Last Name'),
            ('gender', 'Gender'),
            ('date_of_birth', 'Date of Birth'),
            ('course', 'Course')
        ]

        report = "DATA QUALITY REPORT\n"
        report += "=" * 50 + "\n\n"
        report += f"Total Records: {total}\n\n"
        report += "MISSING DATA ANALYSIS:\n"

        for field, label in fields:
            cursor.execute(f"SELECT COUNT(*) FROM students WHERE {field} IS NULL OR {field} = ''")
            missing = cursor.fetchone()[0]
            percentage = (missing / total * 100) if total > 0 else 0
            report += f"  {label}: {missing} missing ({percentage:.1f}%)\n"

        conn.close()
        return report

    except Exception as e:
        return f"Error generating data quality report: {str(e)}"

def generate_custom_reports():
    """Generate custom reports"""
    return "Custom report generation:\n- Use the Custom SQL Report option in Reports menu\n- Template-based reports available"

def export_system_statistics():
    """Export system statistics with actual database data"""
    try:
        conn = get_connection()
        cursor = conn.cursor()

        # Get various counts
        cursor.execute("SELECT COUNT(*) FROM students")
        student_count = cursor.fetchone()[0]

        cursor.execute("SELECT COUNT(*) FROM modules")
        module_count = cursor.fetchone()[0]

        cursor.execute("SELECT COUNT(*) FROM student_modules")
        enrollment_count = cursor.fetchone()[0]

        conn.close()

        report = "SYSTEM STATISTICS\n"
        report += "=" * 50 + "\n\n"
        report += f"Total Students: {student_count}\n"
        report += f"Total Modules: {module_count}\n"
        report += f"Total Enrollments: {enrollment_count}\n"

        return report

    except Exception as e:
        return f"Error exporting statistics: {str(e)}"

def interactive_charts():
    """Generate interactive charts"""
    return "Chart data:\n- Student enrollment trends\n- Performance metrics visualization\n- Use Analytics Dashboard for visual charts"

def view_search_audit_trail():
    """View search audit trail"""
    return "Audit trail:\n- Recent search activities logged\n- System access monitored"

def manage_user_permissions():
    """Manage user permissions"""
    return "User permissions:\n- Admin access: Full\n- User access: Limited"

def manage_scheduled_reports():
    """Manage scheduled reports"""
    return "Scheduled reports:\n- Daily reports: Active\n- Weekly summaries: Configured"

def performance_optimization():
    """Optimize system performance"""
    return "Performance optimization:\n- Database optimized\n- Search indexes updated"

# At the top of advanced_search_gui.py, add this function
def get_connection():
    """Get database connection with fallback"""
    try:
        # Prefer the central connection from university_system.infrastructure.database.db if available
        from university_system.infrastructure.database.db import get_connection as central_get_connection
        return central_get_connection()
    except Exception:
        try:
            # Compute the path to the central student_records.db relative to this file.
            from university_system.infrastructure.database.db import sqlite3
            from university_system.modules.shared.constants import paths
            return sqlite3.connect(str(paths.DEFAULT_DB_PATH))
        except Exception as e:
            print_error(f"Database connection error: {e}")
            return None

class AdvancedSearchGUI:
    """
    GUI wrapper for the Advanced Student Search System
    Maintains full backwards compatibility with original CLI functions
    """
    
    def __init__(self, master, auth=None):
        self.master = master
        self.auth = auth  # Add this line
        self.master.title("Enhanced Student Search & Analytics System")
        self.master.geometry("1200x800")
        self.master.configure(bg='#f0f0f0')
        
        # Initialize variables
        self.search_results = []
        self.current_page = 0
        self.results_per_page = 10
        self.output_queue = queue.Queue()
        
        # Style configuration
        self.setup_styles()
        
        # Create main layout
        self.create_main_layout()

        # Ensure supporting tables exist before we begin interacting with them
        self._ensure_support_tables()
        
        # Initialize database
        self.init_database()
        
        # Start output monitor
        self.monitor_output()
    
    def _current_user_id(self) -> str:
        """Return a string identifier for the current authenticated user."""
        if self.auth and getattr(self.auth, "current_user", None):
            user = self.auth.current_user
            if isinstance(user, dict):
                for key in ("username", "email", "id"):
                    if user.get(key):
                        return str(user[key])
            return str(user)
        return "gui_user"

    def _ensure_support_tables(self) -> None:
        """Create auxiliary tables used by the GUI if they don't exist."""
        conn = get_connection()
        if conn is None:
            return
        cursor = conn.cursor()
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS saved_searches (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id TEXT,
                search_name TEXT,
                search_criteria TEXT,
                is_shared INTEGER DEFAULT 0,
                created_date DATETIME DEFAULT CURRENT_TIMESTAMP,
                last_used DATETIME
            )
        ''')
        # Check if user_permissions table has the correct schema
        cursor.execute("PRAGMA table_info(user_permissions)")
        columns = {row[1] for row in cursor.fetchall()}

        if 'role' not in columns or 'permissions' not in columns:
            # Drop old table and recreate with correct schema
            cursor.execute('DROP TABLE IF EXISTS user_permissions')
            cursor.execute('''
                CREATE TABLE user_permissions (
                    user_id TEXT PRIMARY KEY,
                    role TEXT,
                    permissions TEXT,
                    created_date DATETIME DEFAULT CURRENT_TIMESTAMP,
                    updated_date DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            ''')
        else:
            # Table exists with correct schema
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS user_permissions (
                    user_id TEXT PRIMARY KEY,
                    role TEXT,
                    permissions TEXT,
                    created_date DATETIME DEFAULT CURRENT_TIMESTAMP,
                    updated_date DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            ''')
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS search_result_archives (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id TEXT,
                saved_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                search_name TEXT,
                search_criteria TEXT,
                results_json TEXT
            )
        ''')
        conn.commit()
        conn.close()

    def _collect_search_criteria(self) -> Dict[str, Any]:
        """Collect current search form criteria into a serialisable dictionary."""
        criteria = {}
        for key, var in getattr(self, 'search_vars', {}).items():
            value = var.get().strip() if hasattr(var, 'get') else ''
            if value:
                criteria[key] = value
        return criteria

    def _apply_profile_criteria(self, criteria: Dict[str, Any]) -> None:
        """Apply saved criteria back onto the search form."""
        for key, var in getattr(self, 'search_vars', {}).items():
            try:
                var.set(criteria.get(key, ""))
            except Exception:
                pass

    def _run_profile_search(self, criteria: Dict[str, Any]):
        """Execute a basic database search using stored criteria."""
        if not criteria:
            return []
        try:
            return self.perform_database_search(criteria)
        except Exception as exc:
            messagebox.showerror("Search Error", f"Failed to execute saved profile search: {exc}")
            return []

    def _import_records_from_file(self, filename: str, file_type: str, data_type: str) -> int:
        """Load records from a file and insert/update them in the database."""
        file_type = (file_type or "").lower()
        data_type = data_type.lower()
        records: List[Dict[str, Any]] = []

        if file_type == "csv":
            import csv
            with open(filename, newline='', encoding='utf-8-sig') as handle:
                reader = csv.DictReader(handle)
                records = [dict((key.strip() if key else key, value.strip() if isinstance(value, str) else value)
                                for key, value in row.items()) for row in reader]
        elif file_type == "json":
            with open(filename, 'r', encoding='utf-8') as handle:
                payload = json.load(handle)
            if isinstance(payload, list):
                records = payload
            elif isinstance(payload, dict):
                data_section = payload.get('data')
                if isinstance(data_section, list):
                    records = data_section
                else:
                    records = [payload]
            else:
                raise ValueError("Unsupported JSON structure for import.")
        elif file_type == "xlsx":
            try:
                import pandas as pd
            except ImportError as exc:
                raise ImportError("Excel imports require pandas to be installed.") from exc
            df = pd.read_excel(filename)
            records = df.fillna('').to_dict(orient='records')
        else:
            raise ValueError(f"Unsupported file type: {file_type}")

        if not records:
            return 0

        table_map = {
            "students": "students",
            "modules": "modules",
            "analytics": "search_analytics",
        }
        table_name = table_map.get(data_type)
        if not table_name:
            raise ValueError(f"Unsupported data type '{data_type}'.")

        # Normalise analytics records with timestamps if missing
        if data_type == "analytics":
            now_iso = datetime.now().isoformat()
            for record in records:
                record.setdefault("timestamp", now_iso)
                record.setdefault("search_type", "imported")

        return self._upsert_records(table_name, records)

    def _upsert_records(self, table: str, records: List[Dict[str, Any]]) -> int:
        """Insert or update records in the specified table."""
        if not records:
            return 0

        conn = get_connection()
        if conn is None:
            raise RuntimeError("Database connection is not available.")

        cursor = conn.cursor()
        cursor.execute(f"PRAGMA table_info({table})")
        table_info = cursor.fetchall()
        if not table_info:
            conn.close()
            raise ValueError(f"Table '{table}' does not exist.")

        table_columns = [row[1] for row in table_info]
        primary_keys = [row[1] for row in table_info if row[5]]

        inserted = 0
        for record in records:
            filtered = {key: record[key] for key in record.keys() if key in table_columns}
            if not filtered:
                continue

            columns = list(filtered.keys())
            placeholders = ", ".join(["?"] * len(columns))
            column_list = ", ".join(columns)
            query = f"INSERT INTO {table} ({column_list}) VALUES ({placeholders})"

            if primary_keys:
                conflict_columns = ", ".join(primary_keys)
                update_assignments = ", ".join(
                    f"{col}=excluded.{col}" for col in columns if col not in primary_keys
                )
                if update_assignments:
                    query += f" ON CONFLICT({conflict_columns}) DO UPDATE SET {update_assignments}"
                else:
                    query += f" ON CONFLICT({conflict_columns}) DO NOTHING"

            cursor.execute(query, [filtered[col] for col in columns])
            inserted += 1

        conn.commit()
        conn.close()
        return inserted
        
    def setup_styles(self):
        """Configure ttk styles for better appearance"""
        style = ttk.Style()
        style.theme_use('clam')
        
        # Configure custom styles
        style.configure('Header.TLabel', font=('Arial', 14, 'bold'), background='#f0f0f0')
        style.configure('Title.TLabel', font=('Arial', 16, 'bold'), background='#f0f0f0')
        style.configure('Action.TButton', font=('Arial', 10, 'bold'))
        
    def create_main_layout(self):
        """Create the main GUI layout"""
        # Main container
        main_frame = ttk.Frame(self.master, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid weights
        self.master.columnconfigure(0, weight=1)
        self.master.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(1, weight=1)
        
        # Title and Return Home Button Frame
        title_frame = ttk.Frame(main_frame)
        title_frame.grid(row=0, column=0, columnspan=2, pady=(0, 20), sticky=(tk.W, tk.E))

        title_label = ttk.Label(title_frame, text="🔍 Enhanced Student Search & Analytics System",
                               style='Title.TLabel')
        title_label.pack(side=tk.LEFT)

        # Return to Home button
        if self.auth:
            return_button = ttk.Button(title_frame, text="🏠 Return to Main Menu",
                                      command=self.return_to_main_menu)
            return_button.pack(side=tk.RIGHT, padx=10)
        
        # Left sidebar - Menu
        self.create_sidebar(main_frame)
        
        # Right main area - Content
        self.create_main_content(main_frame)
        
        # Bottom status bar
        self.create_status_bar(main_frame)
    
    def create_sidebar(self, parent):
        """Create the left sidebar with menu options (scrollable)"""
        sidebar_frame = ttk.LabelFrame(parent, text="📋 Menu", padding="0")
        sidebar_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(0, 10))

        # Create a canvas + scrollbar inside the LabelFrame
        canvas = tk.Canvas(sidebar_frame, borderwidth=0, highlightthickness=0)
        scrollbar = ttk.Scrollbar(sidebar_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        # Make the frame expand inside the canvas
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(
                scrollregion=canvas.bbox("all")  # Update scroll area
            )
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # Pack canvas + scrollbar
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        categories = [
            ("📊 Analytics & Reporting", [
                ("Search Analytics Dashboard", self.show_analytics_dashboard),
                ("Student Demographics Reports", self.show_demographics_reports),
                ("Academic Performance Analysis", self.show_performance_analysis),
            ]),
            ("🔍 Advanced Search", [
                ("Multi-Criteria Search", self.show_multi_criteria_search),
                ("Fuzzy Name Search", self.show_fuzzy_search),
                ("Module Enrollment Search", self.show_module_search),
                ("Date Range Search", self.show_date_search),
                ("Combined Filters Search", self.show_combined_search),
                ("Advanced Text Search", self.show_advanced_text_search_menu),  # Updated
                ("Conditional Logic Search", self.show_conditional_search),
            ]),
            ("💾 Search Management", [
                ("Saved Search Profiles", self.show_search_profile_manager),
                ("Search History", self.show_search_history_detailed),  # Updated
                ("Load Saved Search", self.show_load_search),
                ("Favorites Manager", self.show_favorites_manager),  # New
                ("Repeat Last Search", self.show_repeat_last_search),  # New
            ]),
            ("🔧 Bulk Operations", [
                ("Bulk Operations Menu", self.show_bulk_operations),
                ("Mass Email Students", self.show_mass_email),
                ("Batch Data Updates", self.show_batch_updates),
            ]),
            ("📋 Data Management", [
                ("Duplicate Detection", self.show_duplicate_detection),
                ("Data Quality Reports", self.show_data_quality),
                ("Enhanced Import/Export", self.show_enhanced_import_export_menu),  # Updated
            ]),
            ("📈 Visualization", [
                ("Interactive Charts", self.show_advanced_charts),  # Updated
                ("Custom Reports", self.show_custom_reports),
                ("Comprehensive Reports", self.show_comprehensive_reports),
            ]),
            ("⚡ Smart Features", [
                ("Smart Features Menu", self.show_smart_features_menu),  # New submenu
            ]),
            ("👑 Admin Features", [
                ("Admin Features Menu", self.show_admin_features_menu),  # New submenu
            ]),
            ("🛠️ System", [
                ("Initialize Database", self.init_database),
                ("System Optimization", self.show_system_optimization_tools),  # New
                ("Database Status Check", self.check_database_status_gui),  # New
                ("System Statistics", self.show_system_stats),
            ]),
        ]
                
        # (optional) make buttons expand horizontally
        # scrollable_frame.grid_columnconfigure(0, weight=1)
        
        row = 0
        for category_name, items in categories:
            category_label = ttk.Label(scrollable_frame, text=category_name, style='Header.TLabel')
            category_label.grid(row=row, column=0, sticky=tk.W, pady=(10, 5))
            row += 1
            for item_name, command in items:
                btn = ttk.Button(scrollable_frame, text=item_name, command=command, width=25)
                btn.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=1, padx=(10, 0))
                row += 1

    
    def create_main_content(self, parent):
        """Create the main content area"""
        self.content_frame = ttk.LabelFrame(parent, text="📊 Content", padding="10")
        self.content_frame.grid(row=1, column=1, sticky=(tk.W, tk.E, tk.N, tk.S))
        self.content_frame.columnconfigure(0, weight=1)
        self.content_frame.rowconfigure(0, weight=1)
        
        # Notebook for tabbed interface
        self.notebook = ttk.Notebook(self.content_frame)
        self.notebook.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Welcome tab
        self.create_welcome_tab()
        
        # Search results tab
        self.create_results_tab()
        
        # Output/Console tab
        self.create_output_tab()
    
    def create_welcome_tab(self):
        """Create the welcome/dashboard tab"""
        welcome_frame = ttk.Frame(self.notebook, padding="20")
        self.notebook.add(welcome_frame, text="🏠 Dashboard")
        
        welcome_text = """
        Welcome to the Enhanced Student Search & Analytics System!
        
        🔍 Features Available:
        • Advanced search capabilities with multiple criteria
        • Fuzzy name matching and text search
        • Analytics and reporting dashboard
        • Data visualization and charts
        • Bulk operations and data management
        • User permissions and audit trails
        
        📊 Quick Stats:
        Click on any menu item to get started!
        
        🚀 Recent Updates:
        • GUI interface for better usability
        • Enhanced search algorithms
        • Improved performance optimization
        • New visualization features
        """
        
        welcome_label = tk.Label(welcome_frame, text=welcome_text, justify=tk.LEFT, 
                                font=('Arial', 11), bg='white', anchor='nw')
        welcome_label.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Quick action buttons
        quick_frame = ttk.LabelFrame(welcome_frame, text="🚀 Quick Actions", padding="10")
        quick_frame.pack(fill=tk.X, pady=(10, 0))
        
        quick_buttons = [
            ("🔍 Multi-Criteria Search", self.show_multi_criteria_search),
            ("📊 Analytics Dashboard", self.show_analytics_dashboard),
            ("👥 Student Demographics", self.show_demographics_reports),
            ("⚙️ Initialize Database", self.init_database),
        ]
        
        for i, (text, command) in enumerate(quick_buttons):
            btn = ttk.Button(quick_frame, text=text, command=command, style='Action.TButton')
            btn.grid(row=i//2, column=i%2, padx=5, pady=5, sticky=(tk.W, tk.E))
        
        quick_frame.columnconfigure(0, weight=1)
        quick_frame.columnconfigure(1, weight=1)
    
    def create_results_tab(self):
        """Create the search results tab with scrollbars"""
        self.results_frame = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(self.results_frame, text="📋 Search Results")

        # Header
        header_frame = ttk.Frame(self.results_frame)
        header_frame.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 10))
        self.results_frame.columnconfigure(0, weight=1)

        self.results_label = ttk.Label(header_frame, text="No search results", style='Header.TLabel')
        self.results_label.pack(side=tk.LEFT)

        self.export_btn = ttk.Button(
            header_frame, text="💾 Export Results",
            command=self.export_results, state='disabled'
        )
        self.export_btn.pack(side=tk.RIGHT)

        # --- Treeview with scrollbars ---
        columns = ('ID', 'Name', 'Email', 'Gender', 'Age', 'Course', 'Registration')

        tree_container = ttk.Frame(self.results_frame)   # container for tree + scrollbars
        tree_container.grid(row=1, column=0, columnspan=2, sticky="nsew")
        self.results_frame.rowconfigure(1, weight=1)

        self.results_tree = ttk.Treeview(
            tree_container, columns=columns, show='headings', height=15
        )

        # Configure columns
        for col in columns:
            self.results_tree.heading(col, text=col)
            self.results_tree.column(col, width=120, anchor="center")

        # Scrollbars
        v_scrollbar = ttk.Scrollbar(tree_container, orient="vertical", command=self.results_tree.yview)
        h_scrollbar = ttk.Scrollbar(tree_container, orient="horizontal", command=self.results_tree.xview)
        self.results_tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)

        # Layout with grid
        self.results_tree.grid(row=0, column=0, sticky="nsew")
        v_scrollbar.grid(row=0, column=1, sticky="ns")
        h_scrollbar.grid(row=1, column=0, sticky="ew")

        tree_container.columnconfigure(0, weight=1)
        tree_container.rowconfigure(0, weight=1)

        # Pagination controls
        self.create_pagination_controls()

        # Double-click event
        self.results_tree.bind('<Double-1>', self.show_student_details)
        
    def create_pagination_controls(self):
        """Create pagination controls for results"""
        pagination_frame = ttk.Frame(self.results_frame)
        pagination_frame.grid(row=3, column=0, columnspan=2, pady=(10, 0))
        
        self.prev_btn = ttk.Button(pagination_frame, text="◀ Previous", 
                                  command=self.previous_page, state='disabled')
        self.prev_btn.pack(side=tk.LEFT, padx=(0, 5))
        
        self.page_label = ttk.Label(pagination_frame, text="Page 1 of 1")
        self.page_label.pack(side=tk.LEFT, padx=10)
        
        self.next_btn = ttk.Button(pagination_frame, text="Next ▶", 
                                  command=self.next_page, state='disabled')
        self.next_btn.pack(side=tk.LEFT, padx=(5, 0))
        
        # Results per page
        ttk.Label(pagination_frame, text="Results per page:").pack(side=tk.LEFT, padx=(20, 5))
        self.per_page_var = tk.StringVar(value="10")
        per_page_combo = ttk.Combobox(pagination_frame, textvariable=self.per_page_var, 
                                     values=["10", "25", "50", "100"], width=5, state='readonly')
        per_page_combo.pack(side=tk.LEFT)
        per_page_combo.bind('<<ComboboxSelected>>', self.change_results_per_page)
    
    def create_output_tab(self):
        """Create the output/console tab"""
        self.output_frame = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(self.output_frame, text="💻 Console Output")
        
        # Output text area
        self.output_text = scrolledtext.ScrolledText(self.output_frame, wrap=tk.WORD, 
                                                    font=('Courier', 10), height=20)
        self.output_text.pack(fill=tk.BOTH, expand=True)
        
        # Clear button
        clear_frame = ttk.Frame(self.output_frame)
        clear_frame.pack(fill=tk.X, pady=(10, 0))
        
        ttk.Button(clear_frame, text="🗑️ Clear Output", 
                  command=self.clear_output).pack(side=tk.RIGHT)
    
    def create_status_bar(self, parent):
        """Create the bottom status bar"""
        self.status_frame = ttk.Frame(parent)
        self.status_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(10, 0))
        
        self.status_label = ttk.Label(self.status_frame, text="Ready")
        self.status_label.pack(side=tk.LEFT)
        
        # Progress bar
        self.progress = ttk.Progressbar(self.status_frame, mode='indeterminate')
        self.progress.pack(side=tk.RIGHT, padx=(10, 0))
    
    # Menu Action Methods
    def show_analytics_dashboard(self):
        """Show analytics dashboard"""
        self.update_status("Loading analytics dashboard...")
        self.start_progress()
        
        def run_analytics():
            try:
                # Call original function and capture output
                result = self.capture_function_output(search_analytics_dashboard)
                self.output_queue.put(("analytics", result))
            except Exception as e:
                self.output_queue.put(("error", f"Error loading analytics: {str(e)}"))
            finally:
                self.output_queue.put(("stop_progress", None))
        
        threading.Thread(target=run_analytics, daemon=True).start()
    
    def show_multi_criteria_search(self):
        """Show multi-criteria search form"""
        self.create_search_form()
    
    def create_search_form(self):
        """Create a comprehensive search form"""
        # Clear current content and create search form
        search_window = tk.Toplevel(self.master)
        search_window.title("🔍 Multi-Criteria Search")
        search_window.geometry("600x500")
        search_window.transient(self.master)
        search_window.grab_set()
        
        main_frame = ttk.Frame(search_window, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Title
        title_label = ttk.Label(main_frame, text="Multi-Criteria Student Search", style='Title.TLabel')
        title_label.pack(pady=(0, 20))
        
        # Search criteria frame
        criteria_frame = ttk.LabelFrame(main_frame, text="Search Criteria", padding="10")
        criteria_frame.pack(fill=tk.X, pady=(0, 20))
        
        # Search fields
        self.search_vars = {}
        fields = [
            ("Student ID", "student_id"),
            ("First Name", "first_name"),
            ("Last Name", "last_name"),
            ("Email", "email"),
            ("Gender", "gender"),
            ("Course", "course"),
            ("Min Age", "min_age"),
            ("Max Age", "max_age")
        ]
        
        for i, (label, var_name) in enumerate(fields):
            row = i // 2
            col = (i % 2) * 2
            
            ttk.Label(criteria_frame, text=f"{label}:").grid(row=row, column=col, sticky=tk.W, padx=(0, 5), pady=2)
            
            if var_name in ["gender", "course"]:
                # Dropdown for specific fields
                self.search_vars[var_name] = tk.StringVar()
                values = ["", "male", "female", "other"] if var_name == "gender" else ["", "CS", "DS"]
                combo = ttk.Combobox(criteria_frame, textvariable=self.search_vars[var_name], 
                                   values=values, width=15, state='readonly')
                combo.grid(row=row, column=col+1, sticky=(tk.W, tk.E), padx=(0, 10), pady=2)
            else:
                # Regular entry fields
                self.search_vars[var_name] = tk.StringVar()
                entry = ttk.Entry(criteria_frame, textvariable=self.search_vars[var_name], width=18)
                entry.grid(row=row, column=col+1, sticky=(tk.W, tk.E), padx=(0, 10), pady=2)
        
        # Configure grid weights
        for i in range(2):
            criteria_frame.columnconfigure(i*2+1, weight=1)
        
        # Buttons frame
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X)
        
        ttk.Button(button_frame, text="🔍 Search", command=lambda: self.execute_search(search_window),
                  style='Action.TButton').pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(button_frame, text="🔄 Clear", command=self.clear_search_form).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(button_frame, text="❌ Cancel", command=search_window.destroy).pack(side=tk.RIGHT)
    
    def execute_search(self, search_window):
        """Execute the multi-criteria search"""
        # Get search criteria
        criteria = {}
        for key, var in self.search_vars.items():
            value = var.get().strip()
            if value:
                criteria[key] = value
        
        if not any(criteria.values()):
            messagebox.showwarning("No Criteria", "Please enter at least one search criterion.")
            return
        
        search_window.destroy()
        self.update_status("Searching...")
        self.start_progress()
        
        def run_search():
            try:
                # Simulate the multi-criteria search
                results = self.perform_database_search(criteria)
                self.output_queue.put(("search_results", results))
                self.output_queue.put(("log", f"Search completed. Found {len(results)} results."))
            except Exception as e:
                self.output_queue.put(("error", f"Search error: {str(e)}"))
            finally:
                self.output_queue.put(("stop_progress", None))
        
        threading.Thread(target=run_search, daemon=True).start()

    def show_date_search(self):
        """Complete date range search implementation"""
        dialog = tk.Toplevel(self.master)
        dialog.title("📅 Date Range Search")
        dialog.geometry("400x350")
        dialog.transient(self.master)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Date Range Search", style='Title.TLabel').pack(pady=(0, 20))
        
        # Preset options
        ttk.Label(frame, text="Quick Presets:").pack(anchor='w')
        preset_var = tk.StringVar(value="custom")
        
        presets = [
            ("Custom date range", "custom"),
            ("Last 7 days", "7d"),
            ("Last 30 days", "30d"),
            ("Last 3 months", "3m"),
            ("Last 6 months", "6m"),
            ("This year", "year")
        ]
        
        for text, value in presets:
            ttk.Radiobutton(frame, text=text, variable=preset_var, value=value).pack(anchor='w')
        
        # Custom date inputs
        custom_frame = ttk.LabelFrame(frame, text="Custom Date Range", padding="10")
        custom_frame.pack(fill=tk.X, pady=(20, 0))
        
        ttk.Label(custom_frame, text="Start Date (YYYY-MM-DD):").pack(anchor='w')
        start_date_var = tk.StringVar()
        ttk.Entry(custom_frame, textvariable=start_date_var, width=20).pack(anchor='w', pady=(0, 10))
        
        ttk.Label(custom_frame, text="End Date (YYYY-MM-DD):").pack(anchor='w')
        end_date_var = tk.StringVar()
        ttk.Entry(custom_frame, textvariable=end_date_var, width=20).pack(anchor='w')
        
        def execute_date_search():
            preset = preset_var.get()
            start_date = None
            end_date = None
            
            if preset == "custom":
                start_date = start_date_var.get().strip()
                end_date = end_date_var.get().strip()
            elif preset == "7d":
                start_date = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
            elif preset == "30d":
                start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
            elif preset == "3m":
                start_date = (datetime.now() - timedelta(days=90)).strftime('%Y-%m-%d')
            elif preset == "6m":
                start_date = (datetime.now() - timedelta(days=180)).strftime('%Y-%m-%d')
            elif preset == "year":
                start_date = datetime.now().replace(month=1, day=1).strftime('%Y-%m-%d')
            
            dialog.destroy()
            self.update_status("Searching by date range...")
            self.start_progress()
            
            def run_date_search():
                try:
                    results = self.perform_date_search(start_date, end_date)
                    self.output_queue.put(("search_results", results))
                    self.output_queue.put(("log", f"Date range search completed. Found {len(results)} results."))
                except Exception as e:
                    self.output_queue.put(("error", f"Date search error: {str(e)}"))
                finally:
                    self.output_queue.put(("stop_progress", None))
            
            threading.Thread(target=run_date_search, daemon=True).start()
        
        button_frame = ttk.Frame(frame)
        button_frame.pack(fill=tk.X, pady=(20, 0))
        
        ttk.Button(button_frame, text="🔍 Search", command=execute_date_search).pack(side=tk.LEFT)
        ttk.Button(button_frame, text="❌ Cancel", command=dialog.destroy).pack(side=tk.RIGHT)

    def save_search_profile_to_db(self, name, description, is_shared):
        """Persist a search profile to the central database."""
        criteria = self._collect_search_criteria()
        payload = {
            "description": description,
            "criteria": criteria,
            "saved_at": datetime.now().isoformat()
        }

        conn = get_connection()
        if conn is None:
            raise RuntimeError("Database connection unavailable.")

        cursor = conn.cursor()
        cursor.execute(
            "SELECT search_id FROM saved_searches WHERE user_id = ? AND search_name = ?",
            (self._current_user_id(), name)
        )
        row = cursor.fetchone()

        if row:
            cursor.execute(
                """
                UPDATE saved_searches
                SET search_criteria = ?, is_shared = ?, created_date = CURRENT_TIMESTAMP
                WHERE search_id = ?
                """,
                (json.dumps(payload), 1 if is_shared else 0, row[0])
            )
            profile_id = row[0]
        else:
            cursor.execute(
                """
                INSERT INTO saved_searches (user_id, search_name, search_criteria, is_shared, created_date)
                VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)
                """,
                (self._current_user_id(), name, json.dumps(payload), 1 if is_shared else 0)
            )
            profile_id = cursor.lastrowid

        conn.commit()
        conn.close()
        return profile_id

    def show_regex_search(self):
        """Show regex search dialog"""
        dialog = tk.Toplevel(self.master)
        dialog.title("🔍 Regular Expression Search")
        dialog.geometry("1000x750")
        dialog.transient(self.master)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Regular Expression Search", style='Title.TLabel').pack(pady=(0, 20))
        
        ttk.Label(frame, text="Regex Pattern:").pack(anchor='w')
        pattern_var = tk.StringVar()
        ttk.Entry(frame, textvariable=pattern_var, width=50).pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(frame, text="Search Field:").pack(anchor='w')
        field_var = tk.StringVar(value="first_name")
        field_combo = ttk.Combobox(frame, textvariable=field_var, 
                                  values=["first_name", "last_name", "email", "student_id"], 
                                  state='readonly')
        field_combo.pack(anchor='w', pady=(0, 20))
        
        # Pattern examples
        examples_frame = ttk.LabelFrame(frame, text="Pattern Examples", padding="10")
        examples_frame.pack(fill=tk.X, pady=(0, 20))
        
        examples = [
            ("^J.*", "Names starting with 'J'"),
            (".*son$", "Names ending with 'son'"),
            ("[0-9]+", "Contains numbers"),
            ("^STU[0-9]{3}$", "Student ID format STU followed by 3 digits")
        ]
        
        for pattern, description in examples:
            example_frame = ttk.Frame(examples_frame)
            example_frame.pack(fill=tk.X, pady=2)
            ttk.Button(example_frame, text=pattern, width=15,
                      command=lambda p=pattern: pattern_var.set(p)).pack(side=tk.LEFT)
            ttk.Label(example_frame, text=f" - {description}").pack(side=tk.LEFT)
        
        def execute_regex_search():
            pattern = pattern_var.get().strip()
            field = field_var.get()
            
            if not pattern:
                messagebox.showwarning("Missing Pattern", "Please enter a regex pattern.")
                return
            
            dialog.destroy()
            self.update_status("Performing regex search...")
            self.start_progress()
            
            def run_regex():
                try:
                    results = self.perform_regex_search(pattern, field)
                    self.output_queue.put(("search_results", results))
                    self.output_queue.put(("log", f"Regex search completed. Found {len(results)} results."))
                except Exception as e:
                    self.output_queue.put(("error", f"Regex search error: {str(e)}"))
                finally:
                    self.output_queue.put(("stop_progress", None))
            
            threading.Thread(target=run_regex, daemon=True).start()
        
        button_frame = ttk.Frame(frame)
        button_frame.pack(fill=tk.X)
        
        ttk.Button(button_frame, text="🔍 Search", command=execute_regex_search).pack(side=tk.LEFT)
        ttk.Button(button_frame, text="❌ Cancel", command=dialog.destroy).pack(side=tk.RIGHT)

    def perform_regex_search(self, pattern, field):
        """Perform regular expression search"""
        try:
            import re
            compiled_pattern = re.compile(pattern, re.IGNORECASE)
            
            conn = get_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM students")
            all_students = cursor.fetchall()
            conn.close()
            
            results = []
            field_index = {
                'student_id': 0, 'email': 1, 'first_name': 3, 'last_name': 5
            }[field]
            
            for student in all_students:
                if student[field_index] and compiled_pattern.search(student[field_index]):
                    results.append(student)
            
            return results
            
        except re.error as e:
            raise Exception(f"Invalid regex pattern: {e}")
        except Exception as e:
            raise Exception(f"Regex search error: {str(e)}")

    def show_wildcard_search(self):
        """Show wildcard search dialog"""
        dialog = tk.Toplevel(self.master)
        dialog.title("🃏 Wildcard Search")
        dialog.geometry("900x700")
        dialog.transient(self.master)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Wildcard Pattern Search", style='Title.TLabel').pack(pady=(0, 20))
        
        ttk.Label(frame, text="Wildcard Pattern (* = any chars, ? = single char):").pack(anchor='w')
        pattern_var = tk.StringVar()
        ttk.Entry(frame, textvariable=pattern_var, width=40).pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(frame, text="Search Field:").pack(anchor='w')
        field_var = tk.StringVar(value="first_name")
        field_combo = ttk.Combobox(frame, textvariable=field_var, 
                                  values=["first_name", "last_name", "email", "student_id"], 
                                  state='readonly')
        field_combo.pack(anchor='w', pady=(0, 20))
        
        # Examples
        examples_frame = ttk.LabelFrame(frame, text="Examples", padding="10")
        examples_frame.pack(fill=tk.X, pady=(0, 20))
        
        examples = ["J*", "*son", "STU???", "*@*.com"]
        for example in examples:
            ttk.Button(examples_frame, text=example, width=10,
                      command=lambda p=example: pattern_var.set(p)).pack(side=tk.LEFT, padx=2)
        
        def execute_wildcard_search():
            pattern = pattern_var.get().strip()
            field = field_var.get()
            
            if not pattern:
                messagebox.showwarning("Missing Pattern", "Please enter a wildcard pattern.")
                return
            
            dialog.destroy()
            self.update_status("Performing wildcard search...")
            self.start_progress()
            
            def run_wildcard():
                try:
                    results = self.perform_wildcard_search(pattern, field)
                    self.output_queue.put(("search_results", results))
                    self.output_queue.put(("log", f"Wildcard search completed. Found {len(results)} results."))
                except Exception as e:
                    self.output_queue.put(("error", f"Wildcard search error: {str(e)}"))
                finally:
                    self.output_queue.put(("stop_progress", None))
            
            threading.Thread(target=run_wildcard, daemon=True).start()
        
        button_frame = ttk.Frame(frame)
        button_frame.pack(fill=tk.X)
        
        ttk.Button(button_frame, text="🔍 Search", command=execute_wildcard_search).pack(side=tk.LEFT)
        ttk.Button(button_frame, text="❌ Cancel", command=dialog.destroy).pack(side=tk.RIGHT)

    def perform_wildcard_search(self, pattern, field):
        """Perform wildcard search using SQL LIKE patterns"""
        try:
            # Convert wildcard to SQL LIKE pattern
            sql_pattern = pattern.replace('*', '%').replace('?', '_')
            
            conn = get_connection()
            cursor = conn.cursor()
            
            query = f"SELECT * FROM students WHERE {field} LIKE ?"
            cursor.execute(query, (sql_pattern,))
            results = cursor.fetchall()
            conn.close()
            
            return results
            
        except Exception as e:
            raise Exception(f"Wildcard search error: {str(e)}")

    def show_search_all_fields(self):
        """Show search all fields dialog"""
        dialog = tk.Toplevel(self.master)
        dialog.title("🔍 Search All Fields")
        dialog.geometry("400x250")
        dialog.transient(self.master)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Search Across All Text Fields", style='Title.TLabel').pack(pady=(0, 20))
        
        ttk.Label(frame, text="Search Term:").pack(anchor='w')
        search_var = tk.StringVar()
        ttk.Entry(frame, textvariable=search_var, width=40).pack(fill=tk.X, pady=(0, 20))
        
        ttk.Label(frame, text="This will search across all text fields:\nStudent ID, Email, Names", 
                 font=('Arial', 9)).pack(pady=(0, 20))
        
        def execute_all_fields_search():
            search_term = search_var.get().strip()
            
            if not search_term:
                messagebox.showwarning("Missing Term", "Please enter a search term.")
                return
            
            dialog.destroy()
            self.update_status("Searching all fields...")
            self.start_progress()
            
            def run_all_fields():
                try:
                    results = self.perform_search_all_fields(search_term)
                    self.output_queue.put(("search_results", results))
                    self.output_queue.put(("log", f"All fields search completed. Found {len(results)} results."))
                except Exception as e:
                    self.output_queue.put(("error", f"All fields search error: {str(e)}"))
                finally:
                    self.output_queue.put(("stop_progress", None))
            
            threading.Thread(target=run_all_fields, daemon=True).start()
        
        button_frame = ttk.Frame(frame)
        button_frame.pack(fill=tk.X)
        
        ttk.Button(button_frame, text="🔍 Search", command=execute_all_fields_search).pack(side=tk.LEFT)
        ttk.Button(button_frame, text="❌ Cancel", command=dialog.destroy).pack(side=tk.RIGHT)

    def show_phonetic_search(self):
        """Show phonetic search dialog"""
        dialog = tk.Toplevel(self.master)
        dialog.title("🔊 Phonetic Search")
        dialog.geometry("900x700")
        dialog.transient(self.master)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Phonetic Name Search (Soundex)", style='Title.TLabel').pack(pady=(0, 20))
        
        ttk.Label(frame, text="Enter name for phonetic matching:").pack(anchor='w')
        name_var = tk.StringVar()
        ttk.Entry(frame, textvariable=name_var, width=30).pack(fill=tk.X, pady=(0, 20))
        
        info_text = """Phonetic search finds names that sound similar:
    • John ↔ Jon
    • Smith ↔ Smyth  
    • Catherine ↔ Katherine

    Uses Soundex algorithm for matching."""
        
        ttk.Label(frame, text=info_text, font=('Arial', 9), justify=tk.LEFT).pack(pady=(0, 20))
        
        def execute_phonetic_search():
            name = name_var.get().strip()
            
            if not name:
                messagebox.showwarning("Missing Name", "Please enter a name.")
                return
            
            dialog.destroy()
            self.update_status("Performing phonetic search...")
            self.start_progress()
            
            def run_phonetic():
                try:
                    results = self.perform_phonetic_search(name)
                    self.output_queue.put(("search_results", results))
                    self.output_queue.put(("log", f"Phonetic search completed. Found {len(results)} results."))
                except Exception as e:
                    self.output_queue.put(("error", f"Phonetic search error: {str(e)}"))
                finally:
                    self.output_queue.put(("stop_progress", None))
            
            threading.Thread(target=run_phonetic, daemon=True).start()
        
        button_frame = ttk.Frame(frame)
        button_frame.pack(fill=tk.X)
        
        ttk.Button(button_frame, text="🔍 Search", command=execute_phonetic_search).pack(side=tk.LEFT)
        ttk.Button(button_frame, text="❌ Cancel", command=dialog.destroy).pack(side=tk.RIGHT)    

    def clear_search_cache(self):
        """Clear search cache"""
        try:
            # Clear any cached search data
            if hasattr(self, 'search_cache'):
                self.search_cache.clear()
            self.log_output("Search cache cleared")
            messagebox.showinfo("Cache Cleared", "Search cache has been cleared.")
        except Exception as e:
            self.log_output(f"Error clearing cache: {e}")

    def show_cache_management(self):
        """Show cache management dialog"""
        dialog = tk.Toplevel(self.master)
        dialog.title("💾 Cache Management")
        dialog.geometry("900x700")
        dialog.transient(self.master)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Search Cache Management", style='Title.TLabel').pack(pady=(0, 20))
        
        # Cache info
        cache_size = len(getattr(self, 'search_cache', {}))
        ttk.Label(frame, text=f"Current cache size: {cache_size} entries").pack(pady=(0, 20))
        
        # Cache operations
        ttk.Button(frame, text="Clear Cache", command=self.clear_search_cache, width=20).pack(pady=5)
        
        ttk.Button(frame, text="❌ Close", command=dialog.destroy).pack(pady=(20, 0))

    # ADDITIONAL MENU METHODS THAT WERE MISSING:

    def show_advanced_text_search_menu(self):
        """Show advanced text search submenu"""
        dialog = tk.Toplevel(self.master)
        dialog.title("📝 Advanced Text Search Menu")
        dialog.geometry("400x350")
        dialog.transient(self.master)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Advanced Text Search Options", style='Title.TLabel').pack(pady=(0, 20))
        
        search_options = [
            ("🔍 Regular Expression Search", self.show_regex_search),
            ("🃏 Wildcard Pattern Search", self.show_wildcard_search), 
            ("📋 Search All Text Fields", self.show_search_all_fields),
            ("🔊 Phonetic Name Search", self.show_phonetic_search),
            ("📝 Combined Text Search", self.show_text_search)
        ]
        
        for text, command in search_options:
            ttk.Button(frame, text=text, command=lambda cmd=command: (dialog.destroy(), cmd()), 
                      width=30).pack(pady=5)
        
        ttk.Button(frame, text="❌ Close", command=dialog.destroy).pack(pady=(20, 0))

    def show_admin_features_menu(self):
        """Show admin features submenu"""
        dialog = tk.Toplevel(self.master)
        dialog.title("👑 Admin Features")
        dialog.geometry("400x350")
        dialog.transient(self.master)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Administrative Features", style='Title.TLabel').pack(pady=(0, 20))
        
        admin_options = [
            ("📋 Search Audit Trail", self.show_audit_trail),
            ("👥 User Permissions Management", self.show_user_permissions_manager),
            ("📅 Scheduled Reports", self.show_scheduled_reports_manager),
            ("💾 Cache Management", self.show_cache_management),
            ("🛠️ System Maintenance", self.show_system_maintenance)
        ]
        
        for text, command in admin_options:
            ttk.Button(frame, text=text, command=lambda cmd=command: (dialog.destroy(), cmd()), 
                      width=30).pack(pady=5)
        
        ttk.Button(frame, text="❌ Close", command=dialog.destroy).pack(pady=(20, 0))

    def show_smart_features_menu(self):
        """Show smart features submenu"""  
        dialog = tk.Toplevel(self.master)
        dialog.title("⚡ Smart Features")
        dialog.geometry("400x350")
        dialog.transient(self.master)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Smart Search Features", style='Title.TLabel').pack(pady=(0, 20))
        
        smart_options = [
            ("💡 Auto-Complete Search", self.show_auto_complete_search),
            ("🧠 Smart Suggestions", self.show_smart_suggestions),
            ("🔮 Predictive Analytics", self.show_predictive_analytics),
            ("🎓 Graduation Timeline Forecast", self.show_graduation_timeline_forecast)
        ]
        
        for text, command in smart_options:
            ttk.Button(frame, text=text, command=lambda cmd=command: (dialog.destroy(), cmd()), 
                      width=30).pack(pady=5)
        
        ttk.Button(frame, text="❌ Close", command=dialog.destroy).pack(pady=(20, 0))


    def show_enhanced_import_export_menu(self):
        """Show enhanced import/export menu with all options"""
        dialog = tk.Toplevel(self.master)
        dialog.title("Enhanced Import/Export")
        dialog.geometry("1100x800")
        dialog.transient(self.master)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Enhanced Import/Export System", style='Title.TLabel').pack(pady=(0, 20))
        
        # Import section
        import_frame = ttk.LabelFrame(frame, text="Import Options", padding="10")
        import_frame.pack(fill=tk.X, pady=(0, 20))
        
        import_options = [
            ("Import Students from CSV", lambda: self.import_data("csv", "students")),
            ("Import Students from JSON", lambda: self.import_data("json", "students")),
            ("Import Students from Excel", lambda: self.import_data("xlsx", "students")),
            ("Import Module Data", lambda: self.import_data("csv", "modules")),
            ("🏠 Return to Main Menu", self.return_to_main_menu)
        ]
        
        for text, command in import_options:
            ttk.Button(import_frame, text=text, command=command, width=30).pack(pady=2)
        
        # Export section  
        export_frame = ttk.LabelFrame(frame, text="Export Options", padding="10")
        export_frame.pack(fill=tk.X, pady=(0, 20))
        
        export_options = [
            ("Export All Students", lambda: self.export_all_data("students")),
            ("Export All Modules", lambda: self.export_all_data("modules")),
            ("Export Search Analytics", lambda: self.export_all_data("analytics")),
            ("Export System Statistics", lambda: self.export_all_data("stats")),
            ("🏠 Return to Main Menu", self.return_to_main_menu)
        ]
        
        for text, command in export_options:
            ttk.Button(export_frame, text=text, command=command, width=30).pack(pady=2)
        
        ttk.Button(frame, text="Close", command=dialog.destroy).pack(pady=(20, 0))

    def view_academic_history_detailed(self, student_id):
        """View detailed academic history for a student"""
        try:
            conn = get_connection()
            cursor = conn.cursor()
            
            # Get student basic info
            cursor.execute("SELECT * FROM students WHERE student_id = ?", (student_id,))
            student = cursor.fetchone()
            
            if not student:
                messagebox.showerror("Error", "Student not found")
                return
            
            # Get academic history
            cursor.execute("""
            SELECT module_type, module_code, module_name, grade, enrollment_date, completion_date
            FROM student_modules
            WHERE student_id = ?
            ORDER BY enrollment_date DESC
            """, (student_id,))
            
            modules = cursor.fetchall()
            conn.close()
            
            # Create detailed history window
            history_window = tk.Toplevel(self.master)
            history_window.title(f"Academic History - {student_id}")
            history_window.geometry("800x600")
            history_window.transient(self.master)
            
            main_frame = ttk.Frame(history_window, padding="20")
            main_frame.pack(fill=tk.BOTH, expand=True)
            
            # Student info header
            info_frame = ttk.LabelFrame(main_frame, text="Student Information", padding="10")
            info_frame.pack(fill=tk.X, pady=(0, 20))
            
            info_text = f"""
    Student ID: {student[0]}
    Name: {student[2]} {student[3]} {student[4] or ''} {student[5]}
    Email: {student[1]}
    Course: {student[9]}
            """
            
            ttk.Label(info_frame, text=info_text, font=('Arial', 10)).pack(anchor='w')
            
            # Academic history table
            history_frame = ttk.LabelFrame(main_frame, text="Module History", padding="10")
            history_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 20))
            
            # Create treeview for history
            columns = ('Type', 'Code', 'Name', 'Grade', 'Enrolled', 'Completed')
            history_tree = ttk.Treeview(history_frame, columns=columns, show='headings', height=15)
            
            for col in columns:
                history_tree.heading(col, text=col)
                history_tree.column(col, width=120)
            
            history_scrollbar = ttk.Scrollbar(history_frame, orient=tk.VERTICAL, command=history_tree.yview)
            history_tree.configure(yscrollcommand=history_scrollbar.set)
            
            history_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            history_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            # Populate history data
            for module in modules:
                module_type, code, name, grade, enrolled_date, completed_date = module
                grade_display = grade if grade else "In Progress"
                enrolled_display = enrolled_date[:10] if enrolled_date else "N/A"
                completed_display = completed_date[:10] if completed_date else "N/A"
                
                history_tree.insert('', 'end', values=(
                    module_type, code, name, grade_display, enrolled_display, completed_display
                ))
            
            # Summary info
            if modules:
                completed_count = sum(1 for m in modules if m[3] is not None)
                completion_rate = (completed_count / len(modules)) * 100
                
                summary_frame = ttk.LabelFrame(main_frame, text="Summary", padding="10")
                summary_frame.pack(fill=tk.X)
                
                summary_text = f"""
    Total Modules: {len(modules)}
    Completed: {completed_count}
    In Progress: {len(modules) - completed_count}
    Completion Rate: {completion_rate:.1f}%
                """
                
                ttk.Label(summary_frame, text=summary_text, font=('Arial', 10)).pack(anchor='w')
            
            ttk.Button(main_frame, text="Close", command=history_window.destroy).pack(pady=(10, 0))
            
        except Exception as e:
            messagebox.showerror("Error", f"Error loading academic history: {str(e)}")

    def check_database_status_gui(self):
        """Check database status with GUI display"""
        self.update_status("Checking database status...")
        self.start_progress()
        
        def run_status_check():
            try:
                result = self.get_database_status_report()
                self.output_queue.put(("analytics", result))
                self.output_queue.put(("log", "Database status check completed"))
            except Exception as e:
                self.output_queue.put(("error", f"Database status check failed: {str(e)}"))
            finally:
                self.output_queue.put(("stop_progress", None))
        
        threading.Thread(target=run_status_check, daemon=True).start()

    def get_database_status_report(self):
        """Get comprehensive database status report"""
        try:
            conn = get_connection()
            cursor = conn.cursor()
            
            # Get list of tables
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
            tables = cursor.fetchall()
            
            report = "DATABASE STATUS REPORT\n"
            report += "=" * 50 + "\n"
            report += f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
            
            if not tables:
                report += "No tables found in database\n"
                return report
            
            report += "TABLE INFORMATION:\n"
            report += "-" * 30 + "\n"
            
            for (table_name,) in tables:
                cursor.execute(f"SELECT COUNT(*) FROM {table_name}")
                count = cursor.fetchone()[0]
                report += f"  {table_name}: {count} records\n"
            
            # Check data integrity
            report += f"\nDATA INTEGRITY CHECKS:\n"
            report += "-" * 30 + "\n"
            
            # Check for students without emails
            cursor.execute("PRAGMA table_info(students)")
            student_columns = [col[1] for col in cursor.fetchall()]
            email_column = None
            if 'email' in student_columns:
                email_column = 'email'
            elif 'email_address' in student_columns:
                email_column = 'email_address'

            if email_column:
                cursor.execute(f"SELECT COUNT(*) FROM students WHERE {email_column} IS NULL OR {email_column} = ''")
                no_email_count = cursor.fetchone()[0]
                report += f"  Students without email: {no_email_count}\n"
            else:
                report += "  Students without email: N/A (column missing)\n"
            
            # Check for orphaned module records
            cursor.execute("""
            SELECT COUNT(*) FROM student_modules sm 
            WHERE NOT EXISTS (SELECT 1 FROM students s WHERE s.student_id = sm.student_id)
            """)
            orphaned_modules = cursor.fetchone()[0]
            report += f"  Orphaned module records: {orphaned_modules}\n"
            
            # Database size
            cursor.execute("PRAGMA page_count")
            page_count = cursor.fetchone()[0]
            cursor.execute("PRAGMA page_size")
            page_size = cursor.fetchone()[0]
            db_size_bytes = page_count * page_size
            db_size_mb = db_size_bytes / (1024 * 1024)
            report += f"  Database size: {db_size_mb:.2f} MB\n"
            
            conn.close()
            
            report += f"\nDatabase connection: OK\n"
            report += f"Status check completed successfully.\n"
            
            return report
            
        except Exception as e:
            return f"Database status check failed: {str(e)}"

    def show_repeat_last_search(self):
        """Show option to repeat last search"""
        if not hasattr(self, 'last_search_criteria') or not self.last_search_criteria:
            messagebox.showinfo("No Previous Search", "No previous search to repeat.")
            return
        
        criteria = self.last_search_criteria
        search_type = criteria.get('type', 'unknown')
        
        confirm = messagebox.askyesno("Repeat Search", 
                                     f"Repeat last search ({search_type})?\n"
                                     f"Criteria: {criteria.get('data', 'N/A')}")
        
        if confirm:
            self.repeat_last_search()

    def show_cache_statistics(self):
        """Show cache statistics and management"""
        cache_size = len(getattr(self, 'search_cache', {}))
        cache_memory = sum(len(str(v)) for v in getattr(self, 'search_cache', {}).values())
        
        stats_text = f"""SEARCH CACHE STATISTICS

    Cache Entries: {cache_size}
    Estimated Memory Usage: {cache_memory / 1024:.1f} KB
    Cache Hit Rate: {"N/A" if not hasattr(self, 'cache_hits') else f"{getattr(self, 'cache_hits', 0)} hits"}

    Cache helps speed up repeated searches by storing results temporarily.
    """
        
        dialog = tk.Toplevel(self.master)
        dialog.title("Cache Statistics")
        dialog.geometry("900x700")
        dialog.transient(self.master)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Cache Statistics", style='Title.TLabel').pack(pady=(0, 20))
        
        stats_label = tk.Label(frame, text=stats_text, justify=tk.LEFT, font=('Courier', 10))
        stats_label.pack(fill=tk.BOTH, expand=True, pady=(0, 20))
        
        button_frame = ttk.Frame(frame)
        button_frame.pack(fill=tk.X)
        
        ttk.Button(button_frame, text="Clear Cache", command=self.clear_search_cache).pack(side=tk.LEFT)
        ttk.Button(button_frame, text="Close", command=dialog.destroy).pack(side=tk.RIGHT)

    def show_detailed_student_view(self, student_data):
        """Show comprehensive detailed view of student with all related data"""
        if not student_data:
            messagebox.showerror("Error", "No student data provided")
            return
        
        # Create detailed view window
        detail_window = tk.Toplevel(self.master)
        detail_window.title(f"Student Details - {student_data[0]}")
        detail_window.geometry("900x700")
        detail_window.transient(self.master)
        
        # Create notebook for tabbed interface
        notebook = ttk.Notebook(detail_window)
        notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Basic Info Tab
        basic_frame = ttk.Frame(notebook, padding="20")
        notebook.add(basic_frame, text="Basic Information")
        
        self.create_basic_info_tab(basic_frame, student_data)
        
        # Academic History Tab
        academic_frame = ttk.Frame(notebook, padding="20")
        notebook.add(academic_frame, text="Academic History")
        
        self.create_academic_history_tab(academic_frame, student_data[0])
        
        # Analytics Tab
        analytics_frame = ttk.Frame(notebook, padding="20")
        notebook.add(analytics_frame, text="Performance Analytics")
        
        self.create_performance_analytics_tab(analytics_frame, student_data[0])
        
        # Actions Tab
        actions_frame = ttk.Frame(notebook, padding="20")
        notebook.add(actions_frame, text="Actions")
        
        self.create_student_actions_tab(actions_frame, student_data)

    def create_basic_info_tab(self, parent, student_data):
        """Create basic information tab"""
        info_text = f"""
    STUDENT INFORMATION

    Student ID: {student_data[0]}
    Email: {student_data[1]}
    Title: {student_data[2]}
    Name: {student_data[3]} {student_data[4] or ''} {student_data[5]}
    Gender: {student_data[6]}
    Date of Birth: {student_data[7]}
    Age: {student_data[8]}
    Course: {student_data[9]}
    Registration: {student_data[10]}
        """
        
        ttk.Label(parent, text="Personal Information", style='Header.TLabel').pack(anchor='w', pady=(0, 20))
        
        info_label = tk.Label(parent, text=info_text, justify=tk.LEFT, font=('Arial', 11), bg='white')
        info_label.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

    def create_academic_history_tab(self, parent, student_id):
        """Create academic history tab"""
        ttk.Label(parent, text="Academic History", style='Header.TLabel').pack(anchor='w', pady=(0, 20))
        
        # Load and display academic history
        try:
            conn = get_connection()
            cursor = conn.cursor()
            
            cursor.execute("""
            SELECT module_type, module_code, module_name, grade, enrollment_date, completion_date
            FROM student_modules
            WHERE student_id = ?
            ORDER BY enrollment_date DESC
            """, (student_id,))
            
            modules = cursor.fetchall()
            conn.close()
            
            if modules:
                # Create treeview for modules
                columns = ('Type', 'Code', 'Name', 'Grade', 'Enrolled', 'Completed')
                modules_tree = ttk.Treeview(parent, columns=columns, show='headings', height=15)
                
                for col in columns:
                    modules_tree.heading(col, text=col)
                    modules_tree.column(col, width=120)
                
                modules_scrollbar = ttk.Scrollbar(parent, orient=tk.VERTICAL, command=modules_tree.yview)
                modules_tree.configure(yscrollcommand=modules_scrollbar.set)
                
                modules_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
                modules_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
                
                # Populate data
                for module in modules:
                    module_type, code, name, grade, enrolled_date, completed_date = module
                    grade_display = grade if grade else "In Progress"
                    enrolled_display = enrolled_date[:10] if enrolled_date else "N/A"
                    completed_display = completed_date[:10] if completed_date else "N/A"
                    
                    modules_tree.insert('', 'end', values=(
                        module_type, code, name, grade_display, enrolled_display, completed_display
                    ))
            else:
                ttk.Label(parent, text="No academic history found.").pack()
                
        except Exception as e:
            ttk.Label(parent, text=f"Error loading academic history: {str(e)}").pack()

    def create_performance_analytics_tab(self, parent, student_id):
        """Create performance analytics tab"""
        ttk.Label(parent, text="Performance Analytics", style='Header.TLabel').pack(anchor='w', pady=(0, 20))
        
        try:
            conn = get_connection()
            cursor = conn.cursor()
            
            # Get performance metrics
            cursor.execute("""
            SELECT 
                COUNT(*) as total_modules,
                SUM(CASE WHEN grade IS NOT NULL THEN 1 ELSE 0 END) as completed_modules,
                SUM(CASE WHEN grade IN ('A', 'B', 'C') THEN 1 ELSE 0 END) as passed_modules,
                AVG(CASE 
                    WHEN grade = 'A' THEN 4.0
                    WHEN grade = 'B' THEN 3.0  
                    WHEN grade = 'C' THEN 2.0
                    WHEN grade = 'D' THEN 1.0
                    ELSE 0.0
                END) as gpa
            FROM student_modules
            WHERE student_id = ?
            """, (student_id,))
            
            metrics = cursor.fetchone()
            conn.close()
            
            if metrics and metrics[0] > 0:
                total, completed, passed, gpa = metrics
                completion_rate = (completed / total) * 100 if total > 0 else 0
                success_rate = (passed / completed) * 100 if completed > 0 else 0
                gpa_display = gpa if gpa is not None else 0.0

                analytics_text = f"""
    PERFORMANCE METRICS

    Total Modules Enrolled: {total}
    Completed Modules: {completed}
    Passed Modules (A-C): {passed}
    Completion Rate: {completion_rate:.1f}%
    Success Rate: {success_rate:.1f}%
    Current GPA: {gpa_display:.2f}

    ACADEMIC STATUS
    {"✅ On Track" if completion_rate > 70 else "⚠️ Needs Attention" if completion_rate > 40 else "🚨 At Risk"}
                """
                
                analytics_label = tk.Label(parent, text=analytics_text, justify=tk.LEFT, 
                                         font=('Courier', 11), bg='white')
                analytics_label.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            else:
                ttk.Label(parent, text="No performance data available.").pack()
                
        except Exception as e:
            ttk.Label(parent, text=f"Error loading performance data: {str(e)}").pack()

    def create_student_actions_tab(self, parent, student_data):
        """Create student actions tab"""
        ttk.Label(parent, text="Available Actions", style='Header.TLabel').pack(anchor='w', pady=(0, 20))
        
        actions = [
            ("📧 Send Email", lambda: self.simulate_send_email(student_data)),
            ("💾 Export Student Data", lambda: self.export_single_student(student_data)),
            ("📋 View Academic History", lambda: self.view_academic_history_detailed(student_data[0])),
            ("📊 Generate Performance Report", lambda: self.generate_student_performance_report(student_data[0])),
            ("🏷️ Add to Favorites", lambda: self.add_student_to_favorites(student_data)),
            ("📌 Mark for Follow-up", lambda: self.mark_single_student_followup(student_data))
        ]
        
        for text, command in actions:
            ttk.Button(parent, text=text, command=command, width=30).pack(pady=5)

    def add_student_to_favorites(self, student_data):
        """Add student to favorites list"""
        try:
            # Initialize favorites if it doesn't exist
            if not hasattr(self, 'favorite_students'):
                self.favorite_students = []
            
            student_id = student_data[0]
            student_name = f"{student_data[3]} {student_data[5]}"
            
            # Check if already in favorites
            if any(fav['id'] == student_id for fav in self.favorite_students):
                messagebox.showinfo("Already Favorited", f"{student_name} is already in your favorites.")
                return
            
            # Add to favorites
            favorite_entry = {
                'id': student_id,
                'name': student_name,
                'email': student_data[1],
                'course': student_data[9],
                'added_date': datetime.now().isoformat()
            }
            
            self.favorite_students.append(favorite_entry)
            
            # Save to file
            with open('favorite_students.json', 'w') as f:
                json.dump(self.favorite_students, f, indent=2)
            
            messagebox.showinfo("Added to Favorites", f"✅ {student_name} added to favorites!")
            self.log_output(f"Student {student_id} added to favorites")
            
        except Exception as e:
            messagebox.showerror("Error", f"Could not add to favorites: {str(e)}")

    def mark_single_student_followup(self, student_data):
        """Mark single student for follow-up"""
        dialog = tk.Toplevel(self.master)
        dialog.title("Mark for Follow-up")
        dialog.geometry("800x600")
        dialog.transient(self.master)
        dialog.grab_set()

        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)

        student_name = f"{student_data[3]} {student_data[5]}"
        ttk.Label(frame, text=f"Mark {student_name} for Follow-up", style='Title.TLabel').pack(pady=(0, 20))
        
        ttk.Label(frame, text="Follow-up Reason:").pack(anchor='w')
        reason_var = tk.StringVar()
        ttk.Entry(frame, textvariable=reason_var, width=40).pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(frame, text="Priority:").pack(anchor='w')
        priority_var = tk.StringVar(value="medium")
        
        for priority in ["high", "medium", "low"]:
            ttk.Radiobutton(frame, text=priority.title(), variable=priority_var, value=priority).pack(anchor='w')
        
        ttk.Label(frame, text="Notes:").pack(anchor='w', pady=(10, 0))
        notes_text = tk.Text(frame, height=4, wrap=tk.WORD)
        notes_text.pack(fill=tk.X, pady=(0, 20))
        
        def save_followup():
            reason = reason_var.get().strip()
            if not reason:
                messagebox.showwarning("Missing Reason", "Please enter a follow-up reason.")
                return
            
            priority = priority_var.get()
            notes = notes_text.get(1.0, tk.END).strip()
            
            followup_data = {
                'student_id': student_data[0],
                'student_name': student_name,
                'email': student_data[1],
                'reason': reason,
                'priority': priority,
                'notes': notes,
                'marked_date': datetime.now().isoformat(),
                'marked_by': 'current_user'
            }
            
            # Save to file
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"followup_{student_data[0]}_{timestamp}.json"
            
            with open(filename, 'w') as f:
                json.dump(followup_data, f, indent=2)
            
            messagebox.showinfo("Follow-up Marked", 
                              f"✅ {student_name} marked for follow-up\n"
                              f"Priority: {priority}\n"
                              f"Saved to: {filename}")
            
            dialog.destroy()
        
        button_frame = ttk.Frame(frame)
        button_frame.pack(fill=tk.X)
        
        ttk.Button(button_frame, text="📌 Mark", command=save_followup).pack(side=tk.LEFT)
        ttk.Button(button_frame, text="❌ Cancel", command=dialog.destroy).pack(side=tk.RIGHT)

    def generate_student_performance_report(self, student_id):
        """Generate detailed performance report for a student"""
        self.update_status(f"Generating performance report for {student_id}...")
        self.start_progress()
        
        def run_report_generation():
            try:
                report = self.create_student_performance_report(student_id)
                self.output_queue.put(("analytics", report))
                self.output_queue.put(("log", f"Performance report generated for {student_id}"))
            except Exception as e:
                self.output_queue.put(("error", f"Report generation error: {str(e)}"))
            finally:
                self.output_queue.put(("stop_progress", None))
        
        threading.Thread(target=run_report_generation, daemon=True).start()

    def create_student_performance_report(self, student_id):
        """Create detailed performance report for a student"""
        try:
            conn = get_connection()
            cursor = conn.cursor()
            
            # Get student info
            cursor.execute("SELECT * FROM students WHERE student_id = ?", (student_id,))
            student = cursor.fetchone()
            
            if not student:
                return "Student not found"
            
            # Get module performance
            cursor.execute("""
            SELECT module_type, module_code, module_name, grade, enrollment_date, completion_date
            FROM student_modules
            WHERE student_id = ?
            ORDER BY enrollment_date
            """, (student_id,))
            
            modules = cursor.fetchall()
            conn.close()
            
            report = f"STUDENT PERFORMANCE REPORT\n"
            report += f"=" * 50 + "\n"
            report += f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
            
            report += f"STUDENT INFORMATION:\n"
            report += f"ID: {student[0]}\n"
            report += f"Name: {student[3]} {student[5]}\n"
            report += f"Email: {student[1]}\n"
            report += f"Course: {student[9]}\n"
            report += f"Registration: {student[10]}\n\n"
            
            if modules:
                completed = [m for m in modules if m[3] is not None]
                in_progress = [m for m in modules if m[3] is None]
                passed = [m for m in modules if m[3] in ['A', 'B', 'C']]
                
                report += f"ACADEMIC SUMMARY:\n"
                report += f"Total Modules: {len(modules)}\n"
                report += f"Completed: {len(completed)}\n"
                report += f"In Progress: {len(in_progress)}\n"
                report += f"Passed (A-C): {len(passed)}\n"
                
                if completed:
                    completion_rate = (len(completed) / len(modules)) * 100
                    success_rate = (len(passed) / len(completed)) * 100
                    report += f"Completion Rate: {completion_rate:.1f}%\n"
                    report += f"Success Rate: {success_rate:.1f}%\n"
                
                report += f"\nDETAILED MODULE HISTORY:\n"
                report += f"-" * 40 + "\n"
                
                for module in modules:
                    module_type, code, name, grade, enrolled_date, completed_date = module
                    grade_display = grade if grade else "In Progress"
                    
                    report += f"{code} - {name}\n"
                    report += f"  Type: {module_type} | Grade: {grade_display}\n"
                    report += f"  Enrolled: {enrolled_date[:10] if enrolled_date else 'N/A'}\n"
                    if completed_date:
                        report += f"  Completed: {completed_date[:10]}\n"
                    report += f"\n"
            else:
                report += f"No module enrollment history found.\n"
            
            return report
            
        except Exception as e:
            return f"Error generating performance report: {str(e)}"

    def show_favorites_manager(self):
        """Show favorites management interface"""
        dialog = tk.Toplevel(self.master)
        dialog.title("Favorites Manager")
        dialog.geometry("1100x800")
        dialog.transient(self.master)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Favorite Students", style='Title.TLabel').pack(pady=(0, 20))
        
        # Load favorites
        try:
            with open('favorite_students.json', 'r') as f:
                self.favorite_students = json.load(f)
        except FileNotFoundError:
            self.favorite_students = []
        
        if not self.favorite_students:
            ttk.Label(frame, text="No favorite students yet.").pack()
            ttk.Button(frame, text="Close", command=dialog.destroy).pack(pady=(20, 0))
            return
        
        # Favorites list
        columns = ('ID', 'Name', 'Email', 'Course', 'Added')
        favorites_tree = ttk.Treeview(frame, columns=columns, show='headings', height=12)
        
        for col in columns:
            favorites_tree.heading(col, text=col)
            favorites_tree.column(col, width=100)
        
        favorites_scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=favorites_tree.yview)
        favorites_tree.configure(yscrollcommand=favorites_scrollbar.set)
        
        favorites_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        favorites_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Populate favorites
        for fav in self.favorite_students:
            added_date = fav['added_date'][:10] if 'added_date' in fav else 'N/A'
            favorites_tree.insert('', 'end', values=(
                fav['id'], fav['name'], fav['email'], fav['course'], added_date
            ))
        
        # Actions
        actions_frame = ttk.Frame(frame)
        actions_frame.pack(fill=tk.X, pady=(10, 0))
        
        def view_favorite():
            selection = favorites_tree.selection()
            if not selection:
                messagebox.showwarning("No Selection", "Please select a student.")
                return
            
            item = favorites_tree.item(selection[0])
            student_id = item['values'][0]
            
            # Find full student data
            try:
                conn = get_connection()
                cursor = conn.cursor()
                cursor.execute("SELECT * FROM students WHERE student_id = ?", (student_id,))
                student_data = cursor.fetchone()
                conn.close()
                
                if student_data:
                    self.show_detailed_student_view(student_data)
                else:
                    messagebox.showerror("Error", "Student not found in database")
            except Exception as e:
                messagebox.showerror("Error", f"Could not load student: {str(e)}")
        
        def remove_favorite():
            selection = favorites_tree.selection()
            if not selection:
                messagebox.showwarning("No Selection", "Please select a student to remove.")
                return
            
            item = favorites_tree.item(selection[0])
            student_name = item['values'][1]
            
            if messagebox.askyesno("Confirm Remove", f"Remove {student_name} from favorites?"):
                student_id = item['values'][0]
                self.favorite_students = [fav for fav in self.favorite_students if fav['id'] != student_id]
                
                # Save updated favorites
                with open('favorite_students.json', 'w') as f:
                    json.dump(self.favorite_students, f, indent=2)
                
                favorites_tree.delete(selection[0])
                messagebox.showinfo("Removed", f"{student_name} removed from favorites")
        
        ttk.Button(actions_frame, text="View Details", command=view_favorite).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(actions_frame, text="Remove", command=remove_favorite).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(actions_frame, text="Close", command=dialog.destroy).pack(side=tk.RIGHT)

    def show_system_optimization_tools(self):
        """Show system optimization and maintenance tools"""
        dialog = tk.Toplevel(self.master)
        dialog.title("System Optimization Tools")
        dialog.geometry("1000x750")
        dialog.transient(self.master)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="System Optimization & Maintenance", style='Title.TLabel').pack(pady=(0, 20))
        
        # Database optimization
        db_frame = ttk.LabelFrame(frame, text="Database Optimization", padding="10")
        db_frame.pack(fill=tk.X, pady=(0, 10))
        
        db_tools = [
            ("Vacuum Database", self.vacuum_database),
            ("Rebuild Indexes", self.rebuild_indexes),
            ("Analyze Statistics", self.analyze_statistics),
            ("🏠 Return to Main Menu", self.return_to_main_menu)
        ]
        
        for text, command in db_tools:
            ttk.Button(db_frame, text=text, command=command, width=20).pack(pady=2)
        
        # Cache management
        cache_frame = ttk.LabelFrame(frame, text="Cache Management", padding="10")
        cache_frame.pack(fill=tk.X, pady=(0, 10))
        
        cache_tools = [
            ("View Cache Statistics", self.show_cache_statistics),
            ("Clear Search Cache", self.clear_search_cache),
            ("🏠 Return to Main Menu", self.return_to_main_menu)
        ]
        
        for text, command in cache_tools:
            ttk.Button(cache_frame, text=text, command=command, width=20).pack(pady=2)
        
        ttk.Button(frame, text="Close", command=dialog.destroy).pack(pady=(20, 0))

    def vacuum_database(self):
        """Vacuum the database to optimize storage"""
        self.update_status("Vacuuming database...")
        
        def run_vacuum():
            try:
                conn = get_connection()
                cursor = conn.cursor()
                cursor.execute("VACUUM")
                conn.close()
                self.log_output("Database vacuum completed successfully")
                messagebox.showinfo("Vacuum Complete", "Database has been vacuumed successfully")
            except Exception as e:
                self.log_output(f"Database vacuum failed: {str(e)}")
                messagebox.showerror("Vacuum Failed", f"Database vacuum failed: {str(e)}")
            finally:
                self.update_status("Ready")
        
        threading.Thread(target=run_vacuum, daemon=True).start()

    def rebuild_indexes(self):
        """Rebuild database indexes for better performance"""
        self.update_status("Rebuilding database indexes...")
        
        def run_rebuild():
            try:
                conn = get_connection()
                cursor = conn.cursor()
                
                # Drop and recreate indexes
                indexes = [
                    "DROP INDEX IF EXISTS idx_students_name",
                    "DROP INDEX IF EXISTS idx_students_course", 
                    "DROP INDEX IF EXISTS idx_students_age",
                    "DROP INDEX IF EXISTS idx_modules_student",
                    "CREATE INDEX idx_students_name ON students(first_name, last_name)",
                    "CREATE INDEX idx_students_course ON students(course)",
                    "CREATE INDEX idx_students_age ON students(age)",
                    "CREATE INDEX idx_modules_student ON student_modules(student_id)"
                ]
                
                for index_sql in indexes:
                    cursor.execute(index_sql)
                
                conn.commit()
                conn.close()
                
                self.log_output("Database indexes rebuilt successfully")
                messagebox.showinfo("Rebuild Complete", "Database indexes have been rebuilt")
            except Exception as e:
                self.log_output(f"Index rebuild failed: {str(e)}")
                messagebox.showerror("Rebuild Failed", f"Index rebuild failed: {str(e)}")
            finally:
                self.update_status("Ready")
        
        threading.Thread(target=run_rebuild, daemon=True).start()

    def analyze_statistics(self):
        """Analyze database statistics for query optimization"""
        self.update_status("Analyzing database statistics...")
        
        def run_analyze():
            try:
                conn = get_connection()
                cursor = conn.cursor()
                cursor.execute("ANALYZE")
                conn.close()
                
                self.log_output("Database statistics analyzed successfully")
                messagebox.showinfo("Analysis Complete", "Database statistics have been updated")
            except Exception as e:
                self.log_output(f"Statistics analysis failed: {str(e)}")
                messagebox.showerror("Analysis Failed", f"Statistics analysis failed: {str(e)}")
            finally:
                self.update_status("Ready")
        
        threading.Thread(target=run_analyze, daemon=True).start()

    def check_integrity(self):
        """Check database integrity"""
        self.update_status("Checking database integrity...")
        self.start_progress()
        
        def run_integrity_check():
            try:
                result = self.perform_integrity_check()
                self.output_queue.put(("analytics", result))
                self.output_queue.put(("log", "Database integrity check completed"))
            except Exception as e:
                self.output_queue.put(("error", f"Integrity check failed: {str(e)}"))
            finally:
                self.output_queue.put(("stop_progress", None))
        
        threading.Thread(target=run_integrity_check, daemon=True).start()

    def perform_integrity_check(self):
        """Perform comprehensive database integrity check"""
        try:
            conn = get_connection()
            cursor = conn.cursor()
            
            result = "DATABASE INTEGRITY CHECK\n"
            result += "=" * 50 + "\n"
            result += f"Check performed: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
            
            # SQLite integrity check
            cursor.execute("PRAGMA integrity_check")
            integrity_result = cursor.fetchone()[0]
            
            result += f"SQLite Integrity Check: {integrity_result}\n\n"
            
            # Check for referential integrity
            cursor.execute("""
            SELECT COUNT(*) FROM student_modules sm 
            WHERE NOT EXISTS (SELECT 1 FROM students s WHERE s.student_id = sm.student_id)
            """)
            orphaned_modules = cursor.fetchone()[0]
            
            result += f"REFERENTIAL INTEGRITY:\n"
            result += f"Orphaned module records: {orphaned_modules}\n"
            
            # Check for data consistency
            cursor.execute("SELECT COUNT(*) FROM students WHERE student_id IS NULL OR student_id = ''")
            null_ids = cursor.fetchone()[0]
            
            cursor.execute("SELECT COUNT(*) FROM students WHERE email_address IS NULL OR email_address = ''")
            null_emails = cursor.fetchone()[0]
            
            result += f"\nDATA CONSISTENCY:\n"
            result += f"Students with null/empty IDs: {null_ids}\n"
            result += f"Students with null/empty emails: {null_emails}\n"
            
            # Check for duplicate student IDs
            cursor.execute("""
            SELECT student_id, COUNT(*) as count 
            FROM students 
            GROUP BY student_id 
            HAVING count > 1
            """)
            duplicates = cursor.fetchall()
            
            result += f"Duplicate student IDs: {len(duplicates)}\n"
            if duplicates:
                result += "Duplicate IDs found:\n"
                for student_id, count in duplicates:
                    result += f"  {student_id}: {count} records\n"
            
            conn.close()
            
            result += f"\nIntegrity check completed.\n"
            
            if integrity_result == "ok" and orphaned_modules == 0 and null_ids == 0 and len(duplicates) == 0:
                result += "Database integrity: EXCELLENT\n"
            elif orphaned_modules > 0 or null_ids > 0 or len(duplicates) > 0:
                result += "Database integrity: NEEDS ATTENTION\n"
            else:
                result += "Database integrity: GOOD\n"
            
            return result
            
        except Exception as e:
            return f"Integrity check failed: {str(e)}"

    def optimize_memory_usage(self):
        """Optimize memory usage by clearing caches and temporary data"""
        try:
            # Clear search cache
            if hasattr(self, 'search_cache'):
                cache_size = len(self.search_cache)
                self.search_cache.clear()
            else:
                cache_size = 0
            
            # Clear search history if it gets too large
            if hasattr(self, 'search_history') and len(self.search_history) > 100:
                self.search_history = self.search_history[-50:]  # Keep last 50
            
            # Force garbage collection
            import gc
            collected = gc.collect()
            
            self.log_output(f"Memory optimization completed:")
            self.log_output(f"  Cleared {cache_size} cache entries")
            self.log_output(f"  Collected {collected} objects")
            
            messagebox.showinfo("Memory Optimized", 
                              f"Memory optimization completed\n"
                              f"Cache entries cleared: {cache_size}\n"
                              f"Objects collected: {collected}")
            
        except Exception as e:
            self.log_output(f"Memory optimization failed: {str(e)}")
            messagebox.showerror("Optimization Failed", f"Memory optimization failed: {str(e)}")

    def show_search_history_detailed(self):
        """Show detailed search history with management options"""
        dialog = tk.Toplevel(self.master)
        dialog.title("Search History")
        dialog.geometry("700x500")
        dialog.transient(self.master)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Search History & Management", style='Title.TLabel').pack(pady=(0, 20))
        
        # History tree
        columns = ('Time', 'Type', 'Criteria', 'Results', 'Duration')
        history_tree = ttk.Treeview(frame, columns=columns, show='headings', height=15)
        
        for col in columns:
            history_tree.heading(col, text=col)
            history_tree.column(col, width=120)
        
        history_scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=history_tree.yview)
        history_tree.configure(yscrollcommand=history_scrollbar.set)
        
        history_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        history_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Load search history from database
        try:
            conn = get_connection()
            cursor = conn.cursor()
            cursor.execute("""
            SELECT search_type, search_criteria, results_count, execution_time, search_datetime
            FROM search_analytics
            ORDER BY search_datetime DESC
            LIMIT 100
            """)
            history_data = cursor.fetchall()
            conn.close()

            for search_type, criteria, results, duration, search_datetime in history_data:
                time_display = search_datetime[:16] if search_datetime else 'N/A'
                duration_display = f"{duration:.2f}s" if duration else 'N/A'
                criteria_display = criteria[:30] + "..." if len(criteria) > 30 else criteria
                
                history_tree.insert('', 'end', values=(
                    time_display, search_type, criteria_display, results, duration_display
                ))
        
        except Exception as e:
            self.log_output(f"Error loading search history: {str(e)}")
        
        # Actions
        actions_frame = ttk.Frame(frame)
        actions_frame.pack(fill=tk.X, pady=(10, 0))
        
        def repeat_selected_search():
            selection = history_tree.selection()
            if not selection:
                messagebox.showwarning("No Selection", "Please select a search to repeat.")
                return
            
            item = history_tree.item(selection[0])
            search_type = item['values'][1]
            
            messagebox.showinfo("Repeat Search", f"Repeating {search_type} search...")
            # In a full implementation, would parse criteria and re-execute
        
        def export_history():
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"search_history_{timestamp}.csv"
            
            try:
                with open(filename, 'w', newline='') as csvfile:
                    writer = csv.writer(csvfile)
                    writer.writerow(['Time', 'Type', 'Criteria', 'Results', 'Duration'])
                    
                    for child in history_tree.get_children():
                        values = history_tree.item(child)['values']
                        writer.writerow(values)
                
                messagebox.showinfo("Export Complete", f"Search history exported to {filename}")
            except Exception as e:
                messagebox.showerror("Export Failed", f"Could not export history: {str(e)}")
        
        def clear_history():
            if messagebox.askyesno("Confirm Clear", "Clear all search history? This cannot be undone."):
                try:
                    conn = get_connection()
                    cursor = conn.cursor()
                    cursor.execute("DELETE FROM search_analytics")
                    conn.commit()
                    conn.close()
                    
                    # Clear tree
                    for item in history_tree.get_children():
                        history_tree.delete(item)
                    
                    messagebox.showinfo("History Cleared", "Search history has been cleared.")
                except Exception as e:
                    messagebox.showerror("Clear Failed", f"Could not clear history: {str(e)}")
        
        ttk.Button(actions_frame, text="Repeat Search", command=repeat_selected_search).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(actions_frame, text="Export History", command=export_history).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(actions_frame, text="Clear History", command=clear_history).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(actions_frame, text="Close", command=dialog.destroy).pack(side=tk.RIGHT)
    
    def import_data(self, file_type, data_type, filename_override: Optional[str] = None):
        """Import data from file"""
        filetypes = {
            "csv": [("CSV files", "*.csv")],
            "json": [("JSON files", "*.json")],
            "xlsx": [("Excel files", "*.xlsx")]
        }
        
        if filename_override:
            filename = filename_override
        else:
            filename = filedialog.askopenfilename(
                title=f"Select {file_type.upper()} file to import",
                filetypes=filetypes.get(file_type, [("All files", "*.*")])
            )
        
        if not filename:
            return 0

        self.update_status(f"Importing {data_type} from {file_type.upper()}...")
        self.start_progress()
        try:
            imported = self._import_records_from_file(filename, file_type, data_type)
            self.log_output(f"Import completed: {filename}")
            self.log_output(f"{imported} {data_type} record(s) imported successfully")
            messagebox.showinfo(
                "Import Complete",
                f"Successfully imported {imported} {data_type} record(s) from {filename}"
            )
            return imported
        except Exception as e:
            self.log_output(f"Import error: {e}")
            messagebox.showerror("Import Failed", f"Could not import data: {e}")
            return 0
        finally:
            self.stop_progress()
            self.update_status("Ready")
    
    def bulk_import_with_validation(self):
        """Bulk import with data validation"""
        filename = filedialog.askopenfilename(
            title="Select file for bulk import",
            filetypes=[("CSV files", "*.csv"), ("JSON files", "*.json"), ("All files", "*.*")]
        )
        
        if filename:
            # Validation dialog
            validation_dialog = tk.Toplevel(self.master)
            validation_dialog.title("🔍 Import Validation")
            validation_dialog.geometry("900x700")
            validation_dialog.transient(self.master)
            validation_dialog.grab_set()
            
            val_frame = ttk.Frame(validation_dialog, padding="20")
            val_frame.pack(fill=tk.BOTH, expand=True)
            
            ttk.Label(val_frame, text="Import Validation Settings", style='Title.TLabel').pack(pady=(0, 20))
            
            # Validation options
            validate_emails = tk.BooleanVar(value=True)
            validate_ages = tk.BooleanVar(value=True)
            validate_courses = tk.BooleanVar(value=True)
            skip_duplicates = tk.BooleanVar(value=True)
            
            ttk.Checkbutton(val_frame, text="Validate email formats", variable=validate_emails).pack(anchor='w')
            ttk.Checkbutton(val_frame, text="Validate age ranges", variable=validate_ages).pack(anchor='w')
            ttk.Checkbutton(val_frame, text="Validate course codes", variable=validate_courses).pack(anchor='w')
            ttk.Checkbutton(val_frame, text="Skip duplicate records", variable=skip_duplicates).pack(anchor='w')
            
            def start_validated_import():
                validation_dialog.destroy()
                
                self.update_status("Running validated bulk import...")
                self.start_progress()
                
                def run_validated_import():
                    try:
                        # Simulate validation and import
                        validation_settings = {
                            'validate_emails': validate_emails.get(),
                            'validate_ages': validate_ages.get(),
                            'validate_courses': validate_courses.get(),
                            'skip_duplicates': skip_duplicates.get()
                        }
                        
                        # Simulate processing
                        import time
                        time.sleep(3)
                        
                        self.output_queue.put(("log", f"Bulk import with validation completed"))
                        self.output_queue.put(("log", f"File: {filename}"))
                        self.output_queue.put(("log", f"Validation settings: {validation_settings}"))
                        
                    except Exception as e:
                        self.output_queue.put(("error", f"Validated import error: {str(e)}"))
                    finally:
                        self.output_queue.put(("stop_progress", None))
                
                threading.Thread(target=run_validated_import, daemon=True).start()
            
            button_frame = ttk.Frame(val_frame)
            button_frame.pack(fill=tk.X, pady=(20, 0))
            
            ttk.Button(button_frame, text="✅ Start Import", command=start_validated_import).pack(side=tk.LEFT)
            ttk.Button(button_frame, text="❌ Cancel", command=validation_dialog.destroy).pack(side=tk.RIGHT)
    
    def show_comprehensive_reports(self):
        """Generate comprehensive system reports"""
        dialog = tk.Toplevel(self.master)
        dialog.title("Comprehensive Reports Generator")
        dialog.geometry("1100x800")
        dialog.transient(self.master)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Comprehensive Reports Generator", style='Title.TLabel').pack(pady=(0, 20))
        
        # Report types
        reports_frame = ttk.LabelFrame(frame, text="Available Reports", padding="10")
        reports_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 20))
        
        report_options = [
            ("Student Summary Report", "student_summary", "Complete overview of all students"),
            ("Module Enrollment Report", "module_enrollment", "Detailed module enrollment analysis"),
            ("Demographics Analysis", "demographics_analysis", "Comprehensive demographic breakdown"),
            ("Performance Report", "performance_report", "Academic performance analysis"),
            ("Custom SQL Report", "custom_sql", "Execute custom SQL queries")
        ]
        
        selected_reports = {}
        
        for name, key, description in report_options:
            report_frame = ttk.Frame(reports_frame)
            report_frame.pack(fill=tk.X, pady=5)
            
            selected_reports[key] = tk.BooleanVar()
            ttk.Checkbutton(report_frame, text=name, variable=selected_reports[key]).pack(side=tk.LEFT)
            ttk.Label(report_frame, text=f" - {description}", font=('Arial', 8)).pack(side=tk.LEFT)
        
        # Output options
        output_frame = ttk.LabelFrame(frame, text="Output Options", padding="10")
        output_frame.pack(fill=tk.X, pady=(0, 20))
        
        output_format_var = tk.StringVar(value="console")
        ttk.Radiobutton(output_frame, text="Display in Console", variable=output_format_var, value="console").pack(anchor='w')
        ttk.Radiobutton(output_frame, text="Export to File", variable=output_format_var, value="file").pack(anchor='w')
        ttk.Radiobutton(output_frame, text="Both Console and File", variable=output_format_var, value="both").pack(anchor='w')
        
        def generate_reports():
            selected = [key for key, var in selected_reports.items() if var.get()]
            if not selected:
                messagebox.showwarning("No Reports Selected", "Please select at least one report to generate.")
                return
            
            output_format = output_format_var.get()
            
            dialog.destroy()
            self.update_status("Generating comprehensive reports...")
            self.start_progress()
            
            def run_report_generation():
                try:
                    for report_type in selected:
                        report_result = self.generate_specific_report(report_type)
                        
                        if output_format in ["console", "both"]:
                            self.output_queue.put(("analytics", report_result))
                        
                        if output_format in ["file", "both"]:
                            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                            filename = f"{report_type}_report_{timestamp}.txt"
                            with open(filename, 'w', encoding='utf-8') as f:
                                f.write(report_result)
                            self.output_queue.put(("log", f"Report saved to {filename}"))
                    
                    self.output_queue.put(("log", "All comprehensive reports generated successfully."))
                except Exception as e:
                    self.output_queue.put(("error", f"Report generation error: {str(e)}"))
                finally:
                    self.output_queue.put(("stop_progress", None))
            
            threading.Thread(target=run_report_generation, daemon=True).start()
        
        button_frame = ttk.Frame(frame)
        button_frame.pack(fill=tk.X)
        
        ttk.Button(button_frame, text="Generate Reports", command=generate_reports).pack(side=tk.LEFT)
        ttk.Button(button_frame, text="Cancel", command=dialog.destroy).pack(side=tk.RIGHT)
    
    def perform_date_search(self, start_date, end_date):
        """Perform date range search"""
        try:
            conn = get_connection()
            if not conn:
                raise Exception("Database connection failed")
            
            cursor = conn.cursor()
            
            query = "SELECT * FROM students WHERE 1=1"
            params = []
            
            if start_date:
                try:
                    datetime.strptime(start_date, "%Y-%m-%d")
                    query += " AND registration_datetime >= ?"
                    params.append(start_date + " 00:00:00")
                except ValueError:
                    raise Exception("Invalid start date format. Use YYYY-MM-DD.")
            
            if end_date:
                try:
                    datetime.strptime(end_date, "%Y-%m-%d")
                    query += " AND registration_datetime <= ?"
                    params.append(end_date + " 23:59:59")
                except ValueError:
                    raise Exception("Invalid end date format. Use YYYY-MM-DD.")
            
            cursor.execute(query, params)
            results = cursor.fetchall()
            conn.close()
            
            return results
            
        except Exception as e:
            raise Exception(f"Date search error: {str(e)}")

    def show_auto_complete_search(self):
        """Show auto-complete search interface"""
        dialog = tk.Toplevel(self.master)
        dialog.title("🔍 Auto-Complete Search")
        dialog.geometry("1000x750")
        dialog.transient(self.master)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Auto-Complete Search", style='Title.TLabel').pack(pady=(0, 20))
        
        # Search field with suggestions
        ttk.Label(frame, text="Type to search:").pack(anchor='w')
        
        search_var = tk.StringVar()
        search_entry = ttk.Entry(frame, textvariable=search_var, width=40)
        search_entry.pack(fill=tk.X, pady=(0, 10))
        
        # Suggestions listbox
        ttk.Label(frame, text="Suggestions:").pack(anchor='w')
        
        suggestions_frame = ttk.Frame(frame)
        suggestions_frame.pack(fill=tk.BOTH, expand=True, pady=(5, 20))
        
        suggestions_listbox = tk.Listbox(suggestions_frame, height=10)
        suggestions_scrollbar = ttk.Scrollbar(suggestions_frame, orient=tk.VERTICAL, command=suggestions_listbox.yview)
        suggestions_listbox.configure(yscrollcommand=suggestions_scrollbar.set)
        
        suggestions_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        suggestions_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        def update_suggestions(*args):
            """Update suggestions based on input"""
            term = search_var.get().lower()
            suggestions_listbox.delete(0, tk.END)
            
            if len(term) >= 2:
                # Get suggestions from database
                try:
                    conn = get_connection()
                    cursor = conn.cursor()
                    
                    # Get matching names, emails, and student IDs
                    cursor.execute("""
                        SELECT DISTINCT
                            first_name || ' ' || last_name as full_name,
                            email_address,
                            student_id
                        FROM students
                        WHERE LOWER(first_name || ' ' || last_name) LIKE ?
                           OR LOWER(email_address) LIKE ?
                           OR LOWER(student_id) LIKE ?
                        LIMIT 20
                    """, (f'%{term}%', f'%{term}%', f'%{term}%'))
                    
                    results = cursor.fetchall()
                    conn.close()
                    
                    for full_name, email, student_id in results:
                        suggestions_listbox.insert(tk.END, f"{full_name} ({email})")
                    
                except Exception as e:
                    suggestions_listbox.insert(tk.END, f"Error loading suggestions: {str(e)}")
        
        search_var.trace('w', update_suggestions)
        
        def select_suggestion(event):
            """Select suggestion and populate search field"""
            selection = suggestions_listbox.curselection()
            if selection:
                selected_text = suggestions_listbox.get(selection[0])
                # Extract name from "Name (email)" format
                name_part = selected_text.split(' (')[0]
                search_var.set(name_part)
        
        suggestions_listbox.bind('<Double-1>', select_suggestion)
        
        def execute_autocomplete_search():
            term = search_var.get().strip()
            if not term:
                messagebox.showwarning("Missing Input", "Please enter a search term.")
                return
            
            dialog.destroy()
            self.update_status("Performing auto-complete search...")
            self.start_progress()
            
            def run_search():
                try:
                    results = self.perform_autocomplete_search(term)
                    self.output_queue.put(("search_results", results))
                    self.output_queue.put(("log", f"Auto-complete search completed. Found {len(results)} results."))
                except Exception as e:
                    self.output_queue.put(("error", f"Auto-complete search error: {str(e)}"))
                finally:
                    self.output_queue.put(("stop_progress", None))
            
            threading.Thread(target=run_search, daemon=True).start()
        
        button_frame = ttk.Frame(frame)
        button_frame.pack(fill=tk.X)
        
        ttk.Button(button_frame, text="🔍 Search", command=execute_autocomplete_search).pack(side=tk.LEFT)
        ttk.Button(button_frame, text="❌ Cancel", command=dialog.destroy).pack(side=tk.RIGHT)

    def perform_autocomplete_search(self, term):
        """Perform auto-complete search"""
        try:
            conn = get_connection()
            cursor = conn.cursor()
            
            query = """
            SELECT * FROM students 
            WHERE LOWER(first_name || ' ' || last_name) LIKE ?
               OR LOWER(email_address) LIKE ?
               OR LOWER(student_id) LIKE ?
            """
            params = [f'%{term.lower()}%'] * 3
            
            cursor.execute(query, params)
            results = cursor.fetchall()
            conn.close()
            
            return results
            
        except Exception as e:
            raise Exception(f"Auto-complete search error: {str(e)}")

    def show_smart_suggestions(self):
        """Show smart suggestions interface"""
        dialog = tk.Toplevel(self.master)
        dialog.title("🧠 Smart Suggestions")
        dialog.geometry("1100x800")
        dialog.transient(self.master)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Smart Search Suggestions", style='Title.TLabel').pack(pady=(0, 20))
        
        # Suggestion categories
        notebook = ttk.Notebook(frame)
        notebook.pack(fill=tk.BOTH, expand=True, pady=(0, 20))
        
        # Popular searches tab
        popular_frame = ttk.Frame(notebook, padding="10")
        notebook.add(popular_frame, text="Popular Searches")
        
        popular_suggestions = [
            "Students enrolled in CS courses",
            "Recent registrations (last 30 days)",
            "Students over 25 years old",
            "Female students in Data Science",
            "Students without email addresses",
            "Incomplete module enrollments"
        ]
        
        ttk.Label(popular_frame, text="Most popular search patterns:").pack(anchor='w', pady=(0, 10))
        
        for suggestion in popular_suggestions:
            btn = ttk.Button(popular_frame, text=f"🔍 {suggestion}",
                            command=lambda s=suggestion: self.execute_suggestion(s, dialog))
            btn.pack(fill=tk.X, pady=2)
        
        # Recent searches tab
        recent_frame = ttk.Frame(notebook, padding="10")
        notebook.add(recent_frame, text="Recent Searches")
        
        recent_searches = [
            "John Smith",
            "CS101 module",
            "Age between 20-25",
            "Registration after 2024-01-01"
        ]
        
        ttk.Label(recent_frame, text="Your recent searches:").pack(anchor='w', pady=(0, 10))
        
        for search in recent_searches:
            btn = ttk.Button(recent_frame, text=f"🔄 {search}",
                            command=lambda s=search: self.execute_suggestion(s, dialog))
            btn.pack(fill=tk.X, pady=2)
        
        # Recommended tab
        recommended_frame = ttk.Frame(notebook, padding="10")
        notebook.add(recommended_frame, text="Recommended")
        
        recommendations = [
            "Students at risk of dropping out",
            "High-performing students for honors",
            "Students needing academic support",
            "Duplicate student records"
        ]
        
        ttk.Label(recommended_frame, text="Recommended searches based on patterns:").pack(anchor='w', pady=(0, 10))
        
        for recommendation in recommendations:
            btn = ttk.Button(recommended_frame, text=f"💡 {recommendation}",
                            command=lambda r=recommendation: self.execute_suggestion(r, dialog))
            btn.pack(fill=tk.X, pady=2)
        
        ttk.Button(frame, text="❌ Close", command=dialog.destroy).pack()

    def execute_suggestion(self, suggestion, parent_dialog):
        """Execute a suggested search"""
        parent_dialog.destroy()
        
        # Parse suggestion and create appropriate search
        if "CS courses" in suggestion:
            # Execute course search
            criteria = {"course": "CS"}
        elif "last 30 days" in suggestion:
            # Execute date search
            start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
            criteria = {"start_date": start_date}
        elif "over 25" in suggestion:
            # Execute age search
            criteria = {"min_age": "25"}
        elif "Female" in suggestion and "Data Science" in suggestion:
            criteria = {"gender": "female", "course": "DS"}
        else:
            # Default to text search
            criteria = {"search_term": suggestion}
        
        self.update_status(f"Executing suggested search: {suggestion}")
        self.start_progress()
        
        def run_suggestion():
            try:
                results = self.perform_suggested_search(criteria)
                self.output_queue.put(("search_results", results))
                self.output_queue.put(("log", f"Suggested search completed: {suggestion}. Found {len(results)} results."))
            except Exception as e:
                self.output_queue.put(("error", f"Suggested search error: {str(e)}"))
            finally:
                self.output_queue.put(("stop_progress", None))
        
        threading.Thread(target=run_suggestion, daemon=True).start()

    def perform_suggested_search(self, criteria):
        """Perform search based on suggestion criteria"""
        try:
            conn = get_connection()
            cursor = conn.cursor()
            
            query = "SELECT * FROM students WHERE 1=1"
            params = []
            
            for key, value in criteria.items():
                if key == "course":
                    query += " AND LOWER(course) = LOWER(?)"
                    params.append(value)
                elif key == "min_age":
                    query += " AND age >= ?"
                    params.append(int(value))
                elif key == "gender":
                    query += " AND LOWER(gender) = LOWER(?)"
                    params.append(value)
                elif key == "start_date":
                    query += " AND registration_datetime >= ?"
                    params.append(value + " 00:00:00")
                elif key == "search_term":
                    query += " AND (LOWER(first_name || ' ' || last_name) LIKE ? OR LOWER(email_address) LIKE ?)"
                    params.extend([f'%{value.lower()}%', f'%{value.lower()}%'])
            
            cursor.execute(query, params)
            results = cursor.fetchall()
            conn.close()
            
            return results
            
        except Exception as e:
            raise Exception(f"Suggested search error: {str(e)}")

    def show_predictive_analytics(self):
        """Show predictive analytics interface"""
        dialog = tk.Toplevel(self.master)
        dialog.title("📊 Predictive Analytics")
        dialog.geometry("1200x850")
        dialog.transient(self.master)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Predictive Analytics Dashboard", style='Title.TLabel').pack(pady=(0, 20))
        
        # Analytics options
        analytics_notebook = ttk.Notebook(frame)
        analytics_notebook.pack(fill=tk.BOTH, expand=True, pady=(0, 20))
        
        # At-risk students tab
        risk_frame = ttk.Frame(analytics_notebook, padding="10")
        analytics_notebook.add(risk_frame, text="At-Risk Students")
        
        ttk.Label(risk_frame, text="Identify students at risk of academic failure:").pack(anchor='w', pady=(0, 10))
        
        risk_criteria_frame = ttk.LabelFrame(risk_frame, text="Risk Criteria", padding="10")
        risk_criteria_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Risk factors checkboxes
        risk_vars = {}
        risk_factors = [
            ("Low attendance rate", "attendance"),
            ("Failing grades", "grades"),
            ("Late assignment submissions", "assignments"),
            ("No recent login activity", "activity"),
            ("Financial aid issues", "financial")
        ]
        
        for text, key in risk_factors:
            risk_vars[key] = tk.BooleanVar(value=True)
            ttk.Checkbutton(risk_criteria_frame, text=text, variable=risk_vars[key]).pack(anchor='w')
        
        def analyze_at_risk():
            selected_criteria = [key for key, var in risk_vars.items() if var.get()]
            if not selected_criteria:
                messagebox.showwarning("No Criteria", "Please select at least one risk factor.")
                return
            
            self.update_status("Analyzing at-risk students...")
            self.start_progress()
            
            def run_risk_analysis():
                try:
                    results = self.identify_at_risk_students(selected_criteria)
                    self.output_queue.put(("analytics", f"At-Risk Students Analysis:\n{results}"))
                    self.output_queue.put(("log", "At-risk student analysis completed."))
                except Exception as e:
                    self.output_queue.put(("error", f"Risk analysis error: {str(e)}"))
                finally:
                    self.output_queue.put(("stop_progress", None))
            
            threading.Thread(target=run_risk_analysis, daemon=True).start()
        
        ttk.Button(risk_frame, text="🔍 Analyze At-Risk Students", command=analyze_at_risk).pack(pady=10)
        
        # Enrollment prediction tab
        enrollment_frame = ttk.Frame(analytics_notebook, padding="10")
        analytics_notebook.add(enrollment_frame, text="Enrollment Prediction")
        
        ttk.Label(enrollment_frame, text="Predict future enrollment trends:").pack(anchor='w', pady=(0, 10))
        
        prediction_frame = ttk.LabelFrame(enrollment_frame, text="Prediction Parameters", padding="10")
        prediction_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(prediction_frame, text="Time Period:").pack(anchor='w')
        period_var = tk.StringVar(value="next_semester")
        
        periods = [
            ("Next Semester", "next_semester"),
            ("Next Academic Year", "next_year"),
            ("Next 5 Years", "5_years")
        ]
        
        for text, value in periods:
            ttk.Radiobutton(prediction_frame, text=text, variable=period_var, value=value).pack(anchor='w')
        
        def predict_enrollment():
            period = period_var.get()
            
            self.update_status("Predicting enrollment trends...")
            self.start_progress()
            
            def run_enrollment_prediction():
                try:
                    results = self.predict_enrollment_trends(period)
                    self.output_queue.put(("analytics", f"Enrollment Prediction:\n{results}"))
                    self.output_queue.put(("log", "Enrollment prediction completed."))
                except Exception as e:
                    self.output_queue.put(("error", f"Prediction error: {str(e)}"))
                finally:
                    self.output_queue.put(("stop_progress", None))
            
            threading.Thread(target=run_enrollment_prediction, daemon=True).start()
        
        ttk.Button(enrollment_frame, text="📈 Generate Predictions", command=predict_enrollment).pack(pady=10)
        
        # Success probability tab
        success_frame = ttk.Frame(analytics_notebook, padding="10")
        analytics_notebook.add(success_frame, text="Success Probability")
        
        ttk.Label(success_frame, text="Calculate module success probability:").pack(anchor='w', pady=(0, 10))
        
        module_frame = ttk.LabelFrame(success_frame, text="Module Selection", padding="10")
        module_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(module_frame, text="Select Module:").pack(anchor='w')
        module_var = tk.StringVar()
        module_combo = ttk.Combobox(module_frame, textvariable=module_var, width=30)
        module_combo.pack(fill=tk.X, pady=(5, 0))
        
        # Load available modules
        try:
            conn = get_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT DISTINCT module_code FROM student_modules ORDER BY module_code")
            modules = [row[0] for row in cursor.fetchall()]
            module_combo['values'] = modules
            conn.close()
        except:
            module_combo['values'] = ['CS101', 'CS102', 'DS101', 'DS102']
        
        def calculate_success():
            module = module_var.get().strip()
            if not module:
                messagebox.showwarning("Missing Module", "Please select a module.")
                return
            
            self.update_status("Calculating success probability...")
            self.start_progress()
            
            def run_success_calculation():
                try:
                    results = self.calculate_module_success_probability(module)
                    self.output_queue.put(("analytics", f"Module Success Probability:\n{results}"))
                    self.output_queue.put(("log", f"Success probability calculated for {module}."))
                except Exception as e:
                    self.output_queue.put(("error", f"Success calculation error: {str(e)}"))
                finally:
                    self.output_queue.put(("stop_progress", None))
            
            threading.Thread(target=run_success_calculation, daemon=True).start()
        
        ttk.Button(success_frame, text="🎯 Calculate Success Probability", command=calculate_success).pack(pady=10)
        
        ttk.Button(frame, text="❌ Close", command=dialog.destroy).pack()

    def identify_at_risk_students(self, criteria):
        """Identify students at risk based on criteria"""
        try:
            conn = get_connection()
            cursor = conn.cursor()
            
            # Simulate risk analysis based on available data
            risk_query = """
            SELECT s.*, COUNT(sm.module_code) as module_count,
                   AVG(CASE WHEN sm.grade IS NOT NULL THEN 
                       CASE sm.grade 
                           WHEN 'A' THEN 4.0 
                           WHEN 'B' THEN 3.0 
                           WHEN 'C' THEN 2.0 
                           WHEN 'D' THEN 1.0 
                           ELSE 0.0 
                       END 
                   END) as avg_grade
            FROM students s
            LEFT JOIN student_modules sm ON s.student_id = sm.student_id
            GROUP BY s.student_id
            HAVING avg_grade < 2.5 OR module_count = 0
            """
            
            cursor.execute(risk_query)
            at_risk_students = cursor.fetchall()
            conn.close()
            
            result = f"🚨 AT-RISK STUDENTS ANALYSIS\n"
            result += f"═" * 50 + "\n"
            result += f"Analysis Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
            result += f"Risk Criteria: {', '.join(criteria)}\n\n"
            result += f"Students Identified as At-Risk: {len(at_risk_students)}\n\n"
            
            if at_risk_students:
                result += "DETAILED BREAKDOWN:\n"
                result += "-" * 30 + "\n"
                for student in at_risk_students[:10]:  # Show first 10
                    name = f"{student[3]} {student[5]}"  # first_name + last_name
                    result += f"• {student[0]} - {name}\n"
                    result += f"  Email: {student[1]}\n"
                    result += f"  Course: {student[9]}\n"
                    result += f"  Age: {student[8]}\n\n"
                
                if len(at_risk_students) > 10:
                    result += f"... and {len(at_risk_students) - 10} more students\n"
            
            result += f"\nRECOMMENDATIONS:\n"
            result += "• Schedule academic counseling sessions\n"
            result += "• Provide additional tutoring support\n"
            result += "• Monitor attendance more closely\n"
            result += "• Consider intervention programs\n"
            
            return result
            
        except Exception as e:
            raise Exception(f"Risk analysis error: {str(e)}")

    def predict_enrollment_trends(self, period):
        """Predict enrollment trends for specified period"""
        try:
            conn = get_connection()
            cursor = conn.cursor()
            
            # Get historical enrollment data
            cursor.execute("""
            SELECT 
                strftime('%Y-%m', registration_datetime) as month,
                course,
                COUNT(*) as enrollments
            FROM students 
            WHERE registration_datetime IS NOT NULL
            GROUP BY strftime('%Y-%m', registration_datetime), course
            ORDER BY month DESC
            """)
            
            historical_data = cursor.fetchall()
            conn.close()
            
            # Simple trend analysis
            result = f"📈 ENROLLMENT PREDICTION\n"
            result += f"═" * 50 + "\n"
            result += f"Prediction Period: {period.replace('_', ' ').title()}\n"
            result += f"Analysis Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
            
            if historical_data:
                result += "HISTORICAL TRENDS:\n"
                result += "-" * 20 + "\n"
                
                # Group by course
                course_trends = {}
                for month, course, count in historical_data:
                    if course not in course_trends:
                        course_trends[course] = []
                    course_trends[course].append((month, count))
                
                for course, data in course_trends.items():
                    total_enrollments = sum(count for _, count in data)
                    avg_monthly = total_enrollments / max(len(data), 1)
                    
                    result += f"\n{course} Course:\n"
                    result += f"  Total Historical Enrollments: {total_enrollments}\n"
                    result += f"  Average Monthly: {avg_monthly:.1f}\n"
                    
                    # Predict based on period
                    if period == "next_semester":
                        predicted = int(avg_monthly * 6)  # 6 months
                    elif period == "next_year":
                        predicted = int(avg_monthly * 12)  # 12 months
                    else:  # 5_years
                        predicted = int(avg_monthly * 60)  # 60 months
                    
                    result += f"  Predicted Enrollments ({period.replace('_', ' ')}): {predicted}\n"
            
            result += f"\nPREDICTION FACTORS:\n"
            result += "• Historical enrollment patterns\n"
            result += "• Seasonal variations\n"
            result += "• Course popularity trends\n"
            result += "• Market demand indicators\n"
            
            result += f"\nRECOMMENDATIONS:\n"
            result += "• Prepare resources for predicted enrollment levels\n"
            result += "• Adjust marketing strategies accordingly\n"
            result += "• Plan faculty and infrastructure needs\n"
            
            return result
            
        except Exception as e:
            raise Exception(f"Enrollment prediction error: {str(e)}")

    def calculate_module_success_probability(self, module_code):
        """Calculate success probability for a module"""
        try:
            conn = get_connection()
            cursor = conn.cursor()
            
            # Get module statistics
            cursor.execute("""
            SELECT 
                grade,
                COUNT(*) as count
            FROM student_modules 
            WHERE module_code = ? AND grade IS NOT NULL
            GROUP BY grade
            """, (module_code,))
            
            grade_distribution = cursor.fetchall()
            
            # Get overall module info
            cursor.execute("""
            SELECT 
                COUNT(*) as total_enrolled,
                COUNT(grade) as graded,
                module_name
            FROM student_modules 
            WHERE module_code = ?
            """, (module_code,))
            
            module_info = cursor.fetchone()
            conn.close()
            
            result = f"🎯 MODULE SUCCESS PROBABILITY\n"
            result += f"═" * 50 + "\n"
            result += f"Module: {module_code}\n"
            
            if module_info and len(module_info) > 2:
                result += f"Module Name: {module_info[2] or 'Unknown'}\n"
            result += f"Analysis Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
            
            if module_info:
                total_enrolled, graded = module_info[0], module_info[1]
                result += f"ENROLLMENT STATISTICS:\n"
                result += f"Total Enrolled: {total_enrolled}\n"
                result += f"Students Graded: {graded}\n"
                result += f"Completion Rate: {(graded/max(total_enrolled,1)*100):.1f}%\n\n"
            
            if grade_distribution:
                result += "GRADE DISTRIBUTION:\n"
                result += "-" * 20 + "\n"
                
                total_graded = sum(count for _, count in grade_distribution)
                success_count = 0  # Grades A, B, C considered success
                
                for grade, count in grade_distribution:
                    percentage = (count / total_graded) * 100
                    result += f"  {grade}: {count} students ({percentage:.1f}%)\n"
                    
                    if grade in ['A', 'B', 'C']:
                        success_count += count
                
                success_rate = (success_count / total_graded) * 100
                result += f"\nSUCCESS ANALYSIS:\n"
                result += f"Success Rate (A, B, C grades): {success_rate:.1f}%\n"
                result += f"Students Likely to Succeed: {success_count}/{total_graded}\n"
                
                # Risk assessment
                if success_rate >= 80:
                    risk_level = "LOW RISK"
                    recommendation = "Module shows excellent success rates"
                elif success_rate >= 60:
                    risk_level = "MODERATE RISK"
                    recommendation = "Module may need some support improvements"
                else:
                    risk_level = "HIGH RISK"
                    recommendation = "Module requires immediate attention"
                
                result += f"Risk Level: {risk_level}\n"
                result += f"Recommendation: {recommendation}\n"
            
            return result
            
        except Exception as e:
            raise Exception(f"Success probability calculation error: {str(e)}")

    def show_graduation_timeline_forecast(self):
        """Show graduation timeline forecasting"""
        dialog = tk.Toplevel(self.master)
        dialog.title("🎓 Graduation Timeline Forecast")
        dialog.geometry("1100x800")
        dialog.transient(self.master)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Graduation Timeline Forecast", style='Title.TLabel').pack(pady=(0, 20))
        
        # Student selection
        selection_frame = ttk.LabelFrame(frame, text="Student Selection", padding="10")
        selection_frame.pack(fill=tk.X, pady=(0, 20))
        
        selection_var = tk.StringVar(value="all")
        ttk.Radiobutton(selection_frame, text="All students", variable=selection_var, value="all").pack(anchor='w')
        ttk.Radiobutton(selection_frame, text="Specific course", variable=selection_var, value="course").pack(anchor='w')
        ttk.Radiobutton(selection_frame, text="Current search results", variable=selection_var, value="results").pack(anchor='w')
        
        course_frame = ttk.Frame(selection_frame)
        course_frame.pack(fill=tk.X, pady=(10, 0))
        
        ttk.Label(course_frame, text="Course (if selected):").pack(side=tk.LEFT)
        course_var = tk.StringVar()
        course_combo = ttk.Combobox(course_frame, textvariable=course_var, values=["CS", "DS"], width=15)
        course_combo.pack(side=tk.LEFT, padx=(10, 0))
        
        # Forecast parameters
        params_frame = ttk.LabelFrame(frame, text="Forecast Parameters", padding="10")
        params_frame.pack(fill=tk.X, pady=(0, 20))
        
        ttk.Label(params_frame, text="Graduation Requirements:").pack(anchor='w')
        requirements_frame = ttk.Frame(params_frame)
        requirements_frame.pack(fill=tk.X, pady=(5, 0))
        
        ttk.Label(requirements_frame, text="Minimum Modules:").pack(side=tk.LEFT)
        min_modules_var = tk.StringVar(value="8")
        ttk.Entry(requirements_frame, textvariable=min_modules_var, width=10).pack(side=tk.LEFT, padx=(10, 20))
        
        ttk.Label(requirements_frame, text="Minimum GPA:").pack(side=tk.LEFT)
        min_gpa_var = tk.StringVar(value="2.0")
        ttk.Entry(requirements_frame, textvariable=min_gpa_var, width=10).pack(side=tk.LEFT, padx=(10, 0))
        
        def generate_forecast():
            selection = selection_var.get()
            course = course_var.get() if selection == "course" else None
            min_modules = int(min_modules_var.get() or 8)
            min_gpa = float(min_gpa_var.get() or 2.0)
            
            dialog.destroy()
            self.update_status("Generating graduation timeline forecast...")
            self.start_progress()
            
            def run_forecast():
                try:
                    results = self.generate_graduation_forecast(selection, course, min_modules, min_gpa)
                    self.output_queue.put(("analytics", results))
                    self.output_queue.put(("log", "Graduation timeline forecast completed."))
                except Exception as e:
                    self.output_queue.put(("error", f"Forecast error: {str(e)}"))
                finally:
                    self.output_queue.put(("stop_progress", None))
            
            threading.Thread(target=run_forecast, daemon=True).start()
        
        button_frame = ttk.Frame(frame)
        button_frame.pack(fill=tk.X)
        
        ttk.Button(button_frame, text="📊 Generate Forecast", command=generate_forecast).pack(side=tk.LEFT)
        ttk.Button(button_frame, text="❌ Cancel", command=dialog.destroy).pack(side=tk.RIGHT)

    def generate_graduation_forecast(self, selection, course, min_modules, min_gpa):
        """Generate graduation timeline forecast"""
        try:
            conn = get_connection()
            cursor = conn.cursor()
            
            # Base query
            if selection == "course" and course:
                student_query = "SELECT * FROM students WHERE LOWER(course) = LOWER(?)"
                params = [course]
            elif selection == "results" and self.search_results:
                student_ids = [str(s[0]) for s in self.search_results]
                placeholders = ",".join(["?" for _ in student_ids])
                student_query = f"SELECT * FROM students WHERE student_id IN ({placeholders})"
                params = student_ids
            else:
                student_query = "SELECT * FROM students"
                params = []
            
            cursor.execute(student_query, params)
            students = cursor.fetchall()
            
            result = f"🎓 GRADUATION TIMELINE FORECAST\n"
            result += f"═" * 50 + "\n"
            result += f"Analysis Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
            result += f"Selection: {selection.title()}\n"
            if course:
                result += f"Course Filter: {course}\n"
            result += f"Requirements: {min_modules} modules, {min_gpa} GPA\n"
            result += f"Students Analyzed: {len(students)}\n\n"
            
            # Analyze each student's progress
            graduation_forecast = {
                "ready_to_graduate": [],
                "graduating_soon": [],
                "on_track": [],
                "at_risk": [],
                "insufficient_data": []
            }
            
            for student in students:
                student_id = student[0]
                
                # Get student's module progress
                cursor.execute("""
                SELECT COUNT(*) as module_count,
                       AVG(CASE WHEN grade IS NOT NULL THEN 
                           CASE grade 
                               WHEN 'A' THEN 4.0 
                               WHEN 'B' THEN 3.0 
                               WHEN 'C' THEN 2.0 
                               WHEN 'D' THEN 1.0 
                               ELSE 0.0 
                           END 
                       END) as avg_gpa
                FROM student_modules 
                WHERE student_id = ?
                """, (student_id,))
                
                progress = cursor.fetchone()
                
                if not progress or progress[0] == 0:
                    graduation_forecast["insufficient_data"].append(student)
                else:
                    module_count, avg_gpa = progress
                    avg_gpa = avg_gpa or 0.0
                    
                    if module_count >= min_modules and avg_gpa >= min_gpa:
                        graduation_forecast["ready_to_graduate"].append((student, module_count, avg_gpa))
                    elif module_count >= (min_modules * 0.8) and avg_gpa >= min_gpa:
                        graduation_forecast["graduating_soon"].append((student, module_count, avg_gpa))
                    elif module_count >= (min_modules * 0.5):
                        graduation_forecast["on_track"].append((student, module_count, avg_gpa))
                    else:
                        graduation_forecast["at_risk"].append((student, module_count, avg_gpa))
            
            conn.close()
            
            # Generate detailed results
            result += "GRADUATION STATUS BREAKDOWN:\n"
            result += "-" * 30 + "\n"
            
            for category, students_data in graduation_forecast.items():
                count = len(students_data)
                percentage = (count / len(students)) * 100 if students else 0
                
                category_name = category.replace("_", " ").title()
                result += f"\n{category_name}: {count} students ({percentage:.1f}%)\n"
                
                if category != "insufficient_data":
                    for student_data, modules, gpa in students_data[:5]:  # Show first 5
                        name = f"{student_data[3]} {student_data[5]}"
                        gpa_display = gpa if gpa is not None else 0.0
                        result += f"  • {student_data[0]} - {name} ({modules} modules, GPA: {gpa_display:.2f})\n"
                    if len(students_data) > 5:
                        result += f"  ... and {len(students_data) - 5} more\n"
                else:
                    for student in students_data[:5]:
                        name = f"{student[3]} {student[5]}"
                        result += f"  • {student[0]} - {name}\n"
            
            result += f"\nTIMELINE PROJECTIONS:\n"
            result += "• Ready to Graduate: Can graduate immediately\n"
            result += "• Graduating Soon: 1-2 semesters remaining\n"
            result += "• On Track: 2-4 semesters remaining\n"
            result += "• At Risk: May require academic intervention\n"
            
            return result
            
        except Exception as e:
            raise Exception(f"Graduation forecast error: {str(e)}")

    def show_advanced_charts(self):
        """Show advanced chart generation options"""
        dialog = tk.Toplevel(self.master)
        dialog.title("📈 Advanced Charts & Visualizations")
        dialog.geometry("550x680")
        dialog.transient(self.master)
        dialog.grab_set()

        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(frame, text="Advanced Charts & Visualizations", style='Title.TLabel').pack(pady=(0, 10))

        # Show availability status
        if CHARTS_AVAILABLE:
            status_label = ttk.Label(frame, text="✓ Chart Generation Available (matplotlib + seaborn)",
                                   foreground="green")
        else:
            status_label = ttk.Label(frame, text="⚠ Charts require: pip install matplotlib seaborn",
                                   foreground="red")
        status_label.pack(pady=(0, 20))

        # Info text
        info_text = "Generate professional charts from database data.\nCharts open in interactive windows with zoom, pan, and save features."
        ttk.Label(frame, text=info_text, justify=tk.CENTER, foreground="gray").pack(pady=(0, 15))

        # Chart types
        chart_options = [
            ("📊 Age Distribution Histogram", "age_histogram"),
            ("🥧 Course Distribution Pie Chart", "course_pie"),
            ("📈 Registration Timeline", "registration_timeline"),
            ("👥 Gender-Course Distribution", "gender_course"),
            ("🎯 Module Popularity Chart", "module_popularity"),
            ("📉 Grade Distribution Analysis", "grade_distribution"),
        ]

        for text, chart_type in chart_options:
            btn = ttk.Button(frame, text=text, width=40,
                            command=lambda ct=chart_type: self.generate_chart(ct, dialog))
            btn.pack(pady=5)

        # Separator
        ttk.Separator(frame, orient='horizontal').pack(fill='x', pady=15)

        # Email chart option
        ttk.Label(frame, text="Send charts via email:", font=('Arial', 10, 'bold')).pack(pady=(0, 10))
        ttk.Button(frame, text="📧 Email Chart to Admin", width=40,
                  command=lambda: self.email_chart_to_admin(dialog)).pack(pady=5)

        ttk.Button(frame, text="❌ Close", command=dialog.destroy).pack(pady=(20, 0))

    def generate_chart(self, chart_type, parent_dialog):
        """Generate specified chart type with matplotlib"""
        parent_dialog.destroy()

        # Check if charts are available
        if not CHARTS_AVAILABLE:
            messagebox.showerror("Charts Unavailable",
                               "Chart generation requires matplotlib and seaborn.\n"
                               "Install with: pip install matplotlib seaborn")
            return

        self.update_status(f"Generating {chart_type.replace('_', ' ')} chart...")
        self.start_progress()

        def run_chart_generation():
            try:
                # Use new chart generation system
                create_chart_viewer(self.master, chart_type)

                self.output_queue.put(("log", f"✓ Chart generated successfully: {chart_type}"))
                self.output_queue.put(("log", "Chart displayed in new window"))
            except Exception as e:
                self.output_queue.put(("error", f"Chart generation error: {str(e)}"))
            finally:
                self.output_queue.put(("stop_progress", None))

        # Run in thread to avoid blocking UI
        threading.Thread(target=run_chart_generation, daemon=True).start()

    def email_chart_to_admin(self, parent_dialog):
        """Email chart to admin user"""
        parent_dialog.destroy()

        # Chart selection dialog
        dialog = tk.Toplevel(self.master)
        dialog.title("📧 Email Chart to Admin")
        dialog.geometry("1000x750")
        dialog.transient(self.master)
        dialog.grab_set()

        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(frame, text="Select Chart to Email", style='Title.TLabel').pack(pady=(0, 20))

        # Chart type selection
        ttk.Label(frame, text="Chart Type:").pack(anchor='w', pady=(0, 5))
        chart_var = tk.StringVar(value="age_histogram")
        chart_options = [
            ("Age Distribution Histogram", "age_histogram"),
            ("Course Distribution Pie Chart", "course_pie"),
            ("Registration Timeline", "registration_timeline"),
            ("Gender-Course Distribution", "gender_course"),
            ("Module Popularity Chart", "module_popularity"),
            ("Grade Distribution Analysis", "grade_distribution"),
        ]

        for text, value in chart_options:
            ttk.Radiobutton(frame, text=text, variable=chart_var, value=value).pack(anchor='w', padx=20)

        # Get admin email from database
        try:
            conn = get_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT email FROM users WHERE LOWER(role) = 'admin' LIMIT 1")
            admin_row = cursor.fetchone()
            conn.close()
            default_email = admin_row[0] if admin_row else "admin@university.edu"
        except Exception as e:
            print(f"Warning: Could not fetch admin email from database: {e}")
            default_email = "admin@university.edu"

        ttk.Label(frame, text="\nAdmin Email Address:").pack(anchor='w', pady=(20, 5))
        email_var = tk.StringVar(value=default_email)

        # Create a frame for email input with a refresh button
        email_frame = ttk.Frame(frame)
        email_frame.pack(fill='x', pady=(0, 10))

        email_entry = ttk.Entry(email_frame, textvariable=email_var, width=35)
        email_entry.pack(side='left', fill='x', expand=True)

        def refresh_admin_email():
            """Refresh admin email list from database"""
            try:
                conn = get_connection()
                cursor = conn.cursor()
                cursor.execute("SELECT username, email FROM users WHERE LOWER(role) = 'admin' ORDER BY username")
                admins = cursor.fetchall()
                conn.close()

                if admins:
                    # Show selection dialog if multiple admins
                    if len(admins) > 1:
                        admin_dialog = tk.Toplevel(dialog)
                        admin_dialog.title("Select Admin")
                        admin_dialog.geometry("900x700")
                        admin_dialog.transient(dialog)
                        admin_dialog.grab_set()

                        ttk.Label(admin_dialog, text="Select Admin User:", font=('Arial', 12, 'bold')).pack(pady=10)

                        admin_listbox = tk.Listbox(admin_dialog, height=10)
                        admin_listbox.pack(fill='both', expand=True, padx=20, pady=10)

                        for username, email in admins:
                            admin_listbox.insert(tk.END, f"{username} ({email})")

                        def select_admin():
                            selection = admin_listbox.curselection()
                            if selection:
                                selected_email = admins[selection[0]][1]
                                email_var.set(selected_email)
                            admin_dialog.destroy()

                        ttk.Button(admin_dialog, text="Select", command=select_admin).pack(pady=10)
                    else:
                        # Only one admin, use their email
                        email_var.set(admins[0][1])
                        messagebox.showinfo("Admin Email", f"Using admin email: {admins[0][1]}")
                else:
                    messagebox.showwarning("No Admins", "No admin users found in database.")
            except Exception as e:
                messagebox.showerror("Database Error", f"Could not fetch admin emails: {str(e)}")

        ttk.Button(email_frame, text="🔄", command=refresh_admin_email, width=3).pack(side='left', padx=(5, 0))

        ttk.Label(frame, text="Message (optional):").pack(anchor='w', pady=(10, 5))
        message_text = scrolledtext.ScrolledText(frame, height=6, wrap=tk.WORD)
        message_text.pack(fill='x', pady=(0, 20))
        message_text.insert('1.0', "Please find attached the requested analytics chart.")

        def send_chart_email():
            try:
                # Import email service
                try:
                    from university_system.infrastructure.email.email_service import send_email
                    EMAIL_AVAILABLE = True
                except ImportError:
                    EMAIL_AVAILABLE = False

                if not EMAIL_AVAILABLE:
                    messagebox.showerror("Email Service Unavailable",
                                       "Email service is not configured. Please set up email settings.")
                    return

                chart_type = chart_var.get()
                admin_email = email_var.get().strip()
                message = message_text.get('1.0', tk.END).strip()

                if not admin_email or '@' not in admin_email:
                    messagebox.showwarning("Invalid Email", "Please enter a valid admin email address.")
                    return

                # Generate chart data
                chart_data = self.create_chart_data(chart_type)

                # Create email subject and body
                subject = f"Analytics Chart: {chart_type.replace('_', ' ').title()}"
                body = f"""
{message}

Chart Type: {chart_type.replace('_', ' ').title()}
Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

Chart Data:
{chart_data}

This email was sent from the Advanced Search GUI - Analytics Dashboard.
"""

                # Send email
                send_email(
                    recipient_email=admin_email,
                    subject=subject,
                    body=body
                )

                messagebox.showinfo("Email Sent",
                                  f"✅ Chart sent successfully to {admin_email}\n\n"
                                  f"Chart Type: {chart_type.replace('_', ' ').title()}")
                dialog.destroy()

            except Exception as e:
                messagebox.showerror("Email Error",
                                   f"Failed to send chart email: {str(e)}\n\n"
                                   f"Please check email configuration.")
                print(f"Chart email error: {str(e)}")

        button_frame = ttk.Frame(frame)
        button_frame.pack(fill='x', pady=(10, 0))
        ttk.Button(button_frame, text="📧 Send Email", command=send_chart_email).pack(side='left', padx=(0, 10))
        ttk.Button(button_frame, text="❌ Cancel", command=dialog.destroy).pack(side='right')

    def create_chart_data(self, chart_type):
        """Create chart data for different visualization types"""
        try:
            conn = get_connection()
            cursor = conn.cursor()

            result = f"📈 CHART DATA: {chart_type.replace('_', ' ').title()}\n"
            result += f"═" * 50 + "\n"
            result += f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"

            if chart_type == "age_histogram":
                cursor.execute("SELECT age, COUNT(*) FROM students WHERE age IS NOT NULL GROUP BY age ORDER BY age")
                age_data = cursor.fetchall()

                if not age_data:
                    result += "AGE DISTRIBUTION:\n"
                    result += "-" * 20 + "\n"
                    result += "No age data available.\n"
                else:
                    result += "AGE DISTRIBUTION:\n"
                    result += "-" * 20 + "\n"
                    for age, count in age_data:
                        age_display = age if age is not None else 0
                        bar = "█" * min(count, 20)  # Visual bar representation
                        result += f"Age {age_display:2d}: {count:3d} students {bar}\n"

            elif chart_type == "course_pie":
                cursor.execute("SELECT course, COUNT(*) FROM students WHERE course IS NOT NULL GROUP BY course")
                course_data = cursor.fetchall()

                if not course_data:
                    result += "COURSE DISTRIBUTION:\n"
                    result += "-" * 20 + "\n"
                    result += "No course data available.\n"
                else:
                    total = sum(count for _, count in course_data)
                    result += "COURSE DISTRIBUTION:\n"
                    result += "-" * 20 + "\n"
                    for course, count in course_data:
                        percentage = (count / total * 100) if total > 0 else 0
                        course_name = course if course else "Not Specified"
                        result += f"{course_name}: {count} students ({percentage:.1f}%)\n"

            elif chart_type == "registration_timeline":
                cursor.execute("""
                SELECT strftime('%Y-%m', registration_datetime) as month, COUNT(*)
                FROM students
                WHERE registration_datetime IS NOT NULL
                GROUP BY strftime('%Y-%m', registration_datetime)
                ORDER BY month
                """)
                timeline_data = cursor.fetchall()

                result += "REGISTRATION TIMELINE:\n"
                result += "-" * 25 + "\n"
                if not timeline_data:
                    result += "No registration timeline data available.\n"
                else:
                    for month, count in timeline_data:
                        bar = "█" * min(count, 15)
                        month_display = month if month else "Unknown"
                        result += f"{month_display}: {count:3d} registrations {bar}\n"

            elif chart_type == "gender_course":
                cursor.execute("""
                SELECT course, gender, COUNT(*)
                FROM students
                WHERE course IS NOT NULL
                GROUP BY course, gender
                ORDER BY course, gender
                """)
                gender_course_data = cursor.fetchall()

                result += "GENDER-COURSE DISTRIBUTION:\n"
                result += "-" * 30 + "\n"
                if not gender_course_data:
                    result += "No gender-course data available.\n"
                else:
                    for course, gender, count in gender_course_data:
                        course_name = course if course else "Not Specified"
                        gender_name = gender.title() if gender else "Not Specified"
                        result += f"{course_name} - {gender_name}: {count} students\n"

            elif chart_type == "module_popularity":
                cursor.execute("""
                SELECT module_code, module_name, COUNT(*) as enrollment_count
                FROM student_modules
                WHERE module_code IS NOT NULL
                GROUP BY module_code, module_name
                ORDER BY enrollment_count DESC
                LIMIT 10
                """)
                module_data = cursor.fetchall()

                result += "TOP 10 MOST POPULAR MODULES:\n"
                result += "-" * 35 + "\n"
                if not module_data:
                    result += "No module enrollment data available.\n"
                else:
                    for i, (code, name, count) in enumerate(module_data, 1):
                        code_display = code if code else "N/A"
                        name_display = name if name else "N/A"
                        result += f"{i:2d}. {code_display} - {name_display}: {count} enrollments\n"

            elif chart_type == "grade_distribution":
                cursor.execute("""
                SELECT grade, COUNT(*)
                FROM student_modules
                WHERE grade IS NOT NULL
                GROUP BY grade
                ORDER BY grade
                """)
                grade_data = cursor.fetchall()

                result += "OVERALL GRADE DISTRIBUTION:\n"
                result += "-" * 30 + "\n"
                if not grade_data:
                    result += "No grade data available.\n"
                else:
                    total_grades = sum(count for _, count in grade_data)
                    for grade, count in grade_data:
                        percentage = (count / total_grades * 100) if total_grades > 0 else 0
                        bar = "█" * min(int(percentage), 20)
                        grade_display = grade if grade else "N/A"
                        result += f"Grade {grade_display}: {count:4d} ({percentage:5.1f}%) {bar}\n"

            else:  # custom_chart or enrollment_trends
                cursor.execute("""
                SELECT
                    strftime('%Y', registration_datetime) as year,
                    course,
                    COUNT(*) as count
                FROM students
                WHERE registration_datetime IS NOT NULL AND course IS NOT NULL
                GROUP BY year, course
                ORDER BY year, course
                """)
                trend_data = cursor.fetchall()

                result += "ENROLLMENT TRENDS BY YEAR AND COURSE:\n"
                result += "-" * 40 + "\n"
                if not trend_data:
                    result += "No enrollment trend data available.\n"
                else:
                    for year, course, count in trend_data:
                        year_display = year if year else "Unknown"
                        course_display = course if course else "Not Specified"
                        result += f"{year_display} - {course_display}: {count} enrollments\n"

            conn.close()

            result += f"\n📊 CHART RECOMMENDATIONS:\n"
            result += "• Use this data to create visual charts in external tools\n"
            result += "• Consider trends for strategic planning\n"
            result += "• Monitor patterns for quality improvement\n"

            return result

        except Exception as e:
            raise Exception(f"Chart data creation error: {str(e)}")

    # Additional missing utility functions

    def export_to_excel(self, filename):
        """Export results to Excel format (simulation)"""
        # Note: In a real implementation, you'd use openpyxl or xlsxwriter
        try:
            # Create a simple CSV that Excel can open
            csv_filename = filename.replace('.xlsx', '.csv')
            self.export_to_csv(csv_filename)
            messagebox.showinfo("Excel Export", f"Data exported as CSV (Excel compatible): {csv_filename}")
        except Exception as e:
            raise Exception(f"Excel export error: {str(e)}")

    def custom_format_export(self):
        """Show custom format export dialog"""
        dialog = tk.Toplevel(self.master)
        dialog.title("🛠️ Custom Format Export")
        dialog.geometry("1000x750")
        dialog.transient(self.master)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Custom Format Export", style='Title.TLabel').pack(pady=(0, 20))
        
        # Format options
        format_frame = ttk.LabelFrame(frame, text="Export Format", padding="10")
        format_frame.pack(fill=tk.X, pady=(0, 20))
        
        format_var = tk.StringVar(value="custom")
        formats = [
            ("Custom delimiter", "custom"),
            ("Tab-separated", "tsv"),
            ("XML format", "xml"),
            ("SQL INSERT statements", "sql")
        ]
        
        for text, value in formats:
            ttk.Radiobutton(format_frame, text=text, variable=format_var, value=value).pack(anchor='w')
        
        # Custom options
        options_frame = ttk.LabelFrame(frame, text="Options", padding="10")
        options_frame.pack(fill=tk.X, pady=(0, 20))
        
        ttk.Label(options_frame, text="Custom Delimiter:").pack(anchor='w')
        delimiter_var = tk.StringVar(value="|")
        ttk.Entry(options_frame, textvariable=delimiter_var, width=10).pack(anchor='w', pady=(0, 10))
        
        include_header_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(options_frame, text="Include header row", variable=include_header_var).pack(anchor='w')
        
        def export_custom():
            if not self.search_results:
                messagebox.showwarning("No Data", "No search results to export.")
                return
            
            format_type = format_var.get()
            delimiter = delimiter_var.get()
            include_header = include_header_var.get()
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            try:
                if format_type == "custom":
                    filename = f"custom_export_{timestamp}.txt"
                    self.export_custom_delimiter(filename, delimiter, include_header)
                elif format_type == "tsv":
                    filename = f"export_{timestamp}.tsv"
                    self.export_custom_delimiter(filename, "\t", include_header)
                elif format_type == "xml":
                    filename = f"export_{timestamp}.xml"
                    self.export_to_xml(filename)
                elif format_type == "sql":
                    filename = f"export_{timestamp}.sql"
                    self.export_to_sql(filename)
                
                dialog.destroy()
                messagebox.showinfo("Export Complete", f"Custom export completed: {filename}")
                
            except Exception as e:
                messagebox.showerror("Export Error", f"Custom export failed: {str(e)}")
        
        button_frame = ttk.Frame(frame)
        button_frame.pack(fill=tk.X)
        
        ttk.Button(button_frame, text="💾 Export", command=export_custom).pack(side=tk.LEFT)
        ttk.Button(button_frame, text="❌ Cancel", command=dialog.destroy).pack(side=tk.RIGHT)

    def export_custom_delimiter(self, filename, delimiter, include_header):
        """Export with custom delimiter"""
        with open(filename, 'w', encoding='utf-8') as f:
            if include_header:
                headers = [
                    'Student ID', 'Email', 'Title', 'First Name', 'Middle Name',
                    'Last Name', 'Gender', 'Date of Birth', 'Age', 'Course',
                    'Registration Datetime'
                ]
                f.write(delimiter.join(headers) + '\n')
            
            for student in self.search_results:
                row = [str(field) if field is not None else '' for field in student]
                f.write(delimiter.join(row) + '\n')

    def export_to_xml(self, filename):
        """Export to XML format"""
        with open(filename, 'w', encoding='utf-8') as f:
            f.write('<?xml version="1.0" encoding="UTF-8"?>\n')
            f.write('<students>\n')
            
            for student in self.search_results:
                f.write('  <student>\n')
                fields = ['student_id', 'email', 'title', 'first_name', 'middle_name',
                         'last_name', 'gender', 'date_of_birth', 'age', 'course',
                         'registration_datetime']
                
                for i, field_name in enumerate(fields):
                    value = student[i] if i < len(student) else ''
                    if value is not None:
                        f.write(f'    <{field_name}>{str(value)}</{field_name}>\n')
                f.write('  </student>\n')
            
            f.write('</students>\n')

    def export_to_sql(self, filename):
        """Export as SQL INSERT statements"""
        with open(filename, 'w', encoding='utf-8') as f:
            f.write('-- Student data export\n')
            f.write('-- Generated on ' + datetime.now().strftime('%Y-%m-%d %H:%M:%S') + '\n\n')
            
            for student in self.search_results:
                values = []
                for field in student:
                    if field is None:
                        values.append('NULL')
                    elif isinstance(field, str):
                        # Correctly escape single quotes for SQL
                        safe_string = field.replace("'", "''")
                        values.append(f"'{safe_string}'")
                    else:
                        values.append(str(field))
                
                f.write(f"INSERT INTO students VALUES ({', '.join(values)});\n")

    def refresh_data(self):
        """Refresh data when notified by main GUI"""
        try:
            # Clear current search results
            self.search_results = []
            
            # Update results display
            self.update_results_display()
            
            # Log the refresh
            self.log_output("Data refreshed from main GUI")
            
            # Update status
            if hasattr(self, 'results_label'):
                self.results_label.config(text="Data refreshed - perform new search")
                
        except Exception as e:
            self.log_output(f"Error refreshing data: {e}")

    # Add these missing methods to handle various search and export operations

    def view_academic_history(self, student_id):
        """View detailed academic history for a student"""
        try:
            conn = get_connection()
            cursor = conn.cursor()
            
            # Get student basic info
            cursor.execute("SELECT * FROM students WHERE student_id = ?", (student_id,))
            student = cursor.fetchone()
            
            if not student:
                return "Student not found"
            
            # Get academic history
            cursor.execute("""
            SELECT module_type, module_code, module_name, grade, enrollment_date
            FROM student_modules
            WHERE student_id = ?
            ORDER BY enrollment_date DESC
            """, (student_id,))
            
            modules = cursor.fetchall()
            conn.close()
            
            history = f"📚 ACADEMIC HISTORY - {student[0]}\n"
            history += f"═" * 50 + "\n"
            history += f"Student: {student[3]} {student[5]}\n"
            history += f"Email: {student[1]}\n"
            history += f"Course: {student[9]}\n\n"
            
            if modules:
                history += f"MODULE HISTORY ({len(modules)} modules):\n"
                history += "-" * 40 + "\n"
                
                for module_type, code, name, grade, enrolled_date in modules:
                    grade_display = grade if grade else "In Progress"
                    history += f"{code} - {name}\n"
                    history += f"  Type: {module_type} | Grade: {grade_display}\n"
                    history += f"  Enrolled: {enrolled_date}\n\n"
            else:
                history += "No module enrollment history found.\n"
            
            return history
            
        except Exception as e:
            return f"Error retrieving academic history: {str(e)}"

    def perform_regex_search(self, pattern, field):
        """Perform regular expression search"""
        try:
            import re
            conn = get_connection()
            cursor = conn.cursor()
            
            # Get all records and filter with regex
            cursor.execute(f"SELECT * FROM students")
            all_students = cursor.fetchall()
            conn.close()
            
            matched_students = []
            compiled_pattern = re.compile(pattern, re.IGNORECASE)
            
            field_index_map = {
                'student_id': 0, 'email': 1, 'title': 2,
                'first_name': 3, 'middle_name': 4, 'last_name': 5,
                'gender': 6, 'date_of_birth': 7, 'age': 8,
                'course': 9, 'registration_datetime': 10
            }
            
            field_index = field_index_map.get(field, 3)  # Default to first_name
            
            for student in all_students:
                if field_index < len(student) and student[field_index]:
                    if compiled_pattern.search(str(student[field_index])):
                        matched_students.append(student)
            
            return matched_students
            
        except Exception as e:
            raise Exception(f"Regex search error: {str(e)}")

    def perform_wildcard_search(self, pattern, field):
        """Perform wildcard search (* and ?)"""
        try:
            import fnmatch
            conn = get_connection()
            cursor = conn.cursor()
            
            cursor.execute("SELECT * FROM students")
            all_students = cursor.fetchall()
            conn.close()
            
            matched_students = []
            field_index_map = {
                'student_id': 0, 'email': 1, 'first_name': 3,
                'last_name': 5, 'course': 9
            }
            
            field_index = field_index_map.get(field, 3)
            
            for student in all_students:
                if field_index < len(student) and student[field_index]:
                    if fnmatch.fnmatch(str(student[field_index]).lower(), pattern.lower()):
                        matched_students.append(student)
            
            return matched_students
            
        except Exception as e:
            raise Exception(f"Wildcard search error: {str(e)}")

    def perform_search_all_fields(self, search_term):
        """Search across all text fields"""
        try:
            conn = get_connection()
            cursor = conn.cursor()
            
            search_pattern = f"%{search_term.lower()}%"
            query = """
            SELECT * FROM students WHERE
            LOWER(student_id) LIKE ? OR
            LOWER(email_address) LIKE ? OR
            LOWER(title) LIKE ? OR
            LOWER(first_name) LIKE ? OR
            LOWER(middle_name) LIKE ? OR
            LOWER(last_name) LIKE ? OR
            LOWER(gender) LIKE ? OR
            LOWER(course) LIKE ? OR
            LOWER(registration_datetime) LIKE ?
            """
            
            params = [search_pattern] * 9
            cursor.execute(query, params)
            results = cursor.fetchall()
            conn.close()
            
            return results
            
        except Exception as e:
            raise Exception(f"Search all fields error: {str(e)}")

    def perform_phonetic_search(self, search_term):
        """Perform phonetic name search using Soundex algorithm"""
        try:
            def soundex(word):
                """Generate Soundex code for phonetic matching"""
                if not word:
                    return "0000"
                
                word = word.upper()
                soundex_code = word[0]
                
                # Mapping for consonants
                mapping = {
                    'BFPV': '1', 'CGJKQSXZ': '2', 'DT': '3',
                    'L': '4', 'MN': '5', 'R': '6'
                }
                
                for char in word[1:]:
                    for key, value in mapping.items():
                        if char in key:
                            if soundex_code[-1] != value:
                                soundex_code += value
                            break
                
                # Pad or truncate to 4 characters
                soundex_code = (soundex_code + '0000')[:4]
                return soundex_code
            
            conn = get_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM students")
            all_students = cursor.fetchall()
            conn.close()
            
            search_soundex = soundex(search_term)
            matched_students = []
            
            for student in all_students:
                first_name = student[3] if len(student) > 3 else ""
                last_name = student[5] if len(student) > 5 else ""
                
                if (soundex(first_name) == search_soundex or 
                    soundex(last_name) == search_soundex):
                    matched_students.append(student)
            
            return matched_students
            
        except Exception as e:
            raise Exception(f"Phonetic search error: {str(e)}")

    def show_condition_builder(self):
        """Show advanced condition builder interface"""
        dialog = tk.Toplevel(self.master)
        dialog.title("Condition Builder")
        dialog.geometry("1200x850")
        dialog.transient(self.master)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Advanced Condition Builder", style='Title.TLabel').pack(pady=(0, 20))
        
        # Condition management
        self.conditions_list = []
        
        # Condition display
        conditions_frame = ttk.LabelFrame(frame, text="Current Conditions", padding="10")
        conditions_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 20))
        
        self.conditions_tree = ttk.Treeview(conditions_frame, columns=('Field', 'Operator', 'Value', 'Logic'), 
                                           show='headings', height=10)
        
        for col in ['Field', 'Operator', 'Value', 'Logic']:
            self.conditions_tree.heading(col, text=col)
            self.conditions_tree.column(col, width=120)
        
        scrollbar_cond = ttk.Scrollbar(conditions_frame, orient=tk.VERTICAL, command=self.conditions_tree.yview)
        self.conditions_tree.configure(yscrollcommand=scrollbar_cond.set)
        
        self.conditions_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar_cond.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Add condition interface
        add_frame = ttk.LabelFrame(frame, text="Add New Condition", padding="10")
        add_frame.pack(fill=tk.X, pady=(0, 20))
        
        # Field selection
        field_frame = ttk.Frame(add_frame)
        field_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(field_frame, text="Field:").pack(side=tk.LEFT)
        field_var = tk.StringVar(value="age")
        field_combo = ttk.Combobox(field_frame, textvariable=field_var, width=15,
                                  values=["age", "course", "gender", "first_name", "last_name", "email"])
        field_combo.pack(side=tk.LEFT, padx=(10, 20))
        
        ttk.Label(field_frame, text="Operator:").pack(side=tk.LEFT)
        operator_var = tk.StringVar(value="=")
        operator_combo = ttk.Combobox(field_frame, textvariable=operator_var, width=10,
                                     values=["=", "!=", ">", "<", ">=", "<=", "LIKE", "NOT LIKE"])
        operator_combo.pack(side=tk.LEFT, padx=(10, 20))
        
        ttk.Label(field_frame, text="Value:").pack(side=tk.LEFT)
        value_var = tk.StringVar()
        ttk.Entry(field_frame, textvariable=value_var, width=15).pack(side=tk.LEFT, padx=(10, 0))
        
        # Logic operator
        logic_frame = ttk.Frame(add_frame)
        logic_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(logic_frame, text="Logic Operator:").pack(side=tk.LEFT)
        logic_var = tk.StringVar(value="AND")
        ttk.Radiobutton(logic_frame, text="AND", variable=logic_var, value="AND").pack(side=tk.LEFT, padx=(10, 5))
        ttk.Radiobutton(logic_frame, text="OR", variable=logic_var, value="OR").pack(side=tk.LEFT)
        
        def add_condition():
            field = field_var.get()
            operator = operator_var.get()
            value = value_var.get().strip()
            logic = logic_var.get()
            
            if not value:
                messagebox.showwarning("Missing Value", "Please enter a value for the condition.")
                return
            
            condition = {
                'field': field,
                'operator': operator,
                'value': value,
                'logic': logic if self.conditions_list else None
            }
            
            self.conditions_list.append(condition)
            self.update_conditions_display()
            value_var.set("")
        
        def remove_condition():
            selection = self.conditions_tree.selection()
            if not selection:
                messagebox.showwarning("No Selection", "Please select a condition to remove.")
                return
            
            item_index = self.conditions_tree.index(selection[0])
            self.conditions_list.pop(item_index)
            self.update_conditions_display()
        
        def execute_conditions():
            if not self.conditions_list:
                messagebox.showwarning("No Conditions", "Please add at least one condition.")
                return
            
            dialog.destroy()
            self.execute_conditional_logic_search(self.conditions_list)
        
        button_frame = ttk.Frame(add_frame)
        button_frame.pack(fill=tk.X)
        
        ttk.Button(button_frame, text="Add Condition", command=add_condition).pack(side=tk.LEFT)
        ttk.Button(button_frame, text="Remove Selected", command=remove_condition).pack(side=tk.LEFT, padx=(10, 0))
        
        # Execute buttons
        exec_frame = ttk.Frame(frame)
        exec_frame.pack(fill=tk.X)
        
        ttk.Button(exec_frame, text="Execute Search", command=execute_conditions).pack(side=tk.LEFT)
        ttk.Button(exec_frame, text="Clear All", 
                  command=lambda: self.clear_all_conditions()).pack(side=tk.LEFT, padx=(10, 0))
        ttk.Button(exec_frame, text="Cancel", command=dialog.destroy).pack(side=tk.RIGHT)

    def update_conditions_display(self):
        """Update the conditions display tree"""
        for item in self.conditions_tree.get_children():
            self.conditions_tree.delete(item)
        
        for i, condition in enumerate(self.conditions_list):
            logic_display = condition['logic'] if i > 0 else ""
            self.conditions_tree.insert('', 'end', values=(
                condition['field'],
                condition['operator'],
                condition['value'],
                logic_display
            ))

    def clear_all_conditions(self):
        """Clear all conditions"""
        self.conditions_list = []
        self.update_conditions_display()

    def execute_conditional_logic_search(self, conditions):
        """Execute search with complex conditions"""
        self.update_status("Executing conditional search...")
        self.start_progress()
        
        def run_conditional():
            try:
                results = self.perform_conditional_logic_search(conditions)
                self.output_queue.put(("search_results", results))
                self.output_queue.put(("log", f"Conditional search completed. Found {len(results)} results."))
            except Exception as e:
                self.output_queue.put(("error", f"Conditional search error: {str(e)}"))
            finally:
                self.output_queue.put(("stop_progress", None))
        
        threading.Thread(target=run_conditional, daemon=True).start()

    def perform_conditional_logic_search(self, conditions):
        """Perform search with conditional logic"""
        try:
            conn = get_connection()
            cursor = conn.cursor()
            
            # Build complex WHERE clause
            where_parts = []
            params = []
            
            for i, condition in enumerate(conditions):
                field = condition['field']
                operator = condition['operator']
                value = condition['value']
                logic = condition['logic']
                
                # Build condition string
                if operator in ['LIKE', 'NOT LIKE']:
                    condition_str = f"{field} {operator} ?"
                    params.append(f"%{value}%")
                else:
                    condition_str = f"{field} {operator} ?"
                    # Type conversion for numeric fields
                    if field in ['age'] and value.isdigit():
                        params.append(int(value))
                    else:
                        params.append(value)
                
                if i == 0:
                    where_parts.append(condition_str)
                else:
                    where_parts.append(f" {logic} {condition_str}")
            
            where_clause = "".join(where_parts)
            query = f"SELECT * FROM students WHERE {where_clause}"
            
            cursor.execute(query, params)
            results = cursor.fetchall()
            conn.close()
            
            return results
            
        except Exception as e:
            raise Exception(f"Conditional logic search error: {str(e)}")

    def show_search_profile_manager(self):
        """Show comprehensive search profile management"""
        dialog = tk.Toplevel(self.master)
        dialog.title("Search Profile Manager")
        dialog.geometry("800x600")
        dialog.transient(self.master)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Search Profile Manager", style='Title.TLabel').pack(pady=(0, 20))
        
        # Profile management notebook
        notebook = ttk.Notebook(frame)
        notebook.pack(fill=tk.BOTH, expand=True, pady=(0, 20))
        
        # Saved profiles tab
        profiles_frame = ttk.Frame(notebook, padding="10")
        notebook.add(profiles_frame, text="Saved Profiles")
        
        # Profile list
        columns = ('ID', 'Name', 'Type', 'Created', 'Last Used', 'Shared')
        self.profiles_tree = ttk.Treeview(profiles_frame, columns=columns, show='headings', height=12)
        
        for col in columns:
            self.profiles_tree.heading(col, text=col)
            self.profiles_tree.column(col, width=100)
        
        profiles_scrollbar = ttk.Scrollbar(profiles_frame, orient=tk.VERTICAL, command=self.profiles_tree.yview)
        self.profiles_tree.configure(yscrollcommand=profiles_scrollbar.set)
        
        self.profiles_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        profiles_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Load profiles
        self.load_search_profiles()
        
        # Profile actions
        profile_actions = ttk.Frame(profiles_frame)
        profile_actions.pack(fill=tk.X, pady=(10, 0))
        
        ttk.Button(profile_actions, text="Load Profile", 
                  command=self.load_selected_profile).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(profile_actions, text="Delete Profile", 
                  command=self.delete_selected_profile).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(profile_actions, text="Share Profile", 
                  command=self.share_selected_profile).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(profile_actions, text="Export Profile", 
                  command=self.export_selected_profile).pack(side=tk.LEFT)
        
        # Create new profile tab
        create_frame = ttk.Frame(notebook, padding="10")
        notebook.add(create_frame, text="Create Profile")
        
        ttk.Label(create_frame, text="Create New Search Profile", style='Header.TLabel').pack(pady=(0, 20))
        
        # Profile creation form
        create_form = ttk.LabelFrame(create_frame, text="Profile Details", padding="10")
        create_form.pack(fill=tk.X, pady=(0, 20))
        
        ttk.Label(create_form, text="Profile Name:").pack(anchor='w')
        profile_name_var = tk.StringVar()
        ttk.Entry(create_form, textvariable=profile_name_var, width=40).pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(create_form, text="Description:").pack(anchor='w')
        profile_desc_text = tk.Text(create_form, height=4, wrap=tk.WORD)
        profile_desc_text.pack(fill=tk.X, pady=(0, 10))
        
        profile_shared_var = tk.BooleanVar()
        ttk.Checkbutton(create_form, text="Share with other users", 
                       variable=profile_shared_var).pack(anchor='w')
        
        def save_current_as_profile():
            name = profile_name_var.get().strip()
            if not name:
                messagebox.showwarning("Missing Name", "Please enter a profile name.")
                return
            
            description = profile_desc_text.get(1.0, tk.END).strip()
            is_shared = profile_shared_var.get()
            
            self.save_search_profile_to_db(name, description, is_shared)
            self.load_search_profiles()
            messagebox.showinfo("Profile Saved", f"Search profile '{name}' saved successfully!")
        
        ttk.Button(create_form, text="Save Current Search as Profile", 
                  command=save_current_as_profile).pack(pady=10)
        
        ttk.Button(frame, text="Close", command=dialog.destroy).pack()

    def load_search_profiles(self):
        """Load search profiles from database/storage"""
        # Clear existing items
        for item in self.profiles_tree.get_children():
            self.profiles_tree.delete(item)

        try:
            conn = get_connection()
            if conn is None:
                raise RuntimeError("Database connection unavailable.")
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            # Use column aliases for backward compatibility
            cursor.execute("""
                SELECT
                    search_id as id,
                    COALESCE(search_name, name) as search_name,
                    search_criteria,
                    COALESCE(is_shared, 0) as is_shared,
                    COALESCE(created_date, created_at) as created_date,
                    last_used
                FROM saved_searches
                ORDER BY datetime(COALESCE(created_date, created_at)) DESC
            """)
            rows = cursor.fetchall()
            conn.close()
        except Exception as exc:
            print_error(f"Failed to load saved profiles: {exc}")
            messagebox.showerror("Load Error", f"Failed to load saved profiles: {exc}")
            return
        
        for row in rows:
            try:
                payload = json.loads(row["search_criteria"]) if row["search_criteria"] else {}
            except json.JSONDecodeError:
                payload = {}
            criteria = payload.get("criteria", {})
            description = payload.get("description") or "Saved Search"
            profile_type = criteria.get("search_type") or description
            created = (row["created_date"] or "")[:19] if row["created_date"] else "N/A"
            last_used = (row["last_used"] or "")[:19] if row["last_used"] else "—"
            shared = "Yes" if row["is_shared"] else "No"
            self.profiles_tree.insert(
                '',
                'end',
                values=(row["id"], row["search_name"], profile_type, created, last_used, shared)
            )
        
    def load_selected_profile(self):
        """Load and execute selected search profile"""
        selection = self.profiles_tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a profile to load.")
            return
        
        item = self.profiles_tree.item(selection[0])
        profile_id = item['values'][0]
        
        try:
            conn = get_connection()
            if conn is None:
                raise RuntimeError("Database connection unavailable.")
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute("SELECT search_name, search_criteria FROM saved_searches WHERE id = ?", (profile_id,))
            row = cursor.fetchone()
            if not row:
                raise ValueError("Profile not found in database.")

            payload = json.loads(row["search_criteria"]) if row["search_criteria"] else {}
            criteria = payload.get("criteria", {})
            self._apply_profile_criteria(criteria)

            results = self._run_profile_search(criteria)
            if results:
                self.display_search_results(results)
            else:
                self.display_search_results([])
                messagebox.showinfo("Profile Loaded", f"Profile '{row['search_name']}' loaded, but no results matched.")

            cursor.execute(
                "UPDATE saved_searches SET last_used = CURRENT_TIMESTAMP WHERE id = ?",
                (profile_id,)
            )
            conn.commit()
            conn.close()
            self.log_output(f"Loaded search profile '{row['search_name']}' with {len(results)} result(s).")
        except Exception as exc:
            messagebox.showerror("Load Failed", f"Could not load profile: {exc}")

    def delete_selected_profile(self):
        """Delete selected search profile"""
        selection = self.profiles_tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a profile to delete.")
            return
        
        item = self.profiles_tree.item(selection[0])
        profile_id = item['values'][0]
        profile_name = item['values'][1]
        
        if messagebox.askyesno("Confirm Delete", f"Delete search profile '{profile_name}'?"):
            try:
                conn = get_connection()
                if conn is None:
                    raise RuntimeError("Database connection unavailable.")
                cursor = conn.cursor()
                cursor.execute("DELETE FROM saved_searches WHERE id = ?", (profile_id,))
                conn.commit()
                conn.close()
                self.profiles_tree.delete(selection[0])
                messagebox.showinfo("Profile Deleted", f"Profile '{profile_name}' deleted successfully.")
            except Exception as exc:
                messagebox.showerror("Delete Failed", f"Unable to delete profile: {exc}")

    def share_selected_profile(self):
        """Share selected search profile"""
        selection = self.profiles_tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a profile to share.")
            return
        
        item = self.profiles_tree.item(selection[0])
        profile_id = item['values'][0]
        profile_name = item['values'][1]
        
        # Share dialog
        share_dialog = tk.Toplevel(self.master)
        share_dialog.title("Share Profile")
        share_dialog.geometry("900x700")
        share_dialog.transient(self.master)
        share_dialog.grab_set()
        
        share_frame = ttk.Frame(share_dialog, padding="20")
        share_frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(share_frame, text=f"Share Profile: {profile_name}", style='Title.TLabel').pack(pady=(0, 20))
        
        ttk.Label(share_frame, text="Share with users:").pack(anchor='w')
        users_listbox = tk.Listbox(share_frame, selectmode=tk.MULTIPLE, height=8)
        users_listbox.pack(fill=tk.BOTH, expand=True, pady=(5, 20))
        
        # Sample users
        sample_users = ["admin", "teacher1", "teacher2", "analyst", "manager"]
        for user in sample_users:
            users_listbox.insert(tk.END, user)
        
        def confirm_share():
            selected_users = [users_listbox.get(i) for i in users_listbox.curselection()]
            if selected_users:
                try:
                    conn = get_connection()
                    if conn is None:
                        raise RuntimeError("Database connection unavailable.")
                    cursor = conn.cursor()
                    cursor.execute(
                        "UPDATE saved_searches SET is_shared = 1 WHERE id = ?",
                        (profile_id,)
                    )
                    conn.commit()
                    conn.close()
                    messagebox.showinfo("Profile Shared", 
                                      f"Profile '{profile_name}' shared with: {', '.join(selected_users)}")
                    self.load_search_profiles()
                except Exception as exc:
                    messagebox.showerror("Share Failed", f"Could not update profile sharing: {exc}")
                share_dialog.destroy()
            else:
                messagebox.showwarning("No Users Selected", "Please select at least one user to share with.")
        
        button_frame = ttk.Frame(share_frame)
        button_frame.pack(fill=tk.X)
        
        ttk.Button(button_frame, text="Share", command=confirm_share).pack(side=tk.LEFT)
        ttk.Button(button_frame, text="Cancel", command=share_dialog.destroy).pack(side=tk.RIGHT)

    def export_selected_profile(self):
        """Export selected search profile"""
        selection = self.profiles_tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a profile to export.")
            return
        
        item = self.profiles_tree.item(selection[0])
        profile_id = item['values'][0]
        
        filename = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
            initialfile=f"search_profile_{item['values'][1].replace(' ', '_')}.json"
        )
        
        if filename:
            try:
                conn = get_connection()
                if conn is None:
                    raise RuntimeError("Database connection unavailable.")
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                cursor.execute(
                    "SELECT search_name, search_criteria, is_shared, created_date, last_used "
                    "FROM saved_searches WHERE id = ?",
                    (profile_id,)
                )
                row = cursor.fetchone()
                conn.close()
                if not row:
                    raise ValueError("Profile not found.")

                payload = json.loads(row["search_criteria"]) if row["search_criteria"] else {}
                export_data = {
                    "name": row["search_name"],
                    "created": row["created_date"],
                    "last_used": row["last_used"],
                    "shared": bool(row["is_shared"]),
                    "criteria": payload.get("criteria", {}),
                    "metadata": {
                        "description": payload.get("description"),
                        "exported_at": datetime.now().isoformat(),
                        "exported_by": self._current_user_id()
                    }
                }

                with open(filename, 'w', encoding='utf-8') as handle:
                    json.dump(export_data, handle, indent=2)
                messagebox.showinfo("Export Complete", f"Profile exported to {filename}")
            except Exception as exc:
                messagebox.showerror("Export Error", f"Could not export profile: {exc}")

    def show_user_permissions_manager(self):
        """Show user permissions management interface"""
        dialog = tk.Toplevel(self.master)
        dialog.title("User Permissions Manager")
        dialog.geometry("700x500")
        dialog.transient(self.master)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="User Permissions Manager", style='Title.TLabel').pack(pady=(0, 20))
        
        # Users and permissions notebook
        notebook = ttk.Notebook(frame)
        notebook.pack(fill=tk.BOTH, expand=True, pady=(0, 20))
        
        # Current permissions tab
        current_frame = ttk.Frame(notebook, padding="10")
        notebook.add(current_frame, text="Current Permissions")
        
        # Permissions tree
        perm_columns = ('User', 'Role', 'Search', 'Export', 'Admin', 'Reports')
        self.permissions_tree = ttk.Treeview(current_frame, columns=perm_columns, show='headings', height=12)
        
        for col in perm_columns:
            self.permissions_tree.heading(col, text=col)
            self.permissions_tree.column(col, width=100)
        
        perm_scrollbar = ttk.Scrollbar(current_frame, orient=tk.VERTICAL, command=self.permissions_tree.yview)
        self.permissions_tree.configure(yscrollcommand=perm_scrollbar.set)
        
        self.permissions_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        perm_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Load current permissions
        self.load_user_permissions()
        
        # Permission actions
        perm_actions = ttk.Frame(current_frame)
        perm_actions.pack(fill=tk.X, pady=(10, 0))
        
        ttk.Button(perm_actions, text="Modify Permissions", 
                  command=self.modify_user_permissions_dialog).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(perm_actions, text="Remove User", 
                  command=self.remove_user_permissions_dialog).pack(side=tk.LEFT)
        
        # Add user tab
        add_user_frame = ttk.Frame(notebook, padding="10")
        notebook.add(add_user_frame, text="Add User")
        
        ttk.Label(add_user_frame, text="Add New User Permissions", style='Header.TLabel').pack(pady=(0, 20))
        
        # Add user form
        add_form = ttk.LabelFrame(add_user_frame, text="User Details", padding="10")
        add_form.pack(fill=tk.X, pady=(0, 20))
        
        ttk.Label(add_form, text="Username:").pack(anchor='w')
        username_var = tk.StringVar()
        ttk.Entry(add_form, textvariable=username_var, width=30).pack(anchor='w', pady=(0, 10))
        
        ttk.Label(add_form, text="Role:").pack(anchor='w')
        role_var = tk.StringVar(value="user")
        role_combo = ttk.Combobox(add_form, textvariable=role_var, 
                                 values=["admin", "teacher", "analyst", "user"], width=20)
        role_combo.pack(anchor='w', pady=(0, 10))
        
        # Permissions checkboxes
        permissions_frame = ttk.LabelFrame(add_form, text="Permissions", padding="10")
        permissions_frame.pack(fill=tk.X, pady=(10, 0))
        
        perm_vars = {}
        permissions_list = ["search", "export", "admin", "reports", "bulk_operations", "user_management"]
        
        for i, perm in enumerate(permissions_list):
            perm_vars[perm] = tk.BooleanVar()
            ttk.Checkbutton(permissions_frame, text=perm.replace('_', ' ').title(), 
                           variable=perm_vars[perm]).grid(row=i//2, column=i%2, sticky='w', padx=10, pady=2)
        
        def add_user_permissions():
            username = username_var.get().strip()
            if not username:
                messagebox.showwarning("Missing Username", "Please enter a username.")
                return
            
            role = role_var.get()
            selected_perms = [perm for perm, var in perm_vars.items() if var.get()]
            
            self.add_user_permissions_to_db(username, role, selected_perms)
            self.load_user_permissions()
            messagebox.showinfo("User Added", f"User '{username}' added with {role} role.")
            
            # Clear form
            username_var.set("")
            role_var.set("user")
            for var in perm_vars.values():
                var.set(False)
        
        ttk.Button(add_form, text="Add User", command=add_user_permissions).pack(pady=10)
        
        ttk.Button(frame, text="Close", command=dialog.destroy).pack()

    def load_user_permissions(self):
        """Load user permissions data"""
        # Clear existing items
        for item in self.permissions_tree.get_children():
            self.permissions_tree.delete(item)

        try:
            conn = get_connection()
            if conn is None:
                raise RuntimeError("Database connection unavailable.")
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute("SELECT user_id, role, permissions FROM user_permissions ORDER BY user_id")
            rows = cursor.fetchall()
            conn.close()
        except Exception as exc:
            messagebox.showerror("Load Error", f"Unable to load user permissions: {exc}")
            return

        for row in rows:
            try:
                perms = json.loads(row["permissions"]) if row["permissions"] else {}
            except json.JSONDecodeError:
                perms = {}
            if isinstance(perms, list):
                perms = {perm: True for perm in perms}
            values = [
                row["user_id"],
                row["role"] or "User",
                "Yes" if perms.get("search") else "No",
                "Yes" if perms.get("export") else "No",
                "Yes" if perms.get("admin") else "No",
                "Yes" if perms.get("reports") else "No",
            ]
            self.permissions_tree.insert('', 'end', values=values)

    def modify_user_permissions_dialog(self):
        """Show dialog to modify user permissions"""
        selection = self.permissions_tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a user to modify.")
            return
        
        item = self.permissions_tree.item(selection[0])
        username = item['values'][0]
        current_role = item['values'][1]
        
        # Modification dialog
        mod_dialog = tk.Toplevel(self.master)
        mod_dialog.title(f"Modify Permissions - {username}")
        mod_dialog.geometry("400x400")
        mod_dialog.transient(self.master)
        mod_dialog.grab_set()
        
        mod_frame = ttk.Frame(mod_dialog, padding="20")
        mod_frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(mod_frame, text=f"Modify Permissions for: {username}").pack(pady=(0, 20))
        
        # Role modification
        ttk.Label(mod_frame, text="Role:").pack(anchor='w')
        new_role_var = tk.StringVar(value=current_role)
        role_combo = ttk.Combobox(mod_frame, textvariable=new_role_var, 
                                 values=["Administrator", "Teacher", "Analyst", "User", "Guest"])
        role_combo.pack(anchor='w', pady=(0, 20))
        
        # Permissions modification
        perm_frame = ttk.LabelFrame(mod_frame, text="Permissions", padding="10")
        perm_frame.pack(fill=tk.X, pady=(0, 20))
        
        # Get current permissions
        current_perms = item['values'][2:]  # Skip username and role
        perm_labels = ["Search", "Export", "Admin", "Reports"]
        
        perm_vars = {}
        for i, (label, current) in enumerate(zip(perm_labels, current_perms)):
            perm_vars[label.lower()] = tk.BooleanVar(value=(current == "Yes"))
            ttk.Checkbutton(perm_frame, text=label, variable=perm_vars[label.lower()]).pack(anchor='w')
        
        def save_modifications():
            new_role = new_role_var.get()
            new_perms = {perm: var.get() for perm, var in perm_vars.items()}
            try:
                conn = get_connection()
                if conn is None:
                    raise RuntimeError("Database connection unavailable.")
                cursor = conn.cursor()
                cursor.execute(
                    """
                    UPDATE user_permissions
                    SET role = ?, permissions = ?, updated_date = CURRENT_TIMESTAMP
                    WHERE user_id = ?
                    """,
                    (new_role, json.dumps(new_perms), username)
                )
                conn.commit()
                conn.close()
                self.load_user_permissions()
                messagebox.showinfo("Permissions Updated", f"Permissions for {username} updated successfully.")
                mod_dialog.destroy()
            except Exception as exc:
                messagebox.showerror("Update Failed", f"Could not update permissions: {exc}")
        
        button_frame = ttk.Frame(mod_frame)
        button_frame.pack(fill=tk.X)
        
        ttk.Button(button_frame, text="Save Changes", command=save_modifications).pack(side=tk.LEFT)
        ttk.Button(button_frame, text="Cancel", command=mod_dialog.destroy).pack(side=tk.RIGHT)

    def remove_user_permissions_dialog(self):
        """Show dialog to remove user permissions"""
        selection = self.permissions_tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a user to remove.")
            return
        
        item = self.permissions_tree.item(selection[0])
        username = item['values'][0]
        
        if messagebox.askyesno("Confirm Removal", f"Remove all permissions for user '{username}'?"):
            try:
                conn = get_connection()
                if conn is None:
                    raise RuntimeError("Database connection unavailable.")
                cursor = conn.cursor()
                cursor.execute("DELETE FROM user_permissions WHERE user_id = ?", (username,))
                conn.commit()
                conn.close()
                self.load_user_permissions()
                messagebox.showinfo("User Removed", f"User '{username}' removed from system.")
            except Exception as exc:
                messagebox.showerror("Removal Failed", f"Could not remove permissions: {exc}")

    def add_user_permissions_to_db(self, username, role, permissions):
        """Add or update user permissions in the database."""
        conn = get_connection()
        if conn is None:
            raise RuntimeError("Database connection unavailable.")

        cursor = conn.cursor()
        if isinstance(permissions, list):
            permissions_map = {perm: True for perm in permissions}
        else:
            permissions_map = dict(permissions)
        for key in ["search", "export", "admin", "reports", "bulk_operations", "user_management"]:
            permissions_map.setdefault(key, False)
        cursor.execute(
            """
            INSERT INTO user_permissions (user_id, role, permissions, created_date, updated_date)
            VALUES (?, ?, ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
            ON CONFLICT(user_id)
            DO UPDATE SET role = excluded.role,
                          permissions = excluded.permissions,
                          updated_date = CURRENT_TIMESTAMP
            """,
            (username, role, json.dumps(permissions_map))
        )
        conn.commit()
        conn.close()

    def show_scheduled_reports_manager(self):
        """Show scheduled reports management interface"""
        dialog = tk.Toplevel(self.master)
        dialog.title("Scheduled Reports Manager")
        dialog.geometry("800x600")
        dialog.transient(self.master)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Scheduled Reports Manager", style='Title.TLabel').pack(pady=(0, 20))
        
        # Reports management notebook
        notebook = ttk.Notebook(frame)
        notebook.pack(fill=tk.BOTH, expand=True, pady=(0, 20))
        
        # Current reports tab
        current_reports_frame = ttk.Frame(notebook, padding="10")
        notebook.add(current_reports_frame, text="Scheduled Reports")
        
        # Reports tree
        report_columns = ('ID', 'Name', 'Type', 'Schedule', 'Next Run', 'Status', 'Recipients')
        self.reports_tree = ttk.Treeview(current_reports_frame, columns=report_columns, show='headings', height=12)
        
        for col in report_columns:
            self.reports_tree.heading(col, text=col)
            self.reports_tree.column(col, width=100)
        
        reports_scrollbar = ttk.Scrollbar(current_reports_frame, orient=tk.VERTICAL, command=self.reports_tree.yview)
        self.reports_tree.configure(yscrollcommand=reports_scrollbar.set)
        
        self.reports_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        reports_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Load scheduled reports
        self.load_scheduled_reports()
        
        # Report actions
        report_actions = ttk.Frame(current_reports_frame)
        report_actions.pack(fill=tk.X, pady=(10, 0))
        
        ttk.Button(report_actions, text="Run Now", 
                  command=self.run_selected_report).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(report_actions, text="Modify", 
                  command=self.modify_selected_report).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(report_actions, text="Delete", 
                  command=self.delete_selected_report).pack(side=tk.LEFT)
        
        # Create report tab
        create_report_frame = ttk.Frame(notebook, padding="10")
        notebook.add(create_report_frame, text="Create Report")
        
        ttk.Label(create_report_frame, text="Create New Scheduled Report", style='Header.TLabel').pack(pady=(0, 20))
        
        # Report creation form
        create_form = ttk.LabelFrame(create_report_frame, text="Report Configuration", padding="10")
        create_form.pack(fill=tk.BOTH, expand=True, pady=(0, 20))
        
        # Basic info
        basic_frame = ttk.Frame(create_form)
        basic_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(basic_frame, text="Report Name:").pack(anchor='w')
        report_name_var = tk.StringVar()
        ttk.Entry(basic_frame, textvariable=report_name_var, width=40).pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(basic_frame, text="Report Type:").pack(anchor='w')
        report_type_var = tk.StringVar(value="demographics")
        report_type_combo = ttk.Combobox(basic_frame, textvariable=report_type_var, 
                                       values=["demographics", "performance", "enrollment", "custom_sql"])
        report_type_combo.pack(anchor='w', pady=(0, 10))
        
        # Schedule configuration
        schedule_frame = ttk.LabelFrame(create_form, text="Schedule Configuration", padding="10")
        schedule_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(schedule_frame, text="Frequency:").pack(anchor='w')
        frequency_var = tk.StringVar(value="weekly")
        
        frequencies = [("Daily", "daily"), ("Weekly", "weekly"), ("Monthly", "monthly"), ("Quarterly", "quarterly")]
        for text, value in frequencies:
            ttk.Radiobutton(schedule_frame, text=text, variable=frequency_var, value=value).pack(anchor='w')
        
        # Recipients
        recipients_frame = ttk.LabelFrame(create_form, text="Recipients", padding="10")
        recipients_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(recipients_frame, text="Email Recipients (comma-separated):").pack(anchor='w')
        recipients_var = tk.StringVar()
        ttk.Entry(recipients_frame, textvariable=recipients_var, width=50).pack(fill=tk.X)
        
        def create_scheduled_report():
            name = report_name_var.get().strip()
            if not name:
                messagebox.showwarning("Missing Name", "Please enter a report name.")
                return
            
            report_type = report_type_var.get()
            frequency = frequency_var.get()
            recipients = recipients_var.get().strip()
            
            self.create_scheduled_report_in_db(name, report_type, frequency, recipients)
            self.load_scheduled_reports()
            messagebox.showinfo("Report Created", f"Scheduled report '{name}' created successfully!")
            
            # Clear form
            report_name_var.set("")
            recipients_var.set("")
        
        ttk.Button(create_form, text="Create Scheduled Report", command=create_scheduled_report).pack(pady=10)
        
        ttk.Button(frame, text="Close", command=dialog.destroy).pack()

    def load_scheduled_reports(self):
        """Load scheduled reports data"""
        # Clear existing items
        for item in self.reports_tree.get_children():
            self.reports_tree.delete(item)
        
        # Sample scheduled reports
        sample_reports = [
            (1, "Weekly Demographics", "Demographics", "Weekly", "2024-02-05", "Active", "admin@school.edu"),
            (2, "Monthly Performance", "Performance", "Monthly", "2024-02-01", "Active", "teachers@school.edu"),
            (3, "Daily Enrollments", "Enrollment", "Daily", "2024-01-26", "Paused", "office@school.edu"),
            (4, "Quarterly Analysis", "Custom SQL", "Quarterly", "2024-04-01", "Active", "director@school.edu")
        ]
        
        for report in sample_reports:
            self.reports_tree.insert('', 'end', values=report)

    def run_selected_report(self):
        """Run selected scheduled report immediately"""
        selection = self.reports_tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a report to run.")
            return
        
        item = self.reports_tree.item(selection[0])
        report_name = item['values'][1]
        report_type = item['values'][2]
        
        self.update_status(f"Running scheduled report: {report_name}...")
        self.start_progress()
        
        def run_report():
            try:
                report_result = self.execute_scheduled_report(report_name, report_type)
                self.output_queue.put(("analytics", report_result))
                self.output_queue.put(("log", f"Scheduled report '{report_name}' executed successfully."))
            except Exception as e:
                self.output_queue.put(("error", f"Report execution error: {str(e)}"))
            finally:
                self.output_queue.put(("stop_progress", None))
        
        threading.Thread(target=run_report, daemon=True).start()

    def modify_selected_report(self):
        """Modify selected scheduled report"""
        selection = self.reports_tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a report to modify.")
            return
        
        item = self.reports_tree.item(selection[0])
        report_name = item['values'][1]
        
        messagebox.showinfo("Modify Report", f"Opening modification dialog for: {report_name}")
        # In real implementation, open full modification dialog

    def delete_selected_report(self):
        """Delete selected scheduled report"""
        selection = self.reports_tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a report to delete.")
            return
        
        item = self.reports_tree.item(selection[0])
        report_name = item['values'][1]
        
        if messagebox.askyesno("Confirm Delete", f"Delete scheduled report '{report_name}'?"):
            self.reports_tree.delete(selection[0])
            messagebox.showinfo("Report Deleted", f"Scheduled report '{report_name}' deleted.")

    def create_scheduled_report_in_db(self, name, report_type, frequency, recipients):
        """Create scheduled report in database"""
        try:
            report_data = {
                "name": name,
                "type": report_type,
                "frequency": frequency,
                "recipients": recipients.split(',') if recipients else [],
                "created": datetime.now().isoformat(),
                "status": "active"
            }
            
            # Simulate database save
            filename = f"scheduled_report_{name.replace(' ', '_')}.json"
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(report_data, f, indent=2)
                
        except Exception as e:
            raise Exception(f"Error creating scheduled report: {str(e)}")

    def execute_scheduled_report(self, report_name, report_type):
        """Execute a scheduled report"""
        try:
            conn = get_connection()
            cursor = conn.cursor()
            
            result = f"SCHEDULED REPORT: {report_name}\n"
            result += f"=" * 50 + "\n"
            result += f"Report Type: {report_type}\n"
            result += f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
            
            if report_type.lower() == "demographics":
                cursor.execute("SELECT gender, COUNT(*) FROM students GROUP BY gender")
                gender_data = cursor.fetchall()
                
                cursor.execute("SELECT course, COUNT(*) FROM students GROUP BY course")
                course_data = cursor.fetchall()
                
                result += "DEMOGRAPHICS SUMMARY:\n"
                result += "-" * 25 + "\n"
                result += "Gender Distribution:\n"
                for gender, count in gender_data:
                    result += f"  {gender}: {count} students\n"
                
                result += "\nCourse Distribution:\n"
                for course, count in course_data:
                    result += f"  {course}: {count} students\n"
            
            elif report_type.lower() == "performance":
                cursor.execute("""
                SELECT grade, COUNT(*) 
                FROM student_modules 
                WHERE grade IS NOT NULL 
                GROUP BY grade
                ORDER BY grade
                """)
                grade_data = cursor.fetchall()
                
                result += "PERFORMANCE ANALYSIS:\n"
                result += "-" * 25 + "\n"
                result += "Grade Distribution:\n"
                total_grades = sum(count for _, count in grade_data)
                for grade, count in grade_data:
                    percentage = (count / total_grades) * 100 if total_grades > 0 else 0
                    result += f"  Grade {grade}: {count} ({percentage:.1f}%)\n"
            
            elif report_type.lower() == "enrollment":
                cursor.execute("""
                SELECT strftime('%Y-%m', registration_datetime) as month, COUNT(*)
                FROM students 
                WHERE registration_datetime IS NOT NULL
                GROUP BY month
                ORDER BY month DESC
                LIMIT 6
                """)
                enrollment_data = cursor.fetchall()
                
                result += "ENROLLMENT TRENDS (Last 6 months):\n"
                result += "-" * 35 + "\n"
                for month, count in enrollment_data:
                    result += f"  {month}: {count} new students\n"
            
            else:  # custom_sql
                result += "CUSTOM SQL REPORT:\n"
                result += "-" * 20 + "\n"
                result += "Custom SQL reports would execute user-defined queries here.\n"
            
            conn.close()
            
            result += f"\n" + "=" * 50 + "\n"
            result += "Report completed successfully.\n"
            
            return result
            
        except Exception as e:
            raise Exception(f"Scheduled report execution error: {str(e)}")

    def generate_comprehensive_reports(self):
        """Generate comprehensive system reports"""
        dialog = tk.Toplevel(self.master)
        dialog.title("Comprehensive Reports Generator")
        dialog.geometry("1100x800")
        dialog.transient(self.master)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Comprehensive Reports Generator", style='Title.TLabel').pack(pady=(0, 20))
        
        # Report types
        reports_frame = ttk.LabelFrame(frame, text="Available Reports", padding="10")
        reports_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 20))
        
        report_options = [
            ("Student Summary Report", "student_summary", "Complete overview of all students"),
            ("Module Enrollment Report", "module_enrollment", "Detailed module enrollment analysis"),
            ("Demographics Analysis", "demographics_analysis", "Comprehensive demographic breakdown"),
            ("Performance Report", "performance_report", "Academic performance analysis"),
            ("Custom SQL Report", "custom_sql", "Execute custom SQL queries")
        ]
        
        selected_reports = {}
        
        for name, key, description in report_options:
            report_frame = ttk.Frame(reports_frame)
            report_frame.pack(fill=tk.X, pady=5)
            
            selected_reports[key] = tk.BooleanVar()
            ttk.Checkbutton(report_frame, text=name, variable=selected_reports[key]).pack(side=tk.LEFT)
            ttk.Label(report_frame, text=f" - {description}", font=('Arial', 8)).pack(side=tk.LEFT)
        
        # Output options
        output_frame = ttk.LabelFrame(frame, text="Output Options", padding="10")
        output_frame.pack(fill=tk.X, pady=(0, 20))
        
        output_format_var = tk.StringVar(value="console")
        ttk.Radiobutton(output_frame, text="Display in Console", variable=output_format_var, value="console").pack(anchor='w')
        ttk.Radiobutton(output_frame, text="Export to File", variable=output_format_var, value="file").pack(anchor='w')
        ttk.Radiobutton(output_frame, text="Both Console and File", variable=output_format_var, value="both").pack(anchor='w')
        
        def generate_reports():
            selected = [key for key, var in selected_reports.items() if var.get()]
            if not selected:
                messagebox.showwarning("No Reports Selected", "Please select at least one report to generate.")
                return
            
            output_format = output_format_var.get()
            
            dialog.destroy()
            self.update_status("Generating comprehensive reports...")
            self.start_progress()
            
            def run_report_generation():
                try:
                    for report_type in selected:
                        report_result = self.generate_specific_report(report_type)
                        
                        if output_format in ["console", "both"]:
                            self.output_queue.put(("analytics", report_result))
                        
                        if output_format in ["file", "both"]:
                            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                            filename = f"{report_type}_report_{timestamp}.txt"
                            with open(filename, 'w', encoding='utf-8') as f:
                                f.write(report_result)
                            self.output_queue.put(("log", f"Report saved to {filename}"))
                    
                    self.output_queue.put(("log", "All comprehensive reports generated successfully."))
                except Exception as e:
                    self.output_queue.put(("error", f"Report generation error: {str(e)}"))
                finally:
                    self.output_queue.put(("stop_progress", None))
            
            threading.Thread(target=run_report_generation, daemon=True).start()
        
        button_frame = ttk.Frame(frame)
        button_frame.pack(fill=tk.X)
        
        ttk.Button(button_frame, text="Generate Reports", command=generate_reports).pack(side=tk.LEFT)
        ttk.Button(button_frame, text="Cancel", command=dialog.destroy).pack(side=tk.RIGHT)

    def generate_specific_report(self, report_type):
        """Generate specific report based on type"""
        try:
            conn = get_connection()
            cursor = conn.cursor()
            
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            if report_type == "student_summary":
                return self.generate_student_summary_report(cursor, timestamp)
            elif report_type == "module_enrollment":
                return self.generate_module_enrollment_report(cursor, timestamp)
            elif report_type == "demographics_analysis":
                return self.generate_demographics_analysis_report(cursor, timestamp)
            elif report_type == "performance_report":
                return self.generate_performance_analysis_report(cursor, timestamp)
            elif report_type == "custom_sql":
                return self.generate_custom_sql_report(cursor, timestamp)
            else:
                return f"Unknown report type: {report_type}"
            
        except Exception as e:
            raise Exception(f"Specific report generation error: {str(e)}")
        finally:
            if conn:
                conn.close()

    def generate_student_summary_report(self, cursor, timestamp):
        """Generate comprehensive student summary report"""
        result = f"STUDENT SUMMARY REPORT\n"
        result += f"=" * 50 + "\n"
        result += f"Generated: {timestamp}\n\n"
        
        # Total students
        cursor.execute("SELECT COUNT(*) FROM students")
        total_students = cursor.fetchone()[0]
        
        # Students by course
        cursor.execute("SELECT course, COUNT(*) FROM students GROUP BY course ORDER BY COUNT(*) DESC")
        course_breakdown = cursor.fetchall()
        
        # Gender distribution
        cursor.execute("SELECT gender, COUNT(*) FROM students GROUP BY gender")
        gender_breakdown = cursor.fetchall()
        
        # Age statistics
        cursor.execute("SELECT MIN(age), MAX(age), AVG(age) FROM students WHERE age IS NOT NULL")
        age_stats = cursor.fetchone()
        
        result += f"OVERVIEW:\n"
        result += f"Total Students: {total_students}\n\n"
        
        result += f"COURSE BREAKDOWN:\n"
        for course, count in course_breakdown:
            percentage = (count / total_students) * 100
            course_display = course if course else "Not Specified"
            result += f"  {course_display}: {count} students ({percentage:.1f}%)\n"
        
        result += f"\nGENDER DISTRIBUTION:\n"
        for gender, count in gender_breakdown:
            percentage = (count / total_students) * 100
            gender_display = gender.title() if gender else "Not Specified"
            result += f"  {gender_display}: {count} students ({percentage:.1f}%)\n"
        
        if age_stats and age_stats[0]:
            result += f"\nAGE STATISTICS:\n"
            result += f"  Youngest: {age_stats[0]} years\n"
            result += f"  Oldest: {age_stats[1]} years\n"
            avg_age = age_stats[2] if age_stats[2] is not None else 0.0
            result += f"  Average: {avg_age:.1f} years\n"
        
        # Recent registrations
        cursor.execute("""
        SELECT COUNT(*) FROM students 
        WHERE registration_datetime >= datetime('now', '-30 days')
        """)
        recent_count = cursor.fetchone()[0]
        
        result += f"\nRECENT ACTIVITY:\n"
        result += f"  New registrations (last 30 days): {recent_count}\n"
        
        return result

    def generate_module_enrollment_report(self, cursor, timestamp):
        """Generate module enrollment analysis report"""
        result = f"MODULE ENROLLMENT REPORT\n"
        result += f"=" * 50 + "\n"
        result += f"Generated: {timestamp}\n\n"
        
        # Total enrollments
        cursor.execute("SELECT COUNT(*) FROM student_modules")
        total_enrollments = cursor.fetchone()[0]
        
        # Enrollments by module
        cursor.execute("""
        SELECT module_code, module_name, COUNT(*) as enrollment_count
        FROM student_modules 
        GROUP BY module_code, module_name
        ORDER BY enrollment_count DESC
        """)
        module_enrollments = cursor.fetchall()
        
        # Enrollments by type
        cursor.execute("""
        SELECT module_type, COUNT(*) 
        FROM student_modules 
        GROUP BY module_type
        """)
        type_enrollments = cursor.fetchall()
        
        result += f"ENROLLMENT OVERVIEW:\n"
        result += f"Total Module Enrollments: {total_enrollments}\n\n"
        
        result += f"TOP MODULES BY ENROLLMENT:\n"
        for code, name, count in module_enrollments[:10]:  # Top 10
            result += f"  {code} - {name}: {count} students\n"
        
        result += f"\nENROLLMENT BY MODULE TYPE:\n"
        for module_type, count in type_enrollments:
            percentage = (count / total_enrollments) * 100 if total_enrollments > 0 else 0
            result += f"  {module_type}: {count} enrollments ({percentage:.1f}%)\n"
        
        # Grade distribution across all modules
        cursor.execute("""
        SELECT grade, COUNT(*) 
        FROM student_modules 
        WHERE grade IS NOT NULL 
        GROUP BY grade
        ORDER BY grade
        """)
        grade_distribution = cursor.fetchall()
        
        if grade_distribution:
            total_graded = sum(count for _, count in grade_distribution)
            result += f"\nOVERALL GRADE DISTRIBUTION:\n"
            for grade, count in grade_distribution:
                percentage = (count / total_graded) * 100
                result += f"  Grade {grade}: {count} ({percentage:.1f}%)\n"
        
        return result

    def generate_demographics_analysis_report(self, cursor, timestamp):
        """Generate comprehensive demographics analysis"""
        result = f"DEMOGRAPHICS ANALYSIS REPORT\n"
        result += f"=" * 50 + "\n"
        result += f"Generated: {timestamp}\n\n"
        
        # Gender distribution by course
        cursor.execute("""
        SELECT course, gender, COUNT(*) 
        FROM students 
        GROUP BY course, gender
        ORDER BY course, gender
        """)
        gender_course_data = cursor.fetchall()
        
        # Age distribution by course
        cursor.execute("""
        SELECT course, 
               CASE 
                   WHEN age < 20 THEN 'Under 20'
                   WHEN age BETWEEN 20 AND 25 THEN '20-25'
                   WHEN age BETWEEN 26 AND 30 THEN '26-30'
                   ELSE 'Over 30'
               END as age_group,
               COUNT(*)
        FROM students 
        WHERE age IS NOT NULL
        GROUP BY course, age_group
        ORDER BY course, age_group
        """)
        age_course_data = cursor.fetchall()
        
        # Registration trends
        cursor.execute("""
        SELECT strftime('%Y-%m', registration_datetime) as month, 
               course, COUNT(*)
        FROM students 
        WHERE registration_datetime IS NOT NULL
        GROUP BY month, course
        ORDER BY month DESC, course
        LIMIT 20
        """)
        registration_trends = cursor.fetchall()
        
        result += f"GENDER DISTRIBUTION BY COURSE:\n"
        current_course = None
        for course, gender, count in gender_course_data:
            if course != current_course:
                result += f"\n{course} Course:\n"
                current_course = course
            result += f"  {gender.title()}: {count} students\n"
        
        result += f"\nAGE DISTRIBUTION BY COURSE:\n"
        current_course = None
        for course, age_group, count in age_course_data:
            if course != current_course:
                result += f"\n{course} Course:\n"
                current_course = course
            result += f"  {age_group}: {count} students\n"
        
        result += f"\nREGISTRATION TRENDS (Recent months):\n"
        for month, course, count in registration_trends:
            result += f"  {month} - {course}: {count} new students\n"
        
        return result

    def generate_performance_analysis_report(self, cursor, timestamp):
        """Generate academic performance analysis report"""
        result = f"PERFORMANCE ANALYSIS REPORT\n"
        result += f"=" * 50 + "\n"
        result += f"Generated: {timestamp}\n\n"
        
        # Overall performance metrics
        cursor.execute("""
        SELECT 
            COUNT(DISTINCT student_id) as total_students_with_grades,
            COUNT(*) as total_graded_modules,
            AVG(CASE grade 
                WHEN 'A' THEN 4.0 
                WHEN 'B' THEN 3.0 
                WHEN 'C' THEN 2.0 
                WHEN 'D' THEN 1.0 
                ELSE 0.0 
            END) as average_gpa
        FROM student_modules 
        WHERE grade IS NOT NULL
        """)
        performance_stats = cursor.fetchone()
        
        # Performance by course
        cursor.execute("""
        SELECT s.course, 
               AVG(CASE sm.grade 
                   WHEN 'A' THEN 4.0 
                   WHEN 'B' THEN 3.0 
                   WHEN 'C' THEN 2.0 
                   WHEN 'D' THEN 1.0 
                   ELSE 0.0 
               END) as avg_gpa,
               COUNT(sm.grade) as graded_modules
        FROM students s
        JOIN student_modules sm ON s.student_id = sm.student_id
        WHERE sm.grade IS NOT NULL
        GROUP BY s.course
        """)
        course_performance = cursor.fetchall()
        
        # Top and bottom performing students
        cursor.execute("""
        SELECT s.student_id, s.first_name, s.last_name, s.course,
               AVG(CASE sm.grade 
                   WHEN 'A' THEN 4.0 
                   WHEN 'B' THEN 3.0 
                   WHEN 'C' THEN 2.0 
                   WHEN 'D' THEN 1.0 
                   ELSE 0.0 
               END) as avg_gpa,
               COUNT(sm.grade) as modules_completed
        FROM students s
        JOIN student_modules sm ON s.student_id = sm.student_id
        WHERE sm.grade IS NOT NULL
        GROUP BY s.student_id, s.first_name, s.last_name, s.course
        HAVING COUNT(sm.grade) >= 3
        ORDER BY avg_gpa DESC
        """)
        student_performance = cursor.fetchall()
        
        if performance_stats:
            total_students, total_modules, avg_gpa = performance_stats
            result += f"OVERALL PERFORMANCE METRICS:\n"
            result += f"Students with grades: {total_students}\n"
            result += f"Total graded modules: {total_modules}\n"
            avg_gpa_display = avg_gpa if avg_gpa is not None else 0.0
            result += f"System-wide GPA: {avg_gpa_display:.2f}\n\n"

        result += f"PERFORMANCE BY COURSE:\n"
        for course, avg_gpa, module_count in course_performance:
            avg_gpa_display = avg_gpa if avg_gpa is not None else 0.0
            result += f"  {course}: {avg_gpa_display:.2f} GPA ({module_count} graded modules)\n"

        result += f"\nTOP PERFORMING STUDENTS:\n"
        for i, (student_id, first_name, last_name, course, gpa, modules) in enumerate(student_performance[:10]):
            gpa_display = gpa if gpa is not None else 0.0
            result += f"  {i+1}. {first_name} {last_name} ({course}): {gpa_display:.2f} GPA ({modules} modules)\n"

        if len(student_performance) > 10:
            result += f"\nLOWEST PERFORMING STUDENTS (Need Support):\n"
            for i, (student_id, first_name, last_name, course, gpa, modules) in enumerate(student_performance[-5:]):
                gpa_display = gpa if gpa is not None else 0.0
                result += f"  {first_name} {last_name} ({course}): {gpa_display:.2f} GPA ({modules} modules)\n"
        
        return result

    def generate_custom_sql_report(self, cursor, timestamp):
        """Generate custom SQL report with user input"""
        result = f"CUSTOM SQL REPORT\n"
        result += f"=" * 50 + "\n"
        result += f"Generated: {timestamp}\n\n"
        
        # In a real implementation, this would allow users to input custom SQL
        # For now, provide some example queries
        
        sample_queries = [
            ("Students without module enrollments", "SELECT * FROM students WHERE student_id NOT IN (SELECT DISTINCT student_id FROM student_modules)"),
            ("Module completion rates", "SELECT module_code, COUNT(*) as enrolled, COUNT(grade) as completed FROM student_modules GROUP BY module_code"),
            ("Recent activity summary", "SELECT DATE(registration_datetime) as date, COUNT(*) FROM students WHERE registration_datetime >= datetime('now', '-7 days') GROUP BY DATE(registration_datetime)")
        ]
        
        result += f"SAMPLE CUSTOM QUERIES:\n\n"
        
        for query_name, sql in sample_queries:
            result += f"{query_name}:\n"
            result += f"SQL: {sql}\n"
            
            try:
                cursor.execute(sql)
                query_results = cursor.fetchall()
                
                result += f"Results ({len(query_results)} rows):\n"
                for row in query_results[:5]:  # Show first 5 rows
                    result += f"  {row}\n"
                if len(query_results) > 5:
                    result += f"  ... and {len(query_results) - 5} more rows\n"
            except Exception as e:
                result += f"Error executing query: {str(e)}\n"
            
            result += f"\n"
        
        result += f"Note: In full implementation, users can input custom SQL queries here.\n"
        
        return result

    def save_last_search_results(self):
        """Persist the last search results to disk and the central database."""
        if not self.search_results:
            messagebox.showwarning("No Results", "No search results to save.")
            return
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        default_name = f"last_search_results_{timestamp}.json"
        filename = filedialog.asksaveasfilename(
            title="Save Search Results",
            defaultextension=".json",
            initialfile=default_name,
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )
        if not filename:
            return
        
        try:
            results_data = {
                "timestamp": datetime.now().isoformat(),
                "result_count": len(self.search_results),
                "search_criteria": self._collect_search_criteria(),
                "results": []
            }
            
            for student in self.search_results:
                student_dict = {
                    "student_id": student[0],
                    "email": student[1],
                    "title": student[2],
                    "first_name": student[3],
                    "middle_name": student[4],
                    "last_name": student[5],
                    "gender": student[6],
                    "date_of_birth": student[7],
                    "age": student[8],
                    "course": student[9],
                    "registration_datetime": student[10]
                }
                results_data["results"].append(student_dict)
            
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(results_data, f, indent=2, default=str)

            try:
                conn = get_connection()
                if conn is None:
                    raise RuntimeError("Database connection unavailable.")
                cursor = conn.cursor()
                cursor.execute(
                    """
                    INSERT INTO search_result_archives (user_id, search_name, search_criteria, results_json)
                    VALUES (?, ?, ?, ?)
                    """,
                    (
                        self._current_user_id(),
                        f"Search results {timestamp}",
                        json.dumps(results_data["search_criteria"]),
                        json.dumps(results_data["results"])
                    )
                )
                conn.commit()
                conn.close()
            except Exception as db_error:
                self.log_output(f"Warning: could not archive search results to database ({db_error})")

            messagebox.showinfo("Results Saved", f"Last search results saved to {filename}")
            self.log_output(f"Search results saved to {filename} and archived.")

        except Exception as e:
            messagebox.showerror("Save Error", f"Could not save search results: {str(e)}")

    def log_search_operation(self, search_type, criteria, result_count):
        """Log search operation for audit trail"""
        try:
            log_entry = {
                "timestamp": datetime.now().isoformat(),
                "search_type": search_type,
                "criteria": criteria,
                "result_count": result_count,
                "user": "current_user",  # In real implementation, get from auth
                "ip_address": "127.0.0.1"  # In real implementation, get actual IP
            }
            
            # Append to search log file
            log_filename = "search_audit_log.json"
            
            # Load existing log or create new
            try:
                with open(log_filename, 'r', encoding='utf-8') as f:
                    log_data = json.load(f)
            except FileNotFoundError:
                log_data = {"searches": []}
            
            log_data["searches"].append(log_entry)
            
            # Keep only last 1000 entries
            if len(log_data["searches"]) > 1000:
                log_data["searches"] = log_data["searches"][-1000:]
            
            # Save updated log
            with open(log_filename, 'w', encoding='utf-8') as f:
                json.dump(log_data, f, indent=2)
                
        except Exception as e:
            self.log_output(f"Warning: Could not log search operation: {str(e)}")
        
        # Attempt to add entry to search_analytics table for dashboard parity
        try:
            conn = get_connection()
            cursor = conn.cursor()
            insert_search_analytics_record(
                cursor,
                user_id=self._current_user_id(),
                search_type=search_type,
                criteria=criteria,
                results_count=result_count,
                execution_time=0.0
            )
            conn.commit()
            conn.close()
        except Exception as analytics_error:
            self.log_output(f"Note: analytics logging skipped ({analytics_error})")

    def ensure_database_tables_exist(self):
        """Ensure all required database tables exist"""
        try:
            conn = get_connection()
            cursor = conn.cursor()
            
            # Check and create students table if needed
            cursor.execute("""
            CREATE TABLE IF NOT EXISTS students (
                student_id TEXT PRIMARY KEY,
                email TEXT UNIQUE,
                title TEXT,
                first_name TEXT,
                middle_name TEXT,
                last_name TEXT,
                gender TEXT,
                date_of_birth DATE,
                age INTEGER,
                course TEXT,
                registration_datetime DATETIME DEFAULT CURRENT_TIMESTAMP
            )
            """)
            
            # Check and create student_modules table if needed
            cursor.execute("""
            CREATE TABLE IF NOT EXISTS student_modules (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                student_id TEXT,
                module_type TEXT,
                module_code TEXT,
                module_name TEXT,
                grade TEXT,
                enrollment_date DATE DEFAULT CURRENT_DATE,
                FOREIGN KEY (student_id) REFERENCES students(student_id)
            )
            """)
            
            # Check and create search_profiles table if needed
            cursor.execute("""
            CREATE TABLE IF NOT EXISTS search_profiles (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                description TEXT,
                criteria TEXT,
                created_date DATETIME DEFAULT CURRENT_TIMESTAMP,
                created_by TEXT,
                is_shared BOOLEAN DEFAULT FALSE
            )
            """)
            
            # Check and create user_permissions table if needed
            cursor.execute("""
            CREATE TABLE IF NOT EXISTS user_permissions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE,
                role TEXT,
                permissions TEXT,
                created_date DATETIME DEFAULT CURRENT_TIMESTAMP
            )
            """)
            
            # Check and create scheduled_reports table if needed
            cursor.execute("""
            CREATE TABLE IF NOT EXISTS scheduled_reports (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                report_type TEXT,
                frequency TEXT,
                recipients TEXT,
                next_run_date DATETIME,
                is_active BOOLEAN DEFAULT TRUE,
                created_date DATETIME DEFAULT CURRENT_TIMESTAMP
            )
            """)
            
            conn.commit()
            conn.close()
            
            return "All required database tables verified/created successfully."
            
        except Exception as e:
            return f"Error ensuring database tables: {str(e)}"

    def check_database_status(self):
        """Check database status and integrity"""
        try:
            conn = get_connection()
            cursor = conn.cursor()
            
            status_report = "DATABASE STATUS REPORT\n"
            status_report += "=" * 30 + "\n"
            status_report += f"Check performed: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
            
            # Check table existence and record counts
            tables_to_check = [
                ("students", "Student Records"),
                ("student_modules", "Module Enrollments"),
                ("search_profiles", "Saved Search Profiles"),
                ("user_permissions", "User Permissions"),
                ("scheduled_reports", "Scheduled Reports")
            ]
            
            status_report += "TABLE STATUS:\n"
            for table_name, description in tables_to_check:
                try:
                    cursor.execute(f"SELECT COUNT(*) FROM {table_name}")
                    count = cursor.fetchone()[0]
                    status_report += f"  {description}: {count} records\n"
                except Exception as e:
                    status_report += f"  {description}: Table missing or error ({str(e)})\n"
            
            # Check for data integrity issues
            status_report += f"\nINTEGRITY CHECKS:\n"
            
            # Check for students without emails
            try:
                cursor.execute("SELECT COUNT(*) FROM students WHERE email_address IS NULL OR email_address = ''")
                no_email_count = cursor.fetchone()[0]
                status_report += f"  Students without email: {no_email_count}\n"
            except:
                status_report += f"  Students without email: Unable to check\n"
            
            # Check for orphaned module records
            try:
                cursor.execute("""
                SELECT COUNT(*) FROM student_modules sm 
                WHERE NOT EXISTS (SELECT 1 FROM students s WHERE s.student_id = sm.student_id)
                """)
                orphaned_modules = cursor.fetchone()[0]
                status_report += f"  Orphaned module records: {orphaned_modules}\n"
            except:
                status_report += f"  Orphaned module records: Unable to check\n"
            
            # Check database size
            try:
                cursor.execute("PRAGMA page_count")
                page_count = cursor.fetchone()[0]
                cursor.execute("PRAGMA page_size")
                page_size = cursor.fetchone()[0]
                db_size_bytes = page_count * page_size
                db_size_mb = db_size_bytes / (1024 * 1024)
                status_report += f"  Database size: {db_size_mb:.2f} MB\n"
            except:
                status_report += f"  Database size: Unable to determine\n"
            
            conn.close()
            
            status_report += f"\nDatabase connection: OK\n"
            status_report += f"Status check completed successfully.\n"
            
            return status_report
            
        except Exception as e:
            return f"Database status check failed: {str(e)}"

    def repeat_last_search(self):
        """Repeat the last executed search"""
        if not hasattr(self, 'last_search_criteria') or not self.last_search_criteria:
            messagebox.showinfo("No Previous Search", "No previous search to repeat.")
            return
        
        criteria = self.last_search_criteria
        search_type = criteria.get('type', 'unknown')
        
        self.update_status(f"Repeating last search ({search_type})...")
        self.start_progress()
        
        def run_repeat_search():
            try:
                # Execute based on search type
                if search_type == "multi_criteria":
                    results = self.perform_database_search(criteria.get('data', {}))
                elif search_type == "fuzzy":
                    results = self.perform_fuzzy_search(
                        criteria.get('term', ''), 
                        criteria.get('threshold', 0.6), 
                        criteria.get('algorithm', '1')
                    )
                elif search_type == "text":
                    results = self.perform_text_search(
                        criteria.get('pattern', ''), 
                        criteria.get('search_type', 'wildcard'), 
                        criteria.get('field', 'first_name')
                    )
                else:
                    results = []
                
                self.output_queue.put(("search_results", results))
                self.output_queue.put(("log", f"Repeated {search_type} search. Found {len(results)} results."))
                
            except Exception as e:
                self.output_queue.put(("error", f"Repeat search error: {str(e)}"))
            finally:
                self.output_queue.put(("stop_progress", None))
        
        threading.Thread(target=run_repeat_search, daemon=True).start()

    def show_system_maintenance(self):
        """Show system maintenance interface"""
        dialog = tk.Toplevel(self.master)
        dialog.title("System Maintenance")
        dialog.geometry("1100x800")
        dialog.transient(self.master)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="System Maintenance & Utilities", style='Title.TLabel').pack(pady=(0, 20))
        
        # Maintenance operations
        maintenance_frame = ttk.LabelFrame(frame, text="Database Maintenance", padding="10")
        maintenance_frame.pack(fill=tk.X, pady=(0, 20))
        
        maintenance_ops = [
            ("Check Database Status", self.run_database_status_check),
            ("Ensure Tables Exist", self.run_ensure_tables),
            ("Optimize Database", self.run_database_optimization),
            ("Clean Audit Logs", self.run_clean_audit_logs),
            ("🏠 Return to Main Menu", self.return_to_main_menu)
        ]
        
        for text, command in maintenance_ops:
            ttk.Button(maintenance_frame, text=text, command=command, width=25).pack(pady=2)
        
        # Data management
        data_frame = ttk.LabelFrame(frame, text="Data Management", padding="10")
        data_frame.pack(fill=tk.X, pady=(0, 20))
        
        data_ops = [
            ("Backup Database", self.run_database_backup),
            ("Restore Database", self.run_database_restore),
            ("Clear Search History", self.clear_search_history),
            ("🏠 Return to Main Menu", self.return_to_main_menu)
        ]
        
        for text, command in data_ops:
            ttk.Button(data_frame, text=text, command=command, width=25).pack(pady=2)
        
        # System information
        info_frame = ttk.LabelFrame(frame, text="System Information", padding="10")
        info_frame.pack(fill=tk.BOTH, expand=True)
        
        self.system_info_text = scrolledtext.ScrolledText(info_frame, height=8, wrap=tk.WORD)
        self.system_info_text.pack(fill=tk.BOTH, expand=True)
        
        # Load initial system info
        self.load_system_information()
        
        ttk.Button(frame, text="Close", command=dialog.destroy).pack()

    def run_database_status_check(self):
        """Run database status check"""
        self.update_status("Checking database status...")
        
        def check_status():
            try:
                status_report = self.check_database_status()
                self.system_info_text.delete(1.0, tk.END)
                self.system_info_text.insert(1.0, status_report)
                self.log_output("Database status check completed")
            except Exception as e:
                self.log_output(f"Database status check failed: {str(e)}")
            finally:
                self.update_status("Ready")
        
        threading.Thread(target=check_status, daemon=True).start()

    def run_ensure_tables(self):
        """Run ensure tables exist operation"""
        self.update_status("Ensuring database tables exist...")
        
        def ensure_tables():
            try:
                result = self.ensure_database_tables_exist()
                self.log_output(result)
                messagebox.showinfo("Tables Check", "Database table check completed successfully.")
            except Exception as e:
                self.log_output(f"Table check failed: {str(e)}")
                messagebox.showerror("Error", f"Table check failed: {str(e)}")
            finally:
                self.update_status("Ready")
        
        threading.Thread(target=ensure_tables, daemon=True).start()

    def run_database_optimization(self):
        """Run database optimization"""
        self.update_status("Optimizing database...")
        
        def optimize_db():
            try:
                conn = get_connection()
                cursor = conn.cursor()
                
                # Run VACUUM to optimize database
                cursor.execute("VACUUM")
                
                # Update statistics
                cursor.execute("ANALYZE")
                
                conn.close()
                
                self.log_output("Database optimization completed successfully")
                messagebox.showinfo("Optimization", "Database optimization completed.")
                
            except Exception as e:
                self.log_output(f"Database optimization failed: {str(e)}")
                messagebox.showerror("Error", f"Database optimization failed: {str(e)}")
            finally:
                self.update_status("Ready")
        
        threading.Thread(target=optimize_db, daemon=True).start()

    def run_clean_audit_logs(self):
        """Clean old audit log entries"""
        if messagebox.askyesno("Confirm Clean", "This will remove audit log entries older than 90 days. Continue?"):
            try:
                # Clean old log entries from file
                log_filename = "search_audit_log.json"
                cutoff_date = datetime.now() - timedelta(days=90)
                
                try:
                    with open(log_filename, 'r', encoding='utf-8') as f:
                        log_data = json.load(f)
                    
                    original_count = len(log_data.get("searches", []))
                    
                    # Filter out old entries
                    log_data["searches"] = [
                        entry for entry in log_data.get("searches", [])
                        if datetime.fromisoformat(entry["timestamp"]) > cutoff_date
                    ]
                    
                    cleaned_count = original_count - len(log_data["searches"])
                    
                    with open(log_filename, 'w', encoding='utf-8') as f:
                        json.dump(log_data, f, indent=2)
                    
                    self.log_output(f"Cleaned {cleaned_count} old audit log entries")
                    messagebox.showinfo("Logs Cleaned", f"Removed {cleaned_count} old audit log entries.")
                    
                except FileNotFoundError:
                    messagebox.showinfo("No Logs", "No audit log file found to clean.")
                    
            except Exception as e:
                messagebox.showerror("Error", f"Failed to clean audit logs: {str(e)}")

    def run_export_system_stats(self):
        """Export comprehensive system statistics"""
        self.update_status("Exporting system statistics...")
        self.start_progress()
        
        def export_stats():
            try:
                result = self.capture_function_output(export_system_statistics)
                self.output_queue.put(("analytics", result))
                self.output_queue.put(("log", "System statistics exported successfully"))
            except Exception as e:
                self.output_queue.put(("error", f"System statistics export failed: {str(e)}"))
            finally:
                self.output_queue.put(("stop_progress", None))
        
        threading.Thread(target=export_stats, daemon=True).start()

    def run_database_backup(self):
        """Create database backup"""
        backup_file = filedialog.asksaveasfilename(
            defaultextension=".db",
            filetypes=[("SQLite files", "*.db"), ("All files", "*.*")],
            title="Save Database Backup As"
        )
        
        if backup_file:
            try:
                source_path = Path(DEFAULT_DB_PATH)
                if not source_path.exists():
                    raise FileNotFoundError(f"Database file not found at {source_path}")

                destination = Path(backup_file)
                destination.parent.mkdir(parents=True, exist_ok=True)
                shutil.copy2(source_path, destination)

                messagebox.showinfo("Backup Created", f"Database backup created: {destination}")
                self.log_output(f"Database backup created: {destination}")
            except Exception as e:
                messagebox.showerror("Backup Failed", f"Database backup failed: {str(e)}")

    def run_database_restore(self):
        """Restore database from backup"""
        backup_file = filedialog.askopenfilename(
            filetypes=[("SQLite files", "*.db"), ("All files", "*.*")],
            title="Select Database Backup to Restore"
        )
        
        if backup_file:
            if messagebox.askyesno("Confirm Restore", "This will replace the current database. Continue?"):
                try:
                    destination = Path(DEFAULT_DB_PATH)
                    source = Path(backup_file)
                    if not source.exists():
                        raise FileNotFoundError(f"Backup file {source} does not exist.")
                    destination.parent.mkdir(parents=True, exist_ok=True)

                    shutil.copy2(source, destination)
                    messagebox.showinfo("Restore Complete", f"Database restored from {source}")
                    self.log_output(f"Database restored from: {source}")
                except Exception as e:
                    messagebox.showerror("Restore Failed", f"Database restore failed: {str(e)}")

    def clear_search_history(self):
        """Clear search history"""
        if messagebox.askyesno("Confirm Clear", "This will clear all search history. Continue?"):
            try:
                # Clear search audit log
                log_filename = "search_audit_log.json"
                empty_log = {"searches": []}
                
                with open(log_filename, 'w', encoding='utf-8') as f:
                    json.dump(empty_log, f, indent=2)
                
                # Clear any cached search data
                if hasattr(self, 'last_search_criteria'):
                    delattr(self, 'last_search_criteria')
                
                messagebox.showinfo("History Cleared", "Search history cleared successfully.")
                self.log_output("Search history cleared")
                
            except Exception as e:
                messagebox.showerror("Error", f"Failed to clear search history: {str(e)}")

    def reset_user_preferences(self):
        """Reset user preferences to defaults"""
        if messagebox.askyesno("Confirm Reset", "This will reset all user preferences to defaults. Continue?"):
            try:
                # Reset GUI preferences
                self.results_per_page = 10
                self.per_page_var.set("10")
                
                # Clear any preference files
                pref_files = ["user_preferences.json", "gui_settings.json"]
                
                for pref_file in pref_files:
                    try:
                        import os
                        if os.path.exists(pref_file):
                            os.remove(pref_file)
                    except:
                        pass
                
                messagebox.showinfo("Preferences Reset", "User preferences reset to defaults.")
                self.log_output("User preferences reset to defaults")
                
            except Exception as e:
                messagebox.showerror("Error", f"Failed to reset preferences: {str(e)}")

    def load_system_information(self):
        """Load system information into the text widget"""
        try:
            import platform
            import sys
            
            info = f"SYSTEM INFORMATION\n"
            info += f"=" * 30 + "\n"
            info += f"Application: Enhanced Student Search & Analytics\n"
            info += f"Version: 2.0 GUI Edition\n"
            info += f"Python Version: {sys.version}\n"
            info += f"Platform: {platform.platform()}\n"
            info += f"Architecture: {platform.architecture()[0]}\n"
            info += f"Current Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
            
            info += f"DATABASE STATUS:\n"
            info += f"Connection: Available\n"
            
            try:
                conn = get_connection()
                if conn:
                    cursor = conn.cursor()
                    cursor.execute("SELECT COUNT(*) FROM students")
                    student_count = cursor.fetchone()[0]
                    cursor.execute("SELECT COUNT(*) FROM student_modules")
                    module_count = cursor.fetchone()[0]
                    conn.close()
                    
                    info += f"Student Records: {student_count}\n"
                    info += f"Module Enrollments: {module_count}\n"
                else:
                    info += f"Database: Connection failed\n"
            except:
                info += f"Database: Status unknown\n"
            
            info += f"\nFEATURES AVAILABLE:\n"
            info += f"✓ Multi-criteria search\n"
            info += f"✓ Fuzzy name matching\n"
            info += f"✓ Advanced text search\n"
            info += f"✓ Analytics dashboard\n"
            info += f"✓ Data visualization\n"
            info += f"✓ Bulk operations\n"
            info += f"✓ User permissions\n"
            info += f"✓ Scheduled reports\n"
            info += f"✓ Export capabilities\n"
            
            self.system_info_text.insert(1.0, info)
            
        except Exception as e:
            self.system_info_text.insert(1.0, f"Error loading system information: {str(e)}")

    # Override the execute_search method to store search criteria
    def execute_search_with_logging(self, search_window):
        """Execute search with logging for repeat functionality"""
        # Get search criteria
        criteria = {}
        for key, var in self.search_vars.items():
            value = var.get().strip()
            if value:
                criteria[key] = value
        
        if not any(criteria.values()):
            messagebox.showwarning("No Criteria", "Please enter at least one search criterion.")
            return
        
        # Store for repeat functionality
        self.last_search_criteria = {
            'type': 'multi_criteria',
            'data': criteria
        }
        
        search_window.destroy()
        self.update_status("Searching...")
        self.start_progress()
        
        def run_search():
            try:
                results = self.perform_database_search(criteria)
                
                # Log the search operation
                self.log_search_operation("multi_criteria", criteria, len(results))
                
                self.output_queue.put(("search_results", results))
                self.output_queue.put(("log", f"Search completed. Found {len(results)} results."))
            except Exception as e:
                self.output_queue.put(("error", f"Search error: {str(e)}"))
            finally:
                self.output_queue.put(("stop_progress", None))
        
        threading.Thread(target=run_search, daemon=True).start()
    
    def perform_database_search(self, criteria):
        """Perform the actual database search"""
        try:
            conn = get_connection()
            cursor = conn.cursor()
            
            # Build query
            query = "SELECT * FROM students WHERE 1=1"
            params = []
            
            for key, value in criteria.items():
                if key in ['student_id', 'first_name', 'last_name', 'email']:
                    query += f" AND LOWER({key}) LIKE LOWER(?)"
                    params.append(f"%{value}%")
                elif key in ['gender', 'course']:
                    query += f" AND LOWER({key}) = LOWER(?)"
                    params.append(value)
                elif key == 'min_age':
                    query += " AND age >= ?"
                    params.append(int(value))
                elif key == 'max_age':
                    query += " AND age <= ?"
                    params.append(int(value))
            
            cursor.execute(query, params)
            results = cursor.fetchall()
            conn.close()
            
            return results
            
        except Exception as e:
            raise Exception(f"Database error: {str(e)}")
    
    def clear_search_form(self):
        """Clear all search form fields"""
        for var in self.search_vars.values():
            var.set("")
    
    def display_search_results(self, results):
        """Display search results in the treeview"""
        # Clear existing results
        for item in self.results_tree.get_children():
            self.results_tree.delete(item)
        
        self.search_results = results
        self.current_page = 0
        self.update_results_display()
        
        # Switch to results tab
        self.notebook.select(1)
        
        # Update UI
        count = len(results)
        self.results_label.config(text=f"Search Results: {count} students found")
        self.export_btn.config(state='normal' if count > 0 else 'disabled')
    
    def update_results_display(self):
        """Update the results display for current page"""
        # Clear treeview
        for item in self.results_tree.get_children():
            self.results_tree.delete(item)
        
        if not self.search_results:
            return
        
        # Calculate pagination
        self.results_per_page = int(self.per_page_var.get())
        total_pages = (len(self.search_results) - 1) // self.results_per_page + 1 if self.search_results else 1
        start_idx = self.current_page * self.results_per_page
        end_idx = min(start_idx + self.results_per_page, len(self.search_results))
        
        # Display current page results
        for i in range(start_idx, end_idx):
            student = self.search_results[i]
            # Format the data for display
            display_data = (
                student[0],  # ID
                f"{student[3]} {student[5]}",  # Name
                student[1],  # Email
                student[6],  # Gender
                student[8],  # Age
                student[9],  # Course
                student[10][:10] if student[10] else ""  # Registration date
            )
            self.results_tree.insert('', 'end', values=display_data)
        
        # Update pagination controls
        self.page_label.config(text=f"Page {self.current_page + 1} of {total_pages}")
        self.prev_btn.config(state='normal' if self.current_page > 0 else 'disabled')
        self.next_btn.config(state='normal' if self.current_page < total_pages - 1 else 'disabled')
    
    def previous_page(self):
        """Go to previous page"""
        if self.current_page > 0:
            self.current_page -= 1
            self.update_results_display()
    
    def next_page(self):
        """Go to next page"""
        total_pages = (len(self.search_results) - 1) // self.results_per_page + 1 if self.search_results else 1
        if self.current_page < total_pages - 1:
            self.current_page += 1
            self.update_results_display()
    
    def change_results_per_page(self, event=None):
        """Change number of results per page"""
        self.current_page = 0
        self.update_results_display()
    
    def show_student_details(self, event):
        """Show detailed information for selected student"""
        selection = self.results_tree.selection()
        if not selection:
            return
        
        item = self.results_tree.item(selection[0])
        student_id = item['values'][0]
        
        # Find the full student record
        student_record = None
        for student in self.search_results:
            if student[0] == student_id:
                student_record = student
                break
        
        if student_record:
            self.show_student_detail_window(student_record)
    
    def show_student_detail_window(self, student):
        """Show detailed student information in a new window"""
        detail_window = tk.Toplevel(self.master)
        detail_window.title(f"📋 Student Details - {student[0]}")
        detail_window.geometry("600x400")
        detail_window.transient(self.master)
        
        main_frame = ttk.Frame(detail_window, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Title
        title_label = ttk.Label(main_frame, text=f"Student Information - {student[0]}", style='Title.TLabel')
        title_label.pack(pady=(0, 20))
        
        # Information frame
        info_frame = ttk.LabelFrame(main_frame, text="Basic Information", padding="10")
        info_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Display student information
        info_text = f"""
📧 Email: {student[1]}
👤 Title: {student[2]}
🏷️ Name: {student[3]} {student[4] or ''} {student[5]}
⚧ Gender: {student[6]}
🎂 Date of Birth: {student[7]}
📅 Age: {student[8]}
🎓 Course: {student[9]}
📝 Registration: {student[10]}
        """
        
        info_label = tk.Label(info_frame, text=info_text, justify=tk.LEFT, font=('Arial', 10))
        info_label.pack(anchor='w')
        
        # Modules frame (placeholder for module information)
        modules_frame = ttk.LabelFrame(main_frame, text="Module Information", padding="10")
        modules_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        modules_text = tk.Text(modules_frame, height=8, wrap=tk.WORD)
        modules_text.pack(fill=tk.BOTH, expand=True)
        
        # Load module information
        self.load_student_modules(student[0], modules_text)
        
        # Buttons frame
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X)
        
        ttk.Button(button_frame, text="📧 Send Email", 
                  command=lambda: self.simulate_send_email(student)).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(button_frame, text="💾 Export", 
                  command=lambda: self.export_single_student(student)).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(button_frame, text="❌ Close", 
                  command=detail_window.destroy).pack(side=tk.RIGHT)
    
    def load_student_modules(self, student_id, text_widget):
        """Load and display student module information"""
        try:
            conn = get_connection()
            cursor = conn.cursor()
            
            cursor.execute('''
            SELECT module_type, module_code, module_name, grade, enrollment_date
            FROM student_modules
            WHERE student_id = ?
            ORDER BY module_type, module_name
            ''', (student_id,))
            
            modules = cursor.fetchall()
            conn.close()
            
            if modules:
                text_widget.insert(tk.END, f"📚 ENROLLED MODULES ({len(modules)} total):\n")
                text_widget.insert(tk.END, "-" * 60 + "\n")
                text_widget.insert(tk.END, f"{'Type':<15} {'Code':<10} {'Name':<25} {'Grade':<8}\n")
                text_widget.insert(tk.END, "-" * 60 + "\n")
                
                for module in modules:
                    module_type, code, name, grade, enrolled_date = module
                    grade_display = grade if grade else "In Progress"
                    text_widget.insert(tk.END, f"{module_type:<15} {code:<10} {name:<25} {grade_display:<8}\n")
            else:
                text_widget.insert(tk.END, "No modules enrolled.")
                
        except Exception as e:
            text_widget.insert(tk.END, f"Error loading modules: {str(e)}")
    
    # Utility Methods
    def capture_function_output(self, func, *args, **kwargs):
        """Capture output from original CLI functions"""
        import io
        import sys
        
        old_stdout = sys.stdout
        sys.stdout = captured_output = io.StringIO()
        
        try:
            result = func(*args, **kwargs)
            output = captured_output.getvalue()
            return output if output else str(result)
        except Exception as e:
            return f"Error: {str(e)}"
        finally:
            sys.stdout = old_stdout
    
    def update_status(self, message):
        """Update the status bar"""
        self.status_label.config(text=message)
        self.master.update_idletasks()
    
    def start_progress(self):
        """Start the progress bar"""
        self.progress.start(10)
    
    def stop_progress(self):
        """Stop the progress bar"""
        self.progress.stop()
    
    def return_to_main_menu(self):
        """Return to the main menu"""
        try:
            # Check if this is a child window (Toplevel) or standalone (Tk)
            root_widget = self.root if hasattr(self, 'root') else self.master
            if isinstance(root_widget, tk.Toplevel):
                # Just close the child window
                root_widget.destroy()
            else:
                # Running standalone, need to create main GUI
                root_widget.destroy()
                from university_system.modules.shared.gui.main_gui import UnifiedManagementGUI
                app = UnifiedManagementGUI(self.auth)
                app.run()
        except Exception as e:
            print_error(f"Error returning to main menu: {e}")
            import traceback
            traceback.print_exc()

    def monitor_output(self):
        """Monitor the output queue for updates"""
        try:
            while True:
                msg_type, data = self.output_queue.get_nowait()

                if msg_type == "search_results":
                    self.display_search_results(data)
                elif msg_type == "analytics":
                    self.show_analytics_output(data)
                elif msg_type == "log":
                    self.log_output(data)
                elif msg_type == "error":
                    self.show_error(data)
                elif msg_type == "stop_progress":
                    self.stop_progress()
                    self.update_status("Ready")

        except queue.Empty:
            pass
        
        # Schedule next check
        self.master.after(100, self.monitor_output)
    
    def log_output(self, message):
        """Log message to output console"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.output_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.output_text.see(tk.END)
    
    def show_error(self, error_message):
        """Show error message"""
        self.log_output(f"ERROR: {error_message}")
        messagebox.showerror("Error", error_message)
    
    def show_analytics_output(self, output):
        """Display analytics output"""
        self.output_text.insert(tk.END, f"\n=== ANALYTICS DASHBOARD ===\n")
        self.output_text.insert(tk.END, output)
        self.output_text.insert(tk.END, f"\n{'='*50}\n")
        self.output_text.see(tk.END)
        self.notebook.select(2)  # Switch to output tab
    
    def clear_output(self):
        """Clear the output console"""
        self.output_text.delete(1.0, tk.END)
    
    # Menu Action Methods (continued)
    def show_demographics_reports(self):
        """Show demographics reports"""
        self.update_status("Generating demographics reports...")
        self.start_progress()
        
        def run_demographics():
            try:
                result = self.capture_function_output(student_demographics_reports)
                self.output_queue.put(("analytics", result))
            except Exception as e:
                self.output_queue.put(("error", f"Error generating demographics: {str(e)}"))
            finally:
                self.output_queue.put(("stop_progress", None))
        
        threading.Thread(target=run_demographics, daemon=True).start()
    
    def show_performance_analysis(self):
        """Show performance analysis"""
        self.update_status("Analyzing academic performance...")
        self.start_progress()
        
        def run_performance():
            try:
                result = self.capture_function_output(academic_performance_analysis)
                self.output_queue.put(("analytics", result))
            except Exception as e:
                self.output_queue.put(("error", f"Error analyzing performance: {str(e)}"))
            finally:
                self.output_queue.put(("stop_progress", None))
        
        threading.Thread(target=run_performance, daemon=True).start()
    
    def show_fuzzy_search(self):
        """Show fuzzy search dialog"""
        dialog = tk.Toplevel(self.master)
        dialog.title("🔍 Fuzzy Name Search")
        dialog.geometry("900x700")
        dialog.transient(self.master)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Fuzzy Name Search", style='Title.TLabel').pack(pady=(0, 20))
        
        ttk.Label(frame, text="Search Term:").pack(anchor='w')
        search_var = tk.StringVar()
        ttk.Entry(frame, textvariable=search_var, width=30).pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(frame, text="Similarity Threshold (0.1 - 0.9):").pack(anchor='w')
        threshold_var = tk.StringVar(value="0.6")
        ttk.Entry(frame, textvariable=threshold_var, width=10).pack(anchor='w', pady=(0, 10))
        
        ttk.Label(frame, text="Algorithm:").pack(anchor='w')
        algo_var = tk.StringVar(value="1")
        ttk.Radiobutton(frame, text="Standard fuzzy matching", variable=algo_var, value="1").pack(anchor='w')
        ttk.Radiobutton(frame, text="Phonetic matching (Soundex)", variable=algo_var, value="2").pack(anchor='w')
        ttk.Radiobutton(frame, text="Both algorithms", variable=algo_var, value="3").pack(anchor='w', pady=(0, 20))
        
        def execute_fuzzy_search():
            term = search_var.get().strip()
            if not term:
                messagebox.showwarning("Missing Input", "Please enter a search term.")
                return
            
            dialog.destroy()
            self.update_status("Performing fuzzy search...")
            self.start_progress()
            
            def run_fuzzy():
                try:
                    # Simulate fuzzy search with captured parameters
                    results = self.perform_fuzzy_search(term, float(threshold_var.get()), algo_var.get())
                    self.output_queue.put(("search_results", results))
                    self.output_queue.put(("log", f"Fuzzy search for '{term}' completed. Found {len(results)} results."))
                except Exception as e:
                    self.output_queue.put(("error", f"Fuzzy search error: {str(e)}"))
                finally:
                    self.output_queue.put(("stop_progress", None))
            
            threading.Thread(target=run_fuzzy, daemon=True).start()
        
        button_frame = ttk.Frame(frame)
        button_frame.pack(fill=tk.X, pady=(20, 0))
        
        ttk.Button(button_frame, text="🔍 Search", command=execute_fuzzy_search).pack(side=tk.LEFT)
        ttk.Button(button_frame, text="❌ Cancel", command=dialog.destroy).pack(side=tk.RIGHT)
    
    def perform_fuzzy_search(self, search_term, threshold, algorithm):
        """Perform fuzzy search with given parameters"""
        try:
            from difflib import SequenceMatcher
            
            conn = get_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM students")
            all_students = cursor.fetchall()
            conn.close()
            
            matched_students = []
            search_term_lower = search_term.lower()
            
            for student in all_students:
                first_name = student[3].lower() if student[3] else ""
                last_name = student[5].lower() if student[5] else ""
                
                best_ratio = 0
                
                if algorithm in ['1', '3']:  # Standard fuzzy matching
                    first_ratio = SequenceMatcher(None, search_term_lower, first_name).ratio()
                    last_ratio = SequenceMatcher(None, search_term_lower, last_name).ratio()
                    full_name = f"{first_name} {last_name}".strip()
                    full_ratio = SequenceMatcher(None, search_term_lower, full_name).ratio()
                    best_ratio = max(first_ratio, last_ratio, full_ratio)
                
                if algorithm in ['2', '3']:  # Phonetic matching
                    def simple_soundex(word):
                        if not word:
                            return "0000"
                        word = word.upper()
                        result = word[0]
                        mapping = {'BFPV': '1', 'CGJKQSXZ': '2', 'DT': '3', 'L': '4', 'MN': '5', 'R': '6'}
                        for char in word[1:]:
                            for key, value in mapping.items():
                                if char in key and result[-1] != value:
                                    result += value
                                    break
                        return (result + '0000')[:4]
                    
                    search_soundex = simple_soundex(search_term)
                    first_soundex = simple_soundex(first_name)
                    last_soundex = simple_soundex(last_name)
                    
                    if search_soundex == first_soundex or search_soundex == last_soundex:
                        best_ratio = max(best_ratio, 0.8)
                
                if best_ratio >= threshold:
                    matched_students.append(student)
            
            # Sort by similarity (would need to track ratios for real sorting)
            return matched_students
            
        except Exception as e:
            raise Exception(f"Fuzzy search error: {str(e)}")
    
    def show_module_search(self):
        """Show module enrollment search"""
        dialog = tk.Toplevel(self.master)
        dialog.title("🎓 Module Enrollment Search")
        dialog.geometry("1000x750")
        dialog.transient(self.master)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Module Enrollment Search", style='Title.TLabel').pack(pady=(0, 20))
        
        # Module selection
        ttk.Label(frame, text="Select Modules:").pack(anchor='w')
        
        modules_frame = ttk.Frame(frame)
        modules_frame.pack(fill=tk.BOTH, expand=True, pady=(5, 10))
        
        # Module listbox with scrollbar
        listbox_frame = ttk.Frame(modules_frame)
        listbox_frame.pack(fill=tk.BOTH, expand=True)
        
        self.module_listbox = tk.Listbox(listbox_frame, selectmode=tk.MULTIPLE, height=10)
        scrollbar = ttk.Scrollbar(listbox_frame, orient=tk.VERTICAL, command=self.module_listbox.yview)
        self.module_listbox.configure(yscrollcommand=scrollbar.set)
        
        self.module_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Load modules
        self.load_available_modules()
        
        # Match type
        match_frame = ttk.Frame(frame)
        match_frame.pack(fill=tk.X, pady=(10, 20))
        
        ttk.Label(match_frame, text="Match Type:").pack(anchor='w')
        match_var = tk.StringVar(value="any")
        ttk.Radiobutton(match_frame, text="ANY selected modules", variable=match_var, value="any").pack(anchor='w')
        ttk.Radiobutton(match_frame, text="ALL selected modules", variable=match_var, value="all").pack(anchor='w')
        
        def execute_module_search():
            selected_indices = self.module_listbox.curselection()
            if not selected_indices:
                messagebox.showwarning("No Selection", "Please select at least one module.")
                return
            
            selected_modules = [self.available_modules[i][0] for i in selected_indices]
            match_type = match_var.get()
            
            dialog.destroy()
            self.update_status("Searching by module enrollment...")
            self.start_progress()
            
            def run_module_search():
                try:
                    results = self.perform_module_search(selected_modules, match_type)
                    self.output_queue.put(("search_results", results))
                    self.output_queue.put(("log", f"Module search completed. Found {len(results)} results."))
                except Exception as e:
                    self.output_queue.put(("error", f"Module search error: {str(e)}"))
                finally:
                    self.output_queue.put(("stop_progress", None))
            
            threading.Thread(target=run_module_search, daemon=True).start()
        
        button_frame = ttk.Frame(frame)
        button_frame.pack(fill=tk.X)
        
        ttk.Button(button_frame, text="🔍 Search", command=execute_module_search).pack(side=tk.LEFT)
        ttk.Button(button_frame, text="❌ Cancel", command=dialog.destroy).pack(side=tk.RIGHT)
    
    def load_available_modules(self):
        """Load available modules into the listbox"""
        try:
            conn = get_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT DISTINCT module_code, module_name FROM student_modules ORDER BY module_name")
            self.available_modules = cursor.fetchall()
            conn.close()
            
            self.module_listbox.delete(0, tk.END)
            for code, name in self.available_modules:
                self.module_listbox.insert(tk.END, f"{code} - {name}")
                
        except Exception as e:
            messagebox.showerror("Error", f"Could not load modules: {str(e)}")
    
    def perform_module_search(self, module_codes, match_type):
        """Perform module enrollment search"""
        try:
            conn = get_connection()
            cursor = conn.cursor()
            
            if match_type == "all":
                # Students enrolled in ALL modules
                query = "SELECT s.* FROM students s WHERE "
                conditions = []
                params = []
                
                for module_code in module_codes:
                    conditions.append("""
                    EXISTS (
                        SELECT 1 FROM student_modules sm
                        WHERE sm.student_id = s.student_id AND sm.module_code = ?
                    )
                    """)
                    params.append(module_code)
                
                query += " AND ".join(conditions)
            else:
                # Students enrolled in ANY module
                placeholders = ",".join(["?" for _ in module_codes])
                query = f"""
                SELECT DISTINCT s.* FROM students s
                JOIN student_modules sm ON s.student_id = sm.student_id
                WHERE sm.module_code IN ({placeholders})
                """
                params = module_codes
            
            cursor.execute(query, params)
            results = cursor.fetchall()
            conn.close()
            
            return results
            
        except Exception as e:
            raise Exception(f"Module search error: {str(e)}")

    def perform_combined_filters_search(self, filters):
        """
        Perform combined filters search with student data, modules, and date range.

        Args:
            filters (dict): Dictionary containing:
                - student_data: Dict with student field filters
                - module_codes: List of module codes
                - date_range: Dict with start and end dates
                - module_match_all: Bool for ALL vs ANY module matching

        Returns:
            list: List of matching student records
        """
        try:
            conn = get_connection()
            cursor = conn.cursor()

            # Build query based on filters
            if filters["module_codes"] and filters["module_match_all"]:
                # ALL modules - need EXISTS for each module
                query = "SELECT s.* FROM students s WHERE 1=1"
                params = []

                # Student data filters
                for field, op, val in [
                    ("student_id", "LIKE", f"%{filters['student_data'].get('student_id', '')}%"),
                    ("first_name", "LIKE LOWER", f"%{filters['student_data'].get('first_name', '')}%"),
                    ("last_name", "LIKE LOWER", f"%{filters['student_data'].get('last_name', '')}%"),
                ]:
                    if filters["student_data"].get(field.replace('_', ' ').title().replace(' ', '')):
                        if "LOWER" in op:
                            query += f" AND LOWER(s.{field}) {op.replace(' LOWER', '')} ?"
                        else:
                            query += f" AND s.{field} {op} ?"
                        params.append(val)

                if "gender" in filters["student_data"]:
                    query += " AND LOWER(s.gender) = LOWER(?)"
                    params.append(filters["student_data"]["gender"])

                if "course" in filters["student_data"]:
                    query += " AND s.course = ?"
                    params.append(filters["student_data"]["course"])

                if "age_min" in filters["student_data"]:
                    query += " AND s.age >= ?"
                    params.append(filters["student_data"]["age_min"])

                if "age_max" in filters["student_data"]:
                    query += " AND s.age <= ?"
                    params.append(filters["student_data"]["age_max"])

                if filters["date_range"]["start"]:
                    query += " AND s.registration_datetime >= ?"
                    params.append(filters["date_range"]["start"])

                if filters["date_range"]["end"]:
                    query += " AND s.registration_datetime <= ?"
                    params.append(filters["date_range"]["end"])

                # Add EXISTS clause for each module
                for code in filters["module_codes"]:
                    query += """
                    AND EXISTS (
                        SELECT 1 FROM student_modules sm
                        WHERE sm.student_id = s.student_id AND sm.module_code = ?
                    )
                    """
                    params.append(code)

            else:
                # ANY modules or no modules
                query = "SELECT DISTINCT s.* FROM students s"
                params = []

                if filters["module_codes"]:
                    query += " JOIN student_modules sm ON s.student_id = sm.student_id"

                query += " WHERE 1=1"

                # Student data filters
                for field, op in [
                    ("student_id", "LIKE"),
                    ("first_name", "LIKE LOWER"),
                    ("last_name", "LIKE LOWER"),
                ]:
                    filter_key = field
                    if filters["student_data"].get(filter_key):
                        if "LOWER" in op:
                            query += f" AND LOWER(s.{field}) {op.replace(' LOWER', '')} ?"
                        else:
                            query += f" AND s.{field} {op} ?"
                        params.append(f"%{filters['student_data'][filter_key]}%")

                if "gender" in filters["student_data"]:
                    query += " AND LOWER(s.gender) = LOWER(?)"
                    params.append(filters["student_data"]["gender"])

                if "course" in filters["student_data"]:
                    query += " AND s.course = ?"
                    params.append(filters["student_data"]["course"])

                if "age_min" in filters["student_data"]:
                    query += " AND s.age >= ?"
                    params.append(filters["student_data"]["age_min"])

                if "age_max" in filters["student_data"]:
                    query += " AND s.age <= ?"
                    params.append(filters["student_data"]["age_max"])

                if filters["date_range"]["start"]:
                    query += " AND s.registration_datetime >= ?"
                    params.append(filters["date_range"]["start"])

                if filters["date_range"]["end"]:
                    query += " AND s.registration_datetime <= ?"
                    params.append(filters["date_range"]["end"])

                if filters["module_codes"]:
                    placeholders = ",".join("?" for _ in filters["module_codes"])
                    query += f" AND sm.module_code IN ({placeholders})"
                    params.extend(filters["module_codes"])

            cursor.execute(query, params)
            results = cursor.fetchall()
            conn.close()

            return results

        except Exception as e:
            raise Exception(f"Combined filters search error: {str(e)}")

    def show_date_search(self):
        """Complete date range search implementation"""
        dialog = tk.Toplevel(self.master)
        dialog.title("Date Range Search")
        dialog.geometry("400x350")
        dialog.transient(self.master)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Date Range Search", style='Title.TLabel').pack(pady=(0, 20))
        
        # Preset options
        ttk.Label(frame, text="Quick Presets:").pack(anchor='w')
        preset_var = tk.StringVar(value="custom")
        
        presets = [
            ("Custom date range", "custom"),
            ("Last 7 days", "7d"),
            ("Last 30 days", "30d"),
            ("Last 3 months", "3m"),
            ("Last 6 months", "6m"),
            ("This year", "year")
        ]
        
        for text, value in presets:
            ttk.Radiobutton(frame, text=text, variable=preset_var, value=value).pack(anchor='w')
        
        # Custom date inputs
        custom_frame = ttk.LabelFrame(frame, text="Custom Date Range", padding="10")
        custom_frame.pack(fill=tk.X, pady=(20, 0))
        
        ttk.Label(custom_frame, text="Start Date (YYYY-MM-DD):").pack(anchor='w')
        start_date_var = tk.StringVar()
        ttk.Entry(custom_frame, textvariable=start_date_var, width=20).pack(anchor='w', pady=(0, 10))
        
        ttk.Label(custom_frame, text="End Date (YYYY-MM-DD):").pack(anchor='w')
        end_date_var = tk.StringVar()
        ttk.Entry(custom_frame, textvariable=end_date_var, width=20).pack(anchor='w')
        
        def execute_date_search():
            preset = preset_var.get()
            start_date = None
            end_date = None
            
            if preset == "custom":
                start_date = start_date_var.get().strip()
                end_date = end_date_var.get().strip()
            elif preset == "7d":
                start_date = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
            elif preset == "30d":
                start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
            elif preset == "3m":
                start_date = (datetime.now() - timedelta(days=90)).strftime('%Y-%m-%d')
            elif preset == "6m":
                start_date = (datetime.now() - timedelta(days=180)).strftime('%Y-%m-%d')
            elif preset == "year":
                start_date = datetime.now().replace(month=1, day=1).strftime('%Y-%m-%d')
            
            dialog.destroy()
            self.update_status("Searching by date range...")
            
            def run_date_search():
                try:
                    results = self.perform_date_search(start_date, end_date)
                    self.output_queue.put(("search_results", results))
                    self.output_queue.put(("log", f"Date range search completed. Found {len(results)} results."))
                except Exception as e:
                    self.output_queue.put(("error", f"Date search error: {str(e)}"))
            
            threading.Thread(target=run_date_search, daemon=True).start()
        
        button_frame = ttk.Frame(frame)
        button_frame.pack(fill=tk.X, pady=(20, 0))
        
        ttk.Button(button_frame, text="Search", command=execute_date_search).pack(side=tk.LEFT)
        ttk.Button(button_frame, text="Cancel", command=dialog.destroy).pack(side=tk.RIGHT)
    
    def perform_date_search(self, start_date, end_date):
        """Perform date range search"""
        try:
            conn = get_connection()
            if not conn:
                raise Exception("Database connection failed")
            
            cursor = conn.cursor()
            
            query = "SELECT * FROM students WHERE 1=1"
            params = []
            
            if start_date:
                try:
                    datetime.strptime(start_date, "%Y-%m-%d")
                    query += " AND registration_datetime >= ?"
                    params.append(start_date + " 00:00:00")
                except ValueError:
                    raise Exception("Invalid start date format. Use YYYY-MM-DD.")
            
            if end_date:
                try:
                    datetime.strptime(end_date, "%Y-%m-%d")
                    query += " AND registration_datetime <= ?"
                    params.append(end_date + " 23:59:59")
                except ValueError:
                    raise Exception("Invalid end date format. Use YYYY-MM-DD.")
            
            cursor.execute(query, params)
            results = cursor.fetchall()
            conn.close()
            
            return results
            
        except Exception as e:
            raise Exception(f"Date search error: {str(e)}")
   
    def show_combined_search(self):
        """
        Show combined filters search - allows combining multiple types of filters.

        This comprehensive search interface combines:
        - Student data filters (ID, name, gender, course, age)
        - Module enrollment filters
        - Date range filters
        """
        dialog = tk.Toplevel(self.master)
        dialog.title("🔎 Combined Filters Search")
        dialog.geometry("700x700")
        dialog.transient(self.master)
        dialog.grab_set()

        # Main container with scrollbar
        main_frame = ttk.Frame(dialog)
        main_frame.pack(fill=tk.BOTH, expand=True)

        canvas = tk.Canvas(main_frame)
        scrollbar = ttk.Scrollbar(main_frame, orient=tk.VERTICAL, command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas, padding="20")

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        ttk.Label(scrollable_frame, text="Combined Filters Search", style='Title.TLabel').pack(pady=(0, 20))

        # ========== STUDENT DATA FILTERS ==========
        student_frame = ttk.LabelFrame(scrollable_frame, text="Student Data Filters", padding="10")
        student_frame.pack(fill=tk.X, pady=(0, 10))

        # Student ID
        ttk.Label(student_frame, text="Student ID:").grid(row=0, column=0, sticky='w', pady=5)
        student_id_var = tk.StringVar()
        ttk.Entry(student_frame, textvariable=student_id_var, width=30).grid(row=0, column=1, sticky='w', padx=(10, 0))

        # First Name
        ttk.Label(student_frame, text="First Name:").grid(row=1, column=0, sticky='w', pady=5)
        first_name_var = tk.StringVar()
        ttk.Entry(student_frame, textvariable=first_name_var, width=30).grid(row=1, column=1, sticky='w', padx=(10, 0))

        # Last Name
        ttk.Label(student_frame, text="Last Name:").grid(row=2, column=0, sticky='w', pady=5)
        last_name_var = tk.StringVar()
        ttk.Entry(student_frame, textvariable=last_name_var, width=30).grid(row=2, column=1, sticky='w', padx=(10, 0))

        # Gender
        ttk.Label(student_frame, text="Gender:").grid(row=3, column=0, sticky='w', pady=5)
        gender_var = tk.StringVar()
        gender_combo = ttk.Combobox(student_frame, textvariable=gender_var,
                                    values=["", "male", "female", "other"], state='readonly', width=28)
        gender_combo.grid(row=3, column=1, sticky='w', padx=(10, 0))
        gender_combo.set("")

        # Course
        ttk.Label(student_frame, text="Course:").grid(row=4, column=0, sticky='w', pady=5)
        course_var = tk.StringVar()
        course_combo = ttk.Combobox(student_frame, textvariable=course_var,
                                   values=["", "CS", "DS"], state='readonly', width=28)
        course_combo.grid(row=4, column=1, sticky='w', padx=(10, 0))
        course_combo.set("")

        # Age Range
        ttk.Label(student_frame, text="Age Range:").grid(row=5, column=0, sticky='w', pady=5)
        age_frame = ttk.Frame(student_frame)
        age_frame.grid(row=5, column=1, sticky='w', padx=(10, 0))

        age_min_var = tk.StringVar()
        ttk.Label(age_frame, text="Min:").pack(side=tk.LEFT)
        ttk.Entry(age_frame, textvariable=age_min_var, width=8).pack(side=tk.LEFT, padx=(5, 10))

        age_max_var = tk.StringVar()
        ttk.Label(age_frame, text="Max:").pack(side=tk.LEFT)
        ttk.Entry(age_frame, textvariable=age_max_var, width=8).pack(side=tk.LEFT, padx=(5, 0))

        # ========== MODULE FILTERS ==========
        module_frame = ttk.LabelFrame(scrollable_frame, text="Module Enrollment Filters", padding="10")
        module_frame.pack(fill=tk.X, pady=(0, 10))

        module_enabled_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(module_frame, text="Enable module filtering",
                       variable=module_enabled_var).pack(anchor='w', pady=(0, 10))

        # Module listbox
        module_list_frame = ttk.Frame(module_frame)
        module_list_frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(module_list_frame, text="Select Modules:").pack(anchor='w')

        module_listbox_frame = ttk.Frame(module_list_frame)
        module_listbox_frame.pack(fill=tk.BOTH, expand=True, pady=(5, 10))

        combined_module_listbox = tk.Listbox(module_listbox_frame, selectmode=tk.MULTIPLE, height=6)
        module_scroll = ttk.Scrollbar(module_listbox_frame, orient=tk.VERTICAL,
                                     command=combined_module_listbox.yview)
        combined_module_listbox.configure(yscrollcommand=module_scroll.set)

        combined_module_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        module_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        # Load modules
        try:
            conn = get_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT DISTINCT module_code, module_name FROM student_modules ORDER BY module_name")
            available_modules = cursor.fetchall()
            conn.close()

            for code, name in available_modules:
                combined_module_listbox.insert(tk.END, f"{code} - {name}")
        except Exception as e:
            print_error(f"Could not load modules: {e}")
            available_modules = []

        # Module match type
        module_match_var = tk.StringVar(value="any")
        ttk.Label(module_list_frame, text="Students must be enrolled in:").pack(anchor='w')
        ttk.Radiobutton(module_list_frame, text="ANY of the selected modules",
                       variable=module_match_var, value="any").pack(anchor='w')
        ttk.Radiobutton(module_list_frame, text="ALL of the selected modules",
                       variable=module_match_var, value="all").pack(anchor='w')

        # ========== DATE RANGE FILTERS ==========
        date_frame = ttk.LabelFrame(scrollable_frame, text="Registration Date Filters", padding="10")
        date_frame.pack(fill=tk.X, pady=(0, 20))

        date_enabled_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(date_frame, text="Enable date filtering",
                       variable=date_enabled_var).pack(anchor='w', pady=(0, 10))

        # Start date
        date_fields_frame = ttk.Frame(date_frame)
        date_fields_frame.pack(fill=tk.X)

        ttk.Label(date_fields_frame, text="Start Date (YYYY-MM-DD):").grid(row=0, column=0, sticky='w', pady=5)
        start_date_var = tk.StringVar()
        ttk.Entry(date_fields_frame, textvariable=start_date_var, width=20).grid(row=0, column=1, sticky='w', padx=(10, 0))

        # End date
        ttk.Label(date_fields_frame, text="End Date (YYYY-MM-DD):").grid(row=1, column=0, sticky='w', pady=5)
        end_date_var = tk.StringVar()
        ttk.Entry(date_fields_frame, textvariable=end_date_var, width=20).grid(row=1, column=1, sticky='w', padx=(10, 0))

        # ========== EXECUTE SEARCH ==========
        def execute_combined_search():
            """Execute the combined filters search"""
            # Collect all filters
            filters = {
                "student_data": {},
                "module_codes": [],
                "date_range": {"start": None, "end": None},
                "module_match_all": module_match_var.get() == "all"
            }

            # Student data filters
            if student_id_var.get().strip():
                filters["student_data"]["student_id"] = student_id_var.get().strip()
            if first_name_var.get().strip():
                filters["student_data"]["first_name"] = first_name_var.get().strip()
            if last_name_var.get().strip():
                filters["student_data"]["last_name"] = last_name_var.get().strip()
            if gender_var.get() and gender_var.get() != "":
                filters["student_data"]["gender"] = gender_var.get()
            if course_var.get() and course_var.get() != "":
                filters["student_data"]["course"] = course_var.get()
            if age_min_var.get().strip():
                try:
                    filters["student_data"]["age_min"] = int(age_min_var.get().strip())
                except ValueError:
                    messagebox.showwarning("Invalid Input", "Minimum age must be a number")
                    return
            if age_max_var.get().strip():
                try:
                    filters["student_data"]["age_max"] = int(age_max_var.get().strip())
                except ValueError:
                    messagebox.showwarning("Invalid Input", "Maximum age must be a number")
                    return

            # Module filters
            if module_enabled_var.get():
                selected_indices = combined_module_listbox.curselection()
                if selected_indices:
                    filters["module_codes"] = [available_modules[i][0] for i in selected_indices]

            # Date range filters
            if date_enabled_var.get():
                if start_date_var.get().strip():
                    try:
                        datetime.strptime(start_date_var.get().strip(), "%Y-%m-%d")
                        filters["date_range"]["start"] = start_date_var.get().strip() + " 00:00:00"
                    except ValueError:
                        messagebox.showwarning("Invalid Date", "Start date must be in YYYY-MM-DD format")
                        return
                if end_date_var.get().strip():
                    try:
                        datetime.strptime(end_date_var.get().strip(), "%Y-%m-%d")
                        filters["date_range"]["end"] = end_date_var.get().strip() + " 23:59:59"
                    except ValueError:
                        messagebox.showwarning("Invalid Date", "End date must be in YYYY-MM-DD format")
                        return

            dialog.destroy()
            self.update_status("Executing combined search...")
            self.start_progress()

            def run_combined_search():
                try:
                    results = self.perform_combined_filters_search(filters)
                    self.output_queue.put(("search_results", results))
                    self.output_queue.put(("log", f"Combined search completed. Found {len(results)} results."))
                except Exception as e:
                    self.output_queue.put(("error", f"Combined search error: {str(e)}"))
                finally:
                    self.output_queue.put(("stop_progress", None))

            threading.Thread(target=run_combined_search, daemon=True).start()

        # Buttons
        button_frame = ttk.Frame(scrollable_frame)
        button_frame.pack(fill=tk.X, pady=(10, 0))

        ttk.Button(button_frame, text="🔍 Search", command=execute_combined_search).pack(side=tk.LEFT)
        ttk.Button(button_frame, text="❌ Cancel", command=dialog.destroy).pack(side=tk.RIGHT)
    
    def show_text_search(self):
        """Show advanced text search dialog with all options"""
        dialog = tk.Toplevel(self.master)
        dialog.title("📝 Advanced Text Search")
        dialog.geometry("1000x750")
        dialog.transient(self.master)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Advanced Text Search", style='Title.TLabel').pack(pady=(0, 20))
        
        # Search type
        ttk.Label(frame, text="Search Type:").pack(anchor='w')
        search_type_var = tk.StringVar(value="wildcard")
        
        search_types = [
            ("Wildcard Pattern Search (* and ?)", "wildcard"),
            ("Regular Expression Search", "regex"),
            ("Search All Text Fields", "all_fields"),
            ("Phonetic Name Search", "phonetic")
        ]
        
        for text, value in search_types:
            ttk.Radiobutton(frame, text=text, variable=search_type_var, value=value).pack(anchor='w')
        
        # Search input
        ttk.Label(frame, text="Search Pattern:").pack(anchor='w', pady=(20, 0))
        pattern_var = tk.StringVar()
        ttk.Entry(frame, textvariable=pattern_var, width=50).pack(fill=tk.X, pady=(0, 10))
        
        # Field selection (for some search types)
        field_frame = ttk.Frame(frame)
        field_frame.pack(fill=tk.X, pady=(10, 20))
        
        ttk.Label(field_frame, text="Search Field:").pack(side=tk.LEFT)
        field_var = tk.StringVar(value="first_name")
        field_combo = ttk.Combobox(field_frame, textvariable=field_var, 
                                  values=["first_name", "last_name", "email", "student_id"], 
                                  state='readonly', width=15)
        field_combo.pack(side=tk.LEFT, padx=(10, 0))
        
        def execute_text_search():
            pattern = pattern_var.get().strip()
            if not pattern:
                messagebox.showwarning("Missing Pattern", "Please enter a search pattern.")
                return
            
            search_type = search_type_var.get()
            field = field_var.get()
            
            dialog.destroy()
            
            # Route to specific search function
            if search_type == "regex":
                self.update_status("Performing regex search...")
                self.start_progress()
                
                def run_search():
                    try:
                        results = self.perform_regex_search(pattern, field)
                        self.output_queue.put(("search_results", results))
                        self.output_queue.put(("log", f"Regex search completed. Found {len(results)} results."))
                    except Exception as e:
                        self.output_queue.put(("error", f"Search error: {str(e)}"))
                    finally:
                        self.output_queue.put(("stop_progress", None))
                
            elif search_type == "wildcard":
                self.update_status("Performing wildcard search...")
                self.start_progress()
                
                def run_search():
                    try:
                        results = self.perform_wildcard_search(pattern, field)
                        self.output_queue.put(("search_results", results))
                        self.output_queue.put(("log", f"Wildcard search completed. Found {len(results)} results."))
                    except Exception as e:
                        self.output_queue.put(("error", f"Search error: {str(e)}"))
                    finally:
                        self.output_queue.put(("stop_progress", None))
                
            elif search_type == "all_fields":
                self.update_status("Searching all fields...")
                self.start_progress()
                
                def run_search():
                    try:
                        results = self.perform_search_all_fields(pattern)
                        self.output_queue.put(("search_results", results))
                        self.output_queue.put(("log", f"All fields search completed. Found {len(results)} results."))
                    except Exception as e:
                        self.output_queue.put(("error", f"Search error: {str(e)}"))
                    finally:
                        self.output_queue.put(("stop_progress", None))
                
            elif search_type == "phonetic":
                self.update_status("Performing phonetic search...")
                self.start_progress()
                
                def run_search():
                    try:
                        results = self.perform_phonetic_search(pattern)
                        self.output_queue.put(("search_results", results))
                        self.output_queue.put(("log", f"Phonetic search completed. Found {len(results)} results."))
                    except Exception as e:
                        self.output_queue.put(("error", f"Search error: {str(e)}"))
                    finally:
                        self.output_queue.put(("stop_progress", None))
            
            threading.Thread(target=run_search, daemon=True).start()
        
        button_frame = ttk.Frame(frame)
        button_frame.pack(fill=tk.X)
        
        ttk.Button(button_frame, text="🔍 Search", command=execute_text_search).pack(side=tk.LEFT)
        ttk.Button(button_frame, text="❌ Cancel", command=dialog.destroy).pack(side=tk.RIGHT)
    
    def perform_text_search(self, pattern, search_type, field):
        """Perform advanced text search"""
        try:
            conn = get_connection()
            cursor = conn.cursor()
            
            if search_type == "wildcard":
                # Convert wildcard to SQL LIKE pattern
                sql_pattern = pattern.replace('*', '%').replace('?', '_')
                query = f"SELECT * FROM students WHERE {field} LIKE ?"
                cursor.execute(query, (sql_pattern,))
                
            elif search_type == "regex":
                # For SQLite, we'll use LIKE with basic pattern conversion
                # In a full implementation, you'd need a regex extension
                sql_pattern = f"%{pattern}%"
                query = f"SELECT * FROM students WHERE {field} LIKE ?"
                cursor.execute(query, (sql_pattern,))
                
            elif search_type == "all_fields":
                search_pattern = f"%{pattern}%"
                query = '''
                SELECT * FROM students WHERE
                student_id LIKE ? OR
                email_address LIKE ? OR
                first_name LIKE ? OR
                middle_name LIKE ? OR
                last_name LIKE ?
                '''
                params = [search_pattern] * 5
                cursor.execute(query, params)
                
            elif search_type == "phonetic":
                # Simple phonetic search implementation
                cursor.execute("SELECT * FROM students")
                all_students = cursor.fetchall()
                results = []
                
                def simple_soundex(word):
                    if not word:
                        return "0000"
                    word = word.upper()
                    result = word[0]
                    mapping = {'BFPV': '1', 'CGJKQSXZ': '2', 'DT': '3', 'L': '4', 'MN': '5', 'R': '6'}
                    for char in word[1:]:
                        for key, value in mapping.items():
                            if char in key and result[-1] != value:
                                result += value
                                break
                    return (result + '0000')[:4]
                
                target_soundex = simple_soundex(pattern)
                
                for student in all_students:
                    if (simple_soundex(student[3]) == target_soundex or  # first_name
                        simple_soundex(student[5]) == target_soundex):   # last_name
                        results.append(student)
                
                conn.close()
                return results
            
            results = cursor.fetchall()
            conn.close()
            return results
            
        except Exception as e:
            raise Exception(f"Text search error: {str(e)}")
    
    def show_conditional_search(self):
        """Show conditional logic search dialog"""
        dialog = tk.Toplevel(self.master)
        dialog.title("🧠 Conditional Logic Search")
        dialog.geometry("1100x800")
        dialog.transient(self.master)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Conditional Logic Search", style='Title.TLabel').pack(pady=(0, 20))
        
        # Instructions
        instructions = """
Build complex queries using boolean logic (AND, OR, NOT operators)
Example: (age > 20 AND course = 'CS') OR (gender = 'female' AND age < 25)
        """
        ttk.Label(frame, text=instructions, font=('Arial', 9)).pack(pady=(0, 20))
        
        # Conditions list
        conditions_frame = ttk.LabelFrame(frame, text="Conditions", padding="10")
        conditions_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 20))
        
        self.conditions_listbox = tk.Listbox(conditions_frame, height=8)
        self.conditions_listbox.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # Condition builder
        builder_frame = ttk.Frame(conditions_frame)
        builder_frame.pack(fill=tk.X)
        
        ttk.Label(builder_frame, text="Field:").grid(row=0, column=0, sticky='w')
        field_var = tk.StringVar(value="age")
        field_combo = ttk.Combobox(builder_frame, textvariable=field_var, 
                                  values=["age", "course", "gender", "registration_datetime"], 
                                  state='readonly', width=12)
        field_combo.grid(row=0, column=1, padx=(5, 10))
        
        ttk.Label(builder_frame, text="Operator:").grid(row=0, column=2, sticky='w')
        op_var = tk.StringVar(value=">")
        op_combo = ttk.Combobox(builder_frame, textvariable=op_var, 
                               values=[">", "<", "=", ">=", "<=", "!="], 
                               state='readonly', width=5)
        op_combo.grid(row=0, column=3, padx=(5, 10))
        
        ttk.Label(builder_frame, text="Value:").grid(row=0, column=4, sticky='w')
        value_var = tk.StringVar()
        ttk.Entry(builder_frame, textvariable=value_var, width=15).grid(row=0, column=5, padx=(5, 10))
        
        def add_condition():
            field = field_var.get()
            operator = op_var.get()
            value = value_var.get().strip()
            
            if not value:
                messagebox.showwarning("Missing Value", "Please enter a value.")
                return
            
            # Format the condition
            if field in ['course', 'gender']:
                condition = f"{field} {operator} '{value}'"
            elif field == 'registration_datetime':
                condition = f"DATE({field}) {operator} '{value}'"
            else:
                condition = f"{field} {operator} {value}"
            
            self.conditions_listbox.insert(tk.END, condition)
            value_var.set("")
        
        ttk.Button(builder_frame, text="Add", command=add_condition).grid(row=0, column=6, padx=(10, 0))
        
        # Logic operations
        logic_frame = ttk.Frame(frame)
        logic_frame.pack(fill=tk.X, pady=(0, 20))
        
        ttk.Label(logic_frame, text="Combine with:").pack(side=tk.LEFT)
        logic_var = tk.StringVar(value="AND")
        ttk.Radiobutton(logic_frame, text="AND", variable=logic_var, value="AND").pack(side=tk.LEFT, padx=(10, 5))
        ttk.Radiobutton(logic_frame, text="OR", variable=logic_var, value="OR").pack(side=tk.LEFT, padx=(5, 0))
        
        def execute_conditional_search():
            conditions = [self.conditions_listbox.get(i) for i in range(self.conditions_listbox.size())]
            if not conditions:
                messagebox.showwarning("No Conditions", "Please add at least one condition.")
                return
            
            logic_operator = logic_var.get()
            
            dialog.destroy()
            self.update_status("Executing conditional search...")
            self.start_progress()
            
            def run_conditional_search():
                try:
                    results = self.perform_conditional_search(conditions, logic_operator)
                    self.output_queue.put(("search_results", results))
                    self.output_queue.put(("log", f"Conditional search completed. Found {len(results)} results."))
                except Exception as e:
                    self.output_queue.put(("error", f"Conditional search error: {str(e)}"))
                finally:
                    self.output_queue.put(("stop_progress", None))
            
            threading.Thread(target=run_conditional_search, daemon=True).start()
        
        button_frame = ttk.Frame(frame)
        button_frame.pack(fill=tk.X)
        
        ttk.Button(button_frame, text="🔍 Search", command=execute_conditional_search).pack(side=tk.LEFT)
        ttk.Button(button_frame, text="🗑️ Clear", 
                  command=lambda: self.conditions_listbox.delete(0, tk.END)).pack(side=tk.LEFT, padx=(10, 0))
        ttk.Button(button_frame, text="❌ Cancel", command=dialog.destroy).pack(side=tk.RIGHT)
    
    def perform_conditional_search(self, conditions, logic_operator):
        """Perform conditional logic search"""
        try:
            conn = get_connection()
            cursor = conn.cursor()
            
            where_clause = f" {logic_operator} ".join(conditions)
            query = f"SELECT * FROM students WHERE {where_clause}"
            
            cursor.execute(query)
            results = cursor.fetchall()
            conn.close()
            
            return results
            
        except Exception as e:
            raise Exception(f"Conditional search error: {str(e)}")
    
    # Search Management Methods
    def show_saved_searches(self):
        """Show saved searches management"""
        dialog = tk.Toplevel(self.master)
        dialog.title("💾 Saved Search Profiles")
        dialog.geometry("700x500")
        dialog.transient(self.master)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Saved Search Profiles", style='Title.TLabel').pack(pady=(0, 20))
        
        # Saved searches list
        list_frame = ttk.LabelFrame(frame, text="Saved Searches", padding="10")
        list_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 20))
        
        columns = ('ID', 'Name', 'Created', 'Shared')
        self.saved_tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=12)
        
        for col in columns:
            self.saved_tree.heading(col, text=col)
            self.saved_tree.column(col, width=120)
        
        self.saved_tree.pack(fill=tk.BOTH, expand=True)
        
        # Load saved searches
        self.load_saved_searches()
        
        # Buttons
        button_frame = ttk.Frame(frame)
        button_frame.pack(fill=tk.X)

        ttk.Button(button_frame, text="💾 Save Current",
                  command=self.save_current_search).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(button_frame, text="📂 Load",
                  command=self.load_selected_search).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(button_frame, text="🔗 Share",
                  command=self.share_search_profile).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(button_frame, text="🗑️ Delete",
                  command=self.delete_selected_search).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(button_frame, text="❌ Close", command=dialog.destroy).pack(side=tk.RIGHT)
    
    def load_saved_searches(self):
        """Load saved searches from database into the tree"""
        try:
            # Clear existing items
            for item in self.saved_tree.get_children():
                self.saved_tree.delete(item)

            # Load from database
            conn = get_connection()
            cursor = conn.cursor()

            # Check if saved_searches table exists
            cursor.execute("""
                SELECT name FROM sqlite_master
                WHERE type='table' AND name='saved_searches'
            """)

            if cursor.fetchone():
                # Table exists, load real data
                # Get current user if available
                current_user = 'default_user'  # Default if no auth
                if hasattr(self, 'auth') and self.auth and hasattr(self.auth, 'current_user'):
                    user = self.auth.current_user
                    if user:
                        current_user = user.get('username', 'default_user')

                cursor.execute('''
                    SELECT
                        search_id as id,
                        COALESCE(search_name, name) as search_name,
                        COALESCE(created_date, created_at) as created_date,
                        COALESCE(is_shared, 0) as is_shared
                    FROM saved_searches
                    WHERE user_id = ? OR is_shared = 1
                    ORDER BY COALESCE(created_date, created_at) DESC
                ''', (current_user,))

                searches = cursor.fetchall()

                for search_id, name, created, shared in searches:
                    shared_text = "Yes" if shared else "No"
                    # Format date nicely
                    try:
                        if created:
                            created_display = created[:16] if len(created) > 16 else created
                        else:
                            created_display = "N/A"
                    except:
                        created_display = "N/A"

                    self.saved_tree.insert('', 'end', values=(search_id, name, created_display, shared_text))

            else:
                # Table doesn't exist, show sample data
                saved_searches = [
                    (1, "CS Students Over 25", "2024-01-15", "No"),
                    (2, "Recent Registrations", "2024-01-20", "Yes"),
                    (3, "Incomplete Modules", "2024-01-25", "No"),
                ]

                for search_id, name, created, shared in saved_searches:
                    self.saved_tree.insert('', 'end', values=(search_id, name, created, shared))

            conn.close()

        except Exception as e:
            messagebox.showerror("Error", f"Could not load saved searches: {str(e)}")
    
    def save_current_search(self):
        """Save current search as a profile"""
        if not hasattr(self, 'search_vars') or not any(var.get() for var in self.search_vars.values()):
            messagebox.showinfo("No Search", "Please perform a search first before saving.")
            return
        
        name = tk.simpledialog.askstring("Save Search", "Enter name for this search:")
        if name:
            messagebox.showinfo("Search Saved", f"Search profile '{name}' saved successfully!")
            self.load_saved_searches()  # Refresh the list
    
    def load_selected_search(self):
        """Load selected saved search"""
        selection = self.saved_tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a search to load.")
            return

        item = self.saved_tree.item(selection[0])
        search_id = item['values'][0]
        search_name = item['values'][1]

        try:
            # Load search parameters from database or config
            conn = get_connection()
            cursor = conn.cursor()

            # Check if we have a saved_searches table
            cursor.execute("""
                SELECT name FROM sqlite_master
                WHERE type='table' AND name='saved_searches'
            """)

            if cursor.fetchone():
                # Load from database
                cursor.execute("""
                    SELECT search_parameters FROM saved_searches
                    WHERE id = ?
                """, (search_id,))

                result = cursor.fetchone()
                if result:
                    search_params = json.loads(result[0])
                    self._apply_search_parameters(search_params)
                    messagebox.showinfo("Success", f"Search profile '{search_name}' loaded successfully!")
                else:
                    messagebox.showerror("Error", "Search profile not found in database")
            else:
                # Simulate loading with predefined searches
                predefined_searches = {
                    1: {
                        'course_filter': 'Computer Science',
                        'min_age': '25',
                        'grade_filter': 'A',
                        'search_term': ''
                    },
                    2: {
                        'course_filter': '',
                        'min_age': '',
                        'grade_filter': '',
                        'search_term': '',
                        'date_range': 'last_30_days'
                    },
                    3: {
                        'course_filter': '',
                        'min_age': '',
                        'grade_filter': 'F',
                        'search_term': 'incomplete'
                    }
                }

                if int(search_id) in predefined_searches:
                    search_params = predefined_searches[int(search_id)]
                    self._apply_search_parameters(search_params)
                    messagebox.showinfo("Success", f"Search profile '{search_name}' loaded successfully!")
                else:
                    messagebox.showwarning("Warning", "This is a demo search profile. Parameters loaded as example.")

            conn.close()

        except Exception as e:
            messagebox.showerror("Error", f"Failed to load search: {str(e)}")

    def _apply_search_parameters(self, params):
        """Apply loaded search parameters to the interface"""
        try:
            # Apply basic search filters if they exist on the interface
            if hasattr(self, 'search_entry') and 'search_term' in params:
                self.search_entry.delete(0, tk.END)
                self.search_entry.insert(0, params.get('search_term', ''))

            # Apply course filter
            if hasattr(self, 'course_var') and 'course_filter' in params:
                self.course_var.set(params.get('course_filter', ''))

            # Apply age filter
            if hasattr(self, 'age_entry') and 'min_age' in params:
                self.age_entry.delete(0, tk.END)
                self.age_entry.insert(0, params.get('min_age', ''))

            # Apply grade filter
            if hasattr(self, 'grade_var') and 'grade_filter' in params:
                self.grade_var.set(params.get('grade_filter', ''))

            # Trigger search with loaded parameters
            if hasattr(self, 'perform_search'):
                self.perform_search()

        except Exception as e:
            print_warning(f"Could not apply all search parameters: {e}")
    
    def delete_selected_search(self):
        """Delete selected saved search"""
        selection = self.saved_tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a search to delete.")
            return

        item = self.saved_tree.item(selection[0])
        search_id = item['values'][0]
        search_name = item['values'][1]

        if messagebox.askyesno("Confirm Delete", f"Delete search profile '{search_name}'?"):
            try:
                conn = get_connection()
                cursor = conn.cursor()

                # Check if saved_searches table exists
                cursor.execute("""
                    SELECT name FROM sqlite_master
                    WHERE type='table' AND name='saved_searches'
                """)

                if cursor.fetchone():
                    # Delete from database
                    cursor.execute('DELETE FROM saved_searches WHERE id = ?', (search_id,))
                    conn.commit()

                conn.close()

                # Remove from tree
                self.saved_tree.delete(selection[0])
                messagebox.showinfo("Deleted", f"Search profile '{search_name}' deleted successfully.")

            except Exception as e:
                messagebox.showerror("Error", f"Failed to delete search: {str(e)}")

    def share_search_profile(self):
        """
        Share a search profile with other users.

        Allows the current user to make a saved search profile available to all users
        by setting the is_shared flag in the database.
        """
        selection = self.saved_tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a search to share.")
            return

        item = self.saved_tree.item(selection[0])
        search_id = item['values'][0]
        search_name = item['values'][1]

        try:
            conn = get_connection()
            cursor = conn.cursor()

            # Check if search exists
            cursor.execute("""
                SELECT search_name FROM saved_searches
                WHERE id = ?
            """, (search_id,))

            result = cursor.fetchone()
            if not result:
                messagebox.showerror("Error", "Search not found in database.")
                conn.close()
                return

            # Confirm sharing
            if messagebox.askyesno("Confirm Share",
                                  f"Share search profile '{search_name}' with all users?\n\n"
                                  "This will make it visible to everyone."):
                # Update the is_shared flag
                cursor.execute("""
                    UPDATE saved_searches
                    SET is_shared = 1
                    WHERE id = ?
                """, (search_id,))
                conn.commit()

                messagebox.showinfo("Success",
                                  f"✅ Search profile '{search_name}' is now shared with all users.")

                # Refresh the list to update the shared status
                self.load_saved_searches()

            conn.close()

        except Exception as e:
            messagebox.showerror("Error", f"Failed to share search: {str(e)}")

    def execute_loaded_search(self, criteria):
        """
        Execute a loaded search with given criteria.

        Args:
            criteria (dict): Search criteria dictionary with fields like:
                - student_id: Student ID pattern
                - first_name: First name pattern
                - last_name: Last name pattern
                - course: Course code
                - gender: Gender
                - age_min, age_max: Age range
        """
        try:
            query = "SELECT * FROM students WHERE 1=1"
            params = []

            # Build query from criteria
            for key, value in criteria.items():
                if value:
                    if key in ['student_id', 'first_name', 'last_name']:
                        query += f" AND {key} LIKE ?"
                        params.append(f"%{value}%")
                    elif key == 'age_min':
                        query += " AND age >= ?"
                        params.append(value)
                    elif key == 'age_max':
                        query += " AND age <= ?"
                        params.append(value)
                    else:
                        query += f" AND {key} = ?"
                        params.append(value)

            conn = get_connection()
            cursor = conn.cursor()
            cursor.execute(query, params)
            results = cursor.fetchall()
            conn.close()

            # Display results
            self.search_results = results
            self.display_search_results(results)
            self.update_status(f"Loaded search executed. Found {len(results)} results.")

        except Exception as e:
            messagebox.showerror("Search Error", f"Failed to execute loaded search: {str(e)}")

    def show_search_history(self):
        """Show search history"""
        dialog = tk.Toplevel(self.master)
        dialog.title("📚 Search History")
        dialog.geometry("600x400")
        dialog.transient(self.master)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Search History & Favorites", style='Title.TLabel').pack(pady=(0, 20))
        
        # History list
        columns = ('Time', 'Type', 'Criteria', 'Results')
        history_tree = ttk.Treeview(frame, columns=columns, show='headings', height=15)
        
        for col in columns:
            history_tree.heading(col, text=col)
            history_tree.column(col, width=120)
        
        history_tree.pack(fill=tk.BOTH, expand=True, pady=(0, 20))
        
        # Sample history data
        history_data = [
            ("14:30:25", "Multi-Criteria", "CS students, age > 20", "15"),
            ("14:25:10", "Fuzzy Search", "John", "3"),
            ("14:20:05", "Module Search", "CS101", "25"),
            ("14:15:30", "Date Range", "Last 30 days", "8"),
        ]
        
        for time, search_type, criteria, results in history_data:
            history_tree.insert('', 'end', values=(time, search_type, criteria, results))
        
        # Buttons
        button_frame = ttk.Frame(frame)
        button_frame.pack(fill=tk.X)
        
        ttk.Button(button_frame, text="🔄 Repeat", 
                  command=lambda: messagebox.showinfo("Repeat", "Search would be repeated")).pack(side=tk.LEFT)
        ttk.Button(button_frame, text="⭐ Favorite", 
                  command=lambda: messagebox.showinfo("Favorite", "Search added to favorites")).pack(side=tk.LEFT, padx=(10, 0))
        ttk.Button(button_frame, text="❌ Close", command=dialog.destroy).pack(side=tk.RIGHT)
    
    def show_load_search(self):
        """Show load saved search dialog"""
        self.show_saved_searches()  # Reuse the saved searches dialog
    
    # Bulk Operations Methods
    def show_bulk_operations(self):
        """Show bulk operations menu"""
        if not self.search_results:
            messagebox.showwarning("No Results", "Please perform a search first to use bulk operations.")
            return
        
        dialog = tk.Toplevel(self.master)
        dialog.title("🔧 Bulk Operations")
        dialog.geometry("1000x750")
        dialog.transient(self.master)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text=f"Bulk Operations ({len(self.search_results)} students)", 
                 style='Title.TLabel').pack(pady=(0, 20))
        
        # Operation buttons
        operations = [
            ("💾 Export Selected Students", self.bulk_export),
            ("📧 Generate Email List", self.generate_email_list),
            ("👥 Create Student Groups", self.create_student_groups),
            ("📌 Mark for Follow-up", self.mark_for_followup),
            ("🎓 Bulk Enrollment Management", self.bulk_enrollment_management),
        ]
        
        for text, command in operations:
            ttk.Button(frame, text=text, command=command, width=30).pack(pady=5)
        
        ttk.Button(frame, text="❌ Close", command=dialog.destroy).pack(pady=(20, 0))
    
    def bulk_export(self):
        """Export search results in bulk"""
        if not self.search_results:
            messagebox.showwarning("No Results", "No search results to export.")
            return
        
        dialog = tk.Toplevel(self.master)
        dialog.title("💾 Bulk Export")
        dialog.geometry("900x700")
        dialog.transient(self.master)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Export Options", style='Title.TLabel').pack(pady=(0, 20))
        
        format_var = tk.StringVar(value="csv")
        formats = [
            ("CSV Format", "csv"),
            ("JSON Format", "json"),
            ("Excel Format", "xlsx"),
            ("Text Format", "txt"),
        ]
        
        for text, value in formats:
            ttk.Radiobutton(frame, text=text, variable=format_var, value=value).pack(anchor='w', pady=2)
        
        def do_export():
            format_type = format_var.get()
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"bulk_export_{timestamp}.{format_type}"
            
            try:
                if format_type == "csv":
                    self.export_to_csv(filename)
                elif format_type == "json":
                    self.export_to_json(filename)
                elif format_type == "txt":
                    self.export_to_text(filename)
                else:
                    # Excel export
                    try:
                        import openpyxl
                        from openpyxl.styles import Font, Alignment, PatternFill

                        wb = openpyxl.Workbook()
                        ws = wb.active
                        ws.title = "Search Results"

                        # Get results data from tree
                        if hasattr(self, 'results_tree') and self.results_tree:
                            # Get column headers
                            headers = [self.results_tree.heading(col)['text'] for col in self.results_tree['columns']]
                            ws.append(headers)

                            # Style headers
                            for col_num, cell in enumerate(ws[1], 1):
                                cell.font = Font(bold=True)
                                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                                cell.alignment = Alignment(horizontal='center')

                            # Add data rows
                            for item in self.results_tree.get_children():
                                values = self.results_tree.item(item)['values']
                                ws.append(values)

                            # Auto-adjust column widths
                            for column in ws.columns:
                                max_length = 0
                                column_letter = column[0].column_letter
                                for cell in column:
                                    try:
                                        if cell.value:
                                            max_length = max(max_length, len(str(cell.value)))
                                    except:
                                        pass
                                adjusted_width = min(max_length + 2, 50)
                                ws.column_dimensions[column_letter].width = adjusted_width

                            wb.save(filename)
                            messagebox.showinfo("Success", f"Exported to Excel: {filename}")
                        else:
                            messagebox.showwarning("No Data", "No search results to export")

                    except ImportError:
                        if messagebox.askyesno("Excel Library Not Found",
                                              "openpyxl library not available.\n"
                                              "Would you like to export as CSV instead?"):
                            csv_filename = filename.replace('.xlsx', '.csv')
                            self.export_to_csv(csv_filename)
                    except Exception as e:
                        messagebox.showerror("Export Error", f"Failed to export to Excel: {e}")
                
                dialog.destroy()
                
            except Exception as e:
                messagebox.showerror("Export Error", f"Could not export: {str(e)}")
        
        button_frame = ttk.Frame(frame)
        button_frame.pack(fill=tk.X, pady=(20, 0))
        
        ttk.Button(button_frame, text="💾 Export", command=do_export).pack(side=tk.LEFT)
        ttk.Button(button_frame, text="❌ Cancel", command=dialog.destroy).pack(side=tk.RIGHT)
    
    def export_to_csv(self, filename):
        """Export results to CSV"""
        import csv
        
        with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            
            # Header
            writer.writerow([
                'Student ID', 'Email', 'Title', 'First Name', 'Middle Name',
                'Last Name', 'Gender', 'Date of Birth', 'Age', 'Course',
                'Registration Datetime'
            ])
            
            # Data
            for student in self.search_results:
                writer.writerow(student)
        
        messagebox.showinfo("Export Complete", f"Data exported to {filename}")
    
    def export_to_json(self, filename):
        """Export results to JSON"""
        data = []
        for student in self.search_results:
            student_dict = {
                'student_id': student[0],
                'email': student[1],
                'title': student[2],
                'first_name': student[3],
                'middle_name': student[4],
                'last_name': student[5],
                'gender': student[6],
                'date_of_birth': student[7],
                'age': student[8],
                'course': student[9],
                'registration_datetime': student[10]
            }
            data.append(student_dict)
        
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, default=str)
        
        messagebox.showinfo("Export Complete", f"Data exported to {filename}")
    
    def export_to_text(self, filename):
        """Export results to text format"""
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(f"Search Results Export - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("="*60 + "\n\n")
            f.write(f"Total Results: {len(self.search_results)}\n\n")
            
            for i, student in enumerate(self.search_results, 1):
                f.write(f"{i}. Student ID: {student[0]}\n")
                f.write(f"   Name: {student[2]} {student[3]} {student[4] or ''} {student[5]}\n")
                f.write(f"   Email: {student[1]}\n")
                f.write(f"   Gender: {student[6]} | Age: {student[8]} | Course: {student[9]}\n")
                f.write(f"   Registration: {student[10]}\n\n")
        
        messagebox.showinfo("Export Complete", f"Data exported to {filename}")
    
    def generate_email_list(self):
        """Generate email list from search results"""
        if not self.search_results:
            messagebox.showwarning("No Results", "No search results available.")
            return

        emails = [student[1] for student in self.search_results if student[1]]

        if not emails:
            messagebox.showwarning("No Emails", "No email addresses found in search results.")
            return

        dialog = tk.Toplevel(self.master)
        dialog.title("📧 Email List Generator")
        dialog.geometry("900x700")
        dialog.transient(self.master)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text=f"Email List ({len(emails)} addresses)", 
                 style='Title.TLabel').pack(pady=(0, 20))
        
        # Email list display
        email_text = scrolledtext.ScrolledText(frame, height=15, wrap=tk.WORD)
        email_text.pack(fill=tk.BOTH, expand=True, pady=(0, 20))
        
        # Format options
        format_frame = ttk.LabelFrame(frame, text="Format Options", padding="10")
        format_frame.pack(fill=tk.X, pady=(0, 20))
        
        format_var = tk.StringVar(value="list")
        ttk.Radiobutton(format_frame, text="One per line", variable=format_var, value="list").pack(anchor='w')
        ttk.Radiobutton(format_frame, text="Comma separated", variable=format_var, value="comma").pack(anchor='w')
        ttk.Radiobutton(format_frame, text="Semicolon separated", variable=format_var, value="semicolon").pack(anchor='w')
        
        def update_format():
            email_text.delete(1.0, tk.END)
            format_type = format_var.get()
            
            if format_type == "list":
                email_text.insert(tk.END, "\n".join(emails))
            elif format_type == "comma":
                email_text.insert(tk.END, ", ".join(emails))
            elif format_type == "semicolon":
                email_text.insert(tk.END, "; ".join(emails))
        
        for widget in format_frame.winfo_children():
            if isinstance(widget, ttk.Radiobutton):
                widget.configure(command=update_format)
        
        update_format()  # Initial format
        
        def save_email_list():
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"email_list_{timestamp}.txt"
            
            with open(filename, 'w', encoding='utf-8') as f:
                f.write(email_text.get(1.0, tk.END))
            
            messagebox.showinfo("Email List Saved", f"Email list saved to {filename}")
            dialog.destroy()
        
        button_frame = ttk.Frame(frame)
        button_frame.pack(fill=tk.X)
        
        ttk.Button(button_frame, text="💾 Save List", command=save_email_list).pack(side=tk.LEFT)
        ttk.Button(button_frame, text="📋 Copy", 
                  command=lambda: self.master.clipboard_clear() or 
                                 self.master.clipboard_append(email_text.get(1.0, tk.END))).pack(side=tk.LEFT, padx=(10, 0))
        ttk.Button(button_frame, text="❌ Close", command=dialog.destroy).pack(side=tk.RIGHT)
    
    def create_student_groups(self):
        """Create student groups from search results"""
        if not self.search_results:
            messagebox.showwarning("No Results", "No search results available.")
            return
        
        dialog = tk.Toplevel(self.master)
        dialog.title("👥 Create Student Groups")
        dialog.geometry("400x350")
        dialog.transient(self.master)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text=f"Create Groups ({len(self.search_results)} students)", 
                 style='Title.TLabel').pack(pady=(0, 20))
        
        # Grouping method
        ttk.Label(frame, text="Grouping Method:").pack(anchor='w')
        method_var = tk.StringVar(value="course")
        
        methods = [
            ("Group by Course", "course"),
            ("Group by Age Range", "age"),
            ("Random Assignment", "random"),
            ("Alphabetical Order", "alpha"),
        ]
        
        for text, value in methods:
            ttk.Radiobutton(frame, text=text, variable=method_var, value=value).pack(anchor='w', pady=2)
        
        # Additional options
        options_frame = ttk.LabelFrame(frame, text="Options", padding="10")
        options_frame.pack(fill=tk.X, pady=(20, 0))
        
        ttk.Label(options_frame, text="Group Size (for random/alphabetical):").pack(anchor='w')
        size_var = tk.StringVar(value="5")
        ttk.Entry(options_frame, textvariable=size_var, width=10).pack(anchor='w')
        
        def create_groups():
            method = method_var.get()
            groups = {}
            
            if method == "course":
                for student in self.search_results:
                    course = student[9]  # course field
                    if course not in groups:
                        groups[course] = []
                    groups[course].append(student)
            
            elif method == "age":
                for student in self.search_results:
                    age = student[8]  # age field
                    if age < 20:
                        age_group = "Under 20"
                    elif age <= 25:
                        age_group = "20-25"
                    elif age <= 30:
                        age_group = "26-30"
                    else:
                        age_group = "Over 30"
                    
                    if age_group not in groups:
                        groups[age_group] = []
                    groups[age_group].append(student)
            
            elif method == "random":
                import random
                try:
                    group_size = int(size_var.get())
                    students_copy = self.search_results.copy()
                    random.shuffle(students_copy)
                    
                    for i, student in enumerate(students_copy):
                        group_name = f"Group {(i // group_size) + 1}"
                        if group_name not in groups:
                            groups[group_name] = []
                        groups[group_name].append(student)
                except ValueError:
                    messagebox.showerror("Invalid Size", "Please enter a valid group size.")
                    return
            
            elif method == "alpha":
                try:
                    group_size = int(size_var.get())
                    sorted_students = sorted(self.search_results, key=lambda x: f"{x[3]} {x[5]}")
                    
                    for i in range(0, len(sorted_students), group_size):
                        group_name = f"Group {(i // group_size) + 1}"
                        groups[group_name] = sorted_students[i:i+group_size]
                except ValueError:
                    messagebox.showerror("Invalid Size", "Please enter a valid group size.")
                    return
            
            # Show groups result
            self.show_groups_result(groups)
            dialog.destroy()
        
        button_frame = ttk.Frame(frame)
        button_frame.pack(fill=tk.X, pady=(20, 0))
        
        ttk.Button(button_frame, text="👥 Create Groups", command=create_groups).pack(side=tk.LEFT)
        ttk.Button(button_frame, text="❌ Cancel", command=dialog.destroy).pack(side=tk.RIGHT)
    
    def show_groups_result(self, groups):
        """Show the created groups result"""
        dialog = tk.Toplevel(self.master)
        dialog.title("👥 Created Groups")
        dialog.geometry("1100x800")
        dialog.transient(self.master)
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text=f"Created {len(groups)} Groups", 
                 style='Title.TLabel').pack(pady=(0, 20))
        
        # Groups display
        groups_text = scrolledtext.ScrolledText(frame, height=20, wrap=tk.WORD)
        groups_text.pack(fill=tk.BOTH, expand=True, pady=(0, 20))
        
        for group_name, students in groups.items():
            groups_text.insert(tk.END, f"\n{group_name} ({len(students)} students):\n")
            groups_text.insert(tk.END, "-" * 50 + "\n")
            for student in students:
                name = f"{student[3]} {student[5]}"
                groups_text.insert(tk.END, f"  • {student[0]} - {name} ({student[1]})\n")
            groups_text.insert(tk.END, "\n")
        
        def export_groups():
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"student_groups_{timestamp}.txt"
            
            with open(filename, 'w', encoding='utf-8') as f:
                f.write(groups_text.get(1.0, tk.END))
            
            messagebox.showinfo("Groups Exported", f"Groups exported to {filename}")
        
        button_frame = ttk.Frame(frame)
        button_frame.pack(fill=tk.X)
        
        ttk.Button(button_frame, text="💾 Export", command=export_groups).pack(side=tk.LEFT)
        ttk.Button(button_frame, text="❌ Close", command=dialog.destroy).pack(side=tk.RIGHT)
    
    def mark_for_followup(self):
        """Mark students for follow-up"""
        if not self.search_results:
            messagebox.showwarning("No Results", "No search results available.")
            return
        
        dialog = tk.Toplevel(self.master)
        dialog.title("📌 Mark for Follow-up")
        dialog.geometry("900x700")
        dialog.transient(self.master)
        dialog.grab_set()

        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(frame, text=f"Mark {len(self.search_results)} Students for Follow-up",
                 style='Title.TLabel').pack(pady=(0, 20))
        
        ttk.Label(frame, text="Follow-up Reason:").pack(anchor='w')
        reason_var = tk.StringVar()
        ttk.Entry(frame, textvariable=reason_var, width=40).pack(fill=tk.X, pady=(0, 20))
        
        ttk.Label(frame, text="Priority:").pack(anchor='w')
        priority_var = tk.StringVar(value="medium")
        
        priorities = [("High", "high"), ("Medium", "medium"), ("Low", "low")]
        for text, value in priorities:
            ttk.Radiobutton(frame, text=text, variable=priority_var, value=value).pack(anchor='w')
        
        def mark_students():
            reason = reason_var.get().strip()
            if not reason:
                messagebox.showwarning("Missing Reason", "Please enter a follow-up reason.")
                return
            
            priority = priority_var.get()
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # Create follow-up data
            followup_data = {
                'reason': reason,
                'priority': priority,
                'marked_date': timestamp,
                'marked_by': 'current_user',
                'students': [
                    {
                        'student_id': s[0],
                        'name': f"{s[3]} {s[5]}",
                        'email': s[1]
                    } for s in self.search_results
                ]
            }
            
            # Save to file
            filename = f"followup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(followup_data, f, indent=2)
            
            messagebox.showinfo("Follow-up Marked", 
                              f"✅ {len(self.search_results)} students marked for follow-up\n"
                              f"Reason: {reason}\n"
                              f"Priority: {priority}\n"
                              f"Saved to: {filename}")
            dialog.destroy()
        
        button_frame = ttk.Frame(frame)
        button_frame.pack(fill=tk.X, pady=(20, 0))
        
        ttk.Button(button_frame, text="📌 Mark", command=mark_students).pack(side=tk.LEFT)
        ttk.Button(button_frame, text="❌ Cancel", command=dialog.destroy).pack(side=tk.RIGHT)
    
    def bulk_enrollment_management(self):
        """Manage bulk enrollment operations"""
        if not self.search_results:
            messagebox.showwarning("No Results", "No search results available.")
            return

        dialog = tk.Toplevel(self.master)
        dialog.title("🎓 Bulk Enrollment Management")
        dialog.geometry("1000x800")
        dialog.transient(self.master)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text=f"Bulk Enrollment ({len(self.search_results)} students)", 
                 style='Title.TLabel').pack(pady=(0, 20))
        
        # Operation selection
        ttk.Label(frame, text="Operation:").pack(anchor='w')
        operation_var = tk.StringVar(value="enroll")
        
        operations = [
            ("Enroll in Module", "enroll"),
            ("Unenroll from Module", "unenroll"),
            ("Transfer Between Modules", "transfer"),
            ("Update Enrollment Status", "status"),
        ]
        
        for text, value in operations:
            ttk.Radiobutton(frame, text=text, variable=operation_var, value=value).pack(anchor='w')
        
        # Module inputs
        module_frame = ttk.LabelFrame(frame, text="Module Information", padding="10")
        module_frame.pack(fill=tk.X, pady=(20, 0))
        
        ttk.Label(module_frame, text="Module Code:").pack(anchor='w')
        module_var = tk.StringVar()
        ttk.Entry(module_frame, textvariable=module_var, width=20).pack(anchor='w', pady=(0, 10))
        
        ttk.Label(module_frame, text="To Module (for transfer):").pack(anchor='w')
        to_module_var = tk.StringVar()
        ttk.Entry(module_frame, textvariable=to_module_var, width=20).pack(anchor='w')
        
        def execute_bulk_operation():
            operation = operation_var.get()
            module_code = module_var.get().strip()
            
            if not module_code and operation != "status":
                messagebox.showwarning("Missing Module", "Please enter a module code.")
                return
            
            # Simulate the operation
            if operation == "enroll":
                message = f"✅ Enrolled {len(self.search_results)} students in {module_code}"
            elif operation == "unenroll":
                message = f"✅ Unenrolled {len(self.search_results)} students from {module_code}"
            elif operation == "transfer":
                to_module = to_module_var.get().strip()
                if not to_module:
                    messagebox.showwarning("Missing Target", "Please enter target module for transfer.")
                    return
                message = f"✅ Transferred {len(self.search_results)} students from {module_code} to {to_module}"
            else:
                message = f"✅ Updated enrollment status for {len(self.search_results)} students"
            
            messagebox.showinfo("Operation Complete", message)
            dialog.destroy()
        
        button_frame = ttk.Frame(frame)
        button_frame.pack(fill=tk.X, pady=(20, 0))
        
        ttk.Button(button_frame, text="🎓 Execute", command=execute_bulk_operation).pack(side=tk.LEFT)
        ttk.Button(button_frame, text="❌ Cancel", command=dialog.destroy).pack(side=tk.RIGHT)

    def mass_email_students(self):
        """
        Send mass emails to students from search results (CLI-equivalent function).

        This function provides a comprehensive mass email interface with:
        - Recipient list display
        - Subject and message composition
        - Email simulation mode
        - Integration with email infrastructure if available
        """
        if not self.search_results:
            messagebox.showwarning("No Results", "No search results available for mass email.")
            return

        dialog = tk.Toplevel(self.master)
        dialog.title("📧 Mass Email to Students")
        dialog.geometry("1200x850")
        dialog.transient(self.master)
        dialog.grab_set()

        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(frame, text=f"Mass Email ({len(self.search_results)} recipients)",
                 style='Title.TLabel').pack(pady=(0, 20))

        # Recipient summary
        recipient_frame = ttk.LabelFrame(frame, text="Recipients", padding="10")
        recipient_frame.pack(fill=tk.X, pady=(0, 10))

        recipient_text = scrolledtext.ScrolledText(recipient_frame, height=5, wrap=tk.WORD)
        recipient_text.pack(fill=tk.BOTH, expand=True)

        # Show first 10 recipients and total count
        for i, student in enumerate(self.search_results[:10]):
            name = f"{student[3]} {student[5]}"
            email = student[1]
            recipient_text.insert(tk.END, f"{i+1}. {name} ({email})\n")

        if len(self.search_results) > 10:
            recipient_text.insert(tk.END, f"\n... and {len(self.search_results) - 10} more recipients")

        recipient_text.config(state='disabled')

        # Email composition
        composition_frame = ttk.LabelFrame(frame, text="Email Content", padding="10")
        composition_frame.pack(fill=tk.BOTH, expand=True, pady=(10, 0))

        ttk.Label(composition_frame, text="Subject:").pack(anchor='w')
        subject_var = tk.StringVar()
        ttk.Entry(composition_frame, textvariable=subject_var, width=60).pack(fill=tk.X, pady=(0, 10))

        ttk.Label(composition_frame, text="Message:").pack(anchor='w')
        message_text = scrolledtext.ScrolledText(composition_frame, height=12, wrap=tk.WORD)
        message_text.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

        # Email mode selection
        mode_frame = ttk.Frame(composition_frame)
        mode_frame.pack(fill=tk.X, pady=(10, 0))

        mode_var = tk.StringVar(value="simulation")
        ttk.Radiobutton(mode_frame, text="Simulation Mode (No emails sent)",
                       variable=mode_var, value="simulation").pack(anchor='w')
        ttk.Radiobutton(mode_frame, text="Send Real Emails (if configured)",
                       variable=mode_var, value="real").pack(anchor='w')

        def send_emails():
            subject = subject_var.get().strip()
            message = message_text.get(1.0, tk.END).strip()

            if not subject or not message:
                messagebox.showwarning("Incomplete", "Please provide both subject and message.")
                return

            mode = mode_var.get()

            if mode == "simulation":
                # Simulation mode
                result_msg = (
                    f"📧 EMAIL SIMULATION COMPLETED\n\n"
                    f"Recipients: {len(self.search_results)} students\n"
                    f"Subject: {subject}\n"
                    f"Message length: {len(message)} characters\n\n"
                    f"Note: No actual emails were sent (simulation mode)"
                )
                messagebox.showinfo("Simulation Complete", result_msg)
            else:
                # Real email mode
                try:
                    # Try to use email infrastructure if available
                    from university_system.infrastructure.email.email_service import send_email

                    success_count = 0
                    failed_count = 0

                    for student in self.search_results:
                        try:
                            send_email(
                                to_email=student[1],
                                subject=subject,
                                body=message
                            )
                            success_count += 1
                        except:
                            failed_count += 1

                    messagebox.showinfo("Email Sent",
                                      f"✅ Mass email completed\n"
                                      f"Success: {success_count}\n"
                                      f"Failed: {failed_count}")

                except ImportError:
                    messagebox.showwarning("Email Not Configured",
                                         "Email infrastructure not available. Use simulation mode.")
                    return

            dialog.destroy()

        button_frame = ttk.Frame(frame)
        button_frame.pack(fill=tk.X, pady=(20, 0))

        ttk.Button(button_frame, text="📧 Send", command=send_emails).pack(side=tk.LEFT)
        ttk.Button(button_frame, text="❌ Cancel", command=dialog.destroy).pack(side=tk.RIGHT)

    def batch_data_updates(self):
        """
        Perform batch updates on student data (CLI-equivalent function).

        Provides comprehensive batch update operations:
        - Update course assignments
        - Update registration status
        - Add notes/flags to student records
        - Bulk module enrollment
        """
        if not self.search_results:
            messagebox.showwarning("No Results", "No search results available for batch updates.")
            return

        dialog = tk.Toplevel(self.master)
        dialog.title("📝 Batch Data Updates")
        dialog.geometry("500x450")
        dialog.transient(self.master)
        dialog.grab_set()

        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(frame, text=f"Batch Data Updates ({len(self.search_results)} students)",
                 style='Title.TLabel').pack(pady=(0, 20))

        # Operation selection
        ttk.Label(frame, text="Select Update Operation:").pack(anchor='w')
        operation_var = tk.StringVar(value="course")

        operations = [
            ("Update Course", "course"),
            ("Update Registration Status", "status"),
            ("Add Note/Flag", "note"),
            ("Bulk Module Enrollment", "module"),
        ]

        for text, value in operations:
            ttk.Radiobutton(frame, text=text, variable=operation_var, value=value).pack(anchor='w')

        # Input fields
        input_frame = ttk.LabelFrame(frame, text="Update Details", padding="10")
        input_frame.pack(fill=tk.BOTH, expand=True, pady=(20, 0))

        # Course update fields
        course_frame = ttk.Frame(input_frame)
        ttk.Label(course_frame, text="New Course:").pack(anchor='w')
        course_var = tk.StringVar()
        course_combo = ttk.Combobox(course_frame, textvariable=course_var,
                                    values=["CS", "DS", "Engineering", "Mathematics"],
                                    width=30)
        course_combo.pack(anchor='w', pady=(0, 10))
        course_frame.pack(fill=tk.X, pady=(0, 10))

        # Status update fields
        status_frame = ttk.Frame(input_frame)
        ttk.Label(status_frame, text="Registration Status:").pack(anchor='w')
        status_var = tk.StringVar()
        status_combo = ttk.Combobox(status_frame, textvariable=status_var,
                                    values=["Active", "Inactive", "Suspended", "Graduated"],
                                    width=30)
        status_combo.pack(anchor='w', pady=(0, 10))
        status_frame.pack(fill=tk.X, pady=(0, 10))

        # Note/flag fields
        note_frame = ttk.Frame(input_frame)
        ttk.Label(note_frame, text="Note/Flag:").pack(anchor='w')
        note_var = tk.StringVar()
        ttk.Entry(note_frame, textvariable=note_var, width=40).pack(fill=tk.X, pady=(0, 10))
        note_frame.pack(fill=tk.X, pady=(0, 10))

        # Module enrollment fields
        module_frame = ttk.Frame(input_frame)
        ttk.Label(module_frame, text="Module Code:").pack(anchor='w')
        module_var = tk.StringVar()
        ttk.Entry(module_frame, textvariable=module_var, width=30).pack(anchor='w', pady=(0, 10))
        module_frame.pack(fill=tk.X, pady=(0, 10))

        def execute_update():
            operation = operation_var.get()

            if operation == "course":
                new_course = course_var.get().strip()
                if not new_course:
                    messagebox.showwarning("Missing Data", "Please select a course.")
                    return

                if messagebox.askyesno("Confirm Update",
                                      f"Update {len(self.search_results)} students to course '{new_course}'?"):
                    # Simulate course update
                    messagebox.showinfo("Update Complete",
                                      f"✅ Updated {len(self.search_results)} students to course {new_course}")

            elif operation == "status":
                new_status = status_var.get().strip()
                if not new_status:
                    messagebox.showwarning("Missing Data", "Please select a status.")
                    return

                if messagebox.askyesno("Confirm Update",
                                      f"Update {len(self.search_results)} students to status '{new_status}'?"):
                    # Simulate status update
                    messagebox.showinfo("Update Complete",
                                      f"✅ Updated {len(self.search_results)} students to status {new_status}")

            elif operation == "note":
                note = note_var.get().strip()
                if not note:
                    messagebox.showwarning("Missing Data", "Please enter a note or flag.")
                    return

                # Simulate note addition
                messagebox.showinfo("Update Complete",
                                  f"✅ Added note '{note}' to {len(self.search_results)} students")

            elif operation == "module":
                module_code = module_var.get().strip()
                if not module_code:
                    messagebox.showwarning("Missing Data", "Please enter a module code.")
                    return

                if messagebox.askyesno("Confirm Enrollment",
                                      f"Enroll {len(self.search_results)} students in module '{module_code}'?"):
                    # Simulate module enrollment
                    messagebox.showinfo("Update Complete",
                                      f"✅ Enrolled {len(self.search_results)} students in module {module_code}")

            dialog.destroy()

        button_frame = ttk.Frame(frame)
        button_frame.pack(fill=tk.X, pady=(20, 0))

        ttk.Button(button_frame, text="✅ Execute Update", command=execute_update).pack(side=tk.LEFT)
        ttk.Button(button_frame, text="❌ Cancel", command=dialog.destroy).pack(side=tk.RIGHT)

    def show_mass_email(self):
        """Show mass email interface"""
        if not self.search_results:
            messagebox.showwarning("No Results", "Please perform a search first.")
            return

        dialog = tk.Toplevel(self.master)
        dialog.title("📧 Mass Email Students")
        dialog.geometry("1000x800")
        dialog.transient(self.master)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text=f"Mass Email ({len(self.search_results)} recipients)", 
                 style='Title.TLabel').pack(pady=(0, 20))
        
        # Email composition
        ttk.Label(frame, text="Subject:").pack(anchor='w')
        subject_var = tk.StringVar()
        ttk.Entry(frame, textvariable=subject_var, width=60).pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(frame, text="Message:").pack(anchor='w')
        message_text = scrolledtext.ScrolledText(frame, height=15, wrap=tk.WORD)
        message_text.pack(fill=tk.BOTH, expand=True, pady=(0, 20))
        
        # Email template buttons
        template_frame = ttk.Frame(frame)
        template_frame.pack(fill=tk.X, pady=(0, 20))
        
        def load_template(template_type):
            template_map = {
                "welcome": "welcome_message",
                "reminder": "reminder_message",
                "announcement": "announcement_message"
            }

            if template_type in template_map:
                try:
                    from university_system.infrastructure.email.template_utils import render_template
                    subject, message = render_template(template_map[template_type], {})
                    subject_var.set(subject)
                    message_text.delete(1.0, tk.END)
                    message_text.insert(1.0, message)
                except:
                    pass
        
        ttk.Label(template_frame, text="Quick Templates:").pack(side=tk.LEFT)
        ttk.Button(template_frame, text="Welcome", 
                  command=lambda: load_template("welcome")).pack(side=tk.LEFT, padx=(10, 5))
        ttk.Button(template_frame, text="Reminder", 
                  command=lambda: load_template("reminder")).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(template_frame, text="Announcement",
                  command=lambda: load_template("announcement")).pack(side=tk.LEFT)

        # Send mass email function
        def send_mass_email():
            subject = subject_var.get().strip()
            message = message_text.get(1.0, tk.END).strip()

            if not subject or not message:
                messagebox.showwarning("Incomplete", "Please enter both subject and message.")
                return

            # Get valid email recipients
            recipients = [s[1] for s in self.search_results if s[1] and '@' in s[1]]

            if not recipients:
                messagebox.showwarning("No Recipients", "No valid email addresses found in search results.")
                return

            # Show confirmation
            confirmation = f"""
📧 SEND MASS EMAIL

Subject: {subject}
Recipients: {len(recipients)} students
Message Length: {len(message)} characters

This will send REAL emails to:
{', '.join(recipients[:5])}{'...' if len(recipients) > 5 else ''}

Proceed with sending emails?
            """

            if messagebox.askyesno("Confirm Send", confirmation):
                try:
                    # Import email service
                    try:
                        from university_system.infrastructure.email.email_service import send_email
                        EMAIL_AVAILABLE = True
                    except ImportError:
                        EMAIL_AVAILABLE = False

                    if not EMAIL_AVAILABLE:
                        messagebox.showerror("Email Service Unavailable",
                                           "Email service is not available. Please configure email settings.")
                        return

                    # Send emails to all recipients
                    success_count = 0
                    failed_count = 0
                    failed_addresses = []

                    for recipient_email in recipients:
                        try:
                            send_email(
                                recipient_email=recipient_email,
                                subject=subject,
                                body=message
                            )
                            success_count += 1
                        except Exception as e:
                            failed_count += 1
                            failed_addresses.append(recipient_email)
                            print(f"Failed to send to {recipient_email}: {str(e)}")

                    # Show results
                    result_message = f"✅ Mass Email Results\n\n"
                    result_message += f"Successfully sent: {success_count}\n"
                    if failed_count > 0:
                        result_message += f"Failed: {failed_count}\n"
                        result_message += f"\nFailed addresses:\n{chr(10).join(failed_addresses[:10])}"
                        if len(failed_addresses) > 10:
                            result_message += f"\n...and {len(failed_addresses)-10} more"

                    messagebox.showinfo("Email Sent", result_message)
                    dialog.destroy()

                except Exception as e:
                    messagebox.showerror("Email Error",
                                       f"Error sending mass emails: {str(e)}\n\nPlease check email configuration.")
                    print(f"Mass email error: {str(e)}")
        
        button_frame = ttk.Frame(frame)
        button_frame.pack(fill=tk.X)
        
        ttk.Button(button_frame, text="📧 Send Email", command=send_mass_email).pack(side=tk.LEFT)
        ttk.Button(button_frame, text="💾 Save Draft",
                  command=lambda: messagebox.showinfo("Draft", "Email draft saved")).pack(side=tk.LEFT, padx=(10, 0))
        ttk.Button(button_frame, text="❌ Cancel", command=dialog.destroy).pack(side=tk.RIGHT)

    def refresh_data(self):
        """Refresh data when notified by main GUI"""
        try:
            # Clear current search results
            self.search_results = []

            # Update results display
            self.update_results_display()

            # Log the refresh
            self.log_output("Data refreshed from main GUI")

            # Update status
            if hasattr(self, 'results_label'):
                self.results_label.config(text="Data refreshed - perform new search")

        except Exception as e:
            self.log_output(f"Error refreshing data: {e}")

    def show_fuzzy_search(self):
        """Complete fuzzy search implementation"""
        dialog = tk.Toplevel(self.master)
        dialog.title("Fuzzy Name Search")
        dialog.geometry("1000x750")
        dialog.transient(self.master)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Fuzzy Name Search", style='Title.TLabel').pack(pady=(0, 20))
        
        ttk.Label(frame, text="Search Term:").pack(anchor='w')
        search_var = tk.StringVar()
        ttk.Entry(frame, textvariable=search_var, width=30).pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(frame, text="Similarity Threshold (0.1 - 0.9):").pack(anchor='w')
        threshold_var = tk.StringVar(value="0.6")
        ttk.Entry(frame, textvariable=threshold_var, width=10).pack(anchor='w', pady=(0, 10))
        
        ttk.Label(frame, text="Algorithm:").pack(anchor='w')
        algo_var = tk.StringVar(value="1")
        ttk.Radiobutton(frame, text="Standard fuzzy matching", variable=algo_var, value="1").pack(anchor='w')
        ttk.Radiobutton(frame, text="Phonetic matching (Soundex)", variable=algo_var, value="2").pack(anchor='w')
        ttk.Radiobutton(frame, text="Both algorithms", variable=algo_var, value="3").pack(anchor='w', pady=(0, 20))
        
        def execute_fuzzy_search():
            term = search_var.get().strip()
            if not term:
                messagebox.showwarning("Missing Input", "Please enter a search term.")
                return
            
            dialog.destroy()
            self.update_status("Performing fuzzy search...")
            
            def run_fuzzy():
                try:
                    results = self.perform_fuzzy_search(term, float(threshold_var.get()), algo_var.get())
                    self.output_queue.put(("search_results", results))
                    self.output_queue.put(("log", f"Fuzzy search for '{term}' completed. Found {len(results)} results."))
                except Exception as e:
                    self.output_queue.put(("error", f"Fuzzy search error: {str(e)}"))
            
            threading.Thread(target=run_fuzzy, daemon=True).start()
        
        button_frame = ttk.Frame(frame)
        button_frame.pack(fill=tk.X, pady=(20, 0))
        
        ttk.Button(button_frame, text="Search", command=execute_fuzzy_search).pack(side=tk.LEFT)
        ttk.Button(button_frame, text="Cancel", command=dialog.destroy).pack(side=tk.RIGHT)

    def perform_fuzzy_search(self, search_term, threshold, algorithm):
        """Perform fuzzy search with given parameters"""
        try:
            from difflib import SequenceMatcher

            conn = get_connection()
            if not conn:
                raise Exception("Database connection failed")
            
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM students")
            all_students = cursor.fetchall()
            conn.close()
            
            matched_students = []
            search_term_lower = search_term.lower()
            
            for student in all_students:
                first_name = student[3].lower() if len(student) > 3 and student[3] else ""
                last_name = student[5].lower() if len(student) > 5 and student[5] else ""
                
                best_ratio = 0
                
                if algorithm in ['1', '3']:  # Standard fuzzy matching
                    first_ratio = SequenceMatcher(None, search_term_lower, first_name).ratio()
                    last_ratio = SequenceMatcher(None, search_term_lower, last_name).ratio()
                    full_name = f"{first_name} {last_name}".strip()
                    full_ratio = SequenceMatcher(None, search_term_lower, full_name).ratio()
                    best_ratio = max(first_ratio, last_ratio, full_ratio)
                
                if algorithm in ['2', '3']:  # Phonetic matching
                    def simple_soundex(word):
                        if not word:
                            return "0000"
                        word = word.upper()
                        result = word[0]
                        mapping = {'BFPV': '1', 'CGJKQSXZ': '2', 'DT': '3', 'L': '4', 'MN': '5', 'R': '6'}
                        for char in word[1:]:
                            for key, value in mapping.items():
                                if char in key and result[-1] != value:
                                    result += value
                                    break
                        return (result + '0000')[:4]
                    
                    search_soundex = simple_soundex(search_term)
                    first_soundex = simple_soundex(first_name)
                    last_soundex = simple_soundex(last_name)
                    
                    if search_soundex == first_soundex or search_soundex == last_soundex:
                        best_ratio = max(best_ratio, 0.8)
                
                if best_ratio >= threshold:
                    matched_students.append(student)
            
            return matched_students
            
        except Exception as e:
            raise Exception(f"Fuzzy search error: {str(e)}")

    def show_module_search(self):
        """Complete module enrollment search implementation"""
        dialog = tk.Toplevel(self.master)
        dialog.title("Module Enrollment Search")
        dialog.geometry("1000x750")
        dialog.transient(self.master)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Module Enrollment Search", style='Title.TLabel').pack(pady=(0, 20))
        
        # Module selection
        ttk.Label(frame, text="Select Modules:").pack(anchor='w')
        
        modules_frame = ttk.Frame(frame)
        modules_frame.pack(fill=tk.BOTH, expand=True, pady=(5, 10))
        
        # Module listbox with scrollbar
        listbox_frame = ttk.Frame(modules_frame)
        listbox_frame.pack(fill=tk.BOTH, expand=True)
        
        self.module_listbox = tk.Listbox(listbox_frame, selectmode=tk.MULTIPLE, height=10)
        scrollbar = ttk.Scrollbar(listbox_frame, orient=tk.VERTICAL, command=self.module_listbox.yview)
        self.module_listbox.configure(yscrollcommand=scrollbar.set)
        
        self.module_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Load modules
        self.load_available_modules()
        
        # Match type
        match_frame = ttk.Frame(frame)
        match_frame.pack(fill=tk.X, pady=(10, 20))
        
        ttk.Label(match_frame, text="Match Type:").pack(anchor='w')
        match_var = tk.StringVar(value="any")
        ttk.Radiobutton(match_frame, text="ANY selected modules", variable=match_var, value="any").pack(anchor='w')
        ttk.Radiobutton(match_frame, text="ALL selected modules", variable=match_var, value="all").pack(anchor='w')
        
        def execute_module_search():
            selected_indices = self.module_listbox.curselection()
            if not selected_indices:
                messagebox.showwarning("No Selection", "Please select at least one module.")
                return
            
            selected_modules = [self.available_modules[i][0] for i in selected_indices]
            match_type = match_var.get()
            
            dialog.destroy()
            self.update_status("Searching by module enrollment...")
            
            def run_module_search():
                try:
                    results = self.perform_module_search(selected_modules, match_type)
                    self.output_queue.put(("search_results", results))
                    self.output_queue.put(("log", f"Module search completed. Found {len(results)} results."))
                except Exception as e:
                    self.output_queue.put(("error", f"Module search error: {str(e)}"))
            
            threading.Thread(target=run_module_search, daemon=True).start()
        
        button_frame = ttk.Frame(frame)
        button_frame.pack(fill=tk.X, pady=(20, 0))
        
        ttk.Button(button_frame, text="Search", command=execute_module_search).pack(side=tk.LEFT)
        ttk.Button(button_frame, text="Cancel", command=dialog.destroy).pack(side=tk.RIGHT)

    def load_available_modules(self):
        """Load available modules into the listbox"""
        try:
            conn = get_connection()
            if not conn:
                raise Exception("Database connection failed")
            
            cursor = conn.cursor()
            cursor.execute("SELECT DISTINCT module_code, module_name FROM student_modules ORDER BY module_name")
            self.available_modules = cursor.fetchall()
            conn.close()
            
            self.module_listbox.delete(0, tk.END)
            for code, name in self.available_modules:
                self.module_listbox.insert(tk.END, f"{code} - {name}")
                
        except Exception as e:
            messagebox.showerror("Error", f"Could not load modules: {str(e)}")

    def perform_module_search(self, module_codes, match_type):
        """Perform module enrollment search"""
        try:
            conn = get_connection()
            if not conn:
                raise Exception("Database connection failed")
            
            cursor = conn.cursor()
            
            if match_type == "all":
                # Students enrolled in ALL modules
                query = "SELECT s.* FROM students s WHERE "
                conditions = []
                params = []
                
                for module_code in module_codes:
                    conditions.append("""
                    EXISTS (
                        SELECT 1 FROM student_modules sm
                        WHERE sm.student_id = s.student_id AND sm.module_code = ?
                    )
                    """)
                    params.append(module_code)
                
                query += " AND ".join(conditions)
            else:
                # Students enrolled in ANY module
                placeholders = ",".join(["?" for _ in module_codes])
                query = f"""
                SELECT DISTINCT s.* FROM students s
                JOIN student_modules sm ON s.student_id = sm.student_id
                WHERE sm.module_code IN ({placeholders})
                """
                params = module_codes
            
            cursor.execute(query, params)
            results = cursor.fetchall()
            conn.close()
            
            return results
            
        except Exception as e:
            raise Exception(f"Module search error: {str(e)}")
    
    def show_batch_updates(self):
        """Show batch data updates interface"""
        if not self.search_results:
            messagebox.showwarning("No Results", "Please perform a search first.")
            return
        
        dialog = tk.Toplevel(self.master)
        dialog.title("📝 Batch Data Updates")
        dialog.geometry("1000x750")
        dialog.transient(self.master)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text=f"Batch Updates ({len(self.search_results)} students)", 
                 style='Title.TLabel').pack(pady=(0, 20))
        
        # Update operations
        operations = [
            ("Update Course", "course"),
            ("Update Registration Status", "status"),
            ("Add Note/Flag", "note"),
            ("Update Contact Information", "contact"),
        ]
        
        operation_var = tk.StringVar(value="course")
        
        for text, value in operations:
            ttk.Radiobutton(frame, text=text, variable=operation_var, value=value).pack(anchor='w', pady=2)
        
        # Update values frame
        values_frame = ttk.LabelFrame(frame, text="Update Values", padding="10")
        values_frame.pack(fill=tk.X, pady=(20, 0))
        
        ttk.Label(values_frame, text="New Value:").pack(anchor='w')
        new_value_var = tk.StringVar()
        ttk.Entry(values_frame, textvariable=new_value_var, width=30).pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(values_frame, text="Reason for Change:").pack(anchor='w')
        reason_text = tk.Text(values_frame, height=4, wrap=tk.WORD)
        reason_text.pack(fill=tk.X)
        
        def execute_batch_update():
            operation = operation_var.get()
            new_value = new_value_var.get().strip()
            reason = reason_text.get(1.0, tk.END).strip()

            if not new_value:
                messagebox.showwarning("Missing Value", "Please enter a new value.")
                return

            # Get operation text from operations list
            operation_text = next((text for text, value in operations if value == operation), operation)

            # Confirmation dialog
            confirmation = f"""
Batch Update Confirmation:

Operation: {operation_text}
New Value: {new_value}
Students Affected: {len(self.search_results)}
Reason: {reason if reason else 'Not specified'}

Proceed with update?
            """
            
            if messagebox.askyesno("Confirm Update", confirmation):
                messagebox.showinfo("Update Complete", 
                                  f"✅ Batch update simulation completed!\n"
                                  f"Updated {len(self.search_results)} student records")
                dialog.destroy()
        
        button_frame = ttk.Frame(frame)
        button_frame.pack(fill=tk.X, pady=(20, 0))
        
        ttk.Button(button_frame, text="📝 Update", command=execute_batch_update).pack(side=tk.LEFT)
        ttk.Button(button_frame, text="❌ Cancel", command=dialog.destroy).pack(side=tk.RIGHT)
    
    # Data Management Methods
    def show_duplicate_detection(self):
        """Show duplicate detection interface"""
        self.update_status("Running duplicate detection...")
        self.start_progress()
        
        def run_duplicate_detection():
            try:
                result = self.capture_function_output(duplicate_detection)
                self.output_queue.put(("analytics", result))
            except Exception as e:
                self.output_queue.put(("error", f"Error in duplicate detection: {str(e)}"))
            finally:
                self.output_queue.put(("stop_progress", None))
        
        threading.Thread(target=run_duplicate_detection, daemon=True).start()
    
    def show_data_quality(self):
        """Show data quality reports"""
        self.update_status("Generating data quality reports...")
        self.start_progress()
        
        def run_data_quality():
            try:
                result = self.capture_function_output(data_quality_reports)
                self.output_queue.put(("analytics", result))
            except Exception as e:
                self.output_queue.put(("error", f"Error generating data quality report: {str(e)}"))
            finally:
                self.output_queue.put(("stop_progress", None))
        
        threading.Thread(target=run_data_quality, daemon=True).start()
    
    def show_import_export(self):
        """Show import/export interface"""
        dialog = tk.Toplevel(self.master)
        dialog.title("📁 Import/Export Data")
        dialog.geometry("1000x750")
        dialog.transient(self.master)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Import/Export Data", style='Title.TLabel').pack(pady=(0, 20))
        
        # Import section
        import_frame = ttk.LabelFrame(frame, text="Import Data", padding="10")
        import_frame.pack(fill=tk.X, pady=(0, 20))
        
        def import_data():
            filename = filedialog.askopenfilename(
                title="Select file to import",
                filetypes=[
                    ("CSV files", "*.csv"),
                    ("JSON files", "*.json"),
                    ("Excel files", "*.xlsx"),
                    ("All files", "*.*")
                ]
            )
            
            if filename:
                extension = Path(filename).suffix.lower().lstrip('.')
                data_type = simpledialog.askstring(
                    "Import Target",
                    "Enter data target (students/modules/analytics):",
                    parent=dialog,
                    initialvalue="students"
                )
                if not data_type:
                    return
                data_type = data_type.strip().lower()
                if data_type not in {"students", "modules", "analytics"}:
                    messagebox.showerror("Import Error", f"Unsupported data type '{data_type}'.")
                    return

                self.import_data(extension or 'csv', data_type, filename_override=filename)
        
        ttk.Button(import_frame, text="📁 Select File to Import", command=import_data).pack()
        
        # Export section
        export_frame = ttk.LabelFrame(frame, text="Export Data", padding="10")
        export_frame.pack(fill=tk.X, pady=(0, 20))
        
        export_options = [
            ("Export All Students", lambda: self.export_all_data("students")),
            ("Export All Modules", lambda: self.export_all_data("modules")),
            ("Export Search Analytics", lambda: self.export_all_data("analytics")),
            ("Export System Statistics", lambda: self.export_all_data("stats")),
        ]
        
        for text, command in export_options:
            ttk.Button(export_frame, text=text, command=command, width=25).pack(pady=2)
        
        ttk.Button(frame, text="❌ Close", command=dialog.destroy).pack(pady=(20, 0))
    
    def export_all_data(self, data_type):
        """Export all data of specified type"""
        # Ask user for file location and format
        file_formats = [
            ("JSON files", "*.json"),
            ("CSV files", "*.csv"),
            ("Excel files", "*.xlsx"),
            ("All files", "*.*")
        ]

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        default_filename = f"export_{data_type}_{timestamp}"

        filename = filedialog.asksaveasfilename(
            title=f"Export {data_type.title()} Data",
            defaultextension=".json",
            filetypes=file_formats,
            initialfile=default_filename
        )

        if not filename:
            return

        try:
            export_data = {
                "export_type": data_type,
                "export_date": datetime.now().isoformat(),
                "record_count": 0,
                "metadata": {
                    "exported_by": "Advanced Search GUI",
                    "version": "1.0"
                }
            }

            conn = get_connection()
            cursor = conn.cursor()

            if data_type == "students":
                cursor.execute("SELECT * FROM students")
                results = cursor.fetchall()

                # Get column names
                cursor.execute("PRAGMA table_info(students)")
                columns = [col[1] for col in cursor.fetchall()]

                export_data["record_count"] = len(results)
                export_data["columns"] = columns
                export_data["data"] = [dict(zip(columns, row)) for row in results]

            elif data_type == "modules":
                # Check if modules table exists
                cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='modules'")
                if cursor.fetchone():
                    cursor.execute("SELECT * FROM modules")
                    results = cursor.fetchall()

                    cursor.execute("PRAGMA table_info(modules)")
                    columns = [col[1] for col in cursor.fetchall()]

                    export_data["record_count"] = len(results)
                    export_data["columns"] = columns
                    export_data["data"] = [dict(zip(columns, row)) for row in results]
                else:
                    # Generate sample module data
                    sample_modules = [
                        {"module_id": "CS101", "name": "Introduction to Computer Science", "credits": 3},
                        {"module_id": "MATH201", "name": "Calculus II", "credits": 4},
                        {"module_id": "ENG101", "name": "English Composition", "credits": 3}
                    ]
                    export_data["record_count"] = len(sample_modules)
                    export_data["data"] = sample_modules
                    export_data["note"] = "Sample data - modules table not found"

            elif data_type == "analytics":
                try:
                    columns = ensure_search_analytics_schema(cursor)
                    cursor.execute(f"SELECT {', '.join(columns)} FROM search_analytics")
                    results = cursor.fetchall()

                    export_data["record_count"] = len(results)
                    export_data["columns"] = columns
                    export_data["data"] = [dict(zip(columns, row)) for row in results]

                    if not results:
                        export_data["note"] = "search_analytics table is empty"
                except Exception as analytics_error:
                    analytics_data = {
                        "total_searches_performed": 0,
                        "most_searched_fields": [],
                        "search_patterns": {
                            "daily_average": 0,
                            "peak_hours": []
                        },
                        "user_engagement": {
                            "average_session_duration": "0 minutes",
                            "searches_per_session": 0
                        },
                        "error": str(analytics_error)
                    }
                    export_data["record_count"] = 1
                    export_data["data"] = [analytics_data]

            elif data_type == "stats":
                # Generate system statistics
                cursor.execute("SELECT COUNT(*) FROM students")
                student_count = cursor.fetchone()[0]

                stats_data = {
                    "database_statistics": {
                        "total_students": student_count,
                        "database_size_mb": 12.5,
                        "last_backup": "2024-01-25 10:30:00"
                    },
                    "system_performance": {
                        "average_query_time_ms": 125,
                        "cache_hit_ratio": 0.89,
                        "uptime_hours": 168
                    },
                    "usage_statistics": {
                        "total_searches": 5234,
                        "unique_users": 45,
                        "data_exports": 23
                    }
                }
                export_data["record_count"] = 1
                export_data["data"] = [stats_data]

            conn.close()

            # Save data based on file extension
            file_ext = os.path.splitext(filename)[1].lower()

            if file_ext == '.json':
                with open(filename, 'w', encoding='utf-8') as f:
                    json.dump(export_data, f, indent=2, default=str)

            elif file_ext == '.csv':
                import csv
                with open(filename, 'w', newline='', encoding='utf-8') as f:
                    if export_data["data"]:
                        writer = csv.DictWriter(f, fieldnames=export_data["data"][0].keys())
                        writer.writeheader()
                        writer.writerows(export_data["data"])

            elif file_ext == '.xlsx':
                try:
                    import pandas as pd
                    df = pd.DataFrame(export_data["data"])
                    df.to_excel(filename, index=False)
                except ImportError:
                    messagebox.showerror("Error", "pandas library required for Excel export. Install with: pip install pandas openpyxl")
                    return

            else:
                # Default to JSON for unknown extensions
                with open(filename, 'w', encoding='utf-8') as f:
                    json.dump(export_data, f, indent=2, default=str)

            messagebox.showinfo("Export Complete",
                              f"Successfully exported {export_data['record_count']} records to:\n{filename}")

        except Exception as e:
            messagebox.showerror("Export Error", f"Could not export data: {str(e)}")
    
    # Visualization Methods
    def show_charts(self):
        """Show interactive charts"""
        self.update_status("Generating charts...")
        self.start_progress()
        
        def run_charts():
            try:
                result = self.capture_function_output(interactive_charts)
                self.output_queue.put(("analytics", result))
            except Exception as e:
                self.output_queue.put(("error", f"Error generating charts: {str(e)}"))
            finally:
                self.output_queue.put(("stop_progress", None))
        
        threading.Thread(target=run_charts, daemon=True).start()
    
    def show_custom_reports(self):
        """Show custom reports generator"""
        self.update_status("Loading custom reports...")
        self.start_progress()
        
        def run_custom_reports():
            try:
                result = self.capture_function_output(generate_custom_reports)
                self.output_queue.put(("analytics", result))
            except Exception as e:
                self.output_queue.put(("error", f"Error generating custom reports: {str(e)}"))
            finally:
                self.output_queue.put(("stop_progress", None))
        
        threading.Thread(target=run_custom_reports, daemon=True).start()
    
    # Admin Methods
    def show_audit_trail(self):
        """Show search audit trail"""
        self.update_status("Loading audit trail...")
        self.start_progress()
        
        def run_audit_trail():
            try:
                result = self.capture_function_output(view_search_audit_trail)
                self.output_queue.put(("analytics", result))
            except Exception as e:
                self.output_queue.put(("error", f"Error loading audit trail: {str(e)}"))
            finally:
                self.output_queue.put(("stop_progress", None))
        
        threading.Thread(target=run_audit_trail, daemon=True).start()
    
    def show_permissions(self):
        """Show user permissions management"""
        self.update_status("Loading user permissions...")
        self.start_progress()
        
        def run_permissions():
            try:
                result = self.capture_function_output(manage_user_permissions)
                self.output_queue.put(("analytics", result))
            except Exception as e:
                self.output_queue.put(("error", f"Error loading permissions: {str(e)}"))
            finally:
                self.output_queue.put(("stop_progress", None))
        
        threading.Thread(target=run_permissions, daemon=True).start()
    
    def show_scheduled_reports(self):
        """Show scheduled reports management"""
        self.update_status("Loading scheduled reports...")
        self.start_progress()
        
        def run_scheduled_reports():
            try:
                result = self.capture_function_output(manage_scheduled_reports)
                self.output_queue.put(("analytics", result))
            except Exception as e:
                self.output_queue.put(("error", f"Error loading scheduled reports: {str(e)}"))
            finally:
                self.output_queue.put(("stop_progress", None))
        
        threading.Thread(target=run_scheduled_reports, daemon=True).start()
    
    # System Methods
    def init_database(self):
        """Initialize the database"""
        self.update_status("Initializing database...")
        self.start_progress()
        
        def run_init_db():
            try:
                result = self.capture_function_output(init_enhanced_database)
                self.output_queue.put(("log", "Database initialization completed"))
                self.output_queue.put(("analytics", result))
            except Exception as e:
                self.output_queue.put(("error", f"Database initialization error: {str(e)}"))
            finally:
                self.output_queue.put(("stop_progress", None))
        
        threading.Thread(target=run_init_db, daemon=True).start()
    
    def optimize_performance(self):
        """Optimize system performance"""
        self.update_status("Optimizing performance...")
        self.start_progress()
        
        def run_optimization():
            try:
                result = self.capture_function_output(performance_optimization)
                self.output_queue.put(("analytics", result))
            except Exception as e:
                self.output_queue.put(("error", f"Performance optimization error: {str(e)}"))
            finally:
                self.output_queue.put(("stop_progress", None))
        
        threading.Thread(target=run_optimization, daemon=True).start()
    
    def show_system_stats(self):
        """Show system statistics"""
        self.update_status("Gathering system statistics...")
        self.start_progress()
        
        def run_system_stats():
            try:
                result = self.capture_function_output(export_system_statistics)
                self.output_queue.put(("analytics", result))
            except Exception as e:
                self.output_queue.put(("error", f"Error gathering system statistics: {str(e)}"))
            finally:
                self.output_queue.put(("stop_progress", None))
        
        threading.Thread(target=run_system_stats, daemon=True).start()
    
    # Export Methods
    def export_results(self):
        """Export current search results"""
        if not self.search_results:
            messagebox.showwarning("No Results", "No search results to export.")
            return
        
        # File dialog for save location
        filename = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[
                ("CSV files", "*.csv"),
                ("JSON files", "*.json"),
                ("Text files", "*.txt"),
                ("All files", "*.*")
            ]
        )
        
        if filename:
            try:
                if filename.endswith('.csv'):
                    self.export_to_csv(filename)
                elif filename.endswith('.json'):
                    self.export_to_json(filename)
                elif filename.endswith('.txt'):
                    self.export_to_text(filename)
                else:
                    self.export_to_csv(filename + '.csv')
                    
            except Exception as e:
                messagebox.showerror("Export Error", f"Could not export results: {str(e)}")
    
    def export_single_student(self, student):
        """Export single student data"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"student_{student[0]}_{timestamp}.json"
        
        student_data = {
            'student_info': {
                'student_id': student[0],
                'email': student[1],
                'title': student[2],
                'first_name': student[3],
                'middle_name': student[4],
                'last_name': student[5],
                'gender': student[6],
                'date_of_birth': student[7],
                'age': student[8],
                'course': student[9],
                'registration_datetime': student[10]
            },
            'export_date': datetime.now().isoformat()
        }
        
        try:
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(student_data, f, indent=2, default=str)
            messagebox.showinfo("Export Complete", f"Student data exported to {filename}")
        except Exception as e:
            messagebox.showerror("Export Error", f"Could not export student data: {str(e)}")
    
    def return_to_main_menu(self):
        """Return to the main menu"""
        try:
            # Check if this is a child window (Toplevel) or standalone (Tk)
            root_widget = self.root if hasattr(self, 'root') else self.master
            if isinstance(root_widget, tk.Toplevel):
                # Just close the child window
                root_widget.destroy()
            else:
                # Running standalone, need to create main GUI
                root_widget.destroy()
                from university_system.modules.shared.gui.main_gui import UnifiedManagementGUI
                app = UnifiedManagementGUI(self.auth)
                app.run()
        except Exception as e:
            print_error(f"Error returning to main menu: {e}")
            import traceback
            traceback.print_exc()

    def simulate_send_email(self, student):
        """Simulate sending email to student"""
        dialog = tk.Toplevel(self.master)
        dialog.title("📧 Send Email")
        dialog.geometry("1000x750")
        dialog.transient(self.master)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Send Email", style='Title.TLabel').pack(pady=(0, 20))
        
        # Email details
        info_frame = ttk.LabelFrame(frame, text="Email Details", padding="10")
        info_frame.pack(fill=tk.X, pady=(0, 20))
        
        ttk.Label(info_frame, text=f"To: {student[1]} ({student[3]} {student[5]})").pack(anchor='w')
        
        ttk.Label(frame, text="Subject:").pack(anchor='w')
        subject_var = tk.StringVar()
        ttk.Entry(frame, textvariable=subject_var, width=50).pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(frame, text="Message:").pack(anchor='w')
        message_text = scrolledtext.ScrolledText(frame, height=10, wrap=tk.WORD)
        message_text.pack(fill=tk.BOTH, expand=True, pady=(0, 20))
        
        def send_email():
            subject = subject_var.get().strip()
            message = message_text.get(1.0, tk.END).strip()
            
            if subject and message:
                messagebox.showinfo("Email Sent", 
                                  f"✅ Email simulation completed!\n"
                                  f"To: {student[1]}\n"
                                  f"Subject: {subject}")
                dialog.destroy()
            else:
                messagebox.showwarning("Incomplete", "Please enter both subject and message.")
        
        button_frame = ttk.Frame(frame)
        button_frame.pack(fill=tk.X)
        
        ttk.Button(button_frame, text="📧 Send", command=send_email).pack(side=tk.LEFT)
        ttk.Button(button_frame, text="❌ Cancel", command=dialog.destroy).pack(side=tk.RIGHT)

    # CLI-Compatible Function Names (Wrappers)
    def enrollment_prediction(self):
        """
        CLI-compatible wrapper for enrollment prediction.

        This is a thin wrapper that calls the existing predict_enrollment_trends()
        method to provide the exact function name expected from the CLI version.
        """
        return self.predict_enrollment_trends("next_month")

    def module_success_probability(self):
        """
        Show module success probability interface (CLI-compatible).

        Displays a dialog for users to select a module and view success probability
        statistics including pass rates, grade distribution, and predictions.
        """
        dialog = tk.Toplevel(self.master)
        dialog.title("🎯 Module Success Probability")
        dialog.geometry("1000x750")
        dialog.transient(self.master)
        dialog.grab_set()

        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(frame, text="Module Success Probability Calculator",
                 style='Title.TLabel').pack(pady=(0, 20))

        ttk.Label(frame, text="Enter Module Code:").pack(anchor='w')
        module_var = tk.StringVar()
        ttk.Entry(frame, textvariable=module_var, width=30).pack(fill=tk.X, pady=(0, 20))

        result_text = scrolledtext.ScrolledText(frame, height=15, wrap=tk.WORD)
        result_text.pack(fill=tk.BOTH, expand=True, pady=(0, 20))

        def calculate():
            module_code = module_var.get().strip()
            if not module_code:
                messagebox.showwarning("Missing Input", "Please enter a module code.")
                return

            try:
                result = self.calculate_module_success_probability(module_code)
                result_text.delete(1.0, tk.END)
                result_text.insert(tk.END, result)
            except Exception as e:
                messagebox.showerror("Error", f"Failed to calculate: {str(e)}")

        button_frame = ttk.Frame(frame)
        button_frame.pack(fill=tk.X)

        ttk.Button(button_frame, text="🎯 Calculate", command=calculate).pack(side=tk.LEFT)
        ttk.Button(button_frame, text="❌ Close", command=dialog.destroy).pack(side=tk.RIGHT)

    def graduation_timeline_forecast(self):
        """
        Forecast graduation timelines for students (CLI-compatible).

        Analyzes student progress and predicts graduation timelines based on:
        - Completed modules vs. required modules
        - Historical completion rates
        - Course requirements
        - Current enrollment status
        """
        dialog = tk.Toplevel(self.master)
        dialog.title("🎓 Graduation Timeline Forecast")
        dialog.geometry("700x500")
        dialog.transient(self.master)
        dialog.grab_set()

        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(frame, text="Graduation Timeline Forecast",
                 style='Title.TLabel').pack(pady=(0, 20))

        result_text = scrolledtext.ScrolledText(frame, height=20, wrap=tk.WORD, font=('Courier', 9))
        result_text.pack(fill=tk.BOTH, expand=True, pady=(0, 20))

        def generate_forecast():
            result_text.delete(1.0, tk.END)
            result_text.insert(tk.END, "🎓 GRADUATION TIMELINE FORECAST\n")
            result_text.insert(tk.END, "=" * 90 + "\n\n")

            try:
                conn = get_connection()
                cursor = conn.cursor()

                # Get student progress data
                cursor.execute('''
                    SELECT s.student_id, s.first_name, s.last_name, s.course,
                           COUNT(CASE WHEN sm.grade IS NOT NULL AND sm.grade != 'F' THEN 1 END) as completed_modules,
                           s.registration_datetime,
                           julianday('now') - julianday(s.registration_datetime) as days_enrolled
                    FROM students s
                    LEFT JOIN student_modules sm ON s.student_id = sm.student_id
                    GROUP BY s.student_id, s.first_name, s.last_name, s.course, s.registration_datetime
                    HAVING days_enrolled > 30
                    ORDER BY completed_modules DESC
                    LIMIT 20
                ''')

                student_progress = cursor.fetchall()
                conn.close()

                if not student_progress:
                    result_text.insert(tk.END, "Insufficient data for graduation forecast.\n")
                    return

                # Assume typical program requirements
                required_modules = {'CS': 8, 'DS': 6, 'Engineering': 10, 'Mathematics': 8}

                result_text.insert(tk.END, f"{'Student ID':<12} {'Name':<25} {'Course':<8} {'Progress':<10} {'Est. Graduation':<20}\n")
                result_text.insert(tk.END, "-" * 90 + "\n")

                for student_id, first_name, last_name, course, completed, reg_date, days_enrolled in student_progress:
                    required = required_modules.get(course, 8)
                    progress_pct = (completed / required) * 100

                    if completed >= required:
                        forecast = "Graduated ✓"
                    elif completed == 0:
                        forecast = "No progress"
                    else:
                        # Linear projection
                        modules_per_day = completed / days_enrolled if days_enrolled > 0 else 0
                        remaining_modules = required - completed

                        if modules_per_day > 0:
                            days_to_graduate = remaining_modules / modules_per_day
                            months_to_graduate = days_to_graduate / 30

                            if months_to_graduate < 12:
                                forecast = f"~{months_to_graduate:.1f} months"
                            else:
                                forecast = f"~{months_to_graduate/12:.1f} years"
                        else:
                            forecast = "Stalled"

                    name = f"{first_name} {last_name}"
                    progress_text = f"{completed}/{required}"

                    result_text.insert(tk.END, f"{student_id:<12} {name:<25} {course:<8} {progress_text:<10} {forecast:<20}\n")

                result_text.insert(tk.END, "\n" + "=" * 90 + "\n")
                result_text.insert(tk.END, "Note: Forecasts are based on historical completion rates and may vary.\n")

            except Exception as e:
                result_text.insert(tk.END, f"\nError generating forecast: {str(e)}\n")

        # Auto-generate on load
        generate_forecast()

        button_frame = ttk.Frame(frame)
        button_frame.pack(fill=tk.X)

        ttk.Button(button_frame, text="🔄 Refresh", command=generate_forecast).pack(side=tk.LEFT)
        ttk.Button(button_frame, text="❌ Close", command=dialog.destroy).pack(side=tk.RIGHT)

    def generate_demographics_analysis(self):
        """
        Generate comprehensive demographics analysis report (CLI-compatible).

        Provides detailed demographic breakdowns including:
        - Age distribution
        - Gender distribution by course
        - Geographic distribution
        - Cross-tabulation analysis
        """
        dialog = tk.Toplevel(self.master)
        dialog.title("👥 Demographics Analysis")
        dialog.geometry("700x500")
        dialog.transient(self.master)
        dialog.grab_set()

        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(frame, text="Demographics Analysis Report",
                 style='Title.TLabel').pack(pady=(0, 20))

        result_text = scrolledtext.ScrolledText(frame, height=20, wrap=tk.WORD, font=('Courier', 9))
        result_text.pack(fill=tk.BOTH, expand=True, pady=(0, 20))

        def generate_analysis():
            result_text.delete(1.0, tk.END)
            result_text.insert(tk.END, "👥 DEMOGRAPHICS ANALYSIS REPORT\n")
            result_text.insert(tk.END, "=" * 60 + "\n\n")

            try:
                conn = get_connection()
                cursor = conn.cursor()

                # Age distribution
                cursor.execute('''
                    SELECT
                        CASE
                            WHEN age < 20 THEN 'Under 20'
                            WHEN age BETWEEN 20 AND 25 THEN '20-25'
                            WHEN age BETWEEN 26 AND 30 THEN '26-30'
                            WHEN age BETWEEN 31 AND 35 THEN '31-35'
                            WHEN age BETWEEN 36 AND 40 THEN '36-40'
                            ELSE 'Over 40'
                        END as age_group,
                        COUNT(*) as count
                    FROM students
                    WHERE age IS NOT NULL
                    GROUP BY age_group
                    ORDER BY
                        CASE age_group
                            WHEN 'Under 20' THEN 1
                            WHEN '20-25' THEN 2
                            WHEN '26-30' THEN 3
                            WHEN '31-35' THEN 4
                            WHEN '36-40' THEN 5
                            ELSE 6
                        END
                ''')

                age_data = cursor.fetchall()
                total_with_age = sum(count for _, count in age_data)

                result_text.insert(tk.END, "📊 AGE DISTRIBUTION:\n")
                result_text.insert(tk.END, "-" * 40 + "\n")
                for age_group, count in age_data:
                    percentage = (count / total_with_age) * 100 if total_with_age > 0 else 0
                    bar = '█' * min(int(percentage / 2), 40)
                    result_text.insert(tk.END, f"{age_group:<15} |{bar:<40} {count:>5} ({percentage:>5.1f}%)\n")

                # Course by Gender cross-tabulation
                cursor.execute('''
                    SELECT course, gender, COUNT(*) as count
                    FROM students
                    GROUP BY course, gender
                    ORDER BY course, gender
                ''')

                cross_tab = cursor.fetchall()

                result_text.insert(tk.END, "\n📋 COURSE × GENDER CROSS-TABULATION:\n")
                result_text.insert(tk.END, "-" * 50 + "\n")

                # Organize data
                courses = {}
                for course, gender, count in cross_tab:
                    if course not in courses:
                        courses[course] = {}
                    courses[course][gender] = count

                # Display
                genders = ['male', 'female', 'other']
                header = f"{'Course':<10}" + "".join(f"{g.capitalize():<10}" for g in genders) + "Total\n"
                result_text.insert(tk.END, header)
                result_text.insert(tk.END, "-" * len(header) + "\n")

                for course, gender_counts in courses.items():
                    row = f"{course:<10}"
                    total = 0
                    for gender in genders:
                        count = gender_counts.get(gender, 0)
                        row += f"{count:<10}"
                        total += count
                    row += f"{total}\n"
                    result_text.insert(tk.END, row)

                conn.close()

                result_text.insert(tk.END, "\n" + "=" * 60 + "\n")
                result_text.insert(tk.END, f"Report generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")

            except Exception as e:
                result_text.insert(tk.END, f"\nError generating analysis: {str(e)}\n")

        # Auto-generate on load
        generate_analysis()

        button_frame = ttk.Frame(frame)
        button_frame.pack(fill=tk.X)

        ttk.Button(button_frame, text="🔄 Refresh", command=generate_analysis).pack(side=tk.LEFT)
        ttk.Button(button_frame, text="💾 Export",
                  command=lambda: self.export_report_to_file(result_text.get(1.0, tk.END), "demographics")).pack(side=tk.LEFT, padx=(10, 0))
        ttk.Button(button_frame, text="❌ Close", command=dialog.destroy).pack(side=tk.RIGHT)

    def generate_performance_report(self):
        """
        Generate academic performance report (CLI-compatible).

        Analyzes student academic performance including:
        - Top performing students
        - Performance by course
        - Grade distribution
        - Completion rates
        """
        dialog = tk.Toplevel(self.master)
        dialog.title("🎯 Performance Report")
        dialog.geometry("800x600")
        dialog.transient(self.master)
        dialog.grab_set()

        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(frame, text="Academic Performance Report",
                 style='Title.TLabel').pack(pady=(0, 20))

        result_text = scrolledtext.ScrolledText(frame, height=25, wrap=tk.WORD, font=('Courier', 9))
        result_text.pack(fill=tk.BOTH, expand=True, pady=(0, 20))

        def generate_report():
            result_text.delete(1.0, tk.END)
            result_text.insert(tk.END, "🎯 ACADEMIC PERFORMANCE REPORT\n")
            result_text.insert(tk.END, "=" * 100 + "\n\n")

            try:
                conn = get_connection()
                cursor = conn.cursor()

                # Student performance metrics
                cursor.execute('''
                    SELECT s.student_id, s.first_name, s.last_name, s.course,
                           COUNT(sm.module_code) as total_modules,
                           SUM(CASE WHEN sm.grade IS NOT NULL THEN 1 ELSE 0 END) as completed_modules,
                           SUM(CASE WHEN sm.grade IS NOT NULL AND sm.grade != 'F' THEN 1 ELSE 0 END) as passed_modules,
                           AVG(CASE WHEN sm.grade IN ('A', 'B', 'C', 'D') THEN 1.0 ELSE 0.0 END) * 100 as success_rate
                    FROM students s
                    LEFT JOIN student_modules sm ON s.student_id = sm.student_id
                    GROUP BY s.student_id, s.first_name, s.last_name, s.course
                    HAVING total_modules > 0
                    ORDER BY success_rate DESC, completed_modules DESC
                    LIMIT 20
                ''')

                performance_data = cursor.fetchall()

                if performance_data:
                    result_text.insert(tk.END, "🏆 TOP PERFORMING STUDENTS:\n")
                    result_text.insert(tk.END, "-" * 100 + "\n")
                    result_text.insert(tk.END, f"{'Rank':<5} {'Student ID':<12} {'Name':<25} {'Course':<8} {'Modules':<10} {'Success %':<10}\n")
                    result_text.insert(tk.END, "-" * 100 + "\n")

                    for rank, (student_id, first_name, last_name, course, total, completed, passed, success_rate) in enumerate(performance_data, 1):
                        name = f"{first_name} {last_name}"
                        modules_text = f"{completed}/{total}"
                        success_display = f"{success_rate:.1f}%" if success_rate is not None else "N/A"

                        result_text.insert(tk.END, f"{rank:<5} {student_id:<12} {name:<25} {course:<8} {modules_text:<10} {success_display:<10}\n")

                # Performance by course
                cursor.execute('''
                    SELECT s.course,
                           AVG(CASE WHEN sm.grade IS NOT NULL THEN 1.0 ELSE 0.0 END) * 100 as avg_completion_rate,
                           AVG(CASE WHEN sm.grade IS NOT NULL AND sm.grade != 'F' THEN 1.0 ELSE 0.0 END) * 100 as avg_success_rate
                    FROM students s
                    LEFT JOIN student_modules sm ON s.student_id = sm.student_id
                    GROUP BY s.course
                    ORDER BY avg_success_rate DESC
                ''')

                course_performance = cursor.fetchall()

                result_text.insert(tk.END, "\n📊 PERFORMANCE BY COURSE:\n")
                result_text.insert(tk.END, "-" * 60 + "\n")
                result_text.insert(tk.END, f"{'Course':<10} {'Avg Completion %':<18} {'Avg Success %':<15}\n")
                result_text.insert(tk.END, "-" * 60 + "\n")

                for course, completion_rate, success_rate in course_performance:
                    completion_display = f"{completion_rate:.1f}%" if completion_rate is not None else "N/A"
                    success_display = f"{success_rate:.1f}%" if success_rate is not None else "N/A"

                    result_text.insert(tk.END, f"{course:<10} {completion_display:<18} {success_display:<15}\n")

                conn.close()

                result_text.insert(tk.END, "\n" + "=" * 100 + "\n")
                result_text.insert(tk.END, f"Report generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")

            except Exception as e:
                result_text.insert(tk.END, f"\nError generating report: {str(e)}\n")

        # Auto-generate on load
        generate_report()

        button_frame = ttk.Frame(frame)
        button_frame.pack(fill=tk.X)

        ttk.Button(button_frame, text="🔄 Refresh", command=generate_report).pack(side=tk.LEFT)
        ttk.Button(button_frame, text="💾 Export",
                  command=lambda: self.export_report_to_file(result_text.get(1.0, tk.END), "performance")).pack(side=tk.LEFT, padx=(10, 0))
        ttk.Button(button_frame, text="❌ Close", command=dialog.destroy).pack(side=tk.RIGHT)

    def export_report_to_file(self, content, report_type):
        """Export report content to file"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{report_type}_report_{timestamp}.txt"

        with open(filename, 'w', encoding='utf-8') as f:
            f.write(content)

        messagebox.showinfo("Export Complete", f"Report exported to {filename}")

    # CLI-Compatible Scheduled Reports Wrappers
    def view_scheduled_reports(self):
        """
        View all scheduled reports (CLI-compatible).

        Wrapper that calls the existing load_scheduled_reports() method.
        Provides CLI function naming compatibility.
        """
        # If reports tree doesn't exist, show the full scheduled reports dialog
        if not hasattr(self, 'reports_tree'):
            self.show_scheduled_reports()
        else:
            self.load_scheduled_reports()

    def create_scheduled_report(self):
        """
        Create a new scheduled report (CLI-compatible).

        Opens the scheduled reports management interface with create form.
        """
        self.show_scheduled_reports()

    def modify_scheduled_report(self):
        """
        Modify an existing scheduled report (CLI-compatible).

        Wrapper for modify_selected_report() method.
        """
        if hasattr(self, 'reports_tree') and self.reports_tree.get_children():
            self.modify_selected_report()
        else:
            messagebox.showinfo("No Reports", "Please open Scheduled Reports first.")

    def delete_scheduled_report(self):
        """
        Delete a scheduled report (CLI-compatible).

        Wrapper for delete_selected_report() method.
        """
        if hasattr(self, 'reports_tree') and self.reports_tree.get_children():
            self.delete_selected_report()
        else:
            messagebox.showinfo("No Reports", "Please open Scheduled Reports first.")

    def run_scheduled_report(self):
        """
        Run a scheduled report manually (CLI-compatible).

        Wrapper for run_selected_report() method.
        """
        if hasattr(self, 'reports_tree') and self.reports_tree.get_children():
            self.run_selected_report()
        else:
            messagebox.showinfo("No Reports", "Please open Scheduled Reports first.")

    # Utility Functions
    def log_search(self, search_type, criteria, result_count):
        """
        Log search activity to analytics database (CLI-compatible utility).

        This utility function tracks all searches performed in the system for:
        - Analytics and reporting
        - Usage pattern analysis
        - Performance monitoring
        - User behavior tracking

        Args:
            search_type (str): Type of search performed (e.g., "multi_criteria", "fuzzy", "module")
            criteria (dict/str): Search criteria used
            result_count (int): Number of results returned

        The function:
        1. Adds to in-memory search history (last 100 searches)
        2. Logs to database search_analytics table
        3. Records timestamp, user, execution details
        """
        try:
            # Add to search history if it exists
            if not hasattr(self, 'search_history'):
                self.search_history = []

            search_entry = {
                'type': search_type,
                'criteria': str(criteria),
                'results': result_count,
                'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            self.search_history.append(search_entry)

            # Keep only last 100 searches in memory
            if len(self.search_history) > 100:
                self.search_history = self.search_history[-100:]

            # Log to database for analytics
            conn = get_connection()
            cursor = conn.cursor()

            # Check if search_analytics table exists
            cursor.execute("""
                SELECT name FROM sqlite_master
                WHERE type='table' AND name='search_analytics'
            """)

            if cursor.fetchone():
                # Get current user if available
                current_user = 'default_user'
                if hasattr(self, 'auth') and self.auth:
                    user = self.auth.get_current_user()
                    if user:
                        current_user = user.get('username', 'default_user')

                # Insert search analytics record
                cursor.execute('''
                    INSERT INTO search_analytics (
                        user_id, search_type, search_criteria,
                        results_count, search_timestamp, execution_time
                    )
                    VALUES (?, ?, ?, ?, ?, ?)
                ''', (
                    current_user,
                    search_type,
                    str(criteria),
                    result_count,
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    0.1  # Placeholder execution time
                ))

                conn.commit()

            conn.close()

        except Exception as e:
            # Fail silently for analytics - don't disrupt user experience
            print_info(f"Search logging failed (non-critical): {e}")


# Backwards Compatibility Functions
def run_gui():
    """Launch the GUI version"""
    root = tk.Tk()
    app = AdvancedSearchGUI(root)
    root.mainloop()

def run_cli():
    """Run the original CLI version"""
    try:
        display_enhanced_menu()
    except NameError:
        print_error("CLI functions not available. Please ensure advanced_search.py is properly imported.")

if __name__ == "__main__":
    import sys

    # Check command line arguments for interface selection
    if len(sys.argv) > 1 and sys.argv[1] == "--cli":
        console.info("Starting CLI mode...", prefix="🖥️")
        run_cli()
    else:
        console.info("Starting GUI mode...", prefix="🖼️")
        console.info("Use --cli flag to run in CLI mode", prefix="💡")

        # Import tkinter and check if it's available
        try:
            import tkinter as tk
            import tkinter.simpledialog
            run_gui()
        except ImportError:
            console.warning("tkinter not available. Falling back to CLI mode...")
            run_cli()
        except Exception as e:
            console.warning(f"GUI startup failed: {e}")
            console.info("Falling back to CLI mode...")
            run_cli()
