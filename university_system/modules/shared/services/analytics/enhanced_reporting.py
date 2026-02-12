from university_system.infrastructure.database.db import get_db_connection
from university_system.modules.shared.constants import paths
from university_system.modules.shared.utils.i18n import get_text, _
from university_system.infrastructure.database.db import sqlite3
from university_system.core.sql_safety import validate_table_name
import os
import csv
# The following third-party modules are optional. They may not be installed in
# all environments. Where unavailable, stubs are provided later in this file
# to ensure the module still imports successfully.

from datetime import datetime, timedelta
import json
# schedule is an optional dependency; fallback provided later if missing
import time
import threading
# reportlab imports will be attempted later with fallbacks
import hashlib
import logging
import secrets
# flask imports will be attempted later with fallbacks
# sklearn imports will be attempted later with fallbacks
import warnings

# -----------------------------------------------------------------------------
# Optional dependency handling
#
# Many of the analytics features in this module rely on third‑party
# libraries such as pandas, numpy, matplotlib, seaborn, plotly, schedule,
# reportlab, Flask and scikit‑learn. In lightweight environments where
# these packages are not installed, we still want to be able to import
# this module without errors. The following blocks attempt to import each
# optional dependency. If the import fails, a minimal shim is created so
# that references to these modules will not raise attribute errors. In
# these cases, functionality depending on the missing library will
# effectively become no‑ops.

# pandas
try:
    import pandas as pd  # type: ignore
except Exception as e:
    print(f"Optional dependency not installed: {e}")
    class _Dummy:
        def __getattr__(self, name):
            return self
        def __call__(self, *args, **kwargs):
            return None
        def __iter__(self):
            return iter([])
    pd = _Dummy()  # type: ignore

# numpy
try:
    import numpy as np  # type: ignore
except Exception as e:
    print(f"Optional dependency not installed: {e}")
    class _NumPyDummy:
        def __getattr__(self, name):
            def _no_op(*args, **kwargs):
                return None
            return _no_op
        def array(self, *args, **kwargs):
            return []
    np = _NumPyDummy()  # type: ignore

# matplotlib and seaborn
try:
    import matplotlib.pyplot as plt  # type: ignore
except Exception as e:
    print(f"Optional dependency not installed: {e}")
    class _MatPlotDummy:
        def __getattr__(self, name):
            def _no_op(*args, **kwargs):
                return None
            return _no_op
    plt = _MatPlotDummy()  # type: ignore
try:
    import seaborn as sns  # type: ignore
except Exception as e:
    print(f"Optional dependency not installed: {e}")
    class _SeabornDummy:
        def __getattr__(self, name):
            def _no_op(*args, **kwargs):
                return None
            return _no_op
    sns = _SeabornDummy()  # type: ignore

# plotly
try:
    import plotly.graph_objects as go  # type: ignore
    import plotly.express as px  # type: ignore
    import plotly.offline as pyo  # type: ignore
except Exception as e:
    print(f"Optional dependency not installed: {e}")
    class _PlotlyDummy:
        def __getattr__(self, name):
            def _no_op(*args, **kwargs):
                return None
            return _no_op
        def __call__(self, *args, **kwargs):
            return None
    go = _PlotlyDummy()  # type: ignore
    px = _PlotlyDummy()  # type: ignore
    pyo = _PlotlyDummy()  # type: ignore

# schedule
try:
    import schedule  # type: ignore
except Exception as e:
    print(f"Optional dependency not installed: {e}")

    class _ScheduleJobStub:
        """
        Stub implementation for schedule.Job when schedule library is not available.

        This stub allows the code to run without the schedule library installed,
        but scheduled tasks will not actually execute. All method calls return self
        to support method chaining, and job functions are stored but never executed.

        Attributes:
            _jobs (list): List of registered job functions
            _interval (int): Interval value for scheduling
            _unit (str): Time unit for scheduling (seconds, minutes, hours, etc.)
            _at_time (str): Specific time for scheduling
            _tag (str): Tag identifier for the job
        """

        def __init__(self, *args, **kwargs):
            """Initialize the stub with tracking attributes."""
            self._jobs = []
            self._interval = None
            self._unit = None
            self._at_time = None
            self._tag = None
            self._args = args
            self._kwargs = kwargs
            warnings.warn(
                "schedule library not installed. Scheduled tasks will not execute. "
                "Install with: pip install schedule",
                RuntimeWarning
            )

        def do(self, job_func, *args, **kwargs):
            """
            Store the job function but do not schedule it.

            Args:
                job_func: The function to be scheduled
                *args: Positional arguments for the job function
                **kwargs: Keyword arguments for the job function

            Returns:
                self: For method chaining
            """
            self._jobs.append({
                'function': job_func,
                'args': args,
                'kwargs': kwargs,
                'interval': self._interval,
                'unit': self._unit,
                'at_time': self._at_time,
                'tag': self._tag
            })
            logging.debug(
                f"Job '{job_func.__name__}' registered in stub (will not execute). "
                f"Interval: {self._interval} {self._unit}"
            )
            return self

        def at(self, time_str):
            """Set the time for job execution."""
            self._at_time = time_str
            return self

        def tag(self, *tags):
            """Add tags to the job."""
            self._tag = tags
            return self

        def until(self, *args, **kwargs):
            """Set end time for job execution."""
            return self

        def __getattr__(self, name):
            """
            Handle unknown attributes by setting unit and returning self.
            This supports schedule's fluent API like .seconds, .minutes, etc.
            """
            if name in ('second', 'seconds', 'minute', 'minutes', 'hour', 'hours',
                       'day', 'days', 'week', 'weeks', 'monday', 'tuesday',
                       'wednesday', 'thursday', 'friday', 'saturday', 'sunday'):
                self._unit = name
                return self
            return self

        def __call__(self, *args, **kwargs):
            """Support callable behavior for compatibility."""
            return self

    class schedule:  # type: ignore
        """
        Stub implementation for schedule module when not available.

        This stub provides the schedule module's API without actual scheduling
        functionality. It's used to prevent import errors when schedule is not installed.
        """

        _jobs = []  # Class-level job tracking

        @staticmethod
        def every(interval=1):
            """
            Create a new job stub with the specified interval.

            Args:
                interval (int): The interval value for scheduling

            Returns:
                _ScheduleJobStub: A job stub instance
            """
            job = _ScheduleJobStub()
            job._interval = interval
            schedule._jobs.append(job)
            return job

        @staticmethod
        def run_pending():
            """
            Stub for running pending jobs. Does nothing but logs a warning.

            In the real schedule library, this would execute any pending jobs.
            """
            if schedule._jobs:
                logging.debug(
                    f"run_pending() called on stub with {len(schedule._jobs)} registered jobs. "
                    "No jobs executed (schedule library not installed)."
                )
            return None

        @staticmethod
        def clear(tag=None):
            """Clear all jobs or jobs with a specific tag."""
            if tag:
                schedule._jobs = [j for j in schedule._jobs if getattr(j, '_tag', None) != tag]
            else:
                schedule._jobs = []
            return None

        @staticmethod
        def get_jobs(tag=None):
            """Get all scheduled jobs or jobs with a specific tag."""
            if tag:
                return [j for j in schedule._jobs if getattr(j, '_tag', None) == tag]
            return schedule._jobs.copy()

        @staticmethod
        def cancel_job(job):
            """Cancel a specific job."""
            if job in schedule._jobs:
                schedule._jobs.remove(job)
            return None

# reportlab
try:
    from reportlab.lib.pagesizes import letter, A4  # type: ignore
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak  # type: ignore
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle  # type: ignore
    from reportlab.lib import colors  # type: ignore
    from reportlab.lib.units import inch  # type: ignore
except Exception as e:
    print(f"Optional dependency not installed: {e}")
    warnings.warn(
        "reportlab library not installed. PDF generation will not work. "
        "Install with: pip install reportlab",
        RuntimeWarning
    )

    # Provide standard page sizes
    letter = (8.5 * 72, 11 * 72)
    A4 = (8.27 * 72, 11.69 * 72)
    inch = 72

    class _RLDocStub:
        """
        Stub for reportlab.platypus.SimpleDocTemplate.

        Simulates PDF document creation without actually generating files.
        Stores the story elements for inspection and testing.
        """

        def __init__(self, filename, *args, **kwargs):
            """
            Initialize the document stub.

            Args:
                filename (str): Output filename for the PDF
                *args: Additional positional arguments
                **kwargs: Keyword arguments (pagesize, margins, etc.)
            """
            self.filename = filename
            self.pagesize = kwargs.get('pagesize', letter)
            self.leftMargin = kwargs.get('leftMargin', 72)
            self.rightMargin = kwargs.get('rightMargin', 72)
            self.topMargin = kwargs.get('topMargin', 72)
            self.bottomMargin = kwargs.get('bottomMargin', 72)
            self.title = kwargs.get('title', '')
            self.author = kwargs.get('author', '')
            self.story = []
            self._built = False
            logging.debug(f"Created PDF document stub: {filename}")

        def build(self, story, onFirstPage=None, onLaterPages=None, canvasmaker=None):
            """
            Simulate building the PDF document.

            Args:
                story (list): List of flowable elements to add to the document
                onFirstPage: Callback for first page (not used in stub)
                onLaterPages: Callback for later pages (not used in stub)
                canvasmaker: Custom canvas maker (not used in stub)
            """
            self.story = story
            self._built = True
            logging.info(
                f"PDF document stub built: {self.filename} "
                f"({len(story)} elements, reportlab not installed)"
            )
            warnings.warn(
                f"PDF '{self.filename}' not actually created (reportlab not installed). "
                f"Story contains {len(story)} elements.",
                RuntimeWarning
            )

        def multiBuild(self, story, *args, **kwargs):
            """Simulate multi-pass document building."""
            return self.build(story)

    class _RLTableStub:
        """
        Stub for reportlab.platypus.Table.

        Stores table data and styling for inspection without rendering.
        """

        def __init__(self, data, colWidths=None, rowHeights=None, style=None, *args, **kwargs):
            """
            Initialize the table stub.

            Args:
                data (list): 2D list of table data
                colWidths (list): Column widths
                rowHeights (list): Row heights
                style: Table style object
                *args: Additional positional arguments
                **kwargs: Additional keyword arguments
            """
            self.data = data
            self.colWidths = colWidths
            self.rowHeights = rowHeights
            self._style = style
            self.hAlign = kwargs.get('hAlign', 'CENTER')
            self.vAlign = kwargs.get('vAlign', 'MIDDLE')
            self._kwargs = kwargs
            logging.debug(
                f"Created table stub with {len(data) if data else 0} rows, "
                f"{len(data[0]) if data and data[0] else 0} columns"
            )

        def setStyle(self, style):
            """
            Set the table style.

            Args:
                style: TableStyle object or stub
            """
            self._style = style
            logging.debug(f"Table style set: {type(style).__name__}")

        def getKeepWithNext(self):
            """Get keep-with-next property."""
            return False

        def setKeepWithNext(self, val):
            """Set keep-with-next property."""
            pass

    class _RLTableStyleStub:
        """
        Stub for reportlab.platypus.TableStyle.

        Stores table styling commands without applying them.
        """

        def __init__(self, commands=None, parent=None, **kwargs):
            """
            Initialize the table style stub.

            Args:
                commands (list): List of style commands
                parent: Parent style (not used in stub)
                **kwargs: Additional style properties
            """
            self.commands = commands or []
            self.parent = parent
            self._kwargs = kwargs
            logging.debug(f"Created table style stub with {len(self.commands)} commands")

        def add(self, *commands):
            """Add style commands."""
            self.commands.extend(commands)

        def getCommands(self):
            """Get all style commands."""
            return self.commands

    class _RLParagraphStub:
        """
        Stub for reportlab.platypus.Paragraph.

        Stores paragraph text and style without rendering.
        """

        def __init__(self, text, style, *args, **kwargs):
            """
            Initialize the paragraph stub.

            Args:
                text (str): Paragraph text (may contain markup)
                style: ParagraphStyle object or stub
                *args: Additional positional arguments
                **kwargs: Additional keyword arguments
            """
            self.text = text
            self.style = style
            self._args = args
            self._kwargs = kwargs
            logging.debug(f"Created paragraph stub: {text[:50]}...")

        def getPlainText(self):
            """Get plain text without markup."""
            # Simple markup removal
            import re
            return re.sub(r'<[^>]+>', '', self.text)

        def wrap(self, availWidth, availHeight):
            """Simulate text wrapping."""
            return (availWidth, 100)  # Return dummy dimensions

        def split(self, availWidth, availHeight):
            """Simulate paragraph splitting."""
            return [self]

    class _RLSpacerStub:
        """
        Stub for reportlab.platypus.Spacer.

        Represents vertical space in a document.
        """

        def __init__(self, width, height, *args, **kwargs):
            """
            Initialize the spacer stub.

            Args:
                width: Spacer width
                height: Spacer height
                *args: Additional positional arguments
                **kwargs: Additional keyword arguments
            """
            self.width = width
            self.height = height
            self._args = args
            self._kwargs = kwargs
            logging.debug(f"Created spacer stub: {width}x{height}")

        def wrap(self, availWidth, availHeight):
            """Return spacer dimensions."""
            return (self.width, self.height)

    class _RLImageStub:
        """
        Stub for reportlab.platypus.Image.

        Represents an image without actually loading or rendering it.
        """

        def __init__(self, filename, width=None, height=None, *args, **kwargs):
            """
            Initialize the image stub.

            Args:
                filename (str): Path to image file
                width: Image width
                height: Image height
                *args: Additional positional arguments
                **kwargs: Additional keyword arguments
            """
            self.filename = filename
            self.width = width
            self.height = height
            self.hAlign = kwargs.get('hAlign', 'CENTER')
            self._args = args
            self._kwargs = kwargs
            logging.debug(f"Created image stub: {filename} ({width}x{height})")

        def wrap(self, availWidth, availHeight):
            """Return image dimensions."""
            w = self.width if self.width else availWidth
            h = self.height if self.height else 100
            return (w, h)

    class _RLPageBreakStub:
        """
        Stub for reportlab.platypus.PageBreak.

        Represents a page break in a document.
        """

        def __init__(self, *args, **kwargs):
            """Initialize the page break stub."""
            self._args = args
            self._kwargs = kwargs
            logging.debug("Created page break stub")

        def wrap(self, availWidth, availHeight):
            """Return zero dimensions."""
            return (0, 0)

    class _ColorsStub:
        """
        Stub for reportlab.lib.colors module.

        Provides common color constants without actual color objects.
        """

        def __init__(self):
            """Initialize common color constants."""
            # Define common colors as tuples (R, G, B, A)
            self.black = (0, 0, 0, 1)
            self.white = (1, 1, 1, 1)
            self.red = (1, 0, 0, 1)
            self.green = (0, 1, 0, 1)
            self.blue = (0, 0, 1, 1)
            self.yellow = (1, 1, 0, 1)
            self.cyan = (0, 1, 1, 1)
            self.magenta = (1, 0, 1, 1)
            self.grey = (0.5, 0.5, 0.5, 1)
            self.lightgrey = (0.75, 0.75, 0.75, 1)
            self.darkgrey = (0.25, 0.25, 0.25, 1)

        def HexColor(self, hexcode, *args, **kwargs):
            """Create color from hex code."""
            return hexcode

        def Color(self, r, g, b, a=1):
            """Create color from RGBA values."""
            return (r, g, b, a)

    def getSampleStyleSheet():
        """
        Return a sample stylesheet stub.

        Returns:
            dict: Dictionary with common style names
        """
        return {
            'Normal': ParagraphStyle('Normal'),
            'Heading1': ParagraphStyle('Heading1'),
            'Heading2': ParagraphStyle('Heading2'),
            'Heading3': ParagraphStyle('Heading3'),
            'Title': ParagraphStyle('Title'),
            'BodyText': ParagraphStyle('BodyText'),
            'Code': ParagraphStyle('Code'),
        }

    class ParagraphStyle:
        """
        Stub for reportlab.lib.styles.ParagraphStyle.

        Stores paragraph styling properties.
        """

        def __init__(self, name, parent=None, **kwargs):
            """
            Initialize the paragraph style stub.

            Args:
                name (str): Style name
                parent: Parent style (not used in stub)
                **kwargs: Style properties (fontSize, fontName, textColor, etc.)
            """
            self.name = name
            self.parent = parent
            self.fontName = kwargs.get('fontName', 'Helvetica')
            self.fontSize = kwargs.get('fontSize', 12)
            self.leading = kwargs.get('leading', 14)
            self.textColor = kwargs.get('textColor', (0, 0, 0, 1))
            self.alignment = kwargs.get('alignment', 0)  # 0=left, 1=center, 2=right
            self.spaceAfter = kwargs.get('spaceAfter', 0)
            self.spaceBefore = kwargs.get('spaceBefore', 0)
            self.leftIndent = kwargs.get('leftIndent', 0)
            self.rightIndent = kwargs.get('rightIndent', 0)
            self.firstLineIndent = kwargs.get('firstLineIndent', 0)
            self.bulletFontName = kwargs.get('bulletFontName', 'Helvetica')
            self.bulletFontSize = kwargs.get('bulletFontSize', 12)
            self._kwargs = kwargs
            logging.debug(f"Created paragraph style stub: {name}")

        def __repr__(self):
            return f"ParagraphStyle('{self.name}')"

    # Assign stubs to expected names
    colors = _ColorsStub()
    SimpleDocTemplate = _RLDocStub  # type: ignore
    Table = _RLTableStub  # type: ignore
    TableStyle = _RLTableStyleStub  # type: ignore
    Paragraph = _RLParagraphStub  # type: ignore
    Spacer = _RLSpacerStub  # type: ignore
    Image = _RLImageStub  # type: ignore
    PageBreak = _RLPageBreakStub  # type: ignore

# Flask
try:
    from flask import Flask, request, jsonify, session, render_template_string  # type: ignore
except Exception as e:
    print(f"Optional dependency not installed: {e}")
    warnings.warn(
        "Flask library not installed. Web server functionality will not work. "
        "Install with: pip install flask",
        RuntimeWarning
    )

    class _FlaskStub:
        """
        Stub for Flask application when Flask is not installed.

        Simulates Flask app behavior without actually starting a web server.
        Routes and handlers are registered but not served.
        """

        def __init__(self, import_name, *args, **kwargs):
            """
            Initialize the Flask stub.

            Args:
                import_name (str): Name of the application module
                *args: Additional positional arguments
                **kwargs: Configuration options
            """
            self.import_name = import_name
            self.config = kwargs.get('instance_path', {})
            self.debug = False
            self.testing = False
            self._routes = {}
            self._error_handlers = {}
            self._before_request_funcs = []
            self._after_request_funcs = []
            self._template_folder = kwargs.get('template_folder', 'templates')
            self._static_folder = kwargs.get('static_folder', 'static')
            self._secret_key = None
            logging.debug(f"Created Flask app stub: {import_name}")
            logging.warning("Flask not installed - web server will not start")

        def route(self, rule, **options):
            """
            Decorator to register a route handler.

            Args:
                rule (str): URL rule (e.g., '/api/data')
                **options: Additional route options (methods, defaults, etc.)

            Returns:
                function: Decorator function
            """
            def decorator(func):
                methods = options.get('methods', ['GET'])
                self._routes[rule] = {
                    'function': func,
                    'methods': methods,
                    'options': options
                }
                logging.debug(
                    f"Registered route stub: {methods} {rule} -> {func.__name__}"
                )
                return func
            return decorator

        def run(self, host=None, port=None, debug=None, **options):
            """
            Simulate running the Flask development server.

            Args:
                host (str): Server hostname (default: 127.0.0.1)
                port (int): Server port (default: 5000)
                debug (bool): Enable debug mode
                **options: Additional server options
            """
            host = host or '127.0.0.1'
            port = port or 5000
            if debug is not None:
                self.debug = debug

            logging.warning(
                f"Flask app '{self.import_name}' run() called but Flask not installed. "
                f"Would start server at http://{host}:{port}/"
            )
            logging.info(
                f"Registered routes in stub: {list(self._routes.keys())}"
            )
            warnings.warn(
                f"Flask web server not started (Flask not installed). "
                f"Would run at http://{host}:{port}/ with {len(self._routes)} routes.",
                RuntimeWarning
            )

        def errorhandler(self, code_or_exception):
            """
            Decorator to register an error handler.

            Args:
                code_or_exception: HTTP status code or exception class

            Returns:
                function: Decorator function
            """
            def decorator(func):
                self._error_handlers[code_or_exception] = func
                logging.debug(
                    f"Registered error handler stub: {code_or_exception} -> {func.__name__}"
                )
                return func
            return decorator

        def before_request(self, func):
            """
            Register a function to run before each request.

            Args:
                func: Function to execute before requests

            Returns:
                function: The same function
            """
            self._before_request_funcs.append(func)
            logging.debug(f"Registered before_request stub: {func.__name__}")
            return func

        def after_request(self, func):
            """
            Register a function to run after each request.

            Args:
                func: Function to execute after requests

            Returns:
                function: The same function
            """
            self._after_request_funcs.append(func)
            logging.debug(f"Registered after_request stub: {func.__name__}")
            return func

        def context_processor(self, func):
            """Register a template context processor."""
            logging.debug(f"Registered context_processor stub: {func.__name__}")
            return func

        def teardown_appcontext(self, func):
            """Register a function to call when app context tears down."""
            logging.debug(f"Registered teardown_appcontext stub: {func.__name__}")
            return func

        def add_url_rule(self, rule, endpoint=None, view_func=None, **options):
            """Add a URL rule to the application."""
            self._routes[rule] = {
                'function': view_func,
                'endpoint': endpoint,
                'options': options
            }
            logging.debug(f"Added URL rule stub: {rule} -> {endpoint}")

        @property
        def secret_key(self):
            """Get the secret key."""
            return self._secret_key

        @secret_key.setter
        def secret_key(self, value):
            """Set the secret key."""
            self._secret_key = value

        def test_client(self):
            """Return a test client stub."""
            logging.warning("test_client() called but Flask not installed")
            return _FlaskTestClientStub()

    class _FlaskTestClientStub:
        """Stub for Flask test client."""

        def __init__(self):
            """Initialize test client stub."""
            logging.debug("Created Flask test client stub")

        def get(self, *args, **kwargs):
            """Simulate GET request."""
            logging.debug(f"Test client GET stub called: {args}")
            return _FlaskResponseStub()

        def post(self, *args, **kwargs):
            """Simulate POST request."""
            logging.debug(f"Test client POST stub called: {args}")
            return _FlaskResponseStub()

        def put(self, *args, **kwargs):
            """Simulate PUT request."""
            logging.debug(f"Test client PUT stub called: {args}")
            return _FlaskResponseStub()

        def delete(self, *args, **kwargs):
            """Simulate DELETE request."""
            logging.debug(f"Test client DELETE stub called: {args}")
            return _FlaskResponseStub()

    class _FlaskResponseStub:
        """Stub for Flask response object."""

        def __init__(self):
            """Initialize response stub."""
            self.status_code = 200
            self.data = b''
            self.headers = {}

        def get_json(self):
            """Get JSON data from response."""
            return {}

    class _RequestStub:
        """
        Stub for Flask request object.

        Provides empty/None values for all request attributes.
        """

        def __init__(self):
            """Initialize request stub with empty values."""
            self.method = 'GET'
            self.args = {}
            self.form = {}
            self.json = {}
            self.data = b''
            self.files = {}
            self.headers = {}
            self.cookies = {}
            self.path = '/'
            self.url = 'http://localhost:5000/'
            self.base_url = 'http://localhost:5000/'
            self.url_root = 'http://localhost:5000/'
            self.is_json = False
            self.is_secure = False
            self.remote_addr = '127.0.0.1'
            self.environ = {}
            logging.debug("Created Flask request stub")

        def __getattr__(self, name):
            """Return None for any unknown attributes."""
            logging.debug(f"Request stub attribute accessed: {name}")
            return None

        def get_json(self, *args, **kwargs):
            """Get JSON data from request."""
            return self.json

    class _SessionStub(dict):
        """
        Stub for Flask session object.

        Behaves like a dict but logs warnings about Flask not being installed.
        """

        def __init__(self):
            """Initialize session stub."""
            super().__init__()
            logging.debug("Created Flask session stub")

        def __setitem__(self, key, value):
            """Set session item with logging."""
            logging.debug(f"Session stub set: {key} = {value}")
            super().__setitem__(key, value)

        def __getitem__(self, key):
            """Get session item with logging."""
            logging.debug(f"Session stub get: {key}")
            return super().__getitem__(key)

    def jsonify(*args, **kwargs):
        """
        Stub for Flask jsonify function.

        Args:
            *args: Positional arguments (dicts to serialize)
            **kwargs: Keyword arguments (key-value pairs to serialize)

        Returns:
            dict: The input data as-is (not actually JSONified)
        """
        if args:
            data = args[0] if len(args) == 1 else args
        else:
            data = kwargs

        logging.debug(f"jsonify stub called with: {type(data).__name__}")
        return data

    def render_template_string(source, **context):
        """
        Stub for Flask render_template_string function.

        Args:
            source (str): Template source string
            **context: Template context variables

        Returns:
            str: The source string as-is (not actually rendered)
        """
        logging.debug(
            f"render_template_string stub called "
            f"(template length: {len(source)}, context: {list(context.keys())})"
        )
        # Simple variable substitution for basic testing
        result = source
        for key, value in context.items():
            result = result.replace(f'{{{{{key}}}}}', str(value))
        return result

    # Assign stubs to expected names
    Flask = _FlaskStub  # type: ignore
    request = _RequestStub()  # type: ignore
    session = _SessionStub()  # type: ignore

# sklearn
try:
    from sklearn.ensemble import RandomForestClassifier, IsolationForest  # type: ignore
    from sklearn.model_selection import train_test_split  # type: ignore
    from sklearn.preprocessing import StandardScaler  # type: ignore
    from sklearn.metrics import accuracy_score, classification_report  # type: ignore
except Exception as e:
    print(f"Optional dependency not installed: {e}")
    # Provide minimal sklearn stubs
    class _SKLearnDummy:
        def __init__(self, *args, **kwargs):
            pass
        def fit(self, *args, **kwargs):
            return self
        def predict(self, *args, **kwargs):
            return []
        def score(self, *args, **kwargs):
            return 0.0
        def transform(self, *args, **kwargs):
            return []
        def fit_transform(self, *args, **kwargs):
            return []
    RandomForestClassifier = _SKLearnDummy  # type: ignore
    IsolationForest = _SKLearnDummy  # type: ignore
    def train_test_split(*args, **kwargs):  # type: ignore
        return ([], [], [], [])
    StandardScaler = _SKLearnDummy  # type: ignore
    def accuracy_score(*args, **kwargs):  # type: ignore
        return 0.0
    def classification_report(*args, **kwargs):  # type: ignore
        return ""
try:
    from university_system.utils.logging.config import configure_logging
    from university_system.utils.logging.config import get_log_file
except ImportError:
    # Fallback stubs if logging config not available
    def configure_logging(name=None):
        import logging
        return logging.getLogger(name or __name__)
    def get_log_file(filename="system.log"):
        from university_system.modules.shared.constants import paths
        return str(paths.LOG_DIR / filename)

warnings.filterwarnings('ignore')

# Configuration
CONFIG = {
    'database': str(paths.DEFAULT_DB_PATH),
    'reports_dir': str(paths.REPORTS_DIR),
    'templates_dir': str(paths.REPORT_TEMPLATES_DIR),
    'cache_dir': str(paths.REPORT_CACHE_DIR),
    'logs_dir': str(paths.LOG_DIR),
    'scheduled_reports_file': str(paths.REPORTS_DIR / 'scheduled_reports.json'),
    'config_file': str(paths.DATA_DIR / 'system_config.json'),
    'secret_key': os.environ.get('FLASK_SECRET_KEY', secrets.token_hex(32)),
    'cache_expiry_hours': 24,
    'max_cache_size_mb': 500
}

# Ensure all directories exist
for dir_path in [CONFIG['reports_dir'], CONFIG['templates_dir'], CONFIG['cache_dir'], CONFIG['logs_dir']]:
    os.makedirs(dir_path, exist_ok=True)

# Setup logging
logger = configure_logging(name=__name__)

# Flask app for REST API
app = Flask(__name__)
app.secret_key = CONFIG['secret_key']

# Initialize security headers for all responses
try:
    from university_system.infrastructure.security.flask_security_headers import init_security_headers
    init_security_headers(app)
except ImportError:
    pass  # Security headers module not available

@app.errorhandler(404)
def not_found(error):
    return jsonify({'error': 'Endpoint not found'}), 404

@app.errorhandler(500)
def internal_error(error):
    return jsonify({'error': 'Internal server error'}), 500

class SystemConfig:
    """System configuration management"""
    
    @staticmethod
    def load_config():
        try:
            with open(CONFIG['config_file'], 'r') as f:
                return json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            return {
                'security': {
                    'session_timeout': 3600,
                    'max_login_attempts': 5,
                    'require_2fa': False
                },
                'performance': {
                    'enable_caching': True,
                    'max_concurrent_reports': 5,
                    'report_timeout': 1800
                }
            }
    
    @staticmethod
    def save_config(config):
        with open(CONFIG['config_file'], 'w') as f:
            json.dump(config, f, indent=4)

class CacheManager:
    """Report caching for improved performance"""
    
    @staticmethod
    def get_cache_key(template_name, start_date, end_date, filters=None):
        key_data = f"{template_name}_{start_date}_{end_date}_{str(filters)}"
        return hashlib.md5(key_data.encode()).hexdigest()
    
    @staticmethod
    def get_cached_report(cache_key):
        cache_file = os.path.join(CONFIG['cache_dir'], f"{cache_key}.json")
        
        if not os.path.exists(cache_file):
            return None
        
        # Check if cache is expired
        file_time = datetime.fromtimestamp(os.path.getmtime(cache_file))
        if datetime.now() - file_time > timedelta(hours=CONFIG['cache_expiry_hours']):
            os.remove(cache_file)
            return None
        
        try:
            with open(cache_file, 'r') as f:
                return json.load(f)
        except (OSError, IOError, json.JSONDecodeError):
            return None
    
    @staticmethod
    def cache_report(cache_key, report_data):
        cache_file = os.path.join(CONFIG['cache_dir'], f"{cache_key}.json")
        
        try:
            with open(cache_file, 'w') as f:
                json.dump(report_data, f)
            
            # Clean up old cache files if needed
            CacheManager.cleanup_cache()
            
        except Exception as e:
            logger.error(f"Failed to cache report: {str(e)}")
    
    @staticmethod
    def cleanup_cache():
        cache_files = []
        total_size = 0
        
        for file in os.listdir(CONFIG['cache_dir']):
            if file.endswith('.json'):
                file_path = os.path.join(CONFIG['cache_dir'], file)
                size = os.path.getsize(file_path)
                mtime = os.path.getmtime(file_path)
                cache_files.append((file_path, size, mtime))
                total_size += size
        
        # Remove old files if cache is too large
        max_size = CONFIG['max_cache_size_mb'] * 1024 * 1024
        if total_size > max_size:
            cache_files.sort(key=lambda x: x[2])  # Sort by modification time
            
            for file_path, size, _ in cache_files:
                if total_size <= max_size:
                    break
                os.remove(file_path)
                total_size -= size

class DataQualityMonitor:
    """Monitor and report on data quality issues"""
    
    @staticmethod
    def run_quality_checks():
        conn = get_reporting_db_connection()
        quality_report = {
            'timestamp': datetime.now().isoformat(),
            'checks': {}
        }
        
        try:
            # Check for missing data
            missing_data = DataQualityMonitor.check_missing_data(conn)
            quality_report['checks']['missing_data'] = missing_data
            
            # Check for duplicates
            duplicates = DataQualityMonitor.check_duplicates(conn)
            quality_report['checks']['duplicates'] = duplicates
            
            # Check for invalid data
            invalid_data = DataQualityMonitor.check_invalid_data(conn)
            quality_report['checks']['invalid_data'] = invalid_data
            
            # Check data freshness
            freshness = DataQualityMonitor.check_data_freshness(conn)
            quality_report['checks']['data_freshness'] = freshness
            
        finally:
            conn.close()
        
        return quality_report
    
    @staticmethod
    def check_missing_data(conn):
        results = {}
        
        # Check students table
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM students WHERE email_address IS NULL OR email_address = ''")
        missing_emails = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM students WHERE first_name IS NULL OR first_name = ''")
        missing_names = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM students WHERE course IS NULL OR course = ''")
        missing_courses = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM students")
        total_students = cursor.fetchone()[0]
        
        results['students'] = {
            'missing_emails': missing_emails,
            'missing_names': missing_names,
            'missing_courses': missing_courses,
            'total_records': total_students
        }
        
        return results
    
    @staticmethod
    def check_duplicates(conn):
        results = {}
        
        cursor = conn.cursor()
        
        # Check for duplicate emails
        cursor.execute("""
            SELECT email_address, COUNT(*) as count 
            FROM students 
            WHERE email_address IS NOT NULL AND email_address != ''
            GROUP BY email_address 
            HAVING COUNT(*) > 1
        """)
        duplicate_emails = cursor.fetchall()
        
        results['duplicate_emails'] = len(duplicate_emails)
        results['duplicate_email_details'] = [{'email': row[0], 'count': row[1]} for row in duplicate_emails]
        
        return results
    
    @staticmethod
    def check_invalid_data(conn):
        results = {}
        
        cursor = conn.cursor()
        
        # Check for invalid ages
        cursor.execute("SELECT COUNT(*) FROM students WHERE age < 0 OR age > 100")
        invalid_ages = cursor.fetchone()[0]
        
        # Check for invalid email formats
        cursor.execute("""
            SELECT COUNT(*) FROM students 
            WHERE email_address IS NOT NULL 
            AND email_address != ''
            AND email_address NOT LIKE '%@%.%'
        """)
        invalid_emails = cursor.fetchone()[0]
        
        results['invalid_ages'] = invalid_ages
        results['invalid_emails'] = invalid_emails
        
        return results
    
    @staticmethod
    def check_data_freshness(conn):
        results = {}
        
        cursor = conn.cursor()
        
        # Check when data was last updated
        cursor.execute("SELECT MAX(registration_datetime) FROM students")
        last_registration = cursor.fetchone()[0]
        
        if last_registration:
            last_reg_date = datetime.strptime(last_registration, '%Y-%m-%d %H:%M:%S')
            days_since_last = (datetime.now() - last_reg_date).days
            results['days_since_last_registration'] = days_since_last
            results['last_registration_date'] = last_registration
        else:
            results['days_since_last_registration'] = None
            results['last_registration_date'] = None
        
        return results

class PredictiveAnalytics:
    """Machine learning for predictive insights"""
    
    @staticmethod
    def predict_dropout_risk():
        """Predict which students are at risk of dropping out"""
        conn = get_reporting_db_connection()
        
        try:
            # Check which attendance table exists
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='attendance_records'")
            
            if cursor.fetchone():
                attendance_table = 'attendance_records'
                status_column = 'status'
            else:
                # No attendance data available
                return {'error': 'No attendance data available for prediction'}
            
            # Fixed query with proper table aliases
            query = f"""
            SELECT s.student_id, s.age, s.course,
                   COUNT(sm.module_code) as module_count,
                   AVG(CASE WHEN sg.grade IS NOT NULL THEN CAST(sg.grade AS FLOAT) ELSE NULL END) as avg_grade,
                   COUNT(ar.student_id) as attendance_count,
                   SUM(CASE WHEN LOWER(ar.{status_column}) IN ('present', 'attended') THEN 1 ELSE 0 END) as present_count
            FROM students s
            LEFT JOIN student_modules sm ON s.student_id = sm.student_id
            LEFT JOIN student_grades sg ON s.student_id = sg.student_id
            LEFT JOIN {attendance_table} ar ON s.student_id = ar.student_id
            WHERE s.registration_datetime >= date('now', '-1 year')
            GROUP BY s.student_id, s.age, s.course
            """
            
            df = pd.read_sql_query(query, conn)
            
            if df.empty or len(df) < 10:
                return {'error': 'Insufficient data for prediction'}
            
            # Feature engineering
            df['attendance_rate'] = df.apply(
                lambda row: row['present_count'] / row['attendance_count'] if row['attendance_count'] > 0 else 0,
                axis=1
            )
            df['course_numeric'] = df['course'].map({'CS': 1, 'DS': 2}).fillna(0)
            df = df.fillna(0)
            
            # For demonstration, create synthetic dropout labels
            # In reality, you'd have historical dropout data
            np.random.seed(42)
            df['dropout_risk'] = np.random.choice([0, 1], size=len(df), p=[0.8, 0.2])
            
            # Prepare features
            features = ['age', 'module_count', 'avg_grade', 'attendance_rate', 'course_numeric']
            X = df[features]
            y = df['dropout_risk']
            
            # Train model
            if len(np.unique(y)) > 1:
                X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.3, random_state=42)
                
                model = RandomForestClassifier(n_estimators=100, random_state=42)
                model.fit(X_train, y_train)
                
                # Predictions
                y_pred = model.predict(X_test)
                accuracy = accuracy_score(y_test, y_pred)
                
                # Get feature importance
                importance = dict(zip(features, model.feature_importances_))
                
                # Predict for all students
                df['risk_score'] = model.predict_proba(X)[:, 1]
                high_risk_students = df[df['risk_score'] > 0.7][['student_id', 'risk_score']].to_dict('records')
                
                return {
                    'model_accuracy': accuracy,
                    'feature_importance': importance,
                    'high_risk_students': high_risk_students,
                    'total_students_analyzed': len(df)
                }
            else:
                return {'error': 'Insufficient variety in data for meaningful prediction'}
                
        except Exception as e:
            logger.error(f"Error in dropout prediction: {str(e)}")
            return {'error': f'Prediction failed: {str(e)}'}
        finally:
            conn.close()
        
    @staticmethod
    def detect_anomalies():
        """Detect anomalous patterns in student data"""
        conn = get_reporting_db_connection()
        
        try:
            # Fixed query with table aliases to resolve ambiguous column names
            query = """
            SELECT s.student_id, s.age, 
                   COUNT(DISTINCT sm.module_code) as unique_modules,
                   AVG(CASE WHEN sg.grade IS NOT NULL THEN CAST(sg.grade AS FLOAT) ELSE NULL END) as avg_grade
            FROM students s
            LEFT JOIN student_modules sm ON s.student_id = sm.student_id
            LEFT JOIN student_grades sg ON s.student_id = sg.student_id
            GROUP BY s.student_id, s.age
            """
            
            df = pd.read_sql_query(query, conn)
            
            if df.empty or len(df) < 10:
                return {'error': 'Insufficient data for anomaly detection'}
            
            # Prepare features for anomaly detection
            features = ['age', 'unique_modules', 'avg_grade']
            X = df[features].fillna(0)
            
            # Use Isolation Forest for anomaly detection
            iso_forest = IsolationForest(contamination=0.1, random_state=42)
            anomalies = iso_forest.fit_predict(X)
            
            df['is_anomaly'] = anomalies == -1
            anomalous_students = df[df['is_anomaly']][['student_id', 'age', 'unique_modules', 'avg_grade']].to_dict('records')
            
            return {
                'anomalous_students': anomalous_students,
                'total_anomalies': len(anomalous_students),
                'anomaly_rate': len(anomalous_students) / len(df) * 100
            }
            
        except Exception as e:
            logger.error(f"Error in anomaly detection: {str(e)}")
            return {'error': f'Anomaly detection failed: {str(e)}'}
        finally:
            conn.close()
        
class AdvancedVisualization:
    """Advanced chart types and interactive visualizations"""
    
    @staticmethod
    def create_heatmap(data, title, x_col, y_col, value_col):
        """Create a heatmap visualization"""
        plt.figure(figsize=(12, 8))
        
        # Pivot data for heatmap
        pivot_data = data.pivot(index=y_col, columns=x_col, values=value_col)
        
        sns.heatmap(pivot_data, annot=True, cmap='YlOrRd', fmt='.1f')
        plt.title(title)
        plt.tight_layout()
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"heatmap_{timestamp}.png"
        filepath = os.path.join(CONFIG['reports_dir'], filename)
        plt.savefig(filepath, dpi=300, bbox_inches='tight')
        plt.close()
        
        return filepath
    
    @staticmethod
    def create_interactive_dashboard(data_dict):
        """Create an interactive Plotly dashboard"""
        from plotly.subplots import make_subplots
        
        fig = make_subplots(
            rows=2, cols=2,
            subplot_titles=('Course Distribution', 'Age Distribution', 'Registration Trends', 'Module Popularity'),
            specs=[[{"type": "pie"}, {"type": "bar"}],
                   [{"type": "scatter"}, {"type": "bar"}]]
        )
        
        # Course distribution pie chart
        if 'course_distribution' in data_dict:
            course_data = data_dict['course_distribution']
            fig.add_trace(
                go.Pie(labels=course_data['course'], values=course_data['student_count'], name="Courses"),
                row=1, col=1
            )
        
        # Age distribution bar chart
        if 'age_distribution' in data_dict:
            age_data = data_dict['age_distribution']
            fig.add_trace(
                go.Bar(x=age_data['age_group'], y=age_data['student_count'], name="Age Groups"),
                row=1, col=2
            )
        
        # Registration trends line chart
        if 'registration_trends' in data_dict:
            reg_data = data_dict['registration_trends']
            fig.add_trace(
                go.Scatter(x=reg_data['registration_date'], y=reg_data['registration_count'], 
                          mode='lines+markers', name="Registrations"),
                row=2, col=1
            )
        
        # Module popularity bar chart
        if 'module_popularity' in data_dict:
            module_data = data_dict['module_popularity'].head(10)
            fig.add_trace(
                go.Bar(x=module_data['module_code'], y=module_data['student_count'], name="Modules"),
                row=2, col=2
            )
        
        fig.update_layout(height=800, showlegend=False, title_text="Student Analytics Dashboard")
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"dashboard_{timestamp}.html"
        filepath = os.path.join(CONFIG['reports_dir'], filename)
        
        pyo.plot(fig, filename=filepath, auto_open=False)
        
        return filepath
    
    @staticmethod
    def create_correlation_matrix(conn):
        """Create correlation matrix for numeric variables"""
        # Check which attendance table exists
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='attendance_records'")
        
        if cursor.fetchone():
            attendance_table = 'attendance_records'
            status_column = 'status'
        else:
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='student_attendance'")
            if cursor.fetchone():
                attendance_table = 'student_attendance'
                status_column = 'status'
            else:
                return None
        
        # Fixed query with proper table aliases
        query = f"""
        SELECT s.age, 
               COUNT(sm.module_code) as module_count,
               AVG(CASE WHEN sg.grade IS NOT NULL THEN CAST(sg.grade AS FLOAT) ELSE NULL END) as avg_grade,
               COUNT(sa.student_id) as attendance_records,
               SUM(CASE WHEN sa.{status_column} = 'present' THEN 1 ELSE 0 END) as present_count
        FROM students s
        LEFT JOIN student_modules sm ON s.student_id = sm.student_id
        LEFT JOIN student_grades sg ON s.student_id = sg.student_id
        LEFT JOIN {attendance_table} sa ON s.student_id = sa.student_id
        GROUP BY s.student_id, s.age
        """
            
        df = pd.read_sql_query(query, conn)
        df = df.fillna(0)
        
        if len(df) < 2:
            return None
        
        # Calculate attendance rate
        df['attendance_rate'] = df.apply(
            lambda row: row['present_count'] / row['attendance_records'] if row['attendance_records'] > 0 else 0,
            axis=1
        )
        
        # Select numeric columns for correlation
        numeric_cols = ['age', 'module_count', 'avg_grade', 'attendance_rate']
        corr_matrix = df[numeric_cols].corr()
        
        plt.figure(figsize=(10, 8))
        sns.heatmap(corr_matrix, annot=True, cmap='coolwarm', center=0, 
                   square=True, linewidths=0.5)
        plt.title('Student Metrics Correlation Matrix')
        plt.tight_layout()
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"correlation_matrix_{timestamp}.png"
        filepath = os.path.join(CONFIG['reports_dir'], filename)
        plt.savefig(filepath, dpi=300, bbox_inches='tight')
        plt.close()
        
        return filepath

class ReportTemplate:
    """Enhanced report template with new features"""
    
    def __init__(self, name, description, sections, filters=None, 
                 visualization_type='standard', schedule_config=None,
                 security_level='normal', custom_sql=None):
        self.name = name
        self.description = description
        self.sections = sections
        self.filters = filters or {}
        self.visualization_type = visualization_type  # 'standard', 'interactive', 'advanced'
        self.schedule_config = schedule_config or {}
        self.security_level = security_level  # 'normal', 'confidential', 'restricted'
        self.custom_sql = custom_sql or {}
        self.created_at = datetime.now().isoformat()
        self.version = "1.0"
    
    def to_dict(self):
        return {
            "name": self.name,
            "description": self.description,
            "sections": self.sections,
            "filters": self.filters,
            "visualization_type": self.visualization_type,
            "schedule_config": self.schedule_config,
            "security_level": self.security_level,
            "custom_sql": self.custom_sql,
            "created_at": self.created_at,
            "version": self.version
        }
    
    @classmethod
    def from_dict(cls, data):
        template = cls(
            name=data["name"],
            description=data["description"],
            sections=data["sections"],
            filters=data.get("filters", {}),
            visualization_type=data.get("visualization_type", "standard"),
            schedule_config=data.get("schedule_config", {}),
            security_level=data.get("security_level", "normal"),
            custom_sql=data.get("custom_sql", {})
        )
        template.created_at = data.get("created_at", datetime.now().isoformat())
        template.version = data.get("version", "1.0")
        return template

# Enhanced available sections with new analytics
AVAILABLE_SECTIONS = [
    "student_overview",
    "student_list",
    "course_distribution",
    "gender_distribution",
    "age_distribution",
    "module_popularity",
    "registration_trends",
    "grade_distribution",
    "attendance_summary",
    "data_quality_report",
    "predictive_analytics",
    "correlation_analysis",
    "anomaly_detection",
    "performance_benchmarks",
    "trend_analysis"
]

# Database connection helper
def get_reporting_db_connection():
    """Create a connection to the database"""
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    return conn

@app.route('/api/health', methods=['GET'])
def api_health():
    """Health check endpoint"""
    try:
        # Test database connection
        conn = get_reporting_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT 1")
        conn.close()
        
        return jsonify({
            'status': 'healthy',
            'timestamp': datetime.now().isoformat(),
            'version': '2.0',
            'database': 'connected'
        })
    except Exception as e:
        return jsonify({
            'status': 'unhealthy',
            'error': str(e),
            'timestamp': datetime.now().isoformat()
        }), 500

def serialize_dataframe(df):
    """Convert DataFrame to JSON-serializable format"""
    if df is None or df.empty:
        return []
    
    # Replace NaN values with None for JSON serialization
    df_clean = df.fillna('')
    return df_clean.to_dict('records')

@app.route('/api/data/<section>', methods=['GET'])
def api_get_section_data(section):
    """Get data for a specific section"""
    try:
        if section not in AVAILABLE_SECTIONS:
            return jsonify({'error': f'Invalid section: {section}'}), 400
        
        # Get date parameters
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        if not end_date:
            end_date = datetime.now().strftime("%Y-%m-%d")
        if not start_date:
            start_dt = datetime.strptime(end_date, "%Y-%m-%d") - timedelta(days=30)
            start_date = start_dt.strftime("%Y-%m-%d")
        
        # Get data
        df = get_section_dataframe(section, start_date, end_date, {})
        data = serialize_dataframe(df)
        
        return jsonify({
            'section': section,
            'data': data,
            'start_date': start_date,
            'end_date': end_date,
            'record_count': len(data)
        })
        
    except Exception as e:
        logger.error(f"API get section data error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/templates', methods=['GET'])
def api_get_templates():
    templates = load_templates()
    return jsonify(templates)

@app.route('/api/templates', methods=['POST'])
def api_create_template():
    data = request.get_json()
    
    template = ReportTemplate(
        name=data['name'],
        description=data['description'],
        sections=data['sections'],
        filters=data.get('filters', {}),
        visualization_type=data.get('visualization_type', 'standard'),
        security_level=data.get('security_level', 'normal')
    )
    
    saved_template = save_template(template)
    return jsonify(saved_template.to_dict())

@app.route('/api/reports/generate', methods=['POST'])
def api_generate_report():
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'error': 'No JSON data provided'}), 400
        
        template_name = data.get('template_name')
        if not template_name:
            return jsonify({'error': 'template_name is required'}), 400
        
        # Verify template exists
        template = get_template(template_name)
        if not template:
            return jsonify({'error': f'Template "{template_name}" not found'}), 404
        
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        output_format = data.get('format', 'pdf')
        
        # Set default dates if not provided
        if not end_date:
            end_date = datetime.now().strftime("%Y-%m-%d")
        if not start_date:
            start_dt = datetime.strptime(end_date, "%Y-%m-%d") - timedelta(days=30)
            start_date = start_dt.strftime("%Y-%m-%d")
        
        report_path = generate_report(template_name, start_date, end_date, output_format)
        
        if report_path:
            return jsonify({
                'success': True, 
                'report_path': report_path,
                'message': 'Report generated successfully'
            })
        else:
            return jsonify({'success': False, 'error': 'Failed to generate report'}), 500
            
    except Exception as e:
        logger.error(f"API generate report error: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/analytics/quality', methods=['GET'])
def api_data_quality():
    quality_report = DataQualityMonitor.run_quality_checks()
    return jsonify(quality_report)

@app.route('/api/analytics/predictions', methods=['GET'])
def api_predictions():
    predictions = PredictiveAnalytics.predict_dropout_risk()
    return jsonify(predictions)

@app.route('/api/analytics/anomalies', methods=['GET'])
def api_anomalies():
    anomalies = PredictiveAnalytics.detect_anomalies()
    return jsonify(anomalies)

# Enhanced template management functions
def save_template(template):
    """Save a report template to database with versioning"""
    from university_system.infrastructure.database.db import get_connection

    conn = get_connection()
    cursor = conn.cursor()

    # Check if template with this name already exists
    cursor.execute("""
        SELECT template_id, template_content FROM email_templates
        WHERE template_name = ?
    """, (template.name,))
    existing = cursor.fetchone()

    template_dict = template.to_dict()
    template_json = json.dumps(template_dict)

    if existing:
        # Update existing template
        old_content = json.loads(existing[1]) if existing[1] else {}
        old_version = old_content.get("version", "1.0")
        major, minor = old_version.split(".")
        template.version = f"{major}.{int(minor) + 1}"
        template_dict = template.to_dict()
        template_json = json.dumps(template_dict)

        cursor.execute("""
            UPDATE email_templates
            SET template_content = ?, template_type = ?, created_date = ?
            WHERE template_id = ?
        """, (template_json, template_dict.get('sections', [{}])[0].get('type', 'custom') if template_dict.get('sections') else 'custom',
              datetime.now().isoformat(), existing[0]))
        logger.info(f"Template {template.name} updated to version {template.version}")
    else:
        # Insert new template
        cursor.execute("""
            INSERT INTO email_templates (template_name, template_content, template_type, created_date, created_by)
            VALUES (?, ?, ?, ?, ?)
        """, (template.name, template_json,
              template_dict.get('sections', [{}])[0].get('type', 'custom') if template_dict.get('sections') else 'custom',
              datetime.now().isoformat(), 'system'))
        logger.info(f"Template {template.name} created")

    conn.commit()
    conn.close()

    return template

def load_templates():
    """Load all report templates from database"""
    from university_system.infrastructure.database.db import get_connection

    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT template_name, template_content FROM email_templates
            WHERE template_type = 'report_template'
            AND template_content IS NOT NULL AND template_content != ''
        """)
        rows = cursor.fetchall()
        conn.close()

        templates = []
        for row in rows:
            try:
                template_data = json.loads(row[1])
                # Ensure it has the expected structure
                if isinstance(template_data, dict) and 'name' in template_data:
                    templates.append(template_data)
                else:
                    # Try to create a basic template structure from the data
                    templates.append({
                        'name': row[0],
                        'description': 'Legacy template',
                        'sections': [],
                        'version': '1.0'
                    })
            except (json.JSONDecodeError, TypeError):
                # Skip invalid JSON entries
                continue

        return templates
    except Exception as e:
        logger.error(f"Error loading templates from database: {e}")
        return []

def save_template_dict(template_data):
    """Save a template dictionary directly to database (for simple templates)"""
    from university_system.infrastructure.database.db import get_connection

    conn = get_connection()
    cursor = conn.cursor()

    # Check if template with this name already exists
    cursor.execute("""
        SELECT template_id FROM email_templates
        WHERE template_name = ? AND template_type = 'report_template'
    """, (template_data['name'],))
    existing = cursor.fetchone()

    template_json = json.dumps(template_data)

    if existing:
        # Update existing template
        cursor.execute("""
            UPDATE email_templates
            SET template_content = ?, created_date = ?
            WHERE template_id = ?
        """, (template_json, datetime.now().isoformat(), existing[0]))
        logger.info(f"Template {template_data['name']} updated")
    else:
        # Insert new template
        cursor.execute("""
            INSERT INTO email_templates (template_name, template_content, template_type, created_date, created_by)
            VALUES (?, ?, ?, ?, ?)
        """, (template_data['name'], template_json, 'report_template',
              datetime.now().isoformat(), 'system'))
        logger.info(f"Template {template_data['name']} created")

    conn.commit()
    conn.close()

    return template_data

def delete_template_from_db(template_name):
    """Delete a template from the database"""
    from university_system.infrastructure.database.db import get_connection

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        DELETE FROM email_templates
        WHERE template_name = ? AND template_type = 'report_template'
    """, (template_name,))

    conn.commit()
    deleted_count = cursor.rowcount
    conn.close()

    logger.info(f"Deleted {deleted_count} template(s) named {template_name}")
    return deleted_count > 0

def get_template(name):
    """Get a specific template by name"""
    templates = load_templates()
    for template_data in templates:
        if template_data["name"] == name:
            return ReportTemplate.from_dict(template_data)
    return None

# Enhanced report generation with all new features
def generate_report(template, start_date=None, end_date=None, format="pdf", comparison_date=None):
    """Generate a report with enhanced features"""
    if isinstance(template, str):
        template = get_template(template)
        if not template:
            logger.error(f"Template '{template}' not found")
            return None
    
    # Set default date range
    if not end_date:
        end_date = datetime.now().strftime("%Y-%m-%d")
    if not start_date:
        end_dt = datetime.strptime(end_date, "%Y-%m-%d")
        start_dt = end_dt - timedelta(days=30)
        start_date = start_dt.strftime("%Y-%m-%d")
    
    # Check cache first
    config = SystemConfig.load_config()
    if config.get('performance', {}).get('enable_caching', True):
        cache_key = CacheManager.get_cache_key(template.name, start_date, end_date, template.filters)
        cached_data = CacheManager.get_cached_report(cache_key)
        
        if cached_data:
            logger.info(f"Using cached data for report {template.name}")
    
    # Generate report filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{template.name.replace(' ', '_')}_{timestamp}"
    
    try:
        if format == "pdf":
            return generate_enhanced_pdf_report(template, filename, start_date, end_date, comparison_date)
        elif format == "excel":
            return generate_enhanced_excel_report(template, filename, start_date, end_date, comparison_date)
        elif format == "interactive":
            return generate_interactive_report(template, filename, start_date, end_date)
        else:
            logger.error(f"Unsupported format: {format}")
            return None
            
    except Exception as e:
        logger.error(f"Error generating report: {str(e)}")
        raise

def generate_enhanced_pdf_report(template, filename, start_date, end_date, comparison_date=None):
    """Generate an enhanced PDF report with all new features"""
    file_path = os.path.join(CONFIG['reports_dir'], f"{filename}.pdf")
    
    doc = SimpleDocTemplate(file_path, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []
    
    # Enhanced title with security marking
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontSize=24,
        spaceAfter=30,
        textColor=colors.darkblue
    )
    
    title_text = template.name
    if template.security_level != 'normal':
        title_text += f" ({template.security_level.upper()})"
    
    elements.append(Paragraph(title_text, title_style))
    
    # Report metadata
    metadata_style = ParagraphStyle(
        'Metadata',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.grey,
        spaceAfter=20
    )
    
    metadata_text = f"""
    Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>
    Template Version: {template.version}<br/>
    Report Period: {start_date} to {end_date}<br/>
    Security Level: {template.security_level.title()}
    """
    elements.append(Paragraph(metadata_text, metadata_style))
    
    # Description
    elements.append(Paragraph(template.description, styles["Normal"]))
    elements.append(Spacer(1, 0.25*inch))
    
    # Data quality summary
    if 'data_quality_report' in template.sections:
        quality_report = DataQualityMonitor.run_quality_checks()
        elements.extend(generate_quality_section(quality_report, styles))
    
    # Generate each section with enhanced visualizations
    for section in template.sections:
        if section in AVAILABLE_SECTIONS:
            try:
                section_elements = generate_enhanced_section(
                    section, start_date, end_date, comparison_date, 
                    template.filters, template.visualization_type
                )
                elements.extend(section_elements)
            except Exception as e:
                logger.error(f"Error generating section {section}: {str(e)}")
                # Add error message to report
                error_text = f"Error generating {section}: {str(e)}"
                elements.append(Paragraph(error_text, styles["Normal"]))
                elements.append(Spacer(1, 0.25*inch))
    
    # Add predictive analytics if requested
    if 'predictive_analytics' in template.sections:
        predictions = PredictiveAnalytics.predict_dropout_risk()
        elements.extend(generate_predictions_section(predictions, styles))
    
    # Build the PDF
    doc.build(elements)
    logger.info(f"Enhanced PDF report generated: {file_path}")
    return file_path

def generate_enhanced_section(section, start_date, end_date, comparison_date=None, 
                            filters=None, visualization_type='standard'):
    """Generate enhanced report sections with advanced visualizations"""
    elements = []
    styles = getSampleStyleSheet()
    
    # Add section heading
    section_title = section.replace('_', ' ').title()
    section_style = ParagraphStyle(
        'SectionTitle',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.darkblue,
        spaceAfter=15
    )
    elements.append(Paragraph(section_title, section_style))
    
    # Get data for the section
    df = get_section_dataframe(section, start_date, end_date, filters)
    
    if df is None or df.empty:
        elements.append(Paragraph("No data available for this section.", styles["Normal"]))
        elements.append(Spacer(1, 0.25*inch))
        return elements
    
    # Generate appropriate visualization based on type
    chart_path = None
    
    try:
        if visualization_type == 'advanced':
            chart_path = create_advanced_visualization(section, df)
        elif visualization_type == 'interactive':
            # For interactive, create standard chart for PDF
            chart_path = create_standard_chart(section, df)
        else:
            chart_path = create_standard_chart(section, df)
        
        # Add chart to report ONLY if it's a valid image file
        if chart_path and os.path.exists(chart_path) and chart_path.lower().endswith(('.png', '.jpg', '.jpeg')):
            try:
                img = Image(chart_path, width=6*inch, height=4*inch)
                elements.append(img)
                elements.append(Spacer(1, 0.15*inch))
            except Exception as e:
                logger.error(f"Error adding chart to report: {str(e)}")
                # Add text description instead
                elements.append(Paragraph(f"Chart generated: {os.path.basename(chart_path)}", styles["Normal"]))
    except Exception as e:
        logger.error(f"Error creating visualization for {section}: {str(e)}")
        
    # Add statistical summary for numeric data
    if section in ['grade_distribution', 'age_distribution']:
        summary_stats = generate_statistical_summary(df, section)
        if summary_stats:
            elements.append(Paragraph("Statistical Summary:", styles["Heading4"]))
            elements.append(summary_stats)
            elements.append(Spacer(1, 0.15*inch))
    
    # Add data table with enhanced styling
    if len(df) <= 100:  # Only show table for smaller datasets
        data_table = create_enhanced_data_table(df)
        elements.append(data_table)
    else:
        # For large datasets, show summary
        summary_text = f"Dataset contains {len(df)} records. Top 10 records shown in visualization."
        elements.append(Paragraph(summary_text, styles["Italic"]))
    
    elements.append(Spacer(1, 0.25*inch))
    return elements

def create_advanced_visualization(section, df):
    """Create advanced visualizations using seaborn and plotly"""
    if section == "correlation_analysis":
        conn = get_reporting_db_connection()
        try:
            return AdvancedVisualization.create_correlation_matrix(conn)
        finally:
            conn.close()
    
    elif section in ["course_distribution", "gender_distribution"]:
        # Create enhanced pie chart with better styling
        plt.figure(figsize=(12, 8))
        
        if section == "course_distribution":
            labels = df['course']
            sizes = df['student_count']
            colors_list = plt.cm.Set3(np.linspace(0, 1, len(labels)))
        else:
            labels = df['gender']
            sizes = df['student_count']
            colors_list = ['#FF9999', '#66B2FF', '#99FF99', '#FFCC99']
        
        wedges, texts, autotexts = plt.pie(sizes, labels=labels, autopct='%1.1f%%', 
                                          startangle=90, colors=colors_list,
                                          explode=[0.05] * len(labels))
        
        # Enhance text styling
        for autotext in autotexts:
            autotext.set_color('white')
            autotext.set_fontweight('bold')
        
        plt.title(f"{section.replace('_', ' ').title()}", fontsize=16, fontweight='bold')
        
        # Add total count
        total = sizes.sum()
        plt.text(0, -1.3, f"Total: {total}", ha='center', fontsize=12, fontweight='bold')
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{section}_advanced_{timestamp}.png"
        filepath = os.path.join(CONFIG['reports_dir'], filename)
        plt.savefig(filepath, dpi=300, bbox_inches='tight')
        plt.close()
        
        return filepath
    
    return None

def create_interactive_chart(section, df):
    """Create interactive charts using Plotly"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{section}_interactive_{timestamp}.html"
    filepath = os.path.join(CONFIG['reports_dir'], filename)
    
    try:
        if section == "course_distribution":
            fig = px.pie(df, values='student_count', names='course', 
                        title='Course Distribution',
                        hover_data=['student_count'])
            
        elif section == "registration_trends":
            fig = px.line(df, x='registration_date', y='registration_count',
                         title='Registration Trends Over Time',
                         markers=True)
            fig.update_layout(xaxis_title="Date", yaxis_title="Registrations")
            
        elif section == "module_popularity":
            top_modules = df.head(15)
            fig = px.bar(top_modules, x='module_code', y='student_count',
                        title='Module Popularity (Top 15)',
                        hover_data=['module_name'])
            fig.update_layout(xaxis_title="Module Code", yaxis_title="Student Count")
            
        else:
            return None
        
        fig.update_layout(
            font=dict(size=12),
            title_font_size=16,
            showlegend=True
        )
        
        pyo.plot(fig, filename=filepath, auto_open=False)
        return filepath
        
    except Exception as e:
        logger.error(f"Error creating interactive chart: {str(e)}")
        return None

def create_standard_chart(section, df):
    """Create standard charts with enhanced styling"""
    plt.style.use('seaborn-v0_8')
    
    if section in ["course_distribution", "gender_distribution", "age_distribution"]:
        return create_enhanced_pie_chart(df, section)
    elif section == "module_popularity":
        return create_enhanced_bar_chart(df, section)
    elif section == "registration_trends":
        return create_enhanced_line_chart(df, section)
    else:
        return None

def create_enhanced_pie_chart(df, section):
    """Create enhanced pie chart with better styling"""
    plt.figure(figsize=(10, 8))
    
    if section == "course_distribution":
        labels = df['course']
        sizes = df['student_count']
        title = "Course Distribution"
        colors = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4']
    elif section == "gender_distribution":
        labels = df['gender']
        sizes = df['student_count']
        title = "Gender Distribution"
        colors = ['#FF9999', '#66B2FF', '#99FF99']
    elif section == "age_distribution":
        labels = df['age_group']
        sizes = df['student_count']
        title = "Age Distribution"
        colors = ['#FFB347', '#87CEEB', '#DDA0DD', '#98FB98', '#F0E68C']
    
    wedges, texts, autotexts = plt.pie(sizes, labels=labels, autopct='%1.1f%%', 
                                      startangle=90, colors=colors[:len(labels)],
                                      explode=[0.02] * len(labels), shadow=True)
    
    # Style the text
    for text in texts:
        text.set_fontsize(11)
        text.set_fontweight('bold')
    
    for autotext in autotexts:
        autotext.set_color('white')
        autotext.set_fontweight('bold')
        autotext.set_fontsize(10)
    
    plt.title(title, fontsize=16, fontweight='bold', pad=20)
    plt.axis('equal')
    
    # Add total count
    total = sizes.sum()
    plt.figtext(0.5, 0.02, f"Total: {total}", ha='center', fontsize=12, fontweight='bold')
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{section}_enhanced_{timestamp}.png"
    filepath = os.path.join(CONFIG['reports_dir'], filename)
    plt.savefig(filepath, dpi=300, bbox_inches='tight')
    plt.close()
    
    return filepath

def create_enhanced_bar_chart(df, section):
    """Create enhanced bar chart"""
    plt.figure(figsize=(14, 8))
    
    if section == "module_popularity":
        plot_df = df.head(12)  # Show top 12 modules
        
        bars = plt.bar(plot_df['module_code'], plot_df['student_count'], 
                      color='steelblue', alpha=0.8, edgecolor='navy', linewidth=1)
        
        # Add value labels on bars
        for bar in bars:
            height = bar.get_height()
            plt.text(bar.get_x() + bar.get_width()/2., height + 0.5,
                    f'{int(height)}', ha='center', va='bottom', fontweight='bold')
        
        plt.title("Module Popularity (Top 12)", fontsize=16, fontweight='bold', pad=20)
        plt.xlabel("Module Code", fontsize=12, fontweight='bold')
        plt.ylabel("Number of Students", fontsize=12, fontweight='bold')
        plt.xticks(rotation=45, ha='right')
        
        # Add grid for better readability
        plt.grid(axis='y', alpha=0.3, linestyle='--')
        
        # Style the plot
        plt.gca().spines['top'].set_visible(False)
        plt.gca().spines['right'].set_visible(False)
        
        plt.tight_layout()
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{section}_enhanced_{timestamp}.png"
    filepath = os.path.join(CONFIG['reports_dir'], filename)
    plt.savefig(filepath, dpi=300, bbox_inches='tight')
    plt.close()
    
    return filepath

def create_enhanced_line_chart(df, section):
    """Create enhanced line chart"""
    plt.figure(figsize=(14, 8))
    
    if section == "registration_trends":
        plt.plot(df['registration_date'], df['registration_count'],
                marker='o', linewidth=2.5, markersize=6,
                color='steelblue', markerfacecolor='orange', markeredgecolor='navy')

        # Add trend line with error handling
        try:
            if len(df) >= 2:  # Need at least 2 points for a trend line
                x_numeric = np.array(range(len(df)))
                y_numeric = np.array(df['registration_count'])

                # Check for valid data (no NaN, no infinite values)
                if not (np.isnan(y_numeric).any() or np.isinf(y_numeric).any()):
                    # Use polyfit with error handling
                    z = np.polyfit(x_numeric, y_numeric, 1)
                    p = np.poly1d(z)
                    plt.plot(df['registration_date'], p(x_numeric), "--",
                            color='red', alpha=0.8, linewidth=2, label='Trend')
        except (np.linalg.LinAlgError, ValueError) as e:
            # If polyfit fails (SVD convergence, etc.), skip the trend line
            logger.warning(f"Could not generate trend line for registration_trends: {e}")

        plt.title("Registration Trends Over Time", fontsize=16, fontweight='bold', pad=20)
        plt.xlabel("Date", fontsize=12, fontweight='bold')
        plt.ylabel("Number of Registrations", fontsize=12, fontweight='bold')
        plt.xticks(rotation=45)

        # Add grid
        plt.grid(True, alpha=0.3, linestyle='--')

        # Style the plot
        plt.gca().spines['top'].set_visible(False)
        plt.gca().spines['right'].set_visible(False)

        # Add legend if trend line was added
        if plt.gca().get_legend_handles_labels()[0]:
            plt.legend()

        plt.tight_layout()
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{section}_enhanced_{timestamp}.png"
    filepath = os.path.join(CONFIG['reports_dir'], filename)
    plt.savefig(filepath, dpi=300, bbox_inches='tight')
    plt.close()
    
    return filepath

def generate_statistical_summary(df, section):
    """Generate statistical summary for numeric data"""
    styles = getSampleStyleSheet()
    
    if section == "grade_distribution" and 'grade' in df.columns:
        # Convert grades to numeric
        numeric_grades = pd.to_numeric(df['grade'], errors='coerce').dropna()
        
        if len(numeric_grades) > 0:
            stats = {
                'Mean': f"{numeric_grades.mean():.2f}",
                'Median': f"{numeric_grades.median():.2f}",
                'Std Dev': f"{numeric_grades.std():.2f}",
                'Min': f"{numeric_grades.min():.2f}",
                'Max': f"{numeric_grades.max():.2f}",
                'Count': f"{len(numeric_grades)}"
            }
            
            stats_text = " | ".join([f"{k}: {v}" for k, v in stats.items()])
            return Paragraph(stats_text, styles["Normal"])
    
    return None

def create_enhanced_data_table(df):
    """Create enhanced data table with better styling"""
    # Limit columns and rows for readability
    display_df = df.head(50)  # Show max 50 rows
    
    if len(df.columns) > 6:
        # Show only first 6 columns if too many
        display_df = display_df.iloc[:, :6]
    
    # Prepare data for table
    data = [display_df.columns.tolist()]
    for _, row in display_df.iterrows():
        formatted_row = []
        for value in row:
            if pd.isna(value):
                formatted_row.append("N/A")
            elif isinstance(value, float):
                formatted_row.append(f"{value:.2f}")
            else:
                formatted_row.append(str(value))
        data.append(formatted_row)
    
    # Create table with enhanced styling
    table = Table(data)
    table.setStyle(TableStyle([
        # Header styling
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # Data styling
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 1, colors.lightgrey),
        
        # Alternating row colors
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    
    return table

def generate_quality_section(quality_report, styles):
    """Generate data quality section for reports"""
    elements = []
    
    elements.append(Paragraph("Data Quality Report", styles["Heading2"]))
    elements.append(Spacer(1, 0.1*inch))
    
    # Summary of quality issues
    checks = quality_report.get('checks', {})
    
    if 'missing_data' in checks:
        missing = checks['missing_data']['students']
        total = missing['total_records']
        
        if total > 0:
            missing_rate = (missing['missing_emails'] + missing['missing_names'] + missing['missing_courses']) / (total * 3) * 100
            quality_text = f"Data completeness: {100 - missing_rate:.1f}% complete"
            elements.append(Paragraph(quality_text, styles["Normal"]))
    
    if 'duplicates' in checks:
        duplicates = checks['duplicates']['duplicate_emails']
        if duplicates > 0:
            elements.append(Paragraph(f"Warning: {duplicates} duplicate email addresses found", styles["Normal"]))
    
    if 'invalid_data' in checks:
        invalid = checks['invalid_data']
        if invalid['invalid_ages'] > 0 or invalid['invalid_emails'] > 0:
            elements.append(Paragraph(f"Data validation issues: {invalid['invalid_ages']} invalid ages, {invalid['invalid_emails']} invalid emails", styles["Normal"]))
    
    elements.append(Spacer(1, 0.25*inch))
    return elements

def generate_predictions_section(predictions, styles):
    """Generate predictive analytics section"""
    elements = []
    
    elements.append(Paragraph("Predictive Analytics", styles["Heading2"]))
    elements.append(Spacer(1, 0.1*inch))
    
    if 'error' in predictions:
        elements.append(Paragraph(f"Analysis unavailable: {predictions['error']}", styles["Normal"]))
    else:
        if 'model_accuracy' in predictions:
            accuracy_text = f"Model accuracy: {predictions['model_accuracy']:.2%}"
            elements.append(Paragraph(accuracy_text, styles["Normal"]))
        
        if 'high_risk_students' in predictions:
            risk_count = len(predictions['high_risk_students'])
            if risk_count > 0:
                risk_text = f"Students at high risk of dropout: {risk_count}"
                elements.append(Paragraph(risk_text, styles["Normal"]))
        
        if 'feature_importance' in predictions:
            importance = predictions['feature_importance']
            sorted_features = sorted(importance.items(), key=lambda x: x[1], reverse=True)
            
            elements.append(Paragraph("Most important factors:", styles["Heading4"]))
            for feature, score in sorted_features[:3]:
                feature_text = f"• {feature.replace('_', ' ').title()}: {score:.3f}"
                elements.append(Paragraph(feature_text, styles["Normal"]))
    
    elements.append(Spacer(1, 0.25*inch))
    return elements

# Enhanced data retrieval with new sections
def get_section_dataframe(section, start_date, end_date, filters=None):
    """Enhanced data retrieval for all sections including new analytics"""
    conn = get_reporting_db_connection()
    
    try:
        if section == "correlation_analysis":
            return get_correlation_data(conn, start_date, end_date, filters)
        elif section == "trend_analysis":
            return get_trend_data(conn, start_date, end_date, filters)
        elif section == "performance_benchmarks":
            return get_benchmark_data(conn, start_date, end_date, filters)
        else:
            # Use existing data retrieval logic
            return get_original_section_data_complete(conn, section, start_date, end_date, filters)
    
    finally:
        conn.close()

def get_correlation_data(conn, start_date, end_date, filters):
    """Get data for correlation analysis"""
    cursor = conn.cursor()
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='attendance_records'")
    
    if cursor.fetchone():
        attendance_table = 'attendance_records'
    else:
        return pd.DataFrame()
    
    # Fixed query with proper table aliases
    query = f"""
    SELECT s.student_id, s.age, s.course,
           COUNT(DISTINCT sm.module_code) as module_count,
           AVG(CASE WHEN sg.grade IS NOT NULL THEN CAST(sg.grade AS FLOAT) ELSE NULL END) as avg_grade,
           COUNT(ar.student_id) as attendance_records,
           SUM(CASE WHEN LOWER(ar.status) IN ('present', 'attended') THEN 1 ELSE 0 END) as present_count
    FROM students s
    LEFT JOIN student_modules sm ON s.student_id = sm.student_id
    LEFT JOIN student_grades sg ON s.student_id = sg.student_id
    LEFT JOIN {attendance_table} ar ON s.student_id = ar.student_id
    WHERE s.registration_datetime BETWEEN ? AND ?
    GROUP BY s.student_id, s.age, s.course
    """
    
    params = [f"{start_date} 00:00:00", f"{end_date} 23:59:59"]
    df = pd.read_sql_query(query, conn, params=params)
    
    # Calculate derived metrics
    df['attendance_rate'] = df.apply(
        lambda row: row['present_count'] / row['attendance_records'] if row['attendance_records'] > 0 else 0,
        axis=1
    )
    
    return df
    
def get_trend_data(conn, start_date, end_date, filters):
    """Get data for trend analysis"""
    query = """
    SELECT 
        date(registration_datetime) as date,
        course,
        COUNT(*) as daily_registrations,
        AVG(age) as avg_age
    FROM students
    WHERE registration_datetime BETWEEN ? AND ?
    GROUP BY date(registration_datetime), course
    ORDER BY date
    """
    
    params = [f"{start_date} 00:00:00", f"{end_date} 23:59:59"]
    return pd.read_sql_query(query, conn, params=params)

def get_benchmark_data(conn, start_date, end_date, filters):
    """Get data for performance benchmarks"""
    query = """
    SELECT 
        course,
        COUNT(*) as student_count,
        AVG(age) as avg_age,
        MIN(age) as min_age,
        MAX(age) as max_age
    FROM students
    WHERE registration_datetime BETWEEN ? AND ?
    GROUP BY course
    """
    
    params = [f"{start_date} 00:00:00", f"{end_date} 23:59:59"]
    return pd.read_sql_query(query, conn, params=params)

def get_original_section_data_complete(conn, section, start_date, end_date, filters):
    """Complete implementation of original section data retrieval"""
    
    if section == "student_overview":
        query = """
        SELECT COUNT(*) as total_students,
               COUNT(DISTINCT course) as total_courses,
               AVG(age) as avg_age
        FROM students
        WHERE registration_datetime BETWEEN ? AND ?
        """
        params = [f"{start_date} 00:00:00", f"{end_date} 23:59:59"]
        return pd.read_sql_query(query, conn, params=params)

    elif section == "student_list":
        query = """
        SELECT student_id, first_name, last_name, email_address, course,
               age, gender, registration_datetime
        FROM students
        WHERE registration_datetime BETWEEN ? AND ?
        ORDER BY registration_datetime DESC
        """
        params = [f"{start_date} 00:00:00", f"{end_date} 23:59:59"]
        return pd.read_sql_query(query, conn, params=params)

    elif section == "course_distribution":
        query = """
        SELECT course, COUNT(*) as student_count
        FROM students
        WHERE registration_datetime BETWEEN ? AND ?
        GROUP BY course
        ORDER BY student_count DESC
        """
        params = [f"{start_date} 00:00:00", f"{end_date} 23:59:59"]
        return pd.read_sql_query(query, conn, params=params)
    
    elif section == "gender_distribution":
        query = """
        SELECT gender, COUNT(*) as student_count
        FROM students
        WHERE registration_datetime BETWEEN ? AND ?
        AND gender IS NOT NULL
        GROUP BY gender
        ORDER BY student_count DESC
        """
        params = [f"{start_date} 00:00:00", f"{end_date} 23:59:59"]
        return pd.read_sql_query(query, conn, params=params)
    
    elif section == "age_distribution":
        query = """
        SELECT 
            CASE 
                WHEN age < 20 THEN 'Under 20'
                WHEN age BETWEEN 20 AND 25 THEN '20-25'
                WHEN age BETWEEN 26 AND 30 THEN '26-30'
                WHEN age BETWEEN 31 AND 35 THEN '31-35'
                ELSE 'Over 35'
            END as age_group,
            COUNT(*) as student_count
        FROM students
        WHERE registration_datetime BETWEEN ? AND ?
        AND age IS NOT NULL
        GROUP BY age_group
        ORDER BY student_count DESC
        """
        params = [f"{start_date} 00:00:00", f"{end_date} 23:59:59"]
        return pd.read_sql_query(query, conn, params=params)
    
    elif section == "registration_trends":
        query = """
        SELECT 
            DATE(registration_datetime) as registration_date,
            COUNT(*) as registration_count
        FROM students
        WHERE registration_datetime BETWEEN ? AND ?
        GROUP BY DATE(registration_datetime)
        ORDER BY registration_date
        """
        params = [f"{start_date} 00:00:00", f"{end_date} 23:59:59"]
        return pd.read_sql_query(query, conn, params=params)
    
    elif section == "module_popularity":
        # Check if student_modules table exists
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='student_modules'")
        if cursor.fetchone():
            query = """
            SELECT sm.module_code,
                   COALESCE(m.module_name, sm.module_code) as module_name,
                   COUNT(DISTINCT sm.student_id) as student_count
            FROM student_modules sm
            JOIN students s ON sm.student_id = s.student_id
            LEFT JOIN modules m ON sm.module_code = m.module_code
            WHERE s.registration_datetime BETWEEN ? AND ?
            GROUP BY sm.module_code, m.module_name
            ORDER BY student_count DESC
            """
            params = [f"{start_date} 00:00:00", f"{end_date} 23:59:59"]
            return pd.read_sql_query(query, conn, params=params)
        else:
            # Return empty DataFrame if table doesn't exist
            return pd.DataFrame(columns=['module_code', 'module_name', 'student_count'])
    
    elif section == "grade_distribution":
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='student_grades'")
        if cursor.fetchone():
            query = """
            SELECT sg.grade, COUNT(*) as student_count
            FROM student_grades sg
            JOIN students s ON sg.student_id = s.student_id
            WHERE s.registration_datetime BETWEEN ? AND ?
            AND sg.grade IS NOT NULL AND sg.grade != ''
            GROUP BY sg.grade
            ORDER BY sg.grade
            """
            params = [f"{start_date} 00:00:00", f"{end_date} 23:59:59"]
            return pd.read_sql_query(query, conn, params=params)
        else:
            return pd.DataFrame(columns=['grade', 'student_count'])
    
    elif section == "attendance_summary":
        # Check for attendance tables
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='attendance_records'")
        if cursor.fetchone():
            query = """
            SELECT
                CASE
                    WHEN LOWER(ar.status) IN ('present', 'attended') THEN 'Present'
                    WHEN LOWER(ar.status) IN ('absent', 'not attended') THEN 'Absent'
                    WHEN LOWER(ar.status) IN ('late', 'tardy') THEN 'Late'
                    WHEN LOWER(ar.status) IN ('excused', 'excused absence') THEN 'Excused'
                    ELSE ar.status
                END as status,
                COUNT(*) as count
            FROM attendance_records ar
            JOIN students s ON ar.student_id = s.student_id
            WHERE s.registration_datetime BETWEEN ? AND ?
            AND ar.status IS NOT NULL
            GROUP BY
                CASE
                    WHEN LOWER(ar.status) IN ('present', 'attended') THEN 'Present'
                    WHEN LOWER(ar.status) IN ('absent', 'not attended') THEN 'Absent'
                    WHEN LOWER(ar.status) IN ('late', 'tardy') THEN 'Late'
                    WHEN LOWER(ar.status) IN ('excused', 'excused absence') THEN 'Excused'
                    ELSE ar.status
                END
            ORDER BY count DESC
            """
            params = [f"{start_date} 00:00:00", f"{end_date} 23:59:59"]
            return pd.read_sql_query(query, conn, params=params)
        else:
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='student_attendance'")
            if cursor.fetchone():
                query = """
                SELECT sa.status, COUNT(*) as count
                FROM student_attendance sa
                JOIN students s ON sa.student_id = s.student_id
                WHERE s.registration_datetime BETWEEN ? AND ?
                AND sa.status IS NOT NULL
                GROUP BY sa.status
                ORDER BY count DESC
                """
                params = [f"{start_date} 00:00:00", f"{end_date} 23:59:59"]
                return pd.read_sql_query(query, conn, params=params)
            else:
                return pd.DataFrame(columns=['status', 'count'])
    
    # For any unrecognized sections, return empty DataFrame
    return pd.DataFrame()

# Enhanced scheduling with advanced features
class AdvancedScheduledReport:
    """Enhanced scheduled report with advanced features"""
    
    def __init__(self, template_name, schedule_config, recipients=None, 
                 conditions=None, next_run=None):
        self.template_name = template_name
        self.schedule_config = schedule_config  # Complex scheduling configuration
        self.recipients = recipients or []
        self.conditions = conditions or {}  # Conditional execution rules
        self.next_run = next_run or datetime.now().isoformat()
        self.created_at = datetime.now().isoformat()
        self.last_run = None
        self.run_count = 0
        self.is_active = True
    
    def to_dict(self):
        return {
            "template_name": self.template_name,
            "schedule_config": self.schedule_config,
            "recipients": self.recipients,
            "conditions": self.conditions,
            "next_run": self.next_run,
            "created_at": self.created_at,
            "last_run": self.last_run,
            "run_count": self.run_count,
            "is_active": self.is_active
        }
    
    @classmethod
    def from_dict(cls, data):
        report = cls(
            template_name=data["template_name"],
            schedule_config=data["schedule_config"],
            recipients=data.get("recipients", []),
            conditions=data.get("conditions", {}),
            next_run=data.get("next_run")
        )
        report.created_at = data.get("created_at", datetime.now().isoformat())
        report.last_run = data.get("last_run")
        report.run_count = data.get("run_count", 0)
        report.is_active = data.get("is_active", True)
        return report

def generate_enhanced_excel_report(template, filename, start_date, end_date, comparison_date=None):
    """Generate an enhanced Excel report with multiple sheets and formatting"""
    file_path = os.path.join(CONFIG['reports_dir'], f"{filename}.xlsx")
    
    try:
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # Summary sheet
            summary_data = {
                'Report Name': [template.name],
                'Generated': [datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                'Period': [f"{start_date} to {end_date}"],
                'Template Version': [template.version],
                'Security Level': [template.security_level]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Generate data for each section
            for section in template.sections:
                if section in AVAILABLE_SECTIONS:
                    try:
                        df = get_section_dataframe(section, start_date, end_date, template.filters)
                        if df is not None and not df.empty:
                            sheet_name = section[:31]  # Excel sheet name limit
                            df.to_excel(writer, sheet_name=sheet_name, index=False)
                    except Exception as e:
                        logger.error(f"Error adding section {section} to Excel: {str(e)}")
            
            # Format the Excel file
            workbook = writer.book
            
            # Add formatting to summary sheet
            if 'Summary' in workbook.sheetnames:
                summary_sheet = workbook['Summary']
                from openpyxl.styles import Font, PatternFill
                
                # Header formatting
                for cell in summary_sheet[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        logger.info(f"Enhanced Excel report generated: {file_path}")
        return file_path
        
    except Exception as e:
        logger.error(f"Error generating Excel report: {str(e)}")
        raise

def generate_interactive_report(template, filename, start_date, end_date):
    """Generate an interactive HTML report with Plotly charts"""
    file_path = os.path.join(CONFIG['reports_dir'], f"{filename}.html")
    
    try:
        # Collect data for all sections
        data_dict = {}
        
        for section in template.sections:
            if section in AVAILABLE_SECTIONS:
                try:
                    df = get_section_dataframe(section, start_date, end_date, template.filters)
                    if df is not None and not df.empty:
                        data_dict[section] = df
                except Exception as e:
                    logger.error(f"Error getting data for section {section}: {str(e)}")
        
        # Create interactive dashboard
        dashboard_path = AdvancedVisualization.create_interactive_dashboard(data_dict)
        
        if dashboard_path and os.path.exists(dashboard_path):
            # Read the generated HTML and enhance it
            with open(dashboard_path, 'r') as f:
                html_content = f.read()
            
            # Add custom styling and metadata
            enhanced_html = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <title>{template.name} - Interactive Report</title>
                <meta charset="utf-8">
                <style>
                    body {{ font-family: Arial, sans-serif; margin: 20px; }}
                    .header {{ text-align: center; margin-bottom: 30px; }}
                    .metadata {{ background: #f5f5f5; padding: 15px; border-radius: 5px; margin-bottom: 20px; }}
                </style>
            </head>
            <body>
                <div class="header">
                    <h1>{template.name}</h1>
                    <p>Interactive Report - Generated {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                </div>
                <div class="metadata">
                    <p><strong>Report Period:</strong> {start_date} to {end_date}</p>
                    <p><strong>Template Version:</strong> {template.version}</p>
                    <p><strong>Security Level:</strong> {template.security_level.title()}</p>
                </div>
                {html_content.split('<body>')[1] if '<body>' in html_content else html_content}
            </body>
            </html>
            """
            
            with open(file_path, 'w') as f:
                f.write(enhanced_html)
            
            # Clean up temporary dashboard file
            if dashboard_path != file_path:
                os.remove(dashboard_path)
        
        logger.info(f"Interactive report generated: {file_path}")
        return file_path
        
    except Exception as e:
        logger.error(f"Error generating interactive report: {str(e)}")
        raise

# Enhanced system management functions
def run_system_maintenance():
    """Run system maintenance tasks"""
    logger.info("Running system maintenance...")
    
    # Clean up old report files
    cleanup_old_reports()
    
    # Clean up cache
    CacheManager.cleanup_cache()
    
    # Run data quality checks
    quality_report = DataQualityMonitor.run_quality_checks()
    
    # Log maintenance completion
    logger.info("System maintenance completed")
    
    return quality_report

def cleanup_old_reports(days_to_keep=30):
    """Clean up old report files"""
    cutoff_date = datetime.now() - timedelta(days=days_to_keep)
    
    for root, dirs, files in os.walk(CONFIG['reports_dir']):
        for file in files:
            file_path = os.path.join(root, file)
            if os.path.getctime(file_path) < cutoff_date.timestamp():
                try:
                    os.remove(file_path)
                    logger.info(f"Removed old report file: {file}")
                except Exception as e:
                    logger.error(f"Error removing file {file}: {str(e)}")

def load_scheduled_reports():
    """Load scheduled reports from file"""
    try:
        with open(CONFIG['scheduled_reports_file'], 'r') as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return []

def save_scheduled_reports(scheduled_reports):
    """Save scheduled reports to file"""
    with open(CONFIG['scheduled_reports_file'], 'w') as f:
        json.dump(scheduled_reports, f, indent=4)

def start_scheduler():
    """Start the background scheduler for automatic reports"""
    def run_scheduler():
        while True:
            try:
                schedule.run_pending()
                time.sleep(60)  # Check every minute
            except Exception as e:
                logger.error(f"Scheduler error: {str(e)}")
                time.sleep(60)
    
    # Load and schedule all reports
    def schedule_report(report_data):
        """Schedule a single report"""
        def run_report():
            try:
                template_name = report_data['template_name']
                recipients = report_data.get('recipients', [])
                
                # Generate report
                end_date = datetime.now().strftime("%Y-%m-%d")
                start_date = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
                
                report_path = generate_report(template_name, start_date, end_date, 'pdf')
                
                if report_path:
                    # Update run statistics
                    report_data['last_run'] = datetime.now().isoformat()
                    report_data['run_count'] = report_data.get('run_count', 0) + 1
                    
                    logger.info(f"Scheduled report '{template_name}' generated successfully")
                else:
                    logger.error(f"Failed to generate scheduled report '{template_name}'")
                    
            except Exception as e:
                logger.error(f"Error running scheduled report '{template_name}': {str(e)}")
    
    # Schedule all reports
    scheduled_reports = load_scheduled_reports()
    for report_data in scheduled_reports:
        if not report_data['schedule_config'].get('enabled', True):
            continue
            
        frequency = report_data['schedule_config']['frequency']
        hour = report_data['schedule_config']['hour']
        
        if frequency == 'daily':
            schedule.every().day.at(f"{hour:02d}:00").do(lambda r=report_data: schedule_report(r))
        elif frequency == 'weekly':
            schedule.every().monday.at(f"{hour:02d}:00").do(lambda r=report_data: schedule_report(r))
        elif frequency == 'monthly':
            schedule.every().month.at(f"{hour:02d}:00").do(lambda r=report_data: schedule_report(r))
    
    # Start scheduler in background thread
    scheduler_thread = threading.Thread(target=run_scheduler, daemon=True)
    scheduler_thread.start()
    
    logger.info(f"Background scheduler started with {len(scheduled_reports)} scheduled reports")

# Enhanced menu system with all new features
def display_enhanced_reporting_menu():
    """Display the advanced reporting menu with all features"""
    while True:
        print("\n" + "="*60)
        print("🔬 ADVANCED STUDENT REPORTING SYSTEM")
        print("="*60)
        print("📊 REPORT MANAGEMENT:")
        print("1.  Create Report Template")
        print("2.  View Templates")
        print("3.  Generate Report")
        print("4.  Generate Interactive Report")
        print("5.  Delete Template")
        print("")
        print("⏰ SCHEDULING:")
        print("6.  Schedule Regular Report")
        print("7.  View Scheduled Reports")
        print("8.  Manage Schedule")
        print("")
        print("🔍 ANALYTICS & INSIGHTS:")
        print("9.  Data Quality Dashboard")
        print("10. Predictive Analytics")
        print("11. Anomaly Detection")
        print("12. Correlation Analysis")
        print("")
        print("⚙️  SYSTEM:")
        print("17. System Maintenance")
        print("18. Performance Monitor")
        print("19. Export System Logs")
        print("")
        print("🌐 API SERVER:")
        print("20. Start REST API Server")
        print("")
        print("21. Return to Main Menu")
        print("="*60)
        
        choice = input("\nEnter your choice (1-21): ").strip()
        
        try:
            if choice == '1':
                create_advanced_template_menu()
            elif choice == '2':
                view_templates_menu()
                input("\nPress Enter to continue...")
            elif choice == '3':
                generate_advanced_report_menu()
            elif choice == '4':
                generate_interactive_report_menu()
            elif choice == '5':
                delete_template_menu()
            elif choice == '6':
                schedule_advanced_report_menu()
            elif choice == '7':
                view_scheduled_reports_menu()
            elif choice == '8':
                manage_schedule_menu()
            elif choice == '9':
                show_data_quality_dashboard()
            elif choice == '10':
                show_predictive_analytics()
            elif choice == '11':
                show_anomaly_detection()
            elif choice == '12':
                show_correlation_analysis()
            elif choice == '17':
                run_maintenance_menu()
            elif choice == '18':
                show_performance_monitor()
            elif choice == '19':
                export_logs_menu()
            elif choice == '20':
                start_api_server_menu()
            elif choice == '21':
                print("Returning to main menu...")
                break
            else:
                print("❌ Invalid choice. Please try again.")
                
        except KeyboardInterrupt:
            print("\n\nOperation cancelled by user.")
        except Exception as e:
            print(f"❌ An error occurred: {str(e)}")
            logger.error(f"Menu error: {str(e)}")

def create_advanced_template_menu():
    """Enhanced template creation with all new features"""
    print("\n🔧 CREATE ADVANCED REPORT TEMPLATE")
    print("="*50)
    
    # Basic template info
    name = input("📝 Template name: ").strip()
    if not name:
        print("❌ Template name cannot be empty.")
        return
    
    if get_template(name):
        print(f"❌ Template '{name}' already exists.")
        return
    
    description = input("📋 Template description: ").strip()
    
    # Security level
    print("\n🔒 Security Level:")
    print("1. Normal")
    print("2. Confidential") 
    print("3. Restricted")
    
    security_choice = input("Select security level (1-3): ").strip()
    security_levels = {'1': 'normal', '2': 'confidential', '3': 'restricted'}
    security_level = security_levels.get(security_choice, 'normal')
    
    # Visualization type
    print("\n📊 Visualization Type:")
    print("1. Standard (PNG charts)")
    print("2. Advanced (Enhanced styling)")
    print("3. Interactive (HTML charts)")
    
    viz_choice = input("Select visualization type (1-3): ").strip()
    viz_types = {'1': 'standard', '2': 'advanced', '3': 'interactive'}
    visualization_type = viz_types.get(viz_choice, 'standard')
    
    # Select sections
    print("\n📈 Available Report Sections:")
    for i, section in enumerate(AVAILABLE_SECTIONS, 1):
        print(f"{i:2d}. {section.replace('_', ' ').title()}")
    
    sections = []
    while True:
        section_input = input("\nSelect sections (comma-separated numbers, or 'all'): ").strip()
        if section_input.lower() == 'all':
            sections = AVAILABLE_SECTIONS.copy()
            break
        
        try:
            indices = [int(idx.strip()) - 1 for idx in section_input.split(',')]
            sections = [AVAILABLE_SECTIONS[idx] for idx in indices if 0 <= idx < len(AVAILABLE_SECTIONS)]
            if sections:
                break
            else:
                print("❌ No valid sections selected.")
        except (ValueError, IndexError):
            print("❌ Invalid input. Please enter comma-separated numbers.")
    
    # Filters
    filters = {}
    if input("\n🔍 Add filters? (y/n): ").lower() == 'y':
        if input("Filter by course? (y/n): ").lower() == 'y':
            course = input("Enter course (CS/DS): ").upper().strip()
            if course in ['CS', 'DS']:
                filters['course'] = course
        
        if input("Filter by date range? (y/n): ").lower() == 'y':
            days = input("Number of days to include (default 30): ").strip()
            try:
                filters['date_range_days'] = int(days) if days else 30
            except ValueError:
                filters['date_range_days'] = 30
    
    # Create template
    template = ReportTemplate(
        name=name,
        description=description,
        sections=sections,
        filters=filters,
        visualization_type=visualization_type,
        security_level=security_level
    )
    
    save_template(template)
    print(f"\n✅ Template '{name}' created successfully!")
    print(f"   Security Level: {security_level.title()}")
    print(f"   Visualization: {visualization_type.title()}")
    print(f"   Sections: {len(sections)} selected")

def show_data_quality_dashboard():
    """Display data quality dashboard"""
    print("\n🔍 DATA QUALITY DASHBOARD")
    print("="*50)
    
    try:
        quality_report = DataQualityMonitor.run_quality_checks()
        
        print(f"📅 Last Check: {quality_report['timestamp']}")
        print()
        
        checks = quality_report.get('checks', {})
        
        # Missing data summary
        if 'missing_data' in checks:
            missing = checks['missing_data']['students']
            total = missing['total_records']
            print("📊 MISSING DATA ANALYSIS:")
            print(f"   Total Records: {total}")
            print(f"   Missing Emails: {missing['missing_emails']}")
            print(f"   Missing Names: {missing['missing_names']}")
            print(f"   Missing Courses: {missing['missing_courses']}")
            
            if total > 0:
                completeness = ((total * 3) - (missing['missing_emails'] + missing['missing_names'] + missing['missing_courses'])) / (total * 3) * 100
                print(f"   Data Completeness: {completeness:.1f}%")
                
                if completeness < 90:
                    print("   ⚠️  Warning: Data completeness below 90%")
                else:
                    print("   ✅ Good data completeness")
            print()
        
        # Duplicates
        if 'duplicates' in checks:
            duplicates = checks['duplicates']
            print("👥 DUPLICATE ANALYSIS:")
            print(f"   Duplicate Emails: {duplicates['duplicate_emails']}")
            
            if duplicates['duplicate_emails'] > 0:
                print("   📋 Duplicate Email Details:")
                for detail in duplicates.get('duplicate_email_details', [])[:5]:
                    print(f"      {detail['email']}: {detail['count']} occurrences")
                print("   ⚠️  Action Required: Review duplicate emails")
            else:
                print("   ✅ No duplicate emails found")
            print()
        
        # Invalid data
        if 'invalid_data' in checks:
            invalid = checks['invalid_data']
            print("❌ INVALID DATA ANALYSIS:")
            print(f"   Invalid Ages: {invalid['invalid_ages']}")
            print(f"   Invalid Emails: {invalid['invalid_emails']}")
            
            if invalid['invalid_ages'] > 0 or invalid['invalid_emails'] > 0:
                print("   ⚠️  Action Required: Clean invalid data")
            else:
                print("   ✅ No invalid data found")
            print()
        
        # Data freshness
        if 'data_freshness' in checks:
            freshness = checks['data_freshness']
            if freshness['last_registration_date']:
                days_since = freshness['days_since_last_registration']
                print("📆 DATA FRESHNESS:")
                print(f"   Last Registration: {freshness['last_registration_date']}")
                print(f"   Days Since Last: {days_since}")
                
                if days_since > 7:
                    print("   ⚠️  Warning: No recent registrations")
                else:
                    print("   ✅ Recent data available")
            else:
                print("📆 DATA FRESHNESS:")
                print("   ❌ No registration data found")
        
    except Exception as e:
        print(f"❌ Error generating quality report: {str(e)}")
    
    input("\nPress Enter to continue...")

def show_predictive_analytics():
    """Display predictive analytics dashboard"""
    print("\n🔮 PREDICTIVE ANALYTICS")
    print("="*40)
    
    print("🎯 Analyzing dropout risk patterns...")
    
    try:
        predictions = PredictiveAnalytics.predict_dropout_risk()
        
        if 'error' in predictions:
            print(f"❌ Analysis unavailable: {predictions['error']}")
        else:
            print("\n📊 DROPOUT RISK ANALYSIS:")
            
            if 'model_accuracy' in predictions:
                accuracy = predictions['model_accuracy'] * 100
                print(f"   Model Accuracy: {accuracy:.1f}%")
                
                if accuracy > 80:
                    print("   ✅ High confidence predictions")
                elif accuracy > 60:
                    print("   ⚠️  Moderate confidence predictions")
                else:
                    print("   ❌ Low confidence - more data needed")
            
            if 'total_students_analyzed' in predictions:
                print(f"   Students Analyzed: {predictions['total_students_analyzed']}")
            
            if 'high_risk_students' in predictions:
                high_risk = predictions['high_risk_students']
                print(f"   High Risk Students: {len(high_risk)}")
                
                if high_risk:
                    print("\n   🚨 Students requiring attention:")
                    for student in high_risk[:10]:  # Show top 10
                        print(f"      Student ID: {student['student_id']} (Risk: {student['risk_score']:.2%})")
                    
                    if len(high_risk) > 10:
                        print(f"      ... and {len(high_risk) - 10} more")
            
            if 'feature_importance' in predictions:
                importance = predictions['feature_importance']
                sorted_features = sorted(importance.items(), key=lambda x: x[1], reverse=True)
                
                print("\n   📈 Most Important Risk Factors:")
                for feature, score in sorted_features:
                    feature_name = feature.replace('_', ' ').title()
                    print(f"      {feature_name}: {score:.3f}")
    
    except Exception as e:
        print(f"❌ Error in predictive analysis: {str(e)}")
        logger.error(f"Predictive analytics error: {str(e)}")
    
    input("\nPress Enter to continue...")

def show_anomaly_detection():
    """Display anomaly detection results"""
    print("\n🔍 ANOMALY DETECTION")
    print("="*35)
    
    print("🔎 Scanning for unusual patterns...")
    
    try:
        anomalies = PredictiveAnalytics.detect_anomalies()
        
        if 'error' in anomalies:
            print(f"❌ Analysis unavailable: {anomalies['error']}")
        else:
            print("\n📊 ANOMALY DETECTION RESULTS:")
            
            anomaly_count = anomalies.get('total_anomalies', 0)
            anomaly_rate = anomalies.get('anomaly_rate', 0)
            
            print(f"   Anomalous Students: {anomaly_count}")
            print(f"   Anomaly Rate: {anomaly_rate:.2f}%")
            
            if anomaly_rate > 15:
                print("   ⚠️  High anomaly rate - investigate data quality")
            elif anomaly_rate > 5:
                print("   ⚠️  Moderate anomalies detected")
            else:
                print("   ✅ Normal anomaly rate")
            
            if 'anomalous_students' in anomalies and anomalies['anomalous_students']:
                print("\n   🔍 Anomalous Student Profiles:")
                
                for student in anomalies['anomalous_students'][:10]:
                    print(f"      Student ID: {student['student_id']}")
                    print(f"         Age: {student['age']}")
                    print(f"         Modules: {student['unique_modules']}")
                    print(f"         Avg Grade: {student.get('avg_grade', 'N/A')}")
                    print()
                
                if len(anomalies['anomalous_students']) > 10:
                    remaining = len(anomalies['anomalous_students']) - 10
                    print(f"      ... and {remaining} more anomalous profiles")
    
    except Exception as e:
        print(f"❌ Error in anomaly detection: {str(e)}")
        logger.error(f"Anomaly detection error: {str(e)}")
    
    input("\nPress Enter to continue...")

def show_correlation_analysis():
    """Display correlation analysis"""
    print("\n📈 CORRELATION ANALYSIS")
    print("="*40)
    
    try:
        conn = get_reporting_db_connection()
        chart_path = AdvancedVisualization.create_correlation_matrix(conn)
        
        if chart_path:
            print("✅ Correlation matrix generated successfully!")
            print(f"📊 Chart saved to: {chart_path}")
            
            # Open the chart if possible
            try:
                import webbrowser
                webbrowser.open(f"file://{os.path.abspath(chart_path)}")
                print("🔗 Chart opened in default image viewer")
            except Exception:
                print("💡 Please open the chart file manually to view correlations")
        else:
            print("❌ Unable to generate correlation matrix - insufficient data")
    
    except Exception as e:
        print(f"❌ Error generating correlation analysis: {str(e)}")
        logger.error(f"Correlation analysis error: {str(e)}")
    
    input("\nPress Enter to continue...")

def start_api_server_menu():
    """Start the REST API server"""
    print("\n🌐 REST API SERVER")
    print("="*30)
    
    print("API Endpoints available:")
    print("• POST /api/login - User authentication")
    print("• GET  /api/templates - List templates")
    print("• POST /api/templates - Create template")
    print("• POST /api/reports/generate - Generate report")
    print("• GET  /api/analytics/quality - Data quality")
    print("• GET  /api/analytics/predictions - Predictions")
    print("• GET  /api/analytics/anomalies - Anomaly detection")
    
    if input("\nStart API server? (y/n): ").lower() == 'y':
        host = input("Host (localhost): ").strip() or "localhost"
        port_input = input("Port (5000): ").strip()
        port = int(port_input) if port_input else 5000
        
        print(f"\n🚀 Starting API server on http://{host}:{port}")
        print("Press Ctrl+C to stop the server")
        
        try:
            app.run(host=host, port=port, debug=False)
        except KeyboardInterrupt:
            print("\n🛑 API server stopped")

def generate_enhanced_excel_report(template, filename, start_date, end_date, comparison_date=None):
    """Generate an enhanced Excel report with multiple sheets and formatting"""
    file_path = os.path.join(CONFIG['reports_dir'], f"{filename}.xlsx")
    
    try:
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # Summary sheet
            summary_data = {
                'Report Name': [template.name],
                'Generated': [datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                'Period': [f"{start_date} to {end_date}"],
                'Template Version': [template.version],
                'Security Level': [template.security_level]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Generate data for each section
            for section in template.sections:
                if section in AVAILABLE_SECTIONS:
                    try:
                        df = get_section_dataframe(section, start_date, end_date, template.filters)
                        if df is not None and not df.empty:
                            sheet_name = section[:31]  # Excel sheet name limit
                            df.to_excel(writer, sheet_name=sheet_name, index=False)
                    except Exception as e:
                        logger.error(f"Error adding section {section} to Excel: {str(e)}")
            
            # Format the Excel file
            workbook = writer.book
            
            # Add formatting to summary sheet
            if 'Summary' in workbook.sheetnames:
                summary_sheet = workbook['Summary']
                from openpyxl.styles import Font, PatternFill
                
                # Header formatting
                for cell in summary_sheet[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        logger.info(f"Enhanced Excel report generated: {file_path}")
        return file_path
        
    except Exception as e:
        logger.error(f"Error generating Excel report: {str(e)}")
        raise

def generate_interactive_report(template, filename, start_date, end_date):
    """Generate an interactive HTML report with Plotly charts"""
    file_path = os.path.join(CONFIG['reports_dir'], f"{filename}.html")
    
    try:
        # Collect data for all sections
        data_dict = {}
        
        for section in template.sections:
            if section in AVAILABLE_SECTIONS:
                try:
                    df = get_section_dataframe(section, start_date, end_date, template.filters)
                    if df is not None and not df.empty:
                        data_dict[section] = df
                except Exception as e:
                    logger.error(f"Error getting data for section {section}: {str(e)}")
        
        # Create interactive dashboard
        dashboard_path = AdvancedVisualization.create_interactive_dashboard(data_dict)
        
        if dashboard_path and os.path.exists(dashboard_path):
            # Read the generated HTML and enhance it
            with open(dashboard_path, 'r') as f:
                html_content = f.read()
            
            # Add custom styling and metadata
            enhanced_html = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <title>{template.name} - Interactive Report</title>
                <meta charset="utf-8">
                <style>
                    body {{ font-family: Arial, sans-serif; margin: 20px; }}
                    .header {{ text-align: center; margin-bottom: 30px; }}
                    .metadata {{ background: #f5f5f5; padding: 15px; border-radius: 5px; margin-bottom: 20px; }}
                </style>
            </head>
            <body>
                <div class="header">
                    <h1>{template.name}</h1>
                    <p>Interactive Report - Generated {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                </div>
                <div class="metadata">
                    <p><strong>Report Period:</strong> {start_date} to {end_date}</p>
                    <p><strong>Template Version:</strong> {template.version}</p>
                    <p><strong>Security Level:</strong> {template.security_level.title()}</p>
                </div>
                {html_content.split('<body>')[1] if '<body>' in html_content else html_content}
            </body>
            </html>
            """
            
            with open(file_path, 'w') as f:
                f.write(enhanced_html)
            
            # Clean up temporary dashboard file
            if dashboard_path != file_path:
                os.remove(dashboard_path)
        
        logger.info(f"Interactive report generated: {file_path}")
        return file_path
        
    except Exception as e:
        logger.error(f"Error generating interactive report: {str(e)}")
        raise

def view_templates_menu():
    """Display all available templates"""
    print("\n📋 REPORT TEMPLATES")
    print("="*40)
    
    templates = load_templates()
    
    if not templates:
        print("No templates found.")
        return
    
    for i, template_data in enumerate(templates, 1):
        print(f"{i}. {template_data['name']}")
        print(f"   Description: {template_data['description']}")
        print(f"   Sections: {len(template_data['sections'])} sections")
        print(f"   Security: {template_data.get('security_level', 'normal').title()}")
        print(f"   Visualization: {template_data.get('visualization_type', 'standard').title()}")
        print(f"   Version: {template_data.get('version', '1.0')}")
        print(f"   Created: {template_data.get('created_at', 'Unknown')}")
        print()

def generate_advanced_report_menu():
    """Enhanced report generation menu"""
    print("\n📊 GENERATE ADVANCED REPORT")
    print("="*45)
    
    templates = load_templates()
    if not templates:
        print("No templates available. Create a template first.")
        return
    
    # Select template
    print("Available templates:")
    for i, template_data in enumerate(templates, 1):
        print(f"{i}. {template_data['name']} ({template_data.get('visualization_type', 'standard')})")
    
    try:
        choice = int(input(f"\nSelect template (1-{len(templates)}): ")) - 1
        if choice < 0 or choice >= len(templates):
            print("❌ Invalid selection")
            return
        
        template_name = templates[choice]['name']
        
        # Date range
        print(f"\n📅 Date Range (leave empty for last 30 days):")
        start_date = input("Start date (YYYY-MM-DD): ").strip()
        end_date = input("End date (YYYY-MM-DD): ").strip()
        
        if not end_date:
            end_date = datetime.now().strftime("%Y-%m-%d")
        if not start_date:
            start_dt = datetime.strptime(end_date, "%Y-%m-%d") - timedelta(days=30)
            start_date = start_dt.strftime("%Y-%m-%d")
        
        # Format selection
        print(f"\n📄 Output Format:")
        print("1. PDF Report")
        print("2. Excel Spreadsheet")
        print("3. Interactive HTML")
        
        format_choice = input("Select format (1-3): ").strip()
        formats = {'1': 'pdf', '2': 'excel', '3': 'interactive'}
        output_format = formats.get(format_choice, 'pdf')
        
        # Generate report
        print(f"\n🔄 Generating {output_format.upper()} report...")
        
        try:
            report_path = generate_report(template_name, start_date, end_date, output_format)
            
            if report_path:
                print(f"✅ Report generated successfully!")
                print(f"📁 File location: {report_path}")
                
                # Offer to open the report
                if input("Open report now? (y/n): ").lower() == 'y':
                    import webbrowser
                    webbrowser.open(f"file://{os.path.abspath(report_path)}")
            else:
                print("❌ Failed to generate report")
                
        except Exception as e:
            print(f"❌ Error generating report: {str(e)}")
            
    except (ValueError, IndexError):
        print("❌ Invalid selection")

def generate_interactive_report_menu():
    """Generate interactive HTML report menu"""
    print("\n🌐 GENERATE INTERACTIVE REPORT")
    print("="*45)
    
    templates = load_templates()
    if not templates:
        print("No templates available. Create a template first.")
        return
    
    # Filter templates that support interactive visualization
    interactive_templates = [t for t in templates if t.get('visualization_type') in ['interactive', 'advanced']]
    
    if not interactive_templates:
        print("No templates configured for interactive visualization.")
        print("Create a template with 'Interactive' or 'Advanced' visualization type.")
        return
    
    print("Templates with interactive support:")
    for i, template_data in enumerate(interactive_templates, 1):
        print(f"{i}. {template_data['name']} ({template_data.get('visualization_type', 'standard')})")
    
    try:
        choice = int(input(f"\nSelect template (1-{len(interactive_templates)}): ")) - 1
        if choice < 0 or choice >= len(interactive_templates):
            print("❌ Invalid selection")
            return
        
        template_name = interactive_templates[choice]['name']
        
        # Generate with default date range
        end_date = datetime.now().strftime("%Y-%m-%d")
        start_date = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
        
        print(f"\n🔄 Generating interactive report for last 30 days...")
        
        report_path = generate_report(template_name, start_date, end_date, 'interactive')
        
        if report_path:
            print(f"✅ Interactive report generated!")
            print(f"📁 File location: {report_path}")
            
            # Automatically open in browser
            import webbrowser
            webbrowser.open(f"file://{os.path.abspath(report_path)}")
            print("🔗 Report opened in default browser")
        else:
            print("❌ Failed to generate interactive report")
            
    except (ValueError, IndexError):
        print("❌ Invalid selection")

def delete_template_menu():
    """Delete a report template"""
    print("\n🗑️  DELETE TEMPLATE")
    print("="*30)
    
    templates = load_templates()
    if not templates:
        print("No templates to delete.")
        return
    
    print("Available templates:")
    for i, template_data in enumerate(templates, 1):
        print(f"{i}. {template_data['name']}")
    
    try:
        choice = int(input(f"\nSelect template to delete (1-{len(templates)}): ")) - 1
        if choice < 0 or choice >= len(templates):
            print("❌ Invalid selection")
            return
        
        template_name = templates[choice]['name']
        
        confirm = input(f"⚠️  Are you sure you want to delete '{template_name}'? (yes/no): ").strip().lower()
        if confirm == 'yes':
            del templates[choice]
            
            # Save updated templates
            with open(os.path.join(CONFIG['templates_dir'], "templates.json"), 'w') as f:
                json.dump(templates, f, indent=4)
            
            print(f"✅ Template '{template_name}' deleted successfully!")
        else:
            print("❌ Deletion cancelled")
            
    except (ValueError, IndexError):
        print("❌ Invalid selection")

def schedule_advanced_report_menu():
    """Schedule automatic report generation"""
    print("\n⏰ SCHEDULE AUTOMATIC REPORT")
    print("="*45)
    
    templates = load_templates()
    if not templates:
        print("No templates available for scheduling.")
        return
    
    # Select template
    print("Available templates:")
    for i, template_data in enumerate(templates, 1):
        print(f"{i}. {template_data['name']}")
    
    try:
        choice = int(input(f"\nSelect template (1-{len(templates)}): ")) - 1
        if choice < 0 or choice >= len(templates):
            print("❌ Invalid selection")
            return
        
        template_name = templates[choice]['name']
        
        # Schedule configuration
        print(f"\n📅 Schedule Configuration:")
        print("1. Daily")
        print("2. Weekly")
        print("3. Monthly")
        
        schedule_choice = input("Select frequency (1-3): ").strip()
        schedules = {'1': 'daily', '2': 'weekly', '3': 'monthly'}
        frequency = schedules.get(schedule_choice, 'weekly')
        
        # Time configuration
        hour = input("Hour to run (0-23, default 9): ").strip()
        hour = int(hour) if hour.isdigit() and 0 <= int(hour) <= 23 else 9
        
        # Recipients
        recipients = []
        print(f"\n📧 Email Recipients:")
        while True:
            email = input("Enter email address (or press Enter to finish): ").strip()
            if not email:
                break
            if '@' in email:
                recipients.append(email)
            else:
                print("❌ Invalid email format")
        
        if not recipients:
            print("⚠️  No email recipients specified. Report will be generated but not sent.")
        
        # Create schedule configuration
        schedule_config = {
            'frequency': frequency,
            'hour': hour,
            'enabled': True,
            'last_run': None,
            'next_run': None
        }
        
        # Save scheduled report
        scheduled_report = AdvancedScheduledReport(
            template_name=template_name,
            schedule_config=schedule_config,
            recipients=recipients
        )
        
        scheduled_reports = load_scheduled_reports()
        scheduled_reports.append(scheduled_report.to_dict())
        save_scheduled_reports(scheduled_reports)
        
        print(f"✅ Report '{template_name}' scheduled for {frequency} generation at {hour}:00")
        if recipients:
            print(f"📧 Will be sent to: {', '.join(recipients)}")
        
    except (ValueError, IndexError):
        print("❌ Invalid input")

def view_scheduled_reports_menu():
    """View all scheduled reports"""
    print("\n📅 SCHEDULED REPORTS")
    print("="*35)
    
    scheduled_reports = load_scheduled_reports()
    
    if not scheduled_reports:
        print("No scheduled reports found.")
        return
    
    for i, report_data in enumerate(scheduled_reports, 1):
        config = report_data['schedule_config']
        print(f"{i}. {report_data['template_name']}")
        print(f"   Frequency: {config['frequency'].title()}")
        print(f"   Time: {config['hour']}:00")
        print(f"   Status: {'Enabled' if config.get('enabled', True) else 'Disabled'}")
        print(f"   Recipients: {len(report_data.get('recipients', []))}")
        print(f"   Last Run: {report_data.get('last_run', 'Never')}")
        print(f"   Run Count: {report_data.get('run_count', 0)}")
        print()

def manage_schedule_menu():
    """Manage scheduled reports"""
    print("\n⚙️  MANAGE SCHEDULED REPORTS")
    print("="*40)
    
    scheduled_reports = load_scheduled_reports()
    
    if not scheduled_reports:
        print("No scheduled reports to manage.")
        return
    
    print("Scheduled reports:")
    for i, report_data in enumerate(scheduled_reports, 1):
        status = "Enabled" if report_data['schedule_config'].get('enabled', True) else "Disabled"
        print(f"{i}. {report_data['template_name']} ({status})")
    
    try:
        choice = int(input(f"\nSelect report to manage (1-{len(scheduled_reports)}): ")) - 1
        if choice < 0 or choice >= len(scheduled_reports):
            print("❌ Invalid selection")
            return
        
        report = scheduled_reports[choice]
        
        print(f"\n📋 Managing: {report['template_name']}")
        print("1. Enable/Disable")
        print("2. Modify Schedule")
        print("3. Update Recipients")
        print("4. Delete Schedule")
        print("5. Run Now")
        
        action = input("Select action (1-5): ").strip()
        
        if action == '1':
            current_status = report['schedule_config'].get('enabled', True)
            report['schedule_config']['enabled'] = not current_status
            new_status = "Enabled" if not current_status else "Disabled"
            print(f"✅ Schedule {new_status.lower()}")
            
        elif action == '2':
            print("New frequency:")
            print("1. Daily")
            print("2. Weekly") 
            print("3. Monthly")
            
            freq_choice = input("Select (1-3): ").strip()
            frequencies = {'1': 'daily', '2': 'weekly', '3': 'monthly'}
            new_freq = frequencies.get(freq_choice)
            
            if new_freq:
                report['schedule_config']['frequency'] = new_freq
                
                new_hour = input("New hour (0-23): ").strip()
                if new_hour.isdigit() and 0 <= int(new_hour) <= 23:
                    report['schedule_config']['hour'] = int(new_hour)
                
                print(f"✅ Schedule updated to {new_freq} at {report['schedule_config']['hour']}:00")
            
        elif action == '3':
            new_recipients = []
            print("Current recipients:", report.get('recipients', []))
            print("Enter new recipients (press Enter when done):")
            
            while True:
                email = input("Email: ").strip()
                if not email:
                    break
                if '@' in email:
                    new_recipients.append(email)
                    
            report['recipients'] = new_recipients
            print(f"✅ Recipients updated ({len(new_recipients)} recipients)")
            
        elif action == '4':
            confirm = input("⚠️  Delete this schedule? (yes/no): ").strip().lower()
            if confirm == 'yes':
                del scheduled_reports[choice]
                print("✅ Schedule deleted")
            else:
                print("❌ Deletion cancelled")
                return
            
        elif action == '5':
            print("🔄 Running report now...")
            try:
                end_date = datetime.now().strftime("%Y-%m-%d")
                start_date = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
                
                report_path = generate_report(report['template_name'], start_date, end_date, 'pdf')
                
                if report_path and report.get('recipients'):
                    success, message = EmailManager.send_email(
                        report['recipients'],
                        f"Scheduled Report: {report['template_name']}",
                        f"<h2>Scheduled Report</h2><p>Report generated at {datetime.now()}</p>",
                        [report_path]
                    )
                    
                    if success:
                        print("✅ Report generated and sent successfully!")
                    else:
                        print(f"⚠️  Report generated but email failed: {message}")
                else:
                    print("✅ Report generated successfully!")
                    
            except Exception as e:
                print(f"❌ Error running report: {str(e)}")
                return
        
        # Save changes
        save_scheduled_reports(scheduled_reports)
        
    except (ValueError, IndexError):
        print("❌ Invalid selection")

def run_maintenance_menu():
    """Run system maintenance tasks"""
    print("\n🔧 SYSTEM MAINTENANCE")
    print("="*35)
    
    print("Available maintenance tasks:")
    print("1. Clean old reports (30+ days)")
    print("2. Clear cache")
    print("3. Run data quality check")
    print("4. Optimize database")
    print("5. Run all maintenance")
    
    choice = input("Select task (1-5): ").strip()
    
    try:
        if choice == '1':
            print("🗂️  Cleaning old report files...")
            cleanup_old_reports()
            print("✅ Old reports cleaned")
            
        elif choice == '2':
            print("🧹 Clearing cache...")
            CacheManager.cleanup_cache()
            print("✅ Cache cleared")
            
        elif choice == '3':
            print("🔍 Running data quality check...")
            quality_report = DataQualityMonitor.run_quality_checks()
            print("✅ Data quality check completed")
            
            # Show brief summary
            checks = quality_report.get('checks', {})
            if 'missing_data' in checks:
                total = checks['missing_data']['students']['total_records']
                print(f"   📊 {total} total records analyzed")
            
        elif choice == '4':
            print("🗃️  Optimizing database...")
            conn = get_reporting_db_connection()
            try:
                conn.execute("VACUUM")
                conn.execute("ANALYZE")
                print("✅ Database optimized")
            finally:
                conn.close()
                
        elif choice == '5':
            print("🔄 Running all maintenance tasks...")
            quality_report = run_system_maintenance()
            print("✅ All maintenance tasks completed")
            
        else:
            print("❌ Invalid selection")
            
    except Exception as e:
        print(f"❌ Maintenance error: {str(e)}")

def show_performance_monitor():
    """Show system performance metrics"""
    print("\n📊 PERFORMANCE MONITOR")
    print("="*35)
    
    try:
        # Database size
        db_size = os.path.getsize(CONFIG['database']) / (1024 * 1024)  # MB
        print(f"📀 Database size: {db_size:.2f} MB")
        
        # Reports directory size
        reports_size = 0
        for root, dirs, files in os.walk(CONFIG['reports_dir']):
            for file in files:
                reports_size += os.path.getsize(os.path.join(root, file))
        reports_size = reports_size / (1024 * 1024)  # MB
        print(f"📁 Reports size: {reports_size:.2f} MB")
        
        # Cache directory size
        cache_size = 0
        if os.path.exists(CONFIG['cache_dir']):
            for root, dirs, files in os.walk(CONFIG['cache_dir']):
                for file in files:
                    cache_size += os.path.getsize(os.path.join(root, file))
        cache_size = cache_size / (1024 * 1024)  # MB
        print(f"💾 Cache size: {cache_size:.2f} MB")
        
        # Record counts
        conn = get_reporting_db_connection()
        try:
            cursor = conn.cursor()

            cursor.execute("SELECT COUNT(*) FROM students")
            student_count = cursor.fetchone()[0]
            print(f"👥 Total students: {student_count}")

            # Whitelist of allowed tables for SQL injection prevention
            ALLOWED_TABLES = {'student_modules', 'student_grades', 'student_attendance'}

            # Check if other tables exist
            tables = ['student_modules', 'student_grades', 'student_attendance']
            for table in tables:
                # Validate table name against whitelist before using in query
                if table not in ALLOWED_TABLES:
                    continue
                try:
                    safe_table = validate_table_name(table)
                    cursor.execute("SELECT COUNT(*) FROM [" + safe_table + "]")
                    count = cursor.fetchone()[0]
                    table_name = table.replace('_', ' ').title()
                    print(f"📚 {table_name}: {count}")
                except Exception:
                    pass  # Table might not exist
                    
        finally:
            conn.close()
        
        # Template and schedule counts
        templates = load_templates()
        print(f"📋 Templates: {len(templates)}")
        
        scheduled_reports = load_scheduled_reports()
        print(f"⏰ Scheduled reports: {len(scheduled_reports)}")
        
        # Users
        users = UserManager.load_users()
        active_users = sum(1 for u in users.values() if u.get('is_active', True))
        print(f"👤 Active users: {active_users}/{len(users)}")
        
    except Exception as e:
        print(f"❌ Error gathering performance data: {str(e)}")
    
    input("\nPress Enter to continue...")

def export_logs_menu():
    """Export system logs"""
    print("\n📄 EXPORT SYSTEM LOGS")
    print("="*35)
    
    log_file = get_log_file('reporting_system.log')
    
    if not os.path.exists(log_file):
        print("❌ No log file found")
        return
    
    # Get log file info
    log_size = os.path.getsize(log_file) / 1024  # KB
    mod_time = datetime.fromtimestamp(os.path.getmtime(log_file))
    
    print(f"📊 Log file size: {log_size:.2f} KB")
    print(f"📅 Last modified: {mod_time}")
    
    export_choice = input("\nExport options:\n1. Last 100 lines\n2. Last 24 hours\n3. Full log\nSelect (1-3): ").strip()
    
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        export_file = os.path.join(CONFIG['reports_dir'], f"system_logs_{timestamp}.txt")
        
        with open(log_file, 'r') as f:
            lines = f.readlines()
        
        if export_choice == '1':
            # Last 100 lines
            export_lines = lines[-100:] if len(lines) > 100 else lines
            
        elif export_choice == '2':
            # Last 24 hours
            cutoff_time = datetime.now() - timedelta(hours=24)
            export_lines = []
            
            for line in reversed(lines):
                try:
                    # Extract timestamp from log line
                    if line.strip() and ' - ' in line:
                        timestamp_str = line.split(' - ')[0]
                        log_time = datetime.strptime(timestamp_str, '%Y-%m-%d %H:%M:%S,%f')
                        if log_time >= cutoff_time:
                            export_lines.append(line)
                        else:
                            break
                except (ValueError, IndexError):
                    export_lines.append(line)  # Include lines without valid timestamps
            
            export_lines.reverse()
            
        else:
            # Full log
            export_lines = lines
        
        # Write export file
        with open(export_file, 'w') as f:
            f.writelines(export_lines)
        
        print(f"✅ Logs exported successfully!")
        print(f"📁 File: {export_file}")
        print(f"📊 Lines exported: {len(export_lines)}")
        
    except Exception as e:
        print(f"❌ Error exporting logs: {str(e)}")

def load_scheduled_reports():
    """Load scheduled reports from file"""
    try:
        with open(CONFIG['scheduled_reports_file'], 'r') as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return []

def save_scheduled_reports(scheduled_reports):
    """Save scheduled reports to file"""
    with open(CONFIG['scheduled_reports_file'], 'w') as f:
        json.dump(scheduled_reports, f, indent=4)

def start_scheduler():
    """Start the background scheduler for automatic reports"""
    def run_scheduler():
        while True:
            try:
                schedule.run_pending()
                time.sleep(60)  # Check every minute
            except Exception as e:
                logger.error(f"Scheduler error: {str(e)}")
                time.sleep(60)
    
    # Load and schedule all reports
    def schedule_report(report_data):
        """Schedule a single report"""
        def run_report():
            try:
                template_name = report_data['template_name']
                recipients = report_data.get('recipients', [])
                
                # Generate report
                end_date = datetime.now().strftime("%Y-%m-%d")
                start_date = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
                
                report_path = generate_report(template_name, start_date, end_date, 'pdf')
                
                if report_path:
                    # Update run statistics
                    report_data['last_run'] = datetime.now().isoformat()
                    report_data['run_count'] = report_data.get('run_count', 0) + 1
                    
                    # Send email if recipients configured
                    if recipients:
                        from university_system.infrastructure.email.template_utils import render_template
                        subject, body = render_template("scheduled_report", {
                            "template_name": template_name,
                            "generated_date": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                            "start_date": start_date,
                            "end_date": end_date
                        })
                        if subject and body:
                            EmailManager.send_email(recipients, subject, body, [report_path])
                    
                    logger.info(f"Scheduled report '{template_name}' generated successfully")
                else:
                    logger.error(f"Failed to generate scheduled report '{template_name}'")
                    
            except Exception as e:
                logger.error(f"Error running scheduled report '{template_name}': {str(e)}")
    
    # Schedule all reports
    scheduled_reports = load_scheduled_reports()
    for report_data in scheduled_reports:
        if not report_data['schedule_config'].get('enabled', True):
            continue
            
        frequency = report_data['schedule_config']['frequency']
        hour = report_data['schedule_config']['hour']
        
        if frequency == 'daily':
            schedule.every().day.at(f"{hour:02d}:00").do(lambda r=report_data: schedule_report(r))
        elif frequency == 'weekly':
            schedule.every().monday.at(f"{hour:02d}:00").do(lambda r=report_data: schedule_report(r))
        elif frequency == 'monthly':
            schedule.every().month.at(f"{hour:02d}:00").do(lambda r=report_data: schedule_report(r))
    
    # Start scheduler in background thread
    scheduler_thread = threading.Thread(target=run_scheduler, daemon=True)
    scheduler_thread.start()
    
    logger.info(f"Background scheduler started with {len(scheduled_reports)} scheduled reports")
    
# Main function to start the enhanced system
def main():
    """Main function to start the advanced reporting system"""
    print("🔬 Advanced Student Reporting System v2.0")
    print("Initializing enhanced features...")
    
    # Run system checks
    try:
        # Check database connection
        conn = get_reporting_db_connection()
        conn.close()
        print("✅ Database connection verified")
        
        # Initialize system config
        SystemConfig.load_config()
        print("✅ System configuration loaded")
        
        # Start background scheduler
        start_scheduler()
        print("✅ Background scheduler started")
        
        print("\n🎉 All systems ready!")
        
        # Show the main menu
        display_enhanced_reporting_menu()
        
    except Exception as e:
        print(f"❌ System initialization failed: {str(e)}")
        logger.error(f"System initialization error: {str(e)}")

if __name__ == "__main__":
    main()
