"""Content negotiation for API responses (JSON, CSV, Excel).

Lower Priority #12: Support multiple response formats via Accept header
or ?format= query parameter.

Usage in route handlers:
    from education_system.shared.api.content_negotiation import negotiate_response

    @app.route("/api/v1/college/reports/attendance")
    def attendance_report():
        rows = service.get_attendance_data()
        columns = ["student_id", "name", "date", "status"]
        return negotiate_response(rows, columns, filename="attendance_report")

Clients request format via:
    Accept: text/csv
    Accept: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet
    ?format=csv
    ?format=excel
    ?format=json  (default)
"""

import csv
import io

from flask import request, jsonify, Response


def _rows_to_dicts(rows, columns: list[str] | None = None) -> list[dict]:
    """Normalise rows to list of dicts."""
    if not rows:
        return []
    if isinstance(rows[0], dict):
        return rows
    # sqlite3.Row or similar
    if hasattr(rows[0], "keys"):
        return [dict(r) for r in rows]
    # Plain tuples with column names
    if columns:
        return [dict(zip(columns, row)) for row in rows]
    return [{"value": r} for r in rows]


def to_csv(rows, columns: list[str] | None = None) -> str:
    """Convert rows to a CSV string."""
    dicts = _rows_to_dicts(rows, columns)
    if not dicts:
        return ""
    output = io.StringIO()
    keys = columns or list(dicts[0].keys())
    writer = csv.DictWriter(output, fieldnames=keys, extrasaction="ignore")
    writer.writeheader()
    for row in dicts:
        writer.writerow(row)
    return output.getvalue()


def to_excel(rows, columns: list[str] | None = None, sheet_name: str = "Data") -> bytes:
    """Convert rows to an Excel (XLSX) file as bytes.

    Uses openpyxl if available, otherwise falls back to CSV bytes.
    """
    dicts = _rows_to_dicts(rows, columns)
    keys = columns or (list(dicts[0].keys()) if dicts else [])

    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Header row
        for col_idx, header in enumerate(keys, 1):
            ws.cell(row=1, column=col_idx, value=header)

        # Data rows
        for row_idx, row in enumerate(dicts, 2):
            for col_idx, key in enumerate(keys, 1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(key))

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    except ImportError:
        # Fallback: return CSV as bytes
        return to_csv(dicts, keys).encode("utf-8")


def _detect_format() -> str:
    """Detect desired response format from query param or Accept header."""
    # Query parameter takes precedence
    fmt = request.args.get("format", "").lower()
    if fmt in ("csv", "excel", "xlsx", "json"):
        return "excel" if fmt == "xlsx" else fmt

    # Check Accept header
    accept = request.headers.get("Accept", "")
    if "text/csv" in accept:
        return "csv"
    if "spreadsheetml" in accept or "application/vnd.ms-excel" in accept:
        return "excel"

    return "json"


def negotiate_response(rows, columns: list[str] | None = None,
                       filename: str = "export",
                       json_key: str = "data",
                       total: int | None = None):
    """Return data in the format requested by the client.

    Args:
        rows: List of dicts, sqlite3.Row objects, or tuples
        columns: Column names (required if rows are tuples)
        filename: Base filename for CSV/Excel downloads (without extension)
        json_key: Key to wrap data under in JSON response
        total: Total count for pagination (JSON only)

    Returns:
        Flask Response in the requested format
    """
    fmt = _detect_format()

    if fmt == "csv":
        csv_data = to_csv(rows, columns)
        return Response(
            csv_data,
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}.csv"},
        )

    if fmt == "excel":
        excel_data = to_excel(rows, columns, sheet_name=filename)
        return Response(
            excel_data,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"},
        )

    # Default: JSON
    result = {json_key: _rows_to_dicts(rows, columns)}
    if total is not None:
        result["total"] = total
    return jsonify(result)
