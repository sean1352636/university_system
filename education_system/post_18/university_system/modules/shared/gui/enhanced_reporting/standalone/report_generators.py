"""Report generation functions for the enhanced reporting GUI."""

from education_system.post_18.university_system.modules.shared.gui.enhanced_reporting.standalone.constants import (
    logging, os, json, pd, datetime, timedelta,
    paths, get_db_connection,
    ENHANCED_AVAILABLE,
)


def generate_enhanced_pdf_report(template, filename, start_date, end_date, comparison_date=None):
    """Generate enhanced PDF report with advanced features"""
    try:
        if not ENHANCED_AVAILABLE:
            raise Exception("Enhanced reporting not available")

        # Import required libraries
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            from reportlab.graphics.shapes import Drawing
            from reportlab.graphics.charts.linecharts import HorizontalLineChart
            from reportlab.graphics.charts.piecharts import Pie
            from reportlab.graphics.charts.barcharts import VerticalBarChart
        except ImportError:
            # Fallback to basic PDF generation
            raise Exception("ReportLab not available for enhanced PDF generation")

        doc = SimpleDocTemplate(filename, pagesize=A4,
                              leftMargin=0.75*inch, rightMargin=0.75*inch,
                              topMargin=1*inch, bottomMargin=1*inch)

        styles = getSampleStyleSheet()
        story = []

        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=24,
            textColor=colors.HexColor('#2E4E7E'),
            spaceAfter=30
        )
        story.append(Paragraph(f"Enhanced Report: {template['name']}", title_style))
        story.append(Spacer(1, 20))

        # Metadata
        meta_data = [
            ['Report Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Period:', f"{start_date} to {end_date}"],
            ['Template:', template['name']],
            ['Security Level:', template.get('security_level', 'normal').title()]
        ]

        meta_table = Table(meta_data, colWidths=[2*inch, 4*inch])
        meta_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F8F9FA')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(meta_table)
        story.append(Spacer(1, 30))

        # Generate sections
        for section in template.get('sections', []):
            section_data = generate_enhanced_section(section, start_date, end_date,
                                                   comparison_date, template.get('filters'),
                                                   template.get('visualization_type', 'standard'))

            if section_data:
                # Section title
                section_title = section.replace('_', ' ').title()
                story.append(Paragraph(section_title, styles['Heading1']))
                story.append(Spacer(1, 12))

                # Section content
                if 'summary' in section_data:
                    story.append(Paragraph(section_data['summary'], styles['Normal']))
                    story.append(Spacer(1, 10))

                # Add tables if present
                if 'data' in section_data and hasattr(section_data['data'], 'empty') and not section_data['data'].empty:
                    df = section_data['data'].head(10)  # Limit to 10 rows for PDF
                    table_data = [df.columns.tolist()] + df.values.tolist()

                    table = Table(table_data, repeatRows=1)
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E3F2FD')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, -1), 8),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ]))
                    story.append(table)

                story.append(Spacer(1, 20))

        # Build PDF
        doc.build(story)
        return filename

    except Exception as e:
        logging.error(f"Error generating enhanced PDF: {str(e)}")
        return None

def generate_enhanced_section(section, start_date, end_date, comparison_date=None, filters=None, visualization_type='standard'):
    """Generate enhanced section with data and visualizations"""
    try:
        conn = get_db_connection()
        if not conn:
            return None

        section_data = {
            'name': section,
            'summary': '',
            'data': None,
            'visualization': None,
            'statistics': {}
        }

        if section == 'student_overview':
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) as total_students FROM students")
            total_students = cursor.fetchone()[0]

            cursor.execute("SELECT COUNT(DISTINCT course) FROM students WHERE course IS NOT NULL")
            total_courses = cursor.fetchone()[0]

            section_data['summary'] = f"Total Students: {total_students}, Total Courses: {total_courses}"
            section_data['statistics'] = {
                'total_students': total_students,
                'total_courses': total_courses
            }

        elif section == 'course_distribution':
            query = """
            SELECT course, COUNT(*) as count
            FROM students
            WHERE course IS NOT NULL AND registration_datetime BETWEEN ? AND ?
            GROUP BY course
            ORDER BY count DESC
            """
            df = pd.read_sql_query(query, conn, params=[start_date + ' 00:00:00', end_date + ' 23:59:59'])
            section_data['data'] = df
            section_data['summary'] = f"Course distribution across {len(df)} courses"

        elif section == 'gender_distribution':
            query = """
            SELECT gender, COUNT(*) as count
            FROM students
            WHERE gender IS NOT NULL AND registration_datetime BETWEEN ? AND ?
            GROUP BY gender
            """
            df = pd.read_sql_query(query, conn, params=[start_date + ' 00:00:00', end_date + ' 23:59:59'])
            section_data['data'] = df
            section_data['summary'] = f"Gender distribution across {df['count'].sum()} students"

        elif section == 'age_distribution':
            query = """
            SELECT
                CASE
                    WHEN age < 20 THEN 'Under 20'
                    WHEN age BETWEEN 20 AND 25 THEN '20-25'
                    WHEN age BETWEEN 26 AND 30 THEN '26-30'
                    WHEN age BETWEEN 31 AND 35 THEN '31-35'
                    ELSE 'Over 35'
                END as age_group,
                COUNT(*) as count
            FROM students
            WHERE age IS NOT NULL AND registration_datetime BETWEEN ? AND ?
            GROUP BY age_group
            ORDER BY count DESC
            """
            df = pd.read_sql_query(query, conn, params=[start_date + ' 00:00:00', end_date + ' 23:59:59'])
            section_data['data'] = df
            section_data['summary'] = f"Age distribution across {df['count'].sum()} students"

        elif section == 'registration_trends':
            query = """
            SELECT DATE(registration_datetime) as registration_date, COUNT(*) as count
            FROM students
            WHERE registration_datetime BETWEEN ? AND ?
            GROUP BY DATE(registration_datetime)
            ORDER BY registration_date
            """
            df = pd.read_sql_query(query, conn, params=[start_date + ' 00:00:00', end_date + ' 23:59:59'])
            section_data['data'] = df
            section_data['summary'] = f"Registration trends over {len(df)} days"

        # Add more sections as needed...

        conn.close()
        return section_data

    except Exception as e:
        logging.error(f"Error generating section {section}: {str(e)}")
        if conn:
            conn.close()
        return None

def generate_quality_section(quality_report, styles):
    """Generate quality section for PDF report"""
    try:
        from reportlab.platypus import Paragraph, Spacer, Table, TableStyle
        from reportlab.lib import colors

        content = []

        # Title
        content.append(Paragraph("Data Quality Report", styles['Heading1']))
        content.append(Spacer(1, 12))

        # Summary
        checks = quality_report.get('checks', {})

        if 'missing_data' in checks:
            missing = checks['missing_data']['students']
            total = missing['total_records']

            summary_data = [
                ['Metric', 'Value'],
                ['Total Records', str(total)],
                ['Missing Emails', str(missing.get('missing_emails', 0))],
                ['Missing Names', str(missing.get('missing_names', 0))],
                ['Missing Courses', str(missing.get('missing_courses', 0))]
            ]

            if total > 0:
                completeness = ((total * 3) - sum([missing.get('missing_emails', 0),
                                                  missing.get('missing_names', 0),
                                                  missing.get('missing_courses', 0)])) / (total * 3) * 100
                summary_data.append(['Data Completeness', f"{completeness:.1f}%"])

            summary_table = Table(summary_data, colWidths=[2*72, 2*72])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E3F2FD')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            content.append(summary_table)

        return content

    except Exception as e:
        logging.error(f"Error generating quality section: {str(e)}")
        return []

def generate_predictions_section(predictions, styles):
    """Generate predictions section for PDF report"""
    try:
        from reportlab.platypus import Paragraph, Spacer, Table, TableStyle
        from reportlab.lib import colors

        content = []

        # Title
        content.append(Paragraph("Predictive Analytics", styles['Heading1']))
        content.append(Spacer(1, 12))

        if 'error' in predictions:
            content.append(Paragraph(f"Analysis unavailable: {predictions['error']}", styles['Normal']))
        else:
            # Summary data
            summary_data = [['Metric', 'Value']]

            if 'total_students_analyzed' in predictions:
                summary_data.append(['Students Analyzed', str(predictions['total_students_analyzed'])])

            if 'model_accuracy' in predictions:
                accuracy = predictions['model_accuracy'] * 100
                summary_data.append(['Model Accuracy', f"{accuracy:.1f}%"])

            if 'high_risk_students' in predictions:
                high_risk_count = len(predictions['high_risk_students'])
                summary_data.append(['High Risk Students', str(high_risk_count)])

            if len(summary_data) > 1:
                summary_table = Table(summary_data, colWidths=[2*72, 2*72])
                summary_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FFF3E0')),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                content.append(summary_table)

        return content

    except Exception as e:
        logging.error(f"Error generating predictions section: {str(e)}")
        return []

def generate_enhanced_excel_report(template, filename, start_date, end_date, comparison_date=None):
    """Generate enhanced Excel report with multiple sheets and formatting"""
    try:
        if not ENHANCED_AVAILABLE:
            raise Exception("Enhanced reporting not available")

        # Import required libraries
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.chart import BarChart, PieChart, LineChart, Reference
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            raise Exception("openpyxl not available for Excel generation")

        # Create workbook
        wb = openpyxl.Workbook()

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Create summary sheet
        summary_sheet = wb.create_sheet("Summary")
        summary_sheet['A1'] = f"Report: {template['name']}"
        summary_sheet['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        summary_sheet['A3'] = f"Period: {start_date} to {end_date}"
        summary_sheet['A4'] = f"Security Level: {template.get('security_level', 'normal').title()}"

        # Style the summary
        title_font = Font(bold=True, size=16)
        summary_sheet['A1'].font = title_font

        row_num = 6

        # Process each section
        for section in template.get('sections', []):
            try:
                # Get section data
                section_data = generate_enhanced_section(section, start_date, end_date,
                                                       comparison_date, template.get('filters'),
                                                       template.get('visualization_type', 'standard'))

                if section_data and section_data.get('data') is not None and not section_data['data'].empty:
                    df = section_data['data']

                    # Create sheet for this section
                    sheet_name = section.replace('_', ' ').title()[:31]  # Excel sheet name limit
                    section_sheet = wb.create_sheet(sheet_name)

                    # Add section title
                    section_sheet['A1'] = sheet_name
                    section_sheet['A1'].font = Font(bold=True, size=14)

                    # Add summary if available
                    if section_data.get('summary'):
                        section_sheet['A2'] = section_data['summary']
                        section_sheet['A2'].font = Font(italic=True)

                    # Add data starting from row 4
                    for r in dataframe_to_rows(df, index=False, header=True):
                        section_sheet.append(r)

                    # Format headers
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    header_font = Font(color="FFFFFF", bold=True)

                    for cell in section_sheet[4]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center")

                    # Add to summary sheet
                    summary_sheet[f'A{row_num}'] = f"{sheet_name}: {len(df)} records"
                    row_num += 1

            except Exception as e:
                logging.error(f"Error processing section {section}: {str(e)}")
                continue

        # Save workbook
        wb.save(filename)
        return filename

    except Exception as e:
        logging.error(f"Error generating enhanced Excel report: {str(e)}")
        return None

def generate_interactive_report(template, filename, start_date, end_date):
    """Generate interactive HTML report"""
    try:
        html_content = f"""
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Interactive Report: {template['name']}</title>
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    margin: 0;
                    padding: 20px;
                    background-color: #f5f5f5;
                }}
                .header {{
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    color: white;
                    padding: 30px;
                    border-radius: 10px;
                    margin-bottom: 30px;
                    text-align: center;
                }}
                .section {{
                    background: white;
                    padding: 20px;
                    margin-bottom: 20px;
                    border-radius: 10px;
                    box-shadow: 0 2px 10px rgba(0,0,0,0.1);
                }}
                .section h2 {{
                    color: #333;
                    border-bottom: 2px solid #667eea;
                    padding-bottom: 10px;
                }}
                .stats-grid {{
                    display: grid;
                    grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
                    gap: 20px;
                    margin: 20px 0;
                }}
                .stat-card {{
                    background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%);
                    color: white;
                    padding: 20px;
                    border-radius: 10px;
                    text-align: center;
                }}
                .stat-card h3 {{
                    margin: 0;
                    font-size: 2em;
                }}
                .stat-card p {{
                    margin: 5px 0 0 0;
                    opacity: 0.9;
                }}
                table {{
                    width: 100%;
                    border-collapse: collapse;
                    margin-top: 20px;
                }}
                th, td {{
                    padding: 12px;
                    text-align: left;
                    border-bottom: 1px solid #ddd;
                }}
                th {{
                    background-color: #667eea;
                    color: white;
                }}
                tr:hover {{
                    background-color: #f5f5f5;
                }}
                .chart-container {{
                    margin: 20px 0;
                    text-align: center;
                }}
            </style>
            <script src="https://cdn.plot.ly/plotly-latest.min.js"></script>
        </head>
        <body>
            <div class="header">
                <h1>Interactive Report: {template['name']}</h1>
                <p>Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                <p>Period: {start_date} to {end_date}</p>
            </div>
        """

        # Process sections
        for section in template.get('sections', []):
            try:
                section_data = generate_enhanced_section(section, start_date, end_date,
                                                       None, template.get('filters'),
                                                       'interactive')

                if section_data:
                    html_content += f"""
                    <div class="section">
                        <h2>{section.replace('_', ' ').title()}</h2>
                    """

                    if section_data.get('summary'):
                        html_content += f"<p>{section_data['summary']}</p>"

                    if section_data.get('data') is not None and not section_data['data'].empty:
                        df = section_data['data']

                        # Add statistics if numeric data
                        numeric_cols = df.select_dtypes(include=['int64', 'float64']).columns
                        if len(numeric_cols) > 0:
                            html_content += '<div class="stats-grid">'
                            for col in numeric_cols[:4]:  # Limit to 4 stats
                                col_data = df[col].dropna()
                                if len(col_data) > 0:
                                    html_content += f"""
                                    <div class="stat-card">
                                        <h3>{col_data.mean():.1f}</h3>
                                        <p>Avg {col.replace('_', ' ').title()}</p>
                                    </div>
                                    """
                            html_content += '</div>'

                        # Add table
                        html_content += df.head(20).to_html(classes='', escape=False, index=False)

                        # Add interactive chart
                        if len(df) > 1 and len(df.columns) >= 2:
                            chart_id = f"chart_{section}"
                            html_content += f'<div id="{chart_id}" class="chart-container"></div>'

                            # Generate Plotly JavaScript
                            html_content += f"""
                            <script>
                                var data_{section} = [{{
                                    x: {df.iloc[:, 0].tolist()},
                                    y: {df.iloc[:, 1].tolist() if len(df.columns) > 1 else [1] * len(df)},
                                    type: 'bar',
                                    marker: {{color: '#667eea'}}
                                }}];
                                var layout_{section} = {{
                                    title: '{section.replace("_", " ").title()}',
                                    xaxis: {{title: '{df.columns[0]}'}},
                                    yaxis: {{title: '{df.columns[1] if len(df.columns) > 1 else "Count"}'}}
                                }};
                                Plotly.newPlot('{chart_id}', data_{section}, layout_{section});
                            </script>
                            """

                    html_content += "</div>"

            except Exception as e:
                logging.error(f"Error processing section {section}: {str(e)}")
                continue

        # Close HTML
        html_content += """
            </body>
        </html>
        """

        # Save to file
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(html_content)

        return filename

    except Exception as e:
        logging.error(f"Error generating interactive report: {str(e)}")
        return None
