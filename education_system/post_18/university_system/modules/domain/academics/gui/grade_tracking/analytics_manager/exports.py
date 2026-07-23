"""Data export methods for AnalyticsManager."""

import tkinter as tk
from tkinter import messagebox, filedialog
import csv

from education_system.post_18.university_system.infrastructure.database.db import sqlite3

try:
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch
except ImportError:
    SimpleDocTemplate = None


class ExportsMixin:
    """Mixin providing data export methods."""

    def export_reports_pdf(self):
        """Export current report to PDF"""
        if not hasattr(self, 'report_preview') or not self.report_preview:
            messagebox.showwarning("No Report", "Please generate a report first.")
            return

        if SimpleDocTemplate is None:
            messagebox.showerror("Feature Unavailable", "ReportLab is required for PDF export.")
            return

        try:
            report_content = self.report_preview.get('1.0', tk.END)

            filename = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
            )

            if filename:
                doc = SimpleDocTemplate(filename, pagesize=letter)
                styles = getSampleStyleSheet()
                story = []

                for line in report_content.split('\n'):
                    if line.strip():
                        story.append(Paragraph(line, styles['Normal']))
                        story.append(Spacer(1, 0.2*inch))

                doc.build(story)
                messagebox.showinfo("Success", f"Report exported to PDF:\n{filename}")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to export to PDF: {e}")

    def export_reports_csv(self):
        """Export current report to CSV"""
        if not hasattr(self, 'report_preview') or not self.report_preview:
            messagebox.showwarning("No Report", "Please generate a report first.")
            return

        try:
            report_content = self.report_preview.get('1.0', tk.END)

            filename = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
            )

            if filename:
                with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                    writer = csv.writer(csvfile)
                    for line in report_content.split('\n'):
                        if line.strip():
                            writer.writerow([line])

                messagebox.showinfo("Success", f"Report exported to CSV:\n{filename}")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to export to CSV: {e}")

    def export_reports_excel(self):
        """Export current report to Excel"""
        if not hasattr(self, 'report_preview') or not self.report_preview:
            messagebox.showwarning("No Report", "Please generate a report first.")
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            report_content = self.report_preview.get('1.0', tk.END)

            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )

            if filename:
                wb = Workbook()
                ws = wb.active
                ws.title = "Report"

                for idx, line in enumerate(report_content.split('\n'), 1):
                    if line.strip():
                        ws.cell(row=idx, column=1, value=line)

                wb.save(filename)
                messagebox.showinfo("Success", f"Report exported to Excel:\n{filename}")

        except ImportError:
            messagebox.showerror("Feature Unavailable", "openpyxl is required for Excel export.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export to Excel: {e}")
