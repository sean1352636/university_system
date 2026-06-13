"""Import/export methods for IntegrationMarketplaceGUI."""

from __future__ import annotations

from tkinter import messagebox, filedialog
from datetime import datetime
import json
import logging

from education_system.university_system.infrastructure.database.db import get_connection, transaction
from education_system.university_system.core.activity_logger import log_activity
from education_system.university_system.core.i18n import get_text as _t

logger = logging.getLogger(__name__)


class ImportExportMixin:
    """Mixin providing import/export methods."""

    def export_catalog_to_json(self):
        """Export full catalog to JSON file"""
        try:
            filename = filedialog.asksaveasfilename(
                defaultextension=".json",
                filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
                initialfile="integration_catalog_export.json"
            )

            if not filename:
                return

            with get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT integration_id, integration_name, provider_name, integration_type,
                           category, description, version, rating, install_count, is_official,
                           pricing_model, documentation_url
                    FROM integration_catalog
                    WHERE is_active = 1
                ''')
                integrations = cursor.fetchall()

            export_data = {
                'export_date': datetime.now().isoformat(),
                'export_type': 'integration_catalog',
                'version': '1.0',
                'integrations': []
            }

            for row in integrations:
                export_data['integrations'].append({
                    'integration_id': row[0],
                    'integration_name': row[1],
                    'provider_name': row[2],
                    'integration_type': row[3],
                    'category': row[4],
                    'description': row[5],
                    'version': row[6],
                    'rating': row[7],
                    'install_count': row[8],
                    'is_official': bool(row[9]),
                    'pricing_model': row[10],
                    'documentation_url': row[11]
                })

            with open(filename, 'w') as f:
                json.dump(export_data, f, indent=2)

            log_activity('export', 'integration_catalog', None,
                        details={'filename': filename, 'count': len(integrations)})

            messagebox.showinfo(_t("common.success"),
                              f"Exported {len(integrations)} integrations to:\n{filename}")

        except Exception as e:
            logger.error(f"Error exporting catalog: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to export catalog: {e}")

    def import_integrations_from_json(self):
        """Import integrations from JSON backup"""
        try:
            filename = filedialog.askopenfilename(
                filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
            )

            if not filename:
                return

            with open(filename, 'r') as f:
                import_data = json.load(f)

            if 'integrations' not in import_data:
                messagebox.showerror(_t("common.error"), "Invalid import file format")
                return

            imported_count = 0
            skipped_count = 0

            for item in import_data['integrations']:
                try:
                    with get_connection() as conn:
                        cursor = conn.cursor()
                        cursor.execute('''
                            SELECT integration_id FROM integration_catalog
                            WHERE integration_name = ? AND provider_name = ?
                        ''', (item['integration_name'], item['provider_name']))

                        if cursor.fetchone():
                            skipped_count += 1
                            continue

                    with transaction() as conn:
                        cursor = conn.cursor()
                        cursor.execute('''
                            INSERT INTO integration_catalog
                            (integration_name, provider_name, integration_type, category,
                             description, version, is_official, is_active)
                            VALUES (?, ?, ?, ?, ?, ?, ?, 1)
                        ''', (item['integration_name'], item['provider_name'],
                              item.get('integration_type', 'API'), item.get('category', 'Other'),
                              item.get('description', ''), item.get('version', '1.0'),
                              int(item.get('is_official', False))))

                    imported_count += 1

                except Exception as e:
                    logger.error(f"Error importing integration: {e}")
                    skipped_count += 1

            log_activity('import', 'integration_catalog', None,
                        details={'filename': filename, 'imported': imported_count, 'skipped': skipped_count})

            messagebox.showinfo("Import Complete",
                              f"Imported: {imported_count}\nSkipped (duplicates): {skipped_count}")
            self.load_catalog()

        except Exception as e:
            logger.error(f"Error importing integrations: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to import integrations: {e}")

    def export_configuration_bundle(self):
        """Export all configs, credentials (encrypted), and mappings as a bundle"""
        try:
            filename = filedialog.asksaveasfilename(
                defaultextension=".json",
                filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
                initialfile="integration_config_bundle.json"
            )

            if not filename:
                return

            bundle = {
                'export_date': datetime.now().isoformat(),
                'export_type': 'configuration_bundle',
                'version': '1.0',
                'installed_integrations': [],
                'credentials': [],
                'data_mappings': [],
                'webhooks': []
            }

            with get_connection() as conn:
                cursor = conn.cursor()

                # Export installed integrations
                cursor.execute('''
                    SELECT ii.install_id, ic.integration_name, ii.version_installed,
                           ii.configuration, ii.sync_frequency, ii.status
                    FROM installed_integrations ii
                    JOIN integration_catalog ic ON ii.integration_id = ic.integration_id
                    WHERE ii.status != 'uninstalled'
                ''')
                for row in cursor.fetchall():
                    bundle['installed_integrations'].append({
                        'install_id': row[0],
                        'integration_name': row[1],
                        'version': row[2],
                        'configuration': row[3],
                        'sync_frequency': row[4],
                        'status': row[5]
                    })

                # Export credentials (masked)
                cursor.execute('''
                    SELECT credential_id, install_id, credential_type, endpoint_url
                    FROM integration_credentials
                ''')
                for row in cursor.fetchall():
                    bundle['credentials'].append({
                        'credential_id': row[0],
                        'install_id': row[1],
                        'credential_type': row[2],
                        'endpoint_url': row[3],
                        'api_key': '***MASKED***',
                        'api_secret': '***MASKED***'
                    })

                # Export data mappings
                cursor.execute('''
                    SELECT mapping_id, install_id, source_field, target_field,
                           transformation_rule, is_active
                    FROM integration_data_mappings
                ''')
                for row in cursor.fetchall():
                    bundle['data_mappings'].append({
                        'mapping_id': row[0],
                        'install_id': row[1],
                        'source_field': row[2],
                        'target_field': row[3],
                        'transformation_rule': row[4],
                        'is_active': bool(row[5])
                    })

                # Export webhooks
                cursor.execute('''
                    SELECT webhook_id, install_id, webhook_url, event_type, is_active
                    FROM integration_webhooks
                ''')
                for row in cursor.fetchall():
                    bundle['webhooks'].append({
                        'webhook_id': row[0],
                        'install_id': row[1],
                        'webhook_url': row[2],
                        'event_type': row[3],
                        'is_active': bool(row[4])
                    })

            with open(filename, 'w') as f:
                json.dump(bundle, f, indent=2)

            log_activity('export', 'configuration_bundle', None,
                        details={'filename': filename})

            messagebox.showinfo(_t("common.success"),
                              f"Configuration bundle exported to:\n{filename}\n\n"
                              f"Note: Credentials are masked for security.")

        except Exception as e:
            logger.error(f"Error exporting configuration bundle: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to export bundle: {e}")

    def import_configuration_bundle(self):
        """Import and restore configuration bundle"""
        try:
            filename = filedialog.askopenfilename(
                filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
            )

            if not filename:
                return

            with open(filename, 'r') as f:
                bundle = json.load(f)

            if bundle.get('export_type') != 'configuration_bundle':
                messagebox.showerror(_t("common.error"), "Invalid configuration bundle format")
                return

            if not messagebox.askyesno("Confirm Import",
                                       "Import configuration bundle?\n\n"
                                       "This will add new data mappings and webhooks.\n"
                                       "Credentials will need to be reconfigured manually."):
                return

            imported = {'mappings': 0, 'webhooks': 0}

            # Import data mappings
            for mapping in bundle.get('data_mappings', []):
                try:
                    with transaction() as conn:
                        cursor = conn.cursor()
                        cursor.execute('''
                            INSERT INTO integration_data_mappings
                            (install_id, source_field, target_field, transformation_rule, is_active)
                            VALUES (?, ?, ?, ?, ?)
                        ''', (mapping['install_id'], mapping['source_field'],
                              mapping['target_field'], mapping.get('transformation_rule'),
                              int(mapping.get('is_active', True))))
                    imported['mappings'] += 1
                except Exception as e:
                    logger.warning(f"Skipped mapping import: {e}")

            # Import webhooks
            for webhook in bundle.get('webhooks', []):
                try:
                    with transaction() as conn:
                        cursor = conn.cursor()
                        cursor.execute('''
                            INSERT INTO integration_webhooks
                            (install_id, webhook_url, event_type, is_active)
                            VALUES (?, ?, ?, ?)
                        ''', (webhook['install_id'], webhook['webhook_url'],
                              webhook['event_type'], int(webhook.get('is_active', True))))
                    imported['webhooks'] += 1
                except Exception as e:
                    logger.warning(f"Skipped webhook import: {e}")

            log_activity('import', 'configuration_bundle', None,
                        details={'filename': filename, 'imported': imported})

            messagebox.showinfo("Import Complete",
                              f"Imported:\n"
                              f"- Data mappings: {imported['mappings']}\n"
                              f"- Webhooks: {imported['webhooks']}\n\n"
                              f"Note: Credentials must be configured manually.")

            self.load_mappings()
            self.load_webhooks()

        except Exception as e:
            logger.error(f"Error importing configuration bundle: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to import bundle: {e}")

    def export_sync_report_pdf(self):
        """Generate PDF report of sync history"""
        try:
            try:
                from reportlab.lib.pagesizes import letter
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                from reportlab.lib.styles import getSampleStyleSheet
                from reportlab.lib import colors
            except ImportError:
                messagebox.showerror(_t("common.error"),
                                   "PDF export requires reportlab library.\n"
                                   "Install with: pip install reportlab")
                return

            filename = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
                initialfile="sync_report.pdf"
            )

            if not filename:
                return

            with get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT isl.log_id, ic.integration_name, isl.sync_start_time,
                           isl.sync_status, isl.records_synced, isl.errors_encountered
                    FROM integration_sync_logs isl
                    JOIN installed_integrations ii ON isl.install_id = ii.install_id
                    JOIN integration_catalog ic ON ii.integration_id = ic.integration_id
                    ORDER BY isl.sync_start_time DESC
                    LIMIT 100
                ''')
                logs = cursor.fetchall()

            # Create PDF
            doc = SimpleDocTemplate(filename, pagesize=letter)
            styles = getSampleStyleSheet()
            elements = []

            # Title
            elements.append(Paragraph("Integration Sync Report", styles['Title']))
            elements.append(Spacer(1, 20))
            elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
            elements.append(Spacer(1, 20))

            # Summary
            total_syncs = len(logs)
            success_count = sum(1 for log in logs if log[3] == 'success')
            failed_count = sum(1 for log in logs if log[3] == 'failed')

            elements.append(Paragraph("Summary", styles['Heading2']))
            elements.append(Paragraph(f"Total Syncs: {total_syncs}", styles['Normal']))
            elements.append(Paragraph(f"Successful: {success_count}", styles['Normal']))
            elements.append(Paragraph(f"Failed: {failed_count}", styles['Normal']))
            elements.append(Spacer(1, 20))

            # Table data
            table_data = [['Log ID', 'Integration', 'Date', 'Status', 'Records', 'Errors']]
            for log in logs:
                table_data.append([
                    str(log[0]),
                    log[1][:20] if log[1] else 'N/A',
                    log[2][:16] if log[2] else 'N/A',
                    log[3] or _t("common.na"),
                    str(log[4] or 0),
                    str(log[5] or 0)
                ])

            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))

            elements.append(table)

            doc.build(elements)

            log_activity('export', 'sync_report_pdf', None,
                        details={'filename': filename, 'log_count': len(logs)})

            messagebox.showinfo(_t("common.success"), f"PDF report exported to:\n{filename}")

        except Exception as e:
            logger.error(f"Error exporting PDF report: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to export PDF: {e}")

    def export_mappings_to_excel(self):
        """Export data mappings to Excel for review"""
        try:
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill
            except ImportError:
                messagebox.showerror(_t("common.error"),
                                   "Excel export requires openpyxl library.\n"
                                   "Install with: pip install openpyxl")
                return

            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile="data_mappings.xlsx"
            )

            if not filename:
                return

            with get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT idm.mapping_id, ic.integration_name, idm.source_field,
                           idm.target_field, idm.transformation_rule, idm.is_active
                    FROM integration_data_mappings idm
                    JOIN installed_integrations ii ON idm.install_id = ii.install_id
                    JOIN integration_catalog ic ON ii.integration_id = ic.integration_id
                    ORDER BY ic.integration_name, idm.source_field
                ''')
                mappings = cursor.fetchall()

            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Data Mappings"

            # Headers
            headers = ['Mapping ID', 'Integration', 'Source Field', 'Target Field',
                      'Transformation Rule', 'Active']
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font

            # Data
            for row_idx, mapping in enumerate(mappings, 2):
                ws.cell(row=row_idx, column=1, value=mapping[0])
                ws.cell(row=row_idx, column=2, value=mapping[1])
                ws.cell(row=row_idx, column=3, value=mapping[2])
                ws.cell(row=row_idx, column=4, value=mapping[3])
                ws.cell(row=row_idx, column=5, value=mapping[4] or '')
                ws.cell(row=row_idx, column=6, value=_t("common.yes") if mapping[5] else _t("common.no"))

            # Adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                ws.column_dimensions[column].width = min(max_length + 2, 50)

            wb.save(filename)

            log_activity('export', 'data_mappings_excel', None,
                        details={'filename': filename, 'count': len(mappings)})

            messagebox.showinfo(_t("common.success"),
                              f"Exported {len(mappings)} mappings to:\n{filename}")

        except Exception as e:
            logger.error(f"Error exporting mappings to Excel: {e}")
            messagebox.showerror(_t("common.error"), f"Failed to export to Excel: {e}")
