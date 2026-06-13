# Auto-generated module
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import csv
import logging
from datetime import datetime
from education_system.university_system.core.i18n import get_text as _t
from education_system.university_system.modules.shared.gui.main._tk_callback_filter import install_clean_close as _install_clean_close

# Import database connection
from education_system.university_system.infrastructure.database.db import get_db_connection, get_connection, transaction
from education_system.university_system.core.sql_safety import validate_identifier  # nosec B608

logger = logging.getLogger(__name__)


# Shared schema: every comprehensive export formatter iterates this list so all
# four output formats (CSV, XLSX, PDF, TXT) render the same sections in the same
# order, with the same headers. Empty sections are skipped by each formatter.
#
# Each entry:
#   key           — `data_sections[key]` fetched in export_individual_student_data
#   title_key     — i18n key for the section heading
#   header_keys   — i18n keys for the column headers, in row order
#   sheet_name_key — (optional) shorter i18n key used for the Excel sheet tab
#
# Header counts match the column count produced by the SQL in
# export_individual_student_data. `student_info` is handled separately by each
# formatter (it renders as two-column key/value, not a tabular list).
EXPORT_SECTIONS = (
    dict(
        key='modules',
        title_key='student_export.comprehensive.enrolled_modules',
        header_keys=(
            'student_export.headers.module_code',
            'student_export.headers.module_name',
            'student_export.headers.credits',
            'student_export.headers.enrollment_date',
            'student_export.headers.status',
        ),
        sheet_name_key='student_export.sheets.modules',
    ),
    dict(
        key='assignments',
        title_key='student_export.comprehensive.assignments_available',
        header_keys=(
            'student_export.headers.module_code',
            'student_export.headers.title',
            'student_export.headers.due_date',
            'student_export.headers.max_marks',
            'student_export.headers.type',
        ),
        sheet_name_key='student_export.sheets.assignments',
    ),
    dict(
        key='submissions',
        title_key='student_export.comprehensive.assignment_submissions',
        header_keys=(
            'student_export.headers.assignment',
            'student_export.headers.module',
            'student_export.headers.submission_date',
            'student_export.headers.grade',
            'student_export.headers.max_marks',
            'student_export.headers.status',
            'student_export.headers.late',
        ),
        sheet_name_key='student_export.sheets.submissions',
    ),
    dict(
        key='grades',
        title_key='student_export.comprehensive.grades_assessments',
        header_keys=(
            'student_export.headers.module',
            'student_export.headers.assessment',
            'student_export.headers.type',
            'student_export.headers.score',
            'student_export.headers.max_points',
            'student_export.headers.letter_grade',
            'student_export.headers.date',
        ),
        sheet_name_key='student_export.sheets.grades',
    ),
    dict(
        key='attendance',
        title_key='student_export.comprehensive.attendance_records',
        header_keys=(
            'student_export.headers.date',
            'student_export.headers.module',
            'student_export.headers.status',
            'student_export.headers.reason',
        ),
        sheet_name_key='student_export.sheets.attendance',
    ),
    dict(
        key='fees',
        title_key='student_export.comprehensive.fees',
        header_keys=(
            'student_export.headers.fee_type',
            'student_export.headers.amount',
            'student_export.headers.due_date',
            'student_export.headers.status',
        ),
        sheet_name_key='student_export.sheets.fees',
    ),
    dict(
        key='payments',
        title_key='student_export.comprehensive.payments',
        header_keys=(
            'student_export.headers.payment_date',
            'student_export.headers.amount',
            'student_export.headers.method',
            'student_export.headers.transaction_id',
        ),
        sheet_name_key='student_export.sheets.payments',
    ),
    dict(
        key='scholarships',
        title_key='student_export.comprehensive.scholarships',
        header_keys=(
            'student_export.headers.scholarship_name',
            'student_export.headers.amount',
            'student_export.headers.academic_year',
            'student_export.headers.status',
        ),
        sheet_name_key='student_export.sheets.scholarships',
    ),
    dict(
        key='financial_aid',
        title_key='student_export.comprehensive.financial_aid',
        header_keys=(
            'student_export.headers.aid_type',
            'student_export.headers.amount',
            'student_export.headers.academic_year',
            'student_export.headers.status',
        ),
    ),
    dict(
        key='activities',
        title_key='student_export.comprehensive.activities',
        header_keys=(
            'student_export.headers.activity_name',
            'student_export.headers.role',
            'student_export.headers.start_date',
            'student_export.headers.end_date',
            'student_export.headers.status',
        ),
        sheet_name_key='student_export.sheets.activities',
    ),
    dict(
        key='clubs',
        title_key='student_export.comprehensive.club_memberships',
        header_keys=(
            'student_export.headers.club_name',
            'student_export.headers.role',
            'student_export.headers.join_date',
            'student_export.headers.status',
        ),
        sheet_name_key='student_export.sheets.clubs',
    ),
    dict(
        key='health_appointments',
        title_key='student_export.comprehensive.health_appointments',
        header_keys=(
            'student_export.headers.date',
            'student_export.headers.type',
            'student_export.headers.provider',
            'student_export.headers.notes',
        ),
    ),
    dict(
        key='medical_conditions',
        title_key='student_export.comprehensive.medical_conditions',
        header_keys=(
            'student_export.headers.condition',
            'student_export.headers.diagnosed_date',
            'student_export.headers.status',
        ),
    ),
    dict(
        key='housing',
        title_key='student_export.comprehensive.housing_assignments',
        header_keys=(
            'student_export.headers.building',
            'student_export.headers.room_number',
            'student_export.headers.start_date',
            'student_export.headers.end_date',
            'student_export.headers.status',
        ),
        sheet_name_key='student_export.sheets.housing',
    ),
    dict(
        key='internships',
        title_key='student_export.comprehensive.internships_placements',
        header_keys=(
            'student_export.headers.company',
            'student_export.headers.position',
            'student_export.headers.status',
            'student_export.headers.application_date',
            'student_export.headers.start_date',
            'student_export.headers.end_date',
        ),
        sheet_name_key='student_export.sheets.internships',
    ),
    dict(
        key='library',
        title_key='student_export.comprehensive.library_records',
        header_keys=(
            'student_export.headers.book_title',
            'student_export.headers.checkout_date',
            'student_export.headers.due_date',
            'student_export.headers.return_date',
            'student_export.headers.status',
        ),
        sheet_name_key='student_export.sheets.library',
    ),
    dict(
        key='support_tickets',
        title_key='student_export.comprehensive.support_tickets',
        header_keys=(
            'student_export.headers.ticket_id',
            'student_export.headers.subject',
            'student_export.headers.category',
            'student_export.headers.status',
            'student_export.headers.created',
            'student_export.headers.resolved',
        ),
    ),
    dict(
        key='badges',
        title_key='student_export.comprehensive.badges_achievements',
        header_keys=(
            'student_export.headers.badge_name',
            'student_export.headers.date_earned',
            'student_export.headers.description',
        ),
        sheet_name_key='student_export.sheets.badges',
    ),
    dict(
        key='parking',
        title_key='student_export.comprehensive.parking_permits',
        header_keys=(
            'student_export.headers.permit_number',
            'student_export.headers.vehicle_make',
            'student_export.headers.vehicle_model',
            'student_export.headers.start_date',
            'student_export.headers.end_date',
            'student_export.headers.type',
        ),
        sheet_name_key='student_export.sheets.parking',
    ),
    dict(
        key='meal_plan',
        title_key='student_export.comprehensive.meal_plan',
        header_keys=(
            'student_export.headers.plan_type',
            'student_export.headers.balance',
            'student_export.headers.last_transaction_date',
        ),
    ),
    dict(
        key='meal_transactions',
        title_key='student_export.comprehensive.recent_transactions',
        header_keys=(
            'student_export.headers.date',
            'student_export.headers.location',
            'student_export.headers.amount',
            'student_export.headers.description',
        ),
    ),
    dict(
        key='advising',
        title_key='student_export.comprehensive.academic_advising',
        header_keys=(
            'student_export.headers.date',
            'student_export.headers.advisor',
            'student_export.headers.notes',
            'student_export.headers.follow_up_required',
        ),
    ),
)


def _resolved_headers(section):
    """Translate a section's header keys into display strings."""
    return [_t(k) for k in section['header_keys']]


def _excel_sheet_name(section, used_names):
    """Build an Excel-safe sheet name (<=31 chars, unique within a workbook)."""
    raw = _t(section.get('sheet_name_key') or section['title_key'])
    # Excel forbids these characters in sheet names
    for ch in '[]:*?/\\':
        raw = raw.replace(ch, '_')
    name = raw[:31] or section['key'][:31]
    if name not in used_names:
        used_names.add(name)
        return name
    # Deduplicate with a numeric suffix
    base = name[:28]
    for i in range(2, 100):
        candidate = f"{base}_{i}"
        if candidate not in used_names:
            used_names.add(candidate)
            return candidate
    used_names.add(name)
    return name

def export_student_data(self, student_id):
    """Export individual student data"""
    try:
        filename = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Text files", "*.txt"), ("PDF files", "*.pdf"), ("All files", "*.*")]
        )

        if filename:
            # Implementation for exporting student data
            messagebox.showinfo(_t("common.success"), _t("student_export.messages.data_exported", filename=filename))
    except Exception as e:
        messagebox.showerror(_t("common.error"), _t("student_export.errors.export_failed", error=str(e)))
def export_individual_student_data(self, student_id, first_name, last_name):
    """Export individual student data in various formats (PDF, CSV, Excel, TXT)"""
    try:
        # Create export format dialog
        export_dialog = tk.Toplevel(self.root)
        _install_clean_close(export_dialog)
        export_dialog.title(_t("student_export.dialog.title"))
        export_dialog.geometry("400x250")
        export_dialog.transient(self.root)
        export_dialog.grab_set()

        main_frame = ttk.Frame(export_dialog, padding=20)
        main_frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(main_frame, text=_t("student_export.dialog.export_data_for", first_name=first_name, last_name=last_name),
                 font=('TkDefaultFont', 12, 'bold')).pack(pady=(0, 15))

        ttk.Label(main_frame, text=_t("student_export.dialog.select_format")).pack(anchor=tk.W, pady=(0, 10))

        format_var = tk.StringVar(value="csv")
        formats = [
            (_t("student_export.formats.csv"), "csv"),
            (_t("student_export.formats.excel"), "xlsx"),
            (_t("student_export.formats.pdf"), "pdf"),
            (_t("student_export.formats.text"), "txt")
        ]

        for text, value in formats:
            ttk.Radiobutton(main_frame, text=text, variable=format_var, value=value).pack(anchor=tk.W, pady=2)

        def perform_export():
            selected_format = format_var.get()

            # Get file extension and type
            extensions = {
                "csv": (".csv", "CSV files", "*.csv"),
                "xlsx": (".xlsx", "Excel files", "*.xlsx"),
                "pdf": (".pdf", "PDF files", "*.pdf"),
                "txt": (".txt", "Text files", "*.txt")
            }

            ext, file_desc, file_pattern = extensions[selected_format]

            filename = filedialog.asksaveasfilename(
                defaultextension=ext,
                filetypes=[(file_desc, file_pattern), ("All files", "*.*")],
                initialfile=f"student_{student_id}_{first_name}_{last_name}{ext}"
            )

            if not filename:
                return

            # Fetch comprehensive student data
            conn = get_db_connection()
            if not conn:
                messagebox.showerror(_t("common.error"), _t("student_export.errors.db_connection_failed"))
                return

            cursor = conn.cursor()

            # Get all student-related data
            data_sections = {}

            # 1. Basic Student Information
            cursor.execute("SELECT * FROM students WHERE student_id = ?", (student_id,))
            data_sections['student_info'] = cursor.fetchone()

            # 2. Enrolled Modules
            cursor.execute("""
                SELECT sm.module_code, m.module_name, m.credits, sm.enrollment_date, sm.status
                FROM student_modules sm
                LEFT JOIN modules m ON sm.module_code = m.module_code
                WHERE sm.student_id = ?
                ORDER BY sm.enrollment_date DESC
            """, (student_id,))
            data_sections['modules'] = cursor.fetchall()

            # 3. Assignments
            cursor.execute("""
                SELECT a.module_code, a.title, a.due_date, a.max_marks, a.assignment_type
                FROM assignments a
                WHERE a.module_code IN (SELECT module_code FROM student_modules WHERE student_id = ?)
                ORDER BY a.due_date DESC
            """, (student_id,))
            data_sections['assignments'] = cursor.fetchall()

            # 4. Assignment Submissions
            cursor.execute("""
                SELECT a.title, a.module_code, s.submission_date, s.grade,
                       a.max_marks, s.status, s.late_submission
                FROM assignment_submissions s
                JOIN assignments a ON s.assignment_id = a.id
                WHERE s.student_id = ?
                ORDER BY s.submission_date DESC
            """, (student_id,))
            data_sections['submissions'] = cursor.fetchall()

            # 5. Grades/Assessments
            cursor.execute("""
                SELECT a.module_code, a.assessment_name, a.assessment_type,
                       g.score, a.max_points, g.letter_grade, g.submission_date
                FROM grades g
                JOIN assessments a ON g.assessment_id = a.assessment_id
                WHERE g.student_id = ?
                ORDER BY g.submission_date DESC
            """, (student_id,))
            data_sections['grades'] = cursor.fetchall()

            # 6. Attendance Records — primary table is attendance_records (enhanced tracker);
            # fall back to legacy `attendance` table for older deployments.
            try:
                cursor.execute("""
                    SELECT date, module_code, status, notes AS reason
                    FROM attendance_records
                    WHERE student_id = ?
                    ORDER BY date DESC, recorded_at DESC
                """, (student_id,))
                data_sections['attendance'] = cursor.fetchall()
            except Exception:
                data_sections['attendance'] = []
            if not data_sections['attendance']:
                try:
                    cursor.execute("""
                        SELECT date, module_code, status, reason
                        FROM attendance
                        WHERE student_id = ?
                        ORDER BY date DESC
                    """, (student_id,))
                    data_sections['attendance'] = cursor.fetchall()
                except Exception:
                    data_sections['attendance'] = []

            # 7. Financial Information
            try:
                cursor.execute("""
                    SELECT fee_type, amount, due_date, status
                    FROM student_fees
                    WHERE student_id = ?
                    ORDER BY due_date DESC
                """, (student_id,))
                data_sections['fees'] = cursor.fetchall()

                cursor.execute("""
                    SELECT payment_date, amount, payment_method, transaction_id
                    FROM payments
                    WHERE student_id = ?
                    ORDER BY payment_date DESC
                """, (student_id,))
                data_sections['payments'] = cursor.fetchall()
            except Exception:
                data_sections['fees'] = []
                data_sections['payments'] = []

            # 8. Financial Aid & Scholarships
            try:
                cursor.execute("""
                    SELECT s.scholarship_name, ss.amount, ss.academic_year, ss.status
                    FROM student_scholarships ss
                    JOIN scholarships s ON ss.scholarship_id = s.scholarship_id
                    WHERE ss.student_id = ?
                    ORDER BY ss.academic_year DESC
                """, (student_id,))
                data_sections['scholarships'] = cursor.fetchall()

                cursor.execute("""
                    SELECT aid_type, amount, academic_year, status
                    FROM student_financial_aid
                    WHERE student_id = ?
                    ORDER BY academic_year DESC
                """, (student_id,))
                data_sections['financial_aid'] = cursor.fetchall()
            except Exception:
                data_sections['scholarships'] = []
                data_sections['financial_aid'] = []

            # 9. Extracurricular Activities
            try:
                cursor.execute("""
                    SELECT activity_name, role, start_date, end_date, status
                    FROM student_activities
                    WHERE student_id = ?
                    ORDER BY start_date DESC
                """, (student_id,))
                data_sections['activities'] = cursor.fetchall()

                cursor.execute("""
                    SELECT c.club_name, cm.role, cm.join_date, cm.status
                    FROM club_members cm
                    JOIN clubs c ON cm.club_id = c.club_id
                    WHERE cm.student_id = ?
                    ORDER BY cm.join_date DESC
                """, (student_id,))
                data_sections['clubs'] = cursor.fetchall()
            except Exception:
                data_sections['activities'] = []
                data_sections['clubs'] = []

            # 10. Health Records
            try:
                cursor.execute("""
                    SELECT appointment_date, appointment_type, provider, notes
                    FROM health_appointments
                    WHERE student_id = ?
                    ORDER BY appointment_date DESC
                """, (student_id,))
                data_sections['health_appointments'] = cursor.fetchall()

                cursor.execute("""
                    SELECT condition_name, diagnosed_date, status
                    FROM medical_conditions
                    WHERE student_id = ?
                    ORDER BY diagnosed_date DESC
                """, (student_id,))
                data_sections['medical_conditions'] = cursor.fetchall()
            except Exception:
                data_sections['health_appointments'] = []
                data_sections['medical_conditions'] = []

            # 11. Housing Information
            try:
                cursor.execute("""
                    SELECT building_name, room_number, start_date, end_date, status
                    FROM housing_assignments
                    WHERE student_id = ?
                    ORDER BY start_date DESC
                """, (student_id,))
                data_sections['housing'] = cursor.fetchall()
            except Exception:
                data_sections['housing'] = []

            # 12. Internships & Career
            try:
                cursor.execute("""
                    SELECT i.company_name, i.position_title, ia.status,
                           ia.application_date, ip.start_date, ip.end_date
                    FROM internship_applications ia
                    JOIN internships i ON ia.internship_id = i.internship_id
                    LEFT JOIN internship_placements ip ON ia.application_id = ip.application_id
                    WHERE ia.student_id = ?
                    ORDER BY ia.application_date DESC
                """, (student_id,))
                data_sections['internships'] = cursor.fetchall()
            except Exception:
                data_sections['internships'] = []

            # 13. Library Records
            try:
                cursor.execute("""
                    SELECT b.title, bl.checkout_date, bl.due_date, bl.return_date, bl.status
                    FROM book_loans bl
                    JOIN books b ON bl.book_id = b.book_id
                    WHERE bl.student_id = ?
                    ORDER BY bl.checkout_date DESC
                """, (student_id,))
                data_sections['library'] = cursor.fetchall()
            except Exception:
                data_sections['library'] = []

            # 14. Support Tickets
            try:
                cursor.execute("""
                    SELECT ticket_id, subject, category, status, created_date, resolved_date
                    FROM support_tickets
                    WHERE student_id = ?
                    ORDER BY created_date DESC
                """, (student_id,))
                data_sections['support_tickets'] = cursor.fetchall()
            except Exception:
                data_sections['support_tickets'] = []

            # 15. Badges & Achievements
            try:
                cursor.execute("""
                    SELECT ab.badge_name, bi.date_earned, bi.description
                    FROM badge_issuances bi
                    JOIN achievement_badges ab ON bi.badge_id = ab.badge_id
                    WHERE bi.student_id = ?
                    ORDER BY bi.date_earned DESC
                """, (student_id,))
                data_sections['badges'] = cursor.fetchall()
            except Exception:
                data_sections['badges'] = []

            # 16. Parking & Transportation
            try:
                cursor.execute("""
                    SELECT permit_number, vehicle_make, vehicle_model,
                           start_date, end_date, permit_type
                    FROM parking_permits
                    WHERE student_id = ?
                    ORDER BY start_date DESC
                """, (student_id,))
                data_sections['parking'] = cursor.fetchall()
            except Exception:
                data_sections['parking'] = []

            # 17. Meal Plans
            try:
                cursor.execute("""
                    SELECT plan_type, balance, last_transaction_date
                    FROM meal_accounts
                    WHERE student_id = ?
                """, (student_id,))
                data_sections['meal_plan'] = cursor.fetchall()

                cursor.execute("""
                    SELECT created_at, location, amount, description
                    FROM transactions
                    WHERE source_type = 'meal' AND student_id = ?
                    ORDER BY created_at DESC
                    LIMIT 50
                """, (student_id,))
                data_sections['meal_transactions'] = cursor.fetchall()
            except Exception:
                data_sections['meal_plan'] = []
                data_sections['meal_transactions'] = []

            # 18. Academic Advising
            try:
                cursor.execute("""
                    SELECT appointment_date, advisor_name, notes, follow_up_required
                    FROM advising_appointments
                    WHERE student_id = ?
                    ORDER BY appointment_date DESC
                """, (student_id,))
                data_sections['advising'] = cursor.fetchall()
            except Exception:
                data_sections['advising'] = []

            conn.close()

            # Export based on format
            if selected_format == "csv":
                self._export_comprehensive_csv(filename, student_id, first_name, last_name, data_sections)
            elif selected_format == "xlsx":
                self._export_comprehensive_excel(filename, student_id, first_name, last_name, data_sections)
            elif selected_format == "pdf":
                self._export_comprehensive_pdf(filename, student_id, first_name, last_name, data_sections)
            elif selected_format == "txt":
                self._export_comprehensive_txt(filename, student_id, first_name, last_name, data_sections)

            messagebox.showinfo(_t("common.success"), _t("student_export.messages.comprehensive_export_success", filename=filename))
            export_dialog.destroy()

        # Buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(pady=(15, 0))
        ttk.Button(button_frame, text=_t("common.export"), command=perform_export).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text=_t("common.cancel"), command=export_dialog.destroy).pack(side=tk.LEFT, padx=5)

    except Exception as e:
        messagebox.showerror(_t("common.error"), _t("student_export.errors.export_failed", error=str(e)))
def export_data_dialog(self):
    """Create export data dialog"""
    dialog = self.create_themed_toplevel(_t("student_export.dialog.title"), "450x300")
    dialog.grab_set()

    main_frame = ttk.Frame(dialog, padding=20)
    main_frame.pack(fill=tk.BOTH, expand=True)

    ttk.Label(main_frame, text=_t("student_export.dialog.export_format")).grid(row=0, column=0, sticky=tk.W, pady=10)
    format_var = tk.StringVar(value="CSV")

    formats = [(_t("student_export.formats.csv_short"), "CSV"), (_t("student_export.formats.excel_short"), "Excel"), (_t("student_export.formats.pdf_short"), "PDF"), (_t("student_export.formats.text_short"), "Text")]
    for i, (text, value) in enumerate(formats):
        ttk.Radiobutton(main_frame, text=text, variable=format_var, value=value).grid(row=i+1, column=0, sticky=tk.W)

    def export_data():
        """Export data in selected format"""
        try:
            export_format = format_var.get()

            filetypes = {
                "CSV": [("CSV files", "*.csv")],
                "Excel": [("Excel files", "*.xlsx")],
                "PDF": [("PDF files", "*.pdf")],
                "Text": [("Text files", "*.txt")]
            }

            extensions = {
                "CSV": ".csv",
                "Excel": ".xlsx",
                "PDF": ".pdf",
                "Text": ".txt"
            }

            filename = filedialog.asksaveasfilename(
                defaultextension=extensions[export_format],
                filetypes=filetypes[export_format]
            )

            if not filename:
                return

            conn = None
            try:
                conn = get_db_connection()
                if not conn:
                    messagebox.showerror(_t("student_export.errors.export_failed_title"), _t("student_export.errors.db_connection_failed"))
                    return

                cursor = conn.cursor()
                cursor.execute(
                    """
                    SELECT s.student_id,
                           COALESCE(s.title, ''),
                           COALESCE(s.first_name, ''),
                           COALESCE(s.middle_name, ''),
                           COALESCE(s.last_name, ''),
                           COALESCE(s.gender, ''),
                           COALESCE(s.dob, ''),
                           COALESCE(s.age, ''),
                           COALESCE(s.email_address, ''),
                           COALESCE(s.course, ''),
                           COALESCE(s.status, ''),
                           COALESCE(DATE(s.registration_datetime), '')
                    FROM students s
                    ORDER BY s.last_name, s.first_name
                    """
                )

                student_rows = cursor.fetchall()

                if not student_rows:
                    messagebox.showinfo(_t("common.export"), _t("student_export.messages.no_records"))
                    return

                # Build module listings safely without assuming specific schemas
                modules_by_student = {}
                try:
                    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='student_modules'")
                    has_student_modules = cursor.fetchone() is not None

                    if has_student_modules:
                        cursor.execute("PRAGMA table_info(student_modules)")
                        student_module_cols = {row[1] for row in cursor.fetchall()}

                        identifier_col = None
                        if 'module_code' in student_module_cols:
                            identifier_col = 'module_code'
                        elif 'module_id' in student_module_cols:
                            identifier_col = 'module_id'

                        name_col = 'module_name' if 'module_name' in student_module_cols else None

                        # Whitelist of allowed column names for SQL injection prevention
                        ALLOWED_COLUMNS = {
                            'student_id', 'module_id', 'module_code', 'module_name'
                        }

                        module_lookup = {}
                        if identifier_col and 'student_id' in student_module_cols:
                            try:
                                cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='modules'")
                                has_modules_table = cursor.fetchone() is not None
                            except sqlite3.Error:
                                has_modules_table = False

                            if has_modules_table:
                                cursor.execute("PRAGMA table_info(modules)")
                                module_table_cols = {row[1] for row in cursor.fetchall()}

                                if 'module_name' in module_table_cols:
                                    join_key = None
                                    if identifier_col in module_table_cols:
                                        join_key = identifier_col
                                    elif identifier_col == 'module_code' and 'module_id' in module_table_cols:
                                        join_key = 'module_id'

                                    # Validate join_key against whitelist before using in query
                                    if join_key and join_key in ALLOWED_COLUMNS:
                                        safe_join_key = validate_identifier(join_key, "column")
                                        cursor.execute("SELECT " + safe_join_key + ", module_name FROM modules")
                                        module_lookup = {
                                            row[0]: row[1] for row in cursor.fetchall() if row[0] is not None
                                        }

                            select_cols = ["student_id", identifier_col]
                            if name_col:
                                select_cols.append(name_col)

                            # Validate all columns against whitelist before building query
                            if all(col in ALLOWED_COLUMNS for col in select_cols):
                                validated_cols = [validate_identifier(col, "column") for col in select_cols]
                                column_clause = ", ".join(validated_cols)
                                cursor.execute("SELECT " + column_clause + " FROM student_modules")

                                for module_row in cursor.fetchall():
                                    student_id = module_row[0]
                                    identifier = module_row[1]
                                    module_name = None

                                    if name_col and len(module_row) > 2:
                                        module_name = module_row[2]

                                    if not module_name and module_lookup:
                                        module_name = module_lookup.get(identifier)

                                    display_identifier = identifier or 'N/A'
                                    display_name = module_name or module_lookup.get(identifier) or 'Unknown Module'

                                    modules_by_student.setdefault(student_id, []).append(
                                        f"{display_identifier} - {display_name}"
                                    )
                except sqlite3.Error:
                    modules_by_student = {}

                headers = [
                    _t("student_export.headers.student_id"),
                    "Title",
                    _t("student_export.headers.first_name"),
                    "Middle Name",
                    _t("student_export.headers.last_name"),
                    "Gender",
                    "DOB",
                    "Age",
                    _t("student_export.headers.email"),
                    _t("student_export.headers.course"),
                    "Status",
                    _t("student_export.headers.registration_date"),
                    _t("student_export.headers.module_1"),
                    _t("student_export.headers.module_2"),
                    _t("student_export.headers.module_3"),
                    _t("student_export.headers.module_4"),
                    _t("student_export.headers.module_5"),
                    _t("student_export.headers.module_6")
                ]

                processed_rows = []
                for row in student_rows:
                    student_id = row[0]
                    modules_list = modules_by_student.get(student_id, [])

                    # Create row with all student columns + 6 module columns
                    row_data = [
                        row[0],   # Student ID
                        row[1],   # Title
                        row[2],   # First Name
                        row[3],   # Middle Name
                        row[4],   # Last Name
                        row[5],   # Gender
                        row[6],   # DOB
                        row[7],   # Age
                        row[8],   # Email
                        row[9],   # Course
                        row[10],  # Status
                        row[11],  # Registration Date
                    ]

                    # Add up to 6 modules in separate columns
                    for i in range(6):
                        if i < len(modules_list):
                            row_data.append(modules_list[i])
                        else:
                            row_data.append("")

                    processed_rows.append(row_data)

                if export_format == "CSV":
                    with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                        writer = csv.writer(csvfile)  # nosemgrep: python.lang.security.audit.csv-writer-injection
                        writer.writerow(headers)
                        writer.writerows(processed_rows)

                elif export_format == "Excel":
                    try:
                        import pandas as pd
                    except ImportError:
                        messagebox.showerror(
                            _t("student_export.errors.export_failed_title"),
                            _t("student_export.errors.pandas_required")
                        )
                        return

                    df = pd.DataFrame(processed_rows, columns=headers)
                    try:
                        df.to_excel(filename, index=False, engine='openpyxl')
                    except ImportError:
                        messagebox.showerror(
                            _t("student_export.errors.export_failed_title"),
                            _t("student_export.errors.openpyxl_required")
                        )
                        return

                elif export_format == "PDF":
                    try:
                        from reportlab.lib import colors
                        from reportlab.lib.pagesizes import letter, landscape, A4
                        from reportlab.lib.styles import getSampleStyleSheet
                        from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle
                        from reportlab.lib.units import inch
                    except ImportError:
                        messagebox.showerror(
                            _t("student_export.errors.export_failed_title"),
                            _t("student_export.errors.reportlab_required")
                        )
                        return

                    # Use landscape orientation for better column visibility
                    doc = SimpleDocTemplate(
                        filename,
                        pagesize=landscape(letter),
                        leftMargin=0.4*inch,
                        rightMargin=0.4*inch,
                        topMargin=0.5*inch,
                        bottomMargin=0.5*inch
                    )
                    styles = getSampleStyleSheet()
                    body_style = styles['BodyText']
                    body_style.wordWrap = 'CJK'
                    body_style.fontSize = 8

                    # Create a style for wrapping cell text
                    cell_style = styles['Normal'].clone('CellStyle')
                    cell_style.fontSize = 7
                    cell_style.leading = 8
                    cell_style.wordWrap = 'CJK'

                    header_style = styles['Normal'].clone('HeaderStyle')
                    header_style.fontSize = 7
                    header_style.leading = 8
                    header_style.fontName = 'Helvetica-Bold'

                    table_data = [[Paragraph(str(h), header_style) for h in headers]]
                    for row in processed_rows:
                        pdf_row = []
                        for value in row:
                            val_str = str(value) if value else ""
                            pdf_row.append(Paragraph(val_str, cell_style))
                        table_data.append(pdf_row)

                    # Calculate available width
                    page_width = landscape(letter)[0] - (0.8 * inch)  # Subtract margins (0.4 * 2)

                    # Calculate column widths dynamically based on number of columns
                    num_cols = len(headers)
                    # Fixed columns: Student ID, Title, First Name, Middle Name, Last Name,
                    #                Gender, DOB, Age, Email, Course, Status, Reg Date
                    fixed_widths = [
                        0.5*inch,   # Student ID
                        0.35*inch,  # Title
                        0.6*inch,   # First Name
                        0.55*inch,  # Middle Name
                        0.6*inch,   # Last Name
                        0.45*inch,  # Gender
                        0.6*inch,   # DOB
                        0.3*inch,   # Age
                        1.0*inch,   # Email
                        0.55*inch,  # Course
                        0.45*inch,  # Status
                        0.65*inch,  # Registration Date
                    ]
                    fixed_total = sum(fixed_widths)
                    num_module_cols = max(num_cols - 12, 0)
                    if num_module_cols > 0:
                        remaining = page_width - fixed_total
                        module_width = remaining / num_module_cols
                        col_widths = fixed_widths + [module_width] * num_module_cols
                    else:
                        col_widths = fixed_widths[:num_cols]

                    table = Table(table_data, colWidths=col_widths, repeatRows=1)
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('TOPPADDING', (0, 0), (-1, -1), 3),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                        ('LEFTPADDING', (0, 0), (-1, -1), 3),
                        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ]))

                    doc.build([table])

                else:  # Text export
                    with open(filename, 'w', encoding='utf-8') as txtfile:
                        for row in processed_rows:
                            txtfile.write(f"{_t('student_export.text_labels.student_id')}: {row[0]}\n")
                            full_name = " ".join(p for p in (row[1], row[2], row[3], row[4]) if p)
                            txtfile.write(f"{_t('student_export.text_labels.name')}: {full_name}\n")
                            txtfile.write(f"Gender: {row[5]}\n")
                            txtfile.write(f"DOB: {row[6]}\n")
                            txtfile.write(f"Age: {row[7]}\n")
                            txtfile.write(f"{_t('student_export.text_labels.email')}: {row[8]}\n")
                            txtfile.write(f"{_t('student_export.text_labels.course')}: {row[9]}\n")
                            txtfile.write(f"Status: {row[10]}\n")
                            txtfile.write(f"{_t('student_export.text_labels.registration_date')}: {row[11]}\n")
                            txtfile.write(f"{_t('student_export.text_labels.modules')}:\n")
                            # Write each module on a separate line (modules start at index 12)
                            for i in range(12, 18):
                                if i < len(row) and row[i]:
                                    txtfile.write(f"  {_t('student_export.text_labels.module')} {i-11}: {row[i]}\n")
                            txtfile.write("-" * 60 + "\n\n")

                messagebox.showinfo(_t("common.success"), _t("student_export.messages.export_success", filename=filename))
                dialog.destroy()

            finally:
                if conn:
                    conn.close()

        except Exception as e:
            messagebox.showerror(_t("common.error"), _t("student_export.errors.export_failed", error=str(e)))

    # Buttons
    button_frame = ttk.Frame(main_frame)
    button_frame.grid(row=5, column=0, pady=20)

    ttk.Button(button_frame, text=_t("common.export"), command=export_data).pack(side=tk.LEFT, padx=5)
    ttk.Button(button_frame, text=_t("common.cancel"), command=dialog.destroy).pack(side=tk.LEFT, padx=5)
def _export_to_csv(self, filename, student_data, modules_data, grades_data, attendance_data):
    """Export student data to CSV format"""
    with open(filename, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)  # nosemgrep: python.lang.security.audit.csv-writer-injection

        # Convert Row objects to tuples
        student_data = tuple(student_data)
        modules_data = [tuple(m) for m in modules_data]
        grades_data = [tuple(g) for g in grades_data]
        attendance_data = [tuple(a) for a in attendance_data]

        # Student info
        writer.writerow([_t('student_export.sections.student_info')])
        writer.writerow([_t('student_export.headers.student_id'), _t('student_export.headers.email'), _t('student_export.text_labels.name'), _t('student_export.headers.course'), _t('student_export.headers.registration_date')])
        writer.writerow([student_data[0], student_data[1],
                       f"{student_data[3]} {student_data[5]}",
                       student_data[9], student_data[10]])
        writer.writerow([])

        # Modules
        writer.writerow([_t('student_export.sections.enrolled_modules')])
        writer.writerow([_t('student_export.headers.module_code'), _t('student_export.headers.module_name'), _t('student_export.headers.credits')])
        for module in modules_data:
            writer.writerow(module)
        writer.writerow([])

        # Grades
        writer.writerow([_t('student_export.sections.grades')])
        writer.writerow([_t('student_export.headers.module_code'), _t('student_export.headers.assessment'), _t('student_export.headers.grade'), _t('student_export.headers.max_grade')])
        for grade in grades_data:
            writer.writerow(grade)
        writer.writerow([])

        # Attendance
        writer.writerow([_t('student_export.sections.attendance')])
        writer.writerow([_t('student_export.headers.date'), _t('student_export.headers.module_code'), _t('student_export.headers.status')])
        for attendance in attendance_data:
            writer.writerow(attendance)
def _export_to_excel(self, filename, student_id, first_name, last_name,
                    student_data, modules_data, grades_data, attendance_data):
    """Export student data to Excel format"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment

        # Convert Row objects to tuples
        student_data = tuple(student_data)
        modules_data = [tuple(m) for m in modules_data]
        grades_data = [tuple(g) for g in grades_data]
        attendance_data = [tuple(a) for a in attendance_data]

        wb = openpyxl.Workbook()

        # Student Info sheet
        ws1 = wb.active
        ws1.title = _t('student_export.sheets.student_info')
        ws1['A1'] = _t('student_export.sections.student_info_title', first_name=first_name, last_name=last_name)
        ws1['A1'].font = Font(bold=True, size=14)
        ws1.append([_t('student_export.headers.student_id'), student_data[0]])
        ws1.append([_t('student_export.headers.email'), student_data[1]])
        ws1.append([_t('student_export.text_labels.name'), f"{student_data[3]} {student_data[5]}"])
        ws1.append([_t('student_export.headers.course'), student_data[9]])
        ws1.append([_t('student_export.headers.registration_date'), student_data[10]])

        # Modules sheet
        ws2 = wb.create_sheet(_t('student_export.sheets.modules'))
        ws2['A1'] = _t('student_export.sections.enrolled_modules')
        ws2['A1'].font = Font(bold=True, size=12)
        ws2.append([_t('student_export.headers.module_code'), _t('student_export.headers.module_name'), _t('student_export.headers.credits')])
        for module in modules_data:
            ws2.append(list(module))

        # Grades sheet
        ws3 = wb.create_sheet(_t('student_export.sheets.grades'))
        ws3['A1'] = _t('student_export.sections.grades')
        ws3['A1'].font = Font(bold=True, size=12)
        ws3.append([_t('student_export.headers.module_code'), _t('student_export.headers.assessment'), _t('student_export.headers.grade'), _t('student_export.headers.max_grade')])
        for grade in grades_data:
            ws3.append(list(grade))

        # Attendance sheet
        ws4 = wb.create_sheet(_t('student_export.sheets.attendance'))
        ws4['A1'] = _t('student_export.sections.attendance_records')
        ws4['A1'].font = Font(bold=True, size=12)
        ws4.append([_t('student_export.headers.date'), _t('student_export.headers.module_code'), _t('student_export.headers.status')])
        for attendance in attendance_data:
            ws4.append(list(attendance))

        wb.save(filename)
    except ImportError:
        # Fallback to CSV if openpyxl not available
        messagebox.showwarning(_t("common.warning"), _t("student_export.errors.openpyxl_fallback_csv"))
        self._export_to_csv(filename.replace('.xlsx', '.csv'), student_data, modules_data, grades_data, attendance_data)
def _export_to_pdf(self, filename, student_id, first_name, last_name,
                   student_data, modules_data, grades_data, attendance_data):
    """Export student data to PDF format"""
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

        # Convert Row objects to tuples
        student_data = tuple(student_data)
        modules_data = [list(m) for m in modules_data]
        grades_data = [list(g) for g in grades_data]
        attendance_data = [list(a) for a in attendance_data]

        doc = SimpleDocTemplate(filename, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title = Paragraph(f"<b>{_t('student_export.pdf.student_report')}: {first_name} {last_name}</b>", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 20))

        # Student info table
        student_info = [
            [f"{_t('student_export.headers.student_id')}:", student_data[0]],
            [f"{_t('student_export.headers.email')}:", student_data[1]],
            [f"{_t('student_export.text_labels.name')}:", f"{student_data[3]} {student_data[5]}"],
            [f"{_t('student_export.headers.course')}:", student_data[9]],
            [f"{_t('student_export.headers.registration_date')}:", student_data[10]]
        ]
        t1 = Table(student_info, colWidths=[120, 350])
        t1.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, -1), 'Helvetica', 10),
            ('FONT', (0, 0), (0, -1), 'Helvetica-Bold', 10),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(t1)
        elements.append(Spacer(1, 20))

        # Modules
        elements.append(Paragraph(f"<b>{_t('student_export.sections.enrolled_modules')}</b>", styles['Heading2']))
        if modules_data:
            module_table = [[_t('student_export.headers.module_code'), _t('student_export.headers.module_name'), _t('student_export.headers.credits')]] + modules_data
            t2 = Table(module_table)
            t2.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(t2)
        elements.append(Spacer(1, 20))

        # Grades
        elements.append(Paragraph(f"<b>{_t('student_export.sections.grades')}</b>", styles['Heading2']))
        if grades_data:
            grade_table = [[_t('student_export.headers.module'), _t('student_export.headers.assessment'), _t('student_export.headers.grade'), _t('student_export.headers.max')]] + grades_data
            t3 = Table(grade_table)
            t3.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(t3)

        doc.build(elements)
    except ImportError:
        # Fallback to TXT if reportlab not available
        messagebox.showwarning(_t("common.warning"), _t("student_export.errors.reportlab_fallback_txt"))
        self._export_to_txt(filename.replace('.pdf', '.txt'), student_id, first_name, last_name,
                           student_data, modules_data, grades_data, attendance_data)
def _export_to_txt(self, filename, student_id, first_name, last_name,
                   student_data, modules_data, grades_data, attendance_data):
    """Export student data to text format"""
    # Convert Row objects to tuples
    student_data = tuple(student_data)
    modules_data = [tuple(m) for m in modules_data]
    grades_data = [tuple(g) for g in grades_data]
    attendance_data = [tuple(a) for a in attendance_data]

    with open(filename, 'w', encoding='utf-8') as f:
        f.write(f"=" * 80 + "\n")
        f.write(f"{_t('student_export.pdf.student_report').upper()}: {first_name} {last_name}\n")
        f.write(f"=" * 80 + "\n\n")

        # Student info
        f.write(f"{_t('student_export.sections.student_info').upper()}\n")
        f.write("-" * 80 + "\n")
        f.write(f"{_t('student_export.headers.student_id')}: {student_data[0]}\n")
        f.write(f"{_t('student_export.headers.email')}: {student_data[1]}\n")
        f.write(f"{_t('student_export.text_labels.name')}: {student_data[3]} {student_data[5]}\n")
        f.write(f"{_t('student_export.headers.course')}: {student_data[9]}\n")
        f.write(f"{_t('student_export.headers.registration_date')}: {student_data[10]}\n\n")

        # Modules
        f.write(f"{_t('student_export.sections.enrolled_modules').upper()}\n")
        f.write("-" * 80 + "\n")
        for module in modules_data:
            f.write(f"{module[0]} - {module[1]} ({module[2]} {_t('student_export.headers.credits').lower()})\n")
        f.write("\n")

        # Grades
        f.write(f"{_t('student_export.sections.grades').upper()}\n")
        f.write("-" * 80 + "\n")
        for grade in grades_data:
            f.write(f"{grade[0]} - {grade[1]}: {grade[2]}/{grade[3]}\n")
        f.write("\n")

        # Attendance
        f.write(f"{_t('student_export.sections.attendance_records').upper()}\n")
        f.write("-" * 80 + "\n")
        for attendance in attendance_data:
            f.write(f"{attendance[0]} - {attendance[1]}: {attendance[2]}\n")
def _export_comprehensive_csv(self, filename, student_id, first_name, last_name, data_sections):
    """Export comprehensive student data to CSV format with clear sections"""
    with open(filename, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)  # nosemgrep: python.lang.security.audit.csv-writer-injection
        student_data = tuple(data_sections['student_info']) if data_sections['student_info'] else ()

        writer.writerow([f"{_t('student_export.comprehensive.title')} - {first_name} {last_name}"])
        writer.writerow([f"{_t('student_export.comprehensive.generated')}: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        writer.writerow([])

        writer.writerow(['=' * 80])
        writer.writerow([_t('student_export.comprehensive.basic_info')])
        writer.writerow(['=' * 80])
        if student_data:
            writer.writerow([_t('student_export.headers.student_id'), student_data[0]])
            writer.writerow([_t('student_export.headers.email'), student_data[1]])
            writer.writerow([_t('student_export.text_labels.name'), f"{student_data[3]} {student_data[5]}"])
            writer.writerow([_t('student_export.headers.course'), student_data[9]])
            writer.writerow([_t('student_export.headers.registration_date'), student_data[10]])
        writer.writerow([])

        for section in EXPORT_SECTIONS:
            rows = data_sections.get(section['key']) or []
            if not rows:
                continue
            writer.writerow(['=' * 80])
            writer.writerow([_t(section['title_key'])])
            writer.writerow(['=' * 80])
            writer.writerow(_resolved_headers(section))
            for row in rows:
                writer.writerow(tuple(row))
            writer.writerow([])
def _export_comprehensive_excel(self, filename, student_id, first_name, last_name, data_sections):
    """Export comprehensive student data to Excel with multiple sheets"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        used_sheet_names = set()

        def create_sheet(sheet_name, headers, data):
            ws = wb.create_sheet(sheet_name)
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, value in enumerate(tuple(row_data), 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

        student_data = tuple(data_sections['student_info']) if data_sections['student_info'] else ()
        info_sheet = _excel_sheet_name(
            {'sheet_name_key': 'student_export.sheets.student_info',
             'title_key': 'student_export.sheets.student_info', 'key': 'student_info'},
            used_sheet_names,
        )
        ws1 = wb.create_sheet(info_sheet)
        ws1['A1'] = _t('student_export.sections.student_info_title', first_name=first_name, last_name=last_name)
        ws1['A1'].font = Font(bold=True, size=14)
        if student_data:
            ws1.append([_t('student_export.headers.student_id'), student_data[0]])
            ws1.append([_t('student_export.headers.email'), student_data[1]])
            ws1.append([_t('student_export.text_labels.name'), f"{student_data[3]} {student_data[5]}"])
            ws1.append([_t('student_export.headers.course'), student_data[9]])
            ws1.append([_t('student_export.headers.registration_date'), student_data[10]])

        for section in EXPORT_SECTIONS:
            rows = data_sections.get(section['key']) or []
            if not rows:
                continue
            create_sheet(_excel_sheet_name(section, used_sheet_names), _resolved_headers(section), rows)

        wb.save(filename)
    except ImportError:
        messagebox.showwarning(_t("common.warning"), _t("student_export.errors.openpyxl_fallback_csv"))
        self._export_comprehensive_csv(filename.replace('.xlsx', '.csv'), student_id, first_name, last_name, data_sections)
def _export_comprehensive_pdf(self, filename, student_id, first_name, last_name, data_sections):
    """Export comprehensive student data to PDF format"""
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.units import inch

        doc = SimpleDocTemplate(filename, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []
        styles = getSampleStyleSheet()

        heading_style = ParagraphStyle('CustomHeading', parent=styles['Heading2'],
                                      fontSize=14, textColor=colors.HexColor('#003366'))

        elements.append(Paragraph(
            f"<b>{_t('student_export.comprehensive.title').upper()}</b><br/>{first_name} {last_name}",
            styles['Title']))
        elements.append(Spacer(1, 0.3*inch))

        # PDFs can't reasonably hold unlimited rows or very wide cells; cap both.
        PDF_ROW_LIMIT = 20
        PDF_CELL_CHAR_LIMIT = 60

        def add_section(section_title, headers, data):
            if not data:
                return
            elements.append(Paragraph(f"<b>{section_title}</b>", heading_style))
            elements.append(Spacer(1, 0.1*inch))

            truncated = [
                [(str(v)[:PDF_CELL_CHAR_LIMIT] if v is not None else '') for v in tuple(row)]
                for row in data[:PDF_ROW_LIMIT]
            ]
            table_data = [list(headers)] + truncated
            t = Table(table_data, repeatRows=1)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            elements.append(t)
            if len(data) > PDF_ROW_LIMIT:
                elements.append(Paragraph(
                    f"<i>... and {len(data) - PDF_ROW_LIMIT} more rows (see CSV/XLSX/TXT export for full data)</i>",
                    styles['Normal']))
            elements.append(Spacer(1, 0.2*inch))

        student_data = tuple(data_sections['student_info']) if data_sections['student_info'] else ()
        if student_data:
            info_data = [
                [_t('student_export.headers.student_id'), student_data[0]],
                [_t('student_export.headers.email'), student_data[1]],
                [_t('student_export.text_labels.name'), f"{student_data[3]} {student_data[5]}"],
                [_t('student_export.headers.course'), student_data[9]],
            ]
            add_section(_t('student_export.comprehensive.basic_info'),
                        [_t('student_export.headers.field'), _t('student_export.headers.value')],
                        info_data)

        for section in EXPORT_SECTIONS:
            rows = data_sections.get(section['key']) or []
            if not rows:
                continue
            add_section(_t(section['title_key']), _resolved_headers(section), rows)

        doc.build(elements)
    except ImportError:
        messagebox.showwarning(_t("common.warning"), _t("student_export.errors.reportlab_fallback_txt"))
        self._export_comprehensive_txt(filename.replace('.pdf', '.txt'), student_id, first_name, last_name, data_sections)
def _export_comprehensive_txt(self, filename, student_id, first_name, last_name, data_sections):
    """Export comprehensive student data to text format"""
    with open(filename, 'w', encoding='utf-8') as f:
        student_data = tuple(data_sections['student_info']) if data_sections['student_info'] else ()

        f.write("=" * 80 + "\n")
        f.write(f"{_t('student_export.comprehensive.title').upper()}: {first_name} {last_name}\n")
        f.write(f"{_t('student_export.comprehensive.generated')}: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("=" * 80 + "\n\n")

        def write_section(section_title, headers, data):
            if not data:
                return
            f.write("\n" + "=" * 80 + "\n")
            f.write(f"{section_title}\n")
            f.write("=" * 80 + "\n")
            f.write(" | ".join(headers) + "\n")
            f.write("-" * 80 + "\n")
            for row in data:
                f.write(" | ".join(str(v) if v is not None else '' for v in tuple(row)) + "\n")

        f.write(f"{_t('student_export.comprehensive.basic_info')}\n")
        f.write("-" * 80 + "\n")
        if student_data:
            f.write(f"{_t('student_export.headers.student_id')}: {student_data[0]}\n")
            f.write(f"{_t('student_export.headers.email')}: {student_data[1]}\n")
            f.write(f"{_t('student_export.text_labels.name')}: {student_data[3]} {student_data[5]}\n")
            f.write(f"{_t('student_export.headers.course')}: {student_data[9]}\n")
            f.write(f"{_t('student_export.headers.registration_date')}: {student_data[10]}\n")

        for section in EXPORT_SECTIONS:
            rows = data_sections.get(section['key']) or []
            if not rows:
                continue
            write_section(_t(section['title_key']), _resolved_headers(section), rows)
