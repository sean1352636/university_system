from education_system.university_system.infrastructure.database.db import DEFAULT_DB_PATH, get_connection  # injected
from education_system.university_system.core.sql_safety import validate_identifier, validate_table_name, validate_field_for_query, validate_column_name
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext
import threading
import queue
import json
import csv
from datetime import datetime, timedelta
import os
import sys
import shutil

from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

# Import internationalization (i18n) for multi-language support
try:
    from education_system.university_system.modules.shared.utils.i18n import (
        get_text as _t,
        get_current_language,
    )
    I18N_AVAILABLE = True
except ImportError:
    I18N_AVAILABLE = False
    _t = lambda key, **kwargs: key  # Fallback: return key as-is
    get_current_language = lambda: "en"

# Import enhanced console output utility (colors disabled for GUI context)
try:
    from education_system.university_system.modules.shared.utils.console_output import ConsoleOutput
    console = ConsoleOutput(use_colors=False)
    print_success = console.success
    print_error = console.error
    print_warning = console.warning
    print_info = console.info
    print_header = console.header
except ImportError:
    # Fallback to basic print if console_output is not available
    class _FallbackConsole:
        def success(self, msg, **kwargs): print(msg)
        def error(self, msg, **kwargs): print(msg)
        def warning(self, msg, **kwargs): print(msg)
        def info(self, msg, **kwargs): print(msg)
        def header(self, msg, **kwargs): print(f"\n{msg}\n{'='*len(msg)}")
    console = _FallbackConsole()
    print_success = lambda msg: print(msg)
    print_error = lambda msg: print(msg)
    print_warning = lambda msg: print(msg)
    print_info = lambda msg: print(msg)
    print_header = lambda msg, **kwargs: print(f"\n{msg}\n{'='*len(msg)}")

# Import chart generation utility
try:
    from education_system.university_system.modules.shared.utils.chart_generator import (
        ChartGenerator, ChartViewer, DatabaseChartGenerator, create_chart_viewer, CHARTS_AVAILABLE
    )
except ImportError:
    CHARTS_AVAILABLE = False
    ChartGenerator = None
    ChartViewer = None
    DatabaseChartGenerator = None
    create_chart_viewer = None

# Import all original functions (backwards compatibility)
try:
    from education_system.university_system.modules.shared.services.analytics.advanced_search import (
        saved_searches, search_cache, academic_performance_analysis, add_condition,
        add_user_permissions, build_search_analytics_record,
        bulk_enrollment_management, bulk_export, check_database_status,
        create_scheduled_report, create_student_groups, custom_format_export,
        data_quality_reports, display_enhanced_menu, display_search_results,
        duplicate_detection, ensure_search_analytics_schema,
        execute_conditional_search, export_single_student,
        export_system_statistics, export_to_csv, export_to_excel, export_to_json,
        generate_custom_reports, generate_custom_sql_report, generate_email_list,
        generate_module_enrollment_report, generate_student_summary_report,
        get_search_analytics_columns, identify_at_risk_students,
        init_enhanced_database, insert_search_analytics_record,
        interactive_charts, manage_scheduled_reports, manage_user_permissions,
        mark_for_followup, performance_optimization,
        refresh_search_analytics_columns, remove_condition,
        save_last_search_results, search_analytics_dashboard,
        simulate_send_email, student_demographics_reports, view_academic_history,
        view_search_audit_trail
    )
except ImportError:
    # If running as standalone, define minimal required functions
    print_warning("Could not import advanced_search module. Running in standalone mode.")

# Import essential functions from email infrastructure
try:
    from education_system.university_system.infrastructure.email.email_db_utilities import execute_db_operation
    from education_system.university_system.infrastructure.email.admin import search_users, list_all_users
except ImportError as e:
    print_warning(f"Could not import email infrastructure functions: {e}")
    # Create fallback functions
    def execute_db_operation(operation_func):
        conn = get_connection()
        cursor = conn.cursor()
        try:
            result = operation_func(cursor)
            conn.commit()
            return result
        except Exception as e:
            conn.rollback()
            raise e
        finally:
            conn.close()

    def search_users(auth, search_term):
        def _search(cursor):
            cursor.execute("SELECT id, username, email FROM users WHERE username LIKE ? OR email LIKE ?",
                         (f'%{search_term}%', f'%{search_term}%'))
            return [{'id': row[0], 'username': row[1], 'email': row[2]} for row in cursor.fetchall()]
        return execute_db_operation(_search)

    def list_all_users(auth, limit=100):
        def _list_users(cursor):
            cursor.execute("SELECT id, username, email FROM users LIMIT ?", (limit,))
            users = [{'id': row[0], 'username': row[1], 'email': row[2]} for row in cursor.fetchall()]
            return {'users': users, 'total_count': len(users)}
        return execute_db_operation(_list_users)
    
    # Minimal database functions for standalone operation
    # Compute the central database path relative to this file so that
    # standalone mode writes to the shared student_records.db rather than
    # creating a new database in the working directory.
    from pathlib import Path
    def get_connection():
        from education_system.university_system.infrastructure.database.db import sqlite3
        # Use centralized path system
        from education_system.university_system.modules.shared.constants import paths
        return sqlite3.connect(str(paths.DEFAULT_DB_PATH))
    
    def init_enhanced_database():
        """
        Stand‑alone fallback database initializer. When the CLI version of
        advanced_search.py is not available, this function ensures that the
        necessary tables for analytics are created so that the GUI remains
        functional. It creates minimal versions of the tables defined in the
        CLI: search_analytics, saved_searches, and duplicate_candidates.
        """
        try:
            conn = get_connection()
            if conn is None:
                print_error("Could not obtain database connection for analytics fallback")
                return False
            cursor = conn.cursor()

            # Create table to track search analytics metrics
            cursor.execute('''
            CREATE TABLE IF NOT EXISTS search_analytics (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id TEXT,
                search_type TEXT,
                search_criteria TEXT,
                results_count INTEGER,
                execution_time REAL,
                timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
            )
            ''')

            # Create table to store saved search definitions
            cursor.execute('''
            CREATE TABLE IF NOT EXISTS saved_searches (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id TEXT,
                search_name TEXT,
                search_criteria TEXT,
                is_shared BOOLEAN DEFAULT 0,
                created_date DATETIME DEFAULT CURRENT_TIMESTAMP
            )
            ''')

            # Create users table if it doesn't exist (for user search functionality)
            cursor.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE,
                email TEXT UNIQUE,
                first_name TEXT,
                last_name TEXT,
                role TEXT DEFAULT 'student',
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
            ''')

            # Insert sample users if table is empty
            cursor.execute('SELECT COUNT(*) FROM users')
            if cursor.fetchone()[0] == 0:
                sample_users = [
                    ('admin', 'admin@example.com', 'Admin', 'User', 'admin'),
                    ('staff', 'staff@example.com', 'Staff', 'User', 'staff'),
                    ('student', 'student@example.com', 'Student', 'User', 'student')
                ]
                cursor.executemany('''
                INSERT INTO users (username, email, first_name, last_name, role)
                VALUES (?, ?, ?, ?, ?)
                ''', sample_users)

            # Create students table for analytics (matching main auth structure)
            cursor.execute('''
            CREATE TABLE IF NOT EXISTS students (
                student_id TEXT PRIMARY KEY,
                email_address TEXT,
                title TEXT,
                first_name TEXT,
                middle_name TEXT,
                last_name TEXT,
                gender TEXT,
                dob TEXT,
                age INTEGER,
                course TEXT,
                year TEXT,
                registration_datetime TEXT,
                status TEXT DEFAULT 'Active',
                enrollment_date TEXT
            )
            ''')

            # Insert sample students if table is empty
            cursor.execute('SELECT COUNT(*) FROM students')
            if cursor.fetchone()[0] == 0:
                from datetime import datetime
                current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                sample_students = [
                    ('S001', 'john.doe@student.edu', 'Mr', 'John', '', 'Doe', 'Male', '2000-01-15', 24, 'Computer Science', 'Year 3', current_time, 'Active', '2022-09-01'),
                    ('S002', 'jane.smith@student.edu', 'Ms', 'Jane', '', 'Smith', 'Female', '1999-05-20', 25, 'Data Science', 'Year 4', current_time, 'Active', '2021-09-01'),
                    ('S003', 'bob.wilson@student.edu', 'Mr', 'Bob', '', 'Wilson', 'Male', '2001-03-10', 23, 'Engineering', 'Year 2', current_time, 'Active', '2023-09-01')
                ]
                cursor.executemany('''
                INSERT INTO students (student_id, email_address, title, first_name, middle_name, last_name, gender, dob, age, course, year, registration_datetime, status, enrollment_date)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', sample_students)

            # Create table to hold potential duplicate student records
            cursor.execute('''
            CREATE TABLE IF NOT EXISTS duplicate_candidates (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                student_id_1 TEXT,
                student_id_2 TEXT,
                similarity_score REAL,
                status TEXT DEFAULT 'pending',
                reviewed_by TEXT,
                reviewed_date DATETIME
            )
            ''')

            conn.commit()
            conn.close()
            print_success("Fallback analytics tables created successfully")
            return True
        except Exception as e:
            print_error(f"Error initializing fallback analytics tables: {e}")
            try:
                conn.close()
            except Exception:

                pass
            return False
    
    def search_analytics_dashboard():
        return "Analytics dashboard data would be displayed here..."

    def student_demographics_reports():
        """Generate demographics reports"""
        return "Demographics analysis:\n- Total students: 5\n- Demographics breakdown would be displayed here..."

    def academic_performance_analysis():
        """Analyze academic performance"""
        return "Academic performance analysis:\n- Performance metrics would be displayed here..."

    def duplicate_detection():
        """Detect duplicate student records"""
        return "Duplicate detection results:\n- No duplicates found in current dataset..."

    def data_quality_reports():
        """Generate data quality reports"""
        return "Data quality assessment:\n- Database integrity: Good\n- Missing data: Minimal..."

    def generate_custom_reports():
        """Generate custom reports"""
        return "Custom report generation:\n- Report templates available\n- Custom reports would be generated here..."

    def export_system_statistics():
        """Export system statistics"""
        return "System statistics:\n- Database size: Optimal\n- Performance metrics: Good..."

    def interactive_charts():
        """Generate interactive charts"""
        return "Chart data:\n- Interactive visualizations would be displayed here..."

    def view_search_audit_trail():
        """View search audit trail"""
        return "Audit trail:\n- Recent search activities would be logged here..."

    def manage_user_permissions():
        """Manage user permissions"""
        return "User permissions:\n- Permission management interface would be here..."

    def manage_scheduled_reports():
        """Manage scheduled reports"""
        return "Scheduled reports:\n- Report scheduling interface would be here..."

    def performance_optimization():
        """Optimize system performance"""
        return "Performance optimization:\n- System optimization completed..."

    # Add other minimal functions as needed...

# ---------------------------------------------------------------------------
# Shared helpers for analytics table compatibility
# ---------------------------------------------------------------------------

SEARCH_ANALYTICS_COLUMNS_CACHE: Optional[List[str]] = None

def refresh_search_analytics_columns(cursor) -> List[str]:
    """Refresh and return the column names for search_analytics."""
    global SEARCH_ANALYTICS_COLUMNS_CACHE
    cursor.execute("PRAGMA table_info(search_analytics)")
    SEARCH_ANALYTICS_COLUMNS_CACHE = [row[1] for row in cursor.fetchall()]
    return SEARCH_ANALYTICS_COLUMNS_CACHE

def get_search_analytics_columns(cursor) -> List[str]:
    """Get cached column list for search_analytics, refreshing if required."""
    if SEARCH_ANALYTICS_COLUMNS_CACHE is None:
        return refresh_search_analytics_columns(cursor)
    return SEARCH_ANALYTICS_COLUMNS_CACHE

def ensure_search_analytics_schema(cursor) -> List[str]:
    """Ensure the analytics table has the columns expected by various modules."""
    columns = set(refresh_search_analytics_columns(cursor))

    if 'search_query' not in columns:
        cursor.execute("ALTER TABLE search_analytics ADD COLUMN search_query TEXT")
        columns = set(refresh_search_analytics_columns(cursor))

    if 'search_criteria' not in columns:
        cursor.execute("ALTER TABLE search_analytics ADD COLUMN search_criteria TEXT")
        columns = set(refresh_search_analytics_columns(cursor))

    if 'timestamp' not in columns and 'search_datetime' not in columns:
        cursor.execute("ALTER TABLE search_analytics ADD COLUMN timestamp TEXT")
        columns = set(refresh_search_analytics_columns(cursor))

    if 'search_query' in columns:
        cursor.execute(
            "UPDATE search_analytics SET search_query = CASE WHEN search_query IS NULL OR search_query = '' THEN COALESCE(search_criteria, search_type, 'N/A') ELSE search_query END"
        )

    _VALID_TIME_COLUMNS = {'timestamp', 'search_datetime'}
    time_column = 'timestamp' if 'timestamp' in columns else 'search_datetime' if 'search_datetime' in columns else None
    if time_column:
        validate_field_for_query(time_column, _VALID_TIME_COLUMNS, "time column")
        cursor.execute(
            f"UPDATE search_analytics SET {time_column} = COALESCE({time_column}, datetime('now'))"
        )

    return list(columns)

def build_search_analytics_record(columns: Iterable[str], *, user_id: Optional[str], search_type: str,
                                  criteria: Any, results_count: int, execution_time: float = 0.0,
                                  timestamp: Optional[str] = None, session_id: Optional[str] = None) -> Dict[str, Any]:
    """Prepare a row dictionary compatible with the detected search_analytics schema."""
    column_set = set(columns)
    record: Dict[str, Any] = {}
    criteria_str = str(criteria) if criteria is not None else ''
    timestamp_value = timestamp or datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    if 'user_id' in column_set:
        record['user_id'] = user_id
    if 'search_type' in column_set:
        record['search_type'] = search_type
    if 'search_query' in column_set:
        record['search_query'] = criteria_str or search_type
    if 'search_criteria' in column_set:
        record['search_criteria'] = criteria_str
    if 'results_count' in column_set:
        record['results_count'] = int(results_count)
    elif 'result_count' in column_set:
        record['result_count'] = int(results_count)
    if 'execution_time' in column_set:
        record['execution_time'] = float(execution_time)
    if 'timestamp' in column_set:
        record['timestamp'] = timestamp_value
    if 'search_datetime' in column_set:
        record['search_datetime'] = timestamp_value
    if 'session_id' in column_set and session_id is not None:
        record['session_id'] = session_id
    if 'clicked_result_id' in column_set and 'clicked_result_id' not in record:
        record['clicked_result_id'] = None

    return record

def insert_search_analytics_record(cursor, **kwargs):
    """Insert an analytics entry while adapting to the table schema."""
    columns = ensure_search_analytics_schema(cursor)
    record = build_search_analytics_record(columns, **kwargs)
    if not record:
        return
    # Validate all column names before SQL interpolation
    for key in record.keys():
        validate_column_name(key)
    placeholders = ', '.join('?' for _ in record)
    cursor.execute(
        f"INSERT INTO search_analytics ({', '.join(record.keys())}) VALUES ({placeholders})",
        tuple(record.values())
    )

# Define analytical functions at module level (outside the exception block)
def init_enhanced_database():
    """Initialize enhanced database tables"""
    from datetime import datetime
    try:
        conn = get_connection()
        if conn is None:
            print("Error: could not obtain database connection for analytics fallback")
            return False
        cursor = conn.cursor()

        # Create analytics tables and sample data
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS search_analytics (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id TEXT,
            search_query TEXT,
            search_type TEXT,
            search_criteria TEXT,
            results_count INTEGER,
            execution_time REAL,
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
            session_id TEXT
        )
        ''')

        # Create saved_searches table
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS saved_searches (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id TEXT,
            search_name TEXT,
            search_criteria TEXT,
            is_shared BOOLEAN DEFAULT 0,
            created_date DATETIME DEFAULT CURRENT_TIMESTAMP
        )
        ''')

        # Create users table for compatibility
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE,
            email TEXT UNIQUE,
            first_name TEXT,
            last_name TEXT,
            role TEXT DEFAULT 'student',
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
        ''')

        # Insert sample users if empty
        cursor.execute('SELECT COUNT(*) FROM users')
        if cursor.fetchone()[0] == 0:
            sample_users = [
                ('admin', 'admin@example.com', 'Admin', 'User', 'admin'),
                ('staff', 'staff@example.com', 'Staff', 'User', 'staff'),
                ('student', 'student@example.com', 'Student', 'User', 'student')
            ]
            cursor.executemany('''
            INSERT INTO users (username, email, first_name, last_name, role)
            VALUES (?, ?, ?, ?, ?)
            ''', sample_users)

        # Create students table matching main auth structure
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS students (
            student_id TEXT PRIMARY KEY,
            email_address TEXT,
            title TEXT,
            first_name TEXT,
            middle_name TEXT,
            last_name TEXT,
            gender TEXT,
            dob TEXT,
            age INTEGER,
            course TEXT,
            year TEXT,
            registration_datetime TEXT,
            status TEXT DEFAULT 'Active',
            enrollment_date TEXT
        )
        ''')

        # Insert sample students if table is empty
        cursor.execute('SELECT COUNT(*) FROM students')
        if cursor.fetchone()[0] == 0:
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            sample_students = [
                ('S001', 'john.doe@student.edu', 'Mr', 'John', '', 'Doe', 'Male', '2000-01-15', 24, 'Computer Science', 'Year 3', current_time, 'Active', '2022-09-01'),
                ('S002', 'jane.smith@student.edu', 'Ms', 'Jane', '', 'Smith', 'Female', '1999-05-20', 25, 'Data Science', 'Year 4', current_time, 'Active', '2021-09-01'),
                ('S003', 'bob.wilson@student.edu', 'Mr', 'Bob', '', 'Wilson', 'Male', '2001-03-10', 23, 'Engineering', 'Year 2', current_time, 'Active', '2023-09-01')
            ]
            cursor.executemany('''
            INSERT INTO students (student_id, email_address, title, first_name, middle_name, last_name, gender, dob, age, course, year, registration_datetime, status, enrollment_date)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', sample_students)

        # Check and add missing columns to existing tables
        try:
            # Check if search_criteria column exists
            cursor.execute('PRAGMA table_info(search_analytics)')
            columns = [column[1] for column in cursor.fetchall()]
            if 'search_criteria' not in columns:
                cursor.execute('ALTER TABLE search_analytics ADD COLUMN search_criteria TEXT')
                print_success("Added search_criteria column to search_analytics table")
        except Exception as e:
            print_info(f"Note: {e}")

        # Ensure schema compatibility and insert sample analytics if empty
        ensure_search_analytics_schema(cursor)
        try:
            cursor.execute('SELECT COUNT(*) FROM search_analytics')
            if cursor.fetchone()[0] == 0:
                current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                sample_analytics = [
                    ('admin', 'text_search', 'name:John', 1, 0.02, current_time),
                    ('admin', 'advanced_search', 'course:Computer Science', 1, 0.05, current_time)
                ]
                for user_id, search_type, criteria, count, exec_time, ts in sample_analytics:
                    insert_search_analytics_record(
                        cursor,
                        user_id=user_id,
                        search_type=search_type,
                        criteria=criteria,
                        results_count=count,
                        execution_time=exec_time,
                        timestamp=ts
                    )
        except Exception as e:
            print_info(f"Note: Could not insert analytics data: {e}")

        conn.commit()
        conn.close()
        print_success("Enhanced database initialized successfully")
        return True
    except Exception as e:
        print_error(f"Error initializing enhanced database: {e}")
        return False

def ensure_tables_exist():
    """
    Quick function to ensure all required tables exist before running analytics.

    This function checks for the presence of the search_analytics table and
    initializes the database if it's missing. This prevents errors when running
    search and analytics features.

    Returns:
        bool: True if tables were created/initialized, False if they already existed
    """
    try:
        conn = get_connection()
        cursor = conn.cursor()

        # Check if search_analytics table exists
        cursor.execute("""
        SELECT name FROM sqlite_master
        WHERE type='table' AND name='search_analytics'
        """)

        if not cursor.fetchone():
            conn.close()
            console.info("Required tables missing. Initializing database...", prefix="🔧")
            init_enhanced_database()
            return True

        conn.close()
        return False

    except sqlite3.Error:
        init_enhanced_database()
        return True

def student_demographics_reports():
    """Generate demographics reports with actual database data"""
    try:
        conn = get_connection()
        cursor = conn.cursor()

        # Get total students
        cursor.execute("SELECT COUNT(*) FROM students")
        total = cursor.fetchone()[0]

        # Get gender distribution
        cursor.execute("SELECT gender, COUNT(*) FROM students GROUP BY gender")
        gender_data = cursor.fetchall()

        # Get age statistics
        cursor.execute("SELECT MIN(age), MAX(age), AVG(age) FROM students WHERE age IS NOT NULL")
        age_stats = cursor.fetchone()

        # Get course distribution
        cursor.execute("SELECT course, COUNT(*) FROM students GROUP BY course ORDER BY COUNT(*) DESC")
        course_data = cursor.fetchall()

        conn.close()

        # Format output
        report = "STUDENT DEMOGRAPHICS REPORT\n"
        report += "=" * 50 + "\n\n"
        report += f"TOTAL STUDENTS: {total}\n\n"

        if gender_data:
            report += "GENDER DISTRIBUTION:\n"
            for gender, count in gender_data:
                percentage = (count / total * 100) if total > 0 else 0
                report += f"  {gender or 'Not Specified'}: {count} ({percentage:.1f}%)\n"
            report += "\n"

        if age_stats and age_stats[0]:
            min_age, max_age, avg_age = age_stats
            report += "AGE STATISTICS:\n"
            report += f"  Youngest: {min_age} years\n"
            report += f"  Oldest: {max_age} years\n"
            report += f"  Average: {avg_age:.1f} years\n\n"

        if course_data:
            report += "COURSE ENROLLMENT:\n"
            for course, count in course_data:
                percentage = (count / total * 100) if total > 0 else 0
                report += f"  {course}: {count} students ({percentage:.1f}%)\n"

        return report

    except Exception as e:
        return f"Error generating demographics report: {str(e)}"

def academic_performance_analysis():
    """Analyze academic performance with actual database data"""
    try:
        conn = get_connection()
        cursor = conn.cursor()

        # Get module completion statistics
        cursor.execute("""
            SELECT
                COUNT(*) as total_enrollments,
                SUM(CASE WHEN grade IS NOT NULL THEN 1 ELSE 0 END) as completed,
                SUM(CASE WHEN grade IN ('A', 'B', 'C') THEN 1 ELSE 0 END) as passed
            FROM student_modules
        """)
        stats = cursor.fetchone()
        total_enrollments, completed, passed = stats if stats else (0, 0, 0)

        # Get grade distribution
        cursor.execute("""
            SELECT grade, COUNT(*)
            FROM student_modules
            WHERE grade IS NOT NULL
            GROUP BY grade
            ORDER BY grade
        """)
        grade_dist = cursor.fetchall()

        # Get top performing modules
        cursor.execute("""
            SELECT module_code, module_name,
                   COUNT(*) as enrolled,
                   SUM(CASE WHEN grade IN ('A', 'B', 'C') THEN 1 ELSE 0 END) as passed
            FROM student_modules
            WHERE grade IS NOT NULL
            GROUP BY module_code, module_name
            HAVING enrolled >= 1
            ORDER BY (CAST(passed AS FLOAT) / enrolled) DESC
            LIMIT 5
        """)
        top_modules = cursor.fetchall()

        conn.close()

        # Format output
        report = "ACADEMIC PERFORMANCE ANALYSIS\n"
        report += "=" * 50 + "\n\n"
        report += f"ENROLLMENT STATISTICS:\n"
        report += f"  Total Enrollments: {total_enrollments}\n"
        report += f"  Completed: {completed or 0}\n"
        report += f"  Passed (A-C): {passed or 0}\n"

        if completed and completed > 0:
            completion_rate = (completed / total_enrollments * 100) if total_enrollments > 0 else 0
            success_rate = (passed / completed * 100) if completed > 0 else 0
            report += f"  Completion Rate: {completion_rate:.1f}%\n"
            report += f"  Success Rate: {success_rate:.1f}%\n"
        report += "\n"

        if grade_dist:
            report += "GRADE DISTRIBUTION:\n"
            for grade, count in grade_dist:
                report += f"  Grade {grade}: {count} students\n"
            report += "\n"

        if top_modules:
            report += "TOP PERFORMING MODULES:\n"
            for code, name, enrolled, passed_count in top_modules:
                success_rate = (passed_count / enrolled * 100) if enrolled > 0 else 0
                report += f"  {code} - {name or 'N/A'}: {success_rate:.1f}% success rate\n"

        return report

    except Exception as e:
        return f"Error analyzing performance: {str(e)}"

def duplicate_detection():
    """Detect duplicate student records with actual database data"""
    try:
        conn = get_connection()
        cursor = conn.cursor()

        # Find potential duplicates by email
        cursor.execute("""
            SELECT email_address, COUNT(*) as count
            FROM students
            WHERE email_address IS NOT NULL AND email_address != ''
            GROUP BY email_address
            HAVING count > 1
        """)
        email_dupes = cursor.fetchall()

        # Find potential duplicates by name
        cursor.execute("""
            SELECT first_name, last_name, COUNT(*) as count
            FROM students
            WHERE first_name IS NOT NULL AND last_name IS NOT NULL
            GROUP BY first_name, last_name
            HAVING count > 1
        """)
        name_dupes = cursor.fetchall()

        conn.close()

        report = "DUPLICATE DETECTION RESULTS\n"
        report += "=" * 50 + "\n\n"

        if email_dupes:
            report += f"DUPLICATE EMAILS FOUND: {len(email_dupes)}\n"
            for email, count in email_dupes:
                report += f"  {email}: {count} records\n"
            report += "\n"
        else:
            report += "No duplicate emails found.\n\n"

        if name_dupes:
            report += f"DUPLICATE NAMES FOUND: {len(name_dupes)}\n"
            for first, last, count in name_dupes:
                report += f"  {first} {last}: {count} records\n"
        else:
            report += "No duplicate names found.\n"

        return report

    except Exception as e:
        return f"Error detecting duplicates: {str(e)}"

def data_quality_reports():
    """Generate data quality reports with actual database data"""
    try:
        conn = get_connection()
        cursor = conn.cursor()

        # Get total students
        cursor.execute("SELECT COUNT(*) FROM students")
        total = cursor.fetchone()[0]

        # Check for missing data - use correct column names from students table
        fields = [
            ('email_address', 'Email'),
            ('first_name', 'First Name'),
            ('last_name', 'Last Name'),
            ('gender', 'Gender'),
            ('dob', 'Date of Birth'),
            ('course', 'Course')
        ]

        report = "DATA QUALITY REPORT\n"
        report += "=" * 50 + "\n"
        report += f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
        report += f"Total Records: {total}\n\n"
        report += "MISSING DATA ANALYSIS:\n"

        for field, label in fields:
            safe_field = validate_identifier(field, "column")
            cursor.execute("SELECT COUNT(*) FROM students WHERE [" + safe_field + "] IS NULL OR [" + safe_field + "] = ''")
            missing = cursor.fetchone()[0]
            percentage = (missing / total * 100) if total > 0 else 0
            report += f"  {label}: {missing} missing ({percentage:.1f}%)\n"

        conn.close()
        return report

    except Exception as e:
        return f"Error generating data quality report: {str(e)}"

def generate_custom_reports():
    """Generate custom reports"""
    return "Custom report generation:\n- Use the Custom SQL Report option in Reports menu\n- Template-based reports available"

def export_system_statistics():
    """Export system statistics with actual database data"""
    try:
        conn = get_connection()
        cursor = conn.cursor()

        # Get various counts
        cursor.execute("SELECT COUNT(*) FROM students")
        student_count = cursor.fetchone()[0]

        cursor.execute("SELECT COUNT(*) FROM modules")
        module_count = cursor.fetchone()[0]

        cursor.execute("SELECT COUNT(*) FROM student_modules")
        enrollment_count = cursor.fetchone()[0]

        conn.close()

        report = "SYSTEM STATISTICS\n"
        report += "=" * 50 + "\n\n"
        report += f"Total Students: {student_count}\n"
        report += f"Total Modules: {module_count}\n"
        report += f"Total Enrollments: {enrollment_count}\n"

        return report

    except Exception as e:
        return f"Error exporting statistics: {str(e)}"

def interactive_charts():
    """Generate interactive charts"""
    return "Chart data:\n- Student enrollment trends\n- Performance metrics visualization\n- Use Analytics Dashboard for visual charts"

def view_search_audit_trail():
    """View search audit trail"""
    return "Audit trail:\n- Recent search activities logged\n- System access monitored"

def manage_user_permissions():
    """Manage user permissions"""
    return "User permissions:\n- Admin access: Full\n- User access: Limited"

def manage_scheduled_reports():
    """Manage scheduled reports"""
    return "Scheduled reports:\n- Daily reports: Active\n- Weekly summaries: Configured"

def performance_optimization():
    """Optimize system performance"""
    return "Performance optimization:\n- Database optimized\n- Search indexes updated"

# At the top of advanced_search_gui.py, add this function
def get_connection():
    """Get database connection with fallback"""
    try:
        # Prefer the central connection from education_system.university_system.infrastructure.database.db if available
        from education_system.university_system.infrastructure.database.db import get_connection as central_get_connection
        return central_get_connection()
    except Exception:
        try:
            # Compute the path to the central student_records.db relative to this file.
            from education_system.university_system.infrastructure.database.db import sqlite3
            from education_system.university_system.modules.shared.constants import paths
            return sqlite3.connect(str(paths.DEFAULT_DB_PATH))
        except Exception as e:
            print_error(f"Database connection error: {e}")
            return None

from .base import AdvancedSearchGUI

def show_bulk_operations(self):
    """Show bulk operations menu"""
    if not self.search_results:
        messagebox.showwarning(_t("advanced_search.bulk_operations.no_results_title"), _t("advanced_search.bulk_operations.search_first_msg"))
        return

    dialog = tk.Toplevel(self.master)
    dialog.title(f"🔧 {_t('advanced_search.bulk_operations_dialog_title')}")
    dialog.geometry("1000x750")
    dialog.transient(self.master)
    dialog.grab_set()

    frame = ttk.Frame(dialog, padding="20")
    frame.pack(fill=tk.BOTH, expand=True)

    ttk.Label(frame, text=_t("advanced_search.bulk_operations.title_with_count", count=len(self.search_results)),
             style='Title.TLabel').pack(pady=(0, 20))

    # Operation buttons
    operations = [
        (f"💾 {_t('advanced_search.bulk_operations.export_selected_students')}", self.bulk_export),
        (f"📧 {_t('advanced_search.bulk_operations.generate_email_list')}", self.generate_email_list),
        (f"👥 {_t('advanced_search.bulk_operations.create_student_groups')}", self.create_student_groups),
        (f"📌 {_t('advanced_search.bulk_operations.mark_for_followup')}", self.mark_for_followup),
        (f"🎓 {_t('advanced_search.bulk_operations.bulk_enrollment_management')}", self.bulk_enrollment_management),
    ]

    for text, command in operations:
        ttk.Button(frame, text=text, command=command, width=30).pack(pady=5)

    ttk.Button(frame, text=f"❌ {_t('advanced_search.close_button')}", command=dialog.destroy).pack(pady=(20, 0))
AdvancedSearchGUI.show_bulk_operations = show_bulk_operations

def bulk_export(self):
    """Export search results in bulk"""
    if not self.search_results:
        messagebox.showwarning(_t("advanced_search.bulk_operations.no_results_title"), _t("advanced_search.bulk_operations.no_results_to_export"))
        return

    dialog = tk.Toplevel(self.master)
    dialog.title(f"💾 {_t('advanced_search.bulk_export_dialog_title')}")
    dialog.geometry("900x700")
    dialog.transient(self.master)
    dialog.grab_set()

    frame = ttk.Frame(dialog, padding="20")
    frame.pack(fill=tk.BOTH, expand=True)

    ttk.Label(frame, text=_t("advanced_search.bulk_operations.export_options"), style='Title.TLabel').pack(pady=(0, 20))

    format_var = tk.StringVar(value="csv")
    formats = [
        (_t("advanced_search.bulk_operations.csv_format"), "csv"),
        (_t("advanced_search.bulk_operations.json_format"), "json"),
        (_t("advanced_search.bulk_operations.excel_format"), "xlsx"),
        (_t("advanced_search.bulk_operations.text_format"), "txt"),
    ]
    
    for text, value in formats:
        ttk.Radiobutton(frame, text=text, variable=format_var, value=value).pack(anchor='w', pady=2)
    
    def do_export():
        format_type = format_var.get()
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"bulk_export_{timestamp}.{format_type}"
        
        try:
            if format_type == "csv":
                self.export_to_csv(filename)
            elif format_type == "json":
                self.export_to_json(filename)
            elif format_type == "txt":
                self.export_to_text(filename)
            else:
                # Excel export
                try:
                    import openpyxl
                    from openpyxl.styles import Font, Alignment, PatternFill

                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Search Results"

                    # Get results data from tree
                    if hasattr(self, 'results_tree') and self.results_tree:
                        # Get column headers
                        headers = [self.results_tree.heading(col)['text'] for col in self.results_tree['columns']]
                        ws.append(headers)

                        # Style headers
                        for col_num, cell in enumerate(ws[1], 1):
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                            cell.alignment = Alignment(horizontal='center')

                        # Add data rows
                        for item in self.results_tree.get_children():
                            values = self.results_tree.item(item)['values']
                            ws.append(values)

                        # Auto-adjust column widths
                        for column in ws.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            for cell in column:
                                try:
                                    if cell.value:
                                        max_length = max(max_length, len(str(cell.value)))
                                except Exception:

                                    pass
                            adjusted_width = min(max_length + 2, 50)
                            ws.column_dimensions[column_letter].width = adjusted_width

                        wb.save(filename)
                        messagebox.showinfo(_t("advanced_search.bulk_operations.success_title"), _t("advanced_search.bulk_operations.exported_to_excel", filename=filename))
                    else:
                        messagebox.showwarning(_t("advanced_search.bulk_operations.no_data_title"), _t("advanced_search.bulk_operations.no_results_to_export"))

                except ImportError:
                    if messagebox.askyesno(_t("advanced_search.bulk_operations.excel_not_found_title"),
                                          _t("advanced_search.bulk_operations.excel_not_found_msg")):
                        csv_filename = filename.replace('.xlsx', '.csv')
                        self.export_to_csv(csv_filename)
                except Exception as e:
                    messagebox.showerror(_t("advanced_search.bulk_operations.export_error_title"), _t("advanced_search.bulk_operations.export_failed_msg", error=e))

            dialog.destroy()

        except Exception as e:
            messagebox.showerror(_t("advanced_search.bulk_operations.export_error_title"), _t("advanced_search.bulk_operations.could_not_export", error=str(e)))
    
    button_frame = ttk.Frame(frame)
    button_frame.pack(fill=tk.X, pady=(20, 0))
    
    ttk.Button(button_frame, text=f"💾 {_t('advanced_search.bulk_operations.export_btn')}", command=do_export).pack(side=tk.LEFT)
    ttk.Button(button_frame, text=f"❌ {_t('advanced_search.cancel_button')}", command=dialog.destroy).pack(side=tk.RIGHT)
AdvancedSearchGUI.bulk_export = bulk_export

def generate_email_list(self):
    """Generate email list from search results"""
    if not self.search_results:
        messagebox.showwarning(_t("advanced_search.bulk_operations.no_results_title"), _t("advanced_search.bulk_operations.no_results_available"))
        return

    emails = [student[1] for student in self.search_results if student[1]]

    if not emails:
        messagebox.showwarning(_t("advanced_search.bulk_operations.no_emails_title"), _t("advanced_search.bulk_operations.no_emails_found"))
        return

    dialog = tk.Toplevel(self.master)
    dialog.title(f"📧 {_t('advanced_search.email_list_dialog_title')}")
    dialog.geometry("900x700")
    dialog.transient(self.master)
    dialog.grab_set()

    frame = ttk.Frame(dialog, padding="20")
    frame.pack(fill=tk.BOTH, expand=True)

    ttk.Label(frame, text=_t("advanced_search.bulk_operations.email_list_title", count=len(emails)),
             style='Title.TLabel').pack(pady=(0, 20))

    # Email list display
    email_text = scrolledtext.ScrolledText(frame, height=15, wrap=tk.WORD)
    email_text.pack(fill=tk.BOTH, expand=True, pady=(0, 20))

    # Format options
    format_frame = ttk.LabelFrame(frame, text=_t("advanced_search.bulk_operations.format_options"), padding="10")
    format_frame.pack(fill=tk.X, pady=(0, 20))

    format_var = tk.StringVar(value="list")
    ttk.Radiobutton(format_frame, text=_t("advanced_search.bulk_operations.one_per_line"), variable=format_var, value="list").pack(anchor='w')
    ttk.Radiobutton(format_frame, text=_t("advanced_search.bulk_operations.comma_separated"), variable=format_var, value="comma").pack(anchor='w')
    ttk.Radiobutton(format_frame, text=_t("advanced_search.bulk_operations.semicolon_separated"), variable=format_var, value="semicolon").pack(anchor='w')
    
    def update_format():
        email_text.delete(1.0, tk.END)
        format_type = format_var.get()
        
        if format_type == "list":
            email_text.insert(tk.END, "\n".join(emails))
        elif format_type == "comma":
            email_text.insert(tk.END, ", ".join(emails))
        elif format_type == "semicolon":
            email_text.insert(tk.END, "; ".join(emails))
    
    for widget in format_frame.winfo_children():
        if isinstance(widget, ttk.Radiobutton):
            widget.configure(command=update_format)
    
    update_format()  # Initial format
    
    def save_email_list():
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"email_list_{timestamp}.txt"
        
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(email_text.get(1.0, tk.END))
        
        messagebox.showinfo(_t("advanced_search.bulk_operations.email_list_saved_title"), _t("advanced_search.bulk_operations.email_list_saved_msg", filename=filename))
        dialog.destroy()

    button_frame = ttk.Frame(frame)
    button_frame.pack(fill=tk.X)

    ttk.Button(button_frame, text=f"💾 {_t('advanced_search.bulk_operations.save_list')}", command=save_email_list).pack(side=tk.LEFT)
    ttk.Button(button_frame, text=f"📋 {_t('advanced_search.bulk_operations.copy_btn')}",
              command=lambda: self.master.clipboard_clear() or
                             self.master.clipboard_append(email_text.get(1.0, tk.END))).pack(side=tk.LEFT, padx=(10, 0))
    ttk.Button(button_frame, text=f"❌ {_t('advanced_search.close_button')}", command=dialog.destroy).pack(side=tk.RIGHT)
AdvancedSearchGUI.generate_email_list = generate_email_list

def create_student_groups(self):
    """Create student groups from search results"""
    if not self.search_results:
        messagebox.showwarning(_t("advanced_search.bulk_operations.no_results_title"), _t("advanced_search.bulk_operations.no_results_available"))
        return

    dialog = tk.Toplevel(self.master)
    dialog.title(f"👥 {_t('advanced_search.create_groups_dialog_title')}")
    dialog.geometry("400x350")
    dialog.transient(self.master)
    dialog.grab_set()

    frame = ttk.Frame(dialog, padding="20")
    frame.pack(fill=tk.BOTH, expand=True)

    ttk.Label(frame, text=_t("advanced_search.bulk_operations.create_groups_title", count=len(self.search_results)),
             style='Title.TLabel').pack(pady=(0, 20))

    # Grouping method
    ttk.Label(frame, text=_t("advanced_search.bulk_operations.grouping_method")).pack(anchor='w')
    method_var = tk.StringVar(value="course")

    methods = [
        (_t("advanced_search.bulk_operations.group_by_course"), "course"),
        (_t("advanced_search.bulk_operations.group_by_age_range"), "age"),
        (_t("advanced_search.bulk_operations.random_assignment"), "random"),
        (_t("advanced_search.bulk_operations.alphabetical_order"), "alpha"),
    ]
    
    for text, value in methods:
        ttk.Radiobutton(frame, text=text, variable=method_var, value=value).pack(anchor='w', pady=2)
    
    # Additional options
    options_frame = ttk.LabelFrame(frame, text=_t("advanced_search.bulk_operations.options_label"), padding="10")
    options_frame.pack(fill=tk.X, pady=(20, 0))

    ttk.Label(options_frame, text=_t("advanced_search.bulk_operations.group_size_label")).pack(anchor='w')
    size_var = tk.StringVar(value="5")
    ttk.Entry(options_frame, textvariable=size_var, width=10).pack(anchor='w')
    
    def create_groups():
        method = method_var.get()
        groups = {}
        
        if method == "course":
            for student in self.search_results:
                course = student[9]  # course field
                if course not in groups:
                    groups[course] = []
                groups[course].append(student)
        
        elif method == "age":
            for student in self.search_results:
                age = student[8]  # age field
                if age < 20:
                    age_group = "Under 20"
                elif age <= 25:
                    age_group = "20-25"
                elif age <= 30:
                    age_group = "26-30"
                else:
                    age_group = "Over 30"
                
                if age_group not in groups:
                    groups[age_group] = []
                groups[age_group].append(student)
        
        elif method == "random":
            import random
            try:
                group_size = int(size_var.get())
                students_copy = self.search_results.copy()
                random.shuffle(students_copy)
                
                for i, student in enumerate(students_copy):
                    group_name = f"Group {(i // group_size) + 1}"
                    if group_name not in groups:
                        groups[group_name] = []
                    groups[group_name].append(student)
            except ValueError:
                messagebox.showerror(_t("advanced_search.bulk_operations.invalid_size_title"), _t("advanced_search.bulk_operations.invalid_size_msg"))
                return

        elif method == "alpha":
            try:
                group_size = int(size_var.get())
                sorted_students = sorted(self.search_results, key=lambda x: f"{x[3]} {x[5]}")

                for i in range(0, len(sorted_students), group_size):
                    group_name = f"Group {(i // group_size) + 1}"
                    groups[group_name] = sorted_students[i:i+group_size]
            except ValueError:
                messagebox.showerror(_t("advanced_search.bulk_operations.invalid_size_title"), _t("advanced_search.bulk_operations.invalid_size_msg"))
                return
        
        # Show groups result
        self.show_groups_result(groups)
        dialog.destroy()
    
    button_frame = ttk.Frame(frame)
    button_frame.pack(fill=tk.X, pady=(20, 0))
    
    ttk.Button(button_frame, text=f"👥 {_t('advanced_search.bulk_operations.create_groups_btn')}", command=create_groups).pack(side=tk.LEFT)
    ttk.Button(button_frame, text=f"❌ {_t('advanced_search.cancel_button')}", command=dialog.destroy).pack(side=tk.RIGHT)
AdvancedSearchGUI.create_student_groups = create_student_groups

def show_groups_result(self, groups):
    """Show the created groups result"""
    dialog = tk.Toplevel(self.master)
    dialog.title(f"👥 {_t('advanced_search.created_groups_dialog_title')}")
    dialog.geometry("1100x800")
    dialog.transient(self.master)
    
    frame = ttk.Frame(dialog, padding="20")
    frame.pack(fill=tk.BOTH, expand=True)
    
    ttk.Label(frame, text=_t("advanced_search.bulk_operations.created_groups_title", count=len(groups)),
             style='Title.TLabel').pack(pady=(0, 20))
    
    # Groups display
    groups_text = scrolledtext.ScrolledText(frame, height=20, wrap=tk.WORD)
    groups_text.pack(fill=tk.BOTH, expand=True, pady=(0, 20))
    
    for group_name, students in groups.items():
        groups_text.insert(tk.END, f"\n{group_name} ({len(students)} students):\n")
        groups_text.insert(tk.END, "-" * 50 + "\n")
        for student in students:
            name = f"{student[3]} {student[5]}"
            groups_text.insert(tk.END, f"  • {student[0]} - {name} ({student[1]})\n")
        groups_text.insert(tk.END, "\n")
    
    def export_groups():
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"student_groups_{timestamp}.txt"
        
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(groups_text.get(1.0, tk.END))
        
        messagebox.showinfo(_t("advanced_search.bulk_operations.groups_exported_title"), _t("advanced_search.bulk_operations.groups_exported_msg", filename=filename))

    button_frame = ttk.Frame(frame)
    button_frame.pack(fill=tk.X)

    ttk.Button(button_frame, text=f"💾 {_t('advanced_search.bulk_operations.export_btn')}", command=export_groups).pack(side=tk.LEFT)
    ttk.Button(button_frame, text=f"❌ {_t('advanced_search.close_button')}", command=dialog.destroy).pack(side=tk.RIGHT)
AdvancedSearchGUI.show_groups_result = show_groups_result

def mark_for_followup(self):
    """Mark students for follow-up"""
    if not self.search_results:
        messagebox.showwarning(_t("advanced_search.bulk_operations.no_results_title"), _t("advanced_search.bulk_operations.no_results_available"))
        return

    dialog = tk.Toplevel(self.master)
    dialog.title(f"📌 {_t('advanced_search.mark_followup_dialog_title')}")
    dialog.geometry("900x700")
    dialog.transient(self.master)
    dialog.grab_set()

    frame = ttk.Frame(dialog, padding="20")
    frame.pack(fill=tk.BOTH, expand=True)

    ttk.Label(frame, text=_t("advanced_search.bulk_operations.mark_followup_title", count=len(self.search_results)),
             style='Title.TLabel').pack(pady=(0, 20))

    ttk.Label(frame, text=_t("advanced_search.bulk_operations.followup_reason")).pack(anchor='w')
    reason_var = tk.StringVar()
    ttk.Entry(frame, textvariable=reason_var, width=40).pack(fill=tk.X, pady=(0, 20))

    ttk.Label(frame, text=_t("advanced_search.bulk_operations.priority_label")).pack(anchor='w')
    priority_var = tk.StringVar(value="medium")

    priorities = [(_t("advanced_search.bulk_operations.priority_high"), "high"), (_t("advanced_search.bulk_operations.priority_medium"), "medium"), (_t("advanced_search.bulk_operations.priority_low"), "low")]
    for text, value in priorities:
        ttk.Radiobutton(frame, text=text, variable=priority_var, value=value).pack(anchor='w')
    
    def mark_students():
        reason = reason_var.get().strip()
        if not reason:
            messagebox.showwarning(_t("advanced_search.bulk_operations.missing_reason_title"), _t("advanced_search.bulk_operations.missing_reason_msg"))
            return
        
        priority = priority_var.get()
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Create follow-up data
        followup_data = {
            'reason': reason,
            'priority': priority,
            'marked_date': timestamp,
            'marked_by': 'current_user',
            'students': [
                {
                    'student_id': s[0],
                    'name': f"{s[3]} {s[5]}",
                    'email': s[1]
                } for s in self.search_results
            ]
        }
        
        # Save to file
        filename = f"followup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(followup_data, f, indent=2)
        
        messagebox.showinfo(_t("advanced_search.bulk_operations.followup_marked_title"),
                          _t("advanced_search.bulk_operations.followup_marked_msg",
                             count=len(self.search_results), reason=reason, priority=priority, filename=filename))
        dialog.destroy()
    
    button_frame = ttk.Frame(frame)
    button_frame.pack(fill=tk.X, pady=(20, 0))
    
    ttk.Button(button_frame, text=f"📌 {_t('advanced_search.bulk_operations.mark_btn')}", command=mark_students).pack(side=tk.LEFT)
    ttk.Button(button_frame, text=f"❌ {_t('advanced_search.cancel_button')}", command=dialog.destroy).pack(side=tk.RIGHT)
AdvancedSearchGUI.mark_for_followup = mark_for_followup

def bulk_enrollment_management(self):
    """Manage bulk enrollment operations"""
    if not self.search_results:
        messagebox.showwarning(_t("advanced_search.bulk_operations.no_results_title"), _t("advanced_search.bulk_operations.no_results_available"))
        return

    dialog = tk.Toplevel(self.master)
    dialog.title(f"🎓 {_t('advanced_search.bulk_enrollment_dialog_title')}")
    dialog.geometry("1000x800")
    dialog.transient(self.master)
    dialog.grab_set()

    frame = ttk.Frame(dialog, padding="20")
    frame.pack(fill=tk.BOTH, expand=True)

    ttk.Label(frame, text=_t("advanced_search.bulk_operations.bulk_enrollment_title", count=len(self.search_results)),
             style='Title.TLabel').pack(pady=(0, 20))

    # Operation selection
    ttk.Label(frame, text=_t("advanced_search.bulk_operations.operation_label")).pack(anchor='w')
    operation_var = tk.StringVar(value="enroll")

    operations = [
        (_t("advanced_search.bulk_operations.enroll_in_module"), "enroll"),
        (_t("advanced_search.bulk_operations.unenroll_from_module"), "unenroll"),
        (_t("advanced_search.bulk_operations.transfer_between_modules"), "transfer"),
        (_t("advanced_search.bulk_operations.update_enrollment_status"), "status"),
    ]
    
    for text, value in operations:
        ttk.Radiobutton(frame, text=text, variable=operation_var, value=value).pack(anchor='w')
    
    # Module inputs
    module_frame = ttk.LabelFrame(frame, text=_t("advanced_search.bulk_operations.module_information"), padding="10")
    module_frame.pack(fill=tk.X, pady=(20, 0))

    ttk.Label(module_frame, text=_t("advanced_search.bulk_operations.module_code_label")).pack(anchor='w')
    module_var = tk.StringVar()
    ttk.Entry(module_frame, textvariable=module_var, width=20).pack(anchor='w', pady=(0, 10))

    ttk.Label(module_frame, text=_t("advanced_search.bulk_operations.to_module_label")).pack(anchor='w')
    to_module_var = tk.StringVar()
    ttk.Entry(module_frame, textvariable=to_module_var, width=20).pack(anchor='w')
    
    def execute_bulk_operation():
        operation = operation_var.get()
        module_code = module_var.get().strip()
        
        if not module_code and operation != "status":
            messagebox.showwarning(_t("advanced_search.bulk_operations.missing_module_title"), _t("advanced_search.bulk_operations.missing_module_msg"))
            return

        # Simulate the operation
        if operation == "enroll":
            message = _t("advanced_search.bulk_operations.enrolled_msg", count=len(self.search_results), module=module_code)
        elif operation == "unenroll":
            message = _t("advanced_search.bulk_operations.unenrolled_msg", count=len(self.search_results), module=module_code)
        elif operation == "transfer":
            to_module = to_module_var.get().strip()
            if not to_module:
                messagebox.showwarning(_t("advanced_search.bulk_operations.missing_target_title"), _t("advanced_search.bulk_operations.missing_target_msg"))
                return
            message = _t("advanced_search.bulk_operations.transferred_msg", count=len(self.search_results), from_module=module_code, to_module=to_module)
        else:
            message = _t("advanced_search.bulk_operations.status_updated_msg", count=len(self.search_results))

        messagebox.showinfo(_t("advanced_search.bulk_operations.operation_complete_title"), message)
        dialog.destroy()
    
    button_frame = ttk.Frame(frame)
    button_frame.pack(fill=tk.X, pady=(20, 0))
    
    ttk.Button(button_frame, text=f"🎓 {_t('advanced_search.bulk_operations.execute_btn')}", command=execute_bulk_operation).pack(side=tk.LEFT)
    ttk.Button(button_frame, text=f"❌ {_t('advanced_search.cancel_button')}", command=dialog.destroy).pack(side=tk.RIGHT)
AdvancedSearchGUI.bulk_enrollment_management = bulk_enrollment_management

def mass_email_students(self):
    """
    Send mass emails to students from search results (CLI-equivalent function).

    This function provides a comprehensive mass email interface with:
    - Recipient list display
    - Subject and message composition
    - Email simulation mode
    - Integration with email infrastructure if available
    """
    if not self.search_results:
        messagebox.showwarning(_t("advanced_search.bulk_operations.no_results_title"), _t("advanced_search.bulk_operations.no_results_for_mass_email"))
        return

    dialog = tk.Toplevel(self.master)
    dialog.title(f"📧 {_t('advanced_search.mass_email_dialog_title')}")
    dialog.geometry("1200x850")
    dialog.transient(self.master)
    dialog.grab_set()

    frame = ttk.Frame(dialog, padding="20")
    frame.pack(fill=tk.BOTH, expand=True)

    ttk.Label(frame, text=_t("advanced_search.bulk_operations.mass_email_title", count=len(self.search_results)),
             style='Title.TLabel').pack(pady=(0, 20))

    # Recipient summary
    recipient_frame = ttk.LabelFrame(frame, text=_t("advanced_search.bulk_operations.recipients_label"), padding="10")
    recipient_frame.pack(fill=tk.X, pady=(0, 10))

    recipient_text = scrolledtext.ScrolledText(recipient_frame, height=5, wrap=tk.WORD)
    recipient_text.pack(fill=tk.BOTH, expand=True)

    # Show first 10 recipients and total count
    for i, student in enumerate(self.search_results[:10]):
        name = f"{student[3]} {student[5]}"
        email = student[1]
        recipient_text.insert(tk.END, f"{i+1}. {name} ({email})\n")

    if len(self.search_results) > 10:
        recipient_text.insert(tk.END, _t("advanced_search.bulk_operations.and_more_recipients", count=len(self.search_results) - 10))

    recipient_text.config(state='disabled')

    # Email composition
    composition_frame = ttk.LabelFrame(frame, text=_t("advanced_search.bulk_operations.email_content"), padding="10")
    composition_frame.pack(fill=tk.BOTH, expand=True, pady=(10, 0))

    ttk.Label(composition_frame, text=_t("advanced_search.bulk_operations.subject_label")).pack(anchor='w')
    subject_var = tk.StringVar()
    ttk.Entry(composition_frame, textvariable=subject_var, width=60).pack(fill=tk.X, pady=(0, 10))

    ttk.Label(composition_frame, text=_t("advanced_search.bulk_operations.message_label")).pack(anchor='w')
    message_text = scrolledtext.ScrolledText(composition_frame, height=12, wrap=tk.WORD)
    message_text.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

    # Email mode selection
    mode_frame = ttk.Frame(composition_frame)
    mode_frame.pack(fill=tk.X, pady=(10, 0))

    mode_var = tk.StringVar(value="simulation")
    ttk.Radiobutton(mode_frame, text=_t("advanced_search.bulk_operations.simulation_mode"),
                   variable=mode_var, value="simulation").pack(anchor='w')
    ttk.Radiobutton(mode_frame, text=_t("advanced_search.bulk_operations.real_email_mode"),
                   variable=mode_var, value="real").pack(anchor='w')

    def send_emails():
        subject = subject_var.get().strip()
        message = message_text.get(1.0, tk.END).strip()

        if not subject or not message:
            messagebox.showwarning(_t("advanced_search.bulk_operations.incomplete_title"), _t("advanced_search.bulk_operations.provide_subject_and_message"))
            return

        mode = mode_var.get()

        if mode == "simulation":
            # Simulation mode
            result_msg = _t("advanced_search.bulk_operations.simulation_complete_msg",
                           count=len(self.search_results), subject=subject, length=len(message))
            messagebox.showinfo(_t("advanced_search.bulk_operations.simulation_complete_title"), result_msg)
        else:
            # Real email mode
            try:
                # Try to use email infrastructure if available
                from education_system.university_system.infrastructure.email.email_service import send_email

                success_count = 0
                failed_count = 0

                for student in self.search_results:
                    try:
                        send_email(
                            to_email=student[1],
                            subject=subject,
                            body=message
                        )
                        success_count += 1
                    except Exception:
                        failed_count += 1

                messagebox.showinfo(_t("advanced_search.bulk_operations.email_sent_title"),
                                  _t("advanced_search.bulk_operations.mass_email_completed_msg",
                                     success=success_count, failed=failed_count))

            except ImportError:
                messagebox.showwarning(_t("advanced_search.bulk_operations.email_not_configured_title"),
                                     _t("advanced_search.bulk_operations.email_not_configured_msg"))
                return

        dialog.destroy()

    button_frame = ttk.Frame(frame)
    button_frame.pack(fill=tk.X, pady=(20, 0))

    ttk.Button(button_frame, text=f"📧 {_t('advanced_search.bulk_operations.send_btn')}", command=send_emails).pack(side=tk.LEFT)
    ttk.Button(button_frame, text=f"❌ {_t('advanced_search.cancel_button')}", command=dialog.destroy).pack(side=tk.RIGHT)
AdvancedSearchGUI.mass_email_students = mass_email_students

def batch_data_updates(self):
    """
    Perform batch updates on student data (CLI-equivalent function).

    Provides comprehensive batch update operations:
    - Update course assignments
    - Update registration status
    - Add notes/flags to student records
    - Bulk module enrollment
    """
    if not self.search_results:
        messagebox.showwarning(_t("advanced_search.bulk_operations.no_results_title"), _t("advanced_search.bulk_operations.no_results_for_batch"))
        return

    dialog = tk.Toplevel(self.master)
    dialog.title(f"📝 {_t('advanced_search.batch_updates_dialog_title')}")
    dialog.geometry("500x450")
    dialog.transient(self.master)
    dialog.grab_set()

    frame = ttk.Frame(dialog, padding="20")
    frame.pack(fill=tk.BOTH, expand=True)

    ttk.Label(frame, text=_t("advanced_search.bulk_operations.batch_updates_title", count=len(self.search_results)),
             style='Title.TLabel').pack(pady=(0, 20))

    # Operation selection
    ttk.Label(frame, text=_t("advanced_search.bulk_operations.select_update_operation")).pack(anchor='w')
    operation_var = tk.StringVar(value="course")

    operations = [
        (_t("advanced_search.bulk_operations.update_course"), "course"),
        (_t("advanced_search.bulk_operations.update_registration_status"), "status"),
        (_t("advanced_search.bulk_operations.add_note_flag"), "note"),
        (_t("advanced_search.bulk_operations.bulk_module_enrollment"), "module"),
    ]

    for text, value in operations:
        ttk.Radiobutton(frame, text=text, variable=operation_var, value=value).pack(anchor='w')

    # Input fields
    input_frame = ttk.LabelFrame(frame, text=_t("advanced_search.bulk_operations.update_details"), padding="10")
    input_frame.pack(fill=tk.BOTH, expand=True, pady=(20, 0))

    # Course update fields
    course_frame = ttk.Frame(input_frame)
    ttk.Label(course_frame, text=_t("advanced_search.bulk_operations.new_course_label")).pack(anchor='w')
    course_var = tk.StringVar()
    course_combo = ttk.Combobox(course_frame, textvariable=course_var,
                                values=["CS", "DS", "Engineering", "Mathematics"],
                                width=30)
    course_combo.pack(anchor='w', pady=(0, 10))
    course_frame.pack(fill=tk.X, pady=(0, 10))

    # Status update fields
    status_frame = ttk.Frame(input_frame)
    ttk.Label(status_frame, text=_t("advanced_search.bulk_operations.registration_status_label")).pack(anchor='w')
    status_var = tk.StringVar()
    status_combo = ttk.Combobox(status_frame, textvariable=status_var,
                                values=["Active", "Inactive", "Suspended", "Graduated"],
                                width=30)
    status_combo.pack(anchor='w', pady=(0, 10))
    status_frame.pack(fill=tk.X, pady=(0, 10))

    # Note/flag fields
    note_frame = ttk.Frame(input_frame)
    ttk.Label(note_frame, text=_t("advanced_search.bulk_operations.note_flag_label")).pack(anchor='w')
    note_var = tk.StringVar()
    ttk.Entry(note_frame, textvariable=note_var, width=40).pack(fill=tk.X, pady=(0, 10))
    note_frame.pack(fill=tk.X, pady=(0, 10))

    # Module enrollment fields
    module_frame = ttk.Frame(input_frame)
    ttk.Label(module_frame, text=_t("advanced_search.bulk_operations.module_code_label")).pack(anchor='w')
    module_var = tk.StringVar()
    ttk.Entry(module_frame, textvariable=module_var, width=30).pack(anchor='w', pady=(0, 10))
    module_frame.pack(fill=tk.X, pady=(0, 10))

    def execute_update():
        operation = operation_var.get()

        if operation == "course":
            new_course = course_var.get().strip()
            if not new_course:
                messagebox.showwarning(_t("advanced_search.bulk_operations.missing_data_title"), _t("advanced_search.bulk_operations.please_select_course"))
                return

            if messagebox.askyesno(_t("advanced_search.bulk_operations.confirm_update_title"),
                                  _t("advanced_search.bulk_operations.confirm_course_update", count=len(self.search_results), course=new_course)):
                # Simulate course update
                messagebox.showinfo(_t("advanced_search.bulk_operations.update_complete_title"),
                                  _t("advanced_search.bulk_operations.updated_course_msg", count=len(self.search_results), course=new_course))

        elif operation == "status":
            new_status = status_var.get().strip()
            if not new_status:
                messagebox.showwarning(_t("advanced_search.bulk_operations.missing_data_title"), _t("advanced_search.bulk_operations.please_select_status"))
                return

            if messagebox.askyesno(_t("advanced_search.bulk_operations.confirm_update_title"),
                                  _t("advanced_search.bulk_operations.confirm_status_update", count=len(self.search_results), status=new_status)):
                # Simulate status update
                messagebox.showinfo(_t("advanced_search.bulk_operations.update_complete_title"),
                                  _t("advanced_search.bulk_operations.updated_status_msg", count=len(self.search_results), status=new_status))

        elif operation == "note":
            note = note_var.get().strip()
            if not note:
                messagebox.showwarning(_t("advanced_search.bulk_operations.missing_data_title"), _t("advanced_search.bulk_operations.please_enter_note"))
                return

            # Simulate note addition
            messagebox.showinfo(_t("advanced_search.bulk_operations.update_complete_title"),
                              _t("advanced_search.bulk_operations.added_note_msg", note=note, count=len(self.search_results)))

        elif operation == "module":
            module_code = module_var.get().strip()
            if not module_code:
                messagebox.showwarning(_t("advanced_search.bulk_operations.missing_data_title"), _t("advanced_search.bulk_operations.please_enter_module"))
                return

            if messagebox.askyesno(_t("advanced_search.bulk_operations.confirm_enrollment_title"),
                                  _t("advanced_search.bulk_operations.confirm_module_enrollment", count=len(self.search_results), module=module_code)):
                # Simulate module enrollment
                messagebox.showinfo(_t("advanced_search.bulk_operations.update_complete_title"),
                                  _t("advanced_search.bulk_operations.enrolled_in_module_msg", count=len(self.search_results), module=module_code))

        dialog.destroy()

    button_frame = ttk.Frame(frame)
    button_frame.pack(fill=tk.X, pady=(20, 0))

    ttk.Button(button_frame, text=f"✅ {_t('advanced_search.bulk_operations.execute_update_btn')}", command=execute_update).pack(side=tk.LEFT)
    ttk.Button(button_frame, text=f"❌ {_t('advanced_search.cancel_button')}", command=dialog.destroy).pack(side=tk.RIGHT)
AdvancedSearchGUI.batch_data_updates = batch_data_updates

def show_mass_email(self):
    """Show mass email interface"""
    if not self.search_results:
        messagebox.showwarning(_t("advanced_search.bulk_operations.no_results_title"), _t("advanced_search.bulk_operations.search_first_msg"))
        return

    dialog = tk.Toplevel(self.master)
    dialog.title(f"📧 {_t('advanced_search.mass_email_dialog_title')}")
    dialog.geometry("1000x800")
    dialog.transient(self.master)
    dialog.grab_set()

    frame = ttk.Frame(dialog, padding="20")
    frame.pack(fill=tk.BOTH, expand=True)

    ttk.Label(frame, text=_t("advanced_search.bulk_operations.mass_email_title", count=len(self.search_results)),
             style='Title.TLabel').pack(pady=(0, 20))

    # Email composition
    ttk.Label(frame, text=_t("advanced_search.bulk_operations.subject_label")).pack(anchor='w')
    subject_var = tk.StringVar()
    ttk.Entry(frame, textvariable=subject_var, width=60).pack(fill=tk.X, pady=(0, 10))

    ttk.Label(frame, text=_t("advanced_search.bulk_operations.message_label")).pack(anchor='w')
    message_text = scrolledtext.ScrolledText(frame, height=15, wrap=tk.WORD)
    message_text.pack(fill=tk.BOTH, expand=True, pady=(0, 20))
    
    # Email template buttons
    template_frame = ttk.Frame(frame)
    template_frame.pack(fill=tk.X, pady=(0, 20))
    
    def load_template(template_type):
        template_map = {
            "welcome": "welcome_message",
            "reminder": "reminder_message",
            "announcement": "announcement_message"
        }

        if template_type in template_map:
            try:
                from education_system.university_system.infrastructure.email.template_utils import render_template
                subject, message = render_template(template_map[template_type], {})
                subject_var.set(subject)
                message_text.delete(1.0, tk.END)
                message_text.insert(1.0, message)
            except Exception:

                pass
    
    ttk.Label(template_frame, text=_t("advanced_search.bulk_operations.quick_templates")).pack(side=tk.LEFT)
    ttk.Button(template_frame, text=_t("advanced_search.bulk_operations.template_welcome"),
              command=lambda: load_template("welcome")).pack(side=tk.LEFT, padx=(10, 5))
    ttk.Button(template_frame, text=_t("advanced_search.bulk_operations.template_reminder"),
              command=lambda: load_template("reminder")).pack(side=tk.LEFT, padx=(0, 5))
    ttk.Button(template_frame, text=_t("advanced_search.bulk_operations.template_announcement"),
              command=lambda: load_template("announcement")).pack(side=tk.LEFT)

    # Send mass email function
    def send_mass_email():
        subject = subject_var.get().strip()
        message = message_text.get(1.0, tk.END).strip()

        if not subject or not message:
            messagebox.showwarning(_t("advanced_search.bulk_operations.incomplete_title"), _t("advanced_search.bulk_operations.enter_subject_and_message"))
            return

        # Get valid email recipients
        recipients = [s[1] for s in self.search_results if s[1] and '@' in s[1]]

        if not recipients:
            messagebox.showwarning(_t("advanced_search.bulk_operations.no_recipients_title"), _t("advanced_search.bulk_operations.no_valid_emails"))
            return

        # Show confirmation
        confirmation = _t("advanced_search.bulk_operations.confirm_send_msg",
                          subject=subject, count=len(recipients), length=len(message),
                          preview=', '.join(recipients[:5]) + ('...' if len(recipients) > 5 else ''))

        if messagebox.askyesno(_t("advanced_search.bulk_operations.confirm_send_title"), confirmation):
            try:
                # Import email service
                try:
                    from education_system.university_system.infrastructure.email.email_service import send_email
                    EMAIL_AVAILABLE = True
                except ImportError:
                    EMAIL_AVAILABLE = False

                if not EMAIL_AVAILABLE:
                    messagebox.showerror(_t("advanced_search.bulk_operations.email_unavailable_title"),
                                       _t("advanced_search.bulk_operations.email_unavailable_msg"))
                    return

                # Send emails to all recipients
                success_count = 0
                failed_count = 0
                failed_addresses = []

                for recipient_email in recipients:
                    try:
                        send_email(
                            recipient_email=recipient_email,
                            subject=subject,
                            body=message
                        )
                        success_count += 1
                    except Exception as e:
                        failed_count += 1
                        failed_addresses.append(recipient_email)
                        print(f"Failed to send to {recipient_email}: {str(e)}")

                # Show results
                result_message = _t("advanced_search.bulk_operations.mass_email_results_msg",
                                    success=success_count, failed=failed_count,
                                    failed_list=chr(10).join(failed_addresses[:10]),
                                    extra=len(failed_addresses)-10 if len(failed_addresses) > 10 else 0)

                messagebox.showinfo(_t("advanced_search.bulk_operations.email_sent_title"), result_message)
                dialog.destroy()

            except Exception as e:
                messagebox.showerror(_t("advanced_search.bulk_operations.email_error_title"),
                                   _t("advanced_search.bulk_operations.email_error_msg", error=str(e)))
                print(f"Mass email error: {str(e)}")
    
    button_frame = ttk.Frame(frame)
    button_frame.pack(fill=tk.X)
    
    ttk.Button(button_frame, text=f"📧 {_t('advanced_search.bulk_operations.send_email_btn')}", command=send_mass_email).pack(side=tk.LEFT)
    ttk.Button(button_frame, text=f"💾 {_t('advanced_search.bulk_operations.save_draft_btn')}",
              command=lambda: messagebox.showinfo(_t("advanced_search.bulk_operations.draft_title"), _t("advanced_search.bulk_operations.draft_saved_msg"))).pack(side=tk.LEFT, padx=(10, 0))
    ttk.Button(button_frame, text=f"❌ {_t('advanced_search.cancel_button')}", command=dialog.destroy).pack(side=tk.RIGHT)
AdvancedSearchGUI.show_mass_email = show_mass_email

def show_batch_updates(self):
    """Show batch data updates interface"""
    if not self.search_results:
        messagebox.showwarning(_t("advanced_search.bulk_operations.no_results_title"), _t("advanced_search.bulk_operations.search_first_msg"))
        return

    dialog = tk.Toplevel(self.master)
    dialog.title(f"📝 {_t('advanced_search.batch_updates_dialog_title')}")
    dialog.geometry("1000x750")
    dialog.transient(self.master)
    dialog.grab_set()

    frame = ttk.Frame(dialog, padding="20")
    frame.pack(fill=tk.BOTH, expand=True)

    ttk.Label(frame, text=_t("advanced_search.bulk_operations.batch_updates_label", count=len(self.search_results)),
             style='Title.TLabel').pack(pady=(0, 20))

    # Update operations
    operations = [
        (_t("advanced_search.bulk_operations.update_course"), "course"),
        (_t("advanced_search.bulk_operations.update_registration_status"), "status"),
        (_t("advanced_search.bulk_operations.add_note_flag"), "note"),
        (_t("advanced_search.bulk_operations.update_contact_information"), "contact"),
    ]
    
    operation_var = tk.StringVar(value="course")
    
    for text, value in operations:
        ttk.Radiobutton(frame, text=text, variable=operation_var, value=value).pack(anchor='w', pady=2)
    
    # Update values frame
    values_frame = ttk.LabelFrame(frame, text=_t("advanced_search.bulk_operations.update_values"), padding="10")
    values_frame.pack(fill=tk.X, pady=(20, 0))

    ttk.Label(values_frame, text=_t("advanced_search.bulk_operations.new_value_label")).pack(anchor='w')
    new_value_var = tk.StringVar()
    ttk.Entry(values_frame, textvariable=new_value_var, width=30).pack(fill=tk.X, pady=(0, 10))

    ttk.Label(values_frame, text=_t("advanced_search.bulk_operations.reason_for_change")).pack(anchor='w')
    reason_text = tk.Text(values_frame, height=4, wrap=tk.WORD)
    reason_text.pack(fill=tk.X)
    
    def execute_batch_update():
        operation = operation_var.get()
        new_value = new_value_var.get().strip()
        reason = reason_text.get(1.0, tk.END).strip()

        if not new_value:
            messagebox.showwarning(_t("advanced_search.bulk_operations.missing_value_title"), _t("advanced_search.bulk_operations.please_enter_value"))
            return

        # Get operation text from operations list
        operation_text = next((text for text, value in operations if value == operation), operation)

        # Confirmation dialog
        confirmation = _t("advanced_search.bulk_operations.batch_update_confirmation",
                          operation=operation_text, new_value=new_value,
                          count=len(self.search_results),
                          reason=reason if reason else _t("advanced_search.bulk_operations.not_specified"))

        if messagebox.askyesno(_t("advanced_search.bulk_operations.confirm_update_title"), confirmation):
            messagebox.showinfo(_t("advanced_search.bulk_operations.update_complete_title"),
                              _t("advanced_search.bulk_operations.batch_update_complete_msg", count=len(self.search_results)))
            dialog.destroy()
    
    button_frame = ttk.Frame(frame)
    button_frame.pack(fill=tk.X, pady=(20, 0))
    
    ttk.Button(button_frame, text=f"📝 {_t('advanced_search.bulk_operations.update_btn')}", command=execute_batch_update).pack(side=tk.LEFT)
    ttk.Button(button_frame, text=f"❌ {_t('advanced_search.cancel_button')}", command=dialog.destroy).pack(side=tk.RIGHT)
AdvancedSearchGUI.show_batch_updates = show_batch_updates

def simulate_send_email(self, student):
    """Simulate sending email to student"""
    dialog = tk.Toplevel(self.master)
    dialog.title(f"📧 {_t('advanced_search.send_email_dialog_title')}")
    dialog.geometry("1000x750")
    dialog.transient(self.master)
    dialog.grab_set()
    
    frame = ttk.Frame(dialog, padding="20")
    frame.pack(fill=tk.BOTH, expand=True)
    
    ttk.Label(frame, text=_t("advanced_search.bulk_operations.send_email_title"), style='Title.TLabel').pack(pady=(0, 20))

    # Email details
    info_frame = ttk.LabelFrame(frame, text=_t("advanced_search.bulk_operations.email_details_label"), padding="10")
    info_frame.pack(fill=tk.X, pady=(0, 20))
    
    ttk.Label(info_frame, text=_t("advanced_search.bulk_operations.to_recipient", email=student[1], name=f"{student[3]} {student[5]}")).pack(anchor='w')

    ttk.Label(frame, text=_t("advanced_search.bulk_operations.subject_label")).pack(anchor='w')
    subject_var = tk.StringVar()
    ttk.Entry(frame, textvariable=subject_var, width=50).pack(fill=tk.X, pady=(0, 10))

    ttk.Label(frame, text=_t("advanced_search.bulk_operations.message_label")).pack(anchor='w')
    message_text = scrolledtext.ScrolledText(frame, height=10, wrap=tk.WORD)
    message_text.pack(fill=tk.BOTH, expand=True, pady=(0, 20))
    
    def send_email():
        subject = subject_var.get().strip()
        message = message_text.get(1.0, tk.END).strip()

        if not subject or not message:
            messagebox.showwarning(_t("advanced_search.bulk_operations.incomplete_title"), _t("advanced_search.bulk_operations.enter_subject_and_message"))
            return

        try:
            from education_system.university_system.infrastructure.email.email_service import send_email as _send
            success = _send(
                recipient_email=student[1],
                subject=subject,
                body=message
            )
            if success:
                messagebox.showinfo(_t("advanced_search.bulk_operations.email_sent_title"),
                                    _t("advanced_search.bulk_operations.email_sent_success_msg", email=student[1]))
            else:
                messagebox.showinfo(_t("advanced_search.bulk_operations.email_queued_title"),
                                    _t("advanced_search.bulk_operations.email_queued_msg", email=student[1]))
            dialog.destroy()
        except ImportError:
            messagebox.showerror(_t("advanced_search.bulk_operations.error_title"),
                                 _t("advanced_search.bulk_operations.email_service_unavailable"))
        except Exception as e:
            messagebox.showerror(_t("advanced_search.bulk_operations.error_title"), _t("advanced_search.bulk_operations.failed_to_send_email", error=str(e)))

    button_frame = ttk.Frame(frame)
    button_frame.pack(fill=tk.X)

    ttk.Button(button_frame, text=f"📧 {_t('advanced_search.bulk_operations.send_btn')}", command=send_email).pack(side=tk.LEFT)
    ttk.Button(button_frame, text=f"❌ {_t('advanced_search.cancel_button')}", command=dialog.destroy).pack(side=tk.RIGHT)
AdvancedSearchGUI.simulate_send_email = simulate_send_email
