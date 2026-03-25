"""Report generation (PDF, Excel, interactive HTML)."""

import os
from datetime import datetime, timedelta

from education_system.university_system.modules.shared.services.analytics.enhanced_reporting._compat import (
    pd, SimpleDocTemplate, Paragraph, Spacer, Image,
    getSampleStyleSheet, ParagraphStyle, colors, A4, inch,
)
from education_system.university_system.modules.shared.services.analytics.enhanced_reporting.config import CONFIG, AVAILABLE_SECTIONS, SystemConfig, logger
from education_system.university_system.modules.shared.services.analytics.enhanced_reporting.cache import CacheManager
from education_system.university_system.modules.shared.services.analytics.enhanced_reporting.data_quality import DataQualityMonitor
from education_system.university_system.modules.shared.services.analytics.enhanced_reporting.predictive import PredictiveAnalytics
from education_system.university_system.modules.shared.services.analytics.enhanced_reporting.visualization import (
    AdvancedVisualization,
    create_advanced_visualization, create_standard_chart,
    generate_statistical_summary, create_enhanced_data_table,
)
from education_system.university_system.modules.shared.services.analytics.enhanced_reporting.templates_db import get_template as _get_template
from education_system.university_system.modules.shared.services.analytics.enhanced_reporting.data_retrieval import get_section_dataframe


def generate_report(template, start_date=None, end_date=None, format="pdf", comparison_date=None):
    """Generate a report with enhanced features"""
    if isinstance(template, str):
        template = _get_template(template)
        if not template:
            logger.error(f"Template '{template}' not found")
            return None

    # Set default date range
    if not end_date:
        end_date = datetime.now().strftime("%Y-%m-%d")
    if not start_date:
        end_dt = datetime.strptime(end_date, "%Y-%m-%d")
        start_dt = end_dt - timedelta(days=30)
        start_date = start_dt.strftime("%Y-%m-%d")

    # Check cache first
    config = SystemConfig.load_config()
    if config.get('performance', {}).get('enable_caching', True):
        cache_key = CacheManager.get_cache_key(template.name, start_date, end_date, template.filters)
        cached_data = CacheManager.get_cached_report(cache_key)

        if cached_data:
            logger.info(f"Using cached data for report {template.name}")

    # Generate report filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{template.name.replace(' ', '_')}_{timestamp}"

    try:
        if format == "pdf":
            return generate_enhanced_pdf_report(template, filename, start_date, end_date, comparison_date)
        elif format == "excel":
            return generate_enhanced_excel_report(template, filename, start_date, end_date, comparison_date)
        elif format == "interactive":
            return generate_interactive_report(template, filename, start_date, end_date)
        else:
            logger.error(f"Unsupported format: {format}")
            return None

    except Exception as e:
        logger.error(f"Error generating report: {str(e)}")
        raise


def generate_enhanced_pdf_report(template, filename, start_date, end_date, comparison_date=None):
    """Generate an enhanced PDF report with all new features"""
    file_path = os.path.join(CONFIG['reports_dir'], f"{filename}.pdf")

    doc = SimpleDocTemplate(file_path, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    # Enhanced title with security marking
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontSize=24,
        spaceAfter=30,
        textColor=colors.darkblue
    )

    title_text = template.name
    if template.security_level != 'normal':
        title_text += f" ({template.security_level.upper()})"

    elements.append(Paragraph(title_text, title_style))

    # Report metadata
    metadata_style = ParagraphStyle(
        'Metadata',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.grey,
        spaceAfter=20
    )

    metadata_text = f"""
    Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>
    Template Version: {template.version}<br/>
    Report Period: {start_date} to {end_date}<br/>
    Security Level: {template.security_level.title()}
    """
    elements.append(Paragraph(metadata_text, metadata_style))

    # Description
    elements.append(Paragraph(template.description, styles["Normal"]))
    elements.append(Spacer(1, 0.25*inch))

    # Data quality summary
    if 'data_quality_report' in template.sections:
        quality_report = DataQualityMonitor.run_quality_checks()
        elements.extend(generate_quality_section(quality_report, styles))

    # Generate each section with enhanced visualizations
    for section in template.sections:
        if section in AVAILABLE_SECTIONS:
            try:
                section_elements = generate_enhanced_section(
                    section, start_date, end_date, comparison_date,
                    template.filters, template.visualization_type
                )
                elements.extend(section_elements)
            except Exception as e:
                logger.error(f"Error generating section {section}: {str(e)}")
                # Add error message to report
                error_text = f"Error generating {section}: {str(e)}"
                elements.append(Paragraph(error_text, styles["Normal"]))
                elements.append(Spacer(1, 0.25*inch))

    # Add predictive analytics if requested
    if 'predictive_analytics' in template.sections:
        predictions = PredictiveAnalytics.predict_dropout_risk()
        elements.extend(generate_predictions_section(predictions, styles))

    # Build the PDF
    doc.build(elements)
    logger.info(f"Enhanced PDF report generated: {file_path}")
    return file_path


def generate_enhanced_section(section, start_date, end_date, comparison_date=None,
                            filters=None, visualization_type='standard'):
    """Generate enhanced report sections with advanced visualizations"""
    elements = []
    styles = getSampleStyleSheet()

    # Add section heading
    section_title = section.replace('_', ' ').title()
    section_style = ParagraphStyle(
        'SectionTitle',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.darkblue,
        spaceAfter=15
    )
    elements.append(Paragraph(section_title, section_style))

    # Get data for the section
    df = get_section_dataframe(section, start_date, end_date, filters)

    if df is None or df.empty:
        elements.append(Paragraph("No data available for this section.", styles["Normal"]))
        elements.append(Spacer(1, 0.25*inch))
        return elements

    # Generate appropriate visualization based on type
    chart_path = None

    try:
        if visualization_type == 'advanced':
            chart_path = create_advanced_visualization(section, df)
        elif visualization_type == 'interactive':
            # For interactive, create standard chart for PDF
            chart_path = create_standard_chart(section, df)
        else:
            chart_path = create_standard_chart(section, df)

        # Add chart to report ONLY if it's a valid image file
        if chart_path and os.path.exists(chart_path) and chart_path.lower().endswith(('.png', '.jpg', '.jpeg')):
            try:
                img = Image(chart_path, width=6*inch, height=4*inch)
                elements.append(img)
                elements.append(Spacer(1, 0.15*inch))
            except Exception as e:
                logger.error(f"Error adding chart to report: {str(e)}")
                # Add text description instead
                elements.append(Paragraph(f"Chart generated: {os.path.basename(chart_path)}", styles["Normal"]))
    except Exception as e:
        logger.error(f"Error creating visualization for {section}: {str(e)}")

    # Add statistical summary for numeric data
    if section in ['grade_distribution', 'age_distribution']:
        summary_stats = generate_statistical_summary(df, section)
        if summary_stats:
            elements.append(Paragraph("Statistical Summary:", styles["Heading4"]))
            elements.append(summary_stats)
            elements.append(Spacer(1, 0.15*inch))

    # Add data table with enhanced styling
    if len(df) <= 100:  # Only show table for smaller datasets
        data_table = create_enhanced_data_table(df)
        elements.append(data_table)
    else:
        # For large datasets, show summary
        summary_text = f"Dataset contains {len(df)} records. Top 10 records shown in visualization."
        elements.append(Paragraph(summary_text, styles["Italic"]))

    elements.append(Spacer(1, 0.25*inch))
    return elements


def generate_quality_section(quality_report, styles):
    """Generate data quality section for reports"""
    elements = []

    elements.append(Paragraph("Data Quality Report", styles["Heading2"]))
    elements.append(Spacer(1, 0.1*inch))

    # Summary of quality issues
    checks = quality_report.get('checks', {})

    if 'missing_data' in checks:
        missing = checks['missing_data']['students']
        total = missing['total_records']

        if total > 0:
            missing_rate = (missing['missing_emails'] + missing['missing_names'] + missing['missing_courses']) / (total * 3) * 100
            quality_text = f"Data completeness: {100 - missing_rate:.1f}% complete"
            elements.append(Paragraph(quality_text, styles["Normal"]))

    if 'duplicates' in checks:
        duplicates = checks['duplicates']['duplicate_emails']
        if duplicates > 0:
            elements.append(Paragraph(f"Warning: {duplicates} duplicate email addresses found", styles["Normal"]))

    if 'invalid_data' in checks:
        invalid = checks['invalid_data']
        if invalid['invalid_ages'] > 0 or invalid['invalid_emails'] > 0:
            elements.append(Paragraph(f"Data validation issues: {invalid['invalid_ages']} invalid ages, {invalid['invalid_emails']} invalid emails", styles["Normal"]))

    elements.append(Spacer(1, 0.25*inch))
    return elements


def generate_predictions_section(predictions, styles):
    """Generate predictive analytics section"""
    elements = []

    elements.append(Paragraph("Predictive Analytics", styles["Heading2"]))
    elements.append(Spacer(1, 0.1*inch))

    if 'error' in predictions:
        elements.append(Paragraph(f"Analysis unavailable: {predictions['error']}", styles["Normal"]))
    else:
        if 'model_accuracy' in predictions:
            accuracy_text = f"Model accuracy: {predictions['model_accuracy']:.2%}"
            elements.append(Paragraph(accuracy_text, styles["Normal"]))

        if 'high_risk_students' in predictions:
            risk_count = len(predictions['high_risk_students'])
            if risk_count > 0:
                risk_text = f"Students at high risk of dropout: {risk_count}"
                elements.append(Paragraph(risk_text, styles["Normal"]))

        if 'feature_importance' in predictions:
            importance = predictions['feature_importance']
            sorted_features = sorted(importance.items(), key=lambda x: x[1], reverse=True)

            elements.append(Paragraph("Most important factors:", styles["Heading4"]))
            for feature, score in sorted_features[:3]:
                feature_text = f"\u2022 {feature.replace('_', ' ').title()}: {score:.3f}"
                elements.append(Paragraph(feature_text, styles["Normal"]))

    elements.append(Spacer(1, 0.25*inch))
    return elements


def generate_enhanced_excel_report(template, filename, start_date, end_date, comparison_date=None):
    """Generate an enhanced Excel report with multiple sheets and formatting"""
    file_path = os.path.join(CONFIG['reports_dir'], f"{filename}.xlsx")

    try:
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # Summary sheet
            summary_data = {
                'Report Name': [template.name],
                'Generated': [datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                'Period': [f"{start_date} to {end_date}"],
                'Template Version': [template.version],
                'Security Level': [template.security_level]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)

            # Generate data for each section
            for section in template.sections:
                if section in AVAILABLE_SECTIONS:
                    try:
                        df = get_section_dataframe(section, start_date, end_date, template.filters)
                        if df is not None and not df.empty:
                            sheet_name = section[:31]  # Excel sheet name limit
                            df.to_excel(writer, sheet_name=sheet_name, index=False)
                    except Exception as e:
                        logger.error(f"Error adding section {section} to Excel: {str(e)}")

            # Format the Excel file
            workbook = writer.book

            # Add formatting to summary sheet
            if 'Summary' in workbook.sheetnames:
                summary_sheet = workbook['Summary']
                from openpyxl.styles import Font, PatternFill

                # Header formatting
                for cell in summary_sheet[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

        logger.info(f"Enhanced Excel report generated: {file_path}")
        return file_path

    except Exception as e:
        logger.error(f"Error generating Excel report: {str(e)}")
        raise


def generate_interactive_report(template, filename, start_date, end_date):
    """Generate an interactive HTML report with Plotly charts"""
    file_path = os.path.join(CONFIG['reports_dir'], f"{filename}.html")

    try:
        # Collect data for all sections
        data_dict = {}

        for section in template.sections:
            if section in AVAILABLE_SECTIONS:
                try:
                    df = get_section_dataframe(section, start_date, end_date, template.filters)
                    if df is not None and not df.empty:
                        data_dict[section] = df
                except Exception as e:
                    logger.error(f"Error getting data for section {section}: {str(e)}")

        # Create interactive dashboard
        dashboard_path = AdvancedVisualization.create_interactive_dashboard(data_dict)

        if dashboard_path and os.path.exists(dashboard_path):
            # Read the generated HTML and enhance it
            with open(dashboard_path, 'r') as f:
                html_content = f.read()

            # Add custom styling and metadata
            enhanced_html = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <title>{template.name} - Interactive Report</title>
                <meta charset="utf-8">
                <style>
                    body {{ font-family: Arial, sans-serif; margin: 20px; }}
                    .header {{ text-align: center; margin-bottom: 30px; }}
                    .metadata {{ background: #f5f5f5; padding: 15px; border-radius: 5px; margin-bottom: 20px; }}
                </style>
            </head>
            <body>
                <div class="header">
                    <h1>{template.name}</h1>
                    <p>Interactive Report - Generated {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                </div>
                <div class="metadata">
                    <p><strong>Report Period:</strong> {start_date} to {end_date}</p>
                    <p><strong>Template Version:</strong> {template.version}</p>
                    <p><strong>Security Level:</strong> {template.security_level.title()}</p>
                </div>
                {html_content.split('<body>')[1] if '<body>' in html_content else html_content}
            </body>
            </html>
            """

            with open(file_path, 'w') as f:
                f.write(enhanced_html)

            # Clean up temporary dashboard file
            if dashboard_path != file_path:
                os.remove(dashboard_path)

        logger.info(f"Interactive report generated: {file_path}")
        return file_path

    except Exception as e:
        logger.error(f"Error generating interactive report: {str(e)}")
        raise
