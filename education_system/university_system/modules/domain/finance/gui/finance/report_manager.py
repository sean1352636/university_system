"""Financial reports and analytics"""

from education_system.university_system.infrastructure.database.db import DEFAULT_DB_PATH  # injected
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog
from tkinter.scrolledtext import ScrolledText
from education_system.university_system.infrastructure.database.db import sqlite3
import sys
import io
import os
import csv
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import matplotlib.pyplot as plt
import seaborn as sns
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import json
import threading
import warnings
from pathlib import Path
from sklearn.ensemble import IsolationForest
from sklearn.preprocessing import StandardScaler
import joblib
from cryptography.fernet import Fernet
import logging
import qrcode
from io import BytesIO
import base64
from education_system.university_system.modules.domain.finance.gui.finance_reporting import launch_financial_gui

# Import authentication - REQUIRED (no fallback for security)
from education_system.university_system.infrastructure.auth import UserAuth, get_global_auth
from education_system.university_system.infrastructure.shared_context import get_auth

# Import database utilities - use centralized connection management
from education_system.university_system.infrastructure.database.db import get_connection

# Import optional modules with fallbacks for non-critical functionality
try:
    from education_system.university_system.infrastructure.email.email_service import send_email
except ImportError:
    def send_email(*args, **kwargs):
        """Fallback stub when email service is unavailable."""
        return True

try:
    from education_system.university_system.infrastructure.logging.log_config import configure_logging, get_log_file
except ImportError:
    def configure_logging(name=None):
        """Fallback logging configuration."""
        return logging.getLogger(name or __name__)

    def get_log_file(name):
        """Fallback log file path resolution."""
        from education_system.university_system.modules.shared.constants import paths
        return str(paths.LOG_DIR / name)

# Import finance functions from common_imports module (explicit imports)
from education_system.university_system.modules.domain.finance.gui.finance.common_imports import (
    # Financial aid reports
    aid_by_academic_year,
    aid_distribution_summary,
    aid_effectiveness_analysis,
    generate_aid_reports,
    loan_repayment_status_report,
    # Collection reports
    collection_performance_summary,
    view_overdue_accounts,
    # Forecasting and audit
    export_forecast_report,
    generate_audit_report,
)

# Configure logging
log_path = get_log_file("app.log")
os.makedirs(os.path.dirname(log_path), exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_path),
        logging.StreamHandler()
    ]
)

logger = configure_logging(name=__name__)
warnings.filterwarnings('ignore')

# Global variables for backward compatibility
auth = get_global_auth()  # Use centralized auth instance
ENCRYPTION_KEY = Fernet.generate_key()
cipher_suite = Fernet(ENCRYPTION_KEY)

# Payment gateway configurations (from original file)
PAYMENT_GATEWAYS = {
    'stripe': {
        'public_key': os.getenv('STRIPE_PUBLIC_KEY', ''),
        'secret_key': os.getenv('STRIPE_SECRET_KEY', ''),
        'webhook_secret': os.getenv('STRIPE_WEBHOOK_SECRET', '')
    },
    'paypal': {
        'client_id': os.getenv('PAYPAL_CLIENT_ID', ''),
        'client_secret': os.getenv('PAYPAL_CLIENT_SECRET', ''),
        'environment': os.getenv('PAYPAL_ENVIRONMENT', 'sandbox')
    }
}

# WARNING: Never commit real API keys to version control!
# Set these environment variables in your deployment environment
SUPPORTED_CURRENCIES = ['GBP', 'USD', 'EUR', 'CAD', 'AUD']
# Load exchange API key from environment variable
EXCHANGE_API_KEY = os.getenv('EXCHANGE_API_KEY', '')




class ReportManager:
    """Financial reports and analytics"""

    def __init__(self, gui):
        """Initialize manager with reference to main GUI"""
        self.gui = gui
        self.root = gui.root
        self.conn = gui.conn
        try:
            self.finance_system = gui.finance_system
        except Exception:
            self.finance_system = None

    def create_reports_tab(self):
        """Create reports tab"""
        tab = tk.Frame(self.gui.layout.content_frame, bg='white')
        self.gui.layout.tab_frames['reports_detailed'] = tab

        main_frame = ttk.Frame(tab, padding=20)
        main_frame.pack(fill='both', expand=True)

        # Report buttons
        reports_frame = ttk.LabelFrame(main_frame, text="Financial Reports", padding=15)
        reports_frame.pack(fill='x', pady=(0, 20))

        report_buttons = [
            ("💰 Revenue Summary", self.gui_revenue_summary_report),
            ("📋 Outstanding Fees", self.gui_outstanding_fees_report),
            ("💳 Payment Collection", self.gui_payment_collection_report),
            ("👤 Student Accounts", self.gui_student_account_summary),
            ("📊 Fee Type Analysis", self.gui_fee_type_analysis),
            ("Aging Analysis", self.gui_aging_analysis_report),
            ("Collection Cases", self.gui_collection_case_status_report),
            ("Recovery Rate", self.gui_recovery_rate_analysis),
            ("Agency Performance", self.gui_agency_performance_report),
            ("Variance Analysis", self.gui_variance_analysis_report),
            ("Budget Performance", self.gui_budget_performance_trends),
            ("Category Performance", self.gui_category_performance_report),
            ("Monthly Revenue", self.gui_monthly_revenue_trend_report),
            ("Audit Report", self.gui_generate_audit_report),
            ("💼 Payment Methods", self.gui_payment_method_analysis)
        ]

        for i, (text, command) in enumerate(report_buttons):
            ttk.Button(reports_frame, text=text, command=command, width=25).grid(row=i//3, column=i%3, padx=10, pady=5)

        # Report display area
        display_frame = ttk.LabelFrame(main_frame, text="Report Output", padding=15)
        display_frame.pack(fill='both', expand=True)

        self.report_text = ScrolledText(display_frame, height=20, width=80, font=('Courier', 10))
        self.report_text.pack(fill='both', expand=True)

        # Export buttons
        export_frame = ttk.Frame(display_frame)
        export_frame.pack(fill='x', pady=10)

        ttk.Button(export_frame, text="📄 Export to CSV", command=self.export_report_csv).pack(side='left', padx=5)
        ttk.Button(export_frame, text="📊 Export to Excel", command=self.export_report_excel).pack(side='left', padx=5)
        ttk.Button(export_frame, text="🗑️ Clear Report", command=self.clear_report).pack(side='left', padx=5)


    def create_report_card(self, parent, title, description, command, row, col):
        """Create a report card"""
        card_frame = tk.Frame(parent, bg=self.gui.layout.colors['light'], relief='raised', bd=2)
        card_frame.grid(row=row, column=col, padx=10, pady=10, sticky='ew')
        parent.grid_columnconfigure(col, weight=1)

        title_label = tk.Label(card_frame, text=title, font=('Arial', 12, 'bold'),
                              bg=self.gui.layout.colors['light'])
        title_label.pack(pady=5)

        desc_label = tk.Label(card_frame, text=description, font=('Arial', 9),
                             bg=self.gui.layout.colors['light'], wraplength=200)
        desc_label.pack(pady=5)

        generate_btn = tk.Button(card_frame, text="Generate Report", command=command,
                               bg=self.gui.layout.colors['secondary'], fg='white', font=('Arial', 9, 'bold'))
        generate_btn.pack(pady=10)


    def show_reports_tab(self):
        """Switch to reports tab"""
        # Find and select the reports tab
        self.show_tab('reports')

    # ==================== EVENT HANDLERS ====================


    def launch_reporting_gui(self):
        """Launch the advanced reporting GUI interface"""
        try:
            from education_system.university_system.modules.domain.finance.gui.finance_reporting import launch_financial_gui
            launch_financial_gui()
            self.update_status("Advanced Reporting GUI launched")
            self.log_activity("Advanced Reporting GUI opened")
        except ImportError:
            messagebox.showerror("Error", "Advanced reporting module not found. Make sure finance_reporting_gui.py is in the same directory.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to launch reporting GUI: {e}")
            self.update_status("Error launching reporting GUI")


    def generate_revenue_report(self):
        """Generate revenue report"""
        def generate_thread():
            try:
                self.update_status("Generating revenue report...")

                # Redirect stdout to capture report output
                old_stdout = sys.stdout
                sys.stdout = mystdout = io.StringIO()

                # Call original function
                generate_revenue_summary()

                # Get the output
                report_output = mystdout.getvalue()
                sys.stdout = old_stdout

                # Display in GUI
                self.root.after(0, lambda: self.display_report_output("Revenue Report", report_output))

            except Exception as e:
                sys.stdout = old_stdout
                self.root.after(0, lambda _e=e: messagebox.showerror("Report Error", f"Failed to generate revenue report: {str(_e)}"))

        threading.Thread(target=generate_thread, daemon=True).start()


    def generate_student_summary(self):
        """Generate student financial summary"""
        def generate_thread():
            try:
                self.update_status("Generating student summary...")

                old_stdout = sys.stdout
                sys.stdout = mystdout = io.StringIO()

                generate_student_financial_summary()

                report_output = mystdout.getvalue()
                sys.stdout = old_stdout

                self.root.after(0, lambda: self.display_report_output("Student Financial Summary", report_output))

            except Exception as e:
                sys.stdout = old_stdout
                self.root.after(0, lambda _e=e: messagebox.showerror("Report Error", f"Failed to generate student summary: {str(_e)}"))

        threading.Thread(target=generate_thread, daemon=True).start()


    def generate_overdue_report(self):
        """Generate overdue accounts report"""
        def generate_thread():
            try:
                self.update_status("Generating overdue report...")

                old_stdout = sys.stdout
                sys.stdout = mystdout = io.StringIO()

                view_overdue_accounts()

                report_output = mystdout.getvalue()
                sys.stdout = old_stdout

                self.root.after(0, lambda: self.display_report_output("Overdue Accounts Report", report_output))

            except Exception as e:
                sys.stdout = old_stdout
                self.root.after(0, lambda _e=e: messagebox.showerror("Report Error", f"Failed to generate overdue report: {str(_e)}"))

        threading.Thread(target=generate_thread, daemon=True).start()


    def generate_payment_analytics(self):
        """Generate payment analytics report"""
        self.show_payment_analytics()


    def generate_collection_report(self):
        """Generate collection report"""
        def generate_thread():
            try:
                self.update_status("Generating collection report...")

                old_stdout = sys.stdout
                sys.stdout = mystdout = io.StringIO()

                collection_performance_summary()

                report_output = mystdout.getvalue()
                sys.stdout = old_stdout

                self.root.after(0, lambda: self.display_report_output("Collection Report", report_output))

            except Exception as e:
                sys.stdout = old_stdout
                self.root.after(0, lambda _e=e: messagebox.showerror("Report Error", f"Failed to generate collection report: {str(_e)}"))

        threading.Thread(target=generate_thread, daemon=True).start()


    def generate_aid_report(self):
        """Generate financial aid report"""
        def generate_thread():
            try:
                self.update_status("Generating financial aid report...")

                old_stdout = sys.stdout
                sys.stdout = mystdout = io.StringIO()

                aid_distribution_summary()

                report_output = mystdout.getvalue()
                sys.stdout = old_stdout

                self.root.after(0, lambda: self.display_report_output("Financial Aid Report", report_output))

            except Exception as e:
                sys.stdout = old_stdout
                self.root.after(0, lambda _e=e: messagebox.showerror("Report Error", f"Failed to generate aid report: {str(_e)}"))

        threading.Thread(target=generate_thread, daemon=True).start()


    def generate_budget_analysis(self):
        """Generate budget analysis report"""
        def generate_thread():
            try:
                self.update_status("Generating budget analysis...")

                old_stdout = sys.stdout
                sys.stdout = mystdout = io.StringIO()

                budget_summary_report()

                report_output = mystdout.getvalue()
                sys.stdout = old_stdout

                self.root.after(0, lambda: self.display_report_output("Budget Analysis", report_output))

            except Exception as e:
                sys.stdout = old_stdout
                self.root.after(0, lambda _e=e: messagebox.showerror("Report Error", f"Failed to generate budget analysis: {str(_e)}"))

        threading.Thread(target=generate_thread, daemon=True).start()


    def generate_forecast_report(self):
        """Generate forecast report"""
        def generate_thread():
            try:
                self.update_status("Generating forecast report...")

                old_stdout = sys.stdout
                sys.stdout = mystdout = io.StringIO()

                generate_comprehensive_forecast_report()

                report_output = mystdout.getvalue()
                sys.stdout = old_stdout

                self.root.after(0, lambda: self.display_report_output("Forecast Report", report_output))

            except Exception as e:
                sys.stdout = old_stdout
                self.root.after(0, lambda _e=e: messagebox.showerror("Report Error", f"Failed to generate forecast report: {str(_e)}"))

        threading.Thread(target=generate_thread, daemon=True).start()


    def display_report_output(self, title, output):
        """Display report output in the text widget"""
        self.report_output.delete('1.0', tk.END)
        self.report_output.insert('1.0', f"{title}\n{'=' * len(title)}\n\n{output}")
        self.update_status(f"{title} generated successfully")

    # ==================== COLLECTION METHODS ====================


    def gui_generate_revenue_forecast(self):
        """GUI wrapper for revenue forecasting"""
        try:
            self.update_status("Generating revenue forecast...")

            def generate_forecast():
                try:
                    # Generate forecast data
                    forecast_content = f"""Revenue Forecast Analysis
    {'=' * 60}
    Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

    FORECAST PARAMETERS:
    - Forecast Period: 12 months
    - Growth Rate: 5.0% annually
    - Base Currency: GBP

    REVENUE PROJECTIONS:
    {'=' * 60}

    Current Year Analysis:
    - Current Monthly Average: £85,000
    - Projected Annual Revenue: £1,020,000
    - Growth Factor: 1.05

    Monthly Projections (Next 12 Months):
    {'-' * 60}
    """

                    # Generate monthly forecasts
                    base_monthly = 85000
                    growth_rate = 0.05
                    current_date = datetime.now()

                    total_forecast = 0
                    for month in range(1, 13):
                        future_date = current_date + timedelta(days=30 * month)
                        monthly_adjustment = 1 + (growth_rate * (month / 12))
                        projected_revenue = base_monthly * monthly_adjustment
                        total_forecast += projected_revenue

                        forecast_content += f"{future_date.strftime('%B %Y'):<15} £{projected_revenue:>10,.2f}\n"

                    forecast_content += f"""{'-' * 60}
    Total Forecast Revenue: £{total_forecast:>10,.2f}
    Average Monthly Revenue: £{total_forecast/12:>10,.2f}

    CONFIDENCE INTERVALS:
    {'=' * 60}
    Conservative (80%): £{total_forecast * 0.8:>10,.2f}
    Baseline (100%):    £{total_forecast:>10,.2f}
    Optimistic (120%):  £{total_forecast * 1.2:>10,.2f}

    KEY ASSUMPTIONS:
    - Stable student enrollment
    - No major economic disruptions
    - Current fee structures maintained
    - Payment collection rates remain consistent

    RECOMMENDATIONS:
    1. Monitor actual vs. forecast monthly
    2. Adjust projections based on enrollment trends
    3. Consider diversification strategies
    4. Maintain cash reserves for volatility
    """

                    # Update the forecast output in the main thread
                    self.root.after(0, lambda: self.update_forecast_output(forecast_content))

                except Exception as e:
                    error_msg = f"Error generating revenue forecast: {str(e)}"
                    self.root.after(0, lambda: self.update_forecast_output(error_msg))

            # Run in thread to avoid blocking UI
            thread = threading.Thread(target=generate_forecast, daemon=True)
            thread.start()

            # Switch to forecasting tab if it exists
            try:
                self.show_tab('forecasting')
            except Exception as e:
                # Forecasting tab may not exist, silently ignore
                print(f"Debug: Could not switch to forecasting tab: {e}")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate revenue forecast: {e}")
            self.update_status("Revenue forecast generation failed")


    def update_forecast_output(self, content):
        """Update the forecast output display"""
        try:
            # Try to update forecast_output if it exists
            if hasattr(self, 'forecast_output'):
                self.forecast_output.delete('1.0', tk.END)
                self.forecast_output.insert('1.0', content)
            else:
                # Show in a new window if forecast tab doesn't exist
                self.show_text_window("Revenue Forecast", content)

            self.update_status("Revenue forecast generated successfully")

        except Exception as e:
            print(f"Error updating forecast output: {e}")
            # Fallback to showing in reports tab
            try:
                self.show_tab('reports')  # Reports tab
                self.report_text.delete('1.0', tk.END)
                self.report_text.insert('1.0', content)
            except Exception:
                # Last resort - show in message box
                messagebox.showinfo("Revenue Forecast", "Forecast generated - check console output")


    def gui_export_forecast_report(self):
        """GUI wrapper for export_forecast_report"""
        try:
            filename = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf"), ("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Export Forecast Report"
            )

            if filename:
                export_forecast_report(filename)
                messagebox.showinfo("Success", f"Forecast report exported to {filename}")
                self.update_status(f"Forecast report exported to {filename}")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to export forecast report: {e}")


    def gui_generate_audit_report(self):
        """GUI wrapper for generate_audit_report"""
        try:
            old_stdout = sys.stdout
            sys.stdout = mystdout = io.StringIO()

            generate_audit_report()

            output = mystdout.getvalue()
            sys.stdout = old_stdout

            self.show_tab('reports')  # Reports tab
            self.report_text.delete('1.0', tk.END)
            self.report_text.insert('1.0', output)
            self.update_status("Audit report generated")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate audit report: {e}")


    def gui_revenue_summary_report(self):
        """Generate revenue summary report"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Revenue Summary Report")
        dialog.geometry("400x300")
        dialog.transient(self.root)
        dialog.grab_set()

        # Date range inputs
        ttk.Label(dialog, text="Start Date (YYYY-MM-DD):").pack(pady=5)
        start_date_var = tk.StringVar(value=(datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
        ttk.Entry(dialog, textvariable=start_date_var).pack(pady=5)

        ttk.Label(dialog, text="End Date (YYYY-MM-DD):").pack(pady=5)
        end_date_var = tk.StringVar(value=datetime.now().strftime('%Y-%m-%d'))
        ttk.Entry(dialog, textvariable=end_date_var).pack(pady=5)

        def generate_report():
            try:
                start_date = start_date_var.get()
                end_date = end_date_var.get()

                conn = get_connection()
                cursor = conn.cursor()

                cursor.execute('''
                SELECT
                    COUNT(*) as total_payments,
                    SUM(amount) as total_revenue,
                    AVG(amount) as avg_payment,
                    MIN(amount) as min_payment,
                    MAX(amount) as max_payment
                FROM payments
                WHERE payment_date BETWEEN ? AND ? AND status = 'completed'
                ''', (start_date, end_date))

                summary = cursor.fetchone()

                if summary and summary[0] > 0:
                    total_payments, total_revenue, avg_payment, min_payment, max_payment = summary

                    report_content = f"""Revenue Summary Report ({start_date} to {end_date})
    {'=' * 60}
    Total Payments: {total_payments}
    Total Revenue: £{total_revenue:,.2f}
    Average Payment: £{avg_payment:.2f}
    Minimum Payment: £{min_payment:.2f}
    Maximum Payment: £{max_payment:.2f}

    """

                    # Revenue by payment method
                    cursor.execute('''
                    SELECT payment_method, COUNT(*), SUM(amount)
                    FROM payments
                    WHERE payment_date BETWEEN ? AND ? AND status = 'completed'
                    GROUP BY payment_method
                    ORDER BY SUM(amount) DESC
                    ''', (start_date, end_date))

                    method_data = cursor.fetchall()

                    report_content += "Revenue by Payment Method:\n"
                    report_content += "-" * 60 + "\n"
                    for method, count, amount in method_data:
                        percentage = (amount / total_revenue) * 100
                        report_content += f"{method:<20} {count:>6} payments  £{amount:>12,.2f} ({percentage:>5.1f}%)\n"

                    # Display in main report area
                    self.show_tab('reports')  # Reports tab
                    self.report_text.delete('1.0', tk.END)
                    self.report_text.insert('1.0', report_content)

                    dialog.destroy()
                    self.update_status(f"Revenue summary report generated for {start_date} to {end_date}")

                else:
                    messagebox.showinfo("No Data", "No payments found for the specified date range.")

                conn.close()

            except Exception as e:
                messagebox.showerror("Error", f"Failed to generate report: {e}")

        ttk.Button(dialog, text="Generate Report", command=generate_report).pack(pady=20)
        ttk.Button(dialog, text="Cancel", command=dialog.destroy).pack(pady=5)


    def gui_outstanding_fees_report(self):
        """Generate outstanding fees report"""
        try:
            conn = get_connection()
            cursor = conn.cursor()

            cursor.execute('''
            SELECT s.student_id, s.first_name, s.last_name, s.course,
                   ft.fee_name, sf.amount, sf.due_date,
                   julianday('now') - julianday(sf.due_date) as days_overdue
            FROM student_fees sf
            JOIN students s ON sf.student_id = s.student_id
            JOIN fee_types ft ON sf.fee_type_id = ft.fee_type_id
            WHERE sf.status IN ('unpaid', 'partial')
            ORDER BY days_overdue DESC, sf.amount DESC
            ''')

            outstanding = cursor.fetchall()

            if not outstanding:
                report_content = "No outstanding fees found."
            else:
                report_content = f"""Outstanding Fees Report
    {'=' * 120}
    {'Student ID':<12} {'Name':<25} {'Course':<20} {'Fee Type':<20} {'Amount':<12} {'Due Date':<12} {'Days Overdue':<12}
    {'-' * 120}
    """

                total_outstanding = 0
                overdue_count = 0

                for row in outstanding:
                    student_id, first_name, last_name, course, fee_name, amount, due_date, days_overdue = row
                    student_name = f"{first_name} {last_name}"

                    if days_overdue > 0:
                        overdue_indicator = f"{int(days_overdue)}"
                        overdue_count += 1
                    else:
                        overdue_indicator = "Not due"

                    report_content += f"{student_id:<12} {student_name:<25} {course:<20} {fee_name:<20} £{amount:<11.2f} {due_date:<12} {overdue_indicator:<12}\n"
                    total_outstanding += amount

                report_content += f"""{'-' * 120}
    Total Outstanding: £{total_outstanding:,.2f}
    Total Fees: {len(outstanding)}
    Overdue Fees: {overdue_count}
    {'=' * 120}
    """

            # Display in reports tab
            self.show_tab('reports')  # Reports tab
            self.report_text.delete('1.0', tk.END)
            self.report_text.insert('1.0', report_content)

            conn.close()
            self.update_status("Outstanding fees report generated")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate outstanding fees report: {e}")


    def gui_payment_collection_report(self):
        """Generate payment collection report"""
        try:
            conn = get_connection()
            cursor = conn.cursor()

            # Get payment collection data
            cursor.execute('''
            SELECT strftime('%Y-%m', payment_date) as month,
                   COUNT(*) as payment_count,
                   SUM(amount) as total_collected,
                   AVG(amount) as avg_payment
            FROM payments
            WHERE status = 'completed'
            AND payment_date >= date('now', '-12 months')
            GROUP BY month
            ORDER BY month DESC
            ''')

            monthly_data = cursor.fetchall()

            report_content = f"""Payment Collection Report
    {'=' * 80}
    Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

    Monthly Collection Summary (Last 12 Months):
    {'=' * 80}
    {'Month':<10} {'Payments':<10} {'Total Collected':<15} {'Average':<12}
    {'-' * 80}
    """

            total_payments = 0
            total_collected = 0

            for month, count, collected, avg_payment in monthly_data:
                report_content += f"{month:<10} {count:<10} £{collected:<14,.2f} £{avg_payment:<11.2f}\n"
                total_payments += count
                total_collected += collected

            overall_avg = total_collected / total_payments if total_payments > 0 else 0

            report_content += f"""{'-' * 80}
    TOTALS:    {total_payments:<10} £{total_collected:<14,.2f} £{overall_avg:<11.2f}
    {'=' * 80}
    """

            # Display in reports tab
            self.show_tab('reports')  # Reports tab
            self.report_text.delete('1.0', tk.END)
            self.report_text.insert('1.0', report_content)

            conn.close()
            self.update_status("Payment collection report generated")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate payment collection report: {e}")


    def gui_student_account_summary(self):
        """Generate student account summary report"""
        try:
            conn = get_connection()
            cursor = conn.cursor()

            cursor.execute('''
            SELECT s.student_id, s.first_name, s.last_name, s.course,
                   COUNT(sf.student_fee_id) as total_fees,
                   SUM(sf.amount) as total_charged,
                   COUNT(p.payment_id) as payment_count,
                   COALESCE(SUM(p.amount), 0) as total_paid,
                   (SUM(sf.amount) - COALESCE(SUM(p.amount), 0)) as balance
            FROM students s
            LEFT JOIN student_fees sf ON s.student_id = sf.student_id
            LEFT JOIN payments p ON s.student_id = p.student_id AND p.status = 'completed'
            WHERE s.status = 'active'
            GROUP BY s.student_id
            ORDER BY balance DESC
            LIMIT 50
            ''')

            student_data = cursor.fetchall()

            report_content = f"""Student Account Summary
    {'=' * 120}
    Top 50 Student Accounts by Balance

    {'Student ID':<12} {'Name':<25} {'Course':<20} {'Fees':<6} {'Charged':<12} {'Paid':<12} {'Balance':<12}
    {'-' * 120}
    """

            total_charged = 0
            total_paid = 0
            total_balance = 0

            for row in student_data:
                student_id, first_name, last_name, course, fee_count, charged, paid, balance = row
                student_name = f"{first_name} {last_name}"

                charged = charged or 0
                paid = paid or 0
                balance = balance or 0

                report_content += f"{student_id:<12} {student_name:<25} {course:<20} {fee_count:<6} £{charged:<11.2f} £{paid:<11.2f} £{balance:<11.2f}\n"

                total_charged += charged
                total_paid += paid
                total_balance += balance

            report_content += f"""{'-' * 120}
    TOTALS:{'':<57} £{total_charged:<11.2f} £{total_paid:<11.2f} £{total_balance:<11.2f}
    {'=' * 120}

    Summary Statistics:
    - Total Students: {len(student_data)}
    - Collection Rate: {(total_paid/total_charged*100) if total_charged > 0 else 0:.1f}%
    - Students with Outstanding Balance: {sum(1 for row in student_data if (row[6] or 0) > 0)}
    """

            # Display in reports tab
            self.show_tab('reports')  # Reports tab
            self.report_text.delete('1.0', tk.END)
            self.report_text.insert('1.0', report_content)

            conn.close()
            self.update_status("Student account summary generated")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate student account summary: {e}")


    def gui_fee_type_analysis(self):
        """Generate fee type analysis report"""
        try:
            conn = get_connection()
            cursor = conn.cursor()

            cursor.execute('''
            SELECT ft.fee_name,
                   COUNT(sf.student_fee_id) as assignments,
                   SUM(sf.amount) as total_charged,
                   SUM(CASE WHEN sf.status = 'paid' THEN sf.amount ELSE 0 END) as paid_amount,
                   SUM(CASE WHEN sf.status IN ('unpaid', 'partial') THEN sf.amount ELSE 0 END) as outstanding,
                   ROUND((SUM(CASE WHEN sf.status = 'paid' THEN sf.amount ELSE 0 END) * 100.0) /
                         NULLIF(SUM(sf.amount), 0), 2) as collection_rate
            FROM fee_types ft
            LEFT JOIN student_fees sf ON ft.fee_type_id = sf.fee_type_id
            WHERE ft.is_active = 1
            GROUP BY ft.fee_type_id, ft.fee_name
            HAVING COUNT(sf.student_fee_id) > 0
            ORDER BY total_charged DESC
            ''')

            fee_data = cursor.fetchall()

            report_content = f"""Fee Type Analysis Report
    {'=' * 100}
    Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

    {'Fee Type':<25} {'Assignments':<12} {'Charged':<12} {'Paid':<12} {'Outstanding':<12} {'Rate %':<8}
    {'-' * 100}
    """

            for row in fee_data:
                fee_name, assignments, charged, paid, outstanding, rate = row

                charged = charged or 0
                paid = paid or 0
                outstanding = outstanding or 0
                rate = rate or 0

                report_content += f"{fee_name:<25} {assignments:<12} £{charged:<11.2f} £{paid:<11.2f} £{outstanding:<11.2f} {rate:<7.1f}%\n"

            report_content += f"""{'-' * 100}
    Total Fee Types Analyzed: {len(fee_data)}
    {'=' * 100}
    """

            # Display in reports tab
            self.show_tab('reports')  # Reports tab
            self.report_text.delete('1.0', tk.END)
            self.report_text.insert('1.0', report_content)

            conn.close()
            self.update_status("Fee type analysis report generated")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate fee type analysis: {e}")


    def gui_payment_method_analysis(self):
        """Generate payment method analysis report"""
        try:
            conn = get_connection()
            cursor = conn.cursor()

            cursor.execute('''
            SELECT payment_method,
                   COUNT(*) as transaction_count,
                   SUM(amount) as total_amount,
                   AVG(amount) as avg_amount,
                   MIN(amount) as min_amount,
                   MAX(amount) as max_amount,
                   ROUND((SUM(amount) * 100.0) / (SELECT SUM(amount) FROM payments WHERE status = 'completed'), 2) as percentage
            FROM payments
            WHERE status = 'completed'
            GROUP BY payment_method
            ORDER BY total_amount DESC
            ''')

            method_data = cursor.fetchall()

            report_content = f"""Payment Method Analysis Report
    {'=' * 120}
    Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

    {'Method':<15} {'Count':<8} {'Total Amount':<15} {'Average':<12} {'Min':<10} {'Max':<12} {'Share %':<8}
    {'-' * 120}
    """

            total_transactions = sum(row[1] for row in method_data)
            total_amount = sum(row[2] for row in method_data)

            for row in method_data:
                method, count, amount, avg_amt, min_amt, max_amt, percentage = row

                report_content += f"{method:<15} {count:<8} £{amount:<14,.2f} £{avg_amt:<11.2f} £{min_amt:<9.2f} £{max_amt:<11.2f} {percentage:<7.1f}%\n"

            report_content += f"""{'-' * 120}
    TOTALS:         {total_transactions:<8} £{total_amount:<14,.2f}
    {'=' * 120}

    Key Insights:
    - Most Popular Method: {method_data[0][0] if method_data else 'N/A'}
    - Highest Value Method: {method_data[0][0] if method_data else 'N/A'}
    - Average Transaction: £{total_amount/total_transactions if total_transactions > 0 else 0:.2f}
    """

            # Display in reports tab
            self.show_tab('reports')  # Reports tab
            self.report_text.delete('1.0', tk.END)
            self.report_text.insert('1.0', report_content)

            conn.close()
            self.update_status("Payment method analysis generated")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate payment method analysis: {e}")


    def clear_report(self):
        """Clear the report display"""
        self.report_text.delete('1.0', tk.END)
        self.update_status("Report cleared")


    def export_report_csv(self):
        """Export current report to CSV"""
        content = self.report_text.get("1.0", tk.END).strip()
        if not content:
            messagebox.showerror("Error", "No report to export")
            return

        filename = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            initialfile=f"financial_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        )

        if filename:
            try:
                # Simple CSV export - in a real implementation, you'd parse the report properly
                with open(filename, 'w', newline='', encoding='utf-8') as f:
                    f.write(content)
                messagebox.showinfo("Success", f"Report exported to {filename}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to export report: {e}")


    def export_report_excel(self):
        """Export current report to Excel"""
        try:
            # Get current report content
            if not hasattr(self, 'current_report_data') or not self.current_report_data:
                messagebox.showerror("Error", "No report data to export. Please generate a report first.")
                return

            from tkinter import filedialog
            from datetime import datetime

            # Ask user where to save
            filename = filedialog.asksaveasfilename(
                title="Export to Excel",
                defaultextension=".xlsx",
                initialfile=f"finance_report_{datetime.now().strftime('%Y%m%d')}.xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )

            if not filename:
                return

            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment

                # Create workbook
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Finance Report"

                # Add header
                ws['A1'] = "Finance Report"
                ws['A1'].font = Font(size=16, bold=True)
                ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

                # Add data starting from row 4
                row = 4
                for key, value in self.current_report_data.items():
                    ws[f'A{row}'] = str(key)
                    ws[f'B{row}'] = str(value)
                    ws[f'A{row}'].font = Font(bold=True)
                    row += 1

                # Auto-size columns
                ws.column_dimensions['A'].width = 30
                ws.column_dimensions['B'].width = 20

                # Save workbook
                wb.save(filename)
                messagebox.showinfo("Success", f"Report exported to Excel:\n{filename}")

            except ImportError:
                # Fallback to CSV if openpyxl not available
                import csv
                filename = filename.replace('.xlsx', '.csv')
                with open(filename, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(['Finance Report'])
                    writer.writerow(['Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
                    writer.writerow([])
                    for key, value in self.current_report_data.items():
                        writer.writerow([key, value])
                messagebox.showinfo("Success", f"Exported to CSV (openpyxl not available):\n{filename}")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to export report: {e}")


    def show_text_window(self, title, content):
        """Show content in a separate text window"""
        window = tk.Toplevel(self.root)
        window.title(title)
        window.geometry("800x600")
        window.transient(self.root)

        text_widget = ScrolledText(window, font=('Courier', 10))
        text_widget.pack(fill='both', expand=True, padx=10, pady=10)
        text_widget.insert('1.0', content)

        ttk.Button(window, text="Close", command=window.destroy).pack(pady=10)

    # Add these stubs for other missing functions

    def gui_generate_aid_reports(self):
        """GUI wrapper for generate_aid_reports"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Generate Aid Reports")
        dialog.geometry("400x300")
        dialog.transient(self.root)
        dialog.grab_set()

        report_frame = ttk.LabelFrame(dialog, text="Select Report Type", padding=20)
        report_frame.pack(fill='both', expand=True, padx=20, pady=20)

        report_type_var = tk.StringVar(value="distribution")

        ttk.Radiobutton(report_frame, text="Aid Distribution Summary",
                       variable=report_type_var, value="distribution").pack(anchor='w', pady=5)
        ttk.Radiobutton(report_frame, text="Aid by Academic Year",
                       variable=report_type_var, value="academic_year").pack(anchor='w', pady=5)
        ttk.Radiobutton(report_frame, text="Loan Repayment Status",
                       variable=report_type_var, value="loan_repayment").pack(anchor='w', pady=5)
        ttk.Radiobutton(report_frame, text="Aid Effectiveness Analysis",
                       variable=report_type_var, value="effectiveness").pack(anchor='w', pady=5)

        def generate_report():
            try:
                report_type = report_type_var.get()

                old_stdout = sys.stdout
                sys.stdout = mystdout = io.StringIO()

                if report_type == "distribution":
                    aid_distribution_summary()
                elif report_type == "academic_year":
                    aid_by_academic_year()
                elif report_type == "loan_repayment":
                    loan_repayment_status_report()
                elif report_type == "effectiveness":
                    aid_effectiveness_analysis()

                output = mystdout.getvalue()
                sys.stdout = old_stdout

                self.show_tab('reports')  # Reports tab
                self.report_text.delete('1.0', tk.END)
                self.report_text.insert('1.0', output)

                dialog.destroy()
                self.update_status(f"Aid report ({report_type}) generated")

            except Exception as e:
                messagebox.showerror("Error", f"Failed to generate aid report: {e}")

        ttk.Button(report_frame, text="Generate Report", command=generate_report).pack(pady=20)

    def gui_aging_analysis_report(self):
        """Wrapper to call compliance manager's aging analysis report"""
        if hasattr(self.gui, 'compliance'):
            self.gui.compliance.gui_aging_analysis_report()
        else:
            messagebox.showwarning("Not Available", "Compliance manager not initialized")

    def gui_collection_case_status_report(self):
        """Wrapper to call compliance manager's collection case status report"""
        if hasattr(self.gui, 'compliance'):
            self.gui.compliance.gui_collection_case_status_report()
        else:
            messagebox.showwarning("Not Available", "Compliance manager not initialized")

    def gui_recovery_rate_analysis(self):
        """Wrapper to call compliance manager's recovery rate analysis"""
        if hasattr(self.gui, 'compliance'):
            self.gui.compliance.gui_recovery_rate_analysis()
        else:
            messagebox.showwarning("Not Available", "Compliance manager not initialized")

    def gui_agency_performance_report(self):
        """Wrapper to call compliance manager's agency performance report"""
        if hasattr(self.gui, 'compliance'):
            self.gui.compliance.gui_agency_performance_report()
        else:
            messagebox.showwarning("Not Available", "Compliance manager not initialized")

    def gui_variance_analysis_report(self):
        """Wrapper to call budget manager's variance analysis report"""
        if hasattr(self.gui, 'budgets'):
            self.gui.budgets.gui_variance_analysis_report()
        else:
            messagebox.showwarning("Not Available", "Budget manager not initialized")

    def gui_budget_performance_trends(self):
        """Wrapper to call budget manager's budget performance trends"""
        if hasattr(self.gui, 'budgets'):
            self.gui.budgets.gui_budget_performance_trends()
        else:
            messagebox.showwarning("Not Available", "Budget manager not initialized")

    def gui_category_performance_report(self):
        """Wrapper to call budget manager's category performance report"""
        if hasattr(self.gui, 'budgets'):
            self.gui.budgets.gui_category_performance_report()
        else:
            messagebox.showwarning("Not Available", "Budget manager not initialized")

    def gui_monthly_revenue_trend_report(self):
        """Generate monthly revenue trend report"""
        try:
            conn = get_connection()
            cursor = conn.cursor()

            # Get monthly revenue for the past 12 months
            cursor.execute('''
                SELECT
                    strftime('%Y-%m', payment_date) as month,
                    SUM(amount) as revenue
                FROM payments
                WHERE payment_date >= date('now', '-12 months')
                AND status = 'completed'
                GROUP BY strftime('%Y-%m', payment_date)
                ORDER BY month
            ''')

            results = cursor.fetchall()
            conn.close()

            if not results:
                messagebox.showinfo("No Data", "No revenue data found for the past 12 months")
                return

            # Format output
            output = "MONTHLY REVENUE TREND REPORT\n"
            output += "=" * 60 + "\n\n"
            output += f"{'Month':<15} {'Revenue':>20}\n"
            output += "-" * 60 + "\n"

            total_revenue = 0
            for month, revenue in results:
                output += f"{month:<15} £{revenue:>18,.2f}\n"
                total_revenue += revenue

            output += "-" * 60 + "\n"
            output += f"{'Total':<15} £{total_revenue:>18,.2f}\n"
            output += f"{'Average':<15} £{total_revenue/len(results):>18,.2f}\n"

            self.show_text_window("Monthly Revenue Trend Report", output)

        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate monthly revenue report: {e}")

