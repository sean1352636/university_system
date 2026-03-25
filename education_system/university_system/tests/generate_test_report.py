#!/usr/bin/env python3
"""
Generate an XLSX report of all failed, errored, and skipped tests.

Usage:
    python -m education_system.university_system.tests.generate_test_report [pytest args...]

Examples:
    # Run all university tests and generate report
    python -m education_system.university_system.tests.generate_test_report

    # Run specific test directory
    python -m education_system.university_system.tests.generate_test_report education_system/university_system/tests/cli/

    # Pass extra pytest flags
    python -m education_system.university_system.tests.generate_test_report -x --timeout=30
"""

import json
import subprocess
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def run_tests(extra_args: list[str] | None = None) -> dict:
    """Run pytest with JSON report output and return parsed results."""
    report_path = Path("/tmp/pytest_report.json")
    cmd = [
        sys.executable, "-m", "pytest",
        "--tb=long",
        f"--json-report-file={report_path}",
        "--json-report",
        "-q",
    ]
    if extra_args:
        cmd.extend(extra_args)
    else:
        cmd.append("education_system/university_system/tests/")

    subprocess.run(
        cmd,
        cwd=str(Path(__file__).resolve().parents[3]),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
    )
    return json.loads(report_path.read_text())


def collect_problem_tests(report: dict) -> list[dict]:
    """Extract failed, errored, and skipped tests from the JSON report."""
    rows = []
    for test in report.get("tests", []):
        outcome = test.get("outcome", "")
        if outcome not in ("failed", "error", "skipped"):
            continue

        node_id = test.get("nodeid", "")
        # Split into file path and test name
        if "::" in node_id:
            file_path, test_name = node_id.split("::", 1)
        else:
            file_path, test_name = node_id, ""

        # Get the error/skip message
        message = ""
        # Check call phase first (most common for failures)
        for phase in ("call", "setup", "teardown"):
            phase_data = test.get(phase, {})
            if phase_data.get("longrepr"):
                message = phase_data["longrepr"]
                break
            if phase_data.get("message"):
                message = phase_data["message"]
                break

        # For skipped tests, the reason is often just in the top-level message
        if not message and outcome == "skipped":
            for phase in ("setup", "call", "teardown"):
                phase_data = test.get(phase, {})
                if phase_data.get("message"):
                    message = phase_data["message"]
                    break

        # Extract short error type from the message
        error_type = ""
        if outcome == "failed" or outcome == "error":
            for line in message.splitlines():
                stripped = line.strip()
                if stripped.startswith("E   "):
                    error_type = stripped[4:].split(":")[0]
                    break
        elif outcome == "skipped":
            error_type = "Skipped"

        duration = test.get("duration", 0)

        rows.append({
            "status": outcome.upper(),
            "file": file_path,
            "test": test_name,
            "error_type": error_type,
            "message": message[:10000],  # cap very long tracebacks
            "duration": round(duration, 3),
        })

    # Sort: errors first, then failures, then skipped
    order = {"ERROR": 0, "FAILED": 1, "SKIPPED": 2}
    rows.sort(key=lambda r: (order.get(r["status"], 9), r["file"], r["test"]))
    return rows


def write_xlsx(rows: list[dict], summary: dict, output_path: Path) -> None:
    """Write the test results to a formatted XLSX file."""
    wb = openpyxl.Workbook()

    # ── Summary sheet ──
    ws_summary = wb.active
    ws_summary.title = "Summary"

    header_font = Font(bold=True, size=14)
    ws_summary["A1"] = "Test Report"
    ws_summary["A1"].font = header_font

    ws_summary["A3"] = "Generated"
    ws_summary["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws_summary["A4"] = "Total Tests"
    ws_summary["B4"] = summary.get("total", 0)
    ws_summary["A5"] = "Passed"
    ws_summary["B5"] = summary.get("passed", 0)
    ws_summary["A6"] = "Failed"
    ws_summary["B6"] = summary.get("failed", 0)
    ws_summary["A7"] = "Errors"
    ws_summary["B7"] = summary.get("error", 0)
    ws_summary["A8"] = "Skipped"
    ws_summary["B8"] = summary.get("skipped", 0)
    ws_summary["A9"] = "Duration (s)"
    ws_summary["B9"] = round(summary.get("duration", 0), 2)

    for r in range(3, 10):
        ws_summary[f"A{r}"].font = Font(bold=True)

    pass_rate = 0
    total = summary.get("total", 0)
    if total:
        pass_rate = summary.get("passed", 0) / total * 100
    ws_summary["A11"] = "Pass Rate"
    ws_summary["B11"] = f"{pass_rate:.1f}%"
    ws_summary["A11"].font = Font(bold=True)

    ws_summary.column_dimensions["A"].width = 18
    ws_summary.column_dimensions["B"].width = 25

    # ── Details sheet ──
    ws = wb.create_sheet("Failed & Skipped Tests")

    # Styles
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
    skip_fill = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    top_alignment = Alignment(vertical="top")

    headers = ["#", "Status", "File", "Test", "Error Type", "Message", "Duration (s)"]
    col_widths = [5, 10, 50, 45, 30, 80, 12]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"
    # Auto-filter
    ws.auto_filter.ref = f"A1:G{max(len(rows) + 1, 2)}"

    status_fills = {"ERROR": error_fill, "FAILED": fail_fill, "SKIPPED": skip_fill}

    for row_idx, row_data in enumerate(rows, 2):
        fill = status_fills.get(row_data["status"])
        values = [
            row_idx - 1,
            row_data["status"],
            row_data["file"],
            row_data["test"],
            row_data["error_type"],
            row_data["message"],
            row_data["duration"],
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx == 6:  # message column
                cell.alignment = wrap_alignment
            else:
                cell.alignment = top_alignment
            if fill and col_idx == 2:
                cell.fill = fill

    if not rows:
        ws.cell(row=2, column=1, value="All tests passed!")
        ws.merge_cells("A2:G2")
        ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=2, column=1).font = Font(italic=True, size=12)

    wb.save(output_path)


def generate_report(extra_args: list[str] | None = None) -> Path:
    """Run tests and generate the XLSX report. Returns the output path."""
    # Check for pytest-json-report
    try:
        import pytest_jsonreport  # noqa: F401
    except ImportError:
        print("Installing pytest-json-report...")
        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", "pytest-json-report", "-q"]
        )

    print("Running tests...")
    report = run_tests(extra_args)

    summary_data = report.get("summary", {})
    total = sum(
        summary_data.get(k, 0)
        for k in ("passed", "failed", "error", "skipped")
    )
    summary_data["total"] = total
    summary_data["duration"] = report.get("duration", 0)

    rows = collect_problem_tests(report)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = Path(f"test_report_{timestamp}.xlsx")

    write_xlsx(rows, summary_data, output_path)

    n_fail = summary_data.get("failed", 0)
    n_err = summary_data.get("error", 0)
    n_skip = summary_data.get("skipped", 0)
    print(f"\nReport saved to: {output_path.resolve()}")
    print(f"  {n_fail} failed, {n_err} errors, {n_skip} skipped out of {total} tests")
    return output_path


if __name__ == "__main__":
    generate_report(sys.argv[1:] if len(sys.argv) > 1 else None)
