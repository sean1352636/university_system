"""Unified reporting engine for all education systems.

Improvement #14: Shared Reporting Engine

Generates PDF, Excel, and CSV reports with consistent branding across
all 4 education subsystems.
"""

import csv
import io
import logging
from datetime import datetime
from enum import Enum
from pathlib import Path

logger = logging.getLogger(__name__)


class ReportFormat(str, Enum):
    PDF = "pdf"
    EXCEL = "excel"
    CSV = "csv"


class ReportEngine:
    """Generate standardised reports across all education systems.

    Usage:
        engine = ReportEngine(system_name="Sixth Form College")

        # PDF
        pdf_bytes = engine.generate(
            title="Student Attendance Report",
            headers=["Student ID", "Name", "Attendance %"],
            rows=[["SFC0001", "John Doe", "95.2"], ...],
            format=ReportFormat.PDF,
        )

        # Excel
        xlsx_bytes = engine.generate(..., format=ReportFormat.EXCEL)

        # Save to file (auto-detect format from extension)
        engine.save("report.pdf", title=..., headers=..., rows=...)
    """

    def __init__(self, system_name: str = "Education System", logo_path: str | None = None):
        self.system_name = system_name
        self.logo_path = logo_path

    def generate(
        self,
        title: str,
        headers: list[str],
        rows: list[list],
        format: ReportFormat = ReportFormat.PDF,
        *,
        subtitle: str | None = None,
        footer: str | None = None,
        column_widths: list[float] | None = None,
        metadata: dict | None = None,
    ) -> bytes:
        """Generate a report in the specified format. Returns bytes."""
        if format == ReportFormat.PDF:
            return self._generate_pdf(title, headers, rows, subtitle, footer, column_widths)
        elif format == ReportFormat.EXCEL:
            return self._generate_excel(title, headers, rows, subtitle, metadata)
        elif format == ReportFormat.CSV:
            return self._generate_csv(headers, rows)
        else:
            raise ValueError(f"Unsupported format: {format}")

    def save(self, filepath: str | Path, title: str, headers: list[str],
             rows: list[list], format: ReportFormat | None = None, **kwargs) -> Path:
        """Generate and save a report to a file."""
        filepath = Path(filepath)
        if format is None:
            ext_map = {".pdf": ReportFormat.PDF, ".xlsx": ReportFormat.EXCEL, ".csv": ReportFormat.CSV}
            format = ext_map.get(filepath.suffix.lower(), ReportFormat.PDF)

        data = self.generate(title, headers, rows, format, **kwargs)
        filepath.parent.mkdir(parents=True, exist_ok=True)
        filepath.write_bytes(data)
        logger.info("Report saved: %s (%d bytes)", filepath, len(data))
        return filepath

    def _generate_pdf(self, title, headers, rows, subtitle, footer, column_widths):
        """Generate PDF report using reportlab."""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import mm
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        except ImportError:
            raise ImportError("reportlab is required for PDF generation")

        buffer = io.BytesIO()
        page_size = landscape(A4) if len(headers) > 6 else A4

        doc = SimpleDocTemplate(
            buffer, pagesize=page_size,
            topMargin=20 * mm, bottomMargin=20 * mm,
            leftMargin=15 * mm, rightMargin=15 * mm,
        )

        styles = getSampleStyleSheet()
        elements = []

        # Title
        title_style = ParagraphStyle("ReportTitle", parent=styles["Heading1"],
                                     fontSize=16, spaceAfter=6)
        elements.append(Paragraph(title, title_style))

        if subtitle:
            sub_style = ParagraphStyle("ReportSubtitle", parent=styles["Heading2"],
                                       fontSize=11, textColor=colors.grey, spaceAfter=4)
            elements.append(Paragraph(subtitle, sub_style))

        meta_style = ParagraphStyle("ReportMeta", parent=styles["Normal"],
                                     fontSize=8, textColor=colors.grey)
        elements.append(Paragraph(
            f"{self.system_name} | Generated: {datetime.now().strftime('%d %b %Y %H:%M')}",
            meta_style,
        ))
        elements.append(Spacer(1, 8 * mm))

        # Table
        table_data = [headers] + [[str(cell) for cell in row] for row in rows]

        if column_widths:
            table = Table(table_data, colWidths=[w * mm for w in column_widths])
        else:
            table = Table(table_data, repeatRows=1)

        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2c3e50")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("TOPPADDING", (0, 0), (-1, 0), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9fa")]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dee2e6")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(table)

        if footer:
            elements.append(Spacer(1, 10 * mm))
            elements.append(Paragraph(footer, meta_style))

        elements.append(Spacer(1, 3 * mm))
        elements.append(Paragraph(f"Total rows: {len(rows)}", meta_style))

        doc.build(elements)
        return buffer.getvalue()

    def _generate_excel(self, title, headers, rows, subtitle, metadata):
        """Generate Excel report using openpyxl."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise ImportError("openpyxl is required for Excel generation")

        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]

        row_num = 1
        ws.cell(row=row_num, column=1, value=title).font = Font(size=14, bold=True)
        row_num += 1

        if subtitle:
            ws.cell(row=row_num, column=1, value=subtitle).font = Font(size=10, color="666666")
            row_num += 1

        ws.cell(row=row_num, column=1,
                value=f"{self.system_name} | Generated: {datetime.now().strftime('%d %b %Y %H:%M')}"
                ).font = Font(size=8, color="999999")
        row_num += 2

        header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        thin_border = Border(
            left=Side(style="thin", color="dee2e6"),
            right=Side(style="thin", color="dee2e6"),
            top=Side(style="thin", color="dee2e6"),
            bottom=Side(style="thin", color="dee2e6"),
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
        row_num += 1

        alt_fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type="solid")
        for i, row in enumerate(rows):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.border = thin_border
                if i % 2 == 1:
                    cell.fill = alt_fill
            row_num += 1

        # Auto-width columns (limited to 26 columns for simplicity)
        for col_idx in range(1, min(len(headers) + 1, 27)):
            max_len = len(str(headers[col_idx - 1]))
            for row in rows[:100]:  # Sample first 100 rows
                if col_idx - 1 < len(row):
                    max_len = max(max_len, len(str(row[col_idx - 1])))
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

        row_num += 1
        ws.cell(row=row_num, column=1, value=f"Total rows: {len(rows)}").font = Font(size=8, color="999999")

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def _generate_csv(self, headers, rows):
        """Generate CSV report."""
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(headers)
        for row in rows:
            writer.writerow([str(cell) for cell in row])
        return buffer.getvalue().encode("utf-8")
