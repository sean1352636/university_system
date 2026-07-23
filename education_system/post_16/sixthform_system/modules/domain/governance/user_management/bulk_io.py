"""Bulk import / export for the Sixth Form User Management module.

* **Import**: read a CSV or XLSX file, validate every row against the
  same checks ``lifecycle.create_user_full`` uses, then either preview
  the result (dry-run) or commit it. Commit is **atomic at the row
  level** — every row's user-create + extras attach is wrapped in
  ``try/except`` and tracked; the report tells the admin exactly which
  rows landed and which were rejected.

* **Export**: pull every user (optionally filtered by sixth-form
  role), join extras, and write CSV / JSON / XLSX. Default format is
  CSV which has no third-party deps. XLSX requires ``openpyxl``;
  callers can detect that via ``XLSX_AVAILABLE``.

Schema-touching writes go through the existing ``lifecycle`` helpers
so audit + side-table writes stay consistent.
"""

from __future__ import annotations

import csv
import datetime
import json
import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable

from education_system.post_16.sixthform_system.modules.domain.governance.user_management import (
    lifecycle,
)
from education_system.post_16.sixthform_system.modules.shared import user_accounts as _ua
from education_system.post_16.sixthform_system.modules.shared.user_accounts import (
    DEFAULT_ROLE, SYSTEM_KEY, UserAccountError,
)

logger = logging.getLogger(__name__)

try:
    import openpyxl  # type: ignore
    XLSX_AVAILABLE = True
except Exception:                       # pragma: no cover
    XLSX_AVAILABLE = False


# ── Schema ─────────────────────────────────────────────────────────────

REQUIRED_FIELDS = ("username", "password")
OPTIONAL_FIELDS = (
    "display_name", "email", "role",
    "phone", "department", "job_title", "line_manager_id", "notes",
)
ALL_FIELDS = REQUIRED_FIELDS + OPTIONAL_FIELDS


# ── Models ─────────────────────────────────────────────────────────────

@dataclass
class ImportRow:
    """One parsed-and-validated row from the import file."""
    line_no: int
    fields: dict[str, Any] = field(default_factory=dict)
    errors: list[str] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return not self.errors


@dataclass
class ImportPreview:
    rows: list[ImportRow]

    @property
    def valid(self) -> list[ImportRow]:
        return [r for r in self.rows if r.ok]

    @property
    def invalid(self) -> list[ImportRow]:
        return [r for r in self.rows if not r.ok]


@dataclass
class ImportResult:
    created_user_ids: list[int] = field(default_factory=list)
    failed: list[tuple[int, str]] = field(default_factory=list)   # (line_no, error)

    @property
    def total(self) -> int:
        return len(self.created_user_ids) + len(self.failed)


# ── Parsing & validation ──────────────────────────────────────────────

def _normalise_header(name: str) -> str:
    return (name or "").strip().lower().replace(" ", "_")


def _coerce_int(s: Any) -> int | None:
    if s in (None, ""):
        return None
    try:
        return int(str(s).strip())
    except (TypeError, ValueError):
        raise UserAccountError(f"Not an integer: {s!r}")


def parse_file(path: str) -> list[dict[str, Any]]:
    """Read CSV or XLSX into a list of dicts keyed by lowercased header."""
    p = Path(path)
    if not p.is_file():
        raise UserAccountError(f"File not found: {path}")
    suffix = p.suffix.lower()
    if suffix == ".csv":
        return _parse_csv(p)
    if suffix in (".xlsx", ".xlsm"):
        if not XLSX_AVAILABLE:
            raise UserAccountError(
                "openpyxl not installed — install it or export the file "
                "as CSV first.")
        return _parse_xlsx(p)
    raise UserAccountError(
        f"Unsupported extension {suffix!r}. Use .csv or .xlsx.")


def _parse_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames is None:
            raise UserAccountError("CSV has no header row.")
        # Normalise headers
        reader.fieldnames = [_normalise_header(h) for h in reader.fieldnames]
        return [dict(row) for row in reader]


def _parse_xlsx(path: Path) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        raise UserAccountError("Workbook is empty.")
    header = [_normalise_header(h or "") for h in header]
    out: list[dict[str, Any]] = []
    for raw in rows_iter:
        if raw is None or all(c is None for c in raw):
            continue
        out.append({h: ("" if v is None else v)
                     for h, v in zip(header, raw)})
    return out


def validate_rows(records: Iterable[dict[str, Any]],
                   *, default_role: str = DEFAULT_ROLE) -> ImportPreview:
    """Validate parsed records. Header row is line 1; first data row
    is line 2. ``role`` defaults to ``default_role``."""
    rows: list[ImportRow] = []
    seen_usernames: set[str] = set()
    for idx, raw in enumerate(records, start=2):
        r = ImportRow(line_no=idx, fields={})
        # Pull every known field with a stripped default.
        for f in ALL_FIELDS:
            v = raw.get(f, "")
            r.fields[f] = (v.strip() if isinstance(v, str) else v) or ""
        # Required
        for f in REQUIRED_FIELDS:
            if not r.fields.get(f):
                r.errors.append(f"Missing required field: {f}")
        # Email shape (cheap)
        email = r.fields.get("email") or ""
        if email and "@" not in email:
            r.errors.append("Email looks malformed (no '@').")
        # Role
        from education_system.shared.auth.role_manager import ROLE_HIERARCHY
        role = r.fields.get("role") or default_role
        if role not in ROLE_HIERARCHY:
            r.errors.append(f"Unknown role: {role!r}")
        r.fields["role"] = role
        # Line-manager id
        try:
            r.fields["line_manager_id"] = _coerce_int(
                r.fields.get("line_manager_id"))
        except UserAccountError as e:
            r.errors.append(str(e))
        # Duplicate username inside the file
        uname = (r.fields.get("username") or "").strip()
        if uname:
            if uname in seen_usernames:
                r.errors.append(
                    f"Duplicate username inside file: {uname!r}")
            seen_usernames.add(uname)
        rows.append(r)
    return ImportPreview(rows=rows)


def find_existing_usernames(auth: Any,
                              usernames: Iterable[str]) -> set[str]:
    """Return the subset of ``usernames`` that already exist."""
    want = {u for u in usernames if u}
    if not want:
        return set()
    try:
        existing = {u.username for u in _ua.list_users(auth)}
    except Exception:
        logger.exception("find_existing_usernames: list_users failed")
        return set()
    return want & existing


def dry_run(auth: Any, records: Iterable[dict[str, Any]]) -> ImportPreview:
    """Validate + cross-check usernames against the DB."""
    preview = validate_rows(records)
    existing = find_existing_usernames(
        auth, (r.fields.get("username") or "" for r in preview.rows))
    for r in preview.rows:
        if (r.fields.get("username") or "") in existing:
            r.errors.append("Username already exists in the system.")
    return preview


# ── Commit ────────────────────────────────────────────────────────────

def commit_import(auth: Any,
                    preview: ImportPreview,
                    *, system_key: str = SYSTEM_KEY) -> ImportResult:
    """Create every valid row. Invalid rows are skipped (and surfaced
    in ``ImportResult.failed`` so the report shows them too)."""
    result = ImportResult()
    for r in preview.rows:
        if not r.ok:
            result.failed.append((r.line_no, "; ".join(r.errors)))
            continue
        f = r.fields
        try:
            user = lifecycle.create_user_full(
                auth,
                username=f["username"], password=f["password"],
                display_name=f.get("display_name") or None,
                email=f.get("email") or None,
                role=f.get("role") or DEFAULT_ROLE,
                system_key=system_key,
                phone=f.get("phone") or None,
                department=f.get("department") or None,
                job_title=f.get("job_title") or None,
                line_manager_id=f.get("line_manager_id"),
                notes=f.get("notes") or None,
            )
            result.created_user_ids.append(user.id)
        except UserAccountError as e:
            result.failed.append((r.line_no, str(e)))
        except Exception as e:
            logger.exception("commit_import row %d crashed", r.line_no)
            result.failed.append((r.line_no, f"Unexpected: {e}"))
    logger.info("Bulk import: created=%d, failed=%d",
                len(result.created_user_ids), len(result.failed))
    return result


# ── Export ────────────────────────────────────────────────────────────

EXPORT_FIELDS = (
    "id", "username", "display_name", "email",
    "sixthform_role", "is_active", "is_locked",
    "last_login", "password_changed_at",
    "phone", "department", "job_title", "line_manager_id", "notes",
    "created_at", "updated_at",
)


def _user_to_dict(auth: Any, u) -> dict[str, Any]:
    extras = lifecycle.get_extras(auth, u.id)
    return {
        "id": u.id, "username": u.username,
        "display_name": u.display_name or "",
        "email": u.email or "",
        "sixthform_role": u.role_for(SYSTEM_KEY) or "",
        "is_active": int(u.is_active),
        "is_locked": int(u.is_locked),
        "last_login": u.last_login or "",
        "password_changed_at": u.password_changed_at or "",
        "phone": extras.phone or "",
        "department": extras.department or "",
        "job_title": extras.job_title or "",
        "line_manager_id": extras.line_manager_id or "",
        "notes": extras.notes or "",
        "created_at": u.created_at or "",
        "updated_at": u.updated_at or "",
    }


def collect_users(auth: Any, *,
                    sixthform_role: str | None = None) -> list[dict[str, Any]]:
    if sixthform_role:
        rows = _ua.list_users(auth, system_key=SYSTEM_KEY,
                                role=sixthform_role)
    else:
        rows = _ua.list_users(auth)
    return [_user_to_dict(auth, u) for u in rows]


def export(auth: Any, path: str, *,
            fmt: str | None = None,
            sixthform_role: str | None = None) -> int:
    """Write users to ``path``. ``fmt`` defaults to the file extension.
    Returns the number of rows written."""
    p = Path(path)
    chosen = (fmt or p.suffix.lstrip(".") or "csv").lower()
    rows = collect_users(auth, sixthform_role=sixthform_role)
    if chosen == "csv":
        _write_csv(p, rows)
    elif chosen == "json":
        _write_json(p, rows)
    elif chosen == "xlsx":
        if not XLSX_AVAILABLE:
            raise UserAccountError(
                "openpyxl not installed — export as CSV or JSON instead.")
        _write_xlsx(p, rows)
    else:
        raise UserAccountError(
            f"Unsupported format {chosen!r} (csv/json/xlsx).")
    logger.info("Exported %d user(s) to %s (%s)", len(rows), path, chosen)
    return len(rows)


def _write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(EXPORT_FIELDS))
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in EXPORT_FIELDS})


def _write_json(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump({"exported_at":
                   datetime.datetime.now().isoformat(timespec='seconds'),
                   "count": len(rows), "rows": rows},
                  f, indent=2, default=str)


def _write_xlsx(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Users"
    ws.append(list(EXPORT_FIELDS))
    for r in rows:
        ws.append([r.get(k, "") for k in EXPORT_FIELDS])
    wb.save(path)


# ── Template helpers ──────────────────────────────────────────────────

def write_template_csv(path: str) -> None:
    """Write a blank CSV with the expected headers so admins know what
    columns to fill in."""
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    with p.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(ALL_FIELDS)
        w.writerow(["jdoe", "ChangeMe!23", "Jane Doe", "j.doe@example.org",
                    "staff", "07111000000", "Mathematics", "Teacher", "",
                    "Sample row — replace or delete."])
